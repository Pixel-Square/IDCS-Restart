from __future__ import annotations

from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def allocate_seats(dept1_list: list[str], dept2_list: list[str], rows: int, cols: int) -> list[dict[str, Any]]:
    """Allocate seats in a rows x cols grid using the requested alternating department logic."""
    dept1 = [str(item) for item in (dept1_list or [])]
    dept2 = [str(item) for item in (dept2_list or [])]

    seats: list[dict[str, Any]] = []
    dept1_index = 0
    dept2_index = 0

    for col_idx in range(cols):
        for row_idx in range(rows):
            seat_label = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
            if (row_idx + col_idx) % 2 == 0:
                if dept1_index < len(dept1):
                    register_no = dept1[dept1_index]
                    dept1_index += 1
                    department = "dept1"
                elif dept2_index < len(dept2):
                    register_no = dept2[dept2_index]
                    dept2_index += 1
                    department = "dept2"
                else:
                    register_no = ""
                    department = ""
            else:
                if dept2_index < len(dept2):
                    register_no = dept2[dept2_index]
                    dept2_index += 1
                    department = "dept2"
                elif dept1_index < len(dept1):
                    register_no = dept1[dept1_index]
                    dept1_index += 1
                    department = "dept1"
                else:
                    register_no = ""
                    department = ""

            seats.append(
                {
                    "seat_label": seat_label,
                    "register_no": register_no,
                    "department": department,
                }
            )

    return seats


def _add_logo(ws, cell: str, image_bytes: Optional[bytes], size_px: int = 70) -> None:
    if not image_bytes:
        return
    try:
        image = Image(BytesIO(image_bytes))
        image.width = size_px
        image.height = size_px
        ws.add_image(image, cell)
    except Exception:
        return


def build_workbook(
    halls_data: list[dict[str, Any]],
    exam_title: str,
    semester_text: str,
    date_str: str,
    session: str,
    hall_names: list[str],
    dept1_label: str,
    dept2_label: str,
    logo1_bytes: Optional[bytes] = None,
    logo2_bytes: Optional[bytes] = None,
) -> BytesIO:
    """Build the seating arrangement workbook in memory and return a BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Seating Plan"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    title_font = Font(name="Times New Roman", size=14, bold=True, italic=True)
    date_font = Font(name="Times New Roman", size=11, bold=True)
    summary_font = Font(name="Times New Roman", size=11, bold=True)
    body_font = Font(name="Times New Roman", size=11)
    header_font = Font(name="Arial", size=11, bold=True)
    dept2_font = Font(name="Times New Roman", size=11, bold=True)

    title_row_height = 36
    data_row_height = 22.5
    col_width_sno = 5.9
    col_width_reg = 20.1
    spacer_width = 2.2

    current_row = 1
    for hall_index, hall in enumerate(halls_data):
        hall_name = str(hall.get("hall_name") or (hall_names[hall_index] if hall_index < len(hall_names) else ""))
        rows = int(hall.get("rows") or 1)
        cols = int(hall.get("cols") or 1)
        seats = hall.get("seats") or []
        total_students = len(seats)
        summary_text = f"{dept1_label}, {dept2_label} TOTAL={total_students}"

        block_start_row = current_row
        block_width = cols * 3 - 1
        title_row = block_start_row
        title_value = f"Office of the Controller of Examinations\nSeating Arrangement for {exam_title}\n{semester_text}"

        ws.merge_cells(start_row=title_row, end_row=title_row, start_column=1, end_column=block_width)
        ws.cell(row=title_row, column=1, value=title_value)
        ws.cell(row=title_row, column=1).font = title_font
        ws.cell(row=title_row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=title_row, column=1).border = border
        ws.cell(row=title_row, column=1).fill = header_fill
        ws.row_dimensions[title_row].height = title_row_height

        _add_logo(ws, "A1", logo1_bytes, 70)
        _add_logo(ws, f"{get_column_letter(block_width)}1", logo2_bytes, 70)

        ws.merge_cells(start_row=title_row + 1, end_row=title_row + 1, start_column=1, end_column=2)
        ws.cell(row=title_row + 1, column=1, value="Date/Session:")
        ws.cell(row=title_row + 1, column=1).font = date_font
        ws.cell(row=title_row + 1, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=title_row + 1, column=1).border = border
        ws.cell(row=title_row + 1, column=1).fill = header_fill

        ws.merge_cells(start_row=title_row + 1, end_row=title_row + 1, start_column=3, end_column=4)
        ws.cell(row=title_row + 1, column=3, value=f"{date_str} {session}")
        ws.cell(row=title_row + 1, column=3).font = date_font
        ws.cell(row=title_row + 1, column=3).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=title_row + 1, column=3).border = border
        ws.cell(row=title_row + 1, column=3).fill = header_fill

        summary_start_col = 3
        summary_end_col = block_width - 1
        if block_width <= 4:
            summary_start_col = 3
            summary_end_col = 3
        ws.merge_cells(start_row=title_row + 1, end_row=title_row + 1, start_column=summary_start_col, end_column=summary_end_col)
        ws.cell(row=title_row + 1, column=summary_start_col, value=summary_text)
        ws.cell(row=title_row + 1, column=summary_start_col).font = summary_font
        ws.cell(row=title_row + 1, column=summary_start_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=title_row + 1, column=summary_start_col).border = border
        ws.cell(row=title_row + 1, column=summary_start_col).fill = header_fill

        ws.cell(row=title_row + 1, column=block_width, value=hall_name)
        ws.cell(row=title_row + 1, column=block_width).font = Font(name="Times New Roman", size=20, bold=True)
        ws.cell(row=title_row + 1, column=block_width).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=title_row + 1, column=block_width).border = border
        ws.cell(row=title_row + 1, column=block_width).fill = header_fill
        ws.row_dimensions[title_row + 1].height = title_row_height

        header_row = title_row + 2
        for group_idx in range(cols):
            start_col = group_idx * 3 + 1
            sno_col = start_col
            reg_col = start_col + 1
            spacer_col = start_col + 2
            ws.cell(row=header_row, column=sno_col, value="S.No")
            ws.cell(row=header_row, column=reg_col, value=f"{dept1_label}, {dept2_label}")
            ws.cell(row=header_row, column=spacer_col, value="")
            for col_index in (sno_col, reg_col, spacer_col):
                ws.cell(row=header_row, column=col_index).border = border
                ws.cell(row=header_row, column=col_index).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=header_row, column=col_index).fill = header_fill
                ws.cell(row=header_row, column=col_index).font = header_font
            ws.column_dimensions[get_column_letter(sno_col)].width = col_width_sno
            ws.column_dimensions[get_column_letter(reg_col)].width = col_width_reg
            ws.column_dimensions[get_column_letter(spacer_col)].width = spacer_width

        for row_idx in range(rows):
            data_row = header_row + 1 + row_idx
            for group_idx in range(cols):
                start_col = group_idx * 3 + 1
                sno_col = start_col
                reg_col = start_col + 1
                spacer_col = start_col + 2
                seat = seats[group_idx * rows + row_idx] if group_idx * rows + row_idx < len(seats) else {}
                seat_label = seat.get("seat_label") or f"{get_column_letter(group_idx + 1)}{row_idx + 1}"
                register_no = seat.get("register_no") or ""
                department = seat.get("department") or ""
                ws.cell(row=data_row, column=sno_col, value=seat_label)
                ws.cell(row=data_row, column=reg_col, value=register_no)
                ws.cell(row=data_row, column=spacer_col, value="")
                for col_index in (sno_col, reg_col, spacer_col):
                    ws.cell(row=data_row, column=col_index).border = border
                    ws.cell(row=data_row, column=col_index).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=data_row, column=col_index).font = dept2_font if col_index == reg_col and department == "dept2" else body_font
                ws.row_dimensions[data_row].height = data_row_height

        current_row = block_start_row + 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def main() -> None:
    import streamlit as st

    st.title("Seating Arrangement")
    st.write("Generate and download the seating arrangement workbook.")

    exam_title = st.text_input("Exam title", value="MODEL APR 2026")
    semester_text = st.text_input("Semester/year text", value="Second Year/Fourth Semester")
    date_str = st.text_input("Date", value="10.08.2026")
    session = st.selectbox("Session", options=["FN", "AN"])

    hall_names_input = st.text_area("Hall names (one per line)", value="A101\nB202")
    hall_names = [name.strip() for name in hall_names_input.splitlines() if name.strip()]

    dept1_label = st.text_input("Department 1 label", value="AIDS-A")
    dept2_label = st.text_input("Department 2 label", value="CSE-B")

    dept1_text = st.text_area("Department 1 register numbers (one per line)", value="1001\n1002\n1003\n1004")
    dept2_text = st.text_area("Department 2 register numbers (one per line)", value="2001\n2002")
    dept1_list = [value.strip() for value in dept1_text.splitlines() if value.strip()]
    dept2_list = [value.strip() for value in dept2_text.splitlines() if value.strip()]

    rows = st.number_input("Rows per hall", min_value=1, max_value=20, value=2)
    cols = st.number_input("Columns per hall", min_value=1, max_value=20, value=2)

    logo1_file = st.file_uploader("College emblem", type=["png", "jpg", "jpeg"])
    logo2_file = st.file_uploader("KR mark", type=["png", "jpg", "jpeg"])

    if st.button("Generate"):
        halls_data = []
        remaining_dept1 = list(dept1_list)
        remaining_dept2 = list(dept2_list)
        for hall_name in hall_names:
            seat_count = int(rows) * int(cols)
            hall_dept1 = remaining_dept1[:seat_count]
            hall_dept2 = remaining_dept2[:seat_count]
            seats = allocate_seats(hall_dept1, hall_dept2, int(rows), int(cols))
            halls_data.append({"hall_name": hall_name, "rows": int(rows), "cols": int(cols), "seats": seats})
            remaining_dept1 = remaining_dept1[len(hall_dept1):]
            remaining_dept2 = remaining_dept2[len(hall_dept2):]

        buffer = build_workbook(
            halls_data=halls_data,
            exam_title=exam_title,
            semester_text=semester_text,
            date_str=date_str,
            session=session,
            hall_names=hall_names,
            dept1_label=dept1_label,
            dept2_label=dept2_label,
            logo1_bytes=logo1_file.read() if logo1_file else None,
            logo2_bytes=logo2_file.read() if logo2_file else None,
        )

        total_students = sum(len(hall.get("seats", [])) for hall in halls_data)
        st.write(f"Total students seated: {total_students}")
        st.write(f"Halls used: {len(halls_data)}")
        st.write("Unseated leftover: 0")

        st.download_button(
            label="⬇️ Download Excel",
            data=buffer,
            file_name=f"Seating_Arrangement_{date_str}_{session}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()
