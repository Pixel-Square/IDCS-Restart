from __future__ import annotations

import io
import itertools

from django.db import transaction, IntegrityError
from django.db.models.deletion import ProtectedError, RestrictedError
from django.shortcuts import get_object_or_404
from rest_framework import generics, permissions, filters, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from .models import College, FeatureCatalog, CollegeFeature
from .serializers import CollegeSerializer, CollegeUserSerializer, CollegeFeatureSerializer
from accounts.models import User, Role, UserRole
from academics.models import StudentProfile, StaffProfile, Department, Course


def _resolve_college_id(request):
    """Return a college_id from the request: X-College-Id header > POST body > query param.

    This ensures college isolation regardless of how the frontend sends the
    college identifier — explicit header, JSON body field ``college``, or URL
    query param ``college_id``.
    """
    # 1. X-College-Id header
    cid = getattr(request, 'college_id', None)
    if cid is not None:
        return cid
    
    header_cid = request.META.get('HTTP_X_COLLEGE_ID')
    if header_cid:
        try:
            return int(header_cid)
        except (ValueError, TypeError):
            pass
    # 2. Request body (JSON: {"college": 5})
    cid = (request.data or {}).get('college')
    if cid is not None:
        try:
            return int(cid)
        except (TypeError, ValueError):
            pass
    # 3. Query param (?college_id=5)
    raw = request.query_params.get('college_id')
    if raw:
        try:
            return int(raw)
        except (TypeError, ValueError):
            pass
    return None
from accounts.permissions_super_admin import IsSuperAdminOrSuperuser
from .permissions import IsCollegeAdminOrSuperAdmin


# ---------------------------------------------------------------------------
# College CRUD
# ---------------------------------------------------------------------------

class CollegeListCreateView(generics.ListCreateAPIView):
    """List all colleges or create a new one. Accessible to super admins, and college admins see only theirs."""
    serializer_class = CollegeSerializer
    permission_classes = [IsCollegeAdminOrSuperAdmin]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['code', 'name', 'short_name', 'city']
    ordering_fields = ['code', 'name', 'city', 'created_at']
    ordering = ['name']

    def get_queryset(self):
        qs = College.objects.all()
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == 'true')
        
        user = self.request.user
        is_super = user.is_superuser or user.roles.filter(name='SUPER_ADMIN').exists()
        if not is_super:
            from .permissions import _user_college_id
            uid = _user_college_id(user)
            if uid is not None:
                qs = qs.filter(pk=uid)
            else:
                qs = qs.none()
                
        return qs


class CollegeDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete a college. Requires SUPER_ADMIN."""
    serializer_class = CollegeSerializer
    permission_classes = [IsCollegeAdminOrSuperAdmin]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    queryset = College.objects.all()

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except (ProtectedError, RestrictedError) as e:
            protected_objs = list(getattr(e, 'protected_objects', []))
            count = len(protected_objs)
            model_names = sorted(set(obj._meta.verbose_name_plural.title() for obj in protected_objs[:10])) if protected_objs else []
            model_str = f" ({', '.join(model_names)})" if model_names else ""
            msg = f"Cannot delete college because it has {count} dependent protected record(s){model_str}. Please remove or reassign those records first, or deactivate the college."
            return Response({'detail': msg}, status=status.HTTP_409_CONFLICT)
        except IntegrityError as e:
            return Response({'detail': f"Cannot delete college due to database constraint: {str(e)}"}, status=status.HTTP_409_CONFLICT)
        except Exception as e:
            return Response({'detail': f"Failed to delete college: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


# ---------------------------------------------------------------------------
# College Tier — Update
# ---------------------------------------------------------------------------

class CollegeTierUpdateView(APIView):
    """PATCH /api/college/colleges/<id>/tier/
    Update the subscription tier of a college. Body: { "tier": "PRO" }
    Only SUPER_ADMIN or superusers may call this endpoint.
    """
    permission_classes = [IsSuperAdminOrSuperuser]

    VALID_TIERS = {'BASIC', 'PRO', 'PREMIUM'}

    def patch(self, request, pk):
        college = get_object_or_404(College, pk=pk)
        tier = (request.data.get('tier') or '').strip().upper()
        if tier not in self.VALID_TIERS:
            return Response(
                {'detail': f'Invalid tier "{tier}". Must be one of: {", ".join(sorted(self.VALID_TIERS))}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        college.tier = tier
        college.save(update_fields=['tier'])
        return Response({'id': college.pk, 'code': college.code, 'tier': college.tier})


# ---------------------------------------------------------------------------
# College Users — List
# ---------------------------------------------------------------------------

class CollegeUsersListView(APIView):
    """GET /api/college/colleges/<id>/users/  — list all users for a college.

    Query params:
      search     — filter by name, email, reg_no, staff_id
      role       — filter by role name (e.g. STUDENT, FACULTY, SUPER_ADMIN)
      page       — page number (default 1)

      page_size  — results per page (default 50, max 200)

    Response shape:
      {
        "total": <int>,
        "total_students": <int>,
        "total_staff": <int>,
        "page": <int>,
        "page_size": <int>,
        "results": [ ... ]
      }
    """
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request, pk):
        from django.db.models import Q
        from django.core.paginator import Paginator, EmptyPage

        college = get_object_or_404(College, pk=pk)
        self.check_object_permissions(request, college)
        search = request.query_params.get('search', '').strip()
        role_filter = request.query_params.get('role', '').strip().upper()

        try:
            page = max(1, int(request.query_params.get('page', 1)))
        except (TypeError, ValueError):
            page = 1
        try:
            page_size = min(200, max(1, int(request.query_params.get('page_size', 50))))
        except (TypeError, ValueError):
            page_size = 50

        # ── Students ────────────────────────────────────────────────────────
        students_qs = (
            StudentProfile.objects
            .filter(college=college)
            .select_related('user', 'home_department')
            .order_by('user__first_name', 'user__last_name', 'reg_no')
        )

        # ── Staff ───────────────────────────────────────────────────────────
        staff_qs = (
            StaffProfile.objects
            .filter(college=college)
            .select_related('user', 'department')
            .order_by('user__first_name', 'user__last_name', 'staff_id')
        )

        # ── Role filter ─────────────────────────────────────────────────────
        if role_filter:
            roles_list = [r.strip() for r in role_filter.split(',') if r.strip()]
            students_qs = students_qs.filter(
                user__user_roles__role__name__in=roles_list
            ).distinct()
            staff_qs = staff_qs.filter(
                user__user_roles__role__name__in=roles_list
            ).distinct()

        # ── Department filter ───────────────────────────────────────────────
        department_filter = request.query_params.get('department')
        if department_filter:
            try:
                dept_id = int(department_filter)
                students_qs = students_qs.filter(home_department_id=dept_id)
                staff_qs = staff_qs.filter(department_id=dept_id)
            except ValueError:
                pass

        # ── Search filter ───────────────────────────────────────────────────
        if search:
            q_student = (
                Q(user__username__icontains=search)
                | Q(user__email__icontains=search)
                | Q(user__first_name__icontains=search)
                | Q(user__last_name__icontains=search)
                | Q(reg_no__icontains=search)
            )
            q_staff = (
                Q(user__username__icontains=search)
                | Q(user__email__icontains=search)
                | Q(user__first_name__icontains=search)
                | Q(user__last_name__icontains=search)
                | Q(staff_id__icontains=search)
            )
            students_qs = students_qs.filter(q_student)
            staff_qs = staff_qs.filter(q_staff)

        # ── Counts (before pagination) ───────────────────────────────────────
        total_students = students_qs.count()
        total_staff = staff_qs.count()
        total = total_students + total_staff

        # ── Paginate across combined list ───────────────────────────────────
        # Students are listed first, then staff.
        all_profiles = list(students_qs) + list(staff_qs)
        paginator = Paginator(all_profiles, page_size)
        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        serializer = CollegeUserSerializer(page_obj.object_list, many=True)
        return Response({
            'total': total,
            'total_students': total_students,
            'total_staff': total_staff,
            'page': page_obj.number,
            'page_size': page_size,
            'total_pages': paginator.num_pages,
            'results': serializer.data,
        })



# ---------------------------------------------------------------------------
# College Users — Import Template Download
# ---------------------------------------------------------------------------

class CollegeUserImportTemplateView(APIView):
    """GET /api/college/colleges/<id>/users/import-template/?role=STUDENT
    Download an Excel template for the given role type."""
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request, pk):
        college = get_object_or_404(College, pk=pk)
        self.check_object_permissions(request, college)
        role = request.query_params.get('role', 'STUDENT').strip().upper()

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
        except ImportError:
            return Response({'detail': 'openpyxl not installed'}, status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = f'{role} Import'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_align = Alignment(horizontal='center')

        if role == 'STUDENT':
            from academics.models import Batch, Department

            # Fetch college-specific data for dropdowns
            batch_names = list(
                Batch.objects.filter(college=college)
                .order_by('name')
                .values_list('name', flat=True)
                .distinct()
            )
            dept_codes = list(
                Department.objects.filter(college=college)
                .order_by('code')
                .values_list('code', flat=True)
            )
            status_values = ['ACTIVE', 'INACTIVE', 'ALUMNI', 'DETAINED']

            headers = ['Register Number*', 'Email*', 'First Name', 'Last Name',
                       'Batch', 'Status', 'Department Code', 'Phone Number']
            example = ['22CS001', 'student@college.edu', 'John', 'Doe',
                       batch_names[0] if batch_names else '2022',
                       'ACTIVE',
                       dept_codes[0] if dept_codes else 'CSE',
                       '9876543210']

            # Write main headers & example
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                ws.column_dimensions[cell.column_letter].width = max(20, len(h) + 4)
            for col_idx, val in enumerate(example, start=1):
                ws.cell(row=2, column=col_idx, value=val)

            # ---- Build a hidden Lookup sheet for dropdown data ----
            lws = wb.create_sheet(title='_Lookups')
            lws.sheet_state = 'hidden'
            for r, b in enumerate(batch_names, start=1):
                lws.cell(row=r, column=1, value=b)
            for r, d in enumerate(dept_codes, start=1):
                lws.cell(row=r, column=2, value=d)
            for r, s in enumerate(status_values, start=1):
                lws.cell(row=r, column=3, value=s)

            max_rows = 1000  # apply dropdowns to data rows 2..1001

            # Batch dropdown — column E (5)
            if batch_names:
                batch_end = len(batch_names)
                dv_batch = DataValidation(
                    type='list',
                    formula1=f'_Lookups!$A$1:$A${batch_end}',
                    showDropDown=False,
                    allow_blank=True,
                    showErrorMessage=True,
                    error='Choose a batch from the list.',
                    errorTitle='Invalid Batch',
                )
                ws.add_data_validation(dv_batch)
                dv_batch.sqref = f'E2:E{max_rows}'

            # Status dropdown — column F (6)
            dv_status = DataValidation(
                type='list',
                formula1=f'_Lookups!$C$1:$C${len(status_values)}',
                showDropDown=False,
                allow_blank=True,
                showErrorMessage=True,
                error='Choose a status from the list.',
                errorTitle='Invalid Status',
            )
            ws.add_data_validation(dv_status)
            dv_status.sqref = f'F2:F{max_rows}'

            # Department dropdown — column G (7)
            if dept_codes:
                dept_end = len(dept_codes)
                dv_dept = DataValidation(
                    type='list',
                    formula1=f'_Lookups!$B$1:$B${dept_end}',
                    showDropDown=False,
                    allow_blank=True,
                    showErrorMessage=True,
                    error='Choose a department from the list.',
                    errorTitle='Invalid Department',
                )
                ws.add_data_validation(dv_dept)
                dv_dept.sqref = f'G2:G{max_rows}'

        elif role in ('FACULTY', 'STAFF'):
            from academics.models import Department, StaffProfile

            # Fetch college-specific departments
            staff_dept_codes = list(
                Department.objects.filter(college=college)
                .order_by('code')
                .values_list('code', flat=True)
            )
            # Fetch all unique designations from across the system
            designations = list(
                StaffProfile.objects.exclude(designation__isnull=True)
                .exclude(designation='')
                .values_list('designation', flat=True)
                .distinct()
                .order_by('designation')
            )
            # Normalise: deduplicate case-insensitively, keep title-case version
            seen = {}
            for d in designations:
                key = d.strip().lower()
                if key not in seen:
                    seen[key] = d.strip()
            designations = sorted(seen.values(), key=lambda x: x.lower())
            # Ensure common ones are always present
            for common in ['Professor', 'Associate Professor', 'Assistant Professor',
                           'Assistant Professor - SG', 'Assistant Professor - SS',
                           'Teaching Assistant', 'Tutor', 'Research Scholar',
                           'Placement Trainer', 'Attender', 'Lab Assistant']:
                if common.lower() not in {d.lower() for d in designations}:
                    designations.append(common)
            designations = sorted(designations, key=lambda x: x.lower())

            staff_status_values = ['ACTIVE', 'INACTIVE', 'RESIGNED', 'RETIRED']

            headers = ['Staff ID*', 'Email*', 'First Name', 'Last Name',
                       'Department Code', 'Designation', 'Status', 'Phone Number']
            example = ['FAC001', 'faculty@college.edu', 'Jane', 'Smith',
                       staff_dept_codes[0] if staff_dept_codes else 'CSE',
                       'Assistant Professor', 'ACTIVE', '9876543210']

            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                ws.column_dimensions[cell.column_letter].width = max(20, len(h) + 4)
            for col_idx, val in enumerate(example, start=1):
                ws.cell(row=2, column=col_idx, value=val)

            # Hidden Lookup sheet
            lws = wb.create_sheet(title='_Lookups')
            lws.sheet_state = 'hidden'
            for r, d in enumerate(staff_dept_codes, start=1):
                lws.cell(row=r, column=1, value=d)
            for r, d in enumerate(designations, start=1):
                lws.cell(row=r, column=2, value=d)
            for r, s in enumerate(staff_status_values, start=1):
                lws.cell(row=r, column=3, value=s)

            max_rows = 1000

            # Department dropdown — column E (5)
            if staff_dept_codes:
                dv_dept = DataValidation(
                    type='list',
                    formula1=f'_Lookups!$A$1:$A${len(staff_dept_codes)}',
                    showDropDown=False, allow_blank=True,
                    showErrorMessage=True,
                    error='Choose a department from the list.',
                    errorTitle='Invalid Department',
                )
                ws.add_data_validation(dv_dept)
                dv_dept.sqref = f'E2:E{max_rows}'

            # Designation dropdown — column F (6)
            if designations:
                dv_desig = DataValidation(
                    type='list',
                    formula1=f'_Lookups!$B$1:$B${len(designations)}',
                    showDropDown=False, allow_blank=True,
                    showErrorMessage=True,
                    error='Choose a designation from the list.',
                    errorTitle='Invalid Designation',
                )
                ws.add_data_validation(dv_desig)
                dv_desig.sqref = f'F2:F{max_rows}'

            # Status dropdown — column G (7)
            dv_status = DataValidation(
                type='list',
                formula1=f'_Lookups!$C$1:$C${len(staff_status_values)}',
                showDropDown=False, allow_blank=True,
                showErrorMessage=True,
                error='Choose a status from the list.',
                errorTitle='Invalid Status',
            )
            ws.add_data_validation(dv_status)
            dv_status.sqref = f'G2:G{max_rows}'
        else:
            # Generic / other roles
            headers = ['Register Number / Staff ID*', 'Email*', 'First Name', 'Last Name',
                       'Profile Type* (STUDENT/STAFF)', 'Department Code', 'Designation',
                       'Batch', 'Status', 'Phone Number']
            example = ['ID001', 'user@college.edu', 'Alex', 'Kumar',
                       'STAFF', 'CSE', 'Lab Assistant', '', 'ACTIVE', '9876543210']
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                ws.column_dimensions[cell.column_letter].width = max(20, len(h) + 4)
            for col_idx, val in enumerate(example, start=1):
                ws.cell(row=2, column=col_idx, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from django.http import HttpResponse
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{college.code}_{role}_import_template.xlsx"'
        return response


# ---------------------------------------------------------------------------
# College Users — Import
# ---------------------------------------------------------------------------

class CollegeUserImportView(APIView):
    """POST /api/college/colleges/<id>/users/import/
    Upload an Excel file to bulk-create users for this college."""
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def post(self, request, pk):
        college = get_object_or_404(College, pk=pk)
        self.check_object_permissions(request, college)
        role = request.data.get('role', 'STUDENT').strip().upper()
        file = request.FILES.get('file')

        if not file:
            return Response({'detail': 'No file uploaded.'}, status=400)

        name = getattr(file, 'name', '') or ''
        if not name.lower().endswith('.xlsx'):
            return Response({'detail': 'Please upload an .xlsx file.'}, status=400)

        try:
            from openpyxl import load_workbook
            content = file.read()
            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Failed to read Excel file.'}, status=400)

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return Response({'detail': 'No data rows found.'}, status=400)

        created_count = 0
        updated_count = 0
        failed = []
        processed_count = 0  # count only non-blank rows

        for row_idx, row in enumerate(rows, start=2):
            try:
                row = list(row) + [''] * 20  # pad to avoid index errors
                cells = [str(c).strip() if c is not None else '' for c in row]

                # Skip completely blank rows (e.g. empty rows from dropdown range)
                if not any(cells[:8]):
                    continue

                processed_count += 1

                with transaction.atomic():
                    if role == 'STUDENT':
                        user_obj, profile = self._import_student(cells, college)
                    elif role in ('FACULTY', 'STAFF'):
                        user_obj, profile = self._import_staff(cells, college)
                    else:
                        user_obj, profile = self._import_other(cells, college, role)

                    # Assign roles
                    target_role = role if role not in ('FACULTY',) else 'STAFF'
                    if role == 'STUDENT':
                        target_role = 'STUDENT'
                    role_obj, _ = Role.objects.get_or_create(name=target_role)
                    if not user_obj.user_roles.filter(role=role_obj).exists():
                        UserRole.objects.create(user=user_obj, role=role_obj)

                    # For FACULTY specifically, also add FACULTY role
                    if role == 'FACULTY':
                        fac_role, _ = Role.objects.get_or_create(name='FACULTY')
                        if not user_obj.user_roles.filter(role=fac_role).exists():
                            try:
                                UserRole.objects.create(user=user_obj, role=fac_role)
                            except Exception:
                                pass

                    created_count += 1

            except Exception as e:
                failed.append({
                    'row': row_idx,
                    'error': str(e),
                    'data': cells[:5] if 'cells' in dir() else [],
                })

        return Response({
            'created': created_count,
            'updated': updated_count,
            'failed': failed,
            'total_rows': processed_count,
        })

    def _cell(self, cells, idx):
        try:
            v = cells[idx]
            if v is None:
                return ''
            # Normalize float→int for numeric IDs
            try:
                if isinstance(v, float) and v == int(v):
                    return str(int(v))
            except Exception:
                pass
            return str(v).strip()
        except IndexError:
            return ''

    def _import_student(self, cells, college):
        from django.db.models import Q
        reg_no = self._cell(cells, 0)
        email = self._cell(cells, 1)
        first_name = self._cell(cells, 2)
        last_name = self._cell(cells, 3)
        batch = self._cell(cells, 4)
        status_val = self._cell(cells, 5).upper() or 'ACTIVE'
        dept_code = self._cell(cells, 6)
        phone = self._cell(cells, 7)

        if not reg_no:
            raise ValueError('Register Number is required')
        if not email:
            email = f'{reg_no}@imported.local'

        dept = None
        if dept_code:
            dept = Department.objects.filter(
                Q(code__iexact=dept_code) | Q(short_name__iexact=dept_code) | Q(name__iexact=dept_code)
            ).first()

        # Find or create user
        existing = StudentProfile.objects.filter(reg_no=reg_no).select_related('user').first()
        if existing:
            user_obj = existing.user
            user_obj.set_password(reg_no)
            user_obj.must_change_password = True
            if first_name:
                user_obj.first_name = first_name
            if last_name:
                user_obj.last_name = last_name
            if phone:
                user_obj.mobile_no = phone
            user_obj.save()
            existing.college = college
            existing.batch = batch or existing.batch
            existing.status = status_val
            if dept:
                existing.home_department = dept
            if phone:
                existing.mobile_number = phone
            existing.save(update_fields=['college', 'batch', 'status', 'home_department', 'mobile_number'])
            return user_obj, existing
        else:
            user_obj = User(
                username=first_name or reg_no,
                email=email,
                first_name=first_name,
                last_name=last_name,
                must_change_password=True,
            )
            if phone:
                user_obj.mobile_no = phone
            user_obj.set_password(reg_no)
            user_obj.save()

            profile = StudentProfile(
                user=user_obj,
                college=college,
                reg_no=reg_no,
                batch=batch,
                status=status_val,
            )
            if phone:
                profile.mobile_number = phone
            if dept:
                profile.home_department = dept
            profile.save()
            return user_obj, profile

    def _import_staff(self, cells, college):
        from django.db.models import Q
        staff_id = self._cell(cells, 0)
        email = self._cell(cells, 1)
        first_name = self._cell(cells, 2)
        last_name = self._cell(cells, 3)
        dept_code = self._cell(cells, 4)
        designation = self._cell(cells, 5)
        status_val = self._cell(cells, 6).upper() or 'ACTIVE'
        phone = self._cell(cells, 7)

        if not staff_id:
            raise ValueError('Staff ID is required')
        if not email:
            email = f'{staff_id}@imported.local'

        dept = None
        if dept_code:
            dept = Department.objects.filter(
                Q(code__iexact=dept_code) | Q(short_name__iexact=dept_code) | Q(name__iexact=dept_code)
            ).first()

        existing = StaffProfile.objects.filter(staff_id=staff_id).select_related('user').first()
        if existing:
            user_obj = existing.user
            user_obj.set_password(staff_id)
            user_obj.must_change_password = True
            if first_name:
                user_obj.first_name = first_name
            if last_name:
                user_obj.last_name = last_name
            if phone:
                user_obj.mobile_no = phone
            user_obj.save()
            existing.college = college
            existing.designation = designation or existing.designation
            existing.status = status_val
            if dept:
                existing.department = dept
            if phone:
                existing.mobile_number = phone
            existing.save(update_fields=['college', 'designation', 'status', 'department', 'mobile_number'])
            return user_obj, existing
        else:
            user_obj = User(
                username=first_name or staff_id,
                email=email,
                first_name=first_name,
                last_name=last_name,
                must_change_password=True,
            )
            if phone:
                user_obj.mobile_no = phone
            user_obj.set_password(staff_id)
            user_obj.save()

            profile = StaffProfile(
                user=user_obj,
                college=college,
                staff_id=staff_id,
                department=dept,
                designation=designation,
                status=status_val,
            )
            if phone:
                profile.mobile_number = phone
            profile.save()
            return user_obj, profile

    def _import_other(self, cells, college, role_name):
        id_val = self._cell(cells, 0)
        email = self._cell(cells, 1)
        first_name = self._cell(cells, 2)
        last_name = self._cell(cells, 3)
        profile_type = self._cell(cells, 4).upper() or 'STAFF'
        dept_code = self._cell(cells, 5)
        designation = self._cell(cells, 6)
        batch = self._cell(cells, 7)
        status_val = self._cell(cells, 8).upper() or 'ACTIVE'
        phone = self._cell(cells, 9)

        if not id_val:
            raise ValueError('ID is required')
        if not email:
            email = f'{id_val}@imported.local'

        if profile_type == 'STUDENT':
            cells_student = [id_val, email, first_name, last_name, batch, status_val, '', dept_code, phone]
            return self._import_student(cells_student, college)
        else:
            cells_staff = [id_val, email, first_name, last_name, dept_code, designation, status_val, phone]
            return self._import_staff(cells_staff, college)


# ---------------------------------------------------------------------------
# College Users — Delete
# ---------------------------------------------------------------------------

class CollegeUserDeleteView(APIView):
    """DELETE /api/college/colleges/<college_id>/users/<user_id>/"""
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def delete(self, request, pk, user_id):
        college = get_object_or_404(College, pk=pk)
        self.check_object_permissions(request, college)
        user = get_object_or_404(User, pk=user_id)
        self.check_object_permissions(request, user)

        # Remove association, don't delete the user entirely
        sp = getattr(user, 'student_profile', None)
        if sp and sp.college_id == college.pk:
            sp.college = None
            sp.save(update_fields=['college'])

        st = getattr(user, 'staff_profile', None)
        if st and st.college_id == college.pk:
            st.college = None
            st.save(update_fields=['college'])

        return Response(status=204)

class CollegeUserRolesView(APIView):
    """PUT /api/college/colleges/<college_id>/users/<user_id>/roles/"""
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def put(self, request, pk, user_id):
        college = get_object_or_404(College, pk=pk)
        self.check_object_permissions(request, college)
        user = get_object_or_404(User, pk=user_id)
        
        # Verify user belongs to college
        sp = getattr(user, 'student_profile', None)
        st = getattr(user, 'staff_profile', None)
        
        if not (sp and sp.college_id == college.id) and not (st and st.college_id == college.id):
            return Response({"detail": "User does not belong to this college."}, status=status.HTTP_400_BAD_REQUEST)

        role_names = request.data.get('roles', [])
        if not isinstance(role_names, list):
            return Response({"detail": "Roles must be a list of strings."}, status=status.HTTP_400_BAD_REQUEST)
        
        from accounts.models import Role
        roles = []
        for name in role_names:
            clean_name = str(name).strip()
            if not clean_name:
                continue
            role_objs = list(Role.objects.filter(name__iexact=clean_name))
            if role_objs:
                primary_role = next((r for r in role_objs if r.name == clean_name.upper()), role_objs[0])
                roles.append(primary_role)
        
        user.roles.set(roles)
        
        return Response({"detail": "Roles updated successfully.", "roles": [r.name for r in roles]})



# ---------------------------------------------------------------------------
# College Features — List & Bulk Update
# ---------------------------------------------------------------------------

from rest_framework.permissions import BasePermission
class CanManageCollegeFeatures(BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.user.is_superuser:
            return True
        if hasattr(request.user, 'roles'):
            roles = {r.name.upper() for r in request.user.roles.all()}
            if 'SUPER_ADMIN' in roles:
                return True
        return request.user.has_perm('college.change_collegefeature')


class CollegeFeaturesListView(APIView):
    """GET /api/college/colleges/<id>/features/
    Returns all features from the catalog with per-college toggle state.
    Missing CollegeFeature rows are auto-created from the catalog.

    PUT /api/college/colleges/<id>/features/
    Bulk-update feature toggles. Body: { "features": { "obe": true, "coe": false } }
    """
    permission_classes = [CanManageCollegeFeatures]

    def get(self, request, pk):
        college = get_object_or_404(College, pk=pk)
        self.check_object_permissions(request, college)
        catalog = FeatureCatalog.objects.all()

        # Auto-create missing CollegeFeature rows
        existing_codes = set(
            CollegeFeature.objects.filter(college=college).values_list('feature__code', flat=True)
        )
        new_entries = []
        for feat in catalog:
            if feat.code not in existing_codes:
                new_entries.append(
                    CollegeFeature(college=college, feature=feat, is_enabled=feat.is_default)
                )
        if new_entries:
            CollegeFeature.objects.bulk_create(new_entries, ignore_conflicts=True)

        college_features = (
            CollegeFeature.objects
            .filter(college=college)
            .select_related('feature')
            .order_by('feature__sort_order', 'feature__category')
        )
        serializer = CollegeFeatureSerializer(college_features, many=True)
        return Response(serializer.data)

    def put(self, request, pk):
        college = get_object_or_404(College, pk=pk)
        self.check_object_permissions(request, college)
        features_map = request.data.get('features', {})
        if not isinstance(features_map, dict):
            return Response({'detail': 'features must be a dict of {code: bool}'}, status=400)

        from django.utils import timezone
        now = timezone.now()

        updated = 0
        for code, enabled in features_map.items():
            feat = FeatureCatalog.objects.filter(code=code).first()
            if not feat:
                continue
            cf, created = CollegeFeature.objects.get_or_create(
                college=college, feature=feat,
                defaults={'is_enabled': bool(enabled)}
            )
            if not created and cf.is_enabled != bool(enabled):
                cf.is_enabled = bool(enabled)
                if enabled:
                    cf.enabled_at = now
                    cf.disabled_at = None
                else:
                    cf.disabled_at = now
                cf.save(update_fields=['is_enabled', 'enabled_at', 'disabled_at'])
            updated += 1

        # ── Sync CollegeRole entries to match the new feature state ─────────
        from .utils import sync_college_roles
        role_result = sync_college_roles(college)

        return Response({'updated': updated, 'roles': role_result})


# ---------------------------------------------------------------------------
# College Features — Single Toggle
# ---------------------------------------------------------------------------

class CollegeFeatureToggleView(APIView):
    """PATCH /api/college/colleges/<id>/features/<code>/
    Toggle a single feature. Body: { "is_enabled": true }
    """
    permission_classes = [CanManageCollegeFeatures]

    def patch(self, request, pk, code):
        college = get_object_or_404(College, pk=pk)
        self.check_object_permissions(request, college)
        feat = get_object_or_404(FeatureCatalog, code=code)
        enabled = request.data.get('is_enabled')
        if enabled is None:
            return Response({'detail': 'is_enabled is required'}, status=400)

        from django.utils import timezone
        now = timezone.now()

        cf, _ = CollegeFeature.objects.get_or_create(
            college=college, feature=feat,
            defaults={'is_enabled': bool(enabled)}
        )
        cf.is_enabled = bool(enabled)
        if enabled:
            cf.enabled_at = now
            cf.disabled_at = None
        else:
            cf.disabled_at = now
        cf.save(update_fields=['is_enabled', 'enabled_at', 'disabled_at'])

        # ── Sync CollegeRole entries to match the new feature state ─────────
        from .utils import sync_college_roles
        role_result = sync_college_roles(college)

        return Response({
            'code': feat.code,
            'name': feat.name,
            'is_enabled': cf.is_enabled,
            'roles': role_result,
        })


# ---------------------------------------------------------------------------
# Departments — CRUD (Super Admin + College Admin)
# ---------------------------------------------------------------------------

class DepartmentListCreateView(APIView):
    """GET  /api/college/departments/  — list all departments.
       POST /api/college/departments/  — create a new department.
    """
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request):
        qs = Department.objects.all().order_by('code')
        college_id = _resolve_college_id(request)
        if college_id:
            try:
                qs = qs.filter(college_id=int(college_id))
            except (TypeError, ValueError):
                pass
        search = request.query_params.get('search', '').strip()
        if search:
            from django.db.models import Q
            qs = qs.filter(
                Q(code__icontains=search) | Q(name__icontains=search) | Q(short_name__icontains=search)
            )
        data = []
        for d in qs:
            data.append({
                'id': d.id,
                'code': d.code,
                'name': d.name,
                'short_name': d.short_name,
                'is_teaching': d.is_teaching,
                'parent': d.parent_id,
                'parent_name': str(d.parent) if d.parent else None,
                'is_sh_main': d.is_sh_main,
                'college': d.college_id,
            })
        return Response(data)

    def post(self, request):
        code = (request.data.get('code') or '').strip()
        name = (request.data.get('name') or '').strip()
        if not code or not name:
            return Response({'detail': 'code and name are required.'}, status=400)
        college_id = _resolve_college_id(request)
        if Department.objects.filter(code=code, college_id=college_id or None).exists():
            return Response({'detail': f'Department with code "{code}" already exists in this college.'}, status=400)
        d = Department.objects.create(
            code=code,
            name=name,
            short_name=(request.data.get('short_name') or '').strip(),
            is_teaching=request.data.get('is_teaching', True),
            parent_id=request.data.get('parent') or None,
            is_sh_main=request.data.get('is_sh_main', False),
            college_id=college_id or None,
        )
        return Response({
            'id': d.id, 'code': d.code, 'name': d.name,
            'short_name': d.short_name, 'is_teaching': d.is_teaching,
            'parent': d.parent_id, 'is_sh_main': d.is_sh_main,
            'college': d.college_id,
        }, status=201)


class DepartmentDetailView(APIView):
    """GET/PUT/DELETE /api/college/departments/<id>/"""
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request, pk):
        d = get_object_or_404(Department, pk=pk)
        self.check_object_permissions(request, d)
        return Response({
            'id': d.id, 'code': d.code, 'name': d.name,
            'short_name': d.short_name, 'is_teaching': d.is_teaching,
            'parent': d.parent_id, 'parent_name': str(d.parent) if d.parent else None,
            'is_sh_main': d.is_sh_main,
            'college': d.college_id,
        })

    def put(self, request, pk):
        d = get_object_or_404(Department, pk=pk)
        self.check_object_permissions(request, d)
        if 'code' in request.data:
            new_code = request.data['code'].strip()
            if new_code != d.code and Department.objects.filter(code=new_code, college_id=d.college_id).exists():
                return Response({'detail': f'Code "{new_code}" already in use in this college.'}, status=400)
            d.code = new_code
        if 'name' in request.data:
            d.name = request.data['name'].strip()
        if 'short_name' in request.data:
            d.short_name = (request.data['short_name'] or '').strip()
        if 'is_teaching' in request.data:
            d.is_teaching = bool(request.data['is_teaching'])
        if 'parent' in request.data:
            d.parent_id = request.data['parent'] or None
        if 'is_sh_main' in request.data:
            d.is_sh_main = bool(request.data['is_sh_main'])
        if 'college' in request.data:
            d.college_id = request.data['college'] or None
        d.save()
        return Response({
            'id': d.id, 'code': d.code, 'name': d.name,
            'short_name': d.short_name, 'is_teaching': d.is_teaching,
            'parent': d.parent_id, 'is_sh_main': d.is_sh_main,
            'college': d.college_id,
        })

    def delete(self, request, pk):
        d = get_object_or_404(Department, pk=pk)
        self.check_object_permissions(request, d)
        d.delete()
        return Response(status=204)


# ---------------------------------------------------------------------------
# Batches — CRUD (Super Admin only)
# ---------------------------------------------------------------------------

class BatchListCreateView(APIView):
    """GET  /api/college/batches/  — list all batches.
       POST /api/college/batches/  — create a new batch.
    """
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request):
        from academics.models import Batch
        qs = Batch.objects.all().select_related('course', 'course__department', 'department', 'regulation', 'batch_year').order_by('-name')
        college_id = _resolve_college_id(request)
        if college_id:
            try:
                qs = qs.filter(college_id=int(college_id))
            except (TypeError, ValueError):
                pass
        search = request.query_params.get('search', '').strip()
        if search:
            from django.db.models import Q
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(course__name__icontains=search) |
                Q(department__code__icontains=search)
            )
        data = []
        for b in qs:
            data.append({
                'id': b.id,
                'name': b.name,
                'course': b.course_id,
                'course_name': str(b.course) if b.course else None,
                'department': b.department_id,
                'department_name': str(b.department) if b.department else (str(b.course.department) if b.course and b.course.department else None),
                'start_year': b.start_year,
                'end_year': b.end_year,
                'regulation': b.regulation_id,
                'regulation_code': str(b.regulation) if b.regulation else None,
                'is_active': b.is_active,
                'batch_year': b.batch_year_id,
                'college': b.college_id,
            })
        return Response(data)

    def post(self, request):
        from academics.models import Batch
        name = (request.data.get('name') or '').strip()
        if not name:
            return Response({'detail': 'name is required.'}, status=400)
        course_id = request.data.get('course') or None
        department_id = request.data.get('department') or None
        # Auto-derive department from the course if not explicitly supplied
        if department_id is None and course_id:
            from academics.models import Course as CourseModel
            try:
                dept = CourseModel.objects.filter(pk=course_id).values_list('department_id', flat=True).first()
                if dept:
                    department_id = dept
            except Exception:
                pass
        college_id = _resolve_college_id(request)
        try:
            b = Batch.objects.create(
                name=name,
                course_id=course_id,
                department_id=department_id,
                college_id=college_id or None,
                start_year=request.data.get('start_year') or None,
                end_year=request.data.get('end_year') or None,
                regulation_id=request.data.get('regulation') or None,
                is_active=request.data.get('is_active', True),
            )
            return Response({
                'id': b.id, 'name': b.name, 'course': b.course_id,
                'department': b.department_id, 'start_year': b.start_year,
                'end_year': b.end_year, 'regulation': b.regulation_id,
                'is_active': b.is_active,
                'college': b.college_id,
            }, status=201)
        except Exception as e:
            from django.db import IntegrityError
            if isinstance(e, IntegrityError):
                return Response({'detail': f'A batch named "{name}" already exists for this course or department.'}, status=400)
            return Response({'detail': str(e)}, status=500)


# ---------------------------------------------------------------------------
# Batch bulk-create — create many batches in a single request / transaction
# ---------------------------------------------------------------------------

class BatchBulkCreateView(APIView):
    """POST /api/college/batches/bulk/
    Body: { "batches": [ {name, course, department, start_year, end_year, regulation, is_active}, ... ] }
    Creates all supplied batches atomically. Returns list of created objects.
    """
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def post(self, request):
        from academics.models import Batch
        from django.db import transaction, IntegrityError

        items = request.data.get('batches', [])
        if not isinstance(items, list) or len(items) == 0:
            return Response({'detail': 'batches list is required and must be non-empty.'}, status=400)

        college_id = _resolve_college_id(request)

        # Fallback: read 'college' from top-level body key or from the first batch item,
        # in case the X-College-Id header was not sent (e.g. direct navigation).
        if college_id is None:
            top_college = request.data.get('college')
            if top_college:
                try:
                    college_id = int(top_college)
                except (TypeError, ValueError):
                    pass
        if college_id is None and items:
            first_college = items[0].get('college')
            if first_college:
                try:
                    college_id = int(first_college)
                except (TypeError, ValueError):
                    pass

        if not college_id:
            return Response(
                {'detail': 'College context is required. Please navigate to the batch page via the college detail page.'},
                status=400
            )

        # Pre-fetch course → department mapping so we can auto-populate department_id
        # for course-based batches (this makes Django admin and direct DB queries show
        # the department column populated rather than NULL).
        course_ids_in_request = {
            int(item['course']) for item in items if item.get('course')
        }
        course_dept_map: dict[int, int | None] = {}
        if course_ids_in_request:
            from academics.models import Course as CourseModel
            for cobj in CourseModel.objects.filter(pk__in=course_ids_in_request).values('id', 'department_id'):
                course_dept_map[cobj['id']] = cobj['department_id']

        # Validate all items first
        validated = []
        for i, item in enumerate(items):
            name = (item.get('name') or '').strip()
            if not name:
                return Response({'detail': f'Item {i}: name is required.'}, status=400)
            # Use item-level college if it differs from the resolved one (shouldn't happen, but be safe)
            item_college = int(item['college']) if item.get('college') else college_id
            course_id_val = item.get('course') or None
            # Auto-derive department from the course if not explicitly supplied
            dept_id_val = item.get('department') or None
            if dept_id_val is None and course_id_val:
                dept_id_val = course_dept_map.get(int(course_id_val))
            validated.append(Batch(
                name=name,
                course_id=course_id_val,
                department_id=dept_id_val,
                college_id=item_college,
                start_year=item.get('start_year') or None,
                end_year=item.get('end_year') or None,
                regulation_id=item.get('regulation') or None,
                is_active=item.get('is_active', True),
            ))

        try:
            with transaction.atomic():
                created = Batch.objects.bulk_create(validated)
        except IntegrityError:
            # Find which batch names already exist for this college+course/dept combo
            names = list({b.name for b in validated})
            course_ids = list({b.course_id for b in validated if b.course_id})
            dept_ids = list({b.department_id for b in validated if b.department_id})
            from django.db.models import Q as DQ
            q = DQ()
            if course_ids:
                q |= DQ(course_id__in=course_ids)
            if dept_ids:
                q |= DQ(department_id__in=dept_ids)
            existing = Batch.objects.filter(
                name__in=names,
                college_id=college_id or None,
            ).filter(q).values_list('name', flat=True)
            existing_names = ', '.join(sorted(set(existing))) or 'one or more batches'
            return Response({
                'detail': f'Batch{"es" if len(set(existing)) != 1 else ""} already exist for this college: {existing_names}. Each batch name must be unique per course within a college.'
            }, status=400)
        except Exception as e:
            return Response({'detail': str(e)}, status=500)

        # Auto-link BatchYear: find-or-create based on start_year so the
        # Django admin BATCH YEAR column is populated.
        start_year_val = validated[0].start_year if validated else None
        if start_year_val and created:
            from academics.models import BatchYear
            batch_year_name = str(start_year_val)
            try:
                by, _ = BatchYear.objects.get_or_create(
                    name=batch_year_name,
                    defaults={
                        'start_year': start_year_val,
                        'end_year': validated[0].end_year,
                        'college_id': college_id,
                    }
                )
                # Bulk-update all created batches with the batch_year FK
                Batch.objects.filter(pk__in=[b.id for b in created]).update(batch_year=by)
                for b in created:
                    b.batch_year_id = by.pk
            except Exception:
                pass  # Don't fail the whole request for this optional link

        return Response(
            [{'id': b.id, 'name': b.name, 'course': b.course_id,
              'department': b.department_id, 'start_year': b.start_year,
              'end_year': b.end_year, 'regulation': b.regulation_id,
              'is_active': b.is_active, 'college': b.college_id}
             for b in created],
            status=201
        )


class BatchGroupView(APIView):
    """PUT/DELETE /api/college/batches/group/<name>/
    Operates on all batches with the given name within the current college.
    """
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def delete(self, request, name):
        from academics.models import Batch
        college_id = _resolve_college_id(request)
        if not college_id:
            return Response({'detail': 'College context required.'}, status=400)
            
        batches = Batch.objects.filter(name=name, college_id=college_id)
        if not batches.exists():
            return Response(status=404)
            
        batches.delete()
        return Response(status=204)
        
    def put(self, request, name):
        from academics.models import Batch
        college_id = _resolve_college_id(request)
        if not college_id:
            return Response({'detail': 'College context required.'}, status=400)
            
        batches = Batch.objects.filter(name=name, college_id=college_id)
        if not batches.exists():
            return Response(status=404)
            
        # Update fields provided in the request
        update_fields = {}
        if 'start_year' in request.data:
            update_fields['start_year'] = request.data['start_year'] or None
        if 'end_year' in request.data:
            update_fields['end_year'] = request.data['end_year'] or None
        if 'regulation' in request.data:
            update_fields['regulation_id'] = request.data['regulation'] or None
        if 'is_active' in request.data:
            update_fields['is_active'] = request.data['is_active']
            
        if update_fields:
            batches.update(**update_fields)
            
        # Auto-update BatchYear if start_year/end_year changed
        if 'start_year' in request.data:
            from academics.models import BatchYear
            start_year = update_fields.get('start_year')
            if start_year:
                by, _ = BatchYear.objects.get_or_create(
                    name=str(start_year),
                    defaults={
                        'start_year': start_year,
                        'end_year': update_fields.get('end_year'),
                        'college_id': college_id,
                    }
                )
                batches.update(batch_year=by)
            else:
                batches.update(batch_year=None)
                
        return Response({'detail': f'Updated {batches.count()} batches successfully.'})


class BatchDetailView(APIView):
    """GET/PUT/DELETE /api/college/batches/<id>/"""
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request, pk):
        from academics.models import Batch
        b = get_object_or_404(Batch, pk=pk)
        self.check_object_permissions(request, b)
        return Response({
            'id': b.id, 'name': b.name, 'course': b.course_id,
            'course_name': str(b.course) if b.course else None,
            'department': b.department_id,
            'department_name': str(b.department) if b.department else None,
            'start_year': b.start_year, 'end_year': b.end_year,
            'regulation': b.regulation_id,
            'regulation_code': str(b.regulation) if b.regulation else None,
            'is_active': b.is_active,
        })

    def put(self, request, pk):
        from academics.models import Batch
        b = get_object_or_404(Batch, pk=pk)
        self.check_object_permissions(request, b)
        if 'name' in request.data:
            b.name = request.data['name'].strip()
        if 'course' in request.data:
            b.course_id = request.data['course'] or None
        if 'department' in request.data:
            b.department_id = request.data['department'] or None
        if 'start_year' in request.data:
            b.start_year = request.data['start_year'] or None
        if 'end_year' in request.data:
            b.end_year = request.data['end_year'] or None
        if 'regulation' in request.data:
            b.regulation_id = request.data['regulation'] or None
        if 'is_active' in request.data:
            b.is_active = bool(request.data['is_active'])
        if 'college' in request.data:
            b.college_id = request.data['college'] or None
        b.save()
        return Response({
            'id': b.id, 'name': b.name, 'course': b.course_id,
            'department': b.department_id, 'start_year': b.start_year,
            'end_year': b.end_year, 'regulation': b.regulation_id,
            'is_active': b.is_active,
            'college': b.college_id,
        })

    def delete(self, request, pk):
        from academics.models import Batch
        b = get_object_or_404(Batch, pk=pk)
        self.check_object_permissions(request, b)
        b.delete()
        return Response(status=204)


# ---------------------------------------------------------------------------
# Regulations — CRUD (Super Admin only)
# ---------------------------------------------------------------------------

class RegulationListCreateView(APIView):
    """GET  /api/college/regulations/  — list all regulations.
       POST /api/college/regulations/  — create a new regulation.
    """
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request):
        from curriculum.models import Regulation
        qs = Regulation.objects.all().order_by('-code')
        college_id = _resolve_college_id(request)
        if college_id:
            try:
                qs = qs.filter(college_id=int(college_id))
            except (TypeError, ValueError):
                pass
        search = request.query_params.get('search', '').strip()
        if search:
            from django.db.models import Q
            qs = qs.filter(Q(code__icontains=search) | Q(name__icontains=search))
        data = []
        for r in qs:
            data.append({
                'id': r.id,
                'code': r.code,
                'name': r.name,
                'is_active': r.is_active,
                'created_at': r.created_at.isoformat() if r.created_at else None,
                'college': r.college_id,
            })
        return Response(data)

    def post(self, request):
        from curriculum.models import Regulation
        code = (request.data.get('code') or '').strip()
        if not code:
            return Response({'detail': 'code is required.'}, status=400)
        college_id = _resolve_college_id(request)
        if Regulation.objects.filter(code=code, college_id=college_id or None).exists():
            return Response({'detail': f'Regulation "{code}" already exists in this college.'}, status=400)
        r = Regulation.objects.create(
            code=code,
            name=(request.data.get('name') or '').strip(),
            college_id=college_id or None,
            is_active=request.data.get('is_active', True),
        )
        return Response({
            'id': r.id, 'code': r.code, 'name': r.name, 'is_active': r.is_active,
            'college': r.college_id,
        }, status=201)


class RegulationDetailView(APIView):
    """GET/PUT/DELETE /api/college/regulations/<id>/"""
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request, pk):
        from curriculum.models import Regulation
        r = get_object_or_404(Regulation, pk=pk)
        self.check_object_permissions(request, r)
        return Response({
            'id': r.id, 'code': r.code, 'name': r.name, 'is_active': r.is_active,
            'college': r.college_id,
        })

    def put(self, request, pk):
        from curriculum.models import Regulation
        r = get_object_or_404(Regulation, pk=pk)
        self.check_object_permissions(request, r)
        if 'code' in request.data:
            new_code = request.data['code'].strip()
            if new_code != r.code and Regulation.objects.filter(code=new_code, college_id=r.college_id).exists():
                return Response({'detail': f'Code "{new_code}" already in use in this college.'}, status=400)
            r.code = new_code
        if 'name' in request.data:
            r.name = (request.data['name'] or '').strip()
        if 'is_active' in request.data:
            r.is_active = bool(request.data['is_active'])
        if 'college' in request.data:
            r.college_id = request.data['college'] or None
        r.save()
        return Response({
            'id': r.id, 'code': r.code, 'name': r.name, 'is_active': r.is_active,
            'college': r.college_id,
        })

    def delete(self, request, pk):
        from curriculum.models import Regulation
        r = get_object_or_404(Regulation, pk=pk)
        self.check_object_permissions(request, r)
        r.delete()
        return Response(status=204)

# ---------------------------------------------------------------------------
# Programs CRUD
# ---------------------------------------------------------------------------
from academics.models import Program

class ProgramListCreateView(APIView):
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request):
        college_id = _resolve_college_id(request)
        programs = Program.objects.all().order_by('name')
        if college_id:
            try:
                programs = programs.filter(college_id=int(college_id))
            except (TypeError, ValueError):
                pass
        
        data = [{'id': p.id, 'name': p.name, 'college': p.college_id} for p in programs]
        return Response(data)

    def post(self, request):
        college_id = _resolve_college_id(request)
        name = request.data.get('name', '').strip()
        if not name:
            return Response({'detail': 'Program name is required.'}, status=400)
            
        if Program.objects.filter(name=name, college_id=college_id).exists():
            return Response({'detail': f'Program "{name}" already exists in this college.'}, status=400)
            
        p = Program.objects.create(name=name, college_id=college_id)
        return Response({'id': p.id, 'name': p.name, 'college': p.college_id}, status=201)


class ProgramDetailView(APIView):
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request, pk):
        p = get_object_or_404(Program, pk=pk)
        self.check_object_permissions(request, p)
        return Response({'id': p.id, 'name': p.name, 'college': p.college_id})

    def put(self, request, pk):
        p = get_object_or_404(Program, pk=pk)
        self.check_object_permissions(request, p)
        name = request.data.get('name', '').strip()
        if name:
            if name != p.name and Program.objects.filter(name=name, college_id=p.college_id).exists():
                return Response({'detail': f'Program "{name}" already exists in this college.'}, status=400)
            p.name = name
            p.save()
        return Response({'id': p.id, 'name': p.name, 'college': p.college_id})

    def delete(self, request, pk):
        p = get_object_or_404(Program, pk=pk)
        self.check_object_permissions(request, p)
        p.delete()
        return Response(status=204)


# ---------------------------------------------------------------------------
# Courses CRUD
# ---------------------------------------------------------------------------

class CourseListCreateView(APIView):
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request):
        college_id = _resolve_college_id(request)
        courses = Course.objects.all().select_related('department', 'program').order_by('name')
        if college_id:
            try:
                courses = courses.filter(college_id=int(college_id))
            except (TypeError, ValueError):
                pass
        
        data = [
            {
                'id': c.id, 'name': c.name, 
                'department': c.department_id, 'department_name': str(c.department) if c.department else None,
                'program': c.program_id, 'program_name': str(c.program) if c.program else None,
                'college': c.college_id
            }
            for c in courses
        ]
        return Response(data)

    def post(self, request):
        college_id = _resolve_college_id(request)
        name = request.data.get('name', '').strip()
        dept_id = request.data.get('department')
        prog_id = request.data.get('program')
        
        if not name or not dept_id or not prog_id:
            return Response({'detail': 'Course name, department, and program are required.'}, status=400)
            
        if Course.objects.filter(name=name, department_id=dept_id, program_id=prog_id, college_id=college_id).exists():
            return Response({'detail': 'This course already exists in this college.'}, status=400)
            
        c = Course.objects.create(name=name, department_id=dept_id, program_id=prog_id, college_id=college_id)
        return Response({
            'id': c.id, 'name': c.name, 
            'department': c.department_id, 'program': c.program_id, 'college': c.college_id
        }, status=201)


class CourseDetailView(APIView):
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request, pk):
        c = get_object_or_404(Course, pk=pk)
        self.check_object_permissions(request, c)
        return Response({
            'id': c.id, 'name': c.name, 
            'department': c.department_id, 'department_name': str(c.department) if c.department else None,
            'program': c.program_id, 'program_name': str(c.program) if c.program else None,
            'college': c.college_id
        })

    def put(self, request, pk):
        c = get_object_or_404(Course, pk=pk)
        self.check_object_permissions(request, c)
        
        name = request.data.get('name', c.name).strip()
        dept_id = request.data.get('department', c.department_id)
        prog_id = request.data.get('program', c.program_id)
        
        if (name != c.name or dept_id != c.department_id or prog_id != c.program_id):
            if Course.objects.filter(name=name, department_id=dept_id, program_id=prog_id, college_id=c.college_id).exclude(pk=c.pk).exists():
                return Response({'detail': 'This course already exists in this college.'}, status=400)
        
        c.name = name
        if request.data.get('department'):
            c.department_id = dept_id
        if request.data.get('program'):
            c.program_id = prog_id
        c.save()
        
        return Response({
            'id': c.id, 'name': c.name, 
            'department': c.department_id, 'program': c.program_id, 'college': c.college_id
        })

    def delete(self, request, pk):
        c = get_object_or_404(Course, pk=pk)
        self.check_object_permissions(request, c)
        c.delete()
        return Response(status=204)


# ---------------------------------------------------------------------------
# Lookup helpers for frontend dropdowns
# ---------------------------------------------------------------------------

class CourseListView(APIView):
    """GET /api/college/courses/ — list all courses for dropdown use."""
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def get(self, request):
        courses = Course.objects.all().select_related('department', 'program').order_by('name')
        college_id = _resolve_college_id(request)
        if college_id:
            try:
                courses = courses.filter(department__college_id=int(college_id))
            except (TypeError, ValueError):
                pass
        data = [
            {'id': c.id, 'name': c.name, 'department': c.department_id,
             'department_name': str(c.department) if c.department else None}
            for c in courses
        ]
        return Response(data)


# ---------------------------------------------------------------------------
# Department HOD / AHOD assignment
# ---------------------------------------------------------------------------

class DepartmentHodView(APIView):
    """GET  /api/college/departments/<dept_id>/roles/?college_id=<id>
       GET  current HOD & AHOD assignments for a department.

       POST /api/college/departments/<dept_id>/roles/
       Body: { role: 'HOD'|'AHOD', staff_id: <staff profile id>, college_id: <int> }
       Assigns the staff to that role for the current active academic year.
       Deactivates any existing assignment for that role first.
    """
    permission_classes = [IsCollegeAdminOrSuperAdmin]

    def _get_active_year(self, college_id):
        from academics.models import AcademicYear
        ay = AcademicYear.objects.filter(college_id=college_id, is_active=True).first()
        if not ay:
            ay = AcademicYear.objects.filter(college_id=college_id).order_by('-id').first()
        return ay

    def get(self, request, dept_id):
        from academics.models import Department, DepartmentRole, AcademicYear
        college_id = request.query_params.get('college_id') or _resolve_college_id(request)
        dept = get_object_or_404(Department, pk=dept_id)

        qs = DepartmentRole.objects.filter(
            department=dept, is_active=True
        ).select_related('staff', 'staff__user', 'academic_year')

        if college_id:
            try:
                qs = qs.filter(college_id=int(college_id))
            except (TypeError, ValueError):
                pass

        data = []
        for dr in qs:
            data.append({
                'id': dr.id,
                'role': dr.role,
                'role_display': dr.get_role_display(),
                'staff_id': dr.staff_id,
                'staff_profile_id': dr.staff.id,
                'staff_name': f"{dr.staff.user.first_name} {dr.staff.user.last_name}".strip() or dr.staff.staff_id,
                'staff_employee_id': dr.staff.staff_id,
                'academic_year': dr.academic_year_id,
                'academic_year_name': dr.academic_year.name,
            })
        return Response(data)

    def post(self, request, dept_id):
        from academics.models import Department, DepartmentRole, StaffProfile
        from django.db import transaction as db_transaction

        college_id = request.data.get('college_id') or _resolve_college_id(request)
        if not college_id:
            return Response({'detail': 'college_id is required.'}, status=400)
        try:
            college_id = int(college_id)
        except (TypeError, ValueError):
            return Response({'detail': 'Invalid college_id.'}, status=400)

        dept = get_object_or_404(Department, pk=dept_id)
        role = (request.data.get('role') or '').strip().upper()
        if role not in ('HOD', 'AHOD'):
            return Response({'detail': 'role must be HOD or AHOD.'}, status=400)

        staff_profile_id = request.data.get('staff_id')
        if not staff_profile_id:
            return Response({'detail': 'staff_id is required.'}, status=400)
        staff = get_object_or_404(StaffProfile, pk=staff_profile_id)

        ay = self._get_active_year(college_id)
        if not ay:
            return Response({'detail': 'No academic year found for this college. Please create one first.'}, status=400)

        with db_transaction.atomic():
            # Deactivate existing active assignment for this role in this dept/year
            DepartmentRole.objects.filter(
                department=dept,
                role=role,
                academic_year=ay,
                college_id=college_id,
                is_active=True,
            ).update(is_active=False)

            dr = DepartmentRole.objects.create(
                department=dept,
                staff=staff,
                role=role,
                academic_year=ay,
                college_id=college_id,
                is_active=True,
            )

        return Response({
            'id': dr.id,
            'role': dr.role,
            'staff_name': f"{staff.user.first_name} {staff.user.last_name}".strip() or staff.staff_id,
            'academic_year': ay.name,
        }, status=201)

    def delete(self, request, dept_id):
        """DELETE /api/college/departments/<dept_id>/roles/?role=HOD&college_id=<id>"""
        from academics.models import DepartmentRole
        college_id = request.query_params.get('college_id') or _resolve_college_id(request)
        role = (request.query_params.get('role') or '').strip().upper()
        if role not in ('HOD', 'AHOD'):
            return Response({'detail': 'role must be HOD or AHOD.'}, status=400)

        qs = DepartmentRole.objects.filter(
            department_id=dept_id, role=role, is_active=True
        )
        if college_id:
            try:
                qs = qs.filter(college_id=int(college_id))
            except (TypeError, ValueError):
                pass
        qs.update(is_active=False)
        return Response(status=204)


# ---------------------------------------------------------------------------
# Public search endpoints (for external staff registration)
# ---------------------------------------------------------------------------

from django.db.models import Q as _Q
from rest_framework.decorators import api_view


@api_view(['GET'])
def search_colleges(request):
    """Public endpoint to search colleges by name or code."""
    query = request.GET.get('q', '').strip()
    limit = min(int(request.GET.get('limit', 20)), 50)
    if len(query) < 2:
        return Response({'results': [], 'message': 'Enter at least 2 characters'})
    colleges = College.objects.filter(
        _Q(name__icontains=query) | _Q(short_name__icontains=query) | _Q(code__icontains=query)
    ).filter(is_active=True).order_by('name')[:limit]
    results = [
        {'id': c.id, 'code': c.code, 'name': c.name, 'short_name': c.short_name,
         'city': c.city, 'display': f"{c.name}" + (f", {c.city}" if c.city else "")}
        for c in colleges
    ]
    return Response({'results': results})


@api_view(['GET'])
def list_all_colleges(request):
    """Public endpoint to get all active colleges."""
    colleges = College.objects.filter(is_active=True).order_by('name')
    results = [
        {'id': c.id, 'code': c.code, 'name': c.name, 'short_name': c.short_name,
         'city': c.city, 'display': f"{c.name}" + (f", {c.city}" if c.city else "")}
        for c in colleges
    ]
    return Response({'results': results, 'total': len(results)})


# ---------------------------------------------------------------------------
# College Details — Super Admin only, double-auth on every write
# ---------------------------------------------------------------------------


class CollegeDetailsView(APIView):
    """
    GET  /api/college/colleges/<id>/details/
        Returns full college data including logo_url and banner_url.
        Accessible to SUPER_ADMIN only.

    PATCH /api/college/colleges/<id>/details/
        Updates any college field.  Requires the super admin to supply their
        own password in the ``sa_password`` field of the multipart/JSON body.
        Any attempt without correct password is rejected 403.
        Accessible to SUPER_ADMIN only.
    """
    permission_classes = [IsSuperAdminOrSuperuser]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get(self, request, pk):
        college = get_object_or_404(College, pk=pk)
        from .serializers import CollegeSerializer
        ser = CollegeSerializer(college, context={'request': request})
        data = dict(ser.data)
        # Ensure resolution specs are included so UI can show hints
        data['logo_resolution'] = f"{College.LOGO_WIDTH}×{College.LOGO_HEIGHT} px"
        data['banner_resolution'] = f"{College.BANNER_WIDTH}×{College.BANNER_HEIGHT} px"
        return Response(data)

    def patch(self, request, pk):
        college = get_object_or_404(College, pk=pk)

        # ── Double authentication ───────────────────────────────────────────
        sa_password = str(
            request.data.get('sa_password') or
            request.POST.get('sa_password') or
            ''
        ).strip()
        if not sa_password:
            return Response(
                {'detail': 'Super admin password is required to modify college details.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        if not request.user.check_password(sa_password):
            return Response(
                {'detail': 'Incorrect super admin password. Changes not saved.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        # ───────────────────────────────────────────────────────────────────

        # Build mutable data dict excluding the password field
        mutable = {k: v for k, v in request.data.items() if k != 'sa_password'}

        if 'established_year' in mutable and mutable['established_year'] == '':
            mutable['established_year'] = None

        # Handle file uploads
        logo_file = request.FILES.get('logo')
        banner_file = request.FILES.get('banner')

        # Validate image dimensions before saving
        if logo_file:
            try:
                from PIL import Image as PILImage
                img = PILImage.open(logo_file)
                w, h = img.size
                if (w, h) != (College.LOGO_WIDTH, College.LOGO_HEIGHT):
                    return Response(
                        {'detail': f'Logo must be exactly {College.LOGO_WIDTH}×{College.LOGO_HEIGHT} px. '
                                   f'Uploaded image is {w}×{h} px.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                logo_file.seek(0)
            except ImportError:
                pass  # Pillow not available; skip check

        if banner_file:
            try:
                from PIL import Image as PILImage
                img = PILImage.open(banner_file)
                w, h = img.size
                if (w, h) != (College.BANNER_WIDTH, College.BANNER_HEIGHT):
                    return Response(
                        {'detail': f'Banner must be exactly {College.BANNER_WIDTH}×{College.BANNER_HEIGHT} px. '
                                   f'Uploaded image is {w}×{h} px.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                banner_file.seek(0)
            except ImportError:
                pass

        # Partially update the college
        from .serializers import CollegeSerializer
        files_data = {}
        if logo_file:
            files_data['logo'] = logo_file
        if banner_file:
            files_data['banner'] = banner_file

        # Merge text fields with file fields for serializer
        combined = {**mutable, **files_data}

        ser = CollegeSerializer(
            college,
            data=combined,
            partial=True,
            context={'request': request},
        )
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)

        ser.save()
        # Re-serialize with fresh data to return updated URLs
        fresh = CollegeSerializer(college, context={'request': request})
        result = dict(fresh.data)
        result['logo_resolution'] = f"{College.LOGO_WIDTH}×{College.LOGO_HEIGHT} px"
        result['banner_resolution'] = f"{College.BANNER_WIDTH}×{College.BANNER_HEIGHT} px"
        return Response(result)

