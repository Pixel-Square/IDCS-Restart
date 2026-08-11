from __future__ import annotations

from io import BytesIO

from django import forms
from django.contrib import admin, messages
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import path


def _load_workbook(*args, **kwargs):
    # Import lazily because admin modules are imported during Django startup.
    from openpyxl import load_workbook

    return load_workbook(*args, **kwargs)

from .models import College, FeatureCatalog, CollegeFeature


class CollegeUploadExcelForm(forms.Form):
    file = forms.FileField(help_text='Upload .xlsx (Column A: code, Column B: college name)')


def _cell_to_str(v) -> str:
    if v is None:
        return ''
    try:
        if isinstance(v, float) and v.is_integer():
            v = int(v)
    except Exception:
        pass
    return str(v).strip()


def _is_header_row(code: str, name: str) -> bool:
    c = (code or '').strip().lower()
    n = (name or '').strip().lower()
    if c in {'code', 'college code', 'college_code'}:
        return True
    if n in {'name', 'college name', 'college_name'}:
        return True
    return False


class CollegeFeatureInline(admin.TabularInline):
    """Inline editor: toggle features directly from the College detail page."""
    model = CollegeFeature
    extra = 0
    fields = ('feature', 'is_enabled', 'enabled_at', 'disabled_at')
    readonly_fields = ('enabled_at', 'disabled_at')
    autocomplete_fields = ('feature',)


def _safe_count(model_path: str, college) -> int:
    """Safely get a count of records belonging to a college. Returns 0 on error."""
    try:
        app_label, model_name = model_path.split('.')
        from django.apps import apps
        Model = apps.get_model(app_label, model_name)
        return Model.objects.filter(college=college).count()
    except Exception:
        return 0


def _admin_url(app_label: str, model_name: str, college_id: int) -> str:
    """Build a pre-filtered admin changelist URL for a college-scoped model."""
    return f'/admin/{app_label}/{model_name.lower()}/?college__id__exact={college_id}'


@admin.register(College)
class CollegeAdmin(admin.ModelAdmin):
    list_display = ('code', 'short_name', 'name', 'city', 'is_active', 'dashboard_link')
    search_fields = ('code', 'short_name', 'name', 'city')
    list_filter = ('is_active', 'city')
    inlines = [CollegeFeatureInline]

    change_list_template = 'admin/college/college/change_list.html'
    change_form_template = 'admin/college/college/change_form.html'

    def dashboard_link(self, obj):
        from django.utils.html import format_html
        return format_html(
            '<a href="/admin/college/college/{}/dashboard/" '
            'style="background:#417690;color:#fff;padding:3px 10px;border-radius:4px;'
            'text-decoration:none;font-size:12px;font-weight:600;">Dashboard</a>',
            obj.pk,
        )
    dashboard_link.short_description = 'Dashboard'

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path(
                'upload-excel/',
                self.admin_site.admin_view(self.upload_excel_view),
                name='college_college_upload_excel',
            ),
            path(
                '<int:college_id>/dashboard/',
                self.admin_site.admin_view(self.college_dashboard_view),
                name='college_college_dashboard',
            ),
        ]
        return my_urls + urls

    # ------------------------------------------------------------------
    # College Dashboard
    # ------------------------------------------------------------------

    def college_dashboard_view(self, request: HttpRequest, college_id: int) -> HttpResponse:
        if not self.has_view_permission(request):
            raise PermissionError('Forbidden')

        college = get_object_or_404(College, pk=college_id)
        cid = college.pk

        def count(model_path):
            return _safe_count(model_path, college)

        def url(app, model):
            return _admin_url(app, model, cid)

        features_on = list(
            CollegeFeature.objects.filter(college=college, is_enabled=True)
            .select_related('feature')
            .values_list('feature__name', flat=True)
        )
        features_off = list(
            CollegeFeature.objects.filter(college=college, is_enabled=False)
            .select_related('feature')
            .values_list('feature__name', flat=True)
        )

        sections = [
            {
                'title': 'Academics',
                'icon': 'graduation-cap',
                'color': '#417690',
                'models': [
                    {'label': 'Academic Years',      'count': count('academics.AcademicYear'),  'url': url('academics', 'academicyear')},
                    {'label': 'Departments',         'count': count('academics.Department'),    'url': url('academics', 'department')},
                    {'label': 'Batches',             'count': count('academics.Batch'),         'url': url('academics', 'batch')},
                    {'label': 'Batch Years',         'count': count('academics.BatchYear'),     'url': url('academics', 'batchyear')},
                    {'label': 'Sections',            'count': count('academics.Section'),       'url': url('academics', 'section')},
                    {'label': 'Courses',             'count': count('academics.Course'),        'url': url('academics', 'course')},
                    {'label': 'Subjects',            'count': count('academics.Subject'),       'url': url('academics', 'subject')},
                    {'label': 'Programs',            'count': count('academics.Program'),       'url': url('academics', 'program')},
                    {'label': 'Semesters',           'count': count('academics.Semester'),      'url': url('academics', 'semester')},
                ],
            },
            {
                'title': 'Curriculum',
                'icon': 'book-open',
                'color': '#205067',
                'models': [
                    {'label': 'Curriculum Masters',  'count': count('curriculum.CurriculumMaster'),     'url': url('curriculum', 'curriculummaster')},
                    {'label': 'Dept Curricula',      'count': count('curriculum.CurriculumDepartment'), 'url': url('curriculum', 'curriculumdepartment')},
                    {'label': 'Dept Groups',         'count': count('curriculum.DepartmentGroup'),      'url': url('curriculum', 'departmentgroup')},
                    {'label': 'Elective Subjects',   'count': count('curriculum.ElectiveSubject'),      'url': url('curriculum', 'electivesubject')},
                    {'label': 'Regulations',         'count': count('curriculum.Regulation'),           'url': url('curriculum', 'regulation')},
                ],
            },
            {
                'title': 'OBE / Marks',
                'icon': 'chart-bar',
                'color': '#264b5d',
                'models': [
                    {'label': 'CDAP Revisions',      'count': count('OBE.CDAPRevision'),              'url': url('OBE', 'cdaprevision')},
                    {'label': 'LCA Revisions',       'count': count('OBE.LCARevision'),               'url': url('OBE', 'lcarevision')},
                    {'label': 'CO Targets',          'count': count('OBE.COTargetRevision'),          'url': url('OBE', 'cotargetrevision')},
                    {'label': 'Assessment Configs',  'count': count('OBE.OBEAssessmentMasterConfig'), 'url': url('OBE', 'obeassessmentmasterconfig')},
                ],
            },
            {
                'title': 'Feedback',
                'icon': 'clipboard-list',
                'color': '#1a7a4a',
                'models': [
                    {'label': 'Feedback Forms',      'count': count('feedback.FeedbackForm'),           'url': url('feedback', 'feedbackform')},
                    {'label': 'Submissions',         'count': count('feedback.FeedbackFormSubmission'), 'url': url('feedback', 'feedbackformsubmission')},
                ],
            },
            {
                'title': 'Staff & HR',
                'icon': 'users',
                'color': '#6b3a7d',
                'models': [
                    {'label': 'Request Templates',   'count': count('staff_requests.RequestTemplate'),    'url': url('staff_requests', 'requesttemplate')},
                    {'label': 'Staff Requests',      'count': count('staff_requests.StaffRequest'),       'url': url('staff_requests', 'staffrequest')},
                    {'label': 'Approval Flows',      'count': count('staff_requests.ApprovalFlow'),       'url': url('staff_requests', 'approvalflow')},
                    {'label': 'Attendance Records',  'count': count('staff_attendance.AttendanceRecord'), 'url': url('staff_attendance', 'attendancerecord')},
                    {'label': 'Salary Records',      'count': count('staff_salary.StaffSalaryRecord'),    'url': url('staff_salary', 'staffsalaryrecord')},
                ],
            },
            {
                'title': 'Timetable',
                'icon': 'calendar',
                'color': '#7d4a1a',
                'models': [
                    {'label': 'Templates',           'count': count('timetable.TimetableTemplate'),   'url': url('timetable', 'timetabletemplate')},
                    {'label': 'Slots',               'count': count('timetable.TimetableSlot'),       'url': url('timetable', 'timetableslot')},
                    {'label': 'Assignments',         'count': count('timetable.TimetableAssignment'), 'url': url('timetable', 'timetableassignment')},
                ],
            },
            {
                'title': 'LMS & Announcements',
                'icon': 'folder-open',
                'color': '#1a5c7d',
                'models': [
                    {'label': 'Study Materials',     'count': count('lms.StudyMaterial'),             'url': url('lms', 'studymaterial')},
                    {'label': 'Announcements',       'count': count('announcements.Announcement'),    'url': url('announcements', 'announcement')},
                    {'label': 'Calendar Events',     'count': count('academic_calendar.AcademicCalendarEvent'), 'url': url('academic_calendar', 'academiccalendarevent')},
                ],
            },
            {
                'title': 'COE (Exams)',
                'icon': 'target',
                'color': '#7d1a1a',
                'models': [
                    {'label': 'Arrear Students',     'count': count('COE.CoeArrearStudent'),     'url': url('COE', 'coearrearsstudent')},
                    {'label': 'Final Results',       'count': count('COE.CoeFinalResult'),       'url': url('COE', 'coefinalresult')},
                    {'label': 'Assignment Stores',   'count': count('COE.CoeAssignmentStore'),   'url': url('COE', 'coeassignmentstore')},
                ],
            },
            {
                'title': 'Applications',
                'icon': 'file-text',
                'color': '#4a7d1a',
                'models': [
                    {'label': 'Application Types',   'count': count('applications.ApplicationType'),  'url': url('applications', 'applicationtype')},
                    {'label': 'Applications',        'count': count('applications.Application'),      'url': url('applications', 'application')},
                    {'label': 'Approval Flows',      'count': count('applications.ApprovalFlow'),     'url': url('applications', 'approvalflow')},
                ],
            },
            {
                'title': 'Certificates & PBAS',
                'icon': 'award',
                'color': '#7d5a1a',
                'models': [
                    {'label': 'Certificates',        'count': count('certificates.Certificate'),          'url': url('certificates', 'certificate')},
                    {'label': 'PBAS Tickets',        'count': count('pbas.PBASVerificationTicket'),       'url': url('pbas', 'pbasverificationticket')},
                ],
            },
            {
                'title': 'IDCSScan',
                'icon': 'scan-line',
                'color': '#1a4a7d',
                'models': [
                    {'label': 'Fingerprint Enrollments', 'count': count('idcsscan.FingerprintEnrollment'), 'url': url('idcsscan', 'fingerprintenrollment')},
                ],
            },
            {
                'title': 'Question Bank',
                'icon': 'help-circle',
                'color': '#4a1a7d',
                'models': [
                    {'label': 'QP Titles',           'count': count('question_bank.QuestionPaperTitle'),  'url': url('question_bank', 'questionpapertitle')},
                ],
            },
        ]

        context = {
            **self.admin_site.each_context(request),
            'college': college,
            'features_on': features_on,
            'features_off': features_off,
            'sections': sections,
            'title': f'Dashboard — {college.short_name or college.name}',
            'opts': self.model._meta,
        }
        return render(request, 'admin/college/college/dashboard.html', context)

    # ------------------------------------------------------------------
    # Excel upload
    # ------------------------------------------------------------------

    def upload_excel_view(self, request: HttpRequest) -> HttpResponse:
        if not self.has_change_permission(request):
            raise PermissionError('Forbidden')

        if request.method == 'POST':
            form = CollegeUploadExcelForm(request.POST, request.FILES)
            if form.is_valid():
                f = form.cleaned_data['file']
                name = getattr(f, 'name', '') or ''
                if not name.lower().endswith('.xlsx'):
                    messages.error(request, 'Please upload an .xlsx Excel file.')
                    return redirect('admin:college_college_upload_excel')

                try:
                    content = f.read()
                    wb = _load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
                    ws = wb.active
                except Exception:
                    messages.error(request, 'Failed to read Excel file. Ensure it is a valid .xlsx file.')
                    return redirect('admin:college_college_upload_excel')

                created = 0
                updated = 0
                skipped = 0
                total = 0

                first = True
                for row in ws.iter_rows(min_row=1, values_only=True):
                    total += 1
                    code = _cell_to_str(row[0] if len(row) > 0 else '')
                    college_name = _cell_to_str(row[1] if len(row) > 1 else '')

                    if first and _is_header_row(code, college_name):
                        first = False
                        skipped += 1
                        continue
                    first = False

                    if not code or not college_name:
                        skipped += 1
                        continue

                    obj = College.objects.filter(code=code).first()
                    if obj is None:
                        College.objects.create(code=code, name=college_name, is_active=True)
                        created += 1
                    else:
                        if (obj.name or '').strip() != college_name:
                            obj.name = college_name
                            obj.save()
                            updated += 1

                messages.success(
                    request,
                    f'College import complete. Created: {created}, Updated: {updated}, Skipped: {skipped} (Rows: {total}).',
                )
                return redirect('admin:college_college_changelist')
        else:
            form = CollegeUploadExcelForm()

        context = {
            **self.admin_site.each_context(request),
            'opts': self.model._meta,
            'form': form,
            'title': 'Upload Colleges Excel',
        }
        return render(request, 'admin/college/college/upload_excel.html', context)


# ---------------------------------------------------------------------------
# Feature Catalog
# ---------------------------------------------------------------------------

@admin.register(FeatureCatalog)
class FeatureCatalogAdmin(admin.ModelAdmin):
    list_display = ('code', 'name', 'category', 'is_default', 'sort_order', 'applicable_roles')
    list_filter = ('category', 'is_default')
    search_fields = ('code', 'name', 'description', 'applicable_roles', 'sidebar_keys')
    ordering = ('sort_order', 'category', 'name')
    list_editable = ('is_default', 'sort_order')
    filter_horizontal = ('permissions',)
    fieldsets = (
        (None, {
            'fields': ('code', 'name', 'description', 'category', 'icon'),
        }),
        ('Defaults & Ordering', {
            'fields': ('is_default', 'sort_order'),
        }),
        ('Mapping', {
            'fields': ('applicable_roles', 'sidebar_keys', 'permissions'),
            'description': 'Comma-separated values linking features to user roles and sidebar keys.',
        }),
    )


# ---------------------------------------------------------------------------
# College Feature
# ---------------------------------------------------------------------------

@admin.register(CollegeFeature)
class CollegeFeatureAdmin(admin.ModelAdmin):
    list_display = ('college', 'feature', 'is_enabled', 'enabled_at', 'disabled_at')
    list_filter = ('is_enabled', 'college', 'feature__category')
    search_fields = ('college__code', 'college__name', 'feature__code', 'feature__name')
    list_editable = ('is_enabled',)
    autocomplete_fields = ('college', 'feature')
    ordering = ('college__code', 'feature__sort_order')
