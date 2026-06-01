from rest_framework import viewsets, status, serializers
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from django.db.models import Count, Q
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage
from .models import CurriculumMaster, CurriculumDepartment, ElectiveSubject, DepartmentGroup, DepartmentGroupMapping, QuestionPaperType
from .serializers import CurriculumMasterSerializer, CurriculumDepartmentSerializer, ElectiveSubjectSerializer, ElectiveChoiceSerializer, DepartmentGroupSerializer
from .permissions import IsIQACOrReadOnly, IsIQACOnly
from accounts.utils import get_user_permissions
from academics.utils import get_user_effective_departments
from academics.models import StudentProfile
import logging
from rest_framework.views import exception_handler, APIView
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
import csv, io, re

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

logger = logging.getLogger(__name__)


def custom_exception_handler(exc, context):
    response = exception_handler(exc, context)

    if response is not None:
        response.data['status_code'] = response.status_code
        response.data['detail'] = str(exc)

    return response

class CurriculumMasterViewSet(viewsets.ModelViewSet):
    # Order master curriculum entries by semester (ascending) so subjects
    # are arranged sem-wise starting from 1. Tie-break by course_code.
    queryset = CurriculumMaster.objects.all().order_by('semester', 'course_code')
    serializer_class = CurriculumMasterSerializer
    permission_classes = [IsIQACOrReadOnly]

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            import traceback
            logging.getLogger(__name__).error('Error creating CurriculumMaster: %s\n%s', e, traceback.format_exc())
            return Response(
                {'detail': 'Failed to create master entry', 'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            import traceback
            logging.getLogger(__name__).error('Error updating CurriculumMaster: %s\n%s', e, traceback.format_exc())
            return Response(
                {'detail': 'Failed to update master entry', 'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def perform_create(self, serializer):
        try:
            serializer.save(created_by=self.request.user)
        except Exception as e:
            logging.getLogger(__name__).exception('Error in perform_create: %s', e)
            raise

    def perform_update(self, serializer):
        try:
            serializer.save()
        except Exception as e:
            logging.getLogger(__name__).exception('Error in perform_update: %s', e)
            raise

    @action(detail=True, methods=['post'], permission_classes=[IsIQACOrReadOnly])
    def propagate(self, request, pk=None):
        try:
            obj = self.get_object()
            obj.save()  # triggers post_save propagation
            return Response({'status': 'propagation triggered'})
        except Exception as e:
            logging.getLogger(__name__).exception('Error in propagate: %s', e)
            return Response(
                {'detail': 'Failed to propagate', 'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class MasterImportView(APIView):
    """API endpoint to import CurriculumMaster CSV using token/JWT auth.

    Expects multipart/form-data with field `csv_file` containing the CSV.
    Only users with IQAC/HAA group membership or superusers are allowed.
    """
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        user = request.user
        # permission: superuser or IQAC/HAA groups
        if not (user.is_superuser or user.groups.filter(name__in=['IQAC', 'HAA']).exists()):
            return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        uploaded = request.FILES.get('csv_file')
        if not uploaded:
            return Response({'detail': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            data = uploaded.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(data))
        except Exception as e:
            return Response({'detail': f'Failed to read CSV: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        created = 0
        updated = 0
        errors = []
        from academics.models import Semester
        from academics.models import Department

        with transaction.atomic():
            for idx, row in enumerate(reader, start=1):
                try:
                    reg = row.get('regulation') or ''
                    sem_raw = (row.get('semester') or '').strip()
                    m = re.search(r"(\d+)", sem_raw)
                    sem_num = int(m.group(1)) if m else 0
                    if not reg or sem_num <= 0:
                        raise ValueError('regulation and semester required')

                    semester_obj, _ = Semester.objects.get_or_create(number=sem_num)

                    cc = row.get('course_code') or None
                    cname = (row.get('course_name') or '').strip() or None
                    instance = None
                    if cc:
                        instance = CurriculumMaster.objects.filter(regulation=reg, semester__number=sem_num, course_code=cc).first()
                    else:
                        if cname:
                            instance = CurriculumMaster.objects.filter(regulation=reg, semester__number=sem_num, course_code__isnull=True, course_name__iexact=cname).first()

                    vals = {
                        'regulation': reg,
                        'semester': semester_obj,
                        'course_code': cc,
                        'course_name': row.get('course_name') or None,
                        'category': row.get('category') or '',
                        'class_type': row.get('class_type') or 'THEORY',
                        'l': int(row.get('l') or 0),
                        't': int(row.get('t') or 0),
                        'p': int(row.get('p') or 0),
                        's': int(row.get('s') or 0),
                        'c': int(row.get('c') or 0),
                        'internal_mark': int(row.get('internal_mark') or 0),
                        'external_mark': int(row.get('external_mark') or 0),
                        'for_all_departments': (str(row.get('for_all_departments') or '').strip().lower() in ('1','true','yes')),
                        'editable': (str(row.get('editable') or '').strip().lower() in ('1','true','yes')),
                    }

                    if instance:
                        for k, v in vals.items():
                            setattr(instance, k, v)
                        instance.save()
                        updated += 1
                    else:
                        instance = CurriculumMaster.objects.create(**vals)
                        created += 1

                    deps = (row.get('departments') or '')
                    if deps:
                        raw = deps.strip().strip('"').strip("'")
                        dep_list = [d.strip() for d in re.split(r'[;,]\s*', raw) if d.strip()]
                        dep_objs = []
                        unmatched = []
                        for d in dep_list:
                            dep = Department.objects.filter(code__iexact=d).first()
                            if not dep and d.isdigit():
                                dep = Department.objects.filter(id=int(d)).first()
                            if dep:
                                dep_objs.append(dep)
                            else:
                                unmatched.append(d)
                        if dep_objs:
                            instance.departments.set(dep_objs)
                            instance.for_all_departments = False
                            instance.save()
                        if unmatched:
                            errors.append(f'Row {idx}: unmatched departments: {",".join(unmatched)}')
                except Exception as e:
                    errors.append(f'Row {idx}: {e}')

        resp = {'created': created, 'updated': updated, 'errors': errors}
        return Response(resp)

class CurriculumDepartmentViewSet(viewsets.ModelViewSet):
    queryset = CurriculumDepartment.objects.all().select_related('department', 'master')
    serializer_class = CurriculumDepartmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Restrict department rows based on user's profile and role.
        user = self.request.user
        qs = CurriculumDepartment.objects.all().select_related('department', 'master', 'semester')
        
        # Basic filtering
        is_elective = self.request.query_params.get('is_elective')
        if is_elective is not None:
            qs = qs.filter(is_elective=is_elective.lower() in ['true', '1'])
        
        semester = self.request.query_params.get('semester')
        if semester:
            try:
                qs = qs.filter(semester__number=int(semester))
            except (ValueError, TypeError):
                pass

        if not user or not user.is_authenticated:
            return qs.none()

        # Log the user and their groups for debugging
        logger.debug('get_queryset: user=%s, groups=%s', user.username, [g.name for g in user.groups.all()])

        # compute user's effective department ids (includes HOD mappings)
        dept_ids = get_user_effective_departments(user)
        if not dept_ids:
            # fallback: try student section
            student = getattr(user, 'student_profile', None)
            if student:
                try:
                    section = getattr(student, 'current_section', None) or student.get_current_section()
                    if section and getattr(section, 'batch', None) and getattr(section.batch, 'course', None):
                        dept_ids = [section.batch.course.department_id]
                except Exception:
                    dept_ids = []

        # Users with global access (superuser, IQAC/HAA groups, custom roles, or explicit wide perms) see all
        try:
            role_names = {r.name.upper() for r in user.roles.all()}
        except Exception:
            role_names = set()
        if user.is_superuser or user.groups.filter(name__in=['IQAC', 'HAA']).exists() or bool(role_names & {'IQAC', 'HAA', 'IQAC_HEAD', 'OBE_MASTER'}):
            logger.debug('get_queryset: user is superuser or IQAC/HAA; user=%s groups=%s roles=%s', user.username, [g.name for g in user.groups.all()], role_names)
            return qs

        perms = get_user_permissions(user)
        logger.debug('get_queryset: user=%s computed dept_ids=%r perms=%s', getattr(user, 'username', None), dept_ids, perms)
        wide_perms = {'curriculum_master_edit', 'curriculum_master_publish', 'CURRICULUM_MASTER_EDIT', 'CURRICULUM_MASTER_PUBLISH', 'obe.master.manage'}
        if perms & wide_perms:
            logger.debug('get_queryset: user has wide_perms, returning all; user=%s', user.username)
            return qs

        # If we found a department for the user, restrict to it
        if dept_ids:
            logger.debug('get_queryset: restricting to departments=%s for user=%s', dept_ids, user.username)
            return qs.filter(department_id__in=dept_ids)

        # otherwise no rows
        logger.debug('get_queryset: no department found, returning none for user=%s', user.username)
        return qs.none()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        # Log the incoming data for debugging
        logger.debug('perform_update: incoming data=%s', serializer.validated_data)

        user = self.request.user
        try:
            _role_names = {r.name.upper() for r in user.roles.all()}
        except Exception:
            _role_names = set()
        privileged = user.is_superuser or user.groups.filter(name__in=['IQAC', 'HAA']).exists() or bool(_role_names & {'IQAC', 'HAA', 'IQAC_HEAD', 'OBE_MASTER'})
        perms = get_user_permissions(user)
        if perms & {'curriculum_department_approve', 'CURRICULUM_DEPARTMENT_APPROVE', 'curriculum.department.approve'}:
            privileged = True

        try:
            if privileged:
                # set approval on update
                instance = serializer.save()
                instance.approval_status = instance.APPROVAL_APPROVED
                instance.approved_by = user
                from django.utils import timezone
                instance.approved_at = timezone.now()
                instance.save(update_fields=['approval_status', 'approved_by', 'approved_at'])
            else:
                # non-privileged -> handled by serializer to mark PENDING
                # capture returned instance so we can include it in the response
                instance = serializer.save()

            # Return the updated instance as a response
            return Response({
                'status': 'success',
                'data': self.get_serializer(instance).data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error('Error during perform_update: %s', str(e))
            raise serializers.ValidationError({'detail': str(e)})

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def approve(self, request, pk=None):
        """Approve or reject a department row. Body: {"action": "approve"|"reject"} """
        obj = self.get_object()
        user = request.user
        # permission: IQAC/HAA or role-permission 'curriculum.department.approve'
        try:
            _role_names = {r.name.upper() for r in user.roles.all()}
        except Exception:
            _role_names = set()
        privileged = user.is_superuser or user.groups.filter(name__in=['IQAC', 'HAA']).exists() or bool(_role_names & {'IQAC', 'HAA', 'IQAC_HEAD', 'OBE_MASTER'})
        perms = get_user_permissions(user)
        if perms & {'curriculum.department.approve', 'CURRICULUM_DEPARTMENT_APPROVE'}:
            privileged = True
        if not privileged:
            return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        action = (request.data or {}).get('action')
        from django.utils import timezone
        if action == 'approve':
            obj.approval_status = obj.APPROVAL_APPROVED
            obj.approved_by = user
            obj.approved_at = timezone.now()
            obj.overridden = False
            obj.save()
            return Response({'status': 'approved'})
        elif action == 'reject':
            obj.approval_status = obj.APPROVAL_REJECTED
            obj.approved_by = user
            obj.approved_at = timezone.now()
            obj.save()
            return Response({'status': 'rejected'})
        else:
            return Response({'detail': 'action must be "approve" or "reject"'}, status=status.HTTP_400_BAD_REQUEST)


class ElectiveSubjectViewSet(viewsets.ModelViewSet):
    queryset = ElectiveSubject.objects.all().select_related('department', 'parent', 'semester', 'department_group')
    serializer_class = ElectiveSubjectSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = ElectiveSubject.objects.all().select_related('department', 'parent', 'semester', 'department_group')
        qs = qs.annotate(student_count=Count('choices', filter=Q(choices__is_active=True)))
        req = self.request
        dept_id = req.query_params.get('department_id')
        regulation = req.query_params.get('regulation')
        semester = req.query_params.get('semester')
        
        if dept_id:
            try:
                dept_id_int = int(dept_id)
                # Find all department groups that this department is mapped to
                group_ids = DepartmentGroupMapping.objects.filter(
                    department_id=dept_id_int,
                    is_active=True
                ).values_list('group_id', flat=True)
                
                # Filter electives that either:
                # 1. Belong directly to this department, OR
                # 2. Have a department_group that this department is mapped to
                qs = qs.filter(
                    Q(department_id=dept_id_int) | 
                    Q(department_group_id__in=list(group_ids))
                )
            except Exception as e:
                logger.error('Error filtering electives by department_id: %s', e)
                pass
        
        if regulation:
            qs = qs.filter(regulation=regulation)
        if semester:
            try:
                qs = qs.filter(semester__number=int(semester))
            except Exception:
                pass
        return qs.order_by('semester', 'course_code')

    def perform_create(self, serializer):
        user = self.request.user
        perms = get_user_permissions(user)
        # allow if superuser or IQAC/HAA groups or explicit permission
        if user.is_superuser or user.groups.filter(name__in=['IQAC', 'HAA']).exists() or 'academics.change_elective_teaching' in perms or 'academics.manage_curriculum' in perms:
            serializer.save(created_by=user)
            return
        # allow HOD of the department
        try:
            staff_profile = getattr(user, 'staff_profile', None)
            if staff_profile:
                hod_depts = DepartmentRole.objects.filter(staff=staff_profile, role='HOD', is_active=True).values_list('department_id', flat=True)
                dept = serializer.validated_data.get('department') or None
                dept_id = getattr(dept, 'id', None) if dept else None
                if dept_id and dept_id in list(hod_depts):
                    serializer.save(created_by=user)
                    return
        except Exception:
            pass
        raise PermissionDenied('You do not have permission to create elective subjects')

    def perform_update(self, serializer):
        user = self.request.user
        perms = get_user_permissions(user)
        if user.is_superuser or user.groups.filter(name__in=['IQAC', 'HAA']).exists() or 'academics.change_elective_teaching' in perms or 'academics.manage_curriculum' in perms:
            return serializer.save()
        # allow HOD of the department
        try:
            staff_profile = getattr(user, 'staff_profile', None)
            if staff_profile:
                hod_depts = DepartmentRole.objects.filter(staff=staff_profile, role='HOD', is_active=True).values_list('department_id', flat=True)
                inst = getattr(serializer, 'instance', None)
                dept_id = getattr(getattr(inst, 'department', None), 'id', None)
                if dept_id and dept_id in list(hod_depts):
                    return serializer.save()
        except Exception:
            pass
        raise PermissionDenied('You do not have permission to change this elective subject')


class CurriculumDepartmentsView(APIView):
    """Return departments filtered by curriculum permissions.
    
    Uses same permission logic as CurriculumDepartmentViewSet:
    - Superusers, IQAC/HAA groups: see all departments
    - Users with curriculum_master_edit/publish: see all departments
    - HODs/regular staff: see only their effective departments
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        from academics.models import Department
        include_non_teaching = str(request.query_params.get('include_non_teaching', 'false')).strip().lower() in {'1', 'true', 'yes'}
        
        # Users with global access see all departments
        if user.is_superuser or user.groups.filter(name__in=['IQAC', 'HAA']).exists():
            qs = Department.objects.all()
        else:
            perms = get_user_permissions(user)
            wide_perms = {'curriculum_master_edit', 'curriculum_master_publish', 
                         'CURRICULUM_MASTER_EDIT', 'CURRICULUM_MASTER_PUBLISH'}
            if perms & wide_perms:
                # Users with wide curriculum permissions see all
                qs = Department.objects.all()
            else:
                # Regular users see only their effective departments
                dept_ids = get_user_effective_departments(user)
                if not dept_ids:
                    # Try student fallback
                    student = getattr(user, 'student_profile', None)
                    if student:
                        try:
                            section = getattr(student, 'current_section', None) or student.get_current_section()
                            if section and getattr(section, 'batch', None) and getattr(section.batch, 'course', None):
                                dept_ids = [section.batch.course.department_id]
                        except Exception:
                            pass
                
                if not dept_ids:
                    return Response({'results': []})
                
                qs = Department.objects.filter(id__in=dept_ids)

        can_include_non_teaching = bool(
            user.is_superuser
            or user.groups.filter(name__in=['IQAC', 'HAA']).exists()
            or (get_user_permissions(user) & {'curriculum_master_edit', 'curriculum_master_publish', 'CURRICULUM_MASTER_EDIT', 'CURRICULUM_MASTER_PUBLISH'})
        )
        if not (include_non_teaching and can_include_non_teaching):
            qs = qs.filter(is_teaching=True)
        
        results = []
        for d in qs:
            results.append({
                'id': d.id, 
                'code': getattr(d, 'code', None), 
                'name': getattr(d, 'name', None), 
                'short_name': getattr(d, 'short_name', None)
            })
        return Response({'results': results})


class CurriculumPendingCountView(APIView):
    """Return IQAC-only pending counts for department curriculum rows."""

    permission_classes = (IsAuthenticated, IsIQACOnly)

    def get(self, request):
        pending_qs = CurriculumDepartment.objects.filter(
            approval_status=CurriculumDepartment.APPROVAL_PENDING,
            is_elective=False,
        )

        total_pending = pending_qs.count()
        department_counts_qs = (
            pending_qs
            .values('department_id', 'department__code', 'department__short_name', 'department__name')
            .annotate(count=Count('id'))
            .order_by('department__code', 'department__name')
        )

        department_counts = []
        for row in department_counts_qs:
            dept_label = row.get('department__short_name') or row.get('department__code') or row.get('department__name') or 'Unknown'
            department_counts.append({
                'departmentId': row.get('department_id'),
                'department': dept_label,
                'count': row.get('count', 0),
            })

        return Response({
            'totalPending': total_pending,
            'departmentCounts': department_counts,
        })


class ElectiveChoicesView(APIView):
    permission_classes = (IsAuthenticated,)

    def _can_manage(self, user):
        perms = get_user_permissions(user)
        return bool(
            user.is_superuser
            or user.groups.filter(name__in=['IQAC', 'HAA']).exists()
            or 'curriculum.import_elective_choices' in perms
            or 'academics.manage_curriculum' in perms
            or 'academics.change_elective_teaching' in perms
        )

    def get(self, request):
        try:
            from .models import ElectiveChoice
            qs = ElectiveChoice.objects.select_related(
                'student__user',
                'student__section',
                'elective_subject',
                'elective_subject__department',
                'elective_subject__parent',
                'academic_year',
            )

            es_id = request.query_params.get('elective_subject_id') or request.query_params.get('elective')
            parent_id = request.query_params.get('parent_id') or request.query_params.get('parent')
            parent_name = request.query_params.get('parent_name')
            department_id = request.query_params.get('department_id')
            regulation = request.query_params.get('regulation')
            semester = request.query_params.get('semester')
            section_id = request.query_params.get('section_id')
            search = request.query_params.get('search') or request.query_params.get('q')
            academic_year = request.query_params.get('academic_year')
            is_active = request.query_params.get('is_active')
            include_inactive = str(request.query_params.get('include_inactive', '')).strip().lower() in {'1', 'true', 'yes', 'y'}
            page_raw = request.query_params.get('page', '1')
            page_size_raw = request.query_params.get('page_size', '10')

            if es_id:
                try:
                    qs = qs.filter(elective_subject_id=int(es_id))
                except Exception:
                    return Response({'results': []})
            if parent_id:
                try:
                    qs = qs.filter(elective_subject__parent_id=int(parent_id))
                except Exception:
                    return Response({'results': []})
            if parent_name:
                qs = qs.filter(elective_subject__parent__course_name__iexact=str(parent_name).strip())
            if department_id:
                try:
                    dept_id = int(department_id)
                    qs = qs.filter(Q(elective_subject__department_id=dept_id) | Q(elective_subject__parent__department_id=dept_id))
                except Exception:
                    return Response({'results': []})
            if regulation:
                qs = qs.filter(elective_subject__regulation__iexact=regulation)
            if semester:
                try:
                    qs = qs.filter(elective_subject__semester__number=int(semester))
                except Exception:
                    return Response({'results': []})
            if section_id:
                try:
                    qs = qs.filter(student__section_id=int(section_id))
                except Exception:
                    return Response({'results': []})
            if academic_year:
                qs = qs.filter(academic_year__name__icontains=academic_year)
            if not include_inactive and (is_active is None or str(is_active).strip() == ''):
                qs = qs.filter(is_active=True)
            elif is_active is not None and str(is_active).strip() != '':
                active_value = str(is_active).strip().lower() in {'1', 'true', 'yes', 'y'}
                qs = qs.filter(is_active=active_value)
            if search:
                qs = qs.filter(
                    Q(student__reg_no__icontains=search)
                    | Q(student__user__username__icontains=search)
                    | Q(student__user__first_name__icontains=search)
                    | Q(student__user__last_name__icontains=search)
                    | Q(elective_subject__course_code__icontains=search)
                    | Q(elective_subject__course_name__icontains=search)
                )

            ordered_qs = qs.order_by('student__section__name', 'student__reg_no', 'elective_subject__course_code')

            try:
                page = max(1, int(page_raw))
            except Exception:
                page = 1
            try:
                page_size = max(1, min(100, int(page_size_raw)))
            except Exception:
                page_size = 10

            paginator = Paginator(ordered_qs, page_size)
            total_count = paginator.count
            total_pages = max(1, paginator.num_pages)
            try:
                page_obj = paginator.page(page)
            except EmptyPage:
                page = total_pages
                page_obj = paginator.page(page)

            results = ElectiveChoiceSerializer(page_obj.object_list, many=True).data
        except Exception:
            return Response({'results': [], 'count': 0, 'page': 1, 'page_size': 10, 'total_pages': 1})

        return Response({
            'results': results,
            'count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
        })

    def patch(self, request):
        if not self._can_manage(request.user):
            return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        choice_id = request.data.get('choice_id') or request.data.get('id')
        if not choice_id:
            return Response({'detail': 'choice_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from .models import ElectiveChoice
            choice = ElectiveChoice.objects.select_related('student__user', 'student__section', 'elective_subject', 'academic_year').get(pk=int(choice_id))
        except (ValueError, TypeError, ElectiveChoice.DoesNotExist):
            return Response({'detail': 'Elective choice not found'}, status=status.HTTP_404_NOT_FOUND)

        elective_subject_id = request.data.get('elective_subject_id')
        academic_year_id = request.data.get('academic_year_id')
        is_active_raw = request.data.get('is_active')

        if elective_subject_id not in (None, '', 'null'):
            try:
                choice.elective_subject_id = int(elective_subject_id)
            except (ValueError, TypeError):
                return Response({'detail': 'Invalid elective_subject_id'}, status=status.HTTP_400_BAD_REQUEST)

        if academic_year_id not in (None, '', 'null'):
            try:
                choice.academic_year_id = int(academic_year_id)
            except (ValueError, TypeError):
                return Response({'detail': 'Invalid academic_year_id'}, status=status.HTTP_400_BAD_REQUEST)

        if is_active_raw not in (None, ''):
            choice.is_active = str(is_active_raw).strip().lower() in {'1', 'true', 'yes', 'y'}

        duplicate = ElectiveChoice.objects.filter(
            student_id=choice.student_id,
            elective_subject_id=choice.elective_subject_id,
            academic_year_id=choice.academic_year_id,
        ).exclude(pk=choice.pk).exists()
        if duplicate:
            return Response({'detail': 'An elective choice already exists for this student and academic year.'}, status=status.HTTP_400_BAD_REQUEST)

        choice.created_by = choice.created_by or request.user
        choice.save()
        return Response(ElectiveChoiceSerializer(choice).data)


class DepartmentGroupViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing department groups. Read-only for now."""
    queryset = DepartmentGroup.objects.filter(is_active=True).order_by('code')
    serializer_class = DepartmentGroupSerializer
    permission_classes = [IsAuthenticated]



class QuestionPaperTypeListView(APIView):
    """Return list of active Question Paper Types.

    GET /api/curriculum/qp-types/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = QuestionPaperType.objects.filter(is_active=True).order_by('sort_order', 'code')
        data = [{'id': q.id, 'code': q.code, 'label': q.label} for q in qs]
        return Response(data)

from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.db.models import Q
from .models import ElectivePoll, ElectivePollSubject, ElectiveChoice, DepartmentGroupMapping
from .serializers import ElectivePollSerializer, ElectivePollSubjectSerializer

class ElectivePollView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        polls = ElectivePoll.objects.all().order_by('-created_at')
        return Response(ElectivePollSerializer(polls, many=True).data)

    def post(self, request):
        serializer = ElectivePollSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ElectivePollDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return ElectivePoll.objects.get(pk=pk)
        except ElectivePoll.DoesNotExist:
            return None

    def get(self, request, pk):
        poll = self.get_object(pk)
        if not poll:
            return Response({'detail': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        return Response(ElectivePollSerializer(poll).data)

    def patch(self, request, pk):
        poll = self.get_object(pk)
        if not poll:
            return Response({'detail': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        raw_active = request.data.get('is_active') if isinstance(request.data, dict) else None
        has_active = raw_active is not None
        desired_active = None
        if has_active:
            desired_active = str(raw_active).strip().lower() in {'1', 'true', 'yes', 'y'}
        serializer = ElectivePollSerializer(poll, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            with transaction.atomic():
                serializer.save()
                if has_active:
                    ElectivePollSubject.objects.filter(poll=poll).update(is_active=desired_active)
            poll.refresh_from_db()
            return Response(ElectivePollSerializer(poll).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        poll = self.get_object(pk)
        if not poll:
            return Response({'detail': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        poll.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ElectivePollSubjectStatusView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, poll_id, subject_id):
        user = request.user
        perms = get_user_permissions(user)
        if not (user.is_superuser or user.groups.filter(name__in=['IQAC']).exists() or 'curriculum.manage_elective_poll' in perms):
            return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        poll_subject = ElectivePollSubject.objects.filter(poll_id=poll_id, id=subject_id).first()
        if not poll_subject:
            return Response({'detail': 'Subject not found in this poll'}, status=status.HTTP_404_NOT_FOUND)

        payload = request.data or {}
        updated = False

        edit_fields = {'seats', 'course_code', 'course_name'}
        if any(field in payload for field in edit_fields) and poll_subject.is_active:
            return Response({'detail': 'Deactivate the subject before editing.'}, status=status.HTTP_400_BAD_REQUEST)

        if 'is_active' in payload:
            raw_active = payload.get('is_active')
            poll_subject.is_active = str(raw_active).strip().lower() in {'1', 'true', 'yes', 'y'}
            updated = True

        if 'seats' in payload:
            raw_seats = payload.get('seats')
            if raw_seats in (None, '', 'null'):
                poll_subject.seats = None
            else:
                try:
                    seats_val = int(raw_seats)
                except (TypeError, ValueError):
                    return Response({'detail': 'Invalid seats value'}, status=status.HTTP_400_BAD_REQUEST)
                if seats_val < 0:
                    return Response({'detail': 'Seats cannot be negative'}, status=status.HTTP_400_BAD_REQUEST)
                poll_subject.seats = seats_val
            updated = True

        es = poll_subject.elective_subject
        if 'course_code' in payload:
            es.course_code = str(payload.get('course_code') or '').strip() or None
            updated = True
        if 'course_name' in payload:
            es.course_name = str(payload.get('course_name') or '').strip() or None
            updated = True

        if not updated:
            return Response({'detail': 'No fields provided'}, status=status.HTTP_400_BAD_REQUEST)

        if updated:
            if 'course_code' in payload or 'course_name' in payload:
                es.save(update_fields=['course_code', 'course_name', 'updated_at'])
            poll_subject.save(update_fields=['is_active', 'seats'])

        # Keep poll status in sync with subject status.
        if 'is_active' in payload:
            poll = poll_subject.poll
            if poll_subject.is_active:
                if not poll.is_active:
                    poll.is_active = True
                    poll.save(update_fields=['is_active'])
            else:
                has_active = ElectivePollSubject.objects.filter(poll=poll, is_active=True).exists()
                if not has_active and poll.is_active:
                    poll.is_active = False
                    poll.save(update_fields=['is_active'])

        return Response(ElectivePollSubjectSerializer(poll_subject).data)


class ElectivePollSeatCountView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        if not (user.is_superuser or user.groups.filter(name__in=['IQAC']).exists() or 'curriculum.manage_elective_poll' in perms):
            return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        batch_year_id = request.query_params.get('batch_year_id')
        if not batch_year_id:
            return Response({'detail': 'batch_year_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            batch_year_id = int(batch_year_id)
        except (TypeError, ValueError):
            return Response({'detail': 'Invalid batch_year_id'}, status=status.HTTP_400_BAD_REQUEST)

        counts: dict[int, int] = {}
        students = StudentProfile.objects.filter(
            status='ACTIVE',
            section__batch__batch_year_id=batch_year_id,
        ).select_related(
            'home_department',
            'section__batch__course__department',
            'section__batch__department',
        )

        for student in students:
            dept = getattr(student, 'home_department', None)
            if dept is None:
                try:
                    dept = student.section.batch.course.department
                except Exception:
                    dept = None
            if dept is None:
                try:
                    dept = student.section.batch.department
                except Exception:
                    dept = None
            if dept is None:
                continue
            counts[dept.id] = counts.get(dept.id, 0) + 1

        return Response({'batch_year_id': batch_year_id, 'counts': counts})

class ActiveStudentPollsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            student_profile = getattr(request.user, 'student_profile', None)
            if not student_profile:
                return Response({'detail': 'User is not a student'}, status=status.HTTP_403_FORBIDDEN)

            section = student_profile.get_current_section()
            if not section:
                return Response([], status=status.HTTP_200_OK)

            batch = section.batch
            
            home_dept = student_profile.home_department
            if not home_dept and getattr(section, 'batch', None) and getattr(section.batch, 'course', None):
                home_dept = section.batch.course.department

            # Find applicable department groups
            dept_groups = []
            if home_dept:
                dept_groups = list(DepartmentGroupMapping.objects.filter(
                    department=home_dept, is_active=True
                ).values_list('group_id', flat=True))

            # Start with active polls
            query = Q(is_active=True)
            
            # Match batch year if set
            if batch and getattr(batch, 'batch_year_id', None):
                query &= (Q(batch_year_id__isnull=True) | Q(batch_year_id=batch.batch_year_id))
                
            # Match department group if set
            if dept_groups:
                query &= (Q(department_group_id__isnull=True) | Q(department_group_id__in=dept_groups))
            else:
                # If student has no department group mapping, only show polls without group restrictions
                query &= Q(department_group_id__isnull=True)

            polls = ElectivePoll.objects.filter(query).distinct()
            
            # Annotate each poll with the student's existing choice (poll_subject id)
            serialized = ElectivePollSerializer(polls, many=True).data
            existing_choices = {
                ec.elective_subject_id: ec
                for ec in ElectiveChoice.objects.filter(
                    student=student_profile,
                    elective_subject__poll_associations__poll__in=polls
                ).select_related('elective_subject').distinct()
            }
            choice_counts = {
                row['elective_subject_id']: row['count']
                for row in ElectiveChoice.objects.filter(
                    elective_subject__poll_associations__poll__in=polls
                ).values('elective_subject_id').annotate(count=Count('id'))
            }
            # Build a map from elective_subject_id -> poll_subject_id for quick lookup
            elective_to_ps = {
                ps.elective_subject_id: ps.id
                for ps in ElectivePollSubject.objects.filter(poll__in=polls)
            }
            result = []
            for poll_data in serialized:
                poll_data = dict(poll_data)
                
                # Filter out poll_subjects that are blocked for the current student's department
                if home_dept:
                    all_subjects = poll_data.get('poll_subjects', [])
                    filtered_subjects = [
                        ps for ps in all_subjects 
                        if home_dept.id not in (ps.get('blocked_departments') or [])
                    ]
                    poll_data['poll_subjects'] = filtered_subjects

                chosen_ps_id = None
                for ps in poll_data.get('poll_subjects', []):
                    # We need the ES ID to check against existing choices
                    # Check if it was provided in the serialized data or lookup if needed
                    ps_id = ps['id']
                    es_id = ps.get('elective_subject_id')
                    if not es_id:
                        es_id = ElectivePollSubject.objects.filter(id=ps_id).values_list('elective_subject_id', flat=True).first()

                    if es_id:
                        chosen_count = choice_counts.get(es_id, 0)
                        seats_left = ps.get('seats')
                        total_seats = None
                        if seats_left is not None:
                            try:
                                total_seats = int(seats_left) + int(chosen_count)
                            except Exception:
                                total_seats = None
                        ps['chosen_count'] = chosen_count
                        ps['total_seats'] = total_seats

                    if es_id and es_id in existing_choices:
                        chosen_ps_id = ps_id
                        ec = existing_choices.get(es_id)
                        if ec:
                            rank = ElectiveChoice.objects.filter(
                                Q(elective_subject_id=es_id)
                                & (Q(created_at__lt=ec.created_at) | Q(created_at=ec.created_at, id__lte=ec.id))
                            ).count()
                            ps['your_rank'] = rank
                        break
                poll_data['your_choice_poll_subject_id'] = chosen_ps_id
                result.append(poll_data)

            return Response(result)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HodElectivePollStatusView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        if not (user.is_superuser or user.groups.filter(name__in=['IQAC']).exists() or 'curriculum.hod_elective_manage' in perms):
            return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        dept_ids = get_user_effective_departments(user) or []
        if not dept_ids:
            return Response({'departments': []})

        from academics.models import Department, StudentProfile

        dept_qs = Department.objects.filter(id__in=dept_ids)
        dept_map = {d.id: d for d in dept_qs}

        group_map = {}
        for dept_id, group_id in DepartmentGroupMapping.objects.filter(
            department_id__in=dept_ids, is_active=True
        ).values_list('department_id', 'group_id'):
            group_map.setdefault(dept_id, set()).add(group_id)

        polls = list(
            ElectivePoll.objects.select_related('batch_year', 'department_group')
            .order_by('-created_at')
        )

        students = list(
            StudentProfile.objects.select_related(
                'user',
                'home_department',
                'section',
                'section__batch',
                'section__batch__batch_year',
                'section__batch__course',
                'section__batch__course__department',
                'section__batch__department',
                'section__managing_department',
            )
            .filter(status='ACTIVE')
            .filter(
                Q(home_department_id__in=dept_ids)
                | Q(section__batch__course__department_id__in=dept_ids)
                | Q(section__batch__department_id__in=dept_ids)
                | Q(section__managing_department_id__in=dept_ids)
            )
            .distinct()
        )

        def resolve_student_dept_id(student):
            if getattr(student, 'home_department_id', None):
                return student.home_department_id
            section = getattr(student, 'section', None)
            if section and getattr(section, 'managing_department_id', None):
                return section.managing_department_id
            batch = getattr(section, 'batch', None) if section else None
            if batch:
                if getattr(batch, 'course_id', None) and getattr(batch, 'course', None):
                    return getattr(batch.course, 'department_id', None)
                if getattr(batch, 'department_id', None):
                    return batch.department_id
            return None

        def resolve_student_batch_year(student):
            section = getattr(student, 'section', None)
            batch = getattr(section, 'batch', None) if section else None
            if batch and getattr(batch, 'batch_year_id', None):
                return batch.batch_year_id, getattr(getattr(batch, 'batch_year', None), 'name', None)
            return None, None

        students_by_dept = {dept_id: [] for dept_id in dept_ids}
        all_student_ids = []
        for student in students:
            dept_id = resolve_student_dept_id(student)
            if not dept_id or dept_id not in students_by_dept:
                continue
            dept_label = None
            dept_obj = dept_map.get(dept_id)
            if dept_obj:
                dept_label = getattr(dept_obj, 'short_name', None) or getattr(dept_obj, 'code', None) or getattr(dept_obj, 'name', None)
            batch_year_id, batch_year_name = resolve_student_batch_year(student)
            user_obj = getattr(student, 'user', None)
            student_name = None
            if user_obj:
                student_name = (getattr(user_obj, 'get_full_name', lambda: '')() or '').strip() or getattr(user_obj, 'username', None)
            students_by_dept[dept_id].append({
                'student_id': student.id,
                'reg_no': student.reg_no,
                'name': student_name,
                'username': getattr(user_obj, 'username', None) if user_obj else None,
                'section': getattr(getattr(student, 'section', None), 'name', None),
                'department': dept_label,
                'batch_year_id': batch_year_id,
                'batch_year_name': batch_year_name,
            })
            all_student_ids.append(student.id)

        poll_ids = [p.id for p in polls]
        choice_map = {}
        if poll_ids and all_student_ids:
            choice_rows = ElectiveChoice.objects.filter(
                student_id__in=all_student_ids,
                elective_subject__poll_associations__poll_id__in=poll_ids,
            ).values(
                'student_id',
                'elective_subject__poll_associations__poll_id',
                'elective_subject__course_name',
                'elective_subject__course_code',
            )
            for row in choice_rows:
                key = (row.get('student_id'), row.get('elective_subject__poll_associations__poll_id'))
                if key not in choice_map:
                    choice_map[key] = {
                        'course_name': row.get('elective_subject__course_name'),
                        'course_code': row.get('elective_subject__course_code'),
                    }

        department_payloads = []
        for dept_id in dept_ids:
            dept_obj = dept_map.get(dept_id)
            dept_label = None
            if dept_obj:
                dept_label = dept_obj.short_name or dept_obj.code or dept_obj.name
            dept_label = dept_label or f'Dept {dept_id}'
            group_ids = group_map.get(dept_id, set())

            poll_payloads = []
            for poll in polls:
                if poll.department_group_id and poll.department_group_id not in group_ids:
                    continue

                student_rows = []
                for student in students_by_dept.get(dept_id, []):
                    if poll.batch_year_id:
                        if not student.get('batch_year_id') or poll.batch_year_id != student.get('batch_year_id'):
                            continue
                    choice_key = (student['student_id'], poll.id)
                    choice_info = choice_map.get(choice_key)
                    chosen = choice_info is not None
                    student_rows.append({
                        'student_id': student['student_id'],
                        'reg_no': student.get('reg_no'),
                        'name': student.get('name'),
                        'username': student.get('username'),
                        'section': student.get('section'),
                        'batch_year': student.get('batch_year_name'),
                        'chosen': chosen,
                        'chosen_subject_name': choice_info.get('course_name') if choice_info else None,
                        'chosen_subject_code': choice_info.get('course_code') if choice_info else None,
                    })

                poll_payloads.append({
                    'poll_id': poll.id,
                    'parent_elective_name': poll.parent_elective_name,
                    'batch_year': getattr(getattr(poll, 'batch_year', None), 'name', None),
                    'department_group': getattr(getattr(poll, 'department_group', None), 'name', None),
                    'total_students': len(student_rows),
                    'chosen_count': sum(1 for s in student_rows if s['chosen']),
                    'students': student_rows,
                })

            department_payloads.append({
                'department_id': dept_id,
                'department': dept_label,
                'polls': poll_payloads,
            })

        return Response({'departments': department_payloads})

class SubmitElectiveChoiceView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            student_profile = getattr(request.user, 'student_profile', None)
            if not student_profile:
                return Response({'detail': 'User is not a student'}, status=status.HTTP_403_FORBIDDEN)

            poll = ElectivePoll.objects.get(id=pk, is_active=True)
            poll_subject_id = request.data.get('poll_subject_id')
            if not poll_subject_id:
                return Response({'detail': 'poll_subject_id is required'}, status=status.HTTP_400_BAD_REQUEST)

            with transaction.atomic():
                # Lock student row to prevent concurrent double-submits from same user.
                from academics.models import StudentProfile
                StudentProfile.objects.select_for_update().get(pk=student_profile.pk)
                poll_subject = ElectivePollSubject.objects.select_for_update().get(
                    id=poll_subject_id,
                    poll=poll,
                    is_active=True,
                )
                
                # Check if the student has already chosen an elective for this poll
                existing_choice = ElectiveChoice.objects.filter(
                    student=student_profile,
                    elective_subject__poll_associations__poll=poll
                ).first()
                if existing_choice:
                    return Response({'detail': 'You have already submitted a choice for this poll'}, status=status.HTTP_400_BAD_REQUEST)
                
                # Check seats
                if poll_subject.seats is not None:
                    if poll_subject.seats <= 0:
                        return Response({'detail': 'Seats full for this subject'}, status=status.HTTP_400_BAD_REQUEST)
                    poll_subject.seats -= 1
                    poll_subject.save()
                
                # Get the active academic year to associate with the choice
                from academics.models import AcademicYear
                current_year = AcademicYear.objects.filter(is_active=True).first()

                ElectiveChoice.objects.create(
                    student=student_profile,
                    elective_subject=poll_subject.elective_subject,
                    academic_year=current_year,
                    created_by=request.user
                )
            
            return Response({'detail': 'Choice submitted successfully'}, status=status.HTTP_201_CREATED)
        except ElectivePoll.DoesNotExist:
            return Response({'detail': 'Active poll not found'}, status=status.HTTP_404_NOT_FOUND)
        except ElectivePollSubject.DoesNotExist:
            return Response({'detail': 'Subject not found or inactive in this poll'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ElectivePollExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            from openpyxl import Workbook
        except Exception:
            return Response({'detail': 'Excel export requires openpyxl.'}, status=status.HTTP_400_BAD_REQUEST)

        poll = ElectivePoll.objects.filter(pk=pk).first()
        if not poll:
            return Response({'detail': 'Poll not found'}, status=status.HTTP_404_NOT_FOUND)

        poll_subjects = ElectivePollSubject.objects.filter(poll=poll).select_related(
            'elective_subject',
            'elective_subject__department',
            'staff',
            'staff__user'
        )
        subject_staff = {ps.elective_subject_id: ps.staff for ps in poll_subjects}

        choices = ElectiveChoice.objects.filter(
            elective_subject__poll_associations__poll=poll
        ).select_related(
            'student__user',
            'student__section',
            'elective_subject',
            'elective_subject__department',
            'elective_subject__parent',
            'academic_year',
        ).order_by('elective_subject__course_code', 'student__reg_no')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Chosen'

        header = [
            'Poll', 'Batch Year',
            'Subject Code', 'Subject Name',
            'Providing Dept', 'Staff',
            'Student Reg No', 'Student Name', 'Student Username',
            'Section', 'Student Department', 'Academic Year'
        ]
        ws.append(header)

        for choice in choices:
            es = choice.elective_subject
            staff = subject_staff.get(es.id)
            staff_name = None
            if staff and getattr(staff, 'user', None):
                staff_name = staff.user.get_full_name() or staff.user.username
            student = getattr(choice, 'student', None)
            user = getattr(student, 'user', None)
            student_name = None
            if student:
                student_name = getattr(student, 'name', None) or (user.get_full_name() if user else None) or (user.username if user else None)
            dept = getattr(es, 'department', None)
            dept_short = getattr(dept, 'short_name', None) if dept else None
            if not dept_short and dept is not None:
                dept_short = getattr(dept, 'code', None)
            student_dept = None
            student_dept_obj = getattr(student, 'home_department', None)
            if not student_dept_obj and getattr(student, 'section', None):
                sec_batch = getattr(getattr(student, 'section', None), 'batch', None)
                if sec_batch and getattr(sec_batch, 'course', None):
                    student_dept_obj = getattr(sec_batch.course, 'department', None)
            if student_dept_obj:
                student_dept = getattr(student_dept_obj, 'short_name', None) or getattr(student_dept_obj, 'code', None) or getattr(student_dept_obj, 'name', None)
            ws.append([
                poll.parent_elective_name,
                getattr(poll.batch_year, 'name', None),
                getattr(es, 'course_code', None),
                getattr(es, 'course_name', None),
                dept_short,
                staff_name,
                getattr(student, 'reg_no', None),
                student_name,
                getattr(user, 'username', None),
                getattr(getattr(student, 'section', None), 'name', None),
                student_dept,
                getattr(getattr(choice, 'academic_year', None), 'name', None),
            ])

        # Build "Not chosen" sheet with eligible students for this poll
        not_chosen_ws = wb.create_sheet('Not chosen')
        not_chosen_ws.append(header)

        chosen_student_ids = set(choices.values_list('student_id', flat=True))

        from academics.models import StudentProfile

        def resolve_student_dept(student):
            if getattr(student, 'home_department_id', None):
                return student.home_department
            section = getattr(student, 'section', None)
            if section and getattr(section, 'batch', None) and getattr(section.batch, 'course', None):
                return section.batch.course.department
            return None

        students_qs = StudentProfile.objects.select_related(
            'user',
            'section',
            'section__batch',
            'section__batch__batch_year',
            'section__batch__course',
            'section__batch__course__department',
            'home_department',
        ).filter(status='ACTIVE')

        if poll.batch_year_id:
            students_qs = students_qs.filter(section__batch__batch_year_id=poll.batch_year_id)

        for student in students_qs:
            if student.id in chosen_student_ids:
                continue

            dept = resolve_student_dept(student)
            if poll.department_group_id:
                if not dept:
                    continue
                in_group = DepartmentGroupMapping.objects.filter(
                    department=dept,
                    group_id=poll.department_group_id,
                    is_active=True
                ).exists()
                if not in_group:
                    continue
            else:
                # If student has no dept group mapping, allow only polls without group restrictions
                if dept and DepartmentGroupMapping.objects.filter(department=dept, is_active=True).exists():
                    pass

            user_obj = getattr(student, 'user', None)
            student_name = None
            if user_obj:
                student_name = getattr(student, 'name', None) or (user_obj.get_full_name() if user_obj else None) or (user_obj.username if user_obj else None)

            student_dept = None
            if dept:
                student_dept = getattr(dept, 'short_name', None) or getattr(dept, 'code', None) or getattr(dept, 'name', None)
            not_chosen_ws.append([
                poll.parent_elective_name,
                getattr(poll.batch_year, 'name', None),
                None,
                None,
                None,
                None,
                getattr(student, 'reg_no', None),
                student_name,
                getattr(user_obj, 'username', None),
                getattr(getattr(student, 'section', None), 'name', None),
                student_dept,
                None,
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"elective_poll_{poll.id}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
