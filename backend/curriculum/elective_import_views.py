from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from django.http import HttpResponse
from django.db import transaction
from accounts.utils import get_user_permissions
import csv
import io

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


class ElectiveChoiceTemplateDownloadView(APIView):
    """Download an Excel template for importing elective student mappings with department dropdown."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        
        # Check permission
        if not ('curriculum.import_elective_choices' in perms or user.is_staff or user.is_superuser):
            return Response({'error': 'You do not have permission to download elective import template'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        if not EXCEL_SUPPORT:
            return Response({'error': 'Excel support not available. Please install openpyxl.'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Get all departments
            from academics.models import Department
            departments = Department.objects.all().order_by('short_name')
            dept_names = [dept.short_name for dept in departments if dept.short_name]
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Elective Choices"
            
            # Headers
            headers = ['student_reg_no', 'elective_subject_code', 'department', 'semester_type', 'academic_year', 'is_active']
            ws.append(headers)
            
            # Sample row
            sample_dept = dept_names[0] if dept_names else 'AI&DS'
            ws.append(['REG001', 'ELEC001', sample_dept, 'ODD', '2025-2026', 'TRUE'])
            
            # Add data validation (dropdown) for department column (column C)
            if dept_names:
                # Create a comma-separated list of department names
                dept_list = ','.join(dept_names)
                
                # Create data validation for department column
                dv_dept = DataValidation(type="list", formula1=f'"{dept_list}"', allow_blank=True)
                dv_dept.error = 'Please select a department from the dropdown'
                dv_dept.errorTitle = 'Invalid Department'
                dv_dept.prompt = 'Select a department'
                dv_dept.promptTitle = 'Department'
                
                # Apply validation to department column (C2:C1000)
                ws.add_data_validation(dv_dept)
                dv_dept.add(f'C2:C1000')
            
            # Add data validation (dropdown) for semester_type column (column D)
            sem_types = 'ODD,EVEN'
            dv_sem_type = DataValidation(type="list", formula1=f'"{sem_types}"', allow_blank=False)
            dv_sem_type.error = 'Please select ODD or EVEN'
            dv_sem_type.errorTitle = 'Invalid Semester Type'
            dv_sem_type.prompt = 'Select semester type (ODD or EVEN)'
            dv_sem_type.promptTitle = 'Semester Type'
            
            # Apply validation to semester_type column (D2:D1000)
            ws.add_data_validation(dv_sem_type)
            dv_sem_type.add(f'D2:D1000')
            
            # Format academic_year column (E) as text to prevent Excel auto-formatting
            from openpyxl.styles import numbers
            for row in range(2, 1001):  # Apply to rows 2-1000
                ws[f'E{row}'].number_format = numbers.FORMAT_TEXT
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20  # student_reg_no
            ws.column_dimensions['B'].width = 25  # elective_subject_code
            ws.column_dimensions['C'].width = 15  # department
            ws.column_dimensions['D'].width = 15  # semester_type
            ws.column_dimensions['E'].width = 20  # academic_year
            ws.column_dimensions['F'].width = 12  # is_active
            
            # Save to bytes
            from io import BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="elective_choices_template.xlsx"'
            return response
            
        except Exception as e:
            return Response({'error': f'Failed to generate template: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ElectivePollSubjectTemplateDownloadView(APIView):
    """Download an Excel template for importing elective poll subjects."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)

        if not ('curriculum.manage_elective_poll' in perms or user.is_staff or user.is_superuser):
            return Response({'error': 'You do not have permission to download elective poll template'},
                            status=status.HTTP_403_FORBIDDEN)

        if not EXCEL_SUPPORT:
            return Response({'error': 'Excel support not available. Please install openpyxl.'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            from academics.models import Department, StaffProfile
            from io import BytesIO

            departments = Department.objects.all().order_by('short_name', 'code', 'name')
            staff_members = StaffProfile.objects.select_related('user').order_by('staff_id', 'id')

            dept_values = []
            for dept in departments:
                name_label = dept.name or dept.short_name or dept.code or str(dept.id)
                suffix = dept.short_name or dept.code
                dept_values.append(f"{name_label} ({suffix})" if suffix else name_label)

            staff_values = []
            for staff in staff_members:
                user_obj = getattr(staff, 'user', None)
                full_name = ''
                if user_obj:
                    full_name = (getattr(user_obj, 'get_full_name', lambda: '')() or '').strip()
                staff_label = full_name or getattr(user_obj, 'username', '') or staff.staff_id or str(staff.id)
                suffix = f" ({staff.staff_id})" if staff.staff_id else ''
                staff_values.append(f"{staff_label}{suffix}")

            wb = Workbook()
            ws = wb.active
            ws.title = 'Elective Subjects'

            headers = ['code', 'name', 'seats', 'staff', 'dept', 'block_rule']
            ws.append(headers)

            sample_dept = dept_values[0] if dept_values else ''
            sample_staff = staff_values[0] if staff_values else ''
            ws.append(['CS301', 'Advanced Machine Learning', '60', sample_staff, sample_dept, 'Block providing dept'])

            lists_ws = wb.create_sheet('Lists')
            lists_ws.sheet_state = 'hidden'

            for idx, value in enumerate(staff_values, start=1):
                lists_ws.cell(row=idx, column=1, value=value)

            for idx, value in enumerate(dept_values, start=1):
                lists_ws.cell(row=idx, column=2, value=value)

            if staff_values:
                staff_range = f"Lists!$A$1:$A${len(staff_values)}"
                dv_staff = DataValidation(type="list", formula1=staff_range, allow_blank=True)
                dv_staff.error = 'Please select a staff member from the dropdown'
                dv_staff.errorTitle = 'Invalid Staff'
                dv_staff.prompt = 'Select staff'
                dv_staff.promptTitle = 'Staff'
                ws.add_data_validation(dv_staff)
                dv_staff.add('D2:D1000')

            if dept_values:
                dept_range = f"Lists!$B$1:$B${len(dept_values)}"
                dv_dept = DataValidation(type="list", formula1=dept_range, allow_blank=False)
                dv_dept.error = 'Please select a department from the dropdown'
                dv_dept.errorTitle = 'Invalid Department'
                dv_dept.prompt = 'Select department'
                dv_dept.promptTitle = 'Department'
                ws.add_data_validation(dv_dept)
                dv_dept.add('E2:E1000')

            block_values = ['Block providing dept', 'Block outside group']
            for idx, value in enumerate(block_values, start=1):
                lists_ws.cell(row=idx, column=3, value=value)

            block_range = f"Lists!$C$1:$C${len(block_values)}"
            dv_block = DataValidation(type="list", formula1=block_range, allow_blank=True)
            dv_block.error = 'Please select a block rule from the dropdown'
            dv_block.errorTitle = 'Invalid Block Rule'
            dv_block.prompt = 'Select block rule'
            dv_block.promptTitle = 'Block Rule'
            ws.add_data_validation(dv_block)
            dv_block.add('F2:F1000')

            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 36
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 32
            ws.column_dimensions['E'].width = 28
            ws.column_dimensions['F'].width = 24

            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="elective_poll_subjects_template.xlsx"'
            return response

        except Exception as e:
            return Response({'error': f'Failed to generate template: {str(e)}'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ElectiveChoiceBulkImportView(APIView):
    """Bulk import elective student mappings from CSV or Excel file."""
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        user = request.user
        perms = get_user_permissions(user)
        
        # Check permission
        if not ('curriculum.import_elective_choices' in perms or user.is_staff or user.is_superuser):
            return Response({'error': 'You do not have permission to import elective choices'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        uploaded_file = request.FILES.get('csv_file')
        if not uploaded_file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        filename = uploaded_file.name.lower()
        is_excel = filename.endswith(('.xlsx', '.xls'))
        is_csv = filename.endswith('.csv')
        
        if not (is_csv or is_excel):
            return Response({'error': 'File must be CSV or Excel (.xlsx, .xls)'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Parse file based on format
        rows = []
        if is_excel:
            if not EXCEL_SUPPORT:
                return Response({'error': 'Excel support not available. Please install openpyxl or use CSV format.'}, 
                              status=status.HTTP_400_BAD_REQUEST)
            try:
                wb = load_workbook(uploaded_file, read_only=True)
                ws = wb.active
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value)
                
                # Read data rows
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    row_dict = {}
                    for idx, value in enumerate(row):
                        if idx < len(headers) and headers[idx]:
                            row_dict[headers[idx]] = str(value) if value is not None else ''
                    if any(row_dict.values()):  # Skip empty rows
                        rows.append((row_idx, row_dict))
            except Exception as e:
                return Response({'error': f'Failed to parse Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            # Parse CSV
            try:
                decoded_file = uploaded_file.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                rows = [(idx, row) for idx, row in enumerate(reader, start=2)]
            except Exception as e:
                return Response({'error': f'Failed to parse CSV file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Process rows
        try:
            from .models import ElectiveChoice, ElectiveSubject
            from academics.models import StudentProfile, AcademicYear, Department
            
            created_count = 0
            updated_count = 0
            errors = []
            
            with transaction.atomic():
                for idx, row in rows:
                    try:
                        student_reg = row.get('student_reg_no', '').strip()
                        elective_code = row.get('elective_subject_code', '').strip()
                        dept_name = row.get('department', '').strip()
                        semester_type = row.get('semester_type', '').strip().upper()
                        ay_name = row.get('academic_year', '').strip()
                        is_active_str = row.get('is_active', 'TRUE').strip().upper()
                        
                        if not student_reg or not elective_code or not dept_name or not semester_type:
                            errors.append(f'Row {idx}: Missing required fields (student_reg_no, elective_subject_code, department, semester_type)')
                            continue
                        
                        # Validate semester_type
                        if semester_type not in ('ODD', 'EVEN'):
                            errors.append(f'Row {idx}: semester_type must be ODD or EVEN, got "{semester_type}"')
                            continue
                        
                        # Find student
                        try:
                            student = StudentProfile.objects.get(reg_no=student_reg)
                        except StudentProfile.DoesNotExist:
                            errors.append(f'Row {idx}: Student with reg_no "{student_reg}" not found')
                            continue
                        
                        # Find department
                        try:
                            department = Department.objects.get(short_name__iexact=dept_name)
                        except Department.DoesNotExist:
                            errors.append(f'Row {idx}: Department with short_name "{dept_name}" not found')
                            continue
                        
                        # Find academic year with parity
                        academic_year = None
                        if ay_name:
                            try:
                                # Try exact match first
                                academic_year = AcademicYear.objects.filter(name=ay_name, parity=semester_type).first()
                                
                                # If not found, try normalizing the format
                                if not academic_year:
                                    # Convert "2025-26" to "2025-2026" or "2025-2026" to "2025-26"
                                    import re
                                    match = re.match(r'^(\d{4})-(\d{2,4})$', ay_name)
                                    if match:
                                        start_year = match.group(1)
                                        end_year = match.group(2)
                                        
                                        # Try alternative formats
                                        if len(end_year) == 2:
                                            # Convert "2025-26" to "2025-2026"
                                            full_end = start_year[:2] + end_year
                                            academic_year = AcademicYear.objects.filter(name=f'{start_year}-{full_end}', parity=semester_type).first()
                                        elif len(end_year) == 4:
                                            # Convert "2025-2026" to "2025-26"
                                            short_end = end_year[-2:]
                                            academic_year = AcademicYear.objects.filter(name=f'{start_year}-{short_end}', parity=semester_type).first()
                                
                                if not academic_year:
                                    # Get available academic years for better error message
                                    available = AcademicYear.objects.filter(parity=semester_type).values_list('name', flat=True)[:5]
                                    available_str = ', '.join(available) if available else 'none'
                                    errors.append(f'Row {idx}: Academic year "{ay_name}" with semester type "{semester_type}" not found. Available: {available_str}')
                                    continue
                            except Exception as e:
                                errors.append(f'Row {idx}: Error finding academic year: {str(e)}')
                                continue
                        else:
                            # Use active academic year with matching parity
                            academic_year = AcademicYear.objects.filter(is_active=True, parity=semester_type).first()
                            if not academic_year:
                                errors.append(f'Row {idx}: No active academic year found with semester type "{semester_type}". Please specify academic_year in CSV.')
                                continue
                        
                        # Find elective subject by code, department, and semester (which should match the academic year parity)
                        try:
                            # Filter by course_code and department
                            elective = ElectiveSubject.objects.filter(
                                course_code=elective_code,
                                department=department
                            ).first()
                            
                            if not elective:
                                errors.append(f'Row {idx}: Elective subject with code "{elective_code}" in department "{dept_name}" not found')
                                continue
                        except Exception as e:
                            errors.append(f'Row {idx}: Error finding elective subject: {str(e)}')
                            continue
                        
                        is_active = is_active_str in ('TRUE', '1', 'YES', 'Y')
                        
                        # Create or update choice
                        choice, created = ElectiveChoice.objects.update_or_create(
                            student=student,
                            elective_subject=elective,
                            academic_year=academic_year,
                            defaults={
                                'is_active': is_active,
                                'created_by': user
                            }
                        )
                        
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                            
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
            
            result = {
                'message': 'Import completed',
                'created': created_count,
                'updated': updated_count,
                'errors': errors[:50]  # Limit error messages
            }
            
            if errors:
                result['warning'] = f'Import completed with {len(errors)} errors'
            
            return Response(result, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({'error': f'Failed to process file: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)
