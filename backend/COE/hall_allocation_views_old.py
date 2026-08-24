import json
from io import BytesIO

from django.http import HttpResponse
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def allocate_seats_simple(students_list, rows, cols):
    """Simple seat allocation: distribute students in column-major order."""
    seats = []
    student_idx = 0
    
    for col_idx in range(cols):
        for row_idx in range(rows):
            seat_label = f"{chr(65 + col_idx)}{row_idx + 1}"  # A1, A2, ..., B1, B2, ...
            if student_idx < len(students_list):
                student_id = students_list[student_idx]
                seats.append({
                    "seat_label": seat_label,
                    "student_id": student_id,
                })
                student_idx += 1
            else:
                seats.append({
                    "seat_label": seat_label,
                    "student_id": "",
                })
    
    return seats


def seat_positions_for_pattern(rows, cols, pattern):
    """Return a list of (row, col) positions in the order this pattern fills seats."""
    positions = []

    if pattern in (None, '', 'Zigzag'):
        for col_idx in range(cols):
            row_order = list(range(rows)) if col_idx % 2 == 0 else list(reversed(range(rows)))
            for row_idx in row_order:
                positions.append((row_idx, col_idx))
        return positions

    if pattern == 'Straight':
        for row_idx in range(rows):
            for col_idx in range(cols):
                positions.append((row_idx, col_idx))
        return positions

    if pattern == 'Alternate Zigzag':
        for pair_start in range(0, cols, 2):
            pair_direction = 1 if (pair_start // 2) % 2 == 0 else -1
            pair_cols = list(range(pair_start, min(pair_start + 2, cols)))
            for col_offset, col_idx in enumerate(pair_cols):
                direction = pair_direction if col_offset == 0 else -pair_direction
                row_order = list(range(rows)) if direction == 1 else list(reversed(range(rows)))
                for row_idx in row_order:
                    positions.append((row_idx, col_idx))
        return positions

    if pattern == 'U-Shape':
        for row_idx in range(rows):
            for col_idx in range(cols):
                is_boundary = row_idx in (0, rows - 1) or col_idx in (0, cols - 1)
                if is_boundary:
                    positions.append((row_idx, col_idx))
        return positions

    if pattern == 'Circle':
        center_row = rows // 2
        center_col = cols // 2
        max_radius = max(1, max(rows, cols) // 2 + 1)
        for radius in range(max_radius):
            for row_idx in range(center_row - radius, center_row + radius + 1):
                for col_idx in range(center_col - radius, center_col + radius + 1):
                    if 0 <= row_idx < rows and 0 <= col_idx < cols:
                        if abs(row_idx - center_row) + abs(col_idx - center_col) <= radius + 1:
                            positions.append((row_idx, col_idx))
        return positions

    if pattern == 'Clustered':
        for col_idx in range(cols):
            start_row = 0 if col_idx % 2 == 0 else 1
            for row_idx in range(start_row, rows, 2):
                positions.append((row_idx, col_idx))
        return positions

    if pattern == 'Mixed':
        for col_idx in range(cols):
            row_order = list(range(rows)) if col_idx % 2 == 0 else list(reversed(range(rows)))
            for row_idx in row_order:
                positions.append((row_idx, col_idx))
        return positions

    for col_idx in range(cols):
        row_order = list(range(rows)) if col_idx % 2 == 0 else list(reversed(range(rows)))
        for row_idx in row_order:
            positions.append((row_idx, col_idx))
    return positions


class SeatingArrangementExcelView(APIView):
    """Generate and download seating arrangement Excel workbook."""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        Expected JSON payload:
        {
            "exam_title": "CIA-1 August 2026",
            "semester_text": "Final year/ Seventh semester",
            "date_str": "29.11.2025",
            "session": "FN",
            "halls": [
                {
                    "hall_name": "LH-C104",
                    "rows": 7,
                    "cols": 6,
                    "students": [
                        "23038117243210XX",
                        ...
                    ]
                },
                ...
            ]
        }
        """
        try:
            # Handle both request.data (DRF) and request.body (standard Django)
            if hasattr(request, 'data') and isinstance(request.data, dict):
                data = request.data
            else:
                try:
                    data = json.loads(request.body.decode('utf-8'))
                except (json.JSONDecodeError, AttributeError):
                    return Response({'error': 'Invalid JSON payload'}, status=status.HTTP_400_BAD_REQUEST)

            exam_title = data.get('exam_title', 'Exam')
            semester_text = data.get('semester_text', 'Semester')
            date_str = data.get('date_str', '')
            session = data.get('session', 'FN')
            halls = data.get('halls', [])

            if not halls:
                return Response({'error': 'No hall data provided'}, status=status.HTTP_400_BAD_REQUEST)

            # Filter out empty halls and deduplicate student numbers before generating the workbook
            cleaned_halls = []
            for hall in halls:
                students = hall.get('students', []) or []
                deduped_students = []
                seen = set()
                for student in students:
                    reg_no = str(student).strip()
                    if not reg_no or reg_no in seen:
                        continue
                    seen.add(reg_no)
                    deduped_students.append(reg_no)
                if not deduped_students:
                    continue
                hall_copy = dict(hall)
                hall_copy['students'] = deduped_students
                cleaned_halls.append(hall_copy)

            if not cleaned_halls:
                return Response({'error': 'No hall contains students to export'}, status=status.HTTP_400_BAD_REQUEST)

            # Generate workbook
            wb = self._build_workbook(exam_title, semester_text, date_str, session, cleaned_halls)

            # Save to BytesIO buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer_bytes = buffer.getvalue()

            filename = f'seating_arrangement_{date_str.replace("/", "-")}.xlsx'
            response = HttpResponse(
                buffer_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            import traceback
            error_msg = f"{str(e)} | {traceback.format_exc()}"
            return Response({'error': error_msg}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _build_workbook(self, exam_title: str, semester_text: str, date_str: str, session: str, halls: list) -> Workbook:
        """Build the seating arrangement workbook using a proper row/column seat grid."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Seating Arrangement"

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        title_font = Font(name="Times New Roman", size=14, bold=True, italic=True)
        section_font = Font(name="Times New Roman", size=11, bold=True)
        body_font = Font(name="Times New Roman", size=10)
        header_font = Font(name="Arial", size=10, bold=True)

        current_row = 1

        for hall_index, hall in enumerate(halls):
            hall_name = hall.get('hall_name', f'Hall {hall_index + 1}')
            rows = max(1, int(hall.get('rows', 1)))
            cols = max(1, int(hall.get('cols', 1)))
            students = list(hall.get('students', []))
            pattern = hall.get('pattern', 'Zigzag')
            total_students = len(students)

            title_row = current_row
            ws.cell(row=title_row, column=1, value=f"Office of the Controller of Examinations\nSeating Arrangement for {exam_title}\n{semester_text}")
            ws.merge_cells(start_row=title_row, end_row=title_row, start_column=1, end_column=max(cols + 4, 8))
            title_cell = ws.cell(row=title_row, column=1)
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            title_cell.fill = header_fill
            title_cell.border = border
            ws.row_dimensions[title_row].height = 42

            meta_row = title_row + 1
            ws.cell(row=meta_row, column=1, value="Date / Session")
            ws.cell(row=meta_row, column=2, value=f"{date_str} / {session}")
            ws.cell(row=meta_row, column=4, value="Total")
            ws.cell(row=meta_row, column=5, value=total_students)
            ws.cell(row=meta_row, column=7, value="Hall")
            ws.cell(row=meta_row, column=8, value=hall_name)
            for col in (1, 2, 4, 5, 7, 8):
                cell = ws.cell(row=meta_row, column=col)
                cell.font = section_font
                cell.border = border
                cell.fill = header_fill

            pattern_row = meta_row + 1
            ws.cell(row=pattern_row, column=1, value="Pattern")
            ws.cell(row=pattern_row, column=2, value=str(pattern))
            for col in (1, 2):
                cell = ws.cell(row=pattern_row, column=col)
                cell.font = section_font
                cell.border = border
                cell.fill = header_fill

            header_row = pattern_row + 2
            ws.cell(row=header_row, column=1, value="Row / Seat")
            for col_idx in range(cols):
                ws.cell(row=header_row, column=col_idx + 2, value=f"Seat {col_idx + 1}")
            for col in range(1, cols + 3):
                cell = ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.border = border
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            seat_positions = seat_positions_for_pattern(rows, cols, pattern)
            seat_grid = [["" for _ in range(cols)] for _ in range(rows)]
            for index, (seat_row, seat_col) in enumerate(seat_positions[:len(students)]):
                if 0 <= seat_row < rows and 0 <= seat_col < cols:
                    seat_grid[seat_row][seat_col] = str(students[index])

            for row_idx in range(rows):
                data_row = header_row + 1 + row_idx
                ws.cell(row=data_row, column=1, value=f"Row {row_idx + 1}")
                for col_idx in range(cols):
                    cell = ws.cell(row=data_row, column=col_idx + 2, value=seat_grid[row_idx][col_idx])
                    cell.border = border
                    cell.font = body_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.column_dimensions['A'].width = 14
            for col_idx in range(cols):
                ws.column_dimensions[get_column_letter(col_idx + 2)].width = 22

            current_row = header_row + rows + 3

        return wb
