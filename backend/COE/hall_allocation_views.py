import json
from io import BytesIO

from django.http import HttpResponse
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def seat_positions_for_pattern(rows, cols, pattern):
    """Return a list of (row, col) positions in the order this pattern fills seats."""
    positions = []
    seen = set()

    def add_pos(r, c):
        if 0 <= r < rows and 0 <= c < cols and (r, c) not in seen:
            seen.add((r, c))
            positions.append((r, c))

    if pattern == 'Straight':
        for r in range(rows):
            for c in range(cols):
                add_pos(r, c)

    elif pattern in (None, '', 'Zigzag'):
        for c in range(cols):
            row_order = list(range(rows)) if c % 2 == 0 else list(reversed(range(rows)))
            for r in row_order:
                add_pos(r, c)

    elif pattern == 'Alternate Zigzag':
        for pair_start in range(0, cols, 2):
            pair_cols = list(range(pair_start, min(pair_start + 2, cols)))
            if len(pair_cols) == 2:
                c1, c2 = pair_cols[0], pair_cols[1]
                for r in range(rows):
                    if r % 2 == 0:
                        add_pos(r, c1)
                        add_pos(r, c2)
                    else:
                        add_pos(r, c2)
                        add_pos(r, c1)
            else:
                c1 = pair_cols[0]
                for r in range(rows):
                    add_pos(r, c1)

    elif pattern == 'U-Shape':
        top, bottom, left, right = 0, rows - 1, 0, cols - 1
        while top <= bottom and left <= right:
            for r in range(top, bottom + 1):
                add_pos(r, left)
            for c in range(left + 1, right + 1):
                add_pos(bottom, c)
            if left < right:
                for r in range(bottom - 1, top - 1, -1):
                    add_pos(r, right)
            if top < bottom:
                for c in range(right - 1, left, -1):
                    add_pos(top, c)
            top += 1
            bottom -= 1
            left += 1
            right -= 1

    elif pattern == 'Circle':
        center_r = (rows - 1) / 2.0
        center_c = (cols - 1) / 2.0
        cells = []
        for r in range(rows):
            for c in range(cols):
                dist = (r - center_r) ** 2 + (c - center_c) ** 2
                cells.append((dist, r, c))
        cells.sort(key=lambda item: (item[0], item[1], item[2]))
        for _, r, c in cells:
            add_pos(r, c)

    elif pattern == 'Clustered':
        for block_r in range(0, rows, 2):
            for block_c in range(0, cols, 2):
                for r in range(block_r, min(block_r + 2, rows)):
                    for c in range(block_c, min(block_c + 2, cols)):
                        add_pos(r, c)

    elif pattern == 'Mixed':
        for c in range(cols):
            if c % 4 == 0:
                for r in range(rows):
                    add_pos(r, c)
            elif c % 4 == 1:
                for r in range(0, rows, 2):
                    add_pos(r, c)
                for r in range(1, rows, 2):
                    add_pos(r, c)
            elif c % 4 == 2:
                for r in range(rows - 1, -1, -1):
                    add_pos(r, c)
            else:
                for r in range(1, rows, 2):
                    add_pos(r, c)
                for r in range(0, rows, 2):
                    add_pos(r, c)

    # Fallback to guarantee 100% seat coverage
    for r in range(rows):
        for c in range(cols):
            add_pos(r, c)

    return positions


class SeatingArrangementExcelView(APIView):
    """Generate and download seating arrangement Excel workbook."""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        Expected JSON payload:
        {
            "exam_title": "CIA-1 August 2026",
            "semester_text": "Final year/ Seventh semester",
            "date_str": "29.11.2025",
            "session": "FN",
            "halls": [
                {
                    "hall_name": "LH-C104",
                    "rows": 7,
                    "cols": 6,
                    "students": [
                        "23038117243210XX",
                        ...
                    ]
                },
                ...
            ]
        }
        """
        try:
            # Handle both request.data (DRF) and request.body (standard Django)
            if hasattr(request, 'data') and isinstance(request.data, dict):
                data = request.data
            else:
                try:
                    data = json.loads(request.body.decode('utf-8'))
                except (json.JSONDecodeError, AttributeError):
                    return Response({'error': 'Invalid JSON payload'}, status=status.HTTP_400_BAD_REQUEST)

            exam_title = data.get('exam_title', 'Exam')
            semester_text = data.get('semester_text', 'Semester')
            date_str = data.get('date_str', '')
            session = data.get('session', 'FN')
            halls = data.get('halls', [])

            if not halls:
                return Response({'error': 'No hall data provided'}, status=status.HTTP_400_BAD_REQUEST)

            # Filter out empty halls while preserving empty seat placeholders for position grid alignment
            cleaned_halls = []
            for hall in halls:
                students = hall.get('students', []) or []
                actual_students = [str(s).strip() for s in students if str(s).strip()]
                if not actual_students:
                    continue
                cleaned_halls.append(hall)

            if not cleaned_halls:
                return Response({'error': 'No hall contains students to export'}, status=status.HTTP_400_BAD_REQUEST)

            # Generate workbook
            wb = self._build_workbook(exam_title, semester_text, date_str, session, cleaned_halls)

            # Save to BytesIO buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer_bytes = buffer.getvalue()

            filename = f'seating_arrangement_{date_str.replace("/", "-")}.xlsx'
            response = HttpResponse(
                buffer_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            import traceback
            error_msg = f"{str(e)} | {traceback.format_exc()}"
            return Response({'error': error_msg}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _build_workbook(self, exam_title: str, semester_text: str, date_str: str, session: str, halls: list) -> Workbook:
        """Build seating arrangement workbook with Sheet 1 (Master Allocation Summary) and Sheet 2 (Seating Arrangement)."""
        wb = Workbook()

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Color Fills
        green_fill = PatternFill(fill_type="solid", fgColor="549938")
        light_green_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
        header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")

        # Fonts
        banner_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        title_font = Font(name="Times New Roman", size=14, bold=True, italic=True)
        date_font = Font(name="Times New Roman", size=11, bold=True)
        summary_font = Font(name="Times New Roman", size=11, bold=True)
        body_font = Font(name="Times New Roman", size=10)
        header_font = Font(name="Arial", size=11, bold=True)
        dept_summary_font = Font(name="Arial", size=11, bold=True)

        col_width_sno = 6
        col_width_reg = 18
        spacer_width = 0.5

        # =========================================================================
        # SHEET 1: Master Allocation Summary Table
        # =========================================================================
        ws_summary = wb.active
        ws_summary.title = "Master Allocation Summary"

        # Pre-cache department codes for fast register number prefix lookup
        from academics.models import Department, StudentProfile
        dept_mapping = {d.code: (d.short_name or d.code) for d in Department.objects.all()}
        dept_cache = {}

        def get_student_dept(reg, cache, mapping):
            reg = str(reg).strip()
            if not reg:
                return None
            if reg in cache:
                return cache[reg]
            if len(reg) >= 11:
                code_slice = reg[8:11]
                if code_slice in mapping:
                    cache[reg] = mapping[code_slice]
                    return mapping[code_slice]
            # Fallback to StudentProfile lookup
            sp = StudentProfile.objects.filter(reg_no=reg).select_related('home_department').first()
            if sp and sp.home_department:
                val = sp.home_department.short_name or sp.home_department.code
                cache[reg] = val
                return val
            return None

        def resolve_hall_dept_counts(hall):
            """
            Returns a clean dictionary of {dept_name: count} for all non-empty students in this hall.
            Always computes directly from individual student records to guarantee 100% accuracy before and after optimization.
            """
            students = hall.get('students', [])
            student_depts = hall.get('student_depts') or hall.get('studentDepts') or []
            dept_name = hall.get('department', '')

            counts = {}
            for i, s in enumerate(students):
                s_clean = str(s).strip()
                if not s_clean:
                    continue
                # 1. Check if student register number directly identifies department (Anna Univ code)
                d = get_student_dept(s_clean, dept_cache, dept_mapping)
                
                # 2. If not found by register number, check student_depts list
                if not d and i < len(student_depts):
                    cand = str(student_depts[i]).strip()
                    if cand and " / " not in cand:
                        d = cand
                
                # 3. Fallback to single department if hall only has one dept
                if not d and dept_name and " / " not in dept_name:
                    d = dept_name

                if d:
                    counts[d] = counts.get(d, 0) + 1
                else:
                    counts['Other'] = counts.get('Other', 0) + 1

            # Fallback if counts was empty (e.g. dummy test data without dept codes)
            if not counts:
                valid_count = len([s for s in students if str(s).strip()])
                raw_depts = [d.strip() for d in (hall.get('departments') or dept_name.split(' / ')) if d.strip()]
                if not raw_depts:
                    raw_depts = [dept_name] if dept_name else ['General']
                if len(raw_depts) == 1:
                    counts[raw_depts[0]] = valid_count
                elif valid_count > 0:
                    base = valid_count // len(raw_depts)
                    rem = valid_count % len(raw_depts)
                    for idx_d, d_name in enumerate(raw_depts):
                        counts[d_name] = base + (1 if idx_d < rem else 0)

            return counts

        # Calculate Department breakdown per hall dynamically (Pure Dept-wise)
        dept_halls_map = {}
        dept_counts_map = {}

        for hall in halls:
            hall_name = hall.get('hall_name', 'Hall')
            hall_counts = resolve_hall_dept_counts(hall)

            for dept, count in hall_counts.items():
                if count > 0:
                    if dept not in dept_halls_map:
                        dept_halls_map[dept] = []
                        dept_counts_map[dept] = []
                    dept_halls_map[dept].append(hall_name)
                    dept_counts_map[dept].append(count)

        # Banner Row
        banner_text = f"{date_str} {session}   |   {exam_title} Hall Allocation   |   {semester_text}"
        max_summary_cols = max(5, max([len(halls_list) for halls_list in dept_halls_map.values()] + [1]) + 2)
        ws_summary.merge_cells(start_row=1, end_row=1, start_column=1, end_column=max_summary_cols)
        banner_cell = ws_summary.cell(row=1, column=1, value=banner_text)
        banner_cell.font = banner_font
        banner_cell.fill = green_fill
        banner_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.row_dimensions[1].height = 30

        # Table Header Row
        ws_summary.cell(row=2, column=1, value="Dept").font = dept_summary_font
        ws_summary.cell(row=2, column=1).fill = light_green_fill
        ws_summary.cell(row=2, column=1).border = border
        ws_summary.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

        ws_summary.cell(row=2, column=2, value="Count Breakdown").font = dept_summary_font
        ws_summary.cell(row=2, column=2).fill = light_green_fill
        ws_summary.cell(row=2, column=2).border = border
        ws_summary.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="center")

        for col_idx in range(3, max_summary_cols + 1):
            cell = ws_summary.cell(row=2, column=col_idx, value=f"Hall {col_idx - 2}")
            cell.font = dept_summary_font
            cell.fill = light_green_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_summary.row_dimensions[2].height = 24
        ws_summary.column_dimensions['A'].width = 16
        ws_summary.column_dimensions['B'].width = 22

        summary_row_idx = 3
        for dept in sorted(dept_counts_map.keys()):
            counts = dept_counts_map[dept]
            hall_list = dept_halls_map[dept]
            breakdown_str = " + ".join(map(str, counts))
            if len(counts) > 1:
                breakdown_str += f" = {sum(counts)}"

            c1 = ws_summary.cell(row=summary_row_idx, column=1, value=dept)
            c1.font = dept_summary_font
            c1.border = border
            c1.alignment = Alignment(horizontal="left", vertical="center")

            c2 = ws_summary.cell(row=summary_row_idx, column=2, value=breakdown_str)
            c2.font = body_font
            c2.border = border
            c2.alignment = Alignment(horizontal="center", vertical="center")

            for idx, h_name in enumerate(hall_list):
                c_hall = ws_summary.cell(row=summary_row_idx, column=3 + idx, value=h_name)
                c_hall.font = body_font
                c_hall.border = border
                c_hall.alignment = Alignment(horizontal="center", vertical="center")
                ws_summary.column_dimensions[get_column_letter(3 + idx)].width = 16

            ws_summary.row_dimensions[summary_row_idx].height = 20
            summary_row_idx += 1

        # =========================================================================
        # SHEET 2: Seating Arrangement (Individual Halls)
        # =========================================================================
        ws = wb.create_sheet(title="Seating Arrangement")
        current_row = 1

        all_depts = sorted(list(dept_halls_map.keys()))
        global_dept_label = " / ".join(all_depts) if all_depts else "Department"

        def apply_cell_style(cell, border_style=border, fill_style=None, font_style=None, align_style=None):
            if border_style:
                cell.border = border_style
            if fill_style:
                cell.fill = fill_style
            if font_style:
                cell.font = font_style
            if align_style:
                cell.alignment = align_style

        def style_merge_range(sheet, r_start, r_end, c_start, c_end, border_style=border, fill_style=None):
            for r in range(r_start, r_end + 1):
                for c in range(c_start, c_end + 1):
                    apply_cell_style(sheet.cell(row=r, column=c), border_style=border_style, fill_style=fill_style)

        for hall in halls:
            hall_name = hall.get('hall_name', 'Hall')
            rows = int(hall.get('rows', 1))
            cols = int(hall.get('cols', 1))
            students = hall.get('students', [])
            pattern = hall.get('pattern', 'Zigzag')
            dept_name = hall.get('department', global_dept_label)

            # Department split for this specific hall (always computed directly from seated students)
            counts = resolve_hall_dept_counts(hall)
            total_students = sum(counts.values()) if counts else len([s for s in students if str(s).strip()])
            hall_dept_label = " / ".join(sorted(counts.keys())) if counts else dept_name

            if counts:
                breakdown_parts = [f"{d}: {c}" for d, c in sorted(counts.items()) if c > 0]
                summary_value = f"{' | '.join(breakdown_parts)} | TOTAL: {total_students}"
            else:
                summary_value = f"{dept_name}  TOTAL={total_students}"

            # Each seat column consists of exactly 2 Excel columns: S.No + Reg No (no spacer columns)
            total_cols = max(cols * 2, 6)

            # 1. Title row
            title_row = current_row
            title_value = f"Office of the Controller of Examinations\nSeating Arrangement for {exam_title}\n{semester_text}"
            ws.merge_cells(start_row=title_row, end_row=title_row, start_column=1, end_column=total_cols)
            style_merge_range(ws, title_row, title_row, 1, total_cols, border_style=border, fill_style=header_fill)
            
            title_cell = ws.cell(row=title_row, column=1, value=title_value)
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[title_row].height = 45

            # 2. Metadata row
            meta_row = title_row + 1

            if total_cols >= 8:
                # Cols 1-2: Date/Session label
                ws.merge_cells(start_row=meta_row, end_row=meta_row, start_column=1, end_column=2)
                style_merge_range(ws, meta_row, meta_row, 1, 2, fill_style=header_fill)
                cell_d1 = ws.cell(row=meta_row, column=1, value="Date/Session:")
                cell_d1.font = date_font
                cell_d1.alignment = Alignment(horizontal="left", vertical="center")

                # Cols 3-4: Date/Session value
                ws.merge_cells(start_row=meta_row, end_row=meta_row, start_column=3, end_column=4)
                style_merge_range(ws, meta_row, meta_row, 3, 4, fill_style=header_fill)
                cell_d2 = ws.cell(row=meta_row, column=3, value=f"{date_str} {session}")
                cell_d2.font = date_font
                cell_d2.alignment = Alignment(horizontal="left", vertical="center")

                # Cols 5 to (total_cols - 2): Dept & Total
                summary_start = 5
                summary_end = total_cols - 2
                ws.merge_cells(start_row=meta_row, end_row=meta_row, start_column=summary_start, end_column=summary_end)
                style_merge_range(ws, meta_row, meta_row, summary_start, summary_end, fill_style=header_fill)
                summary_cell = ws.cell(row=meta_row, column=summary_start, value=summary_value)
                summary_cell.font = summary_font
                summary_cell.alignment = Alignment(horizontal="center", vertical="center")

                # Cols (total_cols - 1) to total_cols: Hall Number (in the rightmost space)
                hall_start = total_cols - 1
                hall_end = total_cols
                ws.merge_cells(start_row=meta_row, end_row=meta_row, start_column=hall_start, end_column=hall_end)
                style_merge_range(ws, meta_row, meta_row, hall_start, hall_end, fill_style=header_fill)
                hall_cell = ws.cell(row=meta_row, column=hall_start, value=hall_name)
                hall_cell.font = Font(name="Times New Roman", size=13, bold=True)
                hall_cell.alignment = Alignment(horizontal="center", vertical="center")

            else:
                # Cols 1-2: Date/Session
                ws.merge_cells(start_row=meta_row, end_row=meta_row, start_column=1, end_column=2)
                style_merge_range(ws, meta_row, meta_row, 1, 2, fill_style=header_fill)
                cell_d = ws.cell(row=meta_row, column=1, value=f"{date_str} {session}")
                cell_d.font = date_font
                cell_d.alignment = Alignment(horizontal="left", vertical="center")

                # Cols 3-4: Dept & Total
                ws.merge_cells(start_row=meta_row, end_row=meta_row, start_column=3, end_column=4)
                style_merge_range(ws, meta_row, meta_row, 3, 4, fill_style=header_fill)
                summary_cell = ws.cell(row=meta_row, column=3, value=summary_value)
                summary_cell.font = summary_font
                summary_cell.alignment = Alignment(horizontal="center", vertical="center")

                # Cols 5-6: Hall Number
                ws.merge_cells(start_row=meta_row, end_row=meta_row, start_column=5, end_column=6)
                style_merge_range(ws, meta_row, meta_row, 5, 6, fill_style=header_fill)
                hall_cell = ws.cell(row=meta_row, column=5, value=hall_name)
                hall_cell.font = Font(name="Times New Roman", size=13, bold=True)
                hall_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[meta_row].height = 25

            # 3. Header row with S.No and department labels (No spacer columns)
            header_row = meta_row + 1

            for c in range(cols):
                sno_col = c * 2 + 1
                reg_col = c * 2 + 2

                sno_header = ws.cell(row=header_row, column=sno_col, value="S.No")
                reg_header = ws.cell(row=header_row, column=reg_col, value=hall_dept_label)

                for cell in (sno_header, reg_header):
                    apply_cell_style(cell, border_style=border, fill_style=header_fill, font_style=header_font, align_style=Alignment(horizontal="center", vertical="center"))

                ws.column_dimensions[get_column_letter(sno_col)].width = 7
                ws.column_dimensions[get_column_letter(reg_col)].width = 20

            ws.row_dimensions[header_row].height = 20

            # 4. Data rows with student seating grid (Strictly adjacent columns)
            seat_positions = seat_positions_for_pattern(rows, cols, pattern)
            seat_grid = [["" for _ in range(cols)] for _ in range(rows)]
            role_grid = [["" for _ in range(cols)] for _ in range(rows)]
            donor_set = set(hall.get('donor_indices', []) or [])
            shifted_set = set(hall.get('shifted_indices', []) or [])

            donor_fill = PatternFill(fill_type="solid", fgColor="FEF08A")
            shifted_fill = PatternFill(fill_type="solid", fgColor="99F6E4")
            donor_font = Font(name="Times New Roman", size=10, bold=True, color="854D0E")
            shifted_font = Font(name="Times New Roman", size=10, bold=True, color="115E59")

            for index, (row_idx, col_idx) in enumerate(seat_positions[:len(students)]):
                if 0 <= row_idx < rows and 0 <= col_idx < cols:
                    seat_grid[row_idx][col_idx] = str(students[index] or '').strip()
                    if index in donor_set:
                        role_grid[row_idx][col_idx] = "donor"
                    elif index in shifted_set:
                        role_grid[row_idx][col_idx] = "shifted"

            data_row = header_row
            for row_idx in range(rows):
                data_row = header_row + 1 + row_idx
                for c in range(cols):
                    sno_col = c * 2 + 1
                    reg_col = c * 2 + 2

                    seat_label = f"{chr(65 + c)}{row_idx + 1}"
                    student_id = seat_grid[row_idx][c] if row_idx < rows and c < cols else ""
                    role = role_grid[row_idx][c] if row_idx < rows and c < cols else ""

                    sno_cell = ws.cell(row=data_row, column=sno_col, value=seat_label)
                    reg_cell = ws.cell(row=data_row, column=reg_col, value=student_id)

                    apply_cell_style(sno_cell, border_style=border, font_style=body_font, align_style=Alignment(horizontal="center", vertical="center"))
                    apply_cell_style(reg_cell, border_style=border, font_style=body_font, align_style=Alignment(horizontal="center", vertical="center"))

                    if role == "donor":
                        reg_cell.fill = donor_fill
                        reg_cell.font = donor_font
                    elif role == "shifted":
                        reg_cell.fill = shifted_fill
                        reg_cell.font = shifted_font

                ws.row_dimensions[data_row].height = 18

            # Add spacing between halls
            current_row = data_row + 3 if rows > 0 else meta_row + 3

        return wb


class PublishHallPlanView(APIView):
    """
    Publish seating plan so students can view their allocated hall & seat number in their IDCS portal.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import datetime
        from .models import CoeKeyValueStore

        data = request.data if isinstance(request.data, dict) else {}
        if not data and hasattr(request, 'body'):
            try:
                data = json.loads(request.body.decode('utf-8'))
            except Exception:
                data = {}

        exam_title = str(data.get('exam_title') or '').strip() or 'Exam'
        semester_text = str(data.get('semester_text') or '').strip() or 'Semester'
        date_str = str(data.get('date_str') or '').strip()
        session = str(data.get('session') or 'FN').strip()
        halls = data.get('halls', []) or []

        if not halls:
            return Response({'error': 'No hall data provided to publish.'}, status=status.HTTP_400_BAD_REQUEST)

        now_iso = datetime.datetime.now().isoformat()
        student_allocations = {}
        total_published = 0

        for hall in halls:
            hall_name = str(hall.get('hall_name') or 'Hall').strip()
            rows = int(hall.get('rows', 1))
            cols = int(hall.get('cols', 1))
            pattern = hall.get('pattern', 'Zigzag')
            dept = str(hall.get('department') or '').strip()
            students = hall.get('students', []) or []

            positions = seat_positions_for_pattern(rows, cols, pattern)

            for index, (r, c) in enumerate(positions[:len(students)]):
                if index < len(students):
                    s_id = str(students[index] or '').strip()
                    if s_id:
                        seat_label = f"{chr(65 + c)}{r + 1}"
                        normalized_reg = s_id.upper().strip()
                        alloc_record = {
                            'reg_no': s_id,
                            'hall_name': hall_name,
                            'seat_label': seat_label,
                            'row_number': r + 1,
                            'column_letter': chr(65 + c),
                            'department': dept,
                            'exam_title': exam_title,
                            'semester_text': semester_text,
                            'exam_date': date_str,
                            'session': session,
                            'published_at': now_iso,
                        }
                        student_allocations[normalized_reg] = alloc_record
                        total_published += 1

        # Fetch existing published store or initialize
        try:
            store_obj, _ = CoeKeyValueStore.objects.get_or_create(
                store_name='coe_published_hall_plans',
                defaults={'data': {}}
            )
            existing_data = store_obj.data if isinstance(store_obj.data, dict) else {}
        except Exception:
            existing_data = {}

        # Merge or update
        existing_data.update(student_allocations)

        CoeKeyValueStore.objects.update_or_create(
            store_name='coe_published_hall_plans',
            defaults={'data': existing_data}
        )

        return Response({
            'success': True,
            'published_count': total_published,
            'message': f'Successfully published hall plan for {total_published} students.',
            'exam_title': exam_title,
            'date_str': date_str,
            'session': session,
        }, status=status.HTTP_200_OK)


class StudentHallPlanView(APIView):
    """
    Fetch the published hall plan and seating allocation for the logged-in student.
    Returns ONLY this student's allocated hall and seat position.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from .models import CoeKeyValueStore
        from academics.models import StudentProfile

        user = request.user
        student_reg_no = ''

        # 1. Look up student profile reg_no
        try:
            sp = StudentProfile.objects.filter(user=user).first()
            if sp and sp.reg_no:
                student_reg_no = str(sp.reg_no).strip()
        except Exception:
            pass

        # 2. Fallback to username if not found in StudentProfile
        if not student_reg_no:
            student_reg_no = str(user.username or '').strip()

        if not student_reg_no:
            return Response({'published': False, 'allocations': [], 'message': 'Student register number not found.'})

        # 3. Retrieve published plans store
        try:
            store_obj = CoeKeyValueStore.objects.filter(store_name='coe_published_hall_plans').first()
            published_data = store_obj.data if store_obj and isinstance(store_obj.data, dict) else {}
        except Exception:
            published_data = {}

        normalized_reg = student_reg_no.upper().strip()
        matched_alloc = published_data.get(normalized_reg)

        # In case it's stored under raw or partial match
        if not matched_alloc:
            for k, v in published_data.items():
                if str(k).upper().strip() == normalized_reg:
                    matched_alloc = v
                    break

        allocations = [matched_alloc] if matched_alloc else []

        return Response({
            'published': len(allocations) > 0,
            'reg_no': student_reg_no,
            'student_name': user.get_full_name() or user.username,
            'allocations': allocations,
        }, status=status.HTTP_200_OK)


