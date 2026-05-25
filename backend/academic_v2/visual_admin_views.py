"""
Visual Admin API Views
Manages Power BI URL assignments for staff and exposes them to faculties.
"""

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Count, Q

from .models import VisualAdminStaffLink, VisualAdminCourseLink
from academics.models import StaffProfile, TeachingAssignment, AcademicYear


def _is_visual_admin(user):
    return user.roles.filter(name='VISUAL_ADMIN').exists() or user.is_superuser


# ---------------------------------------------------------------------------
# Staff search + link list (for Visual Admin URLs page)
# ---------------------------------------------------------------------------

class VisualAdminStaffListView(APIView):
    """
    GET  /api/academic-v2/visual-admin/staff/?q=<search>
    Returns staff list with their current link configuration.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not _is_visual_admin(request.user):
            return Response({'detail': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

        q = request.query_params.get('q', '').strip()
        qs = StaffProfile.objects.filter(status='ACTIVE').select_related('user', 'department')
        if q:
            qs = qs.filter(
                Q(user__first_name__icontains=q) |
                Q(user__last_name__icontains=q) |
                Q(user__username__icontains=q) |
                Q(staff_id__icontains=q)
            )
        qs = qs.order_by('user__first_name', 'user__last_name')[:100]

        result = []
        for sp in qs:
            try:
                link = sp.visual_admin_link
                overall_url = link.overall_url
                use_course_urls = link.use_course_urls
            except VisualAdminStaffLink.DoesNotExist:
                overall_url = ''
                use_course_urls = False

            result.append({
                'staff_id': sp.staff_id,
                'internal_id': sp.internal_id,
                'name': sp.user.get_full_name() or sp.user.username,
                'email': sp.user.email,
                'department': sp.department.name if sp.department else '',
                'designation': sp.designation,
                'profile_image': sp.user.profile_image or '',
                'overall_url': overall_url,
                'use_course_urls': use_course_urls,
            })
        return Response(result)


# ---------------------------------------------------------------------------
# Per-staff link management
# ---------------------------------------------------------------------------

class VisualAdminStaffLinkView(APIView):
    """
    GET    /api/academic-v2/visual-admin/staff/<staff_id>/
    PUT    /api/academic-v2/visual-admin/staff/<staff_id>/
           Body: { overall_url, use_course_urls, course_links: [{ta_id, url}] }
    """
    permission_classes = [IsAuthenticated]

    def _get_staff(self, staff_id):
        try:
            return StaffProfile.objects.select_related('user', 'department').get(staff_id=staff_id)
        except StaffProfile.DoesNotExist:
            return None

    def get(self, request, staff_id):
        if not _is_visual_admin(request.user):
            return Response({'detail': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

        sp = self._get_staff(staff_id)
        if not sp:
            return Response({'detail': 'Staff not found'}, status=status.HTTP_404_NOT_FOUND)

        try:
            link = sp.visual_admin_link
            use_course_urls = link.use_course_urls
            overall_url = link.overall_url
            course_links_qs = link.course_links.select_related(
                'teaching_assignment__subject',
                'teaching_assignment__curriculum_row__course',
                'teaching_assignment__section',
                'teaching_assignment__academic_year',
            ).all()
        except VisualAdminStaffLink.DoesNotExist:
            use_course_urls = False
            overall_url = ''
            course_links_qs = VisualAdminCourseLink.objects.none()

        # Get all active teaching assignments for this staff
        active_year = AcademicYear.objects.filter(is_current=True).first()
        ta_qs = TeachingAssignment.objects.filter(
            staff=sp,
        ).select_related(
            'subject',
            'curriculum_row__course',
            'curriculum_row__department',
            'section__department',
            'academic_year',
            'elective_subject',
        )
        if active_year:
            ta_qs = ta_qs.filter(academic_year=active_year)

        course_link_map = {cl.teaching_assignment_id: cl.url for cl in course_links_qs}

        courses = []
        for ta in ta_qs:
            course_code = ''
            course_name = ''
            if ta.curriculum_row and ta.curriculum_row.course:
                course_code = ta.curriculum_row.course.code or ''
                course_name = ta.curriculum_row.course.name or ''
            elif ta.subject:
                course_code = ta.subject.code or ''
                course_name = ta.subject.name or ''
            elif ta.elective_subject:
                course_code = ta.elective_subject.course_code or ''
                course_name = ta.elective_subject.course_name or ''
            elif ta.custom_subject:
                course_code = ta.custom_subject
                course_name = ta.get_custom_subject_display()

            dept_name = ''
            if ta.curriculum_row and ta.curriculum_row.department:
                dept_name = ta.curriculum_row.department.name or ''
            elif ta.section and ta.section.department:
                dept_name = ta.section.department.name or ''

            courses.append({
                'ta_id': ta.id,
                'course_code': course_code,
                'course_name': course_name,
                'section': ta.section.name if ta.section else '',
                'department': dept_name,
                'academic_year': ta.academic_year.year_label if hasattr(ta.academic_year, 'year_label') else str(ta.academic_year),
                'url': course_link_map.get(ta.id, ''),
            })

        return Response({
            'staff_id': sp.staff_id,
            'name': sp.user.get_full_name() or sp.user.username,
            'overall_url': overall_url,
            'use_course_urls': use_course_urls,
            'courses': courses,
        })

    def put(self, request, staff_id):
        if not _is_visual_admin(request.user):
            return Response({'detail': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

        sp = self._get_staff(staff_id)
        if not sp:
            return Response({'detail': 'Staff not found'}, status=status.HTTP_404_NOT_FOUND)

        data = request.data
        overall_url = str(data.get('overall_url', '') or '')
        use_course_urls = bool(data.get('use_course_urls', False))
        course_links_data = data.get('course_links', [])

        link, _ = VisualAdminStaffLink.objects.get_or_create(
            staff=sp,
            defaults={'updated_by': request.user},
        )
        link.overall_url = overall_url
        link.use_course_urls = use_course_urls
        link.updated_by = request.user
        link.save()

        # Update course links
        for cl_data in course_links_data:
            ta_id = cl_data.get('ta_id')
            url = str(cl_data.get('url', '') or '')
            if not ta_id:
                continue
            try:
                ta = TeachingAssignment.objects.get(id=ta_id, staff=sp)
            except TeachingAssignment.DoesNotExist:
                continue
            VisualAdminCourseLink.objects.update_or_create(
                staff_link=link,
                teaching_assignment=ta,
                defaults={'url': url},
            )

        return Response({'detail': 'Saved successfully'})


# ---------------------------------------------------------------------------
# Visual Admin users list (for "Contact the Incharge" popup)
# ---------------------------------------------------------------------------

class VisualAdminUsersView(APIView):
    """
    GET /api/academic-v2/visual-admin/users/
    Returns all users with VISUAL_ADMIN role.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from accounts.models import Role, UserRole
        try:
            role = Role.objects.get(name='VISUAL_ADMIN')
        except Role.DoesNotExist:
            return Response([])

        user_roles = UserRole.objects.filter(role=role).select_related(
            'user',
        ).prefetch_related('user__staff_profile__department')

        result = []
        for ur in user_roles:
            u = ur.user
            roles_list = list(u.roles.values_list('name', flat=True))
            dept = ''
            designation = ''
            profile_image = u.profile_image or ''
            try:
                sp = u.staff_profile
                dept = sp.department.name if sp.department else ''
                designation = sp.designation or ''
                if not profile_image and sp.profile_image:
                    profile_image = sp.profile_image.url if hasattr(sp.profile_image, 'url') else str(sp.profile_image)
            except Exception:
                pass
            result.append({
                'id': u.id,
                'name': u.get_full_name() or u.username,
                'email': u.email,
                'profile_image': profile_image,
                'roles': roles_list,
                'department': dept,
                'designation': designation,
            })
        return Response(result)


# ---------------------------------------------------------------------------
# Faculty-facing: get own Power BI URL for a course (by TA id)
# ---------------------------------------------------------------------------

class FacultyVisualAdminLinkView(APIView):
    """
    GET /api/academic-v2/visual-admin/my-link/<ta_id>/
    Called by the faculty dashboard to get their configured Power BI URL.
    Returns: { url, source: 'course'|'overall'|'none' }
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, ta_id):
        try:
            sp = request.user.staff_profile
        except Exception:
            return Response({'url': '', 'source': 'none'})

        try:
            ta = TeachingAssignment.objects.get(id=ta_id, staff=sp)
        except TeachingAssignment.DoesNotExist:
            return Response({'url': '', 'source': 'none'})

        try:
            link = sp.visual_admin_link
        except VisualAdminStaffLink.DoesNotExist:
            return Response({'url': '', 'source': 'none'})

        if link.use_course_urls:
            course_link = VisualAdminCourseLink.objects.filter(
                staff_link=link,
                teaching_assignment=ta,
            ).first()
            if course_link and course_link.url:
                return Response({'url': course_link.url, 'source': 'course'})
            # Fall through to overall
        if link.overall_url:
            return Response({'url': link.overall_url, 'source': 'overall'})

        return Response({'url': '', 'source': 'none'})


# ---------------------------------------------------------------------------
# Dashboard statistics for Visual Admin
# ---------------------------------------------------------------------------

class VisualAdminDashboardStatsView(APIView):
    """
    GET /api/academic-v2/visual-admin/dashboard-stats/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not _is_visual_admin(request.user):
            return Response({'detail': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

        total_staff = StaffProfile.objects.filter(status='ACTIVE').count()
        staff_with_link = VisualAdminStaffLink.objects.filter(
            Q(overall_url__gt='') | Q(use_course_urls=True)
        ).count()
        staff_with_course_urls = VisualAdminStaffLink.objects.filter(use_course_urls=True).count()
        staff_with_overall_only = VisualAdminStaffLink.objects.filter(
            use_course_urls=False, overall_url__gt=''
        ).count()
        total_course_links = VisualAdminCourseLink.objects.filter(url__gt='').count()
        staff_no_link = total_staff - staff_with_link

        from accounts.models import Role, UserRole
        try:
            role = Role.objects.get(name='VISUAL_ADMIN')
            va_count = UserRole.objects.filter(role=role).count()
        except Role.DoesNotExist:
            va_count = 0

        return Response({
            'total_staff': total_staff,
            'staff_with_link': staff_with_link,
            'staff_with_course_urls': staff_with_course_urls,
            'staff_with_overall_only': staff_with_overall_only,
            'staff_no_link': staff_no_link,
            'total_course_links': total_course_links,
            'visual_admin_count': va_count,
        })


# ---------------------------------------------------------------------------
# Export links as Excel / PDF
# ---------------------------------------------------------------------------

from rest_framework.decorators import api_view, permission_classes as drf_permission_classes


@api_view(['GET'])
@drf_permission_classes([IsAuthenticated])
def visual_admin_export(request):
    """
    GET /api/academic-v2/visual-admin/export/?format=excel|pdf
    Returns a downloadable file with columns: staff_id, staff_name, url
    """
    if not _is_visual_admin(request.user):
        return Response({'detail': 'Forbidden'}, status=403)

    fmt = request.query_params.get('export_format', 'excel').lower()

    rows = []
    for sp in StaffProfile.objects.filter(status='ACTIVE').select_related('user').order_by('user__first_name'):
        try:
            link = sp.visual_admin_link
            url = link.overall_url or ''
        except VisualAdminStaffLink.DoesNotExist:
            url = ''
        rows.append({
            'Staff ID': sp.staff_id,
            'Staff Name': sp.user.get_full_name() or sp.user.username,
            'URL': url,
        })

    if fmt == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from django.http import HttpResponse
            import io
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Staff Links'
            headers = ['Staff ID', 'Staff Name', 'URL']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = PatternFill('solid', fgColor='4F46E5')
                cell.font = Font(bold=True, color='FFFFFF')
            for row_idx, row_data in enumerate(rows, 2):
                ws.cell(row=row_idx, column=1, value=row_data['Staff ID'])
                ws.cell(row=row_idx, column=2, value=row_data['Staff Name'])
                ws.cell(row=row_idx, column=3, value=row_data['URL'])
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 80
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HttpResponse(
                buf.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="visual_admin_links.xlsx"'
            return response
        except ImportError:
            return Response({'detail': 'openpyxl not installed'}, status=500)

    # PDF
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib import colors
        from django.http import HttpResponse
        import io
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        table_data = [['Staff ID', 'Staff Name', 'URL']]
        for r in rows:
            table_data.append([r['Staff ID'], r['Staff Name'], r['URL'][:80]])
        t = Table(table_data, colWidths=[80, 160, 400])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F3FF')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        doc.build([t])
        buf.seek(0)
        response = HttpResponse(buf.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="visual_admin_links.pdf"'
        return response
    except ImportError:
        return Response({'detail': 'reportlab not installed'}, status=500)


# ---------------------------------------------------------------------------
# Import links from Excel
# ---------------------------------------------------------------------------

@api_view(['POST'])
@drf_permission_classes([IsAuthenticated])
def visual_admin_import(request):
    """
    POST /api/academic-v2/visual-admin/import/
    Accepts an Excel file with columns: staff_id, staff_name (ignored), url
    """
    if not _is_visual_admin(request.user):
        return Response({'detail': 'Forbidden'}, status=403)

    file = request.FILES.get('file')
    if not file:
        return Response({'detail': 'No file provided'}, status=400)

    try:
        import openpyxl
    except ImportError:
        return Response({'detail': 'openpyxl not installed on server'}, status=500)

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(cell.value or '').strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            sid_col = headers.index('staff id')
            url_col = headers.index('url')
        except ValueError:
            return Response({'detail': 'Excel must have columns: Staff ID, Staff Name, URL'}, status=400)

        updated = 0
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            staff_id = str(row[sid_col] or '').strip()
            url = str(row[url_col] or '').strip()
            if not staff_id:
                continue
            try:
                sp = StaffProfile.objects.get(staff_id=staff_id)
                link, _ = VisualAdminStaffLink.objects.get_or_create(
                    staff=sp, defaults={'updated_by': request.user}
                )
                link.overall_url = url
                link.updated_by = request.user
                link.save()
                updated += 1
            except StaffProfile.DoesNotExist:
                errors.append(f'Staff ID {staff_id} not found')

        msg = f'Imported {updated} record(s).'
        if errors:
            msg += f' {len(errors)} error(s): ' + '; '.join(errors[:5])
        return Response({'message': msg, 'updated': updated, 'errors': errors})
    except Exception as e:
        return Response({'detail': f'Failed to read file: {str(e)}'}, status=400)
