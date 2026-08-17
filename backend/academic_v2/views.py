"""
Academic 2.1 API Views
"""

from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db import transaction
from django.db.models import Q
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
import os
import re
import secrets

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow

from .models import (
    AcV2SemesterConfig,
    AcV2SemesterGroup,
    AcV2ClassType,
    AcV2QpPattern,
    AcV2Course,
    AcV2Section,
    AcV2ExamAssignment,
    AcV2StudentMark,
    AcV2DraftMark,
    AcV2UserPatternOverride,
    AcV2EditRequest,
    AcV2CqiEditRequest,
    AcV2InternalMark,
    AcV2QpType,
    AcV2CourseOutcome,
    AcV2QpAssignment,
    AcV2Cycle,
    AcV2CqiToken,
    AcV2CqiOperator,
    AcV2MyMarksSetting,
    AcV2AcademicNotificationSetting,
    AcV2PublishSetting,
    AcV2GoogleSheetLink,
)
from .serializers import (
    AcV2SemesterConfigSerializer,
    AcV2SemesterGroupSerializer,
    AcV2ClassTypeSerializer,
    AcV2QpPatternSerializer,
    AcV2CourseSerializer,
    AcV2SectionSerializer,
    AcV2ExamAssignmentSerializer,
    AcV2StudentMarkSerializer,
    AcV2StudentMarkBulkSerializer,
    AcV2EditRequestSerializer,
    AcV2CqiEditRequestSerializer,
    AcV2InternalMarkSerializer,
    AcV2UserPatternOverrideSerializer,
    AcV2QpTypeSerializer,
    AcV2CourseOutcomeSerializer,
    AcV2CycleSerializer,
    AcV2PassMarkSettingSerializer,
    AcV2MyMarksSettingSerializer,
    AcV2AcademicNotificationSettingSerializer,
    AcV2CqiTokenSerializer,
    AcV2CqiOperatorSerializer,
    AcV2PublishSettingSerializer,
)
from .models import AcV2PassMarkSetting, AcV2GoogleSheetsOAuthCredential
from .google_sheets_service import (
    GoogleSheetsServiceError,
    create_google_spreadsheet,
    sync_google_sheet_to_backend,
    sync_marks_to_google_sheet,
)


def _extract_spreadsheet_id(sheet_url: str) -> str:
    text = str(sheet_url or '').strip()
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', text)
    return str(match.group(1)) if match else ''


def _render_notification_template(template: str, context: dict) -> str:
    """Simple token replacement for templates like 'Hello {student_name}'."""
    out = str(template or '')
    for k, v in (context or {}).items():
        try:
            out = out.replace('{' + str(k) + '}', str(v if v is not None else ''))
        except Exception:
            continue
    return out


@api_view(['GET', 'POST', 'DELETE'])
@permission_classes([IsAuthenticated])
def google_sheet_links(request):
    """Get and update section-level Google Sheet links and per-exam mapping configs."""
    if request.method == 'POST':
        payload = request.data or {}
        section_id = payload.get('sectionId') or payload.get('section_id')
        if not section_id:
            return Response({'detail': 'sectionId is required.'}, status=status.HTTP_400_BAD_REQUEST)

        section = get_object_or_404(AcV2Section, pk=section_id)
        link, _ = AcV2GoogleSheetLink.objects.get_or_create(
            section=section,
            defaults={
                'sheet_url': '',
                'spreadsheet_id': '',
                'exam_configs': {},
                'is_active': True,
                'updated_by': request.user,
            },
        )

        next_sheet_url = payload.get('sheetUrl') or payload.get('sheet_url')
        if next_sheet_url is not None:
            next_sheet_url = str(next_sheet_url).strip()
            if next_sheet_url:
                spreadsheet_id = _extract_spreadsheet_id(next_sheet_url)
                if not spreadsheet_id:
                    return Response({'detail': 'Invalid Google Sheets URL. Paste a URL like https://docs.google.com/spreadsheets/d/<id>/edit'}, status=status.HTTP_400_BAD_REQUEST)
                link.sheet_url = next_sheet_url
                link.spreadsheet_id = spreadsheet_id
                link.is_active = True

        exam_id = payload.get('examAssignmentId') or payload.get('exam_assignment_id')
        if exam_id:
            exam_configs = link.exam_configs if isinstance(link.exam_configs, dict) else {}
            existing_exam_config = exam_configs.get(str(exam_id), {}) if isinstance(exam_configs.get(str(exam_id), {}), dict) else {}

            mapping_payload = payload.get('mapping') or payload.get('columnMapping') or payload.get('column_mapping') or {}
            if not isinstance(mapping_payload, dict):
                mapping_payload = {}

            merged = {
                **existing_exam_config,
                'sheetTab': str(payload.get('sheetTab') or payload.get('sheet_tab') or existing_exam_config.get('sheetTab') or '').strip(),
                'regNoColumn': str(mapping_payload.get('regNoColumn') or mapping_payload.get('reg_no_column') or existing_exam_config.get('regNoColumn') or 'A').strip().upper(),
                'nameColumn': str(mapping_payload.get('nameColumn') or mapping_payload.get('name_column') or existing_exam_config.get('nameColumn') or 'B').strip().upper(),
                'questionColumns': mapping_payload.get('questionColumns') or mapping_payload.get('question_columns') or existing_exam_config.get('questionColumns') or {},
            }
            exam_configs[str(exam_id)] = merged
            link.exam_configs = exam_configs

        link.updated_by = request.user
        link.save(update_fields=['sheet_url', 'spreadsheet_id', 'exam_configs', 'is_active', 'updated_by', 'updated_at'])

        return Response({
            'id': str(link.id),
            'sectionId': str(section.id),
            'sheetUrl': link.sheet_url,
            'spreadsheetId': link.spreadsheet_id,
            'examConfigs': link.exam_configs,
            'active': bool(link.is_active and link.sheet_url and link.spreadsheet_id),
        })

    if request.method == 'DELETE':
        section_id = request.query_params.get('sectionId') or request.query_params.get('section_id')
        if not section_id:
            return Response({'detail': 'sectionId is required.'}, status=status.HTTP_400_BAD_REQUEST)
        link = AcV2GoogleSheetLink.objects.filter(section_id=section_id).first()
        if link:
            link.delete()
        return Response({'success': True})

    sections_qs = (
        AcV2Section.objects.select_related(
            'course',
            'course__subject',
            'course__subject__course',
            'course__subject__course__department',
            'course__semester',
            'teaching_assignment',
            'teaching_assignment__section',
            'teaching_assignment__section__batch',
            'faculty_user',
        )
        .prefetch_related('exam_assignments', 'exam_assignments__student_marks')
        .order_by('course__subject_code', 'section_name')
    )

    links_by_section_id = {
        str(link.section_id): link
        for link in AcV2GoogleSheetLink.objects.filter(section__in=sections_qs)
    }

    # Build a fallback map from existing configurations across all sections
    # Key: (classType: str, qpType: str, assignmentName: str) -> config dict
    fallback_map = {}
    for link in links_by_section_id.values():
        if not isinstance(link.exam_configs, dict):
            continue
        c_type = getattr(getattr(link.section, 'course', None), 'class_type', None)
        c_type_code = str(getattr(c_type, 'name', 'THEORY') or 'THEORY').strip().upper()
        for exam in link.section.exam_assignments.all():
            cfg = link.exam_configs.get(str(exam.id))
            if cfg and isinstance(cfg, dict) and cfg.get('regNoColumn'):
                qp_type = str(getattr(exam, 'qp_type', '') or '').strip().upper()
                assignment_name = str(exam.exam_display_name or exam.exam or 'Assignment').strip().lower()
                fallback_map[(c_type_code, qp_type, assignment_name)] = cfg

    payload = []
    for section in sections_qs:
        section_link = links_by_section_id.get(str(section.id))
        section_exam_configs = section_link.exam_configs if section_link and isinstance(section_link.exam_configs, dict) else {}
        course = section.course
        subject = getattr(course, 'subject', None)
        subject_course = getattr(subject, 'course', None)
        subject_department = getattr(subject_course, 'department', None)
        semester = getattr(course, 'semester', None)
        teaching_assignment = getattr(section, 'teaching_assignment', None)
        section_model = getattr(teaching_assignment, 'section', None)
        batch = getattr(section_model, 'batch', None)

        department_name = None
        if subject_department is not None:
            department_name = getattr(subject_department, 'short_name', None) or getattr(subject_department, 'name', None)

        class_type_name = None
        if getattr(course, 'class_type', None) is not None:
            class_type_name = getattr(course.class_type, 'name', None) or ''

        faculty_name = None
        if getattr(section, 'faculty_user', None) is not None:
            faculty_name = section.faculty_user.get_full_name() or section.faculty_user.username
        elif getattr(teaching_assignment, 'staff', None) is not None and getattr(teaching_assignment.staff, 'user', None) is not None:
            faculty_name = teaching_assignment.staff.user.get_full_name() or teaching_assignment.staff.user.username

        assignments = []
        exam_assignments = []
        students_map = {}
        for exam_assignment in sorted(section.exam_assignments.all(), key=lambda item: (getattr(item, 'exam', '') or '').lower()):
            assignment_name = exam_assignment.exam_display_name or exam_assignment.exam or 'Assignment'
            assignment_config = section_exam_configs.get(str(exam_assignment.id), {}) if isinstance(section_exam_configs.get(str(exam_assignment.id), {}), dict) else {}
            
            # Apply fallback mapping if missing
            if not assignment_config or not assignment_config.get('regNoColumn'):
                c_type_name_upper = str(class_type_name or 'THEORY').strip().upper()
                qp_type_upper = str(getattr(exam_assignment, 'qp_type', '') or '').strip().upper()
                assign_name_lower = str(assignment_name).strip().lower()
                fallback_key = (c_type_name_upper, qp_type_upper, assign_name_lower)
                if fallback_key in fallback_map:
                    assignment_config = fallback_map[fallback_key]

            assignments.append(assignment_name)

            # Resolve mark_manager info from the QP pattern
            mm_info = None
            try:
                resolved_pattern = exam_assignment.get_qp_pattern()
                if isinstance(resolved_pattern, dict):
                    mm_raw = resolved_pattern.get('mark_manager')
                    if isinstance(mm_raw, dict) and mm_raw.get('enabled'):
                        mm_mode = mm_raw.get('mode', 'admin_defined')  # 'admin_defined' or 'user_define'
                        # Normalise: 'user_define' → 'user_defined' for frontend clarity
                        if mm_mode in ('user_define', 'user_defined'):
                            mm_mode = 'user_defined'
                        else:
                            mm_mode = 'admin_defined'
                        # Build question list from the pattern (titles + marks)
                        mm_questions = []
                        if mm_mode == 'admin_defined':
                            p_titles = resolved_pattern.get('titles') or []
                            p_marks = resolved_pattern.get('marks') or []
                            p_enabled = resolved_pattern.get('enabled') or [True] * len(p_titles)
                            # Also check nested questions array
                            if not p_titles and isinstance(resolved_pattern.get('questions'), list):
                                for qi, q in enumerate(resolved_pattern['questions']):
                                    if isinstance(q, dict):
                                        mm_questions.append({
                                            'id': str(q.get('id') or f'q{qi+1}'),
                                            'title': str(q.get('question_number') or q.get('title') or f'Q{qi+1}'),
                                            'max_marks': q.get('max_marks') or 0,
                                        })
                            else:
                                for qi in range(len(p_titles)):
                                    if qi < len(p_enabled) and not p_enabled[qi]:
                                        continue
                                    mm_questions.append({
                                        'id': f'q{qi+1}',
                                        'title': p_titles[qi] if qi < len(p_titles) else f'Q{qi+1}',
                                        'max_marks': p_marks[qi] if qi < len(p_marks) else 0,
                                    })
                        mm_info = {
                            'enabled': True,
                            'mode': mm_mode,
                            'questions': mm_questions,
                        }
            except Exception:
                mm_info = None

            # classTypeName: prefer FK name (display title), then raw string
            class_type_display = (
                getattr(getattr(course, 'class_type', None), 'name', None)
                or class_type_name
                or 'Theory'
            )

            exam_assignments.append({
                'id': str(exam_assignment.id),
                'sectionId': str(section.id),
                'courseCode': getattr(course, 'subject_code', '') or '',
                'courseName': getattr(course, 'subject_name', '') or '',
                'section': getattr(section, 'section_name', '') or '',
                'assignment': assignment_name,
                'sheetTab': assignment_config.get('sheetTab') or assignment_name,
                'classType': class_type_name or 'THEORY',
                'classTypeName': class_type_display,
                'qpType': getattr(exam_assignment, 'qp_type', '') or '',
                'columnMapping': {
                    'regNoColumn': assignment_config.get('regNoColumn') or 'A',
                    'nameColumn': assignment_config.get('nameColumn') or 'B',
                    'questionColumns': assignment_config.get('questionColumns') if isinstance(assignment_config.get('questionColumns'), dict) else {},
                },
                'markManager': mm_info,
            })

            for mark in exam_assignment.student_marks.all():
                register_no = mark.reg_no or ''
                student_entry = students_map.setdefault(register_no, {
                    'name': mark.student_name or '',
                    'registerNo': register_no,
                })
                if assignment_name:
                    student_entry[assignment_name] = str(mark.total_mark) if mark.total_mark is not None else ''

        students = sorted(students_map.values(), key=lambda item: (item.get('name') or '', item.get('registerNo') or ''))

        payload.append({
            'id': str(section.id),
            'courseCode': getattr(course, 'subject_code', '') or '',
            'courseName': getattr(course, 'subject_name', '') or '',
            'semester': str(getattr(semester, 'number', '') or ''),
            'department': department_name or '',
            'batch': getattr(batch, 'name', '') or '',
            'facultyName': faculty_name or '',
            'section': getattr(section, 'section_name', '') or '',
            'classType': class_type_name or 'THEORY',
            'active': bool(section_link and section_link.is_active and section_link.sheet_url and section_link.spreadsheet_id),
            'sheetUrl': section_link.sheet_url if section_link else '',
            'spreadsheetId': section_link.spreadsheet_id if section_link else '',
            'assignments': assignments,
            'examAssignments': exam_assignments,
            'students': students,
        })

    return Response(payload)


GOOGLE_SHEETS_SCOPES = [
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/script.projects',  # Required for Apps Script injection
]


def _get_google_oauth_client_config() -> dict:
    client_id = getattr(settings, 'GOOGLE_OAUTH_CLIENT_ID', None) or os.environ.get('GOOGLE_OAUTH_CLIENT_ID', '')
    client_secret = getattr(settings, 'GOOGLE_OAUTH_CLIENT_SECRET', None) or os.environ.get('GOOGLE_OAUTH_CLIENT_SECRET', '')
    return {
        'web': {
            'client_id': client_id,
            'client_secret': client_secret,
            'auth_uri': 'https://accounts.google.com/o/oauth2/auth',
            'token_uri': 'https://oauth2.googleapis.com/token',
            'redirect_uris': [
                getattr(settings, 'GOOGLE_OAUTH_REDIRECT_URI', None) or os.environ.get('GOOGLE_OAUTH_REDIRECT_URI', ''),
            ],
        }
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def google_sheets_oauth_status(request):
    credential = AcV2GoogleSheetsOAuthCredential.objects.filter(is_active=True).order_by('-updated_at').first()
    stored_scopes = credential.scopes if credential else []
    has_script_scope = any('script.projects' in str(s) for s in (stored_scopes or []))
    return Response({
        'authorized': bool(credential),
        'userEmail': credential.google_user_email if credential else None,
        'hasScriptScope': has_script_scope,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def google_sheets_oauth_start(request):
    client_config = _get_google_oauth_client_config()
    if not client_config['web']['client_id'] or not client_config['web']['client_secret']:
        return Response({'detail': 'Google OAuth client ID and client secret are not configured.'}, status=status.HTTP_400_BAD_REQUEST)

    state = secrets.token_urlsafe(16)
    request.session['google_sheets_oauth_state'] = state

    env_redirect = getattr(settings, 'GOOGLE_OAUTH_REDIRECT_URI', None) or os.environ.get('GOOGLE_OAUTH_REDIRECT_URI', '')
    redirect_uri = request.GET.get('redirect_uri') or env_redirect or request.build_absolute_uri(reverse('academic_v2:google-sheets-oauth-callback'))
    flow = InstalledAppFlow.from_client_config(client_config, GOOGLE_SHEETS_SCOPES)
    flow.redirect_uri = redirect_uri
    auth_url, _ = flow.authorization_url(access_type='offline', prompt='consent', state=state)
    return Response({'authUrl': auth_url, 'redirectUri': redirect_uri})


@api_view(['GET'])
def google_sheets_oauth_callback(request):
    state = request.GET.get('state')
    expected_state = request.session.get('google_sheets_oauth_state')
    if state != expected_state:
        return HttpResponse('Google authorization failed: invalid state.', status=400)

    code = request.GET.get('code')
    if not code:
        return HttpResponse('Google authorization failed: missing authorization code.', status=400)

    client_config = _get_google_oauth_client_config()
    if not client_config['web']['client_id'] or not client_config['web']['client_secret']:
        return HttpResponse('Google OAuth client ID and client secret are not configured.', status=400)

    env_redirect = getattr(settings, 'GOOGLE_OAUTH_REDIRECT_URI', None) or os.environ.get('GOOGLE_OAUTH_REDIRECT_URI', '')
    redirect_uri = request.GET.get('redirect_uri') or env_redirect or request.build_absolute_uri(reverse('academic_v2:google-sheets-oauth-callback'))
    flow = InstalledAppFlow.from_client_config(client_config, GOOGLE_SHEETS_SCOPES)
    flow.redirect_uri = redirect_uri
    flow.fetch_token(code=code)
    creds = flow.credentials

    credential = AcV2GoogleSheetsOAuthCredential.objects.filter(is_active=True).order_by('-updated_at').first()
    if credential is None:
        credential = AcV2GoogleSheetsOAuthCredential.objects.create(user=request.user if request.user.is_authenticated else None)

    credential.google_user_email = creds.id_token_claims.get('email', '') if getattr(creds, 'id_token_claims', None) else ''
    credential.access_token = creds.token or ''
    credential.refresh_token = creds.refresh_token or credential.refresh_token or ''
    credential.token_uri = creds.token_uri or 'https://oauth2.googleapis.com/token'
    credential.client_id = client_config['web']['client_id']
    credential.client_secret = client_config['web']['client_secret']
    credential.scopes = list(getattr(creds, 'scopes', []) or GOOGLE_SHEETS_SCOPES)
    credential.is_active = True
    credential.save()

    html = '''<html><body><p>Google authorization completed. You can close this window.</p><script>window.opener && window.opener.postMessage({type:'google-oauth-success'}, '*'); window.close();</script></body></html>'''
    return HttpResponse(html)


def _get_qp_specs_for_exam(ea):
    """Retrieve detailed question paper structure (titles, max marks, COs, BTLs) for an exam assignment."""
    ta = ea.section.teaching_assignment
    cr = getattr(ta, 'curriculum_row', None)
    es = getattr(ta, 'elective_subject', None)

    draft = ea.draft_data if isinstance(ea.draft_data, dict) else {}
    user_pattern = draft.get('user_pattern')

    questions = []
    if user_pattern and isinstance(user_pattern, dict):
        titles = user_pattern.get('titles', [])
        marks_list = user_pattern.get('marks', [])
        cos = user_pattern.get('cos', [])
        btls = user_pattern.get('btls', [])
        enabled = user_pattern.get('enabled', [])
        for i in range(len(titles)):
            if i < len(enabled) and not enabled[i]:
                continue
            questions.append({
                'id': f'q{i}',
                'title': titles[i] if i < len(titles) else str(i + 1),
                'max_marks': marks_list[i] if i < len(marks_list) else 0,
                'co': cos[i] if i < len(cos) else 0,
                'btl': btls[i] if i < len(btls) else None,
            })
    else:
        qp_type = ''
        try:
            qp_type = (ea.section.course.question_paper_type or '').strip()
        except Exception:
            qp_type = ''
        if not qp_type:
            qp_type = (getattr(cr, 'question_paper_type', None) or getattr(es, 'question_paper_type', None) or '').strip()
        if not qp_type:
            qp_type = (ea.qp_type or '').strip() or (ea.exam or '').strip() or ''

        exam_key = (ea.exam_display_name or ea.exam or '').strip()

        ct = None
        try:
            ct = ea.section.course.class_type
        except Exception:
            ct = None

        base_qs = AcV2QpPattern.objects.filter(qp_type=qp_type, is_active=True)
        matched_pattern = None

        if ct is not None:
            scoped = base_qs.filter(class_type=ct)
            if exam_key:
                matched_pattern = scoped.filter(name__iexact=exam_key).order_by('-updated_at').first()
            else:
                matched_pattern = scoped.order_by('-updated_at').first()

        if not matched_pattern:
            global_qs = base_qs.filter(class_type__isnull=True)
            if exam_key:
                matched_pattern = global_qs.filter(name__iexact=exam_key).order_by('-updated_at').first()
            else:
                matched_pattern = global_qs.order_by('-updated_at').first()

        if matched_pattern and isinstance(matched_pattern.pattern, dict):
            p = matched_pattern.pattern
            titles = p.get('titles', [])
            marks_list = p.get('marks', [])
            cos = p.get('cos', [])
            btls = p.get('btls', [])
            enabled = p.get('enabled', [])
            for i in range(len(titles)):
                if i < len(enabled) and not enabled[i]:
                    continue
                questions.append({
                    'id': f'q{i}',
                    'title': titles[i] if i < len(titles) else str(i + 1),
                    'max_marks': marks_list[i] if i < len(marks_list) else 0,
                    'co': cos[i] if i < len(cos) else 0,
                    'btl': btls[i] if i < len(btls) else None,
                })
    return questions


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def google_sheets_sync_to_sheet(request):
    payload = request.data or {}
    spreadsheet_id = str(payload.get('spreadsheetId') or payload.get('spreadsheet_id') or '').strip()
    exam_assignment_id = payload.get('examAssignmentId') or payload.get('exam_assignment_id')
    sheet_name = str(payload.get('sheetName') or payload.get('sheet_name') or '').strip() or None
    config = payload.get('config') or {}

    if not spreadsheet_id or not exam_assignment_id:
        return Response({'detail': 'spreadsheetId and examAssignmentId are required.'}, status=status.HTTP_400_BAD_REQUEST)

    exam_assignment = get_object_or_404(AcV2ExamAssignment, pk=exam_assignment_id)
    try:
        result = sync_marks_to_google_sheet(
            exam_assignment=exam_assignment,
            spreadsheet_id=spreadsheet_id,
            config=config,
            sheet_name=sheet_name,
        )
    except GoogleSheetsServiceError as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    return Response(result)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def google_sheets_sync_from_sheet(request):
    payload = request.data or {}
    spreadsheet_id = str(payload.get('spreadsheetId') or payload.get('spreadsheet_id') or '').strip()
    exam_assignment_id = payload.get('examAssignmentId') or payload.get('exam_assignment_id')
    section_id = payload.get('sectionId') or payload.get('section_id')
    sheet_name = str(payload.get('sheetName') or payload.get('sheet_name') or '').strip() or None
    config = payload.get('config') or {}

    if not spreadsheet_id:
        return Response({'detail': 'spreadsheetId is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        if exam_assignment_id:
            exam_assignment = get_object_or_404(AcV2ExamAssignment, pk=exam_assignment_id)
            result = sync_google_sheet_to_backend(
                exam_assignment=exam_assignment,
                spreadsheet_id=spreadsheet_id,
                config=config,
                sheet_name=sheet_name,
            )
            return Response(result)

        if section_id:
            section = get_object_or_404(AcV2Section, pk=section_id)
            results = []
            for exam_assignment in section.exam_assignments.all():
                results.append(sync_google_sheet_to_backend(
                    exam_assignment=exam_assignment,
                    spreadsheet_id=spreadsheet_id,
                    config=config,
                    sheet_name=sheet_name or (exam_assignment.exam_display_name or exam_assignment.exam or 'Marks'),
                ))
            return Response(results)

        return Response({'detail': 'examAssignmentId or sectionId is required.'}, status=status.HTTP_400_BAD_REQUEST)
    except GoogleSheetsServiceError as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def google_sheets_pull_live(request):
    """Pull latest marks from linked Google Sheet into backend for one exam or full section."""
    payload = request.data or {}
    exam_assignment_id = payload.get('examAssignmentId') or payload.get('exam_assignment_id')
    section_id = payload.get('sectionId') or payload.get('section_id')
    config = payload.get('config') or {}

    if not exam_assignment_id and not section_id:
        return Response({'detail': 'examAssignmentId or sectionId is required.'}, status=status.HTTP_400_BAD_REQUEST)

    if exam_assignment_id:
        exam_assignment = get_object_or_404(AcV2ExamAssignment, pk=exam_assignment_id)
        link = AcV2GoogleSheetLink.objects.filter(section=exam_assignment.section, is_active=True).first()
        if not link or not link.spreadsheet_id:
            return Response({'detail': 'No linked Google Sheet found for this section.'}, status=status.HTTP_400_BAD_REQUEST)

        exam_configs = link.exam_configs if isinstance(link.exam_configs, dict) else {}
        exam_config = exam_configs.get(str(exam_assignment.id), {}) if isinstance(exam_configs.get(str(exam_assignment.id), {}), dict) else {}
        sheet_name = str(exam_config.get('sheetTab') or exam_assignment.exam_display_name or exam_assignment.exam or 'Marks').strip()
        result = sync_google_sheet_to_backend(
            exam_assignment=exam_assignment,
            spreadsheet_id=link.spreadsheet_id,
            config=config,
            sheet_name=sheet_name,
            column_mapping=exam_config,
        )
        return Response(result)

    section = get_object_or_404(AcV2Section, pk=section_id)
    link = AcV2GoogleSheetLink.objects.filter(section=section, is_active=True).first()
    if not link or not link.spreadsheet_id:
        return Response({'detail': 'No linked Google Sheet found for this section.'}, status=status.HTTP_400_BAD_REQUEST)

    exam_configs = link.exam_configs if isinstance(link.exam_configs, dict) else {}
    results = []
    for exam_assignment in section.exam_assignments.all():
        exam_config = exam_configs.get(str(exam_assignment.id), {}) if isinstance(exam_configs.get(str(exam_assignment.id), {}), dict) else {}
        sheet_name = str(exam_config.get('sheetTab') or exam_assignment.exam_display_name or exam_assignment.exam or 'Marks').strip()
        results.append(sync_google_sheet_to_backend(
            exam_assignment=exam_assignment,
            spreadsheet_id=link.spreadsheet_id,
            config=config,
            sheet_name=sheet_name,
            column_mapping=exam_config,
        ))
    return Response(results)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def google_sheets_inject_script(request):
    """Inject a Google Apps Script into an already-linked Google Sheet for the given section.

    The Apps Script is auto-generated based on the section's exam assignment configs
    (sheet tabs, register-number column, question columns) and registers an onEdit +
    hourly trigger that POSTs live mark data back to our webhook endpoint.
    """
    from .google_sheets_service import inject_apps_script_to_sheet

    payload = request.data or {}
    section_id = str(payload.get('sectionId') or payload.get('section_id') or '').strip()
    config = payload.get('config') or {}

    if not section_id:
        return Response({'detail': 'sectionId is required.'}, status=status.HTTP_400_BAD_REQUEST)

    section = get_object_or_404(AcV2Section, pk=section_id)
    link = AcV2GoogleSheetLink.objects.filter(section=section, is_active=True).first()
    if not link or not link.spreadsheet_id:
        return Response(
            {'detail': 'No linked Google Sheet found for this section. Paste the sheet URL first.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Build per-exam config list for the Apps Script
    exam_configs_for_script: list[dict] = []
    section_exam_configs = link.exam_configs if isinstance(link.exam_configs, dict) else {}
    for ea in section.exam_assignments.order_by('exam'):
        ea_config = section_exam_configs.get(str(ea.id), {}) if isinstance(section_exam_configs.get(str(ea.id), {}), dict) else {}
        assignment_name = ea.exam_display_name or ea.exam or 'Assignment'
        exam_configs_for_script.append({
            'examId': str(ea.id),
            'sheetTab': ea_config.get('sheetTab') or assignment_name,
            'regNoColumn': ea_config.get('regNoColumn') or 'A',
            'nameColumn': ea_config.get('nameColumn') or 'B',
            'questionColumns': ea_config.get('questionColumns') if isinstance(ea_config.get('questionColumns'), dict) else {},
        })

    # Determine the base backend URL for the webhook
    backend_url = (
        getattr(settings, 'BACKEND_BASE_URL', None)
        or os.environ.get('BACKEND_BASE_URL', '')
        or request.build_absolute_uri('/').rstrip('/')
    )

    try:
        result = inject_apps_script_to_sheet(
            spreadsheet_id=link.spreadsheet_id,
            section_id=section_id,
            exam_configs=exam_configs_for_script,
            backend_url=backend_url,
            config=config,
        )
    except GoogleSheetsServiceError as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    return Response(result)


@api_view(['POST'])
@csrf_exempt
def google_sheets_webhook(request):
    """Webhook endpoint called by the Google Apps Script in a linked sheet.

    Accepts mark rows keyed by register number, matches to students, and saves
    AcV2StudentMark records — identical to the pull-live flow but push-initiated.

    No auth token required since the Apps Script cannot pass session cookies;
    the sectionId + examAssignmentId combination is treated as a shared secret.
    """
    payload = request.data or {}
    section_id = str(payload.get('sectionId') or payload.get('section_id') or '').strip()
    exam_assignment_id = payload.get('examAssignmentId') or payload.get('exam_assignment_id')
    rows = payload.get('rows') or []

    if not section_id:
        return Response({'detail': 'sectionId is required.'}, status=status.HTTP_400_BAD_REQUEST)
    if not exam_assignment_id:
        return Response({'detail': 'examAssignmentId is required.'}, status=status.HTTP_400_BAD_REQUEST)
    if not isinstance(rows, list):
        return Response({'detail': 'rows must be a list.'}, status=status.HTTP_400_BAD_REQUEST)

    # Verify the section + exam assignment exist and are linked
    exam_assignment = get_object_or_404(AcV2ExamAssignment, pk=exam_assignment_id)
    if str(exam_assignment.section_id) != section_id:
        return Response({'detail': 'Section and exam assignment mismatch.'}, status=status.HTTP_400_BAD_REQUEST)

    link = AcV2GoogleSheetLink.objects.filter(section_id=section_id, is_active=True).first()
    if not link:
        return Response({'detail': 'No active Google Sheet link for this section.'}, status=status.HTTP_400_BAD_REQUEST)

    # Process rows directly (they come pre-parsed from the Apps Script)
    from academics.models import StudentProfile
    from .models import AcV2StudentMark

    updated_count = 0
    for row_data in rows:
        if not isinstance(row_data, dict):
            continue
        reg_no = str(row_data.get('regNo') or row_data.get('reg_no') or '').strip()
        student_name = str(row_data.get('name') or '').strip()
        marks_raw = row_data.get('marks') or {}
        if not reg_no:
            continue

        student = StudentProfile.objects.filter(reg_no__iexact=reg_no).first()
        if not student:
            continue

        question_marks: dict = {}
        if isinstance(marks_raw, dict):
            for qkey, qval in marks_raw.items():
                try:
                    question_marks[str(qkey)] = float(qval)
                except (TypeError, ValueError):
                    pass

        mark_obj, _ = AcV2StudentMark.objects.update_or_create(
            exam_assignment=exam_assignment,
            student=student,
            defaults={
                'reg_no': reg_no,
                'student_name': student_name or (student.user.get_full_name() if getattr(student, 'user', None) else ''),
                'question_marks': question_marks,
                'is_absent': False,
                'is_exempted': False,
            },
        )
        try:
            qp_pattern = exam_assignment.get_qp_pattern()
            mark_obj.calculate_co_marks(qp_pattern)
            mark_obj.calculate_total()
            mark_obj.save(update_fields=[
                'reg_no', 'student_name', 'question_marks', 'total_mark',
                'is_absent', 'is_exempted', 'co1_mark', 'co2_mark',
                'co3_mark', 'co4_mark', 'co5_mark', 'co6_mark',
            ])
        except Exception:
            mark_obj.save(update_fields=['reg_no', 'student_name', 'question_marks'])
            
        from .models import AcV2DraftMark
        AcV2DraftMark.objects.update_or_create(
            exam_assignment=exam_assignment,
            student=student,
            defaults={
                'reg_no': reg_no,
                'student_name': student_name or (student.user.get_full_name() if getattr(student, 'user', None) else ''),
                'question_marks': question_marks,
                'total_mark': mark_obj.total_mark,
                'is_absent': False,
            },
        )
        updated_count += 1

    return Response({'success': True, 'updatedRows': updated_count})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def google_sheets_create(request):
    """Create Google spreadsheets for selected Academic 2.1 sections."""
    from academics.models import StudentSectionAssignment

    payload = request.data or {}
    section_ids = payload.get('section_ids') or []
    config = payload.get('config') or {}

    if not isinstance(section_ids, (list, tuple)) or not section_ids:
        return Response({'detail': 'At least one section must be selected.'}, status=status.HTTP_400_BAD_REQUEST)

    results = []
    for section_id in section_ids:
        section = get_object_or_404(AcV2Section, pk=section_id)
        course = section.course
        course_code = getattr(course, 'subject_code', '') or getattr(course, 'subject_name', '') or 'Course'
        course_name = getattr(course, 'subject_name', '') or 'Course'
        title = f"{course_code} - {course_name} Marks Entry"
        
        assignments = []
        assignment_mapping = {}
        seen_names = {}
        exam_assignments_list = sorted(section.exam_assignments.all(), key=lambda item: (getattr(item, 'exam', '') or '').lower())
        for exam_assignment in exam_assignments_list:
            base_name = (exam_assignment.exam_display_name or exam_assignment.exam or 'Assignment').strip()
            if base_name not in seen_names:
                seen_names[base_name] = 1
                unique_name = base_name
            else:
                seen_names[base_name] += 1
                unique_name = f"{base_name} ({seen_names[base_name]})"
            assignments.append(unique_name)
            assignment_mapping[exam_assignment.id] = unique_name

        # Get all active students in the section
        student_assignments = (
            StudentSectionAssignment.objects
            .filter(section=section.teaching_assignment.section, end_date__isnull=True)
            .select_related('student__user')
            .order_by('student__reg_no')
        )

        sheet_data = {}
        for ea in exam_assignments_list:
            unique_name = assignment_mapping[ea.id]
            questions = _get_qp_specs_for_exam(ea)

            # Build Header Row
            headers_row = ["ROLL NO", "NAME"]
            for q in questions:
                header_text = q['title']
                details = []
                if q.get('max_marks'):
                    details.append(f"MAX: {q['max_marks']}")
                if q.get('co'):
                    details.append(f"CO{q['co']}")
                if q.get('btl'):
                    details.append(f"{q['btl']}")
                
                if details:
                    header_text += f" ({', '.join(details)})"
                headers_row.append(header_text)
            headers_row.extend(["TOTAL", "ABSENT"])

            rows = [headers_row]

            # Fetch marks / draft marks
            existing = {
                str(sm.student_id): sm
                for sm in AcV2StudentMark.objects.filter(exam_assignment=ea)
            }
            draft_existing = {
                str(dm.student_id): dm
                for dm in AcV2DraftMark.objects.filter(exam_assignment=ea)
            }

            # Build Student Rows
            for sa in student_assignments:
                sp = sa.student
                sm = existing.get(str(sp.id))
                dm = draft_existing.get(str(sp.id))

                mark_val = None
                co_marks_val = {}
                is_absent_val = False

                if dm:
                    mark_val = float(dm.total_mark) if dm.total_mark is not None else None
                    co_marks_val = dm.question_marks if isinstance(dm.question_marks, dict) else {}
                    is_absent_val = bool(dm.is_absent)
                elif sm:
                    mark_val = float(sm.total_mark) if sm.total_mark is not None else None
                    co_marks_val = sm.question_marks if isinstance(sm.question_marks, dict) else {}
                    is_absent_val = bool(sm.is_absent)

                row = [sp.reg_no or '', str(sp.user) if sp.user else sp.reg_no or '']
                for q in questions:
                    q_id = q['id']
                    val = co_marks_val.get(q_id)
                    if val is None:
                        try:
                            idx = q_id.replace('q', '')
                            val = co_marks_val.get(idx)
                        except Exception:
                            pass
                    row.append(float(val) if val is not None else '')
                
                row.append(mark_val if mark_val is not None else '')
                row.append('Yes' if is_absent_val else 'No')
                rows.append(row)
            
            sheet_data[unique_name] = rows

        # Extract the target folder ID from the configuration and pass it explicitly
        # into the Google Drive create payload as a list element.
        spreadsheet_folder_id = str(config.get('spreadsheetFolderId') or '').strip()
        if not spreadsheet_folder_id:
            return Response(
                {'detail': 'Spreadsheet folder ID is required. Please provide a valid Google Drive folder ID so the service account can create the sheet directly inside that folder.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            created = create_google_spreadsheet(
                title=title,
                assignments=assignments,
                config=config,
                folder_id=spreadsheet_folder_id,
                course_code=course_code,
                webhook_url=str(config.get('webhookUrl') or config.get('webhook_url') or '').strip() or 'https://your-idcs-api.krct.ac.in/api/marks/sync',
                sheet_data=sheet_data,
            )
        except GoogleSheetsServiceError as exc:
            msg = str(exc)
            # If this is a Drive storage quota error, return 507 Insufficient Storage
            if 'storage quota' in msg.lower() or 'storagequotaexceeded' in msg.replace(' ', '').lower():
                return Response({'detail': msg}, status=507)
            return Response({'detail': msg}, status=status.HTTP_400_BAD_REQUEST)

        results.append({
            'id': str(section.id),
            'courseCode': getattr(course, 'subject_code', '') or '',
            'courseName': getattr(course, 'subject_name', '') or '',
            'section': getattr(section, 'section_name', '') or '',
            'sheetUrl': created.get('sheetUrl'),
            'spreadsheetId': created.get('spreadsheetId'),
            'title': created.get('title'),
        })

    return Response(results)


def _resolve_mobile_for_student_profile(sp) -> str:
    """Best-effort mobile resolver for a StudentProfile-like object."""
    mobile = ''
    try:
        mobile = str(getattr(sp, 'mobile_number', '') or '').strip()
    except Exception:
        mobile = ''
    if not mobile:
        try:
            u = getattr(sp, 'user', None)
            mobile = str(getattr(u, 'mobile_number', '') or getattr(u, 'mobile', '') or '').strip() if u else ''
        except Exception:
            mobile = ''
    return str(mobile or '').strip()


def _row_has_meaningful_mark_data(row) -> bool:
    if not isinstance(row, dict):
        return False

    mark = row.get('mark')
    if mark is not None:
        try:
            if float(mark) > 0:
                return True
        except (TypeError, ValueError):
            pass

    co_marks = row.get('co_marks') if isinstance(row.get('co_marks'), dict) else {}
    for value in co_marks.values():
        if value is None:
            continue
        try:
            if float(value) > 0:
                return True
        except (TypeError, ValueError):
            continue
    return False


def _send_student_row_filled_notifications(
    *,
    exam_assignment,
    actor_user,
    previous_marks: dict,
    current_marks: dict,
):
    """Send WhatsApp notifications when a student row is filled with a mark before publish."""
    import logging
    logger = logging.getLogger('academic_v2.notifications')

    try:
        cfg, _ = AcV2AcademicNotificationSetting.objects.get_or_create(key='DEFAULT')
    except Exception as e:
        logger.error(f'Failed to load notification settings: {e}')
        return 0

    if not bool(getattr(cfg, 'notify_on_row_filled', False)):
        return 0

    try:
        from academics.models import StudentSectionAssignment
        from accounts.services.sms import send_whatsapp
    except Exception:
        return 0

    recipients: set[str] = set()
    if isinstance(previous_marks, dict) and isinstance(current_marks, dict):
        all_ids = set(str(k) for k in previous_marks.keys()) | set(str(k) for k in current_marks.keys())
        for sid in all_ids:
            prev_row = previous_marks.get(sid)
            curr_row = current_marks.get(sid)
            prev_present = _row_has_meaningful_mark_data(prev_row)
            curr_present = _row_has_meaningful_mark_data(curr_row)
            if curr_present and not prev_present:
                recipients.add(str(sid))

    if not recipients:
        return 0

    template = str(getattr(cfg, 'row_filled_template', '') or '')
    if not template:
        template = (
            '📝 {course_code} - {course_name}\n'
            '{exam_name} marks row was filled by {faculty_name}.\n'
            'Student: {student_name} ({register_number})\n'
            'Mark: {mark}/{max_mark}'
        )

    course = getattr(getattr(exam_assignment, 'section', None), 'course', None)
    course_code = str(getattr(course, 'subject_code', '') or getattr(course, 'course_code', '') or '')
    course_name = str(getattr(course, 'subject_name', '') or getattr(course, 'course_name', '') or '')
    class_name = str(getattr(course, 'class_type_name', '') or getattr(course, 'class_name', '') or '')
    section_obj = getattr(exam_assignment, 'section', None)
    section_name = str(getattr(section_obj, 'name', '') or '')

    try:
        faculty_name = str(actor_user) if actor_user else ''
    except Exception:
        faculty_name = ''

    exam_name = str(getattr(exam_assignment, 'exam_display_name', '') or getattr(exam_assignment, 'exam', '') or getattr(exam_assignment, 'name', '') or 'Exam')
    max_mark = str(getattr(exam_assignment, 'max_marks', '') or '')

    try:
        ta = exam_assignment.section.teaching_assignment
        acad_sec = ta.section
        assignments = (
            StudentSectionAssignment.objects
            .filter(section=acad_sec, end_date__isnull=True)
            .select_related('student__user')
        )
    except Exception as e:
        logger.error(f'Failed to load student section assignments: {e}')
        return 0

    sent_count = 0
    for sa in assignments:
        sp = getattr(sa, 'student', None)
        if not sp:
            continue
        sid = str(getattr(sp, 'id', '') or '')
        if sid not in recipients:
            continue

        mobile = _resolve_mobile_for_student_profile(sp)
        if not mobile:
            logger.warning(f'No mobile for student {sid}')
            continue

        reg_no = str(getattr(sp, 'reg_no', '') or '')
        student_name = ''
        try:
            u = getattr(sp, 'user', None)
            student_name = str(u) if u else reg_no
        except Exception:
            student_name = reg_no

        row = (current_marks or {}).get(sid) if isinstance(current_marks, dict) else None
        mark = ''
        if isinstance(row, dict):
            mark = row.get('mark')

        msg = _render_notification_template(
            template,
            {
                'course_code': course_code,
                'course_name': course_name,
                'class_name': class_name,
                'section': section_name,
                'exam_name': exam_name,
                'max_mark': max_mark,
                'faculty_name': faculty_name,
                'student_name': student_name,
                'register_number': reg_no,
                'mark': mark if mark is not None else '',
            },
        )

        try:
            result = send_whatsapp(mobile, msg)
            if bool(getattr(result, 'ok', False)):
                sent_count += 1
                logger.info(f'Sent row-filled notification to {reg_no} ({mobile})')
            else:
                logger.error(f'Failed to send row-filled notification to {reg_no} ({mobile}): {getattr(result, "message", "unknown error")}')
        except Exception as e:
            logger.error(f'Failed to send row-filled notification to {reg_no} ({mobile}): {e}')
            continue

    logger.info(f'Row-filled notifications: sent {sent_count}/{len(recipients)} messages')
    return sent_count


def _send_student_publish_notifications(
    *,
    exam_assignment,
    actor_user,
    prev_published_marks: dict,
    new_published_marks: dict,
    was_published_before: bool,
):
    """Send WhatsApp notifications to students when marks are published."""
    import logging
    logger = logging.getLogger('academic_v2.notifications')

    try:
        cfg, _ = AcV2AcademicNotificationSetting.objects.get_or_create(key='DEFAULT')
    except Exception as e:
        logger.error(f'Failed to load notification settings: {e}')
        return

    if not bool(getattr(cfg, 'student_publish_enabled', False)):
        logger.debug('Student publish notifications disabled')
        return

    try:
        from academics.models import StudentSectionAssignment
        from accounts.services.sms import send_whatsapp
    except Exception:
        return

    # Determine changed students (for edited-only notifications)
    changed_ids: set[str] = set()
    if isinstance(prev_published_marks, dict) and isinstance(new_published_marks, dict):
        all_ids = set(str(k) for k in new_published_marks.keys()) | set(str(k) for k in prev_published_marks.keys())
        for sid in all_ids:
            a = prev_published_marks.get(sid)
            b = new_published_marks.get(sid)
            if a != b:
                changed_ids.add(str(sid))

    # Choose the base template for this publish (one message per student per click)
    base_template = ''
    base_recipients: set[str] = set()
    if not was_published_before and bool(getattr(cfg, 'notify_on_first_publish', True)):
        base_template = str(getattr(cfg, 'first_publish_template', '') or '')
        base_recipients = set(str(k) for k in (new_published_marks or {}).keys())
    elif bool(getattr(cfg, 'notify_on_every_publish_click', False)):
        base_template = str(getattr(cfg, 'every_publish_template', '') or '')
        base_recipients = set(str(k) for k in (new_published_marks or {}).keys())

    edited_template = ''
    edited_recipients: set[str] = set()
    if was_published_before and bool(getattr(cfg, 'notify_on_row_edits_only', True)):
        edited_template = str(getattr(cfg, 'edited_rows_template', '') or '')
        edited_recipients = set(changed_ids)

    # If edited-only is enabled and a base (every publish) is also enabled,
    # we replace the base message for edited students to avoid duplicates.
    recipients = set(base_recipients) | set(edited_recipients)
    if not recipients:
        logger.debug(f'No recipients selected for notifications (base={len(base_recipients)}, edited={len(edited_recipients)})')
        return
    
    logger.info(f'Sending notifications to {len(recipients)} students: first_publish={bool(base_recipients)}, edited_only={bool(edited_recipients)}')

    # Course/exam context
    course = getattr(getattr(exam_assignment, 'section', None), 'course', None)
    course_code = str(
        getattr(course, 'subject_code', '')
        or getattr(course, 'course_code', '')
        or ''
    )
    course_name = str(
        getattr(course, 'subject_name', '')
        or getattr(course, 'course_name', '')
        or ''
    )
    class_name = str(getattr(course, 'class_type_name', '') or getattr(course, 'class_name', '') or '')
    section_obj = getattr(exam_assignment, 'section', None)
    course = getattr(section_obj, 'course', None)
    section_name = str(getattr(section_obj, 'name', '') or '')

    try:
        faculty_name = str(actor_user) if actor_user else ''
    except Exception:
        faculty_name = ''

    exam_name = str(getattr(exam_assignment, 'exam_display_name', '') or getattr(exam_assignment, 'exam', '') or getattr(exam_assignment, 'name', '') or 'Exam')
    max_mark = str(getattr(exam_assignment, 'max_marks', '') or '')

    # Load all active students in the academic section.
    try:
        ta = exam_assignment.section.teaching_assignment
        acad_sec = ta.section
        assignments = (
            StudentSectionAssignment.objects
            .filter(section=acad_sec, end_date__isnull=True)
            .select_related('student__user')
        )
        logger.info(f'Found {assignments.count()} students in section for notifications')
    except Exception as e:
        logger.error(f'Failed to load student section assignments: {e}')
        return

    sent_count = 0
    for sa in assignments:
        sp = getattr(sa, 'student', None)
        if not sp:
            continue
        sid = str(getattr(sp, 'id', '') or '')
        if sid not in recipients:
            continue

        mobile = _resolve_mobile_for_student_profile(sp)
        if not mobile:
            logger.warning(f'No mobile for student {sid}')
            continue

        # Pick message template (edited overrides base)
        tpl = edited_template if sid in edited_recipients and edited_template else base_template
        if not tpl:
            logger.warning(f'No template selected for student {sid}')
            continue

        # Student context
        reg_no = str(getattr(sp, 'reg_no', '') or '')
        student_name = ''
        try:
            u = getattr(sp, 'user', None)
            student_name = str(u) if u else reg_no
        except Exception:
            student_name = reg_no

        row = (new_published_marks or {}).get(sid) if isinstance(new_published_marks, dict) else None
        mark = ''
        if isinstance(row, dict):
            mark = row.get('mark')
        msg = _render_notification_template(
            tpl,
            {
                'course_code': course_code,
                'course_name': course_name,
                'class_name': class_name,
                'section': section_name,
                'exam_name': exam_name,
                'max_mark': max_mark,
                'faculty_name': faculty_name,
                'student_name': student_name,
                'register_number': reg_no,
                'mark': mark if mark is not None else '',
            },
        )

        try:
            result = send_whatsapp(mobile, msg)
            if bool(getattr(result, 'ok', False)):
                sent_count += 1
                logger.info(f'Sent notification to {reg_no} ({mobile})')
            else:
                logger.error(f'Failed to send notification to {reg_no} ({mobile}): {getattr(result, "message", "unknown error")}')
        except Exception as e:
            logger.error(f'Failed to send notification to {reg_no} ({mobile}): {e}')
            continue
    
    logger.info(f'Publish notifications: sent {sent_count}/{len(recipients)} messages')


def _has_admin_bypass_access(user):
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    return bool(user.is_staff or user.is_superuser or user.has_perm('academic_v2.page.admin'))


@api_view(['GET', 'PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def admin_pass_mark_settings(request):
    """Get or update the global pass mark setting (singleton — auto-creates if missing)."""
    if not _has_admin_bypass_access(request.user):
        return Response({'detail': 'Permission denied'}, status=403)
    obj, _ = AcV2PassMarkSetting.objects.get_or_create(
        defaults={'out_of': 100, 'pass_mark': 50, 'label': 'Default'},
        label='Default',
    )
    if request.method == 'GET':
        return Response(AcV2PassMarkSettingSerializer(obj).data)
    serializer = AcV2PassMarkSettingSerializer(obj, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


@api_view(['GET', 'PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def admin_academic_notification_settings(request):
    """Get or update Academic 2.1 notification settings (singleton)."""

    if not _has_admin_bypass_access(request.user):
        return Response({'detail': 'Permission denied'}, status=403)

    obj, _ = AcV2AcademicNotificationSetting.objects.get_or_create(
        key='DEFAULT',
        defaults={
            'student_publish_enabled': False,
            'notify_on_first_publish': True,
            'notify_on_row_edits_only': True,
            'notify_on_every_publish_click': False,
            'cqi_announce_enabled': False,
        },
    )

    if request.method == 'GET':
        return Response(AcV2AcademicNotificationSettingSerializer(obj).data)

    serializer = AcV2AcademicNotificationSettingSerializer(obj, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


@api_view(['GET', 'PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def admin_my_marks_settings(request):
    """Get or update My Marks settings (singleton)."""
    if not _has_admin_bypass_access(request.user):
        return Response({'detail': 'Permission denied'}, status=403)

    obj, _ = AcV2MyMarksSetting.objects.get_or_create(
        key='DEFAULT',
        defaults={
            'viewing_enabled': False,
            'require_profile_photo': False,
            'require_mobile_number': False,
        },
    )

    if request.method == 'GET':
        return Response(AcV2MyMarksSettingSerializer(obj).data)

    serializer = AcV2MyMarksSettingSerializer(obj, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


@api_view(['GET', 'PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def admin_publish_settings(request):
    """Get or update Publish settings (singleton). GET is open to all auth users."""
    obj, _ = AcV2PublishSetting.objects.get_or_create(
        key='DEFAULT',
        defaults={'must_fill_all_cells': False, 'publish_progress_duration': 4},
    )
    if request.method == 'GET':
        return Response(AcV2PublishSettingSerializer(obj).data)
    if not _has_admin_bypass_access(request.user):
        return Response({'detail': 'Permission denied'}, status=403)
    serializer = AcV2PublishSettingSerializer(obj, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


def _to_float(value, default=0.0):
    try:
        if value is None:
            return float(default)
        return float(value)
    except Exception:
        return float(default)


def _get_student_profile_for_request(request):
    from academics.models import StudentProfile

    return StudentProfile.objects.select_related('user').filter(user=request.user).first()


def _student_can_access_ta_ids(student_profile):
    from academics.models import StudentSectionAssignment, TeachingAssignment

    section_ids = list(
        StudentSectionAssignment.objects.filter(
            student=student_profile,
            end_date__isnull=True,
        ).values_list('section_id', flat=True)
    )
    if not section_ids:
        return []

    return list(
        TeachingAssignment.objects.filter(
            is_active=True,
            section_id__in=section_ids,
        ).values_list('id', flat=True)
    )


def _course_header_from_section(section):
    course = getattr(section, 'course', None)
    faculty = getattr(section, 'faculty_user', None)
    return {
        'course_code': getattr(course, 'subject_code', '') or '',
        'course_name': getattr(course, 'subject_name', '') or '',
        'class_type': getattr(course, 'class_type_name', '') or '',
        'faculty_name': str(faculty) if faculty else '',
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_my_marks_config(request):
    """Student-facing My Marks access flags and profile requirement status."""
    obj, _ = AcV2MyMarksSetting.objects.get_or_create(
        key='DEFAULT',
        defaults={
            'viewing_enabled': False,
            'require_profile_photo': False,
            'require_mobile_number': False,
        },
    )

    sp = _get_student_profile_for_request(request)
    has_profile_photo = bool(getattr(sp, 'profile_image', None)) if sp else False
    has_mobile_number = bool(str(getattr(sp, 'mobile_number', '') or '').strip()) if sp else False

    return Response({
        'viewing_enabled': bool(obj.viewing_enabled),
        'require_profile_photo': bool(obj.require_profile_photo),
        'require_mobile_number': bool(obj.require_mobile_number),
        'has_profile_photo': has_profile_photo,
        'has_mobile_number': has_mobile_number,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_my_courses(request):
    """Student-facing list of assigned courses with entered-weight aggregates."""
    sp = _get_student_profile_for_request(request)
    if sp is None:
        return Response({'detail': 'Student profile not found.'}, status=404)

    allowed_ta_ids = _student_can_access_ta_ids(sp)
    if not allowed_ta_ids:
        return Response({'courses': []})

    sections = (
        AcV2Section.objects
        .select_related('course', 'faculty_user')
        .filter(teaching_assignment_id__in=allowed_ta_ids)
    )

    exams = list(
        AcV2ExamAssignment.objects.filter(section_id__in=[s.id for s in sections]).select_related('section__course')
    )
    exam_ids = [e.id for e in exams]
    marks_map = {
        str(sm.exam_assignment_id): sm
        for sm in AcV2StudentMark.objects.filter(student=sp, exam_assignment_id__in=exam_ids)
    }

    exams_by_section = {}
    for ea in exams:
        exams_by_section.setdefault(str(ea.section_id), []).append(ea)

    courses = []
    for sec in sections:
        header = _course_header_from_section(sec)
        sec_exams = exams_by_section.get(str(sec.id), [])

        exams_entered = 0
        total_obtained_weight = 0.0
        total_entered_weight = 0.0

        for ea in sec_exams:
            sm = marks_map.get(str(ea.id))
            has_payload = sm is not None
            if not has_payload:
                continue

            exams_entered += 1
            weight = _to_float(getattr(ea, 'weight', 0))
            total_entered_weight += weight

            max_marks = _to_float(getattr(ea, 'max_marks', 0))
            obtained = None if sm.total_mark is None else _to_float(sm.total_mark)
            if (not sm.is_absent) and obtained is not None and max_marks > 0:
                total_obtained_weight += (obtained / max_marks) * weight

        entered_weight_pct = None
        if total_entered_weight > 0:
            entered_weight_pct = (total_obtained_weight / total_entered_weight) * 100.0

        from academics.models import StudentSectionAssignment
        ta = getattr(sec, 'teaching_assignment', None)
        ta_section_id = getattr(ta, 'section_id', None) if ta else None
        
        sec_student_ids = []
        if ta_section_id:
            sec_student_ids = list(
                StudentSectionAssignment.objects.filter(
                    section_id=ta_section_id, end_date__isnull=True
                )
                .values_list('student_id', flat=True)
                .distinct()
            )
        sec_exam_ids = [e.id for e in sec_exams]
        sec_students = _compute_students_entered_weight_pct(sec_student_ids, sec_exam_ids)
        top_students = sec_students[:5]

        courses.append({
            'ta_id': sec.teaching_assignment_id,
            'course_code': header['course_code'],
            'course_name': header['course_name'],
            'class_type': header['class_type'],
            'faculty_name': header['faculty_name'],
            'exams_entered': exams_entered,
            'obtained_weight': round(total_obtained_weight, 2),
            'max_weight': round(total_entered_weight, 2),
            'entered_weight_pct': round(entered_weight_pct, 2) if entered_weight_pct is not None else None,
            'top_students': top_students,
        })

    courses.sort(key=lambda c: (str(c.get('course_code') or ''), str(c.get('course_name') or '')))
    return Response({'courses': courses})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_my_course_detail(request, ta_id: int):
    """Student-facing course detail grouped by cycle with exam-level rows."""
    sp = _get_student_profile_for_request(request)
    if sp is None:
        return Response({'detail': 'Student profile not found.'}, status=404)

    allowed_ta_ids = set(_student_can_access_ta_ids(sp))
    if int(ta_id) not in allowed_ta_ids:
        return Response({'detail': 'Course not found.'}, status=404)

    section = get_object_or_404(
        AcV2Section.objects.select_related('course', 'faculty_user'),
        teaching_assignment_id=ta_id,
    )

    exams = list(
        AcV2ExamAssignment.objects.filter(section=section).order_by('exam_display_name', 'exam')
    )
    marks_map = {
        str(sm.exam_assignment_id): sm
        for sm in AcV2StudentMark.objects.filter(student=sp, exam_assignment_id__in=[e.id for e in exams])
    }

    cycle_buckets = {}
    for ea in exams:
        cycle_state = _get_exam_cycle_state(
            ea,
            semester_id=getattr(getattr(section, 'course', None), 'semester_id', None),
            class_type=getattr(getattr(section, 'course', None), 'class_type', None),
        )

        cid = str(cycle_state.get('cycle_id') or 'uncategorized')
        cname = cycle_state.get('cycle_name') or 'General'
        cycle_obj = AcV2Cycle.objects.filter(id=cycle_state.get('cycle_id')).first() if cycle_state.get('cycle_id') else None
        cdesc = getattr(cycle_obj, 'description', '') if cycle_obj else ''
        corder = getattr(cycle_obj, 'order', 999) if cycle_obj else 999

        bucket = cycle_buckets.setdefault(cid, {
            'cycle_id': cid,
            'cycle_name': cname,
            'cycle_desc': cdesc or '',
            'cycle_order': int(corder),
            'entered_exam_count': 0,
            'total_obtained_weight': 0.0,
            'total_entered_weight': 0.0,
            'entered_weight_pct': None,
            'exams': [],
        })

        sm = marks_map.get(str(ea.id))
        has_payload = sm is not None
        is_absent = bool(sm.is_absent) if sm else False
        total_mark = None if (sm is None or sm.total_mark is None) else _to_float(sm.total_mark)

        max_marks = _to_float(getattr(ea, 'max_marks', 0))
        weight = _to_float(getattr(ea, 'weight', 0))
        obtained_weight = None
        if has_payload:
            bucket['entered_exam_count'] += 1
            bucket['total_entered_weight'] += weight
            if (not is_absent) and total_mark is not None and max_marks > 0:
                obtained_weight = (total_mark / max_marks) * weight
                bucket['total_obtained_weight'] += obtained_weight

        bucket['exams'].append({
            'exam_id': str(ea.id),
            'exam': ea.exam,
            'exam_display': ea.exam_display_name or ea.exam,
            'qp_type': ea.qp_type or '',
            'max_marks': round(max_marks, 2),
            'weight': round(weight, 2),
            'total_mark': round(total_mark, 2) if total_mark is not None else None,
            'obtained_weight': round(obtained_weight, 2) if obtained_weight is not None else None,
            'has_payload': has_payload,
            'is_absent': is_absent,
            'published_at': ea.published_at.isoformat() if ea.published_at else None,
        })

    cycles = []
    for _cid, bucket in cycle_buckets.items():
        if bucket['total_entered_weight'] > 0:
            bucket['entered_weight_pct'] = round((bucket['total_obtained_weight'] / bucket['total_entered_weight']) * 100.0, 2)
        bucket['total_obtained_weight'] = round(bucket['total_obtained_weight'], 2)
        bucket['total_entered_weight'] = round(bucket['total_entered_weight'], 2)
        bucket['exams'].sort(key=lambda x: (x.get('exam_display') or x.get('exam') or ''))
        cycles.append(bucket)

    cycles.sort(key=lambda c: (int(c.get('cycle_order') or 999), str(c.get('cycle_name') or '')))
    header = _course_header_from_section(section)

    return Response({
        'ta_id': int(ta_id),
        'course_code': header['course_code'],
        'course_name': header['course_name'],
        'class_type': header['class_type'],
        'faculty_name': header['faculty_name'],
        'cycles': cycles,
    })




def _compute_students_entered_weight_pct(student_ids, exam_ids):
    try:
        from academics.models import StudentProfile
        from .models import AcV2StudentMark, AcV2ExamAssignment, AcV2InternalMark
        
        profiles = {
            sp.id: sp 
            for sp in StudentProfile.objects.filter(id__in=student_ids).select_related('user')
        }
        
        if not exam_ids or not student_ids:
            return []
        # Group exams by section_id
        exams = list(AcV2ExamAssignment.objects.filter(id__in=exam_ids))
        exams_by_section = {}
        for ea in exams:
            exams_by_section.setdefault(str(ea.section_id), []).append(ea)

        # For each student, we want to calculate (obtained_weight, max_weight) per section
        student_section_weights = {} # student_id -> {section_id_str: (obtained, max)}

        class MockUser:
            is_superuser = True
            is_staff = True
            def is_authenticated(self):
                return True

        class MockRequest:
            user = MockUser()
            method = 'GET'
            query_params = {}

        for sec_id_str, sec_exams in exams_by_section.items():
            # Compute actual total weight (max weight) for this section
            # matching the computedTotalWeight useMemo in InternalMarkPage.tsx
            try:
                sec = AcV2Section.objects.select_related('teaching_assignment__section', 'course__class_type').get(id=sec_id_str)
                class_type = sec.course.class_type
                total_internal = float(class_type.total_internal_marks) if class_type else 40
            except Exception:
                sec = None
                class_type = None
                total_internal = 40

            # Build config lookup map from ClassType exam configs
            config_map = {}
            if class_type and isinstance(class_type.exam_assignments, list):
                for cfg in class_type.exam_assignments:
                    if not isinstance(cfg, dict):
                        continue
                    for key in [cfg.get('exam_display_name'), cfg.get('exam')]:
                        if key:
                            k = re.sub(r'[^a-z0-9]+', '', str(key or '').strip().lower())
                            config_map[k] = cfg

            total_weight = 0.0
            for ex in sec_exams:
                if str(getattr(ex, 'kind', '') or 'exam').lower() == 'cqi':
                    continue

                # Match exam assignment to its ClassType config
                ex_name = re.sub(r'[^a-z0-9]+', '', str(getattr(ex, 'exam_display_name', '') or getattr(ex, 'exam', '') or '').strip().lower())
                cfg = config_map.get(ex_name, {})
                
                # Determine covered COs
                co_count = 5
                try:
                    co_count = ex.section.course.co_count or 5
                except Exception:
                    co_count = 5
                co_numbers = list(range(1, co_count + 1))
                covered_cos = ex.covered_cos if isinstance(ex.covered_cos, list) else co_numbers
                if not covered_cos:
                    covered_cos = co_numbers

                co_weights = cfg.get('co_weights') if isinstance(cfg.get('co_weights'), dict) else {}
                for co_num in covered_cos:
                    w = co_weights.get(str(co_num)) or co_weights.get(int(co_num)) or 0.0
                    total_weight += _to_float(w, 0.0)

                cia_enabled = bool(cfg.get('cia_enabled', False))
                cia_weight_val = cfg.get('cia_weight', None)
                if cia_enabled and cia_weight_val is not None:
                    cia_w = _to_float(cia_weight_val, 0.0)
                    if cia_w > 0:
                        n = len(covered_cos) or 1
                        per_co = bool(cfg.get('cia_weight_per_co', False))
                        effective_w = cia_w if per_co else round(cia_w / n, 2)
                        for _ in range(n):
                            total_weight += effective_w

            rounded_max_weight = round(total_weight, 2)
            sec_max_weight = rounded_max_weight if rounded_max_weight > 0 else float(total_internal)

            # Call the live CO summary endpoint dynamically using MockRequest to get the exact cap-adjusted final mark
            student_final_marks = {}
            if sec and getattr(sec, 'teaching_assignment_id', None):
                try:
                    co_summary_res = faculty_course_co_summary(MockRequest(), sec.teaching_assignment_id)
                    co_summary_data = co_summary_res.data if hasattr(co_summary_res, 'data') else {}
                    if isinstance(co_summary_data, dict) and isinstance(co_summary_data.get('students'), list):
                        for s_data in co_summary_data['students']:
                            student_final_marks[s_data['id']] = _to_float(s_data.get('final_mark'), 0.0)
                except Exception:
                    student_final_marks = {}

            # Query student marks for these exams (for fallback if not in CO summary)
            sec_exam_ids = [e.id for e in sec_exams]
            student_marks = {}
            for sm in AcV2StudentMark.objects.filter(student_id__in=student_ids, exam_assignment_id__in=sec_exam_ids):
                student_marks.setdefault(sm.student_id, []).append(sm)

            exams_map = {str(e.id): e for e in sec_exams}

            for sid in student_ids:
                # Try fetching from the live CO summary first
                if sid in student_final_marks:
                    f_mark = student_final_marks[sid]
                    if f_mark >= 0 and sec_max_weight > 0:
                        student_section_weights.setdefault(sid, {})[sec_id_str] = (f_mark, sec_max_weight)
                        continue

                # Otherwise, fallback to computing from exam weights
                sms = student_marks.get(sid, [])
                obtained_weight = 0.0
                max_weight = 0.0
                for sm in sms:
                    ea = exams_map.get(str(sm.exam_assignment_id))
                    if not ea:
                        continue
                    weight = _to_float(getattr(ea, 'weight', 0) or 0)
                    max_marks = _to_float(getattr(ea, 'max_marks', 0) or 0)
                    obtained = None if sm.total_mark is None else _to_float(sm.total_mark)
                    
                    if (not sm.is_absent) and obtained is not None and max_marks > 0:
                        obtained_weight += (obtained / max_marks) * weight
                        max_weight += weight
                
                if max_weight > 0:
                    # Normalize the obtained weight relative to the computed section max weight
                    normalized_obt = (obtained_weight / max_weight) * sec_max_weight
                    student_section_weights.setdefault(sid, {})[sec_id_str] = (normalized_obt, sec_max_weight)
    except Exception as e:
        import traceback
        with open('/home/iqac2/IDCS-Restart/backend/academic_v2/error.log', 'w') as f:
            f.write(traceback.format_exc())
        raise e

    results = []
    for sid in student_ids:
        prof = profiles.get(sid)
        name = "Student"
        if prof:
            try:
                if prof.user:
                    name = prof.user.first_name or prof.user.username or str(prof.user)
                else:
                    name = prof.reg_no or f"Student {sid}"
            except Exception:
                name = getattr(prof, 'reg_no', '') or f"Student {sid}"

        sec_weights = student_section_weights.get(sid, {})
        pcts = []
        for obt, mx in sec_weights.values():
            if mx > 0:
                pcts.append((obt / mx) * 100.0)

        avg_pct = round(float(sum(pcts) / len(pcts)), 2) if pcts else None
        
        results.append({
            'student_id': sid,
            'student_name': name,
            'average_pct': avg_pct,
        })
        
    # Sort descending
    results.sort(key=lambda x: (x['average_pct'] is None, -float(x['average_pct'] or 0.0)))
    
    # Assign ranks
    for idx, item in enumerate(results):
        item['rank'] = idx + 1
        
    return results


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_my_class_leaderboard(request):
    from academics.models import StudentSectionAssignment, TeachingAssignment

    sp = _get_student_profile_for_request(request)
    if sp is None:
        return Response({'detail': 'Student profile not found.'}, status=404)

    section_ids = list(
        StudentSectionAssignment.objects.filter(student=sp, end_date__isnull=True)
        .values_list('section_id', flat=True)
    )
    if not section_ids:
        return Response({
            'top_students': [],
            'current_student': None,
            'course_count': 0,
        })

    ta_ids = list(
        TeachingAssignment.objects.filter(is_active=True, section_id__in=section_ids)
        .values_list('id', flat=True)
    )
    
    acv2_section_ids = list(
        AcV2Section.objects.filter(teaching_assignment_id__in=ta_ids)
        .values_list('id', flat=True)
    )

    exam_ids = list(
        AcV2ExamAssignment.objects.filter(section_id__in=acv2_section_ids)
        .values_list('id', flat=True)
    )

    student_ids = list(
        StudentSectionAssignment.objects.filter(section_id__in=section_ids, end_date__isnull=True)
        .values_list('student_id', flat=True)
        .distinct()
    )

    students = _compute_students_entered_weight_pct(student_ids, exam_ids)
    current_student = next((item for item in students if item['student_id'] == sp.id), None)

    return Response({
        'top_students': students[:10],
        'current_student': current_student,
        'course_count': len(ta_ids),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_my_course_top_students(request, ta_id: int):
    from academics.models import StudentSectionAssignment

    sp = _get_student_profile_for_request(request)
    if sp is None:
        return Response({'detail': 'Student profile not found.'}, status=404)

    allowed_ta_ids = set(_student_can_access_ta_ids(sp))
    if int(ta_id) not in allowed_ta_ids:
        return Response({'detail': 'Course not found.'}, status=404)

    section = get_object_or_404(
        AcV2Section.objects.select_related('course', 'teaching_assignment'),
        teaching_assignment_id=ta_id,
    )
    header = _course_header_from_section(section)

    ta = getattr(section, 'teaching_assignment', None)
    ta_section_id = getattr(ta, 'section_id', None) if ta else None

    student_ids = []
    if ta_section_id:
        student_ids = list(
            StudentSectionAssignment.objects.filter(
                section_id=ta_section_id, end_date__isnull=True
            )
            .values_list('student_id', flat=True)
            .distinct()
        )

    exam_ids = list(
        AcV2ExamAssignment.objects.filter(section=section)
        .values_list('id', flat=True)
    )

    students = _compute_students_entered_weight_pct(student_ids, exam_ids)
    current_student = next((item for item in students if item['student_id'] == sp.id), None)

    return Response({
        'course_code': header['course_code'],
        'course_name': header['course_name'],
        'top_students': students[:5],
        'current_student': current_student,
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_secure_delete(request):
    """Admin-only secure delete with password confirmation.

    Supports soft-delete for:
    - object_type = 'qp_type' (AcV2QpType)
    - object_type = 'qp_pattern' (AcV2QpPattern)  [includes exam templates where class_type is null]
    """
    if not _has_admin_bypass_access(request.user):
        return Response({'detail': 'Permission denied'}, status=403)

    password = str(request.data.get('password', '') or '')
    if not password or not request.user.check_password(password):
        return Response({'detail': 'Incorrect password. Action cancelled.'}, status=400)

    object_type = str(request.data.get('object_type', '') or '').strip()
    object_id = str(request.data.get('id', '') or '').strip()
    if not object_type or not object_id:
        return Response({'detail': 'object_type and id are required'}, status=400)

    if object_type == 'qp_type':
        obj = get_object_or_404(AcV2QpType, pk=object_id)
        with transaction.atomic():
            obj.is_active = False
            obj.updated_by = request.user
            obj.save(update_fields=['is_active', 'updated_by', 'updated_at'])
            # Deactivate any assignments using this QP type.
            AcV2QpAssignment.objects.filter(qp_type=obj, is_active=True).update(is_active=False, updated_by=request.user)
        return Response({'success': True}, status=200)

    if object_type == 'qp_pattern':
        obj = get_object_or_404(AcV2QpPattern, pk=object_id)
        with transaction.atomic():
            obj.is_active = False
            obj.updated_by = request.user
            obj.save(update_fields=['is_active', 'updated_by', 'updated_at'])
            # Deactivate any QP assignment rows that point at this pattern.
            AcV2QpAssignment.objects.filter(exam_assignment=obj, is_active=True).update(is_active=False, updated_by=request.user)
        return Response({'success': True}, status=200)

    return Response({'detail': 'Unsupported object_type'}, status=400)
from .services.publish_control import check_publish_control, create_edit_request, process_auto_publish, check_cqi_publish_control, create_cqi_edit_request
from .services.mark_calculation import compute_section_internal_marks


PUBLISHED_EXAM_STATUSES = {'PUBLISHED', 'APPROVED', 'LOCKED'}


def _resolve_qp_pattern_for_exam_assignment(exam_assignment, class_type=None):
    qp_type = str(getattr(exam_assignment, 'qp_type', '') or '').strip()
    exam_display_name = str(getattr(exam_assignment, 'exam_display_name', '') or '').strip()
    exam_code = str(getattr(exam_assignment, 'exam', '') or '').strip()
    exam_names = []
    for value in [exam_display_name, exam_code]:
        if value and value not in exam_names:
            exam_names.append(value)

    qp_type_candidates = []
    for value in [qp_type, exam_code, re.sub(r'\s+', '_', exam_display_name).upper() if exam_display_name else '']:
        normalized = str(value or '').strip()
        if normalized and normalized not in qp_type_candidates:
            qp_type_candidates.append(normalized)

    if not qp_type_candidates:
        return None

    # Match using the same fallback order already used while syncing exam assignments:
    # class-specific by display name/code, global by display name/code, then qp_type-only.
    candidate_filters = []
    for qp_type_candidate in qp_type_candidates:
        for exam_name in exam_names:
            if class_type is not None:
                candidate_filters.append({'name__iexact': exam_name, 'qp_type__iexact': qp_type_candidate, 'class_type': class_type, 'is_active': True})
            candidate_filters.append({'name__iexact': exam_name, 'qp_type__iexact': qp_type_candidate, 'is_active': True})
        if class_type is not None:
            candidate_filters.append({'qp_type__iexact': qp_type_candidate, 'class_type': class_type, 'is_active': True})
        candidate_filters.append({'qp_type__iexact': qp_type_candidate, 'is_active': True})

    fallback_pattern = None
    for filters in candidate_filters:
        pattern = AcV2QpPattern.objects.filter(**filters).order_by('-updated_at').first()
        if pattern is None:
            continue
        if getattr(pattern, 'cycle_id', None):
            return pattern
        if fallback_pattern is None:
            fallback_pattern = pattern
    return fallback_pattern


def _get_exam_cycle_state(exam_assignment, semester_id=None, class_type=None):
    pattern = _resolve_qp_pattern_for_exam_assignment(exam_assignment, class_type=class_type)
    cycle = getattr(pattern, 'cycle', None)
    if cycle is None:
        return {
            'cycle_locked': False,
            'cycle_lock_reason': None,
            'cycle_id': None,
            'cycle_name': None,
            'cycle_code': None,
            'cycle_active': None,
            'semester_active': None,
        }

    semester_active = cycle.is_semester_active(semester_id)
    cycle_locked = not semester_active
    reason = None
    if cycle_locked:
        reason = (
            f'{cycle.name} is inactive for this semester.'
            if cycle.is_active else
            f'{cycle.name} is inactive for all semesters.'
        )

    return {
        'cycle_locked': cycle_locked,
        'cycle_lock_reason': reason,
        'cycle_id': str(cycle.id),
        'cycle_name': cycle.name,
        'cycle_code': cycle.code,
        'cycle_active': bool(cycle.is_active),
        'semester_active': semester_active,
    }


# ============================================================================
# SEMESTER CONFIG (Admin)
# ============================================================================

class AcV2SemesterConfigViewSet(viewsets.ModelViewSet):
    """
    Semester configuration CRUD.
    Admin only.
    """
    queryset = AcV2SemesterConfig.objects.all()
    serializer_class = AcV2SemesterConfigSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        semester_id = self.request.query_params.get('semester')
        if semester_id:
            qs = qs.filter(semester_id=semester_id)
        return qs.select_related('semester')
    
    def perform_create(self, serializer):
        serializer.save(updated_by=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)


class AcV2SemesterGroupViewSet(viewsets.ModelViewSet):
    """
    Semester publish-control group CRUD.
    """
    queryset = AcV2SemesterGroup.objects.all().prefetch_related('semesters')
    serializer_class = AcV2SemesterGroupSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
    
    @action(detail=True, methods=['POST'], permission_classes=[IsAuthenticated])
    def reset_requests(self, request, pk=None):
        """
        Admin action: Cancel all pending edit requests for this group.
        Returns count of cancelled requests.
        """
        try:
            group = self.get_object()
        except Exception as e:
            return Response(
                {'detail': f'Group not found: {str(e)}'},
                status=404
            )

        semester_ids = list(group.semesters.values_list('id', flat=True))
        if not semester_ids:
            return Response({'detail': 'Group has no semesters assigned.'}, status=400)
        
        # Verify user is staff/admin
        if not _has_admin_bypass_access(request.user):
            return Response({'detail': 'Permission denied'}, status=403)
        
        # Verify password
        password = request.data.get('password', '')
        if not password or not request.user.check_password(password):
            return Response(
                {'detail': 'Incorrect password. Action cancelled.'},
                status=400
            )
        
        try:
            exams = AcV2ExamAssignment.objects.filter(
                section__course__semester_id__in=semester_ids
            )
            print(f"[RESET_REQUESTS] Found {exams.count()} exams for group {group.id}")

            pending_qs = AcV2EditRequest.objects.filter(
                exam_assignment__in=exams,
                status__in=['PENDING', 'HOD_PENDING', 'IQAC_PENDING']
            )
            print(f"[RESET_REQUESTS] Found {pending_qs.count()} pending requests")

            affected_exam_ids = list(
                pending_qs.values_list('exam_assignment_id', flat=True).distinct()
            )
            print(f"[RESET_REQUESTS] Affected exams: {len(affected_exam_ids)}")

            with transaction.atomic():
                cancelled_count = pending_qs.update(status='REJECTED')
                print(f"[RESET_REQUESTS] Rejected {cancelled_count} requests")
                
                if affected_exam_ids:
                    flag_clear_count = AcV2ExamAssignment.objects.filter(
                        id__in=affected_exam_ids
                    ).update(has_pending_edit_request=False)
                    print(f"[RESET_REQUESTS] Cleared flags on {flag_clear_count} exams")

                reopened_count = AcV2ExamAssignment.objects.filter(
                    section__course__semester_id__in=semester_ids,
                    status__in=['PUBLISHED', 'LOCKED'],
                ).update(
                    status='DRAFT',
                    edit_window_until=None,
                    edit_window_until_publish=False,
                    has_pending_edit_request=False,
                    published_by=None,
                )
                print(f"[RESET_REQUESTS] Reopened {reopened_count} published exams (auto + manually published)")

            print(f"[RESET_REQUESTS] Transaction committed successfully")
            return Response({
                'status': 'success',
                'message': f'Cancelled {cancelled_count} pending edit requests and reopened {reopened_count} auto-published exams',
                'cancelled_count': cancelled_count,
                'reopened_count': reopened_count,
            }, status=200)
            
        except Exception as e:
            print(f"[RESET_REQUESTS] ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'detail': f'Reset failed: {str(e)}'},
                status=500
            )
    
    @action(detail=True, methods=['POST'], permission_classes=[IsAuthenticated])
    def reset_marks(self, request, pk=None):
        """
        Admin action: Clear all marks and revert all exams to DRAFT status for this group.
        Returns count of affected exams.
        """
        try:
            group = self.get_object()
        except Exception as e:
            return Response(
                {'detail': f'Group not found: {str(e)}'},
                status=404
            )

        semester_ids = list(group.semesters.values_list('id', flat=True))
        if not semester_ids:
            return Response({'detail': 'Group has no semesters assigned.'}, status=400)
        
        # Verify user is staff/admin
        if not _has_admin_bypass_access(request.user):
            return Response({'detail': 'Permission denied'}, status=403)
        
        # Verify password
        password = request.data.get('password', '')
        if not password or not request.user.check_password(password):
            return Response(
                {'detail': 'Incorrect password. Action cancelled.'},
                status=400
            )
        
        try:
            exams = AcV2ExamAssignment.objects.filter(
                section__course__semester_id__in=semester_ids
            )
            print(f"[RESET_MARKS] Found {exams.count()} exams for group {group.id}")
            
            affected_count = 0
            with transaction.atomic():
                for exam in exams:
                    exam.draft_data = {}
                    exam.published_data = {}
                    exam.status = 'DRAFT'
                    exam.edit_window_until = None
                    exam.edit_window_until_publish = False
                    exam.has_pending_edit_request = False
                    exam.save(update_fields=[
                        'draft_data', 'published_data', 'status', 
                        'edit_window_until', 'edit_window_until_publish', 
                        'has_pending_edit_request'
                    ])
                    
                    AcV2DraftMark.objects.filter(exam_assignment=exam).delete()
                    AcV2StudentMark.objects.filter(exam_assignment=exam).delete()
                    
                    AcV2EditRequest.objects.filter(
                        exam_assignment=exam,
                        status__in=['PENDING', 'HOD_PENDING', 'IQAC_PENDING']
                    ).update(status='REJECTED')
                    
                    affected_count += 1
            
            print(f"[RESET_MARKS] Transaction committed successfully, affected: {affected_count}")
            return Response({
                'status': 'success',
                'message': f'Reset marks for {affected_count} exam assignments',
                'affected_count': affected_count
            }, status=200)
            
        except Exception as e:
            print(f"[RESET_MARKS] ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'detail': f'Reset failed: {str(e)}'},
                status=500
            )


# ============================================================================
# CLASS TYPE (Admin)
# ============================================================================

class AcV2ClassTypeViewSet(viewsets.ModelViewSet):
    """
    Class type CRUD.
    Admin only.
    """
    queryset = AcV2ClassType.objects.filter(is_active=True)
    serializer_class = AcV2ClassTypeSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        college_id = self.request.query_params.get('college')
        if college_id:
            qs = qs.filter(college_id=college_id)
        return qs.order_by('name')
    
    def perform_create(self, serializer):
        serializer.save(updated_by=self.request.user)
    
    def perform_update(self, serializer):
        layout = self.request.data.get('coattainment_layout')
        instance = serializer.save(updated_by=self.request.user)

        if layout is not None and isinstance(layout, dict):
            vars_data = instance.cqi_global_custom_vars
            if not isinstance(vars_data, dict):
                vars_data = {'_items': vars_data} if isinstance(vars_data, list) else {}
            vars_data['_coattainment_layout'] = layout
            instance.cqi_global_custom_vars = vars_data
            instance.save(update_fields=['cqi_global_custom_vars'])

        # Keep CQI DB config in sync with the JSON-based admin editor.
        # Faculty views prefer AcV2CqiExam (DB) for stable CQI CO/condition symbols.
        try:
            from .models import AcV2CqiExam, AcV2QpType

            def _as_int_list(values):
                out = []
                if not isinstance(values, list):
                    return out
                for v in values:
                    try:
                        n = int(v)
                    except Exception:
                        continue
                    if n not in out:
                        out.append(n)
                return out

            exam_assignments = instance.exam_assignments if isinstance(instance.exam_assignments, list) else []
            global_custom_vars = instance.cqi_global_custom_vars if isinstance(instance.cqi_global_custom_vars, list) else []

            # Track current active CQI codes per qp_type
            active_codes_by_qp: dict[str, set[str]] = {}

            for idx, ea in enumerate(exam_assignments, start=1):
                if not isinstance(ea, dict):
                    continue
                if str(ea.get('kind') or '').strip().lower() != 'cqi':
                    continue

                qp_type_code = str(ea.get('qp_type') or '').strip()
                exam_code = str(ea.get('exam') or '').strip()
                exam_display_name = str(ea.get('exam_display_name') or exam_code or '').strip()
                enabled = bool(ea.get('enabled', True))
                if not qp_type_code or not exam_code:
                    continue

                active_codes_by_qp.setdefault(qp_type_code.lower(), set())
                if enabled:
                    active_codes_by_qp[qp_type_code.lower()].add(exam_code)

                cqi_sub = ea.get('cqi', {}) if isinstance(ea.get('cqi'), dict) else {}

                # Try to resolve qp_type FK (optional)
                qp_obj = (
                    AcV2QpType.objects.filter(code__iexact=qp_type_code).first()
                    if qp_type_code else None
                )

                defaults = {
                    'qp_type_code': qp_type_code,
                    'qp_type': qp_obj,
                    'exam_display_name': exam_display_name,
                    'order': int(ea.get('order') or idx or 0),
                    'is_active': enabled,
                    'cqi_name': str(cqi_sub.get('name') or ''),
                    'cqi_code': str(cqi_sub.get('code') or ''),
                    'cycle_id': str(cqi_sub.get('cycle_id') or ''),
                    'cos': _as_int_list(cqi_sub.get('cos') or []),
                    'considered_exams': cqi_sub.get('exams') if isinstance(cqi_sub.get('exams'), list) else [],
                    'custom_vars': cqi_sub.get('custom_vars') if isinstance(cqi_sub.get('custom_vars'), list) else [],
                    'global_custom_vars': global_custom_vars,
                    'derived_variables': cqi_sub.get('derived_variables') if isinstance(cqi_sub.get('derived_variables'), list) else [],
                    'co_value_expr': str(cqi_sub.get('co_value_expr') or ''),
                    'formula': str(cqi_sub.get('formula') or ''),
                    'conditions': cqi_sub.get('conditions') if isinstance(cqi_sub.get('conditions'), list) else [],
                    'else_formula': str(cqi_sub.get('else_formula') or ''),
                    'updated_by': self.request.user,
                }

                AcV2CqiExam.objects.update_or_create(
                    class_type=instance,
                    qp_type_code=qp_type_code,
                    exam_code=exam_code,
                    defaults=defaults,
                )

            # Deactivate CQI exams that were removed or disabled in the JSON config
            existing = AcV2CqiExam.objects.filter(class_type=instance)
            for row in existing:
                qp_key = str(getattr(row, 'qp_type_code', '') or '').strip().lower()
                code = str(getattr(row, 'exam_code', '') or '').strip()
                allowed = active_codes_by_qp.get(qp_key)
                if allowed is None:
                    # No CQI configured for this qp_type anymore
                    if row.is_active:
                        row.is_active = False
                        row.updated_by = self.request.user
                        row.save(update_fields=['is_active', 'updated_by', 'updated_at'])
                    continue
                if code and code not in allowed and row.is_active:
                    row.is_active = False
                    row.updated_by = self.request.user
                    row.save(update_fields=['is_active', 'updated_by', 'updated_at'])
        except Exception:
            # Never block saving class types due to CQI sync problems.
            pass
    
    def perform_destroy(self, instance):
        # Soft delete
        instance.is_active = False
        instance.save()


# ============================================================================
# QP TYPE (Admin - Master data for exam types)
# ============================================================================

# ============================================================================
# ACADEMIC CYCLE
# ============================================================================

class AcV2CycleViewSet(viewsets.ModelViewSet):
    """
    Academic Cycle CRUD.
    List, create, update, and soft-delete academic cycles.
    """
    queryset = AcV2Cycle.objects.all()
    serializer_class = AcV2CycleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        college_id = self.request.query_params.get('college')
        if college_id:
            qs = qs.filter(college_id=college_id)
        return qs.order_by('order', 'name')

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()

    @action(detail=False, methods=['post'], url_path='reorder')
    def reorder(self, request):
        """Accept [{id, order}, ...] and bulk-update cycle display order."""
        items = request.data if isinstance(request.data, list) else []
        if not items:
            return Response({'detail': 'Expected a list of {id, order} objects.'}, status=400)
        from django.db import transaction
        with transaction.atomic():
            for item in items:
                AcV2Cycle.objects.filter(pk=item.get('id')).update(order=item.get('order', 0))
        return Response({'detail': 'Reordered successfully.'})


# ============================================================================
# QP TYPE (Admin)
# ============================================================================

# ============================================================================
# CQI TOKEN REGISTRY (Admin CRUD + Faculty read)
# ============================================================================

class AcV2CqiTokenViewSet(viewsets.ModelViewSet):
    """
    CQI Token registry.

    GET  (list)  — returns active tokens scoped to the requesting college,
                   plus all global (college=None) tokens.  Supports filters:
                     ?class_type=<uuid>   — include tokens for a specific class type
                     ?condition_only=1    — only tokens with available_in_condition=True
                     ?formula_only=1      — only tokens with available_in_formula=True

    POST / PATCH / DELETE — admin only.  System tokens (is_system=True) cannot
                             be deleted via the API.
    """
    serializer_class = AcV2CqiTokenSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        params = self.request.query_params

        # Base: active tokens
        qs = AcV2CqiToken.objects.filter(is_active=True)

        # Scope to college (global + this college)
        college_id = params.get('college')
        if college_id:
            qs = qs.filter(
                models.Q(college__isnull=True) | models.Q(college_id=college_id)
            )
        else:
            # Default: return global + any college the user belongs to
            qs = qs.filter(college__isnull=True)

        # Optional class-type scope
        class_type_id = params.get('class_type')
        if class_type_id:
            qs = qs.filter(
                models.Q(class_type__isnull=True) | models.Q(class_type_id=class_type_id)
            )
        else:
            qs = qs.filter(class_type__isnull=True)

        # Optional filter: available_in_condition
        if params.get('condition_only') in ('1', 'true', 'yes'):
            qs = qs.filter(available_in_condition=True)

        # Optional filter: available_in_formula
        if params.get('formula_only') in ('1', 'true', 'yes'):
            qs = qs.filter(available_in_formula=True)

        return qs.order_by('order', 'category', 'code')

    def perform_destroy(self, instance):
        if instance.is_system:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('System tokens cannot be deleted.')
        # Soft-delete
        instance.is_active = False
        instance.save()


class AcV2CqiOperatorViewSet(viewsets.ReadOnlyModelViewSet):
    """
    CQI Operator list.  Read-only for all authenticated users.
    Operators are seeded by migration and managed in Django admin.
    """
    queryset = AcV2CqiOperator.objects.filter(is_active=True).order_by('order')
    serializer_class = AcV2CqiOperatorSerializer
    permission_classes = [IsAuthenticated]


# ============================================================================
# QP TYPE (Admin)
# ============================================================================

class AcV2QpTypeViewSet(viewsets.ModelViewSet):
    """
    QP Type (Question Paper Type) CRUD.
    Master data for exam types (SSA, CIA, MODEL, LAB, etc.)
    Admin only.
    """
    queryset = AcV2QpType.objects.filter(is_active=True)
    serializer_class = AcV2QpTypeSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        college_id = self.request.query_params.get('college')
        if college_id:
            qs = qs.filter(college_id=college_id)
        return qs.order_by('name')
    
    def perform_create(self, serializer):
        serializer.save(updated_by=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
    
    def perform_destroy(self, instance):
        # Soft delete
        instance.is_active = False
        instance.save()


class AcV2CourseOutcomeViewSet(viewsets.ModelViewSet):
    """
    Course Outcome CRUD.
    Admin can add/manage CO numbers used by QP editor dropdowns.
    """
    queryset = AcV2CourseOutcome.objects.filter(is_active=True)
    serializer_class = AcV2CourseOutcomeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        college_id = self.request.query_params.get('college')
        if college_id:
            qs = qs.filter(college_id=college_id)
        return qs.order_by('display_order', 'number')

    def perform_create(self, serializer):
        serializer.save(updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def perform_destroy(self, instance):
        # Soft delete
        instance.is_active = False
        instance.save(update_fields=['is_active', 'updated_at'])


# ============================================================================
# QP PATTERN (Admin - Table Creator)
# ============================================================================

class AcV2QpPatternViewSet(viewsets.ModelViewSet):
    """
    QP Pattern CRUD.
    Admin only.
    """
    queryset = AcV2QpPattern.objects.filter(is_active=True)
    serializer_class = AcV2QpPatternSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        qp_type = self.request.query_params.get('qp_type')
        class_type_id = self.request.query_params.get('class_type')
        batch_id = self.request.query_params.get('batch')
        
        if qp_type:
            qs = qs.filter(qp_type=qp_type)
        if class_type_id:
            qs = qs.filter(class_type_id=class_type_id)
        if batch_id:
            qs = qs.filter(batch_id=batch_id)
        
        return qs.select_related('class_type', 'batch')
    
    def perform_create(self, serializer):
        qp_pattern = serializer.save(updated_by=self.request.user)
        self._sync_qp_assignment_from_pattern(qp_pattern)
    
    def perform_update(self, serializer):
        qp_pattern = serializer.save(updated_by=self.request.user)
        self._sync_qp_assignment_from_pattern(qp_pattern)

    def _sync_qp_assignment_from_pattern(self, qp_pattern: AcV2QpPattern):
        """Create/update a QP Assignment row that mirrors the saved QP Pattern."""
        if not qp_pattern or not qp_pattern.is_active:
            return
        if not qp_pattern.class_type_id:
            # Only class-scoped patterns are part of the admin mapping.
            return

        qp_type_obj = (
            AcV2QpType.objects.filter(
                is_active=True,
                code__iexact=(qp_pattern.qp_type or '').strip(),
            )
            .filter(Q(class_type_id=qp_pattern.class_type_id) | Q(class_type__isnull=True))
            .order_by('-class_type_id')
            .first()
        )
        if not qp_type_obj:
            return

        pattern = qp_pattern.pattern or {}
        titles = pattern.get('titles') or []
        marks = pattern.get('marks') or []
        btls = pattern.get('btls') or []
        cos = pattern.get('cos') or []
        enabled = pattern.get('enabled')
        if enabled is None:
            enabled = [True] * len(titles)

        max_len = max(len(titles), len(marks), len(btls), len(cos), len(enabled))
        question_table = []
        for i in range(max_len):
            question_table.append({
                'index': i,
                'title': titles[i] if i < len(titles) else f'Q{i + 1}',
                'max_marks': marks[i] if i < len(marks) else 0,
                'btl_level': btls[i] if i < len(btls) else None,
                'co_number': cos[i] if i < len(cos) else None,
                'enabled': enabled[i] if i < len(enabled) else True,
            })

        defaults = {
            'is_active': True,
            'config': {
                'qp_pattern_id': str(qp_pattern.id),
                'pattern': pattern,
            },
            'question_table': question_table,
            'updated_by': self.request.user,
        }

        AcV2QpAssignment.objects.update_or_create(
            class_type=qp_pattern.class_type,
            qp_type=qp_type_obj,
            exam_assignment=qp_pattern,
            defaults=defaults,
        )


# ============================================================================
# COURSE / SECTION
# ============================================================================

class AcV2CourseViewSet(viewsets.ModelViewSet):
    queryset = AcV2Course.objects.all()
    serializer_class = AcV2CourseSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        semester_id = self.request.query_params.get('semester')
        if semester_id:
            qs = qs.filter(semester_id=semester_id)
        return qs.select_related('class_type', 'semester', 'subject')


class AcV2SectionViewSet(viewsets.ModelViewSet):
    queryset = AcV2Section.objects.all()
    serializer_class = AcV2SectionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        course_id = self.request.query_params.get('course')
        faculty_id = self.request.query_params.get('faculty')
        
        if course_id:
            qs = qs.filter(course_id=course_id)
        if faculty_id:
            qs = qs.filter(faculty_user_id=faculty_id)
        
        return qs.select_related('course', 'faculty_user', 'teaching_assignment')


# ============================================================================
# EXAM ASSIGNMENT
# ============================================================================

class AcV2ExamAssignmentViewSet(viewsets.ModelViewSet):
    queryset = AcV2ExamAssignment.objects.all()
    serializer_class = AcV2ExamAssignmentSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        section_id = self.request.query_params.get('section')
        course_id = self.request.query_params.get('course')
        status_filter = self.request.query_params.get('status')
        
        if section_id:
            qs = qs.filter(section_id=section_id)
        if course_id:
            qs = qs.filter(section__course_id=course_id)
        if status_filter:
            qs = qs.filter(status=status_filter)
        
        return qs.select_related('section__course__class_type')
    
    @action(detail=True, methods=['post'])
    def save_marks(self, request, pk=None):
        """Save draft marks."""
        exam = self.get_object()
        
        if not _has_admin_bypass_access(request.user) and not exam.is_editable():
            return Response(
                {'error': 'This exam is locked and cannot be edited.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        marks_data = request.data.get('marks', {})
        question_btls = request.data.get('question_btls', None)
        marks_map = _save_draft_marks_for_exam(exam, marks_data)

        # Preserve existing draft_data structure; update marks and optional btls
        draft = exam.draft_data if isinstance(exam.draft_data, dict) else {}
        draft['marks'] = marks_map
        if question_btls is not None:
            draft['question_btls'] = question_btls
        
        with transaction.atomic():
            exam.draft_data = draft
            exam.last_saved_at = timezone.now()
            exam.last_saved_by = request.user
            exam.save(update_fields=['draft_data', 'last_saved_at', 'last_saved_by'])
        
        return Response({
            'success': True,
            'last_saved_at': exam.last_saved_at.isoformat(),
        })
    
    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        """Publish marks."""
        exam = self.get_object()
        
        if not _has_admin_bypass_access(request.user) and not exam.is_editable():
            return Response(
                {'error': 'This exam is locked and cannot be edited.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        semester_config = exam.get_semester_config()
        
        with transaction.atomic():
            exam.published_data = exam.draft_data
            exam.published_at = timezone.now()
            exam.published_by = request.user
            
            # Set status based on publish control
            if semester_config and semester_config.publish_control_enabled:
                exam.status = 'PUBLISHED'
            else:
                exam.status = 'DRAFT'  # Keep as draft if no publish control
            
            exam.save()

            # Keep materialized student marks in sync with published snapshot.
            published_marks = exam.published_data.get('marks', {}) if isinstance(exam.published_data, dict) else {}
            if isinstance(published_marks, dict) and published_marks:
                _materialize_student_marks_from_map(exam, published_marks)
            
            # Recompute internal marks
            compute_section_internal_marks(exam.section)
        
        return Response({
            'success': True,
            'status': exam.status,
            'published_at': exam.published_at.isoformat(),
        })
    
    @action(detail=True, methods=['post'])
    def request_edit(self, request, pk=None):
        """Request edit access for published exam."""
        exam = self.get_object()
        reason = request.data.get('reason', '')
        
        if not reason:
            return Response(
                {'error': 'Reason is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        result = create_edit_request(exam, request.user, reason)
        
        if result['success']:
            return Response(result)
        else:
            return Response(result, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def reset_marks(self, request, pk=None):
        """Reset all marks for this exam."""
        exam = self.get_object()
        
        if not exam.is_editable():
            return Response(
                {'error': 'This exam is locked and cannot be edited.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        with transaction.atomic():
            exam.draft_data = {}
            exam.last_saved_at = timezone.now()
            exam.last_saved_by = request.user
            exam.save(update_fields=['draft_data', 'last_saved_at', 'last_saved_by'])
            
            # Delete student marks
            AcV2StudentMark.objects.filter(exam_assignment=exam).delete()
        
        return Response({'success': True})


# ============================================================================
# STUDENT MARKS
# ============================================================================

class AcV2StudentMarkViewSet(viewsets.ModelViewSet):
    queryset = AcV2StudentMark.objects.all()
    serializer_class = AcV2StudentMarkSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        exam_id = self.request.query_params.get('exam_assignment')
        if exam_id:
            qs = qs.filter(exam_assignment_id=exam_id)
        return qs.select_related('exam_assignment', 'student')
    
    @action(detail=False, methods=['post'])
    def bulk_save(self, request):
        """Bulk save student marks."""
        exam_id = request.data.get('exam_assignment')
        marks_list = request.data.get('marks', [])
        
        if not exam_id:
            return Response(
                {'error': 'exam_assignment is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        exam = get_object_or_404(AcV2ExamAssignment, id=exam_id)
        
        if not exam.is_editable():
            return Response(
                {'error': 'This exam is locked.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        qp_pattern = exam.get_qp_pattern()
        
        with transaction.atomic():
            for mark_data in marks_list:
                student_id = mark_data.get('student_id')
                if not student_id:
                    continue
                
                defaults = {
                    'reg_no': mark_data.get('reg_no', ''),
                    'student_name': mark_data.get('student_name', ''),
                    'question_marks': mark_data.get('question_marks', {}),
                    'is_absent': mark_data.get('is_absent', False),
                    'is_exempted': mark_data.get('is_exempted', False),
                    'remarks': mark_data.get('remarks', ''),
                }
                
                sm, created = AcV2StudentMark.objects.update_or_create(
                    exam_assignment=exam,
                    student_id=student_id,
                    defaults=defaults
                )
                
                # Calculate CO marks and total
                sm.calculate_co_marks(qp_pattern)
                sm.calculate_total()
                sm.save()
            
            # Update draft data on exam
            exam.last_saved_at = timezone.now()
            exam.last_saved_by = request.user
            exam.save(update_fields=['last_saved_at', 'last_saved_by'])

        spreadsheet_id = str(request.data.get('spreadsheet_id') or request.data.get('spreadsheetId') or '').strip()
        if spreadsheet_id:
            try:
                sync_marks_to_google_sheet(
                    exam_assignment=exam,
                    spreadsheet_id=spreadsheet_id,
                    config=request.data.get('config') or {},
                    sheet_name=str(request.data.get('sheet_name') or request.data.get('sheetName') or '').strip() or None,
                )
            except GoogleSheetsServiceError:
                pass
        
        return Response({'success': True, 'count': len(marks_list)})


# ============================================================================
# EDIT REQUESTS
# ============================================================================

class AcV2EditRequestViewSet(viewsets.ModelViewSet):
    queryset = AcV2EditRequest.objects.all()
    serializer_class = AcV2EditRequestSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        status_filter = self.request.query_params.get('status')
        requested_by = self.request.query_params.get('requested_by')
        
        if status_filter:
            qs = qs.filter(status=status_filter)

        user = getattr(self.request, 'user', None)

        if requested_by:
            # Non-admin users can only query their own requests
            if not self._has_any_role(user, ['ADMIN']) and not getattr(user, 'is_superuser', False):
                if str(requested_by) != str(getattr(user, 'id', '')):
                    return qs.none()
            qs = qs.filter(requested_by_id=requested_by)

        qs = qs.select_related(
            'exam_assignment__section__course',
            'exam_assignment__section__teaching_assignment__section__batch__department',
            'exam_assignment__section__teaching_assignment__section__managing_department',
            'requested_by',
        )

        # Inbox forwarding: only show items for the *current* approver stage.
        # This is computed from (semester_config.approval_workflow + current_stage).
        if self.action == 'list' and not requested_by:
            if not user:
                return qs.none()

            # Only pending-ish items belong in approval inboxes.
            qs = qs.filter(status__in=['PENDING', 'HOD_PENDING', 'IQAC_PENDING'])

            allowed_ids: list[int] = []
            for er in qs:
                required_role = self._current_required_role(er)
                if not required_role:
                    continue
                # Enforce exact role membership (do NOT treat superuser as all roles)
                if self._has_role_exact(user, required_role):
                    allowed_ids.append(er.id)

            if not allowed_ids:
                return qs.none()
            return qs.filter(id__in=allowed_ids)

        return qs

    def _has_role_exact(self, user, role_name: str) -> bool:
        """Check role membership without superuser override (used for inbox gating)."""
        if not user:
            return False
        role_u = str(role_name or '').strip()
        if not role_u:
            return False
        try:
            if hasattr(user, 'roles'):
                return user.roles.filter(name__iexact=role_u).exists()
        except Exception:
            pass
        try:
            if hasattr(user, 'user_roles'):
                return user.user_roles.filter(role__name__iexact=role_u).exists()
        except Exception:
            pass
        return False

    def _current_required_role(self, edit_request) -> str | None:
        wf_roles = self._workflow_roles(edit_request)
        if not wf_roles:
            return None
        stage_index = max(0, int(getattr(edit_request, 'current_stage', 1) or 1) - 1)
        if stage_index < len(wf_roles):
            return wf_roles[stage_index]
        return wf_roles[-1]

    def _has_any_role(self, user, role_names: list[str]) -> bool:
        if not user:
            return False
        if getattr(user, 'is_superuser', False):
            return True
        wanted = [str(r).strip() for r in (role_names or []) if str(r).strip()]
        if not wanted:
            return False
        try:
            if hasattr(user, 'roles'):
                q = Q()
                for r in wanted:
                    q |= Q(name__iexact=str(r))
                return user.roles.filter(q).exists()
        except Exception:
            pass
        try:
            if hasattr(user, 'user_roles'):
                q = Q()
                for r in wanted:
                    q |= Q(role__name__iexact=str(r))
                return user.user_roles.filter(q).exists()
        except Exception:
            pass
        return False

    def _workflow_roles(self, edit_request) -> list[str]:
        cfg = None
        try:
            cfg = edit_request.exam_assignment.get_semester_config()
        except Exception:
            cfg = None
        wf = getattr(cfg, 'approval_workflow', None) if cfg else None
        roles: list[str] = []
        raw = wf or []
        for item in raw:
            if isinstance(item, str):
                role = item
            elif isinstance(item, dict):
                role = item.get('role')
            else:
                role = None
            role_u = str(role or '').strip().upper()
            if role_u and role_u not in roles:
                roles.append(role_u)
        return roles
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve edit request."""
        edit_request = self.get_object()
        notes = request.data.get('notes', '')

        wf_roles = self._workflow_roles(edit_request)
        if wf_roles:
            required_role = self._current_required_role(edit_request)
            if not required_role:
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)
            if not (getattr(request.user, 'is_superuser', False) or self._has_role_exact(request.user, required_role)):
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)
        else:
            # Fallback legacy behavior when no workflow is configured
            if not self._has_any_role(request.user, ['HOD', 'IQAC', 'ADMIN']):
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)

        if edit_request.status not in ['PENDING', 'HOD_PENDING', 'IQAC_PENDING']:
            return Response(
                {'error': 'This request cannot be approved.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        now = timezone.now()
        history = edit_request.approval_history or []

        # If workflow is configured, enforce stage order strictly (no skipping).
        if wf_roles:
            stage_index = max(0, int(getattr(edit_request, 'current_stage', 1) or 1) - 1)
            required_role = wf_roles[stage_index] if stage_index < len(wf_roles) else wf_roles[-1]

            # Determine acting role for this request/stage
            acting_role = None
            if getattr(request.user, 'is_superuser', False) or self._has_role_exact(request.user, required_role):
                acting_role = required_role

            if acting_role != required_role:
                return Response({'detail': f'Awaiting {required_role} approval.'}, status=status.HTTP_400_BAD_REQUEST)

            history.append({
                'stage': int(getattr(edit_request, 'current_stage', 1) or 1),
                'role': required_role,
                'user_id': getattr(request.user, 'id', None),
                'user_name': str(request.user),
                'action': 'APPROVED',
                'at': now.isoformat(),
                'notes': notes,
            })

            # If there is a next stage, move to it
            if stage_index + 1 < len(wf_roles):
                next_role = wf_roles[stage_index + 1]
                edit_request.current_stage = stage_index + 2
                if next_role == 'HOD':
                    edit_request.status = 'HOD_PENDING'
                elif next_role == 'IQAC':
                    edit_request.status = 'IQAC_PENDING'
                else:
                    edit_request.status = 'PENDING'
                edit_request.approval_history = history
                edit_request.save(update_fields=['current_stage', 'status', 'approval_history'])
                return Response({'success': True, 'status': edit_request.status, 'current_stage': edit_request.current_stage})

        # Final approve (no workflow or last stage)
        cfg = None
        try:
            cfg = edit_request.exam_assignment.get_semester_config()
        except Exception:
            cfg = None

        approval_until_publish = bool(getattr(cfg, 'approval_until_publish', False)) if cfg else False
        try:
            window_minutes = int(request.data.get('window_minutes') or (getattr(cfg, 'approval_window_minutes', 120) if cfg else 120))
        except Exception:
            window_minutes = int(getattr(cfg, 'approval_window_minutes', 120) if cfg else 120)

        # Persist approval history before final approve
        edit_request.approval_history = history
        edit_request.save(update_fields=['approval_history'])

        if approval_until_publish:
            # Grant unlimited edit until next publish
            with transaction.atomic():
                edit_request.status = 'APPROVED'
                edit_request.reviewed_by = request.user
                edit_request.reviewed_at = now
                edit_request.approved_until = None
                edit_request.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'approved_until'])

                ea = edit_request.exam_assignment
                ea.edit_window_until = None
                ea.edit_window_until_publish = True
                ea.has_pending_edit_request = False
                ea.save(update_fields=['edit_window_until', 'edit_window_until_publish', 'has_pending_edit_request'])

            return Response({'success': True, 'status': edit_request.status, 'approved_until': None, 'edit_mode': 'UNTIL_PUBLISH'})

        edit_request.approve(request.user, window_minutes, notes)
        return Response({'success': True, 'status': edit_request.status, 'approved_until': edit_request.approved_until.isoformat() if edit_request.approved_until else None})
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject edit request."""
        edit_request = self.get_object()
        reason = request.data.get('reason', '')

        wf_roles = self._workflow_roles(edit_request)
        if wf_roles:
            required_role = self._current_required_role(edit_request)
            if not required_role:
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)
            if not (getattr(request.user, 'is_superuser', False) or self._has_role_exact(request.user, required_role)):
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)
        else:
            if not self._has_any_role(request.user, ['HOD', 'IQAC', 'ADMIN']):
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)
        
        if edit_request.status not in ['PENDING', 'HOD_PENDING', 'IQAC_PENDING']:
            return Response(
                {'error': 'This request cannot be rejected.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # If workflow is configured, enforce stage order strictly (no skipping).
        if wf_roles:
            stage_index = max(0, int(getattr(edit_request, 'current_stage', 1) or 1) - 1)
            required_role = wf_roles[stage_index] if stage_index < len(wf_roles) else wf_roles[-1]

            # Determine acting role for this request/stage
            acting_role = None
            if getattr(request.user, 'is_superuser', False) or self._has_role_exact(request.user, required_role):
                acting_role = required_role
            if acting_role != required_role:
                return Response({'detail': f'Awaiting {required_role} action.'}, status=status.HTTP_400_BAD_REQUEST)

        # Add role into history as well
        now = timezone.now()
        stage_index = max(0, int(getattr(edit_request, 'current_stage', 1) or 1) - 1)
        required_role = wf_roles[stage_index] if (wf_roles and stage_index < len(wf_roles)) else None
        history = edit_request.approval_history or []
        history.append({
            'stage': int(getattr(edit_request, 'current_stage', 1) or 1),
            'role': required_role,
            'user_id': getattr(request.user, 'id', None),
            'user_name': str(request.user),
            'action': 'REJECTED',
            'at': now.isoformat(),
            'reason': reason,
        })
        edit_request.approval_history = history
        edit_request.save(update_fields=['approval_history'])

        edit_request.reject(request.user, reason)
        return Response({'success': True, 'status': edit_request.status})

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel an edit request (requester only)."""
        edit_request = self.get_object()

        # Only requester (or superuser) can cancel
        if not getattr(request.user, 'is_superuser', False):
            if getattr(edit_request, 'requested_by_id', None) != getattr(request.user, 'id', None):
                return Response({'detail': 'Only the requester can cancel this request.'}, status=status.HTTP_403_FORBIDDEN)

        if edit_request.status not in ['PENDING', 'HOD_PENDING', 'IQAC_PENDING']:
            return Response({'error': 'This request cannot be cancelled.'}, status=status.HTTP_400_BAD_REQUEST)

        now = timezone.now()
        history = edit_request.approval_history or []
        history.append({
            'stage': int(getattr(edit_request, 'current_stage', 1) or 1),
            'role': None,
            'user_id': getattr(request.user, 'id', None),
            'user_name': str(request.user),
            'action': 'CANCELLED',
            'at': now.isoformat(),
        })

        with transaction.atomic():
            edit_request.status = 'CANCELLED'
            edit_request.approval_history = history
            edit_request.save(update_fields=['status', 'approval_history'])

            ea = edit_request.exam_assignment
            ea.has_pending_edit_request = False
            ea.save(update_fields=['has_pending_edit_request'])

        return Response({'success': True, 'status': edit_request.status})


# ============================================================================
# CQI EDIT REQUESTS
# ============================================================================

class AcV2CqiEditRequestViewSet(viewsets.ModelViewSet):
    """ViewSet for CQI edit requests."""
    queryset = AcV2CqiEditRequest.objects.all()
    serializer_class = AcV2CqiEditRequestSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        status_filter = self.request.query_params.get('status')
        requested_by = self.request.query_params.get('requested_by')
        
        if status_filter:
            qs = qs.filter(status=status_filter)

        user = getattr(self.request, 'user', None)

        qs = qs.select_related(
            'cqi_attained__teaching_assignment__staff',
            'cqi_attained__teaching_assignment__course',
            'requested_by',
        )

        if requested_by:
            # Non-admin users can only query their own requests
            if not (_has_admin_bypass_access(user) or getattr(user, 'is_superuser', False)):
                if str(requested_by) != str(getattr(user, 'id', '')):
                    return qs.none()
            qs = qs.filter(requested_by_id=requested_by)
            return qs

        # Inbox gating: only show items for the *current* approver stage
        if self.action == 'list':
            if not user:
                return qs.none()

            # Only pending-ish items belong in approval inboxes
            qs = qs.filter(status__in=['PENDING', 'HOD_PENDING', 'IQAC_PENDING'])

            # Admin inbox should be able to see all pending CQI requests.
            # (Approval is still permission-checked in approve/reject actions.)
            if _has_admin_bypass_access(user) or getattr(user, 'is_superuser', False):
                return qs

            allowed_ids: list = []
            for er in qs:
                required_role = self._current_required_role(er)
                if not required_role:
                    continue
                # Enforce exact role membership (do NOT treat superuser as all roles)
                if self._has_role_exact(user, required_role):
                    allowed_ids.append(er.id)

            if not allowed_ids:
                return qs.none()
            return qs.filter(id__in=allowed_ids)

        return qs

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve CQI edit request."""
        edit_request = self.get_object()
        notes = request.data.get('notes', '')

        wf_roles = self._workflow_roles(edit_request)
        if wf_roles:
            required_role = self._current_required_role(edit_request)
            if not required_role:
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)
            if not (
                getattr(request.user, 'is_superuser', False)
                or _has_admin_bypass_access(request.user)
                or self._has_role_exact(request.user, required_role)
            ):
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)
        else:
            # Fallback legacy behavior when no workflow is configured
            if not (_has_admin_bypass_access(request.user) or self._has_any_role(request.user, ['HOD', 'IQAC', 'ADMIN'])):
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)

        if edit_request.status not in ['PENDING', 'HOD_PENDING', 'IQAC_PENDING']:
            return Response({'error': 'This request cannot be approved.'}, status=status.HTTP_400_BAD_REQUEST)

        now = timezone.now()
        history = edit_request.approval_history or []

        # If workflow is configured, enforce stage order strictly (no skipping).
        if wf_roles:
            stage_index = max(0, int(getattr(edit_request, 'current_stage', 1) or 1) - 1)
            required_role = wf_roles[stage_index] if stage_index < len(wf_roles) else wf_roles[-1]

            acting_role = None
            if (
                getattr(request.user, 'is_superuser', False)
                or _has_admin_bypass_access(request.user)
                or self._has_role_exact(request.user, required_role)
            ):
                acting_role = required_role

            if acting_role != required_role:
                return Response({'detail': f'Awaiting {required_role} approval.'}, status=status.HTTP_400_BAD_REQUEST)

            history.append({
                'stage': int(getattr(edit_request, 'current_stage', 1) or 1),
                'role': required_role,
                'user_id': getattr(request.user, 'id', None),
                'user_name': str(request.user),
                'action': 'APPROVED',
                'at': now.isoformat(),
                'notes': notes,
            })

            # If there is a next stage, move to it
            if stage_index + 1 < len(wf_roles):
                next_role = wf_roles[stage_index + 1]
                edit_request.current_stage = stage_index + 2
                if next_role == 'HOD':
                    edit_request.status = 'HOD_PENDING'
                elif next_role == 'IQAC':
                    edit_request.status = 'IQAC_PENDING'
                else:
                    edit_request.status = 'PENDING'
                edit_request.approval_history = history
                edit_request.save(update_fields=['current_stage', 'status', 'approval_history'])
                return Response({'success': True, 'status': edit_request.status, 'current_stage': edit_request.current_stage})

        # Final approve
        try:
            window_minutes = int(request.data.get('window_minutes') or 120)
        except Exception:
            window_minutes = 120

        # Persist any history we collected (including final-stage role entry above)
        edit_request.approval_history = history
        edit_request.save(update_fields=['approval_history'])

        edit_request.approve(request.user, window_minutes, notes)
        return Response({'success': True, 'status': edit_request.status, 'approved_until': edit_request.approved_until.isoformat() if edit_request.approved_until else None})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject CQI edit request."""
        edit_request = self.get_object()
        reason = request.data.get('reason', '')

        wf_roles = self._workflow_roles(edit_request)
        if wf_roles:
            required_role = self._current_required_role(edit_request)
            if not required_role:
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)
            if not (
                getattr(request.user, 'is_superuser', False)
                or _has_admin_bypass_access(request.user)
                or self._has_role_exact(request.user, required_role)
            ):
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)
        else:
            if not (_has_admin_bypass_access(request.user) or self._has_any_role(request.user, ['HOD', 'IQAC', 'ADMIN'])):
                return Response({'detail': 'Not allowed.'}, status=status.HTTP_403_FORBIDDEN)

        if edit_request.status not in ['PENDING', 'HOD_PENDING', 'IQAC_PENDING']:
            return Response({'error': 'This request cannot be rejected.'}, status=status.HTTP_400_BAD_REQUEST)

        # If workflow is configured, enforce stage order strictly (no skipping).
        if wf_roles:
            stage_index = max(0, int(getattr(edit_request, 'current_stage', 1) or 1) - 1)
            required_role = wf_roles[stage_index] if stage_index < len(wf_roles) else wf_roles[-1]

            acting_role = None
            if (
                getattr(request.user, 'is_superuser', False)
                or _has_admin_bypass_access(request.user)
                or self._has_role_exact(request.user, required_role)
            ):
                acting_role = required_role
            if acting_role != required_role:
                return Response({'detail': f'Awaiting {required_role} action.'}, status=status.HTTP_400_BAD_REQUEST)

        now = timezone.now()
        stage_index = max(0, int(getattr(edit_request, 'current_stage', 1) or 1) - 1)
        required_role = wf_roles[stage_index] if (wf_roles and stage_index < len(wf_roles)) else None
        history = edit_request.approval_history or []
        history.append({
            'stage': int(getattr(edit_request, 'current_stage', 1) or 1),
            'role': required_role,
            'user_id': getattr(request.user, 'id', None),
            'user_name': str(request.user),
            'action': 'REJECTED',
            'at': now.isoformat(),
            'reason': reason,
        })
        edit_request.approval_history = history
        edit_request.save(update_fields=['approval_history'])

        edit_request.reject(request.user, reason)
        return Response({'success': True, 'status': edit_request.status})

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel a CQI edit request (requester only)."""
        edit_request = self.get_object()

        if not getattr(request.user, 'is_superuser', False) and not _has_admin_bypass_access(request.user):
            if getattr(edit_request, 'requested_by_id', None) != getattr(request.user, 'id', None):
                return Response({'detail': 'Only the requester can cancel this request.'}, status=status.HTTP_403_FORBIDDEN)

        if edit_request.status not in ['PENDING', 'HOD_PENDING', 'IQAC_PENDING']:
            return Response({'error': 'This request cannot be cancelled.'}, status=status.HTTP_400_BAD_REQUEST)

        now = timezone.now()
        history = edit_request.approval_history or []
        history.append({
            'stage': int(getattr(edit_request, 'current_stage', 1) or 1),
            'role': None,
            'user_id': getattr(request.user, 'id', None),
            'user_name': str(request.user),
            'action': 'CANCELLED',
            'at': now.isoformat(),
        })

        with transaction.atomic():
            edit_request.status = 'CANCELLED'
            edit_request.approval_history = history
            edit_request.save(update_fields=['status', 'approval_history'])

            try:
                attained = edit_request.cqi_attained
                attained.has_pending_edit_request = False
                attained.save(update_fields=['has_pending_edit_request'])
            except Exception:
                pass

        return Response({'success': True, 'status': edit_request.status})

    def _has_any_role(self, user, role_names: list[str]) -> bool:
        if not user:
            return False
        if getattr(user, 'is_superuser', False):
            return True
        wanted = [str(r).strip() for r in (role_names or []) if str(r).strip()]
        if not wanted:
            return False
        try:
            if hasattr(user, 'roles'):
                q = Q()
                for r in wanted:
                    q |= Q(name__iexact=str(r))
                return user.roles.filter(q).exists()
        except Exception:
            pass
        try:
            if hasattr(user, 'user_roles'):
                q = Q()
                for r in wanted:
                    q |= Q(role__name__iexact=str(r))
                return user.user_roles.filter(q).exists()
        except Exception:
            pass
        return False

    def _has_role_exact(self, user, role_name: str) -> bool:
        """Check role membership without superuser override (used for inbox gating)."""
        if not user:
            return False
        role_u = str(role_name or '').strip()
        if not role_u:
            return False
        try:
            if hasattr(user, 'roles'):
                return user.roles.filter(name__iexact=role_u).exists()
        except Exception:
            pass
        try:
            if hasattr(user, 'user_roles'):
                return user.user_roles.filter(role__name__iexact=role_u).exists()
        except Exception:
            pass
        return False

    def _current_required_role(self, cqi_edit_request) -> str | None:
        wf_roles = self._workflow_roles(cqi_edit_request)
        if not wf_roles:
            return None
        stage_index = max(0, int(getattr(cqi_edit_request, 'current_stage', 1) or 1) - 1)
        if stage_index < len(wf_roles):
            return wf_roles[stage_index]
        return wf_roles[-1] if wf_roles else None

    def _workflow_roles(self, cqi_edit_request) -> list[str]:
        cfg = None
        try:
            cfg = cqi_edit_request.cqi_attained.get_semester_config()
        except Exception:
            cfg = None
        wf = getattr(cfg, 'approval_workflow', None) if cfg else None
        roles: list[str] = []
        raw = wf or []
        for item in raw:
            if isinstance(item, str):
                role = item
            elif isinstance(item, dict):
                role = item.get('role')
            else:
                role = None
            role_u = str(role or '').strip().upper()
            if role_u and role_u not in roles:
                roles.append(role_u)
        return roles


# ============================================================================
# INTERNAL MARKS (Read-Only)
# ============================================================================

class AcV2InternalMarkViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AcV2InternalMark.objects.all()
    serializer_class = AcV2InternalMarkSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        section_id = self.request.query_params.get('section')
        course_id = self.request.query_params.get('course')
        
        if section_id:
            qs = qs.filter(section_id=section_id)
        if course_id:
            qs = qs.filter(section__course_id=course_id)
        
        return qs.select_related('section__course').order_by('reg_no')
    
    @action(detail=False, methods=['post'])
    def recompute(self, request):
        """Recompute internal marks for a section."""
        section_id = request.data.get('section')
        if not section_id:
            return Response(
                {'error': 'section is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        section = get_object_or_404(AcV2Section, id=section_id)
        results = compute_section_internal_marks(section)
        
        return Response({
            'success': True,
            'count': len(results),
        })


# ============================================================================
# USER PATTERN OVERRIDE
# ============================================================================

class AcV2UserPatternOverrideViewSet(viewsets.ModelViewSet):
    queryset = AcV2UserPatternOverride.objects.all()
    serializer_class = AcV2UserPatternOverrideSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        qs = super().get_queryset()
        course_id = self.request.query_params.get('course')
        exam_type = self.request.query_params.get('exam_type')
        
        # Only show user's own overrides
        qs = qs.filter(created_by=self.request.user)
        
        if course_id:
            qs = qs.filter(course_id=course_id)
        if exam_type:
            qs = qs.filter(exam_type=exam_type)
        
        return qs
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


# ============================================================================
# HELPER ENDPOINTS
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def course_internal_summary(request, course_id):
    """
    Get internal mark summary for a course.
    Shows class type, exam assignments, CO coverage, weight matrix.
    """
    course = get_object_or_404(AcV2Course, id=course_id)
    
    # Get class type info
    class_type = course.class_type
    class_type_data = None
    if class_type:
        class_type_data = AcV2ClassTypeSerializer(class_type).data
    
    # Get all sections for this course
    sections = course.sections.all()
    
    # Get exam assignments
    exam_assignments = []
    for section in sections:
        for ea in section.exam_assignments.all():
            exam_assignments.append({
                'id': str(ea.id),
                'exam': ea.exam,
                'exam_display_name': ea.exam_display_name,
                'qp_type': ea.qp_type,
                'weight': float(ea.weight),
                'covered_cos': ea.covered_cos,
                'status': ea.status,
                'section_id': str(section.id),
                'section_name': section.section_name,
            })
    
    # Build CO coverage matrix
    co_coverage = {}
    for ea in exam_assignments:
        for co in ea['covered_cos']:
            co_key = f"CO{co}"
            if co_key not in co_coverage:
                co_coverage[co_key] = []
            co_coverage[co_key].append({
                'exam': ea['exam'],
                'weight': ea['weight'],
            })
    
    # Build weight matrix
    weight_matrix = {}
    exam_types = set(ea['qp_type'] for ea in exam_assignments)
    for et in exam_types:
        weight_matrix[et] = {}
        for co in range(1, course.co_count + 1):
            co_key = f"CO{co}"
            total_weight = sum(
                ea['weight'] / len(ea['covered_cos'])
                for ea in exam_assignments
                if ea['qp_type'] == et and co in ea['covered_cos']
            )
            weight_matrix[et][co_key] = round(total_weight, 2)
    
    return Response({
        'course': AcV2CourseSerializer(course).data,
        'class_type': class_type_data,
        'exam_assignments': exam_assignments,
        'co_coverage': co_coverage,
        'weight_matrix': weight_matrix,
        'total_internal_marks': float(class_type.total_internal_marks) if class_type else 40,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_pattern_for_exam(request, course_id, exam_type):
    """
    Get resolved QP pattern for an exam.
    Follows priority: User Override → Batch Override → Global Pattern
    """
    course = get_object_or_404(AcV2Course, id=course_id)
    user = request.user
    
    # 1. Check user override (if allow_customize_questions)
    if course.class_type and course.class_type.allow_customize_questions:
        user_pattern = AcV2UserPatternOverride.objects.filter(
            course=course,
            exam_type=exam_type,
            created_by=user
        ).first()
        if user_pattern:
            return Response({
                'source': 'user_override',
                'pattern': user_pattern.pattern,
            })
    
    # 2. Check batch override
    # (Need to determine batch from course/semester - simplified here)
    
    # 3. Global pattern
    pattern = AcV2QpPattern.objects.filter(
        qp_type=exam_type,
        class_type=course.class_type,
        is_active=True
    ).first()
    
    if not pattern:
        # Fallback to global without class type
        pattern = AcV2QpPattern.objects.filter(
            qp_type=exam_type,
            class_type__isnull=True,
            is_active=True
        ).first()
    
    if pattern:
        return Response({
            'source': 'global',
            'pattern': pattern.pattern,
        })
    
    return Response({
        'source': 'none',
        'pattern': {},
    })


# ============================================================================
# FACULTY COURSE INFO (for InternalMarkPage)
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_course_info(request, ta_id):
    """Return course information for a teaching assignment, including exam
    assignments configured for the class type. Used by the faculty
    InternalMarkPage. Admins (is_staff) can access any TA by passing bypass_session_id."""
    from academics.models import TeachingAssignment, StudentSectionAssignment

    is_admin_bypass = _has_admin_bypass_access(request.user)

    # Also allow access if the user accessed a bypass session for this TA via share link
    has_share_bypass = not is_admin_bypass and AcV2BypassSession.objects.filter(
        teaching_assignment_id=ta_id,
    ).filter(
        Q(faculty_user=request.user) | Q(shared_accessed_by=request.user)
    ).exists()

    ta_qs = TeachingAssignment.objects.select_related(
        'curriculum_row', 'elective_subject', 'section',
        'section__semester', 'section__managing_department',
        'staff',
    )
    if is_admin_bypass or has_share_bypass:
        ta = get_object_or_404(ta_qs, id=ta_id, is_active=True)
    else:
        ta = get_object_or_404(ta_qs, id=ta_id, staff__user=request.user, is_active=True)

    cr = ta.curriculum_row
    es = ta.elective_subject
    sec = ta.section

    course_code = (cr.course_code if cr else None) or (getattr(es, 'course_code', None)) or '-'
    course_name = (cr.course_name if cr else None) or (getattr(es, 'course_name', None)) or '-'
    class_type_code = (cr.class_type if cr else None) or 'THEORY'

    # QP type: from curriculum row (authoritative) or elective subject
    qp_type_code = (
        getattr(cr, 'question_paper_type', None)
        or getattr(es, 'question_paper_type', None)
        or ''
    ).strip()
    curriculum_qp_type_code = qp_type_code

    # Look up AcV2ClassType - curriculum class_type_code is authoritative.
    # Do NOT fall back to an arbitrary class type; that would hide setup issues
    # and can create wrong exam assignments.
    acv2_ct = (
        AcV2ClassType.objects.filter(is_active=True, short_code__iexact=class_type_code).first()
        or AcV2ClassType.objects.filter(is_active=True, name__iexact=class_type_code).first()
    )
    # NOTE: Always show curriculum class_type_code for consistency with course list
    # Use AcV2ClassType only for exam configurations and total_internal_marks
    class_type_info = {
        'id': str(acv2_ct.id) if acv2_ct else '',
        'name': class_type_code,  # Display curriculum code, not AcV2 display_name
        'total_internal_marks': float(acv2_ct.total_internal_marks) if acv2_ct else 40,
    }

    # Count students in this section
    student_count = 0
    if sec:
        student_count = StudentSectionAssignment.objects.filter(
            section=sec, end_date__isnull=True
        ).count()

    # Build exam list from AcV2ExamAssignment records linked to this TA via AcV2Section
    exams = []
    acv2_sections = AcV2Section.objects.filter(teaching_assignment=ta)
    # Filter by qp_type if set
    ea_qs = AcV2ExamAssignment.objects.filter(section__in=acv2_sections).select_related('section')
    if qp_type_code:
        ea_qs = ea_qs.filter(qp_type__iexact=qp_type_code)
    exam_assignments = list(ea_qs)

    # If curriculum does not have qp_type, infer it from existing exam assignments.
    # This allows faculty pages to show exam components even when curriculum setup is incomplete.
    if not qp_type_code:
        for _ea in exam_assignments:
            _t = (getattr(_ea, 'qp_type', '') or '').strip()
            if _t:
                qp_type_code = _t
                break

    # ClassType exam_assignments filtered to this qp_type (normalized)
    # These configs are the single source of truth for what faculty should see.
    ct_ea_configs = []
    if acv2_ct and acv2_ct.exam_assignments:
        qp_norm = (qp_type_code or '').strip().lower()
        for ea in acv2_ct.exam_assignments:
            if not isinstance(ea, dict):
                continue
            if ea.get('enabled', True) is False:
                continue
            ea_qp = (ea.get('qp_type', '') or '').strip().lower()
            if not qp_norm or ea_qp == qp_norm:
                ct_ea_configs.append(ea)

        # Enforce saved order if provided
        ct_ea_configs.sort(key=lambda x: (
            (x.get('order') if isinstance(x.get('order'), int) else 10**9),
            str(x.get('exam_display_name') or x.get('exam') or ''),
        ))

    # If the class type has no configs for the course qp_type, derive the exam list
    # from QP patterns (class_type + qp_type). IMPORTANT: Do not derive when a
    # ClassType config exists, otherwise removed/stale test patterns can reappear
    # in faculty pages.
    derived_ea_configs = []
    if acv2_ct and qp_type_code and len(ct_ea_configs) == 0:
        patterns_qs = AcV2QpPattern.objects.filter(
            is_active=True,
            qp_type__iexact=qp_type_code,
            class_type=acv2_ct,
        ).order_by('order', 'created_at')
        if not patterns_qs.exists():
            patterns_qs = AcV2QpPattern.objects.filter(
                is_active=True,
                qp_type__iexact=qp_type_code,
                class_type__isnull=True,
            ).order_by('order', 'created_at')

        # index existing weights by exam_display_name (case-insensitive)
        weight_by_name = {}
        if acv2_ct.exam_assignments:
            for ea in acv2_ct.exam_assignments:
                nm = (ea.get('exam_display_name') or ea.get('exam') or '').strip().lower()
                if nm and nm not in weight_by_name:
                    weight_by_name[nm] = ea

        for p in patterns_qs:
            exam_name = (p.name or '').strip()
            if not exam_name:
                continue
            # CQI exams must be configured via the CQI editor (ClassType exam_assignments)
            # and should never be auto-derived from QP patterns.
            if exam_name.strip().lower().startswith('cqi'):
                continue
            w_conf = weight_by_name.get(exam_name.lower())
            derived_ea_configs.append({
                'exam': exam_name,
                'exam_display_name': exam_name,
                'qp_type': qp_type_code,
                'weight': (w_conf.get('weight') if isinstance(w_conf, dict) else None) or float(p.default_weight or 0),
                'co_weights': (w_conf.get('co_weights') if isinstance(w_conf, dict) else None) or {},
                'default_cos': (w_conf.get('default_cos') if isinstance(w_conf, dict) else None) or [],
                'customize_questions': bool((w_conf.get('customize_questions') if isinstance(w_conf, dict) else None) or False),
            })

    def _ea_config_key(ea_conf):
        if not isinstance(ea_conf, dict):
            return ''
        return str(ea_conf.get('exam_display_name') or ea_conf.get('exam') or '').strip().lower()

    effective_ea_configs = []
    seen_ea_config_keys = set()
    for ea_conf in ct_ea_configs + derived_ea_configs:
        cfg_key = _ea_config_key(ea_conf)
        if cfg_key:
            if cfg_key in seen_ea_config_keys:
                continue
            seen_ea_config_keys.add(cfg_key)
        effective_ea_configs.append(ea_conf)

    # Build weight + order lookup from ClassType config (single source of truth)
    norm_exam_key = lambda s: re.sub(r'[^a-z0-9]+', '', str(s or '').strip().lower())

    # Filter out stale AcV2ExamAssignment records that are no longer part of the
    # effective qp_type exam configuration (i.e. deleted/removed in QP pattern).
    # Keep legacy behaviour when we have no effective configs to compare against.
    if effective_ea_configs:
        allowed_keys = set()
        for ea_conf in effective_ea_configs:
            if not isinstance(ea_conf, dict):
                continue
            for key in [ea_conf.get('exam_display_name'), ea_conf.get('exam')]:
                k = norm_exam_key(key)
                if k:
                    allowed_keys.add(k)
        exam_assignments = [
            ea for ea in exam_assignments
            if norm_exam_key(getattr(ea, 'exam_display_name', '') or getattr(ea, 'exam', '') or '') in allowed_keys
        ]

    # DB-backed CQI config lookup (authoritative for CQI CO selections)
    cqi_db_by_key = {}
    try:
        from .models import AcV2CqiExam
        if acv2_ct and qp_type_code:
            for row in AcV2CqiExam.objects.filter(
                class_type=acv2_ct,
                qp_type_code__iexact=qp_type_code,
                is_active=True,
            ).order_by('order', 'created_at'):
                for key in [row.exam_code, row.exam_display_name]:
                    k = norm_exam_key(key)
                    if k:
                        cqi_db_by_key[k] = row
    except Exception:
        cqi_db_by_key = {}
    ct_weight_lookup = {}
    ct_co_weights_lookup = {}  # exam -> {co: weight}
    ct_mm_enabled_lookup = {}
    ct_mm_with_exam_lookup = {}
    ct_mm_without_exam_lookup = {}
    ct_mm_exam_weight_lookup = {}
    ct_mm_exam_weight_per_co_lookup = {}
    ct_order = []
    ct_index = {}
    if effective_ea_configs:
        for i, ea_conf in enumerate(effective_ea_configs):
            exam_code = ea_conf.get('exam', '')
            exam_display = ea_conf.get('exam_display_name', exam_code)
            mm_enabled = bool(ea_conf.get('mark_manager_enabled', False))
            for key in [exam_code, exam_display]:
                k = norm_exam_key(key)
                if not k:
                    continue
                if k not in ct_index:
                    ct_index[k] = i
                    ct_order.append(k)
                ct_weight_lookup[k] = ea_conf.get('weight', 0)
                ct_mm_enabled_lookup[k] = mm_enabled
            co_weights = ea_conf.get('co_weights', {})
            if co_weights:
                w = {int(k): v for k, v in co_weights.items()}
                for key in [exam_code, exam_display]:
                    k = norm_exam_key(key)
                    if k:
                        ct_co_weights_lookup[k] = w

            # Mark Manager configs
            mm_on = ea_conf.get('mm_co_weights_with_exam')
            mm_off = ea_conf.get('mm_co_weights_without_exam')
            mm_exam_weight = ea_conf.get('mm_exam_weight')
            mm_exam_weight_per_co = ea_conf.get('mm_exam_weight_per_co', False)
            if isinstance(mm_on, dict) and mm_on:
                w_on = {int(k): v for k, v in mm_on.items()}
                for key in [exam_code, exam_display]:
                    k = norm_exam_key(key)
                    if k:
                        ct_mm_with_exam_lookup[k] = w_on
            if isinstance(mm_off, dict) and mm_off:
                w_off = {int(k): v for k, v in mm_off.items()}
                for key in [exam_code, exam_display]:
                    k = norm_exam_key(key)
                    if k:
                        ct_mm_without_exam_lookup[k] = w_off
            if mm_exam_weight is not None:
                for key in [exam_code, exam_display]:
                    k = norm_exam_key(key)
                    if k:
                        ct_mm_exam_weight_lookup[k] = float(mm_exam_weight or 0)
            for key in [exam_code, exam_display]:
                k = norm_exam_key(key)
                if k:
                    ct_mm_exam_weight_per_co_lookup[k] = bool(mm_exam_weight_per_co)

    # Sort existing exam assignments using ClassType config order.
    # Any exams not in config come last (stable by created_at).
    exam_assignments.sort(key=lambda ea: (
        ct_index.get(norm_exam_key(getattr(ea, 'exam_display_name', '') or getattr(ea, 'exam', '') or ''), 10**9),
        getattr(ea, 'created_at', None) or timezone.now(),
    ))

    # Deduplicate: keep only the first exam assignment per normalised display name.
    # Stale records can exist with slightly different exam codes (e.g. "CQI2" vs "CQI 2")
    # due to earlier sync runs that used different code normalisation.  Prefer records
    # that have marks entered so progress is not lost.
    exam_assignments.sort(key=lambda ea: (
        ct_index.get(norm_exam_key(getattr(ea, 'exam_display_name', '') or getattr(ea, 'exam', '') or ''), 10**9),
        # Put records with marks first so we keep the one with data when deduping
        -len((ea.draft_data or {}).get('marks', {})) if isinstance(ea.draft_data, dict) else 0,
        getattr(ea, 'created_at', None) or timezone.now(),
    ))
    _seen_ea_display = set()
    _deduped_assignments = []
    for _ea in exam_assignments:
        _dk = norm_exam_key(_ea.exam_display_name or _ea.exam)
        if _dk in _seen_ea_display:
            continue
        _seen_ea_display.add(_dk)
        _deduped_assignments.append(_ea)
    exam_assignments = _deduped_assignments

    for ea in exam_assignments:
        ea_weight = float(ea.weight) if ea.weight else 0
        ea_key = norm_exam_key(ea.exam_display_name or ea.exam)

        # Determine kind from ClassType config (cqi vs regular exam)
        ea_kind = 'exam'
        _cqi_cos = []
        _cqi_name = ''
        for _ea_conf in effective_ea_configs:
            _k = norm_exam_key(_ea_conf.get('exam_display_name', '') or _ea_conf.get('exam', ''))
            if _k == ea_key and str(_ea_conf.get('kind', '')).lower() == 'cqi':
                ea_kind = 'cqi'
                # Prefer DB-backed CQI config when available
                db_cfg = cqi_db_by_key.get(ea_key)
                if db_cfg is not None:
                    _cqi_cos = db_cfg.cos or []
                    _cqi_name = db_cfg.cqi_name or ''
                else:
                    _cqi_sub = _ea_conf.get('cqi', {})
                    _cqi_cos = _cqi_sub.get('cos', [])
                    _cqi_name = _cqi_sub.get('name', '')
                break

        # Resolve weights based on Mark Manager config
        is_mm = bool(ct_mm_enabled_lookup.get(ea_key, False))
        draft = ea.draft_data if isinstance(ea.draft_data, dict) else {}
        user_pattern = draft.get('user_pattern')
        
        p = {}
        if ea_kind != 'cqi':
            qp_type_val = ea.qp_type or ea.exam or ''
            exam_label = (ea.exam_display_name or ea.exam or '').strip()
            qp_match = None
            if not (user_pattern and isinstance(user_pattern, dict)):
                qp_match = (
                    AcV2QpPattern.objects.filter(
                        name__iexact=exam_label, qp_type=qp_type_val,
                        class_type=acv2_ct, is_active=True,
                    ).first()
                    or AcV2QpPattern.objects.filter(
                        name__iexact=exam_label, qp_type=qp_type_val,
                        is_active=True,
                    ).first()
                    or AcV2QpPattern.objects.filter(
                        qp_type=qp_type_val, class_type=acv2_ct, is_active=True,
                    ).first()
                    or AcV2QpPattern.objects.filter(
                        qp_type=qp_type_val, is_active=True,
                    ).first()
                )
            p = user_pattern if (user_pattern and isinstance(user_pattern, dict)) else (qp_match.pattern if (qp_match and isinstance(qp_match.pattern, dict)) else {})

        mm_config = p.get('mark_manager', {}) if isinstance(p, dict) and isinstance(p.get('mark_manager'), dict) else {}
        if mm_config.get('enabled'):
            is_mm = True

        cia_enabled = False
        cia_weight = 0.0
        cia_weight_per_co = False
        if is_mm:
            cia_enabled = bool(mm_config.get('cia_enabled', False))
            if cia_enabled:
                co_weights = ct_mm_with_exam_lookup.get(ea_key, {})
                cia_weight = float(ct_mm_exam_weight_lookup.get(ea_key, 0.0))
                cia_weight_per_co = bool(ct_mm_exam_weight_per_co_lookup.get(ea_key, False))
            else:
                co_weights = ct_mm_without_exam_lookup.get(ea_key, {})
        else:
            co_weights = ct_co_weights_lookup.get(ea_key, {})

        # Resolve weight from ClassType config if ea.weight is 0
        if ea_weight == 0 and ea_key in ct_weight_lookup and ct_weight_lookup[ea_key]:
            if is_mm:
                qp_cos = p.get('cos', []) if isinstance(p, dict) and isinstance(p.get('cos'), list) else []
                derived_set = set()
                for c in qp_cos:
                    if c is not None:
                        for co_num in _extract_co_list(c):
                            derived_set.add(co_num)
                covered_cos_list = sorted(derived_set)
                
                mm_cos_config = mm_config.get('cos', {}) if isinstance(mm_config, dict) and isinstance(mm_config.get('cos'), dict) else {}
                enabled_cos = []
                for co_str, co_cfg in mm_cos_config.items():
                    try:
                        co_num = int(co_str)
                    except Exception:
                        continue
                    if isinstance(co_cfg, dict) and co_cfg.get('enabled'):
                        enabled_cos.append(co_num)
                enabled_cos = sorted(set(enabled_cos))
                if enabled_cos:
                    covered_cos_list = enabled_cos
                
                if cia_enabled:
                    if cia_weight_per_co:
                        ea_weight = sum(float(v or 0) for v in co_weights.values()) + (cia_weight * len(covered_cos_list))
                    else:
                        ea_weight = sum(float(v or 0) for v in co_weights.values()) + cia_weight
                else:
                    ea_weight = sum(float(v or 0) for v in co_weights.values())
            else:
                ea_weight = float(ct_weight_lookup[ea_key])
            
            ea.weight = ea_weight
            ea.save(update_fields=['weight'])

        marks = draft.get('marks', {})
        entered_count = sum(1 for v in marks.values() if v is not None and v != '')
        is_strictly_locked = ea.status in PUBLISHED_EXAM_STATUSES  # PUBLISHED, APPROVED, LOCKED
        has_been_published = bool(is_strictly_locked or ea.published_at)
        cycle_state = _get_exam_cycle_state(
            ea,
            semester_id=getattr(sec, 'semester_id', None),
            class_type=acv2_ct,
        )
        is_locked = bool(is_strictly_locked or cycle_state['cycle_locked'])

        if has_been_published:
            sm_status = 'COMPLETED'
        elif marks:
            sm_status = 'IN_PROGRESS'
        else:
            sm_status = 'NOT_STARTED'

        exams.append({
            'id': str(ea.id),
            'name': ea.exam_display_name or ea.exam or ea.qp_type,
            'short_name': ea.exam or ea.qp_type or '',
            'max_marks': ea.max_marks or 0,
            'weight': ea_weight,
            'co_weights': co_weights,  # Per-CO weights
            'cia_enabled': cia_enabled,
            'cia_weight': cia_weight,
            'cia_weight_per_co': cia_weight_per_co,
            'entered_count': entered_count,
            'total_students': student_count,
            'is_locked': is_locked,
            'cycle_locked': cycle_state['cycle_locked'],
            'lock_reason': cycle_state['cycle_lock_reason'],
            'cycle_name': cycle_state['cycle_name'],
            'cycle_code': cycle_state['cycle_code'],
            'can_view': bool(has_been_published),
            'can_edit': bool((not is_strictly_locked) and (not cycle_state['cycle_locked'])),
            'due_date': None,
            'status': sm_status,
            'kind': ea_kind,
            'cqi_cos': _cqi_cos,
            'cqi_name': _cqi_name,
        })

    # Sync exam assignments from ClassType config.
    # Auto-create AcV2Section + missing AcV2ExamAssignment records.
    if acv2_ct and effective_ea_configs and sec:
        from academics.models import Subject as AcademicsSubject
        # Find or create AcV2Course
        subject = ta.subject or (
            AcademicsSubject.objects.filter(code=course_code).first()
            if course_code and course_code != '-' else None
        )
        semester = sec.semester if sec else None
        acv2_course = None
        if subject and semester:
            acv2_course, created = AcV2Course.objects.get_or_create(
                subject=subject,
                semester=semester,
                defaults={
                    'subject_code': course_code,
                    'subject_name': course_name,
                    'class_type': acv2_ct,
                    'class_type_name': acv2_ct.display_name,
                },
            )
            # Sync question_paper_type from curriculum onto AcV2Course
            if qp_type_code and acv2_course.question_paper_type != qp_type_code:
                acv2_course.question_paper_type = qp_type_code
                acv2_course.save(update_fields=['question_paper_type'])
            # Correct class_type if it was previously set to the wrong one
            if not created and acv2_ct and acv2_course.class_type_id != acv2_ct.id:
                acv2_course.class_type = acv2_ct
                acv2_course.class_type_name = acv2_ct.display_name
                acv2_course.save(update_fields=['class_type', 'class_type_name'])
                # Delete stale exam assignments that have no marks entered
                for acv2_sec_obj in acv2_course.sections.all():
                    stale_eas = AcV2ExamAssignment.objects.filter(section=acv2_sec_obj)
                    for stale_ea in stale_eas:
                        draft = stale_ea.draft_data if isinstance(stale_ea.draft_data, dict) else {}
                        marks = draft.get('marks', {})
                        has_marks = any(v is not None and v != '' for v in marks.values())
                        if not has_marks and stale_ea.status == 'DRAFT':
                            stale_ea.delete()
                # Clear the exams list so they get rebuilt from correct class type
                exams.clear()
        if acv2_course:
            # Create AcV2Section
            acv2_sec, _ = AcV2Section.objects.get_or_create(
                course=acv2_course,
                teaching_assignment=ta,
                defaults={
                    'section_name': sec.name if sec else 'A',
                    'faculty_user': getattr(getattr(ta, 'staff', None), 'user', None) or request.user,
                },
            )
            # Sync faculty_user on retrieve if it differs from the assigned staff user
            assigned_user = getattr(getattr(ta, 'staff', None), 'user', None)
            if assigned_user and acv2_sec.faculty_user != assigned_user:
                acv2_sec.faculty_user = assigned_user
                acv2_sec.save(update_fields=['faculty_user'])
            # Track which exam codes already exist
            existing_exam_codes = set(e['short_name'] for e in exams)

            # --- Clean up stale bad-code records ---
            # WeightagePage previously saved exam = qp_type (e.g., "WD") instead of the
            # exam display name.  These records block correct sync.  Delete them if they
            # have no marks entered and are still in DRAFT status.
            bad_code_eas = AcV2ExamAssignment.objects.filter(
                section=acv2_sec,
                exam=qp_type_code,   # exam code equals the qp_type code – bad data
                status='DRAFT',
            )
            for bad_ea in bad_code_eas:
                draft = bad_ea.draft_data if isinstance(bad_ea.draft_data, dict) else {}
                marks = draft.get('marks', {})
                has_marks = any(v is not None and v != '' for v in marks.values())
                if not has_marks:
                    bad_ea.delete()
                    existing_exam_codes.discard(bad_ea.exam)
            # Rebuild after cleanup
            exams = [e for e in exams if e.get('short_name') != qp_type_code]
            existing_exam_codes = set(e['short_name'] for e in exams)

            # Create/sync exam assignments from effective exam configs (class type qp configs
            # when present; otherwise derived from QP patterns)
            for ea_conf in effective_ea_configs:
                exam_code_raw = ea_conf.get('exam', '')
                display_name = ea_conf.get('exam_display_name', exam_code_raw)
                weight = ea_conf.get('weight', 0)
                qp_type_val = (ea_conf.get('qp_type') or qp_type_code or exam_code_raw or '').strip()

                _new_kind = 'cqi' if str(ea_conf.get('kind', '')).lower() == 'cqi' else 'exam'
                _new_cqi_sub = ea_conf.get('cqi', {}) if isinstance(ea_conf.get('cqi'), dict) else {}
                _new_cqi_cos = _new_cqi_sub.get('cos', []) if isinstance(_new_cqi_sub.get('cos', []), list) else []
                _new_cqi_name = str(_new_cqi_sub.get('name', '') or '')

                # If exam code equals the qp_type (legacy bad data from WeightagePage),
                # use display_name as the unique exam identifier instead.
                exam_code_val = (
                    display_name
                    if (exam_code_raw == qp_type_val and display_name and display_name != qp_type_val)
                    else exam_code_raw
                )
                if not exam_code_val:
                    continue

                # covered_cos/max_marks: for CQI, always take COs from the CQI editor config.
                # For regular exams, derive from QP pattern.
                derived_max = 0
                covered_cos = []
                if _new_kind == 'cqi':
                    # Prefer DB-backed config for CQI to ensure stability.
                    try:
                        db_cfg = (
                            cqi_db_by_key.get(norm_exam_key(exam_code_val))
                            or cqi_db_by_key.get(norm_exam_key(display_name))
                            or cqi_db_by_key.get(norm_exam_key(exam_code_raw))
                        )
                        if db_cfg is not None:
                            _new_cqi_cos = getattr(db_cfg, 'cos', []) or []
                            _new_cqi_name = getattr(db_cfg, 'cqi_name', '') or _new_cqi_name
                    except Exception:
                        pass
                    covered_cos = [int(x) for x in (_new_cqi_cos or []) if str(x).isdigit()]
                else:
                    # Prefer class-specific pattern, then match by display_name, then any match.
                    qp_match = (
                        AcV2QpPattern.objects.filter(
                            name__iexact=display_name, qp_type__iexact=qp_type_val,
                            class_type=acv2_ct, is_active=True,
                        ).first()
                        or AcV2QpPattern.objects.filter(
                            name__iexact=display_name, qp_type__iexact=qp_type_val, is_active=True,
                        ).first()
                        or AcV2QpPattern.objects.filter(
                            qp_type__iexact=qp_type_val, class_type=acv2_ct, is_active=True,
                        ).first()
                        or AcV2QpPattern.objects.filter(
                            qp_type__iexact=qp_type_val, is_active=True,
                        ).first()
                    )
                    if qp_match and isinstance(qp_match.pattern, dict):
                        qp_marks = qp_match.pattern.get('marks', [])
                        qp_cos = qp_match.pattern.get('cos', [])
                        qp_enabled = qp_match.pattern.get('enabled', [True] * len(qp_marks))
                        derived_max = sum(
                            m for i, m in enumerate(qp_marks)
                            if i < len(qp_enabled) and qp_enabled[i]
                        )
                        covered_cos = sorted(set(
                            co
                            for i, c in enumerate(qp_cos)
                            if c is not None and i < len(qp_enabled) and qp_enabled[i]
                            for co in (c if isinstance(c, list) else [c])
                            if isinstance(co, int)
                        ))
                    if not covered_cos:
                        covered_cos = ea_conf.get('default_cos', [])

                final_max = derived_max or weight or 50

                if exam_code_val in existing_exam_codes:
                    # Already in the exams list, skip
                    continue

                ea_obj, created = AcV2ExamAssignment.objects.get_or_create(
                    section=acv2_sec,
                    exam=exam_code_val,
                    defaults={
                        'exam_display_name': display_name,
                        'qp_type': qp_type_val,
                        'max_marks': final_max,
                        'weight': weight,
                        'covered_cos': covered_cos,
                    },
                )
                ea_update_fields = []
                if qp_type_val and (ea_obj.qp_type or '').strip() != qp_type_val:
                    ea_obj.qp_type = qp_type_val
                    ea_update_fields.append('qp_type')
                if display_name and (ea_obj.exam_display_name or '').strip() != display_name:
                    ea_obj.exam_display_name = display_name
                    ea_update_fields.append('exam_display_name')
                # For CQI, covered_cos must always match CQI editor COs.
                if _new_kind == 'cqi' and covered_cos and (ea_obj.covered_cos or []) != covered_cos:
                    ea_obj.covered_cos = covered_cos
                    ea_update_fields.append('covered_cos')
                if ea_update_fields:
                    ea_obj.save(update_fields=ea_update_fields)
                # Get per-CO weights from class type config
                co_weights_for_new = ea_conf.get('co_weights', {})
                if co_weights_for_new:
                    co_weights_for_new = {int(k): v for k, v in co_weights_for_new.items()}
                cycle_state = _get_exam_cycle_state(
                    ea_obj,
                    semester_id=getattr(sec, 'semester_id', None),
                    class_type=acv2_ct,
                )
                exams.append({
                    'id': str(ea_obj.id),
                    'name': display_name,
                    'short_name': exam_code_val,
                    'max_marks': ea_obj.max_marks or 0,
                    'weight': ea_obj.weight or 0,
                    'co_weights': co_weights_for_new,  # Per-CO weights
                    'entered_count': 0,
                    'total_students': student_count,
                    'is_locked': bool(cycle_state['cycle_locked']),
                    'cycle_locked': cycle_state['cycle_locked'],
                    'lock_reason': cycle_state['cycle_lock_reason'],
                    'cycle_name': cycle_state['cycle_name'],
                    'cycle_code': cycle_state['cycle_code'],
                    'can_view': False,
                    'can_edit': not cycle_state['cycle_locked'],
                    'due_date': None,
                    'status': 'NOT_STARTED',
                    'kind': _new_kind,
                    'cqi_cos': _new_cqi_cos if isinstance(_new_cqi_cos, list) else [],
                    'cqi_name': _new_cqi_name,
                })

    # Final ordering for response: respect ClassType config order (when available)
    if ct_index:
        try:
            exams.sort(key=lambda e: (
                ct_index.get(norm_exam_key(e.get('name', '') or e.get('short_name', '') or ''), 10**9),
                e.get('name', '') or '',
            ))
        except Exception:
            pass

    semester_num = sec.semester.number if sec and sec.semester else 0
    dept_name = ''
    if sec and sec.managing_department:
        dept_name = (getattr(sec.managing_department, 'short_name', '')
                     or getattr(sec.managing_department, 'name', ''))

    faculty_display_name = ''
    try:
        faculty_display_name = ta.staff.user.get_full_name() or str(ta.staff.user)
    except Exception:
        pass

    return Response({
        'id': str(ta.id),
        'course_code': course_code,
        'course_name': course_name,
        'class_name': sec.name if sec else '',
        'section': sec.name if sec else '',
        'semester': semester_num,
        'department': dept_name,
        'student_count': student_count,
        'is_elective': bool(ta.elective_subject_id),
        'class_type': class_type_info,
        'qp_type': qp_type_code or None,
        'faculty_name': faculty_display_name,
        'setup_status': {
            'class_type_assigned': bool(acv2_ct),
            # Preserve whether curriculum/elective explicitly has qp_type configured
            'qp_type_assigned': bool(curriculum_qp_type_code),
        },
        'exams': exams,
    })


# ============================================================================
# FACULTY COURSES STATUS (batch, for CourseListPage)
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_courses_status(request):
    """Return lightweight mark-entry status for all active teaching assignments
    of the current user.  Used by CourseListPage to show real completion state.

    Returns: { "<ta_id>": "NOT_STARTED" | "IN_PROGRESS" | "COMPLETED" }
    """
    from academics.models import TeachingAssignment

    ta_ids = list(
        TeachingAssignment.objects.filter(
            staff__user=request.user,
            is_active=True,
        ).values_list('id', flat=True)
    )
    if not ta_ids:
        return Response({})

    # Map section → ta_id
    sections = AcV2Section.objects.filter(
        teaching_assignment_id__in=ta_ids,
    ).values('id', 'teaching_assignment_id')
    section_to_ta: dict = {str(s['id']): s['teaching_assignment_id'] for s in sections}

    if not section_to_ta:
        return Response({str(ta): 'NOT_STARTED' for ta in ta_ids})

    # Fetch lightweight EA info (no draft_data needed – use published_at / status)
    eas = AcV2ExamAssignment.objects.filter(
        section_id__in=section_to_ta.keys(),
    ).values('section_id', 'status', 'published_at', 'draft_data')

    # Initialise all TAs as NOT_STARTED
    result: dict = {str(ta): 'NOT_STARTED' for ta in ta_ids}

    for ea in eas:
        ta_id = str(section_to_ta[str(ea['section_id'])])
        if result.get(ta_id) == 'COMPLETED':
            continue  # Already at max

        is_published = (
            ea['status'] in PUBLISHED_EXAM_STATUSES
            or bool(ea.get('published_at'))
        )
        if is_published:
            result[ta_id] = 'IN_PROGRESS'  # At least one published
            continue

        # Check draft_data for any entered marks
        draft = ea.get('draft_data') or {}
        if isinstance(draft, dict):
            marks = draft.get('marks', {})
            if any(v is not None and v != '' for v in marks.values()):
                result[ta_id] = 'IN_PROGRESS'

    # If ALL exam assignments for a TA are published → COMPLETED
    # Build per-TA counts
    total_per_ta: dict = {}
    published_per_ta: dict = {}
    for ea in eas:
        ta_id = str(section_to_ta[str(ea['section_id'])])
        total_per_ta[ta_id] = total_per_ta.get(ta_id, 0) + 1
        is_pub = ea['status'] in PUBLISHED_EXAM_STATUSES or bool(ea.get('published_at'))
        if is_pub:
            published_per_ta[ta_id] = published_per_ta.get(ta_id, 0) + 1

    for ta_id, total in total_per_ta.items():
        if total > 0 and published_per_ta.get(ta_id, 0) == total:
            result[ta_id] = 'COMPLETED'

    return Response(result)


# ============================================================================
# FACULTY EXAM INFO (for MarkEntryPage)
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_exam_info(request, exam_id):
    """Return exam information for a specific AcV2ExamAssignment."""
    from academics.models import TeachingAssignment

    ea_qs = AcV2ExamAssignment.objects.select_related(
        'section__teaching_assignment__section',
        'section__teaching_assignment__section__semester',
        'section__teaching_assignment__section__managing_department',
        'section__teaching_assignment__curriculum_row',
        'section__teaching_assignment__elective_subject',
    )
    if _has_admin_bypass_access(request.user):
        ea = get_object_or_404(ea_qs, id=exam_id)
    else:
        ea = get_object_or_404(ea_qs, id=exam_id, section__faculty_user=request.user)
    ea = _ensure_due_auto_publish(ea)

    ta = ea.section.teaching_assignment
    acad_sec = ta.section
    cr = ta.curriculum_row
    es = ta.elective_subject

    course_code = (cr.course_code if cr else None) or (getattr(es, 'course_code', None)) or '-'
    course_name = (cr.course_name if cr else None) or (getattr(es, 'course_name', None)) or '-'

    # Publish control / editability (semester open window, due date, publish lock, edit window)
    ctrl = check_publish_control(ea)
    # JSON-safe conversion (timedelta is not serializable)
    if ctrl.get('time_remaining') is not None:
        try:
            ctrl['time_remaining_seconds'] = int(ctrl['time_remaining'].total_seconds())
        except Exception:
            ctrl['time_remaining_seconds'] = None
        try:
            del ctrl['time_remaining']
        except Exception:
            pass
    semester_config = ea.get_semester_config()
    open_from = getattr(semester_config, 'open_from', None) if semester_config else None
    due_at = getattr(semester_config, 'due_at', None) if semester_config else None

    open_remaining_seconds = None
    due_remaining_seconds = None
    try:
        now = timezone.now()
        if open_from and now < open_from:
            open_remaining_seconds = int((open_from - now).total_seconds())
        if due_at:
            if now > due_at:
                due_remaining_seconds = 0
            else:
                due_remaining_seconds = int((due_at - now).total_seconds())
    except Exception:
        open_remaining_seconds = None
        due_remaining_seconds = None

    course_class_type = None
    try:
        course_class_type = ea.section.course.class_type
    except Exception:
        course_class_type = None
    cycle_state = _get_exam_cycle_state(
        ea,
        semester_id=getattr(acad_sec, 'semester_id', None),
        class_type=course_class_type,
    )
    is_locked = bool(ctrl.get('is_locked', False) or cycle_state['cycle_locked'])

    dept_name = ''
    if acad_sec and acad_sec.managing_department:
        dept_name = (getattr(acad_sec.managing_department, 'short_name', '')
                     or getattr(acad_sec.managing_department, 'name', ''))

    draft = ea.draft_data if isinstance(ea.draft_data, dict) else {}
    question_btls = draft.get('question_btls', {})

    # Check if user has a course-specific pattern (from Mark Manager)
    user_pattern = draft.get('user_pattern')

    # Build qp_pattern with questions array
    qp_pattern_response = None
    mark_manager = None

    if user_pattern and isinstance(user_pattern, dict):
        # Use course-specific pattern from Mark Manager (stored in draft_data)
        p = user_pattern
        titles = p.get('titles', [])
        marks_list = p.get('marks', [])
        cos = p.get('cos', [])
        btls = p.get('btls', [])
        enabled = p.get('enabled', [])
        questions = []
        for i in range(len(titles)):
            if i < len(enabled) and not enabled[i]:
                continue
            questions.append({
                'id': f'q{i}',
                'question_number': titles[i] if i < len(titles) else str(i + 1),
                'max_marks': marks_list[i] if i < len(marks_list) else 0,
                'btl_level': btls[i] if i < len(btls) else None,
                'co_number': cos[i] if i < len(cos) else 0,
            })
        if questions:
            qp_pattern_response = {
                'id': 'user_defined',
                'name': 'User Defined',
                'questions': questions,
            }
        # Mark manager config from user's pattern
        mm = p.get('mark_manager')
        if mm and isinstance(mm, dict):
            mark_manager = mm
    else:
        # Fall back to global QP pattern
        qp_type = ''
        try:
            qp_type = (ea.section.course.question_paper_type or '').strip()
        except Exception:
            qp_type = ''
        if not qp_type:
            qp_type = (getattr(cr, 'question_paper_type', None) or getattr(es, 'question_paper_type', None) or '').strip()
        if not qp_type:
            qp_type = (ea.qp_type or '').strip() or (ea.exam or '').strip() or ''

        exam_key = (ea.exam_display_name or ea.exam or '').strip()

        ct = None
        try:
            ct = ea.section.course.class_type
        except Exception:
            ct = None

        base_qs = AcV2QpPattern.objects.filter(qp_type=qp_type, is_active=True)
        matched_pattern = None

        # 1) Class Type + QP Type + Exam name match
        if ct is not None:
            scoped = base_qs.filter(class_type=ct)
            if exam_key:
                matched_pattern = scoped.filter(name__iexact=exam_key).order_by('-updated_at').first()
            else:
                matched_pattern = scoped.order_by('-updated_at').first()

        # 3) Global + QP Type + Exam name match
        if not matched_pattern:
            global_qs = base_qs.filter(class_type__isnull=True)
            if exam_key:
                matched_pattern = global_qs.filter(name__iexact=exam_key).order_by('-updated_at').first()
            else:
                matched_pattern = global_qs.order_by('-updated_at').first()

        if matched_pattern and isinstance(matched_pattern.pattern, dict):
            p = matched_pattern.pattern
            titles = p.get('titles', [])
            marks_list = p.get('marks', [])
            cos = p.get('cos', [])
            btls = p.get('btls', [])
            enabled = p.get('enabled', [])
            questions = []
            for i in range(len(titles)):
                if i < len(enabled) and not enabled[i]:
                    continue
                questions.append({
                    'id': f'q{i}',
                    'question_number': titles[i] if i < len(titles) else str(i + 1),
                    'max_marks': marks_list[i] if i < len(marks_list) else 0,
                    'btl_level': btls[i] if i < len(btls) else None,
                    'co_number': cos[i] if i < len(cos) else 0,
                })
            if questions:
                qp_pattern_response = {
                    'id': str(matched_pattern.id),
                    'name': matched_pattern.name,
                    'questions': questions,
                }
            # Include mark_manager config from QP pattern for user_define mode
            mm = p.get('mark_manager')
            if mm and isinstance(mm, dict) and mm.get('enabled'):
                mark_manager = mm

    # Resolve CQI config for THIS specific exam assignment (not just first CQI).
    # Look up the AcV2CqiExam whose exam_code or exam_display_name matches ea.exam / ea.exam_display_name.
    cqi_config_for_exam = None
    try:
        from .models import AcV2CqiExam
        _ea_exam_key = (ea.exam_display_name or ea.exam or '').strip().lower()
        _ea_qp_type = ''
        try:
            _ea_qp_type = (ea.section.course.question_paper_type or '').strip()
        except Exception:
            _ea_qp_type = (ea.qp_type or '').strip()
        _ea_ct = None
        try:
            _ea_ct = ea.section.course.class_type
        except Exception:
            pass
        if _ea_exam_key and _ea_ct:
            _cqi_qs = AcV2CqiExam.objects.filter(class_type=_ea_ct, is_active=True)
            if _ea_qp_type:
                _cqi_qs = _cqi_qs.filter(qp_type_code__iexact=_ea_qp_type)
            _matched_cqi = None
            for _cqi_row in _cqi_qs.order_by('order', 'created_at'):
                _row_key_code = (_cqi_row.exam_code or '').strip().lower()
                _row_key_disp = (_cqi_row.exam_display_name or '').strip().lower()
                if _ea_exam_key in (_row_key_code, _row_key_disp):
                    _matched_cqi = _cqi_row
                    break
            if _matched_cqi is not None:
                cqi_config_for_exam = {
                    'name': str(_matched_cqi.cqi_name or ''),
                    'code': str(_matched_cqi.cqi_code or ''),
                    'cos': _matched_cqi.cos if isinstance(_matched_cqi.cos, list) else [],
                    'exams': _matched_cqi.considered_exams if isinstance(_matched_cqi.considered_exams, list) else [],
                    'custom_vars': _matched_cqi.custom_vars if isinstance(_matched_cqi.custom_vars, list) else [],
                    'derived_variables': _matched_cqi.derived_variables if isinstance(_matched_cqi.derived_variables, list) else [],
                    'co_value_expr': str(_matched_cqi.co_value_expr or ''),
                    'formula': str(_matched_cqi.formula or ''),
                    'conditions': _matched_cqi.conditions if isinstance(_matched_cqi.conditions, list) else [],
                    'else_formula': str(_matched_cqi.else_formula or ''),
                }
    except Exception:
        cqi_config_for_exam = None

    return Response({
        'id': str(ea.id),
        'name': ea.exam_display_name or ea.exam or ea.qp_type or '',
        'max_marks': float(ea.max_marks) if ea.max_marks else 0,
        'course_id': str(ta.id) if ta else None,
        'course_code': course_code,
        'course_name': course_name,
        'class_name': str(acad_sec) if acad_sec else '',
        'section': acad_sec.name if acad_sec else '',
        'department': dept_name,
        'due_date': ea.edit_window_until.isoformat() if ea.edit_window_until else None,
        'status': ea.status,
        'is_locked': is_locked,
        'cycle_locked': cycle_state['cycle_locked'],
        'lock_reason': cycle_state['cycle_lock_reason'],
        'cycle_name': cycle_state['cycle_name'],
        'cycle_code': cycle_state['cycle_code'],
        'has_pending_edit_request': bool(getattr(ea, 'has_pending_edit_request', False)),
        'publish_control': {
            **ctrl,
            'is_editable': bool(ctrl.get('is_editable', not is_locked)) and not cycle_state['cycle_locked'],
            'open_from': open_from.isoformat() if open_from else None,
            'due_at': due_at.isoformat() if due_at else None,
            'is_open': semester_config.is_open() if semester_config else True,
            'open_remaining_seconds': open_remaining_seconds,
            'due_remaining_seconds': due_remaining_seconds,
        },
        'qp_pattern': qp_pattern_response,
        'question_btls': question_btls,
        'mark_manager': mark_manager,
        'cqi_config': cqi_config_for_exam,
    })


# ==========================================================================
# FACULTY EXAM PUBLISH + REQUEST EDIT (for MarkEntryPage publish control)
# ==========================================================================
# ==========================================================================
# FACULTY EXAM INFO (for MarkEntryPage)
# ============================================================================

def _normalize_mark_number(value):
    if value in (None, ''):
        return None
    try:
        num = float(value)
        # Guard against NaN/Infinity which break JSON/Decimal conversions.
        if num != num or num in (float('inf'), float('-inf')):
            return None
        return num
    except (TypeError, ValueError):
        return None


def _normalize_question_marks(value):
    """Normalize question/CO marks mapping for JSON + DB safety."""
    if not isinstance(value, dict):
        return {}
    out = {}
    for k, v in value.items():
        key = str(k)
        if v in (None, ''):
            out[key] = None
            continue
        try:
            num = float(v)
        except (TypeError, ValueError):
            out[key] = None
            continue
        if num != num or num in (float('inf'), float('-inf')):
            out[key] = None
        else:
            out[key] = num
    return out


def _normalize_marks_payload(marks_payload):
    """Normalize incoming marks payload to list rows + map keyed by student_id."""
    rows = []

    if isinstance(marks_payload, list):
        for item in marks_payload:
            if not isinstance(item, dict):
                continue
            student_id = str(item.get('student_id') or '').strip()
            if not student_id:
                continue
            co_marks = item.get('co_marks') if isinstance(item.get('co_marks'), dict) else item.get('question_marks', {})
            co_marks = _normalize_question_marks(co_marks)
            rows.append({
                'student_id': student_id,
                'mark': _normalize_mark_number(item.get('mark')),
                'co_marks': co_marks,
                'is_absent': bool(item.get('is_absent', False)),
            })
    elif isinstance(marks_payload, dict):
        for student_id, item in marks_payload.items():
            sid = str(student_id or '').strip()
            if not sid:
                continue
            if isinstance(item, dict):
                co_marks = item.get('co_marks') if isinstance(item.get('co_marks'), dict) else item.get('question_marks', {})
                co_marks = _normalize_question_marks(co_marks)
                mark_val = item.get('mark')
                is_absent = bool(item.get('is_absent', False))
            else:
                co_marks = {}
                mark_val = item
                is_absent = False
            rows.append({
                'student_id': sid,
                'mark': _normalize_mark_number(mark_val),
                'co_marks': co_marks,
                'is_absent': is_absent,
            })

    marks_map = {
        row['student_id']: {
            'mark': row['mark'],
            'co_marks': row['co_marks'],
            'is_absent': row['is_absent'],
        }
        for row in rows
    }
    return rows, marks_map


def _save_draft_marks_for_exam(exam_assignment, marks_payload):
    """Persist row-level draft marks into acv2_draft_mark and return normalized map."""
    from academics.models import StudentProfile
    from decimal import Decimal, InvalidOperation

    rows, marks_map = _normalize_marks_payload(marks_payload)
    if not rows:
        return marks_map

    student_ids = [r['student_id'] for r in rows if r.get('student_id')]
    student_map = {
        str(sp.id): sp
        for sp in StudentProfile.objects.filter(id__in=student_ids).select_related('user')
    }

    def _to_decimal_2(value):
        if value in (None, ''):
            return None
        try:
            num = float(value)
        except (TypeError, ValueError):
            return None
        if num != num or num in (float('inf'), float('-inf')):
            return None
        try:
            return Decimal(str(num)).quantize(Decimal('0.01'))
        except (InvalidOperation, TypeError, ValueError):
            return None

    for row in rows:
        sp = student_map.get(row['student_id'])
        if not sp:
            continue
        AcV2DraftMark.objects.update_or_create(
            exam_assignment=exam_assignment,
            student=sp,
            defaults={
                'reg_no': sp.reg_no or '',
                'student_name': str(sp.user) if sp.user else sp.reg_no or '',
                'total_mark': _to_decimal_2(row.get('mark')),
                'question_marks': _normalize_question_marks(row.get('co_marks')),
                'is_absent': bool(row['is_absent']),
            },
        )

    return marks_map


def _materialize_student_marks_from_map(exam_assignment, marks_map):
    """Ensure published/derived marks exist in acv2_student_mark for reporting."""
    from academics.models import StudentProfile

    if not isinstance(marks_map, dict) or not marks_map:
        return

    student_map = {
        str(sp.id): sp
        for sp in StudentProfile.objects.filter(id__in=list(marks_map.keys())).select_related('user')
    }

    for student_id, payload in marks_map.items():
        sp = student_map.get(str(student_id))
        if not sp:
            continue
        co_marks = payload.get('co_marks') if isinstance(payload, dict) else {}
        if not isinstance(co_marks, dict):
            co_marks = {}
        mark_val = payload.get('mark') if isinstance(payload, dict) else None
        if mark_val in ('',):
            mark_val = None

        AcV2StudentMark.objects.update_or_create(
            exam_assignment=exam_assignment,
            student=sp,
            defaults={
                'reg_no': sp.reg_no or '',
                'student_name': str(sp.user) if sp.user else sp.reg_no or '',
                'total_mark': _normalize_mark_number(mark_val),
                'question_marks': co_marks,
                'is_absent': bool(payload.get('is_absent', False)) if isinstance(payload, dict) else False,
            },
        )


def _ensure_due_auto_publish(exam_assignment):
    """Run due-date auto publish lazily for a single exam if configured and overdue."""
    semester_config = exam_assignment.get_semester_config()
    if not semester_config or not semester_config.due_at:
        return exam_assignment
    now = timezone.now()

    # If the due date was extended after an auto-publish, reopen only those auto-published
    # assignments (published_by is NULL). Manually published exams must remain locked
    # until an edit request is approved.
    if now <= semester_config.due_at:
        if exam_assignment.status in ('PUBLISHED', 'LOCKED') and getattr(exam_assignment, 'published_by_id', None) is None:
            exam_assignment.status = 'DRAFT'
            exam_assignment.save(update_fields=['status'])
        return exam_assignment
    if exam_assignment.status in ('PUBLISHED', 'LOCKED'):
        return exam_assignment

    process_auto_publish(semester_config)

    refreshed = AcV2ExamAssignment.objects.get(id=exam_assignment.id)
    if refreshed.status not in ('PUBLISHED', 'LOCKED'):
        return refreshed

    published = refreshed.published_data if isinstance(refreshed.published_data, dict) else {}
    marks_map = published.get('marks', {}) if isinstance(published.get('marks', {}), dict) else {}
    if not marks_map:
        marks_map = {
            str(dm.student_id): {
                'mark': float(dm.total_mark) if dm.total_mark is not None else None,
                'co_marks': dm.question_marks if isinstance(dm.question_marks, dict) else {},
                'is_absent': bool(dm.is_absent),
            }
            for dm in AcV2DraftMark.objects.filter(exam_assignment=refreshed)
        }
    if marks_map:
        _materialize_student_marks_from_map(refreshed, marks_map)
        try:
            compute_section_internal_marks(refreshed.section)
        except Exception:
            pass

    return AcV2ExamAssignment.objects.get(id=exam_assignment.id)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def faculty_exam_publish(request, exam_id):
    """Publish an exam (locks if publish control is enabled)."""
    ea_qs = AcV2ExamAssignment.objects.select_related('section__course__semester')
    if _has_admin_bypass_access(request.user):
        ea = get_object_or_404(ea_qs, id=exam_id)
    else:
        ea = get_object_or_404(ea_qs, id=exam_id, section__faculty_user=request.user)

    cycle_state = _get_exam_cycle_state(
        ea,
        semester_id=getattr(getattr(ea.section, 'course', None), 'semester_id', None),
        class_type=getattr(getattr(ea.section, 'course', None), 'class_type', None),
    )
    if cycle_state['cycle_locked'] and not _has_admin_bypass_access(request.user):
        return Response({'detail': cycle_state['cycle_lock_reason'] or 'This cycle is locked for this semester.'}, status=403)

    if not _has_admin_bypass_access(request.user) and not ea.is_editable():
        return Response({'detail': 'This exam is locked and cannot be published.'}, status=403)

    semester_config = ea.get_semester_config()

    prev_published_data = ea.published_data if isinstance(ea.published_data, dict) else {}
    prev_published_marks = prev_published_data.get('marks', {}) if isinstance(prev_published_data.get('marks', {}), dict) else {}
    was_published_before = bool(ea.published_at) or ea.status in ('PUBLISHED', 'LOCKED') or bool(prev_published_marks)

    with transaction.atomic():
        ea.published_data = ea.draft_data if isinstance(ea.draft_data, dict) else {}
        ea.published_at = timezone.now()
        ea.published_by = request.user
        if semester_config and semester_config.publish_control_enabled:
            ea.status = 'PUBLISHED'
        else:
            ea.status = 'DRAFT'
        # Any approved edit window is consumed by publishing
        ea.edit_window_until = None
        ea.edit_window_until_publish = False
        ea.save(update_fields=['published_data', 'published_at', 'published_by', 'status', 'edit_window_until', 'edit_window_until_publish'])

        published_marks = ea.published_data.get('marks', {}) if isinstance(ea.published_data, dict) else {}
        if isinstance(published_marks, dict) and published_marks:
            _materialize_student_marks_from_map(ea, published_marks)

        # Recompute internal marks for this section
        try:
            compute_section_internal_marks(ea.section)
        except Exception:
            pass

    # Send publish notifications after DB commit.
    try:
        new_published_data = ea.published_data if isinstance(ea.published_data, dict) else {}
        new_published_marks = new_published_data.get('marks', {}) if isinstance(new_published_data.get('marks', {}), dict) else {}
        _send_student_publish_notifications(
            exam_assignment=ea,
            actor_user=request.user,
            prev_published_marks=prev_published_marks,
            new_published_marks=new_published_marks,
            was_published_before=was_published_before,
        )
    except Exception as e:
        import logging
        logging.getLogger('academic_v2.notifications').error(f'Publish notification hook failed: {e}')

    return Response({'success': True, 'status': ea.status, 'published_at': ea.published_at.isoformat()})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def faculty_exam_request_edit(request, exam_id):
    """Request edit access for a published/locked exam."""
    ea = get_object_or_404(
        AcV2ExamAssignment.objects.select_related('section__course__semester'),
        id=exam_id,
        section__faculty_user=request.user,
    )

    reason = request.data.get('reason', '')
    if not reason:
        return Response({'detail': 'Reason is required.'}, status=400)

    result = create_edit_request(ea, request.user, reason)
    if result.get('success'):
        return Response(result)

    # Ensure failures are returned as a proper JSON response
    # (previously this view could fall through and return None, causing a 500)
    return Response(
        {
            'success': False,
            'request_id': result.get('request_id'),
            'error': result.get('error') or 'Request failed',
        },
        status=status.HTTP_400_BAD_REQUEST,
    )
    return Response(result, status=400)


# ============================================================================
# FACULTY EXAM MARKS (GET + POST for MarkEntryPage)
# ============================================================================

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def faculty_exam_marks(request, exam_id):
    """GET: List students with current marks. POST: Save marks."""
    from academics.models import TeachingAssignment, StudentSectionAssignment

    ea_qs = AcV2ExamAssignment.objects.select_related(
        'section__teaching_assignment__section',
    )
    if _has_admin_bypass_access(request.user):
        ea = get_object_or_404(ea_qs, id=exam_id)
    else:
        ea = get_object_or_404(ea_qs, id=exam_id, section__faculty_user=request.user)
    ea = _ensure_due_auto_publish(ea)

    ta = ea.section.teaching_assignment
    acad_sec = ta.section

    if request.method == 'GET':
        # Live pull from linked Google Sheet before rendering the mark entry table.
        # This keeps the faculty table aligned with sheet data when mapping is configured.
        try:
            link = AcV2GoogleSheetLink.objects.filter(section=ea.section, is_active=True).first()
            if link and link.spreadsheet_id:
                exam_configs = link.exam_configs if isinstance(link.exam_configs, dict) else {}
                exam_config = exam_configs.get(str(ea.id), {}) if isinstance(exam_configs.get(str(ea.id), {}), dict) else {}
                sheet_name = str(exam_config.get('sheetTab') or ea.exam_display_name or ea.exam or 'Marks').strip()
                sync_google_sheet_to_backend(
                    exam_assignment=ea,
                    spreadsheet_id=link.spreadsheet_id,
                    config={},
                    sheet_name=sheet_name,
                    column_mapping=exam_config,
                )
        except Exception:
            # Do not block faculty mark entry if live sync fails.
            pass

        # Get all active students in the section
        assignments = (
            StudentSectionAssignment.objects
            .filter(section=acad_sec, end_date__isnull=True)
            .select_related('student__user', 'student__home_department')
            .order_by('student__reg_no')
        )

        # Load existing marks for this exam
        existing = {
            str(sm.student_id): sm
            for sm in AcV2StudentMark.objects.filter(exam_assignment=ea)
        }
        draft_existing = {
            str(dm.student_id): dm
            for dm in AcV2DraftMark.objects.filter(exam_assignment=ea)
        }

        draft_data = ea.draft_data if isinstance(ea.draft_data, dict) else {}
        published_data = ea.published_data if isinstance(ea.published_data, dict) else {}
        draft_marks_map = draft_data.get('marks', {}) if isinstance(draft_data.get('marks', {}), dict) else {}
        published_marks_map = published_data.get('marks', {}) if isinstance(published_data.get('marks', {}), dict) else {}
        # Prefer draft values whenever the exam is currently editable.
        # This fixes the "approved edit request => blank/old table" issue by showing
        # the per-student draft rows (`AcV2DraftMark`) and draft snapshot first.
        is_currently_editable = ea.is_editable()
        if is_currently_editable:
            active_snapshot = draft_marks_map
        else:
            active_snapshot = published_marks_map if ea.status in ('PUBLISHED', 'LOCKED') else draft_marks_map

        students = []
        for sa in assignments:
            sp = sa.student
            sm = existing.get(str(sp.id))
            dm = draft_existing.get(str(sp.id))
            snapshot = active_snapshot.get(str(sp.id)) or draft_marks_map.get(str(sp.id)) or published_marks_map.get(str(sp.id))
            if not isinstance(snapshot, dict):
                snapshot = {}

            mark_val = None
            co_marks_val = {}
            is_absent_val = False

            # Priority: whenever editable, draft overrides published.
            if is_currently_editable and dm:
                mark_val = float(dm.total_mark) if dm.total_mark is not None else None
                co_marks_val = dm.question_marks if isinstance(dm.question_marks, dict) else {}
                is_absent_val = bool(dm.is_absent)
            elif sm:
                mark_val = float(sm.total_mark) if sm.total_mark is not None else None
                co_marks_val = sm.question_marks if isinstance(sm.question_marks, dict) else {}
                is_absent_val = bool(sm.is_absent)
            elif dm:
                mark_val = float(dm.total_mark) if dm.total_mark is not None else None
                co_marks_val = dm.question_marks if isinstance(dm.question_marks, dict) else {}
                is_absent_val = bool(dm.is_absent)
            else:
                mark_val = _normalize_mark_number(snapshot.get('mark'))
                maybe_co_marks = snapshot.get('co_marks', {})
                co_marks_val = maybe_co_marks if isinstance(maybe_co_marks, dict) else {}
                is_absent_val = bool(snapshot.get('is_absent', False))

            students.append({
                'id': str(sp.id),
                'roll_number': sp.reg_no or '',
                'name': str(sp.user) if sp.user else sp.reg_no or '',
                'department': sp.home_department.name if sp.home_department else 'N/A',
                'mark': mark_val,
                'co_marks': co_marks_val,
                'is_absent': is_absent_val,
                'saved': True,
            })

        return Response({'students': students})

    cycle_state = _get_exam_cycle_state(
        ea,
        semester_id=getattr(getattr(ea.section, 'course', None), 'semester_id', None),
        class_type=getattr(getattr(ea.section, 'course', None), 'class_type', None),
    )
    if cycle_state['cycle_locked'] and not _has_admin_bypass_access(request.user):
        return Response({'detail': cycle_state['cycle_lock_reason'] or 'This cycle is locked for this semester.'}, status=403)

    # POST — save marks
    if not _has_admin_bypass_access(request.user) and not ea.is_editable():
        return Response({'detail': 'Exam is locked'}, status=403)

    raw_marks_data = request.data.get('marks', [])
    question_btls = request.data.get('question_btls', {})
    previous_marks_for_notification = {}
    if isinstance(ea.draft_data, dict):
        previous_marks_for_notification = ea.draft_data.get('marks', {}) if isinstance(ea.draft_data.get('marks', {}), dict) else {}
    if not previous_marks_for_notification and isinstance(ea.published_data, dict):
        previous_marks_for_notification = ea.published_data.get('marks', {}) if isinstance(ea.published_data.get('marks', {}), dict) else {}
    # Note: publish action is handled via /exams/<id>/publish/ (publish control)
    # Draft save: persist separately (draft_marks + draft_data). Do not touch
    # published rows here; publishing is handled via /publish/.
    marks_map = _save_draft_marks_for_exam(ea, raw_marks_data)

    # Save question_btls + marks snapshot in draft_data
    draft = ea.draft_data if isinstance(ea.draft_data, dict) else {}
    draft['marks'] = marks_map
    draft['question_btls'] = question_btls
    ea.draft_data = draft
    ea.last_saved_at = timezone.now()
    ea.last_saved_by = request.user
    
    # Auto-publish reopening logic:
    # If exam was auto-published (status=PUBLISHED, edit_window_until_publish=False)
    # and faculty saves new marks, revert to DRAFT to allow continued editing.
    # This differs from manual publish which requires edit request workflow.
    ctrl = check_publish_control(ea)
    if ea.status == 'PUBLISHED' and ctrl.get('publish_control_enabled'):
        # Reopen only auto-published exams (published_by is NULL). Manually published
        # exams must stay published and editable only within approved edit windows.
        if getattr(ea, 'published_by_id', None) is None and not bool(getattr(ea, 'edit_window_until_publish', False)):
            ea.status = 'DRAFT'
    
    ea.save(update_fields=['draft_data', 'last_saved_at', 'last_saved_by', 'status'])

    row_filled_notification_count = 0
    try:
        row_filled_notification_count = _send_student_row_filled_notifications(
            exam_assignment=ea,
            actor_user=request.user,
            previous_marks=previous_marks_for_notification,
            current_marks=marks_map,
        ) or 0
    except Exception as e:
        logger = logging.getLogger('academic_v2.notifications')
        logger.error(f'Row-filled notification hook failed: {e}')

    return Response({
        'status': 'saved',
        'row_filled_notifications_sent': int(row_filled_notification_count),
    })


# ============================================================================
# FACULTY CONFIRM MARK MANAGER (user_define mode)
# ============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def faculty_exam_confirm_mark_manager(request, exam_id):
    """
    Faculty confirms their Mark Manager CO setup (user_define mode).
    Generates question rows from their config and updates the QP pattern + ExamAssignment.
    """
    ea_qs = AcV2ExamAssignment.objects.select_related('section')
    if _has_admin_bypass_access(request.user):
        ea = get_object_or_404(ea_qs, id=exam_id)
    else:
        ea = get_object_or_404(ea_qs, id=exam_id, section__faculty_user=request.user)

    # Use the same editability check as mark saving — respects publish control settings,
    # open_from gating, edit windows, and draft/published status uniformly.
    if not _has_admin_bypass_access(request.user) and not ea.is_editable():
        return Response({'detail': 'Exam is locked'}, status=400)

    config = request.data.get('mark_manager')
    if not config or not isinstance(config, dict):
        return Response({'detail': 'mark_manager config required'}, status=400)

    # Build question rows from the faculty's config
    cos_config = config.get('cos', {})
    cia_enabled = config.get('cia_enabled', False)
    cia_max_marks = config.get('cia_max_marks', 0)

    titles = []
    marks = []
    cos_list = []
    btls = []
    enabled = []

    # CO rows first
    total_max = 0
    covered_cos = []
    for co_str in sorted(cos_config.keys(), key=lambda x: int(x)):
        co_num = int(co_str)
        co_cfg = cos_config[co_str]
        if not co_cfg.get('enabled'):
            continue
        covered_cos.append(co_num)
        num_items = co_cfg.get('num_items', 1)
        per_item_max = co_cfg.get('max_marks', 0)  # max marks PER item
        total_max += per_item_max * num_items
        for i in range(num_items):
            titles.append(f'CO{co_num} - Item {i + 1}')
            marks.append(per_item_max)
            cos_list.append(co_num)
            btls.append(None)
            enabled.append(True)

    # Exam column after all CO items, before Total
    if cia_enabled and cia_max_marks > 0:
        titles.append('Exam')
        marks.append(cia_max_marks)
        cos_list.append(None)
        btls.append(None)
        enabled.append(True)
        total_max += cia_max_marks

    # Store the user-defined pattern in ExamAssignment draft_data (NOT in global QP pattern!)
    # This keeps Mark Manager config per-course/per-exam, independent from other courses.
    new_pattern = {
        'titles': titles,
        'marks': marks,
        'cos': cos_list,
        'btls': btls,
        'enabled': enabled,
        'mark_manager': {
            **config,
            'confirmed': True,
        },
    }

    # Save to draft_data so it's course-specific
    draft = ea.draft_data if isinstance(ea.draft_data, dict) else {}
    draft['user_pattern'] = new_pattern  # Store pattern here, NOT in global QP pattern
    ea.draft_data = draft

    # Update ExamAssignment max_marks and covered_cos
    ea.max_marks = total_max
    ea.covered_cos = covered_cos
    ea.save(update_fields=['draft_data', 'max_marks', 'covered_cos'])

    # Build questions for the response
    questions = []
    for i in range(len(titles)):
        questions.append({
            'id': f'q{i}',
            'question_number': titles[i],
            'max_marks': marks[i],
            'btl_level': btls[i],
            'co_number': cos_list[i] if cos_list[i] is not None else 0,
        })

    return Response({
        'ok': True,
        'max_marks': total_max,
        'questions': questions,
    })


# ============================================================================
# FACULTY CO-WISE SUMMARY (for InternalMarkPage consolidated view)
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_course_co_summary(request, ta_id):
    """
    Return CO-wise, exam-wise mark summary for all students.
    Shows raw marks per CO per exam, weighted marks, CO totals, final mark.
    """
    from academics.models import TeachingAssignment, StudentSectionAssignment, StudentProfile
    from decimal import Decimal
    from .models import AcV2CqiAttained
    import ast
    import math

    ta_qs = TeachingAssignment.objects.select_related(
        'curriculum_row', 'elective_subject', 'section',
        'section__semester', 'section__managing_department',
        'staff',
    )
    if _has_admin_bypass_access(request.user):
        ta = get_object_or_404(ta_qs, id=ta_id, is_active=True)
    else:
        ta = get_object_or_404(ta_qs, id=ta_id, staff__user=request.user, is_active=True)

    cr = ta.curriculum_row
    es = ta.elective_subject
    sec = ta.section

    course_code = (cr.course_code if cr else None) or (getattr(es, 'course_code', None)) or '-'
    course_name = (cr.course_name if cr else None) or (getattr(es, 'course_name', None)) or '-'

    # QP type: from curriculum row (authoritative) or elective subject
    qp_type_code = (
        getattr(cr, 'question_paper_type', None)
        or getattr(es, 'question_paper_type', None)
        or ''
    ).strip()

    # Get AcV2Section(s) for this TA
    acv2_sections = AcV2Section.objects.filter(teaching_assignment=ta).select_related('course__class_type')
    if not acv2_sections.exists():
        return Response({
            'course_code': course_code,
            'course_name': course_name,
            'co_count': 5,
            'total_internal_marks': 40,
            'exams': [],
            'students': [],
        })

    acv2_section = acv2_sections.first()
    acv2_course = acv2_section.course
    class_type = acv2_course.class_type
    co_count = acv2_course.co_count or 5
    total_internal = float(class_type.total_internal_marks) if class_type else 40

    # Get exam assignments for this section, filtered to the course QP type.
    # Otherwise, legacy exam assignments (SSA/CIA/Model etc.) bleed into the table.
    exam_qs = AcV2ExamAssignment.objects.filter(section=acv2_section)
    if qp_type_code:
        exam_qs = exam_qs.filter(qp_type__iexact=qp_type_code)
    exam_assignments = list(exam_qs.order_by('created_at'))

    # Build qp_type-specific effective configs for weight/ordering.
    # Prefer class_type.exam_assignments filtered by qp_type.
    # If missing/mismatched, derive the exam list from QP patterns.
    ct_ea_configs = []
    if class_type and class_type.exam_assignments:
        qp_norm = (qp_type_code or '').strip().lower()
        for ea_conf in class_type.exam_assignments:
            if not isinstance(ea_conf, dict):
                continue
            if ea_conf.get('enabled', True) is False:
                continue
            ea_qp = (ea_conf.get('qp_type', '') or '').strip().lower()
            if not qp_norm or ea_qp == qp_norm:
                ct_ea_configs.append(ea_conf)

        ct_ea_configs.sort(key=lambda x: (
            (x.get('order') if isinstance(x.get('order'), int) else 10**9),
            str(x.get('exam_display_name') or x.get('exam') or ''),
        ))

    derived_ea_configs = []
    if class_type and qp_type_code and len(ct_ea_configs) == 0:
        patterns_qs = AcV2QpPattern.objects.filter(
            is_active=True,
            qp_type__iexact=qp_type_code,
            class_type=class_type,
        ).order_by('order', 'created_at')
        if not patterns_qs.exists():
            patterns_qs = AcV2QpPattern.objects.filter(
                is_active=True,
                qp_type__iexact=qp_type_code,
                class_type__isnull=True,
            ).order_by('order', 'created_at')

        # Optional: reuse existing weights by matching display names
        weight_by_name = {}
        if class_type.exam_assignments:
            for ea_conf in class_type.exam_assignments:
                nm = (ea_conf.get('exam_display_name') or ea_conf.get('exam') or '').strip().lower()
                if nm and nm not in weight_by_name:
                    weight_by_name[nm] = ea_conf

        for p in patterns_qs:
            exam_name = (p.name or '').strip()
            if not exam_name:
                continue
            # CQI exams must be configured via the CQI editor (ClassType exam_assignments)
            # and should never be auto-derived from QP patterns.
            if exam_name.strip().lower().startswith('cqi'):
                continue
            w_conf = weight_by_name.get(exam_name.lower())
            derived_ea_configs.append({
                'exam': exam_name,
                'exam_display_name': exam_name,
                'qp_type': qp_type_code,
                'weight': (w_conf.get('weight') if isinstance(w_conf, dict) else None) or float(p.default_weight or 0),
                'co_weights': (w_conf.get('co_weights') if isinstance(w_conf, dict) else None) or {},
                'default_cos': (w_conf.get('default_cos') if isinstance(w_conf, dict) else None) or [],
                # Mark Manager conditional (optional)
                'mm_co_weights_with_exam': (w_conf.get('mm_co_weights_with_exam') if isinstance(w_conf, dict) else None) or {},
                'mm_co_weights_without_exam': (w_conf.get('mm_co_weights_without_exam') if isinstance(w_conf, dict) else None) or {},
                'mm_exam_weight': (w_conf.get('mm_exam_weight') if isinstance(w_conf, dict) else None) or 0,
            })

    def _ea_config_key(ea_conf):
        if not isinstance(ea_conf, dict):
            return ''
        return str(ea_conf.get('exam_display_name') or ea_conf.get('exam') or '').strip().lower()

    effective_ea_configs = []
    seen_ea_config_keys = set()
    for ea_conf in ct_ea_configs + derived_ea_configs:
        cfg_key = _ea_config_key(ea_conf)
        if cfg_key:
            if cfg_key in seen_ea_config_keys:
                continue
            seen_ea_config_keys.add(cfg_key)
        effective_ea_configs.append(ea_conf)

    # Normalization helper for matching exams between DB and ClassType config
    norm_exam_key = lambda s: re.sub(r'[^a-z0-9]+', '', str(s or '').strip().lower())

    # Filter out stale AcV2ExamAssignment records that are no longer part of the
    # effective qp_type exam configuration (i.e. deleted/removed in QP pattern).
    # Keep legacy behaviour when we have no effective configs to compare against.
    if effective_ea_configs:
        allowed_keys = set()
        for ea_conf in effective_ea_configs:
            if not isinstance(ea_conf, dict):
                continue
            for key in [ea_conf.get('exam_display_name'), ea_conf.get('exam')]:
                k = norm_exam_key(key)
                if k:
                    allowed_keys.add(k)
        exam_assignments = [
            ea for ea in exam_assignments
            if norm_exam_key(getattr(ea, 'exam_display_name', '') or getattr(ea, 'exam', '') or '') in allowed_keys
        ]

    # DB-backed CQI config lookup (authoritative for CQI CO selections/conditions)
    cqi_db_by_key = {}
    try:
        from .models import AcV2CqiExam
        if class_type and qp_type_code:
            for row in AcV2CqiExam.objects.filter(
                class_type=class_type,
                qp_type_code__iexact=qp_type_code,
                is_active=True,
            ).order_by('order', 'created_at'):
                for key in [row.exam_code, row.exam_display_name]:
                    k = norm_exam_key(key)
                    if k:
                        cqi_db_by_key[k] = row
    except Exception:
        cqi_db_by_key = {}

    # Helper: determine kind + CQI subconfig from class-type exam_assignments
    ct_kind_by_key = {}
    if effective_ea_configs:
        for ea_conf in effective_ea_configs:
            try:
                kind = 'cqi' if str(ea_conf.get('kind', '')).lower() == 'cqi' else 'exam'
            except Exception:
                kind = 'exam'
            cqi_sub = ea_conf.get('cqi', {}) if isinstance(ea_conf.get('cqi'), dict) else {}
            for key in [ea_conf.get('exam'), ea_conf.get('exam_display_name')]:
                k = norm_exam_key(key)
                if not k:
                    continue
                # Prefer DB-backed CQI config (when present) to avoid drift in CO/condition symbols.
                if kind == 'cqi' and k in cqi_db_by_key:
                    row = cqi_db_by_key.get(k)
                    if row is not None:
                        cqi_sub = {
                            'name': getattr(row, 'cqi_name', '') or '',
                            'code': getattr(row, 'cqi_code', '') or '',
                            'cycle_id': getattr(row, 'cycle_id', '') or '',
                            'cos': getattr(row, 'cos', []) or [],
                            'exams': getattr(row, 'considered_exams', []) or [],
                            'custom_vars': getattr(row, 'custom_vars', []) or [],
                            'global_custom_vars': getattr(row, 'global_custom_vars', []) or [],
                            'derived_variables': getattr(row, 'derived_variables', []) or [],
                            'co_value_expr': getattr(row, 'co_value_expr', '') or '',
                            'formula': getattr(row, 'formula', '') or '',
                            'conditions': getattr(row, 'conditions', []) or [],
                            'else_formula': getattr(row, 'else_formula', '') or '',
                        }
                if k not in ct_kind_by_key:
                    ct_kind_by_key[k] = { 'kind': kind, 'cqi': cqi_sub }

    # Build weight lookup from effective configs (qp_type-specific)
    ct_weight_map = {}
    ct_co_weights_map = {}  # exam -> {co_num: weight}
    ct_index = {}
    ct_mm_co_weights_with_exam_map = {}  # exam -> {co_num: weight}
    ct_mm_co_weights_without_exam_map = {}  # exam -> {co_num: weight}
    ct_mm_exam_weight_map = {}  # exam -> exam_weight
    ct_mm_exam_weight_per_co_map = {}  # exam -> bool
    ct_mm_enabled_map = {}  # exam -> bool
    if effective_ea_configs:
        for i, ea_conf in enumerate(effective_ea_configs):
            exam_code = ea_conf.get('exam', '')
            exam_display = ea_conf.get('exam_display_name', exam_code)
            mm_enabled = ea_conf.get('mark_manager_enabled', False)
            for key in [exam_code, exam_display]:
                k = norm_exam_key(key)
                if not k:
                    continue
                if k not in ct_index:
                    ct_index[k] = i
                ct_weight_map[k] = ea_conf.get('weight', 0)
                ct_mm_enabled_map[k] = bool(mm_enabled)
            co_weights = ea_conf.get('co_weights', {})
            if co_weights:
                w = {int(k): v for k, v in co_weights.items()}
                for key in [exam_code, exam_display]:
                    k = norm_exam_key(key)
                    if k:
                        ct_co_weights_map[k] = w
 
            # Mark Manager conditional config (optional)
            mm_on = ea_conf.get('mm_co_weights_with_exam')
            mm_off = ea_conf.get('mm_co_weights_without_exam')
            mm_exam_weight = ea_conf.get('mm_exam_weight')
            mm_exam_weight_per_co = ea_conf.get('mm_exam_weight_per_co', False)
            # Backward compatibility: allow nested keys
            if not mm_on and isinstance(ea_conf.get('mm_with_exam'), dict):
                mm_on = ea_conf.get('mm_with_exam', {}).get('co_weights')
                mm_exam_weight = ea_conf.get('mm_with_exam', {}).get('exam_weight', mm_exam_weight)
            if not mm_off and isinstance(ea_conf.get('mm_without_exam'), dict):
                mm_off = ea_conf.get('mm_without_exam', {}).get('co_weights')
 
            if isinstance(mm_on, dict) and mm_on:
                w_on = {int(k): v for k, v in mm_on.items()}
                for key in [exam_code, exam_display]:
                    k = norm_exam_key(key)
                    if k:
                        ct_mm_co_weights_with_exam_map[k] = w_on
            if isinstance(mm_off, dict) and mm_off:
                w_off = {int(k): v for k, v in mm_off.items()}
                for key in [exam_code, exam_display]:
                    k = norm_exam_key(key)
                    if k:
                        ct_mm_co_weights_without_exam_map[k] = w_off
            for key in [exam_code, exam_display]:
                k = norm_exam_key(key)
                if k:
                    ct_mm_exam_weight_per_co_map[k] = bool(mm_exam_weight_per_co)
            if mm_exam_weight is not None:
                try:
                    mm_w = float(mm_exam_weight) or 0
                except Exception:
                    mm_w = 0
                for key in [exam_code, exam_display]:
                    k = norm_exam_key(key)
                    if k:
                        ct_mm_exam_weight_map[k] = mm_w

    # Order exams using effective config order when available
    if ct_index:
        exam_assignments.sort(key=lambda ea: (
            ct_index.get(norm_exam_key(getattr(ea, 'exam_display_name', '') or getattr(ea, 'exam', '') or ''), 10**9),
            getattr(ea, 'created_at', None) or timezone.now(),
        ))

    # Deduplicate exam_assignments by normalised display name (same as faculty_course_info).
    # Prefer records with marks entered so progress is not lost.
    exam_assignments.sort(key=lambda ea: (
        ct_index.get(norm_exam_key(getattr(ea, 'exam_display_name', '') or getattr(ea, 'exam', '') or ''), 10**9),
        -len((ea.draft_data or {}).get('marks', {})) if isinstance(ea.draft_data, dict) else 0,
        getattr(ea, 'created_at', None) or timezone.now(),
    ))
    _co_sum_seen_keys = set()
    _co_sum_deduped = []
    for _co_sum_ea in exam_assignments:
        _co_sum_dk = norm_exam_key(_co_sum_ea.exam_display_name or _co_sum_ea.exam)
        if _co_sum_dk in _co_sum_seen_keys:
            continue
        _co_sum_seen_keys.add(_co_sum_dk)
        _co_sum_deduped.append(_co_sum_ea)
    exam_assignments = _co_sum_deduped

    exams_data = []
    exam_map = {}  # exam_id -> exam info

    first_cqi_sub = None
    first_cqi_exam_id = None

    # Published CQI snapshot for this TA (if any)
    cqi_attained = AcV2CqiAttained.objects.filter(teaching_assignment=ta).first()
    cqi_entries = cqi_attained.entries if cqi_attained and isinstance(cqi_attained.entries, dict) else {}

    # CQI expression evaluation (used for Internal Marks CQI values)
    def _normalize_cqi_expr(expr: str, vars_map: dict) -> str:
        s = str(expr or '').strip()
        if not s:
            return ''
        # CQI editor token picker can build adjacent token expressions that
        # semantically mean addition.
        s = re.sub(r'\]\s+\[', '] + [', s)
        # Replace token form [TOKEN] with its numeric value (or 0 if unknown).
        # This avoids Python identifier limitations for tokens containing '-' (e.g. COX-SSA_1-WEIGHT).
        def repl(m):
            key = str(m.group(1) or '').strip().upper()
            try:
                val = float((vars_map or {}).get(key, 0) or 0)
            except Exception:
                val = 0.0
            # Use a plain numeric literal so ast.parse sees a number.
            return str(val)
        return re.sub(r'\[([A-Za-z0-9_-]+)\]', repl, s)

    def _build_cqi_if_from_clauses(clauses) -> str:
        """
        Build a boolean expression string from if_clauses list.

        Each clause supports two formats:
          NEW: { token, operator, rhs }  — operator is a separate field (e.g. '<', '<=')
          OLD: { token, rhs }            — operator is embedded at start of rhs (e.g. '< 58')

        Backward-compatible: old format is detected automatically.
        Supports any token registered in the system (not limited to 3 core tokens).
        """
        parts = []
        for idx, clause in enumerate(clauses or []):
            if not isinstance(clause, dict):
                continue
            token = str(clause.get('token') or '').strip().upper()
            # RAW clause: token empty => rhs is a full boolean expression.
            if not token:
                rhs_raw = str(clause.get('rhs') or '').strip()
                if rhs_raw:
                    rhs_raw = re.sub(r'\]\s+\[', '] + [', rhs_raw)
                    parts.append(f'({rhs_raw})')
                continue
            if not token:
                continue

            # Determine rhs and operator (new format has explicit 'operator' field)
            operator = str(clause.get('operator') or '').strip()
            if operator == '=':
                operator = '=='
            rhs = str(clause.get('rhs') or '').strip()

            if operator:
                # NEW format: operator and rhs are separate
                combined_rhs = f'{operator} {rhs}'.strip() if rhs else operator
            else:
                # OLD format: operator is embedded in rhs (e.g. '< 58')
                combined_rhs = rhs

            if not combined_rhs:
                continue

            combined_rhs = re.sub(r'\]\s+\[', '] + [', combined_rhs)

            if idx == 0 and token in ('BEFORE_CQI', 'BEFORE_CQI_COX'):
                # Legacy support: some configs keep a full boolean expression in rhs
                # and use '=' as a connector. In that case evaluate rhs directly.
                if operator in ('', '==') and re.search(r'(<=|>=|==|!=|=|<|>)', rhs):
                    parts.append(f'({rhs})')
                    continue
                # First clause legacy behaviour: if rhs starts with a comparator, wrap token+rhs
                is_comparator_only = bool(re.match(r'^(<=|>=|==|!=|=|<|>)', combined_rhs))
                if is_comparator_only:
                    parts.append(f'([{token}] {combined_rhs})')
                else:
                    # Backward-compat for malformed saved clauses with missing operator.
                    parts.append(f'([{token}] < {combined_rhs})')
            else:
                is_comparator_only = bool(re.match(r'^(<=|>=|==|!=|=|<|>)', combined_rhs))
                parts.append(f'([{token}] {combined_rhs})' if is_comparator_only else f'([{token}] < {combined_rhs})')
        return ' && '.join(p for p in parts if p)

    def _resolve_cqi_if_expr(cond: dict) -> str:
        if not isinstance(cond, dict):
            return ''
        direct_if = str(cond.get('if', '') or '').strip()
        if direct_if:
            return direct_if
        clauses = cond.get('if_clauses') if isinstance(cond.get('if_clauses'), list) else []
        if clauses:
            built = _build_cqi_if_from_clauses(clauses)
            if built:
                return built
        return direct_if

    _ALLOWED_FUNCS = {
        'min': min,
        'max': max,
        'abs': abs,
        'round': round,
        'sqrt': math.sqrt,
        'floor': math.floor,
        'ceil': math.ceil,
    }

    def _normalize_cqi_token_code(value: str) -> str:
        return re.sub(r'^_+|_+$', '', re.sub(r'[^A-Z0-9]+', '_', str(value or '').strip().upper()))

    def _normalize_cqi_custom_vars(values) -> list:
        out = []
        if not isinstance(values, list):
            return out
        for item in values:
            if not isinstance(item, dict):
                continue
            code = _normalize_cqi_token_code(item.get('code') or '')
            label = str(item.get('label') or '')
            expr = str(item.get('expr') or '')
            if not code and not label and not expr:
                continue
            out.append({
                'code': code,
                'label': label,
                'expr': expr,
            })
        return out

    def _co_list_from_value(raw_co):
        if isinstance(raw_co, list):
            result = []
            for val in raw_co:
                try:
                    num = int(val)
                except Exception:
                    continue
                if 1 <= num <= co_count:
                    result.append(num)
            return result
        try:
            num = int(raw_co)
        except Exception:
            return []
        return [num] if 1 <= num <= co_count else []

    def _count_mark_manager_items_for_co(exam_info: dict, co_num: int) -> int:
        qp_cos_local = exam_info.get('_qp_cos') or []
        qp_enabled_local = exam_info.get('_qp_enabled') or [True] * len(qp_cos_local)
        count = 0
        for idx, raw_co in enumerate(qp_cos_local):
            if idx < len(qp_enabled_local) and not qp_enabled_local[idx]:
                continue
            if co_num in _co_list_from_value(raw_co):
                count += 1
        return count

    def _safe_eval_cqi_num(expr: str, vars_map: dict) -> float:
        """Safely evaluate a numeric expression for CQI mapping."""
        expr_n = _normalize_cqi_expr(expr, vars_map)
        if not expr_n:
            return float(vars_map.get('CQI', 0) or 0)
        try:
            tree = ast.parse(expr_n, mode='eval')
        except Exception:
            return float(vars_map.get('CQI', 0) or 0)

        def _eval(node):
            if isinstance(node, ast.Expression):
                return _eval(node.body)
            if isinstance(node, ast.Constant):
                return float(node.value) if isinstance(node.value, (int, float)) else 0.0
            if isinstance(node, ast.Num):
                return float(node.n)
            if isinstance(node, ast.Name):
                key = str(node.id or '').upper()
                try:
                    return float(vars_map.get(key, 0) or 0)
                except Exception:
                    return 0.0
            if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
                v = _eval(node.operand)
                return v if isinstance(node.op, ast.UAdd) else -v
            if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Mod, ast.Pow)):
                a = _eval(node.left)
                b = _eval(node.right)
                if isinstance(node.op, ast.Add):
                    return a + b
                if isinstance(node.op, ast.Sub):
                    return a - b
                if isinstance(node.op, ast.Mult):
                    return a * b
                if isinstance(node.op, ast.Div):
                    return a / b if b else 0.0
                if isinstance(node.op, ast.Mod):
                    return a % b if b else 0.0
                if isinstance(node.op, ast.Pow):
                    return a ** b
            if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
                fn = _ALLOWED_FUNCS.get(node.func.id)
                if not fn:
                    return 0.0
                args = [_eval(a) for a in (node.args or [])]
                try:
                    return float(fn(*args))
                except Exception:
                    return 0.0
            return 0.0

        try:
            return float(_eval(tree))
        except Exception:
            return float(vars_map.get('CQI', 0) or 0)

    def _safe_eval_cqi_bool(expr: str, vars_map: dict) -> bool:
        """Safely evaluate a boolean expression (comparisons + AND/OR) for CQI conditions."""
        expr_n = _normalize_cqi_expr(expr, vars_map)
        if not expr_n:
            return False
        # Support JS-style and word-style boolean operators from saved configs.
        expr_n = expr_n.replace('&&', ' and ')
        expr_n = expr_n.replace('||', ' or ')
        expr_n = re.sub(r'\bAND\b', 'and', expr_n, flags=re.IGNORECASE)
        expr_n = re.sub(r'\bOR\b', 'or', expr_n, flags=re.IGNORECASE)
        # Accept single '=' from stored configs as equality.
        expr_n = re.sub(r'(?<![<>!=])=(?!=)', '==', expr_n)
        try:
            tree = ast.parse(expr_n, mode='eval')
        except Exception:
            return False

        def _eval(node):
            if isinstance(node, ast.Expression):
                return _eval(node.body)
            if isinstance(node, ast.Constant):
                return node.value
            if isinstance(node, ast.Num):
                return node.n
            if isinstance(node, ast.Name):
                key = str(node.id or '').upper()
                try:
                    return float(vars_map.get(key, 0) or 0)
                except Exception:
                    return 0.0
            if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub, ast.Not)):
                v = _eval(node.operand)
                if isinstance(node.op, ast.Not):
                    return not bool(v)
                try:
                    v = float(v or 0)
                except Exception:
                    v = 0.0
                return v if isinstance(node.op, ast.UAdd) else -v
            if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Mod, ast.Pow)):
                a = _eval(node.left)
                b = _eval(node.right)
                try:
                    a = float(a or 0)
                except Exception:
                    a = 0.0
                try:
                    b = float(b or 0)
                except Exception:
                    b = 0.0
                if isinstance(node.op, ast.Add):
                    return a + b
                if isinstance(node.op, ast.Sub):
                    return a - b
                if isinstance(node.op, ast.Mult):
                    return a * b
                if isinstance(node.op, ast.Div):
                    return a / b if b else 0.0
                if isinstance(node.op, ast.Mod):
                    return a % b if b else 0.0
                if isinstance(node.op, ast.Pow):
                    return a ** b
            if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
                fn = _ALLOWED_FUNCS.get(node.func.id)
                if not fn:
                    return 0.0
                args = [_eval(a) for a in (node.args or [])]
                try:
                    return float(fn(*args))
                except Exception:
                    return 0.0
            if isinstance(node, ast.BoolOp) and isinstance(node.op, (ast.And, ast.Or)):
                vals = [_eval(v) for v in (node.values or [])]
                if isinstance(node.op, ast.And):
                    return all(bool(v) for v in vals)
                return any(bool(v) for v in vals)
            if isinstance(node, ast.Compare):
                left = _eval(node.left)
                try:
                    left_num = float(left or 0)
                except Exception:
                    left_num = 0.0
                ok = True
                for op, comp in zip(node.ops or [], node.comparators or []):
                    right = _eval(comp)
                    try:
                        right_num = float(right or 0)
                    except Exception:
                        right_num = 0.0
                    if isinstance(op, ast.Eq):
                        ok = ok and (left_num == right_num)
                    elif isinstance(op, ast.NotEq):
                        ok = ok and (left_num != right_num)
                    elif isinstance(op, ast.Lt):
                        ok = ok and (left_num < right_num)
                    elif isinstance(op, ast.LtE):
                        ok = ok and (left_num <= right_num)
                    elif isinstance(op, ast.Gt):
                        ok = ok and (left_num > right_num)
                    elif isinstance(op, ast.GtE):
                        ok = ok and (left_num >= right_num)
                    else:
                        ok = False
                    left_num = right_num
                return ok
            return False

        try:
            return bool(_eval(tree))
        except Exception:
            return False

    for ea in exam_assignments:
        def _extract_co_list(raw_co):
            if raw_co is None:
                return []
            out = []
            if isinstance(raw_co, (list, tuple, set)):
                items = list(raw_co)
            else:
                items = [raw_co]

            for item in items:
                if item is None:
                    continue
                if isinstance(item, str):
                    parts = re.split(r'[^0-9]+', item)
                    for p in parts:
                        if not p:
                            continue
                        try:
                            n = int(p)
                        except Exception:
                            continue
                        if 1 <= n <= co_count and n not in out:
                            out.append(n)
                    continue
                try:
                    n = int(item)
                except Exception:
                    continue
                if 1 <= n <= co_count and n not in out:
                    out.append(n)
            return out

        covered_cos = ea.covered_cos or []
        weight = float(ea.weight) if ea.weight else 0
        max_marks = float(ea.max_marks) if ea.max_marks else 0
        ea_key = norm_exam_key(ea.exam_display_name or ea.exam)

        # Determine kind + CQI metadata from ClassType config
        ea_kind = 'exam'
        cqi_sub = {}
        if ea_key in ct_kind_by_key:
            ea_kind = ct_kind_by_key[ea_key].get('kind') or 'exam'
            cqi_sub = ct_kind_by_key[ea_key].get('cqi') or {}
        if ea_kind == 'cqi':
            try:
                _cqi_cos = cqi_sub.get('cos', []) if isinstance(cqi_sub, dict) else []
            except Exception:
                _cqi_cos = []
            if isinstance(_cqi_cos, list) and _cqi_cos:
                covered_cos = [int(x) for x in _cqi_cos if str(x).isdigit() and 1 <= int(x) <= co_count]

            # Ensure CQI COs are stable and do not depend on stale covered_cos/user patterns.
            # Persist the corrected covered_cos back to the exam assignment to prevent drift.
            try:
                if covered_cos and (ea.covered_cos or []) != covered_cos:
                    ea.covered_cos = covered_cos
                    ea.save(update_fields=['covered_cos'])
            except Exception:
                pass

            if first_cqi_sub is None and isinstance(cqi_sub, dict):
                first_cqi_sub = cqi_sub
                first_cqi_exam_id = str(getattr(ea, 'id', '') or '')

        # Resolve weight from ClassType config if ea.weight is 0
        if weight == 0 and ea_key in ct_weight_map and ct_weight_map[ea_key]:
            weight = float(ct_weight_map[ea_key])
            # Sync back to DB record
            ea.weight = weight
            ea.save(update_fields=['weight'])

        # Check for course-specific pattern from Mark Manager first
        draft = ea.draft_data if isinstance(ea.draft_data, dict) else {}
        user_pattern = draft.get('user_pattern')
        
        co_max_map = {}  # {co_num: total_marks_for_that_co}
        co_weights = {}  # Effective per-CO weights
        cia_enabled = False  # Whether Mark Manager has Exam enabled
        cia_weight = 0  # Exam component weight (admin-defined)
        cia_weight_per_co = False
        exam_max_marks = 0  # Exam component max marks (only when Mark Manager Exam is enabled)
        exam_q_index = None  # Internal: index of Exam question in question_marks (q{index})
        qp_marks = []
        qp_cos = []
        qp_enabled = []
        qp_special_split = []
        qp_special_split_sources = []
        
        if ea_kind == 'cqi':
            # CQI is not pattern-driven. COs come only from the CQI editor config (DB-backed),
            # and weights are computed from CQI evaluation, not from QP pattern CO mapping.
            pass
        elif ea_kind != 'cqi':
            # Resolve QP Pattern (use user-defined custom pattern first, else default pattern)
            qp_type_val = ea.qp_type or ea.exam or ''
            exam_label = (ea.exam_display_name or ea.exam or '').strip()
            qp_match = None
            if not (user_pattern and isinstance(user_pattern, dict)):
                qp_match = (
                    AcV2QpPattern.objects.filter(
                        name__iexact=exam_label, qp_type=qp_type_val,
                        class_type=class_type, is_active=True,
                    ).first()
                    or AcV2QpPattern.objects.filter(
                        name__iexact=exam_label, qp_type=qp_type_val,
                        is_active=True,
                    ).first()
                    or AcV2QpPattern.objects.filter(
                        qp_type=qp_type_val, class_type=class_type, is_active=True,
                    ).first()
                    or AcV2QpPattern.objects.filter(
                        qp_type=qp_type_val, is_active=True,
                    ).first()
                )

            p = user_pattern if (user_pattern and isinstance(user_pattern, dict)) else (qp_match.pattern if (qp_match and isinstance(qp_match.pattern, dict)) else {})
            qp_marks = p.get('marks', []) if isinstance(p.get('marks'), list) else []
            qp_cos = p.get('cos', []) if isinstance(p.get('cos'), list) else []
            qp_enabled = p.get('enabled', []) if isinstance(p.get('enabled'), list) else [True] * len(qp_marks)
            if len(qp_enabled) < len(qp_marks):
                qp_enabled = qp_enabled + [True] * (len(qp_marks) - len(qp_enabled))
            qp_special_split = p.get('special_split', []) if isinstance(p.get('special_split'), list) else []
            qp_special_split_sources = p.get('special_split_sources', []) if isinstance(p.get('special_split_sources'), list) else []

            # Derive max_marks
            derived_max = sum(m for i, m in enumerate(qp_marks) if i < len(qp_enabled) and qp_enabled[i])
            if derived_max > 0:
                max_marks = derived_max

            # Derive covered_cos
            derived_set = set()
            for i, c in enumerate(qp_cos):
                if c is None:
                    continue
                if i < len(qp_enabled) and not qp_enabled[i]:
                    continue
                for co_num in _extract_co_list(c):
                    derived_set.add(co_num)
            derived_cos = sorted(derived_set)
            if derived_cos:
                covered_cos = derived_cos

            # Build per-CO max marks
            for i, c in enumerate(qp_cos):
                if c is not None and i < len(qp_enabled) and qp_enabled[i]:
                    co_list = _extract_co_list(c)
                    if not co_list:
                        continue
                    split_max = (qp_marks[i] if i < len(qp_marks) else 0) / max(len(co_list), 1)
                    for co in co_list:
                        co_max_map[co] = co_max_map.get(co, 0) + split_max

            # Determine if Mark Manager is active for this exam
            is_mm = bool(ct_mm_enabled_map.get(ea_key, False))
            mm_config = p.get('mark_manager', {}) if isinstance(p.get('mark_manager'), dict) else {}
            if mm_config.get('enabled'):
                is_mm = True

            if is_mm:
                cia_enabled = bool(mm_config.get('cia_enabled', False))
                mm_cos_config = mm_config.get('cos', {}) if isinstance(mm_config.get('cos'), dict) else {}

                # Enabled COs (authoritative for Mark Manager)
                enabled_cos = []
                for co_str, co_cfg in mm_cos_config.items():
                    try:
                        co_num = int(co_str)
                    except Exception:
                        continue
                    if isinstance(co_cfg, dict) and co_cfg.get('enabled') and 1 <= co_num <= co_count:
                        enabled_cos.append(co_num)
                enabled_cos = sorted(set(enabled_cos))
                if enabled_cos:
                    covered_cos = enabled_cos

                # Find Exam max marks in user pattern (needed to scale redistributed Exam marks)
                exam_max = 0
                titles = p.get('titles', []) if isinstance(p.get('titles'), list) else []
                for i, t in enumerate(titles):
                    if isinstance(t, str) and t.strip().lower() == 'exam':
                        if i < len(qp_marks) and i < len(qp_enabled) and qp_enabled[i]:
                            try:
                                exam_max = float(qp_marks[i] or 0)
                            except Exception:
                                exam_max = 0
                            exam_q_index = i
                if exam_max == 0 and qp_marks and qp_cos and len(qp_marks) == len(qp_cos):
                    last_idx = len(qp_cos) - 1
                    if last_idx >= 0 and qp_cos[last_idx] is None and (last_idx < len(qp_enabled) and qp_enabled[last_idx]):
                        try:
                            exam_max = float(qp_marks[last_idx] or 0)
                        except Exception:
                            exam_max = 0
                        exam_q_index = last_idx

                exam_max_marks = float(exam_max or 0)

                if cia_enabled:
                    # CONDITION A: WITH Exam -> use admin-defined Mark Manager "with exam" weights
                    base = ct_mm_co_weights_with_exam_map.get(ea_key) or ct_co_weights_map.get(ea_key, {})
                    cia_weight = float(ct_mm_exam_weight_map.get(ea_key, 0) or 0)
                    cia_weight_per_co = bool(ct_mm_exam_weight_per_co_map.get(ea_key, False))

                    # Base CO weights
                    for co_num in covered_cos:
                        co_weights[int(co_num)] = float(base.get(int(co_num), 0) or 0)

                    if cia_weight_per_co:
                        weight = sum(float(v or 0) for v in co_weights.values()) + (float(cia_weight or 0) * len(covered_cos))
                    else:
                        weight = sum(float(v or 0) for v in co_weights.values()) + float(cia_weight or 0)
                else:
                    # CONDITION B: WITHOUT Exam -> use admin-defined Mark Manager "without exam" weights
                    base = ct_mm_co_weights_without_exam_map.get(ea_key) or ct_co_weights_map.get(ea_key, {})
                    for co_num in covered_cos:
                        co_weights[int(co_num)] = float(base.get(int(co_num), 0) or 0)
                    weight = sum(float(v or 0) for v in co_weights.values())
            else:
                # Fall back to global non-Mark Manager ClassType configuration
                co_weights = ct_co_weights_map.get(ea_key, {})
            


        # weight_per_co: for even split fallback when no per-CO weights defined
        if not co_weights and covered_cos:
            weight_per_co = round(weight / len(covered_cos), 2) if covered_cos else 0
        else:
            weight_per_co = 0  # Will use co_weights instead
        # max_per_co: fallback when co_max_map not available
        max_per_co = round(max_marks / len(covered_cos), 2) if covered_cos else max_marks

        combo_questions = []
        if isinstance(qp_cos, list) and qp_cos:
            def _num(x):
                try:
                    return float(x or 0)
                except Exception:
                    return 0.0

            def _safe_bool_at(arr, i: int) -> bool:
                return bool(isinstance(arr, list) and i < len(arr) and arr[i])

            def _safe_sources_at(arr, i: int):
                if not isinstance(arr, list) or i >= len(arr):
                    return []
                v = arr[i]
                return v if isinstance(v, list) else []

            def _special_split_max(i: int) -> float:
                # Derived max preview using QP pattern marks (not student marks)
                if not _safe_bool_at(qp_special_split, i):
                    return _num(qp_marks[i] if isinstance(qp_marks, list) and i < len(qp_marks) else 0)
                sources = _safe_sources_at(qp_special_split_sources, i)
                co_set = set()
                sum_src = 0.0
                for sidx in sources:
                    try:
                        j = int(sidx)
                    except Exception:
                        continue
                    if j == i or j < 0 or j >= len(qp_cos):
                        continue
                    if isinstance(qp_enabled, list) and j < len(qp_enabled) and not qp_enabled[j]:
                        continue
                    sum_src += _num(qp_marks[j] if isinstance(qp_marks, list) and j < len(qp_marks) else 0)
                    for c in _extract_co_list(qp_cos[j]):
                        co_set.add(int(c))
                denom = len(co_set) or 1
                base = _num(qp_marks[i] if isinstance(qp_marks, list) and i < len(qp_marks) else 0)
                return round(sum_src + (base / denom), 2)

            for i, co in enumerate(qp_cos):
                if i < len(qp_enabled) and qp_enabled and not qp_enabled[i]:
                    continue
                co_list = []
                if isinstance(co, (list, tuple, set)):
                    for item in co:
                        try:
                            n = int(item)
                        except Exception:
                            continue
                        if 1 <= n <= co_count and n not in co_list:
                            co_list.append(n)
                elif isinstance(co, str):
                    parts = re.split(r'[^0-9]+', co)
                    for p in parts:
                        if not p:
                            continue
                        try:
                            n = int(p)
                        except Exception:
                            continue
                        if 1 <= n <= co_count and n not in co_list:
                            co_list.append(n)
                else:
                    try:
                        n = int(co)
                    except Exception:
                        n = None
                    if n and 1 <= n <= co_count:
                        co_list.append(n)

                if len(co_list) >= 2:
                    max_q = _special_split_max(i)
                    combo_questions.append({
                        'key': f'combo_q{i}',
                        'co_list': co_list,
                        'max_marks': max_q,
                    })

        exam_info = {
            'id': str(ea.id),
            'name': ea.exam_display_name or ea.exam or ea.qp_type or '',
            'short_name': ea.exam or ea.qp_type or '',
            'max_marks': max_marks,
            'weight': weight,
            'co_weights': {} if ea_kind == 'cqi' else co_weights,  # Per-CO weights (from Mark Manager or admin config)
            'cia_enabled': cia_enabled,  # Whether Mark Manager Exam checkbox is enabled
            'cia_weight': cia_weight,  # Weight for Exam component from Mark Manager
            'cia_weight_per_co': cia_weight_per_co,
            'exam_max_marks': exam_max_marks,
            'covered_cos': covered_cos,
            'weight_per_co': weight_per_co,
            'max_per_co': max_per_co,
            'co_max_map': co_max_map,
            'status': ea.status,
            'combo_questions': combo_questions,
            'kind': ea_kind,
            'cqi_name': str(cqi_sub.get('name', '') or '') if isinstance(cqi_sub, dict) else '',
            'cqi_cos': covered_cos if ea_kind == 'cqi' else [],
        }
        exams_data.append(exam_info)
        # Keep internal fields for per-student recomputation from question_marks.
        # This avoids relying on stale co1..co5 columns when Mark Manager logic changes.
        internal = {
            '_exam_q_index': exam_q_index,
            '_qp_marks': qp_marks if isinstance(qp_marks, list) else [],
        }
        if ea_kind == 'cqi' and isinstance(cqi_sub, dict):
            internal['_cqi_sub'] = cqi_sub
        if isinstance(qp_cos, list) and qp_cos:
            internal['_qp_cos'] = qp_cos
            internal['_qp_enabled'] = qp_enabled
            internal['_qp_special_split'] = qp_special_split if isinstance(qp_special_split, list) else []
            internal['_qp_special_split_sources'] = qp_special_split_sources if isinstance(qp_special_split_sources, list) else []
        exam_map[str(ea.id)] = {**exam_info, **internal}

    # NOTE: The CO-mismatch check that used to wipe cqi_entries here has been removed.
    # With multiple CQI exams (CQI 1 = CO1,CO2 and CQI 2 = CO3,CO4,CO5), entries are
    # merged per-CO on publish, and co_numbers now represents the union of all published
    # CQI COs. Wiping all entries based on a single CQI's CO-set is incorrect.

    # Get all active students in the academic section
    student_assignments = (
        StudentSectionAssignment.objects
        .filter(section=sec, end_date__isnull=True)
        .select_related('student__user')
        .order_by('student__reg_no')
    )

    # Get all student marks across all exams at once
    all_marks = AcV2StudentMark.objects.filter(
        exam_assignment__in=exam_assignments
    ).select_related('exam_assignment')

    # Build mark lookup: student_id -> exam_id -> mark object
    mark_lookup = {}
    for sm in all_marks:
        sid = str(sm.student_id)
        eid = str(sm.exam_assignment_id)
        if sid not in mark_lookup:
            mark_lookup[sid] = {}
        mark_lookup[sid][eid] = sm

    students_data = []
    for sa in student_assignments:
        sp = sa.student
        sid = str(sp.id)
        student_entry = {
            'student_id': sid,
            'reg_no': sp.reg_no or '',
            'name': str(sp.user) if sp.user else sp.reg_no or '',
            'exam_marks': {},
            'weighted_marks': {},
            'co_totals': [0.0] * co_count,
            'final_mark': 0.0,
            'cqi_announce_target': False,
            'cqi_announce_target_cos': [],
        }

        student_marks = mark_lookup.get(sid, {})

        pending_cqi = []

        for ea in exam_assignments:
            eid = str(ea.id)
            einfo = exam_map[eid]
            sm = student_marks.get(eid)

            # CQI-kind exam assignment: derive marks from published CQI snapshot
            if einfo.get('kind') == 'cqi':
                pending_cqi.append(einfo)
                continue

            exam_entry = {
                'is_absent': sm.is_absent if sm else False,
            }
            exam_key = einfo.get('id')

            def _extract_cos(raw_co):
                if raw_co is None:
                    return []
                if isinstance(raw_co, (list, tuple, set)):
                    out = []
                    for item in raw_co:
                        try:
                            n = int(item)
                        except Exception:
                            continue
                        if 1 <= n <= co_count and n not in out:
                            out.append(n)
                    return out
                if isinstance(raw_co, str):
                    parts = re.split(r'[^0-9]+', raw_co)
                    out = []
                    for p in parts:
                        if not p:
                            continue
                        try:
                            n = int(p)
                        except Exception:
                            continue
                        if 1 <= n <= co_count and n not in out:
                            out.append(n)
                    return out
                try:
                    n = int(raw_co)
                except Exception:
                    return []
                return [n] if 1 <= n <= co_count else []

            # For Mark Manager exams, recompute CO marks from question_marks so "Exam" split
            # always applies only to the enabled COs (and stays correct even if older DB rows exist).
            direct_raw = None  # type: ignore
            effective_for_db = None  # type: ignore
            exam_raw_for_split = 0.0
            computed_total_from_questions = None  # type: ignore
            special_combo_value = {}
            if sm and not sm.is_absent and isinstance(sm.question_marks, dict) and isinstance(einfo.get('_qp_cos'), list):
                qp_cos_local = einfo.get('_qp_cos') or []
                qp_enabled_local = einfo.get('_qp_enabled') or [True] * len(qp_cos_local)
                qp_special_local = einfo.get('_qp_special_split') or []
                qp_special_sources_local = einfo.get('_qp_special_split_sources') or []
                qmarks = sm.question_marks

                # Question keys can be either 0-based (q0, q1, ...) or 1-based (q1, q2, ...)
                keys = set(str(k) for k in qmarks.keys())
                q_base = 0
                if 'q0' in keys:
                    q_base = 0
                elif 'q1' in keys and 'q0' not in keys:
                    q_base = 1

                def _qkey(i: int) -> str:
                    return f'q{i + q_base}'

                def _safe_bool_at(arr, i: int) -> bool:
                    return bool(isinstance(arr, list) and i < len(arr) and arr[i])

                def _safe_sources_at(arr, i: int):
                    if not isinstance(arr, list) or i >= len(arr):
                        return []
                    v = arr[i]
                    return v if isinstance(v, list) else []

                # Pre-compute per-student derived values for special_split combo questions.
                special_combo_value = {}
                if isinstance(qp_cos_local, list) and qp_cos_local:
                    for i, _co in enumerate(qp_cos_local):
                        if not _safe_bool_at(qp_special_local, i):
                            continue
                        if i < len(qp_enabled_local) and not qp_enabled_local[i]:
                            continue
                        sources = _safe_sources_at(qp_special_sources_local, i)
                        co_set = set()
                        sum_src = 0.0
                        for sidx in sources:
                            try:
                                j = int(sidx)
                            except Exception:
                                continue
                            if j == i or j < 0 or j >= len(qp_cos_local):
                                continue
                            if j < len(qp_enabled_local) and not qp_enabled_local[j]:
                                continue
                            v = qmarks.get(_qkey(j), 0) or 0
                            if isinstance(v, (int, float)):
                                sum_src += float(v)
                            for c in _extract_cos(qp_cos_local[j]):
                                co_set.add(int(c))
                        denom = len(co_set) or 1
                        base = qmarks.get(_qkey(i), 0) or 0
                        base_num = float(base) if isinstance(base, (int, float)) else 0.0
                        special_combo_value[i] = round(sum_src + (base_num / denom), 2)

                # Build combo question list for UI columns
                combo_cols = []
                for i, co in enumerate(qp_cos_local):
                    if i < len(qp_enabled_local) and not qp_enabled_local[i]:
                        continue
                    co_list = _extract_cos(co)
                    if len(co_list) >= 2:
                        max_marks_list = einfo.get('_qp_marks') or []
                        max_raw = max_marks_list[i] if isinstance(max_marks_list, list) and i < len(max_marks_list) else 0
                        max_q = float(max_raw or 0) if isinstance(max_raw, (int, float)) else 0.0
                        # If this is a special_split question, show derived max in header.
                        if _safe_bool_at(qp_special_local, i):
                            sources = _safe_sources_at(qp_special_sources_local, i)
                            co_set = set()
                            sum_src = 0.0
                            for sidx in sources:
                                try:
                                    j = int(sidx)
                                except Exception:
                                    continue
                                if j == i or j < 0 or j >= len(qp_cos_local):
                                    continue
                                if j < len(qp_enabled_local) and not qp_enabled_local[j]:
                                    continue
                                mr = max_marks_list[j] if isinstance(max_marks_list, list) and j < len(max_marks_list) else 0
                                if isinstance(mr, (int, float)):
                                    sum_src += float(mr)
                                for c in _extract_cos(qp_cos_local[j]):
                                    co_set.add(int(c))
                            denom = len(co_set) or 1
                            max_q = round(sum_src + (max_q / denom), 2)
                        combo_cols.append({
                            'key': f'combo_q{i}',
                            'co_list': co_list,
                            'max_marks': max_q,
                        })
                if combo_cols:
                    einfo['combo_questions'] = combo_cols

                # Base totals from CO-mapped questions
                co_totals_direct = {c: 0.0 for c in range(1, co_count + 1)}
                total_from_questions = 0.0
                for i, co in enumerate(qp_cos_local):
                    if i < len(qp_enabled_local) and not qp_enabled_local[i]:
                        continue
                    q_key = _qkey(i)
                    q_mark = qmarks.get(q_key, 0) or 0
                    if not isinstance(q_mark, (int, float)):
                        continue

                    # Total mark should reflect all enabled question marks (including Exam)
                    total_from_questions += float(q_mark)

                    co_list = _extract_cos(co)
                    if not co_list:
                        continue
                    share = float(q_mark) / len(co_list)
                    for c in co_list:
                        co_totals_direct[c] += share

                # Exam split (only when Mark Manager Exam is enabled)
                if einfo.get('cia_enabled'):
                    exam_idx = einfo.get('_exam_q_index')
                    raw_exam = 0.0
                    if isinstance(exam_idx, int) and exam_idx >= 0:
                        v = qmarks.get(_qkey(exam_idx), 0) or 0
                        if isinstance(v, (int, float)):
                            raw_exam = float(v)
                    exam_entry['exam'] = round(raw_exam, 2)
                    exam_raw_for_split = float(raw_exam or 0)

                    covered = einfo.get('covered_cos') or []
                    enabled_cos = [int(c) for c in covered if isinstance(c, int) and 1 <= int(c) <= co_count]
                else:
                    # Ensure the key exists for UI columns when configured
                    if einfo.get('cia_enabled'):
                        exam_entry['exam'] = 0

                # What we show in the table (direct-only)
                direct_raw = co_totals_direct

                # What we persist to DB (direct + Exam split), so downstream reports remain correct.
                co_totals_effective = dict(co_totals_direct)
                if einfo.get('cia_enabled'):
                    covered = einfo.get('covered_cos') or []
                    enabled_cos = [int(c) for c in covered if isinstance(c, int) and 1 <= int(c) <= co_count]
                    if exam_raw_for_split and enabled_cos:
                        share = float(exam_raw_for_split) / len(enabled_cos)
                        for c in enabled_cos:
                            co_totals_effective[c] = float(co_totals_effective.get(c, 0.0) + share)

                effective_for_db = co_totals_effective
                computed_total_from_questions = round(total_from_questions, 2)

                # Persist back to DB to keep co1..co5 in sync (best-effort)
                try:
                    new_vals = [round(co_totals_effective.get(i, 0.0), 2) for i in range(1, 6)]
                    old_vals = [
                        round(float(getattr(sm, f'co{i}_mark', 0) or 0), 2)
                        for i in range(1, 6)
                    ]
                    if new_vals != old_vals:
                        sm.co1_mark, sm.co2_mark, sm.co3_mark, sm.co4_mark, sm.co5_mark = new_vals
                        sm.save(update_fields=['co1_mark', 'co2_mark', 'co3_mark', 'co4_mark', 'co5_mark'])
                except Exception:
                    pass
            else:
                # If Mark Manager Exam is enabled, still expose exam key for UI
                if einfo.get('cia_enabled'):
                    exam_entry['exam'] = 0
                # Fallback for non-QP single-mark exams: distribute total_mark evenly
                # across covered COs so they appear in CO Summary instead of showing 0.
                if sm and not sm.is_absent and sm.total_mark is not None and not einfo.get('_qp_cos'):
                    covered = einfo.get('covered_cos') or []
                    if covered:
                        total_val = float(sm.total_mark or 0)
                        per_co = round(total_val / len(covered), 4)
                        direct_raw = {co_num: per_co for co_num in covered}
                        computed_total_from_questions = round(total_val, 2)
                        # Persist to DB so downstream reports stay in sync (best-effort)
                        try:
                            new_vals = [round(direct_raw.get(i, 0.0), 2) for i in range(1, 6)]
                            old_vals = [
                                round(float(getattr(sm, f'co{i}_mark', 0) or 0), 2)
                                for i in range(1, 6)
                            ]
                            if new_vals != old_vals:
                                sm.co1_mark, sm.co2_mark, sm.co3_mark, sm.co4_mark, sm.co5_mark = new_vals
                                sm.save(update_fields=['co1_mark', 'co2_mark', 'co3_mark', 'co4_mark', 'co5_mark'])
                        except Exception:
                            pass

            # Raw CO marks from AcV2StudentMark co1..co5 fields
            for co_num in range(1, co_count + 1):
                co_field = f'co{co_num}_mark'
                if direct_raw is not None:
                    raw_val = float(direct_raw.get(co_num, 0) or 0)
                else:
                    raw_val = float(getattr(sm, co_field, None) or 0) if sm else 0
                exam_entry[f'co{co_num}'] = raw_val

            # Combo question raw marks (for UI columns)
            combo_questions = einfo.get('combo_questions') or []
            if combo_questions and sm and isinstance(sm.question_marks, dict):
                for cq in combo_questions:
                    key = cq.get('key')
                    if not key:
                        continue
                    try:
                        idx = int(str(key).replace('combo_q', ''))
                    except Exception:
                        idx = None
                    if idx is None or idx < 0:
                        continue
                    if isinstance(special_combo_value, dict) and idx in special_combo_value:
                        exam_entry[key] = float(special_combo_value[idx] or 0)
                        continue
                    # Use the same base detection as above when possible
                    q_val = sm.question_marks.get(f'q{idx}', None)
                    if q_val is None:
                        q_val = sm.question_marks.get(f'q{idx + 1}', 0)
                    q_val = q_val or 0
                    exam_entry[key] = float(q_val) if isinstance(q_val, (int, float)) else 0.0

            # Total raw mark
            if computed_total_from_questions is not None:
                exam_entry['total'] = float(computed_total_from_questions)
            else:
                exam_entry['total'] = float(sm.total_mark) if sm and sm.total_mark is not None else 0

            if exam_key:
                student_entry['exam_marks'][exam_key] = exam_entry

            # Compute weighted marks for each covered CO
            if sm and not sm.is_absent:
                covered_cos = einfo['covered_cos']
                max_per_co = einfo['max_per_co']
                weight_per_co = einfo['weight_per_co']
                co_weights = einfo.get('co_weights', {})

                # Direct-only weighted marks (for left-side per-exam CO columns)

                for co_num in covered_cos:
                    if co_num < 1 or co_num > co_count:
                        continue
                    co_field = f'co{co_num}_mark'
                    if direct_raw is not None:
                        raw = float(direct_raw.get(co_num, 0) or 0)
                    else:
                        raw = float(getattr(sm, co_field, None) or 0)
                    # Use per-CO max from QP pattern if available, else fall back
                    co_max = einfo['co_max_map'].get(co_num, max_per_co)
                    # Use per-CO weight if defined, else fall back to even split
                    co_weight = co_weights.get(co_num, weight_per_co)
                    if co_max > 0:
                        weighted = round((raw / co_max) * co_weight, 2)
                    else:
                        weighted = 0
                    key = f"{einfo['id']}_CO{co_num}"
                    student_entry['weighted_marks'][key] = weighted
                    student_entry['co_totals'][co_num - 1] += weighted

                # Exam split weighted contribution (ONLY to right-side CO totals)
                if einfo.get('cia_enabled') and exam_raw_for_split:
                    enabled_cos = [
                        int(c) for c in (covered_cos or [])
                        if isinstance(c, int) and 1 <= int(c) <= co_count
                    ]
                    exam_max_marks_local = float(einfo.get('exam_max_marks') or 0)
                    exam_weight_local = float(einfo.get('cia_weight') or 0)
                    per_co_local = bool(einfo.get('cia_weight_per_co', False))
                    if enabled_cos and exam_max_marks_local > 0 and exam_weight_local > 0:
                        share_raw = float(exam_raw_for_split) if per_co_local else (float(exam_raw_for_split) / len(enabled_cos))
                        share_max = float(exam_max_marks_local) if per_co_local else (float(exam_max_marks_local) / len(enabled_cos))
                        share_wt = float(exam_weight_local) if per_co_local else (float(exam_weight_local) / len(enabled_cos))
                        if share_max > 0:
                            for c in enabled_cos:
                                add_w = round((share_raw / share_max) * share_wt, 2)
                                student_entry['co_totals'][c - 1] += add_w
                                # Store per-CO exam split weighted mark for display in CO Summary
                                wm_key = f"{einfo['id']}_exam_CO{c}"
                                student_entry['weighted_marks'][wm_key] = round(add_w, 2)

        # Apply CQI exam(s) after non-CQI totals are known.
        if pending_cqi:
            sid_key = str(sid)
            s_entries = cqi_entries.get(sid_key, {}) if isinstance(cqi_entries, dict) else {}

            before_total_all = float(sum(student_entry.get('co_totals') or []))
            # Precompute sum(CQI inputs) for TOTAL_CQI context.
            sum_raw_cqi = 0.0
            for einfo in pending_cqi:
                for co_num in (einfo.get('covered_cos') or []):
                    try:
                        co_n = int(co_num)
                    except Exception:
                        continue
                    if co_n < 1 or co_n > co_count:
                        continue
                    raw_in = s_entries.get(f'co{co_n}') if isinstance(s_entries, dict) else None
                    try:
                        sum_raw_cqi += float(raw_in) if raw_in is not None else 0.0
                    except Exception:
                        pass

            shared_custom_vars = _normalize_cqi_custom_vars(getattr(class_type, 'cqi_global_custom_vars', []))

            for einfo in pending_cqi:
                cqi_sub = einfo.get('_cqi_sub') if isinstance(einfo, dict) else {}
                conds = cqi_sub.get('conditions', []) if isinstance(cqi_sub, dict) else []
                else_expr = str(cqi_sub.get('else_formula', '') or '') if isinstance(cqi_sub, dict) else ''
                legacy_value_expr = str(cqi_sub.get('co_value_expr', '') or '') if isinstance(cqi_sub, dict) else ''
                selected_exam_codes = set(
                    _normalize_cqi_token_code(x)
                    for x in (cqi_sub.get('exams', []) if isinstance(cqi_sub, dict) and isinstance(cqi_sub.get('exams'), list) else [])
                    if _normalize_cqi_token_code(x)
                )

                # IMPORTANT: CQI Entry page computes condition tokens (CO-RAW, TOTAL_CQI)
                # in InternalMarkPage "weighted space" using only the *considered exams*
                # (cqi_sub.exams filter; if empty => all non-CQI exams).
                # To keep colours/conditions identical between CQI Entry and Internal Mark,
                # build the same per-CO "before" totals from weighted_marks per exam.
                def _before_cqi_total_for_co_ctx(co_num: int) -> float:
                    try:
                        co_n_local = int(co_num)
                    except Exception:
                        return 0.0
                    if co_n_local < 1 or co_n_local > co_count:
                        return 0.0
                    wm_map = student_entry.get('weighted_marks') or {}
                    if not isinstance(wm_map, dict):
                        return 0.0

                    total_local = 0.0
                    for _src_exam in exams_data:
                        if _src_exam.get('kind') == 'cqi':
                            continue
                        exam_code = _normalize_cqi_token_code(
                            _src_exam.get('short_name')
                            or _src_exam.get('name')
                            or _src_exam.get('exam_display_name')
                            or _src_exam.get('exam')
                            or ''
                        )
                        if selected_exam_codes and exam_code not in selected_exam_codes:
                            continue
                        covered = _src_exam.get('covered_cos') or []
                        if not covered:
                            covered = list(range(1, co_count + 1))
                        if co_n_local not in covered:
                            continue
                        try:
                            key = f"{_src_exam.get('id')}_CO{co_n_local}"
                            total_local += float(wm_map.get(key, 0) or 0.0)
                        except Exception:
                            pass
                    return float(total_local or 0.0)
                
                local_custom_vars = _normalize_cqi_custom_vars(cqi_sub.get('custom_vars', []) if isinstance(cqi_sub, dict) else [])
                combined_custom_vars = []
                seen_custom_codes = set()
                for cv in [*shared_custom_vars, *local_custom_vars]:
                    code = cv.get('code') or ''
                    if not code or code in seen_custom_codes:
                        continue
                    seen_custom_codes.add(code)
                    combined_custom_vars.append(cv)

                # CQI threshold tokens (TOTAL_CQI / CQI-TOTAL-MAX):
                # Used for condition evaluation (e.g. TOTAL_CQI < 58).
                # NOTE: We intentionally do NOT enforce an overall 58% budget cap across COs;
                # capping is applied per CO only (when the matched condition has cap_enabled).
                covered_cos_for_cap = []
                for _c in (einfo.get('covered_cos') or []):
                    try:
                        _n = int(_c)
                    except Exception:
                        continue
                    if 1 <= _n <= co_count:
                        covered_cos_for_cap.append(_n)

                total_max_for_cap = 0.0
                before_total_for_cap = 0.0
                if covered_cos_for_cap:
                    for _co_n in covered_cos_for_cap:
                        try:
                            before_total_for_cap += float(_before_cqi_total_for_co_ctx(_co_n) or 0.0)
                        except Exception:
                            pass

                        # Same weighted-space CO-MAX computation used in vars_map['CO-MAX'] (per CO)
                        co_max_for_co_cap = 0.0
                        for _src_exam in exams_data:
                            if _src_exam.get('kind') == 'cqi':
                                continue
                            exam_code = _normalize_cqi_token_code(
                                _src_exam.get('short_name')
                                or _src_exam.get('name')
                                or _src_exam.get('exam_display_name')
                                or _src_exam.get('exam')
                                or ''
                            )
                            if selected_exam_codes and exam_code not in selected_exam_codes:
                                continue
                            covered = _src_exam.get('covered_cos') or []
                            if not covered:
                                covered = list(range(1, co_count + 1))
                            if _co_n not in covered:
                                continue
                            n_cov = len(covered) or 1
                            base_map = _src_exam.get('co_weights') or {}
                            try:
                                base_w = float(base_map.get(_co_n, _src_exam.get('weight_per_co') or 0) or 0)
                            except Exception:
                                base_w = 0.0
                            cia_share = 0.0
                            try:
                                if _src_exam.get('cia_enabled') and float(_src_exam.get('cia_weight') or 0) > 0:
                                    cia_share = float(_src_exam.get('cia_weight') or 0) / n_cov
                            except Exception:
                                cia_share = 0.0
                            co_max_for_co_cap += float(base_w or 0.0) + float(cia_share or 0.0)
                        total_max_for_cap += co_max_for_co_cap

                before_pct_for_cap = 0.0
                try:
                    if total_max_for_cap > 0:
                        before_pct_for_cap = (before_total_for_cap / total_max_for_cap) * 100.0
                except Exception:
                    before_pct_for_cap = 0.0

                exam_entry = { 'is_absent': False }
                total_val = 0.0
                for co_num in (einfo.get('covered_cos') or []):
                    try:
                        co_n = int(co_num)
                    except Exception:
                        continue
                    if co_n < 1 or co_n > co_count:
                        continue
                    raw_in = s_entries.get(f'co{co_n}') if isinstance(s_entries, dict) else None
                    try:
                        cqi_in = float(raw_in) if raw_in is not None else 0.0
                    except Exception:
                        cqi_in = 0.0

                    # Condition context MUST match CQI Entry page:
                    # before-co is computed from considered exams only.
                    before_co = 0.0
                    try:
                        before_co = float(_before_cqi_total_for_co_ctx(co_n) or 0.0)
                    except Exception:
                        before_co = 0.0

                    # Row-level CQI tokens must match CQI Entry page semantics:
                    # - BEFORE_CQI/AFTER_CQI are the row totals across CQI-covered COs
                    # - TOTAL_CQI is the row total percentage BEFORE CQI additions
                    # - CQI-TOTAL-MAX is the max total across CQI-covered COs
                    vars_map = {
                        'CQI': float(cqi_in or 0.0),
                        'X': float(cqi_in or 0.0),
                        'BEFORE_CQI': float(before_total_for_cap or 0.0),
                        'AFTER_CQI': float(before_total_for_cap or 0.0),
                        'TOTAL_CQI': float(before_pct_for_cap or 0.0),
                        'CQI-TOTAL-MAX': float(total_max_for_cap or 0.0),
                    }

                    # New token: BEFORE_CQI_COX is the current CO's pre-CQI raw value.
                    # (Also store a resolved CO{n} alias.)
                    vars_map['BEFORE_CQI_COX'] = float(before_co or 0.0)
                    vars_map[f'BEFORE_CQI_CO{co_n}'] = float(before_co or 0.0)

                    # --- CO-level aliases (current CO context) ---
                    # CO-MAX must be in the same *weighted space* as BEFORE_CQI/CO-RAW.
                    # This matches the frontend CQI entry computation:
                    #   weighted per CO = (raw / co_max_raw) * co_weight
                    # So the maximum weighted contribution per exam/CO is just co_weight
                    # (+ CIA share when Mark Manager Exam split is enabled).
                    co_max_for_co = 0.0
                    for _src_exam in exams_data:
                        if _src_exam.get('kind') == 'cqi':
                            continue
                        exam_code = _normalize_cqi_token_code(
                            _src_exam.get('short_name')
                            or _src_exam.get('name')
                            or _src_exam.get('exam_display_name')
                            or _src_exam.get('exam')
                            or ''
                        )
                        if selected_exam_codes and exam_code not in selected_exam_codes:
                            continue
                        covered = _src_exam.get('covered_cos') or []
                        if not covered:
                            covered = list(range(1, co_count + 1))
                        if co_n not in covered:
                            continue
                        n_cov = len(covered) or 1
                        base_map = _src_exam.get('co_weights') or {}
                        try:
                            base_w = float(base_map.get(co_n, _src_exam.get('weight_per_co') or 0) or 0)
                        except Exception:
                            base_w = 0.0
                        cia_share = 0.0
                        try:
                            if _src_exam.get('cia_enabled') and float(_src_exam.get('cia_weight') or 0) > 0:
                                cia_share = float(_src_exam.get('cia_weight') or 0) / n_cov
                        except Exception:
                            cia_share = 0.0
                        co_max_for_co += float(base_w or 0.0) + float(cia_share or 0.0)

                    vars_map['CO-RAW'] = float(before_co or 0.0)
                    vars_map['CO-TOTAL-RAW'] = vars_map['CO-RAW']
                    vars_map['CO-MAX'] = round(co_max_for_co, 4)
                    # COX-EXAMS-MAX-WEIGHT = same as CO-MAX: total max weight for the current
                    # CO across the checked exam assignments in the CQI config.
                    # Matches the frontend CqiEntryPage value (ctx['COX-EXAMS-MAX-WEIGHT']).
                    vars_map['COX-EXAMS-MAX-WEIGHT'] = round(co_max_for_co, 4)
                    if co_max_for_co > 0:
                        _pct = round((float(before_co or 0.0) / co_max_for_co) * 100, 4)
                    else:
                        _pct = 0.0
                    vars_map['CO-WEIGHT'] = _pct
                    vars_map['CO-TOTAL-WEIGHT'] = _pct

                    # --- Dynamic COx tokens (COX = current CO number) ---
                    vars_map['COX_PERCENT'] = _pct
                    vars_map[f'CO{co_n}_PERCENT'] = _pct
                    vars_map['BEFORE_CQI_COX'] = float(before_co or 0.0)
                    vars_map[f'BEFORE_CQI_CO{co_n}'] = float(before_co or 0.0)
                    vars_map['BEFORE_CQI_COX_TOTAL'] = float(before_co or 0.0)
                    vars_map[f'BEFORE_CQI_CO{co_n}_TOTAL'] = float(before_co or 0.0)
                    vars_map['AFTER_CQI_COX_TOTAL'] = float((before_co or 0.0) + (cqi_in or 0.0))
                    vars_map[f'AFTER_CQI_CO{co_n}_TOTAL'] = vars_map['AFTER_CQI_COX_TOTAL']

                    exam_marks_map = student_entry.get('exam_marks') or {}
                    weighted_marks_map = student_entry.get('weighted_marks') or {}
                    for source_exam in exams_data:
                        if source_exam.get('kind') == 'cqi':
                            continue
                        exam_code = _normalize_cqi_token_code(
                            source_exam.get('short_name')
                            or source_exam.get('name')
                            or source_exam.get('exam_display_name')
                            or source_exam.get('exam')
                            or ''
                        )
                        if not exam_code:
                            continue
                        if selected_exam_codes and exam_code not in selected_exam_codes:
                            continue
                        exam_key = source_exam.get('id')
                        exam_marks = exam_marks_map.get(exam_key, {}) if isinstance(exam_marks_map, dict) else {}
                        if not isinstance(exam_marks, dict):
                            exam_marks = {}
                        co_raw = float(exam_marks.get(f'co{co_n}', 0) or 0)
                        co_weight = float(weighted_marks_map.get(f"{exam_key}_CO{co_n}", 0) or 0) if isinstance(weighted_marks_map, dict) else 0.0
                        # Per-exam CO max must also be in weighted space (== weight contribution ceiling).
                        covered = source_exam.get('covered_cos') or []
                        if not covered:
                            covered = list(range(1, co_count + 1))
                        n_cov = len(covered) or 1
                        max_for_co = 0.0
                        if co_n in covered:
                            base_map = source_exam.get('co_weights') or {}
                            try:
                                max_for_co += float(base_map.get(co_n, source_exam.get('weight_per_co') or 0) or 0)
                            except Exception:
                                pass
                            try:
                                if source_exam.get('cia_enabled') and float(source_exam.get('cia_weight') or 0) > 0:
                                    max_for_co += float(source_exam.get('cia_weight') or 0) / n_cov
                            except Exception:
                                pass

                        vars_map[f'{exam_code}-TOTAL'] = round(max_for_co, 4)
                        vars_map[f'{exam_code}-OBT'] = co_raw
                        vars_map[f'{exam_code}-WEIGHT'] = co_weight
                        vars_map[f'COX-{exam_code}-OBT'] = co_raw
                        vars_map[f'COX-{exam_code}-WEIGHT'] = co_weight
                        # Plain exam code alias: [SSA1] = raw obtained marks for current CO
                        vars_map[exam_code] = co_raw

                        if source_exam.get('mark_manager_enabled') or source_exam.get('cia_enabled'):
                            item_count = _count_mark_manager_items_for_co(source_exam, co_n)
                            vars_map[f'COX-{exam_code}-AVG'] = round((co_raw / item_count), 4) if item_count > 0 else 0.0
                            vars_map[f'{exam_code}-EXAM-OBT'] = float(exam_marks.get('exam', 0) or 0) if source_exam.get('cia_enabled') else 0.0
                            vars_map[f'{exam_code}-EXAM-WEIGHT'] = float(source_exam.get('cia_weight') or 0) if source_exam.get('cia_enabled') else 0.0

                    for custom_var in combined_custom_vars:
                        code = custom_var.get('code') or ''
                        expr = str(custom_var.get('expr') or '').strip()
                        if not code or not expr:
                            continue
                        try:
                            vars_map[code] = _safe_eval_cqi_num(expr, vars_map)
                        except Exception:
                            vars_map[code] = 0.0

                    # Evaluate admin-defined derived variables (COX in name is a placeholder for current CO)
                    # Each derived variable is evaluated and stored as BOTH the COX form and the CO{n} form.
                    # e.g. BEFORE_CQI_COX → vars_map['BEFORE_CQI_COX'] = val
                    #      BEFORE_CQI_CO2 → vars_map['BEFORE_CQI_CO2'] = val  (when co_n=2)
                    cqi_derived_vars = []
                    if isinstance(cqi_sub, dict):
                        cqi_derived_vars = cqi_sub.get('derived_variables') or []
                    if not isinstance(cqi_derived_vars, list):
                        cqi_derived_vars = []
                    for dv in cqi_derived_vars:
                        if not isinstance(dv, dict):
                            continue
                        dv_name = str(dv.get('name') or '').strip().upper()
                        dv_formula = str(dv.get('formula') or '').strip()
                        if not dv_name or not dv_formula:
                            continue
                        try:
                            dv_val = round(_safe_eval_cqi_num(dv_formula, vars_map), 4)
                        except Exception:
                            dv_val = 0.0
                        # Placeholder form (COX literal kept as-is)
                        vars_map[dv_name] = dv_val
                        # Resolved form: replace COX → CO{n}
                        resolved_name = dv_name.replace('COX', f'CO{co_n}')
                        if resolved_name != dv_name:
                            vars_map[resolved_name] = dv_val

                    mapped = None
                    matched_condition_title = ''
                    matched_cond_obj = None  # store for per-condition cap
                    # Condition ladder: first match wins.
                    if isinstance(conds, list) and conds:
                        for cond_idx, cond in enumerate(conds, start=1):
                            if not isinstance(cond, dict):
                                continue
                            if_expr = _resolve_cqi_if_expr(cond)
                            then_expr = str(cond.get('then', '') or '').strip()
                            if if_expr and _safe_eval_cqi_bool(if_expr, vars_map):
                                matched_cond_obj = cond
                                try:
                                    matched_condition_title = str(cond.get('title') or cond.get('name') or '').strip()
                                except Exception:
                                    matched_condition_title = ''
                                if not matched_condition_title:
                                    matched_condition_title = f'CQI Condition {cond_idx}'
                                if then_expr:
                                    mapped = _safe_eval_cqi_num(then_expr, vars_map)
                                break

                    if matched_condition_title:
                        try:
                            titles = student_entry.get('cqi_satisfied_conditions')
                            if not isinstance(titles, list):
                                titles = []
                            if matched_condition_title not in titles:
                                titles.append(matched_condition_title)
                            student_entry['cqi_satisfied_conditions'] = titles

                            # Per-exam, per-CO matched condition index (1-based).
                            # Used by InternalMarkPage to decide cap eligibility per cell.
                            by_exam = student_entry.get('cqi_matched_condition_index_by_exam')
                            if not isinstance(by_exam, dict):
                                by_exam = {}
                            exam_id = str(einfo.get('id') or '')
                            per_exam = by_exam.get(exam_id)
                            if not isinstance(per_exam, dict):
                                per_exam = {}
                            per_exam[str(co_n)] = int(cond_idx) if str(cond_idx).isdigit() else cond_idx
                            by_exam[exam_id] = per_exam
                            student_entry['cqi_matched_condition_index_by_exam'] = by_exam

                            student_entry['cqi_announce_target'] = True
                            target_cos = student_entry.get('cqi_announce_target_cos')
                            if not isinstance(target_cos, list):
                                target_cos = []
                            if co_n not in target_cos:
                                target_cos.append(co_n)
                            student_entry['cqi_announce_target_cos'] = target_cos
                        except Exception:
                            pass

                    if mapped is None:
                        if else_expr.strip():
                            mapped = _safe_eval_cqi_num(else_expr, vars_map)
                        elif legacy_value_expr.strip():
                            mapped = _safe_eval_cqi_num(legacy_value_expr, vars_map)
                        else:
                            mapped = float(cqi_in or 0.0)

                    mapped = round(float(mapped or 0.0), 2)

                    cap_enabled_raw = matched_cond_obj.get('cap_enabled') if isinstance(matched_cond_obj, dict) else False
                    cap_enabled = (
                        cap_enabled_raw is True
                        or cap_enabled_raw == 1
                        or (isinstance(cap_enabled_raw, str) and cap_enabled_raw.strip().lower() in ('true', '1', 'yes', 'y', 'on'))
                    )

                    # Cap rule (58% by default): apply ONLY when the matched condition
                    # has cap_enabled=true. If a condition provides cap_percent, prefer it.
                    _cqi_cap_hit = False
                    try:
                        if cap_enabled:
                            cond_cap_pct = None
                            if isinstance(matched_cond_obj, dict) and matched_cond_obj.get('cap_percent') is not None:
                                cond_cap_pct = float(matched_cond_obj.get('cap_percent') or 0.0)
                            effective_cap_pct = cond_cap_pct if cond_cap_pct and cond_cap_pct > 0 else 58.0
                            co_max_for_cap = float(vars_map.get('CO-MAX') or 0.0)
                            if effective_cap_pct > 0 and co_max_for_cap > 0:
                                cap_ceiling = round((effective_cap_pct / 100.0) * co_max_for_cap, 4)
                                max_add = max(0.0, cap_ceiling - float(before_co or 0.0))
                                uncapped = mapped
                                mapped = round(min(mapped, max_add), 2)
                                if uncapped > max_add + 0.001:
                                    _cqi_cap_hit = True
                    except Exception:
                        pass

                    # Track which COs had the cap actually applied for this student.
                    if _cqi_cap_hit:
                        try:
                            capped_cos = student_entry.get('cqi_capped_cos')
                            if not isinstance(capped_cos, list):
                                capped_cos = []
                            if co_n not in capped_cos:
                                capped_cos.append(co_n)
                            student_entry['cqi_capped_cos'] = capped_cos
                        except Exception:
                            pass

                    exam_entry[f'co{co_n}'] = mapped
                    total_val += mapped

                    # CQI contributes directly to CO totals and MUST also populate weighted_marks,
                    # because InternalMarkPage/CO summary reads weighted_marks keys like:
                    #   "<examAssignmentId>_CO<coNum>"
                    try:
                        student_entry['co_totals'][co_n - 1] = round(float(student_entry['co_totals'][co_n - 1] or 0.0) + mapped, 2)
                    except Exception:
                        pass

                    try:
                        wm_key = f"{einfo['id']}_CO{co_n}"
                        student_entry['weighted_marks'][wm_key] = round(float(mapped or 0.0), 2)
                    except Exception:
                        pass

                exam_entry['total'] = round(total_val, 2)
                student_entry['exam_marks'][einfo.get('id')] = exam_entry

        # Round CO totals
        student_entry['co_totals'] = [round(v, 2) for v in student_entry['co_totals']]
        student_entry['final_mark'] = round(sum(student_entry['co_totals']), 2)

        students_data.append(student_entry)

    # Extract CQI config based on the first CQI exam present for this TA.
    # This keeps CQI entry + InternalMarkPage aligned even if multiple CQI configs exist.
    cqi_config = None
    if isinstance(first_cqi_sub, dict) and first_cqi_sub is not None:
        _raw = first_cqi_sub
        global_vars = getattr(class_type, 'cqi_global_custom_vars', [])
        if not isinstance(global_vars, list):
            global_vars = _raw.get('global_custom_vars', []) if isinstance(_raw.get('global_custom_vars'), list) else []
        cqi_config = {
            'name': str(_raw.get('name', '') or ''),
            'code': str(_raw.get('code', '') or ''),
            'cycle_id': str(_raw.get('cycle_id', '') or ''),
            'cos': _raw.get('cos', []) if isinstance(_raw.get('cos'), list) else [],
            'exams': _raw.get('exams', []) if isinstance(_raw.get('exams'), list) else [],
            'custom_vars': _raw.get('custom_vars', []) if isinstance(_raw.get('custom_vars'), list) else [],
            'global_custom_vars': global_vars,
            'derived_variables': _raw.get('derived_variables', []) if isinstance(_raw.get('derived_variables'), list) else [],
            'co_value_expr': str(_raw.get('co_value_expr', '') or ''),
            'formula': str(_raw.get('formula', '') or ''),
            'conditions': _raw.get('conditions', []) if isinstance(_raw.get('conditions'), list) else [],
            'else_formula': str(_raw.get('else_formula', '') or ''),
        }

    return Response({
        'course_code': course_code,
        'course_name': course_name,
        'co_count': co_count,
        'total_internal_marks': total_internal,
        'class_type': {
            'id': class_type.id if class_type else None,
            'name': class_type.name if class_type else None,
            'code': getattr(class_type, 'code', None) or (class_type.name if class_type else None),
            'coattainment_layout': (
                class_type.cqi_global_custom_vars.get('_coattainment_layout', {})
                if class_type and isinstance(class_type.cqi_global_custom_vars, dict)
                else {}
            ),
        } if class_type else None,
        'qp_type': qp_type_code,
        'exams': exams_data,
        'students': students_data,
        'cqi_config': cqi_config,
    })


# ============================================================================
# EXPORT EXCEL TEMPLATE (for import)
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_exam_export_template(request, exam_id):
    """Export an Excel template with student roster and question columns."""
    from academics.models import StudentSectionAssignment
    import openpyxl
    import re
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from django.http import HttpResponse

    ea_qs = AcV2ExamAssignment.objects.select_related(
        'section__teaching_assignment__section',
        'section__teaching_assignment__curriculum_row',
        'section__teaching_assignment__elective_subject',
    )
    if _has_admin_bypass_access(request.user):
        ea = get_object_or_404(ea_qs, id=exam_id)
    else:
        ea = get_object_or_404(ea_qs, id=exam_id, section__faculty_user=request.user)

    ta = ea.section.teaching_assignment
    acad_sec = ta.section
    cr = ta.curriculum_row
    es = ta.elective_subject
    course_code = (cr.course_code if cr else None) or (getattr(es, 'course_code', None)) or '-'
    course_name = (cr.course_name if cr else None) or (getattr(es, 'course_name', None)) or '-'

    def _norm_header(v) -> str:
        return str(v or '').strip()

    def _question_header(i, q):
        title_raw = _norm_header(q.get('question_number') or q.get('title') or '')
        if not title_raw:
            return f'Q{i + 1}'
        if re.fullmatch(r'\d+', title_raw):
            return f'Q{title_raw}'
        return title_raw

    # Build question columns from the resolved QP pattern for this exam assignment.
    # This matches what the Mark Entry UI shows (supports local overrides).
    qp_type = ea.qp_type or ea.exam or ''
    p = ea.get_qp_pattern() or {}
    question_cols = []
    if isinstance(p, dict) and isinstance(p.get('questions'), list):
        for i, q in enumerate(p.get('questions') or []):
            if not isinstance(q, dict):
                continue
            if q.get('enabled') is False:
                continue
            question_cols.append({
                'key': q.get('id') or f'q{i}',
                'title': _question_header(i, q),
                'max_marks': q.get('max_marks') or q.get('marks') or 0,
                'co': q.get('co_number') or q.get('co') or 0,
            })
    elif isinstance(p, dict):
        titles = p.get('titles', [])
        marks_list = p.get('marks', [])
        cos = p.get('cos', [])
        enabled = p.get('enabled', [])
        for i in range(len(titles)):
            if i < len(enabled) and not enabled[i]:
                continue
            question_cols.append({
                'key': f'q{i}',
                'title': _question_header(i, {'question_number': titles[i] if i < len(titles) else '', 'title': titles[i] if i < len(titles) else ''}),
                'max_marks': marks_list[i] if i < len(marks_list) else 0,
                'co': cos[i] if i < len(cos) else 0,
            })

    # Students
    assignments = (
        StudentSectionAssignment.objects
        .filter(section=acad_sec, end_date__isnull=True)
        .select_related('student__user')
        .order_by('student__reg_no')
    )
    existing = {
        str(sm.student_id): sm
        for sm in AcV2StudentMark.objects.filter(exam_assignment=ea)
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mark Entry'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    sub_font = Font(bold=False, color='6B7280', size=8, italic=True)
    sub_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    locked_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')

    # Row 2: Headers
    base_headers = ['Sl No', 'Register Number', 'Student Name']
    q_headers = [q['title'] for q in question_cols]
    all_headers = base_headers + q_headers + ['Total', 'Absent']

    # Row 1: Info header (merged across full header width)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(4, len(all_headers)))
    info_cell = ws.cell(row=1, column=1, value=f'{course_code} — {course_name} | {ea.exam_display_name or ea.exam or qp_type} | Max: {ea.max_marks}')
    info_cell.font = Font(bold=True, size=11)

    # Row 3: Sub-header (max marks / CO info)
    sub_headers = ['', '', '']
    for q in question_cols:
        co_label = f'CO{q["co"]}' if q['co'] else ''
        sub_headers.append(f'Max:{q["max_marks"]} {co_label}')
    sub_headers += [f'Max:{ea.max_marks}', 'Yes/No']

    editable_header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        if header in q_headers:
            cell.fill = editable_header_fill
        else:
            cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for col_idx, sub in enumerate(sub_headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=sub)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    row_num = 4
    for idx, sa in enumerate(assignments):
        sp = sa.student
        sm = existing.get(str(sp.id))
        reg_no = sp.reg_no or ''
        name = str(sp.user) if sp.user else reg_no

        ws.cell(row=row_num, column=1, value=idx + 1).border = thin_border
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')

        reg_cell = ws.cell(row=row_num, column=2, value=reg_no)
        reg_cell.border = thin_border
        reg_cell.font = Font(bold=True, size=10)
        reg_cell.fill = locked_fill

        name_cell = ws.cell(row=row_num, column=3, value=name)
        name_cell.border = thin_border
        name_cell.fill = locked_fill

        # Question marks
        co_marks = sm.question_marks if sm and isinstance(sm.question_marks, dict) else {}
        for q_idx, q in enumerate(question_cols):
            val = co_marks.get(q['key'])
            cell = ws.cell(row=row_num, column=4 + q_idx, value=val if val is not None else '')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Total
        total_val = float(sm.total_mark) if sm and sm.total_mark is not None else ''
        total_cell = ws.cell(row=row_num, column=4 + len(question_cols), value=total_val)
        total_cell.border = thin_border
        total_cell.alignment = Alignment(horizontal='center')
        total_cell.font = Font(bold=True)

        # Absent
        absent_val = 'Yes' if sm and sm.is_absent else ''
        absent_cell = ws.cell(row=row_num, column=5 + len(question_cols), value=absent_val)
        absent_cell.border = thin_border
        absent_cell.alignment = Alignment(horizontal='center')

        row_num += 1

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    for i in range(len(question_cols)):
        col_letter = openpyxl.utils.get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(4 + len(question_cols))].width = 10
    ws.column_dimensions[openpyxl.utils.get_column_letter(5 + len(question_cols))].width = 10

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_name = f'{course_code}_{ea.exam_display_name or ea.exam or qp_type}'.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{safe_name}.xlsx"'
    return response


# ============================================================================
# CQI (Academic 2.1) - Faculty Draft + Publish
# ============================================================================


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def faculty_exam_cqi_draft(request, exam_id):
    """Backward-compatible CQI draft endpoint keyed by exam_id.

    Maps exam assignment to its teaching assignment and reuses the
    course-level CQI draft handler.
    """
    ea_qs = AcV2ExamAssignment.objects.select_related('section__teaching_assignment')
    if _has_admin_bypass_access(request.user):
        ea = get_object_or_404(ea_qs, id=exam_id)
    else:
        ea = get_object_or_404(ea_qs, id=exam_id, section__faculty_user=request.user)
    return faculty_course_cqi_draft(request, ea.section.teaching_assignment_id)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_exam_cqi_published(request, exam_id):
    """Backward-compatible CQI published endpoint keyed by exam_id.

    Maps exam assignment to its teaching assignment and reuses the
    course-level CQI published handler.
    """
    ea_qs = AcV2ExamAssignment.objects.select_related('section__teaching_assignment')
    if _has_admin_bypass_access(request.user):
        ea = get_object_or_404(ea_qs, id=exam_id)
    else:
        ea = get_object_or_404(ea_qs, id=exam_id, section__faculty_user=request.user)
    return faculty_course_cqi_published(request, ea.section.teaching_assignment_id)


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def faculty_course_cqi_draft(request, ta_id: int):
    """Get/Upsert CQI draft entries for a teaching assignment.

    Payload shape:
      { entries: { "<student_id>": {"co1": number|null, ... } } }
    """
    from academics.models import TeachingAssignment
    from .models import AcV2CqiAssignment

    ta_qs = TeachingAssignment.objects.select_related(
        'curriculum_row', 'elective_subject', 'section', 'staff',
    )
    if _has_admin_bypass_access(request.user):
        ta = get_object_or_404(ta_qs, id=ta_id, is_active=True)
    else:
        ta = get_object_or_404(ta_qs, id=ta_id, staff__user=request.user, is_active=True)

    if request.method == 'GET':
        obj = AcV2CqiAssignment.objects.filter(teaching_assignment=ta).first()
        if obj is None:
            return Response({'draft': None})
        return Response({
            'draft': {
                'co_numbers': obj.co_numbers or [],
                'threshold_percent': float(obj.threshold_percent or 58.0),
                'entries': obj.draft_entries or {},
            },
            'updated_at': obj.draft_updated_at.isoformat() if getattr(obj, 'draft_updated_at', None) else None,
            'updated_by': obj.draft_updated_by,
        })

    # If CQI is already published and publish control locks it, block draft edits.
    try:
        from .models import AcV2CqiAttained
        attained = AcV2CqiAttained.objects.filter(teaching_assignment=ta).first()
        if attained is not None:
            pc = check_cqi_publish_control(attained)
            if pc.get('publish_control_enabled') and pc.get('is_locked') and not pc.get('is_editable'):
                return Response({'detail': 'CQI is locked. Request Edit to modify.'}, status=403)
    except Exception:
        # Fail open for backward compatibility.
        pass

    body = request.data if isinstance(request.data, dict) else {}
    entries = body.get('entries')
    if entries is None or not isinstance(entries, dict):
        return Response({'detail': 'entries must be an object.'}, status=status.HTTP_400_BAD_REQUEST)

    co_numbers = body.get('co_numbers')
    if not isinstance(co_numbers, list):
        co_numbers = None

    threshold = body.get('threshold_percent')
    try:
        threshold_f = float(threshold) if threshold is not None else None
    except Exception:
        threshold_f = None

    user_id = getattr(getattr(request, 'user', None), 'id', None)
    obj, _created = AcV2CqiAssignment.objects.get_or_create(
        teaching_assignment=ta,
        defaults={
            'co_numbers': co_numbers or [],
            'threshold_percent': threshold_f if threshold_f is not None else 58.0,
            'draft_entries': {},
            'draft_updated_by': user_id,
        },
    )

    if co_numbers is not None:
        obj.co_numbers = co_numbers
    if threshold_f is not None:
        obj.threshold_percent = threshold_f

    obj.draft_entries = entries
    obj.draft_updated_by = user_id
    obj.save(update_fields=['co_numbers', 'threshold_percent', 'draft_entries', 'draft_updated_by', 'draft_updated_at'])

    return Response({
        'status': 'ok',
        'updated_at': obj.draft_updated_at.isoformat() if getattr(obj, 'draft_updated_at', None) else None,
        'updated_by': obj.draft_updated_by,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_course_cqi_published(request, ta_id: int):
    """Fetch published CQI snapshot for a teaching assignment."""
    from academics.models import TeachingAssignment
    from .models import AcV2CqiAttained

    _ta_qs_cqi_pub = TeachingAssignment.objects.select_related('staff')
    if _has_admin_bypass_access(request.user):
        ta = get_object_or_404(_ta_qs_cqi_pub, id=ta_id, is_active=True)
    else:
        ta = get_object_or_404(_ta_qs_cqi_pub, id=ta_id, staff__user=request.user, is_active=True)

    obj = AcV2CqiAttained.objects.filter(teaching_assignment=ta).first()
    if obj is None:
        return Response({'published': None})

    pc = check_cqi_publish_control(obj)

    return Response({
        'published': {
            'co_numbers': obj.co_numbers or [],
            'entries': obj.entries or {},
            'published_at': obj.published_at.isoformat() if getattr(obj, 'published_at', None) else None,
            'published_by': obj.published_by,
        },
        'publish_control': pc,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def faculty_course_cqi_publish(request, ta_id: int):
    """Publish CQI snapshot.

    If body.entries is omitted, publishes latest draft.
    """
    from academics.models import TeachingAssignment
    from .models import AcV2CqiAssignment, AcV2CqiAttained
    from .services.mark_calculation import compute_section_internal_marks

    _ta_qs_cqi_publish = TeachingAssignment.objects.select_related('staff')
    if _has_admin_bypass_access(request.user):
        ta = get_object_or_404(_ta_qs_cqi_publish, id=ta_id, is_active=True)
    else:
        ta = get_object_or_404(_ta_qs_cqi_publish, id=ta_id, staff__user=request.user, is_active=True)

    # If there is an existing published CQI and publish control locks it, block re-publish.
    try:
        existing = AcV2CqiAttained.objects.filter(teaching_assignment=ta).first()
        if existing is not None:
            pc_existing = check_cqi_publish_control(existing)
            if pc_existing.get('publish_control_enabled') and pc_existing.get('is_locked') and not pc_existing.get('is_editable'):
                return Response({'detail': 'CQI is locked. Request Edit to re-publish.'}, status=403)
    except Exception:
        # Fail open for backward compatibility.
        pass

    body = request.data if isinstance(request.data, dict) else {}
    entries = body.get('entries')
    co_numbers = body.get('co_numbers')
    if not isinstance(co_numbers, list):
        co_numbers = None

    if entries is None:
        draft = AcV2CqiAssignment.objects.filter(teaching_assignment=ta).first()
        entries = (draft.draft_entries if draft is not None else None)
        if co_numbers is None and draft is not None:
            co_numbers = draft.co_numbers or []

    if entries is None or not isinstance(entries, dict):
        return Response({'detail': 'Missing entries (save a draft first or send entries).'}, status=status.HTTP_400_BAD_REQUEST)

    user_id = getattr(getattr(request, 'user', None), 'id', None)

    # Get or create the attained record, then merge entries per-CO.
    # This preserves entries from other CQI exams (e.g. CQI 1's CO1/CO2 must not be
    # wiped when CQI 2 (CO3/CO4/CO5) is published, and vice-versa).
    obj, _created = AcV2CqiAttained.objects.get_or_create(
        teaching_assignment=ta,
        defaults={
            'entries': {},
            'co_numbers': [],
            'published_by': user_id,
            'edit_window_until': None,
            'has_pending_edit_request': False,
        },
    )

    # Merge new entries into existing at per-student, per-CO level so that
    # publishing CQI 2 doesn't overwrite CQI 1 entries for the same students.
    merged_entries = dict(obj.entries) if isinstance(obj.entries, dict) else {}
    for _sid, _co_updates in entries.items():
        if isinstance(_co_updates, dict):
            _student_entry = dict(merged_entries.get(_sid) or {})
            _student_entry.update(_co_updates)
            merged_entries[_sid] = _student_entry

    # co_numbers = union of existing + new (so both CQI 1 and CQI 2 COs are tracked)
    _existing_co_nums = set(obj.co_numbers or [])
    _existing_co_nums.update(co_numbers or [])

    obj.entries = merged_entries
    obj.co_numbers = sorted(_existing_co_nums)
    obj.published_by = user_id
    obj.edit_window_until = None
    obj.has_pending_edit_request = False
    obj.save(update_fields=['entries', 'co_numbers', 'published_by', 'edit_window_until', 'has_pending_edit_request'])

    try:
        for acv2_section in ta.acv2_sections.select_related('course__class_type').all():
            compute_section_internal_marks(acv2_section)
    except Exception:
        pass

    pc = check_cqi_publish_control(obj)
    return Response({
        'status': 'ok',
        'published_at': obj.published_at.isoformat() if getattr(obj, 'published_at', None) else None,
        'publish_control': pc,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_academic_notification_flags(request):
    """Read-only notification toggles for faculty UI (no templates)."""

    obj, _ = AcV2AcademicNotificationSetting.objects.get_or_create(
        key='DEFAULT',
        defaults={
            'student_publish_enabled': False,
            'notify_on_first_publish': True,
            'notify_on_row_edits_only': True,
            'notify_on_every_publish_click': False,
            'cqi_announce_enabled': False,
        },
    )

    return Response({
        'student_publish_enabled': bool(getattr(obj, 'student_publish_enabled', False)),
        'cqi_announce_enabled': bool(getattr(obj, 'cqi_announce_enabled', False)),
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def faculty_course_cqi_announce(request, ta_id: int):
    """Send CQI announce WhatsApp messages to students who matched any CQI condition."""

    from academics.models import TeachingAssignment, StudentSectionAssignment
    from accounts.services.sms import send_whatsapp
    import logging

    logger = logging.getLogger('academic_v2.notifications')

    try:
        # Access check
        _ta_qs = TeachingAssignment.objects.select_related('curriculum_row', 'elective_subject', 'section', 'staff')
        if _has_admin_bypass_access(request.user):
            ta = get_object_or_404(_ta_qs, id=ta_id, is_active=True)
        else:
            ta = get_object_or_404(_ta_qs, id=ta_id, staff__user=request.user, is_active=True)

        cfg, _ = AcV2AcademicNotificationSetting.objects.get_or_create(key='DEFAULT')
        if not bool(getattr(cfg, 'cqi_announce_enabled', False)):
            return Response({'detail': 'CQI announce is disabled.'}, status=403)

        # Reuse the authoritative CQI condition evaluation from the CO summary logic.
        # Bypass the GET-only DRF wrapper so POST announce requests can still
        # access the same evaluated CQI rows.
        django_request = request._request if hasattr(request, '_request') else request
        summary_func = getattr(faculty_course_co_summary, '__wrapped__', faculty_course_co_summary)
        summary_res = summary_func(django_request, ta_id)
        if isinstance(summary_res, Response):
            summary = summary_res.data if hasattr(summary_res, 'data') else {}
        else:
            summary = {}
        
        if not isinstance(summary, dict):
            logger.error(f'Invalid summary response: {type(summary)}')
            return Response({'detail': 'Failed to compute CQI summary.'}, status=500)

        # Course/faculty context
        course_code = str(summary.get('course_code') or '')
        course_name = str(summary.get('course_name') or '')
        if not course_code or not course_name:
            try:
                cr = getattr(ta, 'curriculum_row', None)
                es = getattr(ta, 'elective_subject', None)
                if not course_code:
                    course_code = str(getattr(cr, 'course_code', None) or getattr(es, 'course_code', None) or '')
                if not course_name:
                    course_name = str(getattr(cr, 'course_name', None) or getattr(es, 'course_name', None) or '')
            except Exception:
                pass
        course_code = course_code or '-'
        course_name = course_name or '-'

        announce_co_numbers: list[int] = []
        try:
            raw_cos = ((summary.get('cqi_config') or {}) if isinstance(summary.get('cqi_config'), dict) else {}).get('cos', [])
            if isinstance(raw_cos, list):
                announce_co_numbers = sorted({int(x) for x in raw_cos if str(x).strip().isdigit() and int(x) > 0})
        except Exception:
            announce_co_numbers = []
        try:
            faculty_name = str(request.user)
        except Exception:
            faculty_name = ''

        # Student roster + mobiles
        sec = ta.section
        assignments = (
            StudentSectionAssignment.objects
            .filter(section=sec, end_date__isnull=True)
            .select_related('student__user')
            .order_by('student__reg_no')
        )
        roster_by_sid = {str(sa.student_id): sa.student for sa in assignments if getattr(sa, 'student_id', None)}

        students = summary.get('students') if isinstance(summary.get('students'), list) else []

        # Build student_id -> condition titles from summary.
        # For announcement we target every student that appears in the CQI table;
        # if the summary comes back empty, fall back to the active roster so the
        # announcement still reaches the students visible in the CQI entry page.
        cond_titles_by_sid: dict[str, list[str]] = {}
        target_sids: set[str] = set()
        for row in students:
            if not isinstance(row, dict):
                continue
            sid = str(row.get('student_id') or '').strip()
            if not sid:
                continue
            target_sids.add(sid)
            titles = row.get('cqi_satisfied_conditions')
            if isinstance(titles, list):
                clean = []
                seen = set()
                for t in titles:
                    s = str(t or '').strip()
                    if not s or s in seen:
                        continue
                    seen.add(s)
                    clean.append(s)
                if clean:
                    cond_titles_by_sid[sid] = clean

        if not target_sids:
            target_sids = set(roster_by_sid.keys())

        matched_sids = set(target_sids)
        matched_count = len(matched_sids)
        if not matched_sids:
            return Response({'status': 'ok', 'sent': 0, 'matched': 0})

        tpl = str(getattr(cfg, 'cqi_announce_template', '') or '')
        if not tpl.strip():
            return Response({'detail': 'CQI announce template is empty.'}, status=400)

        sent = 0
        rows_to_send = students if students else [
            {
                'student_id': sid,
                'co_totals': [],
                'cqi_satisfied_conditions': [],
            }
            for sid in matched_sids
        ]
        preview_rows = []
        for row in rows_to_send:
            if not isinstance(row, dict):
                continue
            sid = str(row.get('student_id') or '').strip()
            if not sid or sid not in matched_sids:
                continue
            sp = roster_by_sid.get(sid)
            if not sp:
                continue
            mobile = _resolve_mobile_for_student_profile(sp)
            if not mobile:
                continue

            co_totals = row.get('co_totals') if isinstance(row.get('co_totals'), list) else []
            # Prefer per-student announce target COs (those that matched CQI conditions)
            per_row_cos = row.get('cqi_announce_target_cos') if isinstance(row.get('cqi_announce_target_cos'), list) else []
            if per_row_cos:
                co_list_for_msg = sorted({int(x) for x in per_row_cos if isinstance(x, int) or (isinstance(x, str) and str(x).strip().isdigit())})
            else:
                if announce_co_numbers:
                    co_list_for_msg = list(announce_co_numbers)
                else:
                    co_list_for_msg = [i + 1 for i in range(len(co_totals)) if i >= 0]

            co_att = []
            for co_num in co_list_for_msg:
                value = None
                try:
                    idx = int(co_num) - 1
                    if 0 <= idx < len(co_totals):
                        value = co_totals[idx]
                except Exception:
                    value = None
                try:
                    num = float(value) if value is not None else None
                except Exception:
                    num = None
                if num is None:
                    co_att.append(f'CO{co_num}')
                else:
                    co_att.append(f'CO{co_num}: {round(num, 2)}')
            co_attainments = ', '.join(co_att).strip()
            satisfied_conditions = ', '.join(cond_titles_by_sid.get(sid, []))

            reg_no = str(getattr(sp, 'reg_no', '') or '')
            try:
                student_name = str(sp.user) if getattr(sp, 'user', None) else reg_no
            except Exception:
                student_name = reg_no

            msg = _render_notification_template(
                tpl,
                {
                    'course_code': course_code,
                    'course_name': course_name,
                    'faculty_name': faculty_name,
                    'student_name': student_name,
                    'register_number': reg_no,
                    'co_attainments': co_attainments,
                    'conditions': satisfied_conditions,
                    'satisfied_conditions': satisfied_conditions,
                },
            )

            try:
                result = send_whatsapp(mobile, msg)
                if bool(getattr(result, 'ok', False)):
                    sent += 1
                else:
                    logger.error(f'Failed to send CQI announce to {reg_no} ({mobile}): {getattr(result, "message", "unknown error")}')
            except Exception as e:
                logger.error(f'Failed to send CQI announce to {reg_no} ({mobile}): {e}')
                # continue sending to other students

            # Collect a small preview for debugging/verification (non-sensitive fields)
            try:
                preview_rows.append({
                    'student_id': sid,
                    'reg_no': reg_no,
                    'co_totals_len': len(co_totals) if isinstance(co_totals, list) else 0,
                    'co_list_for_msg': co_list_for_msg if isinstance(co_list_for_msg, list) else [],
                    'co_attainments_preview': co_attainments,
                    'satisfied_conditions': satisfied_conditions,
                })
            except Exception:
                pass

        return Response({'status': 'ok', 'sent': sent, 'matched': matched_count, 'preview': preview_rows[:20]})
    
    except Exception as e:
        logger.exception('CQI announce endpoint error')
        return Response({'detail': f'Announcement failed: {str(e)}'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def faculty_course_cqi_request_edit(request, ta_id: int):
    """Faculty submits a Request Edit for a published CQI snapshot."""
    from academics.models import TeachingAssignment
    from .models import AcV2CqiAttained

    if _has_admin_bypass_access(request.user):
        ta = get_object_or_404(TeachingAssignment, id=ta_id, is_active=True)
    else:
        ta = get_object_or_404(TeachingAssignment, id=ta_id, staff__user=request.user, is_active=True)

    cqi_attained = get_object_or_404(AcV2CqiAttained, teaching_assignment=ta)

    body = request.data if isinstance(request.data, dict) else {}
    reason = str(body.get('reason', '') or '').strip()
    if not reason:
        return Response({'detail': 'Reason is required.'}, status=400)

    result = create_cqi_edit_request(cqi_attained, request.user, reason)
    if not result['success']:
        return Response({'detail': result['error']}, status=400)

    # Refresh cqi_attained to see newly created request
    cqi_attained.refresh_from_db()

    # Return updated publish_control payload so frontend can refresh state
    pc = check_cqi_publish_control(cqi_attained)
    return Response({'status': 'ok', 'request_id': result['request_id'], 'publish_control': pc})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def faculty_course_cqi_cancel_edit_request(request, ta_id: int):
    """Cancel a pending CQI edit request submitted by the current user."""
    from academics.models import TeachingAssignment
    from .models import AcV2CqiAttained, AcV2CqiEditRequest

    if _has_admin_bypass_access(request.user):
        ta = get_object_or_404(TeachingAssignment, id=ta_id, is_active=True)
    else:
        ta = get_object_or_404(TeachingAssignment, id=ta_id, staff__user=request.user, is_active=True)

    cqi_attained = get_object_or_404(AcV2CqiAttained, teaching_assignment=ta)

    pending = (
        cqi_attained.edit_requests
        .filter(status__in=['PENDING', 'HOD_PENDING', 'IQAC_PENDING'])
        .order_by('-requested_at')
        .first()
    )
    if not pending:
        return Response({'detail': 'No pending edit request found.'}, status=404)

    pending.status = 'CANCELLED'
    pending.save(update_fields=['status'])
    cqi_attained.has_pending_edit_request = False
    cqi_attained.save(update_fields=['has_pending_edit_request'])

    # Refresh to ensure latest state
    cqi_attained.refresh_from_db()

    pc = check_cqi_publish_control(cqi_attained)
    return Response({'status': 'ok', 'publish_control': pc})


# ============================================================================
# IMPORT MARKS FROM EXCEL
# ============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def faculty_exam_import_marks(request, exam_id):
    """Import marks from uploaded Excel file, matching by register number."""
    from academics.models import StudentSectionAssignment, StudentProfile
    import openpyxl
    from io import BytesIO

    ea_qs = AcV2ExamAssignment.objects.select_related('section__teaching_assignment__section')
    if _has_admin_bypass_access(request.user):
        ea = get_object_or_404(ea_qs, id=exam_id)
    else:
        ea = get_object_or_404(ea_qs, id=exam_id, section__faculty_user=request.user)

    # Check if import is allowed: Allow if the exam is editable
    if not _has_admin_bypass_access(request.user) and not ea.is_editable():
        return Response({'detail': 'Exam is locked'}, status=403)

    uploaded = request.FILES.get('file')
    if not uploaded:
        return Response({'detail': 'No file uploaded'}, status=400)

    # Validate file extension
    name = uploaded.name.lower()
    if not name.endswith(('.xlsx', '.xls')):
        return Response({'detail': 'Only .xlsx or .xls files supported'}, status=400)

    # Limit file size (10 MB)
    if uploaded.size > 10 * 1024 * 1024:
        return Response({'detail': 'File too large (max 10 MB)'}, status=400)

    try:
        wb = openpyxl.load_workbook(BytesIO(uploaded.read()), data_only=True)
        ws = wb.active
    except Exception:
        return Response({'detail': 'Failed to read Excel file'}, status=400)

    # Find header row (look for 'Register Number' in first 5 rows)
    header_row = None
    reg_col = None
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value or '').strip().lower()
            if val in ('register number', 'reg no', 'reg_no', 'regno', 'registration number', 'roll no', 'roll_no', 'roll number'):
                header_row = r
                reg_col = c
                break
        if header_row:
            break

    if not header_row or not reg_col:
        return Response({'detail': 'Could not find "Register Number" column in the header'}, status=400)

    # Read headers from header_row
    headers = []
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=c).value or '').strip()
        headers.append(val)

    def _norm_header(v) -> str:
        return str(v or '').strip()

    def _normalize_import_header(v) -> str:
        text = _norm_header(v)
        # Remove trailing parenthesized details like "Q1 (MAX: 10, CO1)" -> "Q1"
        text = re.sub(r'\s*\([^)]*\)$', '', text).strip()
        return text.lower()

    def _question_header(i, q):
        title_raw = _norm_header(q.get('question_number') or q.get('title') or '')
        if not title_raw:
            return f'Q{i + 1}'
        if re.fullmatch(r'\d+', title_raw):
            return f'Q{title_raw}'
        return title_raw

    # Build question columns from the resolved QP pattern for this exam assignment.
    qp_type = ea.qp_type or ea.exam or ''
    p = ea.get_qp_pattern() or {}
    question_cols = []
    if isinstance(p, dict) and isinstance(p.get('questions'), list):
        for i, q in enumerate(p.get('questions') or []):
            if not isinstance(q, dict):
                continue
            if q.get('enabled') is False:
                continue
            question_cols.append({
                'key': q.get('id') or f'q{i}',
                'title': _question_header(i, q),
                'max_marks': q.get('max_marks') or q.get('marks') or 0,
            })
    elif isinstance(p, dict):
        titles = p.get('titles', [])
        marks_list = p.get('marks', [])
        enabled = p.get('enabled', [])
        for i in range(len(titles)):
            if i < len(enabled) and not enabled[i]:
                continue
            question_cols.append({
                'key': f'q{i}',
                'title': _question_header(i, {'question_number': titles[i] if i < len(titles) else '', 'title': titles[i] if i < len(titles) else ''}),
                'max_marks': marks_list[i] if i < len(marks_list) else 0,
            })

    # Map header titles to column indices (0-based).
    # Keep duplicate question titles in order to avoid accidental key overwrite
    # (e.g., repeated labels like "Q1" or numeric headers).
    q_max_by_key = {q['key']: q.get('max_marks') or 0 for q in question_cols}
    header_q_map = {}  # col_index -> question key
    total_col = None
    absent_col = None

    for c_idx, h in enumerate(headers):
        h_norm = _normalize_import_header(h)
        if h_norm in ('total', 'marks', 'total marks', 'total mark'):
            total_col = c_idx
        elif h_norm in ('absent', 'abs', 'absent?'):
            absent_col = c_idx

    # Positional check boundaries
    student_name_col = None
    for idx, h in enumerate(headers):
        if _norm_header(h).lower() in ('student name', 'name'):
            student_name_col = idx
            break

    def _match_header_to_question(header_str, question_title) -> bool:
        h_clean = re.sub(r'\s*\([^)]*\)$', '', str(header_str)).strip().lower()
        q_clean = re.sub(r'\s*\([^)]*\)$', '', str(question_title)).strip().lower()
        h_num = re.sub(r'^q\.?', '', h_clean)
        q_num = re.sub(r'^q\.?', '', q_clean)
        if h_clean == q_clean or h_num == q_num:
            return True
        if h_num.isdigit() and q_num.isdigit() and int(h_num) == int(q_num):
            return True
        return False

    # Perform robust name-based matching
    matched_q_keys = set()
    for c_idx, h in enumerate(headers):
        if c_idx == reg_col - 1 or (student_name_col is not None and c_idx == student_name_col) or c_idx in (total_col, absent_col):
            continue
        h_clean = _normalize_import_header(h)
        if h_clean in ('sl no', 'sl.no', 'slno'):
            continue
        for q in question_cols:
            if q['key'] in matched_q_keys:
                continue
            if _match_header_to_question(h, q['title']):
                header_q_map[c_idx] = q['key']
                matched_q_keys.add(q['key'])
                break

    # Positional fallback for remaining unmatched questions
    if len(header_q_map) < len(question_cols) and student_name_col is not None:
        if total_col is None and len(headers) >= 2:
            total_col = len(headers) - 2
        if absent_col is None and len(headers) >= 1:
            absent_col = len(headers) - 1

        right_bound = total_col if total_col is not None else absent_col
        if right_bound is not None and right_bound > student_name_col:
            candidate_cols = []
            for c_idx in range(student_name_col + 1, right_bound):
                if c_idx not in header_q_map and c_idx != reg_col - 1:
                    candidate_cols.append(c_idx)

            candidate_idx = 0
            for q in question_cols:
                if q['key'] in matched_q_keys:
                    continue
                if candidate_idx < len(candidate_cols):
                    col_idx = candidate_cols[candidate_idx]
                    header_q_map[col_idx] = q['key']
                    matched_q_keys.add(q['key'])
                    candidate_idx += 1

    # Get students in section, build reg_no -> student map
    ta = ea.section.teaching_assignment
    acad_sec = ta.section
    assignments = (
        StudentSectionAssignment.objects
        .filter(section=acad_sec, end_date__isnull=True)
        .select_related('student__user')
    )
    reg_to_student = {}
    for sa in assignments:
        rn = (sa.student.reg_no or '').strip().upper()
        if rn:
            reg_to_student[rn] = sa.student

    # Read data rows (skip header and any sub-header row right after)
    start_row = header_row + 1
    # Skip sub-header row if present (check if it starts with non-numeric)
    first_reg = ws.cell(row=start_row, column=reg_col).value
    first_val = _norm_header(first_reg)
    if not first_val or first_val.lower().startswith('max') or not any(c.isdigit() for c in first_val):
        start_row += 1

    matched = 0
    skipped = 0
    imported_students = []
    unfilled_rows = []

    for r in range(start_row, ws.max_row + 1):
        reg_val = ws.cell(row=r, column=reg_col).value
        if not reg_val:
            continue
        reg_no = str(reg_val).strip().upper()
        if not reg_no or not any(c.isdigit() for c in reg_no):
            continue

        sp = reg_to_student.get(reg_no)
        if not sp:
            skipped += 1
            continue

        # Read question marks
        co_marks = {}
        for c_idx, q_key in header_q_map.items():
            cell_val = ws.cell(row=r, column=c_idx + 1).value  # +1 for openpyxl 1-indexed
            if cell_val is not None:
                try:
                    num = float(cell_val)
                    # Validate against per-question max marks
                    max_m = q_max_by_key.get(q_key, 999)
                    if 0 <= num <= max_m:
                        co_marks[q_key] = num
                except (ValueError, TypeError):
                    pass

        # Read total (only if no question cols, or as fallback)
        total_mark = None
        if total_col is not None:
            tv = ws.cell(row=r, column=total_col + 1).value
            if tv is not None:
                try:
                    total_mark = float(tv)
                    if total_mark < 0 or total_mark > float(ea.max_marks or 999):
                        total_mark = None
                except (ValueError, TypeError):
                    total_mark = None

        # Calculate total from question marks if present
        if co_marks:
            total_mark = round(sum(co_marks.values()), 2)

        # Read absent
        is_absent = False
        if absent_col is not None:
            av = ws.cell(row=r, column=absent_col + 1).value
            if av and str(av).strip().lower() in ('yes', 'y', '1', 'true', 'absent'):
                is_absent = True
                total_mark = None
                co_marks = {}

        imported_students.append({
            'student_id': str(sp.id),
            'roll_number': sp.reg_no or '',
            'name': str(sp.user) if sp.user else sp.reg_no or '',
            'mark': total_mark,
            'co_marks': co_marks,
            'is_absent': is_absent,
        })
        if not is_absent and total_mark is None and not co_marks:
            unfilled_rows.append({
                'roll_number': sp.reg_no or '',
                'name': str(sp.user) if sp.user else sp.reg_no or '',
                'row_number': r,
            })
        matched += 1

    # If caller requested to apply the import, persist marks into AcV2StudentMark
    apply_flag = False
    try:
        ap = None
        if hasattr(request, 'data') and isinstance(request.data, dict):
            ap = request.data.get('apply')
        if ap is None and hasattr(request, 'query_params'):
            ap = request.query_params.get('apply')
        if ap is None:
            ap = request.POST.get('apply') if hasattr(request, 'POST') else None
        apply_flag = str(ap).strip().lower() in ('1', 'true', 'yes', 'y') if ap is not None else False
    except Exception:
        apply_flag = False

    if apply_flag:
        from decimal import Decimal

        applied = 0
        with transaction.atomic():
            for entry in imported_students:
                try:
                    sid = entry.get('student_id')
                    if sid is None:
                        continue
                    sp = StudentProfile.objects.filter(id=int(sid)).first()
                    if not sp:
                        continue

                    co_marks = entry.get('co_marks') if isinstance(entry.get('co_marks'), dict) else {}
                    # normalize numeric marks
                    qm = {}
                    for k, v in co_marks.items():
                        try:
                            if v is None or v == '':
                                continue
                            qm[str(k)] = float(v)
                        except Exception:
                            continue

                    total_val = entry.get('mark')
                    try:
                        total_decimal = Decimal(str(total_val)) if total_val is not None else None
                    except Exception:
                        total_decimal = None

                    defaults = {
                        'reg_no': (entry.get('roll_number') or '')[:50],
                        'student_name': (entry.get('name') or '')[:255],
                        'question_marks': qm,
                        'total_mark': total_decimal,
                        'is_absent': bool(entry.get('is_absent')),
                    }

                    sm, created = AcV2StudentMark.objects.update_or_create(
                        exam_assignment=ea,
                        student=sp,
                        defaults=defaults,
                    )

                    # Recompute CO marks and totals using exam pattern
                    try:
                        qp_pattern = ea.get_qp_pattern() or {}
                        sm.question_marks = qm
                        sm.total_mark = total_decimal
                        sm.is_absent = bool(entry.get('is_absent'))
                        sm.calculate_co_marks(qp_pattern)
                        sm.calculate_total()
                        sm.save()
                    except Exception:
                        # If calculation fails, still keep the raw imported row
                        sm.save()

                    applied += 1
                except Exception:
                    continue

        return Response({'status': 'applied', 'applied_count': applied}, status=200)

    return Response({
        'status': 'preview',
        'matched': matched,
        'skipped': skipped,
        'unfilled_count': len(unfilled_rows),
        'unfilled_rows': unfilled_rows[:20],
        'total_in_file': matched + skipped,
        'total_in_class': len(reg_to_student),
        'students': imported_students,
    })


# ============================================================================
# ADMIN BYPASS API VIEWS
# ============================================================================

from .models import AcV2BypassSession, AcV2BypassLog
from django.utils import timezone as _tz
import os as _os


def _require_admin(request):
    if not request.user or not request.user.is_authenticated:
        return Response({'detail': 'Authentication required.'}, status=401)
    if not _has_admin_bypass_access(request.user):
        return Response({'detail': 'Admin access required.'}, status=403)
    return None


def _serialize_session(session, include_logs=False):
    faculty = session.faculty_user
    shared_by = session.shared_by
    shared_accessed_by = session.shared_accessed_by
    data = {
        'id': str(session.id),
        'admin': {
            'id': session.admin.id,
            'name': session.admin.get_full_name() or session.admin.username,
        },
        'faculty': {
            'id': faculty.id,
            'name': faculty.get_full_name() or faculty.username,
        } if faculty else None,
        'teaching_assignment_id': session.teaching_assignment_id,
        'course_code': session.course_code,
        'course_name': session.course_name,
        'section_name': session.section_name,
        'started_at': session.started_at.isoformat(),
        'ended_at': session.ended_at.isoformat() if session.ended_at else None,
        'duration_seconds': session.duration_seconds,
        'share_token': session.share_token or None,
        'share_expires_at': session.share_expires_at.isoformat() if session.share_expires_at else None,
        'share_max_uses': session.share_max_uses,
        'share_use_count': session.share_use_count,
        'shared_by': {
            'id': shared_by.id,
            'name': shared_by.get_full_name() or shared_by.username,
        } if shared_by else None,
        'shared_accessed_by': {
            'id': shared_accessed_by.id,
            'name': shared_accessed_by.get_full_name() or shared_accessed_by.username,
        } if shared_accessed_by else None,
    }
    if include_logs:
        data['logs'] = [_serialize_log(lg) for lg in session.logs.all()]
    return data


def _serialize_log(log):
    actor = log.actor
    return {
        'id': str(log.id),
        'action': log.action,
        'description': log.description,
        'extra': log.extra,
        'created_at': log.created_at.isoformat(),
        'actor': {
            'id': actor.id,
            'name': actor.get_full_name() or actor.username,
        } if actor else None,
    }


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bypass_start(request):
    """
    Start a bypass session.
    Body: { teaching_assignment_id, course_code, course_name, section_name, faculty_user_id }
    Returns: { session_id }
    """
    err = _require_admin(request)
    if err:
        return err

    ta_id = request.data.get('teaching_assignment_id')
    course_code = request.data.get('course_code', '')
    course_name = request.data.get('course_name', '')
    section_name = request.data.get('section_name', '')
    faculty_user_id = request.data.get('faculty_user_id')

    from django.contrib.auth import get_user_model
    User = get_user_model()
    faculty = None
    if faculty_user_id:
        try:
            faculty = User.objects.get(id=faculty_user_id)
        except User.DoesNotExist:
            pass

    session = AcV2BypassSession.objects.create(
        admin=request.user,
        faculty_user=faculty,
        teaching_assignment_id=ta_id,
        course_code=course_code,
        course_name=course_name,
        section_name=section_name,
    )
    AcV2BypassLog.objects.create(
        session=session,
        actor=request.user,
        action=AcV2BypassLog.ACTION_ENTER,
        description=f"Admin bypass started for {course_code} – {section_name}",
        extra={
            'teaching_assignment_id': ta_id,
            'faculty_user_id': faculty_user_id,
        },
    )
    return Response({'session_id': str(session.id)}, status=201)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bypass_end(request, session_id):
    """End a bypass session."""
    err = _require_admin(request)
    if err:
        return err

    session = get_object_or_404(AcV2BypassSession, id=session_id)
    if not session.ended_at:
        session.ended_at = _tz.now()
        session.save(update_fields=['ended_at'])
    AcV2BypassLog.objects.create(
        session=session,
        actor=request.user,
        action=AcV2BypassLog.ACTION_EXIT,
        description=f"Admin bypass exited. Duration: {session.duration_seconds}s",
        extra={'duration_seconds': session.duration_seconds},
    )
    return Response({'status': 'ended', 'duration_seconds': session.duration_seconds})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bypass_add_log(request, session_id):
    """
    Add a log entry to a bypass session (generic – for mark edits, publishes, etc.)
    Body: { action, description, extra }
    """
    err = _require_admin(request)
    if err:
        return err

    session = get_object_or_404(AcV2BypassSession, id=session_id)
    action = request.data.get('action', AcV2BypassLog.ACTION_OTHER)
    description = request.data.get('description', '')
    extra = request.data.get('extra', {})

    log = AcV2BypassLog.objects.create(
        session=session,
        actor=request.user,
        action=action,
        description=description,
        extra=extra if isinstance(extra, dict) else {},
    )
    return Response(_serialize_log(log), status=201)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bypass_reset_course(request, session_id):
    """
    Reset ALL exam assignments for a teaching assignment (section).
    Deletes student marks and draft marks. Sets exam status back to DRAFT.
    Only affects the specific teaching_assignment_id linked to the bypass session.
    """
    err = _require_admin(request)
    if err:
        return err

    session = get_object_or_404(AcV2BypassSession, id=session_id)
    ta_id = session.teaching_assignment_id
    if not ta_id:
        return Response({'detail': 'No teaching assignment linked to this session.'}, status=400)

    with transaction.atomic():
        # Get the AcV2Section for this teaching assignment
        from .models import AcV2Section
        sections = AcV2Section.objects.filter(teaching_assignment_id=ta_id)
        if not sections.exists():
            return Response({'detail': 'No sections found for this teaching assignment.'}, status=404)

        exam_assignments = AcV2ExamAssignment.objects.filter(section__in=sections)
        exam_count = exam_assignments.count()

        # Delete marks
        deleted_marks, _ = AcV2StudentMark.objects.filter(exam_assignment__in=exam_assignments).delete()
        deleted_drafts, _ = AcV2DraftMark.objects.filter(exam_assignment__in=exam_assignments).delete()

        # Reset exam status to DRAFT
        exam_assignments.update(
            status='DRAFT',
            published_at=None,
            published_by=None,
            published_data={},
            draft_data={},
            has_pending_edit_request=False,
            edit_window_until=None,
            edit_window_until_publish=False,
            last_saved_at=None,
        )

        # Reset CQI (draft + published) for THIS teaching assignment only
        from .models import AcV2CqiAssignment, AcV2CqiAttained

        cqi_draft_cleared = 0
        cqi_published_deleted = 0
        try:
            cqi_asg = AcV2CqiAssignment.objects.filter(teaching_assignment_id=ta_id).first()
            if cqi_asg is not None:
                cqi_asg.draft_entries = {}
                cqi_asg.draft_updated_by = None
                cqi_asg.save(update_fields=['draft_entries', 'draft_updated_by', 'draft_updated_at'])
                cqi_draft_cleared = 1
        except Exception:
            pass

        try:
            cqi_attained_qs = AcV2CqiAttained.objects.filter(teaching_assignment_id=ta_id)
            cqi_published_deleted = cqi_attained_qs.count()
            if cqi_published_deleted:
                cqi_attained_qs.delete()
        except Exception:
            pass

    AcV2BypassLog.objects.create(
        session=session,
        actor=request.user,
        action=AcV2BypassLog.ACTION_RESET_COURSE,
        description=f"Course reset: {session.course_code} – {session.section_name}. {exam_count} exams reset. CQI reset.",
        extra={
            'teaching_assignment_id': ta_id,
            'exam_count': exam_count,
            'deleted_marks': deleted_marks,
            'deleted_drafts': deleted_drafts,
            'cqi_draft_cleared': cqi_draft_cleared,
            'cqi_published_deleted': cqi_published_deleted,
        },
    )
    return Response({
        'status': 'reset',
        'exam_count': exam_count,
        'deleted_marks': deleted_marks,
        'deleted_drafts': deleted_drafts,
        'cqi_draft_cleared': cqi_draft_cleared,
        'cqi_published_deleted': cqi_published_deleted,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bypass_reset_exam(request, session_id, exam_id):
    """
    Reset a single exam assignment (marks + status) by its UUID.
    """
    err = _require_admin(request)
    if err:
        return err

    session = get_object_or_404(AcV2BypassSession, id=session_id)
    exam = get_object_or_404(AcV2ExamAssignment, id=exam_id)

    with transaction.atomic():
        deleted_marks, _ = AcV2StudentMark.objects.filter(exam_assignment=exam).delete()
        deleted_drafts, _ = AcV2DraftMark.objects.filter(exam_assignment=exam).delete()
        exam.status = 'DRAFT'
        exam.published_at = None
        exam.published_by = None
        exam.published_data = {}
        exam.draft_data = {}
        exam.has_pending_edit_request = False
        exam.edit_window_until = None
        exam.edit_window_until_publish = False
        exam.last_saved_at = None
        exam.save()

    AcV2BypassLog.objects.create(
        session=session,
        actor=request.user,
        action=AcV2BypassLog.ACTION_RESET_EXAM,
        description=f"Exam reset: {exam.exam_display_name or exam.exam} (ID: {exam_id})",
        extra={
            'exam_id': str(exam_id),
            'exam_name': exam.exam_display_name or exam.exam,
            'deleted_marks': deleted_marks,
            'deleted_drafts': deleted_drafts,
        },
    )
    return Response({
        'status': 'reset',
        'exam_id': str(exam_id),
        'deleted_marks': deleted_marks,
        'deleted_drafts': deleted_drafts,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bypass_send_message(request, session_id):
    """
    Send a WhatsApp message to the faculty being bypassed.
    Body: { message, mobile? }  — mobile defaults to faculty's profile mobile.
    """
    err = _require_admin(request)
    if err:
        return err

    session = get_object_or_404(AcV2BypassSession, id=session_id)
    message = str(request.data.get('message', '')).strip()
    if not message:
        return Response({'detail': 'message is required.'}, status=400)

    # Resolve mobile number
    mobile = str(request.data.get('mobile', '') or '').strip()
    if not mobile and session.faculty_user:
        try:
            sp = session.faculty_user.staff_profile
            mobile = str(sp.mobile or '').strip()
        except Exception:
            pass
        if not mobile:
            try:
                mobile = str(session.faculty_user.mobile or '').strip()
            except Exception:
                pass

    if not mobile:
        return Response({'detail': 'No mobile number found for faculty.'}, status=400)

    from accounts.services.sms import send_whatsapp
    result = send_whatsapp(mobile, message)

    AcV2BypassLog.objects.create(
        session=session,
        actor=request.user,
        action=AcV2BypassLog.ACTION_MESSAGE,
        description=f"WhatsApp sent to {mobile}: {message[:100]}",
        extra={
            'mobile': mobile,
            'message': message,
            'result': str(result),
        },
    )
    return Response({'status': 'sent', 'mobile': mobile})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bypass_create_share_link(request, session_id):
    """
    Generate a time-limited shared bypass link.
    Body: { expires_at } — ISO datetime string.
    Returns: { share_token, share_url, expires_at }
    """
    err = _require_admin(request)
    if err:
        return err

    session = get_object_or_404(AcV2BypassSession, id=session_id)
    expires_at_str = request.data.get('expires_at')
    if not expires_at_str:
        return Response({'detail': 'expires_at is required.'}, status=400)

    from django.utils.dateparse import parse_datetime
    expires_at = parse_datetime(str(expires_at_str))
    if not expires_at:
        return Response({'detail': 'Invalid expires_at datetime format.'}, status=400)

    token = AcV2BypassSession.generate_share_token()
    max_uses = int(request.data.get('max_uses', 1) or 1)
    max_uses = max(1, min(max_uses, 50))  # clamp 1–50
    session.share_token = token
    session.share_expires_at = expires_at
    session.shared_by = request.user
    session.share_max_uses = max_uses
    session.share_use_count = 0
    session.save(update_fields=['share_token', 'share_expires_at', 'shared_by', 'share_max_uses', 'share_use_count'])

    site_root = str(
        getattr(__import__('django.conf', fromlist=['settings']).settings, 'VITE_API_BASE', '')
        or _os.getenv('VITE_API_BASE')
        or 'https://idcs.zynix.us'
    ).rstrip('/')
    share_url = f"{site_root}/academic-v2/bypass-share/{token}"

    AcV2BypassLog.objects.create(
        session=session,
        actor=request.user,
        action=AcV2BypassLog.ACTION_SHARE,
        description=f"Shared bypass link created. Expires: {expires_at.isoformat()}",
        extra={
            'share_token': token,
            'share_url': share_url,
            'expires_at': expires_at.isoformat(),
        },
    )
    return Response({
        'share_token': token,
        'share_url': share_url,
        'expires_at': expires_at.isoformat(),
    }, status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bypass_session_detail(request, session_id):
    """
    Get a single bypass session by ID.
    Accessible to admins OR the faculty user who is the target of the session.
    This allows faculty who accessed via a shared bypass link to load session info
    without needing admin permissions.
    """
    try:
        session = AcV2BypassSession.objects.select_related(
            'admin', 'faculty_user', 'shared_by', 'shared_accessed_by'
        ).get(id=session_id)
    except (AcV2BypassSession.DoesNotExist, Exception):
        return Response({'detail': 'Session not found.'}, status=404)

    user = request.user
    is_admin = _has_admin_bypass_access(user)
    is_target = session.faculty_user_id == user.id
    if not (is_admin or is_target):
        return Response({'detail': 'Permission denied.'}, status=403)

    return Response(_serialize_session(session))


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bypass_validate_share(request, token):
    """
    Validate a shared bypass token and return the session details.
    If valid and not yet claimed, marks it as accessed by current user.
    """
    session = AcV2BypassSession.objects.filter(share_token=token).first()
    if not session:
        return Response({'valid': False, 'detail': 'Invalid token.'}, status=404)

    if session.share_expires_at and _tz.now() > session.share_expires_at:
        return Response({'valid': False, 'detail': 'Share link has expired.'}, status=410)

    user = request.user

    # Check if this user already accessed this session (allow re-entry without counting again)
    already_accessed = AcV2BypassLog.objects.filter(
        session=session,
        actor=user,
        action=AcV2BypassLog.ACTION_SHARE_ACCESSED,
    ).exists()

    if not already_accessed:
        # New user — check if the usage limit has been reached
        if session.share_max_uses and session.share_use_count >= session.share_max_uses:
            return Response(
                {'valid': False, 'detail': f'This share link has reached its usage limit ({session.share_max_uses} user(s) allowed).'},
                status=403,
            )
        # Increment use count and record who accessed
        session.share_use_count = (session.share_use_count or 0) + 1
        if not session.shared_accessed_by:
            session.shared_accessed_by = user
        session.save(update_fields=['share_use_count', 'shared_accessed_by'])

        site_root = str(
            getattr(__import__('django.conf', fromlist=['settings']).settings, 'VITE_API_BASE', '')
            or _os.getenv('VITE_API_BASE')
            or 'https://idcs.zynix.us'
        ).rstrip('/')
        share_url = f"{site_root}/academic-v2/bypass-share/{session.share_token}"

        AcV2BypassLog.objects.create(
            session=session,
            actor=user,
            action=AcV2BypassLog.ACTION_SHARE_ACCESSED,
            description=f"Shared bypass accessed by {user.get_full_name() or user.username}",
            extra={
                'user_id': user.id,
                'share_url': share_url,
                'share_token': session.share_token,
            },
        )

    return Response({
        'valid': True,
        'session': _serialize_session(session),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bypass_sessions_list(request):
    """List all bypass sessions with logs (admin only)."""
    err = _require_admin(request)
    if err:
        return err

    sessions = AcV2BypassSession.objects.select_related(
        'admin', 'faculty_user', 'shared_by', 'shared_accessed_by'
    ).prefetch_related('logs__actor').order_by('-started_at')

    # Optional filter by faculty_user_id or admin_id
    faculty_id = request.query_params.get('faculty_id')
    if faculty_id:
        sessions = sessions.filter(faculty_user_id=faculty_id)

    result = [_serialize_session(s, include_logs=True) for s in sessions[:200]]
    return Response(result)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_courses_list(request):
    """
    List all AcV2 sections (courses) for admin CourseManager.
    Optionally filter by search query.
    Returns section info with faculty details.
    """
    err = _require_admin(request)
    if err:
        return err

    search = (request.query_params.get('search') or '').strip()

    from .models import AcV2Section, AcV2Course
    from django.db.models import Count, Q as DQ

    qs = AcV2Section.objects.select_related(
        'course', 'course__class_type', 'course__semester',
        'faculty_user', 'teaching_assignment',
        'teaching_assignment__section',
        'teaching_assignment__section__batch__department',
    ).prefetch_related('exam_assignments').order_by(
        'course__subject_code', 'section_name'
    )

    if search:
        qs = qs.filter(
            DQ(course__subject_code__icontains=search)
            | DQ(course__subject_name__icontains=search)
            | DQ(section_name__icontains=search)
            | DQ(faculty_user__first_name__icontains=search)
            | DQ(faculty_user__last_name__icontains=search)
        )

    def _section_data(sec):
        faculty = sec.faculty_user
        course = sec.course
        total_exams = sec.exam_assignments.count()
        published = sum(1 for e in sec.exam_assignments.all() if e.status == 'PUBLISHED')
        ta = sec.teaching_assignment
        dept = ''
        semester_name = ''
        try:
            dept = ta.section.batch.department.name if ta and ta.section and ta.section.batch and ta.section.batch.department else ''
        except Exception:
            pass
        try:
            semester_name = str(course.semester) if course else ''
        except Exception:
            pass

        # Faculty profile photo
        faculty_photo = None
        if faculty:
            try:
                img = faculty.staff_profile.profile_image
                if img:
                    site_root = str(
                        getattr(__import__('django.conf', fromlist=['settings']).settings, 'VITE_API_BASE', '')
                        or _os.getenv('VITE_API_BASE')
                        or 'https://idcs.zynix.us'
                    ).rstrip('/')
                    faculty_photo = f"{site_root}/media/{str(img).lstrip('/')}"
            except Exception:
                pass

        return {
            'section_id': str(sec.id),
            'ta_id': sec.teaching_assignment_id,
            'course_code': course.subject_code if course else '',
            'course_name': course.subject_name if course else '',
            'class_type': course.class_type.name if course and course.class_type else course.class_type_name if course else '',
            'section_name': sec.section_name,
            'semester': semester_name,
            'department': dept,
            'total_exams': total_exams,
            'published_exams': published,
            'faculty': {
                'id': faculty.id,
                'name': faculty.get_full_name() or faculty.username,
                'photo': faculty_photo,
            } if faculty else None,
        }

    data = [_section_data(s) for s in qs[:500]]
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_course_faculty(request, ta_id):
    """
    Get all faculty who are teaching the same course code (same subject_code),
    returned with their sections and exam status.
    Actually returns the specific section info for a given ta_id.
    """
    err = _require_admin(request)
    if err:
        return err

    from .models import AcV2Section
    try:
        section = AcV2Section.objects.select_related(
            'course', 'course__class_type', 'course__semester',
            'faculty_user', 'teaching_assignment',
            'teaching_assignment__staff', 'teaching_assignment__staff__user',
        ).prefetch_related('exam_assignments').get(teaching_assignment_id=ta_id)
    except AcV2Section.DoesNotExist:
        return Response({'detail': 'Section not found.'}, status=404)

    course = section.course
    # Find all sections teaching the same subject_code in the same semester
    from .models import AcV2Course
    sibling_sections = AcV2Section.objects.filter(
        course__subject_code=course.subject_code,
        course__semester=course.semester,
    ).select_related(
        'course', 'faculty_user', 'teaching_assignment',
        'teaching_assignment__staff', 'teaching_assignment__staff__user',
    ).prefetch_related('exam_assignments').exclude(id=section.id)

    def _faculty_card(sec):
        faculty = sec.faculty_user
        # Sync faculty_user on retrieve if it differs from the assigned staff user
        assigned_user = getattr(getattr(sec.teaching_assignment, 'staff', None), 'user', None)
        if assigned_user and faculty != assigned_user:
            sec.faculty_user = assigned_user
            sec.save(update_fields=['faculty_user'])
            faculty = assigned_user
        total_exams = sec.exam_assignments.count()
        published = sum(1 for e in sec.exam_assignments.all() if e.status == 'PUBLISHED')
        faculty_photo = None
        if faculty:
            try:
                img = faculty.staff_profile.profile_image
                if img:
                    site_root = str(
                        getattr(__import__('django.conf', fromlist=['settings']).settings, 'VITE_API_BASE', '')
                        or _os.getenv('VITE_API_BASE')
                        or 'https://idcs.zynix.us'
                    ).rstrip('/')
                    faculty_photo = f"{site_root}/media/{str(img).lstrip('/')}"
            except Exception:
                pass
        return {
            'section_id': str(sec.id),
            'ta_id': sec.teaching_assignment_id,
            'section_name': sec.section_name,
            'total_exams': total_exams,
            'published_exams': published,
            'faculty': {
                'id': faculty.id,
                'name': faculty.get_full_name() or faculty.username,
                'photo': faculty_photo,
            } if faculty else None,
        }

    return Response({
        'course_code': course.subject_code,
        'course_name': course.subject_name,
        'sections': [_faculty_card(section)] + [_faculty_card(s) for s in sibling_sections],
    })


# ============================================================================
# FACULTY RESET NOTICES  (no migration required – queries existing BypassLog)
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def faculty_reset_notices(request, ta_id: int):
    """
    Return recent RESET_COURSE / RESET_EXAM bypass-log entries for this TA.
    Used by InternalMarkPage to show an animated warning popup to the faculty.
    Dismissal is tracked client-side via localStorage (keyed by log id).
    Only returns logs from the last 60 days.
    """
    from datetime import timedelta

    cutoff = _tz.now() - timedelta(days=60)

    logs = (
        AcV2BypassLog.objects
        .filter(
            session__teaching_assignment_id=ta_id,
            action__in=[AcV2BypassLog.ACTION_RESET_COURSE, AcV2BypassLog.ACTION_RESET_EXAM],
            created_at__gte=cutoff,
        )
        .select_related('session__admin', 'actor')
        .order_by('-created_at')[:20]
    )

    result = []
    for log in logs:
        admin_user = log.actor or log.session.admin if log.session else None
        admin_name = (admin_user.get_full_name() or admin_user.username) if admin_user else 'Admin'
        result.append({
            'id': str(log.id),
            'action': log.action,
            'description': log.description,
            'extra': log.extra or {},
            'created_at': log.created_at.isoformat(),
            'reset_by': {
                'name': admin_name,
                'role': 'IQAC / Admin',
            },
            'course_code': log.session.course_code if log.session else '',
            'section_name': log.session.section_name if log.session else '',
        })

    return Response(result)
