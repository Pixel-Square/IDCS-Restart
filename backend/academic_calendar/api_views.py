from __future__ import annotations

import json
import os
import uuid
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional

from django.conf import settings
from django.core.files.storage import default_storage
from django.db import transaction
from django.http import HttpResponse
from django.db.models import Q

from rest_framework import status
from rest_framework.decorators import api_view, parser_classes, permission_classes
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response

from academics.models import Course, Department, DepartmentRole, StudentProfile
from academics.utils import get_user_effective_departments

from .models import (
    AcademicCalendarEvent,
    HodColor,
    AcademicCalendar,
    AcademicCalendarDay,
    AcademicCalendarHoliday,
)
from .n8n_poster_service import fire_n8n_branding_async


def _normalize_text(s: Any) -> str:
    if s is None:
        return ''
    return str(s).strip().lower().replace('\u00a0', ' ').replace('\t', ' ').replace('\n', ' ')


def _normalize_department_name(s: Any) -> str:
    import re

    t = _normalize_text(s)
    t = re.sub(r'[^a-z0-9 ]+', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


TEMPLATE_COLUMNS = ['Date', 'Day', 'Working Days', 'II Year', 'III Year', 'IV Year', 'I Year']
TEMPLATE_DROPDOWN_VALUES = [
    'Placement training',
    'L1',
    'CIA 1',
    'L2',
    'CIA 2',
    'Model',
    'CQI',
    'ESE LAB',
    'ESE Theory',
]

ACADEMIC_CALENDAR_HOLIDAY_NOTE_PREFIX = 'ACADEMIC_CALENDAR:'


def _academic_year_from_dates(start: date, end: date) -> str:
    return f"{start.year}-{str(end.year)[-2:]}"


def _day_abbrev(d: date) -> str:
    # English 3-letter abbreviation
    return d.strftime('%a')


def _week_of_month(d: date) -> int:
    return ((d.day - 1) // 7) + 1


def _is_even_saturday(d: date) -> bool:
    return d.weekday() == 5 and (_week_of_month(d) % 2 == 0)


def _is_odd_saturday(d: date) -> bool:
    return d.weekday() == 5 and (_week_of_month(d) % 2 == 1)


def _normalize_simple_token(val: Any) -> str:
    return str(val or '').strip().lower()


def _is_working_days_holiday(val: Any) -> bool:
    token = _normalize_simple_token(val)
    if not token:
        return False
    for opt in TEMPLATE_DROPDOWN_VALUES:
        if token == opt.strip().lower():
            return False
    return True


def _to_yyyy_mm_dd(d: Optional[date]) -> Optional[str]:
    return d.isoformat() if d else None


def _safe_int(v: Any) -> Optional[int]:
    try:
        if v is None or v == '':
            return None
        return int(str(v).strip())
    except Exception:
        return None


def _clamp_dates(start: date, end: date) -> tuple[date, date]:
    if end < start:
        return start, start
    return start, end


def _department_matches_any(audience_csv: Optional[str], dept_names: List[str]) -> bool:
    if not audience_csv:
        return True
    wanted = {_normalize_department_name(x) for x in dept_names if x}
    if not wanted:
        return False
    parts = [_normalize_department_name(x) for x in str(audience_csv).split(',')]
    return any(p and p in wanted for p in parts)


def _get_staff_profile(user):
    return getattr(user, 'staff_profile', None)


def _get_student_profile(user):
    return getattr(user, 'student_profile', None)


def _is_iqac_user(user) -> bool:
    # Prefer role mapping when available
    try:
        if user.roles.filter(name__iexact='IQAC').exists():
            return True
    except Exception:
        pass

    sp = _get_staff_profile(user)
    if not sp:
        return False

    dept = getattr(sp, 'current_department', None)
    if not dept and hasattr(sp, 'get_current_department'):
        try:
            dept = sp.get_current_department()
        except Exception:
            dept = None
    if not dept:
        dept = getattr(sp, 'department', None)
    if not dept:
        return False

    code = _normalize_text(getattr(dept, 'code', None))
    name = _normalize_department_name(getattr(dept, 'name', None))
    short = _normalize_department_name(getattr(dept, 'short_name', None))
    return (code == 'iqac') or (name == 'iqac') or (short == 'iqac')


def _is_hod_user(user) -> bool:
    sp = _get_staff_profile(user)
    if not sp:
        return False

    try:
        if DepartmentRole.objects.filter(staff=sp, role='HOD', is_active=True).exists():
            return True
    except Exception:
        pass

    try:
        return user.roles.filter(name__iexact='HOD').exists()
    except Exception:
        return False


def _get_primary_department_display(user) -> str:
    sp = _get_staff_profile(user)
    if sp:
        dept = getattr(sp, 'current_department', None)
        if not dept and hasattr(sp, 'get_current_department'):
            try:
                dept = sp.get_current_department()
            except Exception:
                dept = None
        if not dept:
            dept = getattr(sp, 'department', None)
        if dept:
            return getattr(dept, 'name', '') or getattr(dept, 'code', '') or 'Unknown'

    st = _get_student_profile(user)
    if st:
        try:
            sec = st.current_section
            dept = sec and sec.batch and sec.batch.course and sec.batch.course.department
            if dept:
                return getattr(dept, 'name', '') or getattr(dept, 'code', '') or 'Unknown'
        except Exception:
            pass

    return 'Unknown'


def _derive_year_from_semester_number(sem_number: Any) -> Optional[int]:
    try:
        n = int(sem_number)
    except Exception:
        return None
    if n <= 0:
        return None
    return (n + 1) // 2


def _get_student_year(user) -> Optional[int]:
    st = _get_student_profile(user)
    if not st:
        return None
    try:
        sec = st.current_section
        sem = getattr(sec, 'semester', None)
        sem_number = getattr(sem, 'number', None)
        return _derive_year_from_semester_number(sem_number)
    except Exception:
        return None


def _roman_year(y: Optional[int]) -> Optional[str]:
    if not y:
        return None
    mapping = {1: 'I', 2: 'II', 3: 'III', 4: 'IV'}
    return mapping.get(int(y))


def _get_hod_owned_department_names(user) -> List[str]:
    ids = get_user_effective_departments(user)
    names: List[str] = []
    if ids:
        qs = Department.objects.filter(id__in=ids).values('name', 'code', 'short_name')
        for row in qs:
            for k in ('name', 'code', 'short_name'):
                v = row.get(k)
                if v:
                    names.append(str(v))

    primary = _get_primary_department_display(user)
    if primary:
        names.append(primary)

    seen = set()
    out = []
    for n in names:
        nn = _normalize_department_name(n)
        if not nn or nn in seen:
            continue
        seen.add(nn)
        out.append(n)
    return out


def _is_excluded_department(dept: Department) -> bool:
    code = _normalize_text(getattr(dept, 'code', None))
    short = _normalize_text(getattr(dept, 'short_name', None))
    return code in {'scv', 'swe'} or short in {'scv', 'swe'}


def _get_departments_filtered(*, teaching_only: bool | None) -> List[Department]:
    qs = Department.objects.all()
    if teaching_only is True:
        qs = qs.filter(is_teaching=True)
    elif teaching_only is False:
        qs = qs.filter(is_teaching=False)
    return [d for d in qs if not _is_excluded_department(d)]


def _staff_holiday_departments_for_date(d: date, label: str) -> List[Department]:
    token = _normalize_simple_token(label)
    if token in {'sun', 'sunday'}:
        return _get_departments_filtered(teaching_only=None)

    if token in {'sat', 'saturday'}:
        teaching = _get_departments_filtered(teaching_only=True)
        if d.weekday() == 5 and _week_of_month(d) == 3:
            non_teaching = _get_departments_filtered(teaching_only=False)
            return teaching + non_teaching
        return teaching

    return _get_departments_filtered(teaching_only=None)


def _staff_holiday_note(calendar_id: str, name: str) -> str:
    return f"{ACADEMIC_CALENDAR_HOLIDAY_NOTE_PREFIX}{calendar_id}::{name}"


def _clear_staff_holidays_for_calendar(calendar_id: str) -> None:
    from staff_attendance.models import Holiday

    Holiday.objects.filter(notes__startswith=f"{ACADEMIC_CALENDAR_HOLIDAY_NOTE_PREFIX}{calendar_id}::").delete()


def _sync_staff_holidays_for_calendar(
    *,
    calendar_id: str,
    holiday_rows: List[tuple[date, str]],
    actor,
) -> None:
    from staff_attendance.models import Holiday

    _clear_staff_holidays_for_calendar(calendar_id)

    for d, name in holiday_rows:
        dept_list = _staff_holiday_departments_for_date(d, name)
        dept_ids = [x.id for x in dept_list]
        if not dept_ids:
            continue

        note = _staff_holiday_note(calendar_id, name)
        is_sunday = _normalize_simple_token(name) in {'sun', 'sunday'}
        holiday, _ = Holiday.objects.get_or_create(
            date=d,
            defaults={
                'name': name,
                'notes': note,
                'is_sunday': is_sunday,
                'is_removable': True,
                'created_by': actor,
            },
        )
        if holiday.notes != note or holiday.name != name or holiday.is_sunday != is_sunday:
            holiday.name = name
            holiday.notes = note
            holiday.is_sunday = is_sunday
            holiday.is_removable = True
            if not holiday.created_by_id:
                holiday.created_by = actor
            holiday.save(update_fields=['name', 'notes', 'is_sunday', 'is_removable', 'created_by'])

        holiday.departments.set(dept_ids)


# ─────────────────────────────────────────────────────────────
# Timetable sync helper
# ─────────────────────────────────────────────────────────────

# Values that should NOT create special timetable entries
_TIMETABLE_EXCLUDED_VALUES = frozenset([
    '', 'holiday', 'ese lab', 'ese theory', 'sat', 'saturday', 'sun', 'sunday',
])


def _is_timetable_event(val: Any) -> bool:
    """Return True if this calendar cell value should create a special timetable entry."""
    token = _normalize_simple_token(val)
    if not token:
        return False
    # purely numeric values (e.g. working-day counters) are not events
    if token.isdigit():
        return False
    return token not in _TIMETABLE_EXCLUDED_VALUES


def _sync_calendar_to_timetable(calendar: 'AcademicCalendar', actor) -> Dict[str, Any]:
    """
    Read all AcademicCalendarDay rows for *calendar* and create SpecialTimetable /
    SpecialTimetableEntry records for each day that carries a non-excluded event value
    in any of the year columns (II Year, III Year, IV Year, I Year).

    Year mapping (calendar → Batch.start_year):
      I Year  → academic_start_year - 0 (most junior cohort)
      II Year → academic_start_year - 1
      III Year → academic_start_year - 2
      IV Year → academic_start_year - 3

    Returns a summary dict with created/skipped counts.
    """
    summary = {'sections_updated': set(), 'entries_created': 0, 'entries_skipped': 0, 'errors': []}
    try:
        from timetable.models import SpecialTimetable, SpecialTimetableEntry, TimetableSlot, TimetableTemplate
        from academics.models import AcademicYear, Section, Batch, StaffProfile
        from django.db.models import Q as _Q

        # Determine the current academic start year
        active_ay = AcademicYear.objects.filter(is_active=True).order_by('-id').first()
        if not active_ay:
            active_ay = AcademicYear.objects.order_by('-id').first()
        if not active_ay:
            return summary

        try:
            acad_start_year = int(str(active_ay.name).split('-')[0])
        except Exception:
            return summary

        # batch start_year that corresponds to each academic year level:
        # I Year  = started (acad_start_year - 0)  → year_offset = 0
        # II Year = started (acad_start_year - 1)  → year_offset = 1
        # III Year= started (acad_start_year - 2)  → year_offset = 2
        # IV Year = started (acad_start_year - 3)  → year_offset = 3
        year_column_map = [
            ('ii_year',  1, acad_start_year - 1),   # (field, year_label, batch_start_year)
            ('iii_year', 2, acad_start_year - 2),
            ('iv_year',  3, acad_start_year - 3),
            ('i_year',   0, acad_start_year - 0),
        ]

        # Cache: batch_start_year -> list of Section PKs
        section_cache: Dict[int, list] = {}

        def _sections_for_year(batch_start: int) -> list:
            if batch_start in section_cache:
                return section_cache[batch_start]
            secs = list(
                Section.objects.select_related('batch', 'batch__course', 'semester')
                .filter(batch__start_year=batch_start, batch__is_active=True)
            )
            section_cache[batch_start] = secs
            return secs

        # Active timetable template slots cache
        active_template = TimetableTemplate.objects.filter(is_active=True).first()
        if not active_template:
            # Fallback: any template
            active_template = TimetableTemplate.objects.order_by('-id').first()
        if not active_template:
            return summary

        all_slots = list(TimetableSlot.objects.filter(
            template=active_template, is_break=False, is_lunch=False
        ).order_by('index'))
        if not all_slots:
            return summary

        # Get or create SpecialTimetable for a (section, event_name) pair
        st_cache: Dict[tuple, 'SpecialTimetable'] = {}

        def _get_or_create_special_timetable(sec, event_name: str) -> 'SpecialTimetable':
            key = (sec.pk, event_name)
            if key in st_cache:
                return st_cache[key]
            sp = getattr(actor, 'staff_profile', None)
            st, _ = SpecialTimetable.objects.get_or_create(
                section=sec,
                name=event_name,
                defaults={'created_by': sp, 'is_active': True},
            )
            if not st.is_active:
                st.is_active = True
                st.save(update_fields=['is_active'])
            st_cache[key] = st
            return st

        # Iterate all calendar days
        days = list(calendar.days.order_by('date'))
        for day in days:
            for field_name, _year_idx, batch_start in year_column_map:
                event_val = getattr(day, field_name, '') or ''
                sections = _sections_for_year(batch_start)
                if not sections:
                    continue

                if not _is_timetable_event(event_val):
                    # Value cleared/excluded: remove auto-assigned entries (staff=None) for this date & sections
                    for sec in sections:
                        try:
                            deleted_count, _ = SpecialTimetableEntry.objects.filter(
                                timetable__section=sec,
                                date=day.date,
                                staff__isnull=True
                            ).delete()
                            if deleted_count > 0:
                                summary['sections_updated'].add(sec.pk)
                        except Exception as exc:
                            summary['errors'].append(str(exc))
                    continue

                event_name = str(event_val).strip()
                for sec in sections:
                    try:
                        # First, remove existing auto-assigned entries for this date/section if they belong to a DIFFERENT event name.
                        SpecialTimetableEntry.objects.filter(
                            timetable__section=sec,
                            date=day.date,
                            staff__isnull=True
                        ).exclude(subject_text=event_name).delete()

                        st = _get_or_create_special_timetable(sec, event_name)
                        summary['sections_updated'].add(sec.pk)
                        for slot in all_slots:
                            _, created = SpecialTimetableEntry.objects.get_or_create(
                                timetable=st,
                                date=day.date,
                                period=slot,
                                defaults={
                                    'subject_text': event_name,
                                    'staff': None,   # no staff = advisor marks daily attendance
                                    'is_active': True,
                                },
                            )
                            if created:
                                summary['entries_created'] += 1
                            else:
                                summary['entries_skipped'] += 1
                    except Exception as exc:
                        summary['errors'].append(str(exc))
    except Exception as exc:
        summary['errors'].append(str(exc))

    summary['sections_updated'] = len(summary['sections_updated'])
    return summary


def _format_dmy(d: date) -> str:
    return f"{d.day}/{d.month}/{d.year}"


def _calendar_payload(cal: AcademicCalendar, days: Optional[List[AcademicCalendarDay]] = None) -> Dict[str, Any]:
    if days is None:
        days = list(cal.days.order_by('date'))
    return {
        'id': str(cal.id),
        'name': cal.name,
        'from_date': cal.from_date.isoformat(),
        'to_date': cal.to_date.isoformat(),
        'academic_year': cal.academic_year,
        'created_at': cal.created_at.isoformat() if cal.created_at else None,
        'updated_at': cal.updated_at.isoformat() if cal.updated_at else None,
        'dates': [
            {
                'date': _format_dmy(d.date),
                'day': d.day_name,
                'workingDays': d.working_days or '',
                'counter': '',
                'iiYearEvent': d.ii_year or '',
                'iiYearCount': '',
                'iiiYearEvent': d.iii_year or '',
                'iiiYearCount': '',
                'ivYearEvent': d.iv_year or '',
                'ivYearCount': '',
                'iYearText': d.i_year or '',
            }
            for d in days
        ],
    }


def _parse_calendar_date(value: Any) -> Optional[date]:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(value).date()
        except Exception:
            return None
    if isinstance(value, str):
        s = value.strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
    return None


def _parse_calendar_upload(file_bytes: bytes) -> List[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]

    header_row = 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        key = _normalize_text(val)
        if key:
            headers[key] = col

    def col_for(name: str) -> Optional[int]:
        key = _normalize_text(name)
        return headers.get(key)

    date_col = col_for('date')
    work_col = col_for('working days')
    ii_col = col_for('ii year')
    iii_col = col_for('iii year')
    iv_col = col_for('iv year')
    i_col = col_for('i year')

    if not date_col:
        raise ValueError('Date column not found')

    rows: List[Dict[str, Any]] = []
    empty_hits = 0
    for r in range(2, ws.max_row + 1):
        dval = ws.cell(row=r, column=date_col).value
        d = _parse_calendar_date(dval)
        if not d:
            empty_hits += 1
            if empty_hits > 20:
                break
            continue
        empty_hits = 0

        day_name = _day_abbrev(d)
        work_val = ws.cell(row=r, column=work_col).value if work_col else ''
        rows.append({
            'date': d,
            'day': day_name,
            'working_days': str(work_val).strip() if work_val is not None else '',
            'ii_year': str(ws.cell(row=r, column=ii_col).value).strip() if ii_col else '',
            'iii_year': str(ws.cell(row=r, column=iii_col).value).strip() if iii_col else '',
            'iv_year': str(ws.cell(row=r, column=iv_col).value).strip() if iv_col else '',
            'i_year': str(ws.cell(row=r, column=i_col).value).strip() if i_col else '',
        })

    if not rows:
        raise ValueError('No calendar rows detected')
    return rows


def _build_calendar_and_days(*, name: str, from_date: date, to_date: date, rows: List[Dict[str, Any]], user) -> AcademicCalendar:
    cal = AcademicCalendar.objects.create(
        name=name,
        from_date=from_date,
        to_date=to_date,
        academic_year=_academic_year_from_dates(from_date, to_date),
        created_by=user,
    )

    days = [
        AcademicCalendarDay(
            calendar=cal,
            date=r['date'],
            day_name=r['day'],
            working_days=r.get('working_days') or '',
            ii_year=r.get('ii_year') or '',
            iii_year=r.get('iii_year') or '',
            iv_year=r.get('iv_year') or '',
            i_year=r.get('i_year') or '',
        )
        for r in rows
    ]
    AcademicCalendarDay.objects.bulk_create(days, batch_size=500)

    holiday_rows = []
    for r in rows:
        wd = r.get('working_days') or ''
        if _is_working_days_holiday(wd):
            holiday_rows.append((r['date'], str(wd).strip()))

    holiday_models = [
        AcademicCalendarHoliday(calendar=cal, date=d, name=name, source='working_days')
        for d, name in holiday_rows
    ]
    if holiday_models:
        AcademicCalendarHoliday.objects.bulk_create(holiday_models, batch_size=500)

    if holiday_rows:
        _sync_staff_holidays_for_calendar(
            calendar_id=str(cal.id),
            holiday_rows=holiday_rows,
            actor=user,
        )
    return cal


def _event_can_edit_delete(user, event: AcademicCalendarEvent) -> Dict[str, bool]:
    if not user or not user.is_authenticated:
        return {'can_edit': False, 'can_delete': False}

    if _get_student_profile(user) is not None:
        return {'can_edit': False, 'can_delete': False}

    if event.source == AcademicCalendarEvent.Source.IQAC:
        allowed = _is_iqac_user(user)
        return {'can_edit': bool(allowed), 'can_delete': bool(allowed)}

    if event.source == AcademicCalendarEvent.Source.HOD:
        if not _is_hod_user(user) and not _is_iqac_user(user):
            return {'can_edit': False, 'can_delete': False}

        # keep strict: IQAC cannot edit HOD events
        if _is_iqac_user(user):
            return {'can_edit': False, 'can_delete': False}

        if getattr(event, 'created_by_id', None) == getattr(user, 'id', None):
            return {'can_edit': True, 'can_delete': True}

        hod_dept = _get_primary_department_display(user)
        if event.audience_department and _department_matches_any(event.audience_department, [hod_dept]):
            return {'can_edit': True, 'can_delete': True}

    return {'can_edit': False, 'can_delete': False}


def _serialize_events(user, events: List[AcademicCalendarEvent]) -> List[Dict[str, Any]]:
    user_ids = [e.created_by_id for e in events if getattr(e, 'created_by_id', None)]
    colors = {c.hod_id: c.color for c in HodColor.objects.filter(hod_id__in=user_ids)}

    out: List[Dict[str, Any]] = []
    for e in events:
        perms = _event_can_edit_delete(user, e)
        out.append(
            {
                'id': str(e.id),
                'title': e.title,
                'description': e.description,
                'start_date': _to_yyyy_mm_dd(e.start_date),
                'end_date': _to_yyyy_mm_dd(e.end_date),
                'all_day': bool(e.all_day),
                'audience_department': e.audience_department,
                'year': e.year,
                'year_label': e.year_label,
                'source': e.source,
                'created_by': {'id': e.created_by_id, 'username': getattr(e.created_by, 'username', None)},
                'image_url': e.image_url,
                'audience_students': e.audience_students,
                'created_at': e.created_at.isoformat() if e.created_at else None,
                'updated_at': e.updated_at.isoformat() if e.updated_at else None,
                'creator_color': colors.get(e.created_by_id),
                # Branding poster
                'branding_poster_status':    e.branding_poster_status,
                'branding_poster_url':        e.branding_poster_url,
                'branding_poster_design_id':  e.branding_poster_design_id,
                'branding_poster_preview':    e.branding_poster_preview,
                **perms,
            }
        )
    return out


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def calendar_template_download(request):
    if not _is_iqac_user(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    def parse_date_param(key: str) -> Optional[date]:
        raw = request.query_params.get(key)
        if not raw:
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(str(raw).strip(), fmt).date()
            except Exception:
                continue
        return None

    start = parse_date_param('from_date')
    end = parse_date_param('to_date')
    if not start or not end:
        return Response({'error': 'from_date and to_date are required'}, status=status.HTTP_400_BAD_REQUEST)
    if end < start:
        return Response({'error': 'to_date must be after from_date'}, status=status.HTTP_400_BAD_REQUEST)

    odd_sat = str(request.query_params.get('odd_sat') or 'false').lower() in ('1', 'true', 'yes', 'on')
    even_sat = str(request.query_params.get('even_sat') or 'false').lower() in ('1', 'true', 'yes', 'on')

    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Academic Calendar'
    ws.append(TEMPLATE_COLUMNS)

    row_index = 2
    d = start
    while d <= end:
        ws.cell(row=row_index, column=1, value=d)
        ws.cell(row=row_index, column=1).number_format = 'DD/MM/YYYY'
        ws.cell(row=row_index, column=2, value=_day_abbrev(d))

        default_val = ''
        if d.weekday() == 6:
            default_val = 'Sun'
        elif d.weekday() == 5:
            if (odd_sat and _is_odd_saturday(d)) or (even_sat and _is_even_saturday(d)):
                default_val = 'Sat'

        for c in range(3, 8):
            ws.cell(row=row_index, column=c, value=default_val)

        row_index += 1
        d = d + timedelta(days=1)

    dv = DataValidation(
        type='list',
        formula1='"' + ','.join(TEMPLATE_DROPDOWN_VALUES) + '"',
        allow_blank=True,
        showErrorMessage=False,
    )
    ws.add_data_validation(dv)
    if row_index > 2:
        dv.add(f'C2:G{row_index - 1}')

    for idx, col in enumerate(TEMPLATE_COLUMNS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 18 if idx == 1 else 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"academic_calendar_template_{start.isoformat()}_{end.isoformat()}.xlsx"
    resp = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def calendars(request):
    user = request.user
    if request.method == 'GET':
        qs = AcademicCalendar.objects.order_by('-from_date')
        if str(request.query_params.get('current') or '').lower() in ('1', 'true', 'yes', 'on'):
            today = date.today()
            cal = qs.filter(from_date__lte=today, to_date__gte=today).first() or qs.first()
            if not cal:
                return Response({'calendar': None})
            return Response({'calendar': _calendar_payload(cal)})

        return Response(
            {
                'calendars': [
                    {
                        'id': str(c.id),
                        'name': c.name,
                        'from_date': c.from_date.isoformat(),
                        'to_date': c.to_date.isoformat(),
                        'academic_year': c.academic_year,
                        'created_at': c.created_at.isoformat() if c.created_at else None,
                        'updated_at': c.updated_at.isoformat() if c.updated_at else None,
                    }
                    for c in qs
                ]
            }
        )

    if not _is_iqac_user(user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    name = str(request.data.get('name') or '').strip() or None
    if not name:
        return Response({'error': 'name is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        from_date = date.fromisoformat(str(request.data.get('from_date')))
        to_date = date.fromisoformat(str(request.data.get('to_date')))
    except Exception:
        return Response({'error': 'from_date and to_date are required (YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)
    if to_date < from_date:
        return Response({'error': 'to_date must be after from_date'}, status=status.HTTP_400_BAD_REQUEST)

    uploaded = None
    if hasattr(request, 'FILES') and request.FILES:
        uploaded = request.FILES.get('file') or next(iter(request.FILES.values()))
    raw = uploaded.read() if uploaded else request.data.get('file')
    if not raw:
        return Response({'error': 'file is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = _parse_calendar_upload(raw)
    except Exception as exc:
        return Response({'error': f'Invalid file: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

    with transaction.atomic():
        cal = _build_calendar_and_days(name=name, from_date=from_date, to_date=to_date, rows=rows, user=user)

    # Sync calendar events → timetable (fire-and-forget)
    try:
        _sync_calendar_to_timetable(cal, user)
    except Exception:
        pass

    return Response({'calendar': _calendar_payload(cal)})


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def calendar_detail(request, calendar_id):
    try:
        cal = AcademicCalendar.objects.get(id=calendar_id)
    except AcademicCalendar.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        return Response({'calendar': _calendar_payload(cal)})

    if not _is_iqac_user(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'DELETE':
        _clear_staff_holidays_for_calendar(str(cal.id))
        cal.delete()
        return Response({'success': True})

    payload = request.data or {}
    days = payload.get('days') or []
    if not isinstance(days, list):
        return Response({'error': 'days must be a list'}, status=status.HTTP_400_BAD_REQUEST)

    with transaction.atomic():
        for row in days:
            d = _parse_calendar_date(row.get('date'))
            if not d:
                continue
            day = AcademicCalendarDay.objects.filter(calendar=cal, date=d).first()
            if not day:
                continue
            day.day_name = row.get('day') or _day_abbrev(d)
            day.working_days = str(row.get('workingDays') or row.get('working_days') or '').strip()
            day.ii_year = str(row.get('iiYearEvent') or row.get('ii_year') or '').strip()
            day.iii_year = str(row.get('iiiYearEvent') or row.get('iii_year') or '').strip()
            day.iv_year = str(row.get('ivYearEvent') or row.get('iv_year') or '').strip()
            day.i_year = str(row.get('iYearText') or row.get('i_year') or '').strip()
            day.save()

        AcademicCalendarHoliday.objects.filter(calendar=cal).delete()
        all_days = list(AcademicCalendarDay.objects.filter(calendar=cal).order_by('date'))
        holiday_rows = []
        for d in all_days:
            if _is_working_days_holiday(d.working_days):
                holiday_rows.append((d.date, d.working_days))
        if holiday_rows:
            AcademicCalendarHoliday.objects.bulk_create(
                [
                    AcademicCalendarHoliday(calendar=cal, date=dt, name=name, source='working_days')
                    for dt, name in holiday_rows
                ],
                batch_size=500,
            )
            _sync_staff_holidays_for_calendar(
                calendar_id=str(cal.id),
                holiday_rows=holiday_rows,
                actor=request.user,
            )
        else:
            _clear_staff_holidays_for_calendar(str(cal.id))

    # Sync calendar events → timetable (fire-and-forget)
    try:
        _sync_calendar_to_timetable(cal, request.user)
    except Exception:
        pass

    return Response({'calendar': _calendar_payload(cal)})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def config(request):
    mode = (request.query_params.get('mode') or '').strip().lower()
    user = request.user

    if mode == 'iqac':
        if not _is_iqac_user(user):
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        hod_owned_departments: List[str] = []
        student_year: Optional[int] = None
    elif mode == 'hod':
        if _is_iqac_user(user) or not _is_hod_user(user):
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        hod_owned_departments = _get_hod_owned_department_names(user)
        student_year = None
    elif mode == 'student':
        if _get_student_profile(user) is None:
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        hod_owned_departments = []
        student_year = _get_student_year(user)
    else:
        return Response({'error': 'Invalid mode'}, status=status.HTTP_400_BAD_REQUEST)

    departments = list(Department.objects.order_by('code').values_list('name', flat=True))

    return Response(
        {
            'mode': mode,
            'showing_department': _get_primary_department_display(user),
            'student_year': student_year,
            'student_year_roman': _roman_year(student_year),
            'departments': departments,
            'hod_owned_departments': hod_owned_departments,
        }
    )


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def events(request):
    user = request.user

    if request.method == 'GET':
        month_start = request.query_params.get('monthStart')
        month_end = request.query_params.get('monthEnd')
        mode = (request.query_params.get('mode') or '').strip().lower()

        try:
            ms = date.fromisoformat(str(month_start))
            me = date.fromisoformat(str(month_end))
        except Exception:
            return Response({'error': 'Invalid monthStart/monthEnd'}, status=status.HTTP_400_BAD_REQUEST)

        qs = AcademicCalendarEvent.objects.select_related('created_by').filter(start_date__lte=me, end_date__gte=ms)

        if mode == 'iqac':
            if not _is_iqac_user(user):
                return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
            events_list = list(qs)
        elif mode == 'student':
            if _get_student_profile(user) is None:
                return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
            year = _get_student_year(user)
            if year:
                qs = qs.filter(Q(year__isnull=True) | Q(year=year))
            else:
                qs = qs.filter(Q(year__isnull=True))
            events_list = list(qs)
        elif mode == 'hod':
            if _is_iqac_user(user) or not _is_hod_user(user):
                return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
            owned = _get_hod_owned_department_names(user)
            prefilter = Q(audience_department__isnull=True) | Q(audience_department__exact='')
            for n in owned:
                if n:
                    prefilter |= Q(audience_department__icontains=str(n))
            qs = qs.filter(prefilter)
            events_list = [e for e in qs if _department_matches_any(e.audience_department, owned)]
        else:
            return Response({'error': 'Invalid mode'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'events': _serialize_events(user, events_list)})

    if _get_student_profile(user) is not None:
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    mode = str(request.data.get('mode') or '').strip().lower()
    if mode == 'iqac':
        if not _is_iqac_user(user):
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        source = AcademicCalendarEvent.Source.IQAC
    elif mode == 'hod':
        if _is_iqac_user(user) or not _is_hod_user(user):
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        source = AcademicCalendarEvent.Source.HOD
    else:
        return Response({'error': 'Invalid mode'}, status=status.HTTP_400_BAD_REQUEST)

    title = str(request.data.get('title') or '').strip()
    if not title:
        return Response({'error': 'title is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start = date.fromisoformat(str(request.data.get('start_date')))
    except Exception:
        return Response({'error': 'start_date is required (YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        end_raw = request.data.get('end_date')
        end = date.fromisoformat(str(end_raw)) if end_raw else start
    except Exception:
        end = start

    start, end = _clamp_dates(start, end)

    description = str(request.data.get('description') or '').strip() or None
    all_day = str(request.data.get('all_day') or 'true').lower() in ('1', 'true', 'yes', 'on')
    year = _safe_int(request.data.get('year'))
    year_label = str(request.data.get('year_label') or '').strip() or None

    audience_departments: Any = request.data.get('audience_departments')
    if isinstance(audience_departments, str):
        try:
            parsed = json.loads(audience_departments)
            if isinstance(parsed, list):
                audience_departments = parsed
        except Exception:
            audience_departments = [x.strip() for x in audience_departments.split(',') if x.strip()]
    if audience_departments is None:
        audience_departments = []

    if source == AcademicCalendarEvent.Source.HOD:
        global_all = str(request.data.get('audience_all') or '').lower() in ('1', 'true', 'yes', 'on')
        if global_all:
            audience_department_str = None
        else:
            owned = _get_hod_owned_department_names(user)
            owned_norm = {_normalize_department_name(x) for x in owned if x}
            cleaned = []
            for d in (audience_departments or []):
                dn = _normalize_department_name(d)
                if dn and dn in owned_norm:
                    cleaned.append(d)
            if audience_departments and not cleaned:
                return Response({'error': 'Invalid audience_departments for HOD'}, status=status.HTTP_400_BAD_REQUEST)
            audience_department_str = ','.join(cleaned) if cleaned else None
    else:
        cleaned = [str(x).strip() for x in (audience_departments or []) if str(x).strip()]
        audience_department_str = ','.join(cleaned) if cleaned else None

    image_url = str(request.data.get('image_url') or '').strip() or None
    upload = request.FILES.get('image') if hasattr(request, 'FILES') else None
    if upload:
        ext = os.path.splitext(upload.name or '')[1]
        key = f"academic_calendar/{uuid.uuid4().hex}{ext}".strip('/')
        saved_path = default_storage.save(key, upload)
        try:
            image_url = default_storage.url(saved_path)
        except Exception:
            image_url = f"{settings.MEDIA_URL}{saved_path}"

    audience_students: Any = request.data.get('audience_students')
    if isinstance(audience_students, str):
        try:
            audience_students = json.loads(audience_students)
        except Exception:
            audience_students = None

    # Capture any branding-specific fields from the request payload
    branding_data = request.data.get('branding_data')
    if isinstance(branding_data, str):
        import json as _json
        try:
            branding_data = _json.loads(branding_data)
        except Exception:
            branding_data = None
    if not isinstance(branding_data, dict):
        branding_data = None

    event = AcademicCalendarEvent.objects.create(
        title=title,
        description=description,
        start_date=start,
        end_date=end,
        all_day=all_day,
        audience_department=audience_department_str,
        year=year,
        year_label=year_label,
        source=source,
        created_by=user,
        image_url=image_url,
        audience_students=audience_students,
        branding_data=branding_data,
    )

    # Trigger n8n → Canva autofill branding poster (async, non-blocking)
    fire_n8n_branding_async(event)

    return Response({'event': _serialize_events(user, [event])[0]})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def event_update(request, event_id):
    user = request.user
    try:
        event = AcademicCalendarEvent.objects.select_related('created_by').get(id=event_id)
    except AcademicCalendarEvent.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    perms = _event_can_edit_delete(user, event)
    if not perms.get('can_edit'):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    title = str(request.data.get('title') or event.title or '').strip()
    if not title:
        return Response({'error': 'title is required'}, status=status.HTTP_400_BAD_REQUEST)

    description = str(request.data.get('description') or '').strip() or None
    all_day = str(request.data.get('all_day') or event.all_day).lower() in ('1', 'true', 'yes', 'on')
    year = _safe_int(request.data.get('year'))
    year_label = str(request.data.get('year_label') or '').strip() or None

    try:
        start = date.fromisoformat(str(request.data.get('start_date') or event.start_date))
        end = date.fromisoformat(str(request.data.get('end_date') or event.end_date))
    except Exception:
        return Response({'error': 'Invalid date'}, status=status.HTTP_400_BAD_REQUEST)

    start, end = _clamp_dates(start, end)

    audience_departments: Any = request.data.get('audience_departments')
    if isinstance(audience_departments, str):
        try:
            parsed = json.loads(audience_departments)
            if isinstance(parsed, list):
                audience_departments = parsed
        except Exception:
            audience_departments = [x.strip() for x in audience_departments.split(',') if x.strip()]
    if audience_departments is None:
        audience_departments = []

    if event.source == AcademicCalendarEvent.Source.HOD:
        global_all = str(request.data.get('audience_all') or '').lower() in ('1', 'true', 'yes', 'on')
        if global_all:
            audience_department_str = None
        else:
            owned = _get_hod_owned_department_names(user)
            owned_norm = {_normalize_department_name(x) for x in owned if x}
            cleaned = []
            for d in (audience_departments or []):
                dn = _normalize_department_name(d)
                if dn and dn in owned_norm:
                    cleaned.append(d)
            audience_department_str = ','.join(cleaned) if cleaned else None
    else:
        cleaned = [str(x).strip() for x in (audience_departments or []) if str(x).strip()]
        audience_department_str = ','.join(cleaned) if cleaned else None

    image_url = str(request.data.get('image_url') or '').strip() or None
    upload = request.FILES.get('image') if hasattr(request, 'FILES') else None
    if upload:
        ext = os.path.splitext(upload.name or '')[1]
        key = f"academic_calendar/{uuid.uuid4().hex}{ext}".strip('/')
        saved_path = default_storage.save(key, upload)
        try:
            image_url = default_storage.url(saved_path)
        except Exception:
            image_url = f"{settings.MEDIA_URL}{saved_path}"
    if image_url is None:
        image_url = event.image_url

    audience_students: Any = request.data.get('audience_students')
    if isinstance(audience_students, str):
        try:
            audience_students = json.loads(audience_students)
        except Exception:
            audience_students = event.audience_students
    elif audience_students is None:
        audience_students = event.audience_students

    # Allow callers to update branding_data on edit
    branding_data_update = request.data.get('branding_data')
    if isinstance(branding_data_update, str):
        import json as _json
        try:
            branding_data_update = _json.loads(branding_data_update)
        except Exception:
            branding_data_update = None
    if isinstance(branding_data_update, dict):
        # Merge: existing fields are overridden by the incoming values
        merged = dict(event.branding_data or {})
        merged.update(branding_data_update)
        event.branding_data = merged

    event.title = title
    event.description = description
    event.start_date = start
    event.end_date = end
    event.all_day = all_day
    event.audience_department = audience_department_str
    event.year = year
    event.year_label = year_label
    event.image_url = image_url
    event.audience_students = audience_students
    event.save()

    # Re-trigger branding poster generation on every update
    fire_n8n_branding_async(event)

    return Response({'event': _serialize_events(user, [event])[0]})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def event_delete(request, event_id):
    user = request.user
    try:
        event = AcademicCalendarEvent.objects.get(id=event_id)
    except AcademicCalendarEvent.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    perms = _event_can_edit_delete(user, event)
    if not perms.get('can_delete'):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    event.delete()
    return Response({'success': True})


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def hod_colours(request):
    if not _is_iqac_user(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        hod_roles = (
            DepartmentRole.objects.filter(role='HOD', is_active=True)
            .select_related('staff__user', 'department')
            .order_by('department__code')
        )
        users = []
        seen = set()
        for r in hod_roles:
            u = getattr(getattr(r, 'staff', None), 'user', None)
            if not u or u.id in seen:
                continue
            seen.add(u.id)
            users.append({'hod_user_id': u.id, 'username': getattr(u, 'username', ''), 'department': getattr(getattr(r, 'department', None), 'name', None)})

        color_map = {c.hod_id: c.color for c in HodColor.objects.filter(hod_id__in=[x['hod_user_id'] for x in users])}
        for x in users:
            x['color'] = color_map.get(x['hod_user_id'])
        return Response({'hods': users})

    hod_user_id = _safe_int(request.data.get('hod_user_id'))
    color = str(request.data.get('color') or '').strip()
    if not hod_user_id or not color:
        return Response({'error': 'hod_user_id and color are required'}, status=status.HTTP_400_BAD_REQUEST)

    obj, _ = HodColor.objects.get_or_create(hod_id=hod_user_id, defaults={'color': color, 'updated_by': request.user})
    if obj.color != color or obj.updated_by_id != request.user.id:
        obj.color = color
        obj.updated_by = request.user
        obj.save(update_fields=['color', 'updated_by', 'updated_at'])

    return Response({'success': True})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def hod_students(request):
    if _is_iqac_user(request.user) or not _is_hod_user(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    dept_ids = get_user_effective_departments(request.user)
    if not dept_ids:
        return Response({'students': []})

    qs = (
        StudentProfile.objects.select_related('user', 'section__semester', 'section__batch__course__department')
        .filter(section__batch__course__department_id__in=dept_ids)
        .order_by('reg_no')
    )

    students: List[Dict[str, Any]] = []
    for sp in qs[:5000]:
        sec = getattr(sp, 'current_section', None) or getattr(sp, 'section', None)
        dept = None
        try:
            dept = sec and sec.batch and sec.batch.course and sec.batch.course.department
        except Exception:
            dept = None
        sem_number = getattr(getattr(sec, 'semester', None), 'number', None)
        year = _derive_year_from_semester_number(sem_number)
        students.append({'id': sp.id, 'reg_no': sp.reg_no, 'username': getattr(getattr(sp, 'user', None), 'username', None), 'year': year, 'section': getattr(sec, 'name', None), 'department': getattr(dept, 'name', None) if dept else None})

    return Response({'students': students})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_parse(request):
    if not _is_iqac_user(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    uploaded = None
    if hasattr(request, 'FILES') and request.FILES:
        uploaded = request.FILES.get('file') or next(iter(request.FILES.values()))
    raw = uploaded.read() if uploaded else request.body
    if not raw:
        return Response({'success': False, 'events': [], 'errors': ['Empty upload']}, status=status.HTTP_400_BAD_REQUEST)

    try:
        import openpyxl
        from openpyxl.utils.datetime import from_excel
    except Exception:
        return Response({'success': False, 'events': [], 'errors': ['openpyxl not installed']}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    errors: List[str] = []
    events: List[Dict[str, Any]] = []
    try:
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
        ws = wb.worksheets[0]
    except Exception as e:
        return Response({'success': False, 'events': [], 'errors': [f'Failed to read Excel: {e}']}, status=status.HTTP_400_BAD_REQUEST)

    rows = []
    max_rows = min(ws.max_row or 0, 2000)
    max_cols = min(ws.max_column or 0, 60)
    for r in range(1, max_rows + 1):
        row = []
        for c in range(1, max_cols + 1):
            row.append(ws.cell(row=r, column=c).value)
        rows.append(row)

    def header_contains_date(row_vals: List[Any]) -> Optional[int]:
        for idx, v in enumerate(row_vals):
            if v is None:
                continue
            if 'date' in str(v).strip().lower():
                return idx
        return None

    header_row_index = None
    date_col = None
    header_map: Dict[str, int] = {}

    for i in range(min(10, len(rows))):
        dc = header_contains_date(rows[i])
        if dc is not None:
            header_row_index = i
            date_col = dc
            for j, v in enumerate(rows[i]):
                if v is None:
                    continue
                key = _normalize_text(v)
                if key:
                    header_map[key] = j
            break

    if date_col is None:
        best_col = None
        best_score = 0
        scan_rows = rows[: min(30, len(rows))]
        for c in range(max_cols):
            score = 0
            for rv in scan_rows:
                v = rv[c] if c < len(rv) else None
                if v is None:
                    continue
                if isinstance(v, (datetime, date)):
                    score += 2
                elif isinstance(v, (int, float)) and 1 <= float(v) <= 60000:
                    score += 1
                elif isinstance(v, str) and any(ch.isdigit() for ch in v) and ('-' in v or '/' in v):
                    score += 1
            if score > best_score:
                best_score = score
                best_col = c
        date_col = best_col
        header_row_index = 0

    if date_col is None:
        return Response({'success': False, 'events': [], 'errors': ['Could not detect date column']}, status=status.HTTP_400_BAD_REQUEST)

    def parse_date_cell(v: Any) -> Optional[date]:
        if v is None or v == '':
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        if isinstance(v, (int, float)):
            try:
                return from_excel(v).date()
            except Exception:
                return None
        if isinstance(v, str):
            s = v.strip()
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    pass
        return None

    def is_numeric_only_title(title: str) -> bool:
        t = title.strip()
        return bool(t) and all(ch.isdigit() or ch in '.,/- ' for ch in t)

    def find_title_col() -> int:
        for k in ('event', 'placement', 'training'):
            for hk, idx in header_map.items():
                if k in hk:
                    return idx
        day_cols = {idx for hk, idx in header_map.items() if 'day' in hk}
        for c in range(max_cols):
            if c == date_col or c in day_cols:
                continue
            non_empty = 0
            for rv in rows[header_row_index + 1 : header_row_index + 40]:
                v = rv[c] if c < len(rv) else None
                if v is None:
                    continue
                if str(v).strip():
                    non_empty += 1
            if non_empty >= 2:
                return c
        return 0 if date_col != 0 else 1

    title_col = find_title_col()
    desc_col = None
    for k in ('description', 'details', 'remark', 'remarks'):
        for hk, idx in header_map.items():
            if k in hk:
                desc_col = idx
                break
        if desc_col is not None:
            break

    for r_idx in range(header_row_index + 1, len(rows)):
        rv = rows[r_idx]
        d = parse_date_cell(rv[date_col] if date_col < len(rv) else None)
        if not d:
            continue

        title_val = rv[title_col] if title_col < len(rv) else None
        title = str(title_val).strip() if title_val is not None else ''
        if not title:
            for c in range(max_cols):
                if c == date_col:
                    continue
                v = rv[c] if c < len(rv) else None
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    title = s
                    break

        if not title or is_numeric_only_title(title):
            continue

        description = None
        if desc_col is not None and desc_col < len(rv) and rv[desc_col] is not None:
            description = str(rv[desc_col]).strip() or None

        events.append({'title': title, 'description': description, 'start_date': d.isoformat(), 'end_date': d.isoformat(), 'all_day': True, 'audience_department': None, 'year': None, 'source': 'iqac', 'image_url': None, 'audience_students': None})

    return Response({'success': True, 'events': events, 'errors': errors})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_import(request):
    if not _is_iqac_user(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    payload = request.data if isinstance(request.data, dict) else {}
    items = payload.get('events') if isinstance(payload, dict) else []
    if not isinstance(items, list) or not items:
        return Response({'error': 'events list required'}, status=status.HTTP_400_BAD_REQUEST)

    to_create: List[AcademicCalendarEvent] = []
    for it in items:
        try:
            title = str(it.get('title') or '').strip()
            if not title:
                continue
            sd = date.fromisoformat(str(it.get('start_date')))
            ed = date.fromisoformat(str(it.get('end_date') or it.get('start_date')))
            sd, ed = _clamp_dates(sd, ed)
            all_day = bool(it.get('all_day', True))
            audience_department = it.get('audience_department')
            if audience_department == '':
                audience_department = None
            year = _safe_int(it.get('year'))
            year_label = str(it.get('year_label') or '').strip() or None
            image_url = str(it.get('image_url') or '').strip() or None
            audience_students = it.get('audience_students')
            if isinstance(audience_students, str):
                try:
                    audience_students = json.loads(audience_students)
                except Exception:
                    audience_students = None

            to_create.append(
                AcademicCalendarEvent(
                    title=title,
                    description=str(it.get('description') or '').strip() or None,
                    start_date=sd,
                    end_date=ed,
                    all_day=all_day,
                    audience_department=audience_department,
                    year=year,
                    year_label=year_label,
                    source=AcademicCalendarEvent.Source.IQAC,
                    created_by=request.user,
                    image_url=image_url,
                    audience_students=audience_students,
                )
            )
        except Exception:
            continue

    inserted = 0
    with transaction.atomic():
        chunk = 100
        for start in range(0, len(to_create), chunk):
            batch = to_create[start : start + chunk]
            AcademicCalendarEvent.objects.bulk_create(batch, batch_size=chunk)
            inserted += len(batch)

    return Response({'inserted': inserted})


# ── n8n branding-poster callback ──────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([])   # open endpoint – guarded by shared secret
def poster_callback(request, event_id):
    """
    POST /api/academic-calendar/events/<event_id>/poster-callback/

    Called by the n8n workflow after a Canva branding poster has been generated
    and exported.

    Expected JSON body:
      {
        "secret":      "<N8N_WEBHOOK_SECRET>",
        "poster_url":  "https://…/poster.png",
        "design_id":   "DAFxyz…",
        "preview_link": "https://www.canva.com/design/DAFxyz…/view"
      }

    The ``secret`` field must match the ``N8N_WEBHOOK_SECRET`` setting to prevent
    unauthorised writes to event records.
    """
    from django.conf import settings as _settings

    expected_secret = str(getattr(_settings, 'N8N_WEBHOOK_SECRET', '') or '').strip()
    if not expected_secret:
        return Response(
            {'error': 'N8N_WEBHOOK_SECRET is not configured on the server.'},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    body = request.data if isinstance(request.data, dict) else {}
    incoming_secret = str(body.get('secret', '')).strip()

    # Constant-time comparison to avoid timing side-channels
    import hmac as _hmac
    if not _hmac.compare_digest(incoming_secret, expected_secret):
        return Response({'error': 'Invalid secret.'}, status=status.HTTP_403_FORBIDDEN)

    try:
        event = AcademicCalendarEvent.objects.get(id=event_id)
    except (AcademicCalendarEvent.DoesNotExist, Exception):
        return Response({'error': 'Event not found.'}, status=status.HTTP_404_NOT_FOUND)

    poster_url   = str(body.get('poster_url', '')   or '').strip()
    design_id    = str(body.get('design_id', '')    or '').strip()
    preview_link = str(body.get('preview_link', '') or '').strip()

    if not poster_url and not design_id:
        # n8n signals a failure
        event.branding_poster_status = AcademicCalendarEvent.PosterStatus.FAILED
        event.save(update_fields=['branding_poster_status', 'updated_at'])
        return Response({'ok': False, 'detail': 'No poster URL received; marked as failed.'})

    event.branding_poster_url       = poster_url or event.branding_poster_url
    event.branding_poster_design_id = design_id  or event.branding_poster_design_id
    event.branding_poster_preview   = preview_link or event.branding_poster_preview
    event.branding_poster_status    = AcademicCalendarEvent.PosterStatus.READY
    event.save(update_fields=[
        'branding_poster_url',
        'branding_poster_design_id',
        'branding_poster_preview',
        'branding_poster_status',
        'updated_at',
    ])

    logger.info('Poster callback received for event %s — design_id=%s', event_id, design_id)
    return Response({'ok': True})


@api_view(['GET'])
@permission_classes([AllowAny])
def public_events(request):
    """Public, unauthenticated events feed for krgi.co.in.

    Returns AcademicCalendarEvent rows along with branding-generated poster URLs.
    """

    limit_raw = (request.query_params.get('limit') or '').strip()
    scope = (request.query_params.get('scope') or 'upcoming').strip().lower()
    try:
        limit = int(limit_raw) if limit_raw else 50
    except Exception:
        limit = 50
    limit = max(1, min(limit, 200))

    today = date.today()

    qs = AcademicCalendarEvent.objects.all()

    if scope == 'upcoming':
        qs = qs.filter(end_date__gte=today).order_by('start_date', 'title')
    elif scope == 'recent':
        qs = qs.filter(start_date__lte=today).order_by('-start_date', '-created_at')
    else:
        qs = qs.order_by('-start_date', '-created_at')

    events = list(qs[:limit])
    out = []
    for ev in events:
        # Prefer branding poster preview/url, else fall back to event image_url
        poster = (ev.branding_poster_preview or ev.branding_poster_url or ev.image_url or '').strip() or None
        out.append({
            'id': str(ev.id),
            'title': ev.title,
            'description': ev.description,
            'start_date': ev.start_date.isoformat(),
            'end_date': ev.end_date.isoformat() if ev.end_date else None,
            'all_day': bool(ev.all_day),
            'image': poster,
            'poster_status': ev.branding_poster_status,
        })

    return Response({'results': out}, status=200)


@api_view(['GET'])
@permission_classes([AllowAny])
def public_stats(request):
    """Public, unauthenticated stats for krgi.co.in dashboard."""

    # Students
    students_count = StudentProfile.objects.filter(status='ACTIVE').count()

    # Staff
    try:
        from academics.models import StaffProfile

        staff_count = StaffProfile.objects.filter(status='ACTIVE').count()
    except ImportError:
        staff_count = 0

    # Departments
    departments_count = Department.objects.count()

    # Courses
    courses_count = Course.objects.count()

    # Events
    today = date.today()
    active_events = AcademicCalendarEvent.objects.filter(
        Q(end_date__gte=today) | Q(end_date__isnull=True, start_date__gte=today)
    ).count()
    total_events = AcademicCalendarEvent.objects.count()

    return Response(
        {
            'students': students_count,
            'staff': staff_count,
            'departments': departments_count,
            'courses': courses_count,
            'active_events': active_events,
            'total_events': total_events,
        },
        status=200,
    )


# ── Calendar Event Labels ─────────────────────────────────────────────────────

from .models import CalendarEventLabel, CalendarEventAssignment


def _label_to_dict(label: CalendarEventLabel) -> dict:
    return {
        'id': str(label.id),
        'title': label.title,
        'color': label.color,
        'visible_roles': label.visible_roles or [],
        'semesters': label.semesters or [],
        'created_by': label.created_by_id,
        'created_at': label.created_at.isoformat(),
        'updated_at': label.updated_at.isoformat(),
    }


def _assignment_to_dict(a: CalendarEventAssignment) -> dict:
    return {
        'id': str(a.id),
        'event_id': str(a.event_id),
        'event_title': a.event.title,
        'event_color': a.event.color,
        'calendar_ref': a.calendar_ref,
        'start_date': a.start_date.isoformat(),
        'end_date': a.end_date.isoformat(),
        'description': a.description,
        'extra_data': a.extra_data,
        'created_by': a.created_by_id,
        'created_at': a.created_at.isoformat(),
        'updated_at': a.updated_at.isoformat(),
    }


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def event_labels(request):
    """List all event labels or create a new one."""
    if request.method == 'GET':
        labels = CalendarEventLabel.objects.all()
        return Response([_label_to_dict(l) for l in labels])

    # POST – create
    data = request.data
    title = str(data.get('title', '')).strip()
    if not title:
        return Response({'error': 'title is required'}, status=400)
    color = str(data.get('color', '#3B82F6')).strip()
    visible_roles = data.get('visible_roles', [])
    semesters = data.get('semesters', [])
    if not isinstance(visible_roles, list):
        visible_roles = []
    if not isinstance(semesters, list):
        semesters = []
    label = CalendarEventLabel.objects.create(
        title=title,
        color=color,
        visible_roles=visible_roles,
        semesters=semesters,
        created_by=request.user,
    )
    return Response(_label_to_dict(label), status=201)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def event_label_detail(request, label_id):
    try:
        label = CalendarEventLabel.objects.get(pk=label_id)
    except CalendarEventLabel.DoesNotExist:
        return Response({'error': 'Not found'}, status=404)

    if request.method == 'GET':
        return Response(_label_to_dict(label))

    if request.method == 'PUT':
        data = request.data
        if 'title' in data:
            label.title = str(data['title']).strip() or label.title
        if 'color' in data:
            label.color = str(data['color']).strip()
        if 'visible_roles' in data and isinstance(data['visible_roles'], list):
            label.visible_roles = data['visible_roles']
        if 'semesters' in data and isinstance(data['semesters'], list):
            label.semesters = data['semesters']
        label.save()
        return Response(_label_to_dict(label))

    if request.method == 'DELETE':
        label.delete()
        return Response(status=204)


# ── Calendar Event Assignments ────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def event_assignments(request):
    """List all assignments (optionally filter by calendar_ref) or create."""
    if request.method == 'GET':
        qs = CalendarEventAssignment.objects.select_related('event').all()
        calendar_ref = request.query_params.get('calendar_ref')
        if calendar_ref:
            qs = qs.filter(calendar_ref=calendar_ref)
        return Response([_assignment_to_dict(a) for a in qs])

    # POST – create
    data = request.data
    event_id = data.get('event_id')
    calendar_ref = str(data.get('calendar_ref', '')).strip()
    start_date = data.get('start_date')  # YYYY-MM-DD
    end_date = data.get('end_date')
    description = str(data.get('description', '')).strip()
    extra_data = data.get('extra_data') or None

    if not event_id or not calendar_ref or not start_date or not end_date:
        return Response({'error': 'event_id, calendar_ref, start_date, end_date are required'}, status=400)
    try:
        label = CalendarEventLabel.objects.get(pk=event_id)
    except CalendarEventLabel.DoesNotExist:
        return Response({'error': 'Event label not found'}, status=404)
    from datetime import date as _date
    try:
        from datetime import datetime as _dt
        s = _dt.strptime(start_date, '%Y-%m-%d').date()
        e = _dt.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        return Response({'error': 'Dates must be YYYY-MM-DD'}, status=400)
    assignment = CalendarEventAssignment.objects.create(
        event=label,
        calendar_ref=calendar_ref,
        start_date=s,
        end_date=e,
        description=description,
        extra_data=extra_data,
        created_by=request.user,
    )
    return Response(_assignment_to_dict(assignment), status=201)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def event_assignment_detail(request, assignment_id):
    try:
        a = CalendarEventAssignment.objects.select_related('event').get(pk=assignment_id)
    except CalendarEventAssignment.DoesNotExist:
        return Response({'error': 'Not found'}, status=404)

    if request.method == 'GET':
        return Response(_assignment_to_dict(a))

    if request.method == 'PUT':
        data = request.data
        if 'description' in data:
            a.description = str(data['description'])
        if 'extra_data' in data:
            a.extra_data = data['extra_data']
        if 'start_date' in data or 'end_date' in data:
            from datetime import datetime as _dt
            try:
                if 'start_date' in data:
                    a.start_date = _dt.strptime(data['start_date'], '%Y-%m-%d').date()
                if 'end_date' in data:
                    a.end_date = _dt.strptime(data['end_date'], '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Dates must be YYYY-MM-DD'}, status=400)
        a.save()
        return Response(_assignment_to_dict(a))

    if request.method == 'DELETE':
        a.delete()
        return Response(status=204)
