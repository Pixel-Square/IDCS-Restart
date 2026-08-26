import io
import re
import zipfile
from decimal import Decimal
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.db.models import Q
from academics.models import (
    TeachingAssignment,
    StudentSectionAssignment,
    StudentProfile,
    Department,
    Semester,
    AcademicYear,
)
from curriculum.models import ElectiveChoice, ElectiveSubject
from OBE.models import (
    AssessmentDraft,
    Ssa1Mark,
    Ssa2Mark,
    Cia1Mark,
    Cia2Mark,
    Formative1Mark,
    Formative2Mark,
    ModelExamMark,
    Review1Mark,
    Review2Mark,
    LabExamMark,
    FinalInternalMark,
    Cia1PublishedSheet,
    Cia2PublishedSheet,
    ModelPublishedSheet,
    LabPublishedSheet,
)


def _safe_text(value: Any) -> str:
    return str(value or '').strip()


def _safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None


def _safe_filename(value: str, max_len: int = 80) -> str:
    """Sanitize path component to strictly prevent directory traversal and illegal characters."""
    allowed = []
    for ch in str(value or ''):
        if ch.isalnum() or ch in ('_', '-', '.', ' '):
            allowed.append(ch)
        else:
            allowed.append('_')
    out = ''.join(allowed).strip().replace(' ', '_')
    # Prevent leading dots or empty
    out = re.sub(r'^\.+', '', out)
    if not out:
        out = 'unnamed'
    return out[:max_len]


def _parse_semester_value(val: Any) -> Optional[int]:
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if hasattr(val, 'number') and isinstance(getattr(val, 'number'), int):
        return getattr(val, 'number')
    digits = re.findall(r'\d+', str(val))
    if digits:
        return int(digits[0])
    return None


def _resolve_semester_number(ta: TeachingAssignment) -> Optional[int]:
    # 1. Primary: curriculum row semester (true course semester, handles string 'Sem 2' or object)
    cr = getattr(ta, 'curriculum_row', None)
    if cr:
        s = _parse_semester_value(getattr(cr, 'semester', None))
        if s is not None:
            return s
        if getattr(cr, 'master', None):
            s = _parse_semester_value(getattr(cr.master, 'semester', None))
            if s is not None:
                return s

    # 2. Primary: elective subject semester
    es = getattr(ta, 'elective_subject', None)
    if es:
        s = _parse_semester_value(getattr(es, 'semester', None))
        if s is not None:
            return s

    # 3. Fallback: section semester
    sec = getattr(ta, 'section', None)
    if sec:
        s = _parse_semester_value(getattr(sec, 'semester', None))
        if s is not None:
            return s
    return None


def _clean_dept_name(d: Any) -> str:
    if not d:
        return 'GENERAL'
    code = getattr(d, 'code', None) or ''
    name = getattr(d, 'name', None) or ''
    s_name = getattr(d, 'short_name', None) or ''
    if code.upper() in ('S&H', 'SH') or 'humanities' in name.lower():
        return 'S_H'
    raw = s_name or code or name or 'GENERAL'
    return _safe_text(raw).replace('&', '_')


def _resolve_department_info(ta: TeachingAssignment) -> Tuple[Optional[int], str]:
    sem = _resolve_semester_number(ta)
    sec = getattr(ta, 'section', None)
    cr = getattr(ta, 'curriculum_row', None)
    es = getattr(ta, 'elective_subject', None)
    
    code = (getattr(cr, 'course_code', '') or getattr(es, 'course_code', '') or getattr(getattr(ta, 'subject', None), 'code', '') or '').upper()
    name = (getattr(cr, 'course_name', '') or getattr(es, 'course_name', '') or getattr(getattr(ta, 'subject', None), 'name', '') or '').lower()

    # 1. First year / S&H check:
    # If section has managing_department (e.g. Science & Humanities), or course is GEA / S&H, or sem in (1, 2)
    managing_dept = getattr(sec, 'managing_department', None) if sec else None
    if managing_dept and (managing_dept.code in ('S&H', 'SH') or 'humanities' in managing_dept.name.lower()):
        return managing_dept.id, "S_H"
        
    if (sem in (1, 2) or code.startswith('GEA') or code.startswith('FLC') or 'physics' in name or 'chemistry' in name or 'tamil' in name) and managing_dept:
        return managing_dept.id, "S_H"

    if code.startswith('GEA') or code.startswith('FLC') or 'physics' in name or 'chemistry' in name or 'tamil' in name:
        sh_dept = Department.objects.filter(Q(code__in=['S&H', 'SH']) | Q(name__icontains='humanities')).first()
        if sh_dept:
            return sh_dept.id, "S_H"

    # 2. Check curriculum_row department
    if cr and cr.department:
        return cr.department.id, _clean_dept_name(cr.department)

    # 3. Check elective_subject department
    if es and es.department:
        return es.department.id, _clean_dept_name(es.department)

    # 4. Check section batch department
    batch_obj = getattr(sec, 'batch', None) if sec else None
    dept_obj = getattr(getattr(batch_obj, 'course', None), 'department', None) if batch_obj else None
    if dept_obj:
        return dept_obj.id, _clean_dept_name(dept_obj)

    return None, 'GENERAL'


def _resolve_course_info(ta: TeachingAssignment) -> Tuple[str, str]:
    code = ''
    name = ''
    subj = getattr(ta, 'subject', None)
    if subj is not None:
        code = _safe_text(getattr(subj, 'code', ''))
        name = _safe_text(getattr(subj, 'name', ''))
    if not code and getattr(ta, 'curriculum_row', None) is not None:
        row = ta.curriculum_row
        code = _safe_text(getattr(row, 'course_code', ''))
        name = _safe_text(getattr(row, 'course_name', ''))
    if not code and getattr(ta, 'elective_subject', None) is not None:
        es = ta.elective_subject
        code = _safe_text(getattr(es, 'course_code', ''))
        name = _safe_text(getattr(es, 'course_name', ''))
    if not name and getattr(ta, 'curriculum_row', None) is not None:
        name = _safe_text(getattr(ta.curriculum_row, 'course_name', ''))
    if not code and name:
        code = ''.join([w[0] for w in name.split() if w]).upper() or 'COURSE'
    return (code.upper(), name or code)


def _resolve_section_name(ta: TeachingAssignment) -> str:
    sec = getattr(ta, 'section', None)
    sec_name = getattr(sec, 'name', None)
    if sec_name:
        return _safe_text(sec_name)
    
    # Fallback to elective category label if any
    category = None
    if getattr(ta, 'elective_subject', None):
        parent = getattr(ta.elective_subject, 'parent', None)
        if parent and getattr(parent, 'category', None):
            category = str(parent.category).lower()
    elif getattr(ta, 'curriculum_row', None) and getattr(ta.curriculum_row, 'is_elective', False):
        if getattr(ta.curriculum_row, 'category', None):
            category = str(ta.curriculum_row.category).lower()

    if category:
        if 'open elective' in category or 'oe' in category.split():
            return 'OE'
        elif 'professional elective' in category or 'pe' in category.split():
            return 'PE'
        elif 'emerging' in category:
            return 'EE'

    return 'SEC-A'


def _get_students_for_ta(ta: TeachingAssignment) -> List[Dict[str, Any]]:
    students = []
    existing_ids: Set[int] = set()
    sec_id = getattr(ta, 'section_id', None)

    # 1. Section-based students
    if sec_id:
        s_qs = (
            StudentSectionAssignment.objects.filter(section_id=sec_id, end_date__isnull=True)
            .exclude(student__status__in=['INACTIVE', 'DEBAR'])
            .select_related('student__user')
        )
        for s in s_qs:
            sp = s.student
            if not sp or sp.id in existing_ids:
                continue
            u = getattr(sp, 'user', None)
            name = ' '.join([
                _safe_text(getattr(u, 'first_name', '')),
                _safe_text(getattr(u, 'last_name', '')),
            ]).strip() if u else ''
            if not name:
                name = _safe_text(getattr(u, 'username', '')) if u else ''
            
            sid = int(sp.id)
            existing_ids.add(sid)
            students.append({
                'id': sid,
                'reg_no': _safe_text(getattr(sp, 'reg_no', '')),
                'name': name,
            })

        legacy_qs = (
            StudentProfile.objects.filter(section_id=sec_id)
            .exclude(status__in=['INACTIVE', 'DEBAR'])
            .select_related('user')
        )
        for sp in legacy_qs:
            try:
                sid = int(sp.id)
            except Exception:
                continue
            if sid in existing_ids:
                continue
            u = getattr(sp, 'user', None)
            name = ' '.join([
                _safe_text(getattr(u, 'first_name', '')),
                _safe_text(getattr(u, 'last_name', '')),
            ]).strip() if u else ''
            if not name:
                name = _safe_text(getattr(u, 'username', '')) if u else ''
            existing_ids.add(sid)
            students.append({
                'id': sid,
                'reg_no': _safe_text(getattr(sp, 'reg_no', '')),
                'name': name,
            })

    # 2. ElectiveChoice-based students (for Elective Courses)
    es = getattr(ta, 'elective_subject', None)
    if es:
        try:
            ec_qs = ElectiveChoice.objects.filter(elective_subject=es).select_related('student__user')
            for ec in ec_qs:
                sp = ec.student
                if not sp or sp.id in existing_ids:
                    continue
                u = getattr(sp, 'user', None)
                name = ' '.join([
                    _safe_text(getattr(u, 'first_name', '')),
                    _safe_text(getattr(u, 'last_name', '')),
                ]).strip() if u else ''
                if not name:
                    name = _safe_text(getattr(u, 'username', '')) if u else ''
                sid = int(sp.id)
                existing_ids.add(sid)
                students.append({
                    'id': sid,
                    'reg_no': _safe_text(getattr(sp, 'reg_no', '')),
                    'name': name,
                })
        except Exception:
            pass

    # 3. Fallback: Only if students list is empty, extract from AssessmentDraft by teaching_assignment_id
    if not students:
        ta_id = getattr(ta, 'id', None)
        draft_student_ids: Set[int] = set()

        if ta_id:
            try:
                drafts = AssessmentDraft.objects.filter(teaching_assignment_id=ta_id)
                for d in drafts:
                    data = d.data if isinstance(d.data, dict) else {}
                    sheet = data.get('sheet', data) if isinstance(data, dict) else {}
                    
                    # rowsByStudentId
                    rows_by = sheet.get('rowsByStudentId') if isinstance(sheet, dict) else None
                    if isinstance(rows_by, dict):
                        for k in rows_by.keys():
                            k_clean = str(k).replace('id:', '').strip()
                            if k_clean.isdigit():
                                draft_student_ids.add(int(k_clean))
                    
                    # rows list
                    rows = sheet.get('rows') if isinstance(sheet, dict) else None
                    if isinstance(rows, list):
                        for r in rows:
                            if isinstance(r, dict) and r.get('studentId'):
                                try:
                                    draft_student_ids.add(int(r['studentId']))
                                except Exception:
                                    pass
                    
                    # theorySheet / tcplSheet
                    for sub_key in ('theorySheet', 'tcplSheet'):
                        sub_s = data.get(sub_key) or (sheet.get(sub_key) if isinstance(sheet, dict) else None)
                        if isinstance(sub_s, dict):
                            for k in sub_s.keys():
                                k_clean = str(k).replace('id:', '').strip()
                                if k_clean.isdigit():
                                    draft_student_ids.add(int(k_clean))

                new_ids = draft_student_ids - existing_ids
                if new_ids:
                    sp_qs = StudentProfile.objects.filter(id__in=new_ids).select_related('user')
                    for sp in sp_qs:
                        if sp.id in existing_ids:
                            continue
                        u = getattr(sp, 'user', None)
                        name = ' '.join([
                            _safe_text(getattr(u, 'first_name', '')),
                            _safe_text(getattr(u, 'last_name', '')),
                        ]).strip() if u else ''
                        if not name:
                            name = _safe_text(getattr(u, 'username', '')) if u else ''
                        sid = int(sp.id)
                        existing_ids.add(sid)
                        students.append({
                            'id': sid,
                            'reg_no': _safe_text(getattr(sp, 'reg_no', '')),
                            'name': name,
                        })
            except Exception:
                pass

    students.sort(key=lambda r: (_safe_text(r.get('reg_no')), _safe_text(r.get('name'))))
    return students


# Assessment metadata definitions
ASSESSMENT_SPECS = [
    {'key': 'ssa1', 'name': 'SSA1', 'label': 'Assignment 1 (SSA1)', 'model': Ssa1Mark, 'mark_field': 'mark'},
    {'key': 'cia1', 'name': 'CIA1', 'label': 'Continuous Internal Assessment 1 (CIA1)', 'model': Cia1Mark, 'mark_field': 'mark'},
    {'key': 'formative1', 'name': 'Formative1', 'label': 'Formative Assessment 1', 'model': Formative1Mark, 'mark_field': 'mark'},
    {'key': 'review1', 'name': 'Review1', 'label': 'Review 1 (PRBL / TCPR)', 'model': Review1Mark, 'mark_field': 'mark'},
    {'key': 'ssa2', 'name': 'SSA2', 'label': 'Assignment 2 (SSA2)', 'model': Ssa2Mark, 'mark_field': 'mark'},
    {'key': 'cia2', 'name': 'CIA2', 'label': 'Continuous Internal Assessment 2 (CIA2)', 'model': Cia2Mark, 'mark_field': 'mark'},
    {'key': 'formative2', 'name': 'Formative2', 'label': 'Formative Assessment 2', 'model': Formative2Mark, 'mark_field': 'mark'},
    {'key': 'review2', 'name': 'Review2', 'label': 'Review 2 (PRBL / TCPR)', 'model': Review2Mark, 'mark_field': 'mark'},
    {'key': 'model', 'name': 'Model Exam', 'label': 'Model Examination', 'model': ModelExamMark, 'mark_field': 'total_mark'},
    {'key': 'lab', 'name': 'Lab Exam', 'label': 'Laboratory Examination', 'model': LabExamMark, 'mark_field': 'total_mark'},
    {'key': 'final_internal', 'name': 'Final Internal Marks', 'label': 'Final Consolidated Internal Marks', 'model': FinalInternalMark, 'mark_field': 'final_mark'},
]


def _resolve_class_type(ta: TeachingAssignment) -> str:
    ct = ''
    if getattr(ta, 'elective_subject', None) and getattr(ta.elective_subject, 'class_type', None):
        ct = str(ta.elective_subject.class_type).upper().strip()
    elif getattr(ta, 'curriculum_row', None) and getattr(ta.curriculum_row, 'class_type', None):
        ct = str(ta.curriculum_row.class_type).upper().strip()
    return ct or 'THEORY'


def _get_applicable_assessments(class_type: str, draft_assessments: Set[str]) -> List[str]:
    ct = str(class_type or 'THEORY').upper().strip()
    
    if ct == 'TCPR':
        # TCPR strictly uses Review 1 & Review 2 (NO Formative 1 or Formative 2 under any circumstance)
        return ['ssa1', 'cia1', 'review1', 'ssa2', 'cia2', 'review2', 'model']
    elif ct in ('PROJECT', 'PRBL'):
        # Project / PRBL courses use Reviews and Project Model/Viva
        standard = ['review1', 'review2', 'model']
        if 'ssa1' in draft_assessments:
            standard.insert(0, 'ssa1')
        if 'ssa2' in draft_assessments:
            standard.insert(standard.index('review2'), 'ssa2')
        return standard
    elif ct in ('LAB', 'PURE_LAB', 'PURE LAB', 'PRACTICAL', 'LAB_2', 'LAB2'):
        standard = []
        if 'cia1' in draft_assessments or not draft_assessments:
            standard.append('cia1')
        if 'review1' in draft_assessments:
            standard.append('review1')
        if 'cia2' in draft_assessments or not draft_assessments:
            standard.append('cia2')
        if 'review2' in draft_assessments:
            standard.append('review2')
        if 'model' in draft_assessments or not draft_assessments:
            standard.append('model')
        if 'lab' in draft_assessments and 'lab' not in standard:
            standard.append('lab')
        return standard
    elif ct == 'TCPL':
        # TCPL has Theory + Lab
        standard = ['ssa1', 'cia1', 'formative1', 'ssa2', 'cia2', 'formative2', 'model']
        if 'review1' in draft_assessments:
            standard.append('review1')
        if 'review2' in draft_assessments:
            standard.append('review2')
        return standard
    else:
        # Default Theory (THEORY, THEORY_PMBL, ENGLISH, TAMIL, etc.)
        standard = ['ssa1', 'cia1', 'formative1', 'ssa2', 'cia2', 'formative2', 'model']
        for da in draft_assessments:
            if da not in standard and da in ('ssa1', 'cia1', 'formative1', 'review1', 'ssa2', 'cia2', 'formative2', 'review2', 'model', 'lab'):
                standard.append(da)
        return standard


def _extract_assessment_data_for_ta(
    ta: TeachingAssignment,
    asmt_key: str,
    students: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """Extract full question-level and mark data for an assessment and student roster."""
    student_ids = [s['id'] for s in students]
    subject_id = getattr(ta, 'subject_id', None)
    ta_id = getattr(ta, 'id', None)
    class_type = _resolve_class_type(ta)

    # 1. Fetch Draft Data (by TA id first, then subject id)
    draft_obj = None
    if ta_id:
        draft_obj = AssessmentDraft.objects.filter(
            teaching_assignment_id=ta_id,
            assessment=asmt_key,
        ).order_by('-updated_at').first()

    if not draft_obj and subject_id:
        draft_obj = AssessmentDraft.objects.filter(
            subject_id=subject_id,
            assessment=asmt_key,
        ).order_by('-updated_at').first()

    draft_data = draft_obj.data if draft_obj and isinstance(getattr(draft_obj, 'data', None), dict) else {}

    # Always also check published sheet data
    pub_data = {}
    pub = None
    if asmt_key == 'cia1':
        pub = Cia1PublishedSheet.objects.filter(
            Q(teaching_assignment_id=ta_id) | (Q(subject_id=subject_id) if subject_id else Q())
        ).order_by('-updated_at').first()
    elif asmt_key == 'cia2':
        pub = Cia2PublishedSheet.objects.filter(
            Q(teaching_assignment_id=ta_id) | (Q(subject_id=subject_id) if subject_id else Q())
        ).order_by('-updated_at').first()
    elif asmt_key == 'model':
        pub = ModelPublishedSheet.objects.filter(
            Q(teaching_assignment_id=ta_id) | (Q(subject_id=subject_id) if subject_id else Q())
        ).order_by('-updated_at').first()
    elif asmt_key in ('lab', 'review1', 'review2'):
        pub = LabPublishedSheet.objects.filter(
            Q(teaching_assignment_id=ta_id) | (Q(subject_id=subject_id) if subject_id else Q())
        ).order_by('-updated_at').first()
    if pub and isinstance(getattr(pub, 'data', None), dict):
        pub_data = pub.data

    sheet = draft_data.get('sheet', draft_data) if isinstance(draft_data, dict) else {}
    pub_sheet = pub_data.get('sheet', pub_data) if isinstance(pub_data, dict) else {}

    student_records: Dict[int, Dict[str, Any]] = {}
    for s in students:
        student_records[s['id']] = {
            'reg_no': s['reg_no'],
            'name': s['name'],
            'status': 'Present',
            'mark': None,
            'breakdown': {},
        }

    question_cols: List[str] = []

    # Helper to look up all student items across candidate containers
    def _find_all_student_items(containers: List[Any], sid: int, reg_no: str) -> List[Dict[str, Any]]:
        matches = []
        for c in containers:
            if not isinstance(c, dict):
                continue
            # Collect both 'id:1234' and '1234' and 1234 from container
            for key_cand in (f"id:{sid}", str(sid), sid):
                val = c.get(key_cand)
                if isinstance(val, dict) and val not in matches:
                    matches.append(val)
            if reg_no:
                for k, v in c.items():
                    if isinstance(v, dict):
                        if v.get('reg_no') == reg_no or v.get('registerNo') == reg_no or v.get('studentId') == sid or str(v.get('studentId')) == str(sid):
                            if v not in matches:
                                matches.append(v)
        return matches

    if asmt_key in ('cia1', 'cia2'):
        containers = []
        for cand in (
            draft_data.get('theorySheet'),
            draft_data.get('tcplSheet'),
            pub_data.get('theorySheet'),
            pub_data.get('tcplSheet'),
            sheet.get('theorySheet'),
            sheet.get('tcplSheet'),
            sheet.get('rowsByStudentId'),
            pub_data.get('rowsByStudentId'),
            draft_data.get('rowsByStudentId'),
            pub_sheet,
            sheet,
            pub_data,
            draft_data,
        ):
            if isinstance(cand, dict) and any('id:' in str(k) or str(k).isdigit() for k in cand.keys()):
                containers.append(cand)
            elif isinstance(cand, dict) and isinstance(cand.get('rowsByStudentId'), dict):
                containers.append(cand.get('rowsByStudentId'))

        for s in students:
            sid = s['id']
            rec = student_records[sid]
            items = _find_all_student_items(containers, sid, s['reg_no'])
            for item in items:
                if item.get('absent'):
                    rec['status'] = 'Absent'
                
                # Question map (Theory CIA)
                q_map = item.get('q') if isinstance(item.get('q'), dict) else {}
                for qk, qv in q_map.items():
                    col_name = str(qk).upper()
                    q_val = _safe_float(qv)
                    if q_val is not None:
                        rec['breakdown'][col_name] = q_val
                        if col_name not in question_cols:
                            question_cols.append(col_name)

                # Lab experiments (Lab CIA)
                marks_a = item.get('marksA') if isinstance(item.get('marksA'), list) else []
                for idx, mv in enumerate(marks_a, 1):
                    col_name = f"EXP_A{idx}"
                    val = _safe_float(mv)
                    if val is not None:
                        rec['breakdown'][col_name] = val
                        if col_name not in question_cols:
                            question_cols.append(col_name)

                marks_b = item.get('marksB') if isinstance(item.get('marksB'), list) else []
                for idx, mv in enumerate(marks_b, 1):
                    col_name = f"EXP_B{idx}"
                    val = _safe_float(mv)
                    if val is not None:
                        rec['breakdown'][col_name] = val
                        if col_name not in question_cols:
                            question_cols.append(col_name)

                cia_exam = _safe_float(item.get('ciaExam') or item.get('mark') or item.get('total'))
                if cia_exam is not None:
                    rec['mark'] = cia_exam
            
            if rec['mark'] is None and rec['breakdown']:
                num_vals = [v for v in rec['breakdown'].values() if v is not None]
                if num_vals:
                    rec['mark'] = round(sum(num_vals), 2)

    elif asmt_key in ('formative1', 'formative2'):
        containers = []
        for cand in (
            sheet.get('rowsByStudentId'),
            pub_data.get('rowsByStudentId'),
            draft_data.get('rowsByStudentId'),
            draft_data.get('theorySheet'),
            draft_data.get('tcplSheet'),
            sheet,
            pub_data,
            draft_data,
        ):
            if isinstance(cand, dict) and any('id:' in str(k) or str(k).isdigit() for k in cand.keys()):
                containers.append(cand)
            elif isinstance(cand, dict) and isinstance(cand.get('rowsByStudentId'), dict):
                containers.append(cand.get('rowsByStudentId'))

        for s in students:
            sid = s['id']
            rec = student_records[sid]
            items = _find_all_student_items(containers, sid, s['reg_no'])
            for item in items:
                if item.get('absent'):
                    rec['status'] = 'Absent'

                # 1. Lab Experiments by CO (for TCPL formatives)
                marks_by_co = item.get('marksByCo') if isinstance(item.get('marksByCo'), dict) else {}
                for co_num, exp_list in marks_by_co.items():
                    if isinstance(exp_list, list):
                        for idx, ev in enumerate(exp_list, 1):
                            col_name = f"CO{co_num}_EXP{idx}"
                            val = _safe_float(ev)
                            if val is not None:
                                rec['breakdown'][col_name] = val
                                if col_name not in question_cols:
                                    question_cols.append(col_name)

                # 2. Lab Experiments A & B
                marks_a = item.get('marksA') if isinstance(item.get('marksA'), list) else []
                if not marks_by_co:
                    for idx, mv in enumerate(marks_a, 1):
                        col_name = f"EXP_A{idx}"
                        val = _safe_float(mv)
                        if val is not None:
                            rec['breakdown'][col_name] = val
                            if col_name not in question_cols:
                                question_cols.append(col_name)

                marks_b = item.get('marksB') if isinstance(item.get('marksB'), list) else []
                if not marks_by_co:
                    for idx, mv in enumerate(marks_b, 1):
                        col_name = f"EXP_B{idx}"
                        val = _safe_float(mv)
                        if val is not None:
                            rec['breakdown'][col_name] = val
                            if col_name not in question_cols:
                                question_cols.append(col_name)

                # 3. Formative Exam Mark (e.g. 20M / 25M / 50M exam component)
                cia_exam = _safe_float(item.get('ciaExam'))
                if cia_exam is not None:
                    col_name = "EXAM"
                    rec['breakdown'][col_name] = cia_exam
                    if col_name not in question_cols:
                        question_cols.append(col_name)
                    if rec['mark'] is None:
                        rec['mark'] = cia_exam

                # 4. Theory Formative Activities (ATT1, SKILL1, etc.)
                for k, v in item.items():
                    if k in (
                        'studentId', 'absent', 'reg_no', 'registerNo', 'name',
                        'marksA', 'marksB', 'marksByCo', 'caaExamByCo', 'ciaExamByCo',
                        'reviewComponentMarks', 'absentKind', 'ciaExam', 'q', 'total', 'mark'
                    ):
                        continue
                    col_name = str(k).upper()
                    val = _safe_float(v)
                    if val is not None:
                        rec['breakdown'][col_name] = val
                        if col_name not in question_cols:
                            question_cols.append(col_name)

                # 5. Direct Total
                direct = _safe_float(item.get('total') or item.get('mark'))
                if direct is not None and rec['mark'] is None:
                    rec['mark'] = direct

            if rec['mark'] is None and rec['breakdown']:
                num_vals = [v for v in rec['breakdown'].values() if v is not None]
                if num_vals:
                    rec['mark'] = round(sum(num_vals), 2)

            if rec['mark'] is None:
                marks_map = draft_data.get('marks') if isinstance(draft_data, dict) and isinstance(draft_data.get('marks'), dict) else {}
                mv = marks_map.get(sid) or marks_map.get(str(sid))
                if mv is not None:
                    rec['mark'] = _safe_float(mv)

    elif asmt_key in ('ssa1', 'ssa2', 'review1', 'review2'):
        # 1. Check rows list
        rows = sheet.get('rows', []) if isinstance(sheet, dict) else []
        for r in rows:
            if not isinstance(r, dict):
                continue
            sid_raw = r.get('studentId')
            try:
                sid = int(sid_raw)
            except Exception:
                continue
            if sid not in student_records:
                continue
            rec = student_records[sid]
            for co_key in ('co1', 'co2', 'co3', 'co4', 'co5'):
                if co_key in r:
                    val = _safe_float(r.get(co_key))
                    if val is not None:
                        col_name = co_key.upper()
                        rec['breakdown'][col_name] = val
                        if col_name not in question_cols:
                            question_cols.append(col_name)
            
            qmarks = r.get('qMarks')
            if isinstance(qmarks, list) and qmarks:
                for idx, qv in enumerate(qmarks, 1):
                    val = _safe_float(qv)
                    if val is not None:
                        qname = f'Q{idx}'
                        rec['breakdown'][qname] = val
                        if qname not in question_cols:
                            question_cols.append(qname)

            rev_cos = r.get('reviewCoMarks')
            if isinstance(rev_cos, dict):
                for co_name, val_list in rev_cos.items():
                    if isinstance(val_list, list):
                        for idx, qv in enumerate(val_list, 1):
                            val = _safe_float(qv)
                            if val is not None:
                                lbl = f"{co_name.upper()}_P{idx}"
                                rec['breakdown'][lbl] = val
                                if lbl not in question_cols:
                                    question_cols.append(lbl)

            tot = _safe_float(r.get('total'))
            if tot is not None:
                rec['mark'] = tot
            elif rec['breakdown']:
                vals = [v for v in rec['breakdown'].values() if v is not None]
                if vals:
                    rec['mark'] = round(sum(vals), 2)

        # 2. Also check rowsByStudentId for review1/review2
        containers = []
        for cand in (sheet.get('rowsByStudentId'), pub_data.get('rowsByStudentId'), draft_data.get('rowsByStudentId'), sheet, pub_data, draft_data):
            if isinstance(cand, dict) and any('id:' in str(k) or str(k).isdigit() for k in cand.keys()):
                containers.append(cand)
            elif isinstance(cand, dict) and isinstance(cand.get('rowsByStudentId'), dict):
                containers.append(cand.get('rowsByStudentId'))

        for s in students:
            sid = s['id']
            rec = student_records[sid]
            if rec['mark'] is not None:
                continue
            items = _find_all_student_items(containers, sid, s['reg_no'])
            for item in items:
                if item.get('absent'):
                    rec['status'] = 'Absent'
                
                rev_marks = item.get('reviewComponentMarks') if isinstance(item.get('reviewComponentMarks'), dict) else {}
                for rk, rv in rev_marks.items():
                    val = _safe_float(rv)
                    if val is not None:
                        col_name = str(rk).upper()
                        rec['breakdown'][col_name] = val
                        if col_name not in question_cols:
                            question_cols.append(col_name)
                
                cia_exam = _safe_float(item.get('ciaExam'))
                tot = _safe_float(item.get('total') or item.get('mark'))
                if cia_exam is not None:
                    rec['mark'] = cia_exam
                elif tot is not None:
                    rec['mark'] = tot
                elif rev_marks:
                    tot_rev = sum(float(x) for x in rev_marks.values() if _safe_float(x) is not None)
                    rec['mark'] = round(tot_rev, 2)

    elif asmt_key == 'model':
        # Comprehensive model exam search across tcplSheet, theorySheet, rowsByStudentId
        containers = []
        candidate_sources = [
            draft_data.get('theorySheet'),
            draft_data.get('tcplSheet'),
            pub_data.get('theorySheet'),
            pub_data.get('tcplSheet'),
            sheet.get('theorySheet'),
            sheet.get('tcplSheet'),
            sheet.get('rowsByStudentId'),
            pub_data.get('rowsByStudentId'),
            draft_data.get('rowsByStudentId'),
            pub_sheet,
            sheet,
            pub_data,
            draft_data,
        ]

        # Also search sibling drafts for subject or curriculum row
        if subject_id:
            for sd in AssessmentDraft.objects.filter(subject_id=subject_id, assessment='model'):
                if isinstance(sd.data, dict):
                    candidate_sources.extend([
                        sd.data.get('theorySheet'),
                        sd.data.get('tcplSheet'),
                        sd.data.get('rowsByStudentId'),
                        sd.data,
                    ])

        for cand in candidate_sources:
            if isinstance(cand, dict) and any('id:' in str(k) or str(k).isdigit() for k in cand.keys()):
                if cand not in containers:
                    containers.append(cand)
            elif isinstance(cand, dict) and isinstance(cand.get('rowsByStudentId'), dict):
                r_by = cand.get('rowsByStudentId')
                if r_by not in containers:
                    containers.append(r_by)

        for s in students:
            sid = s['id']
            rec = student_records[sid]
            items = _find_all_student_items(containers, sid, s['reg_no'])
            for item in items:
                if item.get('absent'):
                    rec['status'] = 'Absent'
                
                # 1. Questions Q1 to Q16 (only store if numeric value exists)
                q_map = item.get('q') if isinstance(item.get('q'), dict) else {}
                for qk, qv in q_map.items():
                    q_val = _safe_float(qv)
                    if q_val is not None:
                        col_name = str(qk).upper()
                        rec['breakdown'][col_name] = q_val
                        if col_name not in question_cols:
                            question_cols.append(col_name)

                # 2. LAB / REVIEW component (TCPR / TCPL 30M / 50M exam)
                lab_val = _safe_float(item.get('lab') or item.get('review'))
                if lab_val is not None:
                    col_label = 'REVIEW_EXAM' if class_type == 'TCPR' else 'LAB_EXAM'
                    rec['breakdown'][col_label] = lab_val
                    if col_label not in question_cols:
                        question_cols.append(col_label)

                # 3. Record marks (CO5)
                rec_marks = item.get('recordMarksCo5') or item.get('recordMarks')
                if isinstance(rec_marks, list):
                    for idx, rv in enumerate(rec_marks, 1):
                        val = _safe_float(rv)
                        if val is not None:
                            col_name = f"RECORD_CO5_{idx}"
                            rec['breakdown'][col_name] = val
                            if col_name not in question_cols:
                                question_cols.append(col_name)

                # 4. Review component marks
                rev_comp = item.get('reviewComponentMarks') if isinstance(item.get('reviewComponentMarks'), dict) else {}
                for rk, rv in rev_comp.items():
                    val = _safe_float(rv)
                    if val is not None:
                        col_name = str(rk).upper()
                        rec['breakdown'][col_name] = val
                        if col_name not in question_cols:
                            question_cols.append(col_name)

                # 5. Total mark
                direct = _safe_float(item.get('ciaExam') or item.get('modelExam') or item.get('total') or item.get('mark'))
                if direct is not None:
                    rec['mark'] = direct

            # If total mark is still None, check published sheet coMarks for total
            if rec['mark'] is None:
                co_marks_list = pub_data.get('coMarks') or sheet.get('coMarks') or draft_data.get('coMarks')
                if isinstance(co_marks_list, list):
                    for cm_entry in co_marks_list:
                        if isinstance(cm_entry, dict) and (cm_entry.get('studentId') == sid or cm_entry.get('registerNo') == s['reg_no']):
                            rec['mark'] = _safe_float(cm_entry.get('total'))
                            break

            if rec['mark'] is None and rec['breakdown']:
                num_vals = [v for v in rec['breakdown'].values() if v is not None]
                if num_vals:
                    rec['mark'] = round(sum(num_vals), 2)

    elif asmt_key == 'lab':
        containers = []
        for cand in (
            sheet.get('rowsByStudentId'),
            pub_data.get('rowsByStudentId'),
            draft_data.get('rowsByStudentId'),
            draft_data.get('theorySheet'),
            draft_data.get('tcplSheet'),
            sheet,
            pub_data,
            draft_data,
        ):
            if isinstance(cand, dict) and any('id:' in str(k) or str(k).isdigit() for k in cand.keys()):
                containers.append(cand)
            elif isinstance(cand, dict) and isinstance(cand.get('rowsByStudentId'), dict):
                containers.append(cand.get('rowsByStudentId'))

        for s in students:
            sid = s['id']
            rec = student_records[sid]
            items = _find_all_student_items(containers, sid, s['reg_no'])
            for item in items:
                if item.get('absent'):
                    rec['status'] = 'Absent'
                direct = _safe_float(item.get('total') or item.get('mark') or item.get('ciaExam'))
                if direct is not None:
                    rec['mark'] = direct

    # 3. Fallback to Database Model Table if draft mark is missing
    spec = next((s for s in ASSESSMENT_SPECS if s['key'] == asmt_key), None)
    if spec and spec.get('model'):
        model_cls = spec['model']
        mark_f = spec.get('mark_field', 'mark')
        missing_ids = [sid for sid, rec in student_records.items() if rec['mark'] is None]
        if missing_ids and (subject_id or ta_id):
            try:
                qs = model_cls.objects.filter(student_id__in=missing_ids)
                field_names = {f.name for f in model_cls._meta.fields}
                if subject_id and ('subject' in field_names or 'subject_id' in field_names):
                    qs = qs.filter(subject_id=subject_id)
                if ta_id and ('teaching_assignment' in field_names or 'teaching_assignment_id' in field_names):
                    qs = qs.filter(Q(teaching_assignment_id=ta_id) | Q(teaching_assignment__isnull=True))
                
                if mark_f in field_names:
                    for row in qs.values('student_id', mark_f):
                        sid = row.get('student_id')
                        if sid in student_records and student_records[sid]['mark'] is None:
                            student_records[sid]['mark'] = _safe_float(row.get(mark_f))
            except Exception:
                pass

    # Only include question columns that have at least one numeric value across all student records
    question_cols = [
        col for col in question_cols
        if any(r['breakdown'].get(col) is not None for r in student_records.values())
    ]

    # Sort question columns naturally (Q1, Q2.. Q10, CO1, CO2, ATT1, SKILL1, LAB_EXAM, REVIEW_EXAM)
    def _col_sort_key(c: str):
        digits = re.findall(r'\d+', c)
        num = int(digits[0]) if digits else 0
        prefix = re.sub(r'\d+', '', c)
        return (prefix, num, c)

    question_cols.sort(key=_col_sort_key)

    return {
        'question_cols': question_cols,
        'records': student_records,
    }


def _build_final_internal_sheet_data(
    ta: TeachingAssignment,
    students: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """Compile summary components for Final Internal marks."""
    student_records: Dict[int, Dict[str, Any]] = {}
    for s in students:
        student_records[s['id']] = {
            'reg_no': s['reg_no'],
            'name': s['name'],
            'status': 'Present',
            'components': {},
            'final_internal': None,
        }

    # Fetch FinalInternalMark records
    subject_id = getattr(ta, 'subject_id', None)
    ta_id = getattr(ta, 'id', None)
    student_ids = [s['id'] for s in students]

    if subject_id and student_ids:
        try:
            fim_qs = FinalInternalMark.objects.filter(
                student_id__in=student_ids,
                subject_id=subject_id,
            )
            if ta_id:
                fim_qs = fim_qs.filter(Q(teaching_assignment_id=ta_id) | Q(teaching_assignment__isnull=True))
            
            for fim in fim_qs.select_related('student'):
                sid = fim.student_id
                if sid in student_records:
                    student_records[sid]['final_internal'] = _safe_float(fim.final_mark)
        except Exception:
            pass

    return student_records


def _create_styled_course_workbook(
    course_code: str,
    course_name: str,
    section_name: str,
    sem_no: Optional[int],
    dept_name: str,
    ta: TeachingAssignment,
    students: List[Dict[str, Any]],
) -> openpyxl.Workbook:
    """Generate an openpyxl Workbook with separate tabs for every exam tailored to course class_type."""
    wb = openpyxl.Workbook()
    default_sheet = wb.active

    class_type = _resolve_class_type(ta)

    # Styles
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    teal_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    light_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    font_title = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    font_subtitle = Font(name="Calibri", size=11, bold=True, color="374151")
    font_header_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_absent = Font(name="Calibri", size=11, bold=True, color="DC2626")

    thin_border_side = Side(border_style="thin", color="D1D5DB")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom_side = Side(border_style="medium", color="1E3A8A")
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Discover drafts present
    ta_id = getattr(ta, 'id', None)
    draft_asmts = set()
    if ta_id:
        draft_asmts = set(AssessmentDraft.objects.filter(teaching_assignment_id=ta_id).values_list('assessment', flat=True))

    applicable_keys = _get_applicable_assessments(class_type, draft_asmts)

    # Store extracted assessment totals for Final Internal sheet
    extracted_asmt_records: Dict[str, Dict[int, Dict[str, Any]]] = {}

    # 1. Generate Individual Assessment Sheets
    for spec in ASSESSMENT_SPECS:
        asmt_key = spec['key']
        asmt_name = spec['name']
        asmt_label = spec['label']

        if asmt_key == 'final_internal':
            continue

        if asmt_key not in applicable_keys:
            continue

        data = _extract_assessment_data_for_ta(ta, asmt_key, students)
        question_cols = data['question_cols']
        records = data['records']
        extracted_asmt_records[asmt_key] = records

        # For optional assessments not in standard list, skip if no data
        has_any_data = any(r['mark'] is not None or r['breakdown'] for r in records.values())
        if asmt_key not in ('ssa1', 'cia1', 'formative1', 'review1', 'ssa2', 'cia2', 'formative2', 'review2', 'model') and not has_any_data:
            continue

        # Clean sheet title
        sheet_title = asmt_name[:31]
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_view.showGridLines = True

        # Table Headers (Row 4)
        header_row = ['S.No', 'Register No', 'Student Name'] + question_cols + ['Total Mark', 'Status']
        max_col = max(len(header_row), 7)
        max_col_letter = get_column_letter(max_col)

        # Header Info (Rows 1-3)
        ws.merge_cells(f'A1:{max_col_letter}1')
        ws['A1'] = f"{course_code} - {course_name}"
        ws['A1'].font = font_title
        ws['A1'].alignment = align_left

        ws.merge_cells(f'A2:{max_col_letter}2')
        sem_label = f"Semester {sem_no}" if sem_no else "Semester N/A"
        ws['A2'] = f"Department: {dept_name}  |  {sem_label}  |  Section: {section_name}  |  Assessment: {asmt_label} ({class_type})"
        ws['A2'].font = font_subtitle
        ws['A2'].alignment = align_left

        ws.row_dimensions[4].height = 28
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20

        for c_idx, h_text in enumerate(header_row, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h_text)
            cell.font = font_header_white
            cell.fill = navy_fill
            cell.alignment = align_center if c_idx not in (2, 3) else align_left
            cell.border = header_border

        # Student Rows (Rows 5+)
        for s_idx, st in enumerate(students, start=1):
            curr_row = 4 + s_idx
            ws.row_dimensions[curr_row].height = 20
            rec = records.get(st['id'], {})
            
            row_fill = zebra_fill if s_idx % 2 == 0 else PatternFill(fill_type=None)
            
            # S.No
            c1 = ws.cell(row=curr_row, column=1, value=s_idx)
            c1.alignment = align_center

            # Register No
            c2 = ws.cell(row=curr_row, column=2, value=st['reg_no'])
            c2.alignment = align_left
            c2.font = font_regular

            # Student Name
            c3 = ws.cell(row=curr_row, column=3, value=st['name'])
            c3.alignment = align_left
            c3.font = font_regular

            col_pos = 4
            # Question columns
            for q_col in question_cols:
                q_val = rec.get('breakdown', {}).get(q_col)
                qc = ws.cell(row=curr_row, column=col_pos, value=q_val if q_val is not None else '')
                qc.alignment = align_center
                qc.font = font_regular
                col_pos += 1

            # Total Mark
            mark_val = rec.get('mark')
            cm = ws.cell(row=curr_row, column=col_pos, value=mark_val if mark_val is not None else '')
            cm.alignment = align_right
            cm.font = font_bold
            col_pos += 1

            # Status
            stat_val = rec.get('status', 'Present')
            cs = ws.cell(row=curr_row, column=col_pos, value=stat_val)
            cs.alignment = align_center
            cs.font = font_absent if stat_val == 'Absent' else font_regular

            # Apply borders and row fill
            for c_i in range(1, len(header_row) + 1):
                cell_item = ws.cell(row=curr_row, column=c_i)
                cell_item.border = cell_border
                if row_fill.fill_type:
                    cell_item.fill = row_fill

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        for i in range(4, len(header_row) + 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 14

    # 2. Add "Final Internal" Summary Sheet
    ws_final = wb.create_sheet(title="Final Internal Marks"[:31])
    ws_final.sheet_view.showGridLines = True

    # Build tailored Final Internal columns
    final_component_keys = [k for k in applicable_keys if k != 'final_internal']
    key_to_label = {
        'ssa1': 'SSA1',
        'cia1': 'CIA1' if class_type not in ('LAB', 'PURE_LAB', 'PURE LAB', 'PRACTICAL', 'LAB_2') else 'CIA1 LAB',
        'formative1': 'Formative1',
        'review1': 'Review1',
        'ssa2': 'SSA2',
        'cia2': 'CIA2' if class_type not in ('LAB', 'PURE_LAB', 'PURE LAB', 'PRACTICAL', 'LAB_2') else 'CIA2 LAB',
        'formative2': 'Formative2',
        'review2': 'Review2',
        'model': 'Model Exam' if class_type not in ('LAB', 'PURE_LAB', 'PURE LAB', 'PRACTICAL', 'LAB_2') else 'Model Lab',
        'lab': 'Lab Exam',
    }

    final_headers = ['S.No', 'Register No', 'Student Name'] + [key_to_label.get(k, k.upper()) for k in final_component_keys] + ['Final Internal', 'Status']
    max_final_col = max(len(final_headers), 7)
    max_final_letter = get_column_letter(max_final_col)

    ws_final.merge_cells(f'A1:{max_final_letter}1')
    ws_final['A1'] = f"{course_code} - {course_name}"
    ws_final['A1'].font = font_title
    ws_final['A1'].alignment = align_left

    ws_final.merge_cells(f'A2:{max_final_letter}2')
    sem_label = f"Semester {sem_no}" if sem_no else "Semester N/A"
    ws_final['A2'] = f"Department: {dept_name}  |  {sem_label}  |  Section: {section_name}  |  Overall Internal Mark Summary ({class_type})"
    ws_final['A2'].font = font_subtitle
    ws_final['A2'].alignment = align_left

    # Extract final internal marks
    final_records = _build_final_internal_sheet_data(ta, students)

    ws_final.row_dimensions[4].height = 28
    ws_final.row_dimensions[1].height = 24
    ws_final.row_dimensions[2].height = 20

    for c_idx, h_text in enumerate(final_headers, start=1):
        cell = ws_final.cell(row=4, column=c_idx, value=h_text)
        cell.font = font_header_white
        cell.fill = teal_fill
        cell.alignment = align_center if c_idx not in (2, 3) else align_left
        cell.border = header_border

    for s_idx, st in enumerate(students, start=1):
        curr_row = 4 + s_idx
        ws_final.row_dimensions[curr_row].height = 20
        sid = st['id']
        row_fill = zebra_fill if s_idx % 2 == 0 else PatternFill(fill_type=None)

        s_final = final_records.get(sid, {}).get('final_internal')

        row_values = [s_idx, st['reg_no'], st['name']]
        for ck in final_component_keys:
            c_recs = extracted_asmt_records.get(ck)
            if c_recs is None:
                c_recs = _extract_assessment_data_for_ta(ta, ck, students)['records']
                extracted_asmt_records[ck] = c_recs
            mark_v = c_recs.get(sid, {}).get('mark')
            row_values.append(mark_v if mark_v is not None else '')

        row_values.append(s_final if s_final is not None else '')
        row_values.append('Present')

        for c_i, val in enumerate(row_values, start=1):
            cell = ws_final.cell(row=curr_row, column=c_i, value=val)
            cell.border = cell_border
            if row_fill.fill_type:
                cell.fill = row_fill
            if c_i == 1:
                cell.alignment = align_center
            elif c_i in (2, 3):
                cell.alignment = align_left
                cell.font = font_regular
            elif c_i == len(row_values) - 1:
                cell.alignment = align_right
                cell.font = font_bold
                cell.fill = light_blue_fill
            elif c_i == len(row_values):
                cell.alignment = align_center
                cell.font = font_regular
            else:
                cell.alignment = align_right
                cell.font = font_regular

    # Column widths for final sheet
    ws_final.column_dimensions['A'].width = 8
    ws_final.column_dimensions['B'].width = 20
    ws_final.column_dimensions['C'].width = 30
    for i in range(4, len(final_headers) + 1):
        col_letter = get_column_letter(i)
        ws_final.column_dimensions[col_letter].width = 15

    # Safely remove the initial default sheet
    if default_sheet in wb.worksheets and len(wb.worksheets) > 1:
        wb.remove(default_sheet)

    return wb


def generate_semester_courses_marks_zip(
    semesters: List[int],
    academic_year: Optional[str] = None,
    department_id: Optional[int] = None,
    batch: Optional[str] = None,
    regulation: Optional[str] = None,
) -> Tuple[io.BytesIO, int]:
    """Generate structured ZIP archive containing Excel workbooks organized by Semesters & Departments.
    
    Structure:
      Semester {N}/
        {Department}/
          {CourseCode}_{CourseName}_{SectionName}.xlsx
            -> Sheets: SSA1, CIA1, Formative1, SSA2, CIA2, Formative2, Model Exam, Final Internal, ...
    """
    qs = TeachingAssignment.objects.filter(is_active=True).select_related(
        'subject',
        'curriculum_row',
        'curriculum_row__semester',
        'curriculum_row__batch',
        'curriculum_row__department',
        'elective_subject',
        'elective_subject__semester',
        'elective_subject__batch',
        'elective_subject__department',
        'section',
        'section__semester',
        'section__batch',
        'section__batch__regulation',
        'section__batch__course__department',
        'section__managing_department',
    )

    # Filter by semesters
    if semesters:
        sem_q = (
            Q(section__semester_id__in=semesters) |
            Q(section__semester__number__in=semesters) |
            Q(curriculum_row__semester_id__in=semesters) |
            Q(curriculum_row__semester__number__in=semesters) |
            Q(elective_subject__semester_id__in=semesters) |
            Q(elective_subject__semester__number__in=semesters) |
            Q(section__managing_department__code__in=['S&H', 'SH']) |
            Q(section__name__regex=r'^[A-L]$')
        )
        qs = qs.filter(sem_q)

    # Optional department filter
    if department_id:
        dept_q = Q(section__batch__course__department_id=department_id) | \
                 Q(section__managing_department_id=department_id) | \
                 Q(curriculum_row__department_id=department_id) | \
                 Q(elective_subject__department_id=department_id)
        qs = qs.filter(dept_q)

    # Optional batch filter
    if batch:
        qs = qs.filter(
            Q(section__batch__name__iexact=batch) |
            Q(curriculum_row__batch__name__iexact=batch) |
            Q(elective_subject__batch__name__iexact=batch)
        )

    # Optional regulation filter
    if regulation:
        qs = qs.filter(
            Q(section__batch__regulation__name__iexact=regulation) |
            Q(section__batch__regulation__code__iexact=regulation) |
            Q(curriculum_row__regulation__iexact=regulation) |
            Q(elective_subject__regulation__iexact=regulation)
        )

    tas = list(qs.order_by('section__batch__name', 'section__name', 'id'))

    zip_buffer = io.BytesIO()
    file_count = 0

    # Group by (Semester, Department, CourseCode, CourseName, SectionName)
    grouped: Dict[Tuple[int, str, str, str, str], List[TeachingAssignment]] = {}

    for ta in tas:
        sem_no = _resolve_semester_number(ta)
        if semesters and sem_no not in semesters:
            continue
        sem_no_clean = sem_no if sem_no is not None else 0

        _dept_id, dept_name = _resolve_department_info(ta)
        course_code, course_name = _resolve_course_info(ta)
        if not course_code:
            continue

        section_name = _resolve_section_name(ta)
        group_key = (sem_no_clean, dept_name, course_code, course_name, section_name)
        grouped.setdefault(group_key, []).append(ta)

    with zipfile.ZipFile(zip_buffer, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for group_key, group_tas in grouped.items():
            sem_no, dept_name, course_code, course_name, section_name = group_key
            primary_ta = group_tas[0]
            
            # Resolve students
            students = _get_students_for_ta(primary_ta)
            if not students:
                # If primary TA has no students, check if other TAs in group have
                for other_ta in group_tas[1:]:
                    students = _get_students_for_ta(other_ta)
                    if students:
                        primary_ta = other_ta
                        break

            # Create Workbook
            wb = _create_styled_course_workbook(
                course_code=course_code,
                course_name=course_name,
                section_name=section_name,
                sem_no=sem_no if sem_no > 0 else None,
                dept_name=dept_name,
                ta=primary_ta,
                students=students,
            )

            # Build ZIP directory path: Semester <N>/<Department>/<CourseCode>_<CourseName>_<SectionName>.xlsx
            sem_dir = f"Semester {sem_no}" if sem_no > 0 else "Semester_Unassigned"
            dept_dir = _safe_filename(dept_name, max_len=40)
            file_base = _safe_filename(f"{course_code}_{course_name}_{section_name}", max_len=100)
            zip_entry_path = f"{sem_dir}/{dept_dir}/{file_base}.xlsx"

            excel_stream = io.BytesIO()
            wb.save(excel_stream)
            excel_bytes = excel_stream.getvalue()

            zf.writestr(zip_entry_path, excel_bytes)
            file_count += 1

    zip_buffer.seek(0)
    return zip_buffer, file_count


def get_semester_export_preview(semesters: List[int]) -> Dict[str, Any]:
    """Provide summary statistics for the selected semesters."""
    qs = TeachingAssignment.objects.filter(is_active=True).select_related(
        'subject',
        'curriculum_row',
        'curriculum_row__semester',
        'elective_subject',
        'elective_subject__semester',
        'section',
        'section__semester',
        'section__batch__course__department',
        'section__managing_department',
    )

    if semesters:
        sem_q = (
            Q(section__semester_id__in=semesters) |
            Q(section__semester__number__in=semesters) |
            Q(curriculum_row__semester_id__in=semesters) |
            Q(curriculum_row__semester__number__in=semesters) |
            Q(elective_subject__semester_id__in=semesters) |
            Q(elective_subject__semester__number__in=semesters) |
            Q(section__managing_department__code__in=['S&H', 'SH']) |
            Q(section__name__regex=r'^[A-L]$')
        )
        qs = qs.filter(sem_q)

    tas = list(qs)
    courses_set: Set[str] = set()
    dept_set: Set[str] = set()
    section_set: Set[Tuple[str, str]] = set()

    for ta in tas:
        sem_no = _resolve_semester_number(ta)
        if semesters and sem_no not in semesters:
            continue
        code, _name = _resolve_course_info(ta)
        if code:
            courses_set.add(code)
        _did, dept_name = _resolve_department_info(ta)
        if dept_name:
            dept_set.add(dept_name)
        sec = _resolve_section_name(ta)
        if code and sec:
            section_set.add((code, sec))

    return {
        'semesters_selected': semesters,
        'total_teaching_assignments': len(tas),
        'total_unique_courses': len(courses_set),
        'total_course_sections': len(section_set),
        'total_departments': len(dept_set),
        'department_names': sorted(dept_set),
    }
