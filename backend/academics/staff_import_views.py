from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Q
from accounts.utils import get_user_permissions
from accounts.models import User, Role
from academics.models import Department, StaffProfile, STAFF_STATUS_CHOICES
import csv
import io

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


class StaffImportTemplateDownloadView(APIView):
    """Download an Excel template for importing staff with dropdown validations."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        
        # Check permission
        if not ('academics.add_staffprofile' in perms or user.is_staff or user.is_superuser):
            return Response({'error': 'You do not have permission to download staff import template'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        if not EXCEL_SUPPORT:
            return Response({'error': 'Excel support not available. Please install openpyxl.'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Get all departments
            departments = Department.objects.all().order_by('short_name')
            dept_names = [dept.short_name or dept.code for dept in departments if dept.short_name or dept.code]
            
            # Get all roles
            roles = Role.objects.all().order_by('name')
            role_names = [role.name for role in roles]
            
            # Get status choices
            status_values = [choice[0] for choice in STAFF_STATUS_CHOICES]
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Staff Import"
            
            # Headers
            headers = ['Staff ID', 'Department', 'Name', 'Designation', 'Email', 'Password', 'User Roles', 'Status']
            ws.append(headers)
            
            # Sample row
            sample_dept = dept_names[0] if dept_names else 'CSE'
            sample_role = role_names[0] if role_names else 'STAFF'
            ws.append(['STF001', sample_dept, 'John Doe', 'Assistant Professor', 'john.doe@example.com', 'ChangeMe#2026', sample_role, 'ACTIVE'])
            
            # Add data validation (dropdown) for department column (column B)
            if dept_names:
                dept_list = ','.join(dept_names)
                dv_dept = DataValidation(type="list", formula1=f'"{dept_list}"', allow_blank=False)
                dv_dept.error = 'Please select a department from the dropdown'
                dv_dept.errorTitle = 'Invalid Department'
                dv_dept.prompt = 'Select a department'
                dv_dept.promptTitle = 'Department'
                
                ws.add_data_validation(dv_dept)
                dv_dept.add('B2:B1000')
            
            # Add data validation (dropdown) for roles column (column G)
            if role_names:
                roles_list = ','.join(role_names)
                dv_roles = DataValidation(type="list", formula1=f'"{roles_list}"', allow_blank=True)
                dv_roles.error = 'Please select roles from the dropdown. Separate multiple roles with semicolon (;)'
                dv_roles.errorTitle = 'Invalid Role'
                dv_roles.prompt = 'Select one or more roles (separate with semicolon for multiple)'
                dv_roles.promptTitle = 'User Roles'
                
                ws.add_data_validation(dv_roles)
                dv_roles.add('G2:G1000')
            
            # Add data validation (dropdown) for status column (column H)
            if status_values:
                status_list = ','.join(status_values)
                dv_status = DataValidation(type="list", formula1=f'"{status_list}"', allow_blank=False)
                dv_status.error = 'Please select a status from the dropdown'
                dv_status.errorTitle = 'Invalid Status'
                dv_status.prompt = 'Select staff status'
                dv_status.promptTitle = 'Status'
                
                ws.add_data_validation(dv_status)
                dv_status.add('H2:H1000')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15  # Staff ID
            ws.column_dimensions['B'].width = 20  # Department
            ws.column_dimensions['C'].width = 25  # Name
            ws.column_dimensions['D'].width = 25  # Designation
            ws.column_dimensions['E'].width = 30  # Email
            ws.column_dimensions['F'].width = 20  # Password
            ws.column_dimensions['G'].width = 30  # User Roles
            ws.column_dimensions['H'].width = 12  # Status
            
            # Save to bytes
            from io import BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="staff_import_template.xlsx"'
            return response
            
        except Exception as e:
            return Response({'error': f'Failed to generate template: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class StaffBulkImportView(APIView):
    """Bulk import staff from CSV or Excel file."""
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        user = request.user
        perms = get_user_permissions(user)
        
        # Check permission
        if not ('academics.add_staffprofile' in perms or user.is_staff or user.is_superuser):
            return Response({'error': 'You do not have permission to import staff'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        filename = uploaded_file.name.lower()
        is_excel = filename.endswith(('.xlsx', '.xls'))
        is_csv = filename.endswith('.csv')
        
        if not (is_csv or is_excel):
            return Response({'error': 'File must be CSV or Excel (.xlsx, .xls, .csv)'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Parse file based on format
        rows = []
        if is_excel:
            if not EXCEL_SUPPORT:
                return Response({'error': 'Excel support not available. Please install openpyxl or use CSV format.'}, 
                              status=status.HTTP_400_BAD_REQUEST)
            try:
                wb = load_workbook(uploaded_file, read_only=True)
                ws = wb.active
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).strip().lower() if cell.value else '')
                
                # Read data rows
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    row_dict = {}
                    for idx, value in enumerate(row):
                        if idx < len(headers) and headers[idx]:
                            row_dict[headers[idx]] = str(value).strip() if value is not None else ''
                    if any(row_dict.values()):  # Skip empty rows
                        rows.append((row_idx, row_dict))
            except Exception as e:
                return Response({'error': f'Failed to parse Excel file: {str(e)}'}, 
                              status=status.HTTP_400_BAD_REQUEST)
        else:
            # Parse CSV
            try:
                decoded_file = uploaded_file.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                # Normalize header names to lowercase
                normalized_rows = []
                for idx, row in enumerate(reader, start=2):
                    normalized_row = {k.strip().lower(): v.strip() if v else '' for k, v in row.items()}
                    normalized_rows.append((idx, normalized_row))
                rows = normalized_rows
            except Exception as e:
                return Response({'error': f'Failed to parse CSV file: {str(e)}'}, 
                              status=status.HTTP_400_BAD_REQUEST)
        
        # Process rows
        created_count = 0
        updated_count = 0
        errors = []
        
        try:
            with transaction.atomic():
                for idx, row in rows:
                    try:
                        # Extract fields (handle different header variations)
                        staff_id = row.get('staff id', row.get('staffid', '')).strip()
                        dept_name = row.get('department', row.get('dept', '')).strip()
                        full_name = row.get('name', '').strip()
                        designation = row.get('designation', '').strip()
                        email = row.get('email', '').strip()
                        password = row.get('password', '').strip()
                        roles_str = row.get('user roles', row.get('roles', '')).strip()
                        status_value = row.get('status', 'ACTIVE').strip().upper()
                        
                        # Validate required fields
                        if not staff_id:
                            errors.append(f'Row {idx}: Staff ID is required')
                            continue
                        
                        if not dept_name:
                            errors.append(f'Row {idx}: Department is required')
                            continue
                        
                        if not full_name:
                            errors.append(f'Row {idx}: Name is required')
                            continue
                        
                        # Find or create user
                        # Split name into first and last name
                        name_parts = full_name.split(maxsplit=1)
                        first_name = name_parts[0] if name_parts else ''
                        last_name = name_parts[1] if len(name_parts) > 1 else ''
                        
                        # Try to find existing user by email or username (staff_id)
                        existing_user = None
                        if email:
                            existing_user = User.objects.filter(email=email).first()
                        if not existing_user:
                            existing_user = User.objects.filter(username=staff_id).first()
                        
                        if existing_user:
                            # Update existing user
                            user_obj = existing_user
                            user_obj.first_name = first_name
                            user_obj.last_name = last_name
                            if email:
                                user_obj.email = email
                            user_obj.save()
                        else:
                            # Create new user
                            if not password:
                                errors.append(f'Row {idx}: Password is required for new users')
                                continue
                            user_obj = User.objects.create(
                                username=staff_id,
                                first_name=first_name,
                                last_name=last_name,
                                email=email if email else '',
                            )
                            user_obj.set_password(password)
                            user_obj.save()
                        
                        # Find department
                        try:
                            department = Department.objects.get(Q(short_name__iexact=dept_name) | Q(code__iexact=dept_name))
                        except Department.DoesNotExist:
                            errors.append(f'Row {idx}: Department "{dept_name}" not found')
                            continue
                        except Department.MultipleObjectsReturned:
                            # Prefer short_name match
                            department = Department.objects.filter(short_name__iexact=dept_name).first()
                            if not department:
                                department = Department.objects.filter(code__iexact=dept_name).first()
                        
                        # Validate status
                        valid_statuses = [choice[0] for choice in STAFF_STATUS_CHOICES]
                        if status_value not in valid_statuses:
                            errors.append(f'Row {idx}: Invalid status "{status_value}". Must be one of: {", ".join(valid_statuses)}')
                            continue
                        
                        # Create or update staff profile
                        staff_profile, created = StaffProfile.objects.update_or_create(
                            staff_id=staff_id,
                            defaults={
                                'user': user_obj,
                                'department': department,
                                'designation': designation,
                                'status': status_value,
                            }
                        )
                        
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                        
                        # Handle roles
                        if roles_str:
                            role_names = [r.strip() for r in roles_str.split(';') if r.strip()]
                            for role_name in role_names:
                                try:
                                    role_obj = Role.objects.get(name__iexact=role_name)
                                    user_obj.roles.add(role_obj)
                                except Role.DoesNotExist:
                                    errors.append(f'Row {idx}: Role "{role_name}" not found')
                        
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
                        continue
            
            return Response({
                'message': f'Import completed. Created: {created_count}, Updated: {updated_count}',
                'created': created_count,
                'updated': updated_count,
                'errors': errors
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({'error': f'Import failed: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)
