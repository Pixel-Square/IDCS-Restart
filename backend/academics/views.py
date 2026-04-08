from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.db import models
from django.db import transaction
import logging
from django.http import Http404
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
import decimal
import io
import zipfile

from .permissions import IsHODOfDepartment


COE_DEPARTMENT_ACCESS_PERMS = {
    'coe.portal.access',
    'coe.manage.exams',
    'coe.manage.results',
    'coe.manage.circulars',
    'coe.manage.calendar',
}

from .models import (
    TeachingAssignment,
    SectionAdvisor,
    DepartmentRole,
    Section,
    StaffProfile,
    AcademicYear,
    StudentProfile,
    SpecialCourseAssessmentSelection,
    SpecialCourseAssessmentEditRequest,
    Semester,
)
from OBE.models import (
    Cia1Mark,
    Cia2Mark,
    Ssa1Mark,
    Ssa2Mark,
    Review1Mark,
    Review2Mark,
    Formative1Mark,
    Formative2Mark,
    InternalMarkMapping,
)
from .models import PeriodAttendanceSession, PeriodAttendanceRecord
from .models import AttendanceUnlockRequest

from .serializers import (
    SectionAdvisorSerializer,
    TeachingAssignmentSerializer,
    StudentSimpleSerializer,
)
from .serializers import AcademicYearSerializer
from .serializers import PeriodAttendanceSessionSerializer, BulkPeriodAttendanceSerializer, AttendanceUnlockRequestSerializer
from accounts.utils import get_user_permissions
from .utils import get_user_effective_departments, get_user_staff_profile
from .serializers import TeachingAssignmentInfoSerializer
from .serializers import SpecialCourseAssessmentEditRequestSerializer
from rest_framework import routers
from django.db.models import Q
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework import status
from rest_framework.response import Response
from .models import StudentMentorMap
from django.db import transaction


def _ensure_teaching_assignments_from_subject_batches(staff_profile) -> int:
    """Best-effort backfill: create TeachingAssignment rows from StudentSubjectBatch.

    This is used as a fallback when staff have subject batches (i.e., they clearly
    teach something) but their TeachingAssignment table is empty due to missing
    data provisioning.

    Returns the number of TeachingAssignment rows created or re-activated.
    """
    if not staff_profile or not getattr(staff_profile, 'pk', None):
        return 0

    try:
        from .models import StudentSubjectBatch, TeachingAssignment
    except Exception:
        return 0

    # ── Part 1: Curriculum-row based batches (non-elective) ──
    batches_qs = StudentSubjectBatch.objects.filter(
        staff=staff_profile,
        is_active=True,
        academic_year__is_active=True,
        curriculum_row__isnull=False,
    ).select_related('academic_year')

    created_or_updated = 0

    from academics.models import StudentProfile

    if batches_qs.exists():
        # Group by (academic_year, curriculum_row) and create one TeachingAssignment per pair.
        pairs = list(
            batches_qs.values_list('academic_year_id', 'curriculum_row_id').distinct()
        )

        for academic_year_id, curriculum_row_id in pairs:
            if not academic_year_id or not curriculum_row_id:
                continue

            # If any active TA already exists for this staff+course+year (any section), don't create another.
            existing = TeachingAssignment.objects.filter(
                staff=staff_profile,
                academic_year_id=academic_year_id,
                curriculum_row_id=curriculum_row_id,
            )
            if existing.filter(is_active=True).exists():
                continue

            # Infer a representative section if all students in the batches belong to exactly one section.
            section_ids = list(
                StudentProfile.objects.filter(
                    subject_batches__staff=staff_profile,
                    subject_batches__is_active=True,
                    subject_batches__academic_year_id=academic_year_id,
                    subject_batches__curriculum_row_id=curriculum_row_id,
                ).exclude(section_id__isnull=True).values_list('section_id', flat=True).distinct()
            )
            section_id = section_ids[0] if len(section_ids) == 1 else None

            # If an inactive TA exists, reactivate it (prefer one matching inferred section if possible).
            try:
                ta_to_reactivate = None
                if section_id is not None:
                    ta_to_reactivate = existing.filter(section_id=section_id).first()
                if ta_to_reactivate is None:
                    ta_to_reactivate = existing.first()

                if ta_to_reactivate is not None:
                    if not ta_to_reactivate.is_active:
                        ta_to_reactivate.is_active = True
                        ta_to_reactivate.save(update_fields=['is_active'])
                        created_or_updated += 1
                    continue
            except Exception:
                pass

            try:
                TeachingAssignment.objects.create(
                    staff=staff_profile,
                    academic_year_id=academic_year_id,
                    curriculum_row_id=curriculum_row_id,
                    section_id=section_id,
                    is_active=True,
                )
                created_or_updated += 1
            except Exception:
                # Ignore races / integrity errors; this is best-effort.
                continue

    # ── Part 2: Elective batches (no curriculum_row) ──
    # For batches where curriculum_row is NULL, resolve elective_subject
    # from the batch creator's teaching assignments using student overlap.
    elective_batches_qs = StudentSubjectBatch.objects.filter(
        staff=staff_profile,
        is_active=True,
        academic_year__is_active=True,
        curriculum_row__isnull=True,
    ).select_related('academic_year', 'created_by')

    for sb in elective_batches_qs:
        try:
            creator_id = getattr(sb, 'created_by_id', None)
            if not creator_id:
                continue

            academic_year_id = sb.academic_year_id
            if not academic_year_id:
                continue

            # Find elective TAs of the creator
            creator_etas = TeachingAssignment.objects.filter(
                staff_id=creator_id,
                elective_subject__isnull=False,
                is_active=True,
                academic_year_id=academic_year_id,
            ).select_related('elective_subject')

            if not creator_etas.exists():
                continue

            # Determine which elective subject by student overlap
            batch_student_ids = set(sb.students.values_list('id', flat=True))
            if not batch_student_ids:
                continue

            matched_es_id = None
            from curriculum.models import ElectiveChoice
            for eta in creator_etas:
                es_id = eta.elective_subject_id
                if not es_id:
                    continue
                choice_student_ids = set(
                    ElectiveChoice.objects.filter(
                        elective_subject_id=es_id, is_active=True
                    ).values_list('student_id', flat=True)
                )
                overlap = batch_student_ids & choice_student_ids
                if len(overlap) > 0:
                    matched_es_id = es_id
                    break

            # Fallback: if only one creator elective TA exists, use it
            if not matched_es_id and creator_etas.count() == 1:
                matched_es_id = creator_etas.first().elective_subject_id

            if not matched_es_id:
                continue

            # Check if an active TA already exists for this staff + elective_subject + year
            existing = TeachingAssignment.objects.filter(
                staff=staff_profile,
                academic_year_id=academic_year_id,
                elective_subject_id=matched_es_id,
            )
            if existing.filter(is_active=True).exists():
                continue

            # Try reactivating an inactive one
            try:
                inactive = existing.filter(is_active=False).first()
                if inactive:
                    inactive.is_active = True
                    inactive.save(update_fields=['is_active'])
                    created_or_updated += 1
                    continue
            except Exception:
                pass

            try:
                TeachingAssignment.objects.create(
                    staff=staff_profile,
                    academic_year_id=academic_year_id,
                    elective_subject_id=matched_es_id,
                    section_id=None,
                    is_active=True,
                )
                created_or_updated += 1
            except Exception:
                continue
        except Exception:
            continue

    return created_or_updated


# Attendance endpoints removed.


def serializer_check_user_can_manage(user, teaching_assignment):
    # reuse logic from serializers helper if available; basic check here
    staff_profile = getattr(user, 'staff_profile', None)
    if staff_profile and teaching_assignment.staff_id == staff_profile.pk:
        return True
    role_names = {r.name.upper() for r in user.roles.all()}
    if 'HOD' in role_names or 'ADVISOR' in role_names:
        # HOD membership is represented by DepartmentRole entries allowing
        # a staff to be HOD of multiple departments. Check active DepartmentRole
        # records rather than the single department on StaffProfile.
        try:
            if staff_profile:
                hod_depts = DepartmentRole.objects.filter(staff=staff_profile, role='HOD', is_active=True).values_list('department_id', flat=True)
                ta_dept = teaching_assignment.section.batch.course.department_id
                if ta_dept in list(hod_depts):
                    return True
        except Exception:
            pass
    return False


def _user_is_iqac_admin(user) -> bool:
    if user is None or not getattr(user, 'is_authenticated', False):
        return False
    if getattr(user, 'is_superuser', False) or getattr(user, 'is_staff', False):
        return True
    try:
        role_names = {r.name.upper() for r in user.roles.all()}
    except Exception:
        role_names = set()
    if 'IQAC' in role_names:
        return True
    try:
        perms = {str(p or '').lower() for p in (get_user_permissions(user) or [])}
    except Exception:
        perms = set()
    return 'obe.master.manage' in perms


class MyTeachingAssignmentsView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        base_qs = TeachingAssignment.objects.select_related(
            'subject',
            'curriculum_row',
            'curriculum_row__master',
            'curriculum_row__department',
            'curriculum_row__semester',
            'section',
            'academic_year',
            'section__semester',
            'section__batch__course__department',
        )

        qs = base_qs.filter(is_active=True)

        staff_profile = get_user_staff_profile(user)
        if not staff_profile:
            return Response([])

        # staff: only their teaching assignments (do not expand to department-level for HOD/ADVISOR here)
        # Prefer matching by user link; it's stable even if staff profile details change.
        qs_staff = qs.filter(staff__user=user)
        # Fallback: legacy / direct FK match.
        if not qs_staff.exists():
            qs_staff = qs.filter(staff=staff_profile)
        # Final fallback: if assignments exist but are not marked active, include active academic year.
        if not qs_staff.exists():
            qs_staff = base_qs.filter(staff__user=user, academic_year__is_active=True)

        # Backfill from StudentSubjectBatch — always run to pick up newly
        # assigned batches (e.g. elective batches assigned by another staff).
        try:
            _ensure_teaching_assignments_from_subject_batches(staff_profile)
        except Exception:
            pass
        if not qs_staff.exists():
            qs_staff = qs.filter(staff__user=user)
            if not qs_staff.exists():
                qs_staff = qs.filter(staff=staff_profile)

        # Final fallback: do not hide assignments solely due to flags.
        # If the staff has any TeachingAssignment rows at all, return them.
        if not qs_staff.exists():
            qs_staff = base_qs.filter(staff__user=user)
            if not qs_staff.exists():
                qs_staff = base_qs.filter(staff=staff_profile)
        qs = qs_staff

        ser = TeachingAssignmentInfoSerializer(qs.order_by('section__name', 'id'), many=True)
        return Response(ser.data)


class TeachingAssignmentStudentsView(APIView):
    """Return the student roster for a given TeachingAssignment (active students in the section).

    URL: /api/academics/teaching-assignments/<ta_id>/students/
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request, ta_id):
        try:
            ta = TeachingAssignment.objects.select_related(
                'section',
                'section__semester',
                'section__batch__course__department',
                'academic_year',
                'subject',
                'curriculum_row',
                'curriculum_row__department',
                'curriculum_row__semester',
            ).get(pk=ta_id, is_active=True)
        except TeachingAssignment.DoesNotExist:
            raise Http404('Teaching assignment not found')

        # basic permission: allow if user is staff owner, HOD/ADVISOR of the dept, OBE master/IQAC, or staff/admin
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        allowed = False
        if staff_profile and ta.staff_id == staff_profile.pk:
            allowed = True
        else:
            role_names = {r.name.upper() for r in user.roles.all()} if getattr(user, 'roles', None) is not None else set()
            if user.is_staff:
                allowed = True
            elif 'HOD' in role_names or 'ADVISOR' in role_names:
                allowed = True
            else:
                try:
                    perms = {str(p or '').lower() for p in (get_user_permissions(user) or [])}
                except Exception:
                    perms = set()
                if ('obe.master.manage' in perms) or ('IQAC' in role_names) or getattr(user, 'is_superuser', False):
                    allowed = True

        if not allowed:
            return Response({'detail': 'You do not have permission to view this roster.'}, status=403)

        # Prefer active StudentSectionAssignment entries for the section, falling back to StudentProfile.section
        from .models import StudentSectionAssignment, StudentProfile

        section_name = getattr(getattr(ta, 'section', None), 'name', None)

        # Best-effort subject metadata (supports elective_subject, curriculum_row and legacy subject FK)
        subject_code = None
        subject_name = None
        subject_id = getattr(ta, 'subject_id', None)
        class_type = None
        try:
            if (not subject_code or not subject_name) and getattr(ta, 'elective_subject', None):
                es = ta.elective_subject
                subject_code = subject_code or getattr(es, 'course_code', None)
                subject_name = subject_name or getattr(es, 'course_name', None)
                class_type = class_type or getattr(es, 'class_type', None)
            if getattr(ta, 'curriculum_row', None):
                cr = ta.curriculum_row
                subject_code = getattr(cr, 'course_code', None) or getattr(getattr(cr, 'master', None), 'course_code', None)
                subject_name = getattr(cr, 'course_name', None) or getattr(getattr(cr, 'master', None), 'course_name', None)
                class_type = class_type or getattr(cr, 'class_type', None) or getattr(getattr(cr, 'master', None), 'class_type', None)
            if (not subject_code or not subject_name) and getattr(ta, 'subject', None):
                subject_code = subject_code or getattr(ta.subject, 'code', None)
                subject_name = subject_name or getattr(ta.subject, 'name', None)
        except Exception:
            pass

        def _student_display_name(user):
            if not user:
                return None
            try:
                full = ' '.join([
                    str(getattr(user, 'first_name', '') or '').strip(),
                    str(getattr(user, 'last_name', '') or '').strip(),
                ]).strip()
                if full:
                    return full
            except Exception:
                pass
            return getattr(user, 'username', None)

        students = []

        # Elective TA rosters may not have a section; use elective-choices mapping.
        if not getattr(ta, 'section_id', None) and getattr(ta, 'elective_subject_id', None):
            try:
                from curriculum.models import ElectiveChoice

                eqs = (
                    ElectiveChoice.objects.filter(is_active=True, elective_subject_id=int(ta.elective_subject_id))
                    .exclude(student__isnull=True)
                    .select_related('student__user', 'student__section')
                )
                # Try with academic year first, fall back to without if no results
                if getattr(ta, 'academic_year_id', None):
                    eqs_ay = eqs.filter(academic_year_id=ta.academic_year_id)
                    if eqs_ay.exists():
                        eqs = eqs_ay
                for c in eqs:
                    sp = getattr(c, 'student', None)
                    if not sp:
                        continue
                    u = getattr(sp, 'user', None)
                    students.append({
                        'id': sp.id,
                        'reg_no': getattr(sp, 'reg_no', None),
                        'name': _student_display_name(u),
                        'section': str(getattr(sp, 'section', '')) if getattr(sp, 'section_id', None) else None,
                    })
            except Exception:
                students = []
        else:
            s_qs = StudentSectionAssignment.objects.filter(section=ta.section, end_date__isnull=True).exclude(student__status__in=['INACTIVE', 'DEBAR']).select_related('student__user')
            for a in s_qs:
                sp = a.student
                u = getattr(sp, 'user', None)
                students.append({
                    'id': sp.id,
                    'reg_no': getattr(sp, 'reg_no', None),
                    'name': _student_display_name(u),
                    'section': section_name,
                })

        # Also include legacy StudentProfile.section entries (section-based only).
        # IMPORTANT: do this even when StudentSectionAssignment already returned rows,
        # because some deployments still have students mapped only via the legacy
        # StudentProfile.section field (e.g., bulk imports that bypass signals).
        if getattr(ta, 'section', None) is not None:
            try:
                existing_ids = {int(r.get('id')) for r in students if isinstance(r, dict) and r.get('id') is not None}
            except Exception:
                existing_ids = set()

            sp_qs = StudentProfile.objects.filter(section=ta.section).exclude(status__in=['INACTIVE', 'DEBAR']).select_related('user')
            for sp in sp_qs:
                try:
                    sid = int(sp.id)
                except Exception:
                    continue
                if sid in existing_ids:
                    continue
                u = getattr(sp, 'user', None)
                students.append({
                    'id': sp.id,
                    'reg_no': getattr(sp, 'reg_no', None),
                    'name': _student_display_name(u),
                    'section': section_name,
                })

        # ── Batch-based student filtering ──
        # If the TA owner has StudentSubjectBatch entries for this subject/year,
        # restrict the student list to only those in the staff's batches.
        # This ensures batch-assigned staff see only their batch students.
        # Uses ta.staff_id (the TA owner) so HOD/IQAC views also see the correct filtered roster.
        try:
            from academics.models import StudentSubjectBatch as _SSB
            batch_filter_qs = _SSB.objects.filter(
                staff_id=ta.staff_id,
                is_active=True,
            )
            if getattr(ta, 'academic_year_id', None):
                batch_filter_qs = batch_filter_qs.filter(academic_year_id=ta.academic_year_id)
            # For elective TAs (no curriculum_row), match by creator's elective TA overlap
            # For regular TAs, match by curriculum_row
            if getattr(ta, 'curriculum_row_id', None):
                batch_filter_qs = batch_filter_qs.filter(curriculum_row_id=ta.curriculum_row_id)
            else:
                # Elective: match batches without curriculum_row
                batch_filter_qs = batch_filter_qs.filter(curriculum_row__isnull=True)
            user_batches = list(batch_filter_qs)
            if user_batches:
                batch_student_ids = set()
                for ub in user_batches:
                    try:
                        batch_student_ids.update(
                            ub.students.values_list('id', flat=True)
                        )
                    except Exception:
                        pass
                if batch_student_ids:
                    students = [s for s in students if s.get('id') in batch_student_ids]
        except Exception:
            pass

        def _resolve_dept(ta_obj):
            dept = None
            try:
                dept = ta_obj.section.batch.course.department
            except Exception:
                dept = None
            if not dept:
                try:
                    dept = getattr(getattr(ta_obj, 'elective_subject', None), 'department', None)
                except Exception:
                    dept = None
            if not dept:
                try:
                    dept = getattr(getattr(ta_obj, 'curriculum_row', None), 'department', None)
                except Exception:
                    dept = None
            if not dept:
                return None
            return {
                'id': getattr(dept, 'id', None),
                'code': getattr(dept, 'code', None),
                'name': getattr(dept, 'name', None),
                'short_name': getattr(dept, 'short_name', None),
            }

        def _resolve_semester(ta_obj):
            # Prefer section.semester, fall back to curriculum_row.semester or elective_subject.semester
            try:
                sem = getattr(getattr(ta_obj, 'section', None), 'semester', None)
                if sem:
                    return getattr(sem, 'number', None)
            except Exception:
                pass
            try:
                sem = getattr(getattr(ta_obj, 'curriculum_row', None), 'semester', None)
                if sem:
                    return getattr(sem, 'number', None)
            except Exception:
                pass
            try:
                sem = getattr(getattr(ta_obj, 'elective_subject', None), 'semester', None)
                if sem:
                    return getattr(sem, 'number', None)
            except Exception:
                pass
            return None

        return Response({
            'teaching_assignment': {
                'id': ta.id,
                'subject_id': subject_id,
                'subject_code': subject_code,
                'subject_name': subject_name,
                'class_type': class_type,
                'section_id': getattr(ta, 'section_id', None),
                'section_name': section_name,
                'academic_year': getattr(getattr(ta, 'academic_year', None), 'name', None),
                'semester': _resolve_semester(ta),
                'department': _resolve_dept(ta),
            },
            'students': students,
        })


class IqacInternalMarksBulkExportView(APIView):
    """Download filtered internal marks as ZIP of per-course Excel files.

    Query params:
      regulation: string (optional)
      semester: int (optional)
      department_id: int (optional)
      section_id: int (optional)
      batch: string (optional)
      academic_year: string/int (optional)
    """

    permission_classes = (IsAuthenticated,)

    @staticmethod
    def _safe_text(value):
        return str(value or '').strip()

    @staticmethod
    def _safe_float(value):
        if value is None:
            return None
        try:
            return float(value)
        except Exception:
            return None

    @staticmethod
    def _safe_filename(value: str) -> str:
        allowed = []
        for ch in str(value or ''):
            if ch.isalnum() or ch in ('_', '-', '.', ' '):
                allowed.append(ch)
            else:
                allowed.append('_')
        out = ''.join(allowed).strip().replace(' ', '_')
        return out or 'export'

    @staticmethod
    def _study_year_from_semester(semester_number):
        try:
            n = int(semester_number)
            if n <= 0:
                return None
            return str((n + 1) // 2)
        except Exception:
            return None

    def _resolve_semester_number_for_ta(self, ta):
        sec_sem = getattr(getattr(getattr(ta, 'section', None), 'semester', None), 'number', None)
        if sec_sem is not None:
            return sec_sem
        cur_sem = getattr(getattr(getattr(ta, 'curriculum_row', None), 'semester', None), 'number', None)
        if cur_sem is not None:
            return cur_sem
        ele_sem = getattr(getattr(getattr(ta, 'elective_subject', None), 'semester', None), 'number', None)
        return ele_sem

    def _resolve_course(self, ta):
        code = ''
        name = ''
        subj = getattr(ta, 'subject', None)
        if subj is not None:
            code = self._safe_text(getattr(subj, 'code', ''))
            name = self._safe_text(getattr(subj, 'name', ''))
        if not code and getattr(ta, 'curriculum_row', None) is not None:
            row = ta.curriculum_row
            code = self._safe_text(getattr(row, 'course_code', ''))
            name = self._safe_text(getattr(row, 'course_name', ''))
        if not code and getattr(ta, 'elective_subject', None) is not None:
            es = ta.elective_subject
            code = self._safe_text(getattr(es, 'course_code', ''))
            name = self._safe_text(getattr(es, 'course_name', ''))
        return (code.upper(), name or code)

    def _student_rows_for_ta(self, ta):
        from .models import StudentSectionAssignment, StudentProfile

        students = []
        existing_ids = set()
        if getattr(ta, 'section_id', None):
            s_qs = (
                StudentSectionAssignment.objects.filter(section_id=ta.section_id, end_date__isnull=True)
                .exclude(student__status__in=['INACTIVE', 'DEBAR'])
                .select_related('student__user')
            )
            for s in s_qs:
                sp = s.student
                u = getattr(sp, 'user', None)
                name = ' '.join([
                    self._safe_text(getattr(u, 'first_name', '')),
                    self._safe_text(getattr(u, 'last_name', '')),
                ]).strip() if u else ''
                if not name:
                    name = self._safe_text(getattr(u, 'username', '')) if u else ''
                students.append({
                    'id': sp.id,
                    'reg_no': self._safe_text(getattr(sp, 'reg_no', '')),
                    'name': name,
                })
                existing_ids.add(int(sp.id))

            legacy_qs = StudentProfile.objects.filter(section_id=ta.section_id).exclude(status__in=['INACTIVE', 'DEBAR']).select_related('user')
            for sp in legacy_qs:
                try:
                    sid = int(sp.id)
                except Exception:
                    continue
                if sid in existing_ids:
                    continue
                u = getattr(sp, 'user', None)
                name = ' '.join([
                    self._safe_text(getattr(u, 'first_name', '')),
                    self._safe_text(getattr(u, 'last_name', '')),
                ]).strip() if u else ''
                if not name:
                    name = self._safe_text(getattr(u, 'username', '')) if u else ''
                students.append({
                    'id': sp.id,
                    'reg_no': self._safe_text(getattr(sp, 'reg_no', '')),
                    'name': name,
                })

        students.sort(key=lambda r: (self._safe_text(r.get('reg_no')), self._safe_text(r.get('name'))))
        return students

    def _assessment_map(self, model, field_name: str, subject_id: int, student_ids, ta_id: int):
        from django.db.models import Q

        out = {}
        if not subject_id or not student_ids:
            return out

        base = model.objects.filter(subject_id=subject_id, student_id__in=student_ids)
        scoped = base.filter(Q(teaching_assignment_id=ta_id) | Q(teaching_assignment__isnull=True)).values('student_id', 'teaching_assignment_id', field_name)
        for row in scoped:
            sid = int(row.get('student_id'))
            current = out.get(sid)
            val = self._safe_float(row.get(field_name))
            is_ta = row.get('teaching_assignment_id') == ta_id
            if current is None:
                out[sid] = {'value': val, 'is_ta': is_ta}
                continue
            if current.get('is_ta'):
                continue
            if is_ta:
                out[sid] = {'value': val, 'is_ta': True}
            elif current.get('value') is None and val is not None:
                out[sid] = {'value': val, 'is_ta': False}

        missing = [sid for sid in student_ids if sid not in out]
        if missing:
            for row in base.filter(student_id__in=missing).values('student_id', field_name):
                sid = int(row.get('student_id'))
                if sid in out:
                    continue
                out[sid] = {'value': self._safe_float(row.get(field_name)), 'is_ta': False}

        return {sid: data.get('value') for sid, data in out.items()}

    def _extract_model_total_for_student(self, data, student_id):
        if not isinstance(data, dict):
            return None
        sid = str(student_id)

        marks = data.get('marks')
        if isinstance(marks, dict):
            qmarks = marks.get(sid) or marks.get(student_id)
            if isinstance(qmarks, dict):
                total = 0.0
                has_any = False
                for v in qmarks.values():
                    n = self._safe_float(v)
                    if n is not None:
                        total += n
                        has_any = True
                return round(total, 2) if has_any else None

        sheet = data.get('sheet') if isinstance(data, dict) else None
        if isinstance(sheet, dict):
            rows = sheet.get('rowsByStudentId')
            if isinstance(rows, dict):
                row = rows.get(sid) or rows.get(student_id)
                if isinstance(row, dict):
                    direct = self._safe_float(row.get('ciaExam'))
                    if direct is not None:
                        return round(direct, 2)

        return None

    def get(self, request):
        if not _user_is_iqac_admin(request.user):
            return Response({'detail': 'Only IQAC/OBE master can download this export.'}, status=403)

        from openpyxl import Workbook
        from OBE.models import FinalInternalMark
        from OBE.services.final_internal_marks import (
            _compute_weighted_final_total_theory_like,
            recompute_final_internal_marks,
        )

        regulation = self._safe_text(request.query_params.get('regulation'))
        semester = self._safe_text(request.query_params.get('semester'))
        department_id = self._safe_text(request.query_params.get('department_id'))
        section_id = self._safe_text(request.query_params.get('section_id'))
        batch = self._safe_text(request.query_params.get('batch'))
        academic_year = self._safe_text(request.query_params.get('academic_year'))
        ta_ids_raw = self._safe_text(request.query_params.get('ta_ids'))

        ta_ids = []
        if ta_ids_raw:
            for part in ta_ids_raw.split(','):
                part = self._safe_text(part)
                if not part:
                    continue
                try:
                    ta_ids.append(int(part))
                except Exception:
                    return Response({'detail': f'Invalid ta_ids value: {part}'}, status=400)

        ta_ids = sorted(set(ta_ids))

        qs = TeachingAssignment.objects.filter(is_active=True).select_related(
            'subject',
            'curriculum_row',
            'curriculum_row__semester',
            'curriculum_row__batch',
            'curriculum_row__department',
            'elective_subject',
            'elective_subject__semester',
            'elective_subject__batch',
            'elective_subject__department',
            'section',
            'section__semester',
            'section__batch',
            'section__batch__regulation',
            'section__batch__course__department',
        )

        if ta_ids:
            qs = qs.filter(id__in=ta_ids)

        if section_id:
            try:
                qs = qs.filter(section_id=int(section_id))
            except Exception:
                return Response({'detail': 'Invalid section_id'}, status=400)
        if semester:
            try:
                sem_no = int(semester)
                qs = qs.filter(
                    Q(section__semester__number=sem_no)
                    | Q(curriculum_row__semester__number=sem_no)
                    | Q(elective_subject__semester__number=sem_no)
                )
            except Exception:
                return Response({'detail': 'Invalid semester'}, status=400)
        if department_id:
            try:
                dept_no = int(department_id)
                qs = qs.filter(
                    Q(section__batch__course__department_id=dept_no)
                    | Q(curriculum_row__department_id=dept_no)
                    | Q(elective_subject__department_id=dept_no)
                )
            except Exception:
                return Response({'detail': 'Invalid department_id'}, status=400)
        if batch:
            qs = qs.filter(
                Q(section__batch__name__iexact=batch)
                | Q(curriculum_row__batch__name__iexact=batch)
                | Q(elective_subject__batch__name__iexact=batch)
            )
        if regulation:
            qs = qs.filter(
                Q(section__batch__regulation__name__iexact=regulation)
                | Q(section__batch__regulation__code__iexact=regulation)
                | Q(curriculum_row__regulation__iexact=regulation)
                | Q(elective_subject__regulation__iexact=regulation)
            )

        tas = list(qs.order_by('section__batch__name', 'section__name', 'id'))

        if academic_year:
            filtered = []
            for ta in tas:
                sem_num = self._resolve_semester_number_for_ta(ta)
                yr = self._study_year_from_semester(sem_num)
                if yr and yr == academic_year:
                    filtered.append(ta)
            tas = filtered

        if not tas:
            return Response({'detail': 'No teaching assignments found for selected filters.'}, status=404)

        zip_buffer = io.BytesIO()
        file_count = 0

        grouped = {}
        for ta in tas:
            code, course_name = self._resolve_course(ta)
            if not code:
                continue
            section_obj = getattr(ta, 'section', None)
            batch_obj = getattr(section_obj, 'batch', None) if section_obj is not None else None
            dept_obj = getattr(getattr(batch_obj, 'course', None), 'department', None) if batch_obj is not None else None
            dept_id = getattr(dept_obj, 'id', None)
            dept_name = self._safe_text(getattr(dept_obj, 'short_name', None) or getattr(dept_obj, 'code', None) or getattr(dept_obj, 'name', None) or 'N/A')
            sem_no = getattr(getattr(section_obj, 'semester', None), 'number', None)
            reg_obj = getattr(batch_obj, 'regulation', None) if batch_obj is not None else None
            reg_label = self._safe_text(getattr(reg_obj, 'code', None) or getattr(reg_obj, 'name', None) or '')
            batch_name = self._safe_text(getattr(batch_obj, 'name', None) or '')

            key = (code, course_name, sem_no, dept_id, dept_name, reg_label, batch_name)
            grouped.setdefault(key, []).append(ta)

        with zipfile.ZipFile(zip_buffer, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
            for key, course_tas in grouped.items():
                code, course_name, sem_no, _dept_id, dept_name, reg_label, batch_name = key

                subject = Subject.objects.filter(code__iexact=code).first()
                merged_students = {}
                ta_ids_for_course = [int(t.id) for t in course_tas if getattr(t, 'id', None)]
                if not ta_ids_for_course:
                    continue

                fim_qs = (
                    FinalInternalMark.objects.filter(teaching_assignment_id__in=ta_ids_for_course)
                    .select_related('student__user', 'teaching_assignment__section', 'subject')
                    .order_by('student_id', 'teaching_assignment_id')
                )

                for fim in fim_qs:
                    subj_code = self._safe_text(getattr(getattr(fim, 'subject', None), 'code', '')).upper()
                    if subj_code and subj_code != code:
                        continue

                    sp = getattr(fim, 'student', None)
                    if sp is None:
                        continue

                    sid = int(getattr(sp, 'id', 0) or 0)
                    if sid <= 0:
                        continue

                    user = getattr(sp, 'user', None)
                    name = ' '.join([
                        self._safe_text(getattr(user, 'first_name', '')),
                        self._safe_text(getattr(user, 'last_name', '')),
                    ]).strip() if user else ''
                    if not name:
                        name = self._safe_text(getattr(user, 'username', '')) if user else ''

                    ta_obj = getattr(fim, 'teaching_assignment', None)
                    section_name = self._safe_text(getattr(getattr(ta_obj, 'section', None), 'name', None) or '-')
                    total = self._safe_float(getattr(fim, 'final_mark', None))

                    prev = merged_students.get(sid)
                    if prev is None:
                        merged_students[sid] = {
                            'reg_no': self._safe_text(getattr(sp, 'reg_no', '')),
                            'name': name,
                            'section': section_name,
                            'total': total,
                        }
                    else:
                        if prev.get('total') is None and total is not None:
                            prev['total'] = total
                        if self._safe_text(prev.get('section')) in {'', '-'} and section_name not in {'', '-'}:
                            prev['section'] = section_name

                if not merged_students:
                    continue

                rows = sorted(merged_students.values(), key=lambda r: (self._safe_text(r.get('reg_no')), self._safe_text(r.get('name'))))

                wb = Workbook()
                ws = wb.active
                ws.title = 'Internal Marks'
                ws.append(['S.no', 'Register number', 'Student name', 'Section', 'Final Internal mark'])

                for idx, row in enumerate(rows, start=1):
                    ws.append([
                        idx,
                        self._safe_text(row.get('reg_no')),
                        self._safe_text(row.get('name')),
                        self._safe_text(row.get('section')),
                        row.get('total') if row.get('total') is not None else '-',
                    ])

                ws.auto_filter.ref = f"A1:E{ws.max_row}"
                ws.freeze_panes = 'A2'

                meta = []
                if reg_label:
                    meta.append(reg_label)
                if sem_no:
                    meta.append(f"SEM{sem_no}")
                if batch_name:
                    meta.append(batch_name)
                if dept_name:
                    meta.append(dept_name)
                base_name = f"{code}_{course_name}_{'_'.join(meta)}"
                filename = f"{self._safe_filename(base_name)[:140]}.xlsx"

                wb_buf = io.BytesIO()
                wb.save(wb_buf)
                wb_buf.seek(0)
                zf.writestr(filename, wb_buf.read())
                file_count += 1

        if file_count == 0:
            return Response({'detail': 'No exportable internal marks found for selected filters.'}, status=404)

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="internal_marks_export.zip"'
        return response


class IqacInternalMarksCourseExportView(APIView):
    """Download one course internal marks sheet for a teaching assignment.

    Query params:
      ta_id: int (required)
    """

    permission_classes = (IsAuthenticated,)

    @staticmethod
    def _safe_text(value):
        return str(value or '').strip()

    @staticmethod
    def _safe_float(value):
        if value is None:
            return None
        try:
            return float(value)
        except Exception:
            return None

    @staticmethod
    def _safe_filename(value: str) -> str:
        allowed = []
        for ch in str(value or ''):
            if ch.isalnum() or ch in ('_', '-', '.', ' '):
                allowed.append(ch)
            else:
                allowed.append('_')
        out = ''.join(allowed).strip()
        return out or 'internal_marks'

    def _resolve_course(self, ta):
        code = ''
        name = ''
        subj = getattr(ta, 'subject', None)
        if subj is not None:
            code = self._safe_text(getattr(subj, 'code', ''))
            name = self._safe_text(getattr(subj, 'name', ''))
        if not code and getattr(ta, 'curriculum_row', None) is not None:
            row = ta.curriculum_row
            code = self._safe_text(getattr(row, 'course_code', ''))
            name = self._safe_text(getattr(row, 'course_name', ''))
        if not code and getattr(ta, 'elective_subject', None) is not None:
            es = ta.elective_subject
            code = self._safe_text(getattr(es, 'course_code', ''))
            name = self._safe_text(getattr(es, 'course_name', ''))
        code = code.upper()
        return code, (name or code)

    # ──────────────────────────────────────────────────────────────────
    # CO-wise breakdown sheets: Cycle 1 / Cycle 2 / Model
    # ──────────────────────────────────────────────────────────────────
    def _add_co_breakdown_sheets(self, wb, ta, subject, student_list):
        """Append Cycle 1, Cycle 2 and Model Exam CO-breakdown sheets to *wb*.

        Parameters
        ----------
        wb : openpyxl.Workbook
        ta : TeachingAssignment
        subject : Subject (resolved)
        student_list : list[dict] – each dict has 'id', 'name', 'reg_no'; sorted
        """
        from django.db.models import Q
        from OBE.models import (
            Ssa1Mark,
            Ssa2Mark,
            Formative1Mark,
            Formative2Mark,
        )
        from OBE.services.final_internal_marks import (
            _resolve_class_type,
            _resolve_qp_type,
            _extract_ssa_co_splits_for_ta,
            _get_cia_sheet_data,
            _get_model_sheet_data,
            _extract_model_co_marks_for_student,
            _get_qp_pattern,
            _safe_float as _sf,
            _safe_text as _st,
            _parse_co12,
            _parse_co34,
            _parse_question_co_numbers,
            _qp1_final_question_weight,
            _co_weights_12,
            _co_weights_34,
            _clamp,
            _round2,
            _assessment_map,
        )

        if not student_list or subject is None:
            return

        ta_id = ta.id
        subject_id = subject.id
        student_ids = [int(s['id']) for s in student_list]
        reg_map = {int(s['id']): self._safe_text(s.get('reg_no', '')) for s in student_list}

        class_type = _resolve_class_type(ta)
        qp_type = _resolve_qp_type(ta)
        batch_id = getattr(getattr(ta, 'section', None), 'batch_id', None)
        is_qp1_final = 'QP1FINAL' in str(qp_type or '').upper().replace(' ', '')

        # ── helper: display value ────────────────────────────────────
        def _v(x):
            """Round to 2dp or dash."""
            return _round2(x) if x is not None else '-'

        # ═════════════════════════════════════════════════════════════
        # 1. Fetch all raw data (batch queries – one pass)
        # ═════════════════════════════════════════════════════════════

        # --- SSA ---
        ssa1_totals = _assessment_map(Ssa1Mark, 'mark', subject_id, student_ids, ta_id)
        ssa2_totals = _assessment_map(Ssa2Mark, 'mark', subject_id, student_ids, ta_id)
        ssa1_splits_all = _extract_ssa_co_splits_for_ta(subject_id, ta_id, 'ssa1', ['co1', 'co2'])
        ssa2_splits_all = _extract_ssa_co_splits_for_ta(subject_id, ta_id, 'ssa2', ['co3', 'co4'])

        # --- Formative (FA) ---
        def _fetch_formative_bulk(model_cls):
            result = {}
            qs = (
                model_cls.objects.filter(subject_id=subject_id, student_id__in=student_ids)
                .filter(Q(teaching_assignment_id=ta_id) | Q(teaching_assignment__isnull=True))
                .values('student_id', 'teaching_assignment_id', 'skill1', 'skill2', 'att1', 'att2', 'total')
            )
            for row in qs:
                sid = int(row['student_id'])
                is_ta = row.get('teaching_assignment_id') == ta_id
                existing = result.get(sid)
                if existing is None or (not existing.get('_is_ta') and is_ta):
                    result[sid] = {**row, '_is_ta': is_ta}
            return result

        f1_rows_all = _fetch_formative_bulk(Formative1Mark)
        f2_rows_all = _fetch_formative_bulk(Formative2Mark)

        # --- CIA ---
        cia1_sheet = _get_cia_sheet_data(subject_id, ta_id, 'cia1')
        cia2_sheet = _get_cia_sheet_data(subject_id, ta_id, 'cia2')

        cia1_pattern = _get_qp_pattern(class_type=class_type, qp_type=qp_type, exam='CIA1', batch_id=batch_id)
        cia2_pattern = _get_qp_pattern(class_type=class_type, qp_type=qp_type, exam='CIA2', batch_id=batch_id)

        def _build_questions(sheet, pattern, is_cia1):
            qs = sheet.get('questions') if isinstance(sheet.get('questions'), list) else []
            p_marks = pattern.get('marks') if isinstance(pattern, dict) and isinstance(pattern.get('marks'), list) else []
            p_cos = pattern.get('cos') if isinstance(pattern, dict) and isinstance(pattern.get('cos'), list) else []
            out = []
            count = max(len(qs), len(p_marks))
            for i in range(count):
                q = qs[i] if i < len(qs) and isinstance(qs[i], dict) else {}
                key = _st(q.get('key')) or f'q{i + 1}'
                mx = _sf(p_marks[i] if i < len(p_marks) else q.get('max'))
                if mx is None:
                    mx = _sf(q.get('maxMarks'))
                if mx is None:
                    mx = 0.0
                co_raw = p_cos[i] if i < len(p_cos) else q.get('co')
                if is_qp1_final:
                    co = co_raw
                else:
                    co = _parse_co12(co_raw) if is_cia1 else _parse_co34(co_raw)
                out.append({'key': key, 'max': float(mx), 'co': co})
            return out

        cia1_questions = _build_questions(cia1_sheet, cia1_pattern, True)
        cia2_questions = _build_questions(cia2_sheet, cia2_pattern, False)

        cia1_row_map = cia1_sheet.get('rowsByStudentId') if isinstance(cia1_sheet.get('rowsByStudentId'), dict) else {}
        cia2_row_map = cia2_sheet.get('rowsByStudentId') if isinstance(cia2_sheet.get('rowsByStudentId'), dict) else {}

        # CIA2 CO offset for QP1FINAL
        max_seen = 0
        for qq in cia2_questions:
            nums = _parse_question_co_numbers(qq.get('co'))
            if nums:
                max_seen = max(max_seen, max(nums))
        qp1_cia2_offset = 1 if (is_qp1_final and max_seen > 0 and max_seen <= 2) else 0

        def _cia_co_raw(row, questions, is_cia1):
            """Compute raw (unscaled) CO totals for a student from CIA question-level data."""
            if not isinstance(row, dict) or bool(row.get('absent')):
                return None, None, None
            qvals = row.get('q') if isinstance(row.get('q'), dict) else {}
            c_a = 0.0
            c_b = 0.0
            has_any = False
            for q in questions:
                mx = float(q.get('max') or 0)
                n = _sf(qvals.get(q.get('key')))
                if is_qp1_final and is_cia1:
                    raw_nums = _parse_question_co_numbers(q.get('co'))
                    raw_num = raw_nums[0] if raw_nums else None
                    wa = 1.0 if raw_num == 1 else 0.0
                    wb_weight = 1.0 if raw_num == 2 else 0.0
                elif is_qp1_final and not is_cia1:
                    wa = _qp1_final_question_weight(q.get('co'), 2, qp1_cia2_offset)
                    wb_weight = _qp1_final_question_weight(q.get('co'), 3, qp1_cia2_offset)
                elif is_cia1:
                    wa, wb_weight = _co_weights_12(q.get('co'))
                else:
                    wa, wb_weight = _co_weights_34(q.get('co'))
                if n is None:
                    continue
                has_any = True
                mark = _clamp(n, 0, mx)
                c_a += mark * wa
                c_b += mark * wb_weight
            if not has_any:
                return None, None, None
            return _round2(c_a), _round2(c_b), _round2(c_a + c_b)

        # --- MODEL ---
        model_sheet = _get_model_sheet_data(subject_id, ta_id, class_type)
        model_pattern = _get_qp_pattern(class_type=class_type, qp_type=qp_type, exam='MODEL', batch_id=batch_id)

        # ── Determine CO labels based on QP type ─────────────────────
        if is_qp1_final:
            c2_label_a, c2_label_b = 'CO2', 'CO3'
            model_co_keys = ['co1', 'co2', 'co3']
        else:
            c2_label_a, c2_label_b = 'CO3', 'CO4'
            model_co_keys = ['co1', 'co2', 'co3', 'co4', 'co5']

        # ═════════════════════════════════════════════════════════════
        # 2. Cycle 1 sheet – SSA1 + CIA1 + FA1 (CO1, CO2)
        # ═════════════════════════════════════════════════════════════
        ws1 = wb.create_sheet('Cycle 1 (SSA1+CIA1+FA1)')
        ws1.append([
            'S.no', "Student's Name", 'Register Number',
            'SSA1 CO1', 'SSA1 CO2', 'SSA1 Total',
            'CIA1 CO1', 'CIA1 CO2', 'CIA1 Total',
            'FA1 CO1', 'FA1 CO2', 'FA1 Total',
        ])

        for idx, s in enumerate(student_list, start=1):
            sid = int(s['id'])

            # SSA1 CO splits
            sp1 = ssa1_splits_all.get(sid, {})
            s1_co1 = _sf(sp1.get('co1'))
            s1_co2 = _sf(sp1.get('co2'))
            s1_total = _sf(ssa1_totals.get(sid))
            if s1_co1 is None and s1_co2 is None and s1_total is not None:
                s1_co1 = s1_total / 2.0
                s1_co2 = s1_total / 2.0

            # CIA1 CO splits
            c1_row = cia1_row_map.get(str(sid)) or cia1_row_map.get(sid) or {}
            c1_a, c1_b, c1_total = _cia_co_raw(c1_row, cia1_questions, True)

            # FA1 CO splits
            f1 = f1_rows_all.get(sid, {})
            f1_co1 = None
            f1_co2 = None
            f1_total = _sf(f1.get('total'))
            if _sf(f1.get('skill1')) is not None and _sf(f1.get('att1')) is not None:
                f1_co1 = _round2(_sf(f1['skill1']) + _sf(f1['att1']))
            if _sf(f1.get('skill2')) is not None and _sf(f1.get('att2')) is not None:
                f1_co2 = _round2(_sf(f1['skill2']) + _sf(f1['att2']))

            ws1.append([
                idx,
                self._safe_text(s.get('name')),
                self._safe_text(s.get('reg_no')),
                _v(s1_co1), _v(s1_co2), _v(s1_total),
                _v(c1_a), _v(c1_b), _v(c1_total),
                _v(f1_co1), _v(f1_co2), _v(f1_total),
            ])

        ws1.auto_filter.ref = f"A1:L{ws1.max_row}"
        ws1.freeze_panes = 'A2'

        # ═════════════════════════════════════════════════════════════
        # 3. Cycle 2 sheet – SSA2 + CIA2 + FA2
        # ═════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Cycle 2 (SSA2+CIA2+FA2)')
        ws2.append([
            'S.no', "Student's Name", 'Register Number',
            f'SSA2 {c2_label_a}', f'SSA2 {c2_label_b}', 'SSA2 Total',
            f'CIA2 {c2_label_a}', f'CIA2 {c2_label_b}', 'CIA2 Total',
            f'FA2 {c2_label_a}', f'FA2 {c2_label_b}', 'FA2 Total',
        ])

        for idx, s in enumerate(student_list, start=1):
            sid = int(s['id'])

            # SSA2 CO splits
            sp2 = ssa2_splits_all.get(sid, {})
            s2_co_a = _sf(sp2.get('co3'))
            s2_co_b = _sf(sp2.get('co4'))
            s2_total = _sf(ssa2_totals.get(sid))
            if s2_co_a is None and s2_co_b is None and s2_total is not None:
                s2_co_a = s2_total / 2.0
                s2_co_b = s2_total / 2.0

            # QP1FINAL SSA2 special mapping (mirrors _compute_weighted_final_total_theory_like)
            if is_qp1_final and sp2:
                first_v = _sf(sp2.get('co2'))
                if first_v is None:
                    first_v = _sf(sp2.get('co3'))
                if first_v is not None:
                    s2_co_a = first_v
                second_v = None
                if sp2.get('co3') is not None and sp2.get('co2') is not None:
                    second_v = _sf(sp2.get('co3'))
                if second_v is None:
                    second_v = _sf(sp2.get('co4'))
                if second_v is not None:
                    s2_co_b = second_v

            # CIA2 CO splits
            c2_row = cia2_row_map.get(str(sid)) or cia2_row_map.get(sid) or {}
            c2_a, c2_b, c2_total = _cia_co_raw(c2_row, cia2_questions, False)

            # FA2 CO splits
            f2 = f2_rows_all.get(sid, {})
            f2_co_a = None
            f2_co_b = None
            f2_total = _sf(f2.get('total'))
            if _sf(f2.get('skill1')) is not None and _sf(f2.get('att1')) is not None:
                f2_co_a = _round2(_sf(f2['skill1']) + _sf(f2['att1']))
            if _sf(f2.get('skill2')) is not None and _sf(f2.get('att2')) is not None:
                f2_co_b = _round2(_sf(f2['skill2']) + _sf(f2['att2']))

            ws2.append([
                idx,
                self._safe_text(s.get('name')),
                self._safe_text(s.get('reg_no')),
                _v(s2_co_a), _v(s2_co_b), _v(s2_total),
                _v(c2_a), _v(c2_b), _v(c2_total),
                _v(f2_co_a), _v(f2_co_b), _v(f2_total),
            ])

        ws2.auto_filter.ref = f"A1:L{ws2.max_row}"
        ws2.freeze_panes = 'A2'

        # ═════════════════════════════════════════════════════════════
        # 4. Model Exam sheet – all COs
        # ═════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet('Model Exam')
        model_co_headers = [f'MODEL {k.upper()}' for k in model_co_keys]
        ws3.append(['S.no', "Student's Name", 'Register Number'] + model_co_headers + ['MODEL Total'])

        for idx, s in enumerate(student_list, start=1):
            sid = int(s['id'])
            reg_no = reg_map.get(sid, '')

            model_marks = _extract_model_co_marks_for_student(
                model_sheet=model_sheet,
                student_id=sid,
                reg_no=reg_no,
                model_pattern=model_pattern,
            )

            row_data = [idx, self._safe_text(s.get('name')), self._safe_text(s.get('reg_no'))]

            if model_marks:
                m_total = 0.0
                for co_key in model_co_keys:
                    val = _sf(model_marks.get(co_key))
                    row_data.append(_v(val))
                    if val is not None:
                        m_total += val
                row_data.append(_v(m_total))
            else:
                row_data.extend(['-'] * (len(model_co_keys) + 1))

            ws3.append(row_data)

        last_col = len(model_co_headers) + 4  # S.no + Name + Reg + COs + Total
        col_letter = chr(ord('A') + last_col - 1) if last_col <= 26 else 'Z'
        ws3.auto_filter.ref = f"A1:{col_letter}{ws3.max_row}"
        ws3.freeze_panes = 'A2'

    def get(self, request):
        if not _user_is_iqac_admin(request.user):
            return Response({'detail': 'Only IQAC/OBE master can download this export.'}, status=403)

        ta_id_raw = self._safe_text(request.query_params.get('ta_id'))
        if not ta_id_raw:
            return Response({'detail': 'ta_id is required.'}, status=400)
        try:
            ta_id = int(ta_id_raw)
        except Exception:
            return Response({'detail': 'Invalid ta_id.'}, status=400)

        ta = (
            TeachingAssignment.objects.filter(id=ta_id, is_active=True)
            .select_related('subject', 'curriculum_row', 'elective_subject')
            .first()
        )
        if ta is None:
            return Response({'detail': 'Teaching assignment not found.'}, status=404)

        from openpyxl import Workbook
        from OBE.models import FinalInternalMark, LabPublishedSheet, Review1Mark, Review2Mark
        from OBE.services.final_internal_marks import (
            _assessment_map,
            _compute_weighted_final_total_theory_like,
            _resolve_class_type,
            _resolve_qp_type as _rqp,
            recompute_final_internal_marks,
        )

        course_code, course_name = self._resolve_course(ta)
        if not course_code:
            return Response({'detail': 'Unable to resolve course code for this assignment.'}, status=400)
        class_type = str(_resolve_class_type(ta) or '').upper()
        is_project_course = class_type == 'PROJECT'
        _qp_type_raw = _rqp(ta)
        is_qp1_final = 'QP1FINAL' in str(_qp_type_raw or '').upper().replace(' ', '')
        # OE Theory (QP1FINAL) courses convert final mark to 60 instead of 100
        scaled_max = 60.0 if is_qp1_final else 100.0

        sync_result = {}
        try:
            sync_result = recompute_final_internal_marks(
                actor_user_id=getattr(request.user, 'id', None),
                filters={'teaching_assignment_id': ta_id},
            ) or {}
        except Exception:
            sync_result = {'error': 'recompute_failed'}

        fim_qs = (
            FinalInternalMark.objects.filter(teaching_assignment_id=ta_id)
            .select_related('student__user', 'subject')
            .order_by('student_id', 'id')
        )

        student_rows = {}
        for fim in fim_qs:
            subj_code = self._safe_text(getattr(getattr(fim, 'subject', None), 'code', '')).upper()
            if subj_code and subj_code != course_code:
                continue

            sp = getattr(fim, 'student', None)
            if sp is None:
                continue
            sid = int(getattr(sp, 'id', 0) or 0)
            if sid <= 0:
                continue

            user = getattr(sp, 'user', None)
            student_name = ' '.join([
                self._safe_text(getattr(user, 'first_name', '')),
                self._safe_text(getattr(user, 'last_name', '')),
            ]).strip() if user else ''
            if not student_name:
                student_name = self._safe_text(getattr(user, 'username', '')) if user else ''

            live = _compute_weighted_final_total_theory_like(
                ta=ta,
                subject=getattr(fim, 'subject', None),
                student={'id': sid, 'reg_no': self._safe_text(getattr(sp, 'reg_no', ''))},
                ta_id=ta_id,
                return_details=True,
            )

            co_vals = {'co1': None, 'co2': None, 'co3': None, 'co4': None, 'co5': None}
            base_co_vals = {'co1': None, 'co2': None, 'co3': None, 'co4': None, 'co5': None}
            final_mark = None
            total_100 = None
            base_mark = None
            base_total_100 = None
            if isinstance(live, dict):
                final_mark = self._safe_float(live.get('total_40'))
                total_100 = self._safe_float(live.get('total_100'))
                base_mark = self._safe_float(live.get('base_total_40'))
                base_total_100 = self._safe_float(live.get('base_total_100'))
                co_payload = live.get('co_values_40') if isinstance(live.get('co_values_40'), dict) else {}
                co_vals = {
                    'co1': self._safe_float(co_payload.get('co1')),
                    'co2': self._safe_float(co_payload.get('co2')),
                    'co3': self._safe_float(co_payload.get('co3')),
                    'co4': self._safe_float(co_payload.get('co4')),
                    'co5': self._safe_float(co_payload.get('co5')),
                }
                base_co_payload = live.get('base_co_values_40') if isinstance(live.get('base_co_values_40'), dict) else {}
                base_co_vals = {
                    'co1': self._safe_float(base_co_payload.get('co1')),
                    'co2': self._safe_float(base_co_payload.get('co2')),
                    'co3': self._safe_float(base_co_payload.get('co3')),
                    'co4': self._safe_float(base_co_payload.get('co4')),
                    'co5': self._safe_float(base_co_payload.get('co5')),
                }

            if final_mark is None:
                final_mark = self._safe_float(getattr(fim, 'final_mark', None))
            if total_100 is None and final_mark is not None:
                from decimal import Decimal as _D, ROUND_HALF_UP as _RHU
                _raw = (float(final_mark) / 40.0) * scaled_max
                total_100 = int(_D(str(_raw)).quantize(_D('1'), rounding=_RHU))

            prev = student_rows.get(sid)
            if prev is None:
                student_rows[sid] = {
                    'student_id': sid,
                    'name': student_name,
                    'reg_no': self._safe_text(getattr(sp, 'reg_no', '')),
                    'co1': co_vals['co1'],
                    'co2': co_vals['co2'],
                    'co3': co_vals['co3'],
                    'co4': co_vals['co4'],
                    'co5': co_vals['co5'],
                    'fim': final_mark,
                    'total_100': total_100,
                    'base_co1': base_co_vals['co1'],
                    'base_co2': base_co_vals['co2'],
                    'base_co3': base_co_vals['co3'],
                    'base_co4': base_co_vals['co4'],
                    'base_co5': base_co_vals['co5'],
                    'base_fim': base_mark,
                    'base_total_100': base_total_100,
                }
                continue

            if prev.get('fim') is None and final_mark is not None:
                prev['fim'] = final_mark
            if prev.get('total_100') is None and total_100 is not None:
                prev['total_100'] = total_100
            if prev.get('base_fim') is None and base_mark is not None:
                prev['base_fim'] = base_mark
            if prev.get('base_total_100') is None and base_total_100 is not None:
                prev['base_total_100'] = base_total_100
            for key in ('co1', 'co2', 'co3', 'co4', 'co5'):
                if prev.get(key) is None and co_vals.get(key) is not None:
                    prev[key] = co_vals.get(key)
            for key in ('base_co1', 'base_co2', 'base_co3', 'base_co4', 'base_co5'):
                if prev.get(key) is None and base_co_vals.get(key.replace('base_', '')) is not None:
                    prev[key] = base_co_vals.get(key.replace('base_', ''))
            if not self._safe_text(prev.get('name')) and student_name:
                prev['name'] = student_name
            if not self._safe_text(prev.get('reg_no')):
                prev['reg_no'] = self._safe_text(getattr(sp, 'reg_no', ''))
            if not prev.get('student_id'):
                prev['student_id'] = sid

        rows = sorted(student_rows.values(), key=lambda r: (self._safe_text(r.get('reg_no')), self._safe_text(r.get('name'))))
        if not rows:
            return Response({'detail': 'No internal marks found for this teaching assignment.'}, status=404)

        def _extract_review_mark_from_lab_data(lab_data, student_id):
            if not isinstance(lab_data, dict):
                return None
            sid = str(student_id)
            sheet = lab_data.get('sheet') if isinstance(lab_data.get('sheet'), dict) else None
            rows_by_student = None
            if sheet and isinstance(sheet.get('rowsByStudentId'), dict):
                rows_by_student = sheet.get('rowsByStudentId')
            elif isinstance(lab_data.get('rowsByStudentId'), dict):
                rows_by_student = lab_data.get('rowsByStudentId')
            if not isinstance(rows_by_student, dict):
                return None
            row = rows_by_student.get(sid) or rows_by_student.get(student_id)
            if not isinstance(row, dict):
                return None

            direct = self._safe_float(row.get('ciaExam'))
            if direct is not None:
                return round(max(0.0, min(50.0, float(direct))), 2)

            comps = row.get('reviewComponentMarks') if isinstance(row.get('reviewComponentMarks'), dict) else {}
            if not isinstance(comps, dict):
                return None
            total = 0.0
            has_any = False
            for raw in comps.values():
                n = self._safe_float(raw)
                if n is None:
                    continue
                has_any = True
                total += float(n)
            if not has_any:
                return None
            return round(max(0.0, min(50.0, total)), 2)

        wb = Workbook()

        # ════════════════════════════════════════════════════════════════
        # SHEET 1 — Comprehensive Internal Marks (Cycle 1 + 2 + Model + FIM Before/After CQI)
        # ════════════════════════════════════════════════════════════════
        ws = wb.active
        ws.title = 'Internal Marks'

        if is_project_course:
            # ── Project courses: comprehensive layout matching Theory format ──
            # Two-header layout with Review 1/2 Before & After CQI, FIM Before/After CQI
            from django.db.models import Q as _QP
            from .models import Subject as _SubjectModel
            from OBE.models import ObeCqiPublished
            from OBE.services.final_internal_marks import (
                _compute_cqi_add as _cqi_add,
                _pick_scoped_row as _pick_row,
                _safe_float as _sfp,
                _round2 as _r2p,
            )

            _subject_obj = ta.subject
            if _subject_obj is None:
                _subject_obj = _SubjectModel.objects.filter(code__iexact=course_code).first()

            review1_map = {}
            review2_map = {}
            review1_lab_data = {}
            review2_lab_data = {}

            def _pick_lab_data(assessment_key):
                if _subject_obj is None:
                    return {}
                rows_qs = list(
                    LabPublishedSheet.objects.filter(subject_id=_subject_obj.id, assessment=assessment_key)
                    .filter(_QP(teaching_assignment_id=ta_id) | _QP(teaching_assignment__isnull=True))
                    .order_by('-updated_at')
                )
                exact = next((r for r in rows_qs if getattr(r, 'teaching_assignment_id', None) == ta_id), None)
                if exact is not None and isinstance(getattr(exact, 'data', None), dict):
                    return exact.data
                legacy = next((r for r in rows_qs if getattr(r, 'teaching_assignment_id', None) is None), None)
                if legacy is not None and isinstance(getattr(legacy, 'data', None), dict):
                    return legacy.data
                first = rows_qs[0] if rows_qs else None
                return first.data if first is not None and isinstance(getattr(first, 'data', None), dict) else {}

            export_student_ids = []
            for r in rows:
                try:
                    sid = int(r.get('student_id'))
                except Exception:
                    sid = 0
                if sid > 0:
                    export_student_ids.append(sid)

            if _subject_obj is not None and export_student_ids:
                review1_map = _assessment_map(Review1Mark, 'mark', _subject_obj.id, export_student_ids, ta_id)
                review2_map = _assessment_map(Review2Mark, 'mark', _subject_obj.id, export_student_ids, ta_id)
                review1_lab_data = _pick_lab_data('review1')
                review2_lab_data = _pick_lab_data('review2')

            # ── Fetch CQI published snapshot for this project course ──
            cqi_entries = {}
            cqi_co_set = set()
            if _subject_obj is not None:
                cqi_rows = list(
                    ObeCqiPublished.objects.filter(subject_id=_subject_obj.id)
                    .filter(_QP(teaching_assignment_id=ta_id) | _QP(teaching_assignment__isnull=True))
                    .order_by('-published_at')
                )
                cqi_row = _pick_row(cqi_rows, ta_id)
                if cqi_row:
                    cqi_entries = cqi_row.entries if isinstance(getattr(cqi_row, 'entries', None), dict) else {}
                    cqi_nums = cqi_row.co_numbers if isinstance(getattr(cqi_row, 'co_numbers', None), list) else []
                    cqi_co_set = {int(n) for n in cqi_nums if n is not None}

            def _vp(x):
                return _r2p(x) if x is not None else '-'

            # Project weights: Review 1 = 50 (mapped to CO1), Review 2 = 50 (mapped to CO2)
            PROJ_R1_MAX = 50.0
            PROJ_R2_MAX = 50.0
            PROJ_TOTAL_MAX = 100.0

            # ── Build section + column header rows ──
            header_sections = ['', '', '']  # S.no, Name, Reg placeholders
            header_cols = ['S.no', "Student's Name", 'Register Number']

            # Review 1 (Before CQI)
            header_sections.append('Review 1 (Before CQI)')
            header_cols.append('Mark')
            # Review 2 (Before CQI)
            header_sections.append('Review 2 (Before CQI)')
            header_cols.append('Mark')
            # Review 1 (After CQI)
            header_sections.append('Review 1 (After CQI)')
            header_cols.append('Mark')
            # Review 2 (After CQI)
            header_sections.append('Review 2 (After CQI)')
            header_cols.append('Mark')
            # FIM (Before CQI)
            header_sections.append('FIM (Before CQI)')
            header_cols.append('100')
            # FIM (After CQI)
            header_sections.append('FIM (After CQI)')
            header_cols.append('100')

            ws.append(header_sections)
            ws.append(header_cols)

            # ── Merge and style section header cells ──
            try:
                from openpyxl.styles import Font, Alignment
                section_ranges = {}
                for ci, sec in enumerate(header_sections, start=1):
                    if sec:
                        if sec not in section_ranges:
                            section_ranges[sec] = [ci, ci]
                        else:
                            section_ranges[sec][1] = ci
                for sec, (start_col, end_col) in section_ranges.items():
                    if start_col < end_col:
                        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                    cell = ws.cell(row=1, column=start_col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                for ci in range(1, len(header_cols) + 1):
                    ws.cell(row=2, column=ci).font = Font(bold=True)
            except Exception:
                pass

            # ── Write data rows ──
            proj_totals_before = {}
            proj_totals_after = {}

            for idx, row in enumerate(rows, start=1):
                sid = int(row.get('student_id') or 0)

                # Get raw review marks (before CQI)
                r1 = self._safe_float(review1_map.get(sid))
                r2 = self._safe_float(review2_map.get(sid))
                if r1 is None:
                    r1 = _extract_review_mark_from_lab_data(review1_lab_data, sid)
                if r2 is None:
                    r2 = _extract_review_mark_from_lab_data(review2_lab_data, sid)
                if r1 is not None:
                    r1 = round(max(0.0, min(PROJ_R1_MAX, float(r1))), 2)
                if r2 is not None:
                    r2 = round(max(0.0, min(PROJ_R2_MAX, float(r2))), 2)

                # FIM before CQI
                fim_before = round((r1 or 0.0) + (r2 or 0.0), 2) if (r1 is not None or r2 is not None) else None

                # Apply CQI to each review (mapped as CO1=Review1, CO2=Review2)
                cqi_student = cqi_entries.get(str(sid)) or cqi_entries.get(sid) or {}
                r1_after = r1
                r2_after = r2
                r1_add = 0.0
                r2_add = 0.0

                if r1 is not None and 1 in cqi_co_set:
                    inp_co1 = _sfp((cqi_student or {}).get('co1'))
                    r1_add = _cqi_add(co_value=r1, co_max=PROJ_R1_MAX, input_mark=inp_co1)
                    r1_after = _r2p(min(PROJ_R1_MAX, r1 + r1_add))

                if r2 is not None and 2 in cqi_co_set:
                    inp_co2 = _sfp((cqi_student or {}).get('co2'))
                    r2_add = _cqi_add(co_value=r2, co_max=PROJ_R2_MAX, input_mark=inp_co2)
                    r2_after = _r2p(min(PROJ_R2_MAX, r2 + r2_add))

                # FIM after CQI
                fim_after = round((r1_after or 0.0) + (r2_after or 0.0), 2) if (r1_after is not None or r2_after is not None) else None

                # Cap FIM at 100
                if fim_before is not None:
                    fim_before = min(PROJ_TOTAL_MAX, fim_before)
                if fim_after is not None:
                    fim_after = min(PROJ_TOTAL_MAX, fim_after)

                proj_totals_before[sid] = fim_before
                proj_totals_after[sid] = fim_after

                row_data = [
                    idx,
                    self._safe_text(row.get('name')),
                    self._safe_text(row.get('reg_no')),
                    _vp(r1),            # Review 1 Before CQI
                    _vp(r2),            # Review 2 Before CQI
                    _vp(r1_after),      # Review 1 After CQI
                    _vp(r2_after),      # Review 2 After CQI
                    fim_before if fim_before is not None else '-',   # FIM Before CQI /100
                    fim_after if fim_after is not None else '-',     # FIM After CQI /100
                ]
                ws.append(row_data)

            total_cols = len(header_cols)
            from openpyxl.utils import get_column_letter as _gcl_proj
            last_letter = _gcl_proj(total_cols)
            ws.auto_filter.ref = f"A2:{last_letter}{ws.max_row}"
            ws.freeze_panes = 'A3'

        else:
            # ── Theory / Special courses: comprehensive all-in-one sheet ──
            from django.db.models import Q as _Q
            from OBE.models import Ssa1Mark, Ssa2Mark, Formative1Mark, Formative2Mark
            from OBE.services.final_internal_marks import (
                _resolve_qp_type as __rqp,
                _extract_ssa_co_splits_for_ta,
                _get_cia_sheet_data,
                _get_model_sheet_data,
                _extract_model_co_marks_for_student,
                _get_qp_pattern,
                _safe_float as _sf,
                _safe_text as _st,
                _parse_co12,
                _parse_co34,
                _parse_question_co_numbers,
                _qp1_final_question_weight,
                _co_weights_12,
                _co_weights_34,
                _clamp,
                _round2,
            )

            from .models import Subject as _SubjectModel
            _subject_obj = ta.subject
            if _subject_obj is None:
                _subject_obj = _SubjectModel.objects.filter(code__iexact=course_code).first()

            subject_id = _subject_obj.id if _subject_obj else 0
            _student_id_list = [
                {'id': sid, 'name': data.get('name', ''), 'reg_no': data.get('reg_no', '')}
                for sid, data in sorted(
                    student_rows.items(),
                    key=lambda x: (self._safe_text(x[1].get('reg_no')), self._safe_text(x[1].get('name'))),
                )
            ]
            student_ids = [int(s['id']) for s in _student_id_list]
            reg_map = {int(s['id']): self._safe_text(s.get('reg_no', '')) for s in _student_id_list}

            qp_type = __rqp(ta)
            batch_id = getattr(getattr(ta, 'section', None), 'batch_id', None)

            def _v(x):
                return _round2(x) if x is not None else '-'

            # ── Fetch all raw assessment data ──
            ssa1_totals = _assessment_map(Ssa1Mark, 'mark', subject_id, student_ids, ta_id) if subject_id else {}
            ssa2_totals = _assessment_map(Ssa2Mark, 'mark', subject_id, student_ids, ta_id) if subject_id else {}
            ssa1_splits_all = _extract_ssa_co_splits_for_ta(subject_id, ta_id, 'ssa1', ['co1', 'co2']) if subject_id else {}
            ssa2_splits_all = _extract_ssa_co_splits_for_ta(subject_id, ta_id, 'ssa2', ['co3', 'co4']) if subject_id else {}

            def _fetch_formative_bulk(model_cls):
                result = {}
                if not subject_id:
                    return result
                qs = (
                    model_cls.objects.filter(subject_id=subject_id, student_id__in=student_ids)
                    .filter(_Q(teaching_assignment_id=ta_id) | _Q(teaching_assignment__isnull=True))
                    .values('student_id', 'teaching_assignment_id', 'skill1', 'skill2', 'att1', 'att2', 'total')
                )
                for row in qs:
                    sid = int(row['student_id'])
                    is_ta = row.get('teaching_assignment_id') == ta_id
                    existing = result.get(sid)
                    if existing is None or (not existing.get('_is_ta') and is_ta):
                        result[sid] = {**row, '_is_ta': is_ta}
                return result

            f1_rows_all = _fetch_formative_bulk(Formative1Mark)
            f2_rows_all = _fetch_formative_bulk(Formative2Mark)

            cia1_sheet = _get_cia_sheet_data(subject_id, ta_id, 'cia1') if subject_id else {}
            cia2_sheet = _get_cia_sheet_data(subject_id, ta_id, 'cia2') if subject_id else {}

            cia1_pattern = _get_qp_pattern(class_type=class_type, qp_type=qp_type, exam='CIA1', batch_id=batch_id)
            cia2_pattern = _get_qp_pattern(class_type=class_type, qp_type=qp_type, exam='CIA2', batch_id=batch_id)

            def _build_questions(sheet, pattern, is_cia1):
                qs = sheet.get('questions') if isinstance(sheet.get('questions'), list) else []
                p_marks = pattern.get('marks') if isinstance(pattern, dict) and isinstance(pattern.get('marks'), list) else []
                p_cos = pattern.get('cos') if isinstance(pattern, dict) and isinstance(pattern.get('cos'), list) else []
                out = []
                count = max(len(qs), len(p_marks))
                for i in range(count):
                    q = qs[i] if i < len(qs) and isinstance(qs[i], dict) else {}
                    key = _st(q.get('key')) or f'q{i + 1}'
                    mx = _sf(p_marks[i] if i < len(p_marks) else q.get('max'))
                    if mx is None:
                        mx = _sf(q.get('maxMarks'))
                    if mx is None:
                        mx = 0.0
                    co_raw = p_cos[i] if i < len(p_cos) else q.get('co')
                    if is_qp1_final:
                        co = co_raw
                    else:
                        co = _parse_co12(co_raw) if is_cia1 else _parse_co34(co_raw)
                    out.append({'key': key, 'max': float(mx), 'co': co})
                return out

            cia1_questions = _build_questions(cia1_sheet, cia1_pattern, True)
            cia2_questions = _build_questions(cia2_sheet, cia2_pattern, False)

            cia1_row_map = cia1_sheet.get('rowsByStudentId') if isinstance(cia1_sheet.get('rowsByStudentId'), dict) else {}
            cia2_row_map = cia2_sheet.get('rowsByStudentId') if isinstance(cia2_sheet.get('rowsByStudentId'), dict) else {}

            max_seen = 0
            for qq in cia2_questions:
                nums = _parse_question_co_numbers(qq.get('co'))
                if nums:
                    max_seen = max(max_seen, max(nums))
            qp1_cia2_offset = 1 if (is_qp1_final and max_seen > 0 and max_seen <= 2) else 0

            def _cia_co_raw(row, questions, is_cia1):
                if not isinstance(row, dict) or bool(row.get('absent')):
                    return None, None, None
                qvals = row.get('q') if isinstance(row.get('q'), dict) else {}
                c_a = 0.0
                c_b = 0.0
                has_any = False
                for q in questions:
                    mx = float(q.get('max') or 0)
                    n = _sf(qvals.get(q.get('key')))
                    if is_qp1_final and is_cia1:
                        raw_nums = _parse_question_co_numbers(q.get('co'))
                        raw_num = raw_nums[0] if raw_nums else None
                        wa = 1.0 if raw_num == 1 else 0.0
                        wb_weight = 1.0 if raw_num == 2 else 0.0
                    elif is_qp1_final and not is_cia1:
                        wa = _qp1_final_question_weight(q.get('co'), 2, qp1_cia2_offset)
                        wb_weight = _qp1_final_question_weight(q.get('co'), 3, qp1_cia2_offset)
                    elif is_cia1:
                        wa, wb_weight = _co_weights_12(q.get('co'))
                    else:
                        wa, wb_weight = _co_weights_34(q.get('co'))
                    if n is None:
                        continue
                    has_any = True
                    mark = _clamp(n, 0, mx)
                    c_a += mark * wa
                    c_b += mark * wb_weight
                if not has_any:
                    return None, None, None
                return _round2(c_a), _round2(c_b), _round2(c_a + c_b)

            model_sheet = _get_model_sheet_data(subject_id, ta_id, class_type) if subject_id else {}
            model_pattern = _get_qp_pattern(class_type=class_type, qp_type=qp_type, exam='MODEL', batch_id=batch_id)

            if is_qp1_final:
                c2_label_a, c2_label_b = 'CO2', 'CO3'
                model_co_keys = ['co1', 'co2', 'co3']
                fim_co_keys = ['co1', 'co2', 'co3']
            else:
                c2_label_a, c2_label_b = 'CO3', 'CO4'
                model_co_keys = ['co1', 'co2', 'co3', 'co4', 'co5']
                fim_co_keys = ['co1', 'co2', 'co3', 'co4', 'co5']

            scaled_label = str(int(scaled_max))

            # ── Build header rows ──
            header_sections = ['', '', '']  # S.no, Name, Reg placeholders
            header_cols = ['S.no', "Student's Name", 'Register Number']

            # Cycle 1
            for lbl in ['SSA1 CO1', 'SSA1 CO2', 'SSA1 Total', 'CIA1 CO1', 'CIA1 CO2', 'CIA1 Total', 'FA1 CO1', 'FA1 CO2', 'FA1 Total']:
                header_sections.append('Cycle 1')
                header_cols.append(lbl)
            # Cycle 2
            for lbl in [f'SSA2 {c2_label_a}', f'SSA2 {c2_label_b}', 'SSA2 Total',
                        f'CIA2 {c2_label_a}', f'CIA2 {c2_label_b}', 'CIA2 Total',
                        f'FA2 {c2_label_a}', f'FA2 {c2_label_b}', 'FA2 Total']:
                header_sections.append('Cycle 2')
                header_cols.append(lbl)
            # Model
            for k in model_co_keys:
                header_sections.append('Model Exam')
                header_cols.append(f'MODEL {k.upper()}')
            header_sections.append('Model Exam')
            header_cols.append('MODEL Total')
            # FIM Before CQI
            for k in fim_co_keys:
                header_sections.append('FIM (Before CQI)')
                header_cols.append(k.upper())
            header_sections.append('FIM (Before CQI)')
            header_cols.append('40')
            header_sections.append('FIM (Before CQI)')
            header_cols.append(scaled_label)
            # FIM After CQI
            for k in fim_co_keys:
                header_sections.append('FIM (After CQI)')
                header_cols.append(k.upper())
            header_sections.append('FIM (After CQI)')
            header_cols.append('40')
            header_sections.append('FIM (After CQI)')
            header_cols.append(scaled_label)

            # Write section header row
            ws.append(header_sections)
            # Write column header row
            ws.append(header_cols)

            # Merge section header cells
            try:
                from openpyxl.styles import Font, Alignment, PatternFill
                section_ranges = {}
                for ci, sec in enumerate(header_sections, start=1):
                    if sec:
                        if sec not in section_ranges:
                            section_ranges[sec] = [ci, ci]
                        else:
                            section_ranges[sec][1] = ci
                for sec, (start_col, end_col) in section_ranges.items():
                    if start_col < end_col:
                        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                    cell = ws.cell(row=1, column=start_col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                # Bold the column headers
                for ci in range(1, len(header_cols) + 1):
                    ws.cell(row=2, column=ci).font = Font(bold=True)
            except Exception:
                pass

            # ── Write data rows ──
            for idx, s in enumerate(_student_id_list, start=1):
                sid = int(s['id'])
                fim_row = student_rows.get(sid, {})

                # Cycle 1
                sp1 = ssa1_splits_all.get(sid, {})
                s1_co1 = _sf(sp1.get('co1'))
                s1_co2 = _sf(sp1.get('co2'))
                s1_total = _sf(ssa1_totals.get(sid))
                if s1_co1 is None and s1_co2 is None and s1_total is not None:
                    s1_co1 = s1_total / 2.0
                    s1_co2 = s1_total / 2.0
                c1_row = cia1_row_map.get(str(sid)) or cia1_row_map.get(sid) or {}
                c1_a, c1_b, c1_total = _cia_co_raw(c1_row, cia1_questions, True)
                f1 = f1_rows_all.get(sid, {})
                f1_co1 = None
                f1_co2 = None
                f1_total = _sf(f1.get('total'))
                if _sf(f1.get('skill1')) is not None and _sf(f1.get('att1')) is not None:
                    f1_co1 = _round2(_sf(f1['skill1']) + _sf(f1['att1']))
                if _sf(f1.get('skill2')) is not None and _sf(f1.get('att2')) is not None:
                    f1_co2 = _round2(_sf(f1['skill2']) + _sf(f1['att2']))

                # Cycle 2
                sp2 = ssa2_splits_all.get(sid, {})
                s2_co_a = _sf(sp2.get('co3'))
                s2_co_b = _sf(sp2.get('co4'))
                s2_total = _sf(ssa2_totals.get(sid))
                if s2_co_a is None and s2_co_b is None and s2_total is not None:
                    s2_co_a = s2_total / 2.0
                    s2_co_b = s2_total / 2.0
                if is_qp1_final and sp2:
                    first_v = _sf(sp2.get('co2'))
                    if first_v is None:
                        first_v = _sf(sp2.get('co3'))
                    if first_v is not None:
                        s2_co_a = first_v
                    second_v = None
                    if sp2.get('co3') is not None and sp2.get('co2') is not None:
                        second_v = _sf(sp2.get('co3'))
                    if second_v is None:
                        second_v = _sf(sp2.get('co4'))
                    if second_v is not None:
                        s2_co_b = second_v
                c2_row = cia2_row_map.get(str(sid)) or cia2_row_map.get(sid) or {}
                c2_a, c2_b, c2_total = _cia_co_raw(c2_row, cia2_questions, False)
                f2 = f2_rows_all.get(sid, {})
                f2_co_a = None
                f2_co_b = None
                f2_total = _sf(f2.get('total'))
                if _sf(f2.get('skill1')) is not None and _sf(f2.get('att1')) is not None:
                    f2_co_a = _round2(_sf(f2['skill1']) + _sf(f2['att1']))
                if _sf(f2.get('skill2')) is not None and _sf(f2.get('att2')) is not None:
                    f2_co_b = _round2(_sf(f2['skill2']) + _sf(f2['att2']))

                # Model
                model_marks = _extract_model_co_marks_for_student(
                    model_sheet=model_sheet,
                    student_id=sid,
                    reg_no=reg_map.get(sid, ''),
                    model_pattern=model_pattern,
                ) if subject_id else None
                model_vals = []
                m_total = 0.0
                m_has = False
                for k in model_co_keys:
                    val = _sf(model_marks.get(k)) if model_marks else None
                    model_vals.append(_v(val))
                    if val is not None:
                        m_total += val
                        m_has = True
                model_vals.append(_v(m_total) if m_has else '-')

                # FIM Before CQI
                before_co_vals = []
                for k in fim_co_keys:
                    before_co_vals.append(_v(fim_row.get(f'base_{k}')))
                before_co_vals.append(_v(fim_row.get('base_fim')))
                before_co_vals.append(fim_row.get('base_total_100') if fim_row.get('base_total_100') is not None else '-')

                # FIM After CQI
                after_co_vals = []
                for k in fim_co_keys:
                    after_co_vals.append(_v(fim_row.get(k)))
                after_co_vals.append(_v(fim_row.get('fim')))
                after_co_vals.append(fim_row.get('total_100') if fim_row.get('total_100') is not None else '-')

                row_data = [
                    idx,
                    self._safe_text(s.get('name')),
                    self._safe_text(s.get('reg_no')),
                    # Cycle 1
                    _v(s1_co1), _v(s1_co2), _v(s1_total),
                    _v(c1_a), _v(c1_b), _v(c1_total),
                    _v(f1_co1), _v(f1_co2), _v(f1_total),
                    # Cycle 2
                    _v(s2_co_a), _v(s2_co_b), _v(s2_total),
                    _v(c2_a), _v(c2_b), _v(c2_total),
                    _v(f2_co_a), _v(f2_co_b), _v(f2_total),
                ] + model_vals + before_co_vals + after_co_vals

                ws.append(row_data)

            total_cols = len(header_cols)
            last_letter = chr(ord('A') + min(total_cols - 1, 25)) if total_cols <= 26 else 'Z'
            if total_cols > 26:
                # Handle columns beyond Z (AA, AB, etc.)
                from openpyxl.utils import get_column_letter
                last_letter = get_column_letter(total_cols)
            ws.auto_filter.ref = f"A2:{last_letter}{ws.max_row}"
            ws.freeze_panes = 'A3'

        # ════════════════════════════════════════════════════════════════
        # SHEET 2 — Summary (Name, Register No., Total)
        # ════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Summary')
        ws2.append(['Name', 'Register No.', 'Total'])
        try:
            from openpyxl.styles import Font as _F2
            for ci in range(1, 4):
                ws2.cell(row=1, column=ci).font = _F2(bold=True)
        except Exception:
            pass

        for row in rows:
            if is_project_course:
                # For project courses, use the FIM After CQI total (/100)
                sid = int(row.get('student_id') or 0)
                t100 = proj_totals_after.get(sid) if sid > 0 else None
            else:
                t100 = row.get('total_100')
            ws2.append([
                self._safe_text(row.get('name')),
                self._safe_text(row.get('reg_no')),
                t100 if t100 is not None else '-',
            ])
        ws2.auto_filter.ref = f"A1:C{ws2.max_row}"
        ws2.freeze_panes = 'A2'

        filename = f"{self._safe_filename(course_code)} {self._safe_filename(course_name)}.xlsx"
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        response = HttpResponse(
            out.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class SectionAdvisorViewSet(viewsets.ModelViewSet):
    queryset = SectionAdvisor.objects.select_related('section__batch__course__department', 'advisor')
    serializer_class = SectionAdvisorSerializer
    permission_classes = (IsAuthenticated, IsHODOfDepartment)

    def get_queryset(self):
        user = self.request.user
        perms = get_user_permissions(user)
        # users with explicit permission may view advisor assignments
        # but visibility should be limited to departments the user is effective for
        if user.is_superuser:
            return self.queryset

        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return SectionAdvisor.objects.none()

        # compute departments the user effectively represents (own dept + HOD/AHOD mappings)
        allowed_depts = get_user_effective_departments(user)

        # If user has assign permission, allow viewing assignments for their departments
        if 'academics.assign_advisor' in perms:
            if allowed_depts:
                return self.queryset.filter(
                    Q(section__batch__course__department_id__in=allowed_depts) |
                    Q(section__batch__department_id__in=allowed_depts) |
                    Q(section__managing_department_id__in=allowed_depts)
                )
            return SectionAdvisor.objects.none()

        # fallback: HODs (role-based) can view for their HOD departments
        hod_depts = DepartmentRole.objects.filter(staff=staff_profile, role='HOD', is_active=True).values_list('department_id', flat=True)
        return self.queryset.filter(
            Q(section__batch__course__department_id__in=hod_depts) |
            Q(section__batch__department_id__in=hod_depts) |
            Q(section__managing_department_id__in=hod_depts)
        )

    def perform_create(self, serializer):
        user = self.request.user
        # require explicit assign permission (or fallback to model add perm)
        perms = get_user_permissions(user)
        if not (('academics.assign_advisor' in perms) or user.has_perm('academics.add_sectionadvisor')):
            raise PermissionDenied('You do not have permission to assign advisors.')
        # serializer.validate already checks HOD membership and dept match
        serializer.save()

    def create(self, request, *args, **kwargs):
        # Handle duplicate active section+academic_year by updating existing mapping.
        data = request.data or {}
        section_id = data.get('section_id') or data.get('section')
        academic_year = data.get('academic_year')
        advisor_id = data.get('advisor_id') or data.get('advisor')

        # If academic_year missing but section and advisor present, default to active AcademicYear
        if section_id and advisor_id and not academic_year:
            try:
                active_ay = AcademicYear.objects.filter(is_active=True).first() or AcademicYear.objects.order_by('-id').first()
                if active_ay is not None:
                    academic_year = active_ay.pk
            except Exception:
                academic_year = None

        if section_id and academic_year and advisor_id:
            try:
                # Accept numeric IDs or object payloads
                sec_id = int(section_id)
                ay_id = int(academic_year)
            except Exception:
                sec_id = None
                ay_id = None

            if sec_id and ay_id:
                # If provided academic year isn't active, prefer the current active academic year
                try:
                    provided_ay = AcademicYear.objects.filter(pk=ay_id).first()
                    if provided_ay is not None and not provided_ay.is_active:
                        # Prefer an active academic year with the same name (pair like Odd/Even)
                        active_same = AcademicYear.objects.filter(name=provided_ay.name, is_active=True).first()
                        if active_same is not None:
                            ay_id = active_same.pk
                        else:
                            # Fallback to any active academic year
                            active_ay = AcademicYear.objects.filter(is_active=True).first()
                            if active_ay is not None:
                                ay_id = active_ay.pk
                except Exception:
                    pass
                existing = SectionAdvisor.objects.filter(section_id=sec_id, academic_year_id=ay_id, is_active=True).first()
                if existing:
                    # update advisor and return existing
                    try:
                        old_advisor = getattr(existing, 'advisor', None)
                        old_advisor_id = getattr(old_advisor, 'id', None)
                    except Exception:
                        old_advisor = None
                        old_advisor_id = None

                    existing.advisor_id = int(advisor_id)
                    if 'is_active' in data:
                        existing.is_active = bool(data.get('is_active'))
                    existing.save()

                    # If the advisor changed, attempt to remove ADVISOR role from the previous advisor
                    try:
                        if old_advisor_id and int(advisor_id) != int(old_advisor_id):
                            from accounts.models import Role
                            # reload the old advisor instance to be safe
                            old_sp = StaffProfile.objects.filter(pk=old_advisor_id).select_related('user').first()
                            if old_sp:
                                old_user = getattr(old_sp, 'user', None)
                                if old_user:
                                    role_obj = Role.objects.filter(name='ADVISOR').first()
                                    if role_obj and role_obj in old_user.roles.all():
                                        # Check if old advisor has any other active SectionAdvisor mappings
                                        other_active = SectionAdvisor.objects.filter(advisor=old_sp, is_active=True).exclude(pk=existing.pk).exists()
                                        if not other_active:
                                            try:
                                                old_user.roles.remove(role_obj)
                                            except Exception:
                                                # Don't raise from role removal; signal handlers or validations may prevent removal
                                                pass
                    except Exception:
                        # best-effort only; never fail the API because role-sync failed
                        pass

                    serializer = self.get_serializer(existing)
                    return Response(serializer.data, status=status.HTTP_200_OK)

                # No existing mapping -> create using the resolved ay_id
                data_copy = dict(data)
                data_copy['academic_year'] = ay_id
                # ensure section/advisor are integers
                data_copy['section_id'] = sec_id
                data_copy['advisor_id'] = int(advisor_id)
                try:
                    serializer = self.get_serializer(data=data_copy)
                    serializer.is_valid(raise_exception=True)
                    self.perform_create(serializer)
                    headers = self.get_success_headers(serializer.data)
                    return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
                except Exception as e:
                    import logging, traceback
                    logging.getLogger(__name__).exception('Error creating SectionAdvisor: %s', e)
                    tb = traceback.format_exc()
                    return Response({'detail': 'Failed to create advisor assignment.', 'error': str(e), 'trace': tb}, status=status.HTTP_400_BAD_REQUEST)

        return super().create(request, *args, **kwargs)

    def perform_destroy(self, instance):
        """Allow only users with assign_advisor permission or the model delete perm or superuser to delete an advisor assignment."""
        user = self.request.user
        perms = get_user_permissions(user)
        if not (user.is_superuser or ('academics.assign_advisor' in perms) or user.has_perm('academics.delete_sectionadvisor')):
            raise PermissionDenied('You do not have permission to delete advisor assignments.')
        # Deleting the instance will trigger post_delete signal to remove ADVISOR role if no other active mapping exists
        instance.delete()


class MentorStaffListView(APIView):
    permission_classes = (IsAuthenticated,)

    def _has_role(self, user, role_name: str) -> bool:
        try:
            return user and user.is_authenticated and user.roles.filter(name__iexact=role_name).exists()
        except Exception:
            return False

    def _get_allowed_depts_for_assignment(self, user):
        """Derive department IDs the caller is allowed to manage mentor mappings for.

        Primary source: staff current dept + HOD/AHOD (get_user_effective_departments).
        Fallback: departments of sections the staff advises (common when staff
        profiles are missing department/current_department assignments).
        """
        allowed_depts = get_user_effective_departments(user)
        if allowed_depts:
            return allowed_depts

        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return []

        try:
            base_qs = SectionAdvisor.objects.filter(advisor=staff_profile, is_active=True)
            active_year_qs = base_qs.filter(academic_year__is_active=True)
            qs = active_year_qs if active_year_qs.exists() else base_qs

            dept_ids = list(
                qs.exclude(section__batch__course__department_id__isnull=True)
                .values_list('section__batch__course__department_id', flat=True)
                .distinct()
            )
            return [int(d) for d in dept_ids if d]
        except Exception:
            return []

    def get(self, request):
        user = request.user
        # only allow users with assign_mentor permission or superuser
        perms = get_user_permissions(user)
        # Advisors are allowed to use mentor assignment.
        if not (user.is_superuser or ('academics.assign_mentor' in perms) or self._has_role(user, 'ADVISOR')):
            return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        allowed_depts = self._get_allowed_depts_for_assignment(user)

        # If we cannot derive departments (common in partially configured data),
        # don't hide the entire screen; fall back to all staff.
        if not allowed_depts:
            staffs = StaffProfile.objects.all().select_related('user')
        else:
            # Include staff whose legacy department matches OR who has an active
            # department assignment matching.
            staffs = (
                StaffProfile.objects.filter(
                    Q(department__id__in=allowed_depts)
                    | Q(department_assignments__end_date__isnull=True, department_assignments__department_id__in=allowed_depts)
                )
                .select_related('user')
                .distinct()
            )
        data = []
        for s in staffs:
            data.append({'id': s.id, 'user_id': getattr(getattr(s, 'user', None), 'id', None), 'username': getattr(getattr(s, 'user', None), 'username', None), 'staff_id': s.staff_id})
        return Response({'results': data})


class CustomSubjectsListView(APIView):
    """Return the list of allowed custom subject choices for TeachingAssignment.custom_subject."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        try:
            field = TeachingAssignment._meta.get_field('custom_subject')
            choices = getattr(field, 'choices', []) or []
            results = [{'value': c[0], 'label': c[1]} for c in choices]
            return Response({'results': results})
        except Exception as e:
            logging.getLogger(__name__).exception('Failed to get custom subject choices: %s', e)
            try:
                # Fallback: return distinct non-null values present in DB
                qs = TeachingAssignment.objects.exclude(custom_subject__isnull=True).exclude(custom_subject__exact='').values_list('custom_subject', flat=True).distinct()
                results = [{'value': v, 'label': v} for v in qs]
                return Response({'results': results})
            except Exception as e2:
                logging.getLogger(__name__).exception('Fallback failed for custom subject choices: %s', e2)
                # Return safe empty result to avoid 500 in the UI
                return Response({'results': []})


class MentorStudentsForStaffView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request, staff_id: int):
        user = request.user
        perms = get_user_permissions(user)
        has_global = user.is_superuser or ('academics.assign_mentor' in perms)

        # ensure staff exists
        staff = StaffProfile.objects.filter(pk=int(staff_id)).first()
        if not staff:
            return Response({'results': []})

        # Fetch students currently mapped to this mentor (active mappings)
        from .models import StudentMentorMap, StudentSectionAssignment, StudentProfile

        # base mentor mappings
        mentor_maps = StudentMentorMap.objects.filter(mentor=staff, is_active=True).exclude(student__status__in=['INACTIVE', 'DEBAR']).select_related('student__user')

        # Determine target department for the staff (if available)
        target_dept = getattr(getattr(staff, 'current_department', None), 'id', None) or getattr(getattr(staff, 'department', None), 'id', None)

        # If the requester is a plain advisor (has section advisor entries) and
        # is NOT a superuser or HOD for the target department, restrict the
        # mentor mappings to students who are in the requester's advised sections
        requester_staff = getattr(user, 'staff_profile', None)
        if requester_staff and not user.is_superuser:
            # check if requester is HOD of the target department — HODs can view all
            is_requester_hod = False
            try:
                is_requester_hod = DepartmentRole.objects.filter(staff=requester_staff, role='HOD', is_active=True, department_id=target_dept).exists()
            except Exception:
                is_requester_hod = False

            if not is_requester_hod:
                requester_section_ids = list(SectionAdvisor.objects.filter(advisor=requester_staff, is_active=True, academic_year__is_active=True).values_list('section_id', flat=True))
                if requester_section_ids:
                    assigned_student_ids = set(StudentSectionAssignment.objects.filter(section_id__in=requester_section_ids, end_date__isnull=True).values_list('student_id', flat=True))
                    legacy_student_ids = set(StudentProfile.objects.filter(section_id__in=requester_section_ids).values_list('id', flat=True))
                    allowed_student_ids = assigned_student_ids | legacy_student_ids
                    mentor_maps = mentor_maps.filter(student__id__in=allowed_student_ids)

        students = [m.student for m in mentor_maps]

        ser = StudentSimpleSerializer([
            {'id': st.pk, 'reg_no': st.reg_no, 'user': getattr(st, 'user', None), 'section_id': getattr(st, 'section_id', None), 'section_name': str(getattr(st, 'section', ''))}
            for st in students
        ], many=True)
        return Response({'results': ser.data})


class MentorMapCreateView(APIView):
    permission_classes = (IsAuthenticated,)

    def _has_role(self, user, role_name: str) -> bool:
        try:
            return user and user.is_authenticated and user.roles.filter(name__iexact=role_name).exists()
        except Exception:
            return False

    def _get_allowed_depts_for_assignment(self, user):
        allowed_depts = get_user_effective_departments(user)
        if allowed_depts:
            return allowed_depts

        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return []

        try:
            base_qs = SectionAdvisor.objects.filter(advisor=staff_profile, is_active=True)
            active_year_qs = base_qs.filter(academic_year__is_active=True)
            qs = active_year_qs if active_year_qs.exists() else base_qs

            dept_ids = list(
                qs.exclude(section__batch__course__department_id__isnull=True)
                .values_list('section__batch__course__department_id', flat=True)
                .distinct()
            )
            return [int(d) for d in dept_ids if d]
        except Exception:
            return []

    def post(self, request):
        user = request.user
        perms = get_user_permissions(user)
        if not (user.is_superuser or ('academics.assign_mentor' in perms) or self._has_role(user, 'ADVISOR')):
            return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        mentor_id = request.data.get('mentor_id')
        student_ids = request.data.get('student_ids') or request.data.get('student_id')
        if not mentor_id or not student_ids:
            return Response({'detail': 'mentor_id and student_ids required'}, status=status.HTTP_400_BAD_REQUEST)
        if isinstance(student_ids, int):
            student_ids = [student_ids]

        mentor = StaffProfile.objects.filter(pk=int(mentor_id)).first()
        if not mentor:
            return Response({'detail': 'Mentor not found'}, status=status.HTTP_404_NOT_FOUND)

        allowed_depts = self._get_allowed_depts_for_assignment(user)
        target_dept = getattr(getattr(mentor, 'current_department', None), 'id', None) or getattr(getattr(mentor, 'department', None), 'id', None)

        # If we can't derive allowed departments for the caller, allow mapping
        # (otherwise the UI becomes unusable in partially configured setups).
        if not user.is_superuser and allowed_depts and target_dept not in allowed_depts:
            return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        results = {'created': 0, 'skipped': 0, 'errors': []}
        try:
            with transaction.atomic():
                for sid in student_ids:
                    try:
                        sp = StudentProfile.objects.filter(pk=int(sid)).first()
                        if not sp:
                            results['skipped'] += 1
                            continue
                        # deactivate existing active mentor mapping for this student
                        StudentMentorMap.objects.filter(student=sp, is_active=True).update(is_active=False)
                        StudentMentorMap.objects.create(student=sp, mentor=mentor, is_active=True)
                        results['created'] += 1
                    except Exception as e:
                        results['errors'].append(str(e))
        except Exception as e:
            return Response({'detail': 'Failed to create mappings', 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(results)


class MentorUnmapView(APIView):
    permission_classes = (IsAuthenticated,)

    def _has_role(self, user, role_name: str) -> bool:
        try:
            return user and user.is_authenticated and user.roles.filter(name__iexact=role_name).exists()
        except Exception:
            return False

    def post(self, request):
        user = request.user
        perms = get_user_permissions(user)
        if not (user.is_superuser or ('academics.assign_mentor' in perms) or self._has_role(user, 'ADVISOR')):
            return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        mentor_id = request.data.get('mentor_id')
        student_ids = request.data.get('student_ids') or request.data.get('student_id')
        if not student_ids:
            return Response({'detail': 'student_ids required'}, status=status.HTTP_400_BAD_REQUEST)
        if isinstance(student_ids, int):
            student_ids = [student_ids]

        try:
            q = StudentMentorMap.objects.filter(student_id__in=[int(s) for s in student_ids], is_active=True)
            if mentor_id:
                q = q.filter(mentor_id=int(mentor_id))
            updated = q.update(is_active=False)
        except Exception as e:
            return Response({'detail': 'Failed to unmap', 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'unmapped': updated})

class TeachingAssignmentViewSet(viewsets.ModelViewSet):
    queryset = TeachingAssignment.objects.select_related('staff', 'subject', 'section', 'academic_year')
    serializer_class = TeachingAssignmentSerializer
    # Allow authenticated users; detailed authorisation is enforced in methods
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        user = self.request.user

        # Academic Controller (IQAC / OBE master) needs full assignment visibility
        # even when the user does not have a linked staff profile.
        if _user_is_iqac_admin(user):
            return self.queryset

        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return TeachingAssignment.objects.none()
        # Only include assignments for sections the user advises (active mapping)
        # or assignments belonging to the staff themselves. Users with the
        # `academics.view_assigned_subjects` permission (or superusers) are
        # allowed to see elective assignments across departments as well,
        # but should NOT see every regular assignment across the system.
        advisor_section_ids = list(SectionAdvisor.objects.filter(advisor=staff_profile, is_active=True, academic_year__is_active=True).values_list('section_id', flat=True))
        perms = get_user_permissions(user)
        from django.db.models import Q

        # Get HOD-accessible department sections
        hod_department_section_ids = []
        try:
            # Check if user is HOD of any department
            hod_depts = DepartmentRole.objects.filter(
                staff=staff_profile,
                role__iexact='HOD',
                academic_year__is_active=True
            ).values_list('department_id', flat=True)
            
            if hod_depts:
                # Get all sections from HOD's department(s)
                from academics.models import Section
                hod_department_section_ids = list(
                    Section.objects.filter(
                        batch__course__department_id__in=hod_depts
                    ).values_list('id', flat=True)
                )
        except Exception:
            pass

        # If caller has global view permission, expose elective assignments
        # but restrict visibility to assignments whose subject/row department
        # matches the user's effective departments (unless superuser).
        if 'academics.view_assigned_subjects' in perms or user.is_superuser:
            # base: elective assignments
            q = Q(elective_subject__isnull=False)
            # include advisor sections, HOD department sections, and own assignments always
            if advisor_section_ids:
                q |= Q(section_id__in=advisor_section_ids)
            if hod_department_section_ids:
                q |= Q(section_id__in=hod_department_section_ids)
            q |= Q(staff__user=getattr(user, 'id', None))

            # if not superuser, further restrict elective assignments to
            # those belonging to departments the user is effective for
            if not user.is_superuser:
                allowed_depts = get_user_effective_departments(user)
                if allowed_depts:
                    dept_q = (
                        Q(section__batch__course__department_id__in=allowed_depts)
                        | Q(curriculum_row__department_id__in=allowed_depts)
                        # match elective options by their explicit department OR their parent's department
                        | Q(elective_subject__department_id__in=allowed_depts)
                        | Q(elective_subject__parent__department_id__in=allowed_depts)
                    )
                    # apply department filter only to elective assignments part
                    q = (Q(elective_subject__isnull=False) & dept_q) | Q(section_id__in=advisor_section_ids)
                    if hod_department_section_ids:
                        q |= Q(section_id__in=hod_department_section_ids)
                    q |= Q(staff__user=getattr(user, 'id', None))
                else:
                    # no effective departments -> fall back to advisor sections and own assignments
                    q = Q(section_id__in=advisor_section_ids) | Q(staff__user=getattr(user, 'id', None))
                    if hod_department_section_ids:
                        q |= Q(section_id__in=hod_department_section_ids)

            return self.queryset.filter(q)

        # Default: restrict to advisor sections, HOD department sections, and own assignments
        final_q = Q()
        if advisor_section_ids:
            final_q |= Q(section_id__in=advisor_section_ids)
        if hod_department_section_ids:
            final_q |= Q(section_id__in=hod_department_section_ids)
        final_q |= Q(staff__user=getattr(user, 'id', None))

        if final_q:
            return self.queryset.filter(final_q)
        return TeachingAssignment.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        perms = get_user_permissions(user)
        # If user has explicit assign permission or model add perm, allow
        # For elective-specific assignment require separate permission
        is_elective_payload = False
        try:
            if 'elective_subject_id' in getattr(serializer, 'initial_data', {}) or 'elective_subject' in getattr(serializer, 'validated_data', {}):
                is_elective_payload = True
        except Exception:
            is_elective_payload = False

        # If this is an elective payload, serializer.validate() already
        # enforces HOD membership or explicit elective permission. Allow
        # creation when validated (no section required for electives).
        if is_elective_payload:
            serializer.save()
            return
        else:
            if ('academics.assign_teaching' in perms) or user.has_perm('academics.add_teachingassignment'):
                serializer.save()
                return
            serializer.save()
            return

        # Otherwise restrict to advisors for the target section only
        staff_profile = getattr(user, 'staff_profile', None)
        section_obj = None
        try:
            if 'section' in getattr(serializer, 'validated_data', {}):
                section_obj = serializer.validated_data.get('section')
            elif 'section_id' in getattr(serializer, 'validated_data', {}):
                sid = serializer.validated_data.get('section_id')
                from .models import Section as _Section
                section_obj = _Section.objects.filter(pk=int(sid)).first()
        except Exception:
            section_obj = None

        if not section_obj:
            raise PermissionDenied('You do not have permission to assign teaching for this section.')

        is_advisor = SectionAdvisor.objects.filter(section=section_obj, advisor=staff_profile, is_active=True, academic_year__is_active=True).exists() if staff_profile else False

        if not is_advisor:
            raise PermissionDenied('You do not have permission to assign teaching for this section.')

        serializer.save()

    @action(detail=True, methods=['get', 'post'], permission_classes=(IsAuthenticated,), url_path='enabled_assessments', url_name='enabled_assessments')
    def enabled_assessments(self, request, pk=None):
        """Get or set enabled assessments.

        - For normal courses: stored on TeachingAssignment.enabled_assessments.
        - For SPECIAL courses: stored globally on SpecialCourseAssessmentSelection
          (curriculum_row + academic_year), locked after first save.

        GET returns { enabled_assessments: [...], meta: {...} }
        POST accepts { enabled_assessments: ["ssa1","cia1",...] }
        """
        try:
            ta = TeachingAssignment.objects.select_related(
                'section',
                'section__batch',
                'section__batch__course',
                'section__batch__course__department',
                'academic_year',
                'curriculum_row',
                'curriculum_row__master',
                'subject',
                'staff',
            ).get(pk=int(pk), is_active=True)
        except TeachingAssignment.DoesNotExist:
            return Response({'detail': 'Teaching assignment not found'}, status=404)

        def _resolve_curriculum_row(assignment: TeachingAssignment):
            """Best-effort resolve CurriculumDepartment row for this assignment.

            Some legacy TeachingAssignment rows may have `subject` filled but
            `curriculum_row` missing. For SPECIAL course behavior we need the
            curriculum row to find the global lock.
            """
            row = getattr(assignment, 'curriculum_row', None)
            if row is not None:
                return row
            try:
                from curriculum.models import CurriculumDepartment

                dept = None
                try:
                    dept = assignment.section.batch.course.department
                except Exception:
                    dept = None

                subj = getattr(assignment, 'subject', None)
                code = getattr(subj, 'code', None)
                name = getattr(subj, 'name', None)

                qs = CurriculumDepartment.objects.all().select_related('master', 'department')
                if dept is not None:
                    qs = qs.filter(department=dept)

                if code:
                    qs = qs.filter(Q(course_code__iexact=str(code).strip()) | Q(master__course_code__iexact=str(code).strip()))
                elif name:
                    qs = qs.filter(Q(course_name__iexact=str(name).strip()) | Q(master__course_name__iexact=str(name).strip()))
                else:
                    return None

                return qs.order_by('-updated_at', '-id').first()
            except Exception:
                return None

        def _is_special_course_row(row) -> bool:
            try:
                if not row:
                    return False
                # Prefer department row class_type, fall back to master
                ct = getattr(row, 'class_type', None) or getattr(getattr(row, 'master', None), 'class_type', None)
                return str(ct or '').upper() == 'SPECIAL'
            except Exception:
                return False

        def _clean_keys(vals, *, include_model=False):
            allowed = {'ssa1', 'formative1', 'ssa2', 'formative2', 'cia1', 'cia2'}
            if include_model:
                allowed.add('model')
            cleaned = []
            for v in (vals or []):
                try:
                    s = str(v or '').strip().lower()
                except Exception:
                    s = ''
                if s and s in allowed and s not in cleaned:
                    cleaned.append(s)
            return cleaned

        user = request.user
        row = _resolve_curriculum_row(ta)
        if request.method == 'GET':
            meta = {'mode': 'TEACHING_ASSIGNMENT', 'locked': False, 'can_edit': True}
            if _is_special_course_row(row):
                sel = None
                master_id = getattr(row, 'master_id', None) if row is not None else None
                if master_id is not None:
                    # Global selection is shared across all CurriculumDepartment rows
                    # under the same master (course) for the academic year.
                    sel = (
                        SpecialCourseAssessmentSelection.objects.filter(
                            curriculum_row__master_id=master_id,
                            academic_year=ta.academic_year,
                        )
                        .select_related('curriculum_row')
                        .order_by('id')
                        .first()
                    )

                # If there is no global selection yet, fall back to the curriculum
                # configuration so other staff still see consistent enabled exams.
                enabled = None
                if sel is not None:
                    enabled = getattr(sel, 'enabled_assessments', [])
                if enabled is None:
                    enabled = getattr(row, 'enabled_assessments', None) if row is not None else None
                if enabled is None:
                    enabled = []

                locked = bool(sel and sel.locked)

                staff_profile = getattr(user, 'staff_profile', None)
                latest_req = None
                if master_id is not None and staff_profile:
                    latest_req = (
                        SpecialCourseAssessmentEditRequest.objects.filter(
                            selection__curriculum_row__master_id=master_id,
                            selection__academic_year=ta.academic_year,
                            requested_by=staff_profile,
                        )
                        .select_related('selection')
                        .order_by('-requested_at')
                        .first()
                    )

                # Safety net: If the central OBE edit-request queue has already been
                # approved/rejected for this SPECIAL selection, mirror that status here.
                # This ensures faculty immediately sees approval without needing perfect
                # sync in the IQAC approve handler.
                try:
                    if latest_req is not None:
                        from OBE.models import ObeEditRequest

                        staff_user = getattr(staff_profile, 'user', None)
                        subj_code = ''
                        try:
                            subj_code = getattr(row, 'course_code', None) or getattr(getattr(row, 'master', None), 'course_code', None) or ''
                        except Exception:
                            subj_code = ''

                        if staff_user is not None and subj_code:
                            obe_row = (
                                ObeEditRequest.objects.filter(
                                    staff_user=staff_user,
                                    academic_year=ta.academic_year,
                                    subject_code=subj_code,
                                    assessment='model',
                                    scope='MARK_MANAGER',
                                )
                                .order_by('-updated_at', '-id')
                                .first()
                            )

                            if obe_row is not None:
                                # Mirror APPROVED window.
                                if str(getattr(obe_row, 'status', '')).upper() == 'APPROVED':
                                    approved_until = getattr(obe_row, 'approved_until', None)
                                    if approved_until is not None and timezone.now() < approved_until:
                                        if latest_req.status != SpecialCourseAssessmentEditRequest.STATUS_APPROVED or latest_req.can_edit_until != approved_until:
                                            latest_req.status = SpecialCourseAssessmentEditRequest.STATUS_APPROVED
                                            latest_req.can_edit_until = approved_until
                                            latest_req.reviewed_by = getattr(obe_row, 'reviewed_by', None)
                                            latest_req.reviewed_at = getattr(obe_row, 'reviewed_at', None)
                                            latest_req.used_at = None
                                            latest_req.save(update_fields=['status', 'can_edit_until', 'reviewed_by', 'reviewed_at', 'used_at'])

                                # Mirror REJECTED.
                                elif str(getattr(obe_row, 'status', '')).upper() == 'REJECTED':
                                    if latest_req.status != SpecialCourseAssessmentEditRequest.STATUS_REJECTED:
                                        latest_req.status = SpecialCourseAssessmentEditRequest.STATUS_REJECTED
                                        latest_req.can_edit_until = None
                                        latest_req.reviewed_by = getattr(obe_row, 'reviewed_by', None)
                                        latest_req.reviewed_at = getattr(obe_row, 'reviewed_at', None)
                                        latest_req.used_at = None
                                        latest_req.save(update_fields=['status', 'can_edit_until', 'reviewed_by', 'reviewed_at', 'used_at'])
                except Exception:
                    pass

                can_edit = (not locked) or _user_is_iqac_admin(user) or (latest_req.is_edit_granted() if latest_req else False)
                meta = {
                    'mode': 'SPECIAL_GLOBAL',
                    'selection_id': getattr(sel, 'id', None),
                    'locked': locked,
                    'can_edit': can_edit,
                    'edit_request': (
                        {
                            'id': latest_req.id,
                            'status': latest_req.status,
                            'can_edit_until': latest_req.can_edit_until,
                            'used_at': latest_req.used_at,
                        }
                        if latest_req else None
                    ),
                }
                return Response({'enabled_assessments': enabled, 'meta': meta})

            return Response({'enabled_assessments': getattr(ta, 'enabled_assessments', []), 'meta': meta})

        if not serializer_check_user_can_manage(user, ta):
            return Response({'detail': 'You do not have permission to change enabled assessments for this teaching assignment.'}, status=403)

        data = request.data or {}
        vals = data.get('enabled_assessments')
        if vals is None:
            return Response({'detail': 'enabled_assessments is required'}, status=400)
        if not isinstance(vals, (list, tuple)):
            return Response({'detail': 'enabled_assessments must be a list'}, status=400)

        cleaned = _clean_keys(vals, include_model=not _is_special_course_row(row))
        if _is_special_course_row(row):
            if not cleaned:
                return Response({'detail': 'At least one assessment is required for SPECIAL courses.'}, status=400)

            if row is None or getattr(row, 'master_id', None) is None:
                return Response({'detail': 'Unable to resolve the course for this SPECIAL teaching assignment.'}, status=400)

            master_id = row.master_id

            staff_profile = getattr(user, 'staff_profile', None)

            # Find an existing global selection for this master+academic_year.
            sel = (
                SpecialCourseAssessmentSelection.objects.filter(
                    curriculum_row__master_id=master_id,
                    academic_year=ta.academic_year,
                )
                .order_by('id')
                .first()
            )

            if sel is not None and sel.locked and not _user_is_iqac_admin(user):
                latest_req = None
                if staff_profile:
                    latest_req = (
                        SpecialCourseAssessmentEditRequest.objects.filter(
                            selection__curriculum_row__master_id=master_id,
                            selection__academic_year=ta.academic_year,
                            requested_by=staff_profile,
                        )
                        .order_by('-requested_at')
                        .first()
                    )
                if not (latest_req and latest_req.is_edit_granted()):
                    return Response(
                        {
                            'detail': 'Selection is locked for this SPECIAL course. Request IQAC approval to edit.',
                            'enabled_assessments': sel.enabled_assessments,
                            'meta': {
                                'mode': 'SPECIAL_GLOBAL',
                                'selection_id': sel.id,
                                'locked': True,
                                'can_edit': False,
                                'edit_request': (
                                    {
                                        'id': latest_req.id,
                                        'status': latest_req.status,
                                        'can_edit_until': latest_req.can_edit_until,
                                        'used_at': latest_req.used_at,
                                    }
                                    if latest_req else None
                                ),
                            },
                        },
                        status=423,
                    )

                # consume the approval after a successful edit
                try:
                    latest_req.used_at = timezone.now()
                    latest_req.save(update_fields=['used_at'])
                except Exception:
                    pass

            # If no selection exists yet, create it for ALL department rows of this master.
            if sel is None:
                try:
                    from curriculum.models import CurriculumDepartment

                    dept_rows = CurriculumDepartment.objects.filter(master_id=master_id)
                except Exception:
                    dept_rows = []

                created_sel = None
                for r in dept_rows:
                    obj, created = SpecialCourseAssessmentSelection.objects.get_or_create(
                        curriculum_row=r,
                        academic_year=ta.academic_year,
                        defaults={'enabled_assessments': cleaned, 'locked': True, 'created_by': staff_profile},
                    )
                    if created_sel is None:
                        created_sel = obj
                    if not created:
                        obj.enabled_assessments = cleaned
                        obj.locked = True
                        obj.save(update_fields=['enabled_assessments', 'locked', 'updated_at'])

                sel = created_sel

            # Update all selections under this master+year to keep it global.
            try:
                qs = SpecialCourseAssessmentSelection.objects.filter(curriculum_row__master_id=master_id, academic_year=ta.academic_year)
                for obj in qs:
                    obj.enabled_assessments = cleaned
                    obj.locked = True
                    obj.save(update_fields=['enabled_assessments', 'locked', 'updated_at'])
            except Exception as e:
                return Response({'detail': 'Failed to save enabled assessments', 'error': str(e)}, status=500)

            return Response(
                {
                    'enabled_assessments': cleaned,
                    'meta': {
                        'mode': 'SPECIAL_GLOBAL',
                        'selection_id': getattr(sel, 'id', None),
                        'locked': True,
                        'can_edit': _user_is_iqac_admin(user),
                    },
                }
            )

        ta.enabled_assessments = cleaned
        try:
            ta.save(update_fields=['enabled_assessments'])
        except Exception as e:
            return Response({'detail': 'Failed to save enabled assessments', 'error': str(e)}, status=500)

        return Response({'enabled_assessments': cleaned, 'meta': {'mode': 'TEACHING_ASSIGNMENT', 'locked': False, 'can_edit': True}})

    def enabled_assessments_request_edit(self, request, pk=None):
        """Faculty: request IQAC approval to edit SPECIAL_GLOBAL enabled assessments.

        Endpoint: POST /api/academics/teaching-assignments/<pk>/enabled_assessments/request-edit/

        Creates a SpecialCourseAssessmentEditRequest (or returns existing pending/active approval)
        and mirrors it into the central OBE edit queue (ObeEditRequest) so IQAC UIs can review.
        """
        try:
            ta = TeachingAssignment.objects.select_related(
                'section',
                'section__batch',
                'section__batch__course',
                'section__batch__course__department',
                'academic_year',
                'curriculum_row',
                'curriculum_row__master',
                'subject',
                'staff',
            ).get(pk=int(pk), is_active=True)
        except TeachingAssignment.DoesNotExist:
            return Response({'detail': 'Teaching assignment not found'}, status=404)

        staff_profile = getattr(request.user, 'staff_profile', None)
        if not staff_profile:
            return Response({'detail': 'Staff profile not found'}, status=403)

        # Best-effort resolve CurriculumDepartment row.
        row = getattr(ta, 'curriculum_row', None)
        if row is None:
            try:
                from curriculum.models import CurriculumDepartment
                from django.db.models import Q

                dept = None
                try:
                    dept = ta.section.batch.course.department
                except Exception:
                    dept = None

                subj = getattr(ta, 'subject', None)
                code = getattr(subj, 'code', None)
                name = getattr(subj, 'name', None)

                qs = CurriculumDepartment.objects.all().select_related('master', 'department')
                if dept is not None:
                    qs = qs.filter(department=dept)

                if code:
                    qs = qs.filter(Q(course_code__iexact=str(code).strip()) | Q(master__course_code__iexact=str(code).strip()))
                elif name:
                    qs = qs.filter(Q(course_name__iexact=str(name).strip()) | Q(master__course_name__iexact=str(name).strip()))
                else:
                    qs = CurriculumDepartment.objects.none()

                row = qs.order_by('-updated_at', '-id').first()
            except Exception:
                row = None

        if not row:
            return Response({'detail': 'Unable to resolve curriculum row for this teaching assignment.'}, status=400)

        # Ensure SPECIAL course behavior.
        try:
            ct = getattr(row, 'class_type', None) or getattr(getattr(row, 'master', None), 'class_type', None)
            is_special = str(ct or '').upper() == 'SPECIAL'
        except Exception:
            is_special = False

        if not is_special:
            return Response({'detail': 'Edit approval requests are only supported for SPECIAL courses.'}, status=400)

        master_id = getattr(row, 'master_id', None)
        if master_id is None:
            return Response({'detail': 'Unable to resolve course master for this selection.'}, status=400)

        sel = (
            SpecialCourseAssessmentSelection.objects.filter(
                curriculum_row__master_id=master_id,
                academic_year=ta.academic_year,
            )
            .select_related('curriculum_row')
            .order_by('id')
            .first()
        )

        # If nothing is locked yet, there is nothing to request.
        if sel is None or not bool(getattr(sel, 'locked', False)):
            return Response({'detail': 'Selection is not locked yet; no approval needed.'}, status=400)

        existing = (
            SpecialCourseAssessmentEditRequest.objects.filter(
                selection__curriculum_row__master_id=master_id,
                selection__academic_year=ta.academic_year,
                requested_by=staff_profile,
            )
            .select_related('selection')
            .order_by('-requested_at')
            .first()
        )

        def _ensure_obe_backlink():
            """Best-effort mirror into OBE edit-request queue for IQAC screens."""
            try:
                from OBE.models import ObeEditRequest
            except Exception:
                import logging
                logging.getLogger(__name__).exception('Failed to import OBE.models for SPECIAL edit-request backlink')
                return

            staff_user = getattr(staff_profile, 'user', None)
            if staff_user is None:
                return

            subject_code = ''
            subject_name = ''
            try:
                subject_code = (
                    getattr(row, 'course_code', None)
                    or getattr(getattr(row, 'master', None), 'course_code', None)
                    or getattr(getattr(ta, 'subject', None), 'code', None)
                    or getattr(getattr(ta, 'curriculum_row', None), 'course_code', None)
                    or ''
                )
            except Exception:
                subject_code = ''
            try:
                subject_name = (
                    getattr(row, 'course_name', None)
                    or getattr(getattr(row, 'master', None), 'course_name', None)
                    or getattr(getattr(ta, 'subject', None), 'name', None)
                    or ''
                )
            except Exception:
                subject_name = ''

            if not subject_code:
                # If we couldn't resolve a meaningful subject code from the curriculum
                # row or subject, fall back to a stable teaching-assignment based code.
                try:
                    subject_code = (
                        getattr(getattr(ta, 'subject', None), 'code', None)
                        or getattr(getattr(ta, 'curriculum_row', None), 'course_code', None)
                        or f"TA-{getattr(ta, 'id', '')}"
                        or ''
                    )
                except Exception:
                    subject_code = f"TA-{getattr(ta, 'id', '')}"

            if not subject_code:
                return

            section_name = ''
            try:
                section_name = getattr(getattr(ta, 'section', None), 'name', None) or str(getattr(ta, 'section', ''))
            except Exception:
                section_name = ''

            try:
                # Always create a new pending OBE edit request so faculty can re-request
                # multiple times without hitting a limit. IQAC will see each request instantly.
                ObeEditRequest.objects.create(
                    staff_user=staff_user,
                    academic_year=ta.academic_year,
                    subject_code=subject_code,
                    subject_name=subject_name or '',
                    assessment='model',
                    scope='MARK_MANAGER',
                    reason='Edit request: enabled assessments (SPECIAL course global selection)',
                    teaching_assignment=ta,
                    section_name=section_name or '',
                )
            except Exception:
                import logging
                logging.getLogger(__name__).exception('Failed to create ObeEditRequest backlink for subject_code=%s ta=%s', subject_code, getattr(ta, 'id', None))
                return

        if existing:
            # If an existing request is currently granted (approved and within window),
            # return it unchanged so the caller knows edit access is active.
            if existing.is_edit_granted():
                _ensure_obe_backlink()
                ser = SpecialCourseAssessmentEditRequestSerializer(existing)
                return Response(ser.data)

            # If a pending request exists, treat a new request as a "re-send".
            # Update the requested timestamp and mirror into the OBE backlink,
            # then return with 201 so frontends display a fresh "sent" response.
            if existing.status == SpecialCourseAssessmentEditRequest.STATUS_PENDING:
                try:
                    existing.requested_at = timezone.now()
                    existing.save(update_fields=['requested_at'])
                except Exception:
                    pass
                _ensure_obe_backlink()
                ser = SpecialCourseAssessmentEditRequestSerializer(existing)
                return Response(ser.data, status=201)

        req = SpecialCourseAssessmentEditRequest.objects.create(selection=sel, requested_by=staff_profile)

        _ensure_obe_backlink()

        ser = SpecialCourseAssessmentEditRequestSerializer(req)
        return Response(ser.data, status=201)

    def special_qp_pattern(self, request, pk=None):
        """GET/POST custom question paper pattern for SPECIAL courses.

        GET  ?exam=cia1  -> { teaching_assignment_id, exam, pattern }
        POST { exam, pattern: { questions: [{ key, label, max, co, btl }] } }
        """
        try:
            ta = TeachingAssignment.objects.get(pk=int(pk), is_active=True)
        except TeachingAssignment.DoesNotExist:
            return Response({'detail': 'Teaching assignment not found'}, status=404)

        from OBE.models import SpecialCourseQpPattern

        if request.method == 'GET':
            exam = str(request.query_params.get('exam', '')).strip().lower()
            if not exam:
                return Response({'detail': 'exam query parameter required'}, status=400)
            obj = SpecialCourseQpPattern.objects.filter(teaching_assignment=ta, exam=exam).first()
            if obj is None:
                return Response({'teaching_assignment_id': ta.id, 'exam': exam, 'pattern': None})
            return Response({
                'teaching_assignment_id': ta.id,
                'exam': obj.exam,
                'pattern': obj.pattern,
            })

        # POST
        exam = str(request.data.get('exam', '')).strip().lower()
        if not exam:
            return Response({'detail': 'exam field required'}, status=400)

        pattern = request.data.get('pattern', {})
        questions = pattern.get('questions', []) if isinstance(pattern, dict) else []
        if not isinstance(questions, list) or not questions:
            return Response({'detail': 'pattern.questions must be a non-empty list'}, status=400)

        # Validate each question
        clean_qs = []
        for i, q in enumerate(questions):
            if not isinstance(q, dict):
                return Response({'detail': f'Question {i+1} must be an object'}, status=400)
            max_marks = q.get('max', 0)
            try:
                max_marks = int(max_marks)
            except (TypeError, ValueError):
                max_marks = 0
            if max_marks <= 0:
                return Response({'detail': f'Question {i+1}: max marks must be > 0'}, status=400)
            clean_qs.append({
                'key': str(q.get('key', f'q{i+1}')).strip() or f'q{i+1}',
                'label': str(q.get('label', f'Q{i+1}')).strip() or f'Q{i+1}',
                'max': max_marks,
                'co': q.get('co', 1),
                'btl': min(6, max(1, int(q.get('btl', 1) or 1))),
            })

        obj, _ = SpecialCourseQpPattern.objects.update_or_create(
            teaching_assignment=ta,
            exam=exam,
            defaults={
                'pattern': {'questions': clean_qs},
                'updated_by': request.user.id,
            },
        )
        return Response({
            'teaching_assignment_id': ta.id,
            'exam': obj.exam,
            'pattern': obj.pattern,
        })


    def special_co_weights(self, request, pk=None):
        """GET/POST CO attainment weights for a SPECIAL course teaching assignment.

        GET  -> { teaching_assignment_id, weights: { co1: 20.0, co2: 20.0, ... } }
        POST { weights: { co1: 20.0, co2: 20.0, ... } }
        """
        try:
            ta = TeachingAssignment.objects.get(pk=int(pk), is_active=True)
        except TeachingAssignment.DoesNotExist:
            return Response({'detail': 'Teaching assignment not found'}, status=404)

        from OBE.models import SpecialCourseCoWeights

        if request.method == 'GET':
            obj = SpecialCourseCoWeights.objects.filter(teaching_assignment=ta).first()
            return Response({
                'teaching_assignment_id': ta.id,
                'weights': obj.weights if obj else {},
            })

        # POST
        raw = request.data.get('weights', {})
        if not isinstance(raw, dict):
            return Response({'detail': 'weights must be an object'}, status=400)

        clean = {}
        for co_key in ['co1', 'co2', 'co3', 'co4', 'co5']:
            if co_key in raw:
                try:
                    clean[co_key] = float(raw[co_key])
                except (TypeError, ValueError):
                    return Response({'detail': f'{co_key}: must be a number'}, status=400)

        obj, _ = SpecialCourseCoWeights.objects.update_or_create(
            teaching_assignment=ta,
            defaults={'weights': clean, 'updated_by': request.user.id},
        )
        return Response({
            'teaching_assignment_id': ta.id,
            'weights': obj.weights,
        })


class SpecialCourseEnabledAssessmentsView(APIView):
    """Fetch SPECIAL-course global enabled assessments for a course code.

    This reads from `SpecialCourseAssessmentSelection` (global lock) and is used by
    course-level pages (IQAC / course OBE pages) so the UI doesn't rely on stale
    curriculum-row enabled_assessments.

    GET /api/academics/special-courses/<course_code>/enabled_assessments/
      Optional query: ?academic_year_id=<id>
    """

    permission_classes = (IsAuthenticated,)

    def get(self, request, course_code: str):
        code = str(course_code or '').strip()
        if not code:
            return Response({'detail': 'course_code is required.'}, status=400)

        ay_id = request.query_params.get('academic_year_id')
        academic_year = None
        if ay_id:
            try:
                academic_year = AcademicYear.objects.filter(id=int(str(ay_id))).first()
            except Exception:
                academic_year = None
        if academic_year is None:
            academic_year = AcademicYear.objects.filter(is_active=True).order_by('-id').first()

        if academic_year is None:
            return Response({'detail': 'No active academic year found.'}, status=404)

        try:
            from django.db.models import Q
            from .models import SpecialCourseAssessmentSelection

            sel = (
                SpecialCourseAssessmentSelection.objects.filter(academic_year=academic_year)
                .filter(Q(curriculum_row__course_code__iexact=code) | Q(curriculum_row__master__course_code__iexact=code))
                .order_by('id')
                .first()
            )
        except Exception:
            sel = None

        enabled = []
        if sel is not None:
            try:
                enabled = list(getattr(sel, 'enabled_assessments', None) or [])
            except Exception:
                enabled = []

        return Response(
            {
                'course_code': code,
                'academic_year_id': getattr(academic_year, 'id', None),
                'selection_id': getattr(sel, 'id', None),
                'locked': bool(getattr(sel, 'locked', False)) if sel is not None else False,
                'enabled_assessments': enabled,
            }
        )


class SpecialCourseAssessmentEditRequestViewSet(viewsets.ReadOnlyModelViewSet):
    """Read-only viewset for SpecialCourseAssessmentEditRequest.  IQAC can list/review."""

    serializer_class = SpecialCourseAssessmentEditRequestSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        from .models import SpecialCourseAssessmentEditRequest
        user = self.request.user
        try:
            if user.staff_profile.role in ('IQAC', 'HOD', 'PRINCIPAL'):
                return SpecialCourseAssessmentEditRequest.objects.all().order_by('-requested_at')
        except Exception:
            pass
        # Faculty: own requests only
        try:
            return SpecialCourseAssessmentEditRequest.objects.filter(
                requested_by=user.staff_profile
            ).order_by('-requested_at')
        except Exception:
            return SpecialCourseAssessmentEditRequest.objects.none()

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        from .models import SpecialCourseAssessmentEditRequest
        from django.utils import timezone
        import datetime
        try:
            obj = self.get_object()
        except Exception as e:
            return Response({'detail': str(e)}, status=404)
        obj.status = SpecialCourseAssessmentEditRequest.STATUS_APPROVED
        obj.reviewed_by = request.user.staff_profile if hasattr(request.user, 'staff_profile') else None
        obj.reviewed_at = timezone.now()
        hours = int(request.data.get('hours', 24))
        obj.can_edit_until = timezone.now() + datetime.timedelta(hours=max(1, hours))
        obj.save()
        return Response(self.get_serializer(obj).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        from .models import SpecialCourseAssessmentEditRequest
        from django.utils import timezone
        try:
            obj = self.get_object()
        except Exception as e:
            return Response({'detail': str(e)}, status=404)
        obj.status = SpecialCourseAssessmentEditRequest.STATUS_REJECTED
        obj.reviewed_by = request.user.staff_profile if hasattr(request.user, 'staff_profile') else None
        obj.reviewed_at = timezone.now()
        obj.save()
        return Response(self.get_serializer(obj).data)


class AcademicYearViewSet(viewsets.ModelViewSet):
    """Manage AcademicYear objects: create, list, activate/deactivate."""
    queryset = AcademicYear.objects.all().order_by('-id')
    serializer_class = AcademicYearSerializer
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        user = self.request.user
        perms = get_user_permissions(user)
        if not (user.is_staff or 'academics.manage_academicyears' in perms or user.has_perm('academics.add_academicyear')):
            raise PermissionDenied('You do not have permission to create academic years.')
        serializer.save()

    def perform_update(self, serializer):
        user = self.request.user
        perms = get_user_permissions(user)
        if not (user.is_staff or 'academics.manage_academicyears' in perms or user.has_perm('academics.change_academicyear')):
            raise PermissionDenied('You do not have permission to change academic years.')
        serializer.save()

    def perform_destroy(self, instance):
        user = self.request.user
        perms = get_user_permissions(user)
        if not (user.is_staff or 'academics.manage_academicyears' in perms or user.has_perm('academics.delete_academicyear')):
            raise PermissionDenied('You do not have permission to delete academic years.')
        instance.delete()

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except PermissionDenied:
            raise
        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).exception('Error creating TeachingAssignment: %s', e)
            tb = traceback.format_exc()
            return Response({'detail': 'Failed to create teaching assignment.', 'error': str(e), 'trace': tb}, status=status.HTTP_400_BAD_REQUEST)

    def perform_update(self, serializer):
        user = self.request.user
        perms = get_user_permissions(user)

        # Allow if explicit change permission
        if ('academics.change_teaching' in perms) or user.has_perm('academics.change_teachingassignment'):
            serializer.save()
            return

        # Determine whether this is for an elective or regular subject
        is_elective = False
        ta = getattr(serializer, 'instance', None)
        try:
            if 'elective_subject_id' in getattr(serializer, 'initial_data', {}) or 'elective_subject' in getattr(serializer, 'validated_data', {}):
                is_elective = True
            elif ta and getattr(ta, 'elective_subject', None):
                is_elective = True
        except Exception:
            is_elective = False

        # Elective: require elective change permission or HOD of parent dept
        if is_elective:
            if ('academics.change_elective_teaching' in perms) or user.has_perm('academics.change_elective_teaching'):
                serializer.save()
                return
            # check HOD of elective parent department
            try:
                es = None
                if 'elective_subject_id' in getattr(serializer, 'initial_data', {}):
                    from curriculum.models import ElectiveSubject
                    es = ElectiveSubject.objects.filter(pk=int(serializer.initial_data.get('elective_subject_id'))).select_related('parent__department').first()
                elif ta:
                    es = getattr(ta, 'elective_subject', None)
                parent_dept_id = getattr(getattr(es, 'parent', None), 'department_id', None)
                es_dept_id = getattr(es, 'department_id', None)
                staff_profile = getattr(user, 'staff_profile', None)
                if staff_profile:
                    hod_depts = list(DepartmentRole.objects.filter(staff=staff_profile, role='HOD', is_active=True).values_list('department_id', flat=True))
                    # HOD of the parent dept (normal elective) OR HOD of the variant's own dept
                    # (dept-core: parent is S&H but variant belongs to AI&DS HOD etc.)
                    if (parent_dept_id and parent_dept_id in hod_depts) or (es_dept_id and es_dept_id in hod_depts):
                        serializer.save()
                        return
            except Exception:
                pass
            raise PermissionDenied('You do not have permission to change this elective teaching assignment.')

        # Regular subject: advisor for section required
        staff_profile = getattr(user, 'staff_profile', None)
        section_obj = None
        try:
            if 'section' in getattr(serializer, 'validated_data', {}):
                section_obj = serializer.validated_data.get('section')
            elif ta is not None:
                section_obj = getattr(ta, 'section', None)
        except Exception:
            section_obj = getattr(ta, 'section', None) if ta is not None else None

        if not section_obj:
            raise PermissionDenied('You do not have permission to change this teaching assignment.')

        is_advisor = SectionAdvisor.objects.filter(section=section_obj, advisor=staff_profile, is_active=True, academic_year__is_active=True).exists() if staff_profile else False
        if not is_advisor:
            raise PermissionDenied('You do not have permission to change this teaching assignment.')

        serializer.save()


class HODSectionsView(APIView):
    permission_classes = (IsAuthenticated, IsHODOfDepartment)
    def get(self, request):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return Response({'results': []})

        dept_ids = get_user_effective_departments(user) or []
        if not dept_ids:
            return Response({'results': []})

        # Primary own sections (batch belongs to the HOD's department(s))
        own_section_ids = set(Section.objects.filter(
            Q(batch__course__department_id__in=dept_ids) |
            Q(batch__department_id__in=dept_ids) |
            Q(managing_department_id__in=dept_ids)
        ).values_list('pk', flat=True))

        # Secondary: sections where Year-1 students have SECONDARY assignments belonging
        # to this HOD's department (e.g. AI&DS A / B for Year-1 dept-core periods).
        from django.db.models import Exists, OuterRef
        from .models import StudentSectionAssignment as _SSA
        has_secondary_from_dept = _SSA.objects.filter(
            section_id=OuterRef('pk'),
            end_date__isnull=True,
            section_type='SECONDARY',
            student__home_department_id__in=dept_ids,
        )
        secondary_section_ids = set(Section.objects.filter(
            Exists(has_secondary_from_dept)
        ).values_list('pk', flat=True))

        all_section_ids = own_section_ids | secondary_section_ids

        sections = Section.objects.filter(
            pk__in=all_section_ids
        ).select_related(
            'batch__course__department', 'batch__department', 'batch__regulation', 'semester'
        ).order_by('batch__name', 'name')

        results = []
        for s in sections:
            batch = getattr(s, 'batch', None)
            course = getattr(batch, 'course', None) if batch else None
            # For S&H dept-only batches course is None; fall back to batch.department
            # then to section.managing_department
            dept = (
                getattr(course, 'department', None)
                if course is not None
                else (getattr(batch, 'department', None) if batch else None)
                    or getattr(s, 'managing_department', None)
            )
            reg = getattr(batch, 'regulation', None) if batch else None
            sem_obj = getattr(s, 'semester', None)
            sem_val = getattr(sem_obj, 'number', None) if sem_obj else None
            results.append({
                'id': s.id,
                'name': s.name,
                'batch_id': getattr(batch, 'id', None),
                'batch_name': getattr(batch, 'name', None),
                'batch_regulation': {'id': getattr(reg, 'id', None), 'code': getattr(reg, 'code', None), 'name': getattr(reg, 'name', None)} if reg else None,
                'course_id': getattr(course, 'id', None),
                'department_id': getattr(dept, 'id', None),
                'department_code': getattr(dept, 'code', None),
                'department_short_name': getattr(dept, 'short_name', None),
                'semester': sem_val,
            })
        return Response({'results': results})


class SectionsByDeptYearView(APIView):
    """Return sections filtered by departments and years for feedback creation.

    Response items are simple objects of the form:
        {"id": <section_id>, "label": "DEPT - A - Year 3", "year": 3, "department_short_name": "DEPT"}
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        from .models import Section, AcademicYear

        dept_ids_param = request.GET.get('dept_ids', '')
        years_param = request.GET.get('years', '')

        try:
            dept_ids = [int(v) for v in dept_ids_param.split(',') if v.strip()]
        except ValueError:
            dept_ids = []

        try:
            years = [int(v) for v in years_param.split(',') if v.strip()]
        except ValueError:
            years = []

        qs = Section.objects.select_related(
            'batch',
            'batch__course__department',
            'batch__department',
            'managing_department',
        )

        if dept_ids:
            qs = qs.filter(
                Q(batch__course__department_id__in=dept_ids) |
                Q(batch__department_id__in=dept_ids) |
                Q(managing_department_id__in=dept_ids)
            )

        current_acad_year = None
        active_ay = AcademicYear.objects.filter(is_active=True).first()
        if active_ay:
            try:
                current_acad_year = int(str(active_ay.name).split('-')[0])
            except Exception:
                current_acad_year = None

        results = []

        for sec in qs:
            batch = getattr(sec, 'batch', None)
            course = getattr(batch, 'course', None) if batch else None

            dept = (
                getattr(course, 'department', None)
                if course is not None
                else getattr(batch, 'department', None) if batch else None
            ) or getattr(sec, 'managing_department', None)

            dept_short = None
            if dept is not None:
                dept_short = getattr(dept, 'short_name', None) or getattr(dept, 'code', None) or getattr(dept, 'name', None)

            student_year = None
            if batch and getattr(batch, 'start_year', None) and current_acad_year:
                try:
                    delta = current_acad_year - int(batch.start_year)
                    student_year = delta + 1
                except Exception:
                    student_year = None

            if student_year is not None:
                if student_year < 1 or student_year > 4:
                    continue
                if years and student_year not in years:
                    continue
            elif years:
                # When year filter is supplied but we cannot compute year, skip.
                continue

            if dept_ids and dept is None:
                continue

            label_parts = []
            if dept_short:
                label_parts.append(dept_short)
            label_parts.append(str(sec.name))
            if student_year is not None:
                label_parts.append(f"Year {student_year}")
            # Use plain ASCII hyphens to avoid escaped unicode sequences like "\\u2013" in responses
            label = " - ".join(label_parts)

            results.append({
                'id': sec.id,
                'label': label,
                'year': student_year,
                'department_short_name': dept_short,
            })

        results.sort(key=lambda item: (
            item.get('department_short_name') or '',
            item.get('year') or 0,
            item.get('label') or '',
        ))

        return Response(results)


class HODStaffListView(APIView):
    permission_classes = (IsAuthenticated, IsHODOfDepartment)

    def get(self, request):
        # Return staff list limited to the HOD's departments
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return Response({'results': []})
        perms = get_user_permissions(user)
        dept_ids = get_user_effective_departments(user)

        # optionally allow department param; for globally-authorized users this
        # can be ANY department, otherwise it must be one of the user's
        # effective departments.
        dept_param = request.query_params.get('department')
        try:
            dept_filter = int(dept_param) if dept_param else None
        except Exception:
            dept_filter = None

        global_access = (
            user.is_superuser
            or ('academics.view_all_staff' in perms)
            or ('academics.view_all_departments' in perms)
            or ('academics.assign_advisor' in perms)
        )

        if global_access:
            from .models import Department
            if dept_filter:
                dept_ids = [dept_filter]
            else:
                dept_ids = list(Department.objects.values_list('id', flat=True))
        else:
            if dept_filter:
                if dept_filter not in (dept_ids or []):
                    return Response({'results': []})
                dept_ids = [dept_filter]
        # include staff whose `department` FK matches OR who have an active
        # StaffDepartmentAssignment pointing to the department
        from django.db.models import Q
        staff_qs = StaffProfile.objects.filter(
            Q(department_id__in=dept_ids) |
            Q(department_assignments__department_id__in=dept_ids, department_assignments__end_date__isnull=True)
        ).select_related('user').distinct()
        results = []
        for s in staff_qs:
            # Return the full name of the staff member, formatted as "FirstName LastName"
            full_name = None
            if s.user:
                full_name = s.user.get_full_name() or s.user.username
            results.append({
                'id': s.id, 
                'user': full_name,  # Full name instead of just username
                'staff_id': s.staff_id, 
                'department': getattr(s.department, 'id', None)
            })
        return Response({'results': results})


class SemesterViewSet(viewsets.ReadOnlyModelViewSet):
    """Read-only ViewSet for Semester objects."""
    queryset = Semester.objects.all().order_by('number')
    permission_classes = (IsAuthenticated,)

    def get_serializer_class(self):
        from .serializers import SemesterSerializer
        return SemesterSerializer


class DepartmentsListView(APIView):

    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        perms = set(get_user_permissions(user))
        from .models import Department

        has_ps_role = user.roles.filter(name__iexact='PS').exists()

        has_coe_role = False
        try:
            has_coe_role = bool(user.roles.filter(name__iexact='COE').exists())
        except Exception:
            has_coe_role = False

        has_global_access = (
            bool({'academics.view_all_departments', 'academics.view_all_staff'} & perms)
            or bool(COE_DEPARTMENT_ACCESS_PERMS & perms)
            or has_coe_role
            or has_ps_role
            or user.is_superuser
        )

        # accept either view_all_departments or view_all_staff permission as global access
        if has_global_access:
            # Filter for teaching departments only (parent is NULL)
            qs = Department.objects.filter(parent__isnull=True)
        else:
            # Use effective departments (own dept + DepartmentRole HOD/AHOD mappings)
            dept_ids = get_user_effective_departments(user) or []
            if not dept_ids:
                return Response({'results': []})

            # Determine which of the effective departments are themselves teaching departments (parent NULL)
            teaching_ids = set(
                Department.objects.filter(id__in=dept_ids, parent__isnull=True).values_list('id', flat=True)
            )

            # For non-teaching departments, include their parent (the actual teaching department)
            parent_ids = set(
                Department.objects.filter(id__in=dept_ids, parent__isnull=False).values_list('parent_id', flat=True)
            )

            allowed_ids = list(teaching_ids | parent_ids)
            if not allowed_ids:
                return Response({'results': []})

            # Filter for teaching departments only (parent is NULL)
            qs = Department.objects.filter(id__in=allowed_ids, parent__isnull=True)

        include_non_teaching = str(request.query_params.get('include_non_teaching', 'false')).strip().lower() in {'1', 'true', 'yes'}
        can_include_non_teaching = bool(
            user.is_superuser
            or has_ps_role
            or ({'academics.edit_staff', 'academics.view_all_departments', 'academics.view_all_staff'} & perms)
        )
        if not (include_non_teaching and can_include_non_teaching):
            qs = qs.filter(is_teaching=True)

        results = []
        for d in qs:
            results.append({'id': d.id, 'code': getattr(d, 'code', None), 'name': getattr(d, 'name', None), 'short_name': getattr(d, 'short_name', None)})
        return Response({'results': results})


class StaffsPageView(APIView):
    """Return departments along with their staffs according to user's permissions.

    - Requires explicit `academics.view_staffs_page` permission (or staff/superuser).
    - If user has `academics.view_all_staff` (or is staff/superuser), returns all departments.
    - Otherwise HODs and other staff see ONLY their own primary department (not effective departments).

    Response format:
      { results: [ { id, code, name, short_name, staffs: [ {id, staff_id, user: {username, first_name, last_name}, designation, status} ] } ] }
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        has_ps_role = user.roles.filter(name__iexact='PS').exists()

        # require page-view permission unless superuser
        if not (user.is_superuser or has_ps_role or 'academics.view_staffs_page' in perms):
            return Response({'detail': 'You do not have permission to view staffs page.'}, status=403)

        from .models import Department, StaffProfile
        from django.db.models import Q
        import logging
        logger = logging.getLogger(__name__)

        # Debug: Log user permissions
        logger.info(f"StaffsPage - User: {user.username}, Superuser: {user.is_superuser}")
        logger.info(f"StaffsPage - Has PS role: {has_ps_role}")
        logger.info(f"StaffsPage - Permissions: {perms}")
        logger.info(f"StaffsPage - Has view_all_staff: {'academics.view_all_staff' in perms}")
        logger.info(f"StaffsPage - Has edit_staff: {'academics.edit_staff' in perms}")

        # Check if user can edit staff
        can_edit = user.is_superuser or has_ps_role or 'academics.edit_staff' in perms
        
        # Check if user can view all staff (determines if role filter should be shown)
        can_view_all = user.is_superuser or has_ps_role or 'academics.view_all_staff' in perms

        include_non_teaching = str(request.query_params.get('include_non_teaching', 'false')).strip().lower() in {'1', 'true', 'yes'}
        can_include_non_teaching = bool(user.is_superuser or has_ps_role or can_edit or can_view_all)

        # determine departments to include
        if user.is_superuser or has_ps_role or ('academics.view_all_staff' in perms):
            # View all departments
            logger.info(f"StaffsPage - Showing ALL departments (superuser or has view_all_staff)")
            dept_qs = Department.objects.all()
        else:
            # HODs and other staff: show departments they are affiliated with
            # This includes their primary department AND any HOD/AHOD mapped departments
            from academics.utils import get_user_effective_departments
            
            dept_ids = get_user_effective_departments(user)
            if not dept_ids:
                logger.info(f"StaffsPage - No effective departments found, returning empty")
                return Response({'results': []})
            
            logger.info(f"StaffsPage - Showing departments: {dept_ids} (includes primary + HOD mappings)")
            dept_qs = Department.objects.filter(id__in=dept_ids)

        if not (include_non_teaching and can_include_non_teaching):
            dept_qs = dept_qs.filter(is_teaching=True)

        results = []
        for d in dept_qs.order_by('code'):
            # collect staff for department (current FK or active department assignment OR active department role)
            staff_qs = StaffProfile.objects.filter(
                Q(department_id=d.id) |
                Q(department_assignments__department_id=d.id, department_assignments__end_date__isnull=True) |
                Q(department_roles__department_id=d.id, department_roles__is_active=True)
            ).select_related('user').distinct()

            staffs = []
            for s in staff_qs:
                user_data = None
                user_roles = []
                if s.user:
                    user_data = {
                        'username': s.user.username,
                        'first_name': getattr(s.user, 'first_name', ''),
                        'last_name': getattr(s.user, 'last_name', ''),
                        'email': getattr(s.user, 'email', ''),
                    }
                    # Get user roles
                    try:
                        user_roles = [r.name for r in s.user.roles.all()]
                    except Exception:
                        user_roles = []
                
                # Get department roles (HOD, AHOD, etc.)
                department_roles = []
                dept_role_mappings = []
                try:
                    from .models import DepartmentRole
                    dept_roles = DepartmentRole.objects.filter(
                        staff=s, 
                        is_active=True
                    ).select_related('department', 'academic_year')
                    
                    for dept_role in dept_roles:
                        role_display = dept_role.get_role_display()
                        department_roles.append(role_display)
                        dept_role_mappings.append({
                            'department': {
                                'id': dept_role.department.id,
                                'code': dept_role.department.code,
                                'name': dept_role.department.name,
                                'short_name': dept_role.department.short_name,
                            },
                            'role': role_display,
                            'role_code': dept_role.role,
                            'academic_year': dept_role.academic_year.name if dept_role.academic_year else None,
                        })
                except Exception:
                    department_roles = []
                    dept_role_mappings = []
                
                # Combine all roles for filtering
                all_roles = user_roles + department_roles
                
                staffs.append({
                    'id': s.id,
                    'staff_id': s.staff_id,
                    'internal_id': s.internal_id,
                    'user': user_data,
                    'user_id': s.user.id if s.user else None,
                    'designation': getattr(s, 'designation', None),
                    'status': getattr(s, 'status', None),
                    'department': s.department_id if s.department_id else None,
                    'roles': all_roles,  # Combined user roles + department roles
                    'user_roles': user_roles,  # Original user-assigned roles
                    'department_roles': department_roles,  # Department-specific roles
                    'department_role_mappings': dept_role_mappings,  # Detailed role-department mapping
                })

            results.append({
                'id': d.id,
                'code': getattr(d, 'code', None),
                'name': getattr(d, 'name', None),
                'short_name': getattr(d, 'short_name', None),
                'staffs': staffs,
            })

        # Check if user can import staff (HOD, AHOD, or IQAC role required)
        can_import = False
        if user.is_superuser or has_ps_role:
            can_import = True
        else:
            try:
                if user.roles.filter(name__iexact='IQAC').exists():
                    can_import = True
            except Exception:
                pass
            if not can_import:
                from .models import DepartmentRole
                from .utils import get_user_staff_profile
                sp = get_user_staff_profile(user)
                if sp and DepartmentRole.objects.filter(staff=sp, role__in=['HOD', 'AHOD'], is_active=True).exists():
                    can_import = True

        return Response({'results': results, 'can_edit': can_edit, 'can_view_all': can_view_all, 'can_import': can_import})


class DepartmentStaffListView(APIView):
    """Return staff from the same department for attendance swap purposes.
    
    Returns staff members from the requesting user's department who can be assigned to take attendance.
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        
        # Get the staff profile of the requesting user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return Response({'detail': 'You must be a staff member to access this endpoint.'}, status=403)
        
        # Get the user's department
        current_dept = staff_profile.get_current_department()
        if not current_dept:
            return Response({'results': []})
        
        from .models import StaffProfile
        
        # Get all active staff from the same department, excluding the requesting user
        staff_qs = StaffProfile.objects.filter(
            department=current_dept,
            status='ACTIVE'
        ).exclude(id=staff_profile.id).select_related('user').order_by('user__first_name', 'user__last_name')
        
        results = []
        for s in staff_qs:
            results.append({
                'id': s.id,
                'staff_id': s.staff_id,
                'name': f"{s.user.first_name} {s.user.last_name}".strip() or s.user.username,
                'username': s.user.username,
                'designation': s.designation or '',
            })
        
        return Response({'results': results})


class BatchStaffListView(APIView):
    """Return all staff members from all departments for batch creation.
    
    Used when creating student batches - allows selecting staff from any department.
    Supports optional department filtering via ?department_id=<id> parameter.
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        from .models import StaffProfile
        from django.db.models import Q
        
        user = request.user
        
        # Require staff profile
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return Response({'detail': 'You must be a staff member to access this endpoint.'}, status=403)
        
        # Optional department filter
        dept_id = request.query_params.get('department_id')
        
        # Start with all active staff
        staff_qs = StaffProfile.objects.filter(status='ACTIVE').select_related('user', 'department').order_by('user__first_name', 'user__last_name')
        
        # Apply department filter if provided
        if dept_id:
            try:
                dept_id = int(dept_id)
                staff_qs = staff_qs.filter(
                    Q(department_id=dept_id) |
                    Q(department_assignments__department_id=dept_id, department_assignments__end_date__isnull=True)
                ).distinct()
            except (ValueError, TypeError):
                pass
        
        results = []
        for s in staff_qs:
            user_info = None
            if s.user:
                user_info = {
                    'username': s.user.username,
                    'first_name': getattr(s.user, 'first_name', ''),
                    'last_name': getattr(s.user, 'last_name', ''),
                }
            
            dept_info = None
            if s.department:
                dept_info = {
                    'id': s.department.id,
                    'name': s.department.name,
                    'code': getattr(s.department, 'code', ''),
                    'short_name': getattr(s.department, 'short_name', ''),
                }
            
            results.append({
                'id': s.id,
                'staff_id': s.staff_id,
                'user': user_info,
                'name': f"{s.user.first_name} {s.user.last_name}".strip() or s.user.username if s.user else s.staff_id,
                'username': s.user.username if s.user else '',
                'designation': s.designation or '',
                'department': dept_info
            })
        
        return Response({'results': results})


class StaffProfileCreateView(APIView):
    """Create a new staff profile with user account."""
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        from .serializers import StaffProfileSerializer
        from accounts.utils import get_user_permissions
        
        user = request.user
        perms = get_user_permissions(user)
        has_ps_role = user.roles.filter(name__iexact='PS').exists()
        
        # Check permission - require edit_staff or superuser to create staff
        if not (user.is_superuser or has_ps_role or 'academics.edit_staff' in perms):
            return Response(
                {'detail': 'You do not have permission to create staff profiles.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Validate department scope for non-superusers
        if not user.is_superuser and not has_ps_role and 'academics.view_all_staff' not in perms:
            # HODs can only create staff in departments they are mapped to
            from academics.utils import get_user_effective_departments
            
            allowed_dept_ids = get_user_effective_departments(user)
            if allowed_dept_ids:
                requested_dept = request.data.get('department')
                if requested_dept and int(requested_dept) not in allowed_dept_ids:
                    return Response(
                        {'detail': 'You can only create staff in departments you manage.'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                return Response(
                    {'detail': 'You are not mapped to any departments.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        serializer = StaffProfileSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class StaffProfileUpdateView(APIView):
    """Update an existing staff profile."""
    permission_classes = (IsAuthenticated,)

    def put(self, request, pk):
        from .serializers import StaffProfileSerializer
        from accounts.utils import get_user_permissions
        
        user = request.user
        perms = get_user_permissions(user)
        has_ps_role = user.roles.filter(name__iexact='PS').exists()
        
        # Check permission - require edit_staff or superuser to edit staff
        if not (user.is_superuser or has_ps_role or 'academics.edit_staff' in perms):
            return Response(
                {'detail': 'You do not have permission to edit staff profiles.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            staff_profile = StaffProfile.objects.get(pk=pk)
        except StaffProfile.DoesNotExist:
            return Response(
                {'detail': 'Staff profile not found.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Validate department scope for non-superusers without view_all_staff
        if not user.is_superuser and not has_ps_role and 'academics.view_all_staff' not in perms:
            # HODs can only edit staff in departments they are mapped to
            from academics.utils import get_user_effective_departments
            
            allowed_dept_ids = get_user_effective_departments(user)
            if allowed_dept_ids:
                # Check current department of staff being edited
                if staff_profile.department_id not in allowed_dept_ids:
                    return Response(
                        {'detail': 'You can only edit staff in departments you manage.'},
                        status=status.HTTP_403_FORBIDDEN
                    )
                # If they're trying to change department, validate
                new_dept = request.data.get('department')
                if new_dept and int(new_dept) not in allowed_dept_ids:
                    return Response(
                        {'detail': 'You cannot transfer staff to departments you do not manage.'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                return Response(
                    {'detail': 'You are not mapped to any departments.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        serializer = StaffProfileSerializer(staff_profile, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        """Same as PUT but allows partial updates."""
        return self.put(request, pk)


class StaffProfileDeleteView(APIView):
    """Delete a staff profile."""
    permission_classes = (IsAuthenticated,)

    def delete(self, request, pk):
        from accounts.utils import get_user_permissions
        
        user = request.user
        perms = get_user_permissions(user)
        has_ps_role = user.roles.filter(name__iexact='PS').exists()
        
        # Check permission - require edit_staff or superuser to delete staff
        if not (user.is_superuser or has_ps_role or 'academics.edit_staff' in perms):
            return Response(
                {'detail': 'You do not have permission to delete staff profiles.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            staff_profile = StaffProfile.objects.get(pk=pk)
        except StaffProfile.DoesNotExist:
            return Response(
                {'detail': 'Staff profile not found.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Validate department scope for non-superusers without view_all_staff
        if not user.is_superuser and not has_ps_role and 'academics.view_all_staff' not in perms:
            # HODs can only delete staff in departments they are mapped to
            from academics.utils import get_user_effective_departments
            
            allowed_dept_ids = get_user_effective_departments(user)
            if allowed_dept_ids:
                if staff_profile.department_id not in allowed_dept_ids:
                    return Response(
                        {'detail': 'You can only delete staff in departments you manage.'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                return Response(
                    {'detail': 'You are not mapped to any departments.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        # Delete the user (cascade will delete staff profile)
        staff_profile.user.delete()
        
        return Response(status=status.HTTP_204_NO_CONTENT)

class StaffStatusUpdateView(APIView):
    """Update only the status field of a staff profile (Active / Inactive)."""
    permission_classes = (IsAuthenticated,)

    def patch(self, request, pk):
        from accounts.utils import get_user_permissions
        from academics.models import STAFF_STATUS_CHOICES

        user = request.user
        perms = get_user_permissions(user)
        has_ps_role = user.roles.filter(name__iexact='PS').exists()

        if not (user.is_superuser or has_ps_role or 'academics.edit_staff' in perms):
            return Response(
                {'detail': 'You do not have permission to update staff status.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            staff_profile = StaffProfile.objects.get(pk=pk)
        except StaffProfile.DoesNotExist:
            return Response(
                {'detail': 'Staff profile not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Validate department scope for non-superusers without view_all_staff
        if not user.is_superuser and not has_ps_role and 'academics.view_all_staff' not in perms:
            from academics.utils import get_user_effective_departments

            allowed_dept_ids = get_user_effective_departments(user)
            if allowed_dept_ids:
                if staff_profile.department_id not in allowed_dept_ids:
                    return Response(
                        {'detail': 'You can only update status of staff in departments you manage.'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                return Response(
                    {'detail': 'You are not mapped to any departments.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        new_status = request.data.get('status')
        valid_statuses = [choice[0] for choice in STAFF_STATUS_CHOICES]
        if not new_status or new_status not in valid_statuses:
            return Response(
                {'detail': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        staff_profile.status = new_status
        staff_profile.save(update_fields=['status'])

        return Response({'id': staff_profile.pk, 'status': staff_profile.status})
    
class StaffImportView(APIView):
    """Import staff members from an uploaded Excel (.xlsx) or CSV file.

    Only HOD, AHOD, IQAC, PS users, or superusers are allowed to call this endpoint.

    Expected columns (case-insensitive, spaces/underscores ignored):
      Staff ID, Username, Password, First Name, Last Name, Email, Designation, Department, Status

    Required: Staff ID, Username, Email, Department, Status

    Returns:
      { imported: int, total: int, errors: [{ row: int, errors: [str] }] }
    """
    permission_classes = (IsAuthenticated,)

    def _is_allowed(self, user) -> bool:
        if user.is_superuser:
            return True
        try:
            if user.roles.filter(name__iexact='PS').exists():
                return True
            if user.roles.filter(name__iexact='IQAC').exists():
                return True
        except Exception:
            pass
        from .utils import get_user_staff_profile
        sp = get_user_staff_profile(user)
        if sp:
            from .models import DepartmentRole
            if DepartmentRole.objects.filter(staff=sp, role__in=['HOD', 'AHOD'], is_active=True).exists():
                return True
        return False

    @staticmethod
    def _col(row: dict, *names: str) -> str:
        """Extract a value from a row dict using any of the given normalised key names."""
        for name in names:
            needle = name.lower().replace(' ', '').replace('_', '')
            for k, v in row.items():
                if k.lower().replace(' ', '').replace('_', '') == needle:
                    return str(v).strip() if v is not None else ''
        return ''

    def post(self, request):
        import csv
        import io
        import openpyxl

        user = request.user
        if not self._is_allowed(user):
            return Response(
                {'detail': 'Only HOD or IQAC users can import staff.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file_obj.name.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.csv')):
            return Response(
                {'detail': 'Only .xlsx and .csv files are supported.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Parse file ────────────────────────────────────────────────────────
        rows: list[dict] = []
        if filename.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
                ws = wb.active
                headers: list[str] | None = None
                for excel_row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c).strip() if c is not None else '' for c in excel_row]
                        continue
                    row_data = {
                        h: (str(v).strip() if v is not None else '')
                        for h, v in zip(headers, excel_row)
                    }
                    rows.append(row_data)
            except Exception as exc:
                return Response(
                    {'detail': f'Failed to parse Excel file: {exc}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            try:
                content = file_obj.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                for row in reader:
                    rows.append({k.strip(): (v.strip() if v else '') for k, v in row.items()})
            except Exception as exc:
                return Response(
                    {'detail': f'Failed to parse CSV file: {exc}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if not rows:
            return Response(
                {'detail': 'File is empty or has no data rows.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Process rows ──────────────────────────────────────────────────────
        from django.contrib.auth import get_user_model
        from accounts.models import Role
        from .models import Department

        User = get_user_model()
        errors: list[dict] = []
        imported = 0
        seen_staff_ids: set[str] = set()

        for idx, row in enumerate(rows, start=2):  # row 1 is header
            staff_id   = self._col(row, 'staffid', 'staff id', 'staff_id')
            username   = self._col(row, 'username', 'user name', 'user_name')
            password   = self._col(row, 'password', 'pwd')
            first_name = self._col(row, 'firstname', 'first name', 'first_name')
            last_name  = self._col(row, 'lastname', 'last name', 'last_name')
            email      = self._col(row, 'email', 'emailaddress', 'email address', 'email_address')
            designation = self._col(row, 'designation')
            department_name = self._col(row, 'department', 'dept', 'departmentname', 'department name', 'department_name')
            record_status   = self._col(row, 'status')

            # ── Validate required fields ──
            row_errors: list[str] = []
            if not staff_id:
                row_errors.append('Staff ID is required.')
            if not username:
                row_errors.append('Username is required.')
            if not email:
                row_errors.append('Email is required.')
            if not password:
                row_errors.append('Password is required.')
            if not department_name:
                row_errors.append('Department is required.')
            if not record_status:
                row_errors.append('Status is required.')
            elif record_status.upper() not in ('ACTIVE', 'INACTIVE'):
                row_errors.append(f'Status "{record_status}" is invalid. Allowed: ACTIVE, INACTIVE.')

            if staff_id and staff_id in seen_staff_ids:
                row_errors.append(f'Duplicate Staff ID "{staff_id}" in uploaded file.')

            if row_errors:
                errors.append({'row': idx, 'errors': row_errors})
                continue

            seen_staff_ids.add(staff_id)

            # ── Check DB duplicates ──
            if StaffProfile.objects.filter(staff_id=staff_id).exists():
                errors.append({'row': idx, 'errors': [f'Staff ID "{staff_id}" already exists in the system.']})
                continue
            if User.objects.filter(username=username).exists():
                errors.append({'row': idx, 'errors': [f'Username "{username}" already exists in the system.']})
                continue

            # ── Resolve department ──
            # Supports patterns like "103 - CE", "CE", "103", full name.
            dept = None
            from django.db.models import Q as _Q
            dept_search = department_name.strip()
            # Try the full string first, then strip to code/name split on " - "
            dept = Department.objects.filter(
                _Q(name__iexact=dept_search)
                | _Q(code__iexact=dept_search)
                | _Q(short_name__iexact=dept_search)
            ).first()
            if not dept and ' - ' in dept_search:
                # "103 - CE" -> try code part after " - "
                code_part = dept_search.split(' - ', 1)[1].strip()
                dept = Department.objects.filter(
                    _Q(code__iexact=code_part) | _Q(short_name__iexact=code_part)
                ).first()
            if not dept:
                errors.append({'row': idx, 'errors': [f'Department "{department_name}" not found. Use the department code or "<number> - <code>" format.']})
                continue

            # ── Normalise status ──
            norm_status = record_status.upper()  # already validated above

            # ── Create user + staff profile ──
            try:
                with transaction.atomic():
                    user_obj = User.objects.create_user(
                        username=username,
                        password=password,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                    )

                    # Create the StaffProfile BEFORE assigning the STAFF role.
                    # The pre_add M2M signal calls validate_roles_for_user(), which
                    # raises ValidationError if no staff_profile exists yet.
                    # Creating the profile first avoids that and prevents the
                    # TransactionManagementError on subsequent queries.
                    StaffProfile.objects.create(
                        user=user_obj,
                        staff_id=staff_id,
                        department=dept,
                        designation=designation,
                        status=norm_status,
                    )

                    try:
                        staff_role = Role.objects.get(name='STAFF')
                        user_obj.roles.add(staff_role)
                    except Exception:
                        pass  # STAFF role missing or already assigned; non-critical

                    imported += 1
            except Exception as exc:
                errors.append({'row': idx, 'errors': [str(exc)]})

        return Response(
            {'imported': imported, 'total': len(rows), 'errors': errors},
            status=status.HTTP_200_OK,
        )

class AdvisorStaffListView(APIView):
    """Return staff list limited to departments/sections the advisor is assigned to.

    This endpoint is intended for advisors to choose staff when assigning teaching.
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return Response({'results': []})

        # Sections advisor maps to (active advisors for active academic years)
        advisor_qs = SectionAdvisor.objects.filter(advisor=staff_profile, is_active=True, academic_year__is_active=True).select_related('section__batch__course')
        if not advisor_qs.exists():
            return Response({'results': []})

        dept_ids = set()
        for a in advisor_qs:
            try:
                sec = getattr(a, 'section', None)
                batch = getattr(sec, 'batch', None) if sec is not None else None
                course = getattr(batch, 'course', None) if batch is not None else None
                dept = getattr(course, 'department', None) if course is not None else None
                if dept:
                    dept_ids.add(dept.id)
            except Exception:
                continue

        # If caller has explicit permission to view all staff across departments,
        # return full list (optionally filtered by department query param).
        perms = get_user_permissions(user)
        dept_param = request.query_params.get('department')
        try:
            dept_filter = int(dept_param) if dept_param else None
        except Exception:
            dept_filter = None

        if ('academics.view_all_staff' in perms) or ('academics.view_all_departments' in perms) or user.is_staff:
            staff_qs = StaffProfile.objects.all().select_related('user')
            if dept_filter:
                staff_qs = staff_qs.filter(
                    Q(department_id=dept_filter) |
                    Q(department_assignments__department_id=dept_filter, department_assignments__end_date__isnull=True)
                ).distinct()
        else:
            if not dept_ids:
                return Response({'results': []})
            staff_qs = StaffProfile.objects.filter(
                Q(department_id__in=list(dept_ids)) |
                Q(department_assignments__department_id__in=list(dept_ids), department_assignments__end_date__isnull=True)
            ).select_related('user').distinct()

        results = []
        for s in staff_qs:
            user_data = None
            if s.user:
                user_data = {
                    'username': s.user.username,
                    'first_name': getattr(s.user, 'first_name', ''),
                    'last_name': getattr(s.user, 'last_name', '')
                }
            results.append({
                'id': s.id, 
                'user': user_data, 
                'staff_id': s.staff_id, 
                'department': getattr(s.department, 'id', None)
            })
        return Response({'results': results})


class SectionStudentsView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request, section_id: int):
        # Return students in the given section (current active assignments + legacy)
        try:
            sid = int(section_id)
        except Exception:
            return Response({'results': []})

        # current assignments — include both PRIMARY and SECONDARY so that core-dept sections
        # show their Year-1 students (those have SECONDARY assignments to core-dept sections).
        from .models import StudentSectionAssignment, StudentProfile
        assign_qs = StudentSectionAssignment.objects.filter(section_id=sid, end_date__isnull=True).exclude(student__status__in=['INACTIVE', 'DEBAR']).select_related('student__user')
        students = [a.student for a in assign_qs]

        # legacy field
        legacy = StudentProfile.objects.filter(section_id=sid).exclude(status__in=['INACTIVE', 'DEBAR']).select_related('user')
        for s in legacy:
            if not any(x.pk == s.pk for x in students):
                students.append(s)

        ser = StudentSimpleSerializer([
            {'id': st.pk, 'reg_no': st.reg_no, 'user': getattr(st, 'user', None), 'section_id': getattr(st, 'section_id', None), 'section_name': str(getattr(st, 'section', ''))}
            for st in students
        ], many=True)
        return Response({'results': ser.data})


class StaffAssignedSubjectsView(APIView):
    """Return teaching assignments (subjects) for a staff member.

    URL patterns:
    - /api/academics/staff/assigned-subjects/  -> current user's staff_profile
    - /api/academics/staff/<staff_id>/assigned-subjects/ -> specified staff id
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request, staff_id: int = None):
        user = request.user
        # resolve target staff_profile
        target = None
        if staff_id:
            try:
                target = StaffProfile.objects.filter(pk=int(staff_id)).first()
            except Exception:
                target = None
        else:
            target = get_user_staff_profile(user)

        if not target:
            if not staff_id:
                return Response({'results': []})
            return Response({'detail': 'Staff not found'}, status=status.HTTP_404_NOT_FOUND)

        # permission: allow if user is the staff themselves, superuser, has explicit perm,
        # or is HOD for the staff's department
        if user.is_superuser:
            allowed = True
        elif getattr(user, 'id', None) == getattr(getattr(target, 'user', None), 'id', None):
            allowed = True
        else:
            perms = get_user_permissions(user)
            if 'academics.view_assigned_subjects' in perms or user.has_perm('academics.view_assigned_subjects'):
                allowed = True
            else:
                # HODs may view staff in their mapped departments
                hod_dept_ids = get_user_effective_departments(user)
                target_dept_id = None
                try:
                    target_dept = getattr(target, 'current_department', None) or target.get_current_department()
                    if not target_dept:
                        target_dept = getattr(target, 'department', None)
                    target_dept_id = getattr(target_dept, 'id', None) if target_dept else None
                except Exception:
                    target_dept_id = getattr(getattr(target, 'department', None), 'id', None)

                allowed = bool(target_dept_id and target_dept_id in hod_dept_ids)

        if not allowed:
            raise PermissionDenied('You do not have permission to view this staff assignments.')

        # fetch teaching assignments
        # Only return assignments that have an explicit curriculum_row/subject.
        # Prefer the active academic year when available, but avoid returning
        # an empty list just because the active-year flag isn't set.
        qs = TeachingAssignment.objects.filter(
            staff=target,
            is_active=True,
        ).filter(
            Q(curriculum_row__isnull=False) | Q(subject__isnull=False) | Q(elective_subject__isnull=False)
        ).select_related('curriculum_row', 'section', 'academic_year', 'subject')

        try:
            if qs.filter(academic_year__is_active=True).exists():
                qs = qs.filter(academic_year__is_active=True)
        except Exception:
            pass

        # Backfill from StudentSubjectBatch (best-effort) if nothing exists.
        if not qs.exists():
            try:
                _ensure_teaching_assignments_from_subject_batches(target)
            except Exception:
                pass
            qs = TeachingAssignment.objects.filter(
                staff=target,
                is_active=True,
            ).filter(
                Q(curriculum_row__isnull=False) | Q(subject__isnull=False) | Q(elective_subject__isnull=False)
            ).select_related('curriculum_row', 'section', 'academic_year', 'subject')

            try:
                if qs.filter(academic_year__is_active=True).exists():
                    qs = qs.filter(academic_year__is_active=True)
            except Exception:
                pass
        ser = TeachingAssignmentInfoSerializer(qs, many=True)

        # Also include subject batch assignments (StudentSubjectBatch.staff=target).
        # This ensures staff who are assigned only to a batch (not the main subject)
        # still see the subject in their Assigned Subjects page with batch reference.
        results = list(ser.data)
        try:
            from academics.models import StudentSubjectBatch
            from timetable.models import TimetableAssignment
            from academics.models import AcademicYear

            active_ay = None
            try:
                active_ay = AcademicYear.objects.filter(is_active=True).order_by('-id').first()
            except Exception:
                active_ay = None

            bqs = StudentSubjectBatch.objects.filter(
                staff=target,
                is_active=True,
            ).select_related(
                'curriculum_row',
                'curriculum_row__department',
                'curriculum_row__semester',
                'section',
                'academic_year',
                'created_by',
            )
            if active_ay:
                bqs = bqs.filter(academic_year=active_ay)

            # Index existing results by (curriculum_row_id, section_id) so we can attach
            # batch refs without mixing batches across multiple sections.
            by_key: dict = {}
            for r in results:
                try:
                    cid = r.get('curriculum_row_id')
                    if cid is None:
                        continue
                    sid = r.get('section_id')
                    sid_int = int(sid) if sid is not None else None
                    key = (int(cid), sid_int)
                    by_key.setdefault(key, []).append(r)
                except Exception:
                    continue

            def _infer_batch_section(sb_obj):
                section_id = None
                section_name = None

                try:
                    if getattr(sb_obj, 'section_id', None):
                        section_id = getattr(sb_obj, 'section_id', None)
                        try:
                            section_name = getattr(getattr(sb_obj, 'section', None), 'name', None)
                        except Exception:
                            section_name = None
                        return section_id, section_name
                except Exception:
                    section_id = None
                    section_name = None

                try:
                    ta = TimetableAssignment.objects.filter(subject_batch=sb_obj, section__isnull=False).select_related('section').first()
                    if ta and getattr(ta, 'section', None):
                        section_id = getattr(ta.section, 'id', None)
                        section_name = getattr(ta.section, 'name', None)
                        return section_id, section_name
                except Exception:
                    pass

                try:
                    st = sb_obj.students.select_related('section').first()
                    if st and getattr(st, 'section', None):
                        section_id = getattr(st.section, 'id', None)
                        section_name = getattr(st.section, 'name', None)
                        return section_id, section_name
                except Exception:
                    pass

                return None, None

            # Build a cache of creator -> elective TAs so we can resolve subject info
            # for elective batches (curriculum_row is NULL).
            _creator_elective_ta_cache: dict = {}

            def _resolve_elective_subject_for_batch(sb_obj):
                """Look up elective subject info for a batch by checking creator's TAs."""
                creator_id = getattr(sb_obj, 'created_by_id', None)
                if not creator_id:
                    return None, None, None, None, None  # code, name, elective_subject_id, dept_obj, sem_num
                if creator_id not in _creator_elective_ta_cache:
                    etas = TeachingAssignment.objects.filter(
                        staff_id=creator_id,
                        elective_subject__isnull=False,
                        is_active=True,
                    ).select_related(
                        'elective_subject',
                        'elective_subject__department',
                        'elective_subject__semester',
                    )
                    try:
                        if active_ay and etas.filter(academic_year=active_ay).exists():
                            etas = etas.filter(academic_year=active_ay)
                    except Exception:
                        pass
                    _creator_elective_ta_cache[creator_id] = list(etas)

                etas = _creator_elective_ta_cache.get(creator_id, [])
                if not etas:
                    return None, None, None, None, None

                # If there's only one elective TA, use it; otherwise pick first match
                eta = etas[0]
                es = getattr(eta, 'elective_subject', None)
                if not es:
                    return None, None, None, None, None

                dept_obj = None
                try:
                    dept = getattr(es, 'department', None)
                    if dept:
                        dept_obj = {
                            'id': getattr(dept, 'id', None),
                            'code': getattr(dept, 'code', None),
                            'name': getattr(dept, 'name', None),
                            'short_name': getattr(dept, 'short_name', None),
                        }
                except Exception:
                    dept_obj = None

                sem_num = None
                try:
                    sem_num = getattr(getattr(es, 'semester', None), 'number', None)
                except Exception:
                    sem_num = None

                return (
                    getattr(es, 'course_code', None),
                    getattr(es, 'course_name', None),
                    getattr(es, 'pk', None),
                    dept_obj,
                    sem_num,
                )

            for sb in bqs:
                try:
                    cr = getattr(sb, 'curriculum_row', None)
                    batch_info = {'id': sb.pk, 'name': getattr(sb, 'name', None)}
                    section_id, section_name = _infer_batch_section(sb)

                    if cr:
                        # ── Curriculum-row based batch (non-elective) ──
                        key = (int(cr.pk), int(section_id) if section_id is not None else None)
                        existing_rows = by_key.get(key)
                        if not existing_rows and section_id is None:
                            existing_rows = by_key.get((int(cr.pk), None))
                        if existing_rows:
                            for existing in existing_rows:
                                lst = existing.get('subject_batches')
                                if not isinstance(lst, list):
                                    lst = []
                                if not any(int(x.get('id')) == int(sb.pk) for x in lst if isinstance(x, dict) and x.get('id') is not None):
                                    lst.append(batch_info)
                                existing['subject_batches'] = lst
                            continue

                        dept_obj = None
                        try:
                            dept = getattr(cr, 'department', None)
                            if dept:
                                dept_obj = {
                                    'id': getattr(dept, 'id', None),
                                    'code': getattr(dept, 'code', None),
                                    'name': getattr(dept, 'name', None),
                                    'short_name': getattr(dept, 'short_name', None),
                                }
                        except Exception:
                            dept_obj = None

                        sem_num = None
                        try:
                            sem_num = getattr(getattr(cr, 'semester', None), 'number', None)
                        except Exception:
                            sem_num = None

                        results.append({
                            'id': -int(sb.pk),
                            'subject_code': getattr(cr, 'course_code', None),
                            'subject_name': getattr(cr, 'course_name', None),
                            'class_type': 'BATCH',
                            'section_name': section_name,
                            'section_id': section_id,
                            'elective_subject_id': None,
                            'elective_subject_name': None,
                            'curriculum_row_id': getattr(cr, 'pk', None),
                            'batch': None,
                            'semester': sem_num,
                            'academic_year': getattr(getattr(sb, 'academic_year', None), 'name', None),
                            'department': dept_obj,
                            'subject_batches': [batch_info],
                        })
                        by_key.setdefault(key, []).append(results[-1])
                    else:
                        # ── Elective batch (no curriculum_row) ──
                        # Derive subject info from the batch creator's elective teaching assignment.
                        e_code, e_name, e_sub_id, e_dept, e_sem = _resolve_elective_subject_for_batch(sb)
                        if not e_code and not e_name:
                            continue  # can't determine subject at all

                        # Check if an existing result already covers this elective subject
                        matched = False
                        for r in results:
                            if r.get('elective_subject_id') and e_sub_id and int(r.get('elective_subject_id')) == int(e_sub_id):
                                lst = r.get('subject_batches')
                                if not isinstance(lst, list):
                                    lst = []
                                if not any(int(x.get('id')) == int(sb.pk) for x in lst if isinstance(x, dict) and x.get('id') is not None):
                                    lst.append(batch_info)
                                r['subject_batches'] = lst
                                matched = True
                                break
                        if matched:
                            continue

                        results.append({
                            'id': -int(sb.pk),
                            'subject_code': e_code,
                            'subject_name': e_name,
                            'class_type': 'BATCH',
                            'section_name': section_name,
                            'section_id': section_id,
                            'elective_subject_id': e_sub_id,
                            'elective_subject_name': e_name,
                            'curriculum_row_id': None,
                            'batch': None,
                            'semester': e_sem,
                            'academic_year': getattr(getattr(sb, 'academic_year', None), 'name', None),
                            'department': e_dept,
                            'subject_batches': [batch_info],
                        })
                except Exception:
                    continue
        except Exception:
            pass

        return Response({'results': results})


class IQACCourseTeachingMapView(APIView):
    """IQAC/OBE Master: list teaching assignments for a course across sections.

    Returns section + staff mapping as card-friendly rows.
    """

    permission_classes = (IsAuthenticated,)

    def get(self, request, course_code: str):
        user = request.user
        perms = {str(p or '').lower() for p in (get_user_permissions(user) or [])}
        roles = set()
        try:
            roles = {str(r.name or '').upper() for r in user.roles.all()}
        except Exception:
            roles = set()

        # Gate to IQAC/OBE master users only.
        if not (user.is_superuser or user.is_staff or ('obe.master.manage' in perms) or ('IQAC' in roles)):
            raise PermissionDenied('IQAC/OBE Master access only.')

        code = str(course_code or '').strip()
        if not code:
            return Response({'results': []})

        qs = TeachingAssignment.objects.select_related(
            'staff',
            'staff__user',
            'section',
            'academic_year',
            'subject',
            'curriculum_row',
            'curriculum_row__master',
            'section__batch__course__department',
        ).filter(is_active=True)

        # Filter to the requested course first (curriculum_row, master row, legacy subject, or elective).
        qs = qs.filter(
            Q(curriculum_row__course_code__iexact=code)
            | Q(curriculum_row__master__course_code__iexact=code)
            | Q(subject__code__iexact=code)
            | Q(elective_subject__course_code__iexact=code)
        )

        # Prefer active academic year only within this course (avoid hiding results when
        # the course assignments exist but the academic_year.is_active flag isn't set).
        try:
            if qs.filter(academic_year__is_active=True).exists():
                qs_active = qs.filter(academic_year__is_active=True)
                if qs_active.exists():
                    qs = qs_active
        except Exception:
            pass

        results = []
        for ta in qs.order_by('section__name', 'id'):
            sec = getattr(ta, 'section', None)
            ay = getattr(ta, 'academic_year', None)
            staff = getattr(ta, 'staff', None)
            staff_user = getattr(staff, 'user', None) if staff else None

            # Best-effort subject metadata
            subject_code = None
            subject_name = None
            class_type = None
            try:
                if getattr(ta, 'elective_subject', None):
                    es = ta.elective_subject
                    subject_code = getattr(es, 'course_code', None)
                    subject_name = getattr(es, 'course_name', None)
                    class_type = getattr(es, 'class_type', None)
                if getattr(ta, 'curriculum_row', None):
                    cr = ta.curriculum_row
                    subject_code = subject_code or getattr(cr, 'course_code', None) or getattr(getattr(cr, 'master', None), 'course_code', None)
                    subject_name = subject_name or getattr(cr, 'course_name', None) or getattr(getattr(cr, 'master', None), 'course_name', None)
                    class_type = class_type or getattr(cr, 'class_type', None) or getattr(getattr(cr, 'master', None), 'class_type', None)
                if (not subject_code or not subject_name) and getattr(ta, 'subject', None):
                    subject_code = subject_code or getattr(ta.subject, 'code', None)
                    subject_name = subject_name or getattr(ta.subject, 'name', None)
            except Exception:
                pass

            results.append(
                {
                    'teaching_assignment_id': getattr(ta, 'id', None),
                    'course_code': subject_code or code,
                    'course_name': subject_name,
                    'class_type': class_type,
                    'section_id': getattr(sec, 'id', None),
                    'section_name': getattr(sec, 'name', None),
                    'academic_year': getattr(ay, 'name', None) if ay else None,
                    'staff': {
                        'id': getattr(staff, 'id', None),
                        'staff_id': getattr(staff, 'staff_id', None),
                        'username': getattr(staff_user, 'username', None),
                        'name': ' '.join(filter(None, [getattr(staff_user, 'first_name', ''), getattr(staff_user, 'last_name', '')])).strip()
                        or getattr(staff_user, 'username', None),
                    }
                    if staff
                    else None,
                }
            )

        return Response({'results': results})


class SubjectBatchViewSet(viewsets.ModelViewSet):
    """Manage StudentSubjectBatch resources for the current staff user."""
    permission_classes = (IsAuthenticated,)
    serializer_class = None

    def get_serializer_class(self):
        from .serializers import StudentSubjectBatchSerializer
        return StudentSubjectBatchSerializer

    def get_queryset(self):
        user = self.request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return []
        from .models import StudentSubjectBatch
        # staff sees only their own batches; superusers can see all
        qs = StudentSubjectBatch.objects.select_related('staff', 'created_by', 'academic_year', 'section').prefetch_related('students')
        # allow callers to request all batches (useful for timetable editors)
        include_all = str(self.request.query_params.get('include_all') or '').lower() in ('1', 'true', 'yes')

        # optional filter: section_id (critical for multi-section subjects)
        section_id = self.request.query_params.get('section_id') or self.request.query_params.get('section')
        section_id_int = None
        if section_id not in (None, '', 'null'):
            try:
                section_id_int = int(section_id)
                qs = qs.filter(section_id=section_id_int)
            except Exception:
                section_id_int = None

        # Only superusers can truly include_all without a section filter.
        # For non-superusers, include_all is honored only when section_id is provided.
        if include_all and (not user.is_superuser) and not section_id_int:
            include_all = False

        if not user.is_superuser and not include_all:
            # Show batches created by this staff OR assigned to this staff
            qs = qs.filter(Q(staff=staff_profile) | Q(created_by=staff_profile))

        # allow filtering by curriculum_row_id via query param (useful for timetable editor)
        cr = self.request.query_params.get('curriculum_row_id') or self.request.query_params.get('curriculum_row')
        if cr:
            try:
                cr_id = int(cr)
                qs = qs.filter(curriculum_row_id=cr_id)
            except Exception:
                pass
        
        # allow filtering by student_id to find batches containing a specific student
        student_id = self.request.query_params.get('student_id')
        if student_id:
            try:
                student_id = int(student_id)
                qs = qs.filter(students__id=student_id).distinct()
            except Exception:
                pass
        
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Only staff users can create subject batches')
        # Set created_by to current user
        # If staff is not already set by serializer (from staff_id), use current user
        if 'staff' not in serializer.validated_data:
            serializer.save(staff=staff_profile, created_by=staff_profile)
        else:
            serializer.save(created_by=staff_profile)

    def perform_update(self, serializer):
        user = self.request.user
        staff_profile = getattr(user, 'staff_profile', None)
        instance = self.get_object()
        
        # Only the creator can edit the batch
        if instance.created_by and instance.created_by != staff_profile and not user.is_superuser:
            raise PermissionDenied('Only the batch creator can edit this batch')
        
        serializer.save()

    def perform_destroy(self, instance):
        user = self.request.user
        staff_profile = getattr(user, 'staff_profile', None)
        
        # Only the creator can delete the batch
        if instance.created_by and instance.created_by != staff_profile and not user.is_superuser:
            raise PermissionDenied('Only the batch creator can delete this batch')
        
        instance.delete()



class PeriodAttendanceSessionViewSet(viewsets.ModelViewSet):
    queryset = PeriodAttendanceSession.objects.select_related('section', 'period', 'timetable_assignment', 'timetable_assignment__staff', 'teaching_assignment', 'subject_batch', 'subject_batch__staff', 'subject_batch__created_by').prefetch_related('records')
    serializer_class = PeriodAttendanceSessionSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        """Filter queryset by date range and staff assignment (including batch assignments)."""
        queryset = super().get_queryset()
        user = self.request.user
        staff_profile = getattr(user, 'staff_profile', None)
        
        # Filter by staff assignment if staff user
        if staff_profile and not user.is_superuser:
            perms = get_user_permissions(user)
            # If user has special attendance permission, show all
            if 'academics.mark_attendance' not in perms:
                # Show sessions where staff is:
                # 1. Created by this staff
                # 2. Assigned to this staff (swap scenario)
                # 3. Timetable assignment staff matches (advisor who created assignment)
                # 4. Subject batch staff matches
                # 5. Subject batch creator matches
                # 6. Teaching assignment matches for the curriculum_row (actual subject teacher)
                from timetable.models import TimetableAssignment
                from .models import TeachingAssignment
                
                # Build the base filter
                base_filter = Q(created_by=staff_profile) | \
                              Q(assigned_to=staff_profile) | \
                              Q(timetable_assignment__staff=staff_profile) | \
                              Q(subject_batch__staff=staff_profile) | \
                              Q(subject_batch__created_by=staff_profile)
                
                # Add filter for teaching assignment match - when staff is the teaching staff
                # for a curriculum_row assigned in the timetable
                teaching_filter = Q(
                    teaching_assignment__staff=staff_profile
                )
                
                queryset = queryset.filter(base_filter | teaching_filter).distinct()
        
        # Support date filtering for bulk attendance checking
        date_after = self.request.query_params.get('date_after')
        date_before = self.request.query_params.get('date_before')
        
        if date_after:
            try:
                import datetime
                queryset = queryset.filter(date__gte=datetime.date.fromisoformat(date_after))
            except Exception:
                pass
        
        if date_before:
            try:
                import datetime
                queryset = queryset.filter(date__lte=datetime.date.fromisoformat(date_before))
            except Exception:
                pass
        
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Only staff users may create attendance sessions')

        # Check if staff can mark attendance based on their daily attendance
        date = serializer.validated_data.get('date')
        if date:
            from .utils import can_staff_mark_period_attendance
            can_mark, reason, attendance_record = can_staff_mark_period_attendance(user, date)
            if not can_mark:
                raise PermissionDenied(f'Cannot mark period attendance: {reason}')

        # determine day for permission checking
        date = serializer.validated_data.get('date')
        period = serializer.validated_data.get('period')
        section = serializer.validated_data.get('section')
        day = None
        try:
            if date is not None:
                day = date.isoweekday()
        except Exception:
            day = None

        ta = None
        teach_assign = None
        batch_staff_check_passed = False
        try:
            from timetable.models import TimetableAssignment
            if section and period and day:
                # Check if there are ANY batch-specific assignments for this period
                batch_assignments = TimetableAssignment.objects.filter(
                    section=section, period=period, day=day, subject_batch__isnull=False
                )
                
                if batch_assignments.exists():
                    # Batch assignments exist - ONLY batch staff can access, block default staff
                    for ta_candidate in batch_assignments:
                        batch = ta_candidate.subject_batch
                        is_batch_staff = (
                            getattr(batch, 'created_by_id', None) == staff_profile.id or
                            getattr(batch, 'staff_id', None) == staff_profile.id
                        )
                        if is_batch_staff:
                            ta = ta_candidate
                            batch_staff_check_passed = True
                            break
                    # If ta is still None, staff is not in any batch - they cannot access
                else:
                    # No batch assignments - use regular timetable assignment
                    ta = TimetableAssignment.objects.filter(
                        section=section, period=period, day=day, staff=staff_profile
                    ).first()
                
                # If no explicit timetable assignment with this staff, check if they are the teaching staff
                # for any curriculum_row assigned to this period (handles advisor assignment case)
                if ta is None and not batch_assignments.exists():
                    assign = TimetableAssignment.objects.filter(section=section, period=period, day=day).first()
                    if assign and getattr(assign, 'curriculum_row', None):
                        from .models import TeachingAssignment as _TA
                        _acr_name = getattr(getattr(assign, 'curriculum_row', None), 'course_name', None)
                        _sec_dept_id_a = getattr(getattr(getattr(section, 'batch', None), 'course', None), 'department_id', None) if section else None
                        ta_match_qs = _TA.objects.filter(is_active=True, staff=staff_profile)
                        ta_match_qs = ta_match_qs.filter(
                            Q(curriculum_row=assign.curriculum_row) |
                            Q(elective_subject__parent=assign.curriculum_row) |
                            Q(elective_subject__department_group__isnull=False,
                              elective_subject__parent__course_name=_acr_name,
                              elective_subject__department_group__department_mappings__department_id=_sec_dept_id_a,
                              elective_subject__department_group__department_mappings__is_active=True)
                        ).filter(Q(section=section) | Q(section__isnull=True))
                        if ta_match_qs.exists():
                            ta = assign
                            teach_assign = ta_match_qs.order_by(
                                models.Case(
                                    models.When(section=section, then=models.Value(0)),
                                    default=models.Value(1),
                                    output_field=models.IntegerField(),
                                ),
                                'id',
                            ).first()
        except Exception:
            ta = None
            teach_assign = None

        # If teaching_assignment wasn't resolved above, try resolving from timetable assignment.
        if teach_assign is None and staff_profile:
            try:
                from .models import TeachingAssignment as _TA
                ta_qs = _TA.objects.filter(is_active=True, staff=staff_profile).filter(Q(section=section) | Q(section__isnull=True))
                cr = getattr(getattr(ta, 'curriculum_row', None), 'pk', None) if ta is not None else None
                _cr_obj = getattr(ta, 'curriculum_row', None) if ta is not None else None
                _cr_name_fallback = getattr(_cr_obj, 'course_name', None)
                _sec_dept_id_f = getattr(getattr(getattr(section, 'batch', None), 'course', None), 'department_id', None) if section else None
                if cr:
                    # Filter by the period's curriculum_row — only match if staff actually teaches it.
                    ta_qs = ta_qs.filter(
                        Q(curriculum_row_id=cr) |
                        Q(elective_subject__parent_id=cr) |
                        Q(elective_subject__department_group__isnull=False,
                          elective_subject__parent__course_name=_cr_name_fallback,
                          elective_subject__department_group__department_mappings__department_id=_sec_dept_id_f,
                          elective_subject__department_group__department_mappings__is_active=True)
                    )
                    teach_assign = ta_qs.order_by(
                        models.Case(
                            models.When(section=section, then=models.Value(0)),
                            default=models.Value(1),
                            output_field=models.IntegerField(),
                        ),
                        'id',
                    ).first()
                elif ta is None:
                    # No timetable assignment context at all — fall back to any active TA for
                    # this staff+section (e.g. attendance created directly without a timetable slot).
                    teach_assign = ta_qs.order_by(
                        models.Case(
                            models.When(section=section, then=models.Value(0)),
                            default=models.Value(1),
                            output_field=models.IntegerField(),
                        ),
                        'id',
                    ).first()
                # else: ta exists but has no curriculum_row (period uses custom subject_text /
                # substitute subject). Don't pick a random TA — keep teach_assign=None to prevent
                # the wrong subject's TeachingAssignment from being linked to this session.
            except Exception:
                teach_assign = None

        perms = get_user_permissions(user)
        
        # Check if there are batch assignments for this period
        try:
            from timetable.models import TimetableAssignment
            batch_assignments = TimetableAssignment.objects.filter(
                section=section, period=period, day=day, subject_batch__isnull=False
            ) if section and period and day else TimetableAssignment.objects.none()
        except Exception:
            batch_assignments = TimetableAssignment.objects.none()
        
        # If batch assignments exist, ONLY batch staff can access - block everyone else
        if batch_assignments.exists() and not (user.is_superuser or 'academics.mark_attendance' in perms):
            # Check if staff has access to any of the batches
            has_batch_access = False
            if ta and getattr(ta, 'subject_batch', None):
                batch = ta.subject_batch
                is_batch_staff = (
                    getattr(batch, 'created_by_id', None) == staff_profile.id or
                    getattr(batch, 'staff_id', None) == staff_profile.id
                )
                if is_batch_staff:
                    has_batch_access = True
            
            if not has_batch_access:
                # Get batch name for error message
                try:
                    batch_ta = batch_assignments.first()
                    batch_name = batch_ta.subject_batch.name if batch_ta and batch_ta.subject_batch else 'assigned batch'
                    raise PermissionDenied(f'This period is assigned to "{batch_name}". Only staff assigned to that batch can mark attendance.')
                except PermissionDenied:
                    raise
                except Exception:
                    raise PermissionDenied('This period is assigned to a batch. Only staff assigned to that batch can mark attendance.')
        
        # Regular permission check for non-batch periods
        if not batch_assignments.exists():
            if not (ta or 'academics.mark_attendance' in perms or user.is_superuser):
                raise PermissionDenied('You are not assigned to this period and cannot mark attendance')

        # Extract subject_batch from timetable assignment if present
        subject_batch = getattr(ta, 'subject_batch', None) if ta else None
        
        if ta or teach_assign:
            serializer.save(timetable_assignment=ta, teaching_assignment=teach_assign, subject_batch=subject_batch, created_by=staff_profile)
        else:
            serializer.save(created_by=staff_profile)

    @action(detail=False, methods=['post'], url_path='marked-keys')
    def marked_keys(self, request):
        """Return lightweight marked keys for bulk modal.

        Payload:
          {
            month: 'YYYY-MM',
            assignments: [{section_id, period_id, day}, ...]
          }
        Response:
          { results: [{date, section_id, period_id}] }
        """
        import datetime

        data = request.data or {}
        month = str(data.get('month') or '').strip()
        assignments = data.get('assignments') or []
        if not month or len(month) != 7:
            return Response({'detail': 'month is required in YYYY-MM format'}, status=400)

        try:
            year, mon = [int(x) for x in month.split('-')]
            start_date = datetime.date(year, mon, 1)
            if mon == 12:
                end_date = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end_date = datetime.date(year, mon + 1, 1) - datetime.timedelta(days=1)
        except Exception:
            return Response({'detail': 'invalid month format'}, status=400)

        section_ids = set()
        period_ids = set()
        allowed_weekdays = set()
        for a in assignments:
            try:
                section_ids.add(int(a.get('section_id')))
                period_ids.add(int(a.get('period_id')))
                allowed_weekdays.add(int(a.get('day')))
            except Exception:
                continue

        if not section_ids or not period_ids:
            return Response({'results': []})

        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile and not user.is_superuser:
            return Response({'results': []})

        qs = PeriodAttendanceSession.objects.filter(
            date__gte=start_date,
            date__lte=end_date,
            section_id__in=list(section_ids),
            period_id__in=list(period_ids),
        )

        if staff_profile and not user.is_superuser:
            perms = get_user_permissions(user)
            if 'academics.mark_attendance' not in perms:
                qs = qs.filter(
                    Q(created_by=staff_profile) |
                    Q(assigned_to=staff_profile) |
                    Q(timetable_assignment__staff=staff_profile) |
                    Q(subject_batch__staff=staff_profile) |
                    Q(subject_batch__created_by=staff_profile)
                ).distinct()

        results = []
        for row in qs.values('date', 'section_id', 'period_id'):
            dt = row['date']
            if allowed_weekdays and dt.isoweekday() not in allowed_weekdays:
                continue
            results.append({
                'date': dt.isoformat(),
                'section_id': row['section_id'],
                'period_id': row['period_id'],
            })

        return Response({'results': results})

    @action(detail=False, methods=['post'], url_path='bulk-mark')
    def bulk_mark(self, request):
        ser = BulkPeriodAttendanceSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        section_id = data.get('section_id')
        period_id = data.get('period_id')
        teaching_assignment_id = data.get('teaching_assignment_id')
        date = data.get('date')
        records = data.get('records') or []

        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        
        # Check if staff can mark attendance based on their daily attendance
        if date:
            from .utils import can_staff_mark_period_attendance
            can_mark, reason, attendance_record = can_staff_mark_period_attendance(user, date)
            if not can_mark:
                return Response({
                    'error': f'Cannot mark period attendance: {reason}',
                    'attendance_locked': True,
                    'attendance_record': {
                        'date': attendance_record.date.isoformat() if attendance_record else None,
                        'status': attendance_record.status if attendance_record else None
                    } if attendance_record else None
                }, status=status.HTTP_403_FORBIDDEN)
        
        day = None
        try:
            if date is not None:
                day = date.isoweekday()
        except Exception:
            day = None

        ta = None
        teach_assign = None
        try:
            from timetable.models import TimetableAssignment, TimetableSlot
            from .models import Section as _Section
            section = _Section.objects.filter(pk=int(section_id)).first() if section_id is not None else None
            period = TimetableSlot.objects.filter(pk=int(period_id)).first() if period_id is not None else None
            if section and period and day and staff_profile:
                # Check if there are ANY batch-specific assignments for this period
                batch_assignments = TimetableAssignment.objects.filter(
                    section=section, period=period, day=day, subject_batch__isnull=False
                )
                
                if batch_assignments.exists():
                    # Batch assignments exist - ONLY batch staff can access, block default staff
                    for ta_candidate in batch_assignments:
                        batch = ta_candidate.subject_batch
                        is_batch_staff = (
                            getattr(batch, 'created_by_id', None) == staff_profile.id or
                            getattr(batch, 'staff_id', None) == staff_profile.id
                        )
                        if is_batch_staff:
                            ta = ta_candidate
                            break
                    # If ta is still None, staff is not in any batch - they cannot access
                else:
                    # No batch assignments - use regular timetable assignment
                    ta = TimetableAssignment.objects.filter(
                        section=section, period=period, day=day, staff=staff_profile
                    ).first()
                
                if ta is None and not batch_assignments.exists():
                    assign = TimetableAssignment.objects.filter(section=section, period=period, day=day).first()
                    if assign and not getattr(assign, 'staff', None):
                        from .models import TeachingAssignment as _TA
                        _bulk_cr_name = getattr(getattr(assign, 'curriculum_row', None), 'course_name', None)
                        _sec_dept_id_b = getattr(getattr(getattr(section, 'batch', None), 'course', None), 'department_id', None) if section else None
                        ta_match_qs = _TA.objects.filter(is_active=True, staff=staff_profile)
                        ta_match_qs = ta_match_qs.filter(
                            Q(curriculum_row=assign.curriculum_row) |
                            Q(elective_subject__parent=assign.curriculum_row) |
                            Q(elective_subject__department_group__isnull=False,
                              elective_subject__parent__course_name=_bulk_cr_name,
                              elective_subject__department_group__department_mappings__department_id=_sec_dept_id_b,
                              elective_subject__department_group__department_mappings__is_active=True)
                        ).filter(Q(section=section) | Q(section__isnull=True))
                        if ta_match_qs.exists():
                            ta = assign
                            teach_assign = ta_match_qs.order_by(
                                models.Case(
                                    models.When(section=section, then=models.Value(0)),
                                    default=models.Value(1),
                                    output_field=models.IntegerField(),
                                ),
                                'id',
                            ).first()
        except Exception:
            ta = None
            teach_assign = None

        # If client provided a teaching_assignment_id, it takes precedence (subject-wise identity).
        if teaching_assignment_id is not None and staff_profile:
            try:
                from .models import TeachingAssignment as _TA
                ta_obj = _TA.objects.filter(pk=int(teaching_assignment_id), is_active=True, staff=staff_profile).first()
                if not ta_obj:
                    # Check if this is a swap-assigned session: the TA belongs to the original staff,
                    # but current staff has been assigned to take attendance for it.
                    swap_exists = (
                        section is not None and period is not None and date is not None and
                        PeriodAttendanceSession.objects.filter(
                            section=section, period=period, date=date, assigned_to=staff_profile
                        ).exists()
                    )
                    if swap_exists:
                        # Use the TA as-is (belongs to original staff) — swap grants permission
                        ta_obj = _TA.objects.filter(pk=int(teaching_assignment_id), is_active=True).first()
                    else:
                        raise PermissionDenied('Invalid teaching_assignment_id for current staff')
                # Allow section-scoped or department-wide assignments; reject mismatched section.
                if ta_obj and section is not None and getattr(ta_obj, 'section_id', None) not in (None, getattr(section, 'id', None)):
                    raise PermissionDenied('teaching_assignment_id does not match section')
                teach_assign = ta_obj
            except PermissionDenied:
                raise
            except Exception:
                raise PermissionDenied('Invalid teaching_assignment_id')

        # Resolve a specific TeachingAssignment (subject option) for this staff.
        if teach_assign is None and staff_profile:
            try:
                from .models import TeachingAssignment as _TA
                ta_qs = _TA.objects.filter(is_active=True, staff=staff_profile).filter(Q(section=section) | Q(section__isnull=True))
                cr_id = getattr(ta, 'curriculum_row_id', None) if ta is not None else None
                _cr_name_bulk = getattr(getattr(ta, 'curriculum_row', None), 'course_name', None) if ta is not None else None
                _sec_dept_id_bk = getattr(getattr(getattr(section, 'batch', None), 'course', None), 'department_id', None) if section else None
                if cr_id:
                    # Filter by the period's curriculum_row — only match if staff actually teaches it.
                    ta_qs = ta_qs.filter(
                        Q(curriculum_row_id=cr_id) |
                        Q(elective_subject__parent_id=cr_id) |
                        Q(elective_subject__department_group__isnull=False,
                          elective_subject__parent__course_name=_cr_name_bulk,
                          elective_subject__department_group__department_mappings__department_id=_sec_dept_id_bk,
                          elective_subject__department_group__department_mappings__is_active=True)
                    )
                    teach_assign = ta_qs.order_by(
                        models.Case(
                            models.When(section=section, then=models.Value(0)),
                            default=models.Value(1),
                            output_field=models.IntegerField(),
                        ),
                        'id',
                    ).first()
                elif ta is None:
                    # No timetable assignment context at all — fall back to any active TA for
                    # this staff+section (e.g. attendance created directly without a timetable slot).
                    teach_assign = ta_qs.order_by(
                        models.Case(
                            models.When(section=section, then=models.Value(0)),
                            default=models.Value(1),
                            output_field=models.IntegerField(),
                        ),
                        'id',
                    ).first()
                # else: ta exists but curriculum_row is None (custom subject_text / substitute period).
                # Don't pick a random TA — keep teach_assign=None to prevent the wrong subject's
                # TeachingAssignment from being linked to this attendance session.
            except Exception:
                teach_assign = None

        perms = get_user_permissions(user)
        
        # Check if there are batch assignments for this period
        try:
            from timetable.models import TimetableAssignment as _TimetableAssignment
            batch_assignments = _TimetableAssignment.objects.filter(
                section=section, period=period, day=day, subject_batch__isnull=False
            ) if section and period and day else _TimetableAssignment.objects.none()
        except Exception:
            batch_assignments = _TimetableAssignment.objects.none()
        
        # If batch assignments exist, ONLY batch staff can access - block everyone else
        if batch_assignments.exists() and not (user.is_superuser or 'academics.mark_attendance' in perms):
            # Check if staff has access to any of the batches
            has_batch_access = False
            if ta and getattr(ta, 'subject_batch', None):
                batch = ta.subject_batch
                is_batch_staff = (
                    getattr(batch, 'created_by_id', None) == staff_profile.id or
                    getattr(batch, 'staff_id', None) == staff_profile.id
                )
                if is_batch_staff:
                    has_batch_access = True
            
            if not has_batch_access:
                # Get batch name for error message
                try:
                    batch_ta = batch_assignments.first()
                    batch_name = batch_ta.subject_batch.name if batch_ta and batch_ta.subject_batch else 'assigned batch'
                    return Response({
                        'error': f'This period is assigned to "{batch_name}". Only staff assigned to that batch can mark attendance.',
                        'batch_name': batch_name
                    }, status=403)
                except Exception:
                    return Response({
                        'error': 'This period is assigned to a batch. Only staff assigned to that batch can mark attendance.'
                    }, status=403)
        
        # Regular permission check for non-batch periods
        if not batch_assignments.exists():
            if not (ta or teach_assign or 'academics.mark_attendance' in perms or user.is_superuser):
                # Also allow if there's a swap-assigned session for this staff
                is_swap_assigned = (
                    section and period and date and staff_profile and
                    PeriodAttendanceSession.objects.filter(
                        section=section, period=period, date=date, assigned_to=staff_profile
                    ).exists()
                )
                if not is_swap_assigned:
                    raise PermissionDenied('You are not allowed to mark attendance for this period')

        # If this period/date has already been assigned to another staff member,
        # the original staff must be blocked from marking (even through a
        # different teaching_assignment lookup path).
        assigned_conflict = PeriodAttendanceSession.objects.select_related('assigned_to', 'assigned_to__user').filter(
            section=section,
            period=period,
            date=date,
            assigned_to__isnull=False,
        ).exclude(assigned_to=staff_profile).order_by('-id').first()
        if assigned_conflict and not (user.is_superuser or 'analytics.edit_all_analytics' in perms):
            assigned_name = ''
            try:
                assigned_name = assigned_conflict.assigned_to.user.get_full_name() if assigned_conflict.assigned_to and assigned_conflict.assigned_to.user else assigned_conflict.assigned_to.staff_id
            except Exception:
                assigned_name = ''
            return Response({
                'error': f'This period attendance has been assigned to {assigned_name}. You cannot mark attendance.',
                'assigned_to': {
                    'name': assigned_name,
                    'staff_id': getattr(assigned_conflict.assigned_to, 'staff_id', ''),
                },
            }, status=403)

        with transaction.atomic():
            # Optionally create a temporary special timetable entry for this date/period
            try:
                if data.get('create_special'):
                    from timetable.models import SpecialTimetable, SpecialTimetableEntry
                    # create a simple SpecialTimetable container for this section and staff
                    st_name = f"Temp-{section.id}-{period.id}-{str(date)}"
                    special_tt, _ = SpecialTimetable.objects.get_or_create(section=section, name=st_name, defaults={'created_by': staff_profile, 'is_active': True})
                    SpecialTimetableEntry.objects.get_or_create(timetable=special_tt, date=date, period=period, defaults={'staff': staff_profile, 'curriculum_row': getattr(ta, 'curriculum_row', None) if ta is not None else None, 'subject_batch': getattr(ta, 'subject_batch', None) if ta is not None else None, 'subject_text': getattr(ta, 'subject_text', None) if ta is not None else None, 'is_active': True})
            except Exception:
                pass
            # IMPORTANT: session must be subject-wise. Use resolved teaching_assignment
            # (staff+subject, including elective options) to prevent overwrites.
            # Also include subject_batch to allow separate sessions for different batches.
            subject_batch = getattr(ta, 'subject_batch', None) if ta else None
            lookup = {'section': section, 'period': period, 'date': date, 'teaching_assignment': teach_assign, 'subject_batch': subject_batch}
            if teach_assign is None:
                # When teaching_assignment cannot be resolved (common for electives with no-staff
                # timetable entries), use created_by as discriminator to avoid cross-staff session sharing.
                lookup['created_by'] = staff_profile
                # Also include timetable_assignment for metadata, but not as unique key.

            # Before get_or_create: if a session exists that is ASSIGNED to this staff
            # (swap scenario), use that session regardless of created_by, to avoid creating a duplicate.
            assigned_to_me_sess = PeriodAttendanceSession.objects.filter(
                section=section, period=period, date=date, assigned_to=staff_profile
            ).order_by('-id').first()

            if assigned_to_me_sess:
                session, created = assigned_to_me_sess, False
            else:
                session, created = PeriodAttendanceSession.objects.get_or_create(
                    **lookup,
                    defaults={'created_by': staff_profile, 'timetable_assignment': ta, 'teaching_assignment': teach_assign, 'subject_batch': subject_batch}
                )

            # Enforce assigned_to: if session is assigned to someone else, block the original staff
            if not created and session.assigned_to_id and session.assigned_to_id != getattr(staff_profile, 'id', None):
                if not (user.is_superuser or 'analytics.edit_all_analytics' in perms):
                    assigned_name = ''
                    try:
                        session.refresh_from_db(fields=['assigned_to'])
                        assigned_name = session.assigned_to.user.get_full_name() if session.assigned_to and session.assigned_to.user else session.assigned_to.staff_id
                    except Exception:
                        pass
                    return Response({
                        'error': f'This period attendance has been assigned to {assigned_name}. You cannot mark attendance.',
                        'assigned_to': {'name': assigned_name, 'staff_id': getattr(session.assigned_to, 'staff_id', '')}
                    }, status=403)
            # Keep metadata up to date
            dirty_fields = []
            if staff_profile and session.created_by_id is None:
                session.created_by = staff_profile
                dirty_fields.append('created_by')
            if ta is not None and session.timetable_assignment_id != getattr(ta, 'id', None):
                session.timetable_assignment = ta
                dirty_fields.append('timetable_assignment')
            if teach_assign is not None and session.teaching_assignment_id != getattr(teach_assign, 'id', None):
                session.teaching_assignment = teach_assign
                dirty_fields.append('teaching_assignment')
            # Update subject_batch from timetable assignment if needed
            ta_subject_batch = getattr(ta, 'subject_batch', None) if ta is not None else None
            if ta_subject_batch is not None and session.subject_batch_id != getattr(ta_subject_batch, 'id', None):
                session.subject_batch = ta_subject_batch
                dirty_fields.append('subject_batch')
            if dirty_fields:
                session.save(update_fields=dirty_fields)

            out = []
            # If timetable assignment has a subject_batch defined, use that student list
            students_source = None
            if ta and getattr(ta, 'subject_batch', None):
                try:
                    students_source = list(ta.subject_batch.students.all())
                except Exception:
                    students_source = None
            # If there's no subject_batch but staff is assigned to an elective sub-option,
            # use ElectiveChoice mappings to determine students for this elective.
            if students_source is None:
                try:
                    # Prefer the resolved teaching assignment (especially when provided by client)
                    if teach_assign and getattr(teach_assign, 'elective_subject', None):
                        from curriculum.models import ElectiveChoice
                        es = teach_assign.elective_subject
                        choices = ElectiveChoice.objects.filter(elective_subject=es, is_active=True).select_related('student')
                        students_source = [getattr(c, 'student') for c in choices if getattr(c, 'student', None) is not None]
                    elif ta and not getattr(ta, 'subject_batch', None):
                        from .models import TeachingAssignment as _TA
                        # find teaching assignment with elective_subject for this staff matching the curriculum_row parent
                        ta_match = _TA.objects.filter(is_active=True, staff=staff_profile, elective_subject__isnull=False).filter(
                            Q(elective_subject__parent=getattr(ta, 'curriculum_row', None))
                        ).filter(Q(section=section) | Q(section__isnull=True)).first()
                        if ta_match and getattr(ta_match, 'elective_subject', None):
                            from curriculum.models import ElectiveChoice
                            es = ta_match.elective_subject
                            choices = ElectiveChoice.objects.filter(elective_subject=es, is_active=True).select_related('student')
                            students_source = [getattr(c, 'student') for c in choices if getattr(c, 'student', None) is not None]
                except Exception:
                    students_source = None

            from .models import StudentProfile as _StudentProfile
            for rec in records:
                # BulkRecordSerializer provides `student_id` and `status`
                sid = rec.get('student_id')
                status_val = rec.get('status')
                if not sid:
                    continue
                stu = _StudentProfile.objects.filter(pk=int(sid)).first()
                if not stu:
                    continue
                if students_source is not None:
                    # ensure student belongs to subject_batch
                    if not any(getattr(s, 'pk', None) == getattr(stu, 'pk', None) for s in students_source):
                        # skip students not in batch
                        continue

                # ── Daily-attendance override ──────────────────────────────────────────
                # If the student has a daily attendance record for this section/date:
                #   A (Absent)  → force 'A' status (locked; cannot change)
                #   OD / LEAVE  → force the same status (locked; cannot differ per period)
                #   LATE        → force Present ('P') for every period
                # Any other daily status leaves the submitted status_val unchanged.
                daily_override = None
                try:
                    from .models import DailyAttendanceSession as _DAS, DailyAttendanceRecord as _DAR
                    _daily_session = _DAS.objects.filter(
                        section_id=session.section_id, date=session.date
                    ).first()
                    if _daily_session:
                        # Check if daily attendance session is locked - if so, prevent saving period attendance
                        if _daily_session.is_locked:
                            # Skip this record - daily attendance is locked, cannot modify period records
                            continue
                        
                        _daily_rec = _DAR.objects.filter(session=_daily_session, student=stu).first()
                        if _daily_rec:
                            if _daily_rec.status == 'A':
                                # Force Absent when daily attendance is Absent
                                status_val = 'A'
                                daily_override = 'ABSENT'
                            elif _daily_rec.status in ('OD', 'LEAVE'):
                                status_val = _daily_rec.status
                                daily_override = _daily_rec.status
                            elif _daily_rec.status == 'LATE':
                                status_val = 'P'
                                daily_override = 'LATE'
                except Exception:
                    pass  # never block period-attendance saving due to override errors
                # ──────────────────────────────────────────────────────────────────────

                obj, created = PeriodAttendanceRecord.objects.update_or_create(
                    session=session, student=stu,
                    defaults={'status': status_val, 'marked_by': staff_profile}
                )
                out.append({'id': obj.id, 'student_id': getattr(obj.student, 'id', None), 'status': obj.status, 'daily_override': daily_override})

            resp_ser = PeriodAttendanceSessionSerializer(session, context={'request': request})
            return Response(resp_ser.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='bulk-mark-range')
    def bulk_mark_range(self, request):
        """Bulk mark attendance for a date range or explicit date list.

        Expected payload (supports both forms):
          Form A: {section_id, period_id, start_date, end_date, status, student_ids}
          Form B: {assignments: [{section_id, period_id},...], dates: [...], status, student_ids}
        """
        import datetime
        data = request.data or {}
        status_val = data.get('status') or 'P'
        student_ids = data.get('student_ids') or []

        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Only staff may perform bulk range marking')

        # ── 1. Build dates_iter ──────────────────────────────────────────────
        if data.get('dates'):
            try:
                dates_iter = sorted(datetime.date.fromisoformat(d) for d in data['dates'])
            except Exception:
                return Response({'detail': 'Invalid dates list'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            try:
                sd = datetime.date.fromisoformat(data.get('start_date'))
                ed = datetime.date.fromisoformat(data.get('end_date'))
            except Exception:
                return Response({'detail': 'Invalid dates'}, status=status.HTTP_400_BAD_REQUEST)
            if ed < sd:
                return Response({'detail': 'end_date must be >= start_date'}, status=status.HTTP_400_BAD_REQUEST)
            delta = datetime.timedelta(days=1)
            day = sd
            dates_iter = []
            while day <= ed:
                dates_iter.append(day)
                day += delta

        if not dates_iter:
            return Response({'results': []})

        # ── 2. Resolve assignments ────────────────────────────────────────────
        from .models import Section as _Section
        from timetable.models import TimetableSlot, TimetableAssignment, SpecialTimetableEntry
        from .models import (
            PeriodAttendanceSession, PeriodAttendanceRecord,
            StudentSectionAssignment, StudentProfile as _SP,
            TeachingAssignment as _TA,
        )

        assignments_payload = data.get('assignments')
        assignments_list = []
        if assignments_payload and isinstance(assignments_payload, (list, tuple)):
            raw_sids = []
            raw_pids = []
            for item in assignments_payload:
                try:
                    raw_sids.append(int(item['section_id']))
                    raw_pids.append(int(item['period_id']))
                except Exception:
                    pass
            secs_map = {s.pk: s for s in _Section.objects.filter(pk__in=raw_sids)}
            pers_map = {p.pk: p for p in TimetableSlot.objects.filter(pk__in=raw_pids)}
            seen = set()
            for item in assignments_payload:
                try:
                    sid, pid = int(item['section_id']), int(item['period_id'])
                    if (sid, pid) in seen:
                        continue
                    seen.add((sid, pid))
                    s, p = secs_map.get(sid), pers_map.get(pid)
                    if s and p:
                        assignments_list.append((s, p))
                except Exception:
                    pass
        else:
            try:
                s = _Section.objects.filter(pk=int(data.get('section_id'))).first()
                p = TimetableSlot.objects.filter(pk=int(data.get('period_id'))).first()
                if s and p:
                    assignments_list.append((s, p))
            except Exception:
                pass

        if not assignments_list:
            return Response({'results': []})

        perms = get_user_permissions(user)
        all_section_ids = list({s.pk for s, _ in assignments_list})
        all_period_ids  = list({p.pk for _, p in assignments_list})

        # ── 3. PRE-FETCH: TimetableAssignments ───────────────────────────────
        # ta_staff[(sec_id, per_id, dow)] = TA owned by this staff
        # ta_any  [(sec_id, per_id, dow)] = any TA (for permission fallback)
        # ta_anysp[(sec_id, per_id)]      = any TA regardless of day
        ta_staff = {}
        ta_any   = {}
        ta_anysp = {}
        for ta in TimetableAssignment.objects.filter(
            section_id__in=all_section_ids,
            period_id__in=all_period_ids,
        ).select_related('curriculum_row'):
            k3 = (ta.section_id, ta.period_id, ta.day)
            if ta.staff_id == staff_profile.pk:
                ta_staff.setdefault(k3, ta)
            ta_any.setdefault(k3, ta)
            ta_anysp.setdefault((ta.section_id, ta.period_id), ta)

        # ── 4. PRE-FETCH: TeachingAssignments for staff ───────────────────────
        staff_ta_list = list(
            _TA.objects.filter(is_active=True, staff=staff_profile)
            .select_related('curriculum_row', 'elective_subject', 'section')
        )
        # Index by curriculum_row_id (also by elective parent) for fast lookup
        teach_by_cr: dict = {}
        for sta in staff_ta_list:
            if sta.curriculum_row_id:
                teach_by_cr.setdefault(sta.curriculum_row_id, []).append(sta)
            if sta.elective_subject_id and getattr(sta.elective_subject, 'parent_id', None):
                teach_by_cr.setdefault(sta.elective_subject.parent_id, []).append(sta)

        def _staff_ta_matches_cr(cr_id, section_obj):
            """Return True if staff has a TeachingAssignment for this curriculum_row."""
            for sta in teach_by_cr.get(cr_id, []):
                if sta.section_id is None or sta.section_id == section_obj.pk:
                    return True
            return False

        def _best_teach_assign(cr_id, section_obj):
            """Pick the best TeachingAssignment for (staff, cr_id, section)."""
            candidates = teach_by_cr.get(cr_id, []) if cr_id else staff_ta_list
            for sta in candidates:
                if sta.section_id == section_obj.pk:
                    return sta
            for sta in candidates:
                if sta.section_id is None:
                    return sta
            return None

        # ── 5. PRE-FETCH: SpecialTimetableEntries ────────────────────────────
        special_map = {}  # (sec_id, per_id, date) → SpecialTimetableEntry
        for se in SpecialTimetableEntry.objects.filter(
            timetable__section_id__in=all_section_ids,
            period_id__in=all_period_ids,
            date__in=dates_iter,
            is_active=True,
        ).select_related('timetable'):
            key = (se.timetable.section_id, se.period_id, se.date)
            special_map.setdefault(key, se)

        # ── 6. PRE-FETCH: students per section ───────────────────────────────
        section_students: dict = {}
        if not student_ids:
            seen_pks: dict = {}
            for ssa in (
                StudentSectionAssignment.objects
                .filter(section_id__in=all_section_ids, end_date__isnull=True)
                .exclude(student__status__in=['INACTIVE', 'DEBAR'])
                .select_related('student')
            ):
                bucket = section_students.setdefault(ssa.section_id, [])
                seen_s = seen_pks.setdefault(ssa.section_id, set())
                if ssa.student_id not in seen_s:
                    bucket.append(ssa.student)
                    seen_s.add(ssa.student_id)
            for stu in _SP.objects.filter(section_id__in=all_section_ids).exclude(status__in=['INACTIVE', 'DEBAR']).select_related('user'):
                bucket = section_students.setdefault(stu.section_id, [])
                seen_s = seen_pks.setdefault(stu.section_id, set())
                if stu.pk not in seen_s:
                    bucket.append(stu)
                    seen_s.add(stu.pk)

        specific_students = (
            list(_SP.objects.filter(pk__in=[int(sid) for sid in student_ids]))
            if student_ids else []
        )

        # ── 7. PLAN: determine (session_lookup, students) per (day, assignment) ─
        session_plan = []
        for day in dates_iter:
            dow = day.isoweekday()
            for (section_obj, period_obj) in assignments_list:
                se = special_map.get((section_obj.pk, period_obj.pk, day))
                if se:
                    ta = None
                    allow = False

                    # Special timetable entries must only be markable by:
                    # - explicitly assigned staff, OR
                    # - staff tied to the subject batch, OR
                    # - staff who teaches the linked curriculum row.
                    if getattr(se, 'staff_id', None) == staff_profile.pk:
                        allow = True

                    sb = getattr(se, 'subject_batch', None)
                    if not allow and sb:
                        if (
                            getattr(sb, 'created_by_id', None) == staff_profile.pk
                            or getattr(sb, 'staff_id', None) == staff_profile.pk
                        ):
                            allow = True

                    se_cr_id = getattr(se, 'curriculum_row_id', None)
                    if not allow and se_cr_id:
                        if _staff_ta_matches_cr(se_cr_id, section_obj):
                            allow = True

                    if not allow and not (user.is_superuser or 'academics.mark_attendance' in perms):
                        continue
                else:
                    ta = ta_staff.get((section_obj.pk, period_obj.pk, dow))
                    allow = ta is not None
                    if not allow:
                        any_ta = ta_any.get((section_obj.pk, period_obj.pk, dow))
                        if any_ta and _staff_ta_matches_cr(any_ta.curriculum_row_id, section_obj):
                            allow, ta = True, any_ta
                    if not allow:
                        anysp = ta_anysp.get((section_obj.pk, period_obj.pk))
                        if anysp and _staff_ta_matches_cr(anysp.curriculum_row_id, section_obj):
                            allow, ta = True, anysp
                    if not allow and not (user.is_superuser or 'academics.mark_attendance' in perms):
                        continue

                if se:
                    cr_id = getattr(se, 'curriculum_row_id', None)
                    teach_assign = _best_teach_assign(cr_id, section_obj) if cr_id else None
                else:
                    cr_id = getattr(ta, 'curriculum_row_id', None)
                    teach_assign = _best_teach_assign(cr_id, section_obj)

                if specific_students:
                    target_students = specific_students
                else:
                    target_students = section_students.get(section_obj.pk, [])

                lookup = {
                    'section': section_obj,
                    'period': period_obj,
                    'date': day,
                    'teaching_assignment': teach_assign,
                }
                if teach_assign is None:
                    lookup['created_by'] = staff_profile

                session_plan.append({
                    'lookup': lookup,
                    'defaults': {
                        'created_by': staff_profile,
                        'timetable_assignment': ta,
                        'teaching_assignment': teach_assign,
                    },
                    'students': target_students,
                    'day': day,
                    'section_obj': section_obj,
                    'period_obj': period_obj,
                    'teach_assign_id': teach_assign.pk if teach_assign else None,
                })

        if not session_plan:
            return Response({'results': []})

        # ── 8. SESSIONS: pre-fetch existing, create only missing ─────────────
        existing_sessions = {}
        for sess in PeriodAttendanceSession.objects.filter(
            section_id__in=all_section_ids,
            period_id__in=all_period_ids,
            date__in=dates_iter,
        ):
            k = (sess.section_id, sess.period_id, sess.date, sess.teaching_assignment_id)
            existing_sessions[k] = sess
            if sess.teaching_assignment_id is None:
                existing_sessions[(sess.section_id, sess.period_id, sess.date, None, sess.created_by_id)] = sess

        resolved: dict = {}
        for idx, plan in enumerate(session_plan):
            lk = plan['lookup']
            k = (lk['section'].pk, lk['period'].pk, lk['date'], plan['teach_assign_id'])
            sess = existing_sessions.get(k)
            if sess is None and plan['teach_assign_id'] is None:
                sess = existing_sessions.get((lk['section'].pk, lk['period'].pk, lk['date'], None, staff_profile.pk))
            if sess:
                resolved[idx] = sess
            else:
                sess, _ = PeriodAttendanceSession.objects.get_or_create(
                    **lk, defaults=plan['defaults']
                )
                resolved[idx] = sess
                existing_sessions[k] = sess

        # ── 9. RECORDS: bulk create / update ─────────────────────────────────
        all_session_ids = [s.pk for s in resolved.values()]
        existing_records = {
            (r.session_id, r.student_id): r
            for r in PeriodAttendanceRecord.objects.filter(session_id__in=all_session_ids)
        }

        to_create = []
        to_update = []
        for idx, plan in enumerate(session_plan):
            sess = resolved.get(idx)
            if not sess:
                continue
            for stu in plan['students']:
                key = (sess.pk, stu.pk)
                if key in existing_records:
                    rec = existing_records[key]
                    if rec.status != status_val or rec.marked_by_id != staff_profile.pk:
                        rec.status = status_val
                        rec.marked_by = staff_profile
                        to_update.append(rec)
                else:
                    to_create.append(PeriodAttendanceRecord(
                        session=sess, student=stu,
                        status=status_val, marked_by=staff_profile,
                    ))
                    # prevent duplicates within the same batch
                    existing_records[key] = True  # type: ignore[assignment]

        if to_create:
            PeriodAttendanceRecord.objects.bulk_create(to_create, ignore_conflicts=True)
        if to_update:
            PeriodAttendanceRecord.objects.bulk_update(to_update, ['status', 'marked_by'], batch_size=500)

        out_sessions = [
            {
                'date': plan['day'].isoformat(),
                'section_id': plan['section_obj'].pk,
                'period_id': plan['period_obj'].pk,
                'session_id': resolved[idx].pk,
            }
            for idx, plan in enumerate(session_plan)
            if idx in resolved
        ]
        return Response({'results': out_sessions})

    @action(detail=False, methods=['post'], url_path='bulk-mark-statuses')
    def bulk_mark_statuses(self, request):
        """Mark attendance with individual per-student statuses per date.

        Payload: {
            assignments: [{section_id, period_id}],
            date_records: [{date: "YYYY-MM-DD", records: [{student_id: N, status: "P"|"A"}]}]
        }
        """
        import datetime as _dt
        data = request.data or {}
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Only staff may perform bulk attendance marking')

        from .models import (
            Section as _Section, PeriodAttendanceSession as _PAS,
            PeriodAttendanceRecord as _PAR, StudentProfile as _SP,
            TeachingAssignment as _TA,
        )
        from timetable.models import TimetableSlot as _TSlot, TimetableAssignment as _TAssign

        perms = get_user_permissions(user)
        assignments_payload = data.get('assignments') or []
        date_records = data.get('date_records') or []

        if not assignments_payload:
            return Response({'detail': 'assignments required'}, status=status.HTTP_400_BAD_REQUEST)
        if not date_records:
            return Response({'detail': 'date_records required'}, status=status.HTTP_400_BAD_REQUEST)

        # ── 1. Resolve assignments in bulk ────────────────────────────────────
        raw_sids = [int(a['section_id']) for a in assignments_payload if 'section_id' in a]
        raw_pids = [int(a['period_id'])  for a in assignments_payload if 'period_id'  in a]
        secs_map = {s.pk: s for s in _Section.objects.filter(pk__in=raw_sids)}
        pers_map = {p.pk: p for p in _TSlot.objects.filter(pk__in=raw_pids)}
        seen_sp = set()
        assignments_list = []
        for a in assignments_payload:
            try:
                sid, pid = int(a['section_id']), int(a['period_id'])
                if (sid, pid) in seen_sp:
                    continue
                seen_sp.add((sid, pid))
                s, p = secs_map.get(sid), pers_map.get(pid)
                if s and p:
                    assignments_list.append((s, p))
            except Exception:
                continue

        if not assignments_list:
            return Response({'results': []})

        # ── 2. Parse dates ────────────────────────────────────────────────────
        all_dates = []
        for dr in date_records:
            try:
                all_dates.append(_dt.date.fromisoformat(dr['date']))
            except Exception:
                pass
        all_dates = sorted(set(all_dates))

        all_section_ids = list({s.pk for s, _ in assignments_list})
        all_period_ids  = list({p.pk for _, p in assignments_list})

        # ── 3. PRE-FETCH: TimetableAssignments ───────────────────────────────
        ta_staff = {}   # (sec_id, per_id, dow) → TA (this staff)
        ta_any   = {}   # (sec_id, per_id, dow) → TA (any staff)
        ta_anysp = {}   # (sec_id, per_id)      → TA (any day)
        for ta in _TAssign.objects.filter(
            section_id__in=all_section_ids,
            period_id__in=all_period_ids,
        ).select_related('curriculum_row'):
            k3 = (ta.section_id, ta.period_id, ta.day)
            if ta.staff_id == staff_profile.pk:
                ta_staff.setdefault(k3, ta)
            ta_any.setdefault(k3, ta)
            ta_anysp.setdefault((ta.section_id, ta.period_id), ta)

        # ── 4. PRE-FETCH: TeachingAssignments for staff ───────────────────────
        staff_ta_list = list(
            _TA.objects.filter(is_active=True, staff=staff_profile)
            .select_related('curriculum_row', 'elective_subject', 'section')
        )
        teach_by_cr: dict = {}
        for sta in staff_ta_list:
            if sta.curriculum_row_id:
                teach_by_cr.setdefault(sta.curriculum_row_id, []).append(sta)
            if sta.elective_subject_id and getattr(sta.elective_subject, 'parent_id', None):
                teach_by_cr.setdefault(sta.elective_subject.parent_id, []).append(sta)

        def _ta_allowed(cr_id, section_obj):
            for sta in teach_by_cr.get(cr_id, []):
                if sta.section_id is None or sta.section_id == section_obj.pk:
                    return True
            return False

        def _best_teach(cr_id, section_obj):
            candidates = teach_by_cr.get(cr_id, []) if cr_id else staff_ta_list
            for sta in candidates:
                if sta.section_id == section_obj.pk:
                    return sta
            for sta in candidates:
                if sta.section_id is None:
                    return sta
            return None

        # ── 5. PRE-FETCH: students referenced in date_records ────────────────
        all_student_ids = set()
        for dr in date_records:
            for rec in (dr.get('records') or []):
                try:
                    all_student_ids.add(int(rec['student_id']))
                except Exception:
                    pass
        students_map = {s.pk: s for s in _SP.objects.filter(pk__in=all_student_ids)}

        # ── 6. PLAN sessions ─────────────────────────────────────────────────
        # session_plan[key] = {lookup, defaults, records: [(student, status)]}
        session_plan_map: dict = {}  # (sec_id, per_id, date, teach_id) → plan dict
        for dr in date_records:
            try:
                day = _dt.date.fromisoformat(dr['date'])
            except Exception:
                continue
            dow = day.isoweekday()
            records = dr.get('records') or []

            for (section_obj, period_obj) in assignments_list:
                ta = ta_staff.get((section_obj.pk, period_obj.pk, dow))
                allow = ta is not None
                if not allow:
                    any_ta = ta_any.get((section_obj.pk, period_obj.pk, dow))
                    if any_ta and _ta_allowed(any_ta.curriculum_row_id, section_obj):
                        allow, ta = True, any_ta
                if not allow:
                    anysp = ta_anysp.get((section_obj.pk, period_obj.pk))
                    if anysp and _ta_allowed(anysp.curriculum_row_id, section_obj):
                        allow, ta = True, anysp
                if not allow and not (user.is_superuser or 'academics.mark_attendance' in perms):
                    continue

                cr_id = getattr(ta, 'curriculum_row_id', None) if ta else None
                teach_assign = _best_teach(cr_id, section_obj)
                ta_id = teach_assign.pk if teach_assign else None

                plan_key = (section_obj.pk, period_obj.pk, day, ta_id)
                if plan_key not in session_plan_map:
                    lookup = {
                        'section': section_obj, 'period': period_obj,
                        'date': day, 'teaching_assignment': teach_assign,
                    }
                    if teach_assign is None:
                        lookup['created_by'] = staff_profile
                    session_plan_map[plan_key] = {
                        'lookup': lookup,
                        'defaults': {
                            'created_by': staff_profile,
                            'timetable_assignment': ta,
                            'teaching_assignment': teach_assign,
                        },
                        'records': [],
                        'day': day,
                        'section_obj': section_obj,
                        'period_obj': period_obj,
                        'ta_id': ta_id,
                    }

                valid_statuses = {'P', 'A', 'OD', 'LEAVE', 'LATE', 'HD'}
                for rec in records:
                    try:
                        stu = students_map.get(int(rec['student_id']))
                        if not stu:
                            continue
                        sv = rec.get('status', 'P')
                        if sv not in valid_statuses:
                            sv = 'P'
                        session_plan_map[plan_key]['records'].append((stu, sv))
                    except Exception:
                        continue

        if not session_plan_map:
            return Response({'results': []})

        # ── 7. SESSIONS: pre-fetch existing, create only missing ─────────────
        existing_sessions = {}
        for sess in _PAS.objects.filter(
            section_id__in=all_section_ids,
            period_id__in=all_period_ids,
            date__in=all_dates,
        ):
            k = (sess.section_id, sess.period_id, sess.date, sess.teaching_assignment_id)
            existing_sessions[k] = sess
            if sess.teaching_assignment_id is None:
                existing_sessions[(sess.section_id, sess.period_id, sess.date, None, sess.created_by_id)] = sess

        resolved_plans = []
        for plan_key, plan in session_plan_map.items():
            lk = plan['lookup']
            k = (lk['section'].pk, lk['period'].pk, lk['date'], plan['ta_id'])
            sess = existing_sessions.get(k)
            if sess is None and plan['ta_id'] is None:
                sess = existing_sessions.get((lk['section'].pk, lk['period'].pk, lk['date'], None, staff_profile.pk))
            if not sess:
                sess, _ = _PAS.objects.get_or_create(**lk, defaults=plan['defaults'])
                existing_sessions[k] = sess
            resolved_plans.append((sess, plan))

        # ── 8. RECORDS: bulk create / update ─────────────────────────────────
        all_session_ids = [sess.pk for sess, _ in resolved_plans]
        existing_records = {
            (r.session_id, r.student_id): r
            for r in _PAR.objects.filter(session_id__in=all_session_ids)
        }

        to_create = []
        to_update = []
        seen_keys: set = set()
        for sess, plan in resolved_plans:
            for (stu, sv) in plan['records']:
                key = (sess.pk, stu.pk)
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                if key in existing_records:
                    rec = existing_records[key]
                    if rec.status != sv or rec.marked_by_id != staff_profile.pk:
                        rec.status = sv
                        rec.marked_by = staff_profile
                        to_update.append(rec)
                else:
                    to_create.append(_PAR(
                        session=sess, student=stu,
                        status=sv, marked_by=staff_profile,
                    ))

        if to_create:
            _PAR.objects.bulk_create(to_create, ignore_conflicts=True)
        if to_update:
            _PAR.objects.bulk_update(to_update, ['status', 'marked_by'], batch_size=500)

        out_sessions = [
            {
                'date': plan['day'].isoformat(),
                'section_id': plan['section_obj'].pk,
                'period_id': plan['period_obj'].pk,
                'session_id': sess.pk,
                'records_count': len(plan['records']),
            }
            for sess, plan in resolved_plans
        ]
        return Response({'results': out_sessions})

    @action(detail=True, methods=['post'], url_path='lock')
    def lock_session(self, request, pk=None):
        """Lock an attendance session to prevent further edits."""
        session = self.get_object()
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        
        # Check permissions
        perms = get_user_permissions(user)
        is_creator = session.created_by == staff_profile if staff_profile else False
        is_assigned = False
        
        if session.timetable_assignment and staff_profile:
            is_assigned = session.timetable_assignment.staff == staff_profile
        
        if not (is_creator or is_assigned or 'academics.mark_attendance' in perms or user.is_superuser):
            raise PermissionDenied('You do not have permission to lock this attendance session')
        
        session.is_locked = True
        session.save(update_fields=['is_locked'])
        
        serializer = self.get_serializer(session)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='unlock')
    def unlock_session(self, request, pk=None):
        """Unlock an attendance session to allow edits."""
        session = self.get_object()
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        
        # Check permissions - stricter for unlocking
        perms = get_user_permissions(user)
        if not ('academics.mark_attendance' in perms or user.is_superuser):
            raise PermissionDenied('You do not have permission to unlock this attendance session')
        
        session.is_locked = False
        session.save(update_fields=['is_locked'])
        
        serializer = self.get_serializer(session)
        return Response(serializer.data)


class AttendanceUnlockRequestViewSet(viewsets.ModelViewSet):
    """Manage attendance unlock requests: create by staff, list/approve by admins."""
    queryset = AttendanceUnlockRequest.objects.select_related('session__section', 'requested_by', 'reviewed_by').order_by('-requested_at')
    serializer_class = AttendanceUnlockRequestSerializer
    permission_classes = (IsAuthenticated,)
    pagination_class = None  # Disable pagination for simpler response
    logger = logging.getLogger(__name__)

    def get_queryset(self):
        user = self.request.user
        perms = get_user_permissions(user)
        self.logger.info(f"User {user.username} permissions: {perms}")
        # Only users with analytics.view_all_analytics (or superuser) can view all requests
        if 'analytics.view_all_analytics' in perms or user.is_superuser:
            self.logger.info(f"User {user.username} has admin access - returning all requests")
            # Return fresh queryset for admins to see all requests
            qs = AttendanceUnlockRequest.objects.select_related('session__section', 'requested_by', 'reviewed_by').order_by('-requested_at')
            self.logger.info(f"Admin queryset count: {qs.count()}")
            return qs
        staff_profile = getattr(user, 'staff_profile', None)
        if staff_profile:
            self.logger.info(f"Staff profile found: ID={staff_profile.id}, Staff_ID={staff_profile.staff_id}")
            # Return filtered queryset for regular staff to see only their requests
            qs = AttendanceUnlockRequest.objects.select_related('session__section', 'requested_by', 'reviewed_by').filter(requested_by=staff_profile).order_by('-requested_at')
            self.logger.info(f"Staff {user.username} queryset count: {qs.count()}")
            # Log each request for debugging
            for req in qs:
                self.logger.info(f"  - Request #{req.id}: status={req.status}, session={req.session_id}, requested_by={req.requested_by_id}")
            return qs
        self.logger.warning(f"User {user.username} has no staff profile - returning empty queryset")
        return AttendanceUnlockRequest.objects.none()

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        self.logger.info(f"Returning {len(serializer.data)} requests to user {request.user.username}")
        for item in serializer.data:
            self.logger.info(f"  - Request #{item['id']}: status={item['status']}")
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return Response({'detail': 'Only staff may request unlocks'}, status=403)

        session_id = request.data.get('session') or request.data.get('session_id')
        note = request.data.get('note', '')
        try:
            session = PeriodAttendanceSession.objects.filter(pk=int(session_id)).first()
        except Exception:
            session = None
        if not session:
            return Response({'detail': 'Session not found'}, status=404)

        # Check if there's already a pending/in-progress request for this session
        existing_pending = AttendanceUnlockRequest.objects.filter(
            session=session,
            status__in=['PENDING', 'HOD_APPROVED']
        ).first()
        
        if existing_pending:
            # Return existing request instead of creating duplicate
            ser = AttendanceUnlockRequestSerializer(existing_pending, context={'request': request})
            return Response({
                'detail': f'An unlock request for this session is already {existing_pending.status.lower().replace("_", " ")}',
                'existing_request': ser.data
            }, status=400)

        req = AttendanceUnlockRequest.objects.create(session=session, requested_by=staff_profile, note=note)
        self.logger.info(f"Created unlock request #{req.id} for session {session.id} by staff {staff_profile.id} (user: {user.username})")
        ser = AttendanceUnlockRequestSerializer(req, context={'request': request})
        return Response(ser.data, status=201)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        """
        Approve unlock request. Superusers can bypass two-stage approval.
        Regular admins with analytics permission should use the unified-unlock-requests endpoint
        which enforces two-stage approval (HOD first, then final approval).
        """
        user = request.user
        perms = get_user_permissions(user)
        if not ('analytics.view_all_analytics' in perms or user.is_superuser):
            return Response({'detail': 'Permission denied'}, status=403)
        
        # Use direct model lookup instead of get_object() to bypass queryset filtering
        try:
            req = AttendanceUnlockRequest.objects.get(pk=pk)
        except AttendanceUnlockRequest.DoesNotExist:
            return Response({'detail': f'Request with ID {pk} not found'}, status=404)
        
        self.logger.info(f"User {user.username} approving request #{req.id} (status: {req.status}, hod_status: {req.hod_status}, requested_by: {req.requested_by_id})")
        
        # Allow approval if status is PENDING (direct approval) or HOD_APPROVED (final approval)
        if req.status not in ['PENDING', 'HOD_APPROVED']:
            return Response({'detail': f'Request already processed (status: {req.status})'}, status=400)
        
        req.status = 'APPROVED'
        req.reviewed_by = getattr(user, 'staff_profile', None)
        import django.utils.timezone as tz
        req.reviewed_at = tz.now()
        
        # If bypassing HOD approval (superuser direct approval), also update hod_status
        if req.hod_status == 'PENDING':
            req.hod_status = 'APPROVED'
            req.hod_reviewed_by = getattr(user, 'staff_profile', None)
            req.hod_reviewed_at = tz.now()
        
        req.save()
        self.logger.info(f"Request #{req.id} approved successfully")

        try:
            sess = req.session
            sess.is_locked = False
            sess.save(update_fields=['is_locked'])
        except Exception:
            pass

        ser = AttendanceUnlockRequestSerializer(req, context={'request': request})
        return Response(ser.data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        """
        Reject unlock request. Can reject at any stage before final approval.
        """
        user = request.user
        perms = get_user_permissions(user)
        if not ('analytics.view_all_analytics' in perms or user.is_superuser):
            return Response({'detail': 'Permission denied'}, status=403)
        
        # Use direct model lookup instead of get_object() to bypass queryset filtering
        try:
            req = AttendanceUnlockRequest.objects.get(pk=pk)
        except AttendanceUnlockRequest.DoesNotExist:
            return Response({'detail': f'Request with ID {pk} not found'}, status=404)
        
        self.logger.info(f"User {user.username} rejecting request #{req.id} (status: {req.status}, hod_status: {req.hod_status}, requested_by: {req.requested_by_id})")
        
        # Can reject if not already processed as APPROVED or REJECTED
        if req.status in ['APPROVED', 'REJECTED']:
            return Response({'detail': f'Request already {req.status.lower()}'}, status=400)
        
        req.status = 'REJECTED'
        req.reviewed_by = getattr(user, 'staff_profile', None)
        import django.utils.timezone as tz
        req.reviewed_at = tz.now()
        
        # Also update hod_status if it's still pending
        if req.hod_status == 'PENDING':
            req.hod_status = 'REJECTED'
            req.hod_reviewed_by = getattr(user, 'staff_profile', None)
            req.hod_reviewed_at = tz.now()
        
        req.save()
        self.logger.info(f"Request #{req.id} rejected successfully, new status: {req.status}")
        ser = AttendanceUnlockRequestSerializer(req, context={'request': request})
        return Response(ser.data)


class UnifiedUnlockRequestsView(APIView):
    """Unified view that returns both period attendance and daily attendance unlock requests.
    Daily bulk requests (same bulk_group_id) are collapsed into a single grouped row."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        from .models import DailyAttendanceUnlockRequest
        from .serializers import DailyAttendanceUnlockRequestSerializer

        user = request.user
        perms = get_user_permissions(user)

        can_view_all = 'analytics.view_all_analytics' in perms or user.is_superuser
        staff_profile = getattr(user, 'staff_profile', None)

        # ── Period requests ───────────────────────────────────────────────────
        if can_view_all:
            period_requests = AttendanceUnlockRequest.objects.select_related(
                'session__section', 'requested_by', 'reviewed_by', 'hod_reviewed_by'
            ).filter(hod_status='HOD_APPROVED').order_by('-requested_at')
        elif staff_profile:
            period_requests = AttendanceUnlockRequest.objects.select_related(
                'session__section', 'requested_by', 'reviewed_by', 'hod_reviewed_by'
            ).filter(requested_by=staff_profile).order_by('-requested_at')
        else:
            period_requests = AttendanceUnlockRequest.objects.none()

        period_serializer = AttendanceUnlockRequestSerializer(period_requests, many=True, context={'request': request})
        period_data = list(period_serializer.data)
        for item in period_data:
            item['request_type'] = 'period'

        # ── Daily requests (grouped by bulk_group_id) ─────────────────────────
        if can_view_all:
            daily_qs = DailyAttendanceUnlockRequest.objects.select_related(
                'session__section', 'requested_by', 'reviewed_by', 'hod_reviewed_by'
            ).filter(status__in=['PENDING', 'HOD_APPROVED']).order_by('-requested_at')
        elif staff_profile:
            daily_qs = DailyAttendanceUnlockRequest.objects.select_related(
                'session__section', 'requested_by', 'reviewed_by', 'hod_reviewed_by'
            ).filter(requested_by=staff_profile).order_by('-requested_at')
        else:
            daily_qs = DailyAttendanceUnlockRequest.objects.none()

        # Collapse bulk groups into a single representative row
        daily_data = []
        seen_groups = set()
        for req in daily_qs:
            if req.bulk_group_id:
                gid = str(req.bulk_group_id)
                if gid in seen_groups:
                    continue
                seen_groups.add(gid)
                # Fetch all sibling requests in this group
                siblings = list(DailyAttendanceUnlockRequest.objects.filter(
                    bulk_group_id=req.bulk_group_id
                ).select_related('session__section', 'requested_by').order_by('session__date'))
                section = req.session.section if req.session else None
                dates = [str(s.session.date) for s in siblings if s.session]
                daily_data.append({
                    'id': req.id,
                    'bulk_group_id': gid,
                    'request_type': 'daily_bulk',
                    'session_count': len(siblings),
                    'dates': dates,
                    'date_range': f"{dates[0]} → {dates[-1]}" if dates else '',
                    'department': str(getattr(getattr(section, 'batch', None), 'course', None) and
                                      getattr(section.batch.course, 'department', None) and
                                      section.batch.course.department.name or ''),
                    'section_name': str(section) if section else '',
                    'session_display': f"{str(section)} | {len(siblings)} sessions ({dates[0]} → {dates[-1]})" if dates else str(section),
                    'requested_by': {
                        'name': req.requested_by.user.get_full_name() if req.requested_by and req.requested_by.user else '',
                        'staff_id': str(req.requested_by.staff_id) if req.requested_by else '',
                    },
                    'requested_at': req.requested_at.isoformat() if req.requested_at else '',
                    'note': req.note or '',
                    'status': req.status,
                    'hod_status': req.hod_status,
                })
            else:
                # Individual (non-bulk) daily request
                daily_data.append({
                    'id': req.id,
                    'bulk_group_id': None,
                    'request_type': 'daily',
                    'session_count': 1,
                    'dates': [str(req.session.date)] if req.session else [],
                    'department': '',
                    'section_name': str(req.session.section) if req.session else '',
                    'session_display': f"{str(req.session.section) if req.session else ''} | Daily Attendance @ {req.session.date if req.session else ''}",
                    'requested_by': {
                        'name': req.requested_by.user.get_full_name() if req.requested_by and req.requested_by.user else '',
                        'staff_id': str(req.requested_by.staff_id) if req.requested_by else '',
                    },
                    'requested_at': req.requested_at.isoformat() if req.requested_at else '',
                    'note': req.note or '',
                    'status': req.status,
                    'hod_status': req.hod_status,
                })

        combined_data = period_data + daily_data
        combined_data.sort(key=lambda x: x.get('requested_at', ''), reverse=True)

        return Response({
            'results': combined_data,
            'total_period_requests': len(period_data),
            'total_daily_requests': len(daily_data),
            'total_requests': len(combined_data),
        })

    def patch(self, request):
        """Handle approval/rejection for unlock requests. Daily bulk groups are acted on atomically."""
        user = request.user
        perms = get_user_permissions(user)

        if not ('analytics.view_all_analytics' in perms or user.is_superuser):
            return Response({'detail': 'Permission denied'}, status=403)

        request_id = request.data.get('id')
        request_type = request.data.get('request_type', 'period')  # 'period', 'daily', 'daily_bulk'
        bulk_group_id = request.data.get('bulk_group_id')
        action = request.data.get('action')  # 'approve' or 'reject'
        final_note = request.data.get('note', '')

        if not action:
            return Response({'detail': 'Missing required field: action'}, status=400)
        if action not in ['approve', 'reject']:
            return Response({'detail': 'Invalid action. Must be approve or reject'}, status=400)

        staff_profile = getattr(user, 'staff_profile', None)

        try:
            if request_type == 'period':
                if not request_id:
                    return Response({'detail': 'Missing required field: id'}, status=400)
                unlock_req = AttendanceUnlockRequest.objects.select_related('session').get(id=request_id)
                if unlock_req.status != 'HOD_APPROVED':
                    return Response({
                        'detail': f'Request must be HOD-approved first. Current status: {unlock_req.status}'
                    }, status=400)
                if action == 'approve':
                    unlock_req.status = 'APPROVED'
                    unlock_req.reviewed_by = staff_profile
                    unlock_req.reviewed_at = timezone.now()
                    unlock_req.final_note = final_note
                    unlock_req.save()
                    unlock_req.session.is_locked = False
                    unlock_req.session.save(update_fields=['is_locked'])
                    msg = 'Period attendance session unlocked successfully'
                else:
                    unlock_req.status = 'REJECTED'
                    unlock_req.reviewed_by = staff_profile
                    unlock_req.reviewed_at = timezone.now()
                    unlock_req.final_note = final_note
                    unlock_req.save()
                    msg = 'Period unlock request rejected'
                serializer = AttendanceUnlockRequestSerializer(unlock_req, context={'request': request})
                return Response({'success': True, 'message': msg, 'request': serializer.data})

            else:  # daily or daily_bulk
                from .models import DailyAttendanceUnlockRequest, DailyAttendanceRecord
                from django.db import transaction

                # Resolve which records to act on
                if bulk_group_id:
                    reqs = list(DailyAttendanceUnlockRequest.objects.select_related('session').filter(
                        bulk_group_id=bulk_group_id,
                        status__in=['PENDING', 'HOD_APPROVED'],
                    ))
                    if not reqs:
                        return Response({'detail': 'No pending requests found for this group'}, status=404)
                elif request_id:
                    req = DailyAttendanceUnlockRequest.objects.select_related('session').get(id=request_id)
                    if req.status not in ['PENDING', 'HOD_APPROVED']:
                        return Response({'detail': f'Request already finalised: {req.status}'}, status=400)
                    reqs = [req]
                else:
                    return Response({'detail': 'Provide id or bulk_group_id'}, status=400)

                with transaction.atomic():
                    for req in reqs:
                        if action == 'approve':
                            req.status = 'APPROVED'
                            req.reviewed_by = staff_profile
                            req.reviewed_at = timezone.now()
                            req.final_note = final_note
                            req.save()
                            req.session.is_locked = False
                            req.session.save(update_fields=['is_locked'])
                            DailyAttendanceRecord.objects.filter(session=req.session).delete()
                        else:
                            req.status = 'REJECTED'
                            req.reviewed_by = staff_profile
                            req.reviewed_at = timezone.now()
                            req.final_note = final_note
                            req.save()

                count = len(reqs)
                if action == 'approve':
                    msg = f'Daily attendance unlocked and reset for {count} session{"s" if count > 1 else ""}'
                else:
                    msg = f'Daily unlock request rejected for {count} session{"s" if count > 1 else ""}'
                return Response({'success': True, 'message': msg})

        except AttendanceUnlockRequest.DoesNotExist:
            return Response({'detail': 'Request not found'}, status=404)
        except Exception as e:
            return Response({'detail': str(e)}, status=500)

    def delete(self, request):
        """Bulk-delete all unlock requests visible to IQAC/admin."""
        user = request.user
        perms = get_user_permissions(user)

        if not ('analytics.view_all_analytics' in perms or user.is_superuser):
            return Response({'detail': 'Permission denied'}, status=403)

        from .models import DailyAttendanceUnlockRequest

        deleted_period, _ = AttendanceUnlockRequest.objects.all().delete()
        deleted_daily, _ = DailyAttendanceUnlockRequest.objects.all().delete()

        return Response({
            'deleted_period_requests': deleted_period,
            'deleted_daily_requests': deleted_daily,
            'total_deleted': deleted_period + deleted_daily,
        })


class StaffPeriodsView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            return Response({'results': []})

        date_param = request.query_params.get('date')
        import datetime
        try:
            if date_param:
                date = datetime.date.fromisoformat(date_param)
            else:
                date = datetime.date.today()
        except Exception:
            date = datetime.date.today()

        day = date.isoweekday()
        from timetable.models import TimetableAssignment
        # fetch assignments for the day; include assignments that either have staff set
        # to the current user or which are intended to be taught by the staff via
        # an active TeachingAssignment mapping (fallback when TimetableAssignment.staff is null)
        qs = TimetableAssignment.objects.filter(day=day).select_related('period', 'section', 'curriculum_row', 'subject_batch', 'staff')
        # prefetch any existing attendance session for this date so frontend can indicate status
        from .models import PeriodAttendanceSession
        results = []
        for a in qs:
            include = False

            # Reset per-assignment resolved fields (avoid leaking values across loop iterations)
            resolved_subject_display = None
            resolved_elective_id = None
            resolved_elective_name = None
            teach_assign = None
            try:
                # Check for batch-specific staff assignment first
                batch = getattr(a, 'subject_batch', None)
                if batch and getattr(batch, 'staff', None):
                    # If batch has assigned staff, ONLY show to batch staff
                    if getattr(batch.staff, 'id', None) == getattr(staff_profile, 'id', None):
                        include = True
                    else:
                        # This period belongs to a different batch staff, skip it
                        continue
                elif batch and not getattr(batch, 'staff', None):
                    # Batch exists but no staff assigned - only show to batch creator
                    if getattr(batch, 'created_by', None) and getattr(batch.created_by, 'id', None) == getattr(staff_profile, 'id', None):
                        include = True
                    else:
                        # This batch belongs to a different creator, skip it
                        continue
                elif getattr(a, 'staff', None) and getattr(a.staff, 'id', None) == getattr(staff_profile, 'id', None):
                    # No batch, check default timetable staff
                    include = True

                # Resolve a TeachingAssignment for this staff+slot. This is required to correctly
                # identify elective sub-options (and therefore the correct attendance session).
                from .models import TeachingAssignment as _TA
                ta_qs = _TA.objects.filter(is_active=True, staff=staff_profile)

                # Section-scoped or department-wide (section is null) assignments are allowed
                ta_qs = ta_qs.filter(Q(section=a.section) | Q(section__isnull=True))

                # Prefer matching by explicit curriculum_row when present, then subject_batch, then subject_text
                base_cr = getattr(a, 'curriculum_row', None)
                if base_cr is None and getattr(a, 'subject_batch', None) is not None:
                    base_cr = getattr(a.subject_batch, 'curriculum_row', None)

                matched = False
                if base_cr is not None:
                    q_cr = ta_qs.filter(
                        Q(curriculum_row=base_cr) |
                        Q(elective_subject__parent=base_cr) |
                        Q(elective_subject__course_code__iexact=(getattr(base_cr, 'course_code', None) or ''))
                    )
                    if q_cr.exists():
                        ta_qs = q_cr
                        matched = True

                if (not matched) and getattr(a, 'subject_text', None):
                    txt = (a.subject_text or '').strip()
                    ltxt = txt.lower()
                    import re
                    norm = re.sub(r'[^a-z0-9]', '', ltxt)
                    q_text = ta_qs.filter(
                        Q(elective_subject__course_code__iexact=ltxt) |
                        Q(elective_subject__course_code__iexact=norm) |
                        Q(elective_subject__course_name__icontains=txt) |
                        Q(curriculum_row__course_code__iexact=ltxt) |
                        Q(subject__code__iexact=ltxt) |
                        Q(subject__name__icontains=txt)
                    )
                    if q_text.exists():
                        ta_qs = q_text
                        matched = True

                if matched:
                    teach_assign = ta_qs.order_by(
                        models.Case(
                            models.When(section=a.section, then=models.Value(0)),
                            default=models.Value(1),
                            output_field=models.IntegerField(),
                        ),
                        'id',
                    ).first()

                    # Only include fallback TA-resolved entries when timetable.staff is null.
                    if not include and not getattr(a, 'staff', None) and teach_assign is not None:
                        include = True
            except Exception:
                include = False

            # If there already exists a session for this section/period/date that
            # is tied to this staff (either via teaching_assignment.staff or created_by),
            # include this assignment so the staff can view/take that session.
            try:
                from .models import PeriodAttendanceSession as _PAS
                sess_q = _PAS.objects.filter(section=a.section, period=a.period, date=date)
                if staff_profile is not None:
                    if sess_q.filter(teaching_assignment__staff=staff_profile).exists() or sess_q.filter(created_by=staff_profile).exists():
                        include = True
            except Exception:
                pass

            if not include:
                continue

            # If there's a special timetable entry for this section/period/date,
            # prefer the special entry and skip the normal timetable assignment so
            # staff sees only the special period for that date.
            try:
                from timetable.models import SpecialTimetableEntry
                if SpecialTimetableEntry.objects.filter(timetable__section=a.section, period=a.period, date=date, is_active=True).exists():
                    # skip adding the normal assignment — the special entry will be
                    # included separately below (or by separate logic)
                    continue
            except Exception:
                pass

            # Resolve elective display/name from the TeachingAssignment if present
            try:
                if teach_assign is not None and getattr(teach_assign, 'elective_subject', None) is not None:
                    es = teach_assign.elective_subject
                    resolved_subject_display = (getattr(es, 'course_code', None) or getattr(es, 'course_name', None))
                    resolved_elective_id = getattr(es, 'id', None)
                    resolved_elective_name = getattr(es, 'course_name', None) or resolved_subject_display
            except Exception:
                resolved_subject_display = None
                resolved_elective_id = None
                resolved_elective_name = None

            # find existing session for this section/period/date (+ teaching_assignment when available)
            sess_qs = PeriodAttendanceSession.objects.select_related('assigned_to', 'assigned_to__user').filter(section=a.section, period=a.period, date=date)
            if teach_assign is not None:
                sess_qs = sess_qs.filter(teaching_assignment=teach_assign)
            else:
                # teach_assign is None: this is a custom-subject/substitute period where the
                # timetable uses subject_text and no matching TeachingAssignment was found.
                # Look for sessions this staff created OR sessions already linked to any of
                # this staff's teaching assignments (catches historical mis-saves from the old
                # fallback-to-first-TA bug) — but always scoped to created_by=staff_profile
                # to avoid crossing into another staff member's session for the same period.
                if staff_profile is not None:
                    sess_qs = sess_qs.filter(created_by=staff_profile)
                else:
                    # No staff context: can't determine ownership; skip showing any session
                    sess_qs = sess_qs.none()
            session = sess_qs.order_by('-id').first()
            if session is None:
                # If a swap request was approved, assignment may exist on a session
                # whose teaching_assignment does not match this resolved slot.
                # Surface that assigned session so original staff is shown as locked.
                try:
                    session = PeriodAttendanceSession.objects.select_related('assigned_to', 'assigned_to__user').filter(
                        section=a.section,
                        period=a.period,
                        date=date,
                        assigned_to__isnull=False,
                    ).exclude(assigned_to=staff_profile).order_by('-id').first()
                except Exception:
                    session = None
            # determine latest unlock request status for this session (if any)
            unlock_status = None
            unlock_id = None
            try:
                if session:
                    req = AttendanceUnlockRequest.objects.filter(session=session).order_by('-requested_at').first()
                    if req:
                        unlock_status = getattr(req, 'status', None)
                        unlock_id = getattr(req, 'id', None)
            except Exception:
                unlock_status = None
                unlock_id = None


            # compute section strength and any existing attendance counts if session exists
            try:
                from .models import PeriodAttendanceRecord
                from .models import PeriodAttendanceSession as _PAS
                from .models import StudentProfile as _SP
            except Exception:
                PeriodAttendanceRecord = None
                _PAS = None
                _SP = None

            total_strength = None
            present_count = None
            absent_count = None
            leave_count = None
            od_count = None
            late_count = None

            try:
                # always compute section strength for display
                from .models import StudentProfile as SP
                total_strength = SP.objects.filter(section_id=a.section_id).count()
            except Exception:
                total_strength = None

            if session:
                try:
                    records_q = PeriodAttendanceRecord.objects.filter(session=session)
                    present_count = records_q.filter(status__in=['P', 'OD', 'LATE']).count()
                    absent_count = records_q.filter(status='A').count()
                    leave_count = records_q.filter(status='LEAVE').count()
                    od_count = records_q.filter(status='OD').count()
                    late_count = records_q.filter(status='LATE').count()
                except Exception:
                    present_count = absent_count = leave_count = od_count = late_count = None

            # Get batch label if subject_batch_id exists
            batch_label = None
            if getattr(a, 'subject_batch_id', None):
                try:
                    batch_label = getattr(a.subject_batch, 'name', None)
                except Exception:
                    batch_label = None

            results.append({
                'id': a.id,
                'section_id': a.section_id,
                'section_name': str(a.section),
                'period': {'id': a.period.id, 'index': a.period.index, 'label': a.period.label, 'start_time': getattr(a.period, 'start_time', None), 'end_time': getattr(a.period, 'end_time', None)},
                # provide a reliable subject display: prefer curriculum_row code/name, then subject_text
                'subject_id': getattr(getattr(a, 'curriculum_row', None), 'id', None),
                'subject_display': resolved_subject_display or (getattr(getattr(a, 'curriculum_row', None), 'course_code', None) or getattr(getattr(a, 'curriculum_row', None), 'course_name', None) or getattr(a, 'subject_text', None) or None),
                'teaching_assignment_id': getattr(teach_assign, 'id', None),
                'elective_subject_id': resolved_elective_id,
                'elective_subject_name': resolved_elective_name,
                'subject_batch_id': getattr(a, 'subject_batch_id', None),
                'subject_batch_label': batch_label,
                'attendance_session_id': getattr(session, 'id', None),
                'attendance_session_locked': getattr(session, 'is_locked', False) if session else False,
                # assigned_to: staff member assigned to take this period's attendance (via swap)
                'assigned_to': {
                    'id': session.assigned_to.id,
                    'name': session.assigned_to.user.get_full_name() if session.assigned_to.user else '',
                    'staff_id': session.assigned_to.staff_id,
                } if session and session.assigned_to else None,
                # include latest unlock request status (if any) so frontend can show pending/approved/rejected
                'unlock_request_status': unlock_status,
                'unlock_request_id': unlock_id,
                'total_strength': total_strength,
                'present': present_count,
                'absent': absent_count,
                'leave': leave_count,
                'on_duty': od_count,
            })
        # Also include any SpecialTimetableEntry items for this date where the current
        # staff is the assigned staff or is mapped via a TeachingAssignment for the
        # curriculum_row/elective. We present them alongside regular assignments so
        # the staff can open/take attendance for those special periods.
        try:
            from timetable.models import SpecialTimetableEntry
            from .models import PeriodAttendanceSession as _PAS
            special_qs = SpecialTimetableEntry.objects.filter(date=date, is_active=True).select_related('timetable__section', 'period', 'curriculum_row', 'subject_batch', 'staff')
            for se in special_qs:
                include = False
                try:
                    # Check for batch-specific staff assignment first
                    batch = getattr(se, 'subject_batch', None)
                    if batch and getattr(batch, 'staff', None):
                        # If batch has assigned staff, ONLY show to batch staff
                        if getattr(batch.staff, 'id', None) == getattr(staff_profile, 'id', None):
                            include = True
                        else:
                            # This period belongs to a different batch staff, skip it
                            continue
                    elif batch and not getattr(batch, 'staff', None):
                        # Batch exists but no staff assigned - only show to batch creator
                        if getattr(batch, 'created_by', None) and getattr(batch.created_by, 'id', None) == getattr(staff_profile, 'id', None):
                            include = True
                        else:
                            # This batch belongs to a different creator, skip it
                            continue
                    elif getattr(se, 'staff', None) and getattr(se.staff, 'id', None) == getattr(staff_profile, 'id', None):
                        # No batch, check default special entry staff
                        include = True
                    else:
                        # fallback: if special entry has a curriculum_row, check TeachingAssignment mappings
                        if getattr(se, 'curriculum_row', None):
                            from .models import TeachingAssignment as _TA
                            ta_qs = _TA.objects.filter(is_active=True, staff=staff_profile).filter(
                                (Q(curriculum_row=se.curriculum_row) | Q(elective_subject__parent=se.curriculum_row))
                            ).filter(Q(section=se.timetable.section) | Q(section__isnull=True))
                            if ta_qs.exists():
                                include = True
                except Exception:
                    include = False

                if not include:
                    continue

                # resolve teaching assignment for this staff+special slot
                teach_assign = None
                elective_name = None
                try:
                    from .models import TeachingAssignment as _TA
                    ta_qs = _TA.objects.filter(is_active=True, staff=staff_profile).filter(Q(section=se.timetable.section) | Q(section__isnull=True))
                    base_cr = getattr(se, 'curriculum_row', None)
                    if base_cr is None and getattr(se, 'subject_batch', None) is not None:
                        base_cr = getattr(se.subject_batch, 'curriculum_row', None)
                    if base_cr is not None:
                        ta_qs = ta_qs.filter(Q(curriculum_row=base_cr) | Q(elective_subject__parent=base_cr))
                    teach_assign = ta_qs.order_by(
                        models.Case(
                            models.When(section=se.timetable.section, then=models.Value(0)),
                            default=models.Value(1),
                            output_field=models.IntegerField(),
                        ),
                        'id',
                    ).first()
                except Exception:
                    teach_assign = None

                # find existing session for this special entry's section/period/date (+ teaching_assignment when available)
                sess_qs = _PAS.objects.filter(section=se.timetable.section, period=se.period, date=date)
                if teach_assign is not None:
                    sess_qs = sess_qs.filter(teaching_assignment=teach_assign)
                else:
                    sess_qs = sess_qs.filter(teaching_assignment__isnull=True)
                    if staff_profile is not None:
                        sess_qs = sess_qs.filter(created_by=staff_profile)
                    else:
                        sess_qs = sess_qs.none()
                sess = sess_qs.order_by('-id').first()
                subj_disp = None
                subj_id = None
                elective_id = None
                if getattr(se, 'curriculum_row', None):
                    subj_id = se.curriculum_row.id
                    subj_disp = getattr(se.curriculum_row, 'course_code', None) or getattr(se.curriculum_row, 'course_name', None)
                    try:
                        # if this staff is mapped to a sub-elective for this curriculum_row,
                        # prefer that sub-elective's display
                        if teach_assign and getattr(teach_assign, 'elective_subject', None):
                            es = teach_assign.elective_subject
                            subj_disp = (getattr(es, 'course_code', None) or getattr(es, 'course_name', None))
                            elective_id = getattr(es, 'id', None)
                            elective_name = getattr(es, 'course_name', None) or subj_disp
                    except Exception:
                        pass
                else:
                    subj_disp = se.subject_text or None

                # determine latest unlock request for special session (if any)
                special_unlock_status = None
                special_unlock_id = None
                try:
                    if sess:
                        sreq = AttendanceUnlockRequest.objects.filter(session=sess).order_by('-requested_at').first()
                        if sreq:
                            special_unlock_status = getattr(sreq, 'status', None)
                            special_unlock_id = getattr(sreq, 'id', None)
                except Exception:
                    special_unlock_status = None
                    special_unlock_id = None

                # Get batch label for special entry if subject_batch_id exists
                special_batch_label = None
                if getattr(se, 'subject_batch_id', None):
                    try:
                        special_batch_label = getattr(se.subject_batch, 'name', None)
                    except Exception:
                        special_batch_label = None

                results.append({
                    'id': -(se.id),
                    'section_id': se.timetable.section.id,
                    'section_name': str(se.timetable.section),
                    'period': {'id': se.period.id, 'index': se.period.index, 'label': se.period.label, 'start_time': getattr(se.period, 'start_time', None), 'end_time': getattr(se.period, 'end_time', None)},
                    'subject_id': subj_id,
                    'subject_display': subj_disp,
                    'teaching_assignment_id': getattr(teach_assign, 'id', None),
                    'elective_subject_id': elective_id,
                    'elective_subject_name': elective_name,
                    'subject_batch_id': getattr(se, 'subject_batch_id', None),
                    'subject_batch_label': special_batch_label,
                    'attendance_session_id': getattr(sess, 'id', None),
                    'attendance_session_locked': getattr(sess, 'is_locked', False) if sess else False,
                    'unlock_request_status': special_unlock_status,
                    'unlock_request_id': special_unlock_id,
                    'is_special': True,
                    'is_swap': (getattr(getattr(se, 'timetable', None), 'name', '') or '').startswith('[SWAP]'),
                })
        except Exception:
            # non-fatal: if special entries cannot be included, return the standard results
            pass

        # Include sessions that have been ASSIGNED (swapped) to the current staff by another staff member.
        # These won't appear via timetable assignments, so we query PeriodAttendanceSession directly.
        try:
            if staff_profile is not None:
                from .models import PeriodAttendanceSession as _SPAS
                # Collect session IDs already in results to avoid duplicates
                existing_session_ids = {r['attendance_session_id'] for r in results if r.get('attendance_session_id')}
                assigned_sessions = (
                    _SPAS.objects.filter(assigned_to=staff_profile, date=date)
                    .exclude(id__in=existing_session_ids)
                    .select_related(
                        'section', 'period',
                        'teaching_assignment',
                        'teaching_assignment__curriculum_row',
                        'teaching_assignment__elective_subject',
                        'timetable_assignment',
                        'created_by', 'created_by__user',
                        'assigned_to', 'assigned_to__user',
                    )
                )
                for asess in assigned_sessions:
                    ta = getattr(asess, 'teaching_assignment', None)
                    # subject display
                    subj_disp = None
                    subj_id = None
                    elective_id = None
                    elective_name = None
                    batch_label = None
                    batch_id = None
                    try:
                        if ta is not None:
                            cr = getattr(ta, 'curriculum_row', None)
                            if cr:
                                subj_id = cr.id
                                subj_disp = getattr(cr, 'course_code', None) or getattr(cr, 'course_name', None)
                            es = getattr(ta, 'elective_subject', None)
                            if es:
                                subj_disp = getattr(es, 'course_code', None) or getattr(es, 'course_name', None)
                                elective_id = getattr(es, 'id', None)
                                elective_name = getattr(es, 'course_name', None) or subj_disp
                        # batch info lives on TimetableAssignment, not TeachingAssignment
                        tma = getattr(asess, 'timetable_assignment', None)
                        if tma is not None:
                            sb = getattr(tma, 'subject_batch', None)
                            if sb:
                                batch_id = getattr(sb, 'id', None)
                                batch_label = getattr(sb, 'name', None)
                            # fallback subject from timetable assignment if teaching_assignment had none
                            if subj_disp is None:
                                subj_disp = getattr(tma, 'subject_text', None)
                    except Exception:
                        pass
                    # unlock request
                    unlock_status = None
                    unlock_id = None
                    try:
                        req = AttendanceUnlockRequest.objects.filter(session=asess).order_by('-requested_at').first()
                        if req:
                            unlock_status = getattr(req, 'status', None)
                            unlock_id = getattr(req, 'id', None)
                    except Exception:
                        pass
                    # assigned_to info (current staff is the assignee; show who originally created)
                    assignee = asess.assigned_to
                    creator = asess.created_by
                    results.append({
                        'id': -(asess.id + 10000000),  # synthetic negative id to avoid clash with TA ids
                        'section_id': asess.section_id,
                        'section_name': str(asess.section),
                        'period': {
                            'id': asess.period.id,
                            'index': asess.period.index,
                            'label': asess.period.label,
                            'start_time': getattr(asess.period, 'start_time', None),
                            'end_time': getattr(asess.period, 'end_time', None),
                        },
                        'subject_id': subj_id,
                        'subject_display': subj_disp,
                        'teaching_assignment_id': getattr(ta, 'id', None),
                        'elective_subject_id': elective_id,
                        'elective_subject_name': elective_name,
                        'subject_batch_id': batch_id,
                        'subject_batch_label': batch_label,
                        'attendance_session_id': asess.id,
                        'attendance_session_locked': asess.is_locked,
                        'assigned_to': {
                            'id': assignee.id,
                            'name': assignee.user.get_full_name() if assignee and assignee.user else '',
                            'staff_id': getattr(assignee, 'staff_id', ''),
                        } if assignee else None,
                        'original_staff': {
                            'id': creator.id,
                            'name': creator.user.get_full_name() if creator and creator.user else '',
                            'staff_id': getattr(creator, 'staff_id', ''),
                        } if creator else None,
                        'unlock_request_status': unlock_status,
                        'unlock_request_id': unlock_id,
                        'is_swap': True,
                    })
        except Exception:
            # non-fatal: assigned sessions block; standard results are still valid
            pass

        return Response({'results': results})


class AdvisorMyStudentsView(APIView):
    """Return students for sections where the current user is an active advisor.
    Used by the Mentor Assignment page to load the advisor's students.

    Response format:
    { results: [ { section_id, section_name, batch, students: [...] } ] }
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        import traceback, logging
        try:
            user = request.user
            staff_profile = getattr(user, 'staff_profile', None)
            if not staff_profile:
                return Response({'results': []})

            advisor_qs = SectionAdvisor.objects.filter(
                advisor=staff_profile, is_active=True, academic_year__is_active=True
            ).select_related('section', 'section__batch', 'section__batch__course', 'section__batch__regulation')
            sections = [a.section for a in advisor_qs]
            if not sections:
                return Response({'results': []})

            from .models import StudentSectionAssignment, StudentProfile

            section_ids = [s.id for s in sections]
            assign_qs = StudentSectionAssignment.objects.filter(
                section_id__in=section_ids, end_date__isnull=True
            ).exclude(student__status__in=['INACTIVE', 'DEBAR']).select_related('student__user', 'section')
            students_by_section: dict = {}
            for a in assign_qs:
                students_by_section.setdefault(a.section_id, []).append(a.student)

            legacy_qs = StudentProfile.objects.filter(section_id__in=section_ids).exclude(status__in=['INACTIVE', 'DEBAR']).select_related('user', 'section')
            for s in legacy_qs:
                present = students_by_section.setdefault(s.section_id, [])
                if not any(x.pk == s.pk for x in present):
                    present.append(s)

            try:
                student_ids = [st.pk for v in students_by_section.values() for st in v]
                mentor_map: dict = {}
                if student_ids:
                    for mm in StudentMentorMap.objects.filter(student_id__in=student_ids, is_active=True).select_related('mentor'):
                        try:
                            mentor_map[mm.student_id] = {
                                'mentor_id': getattr(mm.mentor, 'id', None),
                                'mentor_name': getattr(getattr(mm.mentor, 'user', None), 'username', None),
                            }
                        except Exception:
                            pass
            except Exception:
                mentor_map = {}

            results = []
            for sec in sections:
                studs = students_by_section.get(sec.id, [])
                ser = StudentSimpleSerializer([
                    {
                        'id': st.pk,
                        'reg_no': st.reg_no,
                        'user': getattr(st, 'user', None),
                        'section_id': getattr(st, 'section_id', None),
                        'section_name': str(getattr(st, 'section', '')),
                        'has_mentor': (st.pk in mentor_map),
                        'mentor_id': mentor_map.get(st.pk, {}).get('mentor_id'),
                        'mentor_name': mentor_map.get(st.pk, {}).get('mentor_name'),
                    }
                    for st in studs
                ], many=True)
                batch = getattr(sec, 'batch', None)
                course = getattr(batch, 'course', None) if batch is not None else None
                dept = getattr(course, 'department', None) if course is not None else None
                reg = getattr(batch, 'regulation', None) if batch else None
                sem_obj = getattr(sec, 'semester', None)
                sem_val = getattr(sem_obj, 'number', None) if sem_obj else None
                results.append({
                    'section_id': sec.id,
                    'section_name': sec.name,
                    'batch': getattr(batch, 'name', None),
                    'batch_regulation': {'id': getattr(reg, 'id', None), 'code': getattr(reg, 'code', None)} if reg else None,
                    'department_id': getattr(dept, 'id', None),
                    'department': {'id': getattr(dept, 'id', None), 'code': getattr(dept, 'code', None)} if dept else None,
                    'department_short_name': (getattr(dept, 'short_name', None) or getattr(dept, 'code', None)) if dept else None,
                    'semester': sem_val,
                    'students': ser.data,
                })

            return Response({'results': results})
        except Exception as e:
            logging.getLogger(__name__).exception('AdvisorMyStudentsView error: %s', e)
            return Response({'detail': 'Internal server error', 'error': str(e)}, status=500)


class MentorMyMenteesView(APIView):
    """Return students for which the current user is the assigned mentor.

    Response format:
    { results: [ { section_id, section_name, batch, department_short_name, students: [...] } ] }
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        import logging
        try:
            from accounts.utils import get_user_permissions
            user = request.user
            perms = get_user_permissions(user)
            if 'academics.view_mentees' not in perms:
                return Response({'error': 'Permission denied'}, status=403)

            staff_profile = getattr(user, 'staff_profile', None)
            if not staff_profile:
                return Response({'results': []})

            from .models import StudentMentorMap
            mentee_maps = StudentMentorMap.objects.filter(
                mentor=staff_profile, is_active=True
            ).exclude(student__status__in=['INACTIVE', 'DEBAR']).select_related(
                'student__user',
                'student__section__batch__course__department',
                'student__section__batch__regulation',
            )

            students_by_section: dict = {}
            for mm in mentee_maps:
                st = mm.student
                sec = getattr(st, 'section', None)
                if sec:
                    students_by_section.setdefault(sec.id, {'section': sec, 'students': []})
                    students_by_section[sec.id]['students'].append(st)

            results = []
            for sec_id, data in students_by_section.items():
                sec = data['section']
                studs = data['students']
                batch = getattr(sec, 'batch', None)
                course = getattr(batch, 'course', None) if batch else None
                dept = getattr(course, 'department', None) if course else None
                ser = StudentSimpleSerializer([
                    {
                        'id': st.pk,
                        'reg_no': st.reg_no,
                        'user': getattr(st, 'user', None),
                        'section_id': sec.id,
                        'section_name': sec.name,
                        'has_mentor': True,
                        'mentor_id': staff_profile.id,
                        'mentor_name': user.username,
                    }
                    for st in studs
                ], many=True)
                results.append({
                    'section_id': sec.id,
                    'section_name': sec.name,
                    'batch': getattr(batch, 'name', None),
                    'department_id': getattr(dept, 'id', None),
                    'department': {'id': getattr(dept, 'id', None), 'code': getattr(dept, 'code', None)} if dept else None,
                    'department_short_name': (getattr(dept, 'short_name', None) or getattr(dept, 'code', None)) if dept else None,
                    'students': ser.data,
                })
            return Response({'results': results})
        except Exception as e:
            logging.getLogger(__name__).exception('MentorMyMenteesView error: %s', e)
            return Response({'detail': 'Internal server error', 'error': str(e)}, status=500)


# DayAttendance endpoints removed as part of attendance feature removal.


class StudentAttendanceView(APIView):
    """Return period-wise attendance records for the current student.

    Query params:
    - start_date (ISO) optional
    - end_date (ISO) optional
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        try:
            from .models import StudentProfile, PeriodAttendanceRecord
            sp = StudentProfile.objects.filter(user=user).first()
            if not sp:
                return Response({'results': []})

            import datetime
            start_param = request.query_params.get('start_date')
            end_param = request.query_params.get('end_date')
            qs = PeriodAttendanceRecord.objects.filter(student=sp).select_related(
                'session__period',
                'session__section',
                'session__timetable_assignment',
                'session__timetable_assignment__curriculum_row',
                'session__timetable_assignment__curriculum_row__master',
                'marked_by',
                'session',
            )
            try:
                if start_param:
                    sd = datetime.date.fromisoformat(start_param)
                    qs = qs.filter(session__date__gte=sd)
                if end_param:
                    ed = datetime.date.fromisoformat(end_param)
                    qs = qs.filter(session__date__lte=ed)
            except Exception:
                pass

            qs = qs.order_by('-session__date', 'session__period__index')

            # overall calculation: only consider marked periods (records) as denominator
            # total_marked_periods = number of PeriodAttendanceRecord entries in the selected range
            total_marked_periods = qs.count()

            # Build output and compute counts
            out = []
            present_count = 0
            status_counts = {}
            # Subject-wise maps: key -> {counts: {status: count}, total: int, display: str}
            subj_map = {}

            # statuses considered present for percentage calculation
            present_statuses = {'P', 'OD', 'LATE'}

            for r in qs:
                sess = getattr(r, 'session', None)
                period = getattr(sess, 'period', None) if sess else None
                section = getattr(sess, 'section', None) if sess else None
                ta = getattr(sess, 'timetable_assignment', None) if sess else None
                # determine subject identifier
                subj_key = None
                subj_disp = None
                subj_code = None
                try:
                    if ta is not None:
                        if getattr(ta, 'curriculum_row', None):
                            cr = ta.curriculum_row
                            master = getattr(cr, 'master', None)
                            subj_key = f"CR:{ta.curriculum_row_id}"
                            subj_code = (
                                getattr(cr, 'course_code', None)
                                or getattr(master, 'course_code', None)
                                or getattr(cr, 'mnemonic', None)
                            )
                            subj_disp = (
                                getattr(cr, 'course_name', None)
                                or getattr(master, 'course_name', None)
                                or getattr(ta, 'subject_text', None)
                                or 'Unassigned'
                            )
                        else:
                            subj_key = f"TXT:{(ta.subject_text or 'Unassigned') }"
                            subj_disp = ta.subject_text or 'Unassigned'
                    else:
                        subj_key = 'Unassigned'
                        subj_disp = 'Unassigned'
                except Exception:
                    subj_key = 'Unassigned'
                    subj_disp = 'Unassigned'
                    subj_code = None

                # update status counters
                status_counts[r.status] = status_counts.get(r.status, 0) + 1
                if r.status in present_statuses:
                    present_count += 1

                # subject-wise totals and status counts
                if subj_key not in subj_map:
                    subj_map[subj_key] = {'counts': {}, 'total': 0, 'display': subj_disp, 'code': subj_code}
                subj_map[subj_key]['total'] += 1
                subj_map[subj_key]['counts'][r.status] = subj_map[subj_key]['counts'].get(r.status, 0) + 1

                out.append({
                    'id': r.id,
                    'date': getattr(sess, 'date', None),
                    'period': {'id': getattr(period, 'id', None), 'index': getattr(period, 'index', None), 'label': getattr(period, 'label', None), 'start_time': getattr(period, 'start_time', None), 'end_time': getattr(period, 'end_time', None)},
                    'section': {'id': getattr(section, 'id', None), 'name': str(section) if section else None},
                    'status': r.status,
                    'marked_at': r.marked_at,
                    'marked_by': getattr(getattr(r, 'marked_by', None), 'staff_id', None),
                    'subject_key': subj_key,
                    'subject_code': subj_code,
                    'subject_display': subj_disp,
                })

            overall_percentage = (present_count / total_marked_periods * 100) if total_marked_periods > 0 else None

            by_subject = []
            for k, v in subj_map.items():
                subject_present = (
                    v['counts'].get('P', 0)
                    + v['counts'].get('OD', 0)
                    + v['counts'].get('LATE', 0)
                )
                perc = (subject_present / v['total'] * 100) if v['total'] > 0 else None
                by_subject.append({
                    'subject_key': k,
                    'subject_code': v.get('code'),
                    'subject_display': v.get('display'),
                    'counts': v.get('counts', {}),
                    'total': v.get('total', 0),
                    'percentage': perc,
                })

            summary = {'overall': {'present': present_count, 'total_marked_periods': total_marked_periods, 'percentage': overall_percentage, 'status_counts': status_counts}, 'by_subject': by_subject}

            return Response({'results': out, 'summary': summary})
        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).exception('StudentAttendanceView error: %s', e)
            tb = traceback.format_exc()
            return Response({'detail': 'Internal server error', 'error': str(e), 'trace': tb}, status=500)


class StudentMarksView(APIView):
    """Return the logged-in student's marks aggregated by subject and cycle.

    URL: /api/academics/student/marks/
    The view derives the student from request.user and will not accept a student_id param.
    """

    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        sp = getattr(user, 'student_profile', None)
        if sp is None:
            return Response({'detail': 'Student profile not found for user.'}, status=status.HTTP_403_FORBIDDEN)

        from django.db.models import Q
        from django.db import OperationalError, ProgrammingError
        from collections import defaultdict

        section = sp.get_current_section() or getattr(sp, 'section', None)
        semester = getattr(section, 'semester', None) if section is not None else None
        course = getattr(getattr(section, 'batch', None), 'course', None) if section is not None else None

        from .models import Subject
        from .models import AcademicYear, TeachingAssignment

        if semester is None:
            return Response({'student': {'id': sp.id, 'reg_no': sp.reg_no}, 'semester': None, 'courses': []})

        # Subjects shown to the student should cover all courses they are enrolled on.
        # In practice, curriculum/elective mapping can be incomplete; the most reliable
        # baseline is Subject(course=student_course, semester=semester).
        # We then UNION any curriculum/elective codes as a best-effort supplement.
        subjects = Subject.objects.none()

        # Best-effort curriculum metadata for class_type/internal max
        try:
            from curriculum.models import CurriculumDepartment
        except Exception:
            CurriculumDepartment = None

        try:
            from curriculum.models import ElectiveChoice
        except Exception:
            ElectiveChoice = None

        regulation_code = None
        try:
            regulation_code = getattr(getattr(getattr(section, 'batch', None), 'regulation', None), 'code', None)
        except Exception:
            regulation_code = None
        if not regulation_code:
            try:
                regulation_code = str(getattr(getattr(section, 'batch', None), 'regulation', '') or '').strip() or None
            except Exception:
                regulation_code = None

        dept = getattr(course, 'department', None) if course is not None else None

        # Build enrolled subject code list from curriculum + elective choices when possible
        allowed_codes = set()
        try:
            if CurriculumDepartment is not None and dept is not None and regulation_code and semester is not None:
                core_codes = list(
                    CurriculumDepartment.objects.filter(
                        department=dept,
                        regulation=regulation_code,
                        semester=semester,
                        is_elective=False,
                    )
                    .exclude(course_code__isnull=True)
                    .exclude(course_code='')
                    .values_list('course_code', flat=True)
                )
                allowed_codes.update([str(c).strip() for c in core_codes if c])

                if ElectiveChoice is not None:
                    # Prefer active academic year electives, but tolerate null academic_year rows.
                    try:
                        ay = AcademicYear.objects.filter(is_active=True).first()
                    except Exception:
                        ay = None

                    eqs = ElectiveChoice.objects.filter(student=sp, is_active=True).select_related('elective_subject')
                    if ay is not None:
                        # include choices bound to active AY OR legacy choices with AY unset
                        eqs = eqs.filter(models.Q(academic_year=ay) | models.Q(academic_year__isnull=True))

                    elective_codes = []
                    for ch in eqs:
                        es = getattr(ch, 'elective_subject', None)
                        if not es:
                            continue
                        if dept is not None and getattr(es, 'department_id', None) != getattr(dept, 'id', None):
                            continue
                        if regulation_code and str(getattr(es, 'regulation', '') or '').strip() != str(regulation_code):
                            continue
                        if getattr(es, 'semester_id', None) != getattr(semester, 'id', None):
                            continue
                        code = str(getattr(es, 'course_code', '') or '').strip()
                        if code:
                            elective_codes.append(code)
                    allowed_codes.update(elective_codes)
        except Exception:
            allowed_codes = set()

        base_subjects = Subject.objects.none()
        if course is not None:
            base_subjects = Subject.objects.filter(semester=semester, course=course)

        code_subjects = Subject.objects.none()
        if allowed_codes:
            code_subjects = Subject.objects.filter(semester=semester, code__in=sorted(list(allowed_codes)))

        # Prefer course+semester baseline; supplement with curriculum/elective code matches.
        subjects = (base_subjects | code_subjects).distinct().order_by('code')

        # If we still have nothing (e.g., course not set on Subject rows), fall back to code-only.
        if not subjects.exists() and allowed_codes:
            subjects = Subject.objects.filter(semester=semester, code__in=sorted(list(allowed_codes))).distinct().order_by('code')

        curriculum_by_code = {}
        try:
            if CurriculumDepartment is not None and dept is not None and regulation_code and semester is not None:
                only_fields = ['course_code', 'course_name', 'class_type', 'internal_mark']
                try:
                    if hasattr(getattr(CurriculumDepartment, '_meta', None), 'get_field'):
                        CurriculumDepartment._meta.get_field('enabled_assessments')
                        only_fields.append('enabled_assessments')
                except Exception:
                    pass

                rows = (
                    CurriculumDepartment.objects.filter(
                        department=dept,
                        regulation=regulation_code,
                        semester=semester,
                    )
                    .exclude(course_code__isnull=True)
                    .exclude(course_code='')
                    .only(*only_fields)
                )
                curriculum_by_code = {str(getattr(r, 'course_code', '') or '').strip(): r for r in rows if str(getattr(r, 'course_code', '') or '').strip()}
        except Exception:
            curriculum_by_code = {}

        # CQI availability: CQI is configured globally by IQAC; the UI shows a CQI marker
        # when CQI is enabled and the subject is not an AUDIT course.
        try:
            from OBE.models import ObeCqiConfig
            cqi_cfg = ObeCqiConfig.objects.first()
            cqi_globally_enabled = bool(cqi_cfg and (cqi_cfg.options or []))
        except Exception:
            cqi_globally_enabled = False

        def _cycle_key_to_int(v):
            s = str(v or '').strip().lower()
            if not s:
                return None
            # tolerate: "1", "cycle 1", "cycle i", "i", "ii"
            if 'ii' in s or s == '2' or 'cycle 2' in s or 'cycle ii' in s:
                return 2
            if '1' in s or s == 'i' or 'cycle 1' in s or 'cycle i' in s:
                return 1
            return None

        def _internal_maxes_from_mapping(mapping_dict):
            if not isinstance(mapping_dict, dict):
                return (None, None, None)
            weights = mapping_dict.get('weights')
            cycles = mapping_dict.get('cycles')
            if not isinstance(weights, list):
                return (None, None, None)

            w_list = []
            for x in weights:
                try:
                    w_list.append(float(x))
                except Exception:
                    w_list.append(0.0)
            total = sum(w_list) if w_list else None

            if not isinstance(cycles, list) or len(cycles) != len(w_list):
                return (total, None, None)

            c1 = 0.0
            c2 = 0.0
            any_cycle = False
            for i, w in enumerate(w_list):
                ck = _cycle_key_to_int(cycles[i])
                if ck == 1:
                    c1 += w
                    any_cycle = True
                elif ck == 2:
                    c2 += w
                    any_cycle = True

            if not any_cycle:
                return (total, None, None)

            return (total, c1 or None, c2 or None)

        def _num(v):
            try:
                if v is None:
                    return None
                return float(v)
            except Exception:
                return None

        def _clamp(n, lo, hi):
            try:
                return max(lo, min(hi, n))
            except Exception:
                return n

        def _to_float_or_none(v):
            if v is None:
                return None
            if v == '':
                return None
            try:
                n = float(v)
                return n
            except Exception:
                return None

        def _extract_lab_total_for_student(data, student_id):
            if not data or not isinstance(data, dict):
                return None
            sheet = data.get('sheet') if isinstance(data, dict) else None
            if not sheet or not isinstance(sheet, dict):
                return None
            rows = sheet.get('rowsByStudentId')
            if not rows or not isinstance(rows, dict):
                return None
            sid = str(student_id)
            row = rows.get(sid) or rows.get(student_id)
            if not row or not isinstance(row, dict):
                return None

            cia_exam = _to_float_or_none(row.get('ciaExam'))
            if cia_exam is not None:
                return _clamp(cia_exam, 0.0, 100.0)

            total = 0.0
            has_any = False
            all_arrays = []
            if isinstance(row.get('marksA'), list):
                all_arrays.append(row.get('marksA'))
            if isinstance(row.get('marksB'), list):
                all_arrays.append(row.get('marksB'))
            marks_by_co = row.get('marksByCo')
            if isinstance(marks_by_co, dict):
                for arr in marks_by_co.values():
                    if isinstance(arr, list):
                        all_arrays.append(arr)

            for arr in all_arrays:
                for v in arr:
                    n = _to_float_or_none(v)
                    if n is not None:
                        total += n
                        has_any = True

            for field in ('caaExamByCo', 'ciaExamByCo'):
                byco = row.get(field)
                if isinstance(byco, dict):
                    for v in byco.values():
                        n = _to_float_or_none(v)
                        if n is not None:
                            total += n
                            has_any = True

            if not has_any:
                return None
            return _clamp(float(round(total)), 0.0, 100.0)

        def _extract_model_total_for_student(data, student_id):
            if not data or not isinstance(data, dict):
                return None
            sid = str(student_id)

            # Preferred structure (as used in frontend result analysis): { marks: { [sid]: { q1: n, ... } } }
            marks = data.get('marks')
            if isinstance(marks, dict):
                qmarks = marks.get(sid) or marks.get(student_id)
                if isinstance(qmarks, dict):
                    total = 0.0
                    has_any = False
                    for v in qmarks.values():
                        n = _to_float_or_none(v)
                        if n is not None:
                            total += n
                            has_any = True
                    return total if has_any else None

            # Fallback: tolerate a lab-style shape (rowsByStudentId) if deployed that way
            return _extract_lab_total_for_student(data, student_id)

        # Resolve candidate teaching assignments for the student's current section
        # (used to scope published sheets / CQI rows) and to infer class_type / enabled_assessments
        # consistent with staff OBE pages.
        ta_ids_by_code = defaultdict(list)
        ta_meta_by_code = {}
        ta_subject_by_code = {}
        try:
            ay_active = AcademicYear.objects.filter(is_active=True).first() or AcademicYear.objects.order_by('-id').first()
        except Exception:
            ay_active = None

        try:
            if ay_active is not None and section is not None:
                tas_qs = (
                    TeachingAssignment.objects.filter(is_active=True, academic_year=ay_active)
                    .filter(Q(section=section) | Q(section__isnull=True))
                    .select_related('subject', 'curriculum_row', 'elective_subject')
                )
                for ta in tas_qs:
                    tcode = None
                    try:
                        if getattr(ta, 'subject_id', None):
                            tcode = getattr(getattr(ta, 'subject', None), 'code', None)
                        if not tcode and getattr(ta, 'curriculum_row_id', None):
                            tcode = getattr(getattr(ta, 'curriculum_row', None), 'course_code', None)
                        if not tcode and getattr(ta, 'elective_subject_id', None):
                            tcode = getattr(getattr(ta, 'elective_subject', None), 'course_code', None)
                    except Exception:
                        tcode = None
                    tcode = str(tcode or '').strip()
                    if tcode:
                        ta_ids_by_code[tcode].append(getattr(ta, 'id', None))

                        # Prefer section-specific teaching assignment metadata.
                        try:
                            is_section_match = bool(section is not None and getattr(ta, 'section_id', None) == getattr(section, 'id', None))
                        except Exception:
                            is_section_match = False

                        existing = ta_meta_by_code.get(tcode)
                        should_set = existing is None or (is_section_match and not existing.get('section_match'))
                        if should_set:
                            try:
                                enabled = getattr(ta, 'enabled_assessments', None)
                            except Exception:
                                enabled = None
                            if not isinstance(enabled, (list, tuple)):
                                enabled = []
                            cleaned_enabled = []
                            for x in enabled:
                                s = str(x or '').strip().lower()
                                if s:
                                    cleaned_enabled.append(s)

                            try:
                                ta_ct = getattr(ta, 'class_type', None)
                            except Exception:
                                ta_ct = None

                            ta_meta_by_code[tcode] = {
                                'class_type': ta_ct,
                                'enabled_assessments': cleaned_enabled,
                                'section_match': is_section_match,
                            }

                            # Use TA.subject as the authoritative Subject row for marks.
                            # This avoids mismatches when multiple Subject rows share the same code.
                            try:
                                ta_subj = getattr(ta, 'subject', None) if getattr(ta, 'subject_id', None) else None
                            except Exception:
                                ta_subj = None
                            if ta_subj is not None:
                                ta_subject_by_code[tcode] = ta_subj
        except Exception:
            ta_ids_by_code = defaultdict(list)
            ta_meta_by_code = {}
            ta_subject_by_code = {}

        try:
            from OBE.models import LabPublishedSheet, ModelPublishedSheet, ObeCqiPublished
        except Exception:
            LabPublishedSheet = None
            ModelPublishedSheet = None
            ObeCqiPublished = None

        # Build final enrolled code list:
        # - curriculum rows (core)
        # - elective choices
        # - any Subject rows resolved for the student
        codes_set = set()
        for c in (allowed_codes or []):
            cc = str(c or '').strip()
            if cc:
                codes_set.add(cc)
        for cc in (curriculum_by_code or {}).keys():
            if cc:
                codes_set.add(cc)
        for s in subjects:
            sc = str(getattr(s, 'code', '') or '').strip()
            if sc:
                codes_set.add(sc)

        # Map code -> Subject (prefer course-specific Subject when available)
        subject_by_code = {}
        try:
            if codes_set:
                cand = Subject.objects.filter(semester=semester, code__in=sorted(list(codes_set)))
                if course is not None:
                    cand = cand.filter(Q(course=course) | Q(course__isnull=True))

                if course is not None:
                    for s in cand.filter(course=course):
                        k = str(getattr(s, 'code', '') or '').strip()
                        if k and k not in subject_by_code:
                            subject_by_code[k] = s

                for s in cand:
                    k = str(getattr(s, 'code', '') or '').strip()
                    if k and k not in subject_by_code:
                        subject_by_code[k] = s
        except Exception:
            subject_by_code = {}


        bi_data_by_subj = {}
        try:
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM bi_obe_student_subject_wide WHERE student_id = %s", [sp.id])
                cols = [col[0] for col in cursor.description]
                for row in cursor.fetchall():
                    d = dict(zip(cols, row))
                    sid = d.get('subject_id')
                    if sid:
                        bi_data_by_subj[sid] = d
        except Exception:
            pass

        out_courses = []

        for code in sorted(list(codes_set)):
            subj = ta_subject_by_code.get(code) or subject_by_code.get(code)
            # curriculum row metadata (class_type, internal max)
            class_type = None
            internal_max_total = None
            enabled_assessments = None
            try:
                row = curriculum_by_code.get(code)
                if row is not None:
                    class_type = getattr(row, 'class_type', None)
                    try:
                        enabled_assessments = getattr(row, 'enabled_assessments', None)
                    except Exception:
                        enabled_assessments = None
                    im = getattr(row, 'internal_mark', None)
                    if im is not None:
                        try:
                            internal_max_total = float(im)
                        except Exception:
                            internal_max_total = None
            except Exception:
                row = None

            # Prefer TA metadata when present (matches staff OBE pages).
            ta_meta = ta_meta_by_code.get(code) or {}
            ta_ct = str(ta_meta.get('class_type') or '').strip() or None
            if ta_ct:
                class_type = ta_ct
            ta_enabled = ta_meta.get('enabled_assessments')
            if isinstance(ta_enabled, list) and ta_enabled:
                enabled_assessments = ta_enabled
            else:
                # normalize curriculum enabled_assessments when present
                if isinstance(enabled_assessments, (list, tuple)):
                    cleaned = []
                    for x in enabled_assessments:
                        s = str(x or '').strip().lower()
                        if s:
                            cleaned.append(s)
                    enabled_assessments = cleaned
                else:
                    enabled_assessments = []

            display_name = None
            try:
                display_name = getattr(subj, 'name', None) if subj is not None else None
            except Exception:
                display_name = None
            if not display_name:
                try:
                    display_name = getattr(row, 'course_name', None) if row is not None else None
                except Exception:
                    display_name = None
            if not display_name:
                display_name = code

            def _safe_total(model, field: str, ta_ids=None):
                if subj is None:
                    return None
                base = model.objects.filter(subject=subj, student=sp)

                # Prefer TA-scoped row when possible.
                if ta_ids:
                    try:
                        v = (
                            base.filter(Q(teaching_assignment_id__in=ta_ids) | Q(teaching_assignment__isnull=True))
                            .values_list(field, flat=True)
                            .first()
                        )
                        if v is not None:
                            return v
                    except (OperationalError, ProgrammingError):
                        # tolerate DBs missing teaching_assignment_id
                        pass
                    except Exception:
                        pass

                # Then prefer legacy (unscoped) row.
                try:
                    v = base.filter(teaching_assignment__isnull=True).values_list(field, flat=True).first()
                    if v is not None:
                        return v
                except (OperationalError, ProgrammingError):
                    pass
                except Exception:
                    pass

                # Final fallback: any row (works even when teaching_assignment_id is missing)
                try:
                    return base.values_list(field, flat=True).first()
                except Exception:
                    return None

            # internal mapping (may be None)
            try:
                if subj is not None:
                    imm = InternalMarkMapping.objects.filter(subject=subj).first()
                    mapping = imm.mapping if imm else None
                else:
                    mapping = None
            except Exception:
                mapping = None

            map_total, map_c1, map_c2 = _internal_maxes_from_mapping(mapping)
            if internal_max_total is None:
                internal_max_total = map_total
            internal_max_cycle1 = map_c1
            internal_max_cycle2 = map_c2
            if internal_max_total is not None and (internal_max_cycle1 is None and internal_max_cycle2 is None):
                # fallback split when only a total is known
                internal_max_cycle1 = internal_max_total / 2.0
                internal_max_cycle2 = internal_max_total / 2.0

            marks_vals = {
                'cia1': _num(_safe_total(Cia1Mark, 'mark', ta_ids=(ta_ids_by_code.get(code) or []))),
                'cia2': _num(_safe_total(Cia2Mark, 'mark', ta_ids=(ta_ids_by_code.get(code) or []))),
                'ssa1': _num(_safe_total(Ssa1Mark, 'mark', ta_ids=(ta_ids_by_code.get(code) or []))),
                'ssa2': _num(_safe_total(Ssa2Mark, 'mark', ta_ids=(ta_ids_by_code.get(code) or []))),
                'review1': _num(_safe_total(Review1Mark, 'mark', ta_ids=(ta_ids_by_code.get(code) or []))),
                'review2': _num(_safe_total(Review2Mark, 'mark', ta_ids=(ta_ids_by_code.get(code) or []))),
                'formative1': _num(_safe_total(Formative1Mark, 'total', ta_ids=(ta_ids_by_code.get(code) or []))),
                'formative2': _num(_safe_total(Formative2Mark, 'total', ta_ids=(ta_ids_by_code.get(code) or []))),
                'model': None,
            }

            # Backfill missing totals from published sheets when the publish flow
            # does not upsert into the per-student totals tables (notably LAB/TCPL).
            ta_ids = [x for x in (ta_ids_by_code.get(code) or []) if x]
            if subj is not None and LabPublishedSheet is not None:
                try:
                    for assessment, key in (
                        ('cia1', 'cia1'),
                        ('cia2', 'cia2'),
                        ('formative1', 'formative1'),
                        ('formative2', 'formative2'),
                        ('review1', 'review1'),
                        ('review2', 'review2'),
                        ('model', 'model'),
                    ):
                        if marks_vals.get(key) is not None:
                            continue
                        qs = LabPublishedSheet.objects.filter(subject=subj, assessment=assessment)
                        if ta_ids:
                            qs = qs.filter(Q(teaching_assignment_id__in=ta_ids) | Q(teaching_assignment__isnull=True))
                        else:
                            qs = qs.filter(Q(teaching_assignment__isnull=True))
                        row = qs.order_by('-updated_at').only('data', 'updated_at').first()
                        if row:
                            v = _extract_lab_total_for_student(getattr(row, 'data', None), sp.id)
                            if v is not None:
                                marks_vals[key] = float(v)
                except Exception:
                    pass

            if subj is not None and marks_vals.get('model') is None and ModelPublishedSheet is not None:
                try:
                    qs = ModelPublishedSheet.objects.filter(subject=subj)
                    if ta_ids:
                        qs = qs.filter(Q(teaching_assignment_id__in=ta_ids) | Q(teaching_assignment__isnull=True))
                    else:
                        qs = qs.filter(Q(teaching_assignment__isnull=True))
                    row = qs.order_by('-updated_at').only('data', 'updated_at').first()
                    if row:
                        v = _extract_model_total_for_student(getattr(row, 'data', None), sp.id)
                        if v is not None:
                            marks_vals['model'] = float(v)
                except Exception:
                    pass

            # best-effort internal marks: sum available internal-like components
            internal_components = [
                marks_vals.get('formative1'),
                marks_vals.get('formative2'),
                marks_vals.get('ssa1'),
                marks_vals.get('ssa2'),
                marks_vals.get('review1'),
                marks_vals.get('review2'),
            ]
            internal_components = [x for x in internal_components if x is not None]
            internal_computed = sum(internal_components) if internal_components else None

            internal_cycle1_components = [
                marks_vals.get('formative1'),
                marks_vals.get('ssa1'),
                marks_vals.get('review1'),
            ]
            internal_cycle1_components = [x for x in internal_cycle1_components if x is not None]
            internal_cycle1 = sum(internal_cycle1_components) if internal_cycle1_components else None

            internal_cycle2_components = [
                marks_vals.get('formative2'),
                marks_vals.get('ssa2'),
                marks_vals.get('review2'),
            ]
            internal_cycle2_components = [x for x in internal_cycle2_components if x is not None]
            internal_cycle2 = sum(internal_cycle2_components) if internal_cycle2_components else None

            # Totals aligned to how users talk about "Cycle" / "Total":
            # - cycle totals include CIA
            # - overall total includes CIA + model
            c1_total_parts = [x for x in [internal_cycle1, marks_vals.get('cia1')] if x is not None]
            c2_total_parts = [x for x in [internal_cycle2, marks_vals.get('cia2')] if x is not None]
            internal_cycle1_total = sum(c1_total_parts) if c1_total_parts else None
            internal_cycle2_total = sum(c2_total_parts) if c2_total_parts else None

            all_total_parts = [x for x in [internal_cycle1_total, internal_cycle2_total, marks_vals.get('model')] if x is not None]
            internal_total = sum(all_total_parts) if all_total_parts else None

            ct_norm = str(class_type or '').upper()
            # In this codebase, class_type values include THEORY/LAB/TCPR/TCPL/PRACTICAL/PROJECT/SPECIAL.
            # CQI is configured globally by IQAC; show it for all academic class types except AUDIT.
            has_cqi = bool(cqi_globally_enabled and ct_norm != 'AUDIT')

            # CIA max defaults (no authoritative config in DB yet)
            cia_max = 30.0

            # Optional CO attainment values when CQI is published.
            cos = None
            if subj is not None and ObeCqiPublished is not None:
                try:
                    qs = ObeCqiPublished.objects.filter(subject=subj)
                    if ta_ids:
                        qs = qs.filter(teaching_assignment_id__in=ta_ids)
                    row = qs.order_by('-published_at').only('entries', 'published_at').first()
                    if row and isinstance(getattr(row, 'entries', None), dict):
                        ent = row.entries.get(str(sp.id)) or row.entries.get(sp.id)
                        if isinstance(ent, dict):
                            cos = {str(k): _num(v) for k, v in ent.items()}
                except Exception:
                    cos = None

            out_courses.append(
                {
                    'id': getattr(subj, 'id', None),
                    'code': code,
                    'name': display_name,
                    'class_type': ct_norm or None,
                    'enabled_assessments': enabled_assessments,
                    'marks': {
                        'cia1': marks_vals.get('cia1'),
                        'cia2': marks_vals.get('cia2'),
                        'cia_max': cia_max,
                        'ssa1': marks_vals.get('ssa1'),
                        'ssa2': marks_vals.get('ssa2'),
                        'review1': marks_vals.get('review1'),
                        'review2': marks_vals.get('review2'),
                        'formative1': marks_vals.get('formative1'),
                        'formative2': marks_vals.get('formative2'),
                        'model': marks_vals.get('model'),
                        'internal': {
                            'computed': internal_computed,
                            'cycle1': internal_cycle1,
                            'cycle2': internal_cycle2,
                            'cycle1_total': internal_cycle1_total,
                            'cycle2_total': internal_cycle2_total,
                            'total': internal_total,
                            'max_total': internal_max_total,
                            'max_cycle1': internal_max_cycle1,
                            'max_cycle2': internal_max_cycle2,
                            'mapping': mapping,
                        },
                        'has_cqi': has_cqi,
                        **({'cos': cos} if cos is not None else {}),
                        'bi': {k: (float(v) if isinstance(v, decimal.Decimal) else v) for k, v in bi_data_by_subj.get(getattr(subj, 'id', None), {}).items() if v is not None} if getattr(subj, 'id', None) else {},
                    },
                }
            )

        resp = {
            'student': {
                'id': sp.id,
                'reg_no': sp.reg_no,
                'name': getattr(getattr(sp, 'user', None), 'username', None),
            },
            'semester': {'id': getattr(semester, 'id', None), 'number': getattr(semester, 'number', None)},
            'courses': out_courses,
        }

        return Response(resp)


class StudentSectionSubjectsView(APIView):
    """Return students in the logged-in student's section + their enrolled subjects.

    URL: /api/academics/student/section-subjects/
    Scope: only the student's current section; minimal identity fields.
    """

    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        sp = getattr(user, 'student_profile', None)
        if sp is None:
            return Response({'detail': 'Student profile not found for user.'}, status=status.HTTP_403_FORBIDDEN)

        section = sp.get_current_section() or getattr(sp, 'section', None)
        if section is None:
            return Response({'section': None, 'semester': None, 'students': []})

        semester = getattr(section, 'semester', None)
        course = getattr(getattr(section, 'batch', None), 'course', None)
        dept = getattr(course, 'department', None) if course is not None else None

        from django.db.models import Q
        from collections import defaultdict
        from .models import StudentProfile, Subject, AcademicYear

        # Prefer active assignment mapping but tolerate legacy `section` field.
        students_qs = (
            StudentProfile.objects.filter(
                Q(section=section) |
                Q(section_assignments__section=section, section_assignments__end_date__isnull=True)
            )
            .select_related('user')
            .distinct()
            .order_by('reg_no')
        )

        # If semester is unknown, return students with empty subject lists.
        if semester is None:
            out = []
            for s in students_qs:
                u = getattr(s, 'user', None)
                display_name = ' '.join([x for x in [getattr(u, 'first_name', ''), getattr(u, 'last_name', '')] if x]).strip()
                if not display_name:
                    display_name = getattr(u, 'username', '') or s.reg_no
                out.append({'id': s.id, 'reg_no': s.reg_no, 'name': display_name, 'subjects': []})
            return Response({
                'section': {'id': getattr(section, 'id', None), 'name': getattr(section, 'name', None)},
                'semester': None,
                'students': out,
            })

        # Curriculum-based enrollment resolution (core + per-student electives)
        try:
            from curriculum.models import CurriculumDepartment
        except Exception:
            CurriculumDepartment = None

        try:
            from curriculum.models import ElectiveChoice
        except Exception:
            ElectiveChoice = None

        regulation_code = None
        try:
            regulation_code = getattr(getattr(getattr(section, 'batch', None), 'regulation', None), 'code', None)
        except Exception:
            regulation_code = None
        if not regulation_code:
            try:
                regulation_code = str(getattr(getattr(section, 'batch', None), 'regulation', '') or '').strip() or None
            except Exception:
                regulation_code = None

        core_codes = []
        try:
            if CurriculumDepartment is not None and dept is not None and regulation_code and semester is not None:
                core_codes = list(
                    CurriculumDepartment.objects.filter(
                        department=dept,
                        regulation=regulation_code,
                        semester=semester,
                        is_elective=False,
                    )
                    .exclude(course_code__isnull=True)
                    .exclude(course_code='')
                    .values_list('course_code', flat=True)
                )
        except Exception:
            core_codes = []

        core_set = set([str(c).strip() for c in core_codes if c])

        electives_by_student_id = defaultdict(set)
        all_elective_codes = set()
        try:
            if ElectiveChoice is not None:
                try:
                    ay = AcademicYear.objects.filter(is_active=True).first()
                except Exception:
                    ay = None

                eqs = ElectiveChoice.objects.filter(student__in=students_qs, is_active=True).select_related('elective_subject', 'student')
                if ay is not None:
                    eqs = eqs.filter(models.Q(academic_year=ay) | models.Q(academic_year__isnull=True))

                for ch in eqs:
                    st = getattr(ch, 'student', None)
                    es = getattr(ch, 'elective_subject', None)
                    if not st or not es:
                        continue
                    if dept is not None and getattr(es, 'department_id', None) != getattr(dept, 'id', None):
                        continue
                    if regulation_code and str(getattr(es, 'regulation', '') or '').strip() != str(regulation_code):
                        continue
                    if getattr(es, 'semester_id', None) != getattr(semester, 'id', None):
                        continue
                    code = str(getattr(es, 'course_code', '') or '').strip()
                    if not code:
                        continue
                    electives_by_student_id[getattr(st, 'id', None)].add(code)
                    all_elective_codes.add(code)
        except Exception:
            electives_by_student_id = defaultdict(set)
            all_elective_codes = set()

        # Decide subject resolution strategy
        subject_source = 'curriculum' if (core_set or all_elective_codes) else 'course_fallback'

        base_subjects_list = None
        subject_by_code = {}
        if subject_source == 'curriculum':
            union_codes = set(core_set) | set(all_elective_codes)
            subjects_qs = Subject.objects.filter(semester=semester, code__in=sorted(list(union_codes))).only('code', 'name').order_by('code')
            subject_by_code = {getattr(s, 'code', None): {'code': s.code, 'name': s.name} for s in subjects_qs}
        else:
            if course is not None:
                base_subjects_qs = Subject.objects.filter(semester=semester, course=course).only('code', 'name').order_by('code')
                base_subjects_list = [{'code': s.code, 'name': s.name} for s in base_subjects_qs]
            else:
                base_subjects_list = []

        out = []
        for s in students_qs:
            u = getattr(s, 'user', None)
            display_name = ' '.join([x for x in [getattr(u, 'first_name', ''), getattr(u, 'last_name', '')] if x]).strip()
            if not display_name:
                display_name = getattr(u, 'username', '') or s.reg_no

            if subject_source == 'curriculum':
                codes = set(core_set) | set(electives_by_student_id.get(s.id, set()))
                subjects_list = [subject_by_code[c] for c in sorted(list(codes)) if c in subject_by_code]
            else:
                subjects_list = base_subjects_list

            out.append({
                'id': s.id,
                'reg_no': s.reg_no,
                'name': display_name,
                'subjects': subjects_list,
            })

        return Response({
            'section': {
                'id': getattr(section, 'id', None),
                'name': getattr(section, 'name', None),
            },
            'semester': {
                'id': getattr(semester, 'id', None),
                'number': getattr(semester, 'number', None),
            },
            'subject_source': subject_source,
            'students': out,
        })


class DepartmentStudentsView(APIView):
    """Return students from the same department as the current user.
    
    Requires 'students.view_department_students' permission.

    Without ?section_id  → returns section list (metadata, no students)
    With    ?section_id=X → returns students for that one section only
    """
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):
        import traceback, logging
        try:
            user = request.user
            
            from accounts.utils import get_user_permissions
            perms = get_user_permissions(user)
            
            if not ('students.view_department_students' in perms or user.has_perm('students.view_department_students')):
                return Response({'error': 'Permission denied'}, status=403)
                
            staff_profile = getattr(user, 'staff_profile', None)
            if not staff_profile:
                return Response({'sections': [], 'results': []})

            from .models import Section, StudentSectionAssignment, StudentProfile

            # Collect all departments the user is associated with:
            # primary department + any active DepartmentRole entries (HOD/AHOD of multiple depts)
            dept_ids = set()
            user_dept = getattr(staff_profile, 'department', None)
            if user_dept:
                dept_ids.add(user_dept.id)
            role_dept_ids = DepartmentRole.objects.filter(
                staff=staff_profile, is_active=True
            ).values_list('department_id', flat=True)
            dept_ids.update(role_dept_ids)

            if not dept_ids:
                return Response({'sections': [], 'results': []})

            # Own sections: batch belongs to the user's departments (via course or direct dept FK)
            own_sections = Section.objects.filter(
                Q(batch__course__department_id__in=dept_ids) | Q(batch__department_id__in=dept_ids)
            ).select_related(
                'batch', 'batch__course', 'batch__course__department', 'batch__department'
            )

            # Shared sections: dept-only batches (e.g. S&H Year-1) that contain students
            # whose home_department is one of the user's departments. These students are
            # temporarily housed in the shared batch for Year-1 but belong to the core dept.
            from django.db.models import Exists, OuterRef
            has_home_dept_student = StudentSectionAssignment.objects.filter(
                section_id=OuterRef('pk'),
                end_date__isnull=True,
                student__home_department_id__in=dept_ids,
            )
            shared_sections = Section.objects.filter(
                batch__course__isnull=True,       # dept-only batch (no course → S&H style)
                batch__department__isnull=False,   # has explicit batch department
            ).exclude(
                batch__department_id__in=dept_ids  # exclude own dept-only batches
            ).filter(
                Exists(has_home_dept_student)
            ).select_related(
                'batch', 'batch__department'
            )

            # Combined set IDs for fast lookup
            own_section_ids = set(own_sections.values_list('id', flat=True))
            shared_section_ids = set(shared_sections.values_list('id', flat=True))
            all_accessible_ids = own_section_ids | shared_section_ids

            section_id_param = request.query_params.get('section_id')

            # ── Section list mode (no section_id param) ──────────────────────
            if not section_id_param:
                section_list = []
                # Own sections
                for sec in own_sections.order_by('batch__name', 'name'):
                    batch = getattr(sec, 'batch', None)
                    course = getattr(batch, 'course', None) if batch else None
                    dept = (getattr(course, 'department', None) if course else None) or getattr(batch, 'department', None)
                    section_list.append({
                        'section_id': sec.id,
                        'section_name': sec.name,
                        'batch_name': getattr(batch, 'name', None),
                        'department_code': getattr(dept, 'code', None) if dept else None,
                        'department_short_name': (getattr(dept, 'short_name', None) or getattr(dept, 'code', None)) if dept else None,
                        'department_name': getattr(dept, 'name', None) if dept else None,
                        'is_shared_section': False,
                    })
                # Shared sections (Year-1 / S&H style)
                for sec in shared_sections.order_by('batch__name', 'name'):
                    batch = getattr(sec, 'batch', None)
                    shared_dept = getattr(batch, 'department', None) if batch else None
                    section_list.append({
                        'section_id': sec.id,
                        'section_name': sec.name,
                        'batch_name': getattr(batch, 'name', None),
                        'department_code': getattr(shared_dept, 'code', None) if shared_dept else None,
                        'department_short_name': (getattr(shared_dept, 'short_name', None) or getattr(shared_dept, 'code', None)) if shared_dept else None,
                        'department_name': getattr(shared_dept, 'name', None) if shared_dept else None,
                        'is_shared_section': True,
                    })
                return Response({'sections': section_list})

            # ── Single section students mode ─────────────────────────────────
            try:
                section_id_int = int(section_id_param)
            except ValueError:
                return Response({'error': 'Invalid section_id'}, status=400)

            if section_id_int not in all_accessible_ids:
                return Response({'error': 'Section not found or not in your departments'}, status=404)

            is_shared = section_id_int in shared_section_ids

            # Fetch the section object
            sec = (shared_sections if is_shared else own_sections).filter(id=section_id_int).first()
            if not sec:
                return Response({'error': 'Section not found or not in your departments'}, status=404)

            batch = getattr(sec, 'batch', None)
            course = getattr(batch, 'course', None) if batch else None
            dept = (getattr(course, 'department', None) if course else None) or getattr(batch, 'department', None)

            # For shared sections only return students belonging to the user's home departments
            if is_shared:
                assign_qs = StudentSectionAssignment.objects.filter(
                    section_id=section_id_int,
                    end_date__isnull=True,
                    student__home_department_id__in=dept_ids,
                ).select_related('student__user', 'student__home_department')
                legacy_qs = StudentProfile.objects.filter(
                    section_id=section_id_int,
                    home_department_id__in=dept_ids,
                ).select_related('user', 'home_department')
            else:
                assign_qs = StudentSectionAssignment.objects.filter(
                    section_id=section_id_int, end_date__isnull=True
                ).select_related('student__user')
                legacy_qs = StudentProfile.objects.filter(section_id=section_id_int).select_related('user')

            studs = []
            for a in assign_qs:
                studs.append(a.student)

            present_pks = {s.pk for s in studs}
            for s in legacy_qs:
                if s.pk not in present_pks:
                    studs.append(s)

            students_out = []
            for st in studs:
                user_obj = getattr(st, 'user', None)
                full_name = (
                    f"{getattr(user_obj, 'first_name', '')} {getattr(user_obj, 'last_name', '')}".strip()
                    if user_obj else ''
                )
                home_dept_obj = getattr(st, 'home_department', None)
                students_out.append({
                    'id': st.pk,
                    'reg_no': st.reg_no,
                    'name': full_name or getattr(user_obj, 'username', None),
                    'username': getattr(user_obj, 'username', None),
                    'first_name': getattr(user_obj, 'first_name', '') if user_obj else '',
                    'last_name': getattr(user_obj, 'last_name', '') if user_obj else '',
                    'email': getattr(user_obj, 'email', '') if user_obj else '',
                    'status': getattr(st, 'status', 'ACTIVE').lower(),
                    'section_id': sec.id,
                    'section_name': sec.name,
                    'department_code': getattr(dept, 'code', None) if dept else None,
                    'department_name': getattr(dept, 'name', None) if dept else None,
                    'home_department_code': getattr(home_dept_obj, 'code', None) if home_dept_obj else None,
                    'home_department_short_name': getattr(home_dept_obj, 'short_name', None) if home_dept_obj else None,
                    'batch': getattr(batch, 'name', None),
                    'is_shared_section': is_shared,
                })

            return Response({
                'section_id': sec.id,
                'section_name': sec.name,
                'batch_name': getattr(batch, 'name', None),
                'department_code': getattr(dept, 'code', None) if dept else None,
                'department_name': getattr(dept, 'name', None) if dept else None,
                'is_shared_section': is_shared,
                'students': sorted(students_out, key=lambda x: x['reg_no'] or ''),
            })
            
        except Exception as e:
            logging.getLogger(__name__).exception('DepartmentStudentsView error: %s', e)
            return Response({'error': 'Server error', 'detail': str(e)}, status=500)


class AllStudentsView(APIView):
    """Return students from all departments.
    
    Requires 'students.view_all_students' permission.

    Without ?section_id  → returns section list (metadata, no students)
    With    ?section_id=X → returns students for that one section only
    """
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):
        import logging
        try:
            user = request.user
            
            from accounts.utils import get_user_permissions
            perms = get_user_permissions(user)
            
            if not ('students.view_all_students' in perms or user.has_perm('students.view_all_students')):
                return Response({'error': 'Permission denied'}, status=403)
                
            from .models import Section, StudentSectionAssignment, StudentProfile
            
            sections = Section.objects.filter(
                Q(batch__course__department__isnull=False) | Q(batch__department__isnull=False)
            ).select_related(
                'batch', 'batch__course', 'batch__course__department', 'batch__department'
            ).order_by('batch__name', 'name')

            section_id_param = request.query_params.get('section_id')

            # ── Section list mode ────────────────────────────────────────────
            if not section_id_param:
                section_list = []
                for sec in sections:
                    batch = getattr(sec, 'batch', None)
                    course = getattr(batch, 'course', None) if batch else None
                    dept = (getattr(course, 'department', None) if course else None) or getattr(batch, 'department', None)
                    section_list.append({
                        'section_id': sec.id,
                        'section_name': sec.name,
                        'batch_name': getattr(batch, 'name', None),
                        'department_code': getattr(dept, 'code', None) if dept else None,
                        'department_short_name': (getattr(dept, 'short_name', None) or getattr(dept, 'code', None)) if dept else None,
                        'department_name': getattr(dept, 'name', None) if dept else None,
                    })
                return Response({'sections': section_list})

            # ── Single section students mode ─────────────────────────────────
            try:
                section_id_int = int(section_id_param)
            except ValueError:
                return Response({'error': 'Invalid section_id'}, status=400)

            sec = sections.filter(id=section_id_int).first()
            if not sec:
                return Response({'error': 'Section not found'}, status=404)

            batch = getattr(sec, 'batch', None)
            course = getattr(batch, 'course', None) if batch else None
            dept = (getattr(course, 'department', None) if course else None) or getattr(batch, 'department', None)

            studs = []
            assign_qs = StudentSectionAssignment.objects.filter(
                section_id=section_id_int, end_date__isnull=True
            ).select_related('student__user')
            for a in assign_qs:
                studs.append(a.student)

            legacy_qs = StudentProfile.objects.filter(section_id=section_id_int).select_related('user')
            present_pks = {s.pk for s in studs}
            for s in legacy_qs:
                if s.pk not in present_pks:
                    studs.append(s)

            students_out = []
            for st in studs:
                user_obj = getattr(st, 'user', None)
                full_name = (
                    f"{getattr(user_obj, 'first_name', '')} {getattr(user_obj, 'last_name', '')}".strip()
                    if user_obj else ''
                )
                students_out.append({
                    'id': st.pk,
                    'reg_no': st.reg_no,
                    'name': full_name or getattr(user_obj, 'username', None),
                    'username': getattr(user_obj, 'username', None),
                    'first_name': getattr(user_obj, 'first_name', '') if user_obj else '',
                    'last_name': getattr(user_obj, 'last_name', '') if user_obj else '',
                    'email': getattr(user_obj, 'email', '') if user_obj else '',
                    'status': getattr(st, 'status', 'ACTIVE').lower(),
                    'section_id': sec.id,
                    'section_name': sec.name,
                    'department_code': getattr(dept, 'code', None) if dept else None,
                    'department_name': getattr(dept, 'name', None) if dept else None,
                    'batch': getattr(batch, 'name', None),
                })

            return Response({
                'section_id': sec.id,
                'section_name': sec.name,
                'batch_name': getattr(batch, 'name', None),
                'department_code': getattr(dept, 'code', None) if dept else None,
                'department_name': getattr(dept, 'name', None) if dept else None,
                'students': sorted(students_out, key=lambda x: x['reg_no'] or ''),
            })
            
        except Exception as e:
            logging.getLogger(__name__).exception('AllStudentsView error: %s', e)
            return Response({'error': 'Server error', 'detail': str(e)}, status=500)


class StudentProfileUpdateView(APIView):
    """Update editable student details (username/email/status/section)."""
    permission_classes = (IsAuthenticated,)

    def patch(self, request, student_id: int):
        user = request.user
        perms = set(get_user_permissions(user) or [])
        allowed = user.is_superuser or bool(
            perms.intersection({
                'students.view_all_students',
                'students.view_department_students',
                'academics.view_my_students',
                'academics.view_mentees',
            })
        )
        if not allowed:
            return Response({'error': 'Permission denied'}, status=403)

        student = StudentProfile.objects.select_related(
            'user',
            'section__batch__course__department',
            'section__batch__department',
            'home_department',
        ).filter(pk=student_id).first()
        if not student:
            return Response({'error': 'Student not found'}, status=404)

        data = request.data or {}
        user_obj = student.user

        username_changed = False
        email_changed = False

        username = str(data.get('username', user_obj.username or '')).strip()
        if username and username != user_obj.username:
            from accounts.models import User
            exists = User.objects.filter(username__iexact=username).exclude(pk=user_obj.pk).exists()
            if exists:
                return Response({'error': 'Username already exists'}, status=400)
            user_obj.username = username
            username_changed = True

        if 'email' in data:
            email = str(data.get('email') or '').strip()
            if email != (user_obj.email or ''):
                from accounts.models import User
                exists = User.objects.filter(email__iexact=email).exclude(pk=user_obj.pk).exists() if email else False
                if exists:
                    return Response({'error': 'Email already exists'}, status=400)
                user_obj.email = email
                email_changed = True

        if username_changed or email_changed:
            update_fields = []
            if username_changed:
                update_fields.append('username')
            if email_changed:
                update_fields.append('email')
            user_obj.save(update_fields=update_fields)

        if 'status' in data:
            status_value = str(data.get('status') or '').strip().upper()
            valid_statuses = {'ACTIVE', 'INACTIVE', 'ALUMNI', 'DEBAR'}
            if status_value not in valid_statuses:
                return Response({'error': 'Invalid status value'}, status=400)
            student.status = status_value

        if 'section_id' in data or 'section' in data:
            raw_section = data.get('section_id', data.get('section'))
            if raw_section in (None, '', 'null'):
                student.section = None
            else:
                try:
                    section_id = int(raw_section)
                except (TypeError, ValueError):
                    return Response({'error': 'Invalid section_id'}, status=400)
                section_obj = Section.objects.filter(pk=section_id).first()
                if not section_obj:
                    return Response({'error': 'Section not found'}, status=404)
                student.section = section_obj

        try:
            student.save()
        except Exception as exc:
            return Response({'error': str(exc)}, status=400)

        student.refresh_from_db()

        section_obj = getattr(student, 'section', None)
        batch = getattr(section_obj, 'batch', None) if section_obj else None
        course = getattr(batch, 'course', None) if batch else None
        dept = (getattr(course, 'department', None) if course else None) or (getattr(batch, 'department', None) if batch else None)
        home_dept = getattr(student, 'home_department', None)

        display_name = f"{getattr(user_obj, 'first_name', '')} {getattr(user_obj, 'last_name', '')}".strip() or user_obj.username

        return Response({
            'id': student.pk,
            'reg_no': student.reg_no,
            'name': display_name,
            'username': user_obj.username,
            'first_name': getattr(user_obj, 'first_name', '') or '',
            'last_name': getattr(user_obj, 'last_name', '') or '',
            'email': getattr(user_obj, 'email', '') or '',
            'status': (student.status or 'ACTIVE').lower(),
            'section_id': getattr(section_obj, 'id', None),
            'section_name': getattr(section_obj, 'name', None),
            'batch': getattr(batch, 'name', None),
            'department_code': getattr(dept, 'code', None) if dept else None,
            'department_short_name': (getattr(dept, 'short_name', None) or getattr(dept, 'code', None)) if dept else None,
            'department_name': getattr(dept, 'name', None) if dept else None,
            'home_department_code': getattr(home_dept, 'code', None) if home_dept else None,
            'home_department_short_name': getattr(home_dept, 'short_name', None) if home_dept else None,
        }, status=200)


class BatchYearViewSet(viewsets.ModelViewSet):
    """CRUD for common BatchYear labels (e.g. '2023') that are department-agnostic."""
    permission_classes = (IsAuthenticated,)

    def get_serializer_class(self):
        from rest_framework import serializers
        from .models import BatchYear

        class BatchYearSerializer(serializers.ModelSerializer):
            class Meta:
                model = BatchYear
                fields = ('id', 'name', 'start_year', 'end_year')

        return BatchYearSerializer

    def get_queryset(self):
        from .models import BatchYear
        return BatchYear.objects.all().order_by('-name')


class BatchListView(APIView):
    """List all Batch objects with department info and is_active status.

    Used by frontend dropdowns that need per-department active flags.
    Returns: id, name, department_code, department_short_name, is_active
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        from .models import Batch
        qs = Batch.objects.select_related('course__department', 'department').order_by('-name')
        results = []
        for b in qs:
            dept = b.effective_department
            results.append({
                'id': b.id,
                'name': b.name,
                'department_code': getattr(dept, 'code', None),
                'department_short_name': getattr(dept, 'short_name', None),
                'is_active': b.is_active,
            })
        return Response(results)


class AllStaffListView(APIView):
    """Return all staff members from the database (not department-filtered).
    
    Used for listing all available staff to add to a department.
    Requires academics.view_staffs_page permission.
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        has_ps_role = user.roles.filter(name__iexact='PS').exists()

        # Require page-view permission unless superuser
        if not (user.is_superuser or has_ps_role or 'academics.view_staffs_page' in perms):
            return Response({'detail': 'You do not have permission to view staff list.'}, status=403)

        from .models import StaffProfile, DepartmentRole, AcademicYear

        # Get all staff members ordered by staff_id
        staff_qs = StaffProfile.objects.select_related('user', 'department').order_by('staff_id')

        # Get active academic year
        active_year = AcademicYear.objects.filter(is_active=True).first()

        # Get all active department roles for current academic year
        dept_roles = {}
        if active_year:
            for dr in DepartmentRole.objects.filter(
                academic_year=active_year,
                is_active=True
            ).select_related('department', 'staff'):
                staff_id = dr.staff_id
                if staff_id not in dept_roles:
                    dept_roles[staff_id] = []
                dept_roles[staff_id].append({
                    'department': {
                        'id': dr.department.id,
                        'code': dr.department.code,
                        'name': dr.department.name,
                        'short_name': dr.department.short_name,
                    },
                    'role': dr.role,
                    'academic_year': active_year.name,
                })

        results = []
        for s in staff_qs:
            user_data = None
            user_roles = []
            if s.user:
                user_data = {
                    'username': s.user.username,
                    'first_name': getattr(s.user, 'first_name', ''),
                    'last_name': getattr(s.user, 'last_name', ''),
                    'email': getattr(s.user, 'email', ''),
                }
                # Get user roles
                try:
                    user_roles = [r.name for r in s.user.roles.all()]
                except Exception:
                    user_roles = []

            dept_data = None
            if s.department:
                dept_data = {
                    'id': s.department.id,
                    'code': s.department.code,
                    'name': s.department.name,
                    'short_name': s.department.short_name,
                }

            results.append({
                'id': s.id,
                'staff_id': s.staff_id,
                'internal_id': s.internal_id,
                'user': user_data,
                'user_roles': user_roles,
                'designation': getattr(s, 'designation', None),
                'status': getattr(s, 'status', None),
                'current_department': dept_data,
                'department_roles': dept_roles.get(s.id, []),
            })

        return Response({'results': results})


class StaffInternalIdShuffleView(APIView):
    """Shuffle internal IDs for staff members after password confirmation."""
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        user = request.user
        perms = get_user_permissions(user)
        has_ps_role = user.roles.filter(name__iexact='PS').exists()

        if not (user.is_superuser or has_ps_role or 'academics.edit_staff' in perms):
            return Response(
                {'detail': 'You do not have permission to shuffle internal IDs.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        password = str(request.data.get('password') or '').strip()
        if not password:
            return Response(
                {'detail': 'Password is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not user.check_password(password):
            return Response(
                {'detail': 'Invalid password.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        department_id = request.data.get('department_id')

        staff_qs = StaffProfile.objects.all().order_by('id')

        if department_id not in (None, '', 'all'):
            try:
                department_id = int(department_id)
            except (TypeError, ValueError):
                return Response(
                    {'detail': 'Invalid department_id.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Validate department scope for non-superusers without view_all_staff.
            if not user.is_superuser and not has_ps_role and 'academics.view_all_staff' not in perms:
                allowed_dept_ids = get_user_effective_departments(user)
                if not allowed_dept_ids or department_id not in allowed_dept_ids:
                    return Response(
                        {'detail': 'You can only shuffle IDs for departments you manage.'},
                        status=status.HTTP_403_FORBIDDEN,
                    )

            staff_qs = staff_qs.filter(
                Q(department_id=department_id) |
                Q(department_assignments__department_id=department_id, department_assignments__end_date__isnull=True) |
                Q(department_roles__department_id=department_id, department_roles__is_active=True)
            ).distinct().order_by('id')
        elif not user.is_superuser and not has_ps_role and 'academics.view_all_staff' not in perms:
            # For scoped users, shuffle only departments they effectively manage.
            allowed_dept_ids = get_user_effective_departments(user)
            if not allowed_dept_ids:
                return Response(
                    {'detail': 'You are not mapped to any departments.'},
                    status=status.HTTP_403_FORBIDDEN,
                )

            staff_qs = staff_qs.filter(
                Q(department_id__in=allowed_dept_ids) |
                Q(department_assignments__department_id__in=allowed_dept_ids, department_assignments__end_date__isnull=True) |
                Q(department_roles__department_id__in=allowed_dept_ids, department_roles__is_active=True)
            ).distinct().order_by('id')

        staff_list = list(staff_qs)
        if not staff_list:
            return Response({'detail': 'No staff records found to shuffle.', 'updated': 0}, status=status.HTTP_200_OK)

        existing_ids = set(
            StaffProfile.objects.exclude(internal_id__isnull=True).exclude(internal_id='').values_list('internal_id', flat=True)
        )

        with transaction.atomic():
            for s in staff_list:
                # Remove current value from occupied set so it can be reassigned.
                if s.internal_id:
                    existing_ids.discard(s.internal_id)

            updates = []
            for s in staff_list:
                candidate = None
                for _ in range(200):
                    maybe = StaffProfile.generate_unique_internal_id()
                    if maybe not in existing_ids:
                        candidate = maybe
                        break
                if not candidate:
                    return Response(
                        {'detail': 'Unable to generate unique internal IDs. Try again.'},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )

                existing_ids.add(candidate)
                s.internal_id = candidate
                updates.append(s)

            StaffProfile.objects.bulk_update(updates, ['internal_id'])

        return Response(
            {
                'detail': f'Internal IDs shuffled successfully for {len(staff_list)} staff member(s).',
                'updated': len(staff_list),
            },
            status=status.HTTP_200_OK,
        )


class StaffDepartmentAssignView(APIView):
    """Assign a staff member to a department with a specific role.
    
    POST /api/academics/staff-department-assign/
    Body: { 
        staff_id: <int>, 
        department_id: <int>,
        role: 'STAFF' | 'HOD' | 'AHOD' (default: 'STAFF')
    }
    
    Requires academics.edit_staff permission.
    """
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        user = request.user
        perms = get_user_permissions(user)
        has_ps_role = user.roles.filter(name__iexact='PS').exists()

        # Check permission
        if not (user.is_superuser or has_ps_role or 'academics.edit_staff' in perms):
            return Response(
                {'detail': 'You do not have permission to assign staff to departments.'},
                status=status.HTTP_403_FORBIDDEN
            )

        staff_id = request.data.get('staff_id')
        department_id = request.data.get('department_id')
        role = request.data.get('role', 'STAFF').upper()

        if not staff_id or not department_id:
            return Response(
                {'detail': 'Both staff_id and department_id are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if role not in ['STAFF', 'HOD', 'AHOD']:
            return Response(
                {'detail': 'Invalid role. Must be STAFF, HOD, or AHOD.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from .models import StaffProfile, Department, DepartmentRole, AcademicYear

        try:
            staff = StaffProfile.objects.get(id=staff_id)
        except StaffProfile.DoesNotExist:
            return Response(
                {'detail': 'Staff member not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            department = Department.objects.get(id=department_id)
        except Department.DoesNotExist:
            return Response(
                {'detail': 'Department not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Validate department scope for non-superusers
        if not user.is_superuser and not has_ps_role and 'academics.view_all_staff' not in perms:
            from academics.utils import get_user_effective_departments
            
            allowed_dept_ids = get_user_effective_departments(user)
            if allowed_dept_ids:
                if department_id not in allowed_dept_ids:
                    return Response(
                        {'detail': 'You can only assign staff to departments you manage.'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                return Response(
                    {'detail': 'You are not mapped to any departments.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        # Get active academic year for HOD/AHOD assignments
        active_year = AcademicYear.objects.filter(is_active=True).first()
        if not active_year and role in ['HOD', 'AHOD']:
            return Response(
                {'detail': 'No active academic year found. Cannot assign HOD/AHOD roles.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if role == 'STAFF':
            # For STAFF role: update primary department
            # Allow swapping departments - if staff already has a department, update it
            if staff.department and staff.department.id == department_id:
                # Already assigned to this department
                return Response({
                    'detail': f'Staff is already assigned as STAFF to {department.name}.',
                    'staff_id': staff.id,
                    'department_id': department.id,
                    'role': 'STAFF',
                }, status=status.HTTP_200_OK)

            # Assign or swap department
            old_dept_name = staff.department.name if staff.department else None
            staff.department = department
            staff.save(update_fields=['department'])

            if old_dept_name:
                detail_msg = f'Staff department changed from {old_dept_name} to {department.name} successfully.'
            else:
                detail_msg = f'Staff assigned as STAFF to {department.name} successfully.'

            return Response({
                'detail': detail_msg,
                'staff_id': staff.id,
                'department_id': department.id,
                'role': 'STAFF',
            }, status=status.HTTP_200_OK)

        else:
            # For HOD/AHOD roles: create or update DepartmentRole
            dept_role, created = DepartmentRole.objects.get_or_create(
                department=department,
                staff=staff,
                role=role,
                academic_year=active_year,
                defaults={'is_active': True}
            )

            if not created:
                # Update existing role to active
                dept_role.is_active = True
                dept_role.save(update_fields=['is_active'])

            # Auto-assign corresponding user role (HOD or AHOD)
            from accounts.models import Role, UserRole
            try:
                role_obj = Role.objects.get(name=role)
                # Check if user already has this role
                if not UserRole.objects.filter(user=staff.user, role=role_obj).exists():
                    UserRole.objects.create(user=staff.user, role=role_obj)
            except Role.DoesNotExist:
                # Role doesn't exist in system, skip user role assignment
                pass
            except Exception as e:
                # Log but don't fail the department role assignment
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Failed to assign user role {role} to {staff.user.username}: {e}")

            action = 'assigned' if created else 'updated'
            return Response({
                'detail': f'Staff {action} as {role} to {department.name} successfully.',
                'staff_id': staff.id,
                'department_id': department.id,
                'role': role,
                'academic_year': active_year.name if active_year else None,
            }, status=status.HTTP_200_OK)


class StaffDepartmentRoleRemoveView(APIView):
    """Remove a staff member's department role (HOD/AHOD) or primary STAFF assignment.
    
    POST /api/academics/staff-department-role-remove/
    Body: { 
        staff_id: <int>, 
        department_id: <int>,
        role: 'STAFF' | 'HOD' | 'AHOD'
    }
    
    Requires academics.edit_staff permission.
    """
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        user = request.user
        perms = get_user_permissions(user)
        has_ps_role = user.roles.filter(name__iexact='PS').exists()

        # Check permission
        if not (user.is_superuser or has_ps_role or 'academics.edit_staff' in perms):
            return Response(
                {'detail': 'You do not have permission to modify staff department assignments.'},
                status=status.HTTP_403_FORBIDDEN
            )

        staff_id = request.data.get('staff_id')
        department_id = request.data.get('department_id')
        role = request.data.get('role', '').upper()

        if not staff_id or not department_id or not role:
            return Response(
                {'detail': 'staff_id, department_id, and role are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if role not in ['STAFF', 'HOD', 'AHOD']:
            return Response(
                {'detail': 'Invalid role. Must be STAFF, HOD, or AHOD.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from .models import StaffProfile, Department, DepartmentRole, AcademicYear

        try:
            staff = StaffProfile.objects.get(id=staff_id)
        except StaffProfile.DoesNotExist:
            return Response(
                {'detail': 'Staff member not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            department = Department.objects.get(id=department_id)
        except Department.DoesNotExist:
            return Response(
                {'detail': 'Department not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Validate department scope for non-superusers
        if not user.is_superuser and not has_ps_role and 'academics.view_all_staff' not in perms:
            from academics.utils import get_user_effective_departments
            
            allowed_dept_ids = get_user_effective_departments(user)
            if allowed_dept_ids:
                if department_id not in allowed_dept_ids:
                    return Response(
                        {'detail': 'You can only modify staff in departments you manage.'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                return Response(
                    {'detail': 'You are not mapped to any departments.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        if role == 'STAFF':
            # Remove primary department assignment
            if staff.department and staff.department.id == department_id:
                staff.department = None
                staff.save(update_fields=['department'])
                return Response({
                    'detail': f'Staff removed from {department.name} (STAFF role).',
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'detail': 'Staff is not assigned as STAFF to this department.',
                }, status=status.HTTP_400_BAD_REQUEST)

        else:
            # Remove HOD/AHOD role
            active_year = AcademicYear.objects.filter(is_active=True).first()
            if not active_year:
                return Response(
                    {'detail': 'No active academic year found.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            dept_roles = DepartmentRole.objects.filter(
                department=department,
                staff=staff,
                role=role,
                academic_year=active_year,
                is_active=True
            )

            if dept_roles.exists():
                dept_roles.update(is_active=False)

                # Check if staff still has any other active roles of this type (HOD or AHOD)
                # If not, remove the corresponding user role
                other_active_roles = DepartmentRole.objects.filter(
                    staff=staff,
                    role=role,
                    is_active=True
                ).exclude(
                    department=department,
                    academic_year=active_year
                ).exists()

                if not other_active_roles:
                    # No other active roles of this type - remove user role
                    from accounts.models import Role, UserRole
                    try:
                        role_obj = Role.objects.get(name=role)
                        UserRole.objects.filter(user=staff.user, role=role_obj).delete()
                    except Role.DoesNotExist:
                        pass
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(f"Failed to remove user role {role} from {staff.user.username}: {e}")

                return Response({
                    'detail': f'Removed {role} role from {department.name}.',
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'detail': f'Staff does not have an active {role} role in this department.',
                }, status=status.HTTP_400_BAD_REQUEST)


class BulkAssignSecondarySectionView(APIView):
    """Bulk-assign Year-1 students to their core-dept (SECONDARY) sections.

    POST /api/academics/bulk-assign-secondary-section/
    {
      "section_id": <int>,          // target core-dept section (AI&DS A, AI&DS B, …)
      "student_ids": [<int>, ...]   // student PKs to assign
    }

    The endpoint creates a SECONDARY StudentSectionAssignment for each student,
    ending any previous SECONDARY assignment for the same student first.
    The students' PRIMARY (S&H) assignments are untouched.

    Requires the caller to be a HOD/AHOD for the section's department or a superuser.
    """
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        from .models import Section, StudentSectionAssignment, StudentProfile, DepartmentRole
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)

        section_id = request.data.get('section_id')
        student_ids = request.data.get('student_ids', [])

        if not section_id or not student_ids:
            return Response({'detail': 'section_id and student_ids are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            section = Section.objects.select_related(
                'batch__course__department', 'batch__department'
            ).get(pk=int(section_id))
        except Section.DoesNotExist:
            return Response({'detail': 'Section not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Determine the section's department
        batch = getattr(section, 'batch', None)
        course = getattr(batch, 'course', None) if batch else None
        dept = (getattr(course, 'department', None) if course else None) or getattr(batch, 'department', None)

        # Permission check: user must be HOD/AHOD for this dept, or superuser
        if not user.is_superuser:
            if not staff_profile:
                return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
            dept_id = getattr(dept, 'pk', None)
            is_hod = DepartmentRole.objects.filter(
                staff=staff_profile, role__in=['HOD', 'AHOD'], is_active=True,
                department_id=dept_id
            ).exists() if dept_id else False
            if not is_hod:
                return Response({'detail': 'Only the HOD/AHOD of the section department can assign secondary sections.'}, status=status.HTTP_403_FORBIDDEN)

        from django.utils import timezone
        today = timezone.now().date()
        created_count = 0
        updated_count = 0
        errors = []

        for sid in student_ids:
            try:
                student = StudentProfile.objects.get(pk=int(sid))
                # End existing active SECONDARY assignment for this student
                existing = StudentSectionAssignment.objects.filter(
                    student=student,
                    end_date__isnull=True,
                    section_type=StudentSectionAssignment.SECTION_TYPE_SECONDARY,
                ).first()
                if existing:
                    if existing.section_id == section.pk:
                        # Already assigned to this section; skip
                        continue
                    existing.end_date = today
                    existing.save(update_fields=['end_date'])
                    updated_count += 1

                StudentSectionAssignment.objects.create(
                    student=student,
                    section=section,
                    section_type=StudentSectionAssignment.SECTION_TYPE_SECONDARY,
                    start_date=today,
                )
                created_count += 1
            except StudentProfile.DoesNotExist:
                errors.append(f'Student {sid} not found.')
            except Exception as e:
                errors.append(f'Error assigning student {sid}: {str(e)}')

        return Response({
            'created': created_count,
            'updated': updated_count,
            'errors': errors,
        }, status=status.HTTP_200_OK)


class RemoveSecondarySectionView(APIView):
    """Remove a student's SECONDARY section assignment.

    POST /api/academics/remove-secondary-section/
    {
      "student_ids": [<int>, ...],   // optional; if omitted + section_id provided, clears all in section
      "section_id": <int>            // optional filter
    }
    """
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        from .models import StudentSectionAssignment, StudentProfile, DepartmentRole, Section
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)

        student_ids = request.data.get('student_ids', [])
        section_id = request.data.get('section_id')

        if not student_ids and not section_id:
            return Response({'detail': 'Provide student_ids or section_id.'}, status=status.HTTP_400_BAD_REQUEST)

        if not user.is_superuser and not staff_profile:
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        from django.utils import timezone
        today = timezone.now().date()

        qs = StudentSectionAssignment.objects.filter(
            end_date__isnull=True,
            section_type=StudentSectionAssignment.SECTION_TYPE_SECONDARY,
        )
        if student_ids:
            qs = qs.filter(student_id__in=[int(s) for s in student_ids])
        if section_id:
            qs = qs.filter(section_id=int(section_id))

        count = qs.count()
        qs.update(end_date=today)
        return Response({'ended': count})


# ---------------------------------------------------------------------------
# ExtStaffProfile views
# ---------------------------------------------------------------------------

class ExtStaffProfileListCreateView(APIView):
    """
    GET  /api/academics/ext-staff-profiles/   - list all ext staff profiles
    POST /api/academics/ext-staff-profiles/   - create a new ext staff profile
    """
    permission_classes = (IsAuthenticated,)

    def _check_permission(self, request):
        user = request.user
        if user.is_superuser or user.has_perm('academics.view_staffs_page'):
            return
        try:
            if user.roles.filter(name__iexact='IQAC').exists() or user.roles.filter(name__iexact='COE').exists():
                return
        except Exception:
            pass
        raise PermissionDenied('You do not have permission to manage ext staff profiles.')

    def get(self, request):
        self._check_permission(request)
        from .models import ExtStaffProfile
        from .serializers import ExtStaffProfileSerializer
        qs = ExtStaffProfile.objects.select_related('user').order_by('-created_at')
        return Response(ExtStaffProfileSerializer(qs, many=True).data)

    def post(self, request):
        self._check_permission(request)
        from .serializers import ExtStaffProfileSerializer
        serializer = ExtStaffProfileSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ExtStaffProfileDetailView(APIView):
    """
    PATCH  /api/academics/ext-staff-profiles/<pk>/  - partial update
    DELETE /api/academics/ext-staff-profiles/<pk>/  - delete
    """
    permission_classes = (IsAuthenticated,)

    def _check_permission(self, request):
        user = request.user
        if user.is_superuser or user.has_perm('academics.view_staffs_page'):
            return
        try:
            if user.roles.filter(name__iexact='IQAC').exists() or user.roles.filter(name__iexact='COE').exists():
                return
        except Exception:
            pass
        raise PermissionDenied('You do not have permission to manage ext staff profiles.')

    def _get_obj(self, pk):
        from .models import ExtStaffProfile
        return get_object_or_404(ExtStaffProfile, pk=pk)

    def patch(self, request, pk):
        self._check_permission(request)
        from .serializers import ExtStaffProfileSerializer
        obj = self._get_obj(pk)
        serializer = ExtStaffProfileSerializer(obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, pk):
        self._check_permission(request)
        obj = self._get_obj(pk)
        # Also delete the associated user
        user = obj.user
        obj.delete()
        if user:
            user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ExtStaffProfileBulkDeleteView(APIView):
    """
    POST /api/academics/ext-staff-profiles/bulk-delete/
    Body: { "ids": [1, 2, 3] }
    Deletes multiple ExtStaffProfile records at once.
    """
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        current_user = request.user
        if not current_user.is_superuser and not current_user.has_perm('academics.view_staffs_page'):
            # Also allow IQAC and COE roles
            try:
                if not (current_user.roles.filter(name__iexact='IQAC').exists() or current_user.roles.filter(name__iexact='COE').exists()):
                    return Response(
                        {'detail': 'You do not have permission to delete ext staff profiles.'},
                        status=status.HTTP_403_FORBIDDEN,
                    )
            except Exception:
                return Response(
                    {'detail': 'You do not have permission to delete ext staff profiles.'},
                    status=status.HTTP_403_FORBIDDEN,
                )

        from .models import ExtStaffProfile

        ids = request.data.get('ids', [])
        if not ids or not isinstance(ids, list):
            return Response(
                {'detail': 'No IDs provided. Expected {"ids": [1, 2, 3]}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        deleted_count = 0
        errors = []
        for profile_id in ids:
            try:
                profile = ExtStaffProfile.objects.get(pk=profile_id)
                # Also delete the associated user
                profile_user = profile.user
                profile.delete()
                if profile_user:
                    profile_user.delete()
                deleted_count += 1
            except ExtStaffProfile.DoesNotExist:
                errors.append(f'Profile with ID {profile_id} not found.')
            except Exception as exc:
                errors.append(f'Error deleting ID {profile_id}: {exc}')

        return Response({
            'deleted': deleted_count,
            'total': len(ids),
            'errors': errors,
        })


class ExtStaffProfileUsersView(APIView):
    """
    GET /api/academics/ext-staff-profiles/available-users/
    Returns all Users that do NOT yet have an ExtStaffProfile, for the picker.
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        if user.is_superuser or user.has_perm('academics.view_staffs_page'):
            pass  # allowed
        else:
            try:
                if not (user.roles.filter(name__iexact='IQAC').exists() or user.roles.filter(name__iexact='COE').exists()):
                    raise PermissionDenied()
            except Exception:
                raise PermissionDenied()
        from django.contrib.auth import get_user_model
        from .models import ExtStaffProfile
        User = get_user_model()
        already_assigned = ExtStaffProfile.objects.values_list('user_id', flat=True)
        qs = User.objects.exclude(id__in=already_assigned).order_by('username')
        data = [
            {
                'id': u.id,
                'username': u.username,
                'email': u.email,
                'full_name': u.get_full_name() or u.username,
            }
            for u in qs
        ]
        return Response(data)


class ExtStaffProfileBulkImportView(APIView):
    """Bulk import external staff profiles from an uploaded Excel (.xlsx) or CSV file.

    POST /api/academics/ext-staff-profiles/import/

    Expected columns (case-insensitive, spaces/underscores ignored):
        Username, Email, Password, First Name, Last Name, Designation, College Name,
        Department, Mobile, Gender, PhD Status, Total Experience,
        Date of Birth, Notes

    Required: Username, Email, First Name

    Returns:
        { imported: int, total: int, errors: [{ row: int, errors: [str] }] }
    """
    permission_classes = (IsAuthenticated,)

    def _is_allowed(self, user) -> bool:
        if user.is_superuser:
            return True
        if user.has_perm('academics.view_staffs_page'):
            return True
        try:
            if user.roles.filter(name__iexact='IQAC').exists() or user.roles.filter(name__iexact='COE').exists():
                return True
        except Exception:
            pass
        return False

    @staticmethod
    def _col(row: dict, *names: str) -> str:
        """Extract a value from a row dict using any of the given normalised key names."""
        for name in names:
            needle = name.lower().replace(' ', '').replace('_', '').replace('-', '')
            for k, v in row.items():
                if k.lower().replace(' ', '').replace('_', '').replace('-', '') == needle:
                    return str(v).strip() if v is not None else ''
        return ''

    def post(self, request):
        import csv
        import io
        import openpyxl

        user = request.user
        if not self._is_allowed(user):
            return Response(
                {'detail': 'You do not have permission to import external staff profiles.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file_obj.name.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.csv')):
            return Response(
                {'detail': 'Only .xlsx and .csv files are supported.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Parse file ────────────────────────────────────────────────────────
        rows: list[dict] = []
        if filename.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
                ws = wb.active
                headers: list[str] | None = None
                for excel_row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c).strip() if c is not None else '' for c in excel_row]
                        continue
                    row_data = {
                        h: (str(v).strip() if v is not None else '')
                        for h, v in zip(headers, excel_row)
                    }
                    rows.append(row_data)
            except Exception as exc:
                return Response(
                    {'detail': f'Failed to parse Excel file: {exc}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            try:
                content = file_obj.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                for row in reader:
                    rows.append({k.strip(): (v.strip() if v else '') for k, v in row.items()})
            except Exception as exc:
                return Response(
                    {'detail': f'Failed to parse CSV file: {exc}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if not rows:
            return Response(
                {'detail': 'File is empty or has no data rows.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Process rows ──────────────────────────────────────────────────────
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Group
        from .models import ExtStaffProfile

        User = get_user_model()
        errors: list[dict] = []
        imported = 0
        seen_usernames: set[str] = set()
        seen_emails: set[str] = set()

        # Get or create Ext_staff group
        ext_staff_group, _ = Group.objects.get_or_create(name='EXT_STAFF')

        for idx, row in enumerate(rows, start=2):  # row 1 is header
            salutation   = self._col(row, 'salutation', 'title', 'prefix')
            full_name    = self._col(row, 'fullname', 'full name', 'full_name', 'name', 'faculty name', 'name of the faculty')
            username     = self._col(row, 'username', 'user name', 'user_name')
            email        = self._col(row, 'email', 'emailaddress', 'email address', 'email_address')
            password     = self._col(row, 'password', 'pwd')
            first_name   = self._col(row, 'firstname', 'first name', 'first_name')
            last_name    = self._col(row, 'lastname', 'last name', 'last_name')
            designation  = self._col(row, 'designation', 'desig')
            college_name = self._col(row, 'collegename', 'college name', 'college', 'nameofcollege', 'name of college', 'institute')
            department   = self._col(row, 'department', 'dept', 'department working in')
            mobile       = self._col(row, 'mobile', 'mobilenumber', 'mobile number', 'phone', 'contact')
            gender       = self._col(row, 'gender', 'genderspecification', 'gender specification')
            ug_spec      = self._col(row, 'ugspecialization', 'ug specialization', 'ug with specialization', 'ug')
            pg_spec      = self._col(row, 'pgspecialization', 'pg specialization', 'pg with specialization', 'pg')
            phd_status   = self._col(row, 'phdstatus', 'phd status', 'phd')
            experience   = self._col(row, 'totalexperience', 'total experience', 'experience')
            engg_exp     = self._col(row, 'engineeringcollegeexperience', 'engineering college experience', 'engg college experience', 'engg exp')
            dob          = self._col(row, 'dateofbirth', 'date of birth', 'dob')
            notes_text   = self._col(row, 'notes', 'remarks')
            teaching     = self._col(row, 'teaching', 'facultytype', 'type of faculty')
            faculty_id   = self._col(row, 'facultyid', 'faculty id', 'staffid', 'staff id')
            # Bank details
            acc_holder   = self._col(row, 'accountholdername', 'account holder name', 'holder name')
            acc_number   = self._col(row, 'accountnumber', 'account number', 'account no').lstrip("'")  # Remove leading apostrophe if present
            bank_name    = self._col(row, 'bankname', 'bank name', 'name of bank')
            branch_name  = self._col(row, 'bankbranchname', 'bank branch name', 'branch name', 'branch')
            ifsc         = self._col(row, 'ifsccode', 'ifsc code', 'ifsc')

            # If no first_name but full_name provided, use full_name as first_name
            if not first_name and full_name:
                first_name = full_name

            # Auto-generate username if not provided: salutation + space + full_name
            if not username:
                name_part = full_name or first_name or ''
                if salutation and name_part:
                    username = f"{salutation} {name_part}"
                elif name_part:
                    username = name_part.replace(' ', '_').lower()

            # Only keep non-structured notes in the notes field
            full_notes = notes_text or ''

            # Parse date_of_birth
            parsed_dob = None
            if dob:
                from datetime import datetime
                for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y'):
                    try:
                        parsed_dob = datetime.strptime(dob, fmt).date()
                        break
                    except ValueError:
                        continue

            # ── Validate required fields ──
            row_errors: list[str] = []
            if not username:
                row_errors.append('Username could not be generated. Provide Salutation and Full Name, or a Username.')
            if not email:
                row_errors.append('Email is required.')
            if not first_name:
                row_errors.append('Full Name (or First Name) is required.')

            if username and username.lower() in seen_usernames:
                row_errors.append(f'Duplicate username "{username}" in uploaded file.')
            if email and email.lower() in seen_emails:
                row_errors.append(f'Duplicate email "{email}" in uploaded file.')

            if row_errors:
                errors.append({'row': idx, 'errors': row_errors})
                continue

            seen_usernames.add(username.lower())
            seen_emails.add(email.lower())

            # ── Check for existing user by email - UPDATE instead of error ──
            existing_user = User.objects.filter(email__iexact=email).first()
            
            if existing_user:
                # Update existing user and their ExtStaffProfile
                try:
                    existing_user.first_name = first_name
                    existing_user.last_name = last_name or ''
                    if password:
                        existing_user.set_password(password)
                    existing_user.save()
                    
                    # Add to Ext_staff group if not already
                    existing_user.groups.add(ext_staff_group)
                    
                    # Update or create ExtStaffProfile
                    profile, created = ExtStaffProfile.objects.update_or_create(
                        user=existing_user,
                        defaults={
                            'salutation': salutation,
                            'designation': designation,
                            'college_name': college_name,
                            'teaching': teaching,
                            'faculty_id': faculty_id,
                            'department': department,
                            'mobile': mobile,
                            'gender': gender,
                            'ug_specialization': ug_spec,
                            'pg_specialization': pg_spec,
                            'phd_status': phd_status,
                            'total_experience': experience,
                            'engg_college_experience': engg_exp,
                            'date_of_birth': parsed_dob,
                            'account_holder_name': acc_holder,
                            'account_number': acc_number,
                            'bank_name': bank_name,
                            'bank_branch_name': branch_name,
                            'ifsc_code': ifsc,
                            'notes': full_notes,
                            'is_active': True,
                        }
                    )
                    imported += 1
                except Exception as exc:
                    errors.append({'row': idx, 'errors': [f'Failed to update: {exc}']})
                continue

            # ── Check if username is taken by another user ──
            if User.objects.filter(username__iexact=username).exists():
                errors.append({'row': idx, 'errors': [f'Username "{username}" already exists in the system.']})
                continue

            # ── Create User and ExtStaffProfile ──
            try:
                new_user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password or 'changeme123',
                    first_name=first_name,
                    last_name=last_name or '',
                )
                # Add user to Ext_staff group
                new_user.groups.add(ext_staff_group)

                ExtStaffProfile.objects.create(
                    user=new_user,
                    salutation=salutation,
                    designation=designation,
                    college_name=college_name,
                    teaching=teaching,
                    faculty_id=faculty_id,
                    department=department,
                    mobile=mobile,
                    gender=gender,
                    ug_specialization=ug_spec,
                    pg_specialization=pg_spec,
                    phd_status=phd_status,
                    total_experience=experience,
                    engg_college_experience=engg_exp,
                    date_of_birth=parsed_dob,
                    account_holder_name=acc_holder,
                    account_number=acc_number,
                    bank_name=bank_name,
                    bank_branch_name=branch_name,
                    ifsc_code=ifsc,
                    notes=full_notes,
                    is_active=True,
                )
                imported += 1
            except Exception as exc:
                errors.append({'row': idx, 'errors': [str(exc)]})

        return Response({
            'imported': imported,
            'total': len(rows),
            'errors': errors,
        })


# ---------------------------------------------------------------------------
# ExtStaffFormSettings views - Registration form management
# ---------------------------------------------------------------------------

class ExtStaffFormSettingsView(APIView):
    """
    GET  /api/academics/ext-staff-form/settings/  - Get form settings
    PUT  /api/academics/ext-staff-form/settings/  - Update form settings
    """
    permission_classes = (IsAuthenticated,)

    def _check_permission(self, request):
        user = request.user
        if user.is_superuser or user.has_perm('academics.view_staffs_page'):
            return
        try:
            if user.roles.filter(name__iexact='IQAC').exists() or user.roles.filter(name__iexact='COE').exists():
                return
        except Exception:
            pass
        raise PermissionDenied('You do not have permission to manage ext staff form settings.')

    def get(self, request):
        self._check_permission(request)
        from .models import ExtStaffFormSettings
        settings_obj = ExtStaffFormSettings.get_or_create_settings()
        
        # Build the share URL
        share_url = request.build_absolute_uri(f'/ext-register/{settings_obj.form_code}/')
        
        return Response({
            'form_code': settings_obj.form_code,
            'form_title': settings_obj.form_title,
            'form_description': settings_obj.form_description,
            'is_accepting_responses': settings_obj.is_accepting_responses,
            'field_config': settings_obj.field_config,
            'share_url': share_url,
            'available_fields': ExtStaffFormSettings.get_available_fields(),
            'updated_at': settings_obj.updated_at,
        })

    def put(self, request):
        self._check_permission(request)
        from .models import ExtStaffFormSettings
        settings_obj = ExtStaffFormSettings.get_or_create_settings()
        
        data = request.data
        if 'form_title' in data:
            settings_obj.form_title = data['form_title']
        if 'form_description' in data:
            settings_obj.form_description = data['form_description']
        if 'is_accepting_responses' in data:
            settings_obj.is_accepting_responses = data['is_accepting_responses']
        if 'field_config' in data:
            settings_obj.field_config = data['field_config']
        
        settings_obj.updated_by = request.user
        settings_obj.save()
        
        share_url = request.build_absolute_uri(f'/ext-register/{settings_obj.form_code}/')
        
        return Response({
            'form_code': settings_obj.form_code,
            'form_title': settings_obj.form_title,
            'form_description': settings_obj.form_description,
            'is_accepting_responses': settings_obj.is_accepting_responses,
            'field_config': settings_obj.field_config,
            'share_url': share_url,
            'updated_at': settings_obj.updated_at,
        })


class ExtStaffPublicFormView(APIView):
    """
    GET  /api/academics/ext-staff-form/public/<form_code>/  - Get public form config
    POST /api/academics/ext-staff-form/public/<form_code>/  - Submit registration
    """
    permission_classes = []  # Public endpoint
    authentication_classes = []  # No auth required

    def get(self, request, form_code):
        from .models import ExtStaffFormSettings
        try:
            settings_obj = ExtStaffFormSettings.objects.get(form_code=form_code)
        except ExtStaffFormSettings.DoesNotExist:
            return Response({'detail': 'Form not found.'}, status=status.HTTP_404_NOT_FOUND)
        
        if not settings_obj.is_accepting_responses:
            return Response({
                'form_code': form_code,
                'form_title': settings_obj.form_title,
                'is_accepting_responses': False,
                'message': 'This form is currently not accepting responses.',
            })
        
        # Return only enabled fields
        enabled_fields = [f for f in settings_obj.field_config if f.get('enabled', False)]
        enabled_fields.sort(key=lambda x: x.get('order', 999))
        
        return Response({
            'form_code': form_code,
            'form_title': settings_obj.form_title,
            'form_description': settings_obj.form_description,
            'is_accepting_responses': True,
            'fields': enabled_fields,
        })

    def post(self, request, form_code):
        from .models import ExtStaffFormSettings, ExtStaffProfile
        from django.contrib.auth import get_user_model
        
        User = get_user_model()
        
        try:
            settings_obj = ExtStaffFormSettings.objects.get(form_code=form_code)
        except ExtStaffFormSettings.DoesNotExist:
            return Response({'detail': 'Form not found.'}, status=status.HTTP_404_NOT_FOUND)
        
        if not settings_obj.is_accepting_responses:
            return Response(
                {'detail': 'This form is not accepting responses.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        data = request.data
        errors = {}
        
        # Get user_id from signup step
        user_id = data.get('user_id')
        if not user_id:
            return Response(
                {'detail': 'User ID is required. Please complete signup first.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            return Response(
                {'detail': 'Invalid user. Please signup again.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get enabled fields for validation
        enabled_fields = [f for f in settings_obj.field_config if f.get('enabled', False)]
        
        # Validate required fields (exclude email since already captured in signup)
        for field in enabled_fields:
            field_name = field['field']
            if field_name == 'email':  # Email already captured
                continue
            if field.get('required', False):
                value = data.get(field_name, '')
                if not value or (isinstance(value, str) and not value.strip()):
                    errors[field_name] = f"{field.get('label', field_name)} is required."
        
        # Validate full_name
        full_name = data.get('full_name', '').strip()
        if not full_name:
            errors['full_name'] = 'Full Name is required.'
        
        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)
        
        # Update user's name
        salutation = data.get('salutation', '').strip()
        if salutation and full_name:
            username = f"{salutation} {full_name}"
        else:
            username = full_name
        
        # Parse date_of_birth
        dob = data.get('date_of_birth', '')
        parsed_dob = None
        if dob:
            from datetime import datetime
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
                try:
                    parsed_dob = datetime.strptime(dob, fmt).date()
                    break
                except ValueError:
                    continue
        
        try:
            # Update user info
            user.username = username
            user.first_name = full_name
            user.save()
            
            # Handle file upload for passbook_proof
            passbook_file = request.FILES.get('passbook_proof')
            
            # Get or create ExtStaffProfile
            profile, _ = ExtStaffProfile.objects.get_or_create(user=user)
            
            # Update profile fields
            profile.salutation = salutation
            profile.designation = data.get('designation', '')
            profile.college_name = data.get('college_name', '')
            profile.teaching = data.get('teaching', '')
            profile.faculty_id = data.get('faculty_id', '')
            profile.department = data.get('department', '')
            profile.mobile = data.get('mobile', '')
            profile.gender = data.get('gender', '')
            profile.ug_specialization = data.get('ug_specialization', '')
            profile.pg_specialization = data.get('pg_specialization', '')
            profile.phd_status = data.get('phd_status', '')
            profile.total_experience = data.get('total_experience', '')
            profile.engg_college_experience = data.get('engg_college_experience', '')
            profile.date_of_birth = parsed_dob
            profile.account_holder_name = data.get('account_holder_name', '')
            profile.account_number = data.get('account_number', '')
            profile.bank_name = data.get('bank_name', '')
            profile.bank_branch_name = data.get('bank_branch_name', '')
            profile.ifsc_code = data.get('ifsc_code', '')
            profile.is_active = True
            
            if passbook_file:
                profile.passbook_proof = passbook_file
            
            profile.save()
            
            return Response({
                'success': True,
                'message': 'Registration successful! Your profile has been saved.',
                'ext_uid': profile.external_id,
                'username': username,
                'email': user.email,
            }, status=status.HTTP_201_CREATED)
            
        except Exception as exc:
            return Response(
                {'detail': f'Registration failed: {str(exc)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ExtStaffCheckEmailView(APIView):
    """
    POST /api/academics/ext-staff-form/public/<form_code>/check-email/
    Check if email already exists in the system.
    """
    permission_classes = []
    authentication_classes = []

    def post(self, request, form_code):
        from .models import ExtStaffFormSettings
        from django.contrib.auth import get_user_model
        
        User = get_user_model()
        
        # Verify form exists and is accepting responses
        try:
            settings_obj = ExtStaffFormSettings.objects.get(form_code=form_code)
        except ExtStaffFormSettings.DoesNotExist:
            return Response({'detail': 'Form not found.'}, status=status.HTTP_404_NOT_FOUND)
        
        if not settings_obj.is_accepting_responses:
            return Response(
                {'detail': 'This form is not accepting responses.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        email = request.data.get('email', '').strip().lower()
        if not email:
            return Response({'error': 'Email is required.'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if email exists
        exists = User.objects.filter(email__iexact=email).exists()
        
        return Response({
            'email': email,
            'exists': exists,
            'message': 'Email already registered. Please login to update your profile.' if exists else 'Email available for registration.',
        })


class ExtStaffSignupView(APIView):
    """
    POST /api/academics/ext-staff-form/public/<form_code>/signup/
    Create a new user with email and password, or without email (username only).
    """
    permission_classes = []
    authentication_classes = []

    def post(self, request, form_code):
        from .models import ExtStaffFormSettings, ExtStaffProfile
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Group
        import uuid
        
        User = get_user_model()
        
        # Verify form exists and is accepting responses
        try:
            settings_obj = ExtStaffFormSettings.objects.get(form_code=form_code)
        except ExtStaffFormSettings.DoesNotExist:
            return Response({'detail': 'Form not found.'}, status=status.HTTP_404_NOT_FOUND)
        
        if not settings_obj.is_accepting_responses:
            return Response(
                {'detail': 'This form is not accepting responses.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        email = request.data.get('email', '').strip().lower()
        password = request.data.get('password', '')
        confirm_password = request.data.get('confirm_password', '')
        full_name = request.data.get('full_name', '').strip()
        skip_email = request.data.get('skip_email', False)
        
        errors = {}
        
        if skip_email:
            # Signup without email - require full_name
            if not full_name:
                errors['full_name'] = 'Full name is required.'
        else:
            # Signup with email
            if not email:
                errors['email'] = 'Email is required.'
            elif User.objects.filter(email__iexact=email).exists():
                errors['email'] = 'Email already registered.'
        
        # Validate password
        if not password:
            errors['password'] = 'Password is required.'
        elif len(password) < 6:
            errors['password'] = 'Password must be at least 6 characters.'
        
        # Validate confirm password
        if password != confirm_password:
            errors['confirm_password'] = 'Passwords do not match.'
        
        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            if skip_email:
                # Generate a unique username from full_name
                base_username = ''.join(c.lower() for c in full_name if c.isalnum())[:20]
                if not base_username:
                    base_username = 'ext'
                username = base_username
                suffix = 1
                while User.objects.filter(username=username).exists():
                    username = f"{base_username}{suffix}"
                    suffix += 1
                
                # Create user without email
                new_user = User.objects.create_user(
                    username=username,
                    email='',  # No email
                    password=password,
                    first_name=full_name.split()[0] if full_name else '',
                    last_name=' '.join(full_name.split()[1:]) if len(full_name.split()) > 1 else '',
                )
            else:
                # Create user with email as username
                new_user = User.objects.create_user(
                    username=email,
                    email=email,
                    password=password,
                )
            
            # Add to Ext_staff group
            ext_staff_group, _ = Group.objects.get_or_create(name='EXT_STAFF')
            new_user.groups.add(ext_staff_group)
            
            # Create empty ExtStaffProfile
            profile = ExtStaffProfile.objects.create(
                user=new_user,
                is_active=True,
            )
            
            return Response({
                'success': True,
                'message': 'Account created successfully. Please complete your profile.',
                'user_id': new_user.id,
                'username': new_user.username,
                'email': email or None,
                'ext_uid': profile.external_id,
            }, status=status.HTTP_201_CREATED)
            
        except Exception as exc:
            return Response(
                {'detail': f'Signup failed: {str(exc)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
