import csv
import io
import re
import zipfile

from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from rest_framework import status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.utils import get_user_permissions

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

User = get_user_model()

STUDENT_STATUS_CHOICES = ['ACTIVE', 'INACTIVE', 'ALUMNI', 'DEBAR']
_EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')


def _has_import_permission(user, perms):
    return user.is_superuser or user.is_staff or 'students.view_all_students' in perms


class StudentImportTemplateDownloadView(APIView):
    """Download an Excel (or CSV fallback) template for bulk student import."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        perms = get_user_permissions(request.user)
        if not _has_import_permission(request.user, perms):
            return Response(
                {'error': 'You do not have permission to download the student import template.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        if not EXCEL_SUPPORT:
            # Fallback: return a CSV template
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['student_reg_no', 'name', 'department', 'section', 'email', 'batch', 'status', 'core_department'])
            writer.writerow(['REG2024001', 'John Doe', 'AI&DS', 'A', 'john.doe@example.com', '2024', 'ACTIVE', ''])
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="students_import_template.csv"'
            return response

        try:
            from io import BytesIO
            from academics.models import Department

            departments = list(
                Department.objects.order_by('short_name')
                .values_list('short_name', flat=True)
            )
            departments = [d for d in departments if d]

            wb = Workbook()
            ws = wb.active
            ws.title = 'Students Import'

            headers = ['student_reg_no', 'name', 'department', 'section', 'email', 'batch', 'status', 'core_department']
            ws.append(headers)

            sample_dept = departments[0] if departments else 'AI&DS'
            ws.append(['REG2024001', 'John Doe', sample_dept, 'A', 'john.doe@example.com', '2024', 'ACTIVE', ''])

            # Build a hidden "_Lists" sheet for dropdown reference data.
            # Cross-sheet references avoid the 255-char inline list limit.
            lists_ws = wb.create_sheet(title='_Lists')
            lists_ws.sheet_state = 'hidden'
            for i, dept in enumerate(departments, start=1):
                lists_ws.cell(row=i, column=1, value=dept)
            for i, s in enumerate(STUDENT_STATUS_CHOICES, start=1):
                lists_ws.cell(row=i, column=2, value=s)

            if departments:
                dept_ref = f"_Lists!$A$1:$A${len(departments)}"
                dv_dept = DataValidation(type='list', formula1=dept_ref, allow_blank=True)
                ws.add_data_validation(dv_dept)
                dv_dept.add('C2:C1000')

                dv_core_dept = DataValidation(type='list', formula1=dept_ref, allow_blank=True)
                ws.add_data_validation(dv_core_dept)
                dv_core_dept.add('H2:H1000')

            status_ref = f"_Lists!$B$1:$B${len(STUDENT_STATUS_CHOICES)}"
            dv_status = DataValidation(type='list', formula1=status_ref, allow_blank=False)
            ws.add_data_validation(dv_status)
            dv_status.add('G2:G1000')

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 32
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 20

            # Save to buffer first
            raw_buf = BytesIO()
            wb.save(raw_buf)
            raw_buf.seek(0)

            # Post-process: strip <workbookProtection/> which WPS rejects
            cleaned_buf = BytesIO()
            with zipfile.ZipFile(raw_buf, 'r') as zin:
                with zipfile.ZipFile(cleaned_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        data = zin.read(item.filename)
                        if item.filename == 'xl/workbook.xml':
                            data = data.replace(b'<workbookProtection/>', b'')
                        zout.writestr(item, data)

            xlsx_bytes = cleaned_buf.getvalue()
            response = HttpResponse(
                xlsx_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="students_import_template.xlsx"'
            response['Content-Length'] = str(len(xlsx_bytes))
            return response

        except Exception as exc:
            return Response(
                {'error': f'Failed to generate template: {exc}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class StudentBulkImportView(APIView):
    """Bulk-import students from a CSV or Excel file.

    Existing students (matched by ``student_reg_no``) are updated; new ones are
    created.  A default *unusable* password is set so the account cannot be used
    until an admin assigns a password.
    """
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        perms = get_user_permissions(request.user)
        if not _has_import_permission(request.user, perms):
            return Response(
                {'error': 'You do not have permission to import students.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        filename = uploaded_file.name.lower()
        is_excel = filename.endswith(('.xlsx', '.xls'))
        is_csv = filename.endswith('.csv')

        if not (is_csv or is_excel):
            return Response(
                {'error': 'File must be CSV (.csv) or Excel (.xlsx).'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # -- Parse file --
        rows = []  # list of (row_number, dict)

        if is_excel:
            if not EXCEL_SUPPORT:
                return Response(
                    {'error': 'Excel support unavailable on this server. Please use CSV.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                wb = load_workbook(uploaded_file, read_only=True, data_only=True)
                ws = wb.active
                raw_headers = [cell.value for cell in ws[1]]
                headers = [str(h).strip() if h is not None else '' for h in raw_headers]
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    row_dict = {
                        headers[i]: (str(v).strip() if v is not None else '')
                        for i, v in enumerate(row)
                        if i < len(headers) and headers[i]
                    }
                    if any(row_dict.values()):
                        rows.append((row_idx, row_dict))
            except Exception as exc:
                return Response(
                    {'error': f'Failed to parse Excel file: {exc}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            try:
                decoded = uploaded_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = [
                    (idx, {k: (v or '').strip() for k, v in row.items()})
                    for idx, row in enumerate(reader, start=2)
                ]
            except Exception as exc:
                return Response(
                    {'error': f'Failed to parse CSV file: {exc}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if not rows:
            return Response(
                {'error': 'The uploaded file is empty or has no data rows.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # -- Import --
        from academics.models import Department, Section, StudentProfile, StudentSectionAssignment

        created_count = 0
        updated_count = 0
        errors = []

        # Pre-fetch departments once
        dept_map = {
            d.short_name.upper(): d
            for d in Department.objects.all()
            if d.short_name
        }

        with transaction.atomic():
            for row_idx, row in rows:
                try:
                    reg_no = row.get('student_reg_no', '').strip()
                    name = row.get('name', '').strip()
                    dept_name = row.get('department', '').strip()
                    section_name = row.get('section', '').strip()
                    email = row.get('email', '').strip()
                    batch_year = row.get('batch', '').strip()
                    row_status = row.get('status', 'ACTIVE').strip().upper()
                    core_dept_name = row.get('core_department', '').strip()

                    # Required field
                    if not reg_no:
                        errors.append(f'Row {row_idx}: student_reg_no is required.')
                        continue

                    # Status validation
                    if row_status not in STUDENT_STATUS_CHOICES:
                        errors.append(
                            f'Row {row_idx}: Invalid status "{row_status}". '
                            f'Must be one of: {", ".join(STUDENT_STATUS_CHOICES)}'
                        )
                        continue

                    # Email format validation
                    if email and not _EMAIL_RE.match(email):
                        errors.append(f'Row {row_idx}: Invalid email format "{email}".')
                        continue

                    # Resolve department
                    department = None
                    if dept_name:
                        department = dept_map.get(dept_name.upper())
                        if not department:
                            errors.append(
                                f'Row {row_idx}: Department "{dept_name}" not found.'
                            )
                            continue

                    # Resolve core_department (optional – for Year 1 S&H students whose
                    # home_department differs from their section's department)
                    core_department = None
                    if core_dept_name:
                        core_department = dept_map.get(core_dept_name.upper())
                        if not core_department:
                            errors.append(
                                f'Row {row_idx}: core_department "{core_dept_name}" not found.'
                            )
                            continue

                    # Resolve section
                    section = None
                    if section_name and department:
                        qs = Section.objects.filter(
                            name__iexact=section_name,
                        ).filter(
                            Q(batch__course__department=department)
                            | Q(batch__department=department)
                        )
                        if batch_year:
                            # Prefer exact start_year match when batch_year is numeric
                            if batch_year.isdigit():
                                by_year = qs.filter(batch__start_year=int(batch_year))
                                if by_year.exists():
                                    qs = by_year
                                else:
                                    # fall back to name contains
                                    qs = qs.filter(
                                        Q(batch__name__icontains=batch_year)
                                        | Q(batch__batch_year__name__icontains=batch_year)
                                    )
                            else:
                                qs = qs.filter(
                                    Q(batch__name__icontains=batch_year)
                                    | Q(batch__batch_year__name__icontains=batch_year)
                                )
                        section = qs.order_by('-batch__start_year').first()

                        if not section:
                            errors.append(
                                f'Row {row_idx}: Section "{section_name}" not found '
                                f'for department "{dept_name}"'
                                + (f' and batch "{batch_year}"' if batch_year else '') + '.'
                            )
                            continue

                    # -- Update existing student --
                    existing = StudentProfile.objects.filter(reg_no=reg_no).first()
                    if existing:
                        existing.status = row_status
                        # Use core_department for home_department if provided; otherwise use section's department
                        if core_department:
                            existing.home_department = core_department
                        elif department:
                            existing.home_department = department
                        if section:
                            existing.section = section
                        if batch_year:
                            existing.batch = batch_year
                        existing.save(update_fields=['status', 'home_department', 'section', 'batch'])

                        # Create/update active section assignment so student appears in list
                        if section:
                            active = StudentSectionAssignment.objects.filter(
                                student=existing, end_date__isnull=True
                            ).first()
                            if not active or active.section_id != section.id:
                                # save() auto-closes previous active assignment
                                StudentSectionAssignment.objects.create(
                                    student=existing,
                                    section=section,
                                )

                        if existing.user:
                            user_obj = existing.user
                            changed = False
                            if email and user_obj.email != email:
                                user_obj.email = email
                                changed = True
                            if name:
                                parts = name.split()
                                user_obj.first_name = parts[0]
                                user_obj.last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''
                                changed = True
                            if changed:
                                user_obj.save(update_fields=['email', 'first_name', 'last_name'])
                            # Ensure STUDENT role is assigned
                            try:
                                from accounts.models import Role, UserRole
                                student_role = Role.objects.filter(name='STUDENT').first()
                                if student_role:
                                    UserRole.objects.get_or_create(user=user_obj, role=student_role)
                            except Exception:
                                pass
                        updated_count += 1

                    # -- Create new student --
                    else:
                        # Name is required to create a meaningful student account
                        if not name:
                            errors.append(
                                f'Row {row_idx}: name is required for new student (reg_no: "{reg_no}").'
                            )
                            continue

                        # Use the student's name as username (matching existing student convention)
                        base_username = name
                        username = base_username
                        counter = 1
                        while User.objects.filter(username=username).exists():
                            username = f'{base_username}{counter}'
                            counter += 1

                        parts = name.split()
                        first_name = parts[0]
                        last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''

                        new_user = User(
                            username=username,
                            email=email,
                            first_name=first_name,
                            last_name=last_name,
                        )
                        new_user.set_unusable_password()
                        new_user.save()

                        new_profile = StudentProfile.objects.create(
                            user=new_user,
                            reg_no=reg_no,
                            status=row_status,
                            home_department=core_department if core_department else department,
                            section=section,
                            batch=batch_year,
                        )

                        # Create active section assignment so student appears in list
                        if section:
                            StudentSectionAssignment.objects.create(
                                student=new_profile,
                                section=section,
                            )

                        # Assign STUDENT role so the account has proper access
                        try:
                            from accounts.models import Role, UserRole
                            student_role = Role.objects.filter(name='STUDENT').first()
                            if student_role:
                                UserRole.objects.get_or_create(user=new_user, role=student_role)
                        except Exception:
                            pass

                        created_count += 1

                except Exception as exc:
                    errors.append(f'Row {row_idx}: Unexpected error - {exc}')

        return Response({
            'message': 'Import completed.',
            'created': created_count,
            'updated': updated_count,
            'errors': errors[:50],
            'total_errors': len(errors),
        })
