import logging
import datetime
from io import BytesIO
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.http import HttpResponse

logger = logging.getLogger(__name__)
from django.db.models import Count, Q, Avg, F
from django.utils import timezone
from datetime import date, timedelta
from .models import (
    PeriodAttendanceRecord, 
    PeriodAttendanceSession,
    StudentProfile, 
    Section, 
    Department,
    StaffProfile,
    DepartmentRole
)
from accounts.utils import get_user_permissions


class AttendanceAnalyticsView(APIView):
    """
    Attendance analytics with three permission levels:
    1. 'analytics.view_all_analytics' - View all departments, classes, students
    2. 'analytics.view_department_analytics' - View own department only
    3. 'analytics.view_class_analytics' - View own class only (advisors)
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        staff_profile = getattr(user, 'staff_profile', None)
        
        # Determine permission level
        can_view_all = 'analytics.view_all_analytics' in perms or user.is_superuser
        can_view_department = 'analytics.view_department_analytics' in perms or can_view_all
        can_view_class = 'analytics.view_class_analytics' in perms or can_view_department
        
        if not (can_view_all or can_view_department or can_view_class):
            raise PermissionDenied('You do not have permission to view analytics')
        
        # Get query parameters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        department_id = request.query_params.get('department_id')
        section_id = request.query_params.get('section_id')
        view_type = request.query_params.get('view_type', 'overview')  # overview, department, class, student
        
        # Set default date range (last 30 days)
        if not end_date:
            end_date = date.today()
        else:
            try:
                end_date = date.fromisoformat(end_date)
            except:
                end_date = date.today()
        
        if not start_date:
            start_date = end_date - timedelta(days=30)
        else:
            try:
                start_date = date.fromisoformat(start_date)
            except:
                start_date = end_date - timedelta(days=30)
        
        # Build base queryset
        sessions_qs = PeriodAttendanceSession.objects.filter(
            date__gte=start_date,
            date__lte=end_date
        ).select_related('section', 'section__batch', 'section__batch__course', 'section__batch__course__department')
        
        records_qs = PeriodAttendanceRecord.objects.filter(
            session__date__gte=start_date,
            session__date__lte=end_date
        ).select_related('student', 'session__section', 'session__section__batch', 'session__section__batch__course', 'session__section__batch__course__department')
        
        # Apply permission-based filtering
        # Note: do not force only Period 1 here — allow analytics across periods
        # If a caller needs period-specific data they can pass a `period_index` query param
        period_index = request.query_params.get('period_index')
        try:
            if period_index:
                sessions_qs = sessions_qs.filter(period__index=int(period_index))
                records_qs = records_qs.filter(session__period__index=int(period_index))
        except Exception:
            # ignore any parse/filter errors and continue without period filtering
            pass
        if not can_view_all:
            if can_view_department and staff_profile:
                # Filter by staff's department
                dept_roles = DepartmentRole.objects.filter(staff=staff_profile, is_active=True)
                dept_ids = [dr.department_id for dr in dept_roles if dr.department_id]
                
                # Also check if staff has teaching assignments in other departments
                from .models import TeachingAssignment
                teaching_depts = TeachingAssignment.objects.filter(
                    staff=staff_profile, 
                    is_active=True
                ).values_list('section__batch__course__department_id', flat=True).distinct()
                dept_ids.extend(list(teaching_depts))
                dept_ids = list(set(filter(None, dept_ids)))
                
                if dept_ids:
                    sessions_qs = sessions_qs.filter(section__batch__course__department_id__in=dept_ids)
                    records_qs = records_qs.filter(session__section__batch__course__department_id__in=dept_ids)
                else:
                    # No department access
                    sessions_qs = sessions_qs.none()
                    records_qs = records_qs.none()
            
            elif can_view_class and staff_profile:
                # Filter by advisor's sections only
                from .models import SectionAdvisor
                advisor_sections = SectionAdvisor.objects.filter(
                    advisor=staff_profile,
                    is_active=True
                ).values_list('section_id', flat=True)
                
                if advisor_sections:
                    sessions_qs = sessions_qs.filter(section_id__in=advisor_sections)
                    records_qs = records_qs.filter(session__section_id__in=advisor_sections)
                else:
                    sessions_qs = sessions_qs.none()
                    records_qs = records_qs.none()
        
        # Apply user-selected filters (if permitted)
        if department_id and can_view_all:
            sessions_qs = sessions_qs.filter(section__batch__course__department_id=department_id)
            records_qs = records_qs.filter(session__section__batch__course__department_id=department_id)
        
        if section_id and (can_view_all or can_view_department):
            sessions_qs = sessions_qs.filter(section_id=section_id)
            records_qs = records_qs.filter(session__section_id=section_id)
        
        # Calculate statistics based on view_type
        if view_type == 'overview':
            data = self._get_overview_stats(records_qs, sessions_qs, start_date, end_date)
        elif view_type == 'department':
            data = self._get_department_stats(records_qs, can_view_all)
        elif view_type == 'class':
            data = self._get_class_stats(records_qs, can_view_all or can_view_department)
        elif view_type == 'student':
            data = self._get_student_stats(records_qs, section_id)
        else:
            data = self._get_overview_stats(records_qs, sessions_qs, start_date, end_date)
        
        return Response({
            'permission_level': 'all' if can_view_all else ('department' if can_view_department else 'class'),
            'date_range': {'start': start_date.isoformat(), 'end': end_date.isoformat()},
            'data': data
        })
    
    def _get_overview_stats(self, records_qs, sessions_qs, start_date, end_date):
        """Overall statistics summary"""
        total_sessions = sessions_qs.count()
        total_records = records_qs.count()
        
        status_breakdown = records_qs.values('status').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Treat 'P', 'OD' and 'LATE' as present for attendance percentage
        present_count = records_qs.filter(status__in=['P', 'OD', 'LATE']).count()
        absent_count = records_qs.filter(status='A').count()
        
        attendance_rate = (present_count / total_records * 100) if total_records > 0 else 0
        
        # Trend over time (daily) - use simple date grouping
        daily_stats = records_qs.values('session__date').annotate(
            total=Count('id'),
            present=Count('id', filter=Q(status__in=['P', 'OD', 'LATE'])),
            absent=Count('id', filter=Q(status='A'))
        ).order_by('session__date')
        
        # Format dates as strings for JSON serialization
        daily_trend = []
        for stat in daily_stats:
            daily_trend.append({
                'day': stat['session__date'].isoformat() if stat['session__date'] else None,
                'total': stat['total'],
                'present': stat['present'],
                'absent': stat['absent']
            })
        
        return {
            'summary': {
                'total_sessions': total_sessions,
                'total_records': total_records,
                'present_count': present_count,
                'absent_count': absent_count,
                'attendance_rate': round(attendance_rate, 2)
            },
            'status_breakdown': list(status_breakdown),
            'daily_trend': daily_trend
        }
    
    def _get_department_stats(self, records_qs, can_filter=True):
        """Department-wise statistics"""
        dept_stats = records_qs.values(
            'session__section__batch__course__department__id',
            'session__section__batch__course__department__name',
            'session__section__batch__course__department__short_name'
        ).annotate(
            total_records=Count('id'),
            present_count=Count('id', filter=Q(status__in=['P', 'OD', 'LATE'])),
            absent_count=Count('id', filter=Q(status='A')),
            leave_count=Count('id', filter=Q(status='LEAVE')),
            od_count=Count('id', filter=Q(status='OD'))
        ).order_by('-total_records')
        
        # Calculate attendance rate for each department
        result = []
        for dept in dept_stats:
            rate = (dept['present_count'] / dept['total_records'] * 100) if dept['total_records'] > 0 else 0
            result.append({
                'department_id': dept['session__section__batch__course__department__id'],
                'department_name': dept['session__section__batch__course__department__name'],
                'department_short': dept['session__section__batch__course__department__short_name'],
                'total_records': dept['total_records'],
                'present': dept['present_count'],
                'absent': dept['absent_count'],
                'leave': dept['leave_count'],
                'on_duty': dept['od_count'],
                'attendance_rate': round(rate, 2)
            })
        
        return {'departments': result}
    
    def _get_class_stats(self, records_qs, can_filter=True):
        """Class/Section-wise statistics"""
        class_stats = records_qs.values(
            'session__section__id',
            'session__section__name',
            'session__section__batch__course__name',
            'session__section__batch__course__department__short_name'
        ).annotate(
            total_records=Count('id'),
            present_count=Count('id', filter=Q(status__in=['P', 'OD', 'LATE'])),
            absent_count=Count('id', filter=Q(status='A')),
            leave_count=Count('id', filter=Q(status='LEAVE')),
            od_count=Count('id', filter=Q(status='OD')),
            late_count=Count('id', filter=Q(status='LATE'))
        ).order_by('-total_records')
        
        result = []
        for cls in class_stats:
            rate = (cls['present_count'] / cls['total_records'] * 100) if cls['total_records'] > 0 else 0
            result.append({
                'section_id': cls['session__section__id'],
                'section_name': cls['session__section__name'],
                'course_name': cls['session__section__batch__course__name'],
                'department': cls['session__section__batch__course__department__short_name'],
                'total_records': cls['total_records'],
                'present': cls['present_count'],
                'absent': cls['absent_count'],
                'leave': cls['leave_count'],
                'on_duty': cls['od_count'],
                'attendance_rate': round(rate, 2)
            })
        
        return {'classes': result}
    
    def _get_student_stats(self, records_qs, section_id=None):
        """Student-wise statistics"""
        student_stats = records_qs.values(
            'student__id',
            'student__reg_no',
            'student__user__username',
            'student__section__name'
        ).annotate(
            total_records=Count('id'),
            present_count=Count('id', filter=Q(status__in=['P', 'OD', 'LATE'])),
            absent_count=Count('id', filter=Q(status='A')),
            leave_count=Count('id', filter=Q(status='LEAVE')),
            od_count=Count('id', filter=Q(status='OD')),
            late_count=Count('id', filter=Q(status='LATE'))
        ).order_by('-total_records')
        
        result = []
        for student in student_stats:
            rate = (student['present_count'] / student['total_records'] * 100) if student['total_records'] > 0 else 0
            result.append({
                'student_id': student['student__id'],
                'reg_no': student['student__reg_no'],
                'name': student['student__user__username'],
                'section': student['student__section__name'],
                'total_records': student['total_records'],
                'present': student['present_count'],
                'absent': student['absent_count'],
                'leave': student['leave_count'],
                'on_duty': student['od_count'],
                'late': student['late_count'],
                'attendance_rate': round(rate, 2)
            })
        
        return {'students': result}


class AnalyticsFiltersView(APIView):
    """
    Get available filters based on user's permission level
    """
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        staff_profile = getattr(user, 'staff_profile', None)
        
        # Determine permission level
        can_view_all = 'analytics.view_all_analytics' in perms or user.is_superuser
        can_view_department = 'analytics.view_department_analytics' in perms or can_view_all
        
        # ALSO check if user is an HOD via DepartmentRole - they should have department-level access
        is_hod = False
        if staff_profile and not can_view_department:
            is_hod = DepartmentRole.objects.filter(
                staff=staff_profile,
                role='HOD',
                is_active=True
            ).exists()
            if is_hod:
                can_view_department = True
        
        can_view_class = 'analytics.view_class_analytics' in perms or can_view_department
        
        if not (can_view_all or can_view_department or can_view_class):
            raise PermissionDenied('You do not have permission to view analytics')
        
        departments = []
        sections = []
        
        if can_view_all:
            # Get all departments
            departments = list(Department.objects.all().values('id', 'name', 'short_name').order_by('name'))
            # Get all sections
            sections = list(Section.objects.select_related('batch', 'batch__course', 'batch__course__department').values(
                'id', 'name', 
                'batch__course__name',
                'batch__course__department__id',
                'batch__course__department__short_name'
            ).order_by('batch__course__department__short_name', 'name'))
            
        elif can_view_department and staff_profile:
            # Get only staff's departments
            dept_roles = DepartmentRole.objects.filter(staff=staff_profile, is_active=True)
            dept_ids = [dr.department_id for dr in dept_roles if dr.department_id]
            
            # Also check teaching assignments
            from .models import TeachingAssignment
            teaching_depts = TeachingAssignment.objects.filter(
                staff=staff_profile, 
                is_active=True
            ).values_list('section__batch__course__department_id', flat=True).distinct()
            dept_ids.extend(list(teaching_depts))
            dept_ids = list(set(filter(None, dept_ids)))
            
            if dept_ids:
                departments = list(Department.objects.filter(id__in=dept_ids).values('id', 'name', 'short_name').order_by('name'))
                sections = list(Section.objects.filter(
                    batch__course__department_id__in=dept_ids
                ).select_related('batch', 'batch__course', 'batch__course__department').values(
                    'id', 'name',
                    'batch__course__name',
                    'batch__course__department__id',
                    'batch__course__department__short_name'
                ).order_by('batch__course__department__short_name', 'name'))
        
        elif can_view_class and staff_profile:
            # Get only advisor's sections
            from .models import SectionAdvisor
            advisor_sections_qs = SectionAdvisor.objects.filter(
                advisor=staff_profile,
                is_active=True
            ).select_related('section', 'section__batch', 'section__batch__course', 'section__batch__course__department')
            
            section_ids = [sa.section_id for sa in advisor_sections_qs]
            
            if section_ids:
                sections = list(Section.objects.filter(
                    id__in=section_ids
                ).select_related('batch', 'batch__course', 'batch__course__department').values(
                    'id', 'name',
                    'batch__course__name',
                    'batch__course__department__id',
                    'batch__course__department__short_name'
                ).order_by('batch__course__department__short_name', 'name'))
                
                # Get departments of these sections
                dept_ids = list(set([s['batch__course__department__id'] for s in sections]))
                departments = list(Department.objects.filter(id__in=dept_ids).values('id', 'name', 'short_name').order_by('name'))
        
        return Response({
            'permission_level': 'all' if can_view_all else ('department' if can_view_department else 'class'),
            'departments': departments,
            'sections': sections
        })


class ClassAttendanceReportView(APIView):
    """Return a compact report for a given section and date (defaults to today)."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        staff_profile = getattr(user, 'staff_profile', None)

        # permission: allow if user can view class or department or all
        can_view_all = 'analytics.view_all_analytics' in perms or user.is_superuser
        can_view_department = 'analytics.view_department_analytics' in perms or can_view_all
        can_view_class = 'analytics.view_class_analytics' in perms or can_view_department
        if not (can_view_all or can_view_department or can_view_class):
            raise PermissionDenied('You do not have permission to view analytics')

        section_id = request.query_params.get('section_id')
        date_str = request.query_params.get('date')
        try:
            if date_str:
                target_date = date.fromisoformat(date_str)
            else:
                target_date = date.today()
        except Exception:
            target_date = date.today()

        if not section_id:
            return Response({'detail': 'section_id required'}, status=400)

        # permission-based section scoping
        allowed_section_ids = None
        if not can_view_all:
            if can_view_department and staff_profile:
                from .models import TeachingAssignment
                dept_roles = DepartmentRole.objects.filter(staff=staff_profile, is_active=True)
                dept_ids = [dr.department_id for dr in dept_roles if dr.department_id]
                teaching_depts = TeachingAssignment.objects.filter(staff=staff_profile, is_active=True).values_list('section__batch__course__department_id', flat=True).distinct()
                dept_ids.extend(list(teaching_depts))
                dept_ids = list(set(filter(None, dept_ids)))
                if dept_ids:
                    allowed_section_ids = list(Section.objects.filter(batch__course__department_id__in=dept_ids).values_list('id', flat=True))
                else:
                    allowed_section_ids = []
            elif can_view_class and staff_profile:
                from .models import SectionAdvisor
                allowed_section_ids = list(SectionAdvisor.objects.filter(advisor=staff_profile, is_active=True).values_list('section_id', flat=True))

        if allowed_section_ids is not None and int(section_id) not in allowed_section_ids:
            return Response({'detail': 'Permission denied for this section'}, status=403)

        # Build records for the given section/date. Allow optional `period_index` param.
        period_index = request.query_params.get('period_index')
        recs_q = PeriodAttendanceRecord.objects.filter(
            session__section_id=int(section_id),
            session__date=target_date
        ).select_related('student').order_by('-id')
        if period_index:
            try:
                recs_q = recs_q.filter(session__period__index=int(period_index))
            except Exception:
                pass
        recs = recs_q

        # total strength: count students in the section (active only)
        total_strength = StudentProfile.objects.filter(section_id=int(section_id)).exclude(status__in=['INACTIVE', 'DEBAR']).count()

        # counts
        present_count = recs.filter(status__in=['P', 'OD', 'LATE']).count()
        absent_count = recs.filter(status='A').count()
        leave_count = recs.filter(status='LEAVE').count()
        od_count = recs.filter(status='OD').count()
        late_count = recs.filter(status='LATE').count()

        # build lists of all students' reg_no last-3-digits per status (unique, preserve order)
        def last3_list_for(status_code):
            seen = set()
            out = []
            for r in recs.filter(status=status_code).select_related('student'):
                reg = getattr(getattr(r, 'student', None), 'reg_no', None)
                if not reg:
                    continue
                suffix = reg[-3:]
                if suffix not in seen:
                    out.append(suffix)
                    seen.add(suffix)
            return out

        absent_list = last3_list_for('A')
        leave_list = last3_list_for('LEAVE')
        od_list = last3_list_for('OD')
        late_list = last3_list_for('LATE')

        attendance_pct = (present_count / total_strength * 100) if total_strength > 0 else 0

        # section display + batch/department
        section_obj = Section.objects.filter(pk=int(section_id)).select_related('batch', 'batch__course', 'batch__course__department').first()
        section_name = section_obj.name if section_obj else ''
        # Group sessions by a canonical group key so multi-section assignments
        # for the same subject/parent curriculum_row/subject_batch appear as a
        # single aggregated card.
        groups = {}
        import re
        from .models import TeachingAssignment as _TA

        for session in sessions:
            ta = session.timetable_assignment
            # ensure the staff is responsible for this session
            if ta and ta.staff_id and ta.staff_id != staff_profile.id:
                is_teaching = teaching_assignments.filter(section_id=session.section_id).exists()
                if not is_teaching:
                    continue

            period_idx = getattr(session.period, 'index', 0) if session.period else 0

            # determine canonical group id
            group_id = None
            group_type = None

            if ta and getattr(ta, 'curriculum_row', None):
                group_type = 'curr'
                group_id = ta.curriculum_row_id
            elif ta and getattr(ta, 'subject_batch', None):
                group_type = 'batch'
                group_id = ta.subject_batch_id
            elif ta and getattr(ta, 'subject_text', None):
                raw = (ta.subject_text or '').strip()
                subj_code = re.sub(r'[^A-Za-z0-9]', '', raw).lower()
                if subj_code:
                    group_type = 'subj'
                    group_id = subj_code
            else:
                # fallback: try to resolve via TeachingAssignment mappings
                ta_match = _TA.objects.filter(staff=staff_profile, is_active=True).filter(
                    Q(section_id=session.section_id) | Q(section__isnull=True)
                ).filter(
                    Q(curriculum_row__isnull=False) | Q(elective_subject__isnull=False) | Q(subject__isnull=False)
                ).first()
                if ta_match:
                    if getattr(ta_match, 'curriculum_row', None):
                        group_type = 'curr'
                        group_id = ta_match.curriculum_row_id
                    elif getattr(ta_match, 'elective_subject', None):
                        # prefer parent curriculum_row for elective
                        parent_id = getattr(ta_match.elective_subject, 'parent_id', None)
                        if parent_id:
                            group_type = 'elective_parent'
                            group_id = parent_id
                        else:
                            group_type = 'elective'
                            group_id = ta_match.elective_subject_id
                    elif getattr(ta_match, 'subject', None) and getattr(ta_match.subject, 'code', None):
                        group_type = 'subj'
                        group_id = re.sub(r'[^A-Za-z0-9]', '', getattr(ta_match.subject, 'code', '')).lower()

            if group_type is None:
                group_type = 'section'
                group_id = session.section_id

            key = (period_idx, group_type, group_id)
            groups.setdefault(key, []).append(session)

            # Build period_stats from grouped sessions
            for key, sess_list in groups.items():
                period_idx, gtype, gid = key
                sess_ids = [s.id for s in sess_list]
                records = PeriodAttendanceRecord.objects.filter(session_id__in=sess_ids)

                # determine involved sections
                section_ids = list({s.section_id for s in sess_list})

                if gtype in ('curr', 'elective_parent', 'batch', 'subj', 'elective'):
                    total_strength = StudentProfile.objects.filter(section_id__in=section_ids).exclude(status__in=['INACTIVE', 'DEBAR']).count()
                else:
                    total_strength = StudentProfile.objects.filter(section_id=section_ids[0]).exclude(status__in=['INACTIVE', 'DEBAR']).count() if section_ids else 0

                total_records = records.count()
                present_count = records.filter(status__in=['P', 'OD', 'LATE']).count()
                absent_count = records.filter(status='A').count()
                leave_count = records.filter(status='LEAVE').count()
                od_count = records.filter(status='OD').count()
                late_count = records.filter(status='LATE').count()
                attendance_pct = (present_count / total_records * 100) if total_records > 0 else 0

                # subject display name: try curriculum_row -> subject_batch -> subject_text
                subject_name = ''
                try:
                    if gtype in ('curr', 'elective_parent'):
                        from curriculum.models import CurriculumDepartment
                        parent = CurriculumDepartment.objects.filter(pk=gid).first()
                        if parent:
                            code = getattr(parent, 'course_code', '') or ''
                            name = getattr(parent, 'course_name', '') or ''
                            subject_name = f"{code} - {name}".strip(' -')
                    elif gtype == 'batch':
                        # use first session's timetable_assignment subject_batch name
                        ta0 = sess_list[0].timetable_assignment
                        subject_name = getattr(getattr(ta0, 'subject_batch', None), 'name', '') if ta0 else ''
                    elif gtype in ('subj', 'elective'):
                        # show normalized subject code
                        subject_name = str(gid).upper()
                    else:
                        ta0 = sess_list[0].timetable_assignment
                        if ta0 and getattr(ta0, 'curriculum_row', None):
                            code = getattr(ta0.curriculum_row, 'course_code', '') or ''
                            name = getattr(ta0.curriculum_row, 'course_name', '') or ''
                            subject_name = f"{code} - {name}".strip(' -')
                        elif ta0 and getattr(ta0, 'subject_text', None):
                            subject_name = ta0.subject_text
                except Exception:
                    subject_name = ''

                # Build card
                period_stats.append({
                    'session_id': None if len(sess_list) > 1 else sess_list[0].id,
                    'period_index': period_idx,
                    'period_label': getattr(sess_list[0].period, 'label', '') if sess_list and getattr(sess_list[0], 'period', None) else '',
                    'period_start': str(getattr(sess_list[0].period, 'start_time', '')) if sess_list and getattr(sess_list[0], 'period', None) else '',
                    'period_end': str(getattr(sess_list[0].period, 'end_time', '')) if sess_list and getattr(sess_list[0], 'period', None) else '',
                    'section_id': None if len(section_ids) > 1 else (section_ids[0] if section_ids else None),
                    'section_name': '' if len(section_ids) > 1 else (getattr(sess_list[0].section, 'name', '') if sess_list and getattr(sess_list[0], 'section', None) else ''),
                    'sections': section_ids if len(section_ids) > 1 else None,
                    'section_names': [getattr(s.section, 'name', '') for s in sess_list] if len(section_ids) > 1 else None,
                    'session_ids': sess_ids,
                    'group_key': f"{gtype}:{gid}",
                    'subject': subject_name,
                    'total_strength': total_strength,
                    'total_marked': total_records,
                    'present': present_count,
                    'absent': absent_count,
                    'leave': leave_count,
                    'late': late_count,
                    'on_duty': od_count,
                    'attendance_percentage': round(attendance_pct, 2),
                    'is_locked': any(getattr(s, 'is_locked', False) for s in sess_list),
                    'marked_at': min((getattr(s, 'created_at') for s in sess_list if getattr(s, 'created_at', None)), default=None).isoformat() if any(getattr(s, 'created_at', None) for s in sess_list) else None,
                    'marked_by': getattr(getattr(sess_list[0], 'created_by', None), 'user', None).username if getattr(sess_list[0], 'created_by', None) and getattr(getattr(sess_list[0], 'created_by', None), 'user', None) else ''
                })

                total_records = records.count()
                present_count = records.filter(status__in=['P', 'OD', 'LATE']).count()
                absent_count = records.filter(status='A').count()
                leave_count = records.filter(status='LEAVE').count()
                od_count = records.filter(status='OD').count()
                late_count = records.filter(status='LATE').count()
                
                attendance_pct = (present_count / total_records * 100) if total_records > 0 else 0

                # `total_strength` was computed above for aggregated keys
                # (curr, batch, elective_parent). For single-session keys the
                # default was already set above; do not overwrite it here.

                # Get subject info - defensive attribute access
                subject_name = ''
                try:
                    if ta and getattr(ta, 'curriculum_row', None):
                        code = getattr(ta.curriculum_row, 'course_code', '') or ''
                        name = getattr(ta.curriculum_row, 'course_name', '') or ''
                        subject_name = f"{code} - {name}".strip(' -') if code or name else ''
                    elif ta and getattr(ta, 'subject_batch', None):
                        subject_name = getattr(ta.subject_batch, 'name', '') or 'Subject'
                    # If we deduped by elective_parent, prefer showing the parent curriculum row
                    # course code/name if available.
                    if key[0] == 'elective_parent':
                        try:
                            from curriculum.models import CurriculumDepartment
                            parent = CurriculumDepartment.objects.filter(pk=key[2]).first()
                            if parent:
                                code = getattr(parent, 'course_code', '') or ''
                                name = getattr(parent, 'course_name', '') or ''
                                subject_name = f"{code} - {name}".strip(' -') if code or name else subject_name
                        except Exception:
                            pass
                except Exception:
                    subject_name = ''
                
                try:
                    # determine aggregated sections for this card (if any)
                    aggregated_sections = None
                    if key[0] in ('curr', 'elective_parent'):
                        aggregated_sections = list(section_ids_for_curr) if 'section_ids_for_curr' in locals() else None
                    elif key[0] == 'batch':
                        aggregated_sections = list(section_ids_for_batch) if 'section_ids_for_batch' in locals() else None

                    period_stats.append({
                        'session_id': session.id,
                        'period_index': getattr(session.period, 'index', 0) if session.period else 0,
                        'period_label': getattr(session.period, 'label', '') if session.period else '',
                        'period_start': str(getattr(session.period, 'start_time', '')) if session.period and getattr(session.period, 'start_time', None) else '',
                        'period_end': str(getattr(session.period, 'end_time', '')) if session.period and getattr(session.period, 'end_time', None) else '',
                        'section_id': None if aggregated_sections else session.section_id,
                        'section_name': '' if aggregated_sections else (getattr(session.section, 'name', '') if session.section else ''),
                        'sections': aggregated_sections,
                        'subject': subject_name,
                        'total_strength': total_strength,
                        'total_marked': total_records,
                        'present': present_count,
                        'absent': absent_count,
                        'leave': leave_count,
                        'late': late_count,
                        'on_duty': od_count,
                        'attendance_percentage': round(attendance_pct, 2),
                        'is_locked': getattr(session, 'is_locked', False),
                        'marked_at': session.created_at.isoformat() if getattr(session, 'created_at', None) else None,
                        # `PeriodAttendanceSession` uses `created_by` (StaffProfile). Use that for marked_by display.
                        'marked_by': getattr(getattr(session, 'created_by', None), 'user', None).username if getattr(session, 'created_by', None) and getattr(getattr(session, 'created_by', None), 'user', None) else ''
                    })
                except Exception as e:
                    # Log but don't crash if one session fails
                    import logging
                    logging.error(f"Failed to add period stat for session {session.id}: {e}")
                    continue
            
            # Sort by period index
            period_stats.sort(key=lambda x: x['period_index'])
            
            return Response({
                'date': target_date.isoformat(),
                'periods': period_stats,
                'total_periods': len(period_stats),
                'staff_name': user.username if user else ''
            })


class TodayPeriodAttendanceView(APIView):
    """Return period-wise attendance cards for the given date (defaults to today)."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        staff_profile = getattr(user, 'staff_profile', None)

        # permission: allow if user can view class or department or all
        can_view_all = 'analytics.view_all_analytics' in perms or user.is_superuser
        can_view_department = 'analytics.view_department_analytics' in perms or can_view_all
        can_view_class = 'analytics.view_class_analytics' in perms or can_view_department
        if not (can_view_all or can_view_department or can_view_class):
            raise PermissionDenied('You do not have permission to view analytics')

        date_str = request.query_params.get('date')
        try:
            target_date = date.fromisoformat(date_str) if date_str else date.today()
        except Exception:
            target_date = date.today()

        # base sessions for the date
        sessions_q = PeriodAttendanceSession.objects.filter(date=target_date).select_related(
            'period', 'section', 'section__batch', 'section__batch__course', 'section__batch__course__department',
            'timetable_assignment', 'timetable_assignment__curriculum_row', 'timetable_assignment__subject_batch',
            'teaching_assignment', 'teaching_assignment__curriculum_row', 'teaching_assignment__elective_subject', 'teaching_assignment__elective_subject__parent',
            'teaching_assignment__subject'
        ).order_by('period__index')

        # apply permission scoping
        if not can_view_all:
            if can_view_department and staff_profile:
                from .models import TeachingAssignment
                dept_roles = DepartmentRole.objects.filter(staff=staff_profile, is_active=True)
                dept_ids = [dr.department_id for dr in dept_roles if dr.department_id]
                teaching_depts = TeachingAssignment.objects.filter(staff=staff_profile, is_active=True).values_list('section__batch__course__department_id', flat=True).distinct()
                dept_ids.extend(list(teaching_depts))
                dept_ids = list(set(filter(None, dept_ids)))
                if dept_ids:
                    sessions_q = sessions_q.filter(section__batch__course__department_id__in=dept_ids)
                else:
                    sessions_q = sessions_q.none()
            elif can_view_class and staff_profile:
                from .models import SectionAdvisor
                advisor_sections = SectionAdvisor.objects.filter(advisor=staff_profile, is_active=True).values_list('section_id', flat=True)
                if advisor_sections:
                    sessions_q = sessions_q.filter(section_id__in=advisor_sections)
                else:
                    sessions_q = sessions_q.none()

        period_stats = []
        for session in sessions_q:
            try:
                recs = PeriodAttendanceRecord.objects.filter(session=session)
                total_strength = StudentProfile.objects.filter(section_id=session.section_id).exclude(status__in=['INACTIVE', 'DEBAR']).count() if session.section_id else 0
                total_records = recs.count()
                present_count = recs.filter(status__in=['P', 'OD', 'LATE']).count()
                absent_count = recs.filter(status='A').count()
                leave_count = recs.filter(status='LEAVE').count()
                od_count = recs.filter(status='OD').count()
                late_count = recs.filter(status='LATE').count()
                attendance_pct = (present_count / total_strength * 100) if total_strength > 0 else 0

                # subject display
                subject_name = ''
                ta = getattr(session, 'timetable_assignment', None)
                teach = getattr(session, 'teaching_assignment', None)
                try:
                    if teach and getattr(teach, 'elective_subject', None):
                        es = teach.elective_subject
                        code = getattr(es, 'course_code', '') or ''
                        name = getattr(es, 'course_name', '') or ''
                        subject_name = f"{code} - {name}".strip(' -') if code or name else str(es)
                    elif teach and getattr(teach, 'curriculum_row', None):
                        cr = teach.curriculum_row
                        code = getattr(cr, 'course_code', '') or ''
                        name = getattr(cr, 'course_name', '') or ''
                        subject_name = f"{code} - {name}".strip(' -') if code or name else ''
                    elif ta and getattr(ta, 'curriculum_row', None):
                        code = getattr(ta.curriculum_row, 'course_code', '') or ''
                        name = getattr(ta.curriculum_row, 'course_name', '') or ''
                        subject_name = f"{code} - {name}".strip(' -') if code or name else ''
                    elif ta and getattr(ta, 'subject_batch', None):
                        subject_name = getattr(ta.subject_batch, 'name', '') or 'Subject'
                    elif ta and getattr(ta, 'subject_text', None):
                        subject_name = ta.subject_text
                except Exception:
                    subject_name = ''

                period_stats.append({
                    'session_id': session.id,
                    'period_index': getattr(session.period, 'index', 0) if session.period else 0,
                    'period_label': getattr(session.period, 'label', '') if session.period else '',
                    'period_start': str(getattr(session.period, 'start_time', '')) if session.period and getattr(session.period, 'start_time', None) else '',
                    'period_end': str(getattr(session.period, 'end_time', '')) if session.period and getattr(session.period, 'end_time', None) else '',
                    'section_id': session.section_id,
                    'section_name': getattr(session.section, 'name', '') if session.section else '',
                    'subject': subject_name,
                    'total_strength': total_strength,
                    'total_marked': total_records,
                    'present': present_count,
                    'absent': absent_count,
                    'leave': leave_count,
                    'late': late_count,
                    'on_duty': od_count,
                    'attendance_percentage': round(attendance_pct, 2),
                    'is_locked': getattr(session, 'is_locked', False),
                    'marked_at': session.created_at.isoformat() if getattr(session, 'created_at', None) else None,
                    'marked_by': getattr(getattr(session, 'created_by', None), 'user', None).username if getattr(session, 'created_by', None) and getattr(getattr(session, 'created_by', None), 'user', None) else ''
                })
            except Exception:
                # don't let one session break the whole response
                continue

        period_stats.sort(key=lambda x: x['period_index'])
        return Response({'date': target_date.isoformat(), 'periods': period_stats, 'total_periods': len(period_stats), 'staff_name': user.username if user else ''})


class PeriodAttendanceReportView(APIView):
    """Return a detailed attendance report for a specific period session."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        staff_profile = getattr(user, 'staff_profile', None)

        # permission: allow if user can view class or department or all
        can_view_all = 'analytics.view_all_analytics' in perms or user.is_superuser
        can_view_department = 'analytics.view_department_analytics' in perms or can_view_all
        can_view_class = 'analytics.view_class_analytics' in perms or can_view_department
        if not (can_view_all or can_view_department or can_view_class):
            raise PermissionDenied('You do not have permission to view analytics')

        session_id = request.query_params.get('session_id')
        if not session_id:
            return Response({'detail': 'session_id required'}, status=400)

        try:
            session = PeriodAttendanceSession.objects.select_related(
                'period', 'section', 'section__batch', 'section__batch__course',
                'section__batch__course__department', 'timetable_assignment',
                'timetable_assignment__curriculum_row', 'timetable_assignment__subject_batch',
                'created_by'
            ).get(pk=int(session_id))
        except PeriodAttendanceSession.DoesNotExist:
            return Response({'detail': 'Session not found'}, status=404)

        # permission-based section scoping
        allowed_section_ids = None
        if not can_view_all:
            if can_view_department and staff_profile:
                from .models import TeachingAssignment
                dept_roles = DepartmentRole.objects.filter(staff=staff_profile, is_active=True)
                dept_ids = [dr.department_id for dr in dept_roles if dr.department_id]
                teaching_depts = TeachingAssignment.objects.filter(staff=staff_profile, is_active=True).values_list('section__batch__course__department_id', flat=True).distinct()
                dept_ids.extend(list(teaching_depts))
                dept_ids = list(set(filter(None, dept_ids)))
                if dept_ids:
                    allowed_section_ids = list(Section.objects.filter(batch__course__department_id__in=dept_ids).values_list('id', flat=True))
                else:
                    allowed_section_ids = []
            elif can_view_class and staff_profile:
                from .models import SectionAdvisor
                allowed_section_ids = list(SectionAdvisor.objects.filter(advisor=staff_profile, is_active=True).values_list('section_id', flat=True))

        if allowed_section_ids is not None and session.section_id not in allowed_section_ids:
            return Response({'detail': 'Permission denied for this session'}, status=403)

        # Get attendance records for this session
        recs = PeriodAttendanceRecord.objects.filter(
            session=session
        ).select_related('student').order_by('-id')

        # total strength: count students in the section (active only)
        total_strength = StudentProfile.objects.filter(section_id=session.section_id).exclude(status__in=['INACTIVE', 'DEBAR']).count()

        # counts
        present_count = recs.filter(status__in=['P', 'OD', 'LATE']).count()
        absent_count = recs.filter(status='A').count()
        leave_count = recs.filter(status='LEAVE').count()
        od_count = recs.filter(status='OD').count()
        late_count = recs.filter(status='LATE').count()

        # build lists of all students' reg_no last-3-digits per status (unique, preserve order)
        def last3_list_for(status_code):
            seen = set()
            out = []
            for r in recs.filter(status=status_code).select_related('student'):
                reg = getattr(getattr(r, 'student', None), 'reg_no', None)
                if not reg:
                    continue
                suffix = reg[-3:]
                if suffix not in seen:
                    out.append(suffix)
                    seen.add(suffix)
            return out

        absent_list = last3_list_for('A')
        leave_list = last3_list_for('LEAVE')
        od_list = last3_list_for('OD')
        late_list = last3_list_for('LATE')

        attendance_pct = (present_count / total_strength * 100) if total_strength > 0 else 0

        # Get period info
        period_label = getattr(session.period, 'label', '') if session.period else ''
        period_start = str(getattr(session.period, 'start_time', '')) if session.period and getattr(session.period, 'start_time', None) else ''
        period_end = str(getattr(session.period, 'end_time', '')) if session.period and getattr(session.period, 'end_time', None) else ''

        # section display + batch/department
        section_name = getattr(session.section, 'name', '') if session.section else ''
        batch_name = getattr(getattr(session.section, 'batch', None), 'name', '') if session.section and getattr(session.section, 'batch', None) else ''
        dept = getattr(getattr(getattr(session.section, 'batch', None), 'course', None), 'department', None) if session.section else None
        department_name = getattr(dept, 'name', '') if dept else ''
        department_short = getattr(dept, 'short_name', '') if dept else ''

        # Get subject info - defensive attribute access
        subject_name = ''
        ta = session.timetable_assignment
        try:
            if ta and getattr(ta, 'curriculum_row', None):
                code = getattr(ta.curriculum_row, 'course_code', '') or ''
                name = getattr(ta.curriculum_row, 'course_name', '') or ''
                subject_name = f"{code} - {name}".strip(' -') if code or name else ''
            elif ta and getattr(ta, 'subject_batch', None):
                subject_name = getattr(ta.subject_batch, 'name', '') or 'Subject'
        except Exception:
            subject_name = ''

        return Response({
            'session_id': session.id,
            'date': session.date.isoformat(),
            'period_label': period_label,
            'period_start': period_start,
            'period_end': period_end,
            'subject': subject_name,
            'section_id': session.section_id,
            'section_name': section_name,
            'batch_name': batch_name,
            'department_name': department_name,
            'department_short': department_short,
            'total_strength': total_strength,
            'total_marked': recs.count(),
            'present': present_count,
            'absent': absent_count,
            'leave': leave_count,
            'late': late_count,
            'on_duty': od_count,
            'absent_list': absent_list,
            'leave_list': leave_list,
            'od_list': od_list,
            'late_list': late_list,
            'attendance_percentage': round(attendance_pct, 2),
            'is_locked': getattr(session, 'is_locked', False),
            'marked_by': getattr(getattr(session, 'created_by', None), 'user', None).username if getattr(session, 'created_by', None) and getattr(getattr(session, 'created_by', None), 'user', None) else '',
            'marked_at': session.created_at.isoformat() if getattr(session, 'created_at', None) else None
        })


class OverallSectionView(APIView):
    """Return overall section-level attendance summary for the given date (defaults to today).

    Permissioning mirrors other analytics endpoints: users with
    'analytics.view_all_analytics' can view all, 'analytics.view_department_analytics'
    can view their departments, and 'analytics.view_class_analytics' can view advisor sections.
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        perms = get_user_permissions(user)
        staff_profile = getattr(user, 'staff_profile', None)

        can_view_all = 'analytics.view_all_analytics' in perms or user.is_superuser
        can_view_department = 'analytics.view_department_analytics' in perms or can_view_all
        can_view_class = 'analytics.view_class_analytics' in perms or can_view_department
        if not (can_view_all or can_view_department or can_view_class):
            raise PermissionDenied('You do not have permission to view analytics')

        date_str = request.query_params.get('date')
        date_from_str = request.query_params.get('date_from')
        date_to_str = request.query_params.get('date_to')
        complete = request.query_params.get('complete', '').lower() in ('true', '1')
        try:
            if date_from_str and date_to_str:
                date_from = date.fromisoformat(date_from_str)
                date_to = date.fromisoformat(date_to_str)
            elif date_str:
                date_from = date_to = date.fromisoformat(date_str)
            else:
                date_from = date_to = date.today()
        except Exception:
            date_from = date_to = date.today()
        is_range = complete or (date_from != date_to)

        # Use DailyAttendanceSession instead of Period 1 attendance
        from .models import DailyAttendanceSession, DailyAttendanceRecord

        base_qs = DailyAttendanceSession.objects.all() if complete else DailyAttendanceSession.objects.filter(date__range=(date_from, date_to))
        sessions_q = base_qs.select_related(
            'section', 'section__batch', 'section__batch__course', 'section__batch__course__department'
        )

        # apply permission scoping and build section map
        section_map = {}
        
        if not can_view_all:
            if can_view_department and staff_profile:
                # For department view, show ALL sections in the department (not just assigned ones)
                from .models import TeachingAssignment
                dept_roles = DepartmentRole.objects.filter(staff=staff_profile, is_active=True)
                dept_ids = [dr.department_id for dr in dept_roles if dr.department_id]
                # Only use DepartmentRole, not teaching assignments, to avoid showing only assigned sections
                dept_ids = list(set(filter(None, dept_ids)))
                
                if dept_ids:
                    # Get all sections in these departments
                    all_dept_sections = Section.objects.filter(
                        batch__course__department_id__in=dept_ids
                    ).select_related('batch', 'batch__course', 'batch__course__department')
                    
                    # Pre-populate section_map with all department sections
                    for section in all_dept_sections:
                        section_map[section.id] = {
                            'section_id': section.id,
                            'section_name': section.name,
                            'department_name': section.batch.course.department.name if section.batch and section.batch.course and section.batch.course.department else '',
                            'department_short': section.batch.course.department.short_name if section.batch and section.batch.course and section.batch.course.department else '',
                            'batch': section.batch.name if section.batch else '',
                            'batch_id': section.batch.id if section.batch else None,
                            'total_strength': StudentProfile.objects.filter(section_id=section.id).exclude(status__in=['INACTIVE', 'DEBAR']).count(),
                            'total_marked': 0,
                            'present': 0,
                            'absent': 0,
                            'on_duty': 0,
                            'leave': 0,
                            'is_locked': False,
                            'marked_at': None,
                            'attendance_session_id': None,
                            'session_id': None,
                        }
                    
                    # Filter sessions to these departments
                    sessions_q = sessions_q.filter(section__batch__course__department_id__in=dept_ids)
                else:
                    sessions_q = sessions_q.none()
            elif can_view_class and staff_profile:
                from .models import SectionAdvisor
                advisor_sections = SectionAdvisor.objects.filter(advisor=staff_profile, is_active=True).values_list('section_id', flat=True)
                if advisor_sections:
                    sessions_q = sessions_q.filter(section_id__in=advisor_sections)
                else:
                    sessions_q = sessions_q.none()

        # aggregate per-section stats from sessions
        for session in sessions_q:
            sec_id = session.section_id
            sec_name = getattr(session.section, 'name', '') if session.section else ''
            
            # Only create new entry if not already in map (for non-department views)
            if sec_id not in section_map:
                section_map[sec_id] = {
                        'section_id': sec_id,
                            'section_name': sec_name,
                            'department_name': getattr(getattr(getattr(session.section, 'batch', None), 'course', None), 'department', None) and getattr(getattr(getattr(session.section, 'batch', None), 'course', None), 'department', None).name or '',
                            'department_short': getattr(getattr(getattr(session.section, 'batch', None), 'course', None), 'department', None) and getattr(getattr(getattr(session.section, 'batch', None), 'course', None), 'department', None).short_name or '',
                            'batch': getattr(getattr(session.section, 'batch', None), 'name', '') if getattr(session, 'section', None) and getattr(session.section, 'batch', None) else '',
                            'batch_id': getattr(getattr(session.section, 'batch', None), 'id', None) if getattr(session, 'section', None) and getattr(session.section, 'batch', None) else None,
                    'total_strength': 0,
                    'total_marked': 0,
                    'present': 0,
                    'absent': 0,
                    'on_duty': 0,
                    'leave': 0,
                    'is_locked': False,
                    'marked_at': None,
                    'attendance_session_id': None,
                    'session_id': None,
                }
            
            try:
                recs = DailyAttendanceRecord.objects.filter(session=session)
                total_strength = StudentProfile.objects.filter(section_id=session.section_id).exclude(status__in=['INACTIVE', 'DEBAR']).count() if session.section_id else 0
                total_records = recs.count()
                present_count = recs.filter(status__in=['P', 'OD', 'LATE']).count()
                absent_count = recs.filter(status='A').count()
                od_count = recs.filter(status='OD').count()
                leave_count = recs.filter(status='LEAVE').count()

                entry = section_map[sec_id]
                # Update with attendance data
                entry['total_strength'] = max(entry['total_strength'], total_strength)
                entry['total_marked'] += total_records
                entry['present'] += present_count
                entry['absent'] += absent_count
                entry['on_duty'] += od_count
                entry['leave'] += leave_count
                entry['is_locked'] = entry['is_locked'] or getattr(session, 'is_locked', False)
                # Store the session ID for fetching detailed records later
                entry['attendance_session_id'] = session.id
                entry['session_id'] = session.id
                if getattr(session, 'created_at', None):
                    if not entry['marked_at']:
                        entry['marked_at'] = session.created_at.isoformat()
                    else:
                        # keep earliest
                        try:
                            if session.created_at.isoformat() < entry['marked_at']:
                                entry['marked_at'] = session.created_at.isoformat()
                        except Exception:
                            pass
            except Exception:
                continue

        # finalize attendance_percentage
        result = []
        for sec in section_map.values():
            total_marked = sec.get('total_marked', 0)
            present = sec.get('present', 0)
            attendance_pct = (present / sec['total_strength'] * 100) if sec['total_strength'] > 0 else 0
            sec['attendance_percentage'] = round(attendance_pct, 2)
            result.append(sec)

        # sort by section name
        result.sort(key=lambda x: (x['section_name'] or ''))

        return Response({'date': date_from.isoformat(), 'date_from': date_from.isoformat(), 'date_to': date_to.isoformat(), 'is_range': is_range, 'sections': result, 'total_sections': len(result)})


class MyClassStudentsView(APIView):
    """
    Get students from advisor's assigned sections for daily attendance marking.
    Also includes sections where daily attendance has been assigned to this staff (via swap).
    
    Access is granted to:
    - Section advisors for their assigned sections
    - Staff members with sections assigned to them via daily attendance swap
    - Users with appropriate analytics permissions
    """
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')
        
        from .models import SectionAdvisor, StudentProfile, DailyAttendanceSession, DailyAttendanceRecord
        from datetime import date as date_class
        
        # Get advisor's assigned sections
        advisor_section_ids = set(SectionAdvisor.objects.filter(
            advisor=staff_profile,
            is_active=True
        ).values_list('section_id', flat=True))
        
        # Get sections with daily attendance assigned to this staff (via swap)
        # Only include future or today's sessions
        today = date_class.today()
        assigned_section_ids = set(DailyAttendanceSession.objects.filter(
            assigned_to=staff_profile,
            date__gte=today
        ).values_list('section_id', flat=True).distinct())
        
        # Keep all advisor sections visible (even if assigned to others)
        # Original advisors should always see their sections but in read-only mode
        
        # Combine both sets
        all_section_ids = advisor_section_ids | assigned_section_ids
        
        if not all_section_ids:
            return Response({'sections': [], 'message': 'No assigned sections found'})
        
        # Get all students from these sections (exclude inactive/debarred)
        students = StudentProfile.objects.filter(
            section_id__in=all_section_ids
        ).exclude(status__in=['INACTIVE', 'DEBAR']).select_related(
            'user', 
            'section', 
            'section__batch', 
            'section__batch__course__department'
        ).order_by('section__name', 'reg_no')
        
        # Get session status for the requested date (defaults to today)
        date_str = request.query_params.get('date')
        try:
            target_date = date_class.fromisoformat(date_str) if date_str else date_class.today()
        except (ValueError, TypeError):
            target_date = date_class.today()
        sessions_status = {}
        daily_sessions = DailyAttendanceSession.objects.filter(
            section_id__in=all_section_ids,
            date=target_date
        ).select_related('assigned_to', 'assigned_to__user')
        
        session_id_to_section: dict = {}
        for session in daily_sessions:
            # Check if session has any attendance records (marked)
            has_records = session.records.exists()
            sessions_status[session.section_id] = {
                'session_id': session.id,
                'is_locked': session.is_locked,
                'has_attendance': has_records,
                'assigned_to': {
                    'id': session.assigned_to.id,
                    'name': session.assigned_to.user.get_full_name() if session.assigned_to.user else '',
                    'staff_id': session.assigned_to.staff_id
                } if session.assigned_to else None,
                'unlock_request_status': None,
                'unlock_request_hod_status': None,
            }
            session_id_to_section[session.id] = session.section_id

        # Attach unlock request status for each session in one query
        if session_id_to_section:
            from .models import DailyAttendanceUnlockRequest as _DUR
            unlock_qs = _DUR.objects.filter(
                session_id__in=session_id_to_section.keys()
            ).order_by('-requested_at').values('session_id', 'status', 'hod_status')
            seen_sessions: set = set()
            for row in unlock_qs:
                sid = row['session_id']
                if sid in seen_sessions:
                    continue
                seen_sessions.add(sid)
                sec_id = session_id_to_section.get(sid)
                if sec_id and sec_id in sessions_status:
                    sessions_status[sec_id]['unlock_request_status'] = row['status']
                    sessions_status[sec_id]['unlock_request_hod_status'] = row['hod_status']
        
        # Group by section
        sections_data = {}
        for student in students:
            section_id = student.section_id
            if section_id not in sections_data:
                # Check if this is an advisor section or assigned section
                is_advisor = section_id in advisor_section_ids
                is_assigned = section_id in assigned_section_ids
                
                # Extract department information
                department_name = ''
                department_short_name = ''
                if student.section and student.section.batch and student.section.batch.course and student.section.batch.course.department:
                    dept = student.section.batch.course.department
                    department_name = dept.name
                    department_short_name = dept.short_name or dept.code or dept.name
                
                # Get session status
                session_status = sessions_status.get(section_id, {})
                
                sections_data[section_id] = {
                    'section_id': section_id,
                    'section_name': student.section.name if student.section else '',
                    'batch_name': student.section.batch.name if student.section and student.section.batch else '',
                    'department_name': department_name,
                    'department_short_name': department_short_name,
                    'is_advisor': is_advisor,
                    'is_assigned_via_swap': is_assigned,
                    'session_status': session_status,
                    'students': []
                }
            
            sections_data[section_id]['students'].append({
                'id': student.id,
                'reg_no': student.reg_no,
                'name': student.user.get_full_name() if student.user else '',
                'username': student.user.username if student.user else '',
                'section_id': section_id,
            })
        
        return Response({
            'sections': list(sections_data.values()),
            'total_sections': len(sections_data),
            'total_students': students.count()
        })


def _parse_daily_bulk_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value or '').strip()
    if not text:
        return None
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%d.%m.%Y'):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except Exception:
            continue
    try:
        return datetime.date.fromisoformat(text)
    except Exception:
        return None


def _normalize_daily_bulk_status(value):
    text = str(value or '').strip().upper()
    if not text:
        return None
    mapping = {
        'P': 'P',
        'PRESENT': 'P',
        'A': 'A',
        'ABSENT': 'A',
        'OD': 'OD',
        'ON DUTY': 'OD',
        'ONDUTY': 'OD',
        'LEAVE': 'LEAVE',
        'L': 'LEAVE',
    }
    return mapping.get(text, '__INVALID__')


def _excel_status_from_code(status_code):
    mapping = {
        'P': 'Present',
        'A': 'Absent',
        'OD': 'OD',
        'LEAVE': 'Leave',
        'L': 'Leave',
    }
    return mapping.get((status_code or '').upper(), 'Present')


def _daily_bulk_access_context(user, staff_profile):
    from .models import SectionAdvisor, DailyAttendanceSession, DepartmentRole, TeachingAssignment

    perms = get_user_permissions(user)
    can_view_all = 'analytics.view_all_analytics' in perms or user.is_superuser
    can_view_department = 'analytics.view_department_analytics' in perms or can_view_all

    advisor_section_ids = set(SectionAdvisor.objects.filter(
        advisor=staff_profile,
        is_active=True,
    ).values_list('section_id', flat=True))

    assigned_section_ids = set(DailyAttendanceSession.objects.filter(
        assigned_to=staff_profile,
    ).values_list('section_id', flat=True).distinct())

    department_ids = set()
    if can_view_department:
        department_ids |= set(DepartmentRole.objects.filter(
            staff=staff_profile,
            is_active=True,
        ).values_list('department_id', flat=True))
        department_ids |= set(TeachingAssignment.objects.filter(
            staff=staff_profile,
            is_active=True,
        ).values_list('section__batch__course__department_id', flat=True))
        department_ids = set([d for d in department_ids if d])

    return {
        'can_view_all': can_view_all,
        'can_view_department': can_view_department,
        'advisor_section_ids': advisor_section_ids,
        'assigned_section_ids': assigned_section_ids,
        'department_ids': department_ids,
    }


def _can_access_daily_bulk_section(user, staff_profile, section, start_date=None, end_date=None):
    from .models import DailyAttendanceSession

    ctx = _daily_bulk_access_context(user, staff_profile)
    if ctx['can_view_all']:
        return True
    if section.id in ctx['advisor_section_ids']:
        return True

    if ctx['can_view_department'] and section.batch and section.batch.course:
        if section.batch.course.department_id in ctx['department_ids']:
            return True

    assigned_qs = DailyAttendanceSession.objects.filter(
        section=section,
        assigned_to=staff_profile,
    )
    if start_date:
        assigned_qs = assigned_qs.filter(date__gte=start_date)
    if end_date:
        assigned_qs = assigned_qs.filter(date__lte=end_date)
    if assigned_qs.exists():
        return True

    return section.id in ctx['assigned_section_ids']


class BulkAttendanceSectionsView(APIView):
    """List daily-attendance sections accessible for bulk Excel operations."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')

        from .models import Section

        ctx = _daily_bulk_access_context(user, staff_profile)
        if ctx['can_view_all']:
            qs = Section.objects.all()
        else:
            section_ids = set(ctx['advisor_section_ids']) | set(ctx['assigned_section_ids'])
            if ctx['can_view_department'] and ctx['department_ids']:
                dept_ids = list(ctx['department_ids'])
                dept_sections = Section.objects.filter(
                    batch__course__department_id__in=dept_ids
                ).values_list('id', flat=True)
                section_ids |= set(dept_sections)
            if not section_ids:
                return Response([])
            qs = Section.objects.filter(id__in=list(section_ids))

        sections = qs.select_related('batch', 'batch__course__department').order_by(
            'batch__course__department__short_name', 'batch__name', 'name'
        )

        data = []
        for s in sections:
            dept = getattr(getattr(getattr(s, 'batch', None), 'course', None), 'department', None)
            data.append({
                'section_id': s.id,
                'section_name': s.name,
                'batch_name': s.batch.name if s.batch else '',
                'department_name': dept.name if dept else '',
                'department_short_name': (dept.short_name or dept.code or dept.name) if dept else '',
            })
        return Response(data)


class BulkAttendanceDownloadView(APIView):
    """Download daily attendance Excel template for a section and date range."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')

        section_id = request.query_params.get('section_id')
        start_date = _parse_daily_bulk_date(request.query_params.get('start_date'))
        end_date = _parse_daily_bulk_date(request.query_params.get('end_date'))

        if not section_id:
            return Response({'error': 'section_id is required'}, status=400)
        if not start_date or not end_date:
            return Response({'error': 'Valid start_date and end_date are required'}, status=400)
        if end_date < start_date:
            return Response({'error': 'end_date must be greater than or equal to start_date'}, status=400)

        try:
            section_id_int = int(section_id)
        except Exception:
            return Response({'error': 'Invalid section_id'}, status=400)

        from .models import Section, StudentProfile, DailyAttendanceRecord

        try:
            section = Section.objects.select_related('batch', 'batch__course__department').get(id=section_id_int)
        except Section.DoesNotExist:
            return Response({'error': 'Section not found'}, status=404)

        if not _can_access_daily_bulk_section(user, staff_profile, section, start_date, end_date):
            raise PermissionDenied('You do not have access to this section for bulk attendance')

        try:
            from openpyxl import Workbook
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response({'error': 'Excel support not available. Please install openpyxl.'}, status=500)

        students = list(
            StudentProfile.objects.filter(section=section)
            .exclude(status__in=['INACTIVE', 'DEBAR'])
            .select_related('user')
            .order_by('reg_no')
        )

        days = []
        cur = start_date
        while cur <= end_date:
            days.append(cur)
            cur += datetime.timedelta(days=1)

        # Apply excluded_dates filter
        excluded_dates_raw = request.query_params.get('excluded_dates', '')
        if excluded_dates_raw:
            excluded_set = set()
            for d in excluded_dates_raw.split(','):
                parsed = _parse_daily_bulk_date(d.strip())
                if parsed:
                    excluded_set.add(parsed)
            days = [d for d in days if d not in excluded_set]

        if not days:
            return Response({'error': 'No dates remaining after exclusions'}, status=400)

        existing = DailyAttendanceRecord.objects.filter(
            session__section=section,
            session__date__gte=start_date,
            session__date__lte=end_date,
            student_id__in=[s.id for s in students],
        ).values('student_id', 'session__date', 'status')
        status_map = {
            (row['student_id'], row['session__date']): _excel_status_from_code(row['status'])
            for row in existing
        }

        wb = Workbook()
        ws = wb.active
        ws.title = 'Daily Attendance'

        headers = ['Register Number', 'Name'] + [d.strftime('%Y-%m-%d') for d in days]
        ws.append(headers)

        for student in students:
            if student.user:
                name = student.user.get_full_name().strip() or student.user.username
            else:
                name = student.reg_no
            row = [student.reg_no, name]
            for d in days:
                row.append(status_map.get((student.id, d), 'Present'))
            ws.append(row)

        lists_ws = wb.create_sheet('_Lists')
        lists_ws.sheet_state = 'hidden'
        for idx, label in enumerate(['Present', 'Absent', 'OD', 'Leave'], start=1):
            lists_ws.cell(row=idx, column=1, value=label)

        if days:
            dv = DataValidation(type='list', formula1='=_Lists!$A$1:$A$4', allow_blank=False)
            ws.add_data_validation(dv)
            start_col = 3
            end_col = 2 + len(days)
            end_row = max(2, len(students) + 1)
            dv.add(f"{get_column_letter(start_col)}2:{get_column_letter(end_col)}{end_row}")

        ws.freeze_panes = 'C2'
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 32
        for i in range(3, 3 + len(days)):
            ws.column_dimensions[get_column_letter(i)].width = 14

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        filename = f"daily_attendance_{section.name}_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
        response = HttpResponse(
            out.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class BulkAttendanceLockedSessionsView(APIView):
    """List locked daily attendance sessions for a section (optionally in date range)."""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')

        section_id = request.query_params.get('section_id')
        start_date = _parse_daily_bulk_date(request.query_params.get('start_date'))
        end_date = _parse_daily_bulk_date(request.query_params.get('end_date'))

        if not section_id:
            return Response({'error': 'section_id is required'}, status=400)

        if request.query_params.get('start_date') and not start_date:
            return Response({'error': 'Invalid start_date'}, status=400)
        if request.query_params.get('end_date') and not end_date:
            return Response({'error': 'Invalid end_date'}, status=400)
        if start_date and end_date and end_date < start_date:
            return Response({'error': 'end_date must be greater than or equal to start_date'}, status=400)

        try:
            section_id_int = int(section_id)
        except Exception:
            return Response({'error': 'Invalid section_id'}, status=400)

        from .models import Section, DailyAttendanceSession, DailyAttendanceUnlockRequest

        try:
            section = Section.objects.select_related('batch', 'batch__course__department').get(id=section_id_int)
        except Section.DoesNotExist:
            return Response({'error': 'Section not found'}, status=404)

        if not _can_access_daily_bulk_section(user, staff_profile, section, start_date, end_date):
            raise PermissionDenied('You do not have access to this section for bulk attendance')

        sessions_qs = DailyAttendanceSession.objects.filter(section=section, is_locked=True)
        if start_date:
            sessions_qs = sessions_qs.filter(date__gte=start_date)
        if end_date:
            sessions_qs = sessions_qs.filter(date__lte=end_date)

        sessions = list(sessions_qs.order_by('date').only('id', 'date'))
        session_ids = [s.id for s in sessions]

        latest_requests = {}
        if session_ids:
            for unlock_request in DailyAttendanceUnlockRequest.objects.filter(
                session_id__in=session_ids
            ).order_by('session_id', '-requested_at'):
                if unlock_request.session_id not in latest_requests:
                    latest_requests[unlock_request.session_id] = unlock_request

        results = []
        for sess in sessions:
            unlock_request = latest_requests.get(sess.id)
            results.append({
                'session_id': sess.id,
                'section_id': section.id,
                'section_name': str(section),
                'date': sess.date.isoformat(),
                'unlock_request_id': unlock_request.id if unlock_request else None,
                'unlock_request_status': unlock_request.status if unlock_request else None,
                'unlock_request_hod_status': unlock_request.hod_status if unlock_request else None,
            })

        return Response({'results': results, 'count': len(results)})


class BulkAttendanceImportView(APIView):
    """Import daily attendance from Excel or JSON preview data for a section."""
    permission_classes = (IsAuthenticated,)
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def post(self, request):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')

        # Accept JSON body (from preview) or multipart file upload
        json_attendance = request.data.get('attendance') if isinstance(request.data.get('attendance'), list) else None
        uploaded = request.FILES.get('file')
        section_id = request.data.get('section_id')
        lock_session_raw = request.data.get('lock_session', False)
        if isinstance(lock_session_raw, bool):
            lock_session = lock_session_raw
        else:
            lock_session = str(lock_session_raw).strip().lower() in ('1', 'true', 'yes', 'y')

        if not section_id:
            return Response({'error': 'section_id is required'}, status=400)
        if json_attendance is None and not uploaded:
            return Response({'error': 'Either file or attendance JSON is required'}, status=400)

        try:
            section_id_int = int(section_id)
        except Exception:
            return Response({'error': 'Invalid section_id'}, status=400)

        from django.db import transaction
        from .models import Section, StudentProfile, DailyAttendanceSession, DailyAttendanceRecord

        try:
            section = Section.objects.select_related('batch', 'batch__course__department').get(id=section_id_int)
        except Section.DoesNotExist:
            return Response({'error': 'Section not found'}, status=404)

        if not _can_access_daily_bulk_section(user, staff_profile, section):
            raise PermissionDenied('You do not have access to this section for bulk attendance import')

        from django.db import transaction
        from .models import Section, StudentProfile, DailyAttendanceSession, DailyAttendanceRecord

        students = list(
            StudentProfile.objects.filter(section=section)
            .exclude(status__in=['INACTIVE', 'DEBAR'])
            .select_related('user')
        )
        students_by_reg = {str(s.reg_no).strip().upper(): s for s in students if s.reg_no}

        # ── Build rows_data from JSON (preview path) or Excel (file path) ──────────
        # rows_data is a list of dicts: {reg_no, dates: {date_str: status}, remarks: {date_str: remark}}
        rows_data = []
        errors = []

        if json_attendance is not None:
            # JSON path: sent from the frontend preview
            for entry in json_attendance:
                reg_no_raw = entry.get('reg_no', '')
                dates_map = entry.get('dates', {})    # {date_str: status}
                remarks_map = entry.get('remarks', {})  # {date_str: remark}
                rows_data.append({
                    'reg_no': str(reg_no_raw).strip(),
                    'dates': dates_map,
                    'remarks': remarks_map,
                })
        else:
            # Excel file path
            try:
                from openpyxl import load_workbook
            except ImportError:
                return Response({'error': 'Excel support not available. Please install openpyxl.'}, status=500)

            try:
                wb = load_workbook(uploaded, data_only=True)
                ws = wb.active
            except Exception as exc:
                return Response({'error': f'Invalid Excel file: {exc}'}, status=400)

            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            if max_row < 2 or max_col < 3:
                return Response({'error': 'Excel file must contain header row and at least one date column (starting column C)'}, status=400)

            headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
            # Format: col A = Register Number, col B = Name, col C onward = dates
            date_columns = []
            for col in range(3, max_col + 1):
                parsed_date = _parse_daily_bulk_date(headers[col - 1])
                if parsed_date:
                    date_columns.append((col, parsed_date))

            if not date_columns:
                return Response({'error': 'No valid date columns found from column C onward'}, status=400)

            for row in range(2, max_row + 1):
                reg_no_raw = ws.cell(row=row, column=1).value
                if reg_no_raw is None or str(reg_no_raw).strip() == '':
                    continue
                # Excel may auto-cast integer reg-nos to float
                if isinstance(reg_no_raw, float) and reg_no_raw == int(reg_no_raw):
                    reg_no_raw = int(reg_no_raw)
                dates_map = {}
                for col, att_date in date_columns:
                    raw_status = ws.cell(row=row, column=col).value
                    dates_map[att_date.isoformat()] = raw_status
                rows_data.append({'reg_no': str(reg_no_raw).strip(), 'dates': dates_map, 'remarks': {}})

        # ── Shared save loop ────────────────────────────────────────────────────────
        created_count = 0
        updated_count = 0
        locked_count = 0
        period_records_updated = 0
        processed_session_ids = set()
        skipped_locked_sessions = {}
        session_meta = {}
        period_overrides = {}

        with transaction.atomic():
            for entry in rows_data:
                reg_no_val = entry['reg_no']
                if not reg_no_val:
                    continue

                student = students_by_reg.get(reg_no_val.upper())
                if student is None:
                    if len(errors) < 200:
                        errors.append(f'Student not found in section: {reg_no_val}')
                    continue

                dates_map = entry.get('dates', {})
                remarks_map = entry.get('remarks', {})

                for date_str, raw_status in dates_map.items():
                    norm_status = _normalize_daily_bulk_status(raw_status)
                    if norm_status is None:
                        continue
                    if norm_status == '__INVALID__':
                        if len(errors) < 200:
                            errors.append(f"{reg_no_val}, {date_str}: invalid status '{raw_status}'")
                        continue

                    att_date = _parse_daily_bulk_date(date_str)
                    if att_date is None:
                        if len(errors) < 200:
                            errors.append(f'{reg_no_val}: could not parse date "{date_str}"')
                        continue

                    session, _ = DailyAttendanceSession.objects.get_or_create(
                        section=section,
                        date=att_date,
                        defaults={'created_by': staff_profile},
                    )

                    if session.assigned_to and session.assigned_to != staff_profile and not user.is_superuser:
                        if len(errors) < 200:
                            errors.append(f'{reg_no_val}, {date_str}: session assigned to another staff; skipped')
                        continue

                    if session.is_locked:
                        skipped_locked_sessions[session.id] = {
                            'session_id': session.id,
                            'section_id': section.id,
                            'section_name': str(section),
                            'date': att_date.isoformat(),
                        }
                        continue

                    processed_session_ids.add(session.id)
                    session_meta[session.id] = {
                        'section_id': section.id,
                        'date': att_date,
                    }

                    period_status = None
                    if norm_status in ('OD', 'LEAVE'):
                        period_status = norm_status
                    elif norm_status == 'LATE':
                        period_status = 'P'
                    if period_status:
                        session_overrides = period_overrides.setdefault(session.id, {})
                        session_overrides[student.id] = period_status

                    remark = str(remarks_map.get(date_str, '') or '').strip()

                    record, created = DailyAttendanceRecord.objects.get_or_create(
                        session=session,
                        student=student,
                        defaults={
                            'status': norm_status,
                            'marked_by': staff_profile,
                            'remarks': remark or None,
                        },
                    )
                    if created:
                        created_count += 1
                        continue

                    changed = record.status != norm_status or record.marked_by_id != staff_profile.id
                    if remark:
                        changed = changed or (record.remarks or '') != remark
                    if changed:
                        record.status = norm_status
                        record.marked_by = staff_profile
                        if remark:
                            record.remarks = remark
                        record.save(update_fields=['status', 'marked_by', 'marked_at', 'remarks'])
                        updated_count += 1

            if period_overrides:
                from .models import PeriodAttendanceSession, PeriodAttendanceRecord

                for session_id, student_statuses in period_overrides.items():
                    meta = session_meta.get(session_id)
                    if not meta:
                        continue
                    period_sessions = PeriodAttendanceSession.objects.filter(
                        section_id=meta['section_id'],
                        date=meta['date'],
                    )
                    for student_id, period_status in student_statuses.items():
                        period_records_updated += PeriodAttendanceRecord.objects.filter(
                            session__in=period_sessions,
                            student_id=student_id,
                        ).update(status=period_status)

            if lock_session and processed_session_ids:
                locked_count = DailyAttendanceSession.objects.filter(
                    id__in=list(processed_session_ids),
                    is_locked=False,
                ).update(is_locked=True)

        latest_request_session_ids = set(skipped_locked_sessions.keys())
        if lock_session and processed_session_ids:
            latest_request_session_ids.update(processed_session_ids)

        latest_requests = {}
        if latest_request_session_ids:
            from .models import DailyAttendanceUnlockRequest

            for unlock_request in DailyAttendanceUnlockRequest.objects.filter(
                session_id__in=list(latest_request_session_ids)
            ).order_by('session_id', '-requested_at'):
                if unlock_request.session_id not in latest_requests:
                    latest_requests[unlock_request.session_id] = unlock_request

        locked_session_list = []
        if lock_session and processed_session_ids:
            for session_id in sorted(processed_session_ids, key=lambda item: session_meta[item]['date']):
                meta = session_meta[session_id]
                unlock_request = latest_requests.get(session_id)
                locked_session_list.append({
                    'session_id': session_id,
                    'section_id': meta['section_id'],
                    'section_name': str(section),
                    'date': meta['date'].isoformat(),
                    'unlock_request_id': unlock_request.id if unlock_request else None,
                    'unlock_request_status': unlock_request.status if unlock_request else None,
                    'unlock_request_hod_status': unlock_request.hod_status if unlock_request else None,
                })

        if skipped_locked_sessions:
            skipped_locked_session_list = []
            for session_id, session_info in sorted(skipped_locked_sessions.items(), key=lambda item: item[1]['date']):
                unlock_request = latest_requests.get(session_id)
                skipped_locked_session_list.append({
                    **session_info,
                    'unlock_request_id': unlock_request.id if unlock_request else None,
                    'unlock_request_status': unlock_request.status if unlock_request else None,
                    'unlock_request_hod_status': unlock_request.hod_status if unlock_request else None,
                })
            errors.append(f'Skipped {len(skipped_locked_session_list)} locked sessions.')
        else:
            skipped_locked_session_list = []

        return Response({
            'created': created_count,
            'updated': updated_count,
            'locked': locked_count,
            'period_records_updated': period_records_updated,
            'locked_sessions': locked_session_list,
            'skipped_locked_sessions': skipped_locked_session_list,
            'errors': errors,
        })


class MyClassAttendanceAnalyticsView(APIView):
    """
    Get period attendance analytics for advisor's assigned sections.
    Requires 'analytics.view_class_analytics' permission or advisor role.
    """
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):
        try:
            user = request.user
            staff_profile = getattr(user, 'staff_profile', None)
            if not staff_profile:
                return Response({'error': 'Staff profile required', 'sections': []}, status=400)
            
            from .models import SectionAdvisor
            from datetime import date as date_class
            
            # Get query parameters
            date_str = request.query_params.get('date')
            date_from_str = request.query_params.get('date_from')
            date_to_str = request.query_params.get('date_to')
            complete = request.query_params.get('complete', '').lower() in ('true', '1')
            view_mode = request.query_params.get('view_mode', 'class')
            
            if view_mode != 'class':
                return Response({'error': 'This endpoint only supports view_mode=class'}, status=400)
            
            # Parse date range
            try:
                if date_from_str and date_to_str:
                    date_from = date_class.fromisoformat(date_from_str)
                    date_to = date_class.fromisoformat(date_to_str)
                elif date_str:
                    date_from = date_to = date_class.fromisoformat(date_str)
                else:
                    date_from = date_to = date_class.today()
            except:
                date_from = date_to = date_class.today()
            target_date = date_from  # for backward compat in response
            is_range = complete or (date_from != date_to)
            
            # Get advisor's assigned sections
            advisor_sections = SectionAdvisor.objects.filter(
                advisor=staff_profile,
                is_active=True
            ).values_list('section_id', flat=True)
            
            if not advisor_sections:
                return Response({
                    'sections': [], 
                    'message': 'No assigned sections found',
                    'debug': {
                        'staff_profile_id': staff_profile.id if staff_profile else None,
                        'total_advisors': SectionAdvisor.objects.count(),
                        'active_advisors': SectionAdvisor.objects.filter(is_active=True).count()
                    }
                })
            
            # Get period attendance sessions for these sections on date range (or all if complete)
            _period_base = PeriodAttendanceSession.objects.filter(section_id__in=list(advisor_sections))
            sessions = (_period_base if complete else _period_base.filter(date__range=(date_from, date_to))).select_related(
                'section', 'section__batch', 'period', 'period__template',
                'teaching_assignment', 'teaching_assignment__subject',
                'teaching_assignment__curriculum_row', 'teaching_assignment__elective_subject'
            )
            
            # Also get daily attendance sessions for these sections on date range (or all if complete)
            from .models import DailyAttendanceSession, DailyAttendanceRecord
            _daily_base = DailyAttendanceSession.objects.filter(section_id__in=list(advisor_sections))
            daily_sessions = (_daily_base if complete else _daily_base.filter(date__range=(date_from, date_to))).select_related('section', 'section__batch')
            
            # Build a map of daily attendance data by section_id (aggregated across date range)
            daily_attendance_map = {}
            for daily_session in daily_sessions:
                section_id = daily_session.section_id
                records = DailyAttendanceRecord.objects.filter(session=daily_session)
                present_count = records.filter(status__in=['P', 'LATE', 'OD']).count()
                absent_count = records.filter(status='A').count()
                leave_count_daily = records.filter(status='LEAVE').count()
                od_count_daily = records.filter(status='OD').count()
                
                # Get section details
                section = daily_session.section
                section_name = section.name if section else 'Unknown'
                department_name = 'Unknown'
                batch_name = 'Unknown'
                
                if section and section.batch:
                    batch_name = section.batch.name or 'Unknown'
                    if section.batch.course and section.batch.course.department:
                        department_name = section.batch.course.department.name or 'Unknown'

                if section_id not in daily_attendance_map:
                    daily_attendance_map[section_id] = {
                        'session_id': daily_session.id,
                        'section_id': section_id,
                        'section_name': section_name,
                        'department': department_name,
                        'department_name': department_name,
                        'batch': batch_name,
                        'batch_name': batch_name,
                        'present_count': 0,
                        'absent_count': 0,
                        'leave_count': 0,
                        'od_count': 0,
                        'total_count': 0,
                        'is_locked': False,
                        'attendance_marked': False,
                        'days_count': 0,
                    }
                # Aggregate
                entry = daily_attendance_map[section_id]
                entry['present_count'] += present_count
                entry['absent_count'] += absent_count
                entry['leave_count'] += leave_count_daily
                entry['od_count'] += od_count_daily
                entry['total_count'] += present_count + absent_count + leave_count_daily
                entry['is_locked'] = entry['is_locked'] or getattr(daily_session, 'is_locked', False)
                entry['attendance_marked'] = entry['attendance_marked'] or ((present_count + absent_count) > 0)
                entry['days_count'] += 1
            
            sections_data = []
            session_errors = []

            # Pre-build a cache: template_id -> {slot_index -> actual_period_number}
            # Actual period number = rank among non-break, non-lunch slots ordered by index
            from timetable.models import TimetableSlot
            _period_num_cache: dict = {}  # (template_id, slot_index) -> actual period number
            _template_ids = set()
            for s in sessions:
                if s.period and s.period.template_id:
                    _template_ids.add(s.period.template_id)
            for tmpl_id in _template_ids:
                slots = TimetableSlot.objects.filter(
                    template_id=tmpl_id,
                    is_break=False,
                    is_lunch=False
                ).order_by('index')
                for rank, slot in enumerate(slots, start=1):
                    _period_num_cache[(tmpl_id, slot.index)] = rank

            def get_actual_period_number(period_obj):
                """Return the actual teaching period number (breaks excluded)."""
                if not period_obj:
                    return 1
                template_id = period_obj.template_id
                idx = period_obj.index
                return _period_num_cache.get((template_id, idx), idx)

            # When date range spans multiple days, aggregate by (section_id, subject_code, period_number)
            period_agg_map = {}
            for session in sessions:
                try:
                    # Calculate attendance counts safely
                    records = session.records.all()
                    present_count = sum(1 for r in records if r.status in ['P', 'LATE', 'OD'])
                    absent_count = sum(1 for r in records if r.status == 'A')
                    leave_count = sum(1 for r in records if r.status == 'LEAVE')
                    od_count = sum(1 for r in records if r.status == 'OD')
                    
                    section = session.section
                    department_name = 'Unknown'
                    batch_name = 'Unknown'
                    if section and section.batch:
                        batch_name = section.batch.name or 'Unknown'
                        if section.batch.course and section.batch.course.department:
                            department_name = section.batch.course.department.name or 'Unknown'
                    
                    subject_name = 'N/A'
                    subject_code = 'N/A'
                    if session.teaching_assignment:
                        ta = session.teaching_assignment
                        if ta.subject:
                            subject_name = ta.subject.name
                            subject_code = ta.subject.code
                        elif ta.curriculum_row:
                            subject_name = ta.curriculum_row.course_name or 'N/A'
                            subject_code = ta.curriculum_row.course_code or 'N/A'
                        elif ta.elective_subject:
                            subject_name = ta.elective_subject.course_name or 'N/A'
                            subject_code = ta.elective_subject.course_code or 'N/A'
                        elif ta.custom_subject:
                            subject_name = ta.custom_subject
                            subject_code = ta.custom_subject
                    
                    period_num = get_actual_period_number(session.period)
                    agg_key = (session.section_id, subject_code, period_num)

                    if agg_key not in period_agg_map:
                        period_agg_map[agg_key] = {
                            'id': session.id,
                            'session_id': session.id,
                            'section_id': session.section_id,
                            'section_name': section.name if section else 'Unknown',
                            'department': department_name,
                            'department_name': department_name,
                            'batch': batch_name,
                            'batch_name': batch_name,
                            'period_number': period_num,
                            'subject_name': subject_name,
                            'subject_code': subject_code,
                            'start_time': session.period.start_time.strftime('%H:%M') if (session.period and session.period.start_time) else 'N/A',
                            'end_time': session.period.end_time.strftime('%H:%M') if (session.period and session.period.end_time) else 'N/A',
                            'present_count': 0,
                            'absent_count': 0,
                            'leave_count': 0,
                            'od_count': 0,
                            'total_count': 0,
                            'is_locked': False,
                            'attendance_marked': False,
                            'sessions_count': 0,
                        }
                    agg = period_agg_map[agg_key]
                    agg['present_count'] += present_count
                    agg['absent_count'] += absent_count
                    agg['leave_count'] += leave_count
                    agg['od_count'] += od_count
                    agg['total_count'] += present_count + absent_count + leave_count
                    agg['is_locked'] = agg['is_locked'] or getattr(session, 'is_locked', False)
                    agg['attendance_marked'] = agg['attendance_marked'] or ((present_count + absent_count) > 0)
                    agg['sessions_count'] += 1
                except Exception as e:
                    session_errors.append({
                        'session_id': getattr(session, 'id', 'unknown'),
                        'error': str(e),
                        'error_type': type(e).__name__
                    })
                    continue

            for agg in period_agg_map.values():
                tc = agg['total_count']
                agg['attendance_percentage'] = round((agg['present_count'] / tc) * 100, 1) if tc > 0 else 0
                agg['date'] = date_from.isoformat()
                sections_data.append(agg)
            
            return Response({
                'sections': sections_data,
                'daily_attendance': daily_attendance_map,
                'date': date_from.isoformat(),
                'date_from': date_from.isoformat(),
                'date_to': date_to.isoformat(),
                'is_range': is_range,
                'total_sections': len(set(s['section_id'] for s in sections_data)) if sections_data else 0,
                'total_periods': len(sections_data),
                'debug': {
                    'advisor_sections_count': len(list(advisor_sections)),
                    'sessions_found': sessions.count(),
                    'daily_sessions_found': daily_sessions.count(),
                    'staff_profile_id': staff_profile.id,
                    'session_errors': session_errors if session_errors else []
                }
            })
        
        except Exception as e:
            logger.exception('AttendanceAnalyticsView unhandled error for user=%s', getattr(request.user, 'id', None))
            return Response({
                'error': 'Server error. Please try again.',
                'sections': [],
            }, status=500)


class DailyAttendanceView(APIView):
    """
    GET: Fetch daily attendance for a section on a date
    POST: Save/update daily attendance for students
    """
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')
        
        from .models import SectionAdvisor, DailyAttendanceSession, DailyAttendanceRecord, StudentProfile
        from datetime import date as date_class
        
        section_id = request.query_params.get('section_id')
        date_str = request.query_params.get('date')
        
        if not section_id:
            return Response({'error': 'section_id required'}, status=400)
        
        try:
            target_date = date_class.fromisoformat(date_str) if date_str else date_class.today()
        except Exception:
            target_date = date_class.today()
        
        # Check if a session exists for this section and date
        existing_session = DailyAttendanceSession.objects.filter(
            section_id=section_id,
            date=target_date
        ).select_related('assigned_to', 'created_by').first()
        
        # Allow access if any of:
        #   1. Session is assigned to this staff member (after swap)
        #   2. Section advisor for this section
        #   3. Has a teaching assignment for this section (period teacher)
        #   4. Has department-level or all-level analytics permission
        #   5. Superuser
        perms = get_user_permissions(user)
        can_view_all = 'analytics.view_all_analytics' in perms or user.is_superuser
        can_view_department = 'analytics.view_department_analytics' in perms or can_view_all

        # Check if session is specifically assigned to this user
        is_assigned = existing_session and existing_session.assigned_to == staff_profile

        is_advisor = SectionAdvisor.objects.filter(
            advisor=staff_profile,
            section_id=section_id,
            is_active=True
        ).exists()

        has_teaching = False
        has_elective_for_section = False
        if not is_advisor and not can_view_department and not is_assigned:
            from .models import TeachingAssignment
            # Direct section assignment (regular subjects)
            has_teaching = TeachingAssignment.objects.filter(
                staff=staff_profile,
                section_id=section_id,
                is_active=True
            ).exists()
            # Elective assignment: staff teaches an elective that has students enrolled
            # from the requested section (students from cross-dept/cross-section groups)
            if not has_teaching:
                has_elective_for_section = TeachingAssignment.objects.filter(
                    staff=staff_profile,
                    elective_subject__isnull=False,
                    is_active=True,
                    elective_subject__choices__student__section_id=section_id,
                    elective_subject__choices__is_active=True,
                ).exists()

        if not (is_assigned or is_advisor or has_teaching or has_elective_for_section or can_view_department):
            raise PermissionDenied('You are not assigned to this section')
        
        # Only advisors (or admins or assigned staff) may auto-create the session.
        # Period-only teachers just read whatever session already exists.
        if is_advisor or can_view_department or is_assigned:
            session, created = DailyAttendanceSession.objects.get_or_create(
                section_id=section_id,
                date=target_date,
                defaults={'created_by': staff_profile}
            )
        else:
            session = existing_session
            # No daily session yet means no overrides – return empty gracefully
            if session is None:
                students = StudentProfile.objects.filter(section_id=section_id).exclude(status__in=['INACTIVE', 'DEBAR']).select_related('user').order_by('reg_no')
                return Response({
                    'session_id': None,
                    'section_id': section_id,
                    'date': target_date.isoformat(),
                    'is_locked': False,
                    'unlock_request_status': None,
                    'unlock_request_id': None,
                    'assigned_to': None,
                    'created_by': None,
                    'students': [
                        {
                            'student_id': s.id,
                            'reg_no': s.reg_no,
                            'name': s.user.get_full_name() if s.user else '',
                            'username': s.user.username if s.user else '',
                            'status': 'P',
                            'remarks': '',
                            'marked_at': None,
                        }
                        for s in students
                    ],
                    'total_students': students.count(),
                })
        
        # ══════════════════════════════════════════════════════════════════════
        # ACCESS CONTROL FOR SWAPPED SESSIONS
        # If session has been assigned to another staff:
        #   - Assigned staff can view and edit
        #   - Original advisor can view (read-only) but not edit
        #   - Admin/superuser can do everything
        # ══════════════════════════════════════════════════════════════════════
        is_original_advisor = is_advisor and session.assigned_to and session.assigned_to != staff_profile
        
        if session.assigned_to:
            # Session has been swapped/assigned to another staff member
            # Allow original advisor to VIEW but they can't EDIT (checked in POST)
            if not (session.assigned_to == staff_profile or is_advisor or can_view_all):
                raise PermissionDenied('This attendance has been assigned to another staff member')
        
        # Get existing records
        records = DailyAttendanceRecord.objects.filter(session=session).select_related('student', 'student__user')
        records_map = {rec.student_id: rec for rec in records}
        
        # Get all students in section
        students = StudentProfile.objects.filter(section_id=section_id).exclude(status__in=['INACTIVE', 'DEBAR']).select_related('user').order_by('reg_no')
        
        students_data = []
        for student in students:
            record = records_map.get(student.id)
            students_data.append({
                'student_id': student.id,
                'reg_no': student.reg_no,
                'name': student.user.get_full_name() if student.user else '',
                'username': student.user.username if student.user else '',
                'status': record.status if record else 'P',
                'remarks': record.remarks if record else '',
                'marked_at': record.marked_at.isoformat() if record and record.marked_at else None,
                'marked_by': {
                    'id': record.marked_by.id,
                    'name': record.marked_by.user.get_full_name() if record.marked_by.user else '',
                    'staff_id': record.marked_by.staff_id
                } if record and record.marked_by else None,
            })
        
        # Get swap history for this session
        from .models import DailyAttendanceSwapRecord
        swap_records = DailyAttendanceSwapRecord.objects.filter(
            session=session
        ).select_related('assigned_by', 'assigned_by__user', 'assigned_to', 'assigned_to__user').order_by('-assigned_at')
        
        swap_history = []
        for swap in swap_records:
            # Better name resolution with fallbacks
            assigned_by_name = ''
            assigned_to_name = ''
            
            if swap.assigned_by:
                if swap.assigned_by.user and swap.assigned_by.user.get_full_name().strip():
                    assigned_by_name = swap.assigned_by.user.get_full_name()
                elif swap.assigned_by.user and (swap.assigned_by.user.first_name or swap.assigned_by.user.last_name):
                    assigned_by_name = f"{swap.assigned_by.user.first_name} {swap.assigned_by.user.last_name}".strip()
                else:
                    assigned_by_name = swap.assigned_by.staff_id or f'Staff ID {swap.assigned_by.id}'
            
            if swap.assigned_to:
                if swap.assigned_to.user and swap.assigned_to.user.get_full_name().strip():
                    assigned_to_name = swap.assigned_to.user.get_full_name()
                elif swap.assigned_to.user and (swap.assigned_to.user.first_name or swap.assigned_to.user.last_name):
                    assigned_to_name = f"{swap.assigned_to.user.first_name} {swap.assigned_to.user.last_name}".strip()
                else:
                    assigned_to_name = swap.assigned_to.staff_id or f'Staff ID {swap.assigned_to.id}'
            
            swap_history.append({
                'id': swap.id,
                'assigned_by': {
                    'id': swap.assigned_by.id,
                    'name': assigned_by_name,
                    'staff_id': swap.assigned_by.staff_id if swap.assigned_by else ''
                } if swap.assigned_by else None,
                'assigned_to': {
                    'id': swap.assigned_to.id,
                    'name': assigned_to_name,
                    'staff_id': swap.assigned_to.staff_id if swap.assigned_to else ''
                } if swap.assigned_to else None,
                'assigned_at': swap.assigned_at.isoformat() if swap.assigned_at else None,
                'reason': swap.reason or ''
            })
        
        # Check for any unlock request for this session (include all statuses so REJECTED is also returned)
        from .models import DailyAttendanceUnlockRequest
        unlock_request = DailyAttendanceUnlockRequest.objects.filter(
            session=session
        ).order_by('-requested_at').first()
        
        return Response({
            'session_id': session.id,
            'section_id': section_id,
            'date': target_date.isoformat(),
            'is_locked': session.is_locked,
            'unlock_request_status': unlock_request.status if unlock_request else None,
            'unlock_request_id': unlock_request.id if unlock_request else None,
            'unlock_request_hod_status': unlock_request.hod_status if unlock_request else None,
            'assigned_to': {
                'id': session.assigned_to.id,
                'name': session.assigned_to.user.get_full_name() if session.assigned_to.user else '',
                'staff_id': session.assigned_to.staff_id
            } if session.assigned_to else None,
            'created_by': {
                'id': session.created_by.id,
                'name': session.created_by.user.get_full_name() if session.created_by.user else '',
                'staff_id': session.created_by.staff_id
            } if session.created_by else None,
            'is_read_only': is_original_advisor,  # True if original advisor viewing after assignment
            'swap_history': swap_history,
            'students': students_data,
            'total_students': len(students_data),
        })
    
    def post(self, request):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')
        
        from .models import SectionAdvisor, DailyAttendanceSession, DailyAttendanceRecord, StaffProfile, DailyAttendanceSwapRecord
        from datetime import date as date_class
        from django.db import transaction
        
        section_id = request.data.get('section_id')
        date_str = request.data.get('date')
        attendance_data = request.data.get('attendance', [])  # List of {student_id, status, remarks}
        taken_by_staff_id = request.data.get('taken_by_staff_id')  # Optional: staff who actually took attendance
        
        if not section_id:
            return Response({'error': 'section_id required'}, status=400)
        
        try:
            target_date = date_class.fromisoformat(date_str) if date_str else date_class.today()
        except Exception:
            target_date = date_class.today()
        
        # Get or create session first to check assignment status
        session, created = DailyAttendanceSession.objects.select_related('assigned_to', 'created_by').get_or_create(
            section_id=section_id,
            date=target_date,
            defaults={'created_by': staff_profile}
        )
        
        # Check if session is locked
        if session.is_locked:
            return Response({'error': 'Attendance is locked for this date'}, status=403)
        
        # ══════════════════════════════════════════════════════════════════════
        # ACCESS CONTROL: Determine who can save based on assignment status
        # ══════════════════════════════════════════════════════════════════════
        perms = get_user_permissions(user)
        can_edit_all = 'analytics.edit_all_analytics' in perms or user.is_superuser
        
        # Check if user is advisor of this section
        is_advisor = SectionAdvisor.objects.filter(
            advisor=staff_profile,
            section_id=section_id,
            is_active=True
        ).exists()
        
        if session.assigned_to:
            # Session has been assigned to another staff
            if session.assigned_to == staff_profile:
                # Assigned staff can save
                can_save = True
            elif is_advisor:
                # Original advisor trying to save - BLOCK IT
                return Response({
                    'error': 'This attendance has been assigned to another staff member. You cannot make changes.',
                    'assigned_to': {
                        'name': session.assigned_to.user.get_full_name() if session.assigned_to.user else '',
                        'staff_id': session.assigned_to.staff_id
                    }
                }, status=403)
            elif can_edit_all:
                # Admin can save
                can_save = True
            else:
                raise PermissionDenied('This attendance has been assigned to another staff member')
        else:
            # No assignment yet - verify user is advisor of this section
            if not (is_advisor or can_edit_all):
                raise PermissionDenied('You are not assigned to this section')
            
            if not (is_advisor or can_edit_all):
                raise PermissionDenied('You are not an advisor of this section')
            can_save = True
        
        # Handle staff swap: if taken_by_staff_id is provided, verify it's valid and set assignment
        marking_staff = staff_profile  # Default to current user
        swap_record_created = False
        swap_staff = None  # Initialize for later use in response message
        if taken_by_staff_id:
            try:
                swap_staff = StaffProfile.objects.select_related('user').get(id=taken_by_staff_id, status='ACTIVE')
                # Verify the swap staff is from the same department (security check)
                current_dept = staff_profile.get_current_department()
                swap_dept = swap_staff.get_current_department()
                if current_dept and swap_dept and current_dept.id == swap_dept.id:
                    marking_staff = swap_staff
                    # Set the assignment in the session
                    old_assigned_to = session.assigned_to
                    session.assigned_to = swap_staff
                    session.save(update_fields=['assigned_to'])
                    
                    # Refresh session from database to ensure we have the latest data
                    session.refresh_from_db()
                    
                    # Create swap record for audit trail (only if this is a new assignment or reassignment)
                    if old_assigned_to != swap_staff:
                        DailyAttendanceSwapRecord.objects.create(
                            session=session,
                            assigned_by=staff_profile,
                            assigned_to=swap_staff,
                            reason=f'Attendance assigned from {staff_profile.user.get_full_name() if staff_profile.user else staff_profile.staff_id} to {swap_staff.user.get_full_name() if swap_staff.user else swap_staff.staff_id}'
                        )
                        swap_record_created = True
                else:
                    return Response({'error': 'Swap staff must be from the same department'}, status=400)
            except StaffProfile.DoesNotExist:
                return Response({'error': 'Invalid staff selected for swap'}, status=400)
            except Exception as e:
                return Response({'error': f'Failed to assign staff: {str(e)}'}, status=500)
        
        try:
            with transaction.atomic():
                # Update or create records
                updated_students = []
                for item in attendance_data:
                    student_id = item.get('student_id')
                    status = item.get('status', 'P')
                    remarks = item.get('remarks', '')
                    
                    if student_id:
                        DailyAttendanceRecord.objects.update_or_create(
                            session=session,
                            student_id=student_id,
                            defaults={
                                'status': status,
                                'remarks': remarks,
                                'marked_by': marking_staff  # Use swap staff if provided, else current staff
                            }
                        )
                        updated_students.append({'student_id': student_id, 'status': status})
                
                # ── Update existing period attendance records ──────────────────────────
                # After saving daily attendance, update all existing period attendance records
                # for the same students on the same date to reflect daily attendance overrides:
                #   OD/LEAVE → force same status in all period records
                #   LATE     → force Present ('P') in all period records  
                #   P/A      → no override needed for period records
                from .models import PeriodAttendanceSession, PeriodAttendanceRecord
                
                period_sessions = PeriodAttendanceSession.objects.filter(
                    section_id=section_id,
                    date=target_date
                )
                
                period_records_updated = 0
                for student_info in updated_students:
                    student_id = student_info['student_id']
                    daily_status = student_info['status']
                    
                    # Determine what the period status should be
                    period_status = None
                    if daily_status in ('OD', 'LEAVE'):
                        # Force OD/LEAVE in all period records
                        period_status = daily_status
                    elif daily_status == 'LATE':
                        # Force Present in all period records
                        period_status = 'P'
                    # For 'P' or 'A' daily status, don't override period records
                    
                    if period_status:
                        # Update all period records for this student on this date
                        updated = PeriodAttendanceRecord.objects.filter(
                            session__in=period_sessions,
                            student_id=student_id
                        ).update(status=period_status)
                        period_records_updated += updated
                # ───────────────────────────────────────────────────────────────────────
                
                # Determine appropriate success message
                if taken_by_staff_id and len(attendance_data) == 0:
                    # This was an assignment-only operation
                    if swap_staff and swap_staff.user:
                        staff_name = swap_staff.user.get_full_name()
                    elif swap_staff:
                        staff_name = swap_staff.staff_id
                    else:
                        staff_name = f'Staff ID {taken_by_staff_id}'
                    message = f'Attendance successfully assigned to {staff_name}'
                elif taken_by_staff_id and swap_staff:
                    # Assignment with attendance marking
                    if swap_staff.user:
                        staff_name = swap_staff.user.get_full_name()
                    else:
                        staff_name = swap_staff.staff_id
                    message = f'Attendance saved and assigned to {staff_name}'
                else:
                    # Normal attendance save
                    message = 'Attendance saved successfully'
                
                return Response({
                    'success': True,
                    'message': message,
                    'session_id': session.id,
                    'records_updated': len(attendance_data),
                    'period_records_updated': period_records_updated,
                    'swap_record_created': swap_record_created,
                    # Include updated session data for frontend
                    'assigned_to': {
                        'id': session.assigned_to.id,
                        'name': session.assigned_to.user.get_full_name() if session.assigned_to.user else '',
                        'staff_id': session.assigned_to.staff_id
                    } if session.assigned_to else None,
                    'is_read_only': False,  # Assigned staff has full access
                })
        
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class DailyAttendanceLockView(APIView):
    """
    POST /api/academics/analytics/daily-attendance-lock/{session_id}/
    Lock a daily attendance session
    """
    permission_classes = (IsAuthenticated,)
    
    def post(self, request, session_id):
        from .models import DailyAttendanceSession, SectionAdvisor
        
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')
        
        try:
            session = DailyAttendanceSession.objects.select_related('section', 'assigned_to').get(id=session_id)
        except DailyAttendanceSession.DoesNotExist:
            return Response({'error': 'Session not found'}, status=404)
        
        # Check if user is advisor for this section
        is_advisor = SectionAdvisor.objects.filter(
            advisor=staff_profile,
            section=session.section,
            is_active=True
        ).exists()
        
        # Check if user is the assigned staff member for this session
        is_assigned = session.assigned_to == staff_profile if session.assigned_to else False
        
        if not (is_advisor or is_assigned or session.created_by == staff_profile or user.is_superuser):
            raise PermissionDenied('You do not have permission to lock this session')
        
        session.is_locked = True
        session.save(update_fields=['is_locked'])
        
        return Response({
            'success': True,
            'message': 'Daily attendance session locked successfully',
            'session_id': session.id,
            'is_locked': session.is_locked
        })


class DailyAttendanceUnlockView(APIView):
    """
    POST /api/academics/analytics/daily-attendance-unlock/{session_id}/
    Unlock a daily attendance session (admin only)
    """
    permission_classes = (IsAuthenticated,)
    
    def post(self, request, session_id):
        from .models import DailyAttendanceSession
        
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')
        
        # Only admins can directly unlock
        if not user.is_superuser:
            raise PermissionDenied('Only administrators can unlock daily attendance sessions')
        
        try:
            session = DailyAttendanceSession.objects.get(id=session_id)
        except DailyAttendanceSession.DoesNotExist:
            return Response({'error': 'Session not found'}, status=404)
        
        session.is_locked = False
        session.save(update_fields=['is_locked'])
        
        return Response({
            'success': True,
            'message': 'Daily attendance session unlocked successfully',
            'session_id': session.id,
            'is_locked': session.is_locked
        })


class DailyAttendanceUnlockRequestView(APIView):
    """
    POST /api/academics/analytics/daily-attendance-unlock-request/
    Create an unlock request for daily attendance
    """
    permission_classes = (IsAuthenticated,)
    
    def post(self, request):
        from .models import DailyAttendanceSession, DailyAttendanceUnlockRequest, SectionAdvisor
        
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')
        
        session_id = request.data.get('session')
        note = request.data.get('note', '')
        
        if not session_id:
            return Response({'error': 'session_id required'}, status=400)
        
        try:
            session = DailyAttendanceSession.objects.select_related('section', 'assigned_to').get(id=session_id)
        except DailyAttendanceSession.DoesNotExist:
            return Response({'error': 'Session not found'}, status=404)
        
        # Check if user is advisor for this section
        is_advisor = SectionAdvisor.objects.filter(
            advisor=staff_profile,
            section=session.section,
            is_active=True
        ).exists()
        
        # Check if user is the assigned staff member for this session
        is_assigned = session.assigned_to == staff_profile if session.assigned_to else False
        
        if not (is_advisor or is_assigned or session.created_by == staff_profile):
            raise PermissionDenied('You do not have permission to request unlock for this session')
        
        # Check if there's already a pending request
        existing = DailyAttendanceUnlockRequest.objects.filter(
            session=session,
            status__in=['PENDING', 'HOD_APPROVED']
        ).first()
        
        if existing:
            return Response({
                'error': f'An unlock request for this session is already {existing.status.lower().replace("_", " ")}'
            }, status=400)
        
        # Create the unlock request
        unlock_request = DailyAttendanceUnlockRequest.objects.create(
            session=session,
            requested_by=staff_profile,
            note=note
        )
        
        return Response({
            'success': True,
            'message': 'Unlock request submitted successfully. It will first be reviewed by your HOD.',
            'id': unlock_request.id,
            'status': unlock_request.status,
            'hod_status': unlock_request.hod_status,
            'session_id': session.id
        })


class BulkDailyAttendanceUnlockRequestView(APIView):
    """
    POST /api/academics/bulk-attendance/unlock-request/
    Submit a single grouped unlock request for multiple daily attendance sessions at once.
    Body: { session_ids: [1, 2, ...], note: '' }
    Creates individual DailyAttendanceUnlockRequest records atomically and returns a
    unified summary so the frontend only needs one HTTP round-trip.
    """
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        from django.db import transaction
        from .models import DailyAttendanceSession, DailyAttendanceUnlockRequest, SectionAdvisor

        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')

        session_ids = request.data.get('session_ids', [])
        note = str(request.data.get('note', '') or '')

        if not session_ids or not isinstance(session_ids, list):
            return Response({'error': 'session_ids must be a non-empty list'}, status=400)

        import uuid
        bulk_group_id = uuid.uuid4()

        sessions = list(
            DailyAttendanceSession.objects.select_related('section', 'assigned_to')
            .filter(id__in=session_ids)
        )
        found_ids = {s.id for s in sessions}
        missing_ids = [sid for sid in session_ids if sid not in found_ids]

        created_requests = []
        already_pending = []
        skipped_ids = list(missing_ids)  # not found

        with transaction.atomic():
            for session in sessions:
                # Permission: advisor, assigned, creator, or superuser
                is_advisor = SectionAdvisor.objects.filter(
                    advisor=staff_profile, section=session.section, is_active=True
                ).exists()
                is_assigned = session.assigned_to == staff_profile if session.assigned_to else False
                if not (is_advisor or is_assigned or session.created_by == staff_profile or user.is_superuser):
                    skipped_ids.append(session.id)
                    continue

                existing = DailyAttendanceUnlockRequest.objects.filter(
                    session=session, status__in=['PENDING', 'HOD_APPROVED']
                ).first()
                if existing:
                    already_pending.append({
                        'session_id': session.id,
                        'unlock_request_id': existing.id,
                        'unlock_request_status': existing.status,
                        'unlock_request_hod_status': existing.hod_status,
                    })
                    continue

                req = DailyAttendanceUnlockRequest.objects.create(
                    session=session,
                    requested_by=staff_profile,
                    note=note,
                    bulk_group_id=bulk_group_id,
                )
                created_requests.append({
                    'session_id': session.id,
                    'unlock_request_id': req.id,
                    'unlock_request_status': req.status,
                    'unlock_request_hod_status': req.hod_status,
                })

        return Response({
            'created': created_requests,
            'already_pending': already_pending,
            'skipped_ids': skipped_ids,
            'total_created': len(created_requests),
            'total_already_pending': len(already_pending),
        })


class PeriodAttendanceUnlockRequestView(APIView):
    """
    POST /api/academics/analytics/period-attendance-unlock-request/
    Create an unlock request for period attendance
    """
    permission_classes = (IsAuthenticated,)
    
    def post(self, request):
        from .models import PeriodAttendanceSession, AttendanceUnlockRequest, TeachingAssignment
        
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')
        
        session_id = request.data.get('session')
        note = request.data.get('note', '')
        
        if not session_id:
            return Response({'error': 'session_id required'}, status=400)
        
        try:
            session = PeriodAttendanceSession.objects.select_related('section', 'teaching_assignment').get(id=session_id)
        except PeriodAttendanceSession.DoesNotExist:
            return Response({'error': 'Session not found'}, status=404)
        
        # Check if user has permission for this session
        is_teacher = session.teaching_assignment and session.teaching_assignment.staff == staff_profile
        
        if not (is_teacher or session.created_by == staff_profile):
            raise PermissionDenied('You do not have permission to request unlock for this session')
        
        # Check if there's already a pending request
        existing = AttendanceUnlockRequest.objects.filter(
            session=session,
            status__in=['PENDING', 'HOD_APPROVED']
        ).first()
        
        if existing:
            return Response({
                'error': f'An unlock request for this session is already {existing.status.lower().replace("_", " ")}'
            }, status=400)
        
        # Create the unlock request
        unlock_request = AttendanceUnlockRequest.objects.create(
            session=session,
            requested_by=staff_profile,
            note=note
        )
        
        return Response({
            'success': True,
            'message': 'Unlock request submitted successfully. It will first be reviewed by your HOD.',
            'id': unlock_request.id,
            'status': unlock_request.status,
            'hod_status': unlock_request.hod_status,
            'session_id': session.id
        })


class HODUnlockRequestsView(APIView):
    """
    GET: List all unlock requests pending HOD approval for user's department
    POST: Approve or reject as HOD
    """
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):
        from .models import AttendanceUnlockRequest, DailyAttendanceUnlockRequest, Department, DepartmentRole
        from academics.serializers import AttendanceUnlockRequestSerializer
        
        try:
            user = request.user
            staff_profile = getattr(user, 'staff_profile', None)
            
            if not staff_profile:
                return Response({'error': 'Staff profile required'}, status=403)
            
            # Get departments where user is HOD through DepartmentRole
            hod_roles = DepartmentRole.objects.filter(
                staff=staff_profile,
                role='HOD',
                is_active=True
            ).values_list('department_id', flat=True)
            
            hod_departments = list(hod_roles)
            
            if not hod_departments:
                return Response({'error': 'You are not an HOD of any department'}, status=403)
            
            # Get period attendance unlock requests
            period_requests = AttendanceUnlockRequest.objects.select_related(
                'session__section__batch__course__department', 'requested_by', 'requested_by__user', 'hod_reviewed_by'
            ).filter(
                session__section__batch__course__department_id__in=hod_departments,
                hod_status='PENDING'
            ).order_by('-requested_at')
            period_requests_list = list(period_requests)
            
            # Get daily attendance unlock requests
            daily_requests = DailyAttendanceUnlockRequest.objects.select_related(
                'session__section__batch__course__department', 'requested_by', 'requested_by__user', 'hod_reviewed_by'
            ).filter(
                session__section__batch__course__department_id__in=hod_departments,
                hod_status='PENDING'
            ).order_by('-requested_at')
            daily_requests_list = list(daily_requests)
            
            # Combine both lists
            all_requests = period_requests_list + daily_requests_list
            
            # Sort by requested_at
            all_requests.sort(key=lambda x: x.requested_at, reverse=True)
            
            # Format results manually
            results = []
            for req in all_requests:
                # Determine request type
                is_daily = isinstance(req, DailyAttendanceUnlockRequest)
                
                data = {
                    'id': req.id,
                    'request_type': 'daily' if is_daily else 'period',
                    'status': req.status,
                    'hod_status': req.hod_status,
                    'requested_at': req.requested_at.strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
                    'requested_by': {
                        'id': req.requested_by.id if req.requested_by else None,
                        'name': req.requested_by.user.get_full_name() if (req.requested_by and hasattr(req.requested_by, 'user') and req.requested_by.user) else '',
                        'staff_id': req.requested_by.staff_id if req.requested_by else ''
                    },
                    'note': req.note or '',
                    'session': {
                        'id': req.session.id,
                        'section': str(req.session.section),
                        'date': req.session.date.strftime('%Y-%m-%d'),
                    }
                }
                
                # Add period field only for period attendance  
                if not is_daily:
                    data['session']['period'] = str(req.session.period) if req.session.period else None
                    
                results.append(data)
            
            return Response({
                'results': results,
                'count': len(results)
            })
        except Exception as e:
            import traceback
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error in HODUnlockRequestsView.get: {str(e)}")
            logger.error(traceback.format_exc())
            return Response({'error': f'Internal server error: {str(e)}'}, status=500)
    
    def post(self, request):
        from .models import AttendanceUnlockRequest, DailyAttendanceUnlockRequest, DepartmentRole
        
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        
        if not staff_profile:
            return Response({'error': 'Staff profile required'}, status=403)
        
        request_id = request.data.get('id')
        request_type = request.data.get('request_type')  # 'period' or 'daily'
        action = request.data.get('action')  # 'approve' or 'reject'
        hod_note = request.data.get('note', '')
        
        if not all([request_id, request_type, action]):
            return Response({'error': 'Missing required fields'}, status=400)
        
        if action not in ['approve', 'reject']:
            return Response({'error': 'Invalid action'}, status=400)
        
        # Get departments where user is HOD through DepartmentRole
        hod_roles = DepartmentRole.objects.filter(
            staff=staff_profile,
            role='HOD',
            is_active=True
        ).values_list('department_id', flat=True)
        
        hod_departments = list(hod_roles)
        
        if not hod_departments:
            return Response({'error': 'You are not an HOD of any department'}, status=403)
        
        try:
            if request_type == 'period':
                unlock_req = AttendanceUnlockRequest.objects.select_related('session__section__batch__course__department').get(id=request_id)
            else:
                unlock_req = DailyAttendanceUnlockRequest.objects.select_related('session__section__batch__course__department').get(id=request_id)
            
            # Verify this request is for user's department
            req_dept_id = unlock_req.session.section.batch.course.department_id if unlock_req.session and unlock_req.session.section and unlock_req.session.section.batch and unlock_req.session.section.batch.course else None
            if req_dept_id not in hod_departments:
                return Response({'error': 'This request is not for your department'}, status=403)
            
            if unlock_req.hod_status != 'PENDING':
                return Response({'error': 'This request has already been reviewed by HOD'}, status=400)
            
            if action == 'approve':
                unlock_req.hod_status = 'HOD_APPROVED'
                unlock_req.status = 'HOD_APPROVED'  # Update overall status too
                unlock_req.hod_reviewed_by = staff_profile
                unlock_req.hod_reviewed_at = timezone.now()
                unlock_req.hod_note = hod_note
                unlock_req.save()
                
                message = f'{request_type.title()} attendance unlock request approved by HOD. Forwarded to final approval.'
            else:
                unlock_req.hod_status = 'REJECTED'
                unlock_req.status = 'REJECTED'
                unlock_req.hod_reviewed_by = staff_profile
                unlock_req.hod_reviewed_at = timezone.now()
                unlock_req.hod_note = hod_note
                unlock_req.save()
                
                message = f'{request_type.title()} attendance unlock request rejected by HOD.'
            
            return Response({
                'success': True,
                'message': message
            })
            
        except (AttendanceUnlockRequest.DoesNotExist, DailyAttendanceUnlockRequest.DoesNotExist):
            return Response({'error': 'Request not found'}, status=404)


class DailyAttendanceSessionDetailView(APIView):
    """
    GET: Fetch detailed records for a specific daily attendance session
    Similar to period-attendance detail endpoint
    """
    permission_classes = (IsAuthenticated,)
    
    def get(self, request, session_id):
        try:
            from .models import DailyAttendanceSession, DailyAttendanceRecord
            
            session = DailyAttendanceSession.objects.select_related(
                'section', 'section__batch', 'created_by'
            ).get(id=session_id)
            
            # Get all records for this session
            records = DailyAttendanceRecord.objects.filter(
                session=session
            ).select_related('student', 'student__user').order_by('student__reg_no')
            
            records_data = []
            for record in records:
                student = record.student
                reg_no = student.reg_no if student else ''
                records_data.append({
                    'id': record.id,
                    'student_id': student.id if student else None,
                    'student_pk': student.id if student else None,
                    'reg_no': reg_no,
                    'regno': reg_no,
                    'student': {
                        'id': student.id if student else None,
                        'pk': student.id if student else None,
                        'reg_no': reg_no,
                        'regno': reg_no,
                        'registration_number': reg_no,
                        'name': student.user.get_full_name() if student and student.user else '',
                        'username': student.user.username if student and student.user else '',
                    },
                    'status': record.status,
                    'attendance': record.status,
                    'type': record.status,
                    'remarks': record.remarks or '',
                    'marked_at': record.marked_at.isoformat() if record.marked_at else None,
                })
            
            return Response({
                'session_id': session.id,
                'id': session.id,
                'section_id': session.section_id,
                'section_name': session.section.name if session.section else '',
                'date': session.date.isoformat(),
                'session_date': session.date.isoformat(),
                'is_locked': session.is_locked,
                'created_at': session.created_at.isoformat() if session.created_at else None,
                'records': records_data,
                'total_records': len(records_data),
            })
        except DailyAttendanceSession.DoesNotExist:
            return Response({'error': 'Session not found'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class SectionStudentAttendanceDayView(APIView):
    """
    Returns per-student attendance details for a section on a specific date.

    Query params:
      section_id  (required)  – section to inspect
      date        (required)  – ISO date string YYYY-MM-DD
      session_id  (optional)  – if supplied, filters to ONE specific period session
                                (used by My Class subject row expansions)

    For each student the response includes:
      reg_no, name, daily_status,
      total_periods  – how many PeriodAttendanceSessions exist for this section/date
      present_periods – periods where the student was P / OD / LATE
      absent_periods  – periods where the student was A
      leave_periods   – periods where the student was LEAVE
      percentage      – present_periods / total_periods * 100  (0 if no sessions)
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        section_id = request.query_params.get('section_id')
        date_str = request.query_params.get('date')
        date_from_str = request.query_params.get('date_from')
        date_to_str = request.query_params.get('date_to')
        complete = request.query_params.get('complete', '').lower() in ('true', '1')
        session_id = request.query_params.get('session_id')  # optional – filter to one period

        if not section_id:
            return Response({'error': 'section_id is required'}, status=400)

        from datetime import date as date_class
        # Resolve date range: prefer complete (no filter), then date_from/date_to, then single date
        try:
            if complete:
                target_from = target_to = None  # no date constraint
            elif date_from_str and date_to_str:
                target_from = date_class.fromisoformat(date_from_str)
                target_to = date_class.fromisoformat(date_to_str)
            elif date_str:
                target_from = target_to = date_class.fromisoformat(date_str)
            else:
                return Response({'error': 'date or date_from+date_to are required'}, status=400)
        except ValueError:
            return Response({'error': 'Invalid date format, expected YYYY-MM-DD'}, status=400)

        try:
            section_id = int(section_id)
        except (TypeError, ValueError):
            return Response({'error': 'section_id must be an integer'}, status=400)

        # --- Determine period sessions in range ---
        sessions_qs = PeriodAttendanceSession.objects.filter(section_id=section_id)
        if not complete and target_from and target_to:
            sessions_qs = sessions_qs.filter(date__range=(target_from, target_to))
        if session_id:
            try:
                sessions_qs = sessions_qs.filter(pk=int(session_id))
            except (TypeError, ValueError):
                pass

        # Fetch session_id → (date, period_id) mapping so we can deduplicate correctly.
        #
        # Two kinds of duplication to handle:
        #   1. Same-day elective splits: multiple sessions share the same period_id on the
        #      same date (3 staff groups marking period 4 → should count as 1 period-slot that day).
        #   2. Multi-day range: period 4 on Mon + period 4 on Tue = 2 period-slots total (correct).
        #
        # Deduplication key = (date, period_id)  — collapses same-day elective splits,
        # but keeps each day's occurrence as a distinct slot.
        session_slot_map: dict = {}  # session_id -> (date, period_id)
        for row in sessions_qs.values('id', 'period_id', 'date'):
            session_slot_map[row['id']] = (row['date'], row['period_id'])

        session_ids = list(session_slot_map.keys())

        # Unique (date, period_id) slots across the range
        unique_slots = set(session_slot_map.values())
        total_periods = len(unique_slots)

        # --- Get all students in the section (exclude inactive/debarred) ---
        students = StudentProfile.objects.filter(
            section_id=section_id
        ).exclude(status__in=['INACTIVE', 'DEBAR']).select_related('user').order_by('reg_no')

        # --- Get all period records for these sessions in one query ---
        records_qs = PeriodAttendanceRecord.objects.filter(
            session_id__in=session_ids
        ).values('student_id', 'status', 'session_id')

        from collections import defaultdict
        # Status priority for deduplication within one slot (higher = better)
        _STATUS_PRIORITY = {'P': 5, 'OD': 4, 'LATE': 3, 'LEAVE': 2, 'A': 1}

        # student_id -> {(date, period_id) -> best_status}
        # Elective splits on the same day: keep the best status across all split sessions.
        student_period_status: dict = defaultdict(dict)  # {student_id: {slot_key: status}}
        for rec in records_qs:
            slot_key = session_slot_map.get(rec['session_id'])
            sid = rec['student_id']
            new_status = rec['status']
            existing = student_period_status[sid].get(slot_key)
            if existing is None or _STATUS_PRIORITY.get(new_status, 0) > _STATUS_PRIORITY.get(existing, 0):
                student_period_status[sid][slot_key] = new_status

        # Flatten: student_id -> list[status] (one entry per unique (date, period_id) slot)
        student_records: dict = {
            sid: list(slot_map.values())
            for sid, slot_map in student_period_status.items()
        }

        # --- Get daily attendance aggregated across the date range ---
        from .models import DailyAttendanceSession, DailyAttendanceRecord
        # student_id -> {present_days, absent_days, leave_days, last_status}
        daily_agg: dict = defaultdict(lambda: {'present_days': 0, 'absent_days': 0, 'leave_days': 0, 'last_status': None, 'last_date': None})
        try:
            daily_sessions = DailyAttendanceSession.objects.filter(
                section_id=section_id,
            )
            if not complete and target_from and target_to:
                daily_sessions = daily_sessions.filter(date__range=(target_from, target_to))
            daily_sessions = daily_sessions.order_by('date')
            daily_session_ids = list(daily_sessions.values_list('id', flat=True))
            # Map session id -> date
            session_date_map = {ds.id: ds.date for ds in daily_sessions}
            for dr in DailyAttendanceRecord.objects.filter(session_id__in=daily_session_ids).values('student_id', 'status', 'session_id'):
                entry = daily_agg[dr['student_id']]
                ds_date = session_date_map.get(dr['session_id'])
                if dr['status'] in ('P', 'OD', 'LATE'):
                    entry['present_days'] += 1
                elif dr['status'] == 'A':
                    entry['absent_days'] += 1
                elif dr['status'] == 'LEAVE':
                    entry['leave_days'] += 1
                # track latest-date status for display
                if ds_date and (entry['last_date'] is None or ds_date >= entry['last_date']):
                    entry['last_date'] = ds_date
                    entry['last_status'] = dr['status']
        except Exception:
            pass

        is_range = complete or (target_from != target_to if (target_from and target_to) else False)

        # --- Build response ---
        result = []
        for stu in students:
            statuses = student_records.get(stu.id, [])
            present_p = sum(1 for s in statuses if s in ('P', 'OD', 'LATE'))
            absent_p = sum(1 for s in statuses if s == 'A')
            leave_p = sum(1 for s in statuses if s == 'LEAVE')
            pct = round(present_p / total_periods * 100, 1) if total_periods > 0 else None
            name = stu.user.get_full_name() if stu.user else ''
            if not name and stu.user:
                name = stu.user.username or ''
            da = daily_agg.get(stu.id, {})
            result.append({
                'student_id': stu.id,
                'reg_no': stu.reg_no or '',
                'name': name,
                'daily_status': da.get('last_status', None) if not is_range else None,
                'daily_present_days': da.get('present_days', 0),
                'daily_absent_days': da.get('absent_days', 0),
                'daily_leave_days': da.get('leave_days', 0),
                'total_periods': total_periods,
                'present_periods': present_p,
                'absent_periods': absent_p,
                'leave_periods': leave_p,
                'percentage': pct,
            })

        return Response({
            'section_id': section_id,
            'date': target_from.isoformat() if target_from else None,
            'date_from': target_from.isoformat() if target_from else None,
            'date_to': target_to.isoformat() if target_to else None,
            'is_range': is_range,
            'total_periods': total_periods,
            'students': result,
        })


class DailyAttendanceRevertAssignmentView(APIView):
    """
    POST /api/academics/analytics/daily-attendance-revert/{session_id}/
    Revert assignment back to original advisor (only if assigned staff hasn't marked attendance)
    """
    permission_classes = (IsAuthenticated,)
    
    def post(self, request, session_id):
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')
        
        from .models import DailyAttendanceSession, DailyAttendanceRecord, SectionAdvisor, DailyAttendanceSwapRecord
        from django.db import transaction
        
        try:
            session = DailyAttendanceSession.objects.select_related(
                'assigned_to', 'assigned_to__user', 'created_by', 'section'
            ).get(id=session_id)
        except DailyAttendanceSession.DoesNotExist:
            return Response({'error': 'Session not found'}, status=404)
        
        # Check if user is the original advisor for this section
        is_advisor = SectionAdvisor.objects.filter(
            advisor=staff_profile,
            section=session.section,
            is_active=True
        ).exists()
        
        if not (is_advisor or user.is_superuser):
            raise PermissionDenied('Only the original advisor can revert assignments')
        
        # Check if session is actually assigned to someone else
        if not session.assigned_to:
            return Response({'error': 'This session is not assigned to anyone'}, status=400)
        
        if session.assigned_to == staff_profile:
            return Response({'error': 'This session is already assigned to you'}, status=400)
        
        # Check if assigned staff has marked any attendance
        assigned_staff_records = DailyAttendanceRecord.objects.filter(
            session=session,
            marked_by=session.assigned_to
        ).exists()
        
        if assigned_staff_records:
            return Response({
                'error': f'Cannot revert assignment. {session.assigned_to.user.get_full_name() if session.assigned_to.user else session.assigned_to.staff_id} has already marked attendance.',
                'assigned_staff': {
                    'name': session.assigned_to.user.get_full_name() if session.assigned_to.user else '',
                    'staff_id': session.assigned_to.staff_id
                }
            }, status=400)
        
        try:
            with transaction.atomic():
                # Store the assigned staff info for the response
                previously_assigned_to = session.assigned_to
                previously_assigned_name = previously_assigned_to.user.get_full_name() if previously_assigned_to.user else previously_assigned_to.staff_id
                
                # Revert the assignment
                session.assigned_to = None
                session.save(update_fields=['assigned_to'])
                
                # Create swap record for audit trail
                DailyAttendanceSwapRecord.objects.create(
                    session=session,
                    assigned_by=staff_profile,
                    assigned_to=None,  # Reverting to original (no assignment)
                    reason=f'Assignment reverted by {staff_profile.user.get_full_name() if staff_profile.user else staff_profile.staff_id}. Was assigned to {previously_assigned_name}'
                )
                
                return Response({
                    'success': True,
                    'message': f'Assignment successfully reverted from {previously_assigned_name}. You can now mark attendance for this section.',
                    'session_id': session.id,
                    'reverted_from': {
                        'name': previously_assigned_to.user.get_full_name() if previously_assigned_to.user else '',
                        'staff_id': previously_assigned_to.staff_id
                    }
                })
                
        except Exception as e:
            return Response({'error': f'Failed to revert assignment: {str(e)}'}, status=500)


class PeriodAttendanceSwapView(APIView):
    """
    POST /api/academics/analytics/period-attendance-swap/
    Assign a period attendance session to another staff member.
    Body: { section_id, period_id, date, teaching_assignment_id (optional), taken_by_staff_id }
    """
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        from .models import PeriodAttendanceSession, PeriodAttendanceSwapRecord, TeachingAssignment
        from .models import StaffProfile, SectionAdvisor, Section
        from timetable.models import TimetableSlot
        from django.db import transaction
        import datetime

        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')

        data = request.data
        section_id = data.get('section_id')
        period_id = data.get('period_id')
        date_str = data.get('date')
        teaching_assignment_id = data.get('teaching_assignment_id')
        taken_by_staff_id = data.get('taken_by_staff_id')

        if not all([section_id, period_id, date_str, taken_by_staff_id]):
            return Response({'error': 'section_id, period_id, date, and taken_by_staff_id are required'}, status=400)

        try:
            target_date = datetime.date.fromisoformat(date_str)
        except (ValueError, TypeError):
            return Response({'error': 'Invalid date format'}, status=400)

        try:
            section = Section.objects.get(id=section_id)
        except Section.DoesNotExist:
            return Response({'error': 'Section not found'}, status=404)

        try:
            period = TimetableSlot.objects.get(id=period_id)
        except TimetableSlot.DoesNotExist:
            return Response({'error': 'Period not found'}, status=404)

        # Verify requesting user has a marking assignment for this period or has the mark_attendance perm
        perms = get_user_permissions(user)
        can_mark_all = 'academics.mark_attendance' in perms or user.is_superuser
        teach_assign = None
        if teaching_assignment_id:
            try:
                teach_assign = TeachingAssignment.objects.get(id=teaching_assignment_id, is_active=True)
            except TeachingAssignment.DoesNotExist:
                pass

        # Validate target staff
        try:
            swap_staff = StaffProfile.objects.select_related('user').get(id=taken_by_staff_id, status='ACTIVE')
        except StaffProfile.DoesNotExist:
            return Response({'error': 'Target staff not found or inactive'}, status=404)

        # Same department check
        current_dept = staff_profile.get_current_department()
        swap_dept = swap_staff.get_current_department()
        if current_dept and swap_dept and current_dept.id != swap_dept.id:
            return Response({'error': 'Swap staff must be from the same department'}, status=400)

        try:
            with transaction.atomic():
                lookup = {'section': section, 'period': period, 'date': target_date}
                if teach_assign:
                    lookup['teaching_assignment'] = teach_assign
                else:
                    lookup['created_by'] = staff_profile

                session, created = PeriodAttendanceSession.objects.get_or_create(
                    **lookup,
                    defaults={'created_by': staff_profile, 'teaching_assignment': teach_assign}
                )

                old_assigned = session.assigned_to
                if old_assigned == swap_staff:
                    return Response({'error': 'Already assigned to this staff member'}, status=400)

                session.assigned_to = swap_staff
                session.save(update_fields=['assigned_to'])

                PeriodAttendanceSwapRecord.objects.create(
                    session=session,
                    assigned_by=staff_profile,
                    assigned_to=swap_staff,
                    reason=f'Period attendance assigned from {staff_profile.user.get_full_name() if staff_profile.user else staff_profile.staff_id} to {swap_staff.user.get_full_name() if swap_staff.user else swap_staff.staff_id}'
                )

                return Response({
                    'success': True,
                    'message': f'Period attendance assigned to {swap_staff.user.get_full_name() if swap_staff.user else swap_staff.staff_id}',
                    'session_id': session.id,
                    'assigned_to': {
                        'id': swap_staff.id,
                        'name': swap_staff.user.get_full_name() if swap_staff.user else '',
                        'staff_id': swap_staff.staff_id
                    }
                })
        except Exception as e:
            return Response({'error': f'Failed to assign: {str(e)}'}, status=500)


class PeriodAttendanceRevertAssignmentView(APIView):
    """
    POST /api/academics/analytics/period-attendance-revert/{session_id}/
    Revert period assignment back to original staff (only if assigned staff hasn't marked).
    """
    permission_classes = (IsAuthenticated,)

    def post(self, request, session_id):
        from .models import PeriodAttendanceSession, PeriodAttendanceRecord, PeriodAttendanceSwapRecord
        from django.db import transaction

        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')

        try:
            session = PeriodAttendanceSession.objects.select_related(
                'assigned_to', 'assigned_to__user', 'created_by', 'section'
            ).get(id=session_id)
        except PeriodAttendanceSession.DoesNotExist:
            return Response({'error': 'Session not found'}, status=404)

        if not session.assigned_to:
            return Response({'error': 'This session is not assigned to anyone'}, status=400)

        if session.assigned_to == staff_profile:
            return Response({'error': 'This session is already assigned to you'}, status=400)

        # Only original creator / superuser can revert
        perms = get_user_permissions(user)
        is_creator = session.created_by == staff_profile
        if not (is_creator or user.is_superuser or 'analytics.edit_all_analytics' in perms):
            raise PermissionDenied('Only the original staff can revert assignments')

        # Check if assigned staff has already marked any attendance
        assigned_staff_records = PeriodAttendanceRecord.objects.filter(
            session=session, marked_by=session.assigned_to
        ).exists()

        if assigned_staff_records:
            return Response({
                'error': f'Cannot revert. {session.assigned_to.user.get_full_name() if session.assigned_to.user else session.assigned_to.staff_id} has already marked attendance for this period.',
                'assigned_staff': {
                    'name': session.assigned_to.user.get_full_name() if session.assigned_to.user else '',
                    'staff_id': session.assigned_to.staff_id
                }
            }, status=400)

        try:
            with transaction.atomic():
                prev = session.assigned_to
                prev_name = prev.user.get_full_name() if prev.user else prev.staff_id

                session.assigned_to = None
                session.save(update_fields=['assigned_to'])

                PeriodAttendanceSwapRecord.objects.create(
                    session=session,
                    assigned_by=staff_profile,
                    assigned_to=None,
                    reason=f'Assignment reverted by {staff_profile.user.get_full_name() if staff_profile.user else staff_profile.staff_id}. Was assigned to {prev_name}'
                )

                return Response({
                    'success': True,
                    'message': f'Assignment reverted from {prev_name}. You can now mark attendance.',
                    'session_id': session.id,
                    'reverted_from': {
                        'name': prev.user.get_full_name() if prev.user else '',
                        'staff_id': prev.staff_id
                    }
                })
        except Exception as e:
            return Response({'error': f'Failed to revert: {str(e)}'}, status=500)



class AttendanceAssignmentRequestView(APIView):
    """
    GET  /api/academics/attendance-assignment-requests/?status=PENDING
    POST /api/academics/attendance-assignment-requests/
    """
    permission_classes = (IsAuthenticated,)

    def _serialize(self, req):
        return {
            'id': req.id,
            'assignment_type': req.assignment_type,
            'status': req.status,
            'date': req.date.isoformat() if req.date else None,
            'section_id': req.section_id,
            'section_name': req.section.name if req.section else '',
            'period_id': req.period_id,
            'period_label': (req.period.label or f'Period {req.period.index}') if req.period else None,
            'requested_by_id': req.requested_by_id,
            'requested_by_name': req.requested_by.user.get_full_name() if req.requested_by and req.requested_by.user else '',
            'requested_to_id': req.requested_to_id,
            'requested_to_name': req.requested_to.user.get_full_name() if req.requested_to and req.requested_to.user else '',
            'reason': req.reason or '',
            'created_at': req.created_at.isoformat() if req.created_at else None,
            'responded_at': req.responded_at.isoformat() if req.responded_at else None,
            'response_message': req.response_message or '',
        }

    def get(self, request):
        from .models import AttendanceAssignmentRequest
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')
        status_filter = request.query_params.get('status')
        base_qs = dict(requested_by=staff_profile) if True else {}
        sent_qs = AttendanceAssignmentRequest.objects.filter(
            requested_by=staff_profile
        ).select_related('requested_by__user', 'requested_to__user', 'section')
        received_qs = AttendanceAssignmentRequest.objects.filter(
            requested_to=staff_profile
        ).select_related('requested_by__user', 'requested_to__user', 'section')
        if status_filter:
            sent_qs = sent_qs.filter(status=status_filter)
            received_qs = received_qs.filter(status=status_filter)
        return Response({
            'sent': [self._serialize(r) for r in sent_qs],
            'received': [self._serialize(r) for r in received_qs],
        })

    def post(self, request):
        from .models import AttendanceAssignmentRequest, StaffProfile, Section, SectionAdvisor
        from datetime import date as date_class
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')
        section_id = request.data.get('section_id')
        date_str = request.data.get('date')
        requested_to_id = request.data.get('requested_to_id')
        assignment_type = request.data.get('assignment_type', 'DAILY')
        if not section_id or not date_str or not requested_to_id:
            return Response({'error': 'section_id, date, and requested_to_id are required'}, status=400)
        try:
            date_obj = date_class.fromisoformat(str(date_str))
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
        try:
            section = Section.objects.get(id=section_id)
        except Section.DoesNotExist:
            return Response({'error': 'Section not found'}, status=404)
        try:
            requested_to = StaffProfile.objects.select_related('user').get(id=requested_to_id, status='ACTIVE')
        except StaffProfile.DoesNotExist:
            return Response({'error': 'Target staff not found or inactive'}, status=404)
        if requested_to == staff_profile:
            return Response({'error': 'Cannot request yourself'}, status=400)
        # Resolve period for PERIOD type
        period_obj = None
        if assignment_type == 'PERIOD':
            period_id_val = request.data.get('period_id')
            if not period_id_val:
                return Response({'error': 'period_id is required for PERIOD type'}, status=400)
            try:
                from timetable.models import TimetableSlot
                period_obj = TimetableSlot.objects.get(id=period_id_val)
            except TimetableSlot.DoesNotExist:
                return Response({'error': 'Period not found'}, status=404)
        # Duplicate check (for PERIOD also match on period)
        dup_qs = AttendanceAssignmentRequest.objects.filter(
            section=section, date=date_obj, assignment_type=assignment_type,
            status='PENDING', requested_by=staff_profile
        )
        if period_obj:
            dup_qs = dup_qs.filter(period=period_obj)
        existing = dup_qs.first()
        if existing:
            to_name = existing.requested_to.user.get_full_name() if existing.requested_to and existing.requested_to.user else ''
            return Response({'error': f'You already have a pending request for this session (sent to {to_name}).'}, status=400)
        req = AttendanceAssignmentRequest.objects.create(
            assignment_type=assignment_type,
            section=section,
            date=date_obj,
            requested_by=staff_profile,
            requested_to=requested_to,
            status='PENDING',
            period=period_obj,
        )
        to_name = requested_to.user.get_full_name() if requested_to.user else requested_to.staff_id
        return Response({
            'success': True,
            'id': req.id,
            'message': f'Request sent to {to_name}. They can approve or reject it.',
            'request': self._serialize(req),
        }, status=201)


class AttendanceAssignmentRequestActionView(APIView):
    """
    POST /api/academics/attendance-assignment-requests/<pk>/approve/
    POST /api/academics/attendance-assignment-requests/<pk>/reject/
    """
    permission_classes = (IsAuthenticated,)

    def post(self, request, pk, action):
        from .models import AttendanceAssignmentRequest, DailyAttendanceSession, DailyAttendanceSwapRecord
        from .models import PeriodAttendanceSession, PeriodAttendanceSwapRecord
        from django.utils import timezone as tz
        from django.db import transaction
        user = request.user
        staff_profile = getattr(user, 'staff_profile', None)
        if not staff_profile:
            raise PermissionDenied('Staff profile required')
        if action not in ('approve', 'reject'):
            return Response({'error': 'Invalid action'}, status=400)
        try:
            req = AttendanceAssignmentRequest.objects.select_related(
                'requested_by__user', 'requested_to__user', 'section'
            ).get(id=pk)
        except AttendanceAssignmentRequest.DoesNotExist:
            return Response({'error': 'Request not found'}, status=404)
        if req.requested_to != staff_profile:
            raise PermissionDenied('Only the requested staff can respond')
        if req.status != 'PENDING':
            return Response({'error': f'Request is already {req.status.lower()}'}, status=400)
        with transaction.atomic():
            req.responded_at = tz.now()
            if action == 'approve':
                req.status = 'APPROVED'
                req.response_message = request.data.get('message', '')
                req.save()
                if req.assignment_type == 'DAILY':
                    session, _ = DailyAttendanceSession.objects.get_or_create(
                        section=req.section,
                        date=req.date,
                        defaults={'created_by': req.requested_by}
                    )
                    old_assigned = session.assigned_to
                    session.assigned_to = req.requested_to
                    session.save(update_fields=['assigned_to'])
                    if old_assigned != req.requested_to:
                        by_name = req.requested_by.user.get_full_name() if req.requested_by.user else req.requested_by.staff_id
                        to_name = req.requested_to.user.get_full_name() if req.requested_to.user else req.requested_to.staff_id
                        DailyAttendanceSwapRecord.objects.create(
                            session=session,
                            assigned_by=req.requested_by,
                            assigned_to=req.requested_to,
                            reason=f'Assignment request approved by {to_name}'
                        )
                    req.daily_session = session
                    req.save(update_fields=['daily_session'])
                elif req.assignment_type == 'PERIOD' and req.period:
                    # Assign all existing period sessions for section+period+date; create one if none exist
                    period_sessions = list(PeriodAttendanceSession.objects.filter(
                        section=req.section, period=req.period, date=req.date
                    ))
                    if period_sessions:
                        for ps in period_sessions:
                            ps.assigned_to = req.requested_to
                            ps.save(update_fields=['assigned_to'])
                            PeriodAttendanceSwapRecord.objects.create(
                                session=ps,
                                assigned_by=req.requested_by,
                                assigned_to=req.requested_to,
                                reason='Period attendance request approved'
                            )
                        req.period_session = period_sessions[0]
                        req.save(update_fields=['period_session'])
                    else:
                        # No session yet — create a placeholder so the assignee sees it
                        try:
                            ps = PeriodAttendanceSession.objects.create(
                                section=req.section, period=req.period, date=req.date,
                                created_by=req.requested_by, assigned_to=req.requested_to,
                            )
                            PeriodAttendanceSwapRecord.objects.create(
                                session=ps,
                                assigned_by=req.requested_by,
                                assigned_to=req.requested_to,
                                reason='Period attendance request approved'
                            )
                            req.period_session = ps
                            req.save(update_fields=['period_session'])
                        except Exception:
                            pass  # Unique constraint edge case — sessions will be updated on next load
                # Cancel other pending requests for the same session
                cancel_qs = AttendanceAssignmentRequest.objects.filter(
                    section=req.section, date=req.date,
                    assignment_type=req.assignment_type, status='PENDING'
                ).exclude(id=req.id)
                if req.assignment_type == 'PERIOD' and req.period:
                    cancel_qs = cancel_qs.filter(period=req.period)
                cancel_qs.update(status='CANCELLED')
                to_name = req.requested_to.user.get_full_name() if req.requested_to.user else req.requested_to.staff_id
                return Response({'success': True, 'message': f'Approved. Daily attendance for {req.section} on {req.date} is now assigned to {to_name}.'})
            else:
                req.status = 'REJECTED'
                req.response_message = request.data.get('message', '')
                req.save()
                by_name = req.requested_by.user.get_full_name() if req.requested_by.user else req.requested_by.staff_id
                return Response({'success': True, 'message': f'Request from {by_name} has been rejected.'})

class AttendanceNotificationCountView(APIView):
    """
    GET /api/academics/analytics/attendance-notification-count/
    Returns the count of pending attendance unlock requests visible to the current user.
    - HOD: counts PENDING requests in their department (awaiting HOD approval)
    - IQAC / view_all_analytics: counts HOD_APPROVED period requests + PENDING/HOD_APPROVED daily requests
      (awaiting final approval)
    Only returns a count > 0 for users who have something to action; all others receive 0.
    """
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        from .models import AttendanceUnlockRequest, DailyAttendanceUnlockRequest, DepartmentRole

        user = request.user
        perms = get_user_permissions(user)
        staff_profile = getattr(user, 'staff_profile', None)

        can_view_all = 'analytics.view_all_analytics' in perms or user.is_superuser

        if can_view_all:
            # IQAC / admin: count requests that have passed HOD and are awaiting final approval
            period_count = AttendanceUnlockRequest.objects.filter(
                hod_status='HOD_APPROVED', status='HOD_APPROVED'
            ).count()
            daily_count = DailyAttendanceUnlockRequest.objects.filter(
                status__in=['PENDING', 'HOD_APPROVED']
            ).count()
            total = period_count + daily_count
            return Response({'count': total, 'role': 'iqac'})

        # Check HOD role via DepartmentRole
        if staff_profile:
            hod_dept_ids = list(
                DepartmentRole.objects.filter(
                    staff=staff_profile, role='HOD', is_active=True
                ).values_list('department_id', flat=True)
            )
            if hod_dept_ids:
                period_count = AttendanceUnlockRequest.objects.filter(
                    session__section__batch__course__department_id__in=hod_dept_ids,
                    hod_status='PENDING'
                ).count()
                daily_count = DailyAttendanceUnlockRequest.objects.filter(
                    session__section__batch__course__department_id__in=hod_dept_ids,
                    hod_status='PENDING'
                ).count()
                total = period_count + daily_count
                return Response({'count': total, 'role': 'hod'})

        # User has no actionable requests
        return Response({'count': 0, 'role': 'none'})
