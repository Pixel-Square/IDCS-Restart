from django.contrib import admin
from django.urls import path, reverse
from django.http import JsonResponse
from django.shortcuts import render
from django.contrib import messages
from django.db import transaction
from django.utils import timezone
import json
from .models import (
    AcademicYear,
    Department,
    Program,
    Course,
    Semester,
    Section,
    Batch,
    Subject,
    PROFILE_STATUS_CHOICES,
    StudentProfile,
    StaffProfile,
    StudentSectionAssignment,
    StaffDepartmentAssignment,
    RoleAssignment,
    SpecialCourseAssessmentSelection,
    SpecialCourseAssessmentEditRequest,
)
from .models import TeachingAssignment
from .models import StudentMentorMap, SectionAdvisor, DepartmentRole
from .models import StudentSubjectBatch
from .models import PeriodAttendanceSession, PeriodAttendanceRecord, AttendanceUnlockRequest
from django import forms
from django.db import models
from django.core.exceptions import ValidationError
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from accounts.models import Role
from django.contrib.auth.models import Group
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone

import openpyxl

try:
    from curriculum.models import SPECIAL_ASSESSMENT_CHOICES
except Exception:
    SPECIAL_ASSESSMENT_CHOICES = (
        ('ssa1', 'SSA 1'),
        ('ssa2', 'SSA 2'),
        ('formative1', 'Formative 1'),
        ('formative2', 'Formative 2'),
        ('cia1', 'CIA 1'),
        ('cia2', 'CIA 2'),
    )


class StudentProfileForm(forms.ModelForm):
    class Meta:
        model = StudentProfile
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk:
            # disable reg_no editing for existing records in admin
            if 'reg_no' in self.fields:
                self.fields['reg_no'].disabled = True
            # deprecate editing section here; assignments should be used
            if 'section' in self.fields:
                self.fields['section'].disabled = True

    def clean_reg_no(self):
        val = self.cleaned_data.get('reg_no')
        if self.instance and self.instance.pk:
            # ensure immutability is enforced at form level with a friendly error
            if val != self.instance.reg_no:
                raise ValidationError('Student reg_no is immutable and cannot be changed.')
        return val


class StaffProfileForm(forms.ModelForm):
    class Meta:
        model = StaffProfile
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk:
            # disable staff_id editing for existing records in admin
            if 'staff_id' in self.fields:
                self.fields['staff_id'].disabled = True
            # deprecate editing department here; assignments should be used
            if 'department' in self.fields:
                self.fields['department'].disabled = True

    def clean_staff_id(self):
        val = self.cleaned_data.get('staff_id')
        if self.instance and self.instance.pk:
            # ensure immutability is enforced at form level with a friendly error
            if val != self.instance.staff_id:
                raise ValidationError('Staff staff_id is immutable and cannot be changed.')
        return val


@admin.register(StudentProfile)
class StudentProfileAdmin(admin.ModelAdmin):
    form = StudentProfileForm
    change_list_template = 'admin/academics/studentprofile_change_list.html'
    list_display = ('user', 'reg_no', 'get_department', 'batch', 'current_section_display', 'status')
    search_fields = ('reg_no', 'user__username', 'user__email')
    # filter by the department through the section->semester->course relation
    list_filter = ('section__batch__course__department', 'batch')
    actions = ('deactivate_students', 'mark_alumni', 'delete_profiles_and_users')

    change_list_template = 'admin/academics/studentprofile/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('sheets/', self.admin_site.admin_view(self.sheets_view), name='academics_studentprofile_sheets'),
            path('sheets/data/', self.admin_site.admin_view(self.sheets_data_view), name='academics_studentprofile_sheets_data'),
            path('sheets/save/', self.admin_site.admin_view(self.sheets_save_view), name='academics_studentprofile_sheets_save'),
        ]
        return custom + urls

    def sheets_view(self, request):
        if not self.has_add_permission(request):
            messages.error(request, 'You do not have permission to use Sheets.')
            return render(request, 'admin/academics/studentprofile/sheets.html', {
                **self.admin_site.each_context(request),
                'title': 'Student Profiles Sheets',
                'sections': [],
                'status_choices': list(PROFILE_STATUS_CHOICES),
                'data_url': reverse('admin:academics_studentprofile_sheets_data'),
                'save_url': reverse('admin:academics_studentprofile_sheets_save'),
            })

        sections_qs = Section.objects.select_related('semester__course__department').all().order_by(
            'semester__course__department__code', 'semester__course__name', 'semester__number', 'name'
        )
        sections = [{'id': s.pk, 'label': str(s)} for s in sections_qs]

        return render(request, 'admin/academics/studentprofile/sheets.html', {
            **self.admin_site.each_context(request),
            'title': 'Student Profiles Sheets',
            'sections': sections,
            'status_choices': list(PROFILE_STATUS_CHOICES),
            'data_url': reverse('admin:academics_studentprofile_sheets_data'),
            'save_url': reverse('admin:academics_studentprofile_sheets_save'),
        })

    def sheets_data_view(self, request):
        if request.method != 'GET':
            return JsonResponse({'detail': 'Method not allowed'}, status=405)

        if not self.has_add_permission(request):
            return JsonResponse({'detail': 'Forbidden'}, status=403)

        # show all users (include existing students). Frontend will mark existing profiles as read-only.
        from accounts.models import User

        qs = User.objects.all().order_by('username').select_related('student_profile__section', 'staff_profile')
        rows = []
        for u in qs:
            sp = getattr(u, 'student_profile', None)
            rows.append({
                'id': u.pk,
                'username': u.username,
                'email': u.email or '',
                'has_student_profile': sp is not None,
                'reg_no': sp.reg_no if sp is not None else '',
                'section_id': sp.section.pk if (sp is not None and sp.section) else '',
                'section_label': str(sp.section) if (sp is not None and sp.section) else '',
                'batch': sp.batch if sp is not None else '',
                'status': sp.status if sp is not None else 'ACTIVE',
            })
        return JsonResponse({'rows': rows})

    def sheets_save_view(self, request):
        if request.method != 'POST':
            return JsonResponse({'detail': 'Method not allowed'}, status=405)

        if not self.has_add_permission(request):
            return JsonResponse({'detail': 'Forbidden'}, status=403)

        try:
            payload = json.loads((request.body or b'{}').decode('utf-8'))
        except Exception:
            return JsonResponse({'detail': 'Invalid JSON body'}, status=400)

        rows = payload.get('rows') or []
        if not isinstance(rows, list) or not rows:
            return JsonResponse({'detail': 'No rows provided'}, status=400)

        from accounts.models import User

        created = 0
        failed = []
        today = timezone.now().date()

        for idx, r in enumerate(rows, start=1):
            try:
                user_id = int(r.get('user_id'))
            except Exception:
                failed.append({'index': idx, 'user_id': r.get('user_id'), 'error': 'Invalid user_id'})
                continue

            reg_no = (r.get('reg_no') or '').strip()
            batch = (r.get('batch') or '').strip()
            status = (r.get('status') or 'ACTIVE').strip().upper()
            section_id = (r.get('section_id') or '').strip()

            if not reg_no:
                failed.append({'index': idx, 'user_id': user_id, 'error': 'reg_no is required'})
                continue

            if status not in {c[0] for c in PROFILE_STATUS_CHOICES}:
                failed.append({'index': idx, 'user_id': user_id, 'error': f'Invalid status: {status}'})
                continue

            try:
                user = User.objects.select_related('student_profile', 'staff_profile').get(pk=user_id)
            except User.DoesNotExist:
                failed.append({'index': idx, 'user_id': user_id, 'error': 'User not found'})
                continue

            if getattr(user, 'student_profile', None) is not None:
                failed.append({'index': idx, 'user_id': user_id, 'error': 'User already has a StudentProfile'})
                continue

            if getattr(user, 'staff_profile', None) is not None:
                failed.append({'index': idx, 'user_id': user_id, 'error': 'User already has a StaffProfile'})
                continue

            section_obj = None
            if section_id:
                try:
                    section_obj = Section.objects.filter(pk=int(section_id)).first()
                except Exception:
                    section_obj = None

            try:
                with transaction.atomic():
                    sp = StudentProfile.objects.create(
                        user=user,
                        reg_no=reg_no,
                        section=section_obj,
                        batch=batch,
                        status=status,
                    )
                    # keep history assignment too (optional, but better for later changes)
                    if section_obj is not None:
                        StudentSectionAssignment.objects.create(student=sp, section=section_obj, start_date=today)
                    created += 1
            except Exception as e:
                failed.append({'index': idx, 'user_id': user_id, 'error': str(e)})

        return JsonResponse({'created': created, 'failed': failed})

    def deactivate_students(self, request, queryset):
        from accounts.services import deactivate_user
        for p in queryset:
            deactivate_user(p.user, profile_status='INACTIVE', reason='deactivated via admin', actor=request.user)
    deactivate_students.short_description = 'Deactivate selected student profiles'

    def mark_alumni(self, request, queryset):
        for p in queryset:
            p.status = 'ALUMNI'
            p.save(update_fields=['status'])
    mark_alumni.short_description = 'Mark selected students as ALUMNI'

    def get_department(self, obj):
        # obj.section -> semester -> course -> department
        sec = getattr(obj, 'section', None)
        if not sec:
            return None
        # section is now batch-wise; resolve course via batch
        batch = getattr(sec, 'batch', None)
        if not batch:
            return None
        course = getattr(batch, 'course', None)
        if not course:
            return None
        dept = getattr(course, 'department', None)
        return dept

    get_department.short_description = 'Department'

    def current_section_display(self, obj):
        sec = obj.current_section
        if sec is None:
            return None
        return getattr(sec, 'name', str(sec))
    current_section_display.short_description = 'Current Section'

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('import-students/', self.admin_site.admin_view(self.import_students), name='academics_student_import'),
            path('download-template/', self.admin_site.admin_view(self.download_template), name='academics_student_template'),
        ]
        return custom + urls

    def import_students(self, request):
        """Admin view to import students from an Excel (.xlsx) file.

        Expected columns (first row header):
          reg_no, username, email, first_name, last_name, batch, section, password

        'batch' can be:
          - Just the batch name (e.g., '2023')
          - Department and batch: 'DEPT_SHORT :: BATCH_NAME' (e.g., 'CSE :: 2023')

        'section' should be in format:
          - 'DEPT_SHORT :: BATCH_NAME :: SECTION_NAME' (e.g., 'CSE :: 2023 :: A')
          - Fallback: 'BATCH_NAME :: SECTION_NAME' (e.g., '2023 :: A')
        """
        if request.method == 'POST':
            uploaded = request.FILES.get('xlsx_file')
            if not uploaded:
                messages.error(request, 'No file uploaded')
                return redirect(request.path)

            User = get_user_model()
            wb = None
            try:
                wb = openpyxl.load_workbook(uploaded)
            except Exception as e:
                messages.error(request, f'Failed to read Excel file: {e}')
                return redirect(request.path)

            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                messages.error(request, 'Excel file contains no data')
                return redirect(request.path)

            headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
            col_index = {name: idx for idx, name in enumerate(headers)}
            required = ['reg_no']
            missing = [r for r in required if r not in col_index]
            if missing:
                messages.error(request, f'Missing required columns: {missing}')
                return redirect(request.path)

            created = 0
            skipped = 0
            errors = []
            for i, row in enumerate(rows[1:], start=2):
                try:
                    reg_no = row[col_index.get('reg_no')] if 'reg_no' in col_index else None
                    if not reg_no:
                        skipped += 1
                        continue
                    reg_no = str(reg_no).strip()
                    username = None
                    if 'username' in col_index:
                        username = row[col_index.get('username')]
                    email = None
                    if 'email' in col_index:
                        email = row[col_index.get('email')]
                    first_name = row[col_index.get('first_name')] if 'first_name' in col_index else ''
                    last_name = row[col_index.get('last_name')] if 'last_name' in col_index else ''
                    batch_name = row[col_index.get('batch')] if 'batch' in col_index else None
                    section_name = row[col_index.get('section')] if 'section' in col_index else None
                    password = row[col_index.get('password')] if 'password' in col_index else None

                    username = str(username).strip() if username else reg_no
                    email = str(email).strip() if email else ''
                    first_name = str(first_name).strip() if first_name else ''
                    last_name = str(last_name).strip() if last_name else ''
                    batch_name = str(batch_name).strip() if batch_name else None
                    section_name = str(section_name).strip() if section_name else None
                    password = str(password) if password else reg_no

                    # Parse batch_name if it's in the format "DEPT :: BATCH_NAME" (extract just the batch name)
                    batch_name_only = batch_name
                    if batch_name and '::' in batch_name:
                        parts = [p.strip() for p in batch_name.split('::')]
                        if len(parts) == 2:
                            # Format is "DEPT :: BATCH_NAME", use the batch name part
                            batch_name_only = parts[1]
                        else:
                            # Unexpected format, use last part
                            batch_name_only = parts[-1]

                    with transaction.atomic():
                        # Prefer de-duplication by reg_no/email (not by name/spacing).
                        email_norm = email.strip().lower() if email else ''
                        existing_sp = StudentProfile.objects.select_related('user').filter(reg_no__iexact=reg_no).first()
                        user = None
                        created_user = False

                        if existing_sp:
                            user = existing_sp.user
                            if email_norm and user.email and user.email.strip().lower() != email_norm:
                                errors.append(f'Row {i}: reg_no {reg_no} has different email ({user.email})')
                                skipped += 1
                                continue
                        else:
                            if email_norm:
                                user = User.objects.filter(email__iexact=email_norm).first()
                            if user:
                                # If user exists by email, ensure reg_no is not mapped to a different profile.
                                sp_existing = StudentProfile.objects.filter(user=user).first()
                                if sp_existing and sp_existing.reg_no.strip().lower() != reg_no.strip().lower():
                                    errors.append(f'Row {i}: email {email} already linked to reg_no {sp_existing.reg_no}')
                                    skipped += 1
                                    continue
                            else:
                                # Check username collision - verify email to determine if truly different student
                                if User.objects.filter(username=username).exists():
                                    existing_user = User.objects.filter(username=username).first()
                                    # Check if different person by comparing emails
                                    if email_norm and existing_user.email and existing_user.email.strip().lower() != email_norm:
                                        # Different email = different person, use fallback username
                                        fallback_username = f"{username}-{reg_no}"
                                        if User.objects.filter(username=fallback_username).exists():
                                            errors.append(f'Row {i}: username collision - {username} and {fallback_username} both exist')
                                            skipped += 1
                                            continue
                                        username = fallback_username
                                    elif not email_norm or not existing_user.email:
                                        # Cannot verify email, treat as different person for safety
                                        fallback_username = f"{username}-{reg_no}"
                                        if User.objects.filter(username=fallback_username).exists():
                                            errors.append(f'Row {i}: username collision - cannot verify email match')
                                            skipped += 1
                                            continue
                                        username = fallback_username
                                    # If emails match, reuse existing user below

                                user, created_user = User.objects.get_or_create(
                                    username=username,
                                    defaults={'email': email, 'first_name': first_name, 'last_name': last_name}
                                )

                        if user and not created_user:
                            updated = False
                            if email and (not user.email):
                                user.email = email; updated = True
                            if first_name and (not user.first_name):
                                user.first_name = first_name; updated = True
                            if last_name and (not user.last_name):
                                user.last_name = last_name; updated = True
                            if updated:
                                user.save()
                        if user and (created_user or (not user.has_usable_password())):
                            user.set_password(password)
                            user.save()

                        sp, sp_created = StudentProfile.objects.get_or_create(
                            reg_no=reg_no,
                            defaults={'user': user, 'batch': batch_name_only or ''}
                        )
                        if not sp_created:
                            if sp.user != user:
                                errors.append(f'Row {i}: reg_no {reg_no} already assigned to a different user')
                                skipped += 1
                                continue
                            if batch_name_only and (not sp.batch):
                                sp.batch = batch_name_only
                                sp.save(update_fields=['batch'])

                        # section may be provided as one of:
                        #  - 'Batch :: Section'
                        #  - 'DEPT_SHORT :: Batch :: Section'  (new template format)
                        sec = None
                        if section_name:
                            if '::' in section_name:
                                parts = [p.strip() for p in section_name.split('::')]
                                if len(parts) == 3:
                                    # Dept short, Batch, Section
                                    dept_short, bpart, spart = parts
                                    sec = Section.objects.filter(
                                        name=spart,
                                        batch__name=bpart,
                                        batch__course__department__short_name__iexact=dept_short,
                                    ).select_related('batch').first()
                                elif len(parts) == 2:
                                    bpart, spart = parts
                                    sec = Section.objects.filter(name=spart, batch__name=bpart).select_related('batch').first()
                                else:
                                    # unexpected format; try best-effort match by section name
                                    sec = Section.objects.filter(name=parts[-1]).select_related('batch').first()
                            else:
                                # fallback: if batch provided, try match by both
                                if batch_name_only:
                                    sec = Section.objects.filter(name=section_name, batch__name=batch_name_only).select_related('batch').first()
                                else:
                                    sec = Section.objects.filter(name=section_name).select_related('batch').first()
                        if sec:
                            sp.section = sec
                            sp.save(update_fields=['section'])
                            # create or update a StudentSectionAssignment for this student
                            try:
                                today = timezone.now().date()
                                existing = StudentSectionAssignment.objects.filter(student=sp, end_date__isnull=True).first()
                                if not existing:
                                    StudentSectionAssignment.objects.create(student=sp, section=sec, start_date=today)
                            except Exception:
                                pass

                        # ensure user is in 'students' group
                        try:
                            grp, _ = Group.objects.get_or_create(name='students')
                            if grp not in user.groups.all():
                                user.groups.add(grp)
                        except Exception:
                            pass

                        # ensure user has logical Role 'STUDENT'
                        try:
                            role_obj, _ = Role.objects.get_or_create(name='STUDENT')
                            # add via m2m so validation hooks run
                            if role_obj not in user.roles.all():
                                user.roles.add(role_obj)
                        except ValidationError as e:
                            # validation may fail if profile not yet attached; ignore and continue
                            pass
                        except Exception:
                            pass

                        created += 1
                except Exception as e:
                    errors.append(f'Row {i}: {e}')

            msg = f'Import completed: created_or_updated={created} skipped={skipped} errors={len(errors)}'
            if errors:
                for err in errors[:10]:
                    messages.error(request, err)
            messages.success(request, msg)
            return redirect(request.path)

        context = dict(self.admin_site.each_context(request))
        context.update({
            'title': 'Import Students from Excel',
        })
        return render(request, 'admin/academics/import_students.html', context)

    def download_template(self, request):
        # generate a simple Excel template with headers and a sample row
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'import'
        headers = ['reg_no', 'username', 'email', 'first_name', 'last_name', 'batch', 'section', 'password']
        ws.append(headers)
        # sample row uses composite format: batch='DEPT :: BATCH_NAME', section='DEPT :: BATCH :: SECTION'
        ws.append(['REG2026001', 'reg2026001', 'student@example.edu', 'First', 'Last', 'CSE :: 2026', 'CSE :: 2026 :: A', 'changeme'])

        # prepare lookup lists for batches and composite sections
        # Fetch batches with their full course name
        batches_qs = Batch.objects.select_related('course__department').order_by('course__department__short_name', 'name')
        batches = []
        for b in batches_qs:
            dept_short = b.course.department.short_name if b.course and b.course.department else ''
            # format: "DEPT_SHORT :: BATCH_NAME" (e.g., "CSE :: 2023")
            batches.append(f"{dept_short} :: {b.name}")
        
        # build composite 'DEPT :: Batch :: Section' entries (department short_name)
        sections_qs = Section.objects.select_related('batch__course__department').order_by('batch__course__department__short_name', 'batch__name', 'name')
        sections = []
        for s in sections_qs:
            dept_short = s.batch.course.department.short_name if s.batch and s.batch.course and s.batch.course.department else ''
            # format: "DEPT_SHORT :: BATCH :: SECTION" (e.g., "CSE :: 2023 :: A")
            sections.append(f"{dept_short} :: {s.batch.name} :: {s.name}")

        # create hidden sheet with lists
        lists = wb.create_sheet(title='lists')
        for i, b in enumerate(batches, start=1):
            lists.cell(row=i, column=1, value=b)
        for i, s in enumerate(sections, start=1):
            lists.cell(row=i, column=2, value=s)
        lists.sheet_state = 'hidden'

        # add data validation for batch (column F) and section (column G)
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            wb_refs = len(batches) or 1
            ws_batch_range = f"=lists!$A$1:$A${wb_refs}"
            dv_batch = DataValidation(type="list", formula1=ws_batch_range, allow_blank=True)
            ws.add_data_validation(dv_batch)
            dv_batch.add("F2:F500")

            ws_secs = len(sections) or 1
            ws_section_range = f"=lists!$B$1:$B${ws_secs}"
            dv_section = DataValidation(type="list", formula1=ws_section_range, allow_blank=True)
            ws.add_data_validation(dv_section)
            dv_section.add("G2:G500")
        except Exception:
            # if validation can't be added, continue — template still useful
            pass

        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        return resp


@admin.register(StaffProfile)
class StaffProfileAdmin(admin.ModelAdmin):
    form = StaffProfileForm
    change_list_template = 'admin/academics/staffprofile_change_list.html'
    list_display = ('user', 'staff_id', 'current_department_display', 'designation', 'status')
    search_fields = ('staff_id', 'user__username', 'user__email')
    list_filter = ('department', 'designation')

    def has_delete_permission(self, request, obj=None):
        return False

    actions = ('deactivate_staff', 'mark_resigned', 'delete_profiles_and_users')

    def deactivate_staff(self, request, queryset):
        from accounts.services import deactivate_user
        for p in queryset:
            deactivate_user(p.user, profile_status='INACTIVE', reason='deactivated via admin', actor=request.user)
    deactivate_staff.short_description = 'Deactivate selected staff profiles'

    def mark_resigned(self, request, queryset):
        for p in queryset:
            p.status = 'RESIGNED'
            p.save(update_fields=['status'])
    mark_resigned.short_description = 'Mark selected staff as RESIGNED'

    def delete_profiles_and_users(self, request, queryset):
        from django.db import transaction
        for p in queryset:
            try:
                with transaction.atomic():
                    user = p.user
                    p.delete()
                    if user:
                        user.delete()
            except Exception:
                pass
    delete_profiles_and_users.short_description = 'Permanently delete selected profiles and their users'

    def current_department_display(self, obj):
        dept = obj.current_department
        if dept is None:
            return None
        return getattr(dept, 'code', str(dept))
    current_department_display.short_description = 'Current Department'

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('import-staff/', self.admin_site.admin_view(self.import_staff), name='academics_staff_import'),
            path('download-template-staff/', self.admin_site.admin_view(self.download_template_staff), name='academics_staff_template'),
        ]
        return custom + urls

    def import_staff(self, request):
        """Admin view to import staff profiles from an Excel (.xlsx) file.

        Expected columns (first row header):
          staff_id, username, email, first_name, last_name, department, designation, password
        """
        if request.method == 'POST':
            uploaded = request.FILES.get('xlsx_file')
            if not uploaded:
                messages.error(request, 'No file uploaded')
                return redirect(request.path)

            User = get_user_model()
            wb = None
            try:
                wb = openpyxl.load_workbook(uploaded)
            except Exception as e:
                messages.error(request, f'Failed to read Excel file: {e}')
                return redirect(request.path)

            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                messages.error(request, 'Excel file contains no data')
                return redirect(request.path)

            headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
            col_index = {name: idx for idx, name in enumerate(headers)}
            required = ['staff_id']
            missing = [r for r in required if r not in col_index]
            if missing:
                messages.error(request, f'Missing required columns: {missing}')
                return redirect(request.path)

            created = 0
            skipped = 0
            errors = []
            for i, row in enumerate(rows[1:], start=2):
                try:
                    staff_id = row[col_index.get('staff_id')] if 'staff_id' in col_index else None
                    if not staff_id:
                        skipped += 1
                        continue
                    staff_id = str(staff_id).strip()
                    username = None
                    if 'username' in col_index:
                        username = row[col_index.get('username')]
                    email = None
                    if 'email' in col_index:
                        email = row[col_index.get('email')]
                    first_name = row[col_index.get('first_name')] if 'first_name' in col_index else ''
                    last_name = row[col_index.get('last_name')] if 'last_name' in col_index else ''
                    department_name = row[col_index.get('department')] if 'department' in col_index else None
                    designation = row[col_index.get('designation')] if 'designation' in col_index else ''
                    password = row[col_index.get('password')] if 'password' in col_index else None
                    role_value = row[col_index.get('role')] if 'role' in col_index else None

                    username = str(username).strip() if username else staff_id
                    email = str(email).strip() if email else ''
                    first_name = str(first_name).strip() if first_name else ''
                    last_name = str(last_name).strip() if last_name else ''
                    department_name = str(department_name).strip() if department_name else None
                    designation = str(designation).strip() if designation else ''
                    password = str(password) if password else staff_id

                    with transaction.atomic():
                        user, created_user = User.objects.get_or_create(username=username, defaults={'email': email, 'first_name': first_name, 'last_name': last_name})
                        if not created_user:
                            updated = False
                            if email and (not user.email):
                                user.email = email; updated = True
                            if first_name and (not user.first_name):
                                user.first_name = first_name; updated = True
                            if last_name and (not user.last_name):
                                user.last_name = last_name; updated = True
                            if updated:
                                user.save()
                        if created_user or (not user.has_usable_password()):
                            user.set_password(password)
                            user.save()

                        sp, sp_created = StaffProfile.objects.get_or_create(user=user, defaults={'staff_id': staff_id, 'designation': designation})
                        if not sp_created:
                            if sp.staff_id != staff_id:
                                errors.append(f'Row {i}: existing user {username} has different staff_id ({sp.staff_id})')
                                skipped += 1
                                continue

                        # resolve department by code or name
                        if department_name:
                            dept = Department.objects.filter(models.Q(code__iexact=department_name) | models.Q(name__iexact=department_name) | models.Q(short_name__iexact=department_name)).first()
                            if dept:
                                sp.department = dept
                                sp.save(update_fields=['department'])

                        # If a logical role was provided, handle role assignment
                        try:
                            rv = str(role_value).strip().upper() if role_value is not None else ''
                        except Exception:
                            rv = ''
                        
                        if rv:
                            try:
                                # Ensure the role exists in the system
                                logical_role, _ = Role.objects.get_or_create(name=rv)
                                
                                # Add the logical role to user if not already assigned
                                if logical_role not in user.roles.all():
                                    user.roles.add(logical_role)
                                    
                                # Handle department-specific roles (HOD, AHOD) - update DepartmentRole table
                                if rv in ('HOD', 'AHOD'):
                                    # Re-resolve department if not already resolved
                                    if not dept and department_name:
                                        dept = Department.objects.filter(
                                            models.Q(code__iexact=department_name) | 
                                            models.Q(name__iexact=department_name) | 
                                            models.Q(short_name__iexact=department_name)
                                        ).first()
                                    
                                    if dept:
                                        # Get active academic year
                                        ay = AcademicYear.objects.filter(is_active=True).first() or AcademicYear.objects.order_by('-id').first()
                                        if ay:
                                            # For HOD role, deactivate existing HOD for the department/year
                                            if rv == 'HOD':
                                                DepartmentRole.objects.filter(
                                                    department=dept, 
                                                    academic_year=ay, 
                                                    role='HOD', 
                                                    is_active=True
                                                ).update(is_active=False)
                                            
                                            # Create or update the department role record
                                            dept_role, created = DepartmentRole.objects.get_or_create(
                                                department=dept,
                                                staff=sp,
                                                role=rv,
                                                academic_year=ay,
                                                defaults={'is_active': True}
                                            )
                                            if not created:
                                                dept_role.is_active = True
                                                dept_role.save()
                                    else:
                                        errors.append(f'Row {i}: Department required for role {rv} but not found: {department_name}')
                                            
                                # Handle advisor role assignment
                                elif rv == 'ADVISOR':
                                    # Advisors are handled through section assignments, not department roles
                                    pass
                                    
                                # Handle other institutional roles (IQAC, PRINCIPAL, etc.)
                                # These don't need department role entries, just the role assignment
                                
                            except ValidationError as ve:
                                errors.append(f'Row {i}: failed to assign role {rv}: {ve}')
                            except Exception as e:
                                errors.append(f'Row {i}: error assigning role {rv}: {e}')

                        if designation and (not sp.designation):
                            sp.designation = designation
                            sp.save(update_fields=['designation'])

                        # ensure user is in 'staff' group
                        try:
                            grp, _ = Group.objects.get_or_create(name='staff')
                            if grp not in user.groups.all():
                                user.groups.add(grp)
                        except Exception:
                            pass

                        # ensure logical Role 'STAFF' - always assign STAFF role by default
                        try:
                            staff_role, _ = Role.objects.get_or_create(name='STAFF')
                            if staff_role not in user.roles.all():
                                user.roles.add(staff_role)
                        except ValidationError as ve:
                            errors.append(f'Row {i}: failed to assign role STAFF: {ve}')
                        except Exception as e:
                            errors.append(f'Row {i}: error assigning role STAFF: {e}')

                        created += 1
                except Exception as e:
                    errors.append(f'Row {i}: {e}')

            msg = f'Staff import completed: created_or_updated={created} skipped={skipped} errors={len(errors)}'
            if errors:
                for err in errors[:10]:
                    messages.error(request, err)
            messages.success(request, msg)
            return redirect(request.path)

        context = dict(self.admin_site.each_context(request))
        context.update({'title': 'Import Staff from Excel'})
        return render(request, 'admin/academics/import_staff.html', context)

    def download_template_staff(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'import'
        headers = ['staff_id', 'username', 'email', 'first_name', 'last_name', 'department', 'designation', 'role', 'password']
        ws.append(headers)
        ws.append(['STAFF001', 'jsmith', 'jsmith@example.edu', 'John', 'Smith', 'CSE', 'Lecturer', 'STAFF', 'changeme'])

        # departments list (use short_name for dropdown values)
        depts = list(Department.objects.values_list('short_name', flat=True))
        # available roles
        from accounts.models import Role
        roles = list(Role.objects.values_list('name', flat=True))
        if not roles:  # fallback if no roles exist
            roles = ['STAFF', 'HOD', 'AHOD', 'ADVISOR', 'IQAC', 'PRINCIPAL']
            
        lists = wb.create_sheet(title='lists')
        for i, d in enumerate(depts, start=1):
            lists.cell(row=i, column=1, value=d)
        for i, r in enumerate(roles, start=1):
            lists.cell(row=i, column=2, value=r)
        lists.sheet_state = 'hidden'

        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            # Department validation (column F)
            dv_dept = DataValidation(type='list', formula1=f"=lists!$A$1:$A${len(depts) or 1}", allow_blank=True)
            ws.add_data_validation(dv_dept)
            dv_dept.add('F2:F500')
            # Role validation (column H)
            dv_role = DataValidation(type='list', formula1=f"=lists!$B$1:$B${len(roles) or 1}", allow_blank=True)
            ws.add_data_validation(dv_role)
            dv_role.add('H2:H500')
        except Exception:
            pass

        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="staff_import_template.xlsx"'
        return resp


@admin.register(StudentSectionAssignment)
class StudentSectionAssignmentAdmin(admin.ModelAdmin):
    list_display = ('student', 'section', 'start_date', 'end_date', 'created_at')
    search_fields = ('student__reg_no', 'student__user__username')
    list_filter = ('section',)
    actions = ('end_assignments',)

    def end_assignments(self, request, queryset):
        from django.utils import timezone
        today = timezone.now().date()
        for a in queryset:
            if a.end_date is None:
                a.end_date = today
                a.save(update_fields=['end_date'])
    end_assignments.short_description = 'End selected student section assignments (set end_date to today)'


@admin.register(StaffDepartmentAssignment)
class StaffDepartmentAssignmentAdmin(admin.ModelAdmin):
    list_display = ('staff', 'department', 'start_date', 'end_date', 'created_at')
    search_fields = ('staff__staff_id', 'staff__user__username')
    list_filter = ('department',)
    actions = ('end_assignments',)

    def end_assignments(self, request, queryset):
        from django.utils import timezone
        today = timezone.now().date()
        for a in queryset:
            if a.end_date is None:
                a.end_date = today
                a.save(update_fields=['end_date'])
    end_assignments.short_description = 'End selected staff department assignments (set end_date to today)'


@admin.register(RoleAssignment)
class RoleAssignmentAdmin(admin.ModelAdmin):
    list_display = ('staff', 'role_name', 'start_date', 'end_date', 'created_at')
    search_fields = ('staff__staff_id', 'role_name', 'staff__user__username')
    list_filter = ('role_name',)
    actions = ('end_assignments',)

    def end_assignments(self, request, queryset):
        from django.utils import timezone
        today = timezone.now().date()
        for a in queryset:
            if a.end_date is None:
                a.end_date = today
                a.save(update_fields=['end_date'])
    end_assignments.short_description = 'End selected role assignments (set end_date to today)'



@admin.register(AcademicYear)
class AcademicYearAdmin(admin.ModelAdmin):
    list_display = ('name', 'parity', 'is_active')
    list_editable = ('is_active',)
    list_filter = ('parity', 'is_active')
    search_fields = ('name',)


@admin.register(Department)
class DepartmentAdmin(admin.ModelAdmin):
    list_display = ('code', 'short_name', 'name')
    search_fields = ('code', 'short_name', 'name')


@admin.register(Program)
class ProgramAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)


@admin.register(Course)
class CourseAdmin(admin.ModelAdmin):
    list_display = ('name', 'department', 'program')
    search_fields = ('name',)
    list_filter = ('department', 'program')


@admin.register(Semester)
class SemesterAdmin(admin.ModelAdmin):
    # `course` was removed from Semester; show the semester number only.
    list_display = ('number',)
    search_fields = ()
    list_filter = ('number',)


@admin.register(Section)
class SectionAdmin(admin.ModelAdmin):
    list_display = ('batch', 'name', 'semester')
    search_fields = ('name',)
    list_filter = ('batch',)


@admin.register(Subject)
class SubjectAdmin(admin.ModelAdmin):
    list_display = ('code', 'name', 'course', 'semester')
    search_fields = ('code', 'name')
    list_filter = ('course', 'semester')


@admin.register(Batch)
class BatchAdmin(admin.ModelAdmin):
    list_display = ('name', 'course', 'regulation_display', 'start_year', 'end_year')
    search_fields = ('name', 'course__name', 'regulation__code', 'regulation__name')
    list_filter = ('course', 'regulation')
    raw_id_fields = ('regulation',)
    
    def regulation_display(self, obj):
        reg = getattr(obj, 'regulation', None)
        return getattr(reg, 'code', '—') if reg else '—'
    regulation_display.short_description = 'Regulation'


@admin.register(TeachingAssignment)
class TeachingAssignmentAdmin(admin.ModelAdmin):
    list_display = ('staff', 'subject_display', 'section', 'academic_year', 'is_active')
    search_fields = (
        'staff__staff_id', 'staff__user__username', 'subject__code', 'subject__name', 'section__name'
    )
    list_filter = ('academic_year', 'is_active', 'section__batch__course__department')
    raw_id_fields = ('staff', 'curriculum_row', 'section', 'academic_year')

    class TeachingAssignmentForm(forms.ModelForm):
        class Meta:
            model = TeachingAssignment
            fields = '__all__'

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            try:
                from curriculum.models import ElectiveSubject

                class ElectiveChoiceField(forms.ModelChoiceField):
                    def label_from_instance(self, obj):
                        try:
                            parent = getattr(obj, 'parent', None)
                            dept = getattr(obj, 'department', None)
                            label_parts = []
                            # elective option code/name
                            opt = (getattr(obj, 'course_code', None) or getattr(obj, 'course_name', None) or str(obj.pk))
                            label_parts.append(opt)
                            # parent curriculum row (if present)
                            if parent is not None:
                                ptxt = (getattr(parent, 'course_code', None) or getattr(parent, 'course_name', None))
                                if ptxt:
                                    label_parts.append(f"(Parent: {ptxt})")
                            # department
                            if dept is not None:
                                dtxt = getattr(dept, 'code', None) or getattr(dept, 'name', None)
                                if dtxt:
                                    label_parts.append(f"[{dtxt}]")
                            return ' '.join(label_parts)
                        except Exception:
                            return str(obj)

                self.fields['elective_subject'] = ElectiveChoiceField(
                    queryset=ElectiveSubject.objects.select_related('parent', 'department').all().order_by('department__code', 'parent__course_code', 'course_code'),
                    required=False,
                )
            except Exception:
                # If elective model is unavailable, leave default field
                pass

        def clean(self):
            cleaned = super().clean()
            elective = cleaned.get('elective_subject')
            section = cleaned.get('section')
            if elective and section:
                try:
                    sec_dept = getattr(getattr(section, 'batch', None), 'course', None)
                    sec_dept = getattr(sec_dept, 'department', None)
                except Exception:
                    sec_dept = None

                # resolve elective's department
                try:
                    elect_dept = getattr(elective, 'department', None)
                except Exception:
                    elect_dept = None

                # If elective's department doesn't match the section's department, try to find an option for the section's dept
                if sec_dept and elect_dept and sec_dept.pk != elect_dept.pk:
                    try:
                        from curriculum.models import ElectiveSubject
                        parent = getattr(elective, 'parent', None)
                        if parent is not None:
                            alt = ElectiveSubject.objects.filter(parent=parent, department=sec_dept).first()
                            if alt:
                                cleaned['elective_subject'] = alt
                                return cleaned
                    except Exception:
                        pass

                    # no matching elective found; raise validation error to avoid cross-dept assignment
                    raise ValidationError({'elective_subject': 'Selected elective does not belong to the same department as the chosen section. Choose an elective option for the section\'s department.'})

            return cleaned

    form = TeachingAssignmentForm

    def get_form(self, request, obj=None, **kwargs):
        # Obtain default form
        form = super().get_form(request, obj, **kwargs)
        try:
            from curriculum.models import ElectiveSubject
            # Superusers see all elective options
            if request.user.is_superuser:
                qs = ElectiveSubject.objects.select_related('parent', 'department').all().order_by('department__code', 'parent__course_code', 'course_code')
            else:
                staff_profile = getattr(request.user, 'staff_profile', None)
                dept_ids = []
                if staff_profile is not None:
                    # Departments where the staff has DepartmentRole (e.g. HOD) and is active
                    from .models import DepartmentRole
                    roles = DepartmentRole.objects.filter(staff=staff_profile, is_active=True).select_related('department')
                    for r in roles:
                        if getattr(r, 'department', None):
                            dept_ids.append(r.department.pk)
                    # fallback to staff_profile.department if set
                    sp_dept = getattr(staff_profile, 'department', None)
                    if sp_dept is not None and sp_dept.pk not in dept_ids:
                        dept_ids.append(sp_dept.pk)

                if dept_ids:
                    qs = ElectiveSubject.objects.select_related('parent', 'department').filter(department__pk__in=dept_ids).order_by('department__code', 'parent__course_code', 'course_code')
                else:
                    # no known department context: restrict nothing but keep ordering
                    qs = ElectiveSubject.objects.select_related('parent', 'department').all().order_by('department__code', 'parent__course_code', 'course_code')

            if 'elective_subject' in form.base_fields:
                form.base_fields['elective_subject'].queryset = qs
        except Exception:
            pass
        return form

    def subject_display(self, obj):
        try:
            # Prefer elective_subject when present
            if getattr(obj, 'elective_subject', None):
                es = obj.elective_subject
                # If this elective option has a parent curriculum row, show parent + dept and list all options
                parent = getattr(es, 'parent', None)
                dept = getattr(es, 'department', None)
                parts = []
                if parent:
                    parts.append(f"Parent: {getattr(parent, 'course_code', '') or getattr(parent, 'course_name', '')}".strip())
                if dept:
                    # show department code if available, else name
                    parts.append(f"Dept: {getattr(dept, 'code', None) or getattr(dept, 'name', '')}")

                # list all elective options belonging to the parent (show code or name)
                try:
                    options = []
                    if parent is not None:
                        opts_qs = getattr(parent, 'elective_options', None)
                        if opts_qs is None:
                            from curriculum.models import ElectiveSubject
                            opts_qs = ElectiveSubject.objects.filter(parent=parent)
                        for o in opts_qs.all():
                            options.append(getattr(o, 'course_code', None) or getattr(o, 'course_name', None) or str(o.pk))
                    else:
                        # fallback to showing the single elective option
                        options.append(getattr(es, 'course_code', None) or getattr(es, 'course_name', None) or str(es.pk))
                    if options:
                        parts.append('Options: ' + ', '.join(options))
                except Exception:
                    # best-effort; ignore option listing failures
                    parts.append(f"Option: {getattr(es, 'course_code', '') or getattr(es, 'course_name', '')}".strip())

                return ' | '.join([p for p in parts if p])

            from curriculum.models import CurriculumDepartment
            # prefer explicit curriculum_row if set
            if getattr(obj, 'curriculum_row', None):
                cr = obj.curriculum_row
                return f"{cr.course_code or ''} - {cr.course_name or ''}".strip(' -')

            # if subject exists, show its name/code
            if getattr(obj, 'subject', None):
                try:
                    return getattr(obj.subject, 'name') or getattr(obj.subject, 'code') or str(obj.subject)
                except Exception:
                    return str(obj.subject)

            # fallback: return first curriculum row for section's department
            dept = None
            try:
                if getattr(obj, 'section', None):
                    dept = obj.section.batch.course.department
                # If no section (department-wide elective), try elective parent dept
                elif getattr(obj, 'elective_subject', None):
                    dept = getattr(getattr(obj.elective_subject, 'parent', None), 'department', None)
            except Exception:
                dept = None
            if dept is not None:
                row = CurriculumDepartment.objects.filter(department=dept).first()
            else:
                row = CurriculumDepartment.objects.first()
            if row:
                return f"{row.course_code or ''} - {row.course_name or ''}".strip(' -')
        except Exception:
            pass
        return 'No subject'

    subject_display.short_description = 'Subject (Curriculum)'


@admin.register(StudentMentorMap)
class StudentMentorMapAdmin(admin.ModelAdmin):
    list_display = ('student', 'mentor', 'is_active')
    list_filter = ('is_active', 'student__section__batch__course__department')
    search_fields = ('student__reg_no', 'mentor__staff_id', 'mentor__user__username')
    raw_id_fields = ('student', 'mentor')


@admin.register(SectionAdvisor)
class SectionAdvisorAdmin(admin.ModelAdmin):
    list_display = ('section', 'advisor', 'academic_year', 'is_active')
    list_filter = ('academic_year', 'is_active', 'section__batch__course__department')
    search_fields = ('section__name', 'advisor__staff_id', 'advisor__user__username')
    raw_id_fields = ('section', 'advisor', 'academic_year')


@admin.register(DepartmentRole)
class DepartmentRoleAdmin(admin.ModelAdmin):
    list_display = ('department', 'role', 'staff', 'academic_year', 'is_active')
    list_filter = ('academic_year', 'is_active', 'department', 'role')
    search_fields = ('staff__staff_id', 'staff__user__username')
    raw_id_fields = ('staff', 'academic_year')


# DayAttendance admin removed along with models


@admin.register(StudentSubjectBatch)
class StudentSubjectBatchAdmin(admin.ModelAdmin):
    list_display = ('name', 'staff', 'academic_year', 'curriculum_row', 'is_active', 'created_at')
    search_fields = ('name', 'staff__staff_id', 'staff__user__username')
    list_filter = ('academic_year', 'is_active')
    raw_id_fields = ('staff', 'curriculum_row')



class PeriodAttendanceRecordInline(admin.TabularInline):
    model = PeriodAttendanceRecord
    extra = 0
    readonly_fields = ('marked_at',)


@admin.register(PeriodAttendanceSession)
class PeriodAttendanceSessionAdmin(admin.ModelAdmin):
    list_display = ('section', 'period', 'date', 'teaching_assignment', 'timetable_assignment', 'created_by', 'is_locked', 'created_at')
    list_filter = ('date', 'is_locked')
    raw_id_fields = ('section', 'period', 'teaching_assignment', 'timetable_assignment', 'created_by')
    inlines = (PeriodAttendanceRecordInline,)


@admin.register(PeriodAttendanceRecord)
class PeriodAttendanceRecordAdmin(admin.ModelAdmin):
    list_display = ('session', 'student', 'status', 'marked_at', 'marked_by')
    search_fields = ('student__reg_no', 'student__user__username')
    list_filter = ('status',)
    raw_id_fields = ('session', 'student', 'marked_by')



@admin.register(AttendanceUnlockRequest)
class AttendanceUnlockRequestAdmin(admin.ModelAdmin):
    list_display = ('id', 'session', 'requested_by', 'requested_at', 'status', 'reviewed_by', 'reviewed_at')
    list_filter = ('status', 'requested_at', 'reviewed_at')
    search_fields = ('requested_by__user__username', 'requested_by__staff_id', 'reviewed_by__user__username')
    raw_id_fields = ('session', 'requested_by', 'reviewed_by')
    readonly_fields = ('requested_at', 'reviewed_at')
    date_hierarchy = 'requested_at'

class SpecialCourseAssessmentEditRequestInline(admin.TabularInline):
    model = SpecialCourseAssessmentEditRequest
    extra = 0
    fields = ('requested_by', 'status', 'requested_at', 'reviewed_by', 'reviewed_at', 'can_edit_until', 'used_at')
    readonly_fields = ('requested_at', 'reviewed_at', 'used_at')
    raw_id_fields = ('requested_by', 'reviewed_by')


class SpecialCourseAssessmentSelectionAdminForm(forms.ModelForm):
    enabled_assessments = forms.MultipleChoiceField(
        choices=SPECIAL_ASSESSMENT_CHOICES,
        required=False,
        widget=forms.CheckboxSelectMultiple,
        help_text='Enabled assessment tables for this SPECIAL course (saved globally per academic year).',
    )

    class Meta:
        model = SpecialCourseAssessmentSelection
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        inst = getattr(self, 'instance', None)
        if inst and getattr(inst, 'enabled_assessments', None):
            self.fields['enabled_assessments'].initial = list(inst.enabled_assessments or [])

    def clean_enabled_assessments(self):
        vals = self.cleaned_data.get('enabled_assessments') or []
        return [str(v).strip().lower() for v in vals if str(v).strip()]

    def clean(self):
        cleaned = super().clean()
        ea = cleaned.get('enabled_assessments') or []
        if not ea:
            self.add_error('enabled_assessments', 'Select at least one assessment for Special courses.')
        return cleaned


@admin.register(SpecialCourseAssessmentSelection)
class SpecialCourseAssessmentSelectionAdmin(admin.ModelAdmin):
    form = SpecialCourseAssessmentSelectionAdminForm
    list_display = (
        'id',
        'master_course_code',
        'master_course_name',
        'department_code',
        'academic_year',
        'locked',
        'enabled_assessments_display',
        'updated_at',
    )
    list_filter = (
        'academic_year',
        'locked',
        'curriculum_row__department',
        'curriculum_row__master__regulation',
        'curriculum_row__master__semester',
    )
    search_fields = (
        'curriculum_row__course_code',
        'curriculum_row__course_name',
        'curriculum_row__master__course_code',
        'curriculum_row__master__course_name',
        'curriculum_row__department__code',
    )
    raw_id_fields = ('curriculum_row', 'academic_year', 'created_by')
    readonly_fields = ('created_at', 'updated_at')
    inlines = (SpecialCourseAssessmentEditRequestInline,)

    def master_course_code(self, obj):
        try:
            m = getattr(getattr(obj, 'curriculum_row', None), 'master', None)
            return getattr(m, 'course_code', None) or getattr(obj.curriculum_row, 'course_code', None)
        except Exception:
            return None

    master_course_code.short_description = 'Course Code'

    def master_course_name(self, obj):
        try:
            m = getattr(getattr(obj, 'curriculum_row', None), 'master', None)
            return getattr(m, 'course_name', None) or getattr(obj.curriculum_row, 'course_name', None)
        except Exception:
            return None

    master_course_name.short_description = 'Course Name'

    def department_code(self, obj):
        try:
            return getattr(getattr(obj.curriculum_row, 'department', None), 'code', None)
        except Exception:
            return None

    department_code.short_description = 'Dept'

    def enabled_assessments_display(self, obj):
        vals = getattr(obj, 'enabled_assessments', None) or []
        if not vals:
            return '(none)'
        order = [k for k, _ in SPECIAL_ASSESSMENT_CHOICES]
        label_map = {k: v for k, v in SPECIAL_ASSESSMENT_CHOICES}
        normalized = {str(x).strip().lower() for x in vals}
        display = [label_map.get(k, k) for k in order if k in normalized]
        return ', '.join(display) if display else ', '.join(sorted(normalized))

    enabled_assessments_display.short_description = 'Assessments'

    def save_model(self, request, obj, form, change):
        """Keep SPECIAL selection global across all departments of the same master.

        The faculty UI treats enabled assessments as a *global* selection for a
        SPECIAL course (master) per academic year. Editing a single row in admin
        should propagate to all department rows to avoid mismatches.
        """
        actor_staff = getattr(getattr(request, 'user', None), 'staff_profile', None)
        with transaction.atomic():
            if not getattr(obj, 'created_by', None) and actor_staff:
                obj.created_by = actor_staff
            super().save_model(request, obj, form, change)

            master_id = None
            try:
                master_id = obj.curriculum_row.master_id
            except Exception:
                master_id = None

            if not master_id:
                return

            try:
                from curriculum.models import CurriculumDepartment

                dept_rows = CurriculumDepartment.objects.filter(master_id=master_id)
            except Exception:
                dept_rows = []

            for row in dept_rows:
                if not row:
                    continue
                sel, created = SpecialCourseAssessmentSelection.objects.get_or_create(
                    curriculum_row=row,
                    academic_year=obj.academic_year,
                    defaults={
                        'enabled_assessments': list(obj.enabled_assessments or []),
                        'locked': bool(obj.locked),
                        'created_by': obj.created_by,
                    },
                )
                if created:
                    continue
                if sel.enabled_assessments != obj.enabled_assessments or sel.locked != obj.locked:
                    sel.enabled_assessments = list(obj.enabled_assessments or [])
                    sel.locked = bool(obj.locked)
                    sel.save(update_fields=['enabled_assessments', 'locked', 'updated_at'])


@admin.register(SpecialCourseAssessmentEditRequest)
class SpecialCourseAssessmentEditRequestAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'course_code',
        'academic_year',
        'requested_by',
        'status',
        'requested_at',
        'can_edit_until',
        'used_at',
    )
    list_filter = ('status', 'selection__academic_year')
    search_fields = (
        'requested_by__staff_id',
        'requested_by__user__username',
        'selection__curriculum_row__course_code',
        'selection__curriculum_row__course_name',
        'selection__curriculum_row__master__course_code',
        'selection__curriculum_row__master__course_name',
    )
    raw_id_fields = ('selection', 'requested_by', 'reviewed_by')
    readonly_fields = ('requested_at',)

    def academic_year(self, obj):
        try:
            return obj.selection.academic_year
        except Exception:
            return None

    academic_year.short_description = 'Academic Year'

    def course_code(self, obj):
        try:
            cr = obj.selection.curriculum_row
            m = getattr(cr, 'master', None)
            return getattr(m, 'course_code', None) or getattr(cr, 'course_code', None)
        except Exception:
            return None

    course_code.short_description = 'Course Code'
