import ast
import calendar
import csv
import io
import math
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.db import IntegrityError, connection
from django.db.models import Q
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from academics.models import Department
from staff_requests.models import StaffRequest
from staff_requests.permissions import IsAdminOrHR

from .models import (
    SalaryBankDeclaration,
    SalaryDeductionType,
    SalaryEMIPlan,
    SalaryEarnType,
    SalaryFormulaConfig,
    SalaryMonthPublish,
    SalaryMonthlyInput,
    SalaryPFConfig,
    SalaryPublishedReceipt,
    StaffSalaryDeclaration,
)


DEFAULT_FORMULAS = {
    'working_days': 'days_in_month - lop_days',
    'lop_amount': '((basic_salary + allowance) / days_in_month) * lop_days if days_in_month > 0 else 0',
    'gross_salary': '(basic_salary + allowance) - lop_amount',
    'total_salary': 'gross_salary + total_earn',
    'net_salary': 'total_salary + pf_amount - od_new - total_deduction - others',
}


class SafeFormulaEvaluator:
    ALLOWED_NODES = {
        ast.Expression,
        ast.BinOp,
        ast.UnaryOp,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.Mod,
        ast.Pow,
        ast.USub,
        ast.UAdd,
        ast.Load,
        ast.Name,
        ast.Constant,
        ast.IfExp,
        ast.Compare,
        ast.Gt,
        ast.GtE,
        ast.Lt,
        ast.LtE,
        ast.Eq,
        ast.NotEq,
        ast.BoolOp,
        ast.And,
        ast.Or,
    }

    @classmethod
    def evaluate(cls, expression, variables, default=0.0):
        if not expression:
            return float(default)
        try:
            tree = ast.parse(expression, mode='eval')
            for node in ast.walk(tree):
                if type(node) not in cls.ALLOWED_NODES:
                    return float(default)
                if isinstance(node, ast.Name) and node.id not in variables:
                    return float(default)
            value = eval(compile(tree, '<formula>', 'eval'), {'__builtins__': {}}, variables)
            return float(value)
        except Exception:
            return float(default)


class StaffSalaryViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def _to_float_safe(self, value, default=0.0):
        try:
            if value in [None, '']:
                return float(default)
            return float(value)
        except (TypeError, ValueError):
            return float(default)

    def _get_or_create_monthly_input(self, staff, month_date):
        row = SalaryMonthlyInput.objects.filter(staff=staff, month=month_date).first()
        if row:
            return row

        try:
            row, _ = SalaryMonthlyInput.objects.get_or_create(staff=staff, month=month_date)
            return row
        except IntegrityError as exc:
            # Some environments still have a legacy NOT NULL lop_days column in DB.
            if 'lop_days' not in str(exc).lower():
                raise

            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO staff_salary_salarymonthlyinput
                    (staff_id, month, earn_values, deduction_values, include_in_salary, is_cash, od_new, others, lop_days, updated_at)
                    VALUES (%s, %s, %s::jsonb, %s::jsonb, %s, %s, %s, %s, %s, NOW())
                    ON CONFLICT (staff_id, month) DO NOTHING
                    """,
                    [staff.id, month_date, '{}', '{}', True, False, 0.0, 0.0, 0.0],
                )

            return SalaryMonthlyInput.objects.filter(staff=staff, month=month_date).first()

    def _attendance_absent_units(self, record):
        fn_status = (record.fn_status or '').strip().lower()
        an_status = (record.an_status or '').strip().lower()

        if fn_status or an_status:
            units = 0.0
            if fn_status == 'absent':
                units += 0.5
            if an_status == 'absent':
                units += 0.5
            return units

        return 1.0 if (record.status or '').strip().lower() == 'absent' else 0.0

    def _normalize_shift_value(self, value):
        token = str(value or '').strip().upper()
        if token == 'FULL DAY':
            token = 'FULL'
        return token

    def _single_day_units(self, from_noon, to_noon):
        if from_noon in ['FN', 'AN'] and to_noon in ['FN', 'AN']:
            return 0.5 if from_noon == to_noon else 1.0
        if from_noon in ['FN', 'AN'] and not to_noon:
            return 0.5
        if to_noon in ['FN', 'AN'] and not from_noon:
            return 0.5
        if from_noon == 'FULL' or to_noon == 'FULL':
            return 1.0
        return 1.0

    def _parse_form_date(self, raw_value):
        if isinstance(raw_value, datetime):
            return raw_value.date()
        if isinstance(raw_value, date):
            return raw_value
        if not isinstance(raw_value, str):
            return None

        token = raw_value.strip()
        if not token:
            return None

        try:
            return datetime.fromisoformat(token.replace('Z', '+00:00')).date()
        except Exception:
            pass

        try:
            return datetime.strptime(token[:10], '%Y-%m-%d').date()
        except Exception:
            return None

    def _build_monthly_holiday_map(self, month_start, month_end):
        from staff_attendance.models import Holiday

        holiday_map = {}
        holidays = Holiday.objects.filter(date__gte=month_start, date__lte=month_end).prefetch_related('departments')
        for holiday in holidays:
            dept_ids = set(holiday.departments.values_list('id', flat=True))
            holiday_map[holiday.date] = {
                'all_departments': len(dept_ids) == 0,
                'department_ids': dept_ids,
            }
        return holiday_map

    def _is_holiday_for_department(self, target_date, department_id, holiday_map):
        holiday_entry = holiday_map.get(target_date)
        if not holiday_entry:
            return False
        if holiday_entry['all_departments']:
            return True
        return department_id is not None and department_id in holiday_entry['department_ids']

    def _extract_requested_units_by_date(self, form_data, month_start, month_end, is_holiday):
        dates = {}

        start_date = None
        end_date = None
        for start_key in ['start_date', 'from_date', 'startDate', 'fromDate', 'from']:
            if start_key in form_data:
                start_date = form_data[start_key]
                break

        for end_key in ['end_date', 'to_date', 'endDate', 'toDate', 'to']:
            if end_key in form_data:
                end_date = form_data[end_key]
                break

        if not start_date and 'date' in form_data:
            start_date = form_data['date']
        if not end_date and 'date' in form_data:
            end_date = form_data['date']

        start = self._parse_form_date(start_date)
        end = self._parse_form_date(end_date)
        if not start or not end:
            return dates

        from_noon = self._normalize_shift_value(
            form_data.get('from_noon', form_data.get('from_shift', form_data.get('shift', '')))
        )
        to_noon = self._normalize_shift_value(
            form_data.get('to_noon', form_data.get('to_shift', form_data.get('shift', '')))
        )

        if start > end:
            start, end = end, start

        start = max(start, month_start)
        end = min(end, month_end)
        if start > end:
            return dates

        if start == end:
            if start.weekday() == 6 or is_holiday(start):
                return dates
            return {start: self._single_day_units(from_noon, to_noon)}

        current = start
        while current <= end:
            if current.weekday() == 6 or is_holiday(current):
                current += timedelta(days=1)
                continue

            units = 1.0
            if current == start and from_noon == 'AN':
                units = 0.5
            if current == end and to_noon == 'FN':
                units = 0.5
            dates[current] = units
            current += timedelta(days=1)

        return dates

    def _build_monthly_lop_map(self, staff_users, month_date):
        from staff_attendance.models import AttendanceRecord

        if not staff_users:
            return {}

        month_start = month_date
        month_end = date(month_date.year, month_date.month, calendar.monthrange(month_date.year, month_date.month)[1])
        holiday_map = self._build_monthly_holiday_map(month_start, month_end)

        staff_ids = [u.id for u in staff_users]
        dept_map = {}
        for user in staff_users:
            profile = getattr(user, 'staff_profile', None)
            dept = getattr(profile, 'department', None) if profile else None
            dept_map[user.id] = dept.id if dept else None

        absent_units_map = {staff_id: {} for staff_id in staff_ids}
        attendance_records = AttendanceRecord.objects.filter(
            user_id__in=staff_ids,
            date__gte=month_start,
            date__lte=month_end,
        )

        for record in attendance_records:
            if record.date.weekday() == 6:
                continue

            department_id = dept_map.get(record.user_id)
            if self._is_holiday_for_department(record.date, department_id, holiday_map):
                continue

            absent_units = self._attendance_absent_units(record)
            if absent_units > 0:
                absent_units_map[record.user_id][record.date] = absent_units

        approved_requests = StaffRequest.objects.filter(
            applicant_id__in=staff_ids,
            status='approved',
            template__leave_policy__action__in=['deduct', 'neutral'],
        )

        for request in approved_requests:
            department_id = dept_map.get(request.applicant_id)

            def is_holiday_for_request(target_date):
                return self._is_holiday_for_department(target_date, department_id, holiday_map)

            requested_units = self._extract_requested_units_by_date(
                request.form_data or {},
                month_start,
                month_end,
                is_holiday_for_request,
            )
            if not requested_units:
                continue

            remaining_for_staff = absent_units_map.get(request.applicant_id, {})
            for request_date, request_units in requested_units.items():
                absent_left = float(remaining_for_staff.get(request_date, 0.0))
                if absent_left <= 0:
                    continue
                covered_now = min(absent_left, float(request_units or 0.0))
                if covered_now > 0:
                    remaining_for_staff[request_date] = round(absent_left - covered_now, 2)

        return {
            staff_id: round(sum(absent_units_map.get(staff_id, {}).values()), 2)
            for staff_id in staff_ids
        }

    def _check_hr(self, request):
        if IsAdminOrHR().has_permission(request, self):
            return True
        user = request.user
        if not user or not user.is_authenticated:
            return False
        if user.is_superuser:
            return True
        # Fallback for deployments where roles are exposed via user.roles M2M
        if hasattr(user, 'roles'):
            try:
                return user.roles.filter(name__in=['HR', 'ADMIN']).exists()
            except Exception:
                return False
        return False

    def _month_to_date(self, month_str):
        parts = (month_str or '').split('-')
        if len(parts) != 2:
            raise ValueError('Invalid month format, use YYYY-MM')
        year = int(parts[0])
        month = int(parts[1])
        return date(year, month, 1)

    def _get_staff_queryset(self, department_id=None):
        User = get_user_model()
        qs = User.objects.filter(is_active=True, staff_profile__isnull=False).select_related('staff_profile', 'staff_profile__department')
        if department_id:
            try:
                dept_id = int(department_id)
                qs = qs.filter(staff_profile__department_id=dept_id)
            except (ValueError, TypeError):
                pass
        return qs.order_by('staff_profile__staff_id')

    def _get_pf_config(self):
        config, _ = SalaryPFConfig.objects.get_or_create(id=1)
        return config

    def _get_formula_config(self):
        formula_obj, _ = SalaryFormulaConfig.objects.get_or_create(id=1, defaults={'expressions': DEFAULT_FORMULAS})
        expr = dict(DEFAULT_FORMULAS)
        saved = dict(formula_obj.expressions or {})

        # Auto-correct legacy buggy formulas that force gross/net salary to zero when lop_days > 0.
        changed = False
        if str(saved.get('lop_amount', '')).strip() == '(basic_salary + allowance) / lop_days if lop_days > 0 else 0':
            saved['lop_amount'] = DEFAULT_FORMULAS['lop_amount']
            changed = True
        if str(saved.get('gross_salary', '')).strip() == '(basic_salary + allowance) - (lop_amount * lop_days)':
            saved['gross_salary'] = DEFAULT_FORMULAS['gross_salary']
            changed = True

        if changed:
            formula_obj.expressions = saved
            formula_obj.save(update_fields=['expressions'])

        expr.update(saved)
        return formula_obj, expr

    def _add_months(self, dt, months):
        y = dt.year + (dt.month - 1 + months) // 12
        m = (dt.month - 1 + months) % 12 + 1
        return date(y, m, 1)

    def _build_monthly_sheet_data(self, month_date, department_id=None):
        month_str = month_date.strftime('%Y-%m')
        days_in_month = calendar.monthrange(month_date.year, month_date.month)[1]
        pf_config = self._get_pf_config()
        _, formulas = self._get_formula_config()

        earn_types = list(SalaryEarnType.objects.filter(is_active=True).order_by('sort_order', 'id'))
        deduction_types = list(SalaryDeductionType.objects.filter(is_active=True).order_by('sort_order', 'id'))
        emi_type_ids = {d.id for d in deduction_types if d.mode == 'emi'}

        staff_users = list(self._get_staff_queryset(department_id=department_id))
        staff_ids = [u.id for u in staff_users]

        declaration_map = {
            d.staff_id: d
            for d in StaffSalaryDeclaration.objects.filter(staff_id__in=staff_ids).select_related('bank')
        }
        monthly_map = {
            m.staff_id: m
            for m in SalaryMonthlyInput.objects.filter(staff_id__in=staff_ids, month=month_date)
        }

        lop_map = self._build_monthly_lop_map(staff_users, month_date)

        historical_inputs = SalaryMonthlyInput.objects.filter(
            staff_id__in=staff_ids,
            month__lte=month_date,
        ).values('staff_id', 'month', 'include_in_salary')
        include_flag_map = {
            (row['staff_id'], row['month']): bool(row['include_in_salary'])
            for row in historical_inputs
        }

        emi_plans = SalaryEMIPlan.objects.filter(
            staff_id__in=staff_ids,
            deduction_type_id__in=list(emi_type_ids),
            is_active=True,
        ).select_related('deduction_type')

        emi_amount_map = {}
        for plan in emi_plans:
            if month_date < plan.start_month:
                continue

            payable_before = 0
            month_cursor = plan.start_month
            while month_cursor < month_date:
                include_before = include_flag_map.get((plan.staff_id, month_cursor), True)
                if include_before:
                    payable_before += 1
                month_cursor = self._add_months(month_cursor, 1)

            include_current = include_flag_map.get((plan.staff_id, month_date), True)
            if include_current and payable_before < plan.months:
                key = (plan.staff_id, plan.deduction_type_id)
                emi_amount_map[key] = round(float(plan.total_amount or 0) / max(1, plan.months), 2)

        published_obj = SalaryMonthPublish.objects.filter(month=month_date, is_active=True).first()

        rows = []
        for idx, staff in enumerate(staff_users, start=1):
            profile = getattr(staff, 'staff_profile', None)
            dept = getattr(profile, 'department', None) if profile else None
            declaration = declaration_map.get(staff.id)
            monthly = monthly_map.get(staff.id)

            include_in_salary = bool(monthly.include_in_salary) if monthly else True
            is_cash = bool(monthly.is_cash) if monthly else False
            basic_salary = self._to_float_safe(declaration.basic_salary if declaration else 0, 0)
            allowance = self._to_float_safe(declaration.allowance if declaration else 0, 0)
            pf_enabled = bool(declaration.pf_enabled) if declaration else True
            lop_days = max(0.0, self._to_float_safe(lop_map.get(staff.id, 0.0), 0.0))

            earn_values = {}
            total_earn = 0.0
            monthly_earn_values = (monthly.earn_values if monthly else {}) or {}
            if not isinstance(monthly_earn_values, dict):
                monthly_earn_values = {}
            for e in earn_types:
                amount = self._to_float_safe(monthly_earn_values.get(str(e.id), 0), 0)
                earn_values[str(e.id)] = amount
                total_earn += amount

            deduction_values = {}
            total_deduction = 0.0
            monthly_deduction_values = (monthly.deduction_values if monthly else {}) or {}
            if not isinstance(monthly_deduction_values, dict):
                monthly_deduction_values = {}
            for d in deduction_types:
                if d.mode == 'emi':
                    amount = self._to_float_safe(emi_amount_map.get((staff.id, d.id), 0), 0)
                else:
                    amount = self._to_float_safe(monthly_deduction_values.get(str(d.id), 0), 0)
                deduction_values[str(d.id)] = amount
                total_deduction += amount

            od_new = self._to_float_safe(monthly.od_new if monthly else 0, 0)
            others = self._to_float_safe(monthly.others if monthly else 0, 0)

            context = {
                'days_in_month': float(days_in_month),
                'lop_days': lop_days,
                'basic_salary': basic_salary,
                'allowance': allowance,
                'total_earn': total_earn,
                'total_deduction': total_deduction,
                'od_new': od_new,
                'others': others,
            }

            working_days = SafeFormulaEvaluator.evaluate(formulas.get('working_days'), context, default=context['days_in_month'] - lop_days)
            context['working_days'] = working_days
            lop_amount = SafeFormulaEvaluator.evaluate(
                formulas.get('lop_amount'),
                context,
                default=(((basic_salary + allowance) / days_in_month) * lop_days if days_in_month > 0 else 0),
            )
            context['lop_amount'] = lop_amount
            gross_salary = SafeFormulaEvaluator.evaluate(
                formulas.get('gross_salary'),
                context,
                default=(basic_salary + allowance) - lop_amount,
            )
            context['gross_salary'] = gross_salary
            total_salary = SafeFormulaEvaluator.evaluate(formulas.get('total_salary'), context, default=gross_salary + total_earn)
            context['total_salary'] = total_salary

            pf_amount = 0.0
            dept_id = dept.id if dept else None
            if pf_enabled and dept_id in (pf_config.type1_department_ids or []):
                if total_salary >= float(pf_config.threshold_amount):
                    pf_amount = float(pf_config.fixed_pf_amount)
                else:
                    pf_amount = total_salary * float(pf_config.percentage_rate) / 100.0
            elif pf_enabled and dept_id in (pf_config.type2_department_ids or []):
                try:
                    declaration = StaffSalaryDeclaration.objects.get(staff_id=staff.id)
                    pf_amount = float(declaration.type2_pf_value)
                except StaffSalaryDeclaration.DoesNotExist:
                    pf_amount = 0.0

            context['pf_amount'] = pf_amount
            net_salary = SafeFormulaEvaluator.evaluate(
                formulas.get('net_salary'),
                context,
                default=(total_salary + pf_amount - od_new - total_deduction - others),
            )

            computed_gross_salary = gross_salary
            computed_salary = basic_salary + allowance
            computed_total_salary = total_salary
            computed_pf_amount = pf_amount
            computed_total_deduction = total_deduction + od_new + others
            computed_net_salary = net_salary

            if not include_in_salary:
                working_days = 0.0
                gross_salary = 0.0
                lop_amount = 0.0
                total_salary = 0.0
                pf_amount = 0.0
                total_deduction = 0.0
                net_salary = 0.0
                deduction_values = {str(d.id): 0.0 for d in deduction_types}

            rows.append({
                's_no': idx,
                'staff_user_id': staff.id,
                'staff_id': getattr(profile, 'staff_id', None) or staff.username,
                'staff_name': staff.get_full_name() or staff.username,
                'department': {'id': dept.id if dept else None, 'name': dept.name if dept else 'N/A'},
                'bank_name': declaration.bank.name if declaration and declaration.bank else '',
                'account_no': declaration.account_no if declaration else '',
                'ifsc_code': declaration.ifsc_code if declaration else '',
                'include_in_salary': include_in_salary,
                'is_cash': is_cash,
                'computed_salary': round(computed_salary, 2),
                'computed_gross_salary': round(computed_gross_salary, 2),
                'computed_total_salary': round(computed_total_salary, 2),
                'computed_pf_amount': round(computed_pf_amount, 2),
                'computed_total_deduction': round(computed_total_deduction, 2),
                'computed_net_salary': round(computed_net_salary, 2),
                'basic_salary': round(basic_salary if include_in_salary else 0.0, 2),
                'allowance': round(allowance if include_in_salary else 0.0, 2),
                'days': round(working_days, 2),
                'gross_salary': round(gross_salary, 2),
                'lop_days': round(lop_days, 2),
                'lop_amount': round(lop_amount, 2),
                'earn_values': earn_values,
                'total_salary': round(total_salary, 2),
                'pf_amount': round(pf_amount, 2),
                'od_new': round(od_new, 2),
                'deduction_values': deduction_values,
                'others': round(others, 2),
                'net_salary': round(net_salary, 2),
            })

        return {
            'month': month_str,
            'days_in_month': days_in_month,
            'earn_types': [{'id': e.id, 'name': e.name} for e in earn_types],
            'deduction_types': [{'id': d.id, 'name': d.name, 'mode': d.mode} for d in deduction_types],
            'results': rows,
            'formulas': formulas,
            'published': bool(published_obj),
            'published_at': published_obj.published_at.isoformat() if published_obj else None,
        }

    def _staff_type_label(self, department_id, pf_config):
        type2_ids = set(pf_config.type2_department_ids or [])
        if department_id in type2_ids:
            return 'Non-Teaching Staff'
        return 'Teaching Staff'

    def _build_payroll_report(self, month_date):
        payload = self._build_monthly_sheet_data(month_date, department_id=None)
        pf_config = self._get_pf_config()
        earn_types = payload.get('earn_types', [])
        deduction_types = payload.get('deduction_types', [])

        def blank_group():
            return {
                'salary': 0.0,
                'lop': 0.0,
                'earn': {str(e['id']): 0.0 for e in earn_types},
                'gross_salary': 0.0,
                'pf_amount': 0.0,
                'deduction': {str(d['id']): 0.0 for d in deduction_types},
                'total_deduction': 0.0,
                'net_salary': 0.0,
            }

        section1_groups = {
            'Teaching Staff': blank_group(),
            'Non-Teaching Staff': blank_group(),
        }

        bank_names = list(
            SalaryBankDeclaration.objects.filter(is_active=True)
            .order_by('sort_order', 'name', 'id')
            .values_list('name', flat=True)
        )
        if 'Other Bank' not in bank_names:
            bank_names.append('Other Bank')

        section2_groups = {
            'Teaching Staff': {
                'banks': {name: {'total_request': 0, 'amount': 0.0} for name in bank_names},
                'cash': 0.0,
                'total_salary': 0.0,
            },
            'Non-Teaching Staff': {
                'banks': {name: {'total_request': 0, 'amount': 0.0} for name in bank_names},
                'cash': 0.0,
                'total_salary': 0.0,
            },
        }

        for row in payload.get('results', []):
            dept_id = (row.get('department') or {}).get('id')
            staff_type = self._staff_type_label(dept_id, pf_config)

            salary = float(row.get('computed_salary', 0) or 0)
            lop = float(row.get('lop_amount', 0) or 0)
            gross = float(row.get('computed_gross_salary', row.get('gross_salary', 0)) or 0)
            pf_amt = float(row.get('computed_pf_amount', row.get('pf_amount', 0)) or 0)
            others = float(row.get('others', 0) or 0)
            od_new = float(row.get('od_new', 0) or 0)
            net = float(row.get('computed_net_salary', row.get('net_salary', 0)) or 0)

            deduction_sum = 0.0
            for d in deduction_types:
                d_key = str(d['id'])
                d_value = float((row.get('deduction_values') or {}).get(d_key, 0) or 0)
                section1_groups[staff_type]['deduction'][d_key] += d_value
                deduction_sum += d_value

            for e in earn_types:
                e_key = str(e['id'])
                e_value = float((row.get('earn_values') or {}).get(e_key, 0) or 0)
                section1_groups[staff_type]['earn'][e_key] += e_value

            section1_groups[staff_type]['salary'] += salary
            section1_groups[staff_type]['lop'] += lop
            section1_groups[staff_type]['gross_salary'] += gross
            section1_groups[staff_type]['pf_amount'] += pf_amt
            section1_groups[staff_type]['total_deduction'] += float(row.get('computed_total_deduction', deduction_sum + others + od_new) or 0)
            section1_groups[staff_type]['net_salary'] += net

            include_in_salary = bool(row.get('include_in_salary', True))
            is_cash = bool(row.get('is_cash', False))
            bank_name = str(row.get('bank_name') or '').strip() or 'Other Bank'
            if bank_name not in section2_groups[staff_type]['banks']:
                section2_groups[staff_type]['banks'][bank_name] = {'total_request': 0, 'amount': 0.0}

            if is_cash:
                section2_groups[staff_type]['cash'] += net
            elif include_in_salary:
                section2_groups[staff_type]['banks'][bank_name]['total_request'] += 1
                section2_groups[staff_type]['banks'][bank_name]['amount'] += net

            section2_groups[staff_type]['total_salary'] += net

        section1_rows = []
        section1_grand = blank_group()
        for idx, label in enumerate(['Teaching Staff', 'Non-Teaching Staff'], start=1):
            g = section1_groups[label]
            section1_rows.append({
                's_no': idx,
                'staff_type': label,
                'salary': round(g['salary'], 2),
                'lop': round(g['lop'], 2),
                'earn': {k: round(v, 2) for k, v in g['earn'].items()},
                'gross_salary': round(g['gross_salary'], 2),
                'pf_amount': round(g['pf_amount'], 2),
                'deduction': {k: round(v, 2) for k, v in g['deduction'].items()},
                'total_deduction': round(g['total_deduction'], 2),
                'net_salary': round(g['net_salary'], 2),
            })

            section1_grand['salary'] += g['salary']
            section1_grand['lop'] += g['lop']
            section1_grand['gross_salary'] += g['gross_salary']
            section1_grand['pf_amount'] += g['pf_amount']
            section1_grand['total_deduction'] += g['total_deduction']
            section1_grand['net_salary'] += g['net_salary']
            for e_key, e_val in g['earn'].items():
                section1_grand['earn'][e_key] += e_val
            for d_key, d_val in g['deduction'].items():
                section1_grand['deduction'][d_key] += d_val

        section2_rows = []
        section2_grand = {
            'banks': defaultdict(lambda: {'total_request': 0, 'amount': 0.0}),
            'cash': 0.0,
            'total_salary': 0.0,
        }
        for idx, label in enumerate(['Teaching Staff', 'Non-Teaching Staff'], start=1):
            g = section2_groups[label]
            section2_rows.append({
                's_no': idx,
                'staff_type': label,
                'banks': {
                    name: {
                        'total_request': int(g['banks'].get(name, {}).get('total_request', 0)),
                        'amount': round(float(g['banks'].get(name, {}).get('amount', 0) or 0), 2),
                    }
                    for name in sorted(g['banks'].keys())
                },
                'cash': round(g['cash'], 2),
                'total_salary': round(g['total_salary'], 2),
            })

            for bank_name, bank_val in g['banks'].items():
                section2_grand['banks'][bank_name]['total_request'] += int(bank_val.get('total_request', 0) or 0)
                section2_grand['banks'][bank_name]['amount'] += float(bank_val.get('amount', 0) or 0)
            section2_grand['cash'] += float(g['cash'] or 0)
            section2_grand['total_salary'] += float(g['total_salary'] or 0)

        return {
            'month': payload.get('month'),
            'earn_types': earn_types,
            'deduction_types': deduction_types,
            'section1': {
                'rows': section1_rows,
                'grand_total': {
                    'staff_type': 'Grand Total',
                    'salary': round(section1_grand['salary'], 2),
                    'lop': round(section1_grand['lop'], 2),
                    'earn': {k: round(v, 2) for k, v in section1_grand['earn'].items()},
                    'gross_salary': round(section1_grand['gross_salary'], 2),
                    'pf_amount': round(section1_grand['pf_amount'], 2),
                    'deduction': {k: round(v, 2) for k, v in section1_grand['deduction'].items()},
                    'total_deduction': round(section1_grand['total_deduction'], 2),
                    'net_salary': round(section1_grand['net_salary'], 2),
                },
            },
            'section2': {
                'bank_columns': sorted(section2_grand['banks'].keys()),
                'rows': section2_rows,
                'grand_total': {
                    'staff_type': 'Grand Total',
                    'banks': {
                        name: {
                            'total_request': int(vals['total_request']),
                            'amount': round(float(vals['amount'] or 0), 2),
                        }
                        for name, vals in sorted(section2_grand['banks'].items(), key=lambda x: x[0])
                    },
                    'cash': round(section2_grand['cash'], 2),
                    'total_salary': round(section2_grand['total_salary'], 2),
                },
            },
        }

    def _build_bank_staff_report(self, month_date, bank_name=None):
        payload = self._build_monthly_sheet_data(month_date, department_id=None)
        rows = payload.get('results', [])

        bank_options = sorted({(str(r.get('bank_name') or '').strip() or 'Other Bank') for r in rows})
        if 'Cash' not in bank_options:
            bank_options.append('Cash')

        filtered = []
        for row in rows:
            row_bank = 'Cash' if bool(row.get('is_cash', False)) else (str(row.get('bank_name') or '').strip() or 'Other Bank')
            if bank_name and bank_name.strip() and row_bank.lower() != bank_name.strip().lower():
                continue
            filtered.append({
                'staff_user_id': row.get('staff_user_id'),
                'staff_id': row.get('staff_id', ''),
                'staff_name': row.get('staff_name', ''),
                'department': (row.get('department') or {}).get('name', 'N/A'),
                'bank': row_bank,
                'account_no': row.get('account_no', ''),
                'ifsc_code': row.get('ifsc_code', ''),
                'gross_salary': round(float(row.get('computed_gross_salary', row.get('gross_salary', 0)) or 0), 2),
            })

        for idx, item in enumerate(filtered, start=1):
            item['s_no'] = idx

        return {
            'month': payload.get('month'),
            'bank_filter': bank_name or '',
            'bank_options': bank_options,
            'rows': filtered,
            'count': len(filtered),
        }

    def _apply_excel_header_style(self, ws, row_number=1):
        fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        font = Font(bold=True, color='FFFFFF')
        for cell in ws[row_number]:
            cell.fill = fill
            cell.font = font

    def _excel_response(self, wb, filename):
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', cell.value)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _safe_excel_number(self, value, default=0.0):
        num = self._to_float_safe(value, default)
        if not math.isfinite(num):
            return float(default)
        return num

    def _export_monthly_sheet_excel(self, payload):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Monthly Sheet'

        ws.append(['Staff Salary Monthly Sheet'])
        ws.append(['Month', payload.get('month')])
        ws.append([])

        header = [
            'Staff ID', 'Staff Name', 'Department', 'Include In Salary', 'Cash', 'Basic Salary', 'Allowance',
            'Days', 'Gross Salary', 'LOP Amount'
        ]
        for e in payload.get('earn_types', []):
            header.append(e['name'])
        header.extend(['Total Salary', 'PF Amount', 'OD New'])
        for d in payload.get('deduction_types', []):
            header.append(d['name'])
        header.extend(['Others', 'Net Salary'])
        ws.append(header)
        self._apply_excel_header_style(ws, ws.max_row)

        for r in payload.get('results', []):
            row = [
                r.get('staff_id', ''),
                r.get('staff_name', ''),
                (r.get('department') or {}).get('name', ''),
                'Yes' if r.get('include_in_salary', True) else 'No',
                'Yes' if r.get('is_cash', False) else 'No',
                self._safe_excel_number(r.get('basic_salary', 0)),
                self._safe_excel_number(r.get('allowance', 0)),
                self._safe_excel_number(r.get('days', 0)),
                self._safe_excel_number(r.get('gross_salary', 0)),
                self._safe_excel_number(r.get('lop_amount', 0)),
            ]
            earn_values = r.get('earn_values') or {}
            for e in payload.get('earn_types', []):
                row.append(self._safe_excel_number(earn_values.get(str(e['id']), 0)))
            row.extend([
                self._safe_excel_number(r.get('total_salary', 0)),
                self._safe_excel_number(r.get('pf_amount', 0)),
                self._safe_excel_number(r.get('od_new', 0)),
            ])
            deduction_values = r.get('deduction_values') or {}
            for d in payload.get('deduction_types', []):
                row.append(self._safe_excel_number(deduction_values.get(str(d['id']), 0)))
            row.extend([
                self._safe_excel_number(r.get('others', 0)),
                self._safe_excel_number(r.get('net_salary', 0)),
            ])
            ws.append(row)

        return self._excel_response(wb, f"salary_monthly_sheet_{payload.get('month')}.xlsx")

    def _export_payroll_report_excel(self, report):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Payroll Section 1'

        header1 = ['S.No', 'Staff Type', 'Salary', 'LOP']
        for e in report.get('earn_types', []):
            header1.append(e['name'])
        header1.extend(['Gross Salary', 'PF Amount'])
        for d in report.get('deduction_types', []):
            header1.append(d['name'])
        header1.extend(['Total Deduction', 'Net Salary'])
        ws1.append(header1)
        self._apply_excel_header_style(ws1, 1)

        for row in report.get('section1', {}).get('rows', []):
            values = [row.get('s_no'), row.get('staff_type'), row.get('salary', 0), row.get('lop', 0)]
            for e in report.get('earn_types', []):
                values.append((row.get('earn') or {}).get(str(e['id']), 0))
            values.extend([row.get('gross_salary', 0), row.get('pf_amount', 0)])
            for d in report.get('deduction_types', []):
                values.append((row.get('deduction') or {}).get(str(d['id']), 0))
            values.extend([row.get('total_deduction', 0), row.get('net_salary', 0)])
            ws1.append(values)

        grand = report.get('section1', {}).get('grand_total') or {}
        if grand:
            values = ['', 'Grand Total', grand.get('salary', 0), grand.get('lop', 0)]
            for e in report.get('earn_types', []):
                values.append((grand.get('earn') or {}).get(str(e['id']), 0))
            values.extend([grand.get('gross_salary', 0), grand.get('pf_amount', 0)])
            for d in report.get('deduction_types', []):
                values.append((grand.get('deduction') or {}).get(str(d['id']), 0))
            values.extend([grand.get('total_deduction', 0), grand.get('net_salary', 0)])
            ws1.append(values)

        ws2 = wb.create_sheet(title='Payroll Section 2')
        bank_columns = report.get('section2', {}).get('bank_columns', [])
        header2 = ['S.No', 'Staff Type']
        for bank_name in bank_columns:
            header2.extend([f'{bank_name} Total Request', f'{bank_name} Amount'])
        header2.extend(['Cash', 'Total Salary'])
        ws2.append(header2)
        self._apply_excel_header_style(ws2, 1)

        for row in report.get('section2', {}).get('rows', []):
            values = [row.get('s_no'), row.get('staff_type')]
            for bank_name in bank_columns:
                values.extend([
                    (row.get('banks') or {}).get(bank_name, {}).get('total_request', 0),
                    (row.get('banks') or {}).get(bank_name, {}).get('amount', 0),
                ])
            values.extend([row.get('cash', 0), row.get('total_salary', 0)])
            ws2.append(values)

        grand2 = report.get('section2', {}).get('grand_total') or {}
        if grand2:
            values = ['', 'Grand Total']
            for bank_name in bank_columns:
                values.extend([
                    (grand2.get('banks') or {}).get(bank_name, {}).get('total_request', 0),
                    (grand2.get('banks') or {}).get(bank_name, {}).get('amount', 0),
                ])
            values.extend([grand2.get('cash', 0), grand2.get('total_salary', 0)])
            ws2.append(values)

        return self._excel_response(wb, f"salary_payroll_report_{report.get('month')}.xlsx")

    def _export_bank_staff_report_excel(self, report):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Bank Staff Report'

        ws.append(['S.No', 'Staff ID', 'Staff Name', 'Department', 'Bank', 'A/C No', 'IFSC Code', 'Gross Salary'])
        self._apply_excel_header_style(ws, 1)

        for row in report.get('rows', []):
            ws.append([
                row.get('s_no'),
                row.get('staff_id', ''),
                row.get('staff_name', ''),
                row.get('department', ''),
                row.get('bank', ''),
                row.get('account_no', ''),
                row.get('ifsc_code', ''),
                self._safe_excel_number(row.get('gross_salary', 0)),
            ])

        return self._excel_response(wb, f"salary_bank_staff_report_{report.get('month')}.xlsx")

    @action(detail=False, methods=['get', 'post'])
    def bank_declarations(self, request):
        if request.method.lower() == 'post' and not self._check_hr(request):
            return Response({'error': 'Only HR/Admin can modify bank declarations'}, status=status.HTTP_403_FORBIDDEN)

        if request.method.lower() == 'post':
            items = request.data.get('items', [])
            for idx, item in enumerate(items, start=1):
                name = str(item.get('name') or '').strip()
                if not name:
                    continue

                obj = None
                if item.get('id'):
                    obj = SalaryBankDeclaration.objects.filter(id=item['id']).first()
                if not obj:
                    obj = SalaryBankDeclaration.objects.filter(name__iexact=name).first()
                if not obj:
                    obj = SalaryBankDeclaration()

                obj.name = name
                obj.is_active = bool(item.get('is_active', True))
                obj.sort_order = int(item.get('sort_order') or idx)
                obj.save()

            return Response({'message': 'Bank declarations saved'})

        if not SalaryBankDeclaration.objects.exists():
            default_names = ['SBI', 'ICICI BANK', 'HDFC BANK', 'CANARA BANK', 'INDIAN BANK']
            for idx, name in enumerate(default_names, start=1):
                SalaryBankDeclaration.objects.create(name=name, is_active=True, sort_order=idx)

        rows = SalaryBankDeclaration.objects.all().values('id', 'name', 'is_active', 'sort_order')
        return Response({'count': len(rows), 'results': list(rows)})

    @action(detail=False, methods=['get', 'post'])
    def declarations(self, request):
        # Allow any authenticated user to view declarations, but only HR/Admin can modify
        if request.method.lower() == 'post' and not self._check_hr(request):
            return Response({'error': 'Only HR/Admin can modify declarations'}, status=status.HTTP_403_FORBIDDEN)

        if request.method.lower() == 'post':
            updates = request.data.get('items', [])
            for item in updates:
                staff_user_id = item.get('staff_user_id')
                if not staff_user_id:
                    continue
                User = get_user_model()
                try:
                    staff_user = User.objects.get(id=staff_user_id)
                except User.DoesNotExist:
                    continue
                obj, _ = StaffSalaryDeclaration.objects.get_or_create(staff=staff_user)
                obj.basic_salary = float(item.get('basic_salary') or 0)
                obj.allowance = float(item.get('allowance') or 0)
                obj.pf_enabled = bool(item.get('pf_enabled', True))
                obj.type2_pf_value = float(item.get('type2_pf_value') or 0)
                bank_id = item.get('bank_id')
                if bank_id in [None, '', 0, '0']:
                    obj.bank = None
                else:
                    obj.bank = SalaryBankDeclaration.objects.filter(id=bank_id).first()
                obj.account_no = str(item.get('account_no') or '').strip()
                obj.ifsc_code = str(item.get('ifsc_code') or '').strip().upper()
                obj.save()
            return Response({'message': 'Declarations saved'})

        department_id = request.query_params.get('department_id')
        staff_users = list(self._get_staff_queryset(department_id=department_id))
        declaration_map = {d.staff_id: d for d in StaffSalaryDeclaration.objects.filter(staff_id__in=[u.id for u in staff_users])}
        bank_options = list(
            SalaryBankDeclaration.objects.filter(is_active=True)
            .order_by('sort_order', 'name', 'id')
            .values('id', 'name')
        )

        rows = []
        for idx, staff in enumerate(staff_users, start=1):
            dec = declaration_map.get(staff.id)
            profile = getattr(staff, 'staff_profile', None)
            dept = getattr(profile, 'department', None) if profile else None
            rows.append({
                's_no': idx,
                'staff_user_id': staff.id,
                'staff_id': getattr(profile, 'staff_id', None) or staff.username,
                'name': staff.get_full_name() or staff.username,
                'department': {'id': dept.id if dept else None, 'name': dept.name if dept else 'N/A'},
                'basic_salary': float(dec.basic_salary if dec else 0),
                'allowance': float(dec.allowance if dec else 0),
                'pf_enabled': bool(dec.pf_enabled) if dec else True,
                'type2_pf_value': float(dec.type2_pf_value if dec else 0),
                'bank_id': dec.bank_id if dec else None,
                'bank_name': dec.bank.name if dec and dec.bank else '',
                'account_no': dec.account_no if dec else '',
                'ifsc_code': dec.ifsc_code if dec else '',
                'is_new': not bool(dec),
            })

        return Response({'count': len(rows), 'results': rows, 'bank_options': bank_options})

    @action(detail=False, methods=['get', 'post'])
    def pf_config(self, request):
        # Allow any authenticated user to view config, but only HR/Admin can modify
        if request.method.lower() == 'post' and not self._check_hr(request):
            return Response({'error': 'Only HR/Admin can modify PF config'}, status=status.HTTP_403_FORBIDDEN)

        config = self._get_pf_config()
        if request.method.lower() == 'post':
            payload = request.data or {}
            config.threshold_amount = float(payload.get('threshold_amount', config.threshold_amount))
            config.fixed_pf_amount = float(payload.get('fixed_pf_amount', config.fixed_pf_amount))
            config.percentage_rate = float(payload.get('percentage_rate', config.percentage_rate))
            config.type1_department_ids = [int(x) for x in payload.get('type1_department_ids', config.type1_department_ids or [])]
            config.type2_department_ids = [int(x) for x in payload.get('type2_department_ids', config.type2_department_ids or [])]
            config.save()

        departments = Department.objects.all().order_by('name').values('id', 'name')
        return Response({
            'threshold_amount': float(config.threshold_amount),
            'fixed_pf_amount': float(config.fixed_pf_amount),
            'percentage_rate': float(config.percentage_rate),
            'type1_department_ids': config.type1_department_ids or [],
            'type2_department_ids': config.type2_department_ids or [],
            'departments': list(departments),
        })

    @action(detail=False, methods=['get', 'post'])
    def deduction_types(self, request):
        if request.method.lower() == 'post' and not self._check_hr(request):
            return Response({'error': 'Only HR/Admin can modify deduction types'}, status=status.HTTP_403_FORBIDDEN)

        if request.method.lower() == 'post':
            items = request.data.get('items', [])
            for idx, item in enumerate(items, start=1):
                obj = None
                if item.get('id'):
                    obj = SalaryDeductionType.objects.filter(id=item['id']).first()
                if not obj:
                    obj = SalaryDeductionType()
                obj.name = str(item.get('name') or '').strip() or f'Deduction {idx}'
                obj.mode = item.get('mode') if item.get('mode') in ['emi', 'monthly'] else 'monthly'
                obj.is_active = bool(item.get('is_active', True))
                obj.sort_order = int(item.get('sort_order') or idx)
                obj.save()
            return Response({'message': 'Deduction types saved'})

        if not SalaryDeductionType.objects.exists():
            defaults = [
                ('Type 1', 'emi'),
                ('Type 2', 'emi'),
                ('Type 3', 'emi'),
                ('Type 4', 'monthly'),
                ('Type 5', 'monthly'),
                ('Type 6', 'monthly'),
            ]
            for idx, (name, mode) in enumerate(defaults, start=1):
                SalaryDeductionType.objects.create(name=name, mode=mode, is_active=True, sort_order=idx)

        rows = SalaryDeductionType.objects.all().values('id', 'name', 'mode', 'is_active', 'sort_order')
        return Response({'results': list(rows)})

    @action(detail=False, methods=['get', 'post'])
    def earn_types(self, request):
        if request.method.lower() == 'post' and not self._check_hr(request):
            return Response({'error': 'Only HR/Admin can modify earn types'}, status=status.HTTP_403_FORBIDDEN)

        if request.method.lower() == 'post':
            items = request.data.get('items', [])
            for idx, item in enumerate(items, start=1):
                obj = None
                if item.get('id'):
                    obj = SalaryEarnType.objects.filter(id=item['id']).first()
                if not obj:
                    obj = SalaryEarnType()
                obj.name = str(item.get('name') or '').strip() or f'Earn {idx}'
                obj.is_active = bool(item.get('is_active', True))
                obj.sort_order = int(item.get('sort_order') or idx)
                obj.save()
            return Response({'message': 'Earn types saved'})

        if not SalaryEarnType.objects.exists():
            defaults = ['Earn Type 1', 'Earn Type 2', 'Earn Type 3']
            for idx, name in enumerate(defaults, start=1):
                SalaryEarnType.objects.create(name=name, is_active=True, sort_order=idx)

        rows = SalaryEarnType.objects.all().values('id', 'name', 'is_active', 'sort_order')
        return Response({'results': list(rows)})

    @action(detail=False, methods=['get', 'post'])
    def emi_plans(self, request):
        if request.method.lower() == 'post' and not self._check_hr(request):
            return Response({'error': 'Only HR/Admin can modify EMI plans'}, status=status.HTTP_403_FORBIDDEN)

        if request.method.lower() == 'post':
            items = request.data.get('items', [])
            User = get_user_model()
            for item in items:
                staff_user_id = item.get('staff_user_id')
                deduction_type_id = item.get('deduction_type_id')
                start_month = item.get('start_month')
                if not staff_user_id or not deduction_type_id or not start_month:
                    continue
                try:
                    staff = User.objects.get(id=staff_user_id)
                    dtype = SalaryDeductionType.objects.get(id=deduction_type_id)
                    month_date = self._month_to_date(start_month)
                except Exception:
                    continue
                plan, _ = SalaryEMIPlan.objects.get_or_create(
                    staff=staff,
                    deduction_type=dtype,
                    start_month=month_date,
                    defaults={'months': 1, 'total_amount': 0.0},
                )
                plan.total_amount = float(item.get('total_amount') or 0)
                plan.months = max(1, int(item.get('months') or 1))
                plan.is_active = bool(item.get('is_active', True))
                plan.save()
            return Response({'message': 'EMI plans saved'})

        staff_user_id = request.query_params.get('staff_user_id')
        qs = SalaryEMIPlan.objects.select_related('staff', 'staff__staff_profile', 'deduction_type')
        if staff_user_id:
            qs = qs.filter(staff_id=staff_user_id)

        results = []
        for p in qs.order_by('-start_month'):
            profile = getattr(p.staff, 'staff_profile', None)
            results.append({
                'id': p.id,
                'staff_user_id': p.staff_id,
                'staff_id': getattr(profile, 'staff_id', None) or p.staff.username,
                'staff_name': p.staff.get_full_name() or p.staff.username,
                'deduction_type_id': p.deduction_type_id,
                'deduction_type_name': p.deduction_type.name,
                'total_amount': float(p.total_amount),
                'months': p.months,
                'start_month': p.start_month.strftime('%Y-%m'),
                'is_active': p.is_active,
            })
        return Response({'results': results})

    @action(detail=False, methods=['get', 'post'])
    def formulas(self, request):
        if request.method.lower() == 'post' and not self._check_hr(request):
            return Response({'error': 'Only HR/Admin can modify formulas'}, status=status.HTTP_403_FORBIDDEN)

        obj, current = self._get_formula_config()
        if request.method.lower() == 'post':
            incoming = request.data.get('expressions', {}) or {}
            merged = dict(DEFAULT_FORMULAS)
            merged.update(current)
            for key, value in incoming.items():
                merged[key] = str(value)
            obj.expressions = merged
            obj.save()
            current = merged

        return Response({'expressions': current, 'defaults': DEFAULT_FORMULAS})

    @action(detail=False, methods=['get', 'post'])
    def monthly_sheet(self, request):
        if request.method.lower() == 'post' and not self._check_hr(request):
            return Response({'error': 'Only HR/Admin can modify monthly sheet'}, status=status.HTTP_403_FORBIDDEN)

        month_str = request.query_params.get('month') or request.data.get('month')
        if not month_str:
            return Response({'error': 'month is required (YYYY-MM)'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            month_date = self._month_to_date(month_str)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        if request.method.lower() == 'post':
            updates = request.data.get('items', [])
            User = get_user_model()
            for item in updates:
                staff_user_id = item.get('staff_user_id')
                if not staff_user_id:
                    continue
                try:
                    staff = User.objects.get(id=staff_user_id)
                except User.DoesNotExist:
                    continue
                row = self._get_or_create_monthly_input(staff, month_date)
                if not row:
                    continue
                incoming_earn_values = item.get('earn_values')
                incoming_deduction_values = item.get('deduction_values')
                if isinstance(incoming_earn_values, dict):
                    row.earn_values = incoming_earn_values
                elif row.earn_values is None or not isinstance(row.earn_values, dict):
                    row.earn_values = {}

                if isinstance(incoming_deduction_values, dict):
                    row.deduction_values = incoming_deduction_values
                elif row.deduction_values is None or not isinstance(row.deduction_values, dict):
                    row.deduction_values = {}
                if 'include_in_salary' in item:
                    row.include_in_salary = bool(item.get('include_in_salary'))
                if 'is_cash' in item:
                    row.is_cash = bool(item.get('is_cash'))
                row.od_new = self._to_float_safe(item.get('od_new', row.od_new), row.od_new or 0)
                row.others = self._to_float_safe(item.get('others', row.others), row.others or 0)
                row.save()
        payload = self._build_monthly_sheet_data(
            month_date,
            department_id=request.query_params.get('department_id') or request.data.get('department_id'),
        )
        return Response(payload)

    @action(detail=False, methods=['get'])
    def monthly_sheet_download(self, request):
        month_str = request.query_params.get('month')
        if not month_str:
            return Response({'error': 'month is required (YYYY-MM)'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            month_date = self._month_to_date(month_str)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        payload = self._build_monthly_sheet_data(month_date, department_id=request.query_params.get('department_id'))

        return self._export_monthly_sheet_excel(payload)

    @action(detail=False, methods=['post'])
    def publish_month(self, request):
        if not self._check_hr(request):
            return Response({'error': 'Only HR/Admin can publish salary receipts'}, status=status.HTTP_403_FORBIDDEN)

        month_str = request.data.get('month')
        if not month_str:
            return Response({'error': 'month is required (YYYY-MM)'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            month_date = self._month_to_date(month_str)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        requested_state = request.data.get('is_published', True)
        if isinstance(requested_state, str):
            is_published = requested_state.strip().lower() in ['1', 'true', 'yes', 'on']
        else:
            is_published = bool(requested_state)

        payload = self._build_monthly_sheet_data(month_date, department_id=None)

        published_obj, _ = SalaryMonthPublish.objects.get_or_create(month=month_date)
        published_obj.is_active = is_published
        published_obj.published_by = request.user
        published_obj.save()

        if is_published:
            for row in payload.get('results', []):
                staff_id = row.get('staff_user_id')
                if not staff_id:
                    continue
                SalaryPublishedReceipt.objects.update_or_create(
                    month=month_date,
                    staff_id=staff_id,
                    defaults={
                        'is_salary_included': bool(row.get('include_in_salary', True)),
                        'receipt_data': row,
                        'published_by': request.user,
                    },
                )

        return Response({
            'message': 'Salary publish state updated',
            'month': month_str,
            'count': len(payload.get('results', [])) if is_published else 0,
            'is_published': is_published,
            'published_at': published_obj.published_at.isoformat(),
        })

    @action(detail=False, methods=['get'])
    def my_receipts(self, request):
        month_str = request.query_params.get('month')
        receipts = SalaryPublishedReceipt.objects.filter(staff=request.user)

        if month_str:
            try:
                month_date = self._month_to_date(month_str)
                receipts = receipts.filter(month=month_date)
            except ValueError as exc:
                return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        results = []
        for r in receipts.order_by('-month'):
            results.append({
                'id': r.id,
                'month': r.month.strftime('%Y-%m'),
                'is_salary_included': r.is_salary_included,
                'published_at': r.published_at.isoformat() if r.published_at else None,
                'receipt': r.receipt_data or {},
            })
        return Response({'results': results})

    @action(detail=False, methods=['get'])
    def salary_reports(self, request):
        if not self._check_hr(request):
            return Response({'error': 'Only HR/Admin can access salary reports'}, status=status.HTTP_403_FORBIDDEN)

        month_str = request.query_params.get('month')
        report_type = str(request.query_params.get('report_type') or 'payroll').strip().lower()
        bank_name = request.query_params.get('bank')
        export_format = str(
            request.query_params.get('export')
            or request.query_params.get('format')
            or 'json'
        ).strip().lower()

        if not month_str:
            return Response({'error': 'month is required (YYYY-MM)'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month_date = self._month_to_date(month_str)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        if report_type == 'payroll':
            report = self._build_payroll_report(month_date)
            if export_format == 'excel':
                return self._export_payroll_report_excel(report)
            return Response({
                'report_type': 'payroll',
                'report': report,
            })

        if report_type == 'bank_staff':
            report = self._build_bank_staff_report(month_date, bank_name=bank_name)
            if export_format == 'excel':
                return self._export_bank_staff_report_excel(report)
            return Response({
                'report_type': 'bank_staff',
                'report': report,
            })

        return Response(
            {'error': "Invalid report_type. Use 'payroll' or 'bank_staff'"},
            status=status.HTTP_400_BAD_REQUEST,
        )
