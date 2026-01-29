from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from .articulation_parser import parse_articulation_matrix_excel


def _to_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


_NUMERIC_TITLE_RE = re.compile(r"^\s*\d+(?:\.0+)?\s*$")


def _as_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return v != 0
    if isinstance(v, str):
        s = v.strip().lower()
        if s in ("true", "t", "yes", "y", "1", "x", "✓", "✔"):
            return True
        if s in ("false", "f", "no", "n", "0", ""):
            return False
    return bool(v)


def _col_letter_to_index(col: str) -> int:
    """A -> 1, Z -> 26, AA -> 27, ..."""
    col = col.strip().upper()
    n = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            continue
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def parse_cdap_excel(file_obj) -> Dict[str, Any]:
    """Parse CDAP Excel file (Revised CDP sheet) into the expected payload."""
    file_bytes = file_obj.read() if hasattr(file_obj, "read") else bytes(file_obj)
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    if not wb.worksheets:
        return {
            "rows": [],
            "books": {"textbook": "", "reference": ""},
            "active_learning": {"grid": [], "dropdowns": [], "optionsByRow": []},
            "articulation_extras": {},
        }

    all_sheet_titles = [ws.title for ws in wb.worksheets]

    def _looks_like_revised_cdp(ws) -> bool:
        return _find_cdp_header_row(ws) is not None

    ws_cdap = None
    # Prefer a name match like "REVISED CDP" / "1.REVISED CDP"
    for ws in wb.worksheets:
        name_norm = re.sub(r"[^a-z0-9]+", " ", _to_text(ws.title).lower()).strip()
        if "revised" in name_norm and "cdp" in name_norm:
            ws_cdap = ws
            break

    # Fallback: header detection
    if ws_cdap is None:
        for ws in wb.worksheets:
            if _looks_like_revised_cdp(ws):
                ws_cdap = ws
                break

    # Final fallback: first sheet
    if ws_cdap is None:
        ws_cdap = wb.worksheets[0]

    col_start_base = _col_letter_to_index("A")
    col_end_base = _col_letter_to_index("AG")

    def _norm_label(v: Any) -> str:
        s = _to_text(v).lower()
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _effective_cell_text(ws, r: int, c: int) -> str:
        v = ws.cell(row=r, column=c).value
        if v not in (None, ""):
            return _norm_label(v)
        try:
            ranges = list(getattr(ws, "merged_cells", None).ranges)  # type: ignore[attr-defined]
        except Exception:
            ranges = []
        if not ranges:
            return ""
        for rng in ranges:
            try:
                if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                    tl = ws.cell(row=rng.min_row, column=rng.min_col).value
                    return _norm_label(tl)
            except Exception:
                continue
        return ""

    def _find_cdp_header_row(ws) -> Optional[int]:
        max_scan_row = min(ws.max_row or 0, 80)
        for r in range(1, max_scan_row + 1):
            for c in range(col_start_base, col_end_base + 1):
                v = _effective_cell_text(ws, r, c)
                if re.fullmatch(r"content\s*[- ]?\s*type", v or ""):
                    return r
        return None

    def _find_header_col(ws, header_row: int, label_pattern: str) -> Optional[int]:
        if not header_row:
            return None
        for c in range(col_start_base, col_end_base + 1):
            v = _effective_cell_text(ws, header_row, c)
            if re.search(label_pattern, v or ""):
                return c
        return None

    detected_header_row_2 = _find_cdp_header_row(ws_cdap)
    header_row_2 = detected_header_row_2 or 12
    row_delta = header_row_2 - 12

    # Fixed row ranges per user-provided template
    unit_ranges = [
        (13, 21),
        (23, 31),
        (33, 41),
        (43, 51),
        (53, 61),
    ]
    unit_ranges = [(start + row_delta, end + row_delta) for start, end in unit_ranges]

    # Column mapping (1-based indices)
    col_unit = _col_letter_to_index("A")
    col_unit_name = _col_letter_to_index("B")
    col_co = _col_letter_to_index("C")
    col_content_type = _col_letter_to_index("D")
    col_part_no = _col_letter_to_index("E")
    col_topics = _col_letter_to_index("F")
    col_sub_topics = _col_letter_to_index("G")
    col_bt_level = _col_letter_to_index("H")
    col_total_hours = _col_letter_to_index("I")

    po_start = _col_letter_to_index("J")  # PO1
    po_end = _col_letter_to_index("T")    # PO11
    pso_start = _col_letter_to_index("U") # PSO1
    pso_end = _col_letter_to_index("W")   # PSO3

    parsed_rows: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []

    for unit_idx, (start_row, end_row) in enumerate(unit_ranges, start=1):
        raw_unit_value = _to_text(ws_cdap.cell(row=start_row, column=col_unit).value)
        unit_value = raw_unit_value
        # Normalize values like 5.0 -> 5
        if re.fullmatch(r"\d+(?:\.0+)?", unit_value):
            unit_value = str(int(float(unit_value)))

        unit_title = _to_text(ws_cdap.cell(row=start_row, column=col_unit_name).value)
        unit_co = _to_text(ws_cdap.cell(row=start_row, column=col_co).value)
        if not unit_title:
            warnings.append({"unit": unit_idx, "row": start_row, "warning": "Missing unit title in column B"})

        for r in range(start_row, end_row + 1):
            content_type = _to_text(ws_cdap.cell(row=r, column=col_content_type).value)
            if content_type and "#n/a" in content_type.replace(" ", "").lower():
                continue

            record: Dict[str, Any] = {
                "excel_row": r,
                "unit_index": unit_idx,
                "unit": unit_value if r == start_row else "",
                "unit_name": unit_title if r == start_row else "",
                "co": unit_co if r == start_row else "",
                "content_type": content_type,
                "part_no": _to_text(ws_cdap.cell(row=r, column=col_part_no).value),
                "topics": _to_text(ws_cdap.cell(row=r, column=col_topics).value),
                "sub_topics": _to_text(ws_cdap.cell(row=r, column=col_sub_topics).value),
                "bt_level": _to_text(ws_cdap.cell(row=r, column=col_bt_level).value),
                "total_hours_required": _to_text(ws_cdap.cell(row=r, column=col_total_hours).value),
            }

            # PO1..PO11 (J..T)
            for i, col in enumerate(range(po_start, po_end + 1), start=1):
                record[f"po{i}"] = _as_bool(ws_cdap.cell(row=r, column=col).value)

            # PSO1..PSO3 (U..W)
            for i, col in enumerate(range(pso_start, pso_end + 1), start=1):
                record[f"pso{i}"] = _as_bool(ws_cdap.cell(row=r, column=col).value)

            parsed_rows.append(record)

    def best_text_in_row(r: int, col_a: str = "A", col_b: str = "E") -> str:
        start = _col_letter_to_index(col_a)
        end = _col_letter_to_index(col_b)
        candidates: List[str] = []
        for c in range(start, end + 1):
            v = _to_text(ws_cdap.cell(row=r, column=c).value)
            if not v:
                continue
            if _NUMERIC_TITLE_RE.match(v):
                continue
            low = v.strip().lower()
            if low in {"textbook", "text book", "reference", "reference book", "reference books"}:
                continue
            candidates.append(v)
        if not candidates:
            return ""
        return max(candidates, key=len)

    def best_text_near_row(target_row: int) -> str:
        v = best_text_in_row(target_row)
        if v:
            return v
        for d in range(1, 6):
            down = best_text_in_row(target_row + d)
            if down:
                return down
            up = best_text_in_row(target_row - d)
            if up:
                return up
        return ""

    # Fixed cells per template: textbook in B64, reference in B68
    textbook_text = _to_text(ws_cdap.cell(row=64 + row_delta, column=_col_letter_to_index("B")).value)
    if not textbook_text:
        textbook_text = best_text_near_row(64 + row_delta)

    reference_text = _to_text(ws_cdap.cell(row=68 + row_delta, column=_col_letter_to_index("B")).value)

    def split_dropdown_options(text: str) -> List[str]:
        s = _to_text(text)
        if not s:
            return []
        s = s.replace("\r\n", "\n").replace("\r", "\n")
        parts = [p.strip() for p in re.split(r"[\n,;]+", s) if p and p.strip()]
        out: List[str] = []
        seen = set()
        for p in parts:
            if p not in seen:
                out.append(p)
                seen.add(p)
        return out

    g_col = _col_letter_to_index("G")
    active_learning_dropdown_rows = list(range(64 + row_delta, 71 + row_delta))
    active_learning_dropdown_options: List[List[str]] = []
    for r in active_learning_dropdown_rows:
        cell_text = ws_cdap.cell(row=r, column=g_col).value
        active_learning_dropdown_options.append(split_dropdown_options(_to_text(cell_text)))

    # Extract assessment rows from the Articulation Matrix sheet (page-2)
    articulation_extras: Dict[str, List[Dict[str, Any]]] = {}
    try:
        art = parse_articulation_matrix_excel(io.BytesIO(file_bytes))
        for u in art.get('units') or []:
            unit_label = _to_text(u.get('unit'))
            rows_list = u.get('rows') or []
            picked: List[Dict[str, Any]] = []
            for rr in rows_list:
                label = _to_text(rr.get('co_mapped')).strip().lower()
                if label.startswith('ssa') or label.startswith('active learning') or label.startswith('special activity'):
                    picked.append(rr)
            if picked:
                articulation_extras[unit_label] = picked
    except Exception:
        articulation_extras = {}

    return {
        "rows": parsed_rows,
        "books": {"textbook": textbook_text, "reference": reference_text},
        "active_learning": {
            "grid": [],
            "dropdowns": ["" for _ in range(7)],
            "optionsByRow": active_learning_dropdown_options,
        },
        "articulation_extras": articulation_extras,
        "meta": {
            "sheet_used": ws_cdap.title,
            "sheet_index_1based": (all_sheet_titles.index(ws_cdap.title) + 1) if ws_cdap.title in all_sheet_titles else None,
        },
        "warnings": warnings,
    }
