from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


def _to_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _norm_header(v: Any) -> str:
    s = _to_text(v).upper()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _as_number_or_dash(v: Any):
    """Keep '-' as '-', convert numeric-like to float/int, else return text."""
    if v is None:
        return "-"
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s == "-":
            return "-"
        # Sometimes topic no like '5.5.' should remain text
        if re.fullmatch(r"\d+(?:\.\d+)?\.?", s) and s.endswith("."):
            return s
        try:
            n = float(s)
            if n.is_integer():
                return int(n)
            return n
        except Exception:
            return s
    if isinstance(v, bool):
        return 1 if v else "-"
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            return int(v)
        return v
    return _to_text(v)


_PO_RE = re.compile(r"^PO\s*\d+$", re.I)
_PSO_RE = re.compile(r"^PSO\s*\d+$", re.I)


def _find_sheet_articulation(wb) -> Optional[Any]:
    # Prefer explicit title match
    for ws in wb.worksheets:
        title = _to_text(ws.title).lower()
        if "articulation" in title and "matrix" in title:
            return ws
    # Fallback: any sheet with PO1.. and PSO1 headers
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row or 0, 120) + 1):
            values = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column or 0, 40) + 1)]
            headers = {_norm_header(v) for v in values if not _is_blank(v)}
            if "S. NO" in headers and "CO MAPPED" in headers and any(_PO_RE.match(h) for h in headers):
                return ws
    return None


def _find_unit_blocks(ws) -> List[Tuple[str, int, int]]:
    """Return list of (unit_label, header_row, data_start_row)."""
    unit_blocks: List[Tuple[str, int, int]] = []

    for r in range(1, (ws.max_row or 0) + 1):
        unit_label = ws.cell(row=r, column=2).value  # Column B
        if isinstance(unit_label, str) and unit_label.strip().upper().startswith("UNIT"):
            # find header row after unit label
            header_row = None
            for rr in range(r, min(r + 10, ws.max_row or 0) + 1):
                if _norm_header(ws.cell(row=rr, column=2).value) == "S. NO":
                    header_row = rr
                    break
            if header_row:
                unit_blocks.append((unit_label.strip(), header_row, header_row + 1))

    return unit_blocks


def parse_articulation_matrix_excel(file_obj) -> Dict[str, Any]:
    file_bytes = file_obj.read() if hasattr(file_obj, "read") else bytes(file_obj)
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    if not wb.worksheets:
        return {"units": [], "summaries": {}, "meta": {"sheet_used": None}}

    ws = _find_sheet_articulation(wb) or wb.worksheets[0]

    units: List[Dict[str, Any]] = []

    # Columns (based on observed template)
    # B: S.NO, C: CO MAPPED, D: TOPIC NO, E: TOPIC NAME
    # F-P: PO1..PO11, Q-S: PSO1..PSO3, T: hours
    col_sno = 2
    col_co = 3
    col_topic_no = 4
    col_topic_name = 5
    po_start = 6
    po_count = 11
    pso_start = po_start + po_count
    pso_count = 3
    col_hours = pso_start + pso_count

    unit_blocks = _find_unit_blocks(ws)
    # If unit blocks are missing, fall back to first header row
    if not unit_blocks:
        for r in range(1, min(ws.max_row or 0, 150) + 1):
            if _norm_header(ws.cell(row=r, column=2).value) == "S. NO":
                unit_blocks = [("UNIT", r, r + 1)]
                break

    # Build a quick index of where each unit ends
    unit_starts = [b[1] for b in unit_blocks]
    unit_end_by_header = {}
    for idx, (_, header_row, _) in enumerate(unit_blocks):
        end_row = (unit_blocks[idx + 1][1] - 1) if idx + 1 < len(unit_blocks) else (ws.max_row or 0)
        unit_end_by_header[header_row] = end_row

    for (unit_label, header_row, data_start) in unit_blocks:
        end_row = unit_end_by_header.get(header_row, ws.max_row or 0)
        rows: List[Dict[str, Any]] = []

        for r in range(data_start, end_row + 1):
            # Stop if next unit header detected or we hit summary section
            b = ws.cell(row=r, column=2).value
            if isinstance(b, str) and b.strip().upper().startswith("UNIT"):
                break

            # Summary section usually has 'COs' at col E (5)
            e = ws.cell(row=r, column=5).value
            if isinstance(e, str) and e.strip().upper() in {"COS", "COs".upper()}:
                break

            sno = ws.cell(row=r, column=col_sno).value
            co = ws.cell(row=r, column=col_co).value
            topic_no = ws.cell(row=r, column=col_topic_no).value
            topic_name = ws.cell(row=r, column=col_topic_name).value

            # Consider row empty if it has no identifying fields
            if _is_blank(sno) and _is_blank(co) and _is_blank(topic_no) and _is_blank(topic_name):
                continue

            po_vals = []
            for i in range(po_count):
                po_vals.append(_as_number_or_dash(ws.cell(row=r, column=po_start + i).value))

            pso_vals = []
            for i in range(pso_count):
                pso_vals.append(_as_number_or_dash(ws.cell(row=r, column=pso_start + i).value))

            hours = _as_number_or_dash(ws.cell(row=r, column=col_hours).value)

            rows.append(
                {
                    "excel_row": r,
                    "s_no": _as_number_or_dash(sno),
                    "co_mapped": _to_text(co),
                    "topic_no": _as_number_or_dash(topic_no),
                    "topic_name": _to_text(topic_name),
                    "po": po_vals,
                    "pso": pso_vals,
                    "hours": hours,
                }
            )

        if rows:
            units.append({"unit": unit_label, "header_row": header_row, "rows": rows})

    # Parse CO vs PO/PSO summary table if present (typically around row ~92)
    summaries: Dict[str, Any] = {}
    # Find a row where col E is 'COs' and col F is 'PO1'
    summary_header_row = None
    for r in range(1, min(ws.max_row or 0, 250) + 1):
        if _norm_header(ws.cell(row=r, column=5).value) == "COS" and _norm_header(ws.cell(row=r, column=6).value) == "PO1":
            summary_header_row = r
            break

    if summary_header_row:
        # rows until we hit 'Average' in col E
        co_rows = []
        r = summary_header_row + 1
        while r <= (ws.max_row or 0):
            label = _to_text(ws.cell(row=r, column=5).value)
            if not label:
                r += 1
                continue
            if label.strip().upper() == "AVERAGE":
                avg_row = {
                    "label": "Average",
                    "po": [_as_number_or_dash(ws.cell(row=r, column=6 + i).value) for i in range(po_count)],
                    "pso": [_as_number_or_dash(ws.cell(row=r, column=6 + po_count + i).value) for i in range(pso_count)],
                    "average": _as_number_or_dash(ws.cell(row=r, column=6 + po_count + pso_count).value),
                }
                summaries["co_po_summary"] = {"rows": co_rows, "average": avg_row}
                break

            co_rows.append(
                {
                    "label": label,
                    "po": [_as_number_or_dash(ws.cell(row=r, column=6 + i).value) for i in range(po_count)],
                    "pso": [_as_number_or_dash(ws.cell(row=r, column=6 + po_count + i).value) for i in range(pso_count)],
                    "average": _as_number_or_dash(ws.cell(row=r, column=6 + po_count + pso_count).value),
                }
            )
            r += 1

    return {
        "units": units,
        "summaries": summaries,
        "meta": {
            "sheet_used": ws.title,
            "sheet_index_1based": wb.sheetnames.index(ws.title) + 1 if ws.title in wb.sheetnames else None,
        },
    }
