from decimal import Decimal, InvalidOperation
from datetime import datetime
import logging
import re

from django.contrib.auth.decorators import login_required
from django.db import transaction, connection
from django.db.models import Q
from django.db.utils import OperationalError, ProgrammingError
from django.http import HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.core.exceptions import FieldError

from academics.models import Subject, StudentProfile, TeachingAssignment, Semester

logger = logging.getLogger(__name__)

# Mark Entry Tabs View (Faculty OBE Section)
@login_required
def mark_entry_tabs(request, subject_id):
    # Faculty-only page (must have a staff profile)
    if not hasattr(request.user, 'staff_profile'):
        return HttpResponseForbidden('Faculty access only.')

    subject = _get_subject(subject_id, request)
    tab = request.GET.get('tab', 'dashboard')
    saved = request.GET.get('saved') == '1'

    # Basic student list for the subject semester (fallback to all students)
    # Order students by name (last, first, username) for SSA1/CIA1/Formative1 display
    students = (
        StudentProfile.objects.select_related('user', 'section')
        .filter(section__semester=subject.semester)
        .order_by('user__last_name', 'user__first_name', 'user__username')
    )
    if not students.exists():
        students = StudentProfile.objects.select_related('user', 'section').all().order_by('user__last_name', 'user__first_name', 'user__username')

    from .models import Cia1Mark, ProjectMark

    ta_id = _parse_int(request.GET.get('teaching_assignment_id') or request.POST.get('teaching_assignment_id'))
    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
    strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)

    errors: list[str] = []

    if request.method == 'POST' and tab in {'cia1', 'project'}:
        is_project_tab = tab == 'project'
        mark_model = ProjectMark if is_project_tab else Cia1Mark
        input_prefix = 'project_mark_' if is_project_tab else 'mark_'
        with transaction.atomic():
            for s in students:
                key = f'{input_prefix}{s.id}'
                raw = (request.POST.get(key) or '').strip()

                if raw == '':
                    # Blank => clear stored mark
                    _delete_scoped_mark(mark_model, subject=subject, student=s, teaching_assignment=ta)
                    continue

                try:
                    mark = Decimal(raw)
                except (InvalidOperation, ValueError):
                    errors.append(f'Invalid mark for {s.reg_no}: {raw}')
                    continue

                _upsert_scoped_mark(
                    mark_model,
                    subject=subject,
                    student=s,
                    teaching_assignment=ta,
                    mark_defaults={'mark': mark},
                )

        if not errors:
            return redirect(f"{request.path}?tab={tab}&saved=1")

    try:
        rows = _filter_marks_queryset_for_teaching_assignment(
            Cia1Mark.objects.filter(subject=subject, student__in=students),
            ta,
            strict_scope=strict_scope,
        )
        marks = {
            m.student_id: m.mark
            for m in rows
        }
    except OperationalError:
        marks = {}
    cia1_rows = [{'student': s, 'mark': marks.get(s.id)} for s in students]

    try:
        p_rows = _filter_marks_queryset_for_teaching_assignment(
            ProjectMark.objects.filter(subject=subject, student__in=students),
            ta,
            strict_scope=strict_scope,
        )
        project_marks = {
            m.student_id: m.mark
            for m in p_rows
        }
    except OperationalError:
        project_marks = {}

    if not project_marks:
        try:
            _backfill_project_marks_from_lab_published(
                subject=subject,
                teaching_assignment=ta,
                strict_scope=strict_scope,
            )
            p_rows = _filter_marks_queryset_for_teaching_assignment(
                ProjectMark.objects.filter(subject=subject, student__in=students),
                ta,
                strict_scope=strict_scope,
            )
            project_marks = {
                m.student_id: m.mark
                for m in p_rows
            }
        except Exception:
            pass

    project_rows = [{'student': s, 'mark': project_marks.get(s.id)} for s in students]

    context = {
        'subject': subject,
        'students': students,
        'tab': tab,
        'saved': saved,
        'cia1_rows': cia1_rows,
        'project_rows': project_rows,
        'errors': errors,
    }
    return render(request, 'OBE/mark_entry_tabs.html', context)
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.apps import apps

from .models import CdapRevision, CdapActiveLearningAnalysisMapping, ObeAssessmentMasterConfig
from .services.cdap_parser import parse_cdap_excel
from .services.articulation_parser import parse_articulation_matrix_excel
from .services.articulation_from_revision import build_articulation_matrix_from_revision_rows
from accounts.utils import get_user_permissions
from django.core.files.storage import default_storage
from django.conf import settings
import os
from .services.final_internal_marks import recompute_final_internal_marks


def _student_display_name(user) -> str:
    if not user:
        return ''
    try:
        full = ' '.join([
            str(getattr(user, 'first_name', '') or '').strip(),
            str(getattr(user, 'last_name', '') or '').strip(),
        ]).strip()
        if full:
            return full
    except Exception:
        pass
    return str(getattr(user, 'username', '') or '').strip()


def _get_students_for_teaching_assignment(ta):
    """Best-effort roster resolution (mirrors academics.TeachingAssignmentStudentsView)."""
    if not ta:
        return []

    students = []
    section_name = getattr(getattr(ta, 'section', None), 'name', None)

    # Elective TA rosters may not have a section; use elective-choices mapping.
    if not getattr(ta, 'section_id', None) and getattr(ta, 'elective_subject_id', None):
        try:
            from curriculum.models import ElectiveChoice

            eqs = (
                ElectiveChoice.objects.filter(is_active=True, elective_subject_id=int(ta.elective_subject_id))
                .exclude(student__isnull=True)
                .select_related('student__user', 'student__section')
            )
            if getattr(ta, 'academic_year_id', None):
                # Restrict by academic year only when it does not truncate
                # the active elective roster. Legacy/migrated data can spread
                # valid choices across year rows for the same elective.
                year_qs = eqs.filter(academic_year_id=ta.academic_year_id)
                if year_qs.exists() and year_qs.count() == eqs.count():
                    eqs = year_qs
            for c in eqs:
                sp = getattr(c, 'student', None)
                if not sp:
                    continue
                students.append(sp)
        except Exception:
            students = []
    elif getattr(ta, 'section', None) is not None:
        try:
            from academics.models import StudentSectionAssignment, StudentProfile

            s_qs = (
                StudentSectionAssignment.objects.filter(section=ta.section, end_date__isnull=True)
                .exclude(student__status__in=['INACTIVE', 'DEBAR'])
                .select_related('student__user')
                .order_by('student__reg_no')
            )
            for a in s_qs:
                students.append(a.student)

            # Also include legacy StudentProfile.section entries even when we already
            # have students from StudentSectionAssignment, to avoid silently dropping
            # students whose section assignment rows were never backfilled.
            try:
                existing_ids = {int(getattr(s, 'id', None)) for s in students if getattr(s, 'id', None) is not None}
            except Exception:
                existing_ids = set()

            sp_qs = (
                StudentProfile.objects.filter(section=ta.section)
                .exclude(status__in=['INACTIVE', 'DEBAR'])
                .select_related('user')
                .order_by('reg_no')
            )
            for sp in sp_qs:
                try:
                    sid = int(getattr(sp, 'id', None))
                except Exception:
                    continue
                if sid in existing_ids:
                    continue
                students.append(sp)
        except Exception:
            students = []

    # ── Batch-based student filtering ──
    # If the TA's staff has StudentSubjectBatch entries for this subject/year,
    # restrict the student list to only those in the staff's batches.
    # This ensures batch-assigned staff see only their batch students in marks pages.
    if students and getattr(ta, 'staff_id', None):
        try:
            from academics.models import StudentSubjectBatch as _SSB
            batch_qs = _SSB.objects.filter(
                staff_id=ta.staff_id,
                is_active=True,
            )
            if getattr(ta, 'academic_year_id', None):
                batch_qs = batch_qs.filter(academic_year_id=ta.academic_year_id)
            if getattr(ta, 'curriculum_row_id', None):
                batch_qs = batch_qs.filter(curriculum_row_id=ta.curriculum_row_id)
            else:
                batch_qs = batch_qs.filter(curriculum_row__isnull=True)
            user_batches = list(batch_qs)
            if user_batches:
                batch_student_ids = set()
                for ub in user_batches:
                    try:
                        batch_student_ids.update(
                            ub.students.values_list('id', flat=True)
                        )
                    except Exception:
                        pass
                if batch_student_ids:
                    students = [s for s in students if getattr(s, 'id', None) in batch_student_ids]
        except Exception:
            pass

    return students


def _subject_teaching_assignment_q(subject_code: str):
    code = str(subject_code or '').strip()
    return (
        Q(subject__code__iexact=code)
        | Q(curriculum_row__course_code__iexact=code)
        | Q(curriculum_row__master__course_code__iexact=code)
    )


def _has_parallel_teaching_assignments(*, subject_code: str, teaching_assignment=None) -> bool:
    ta_id = getattr(teaching_assignment, 'id', None)
    if ta_id is None:
        return False
    try:
        qs = TeachingAssignment.objects.filter(is_active=True).filter(_subject_teaching_assignment_q(subject_code))
        return qs.exclude(id=ta_id).exists()
    except Exception:
        return False


def _strict_assignment_scope(*, subject_code: str, teaching_assignment=None) -> bool:
    return teaching_assignment is not None and _has_parallel_teaching_assignments(
        subject_code=subject_code,
        teaching_assignment=teaching_assignment,
    )


def _get_teaching_assignment_student_ids(ta) -> list[int]:
    ids: list[int] = []
    for student in _get_students_for_teaching_assignment(ta):
        student_id = getattr(student, 'id', None)
        if isinstance(student_id, int):
            ids.append(student_id)
    return ids


def _get_studentprofile_queryset_for_teaching_assignment(ta):
    """Return a stable StudentProfile queryset for the TA's roster."""
    from academics.models import StudentProfile

    ids = _get_teaching_assignment_student_ids(ta)
    if not ids:
        return StudentProfile.objects.none()
    return (
        StudentProfile.objects.select_related('user', 'section')
        .filter(id__in=ids)
        .order_by('user__last_name', 'user__first_name', 'user__username')
    )


_DB_COLUMN_EXISTS_CACHE: dict[tuple[str, str], bool] = {}


def _db_table_has_column(db_table: str, column_name: str) -> bool:
    key = (db_table, column_name)
    cached = _DB_COLUMN_EXISTS_CACHE.get(key)
    if cached is not None:
        return cached
    try:
        with connection.cursor() as cursor:
            desc = connection.introspection.get_table_description(cursor, db_table)
        present = any(getattr(c, 'name', None) == column_name for c in desc)
    except Exception:
        present = False
    _DB_COLUMN_EXISTS_CACHE[key] = present
    return present


def _filter_marks_queryset_for_teaching_assignment(qs, ta, *, strict_scope: bool = False):
    model = getattr(qs, 'model', None)
    has_ta_field = bool(model) and any(getattr(f, 'name', '') == 'teaching_assignment' for f in model._meta.get_fields())
    has_ta_db_column = False
    if has_ta_field and model is not None:
        try:
            has_ta_db_column = _db_table_has_column(model._meta.db_table, 'teaching_assignment_id')
        except Exception:
            has_ta_db_column = False

    if has_ta_field and has_ta_db_column:
        if ta is not None:
            scoped_qs = qs.filter(teaching_assignment=ta)
            if scoped_qs.exists():
                return scoped_qs
            if strict_scope:
                # In strict scope mode (multiple parallel TAs for the same subject),
                # we still want to show legacy/unscoped marks (NULL teaching_assignment)
                # that belong to the selected TA roster. Many older DBs stored CIA marks
                # without TA linkage; returning qs.none() hides marks entirely.
                student_ids = _get_teaching_assignment_student_ids(ta)
                if student_ids:
                    null_scoped = qs.filter(teaching_assignment__isnull=True, student_id__in=student_ids)
                    if null_scoped.exists():
                        return null_scoped
                    # As a last resort (when TA linkage is inconsistent), return roster-filtered marks.
                    roster_scoped = qs.filter(student_id__in=student_ids)
                    if roster_scoped.exists():
                        return roster_scoped
                return qs.none()
            return qs.filter(teaching_assignment__isnull=True)
        return qs.filter(teaching_assignment__isnull=True)

    # If the model expects teaching-assignment scoping but the DB column is missing,
    # the marks stored in the table are effectively unscoped. We *can* try to
    # approximate scoping by filtering by the TA roster, but that roster can be
    # incomplete or unavailable. To avoid hiding published marks, fall back to
    # unscoped rows when roster scoping yields no results.
    if has_ta_field and not has_ta_db_column:
        if ta is None:
            return qs

        student_ids = _get_teaching_assignment_student_ids(ta)
        if not student_ids:
            return qs

        try:
            scoped = qs.filter(student_id__in=student_ids)
            return scoped if scoped.exists() else qs
        except Exception:
            return qs

    if ta is None:
        return qs

    student_ids = _get_teaching_assignment_student_ids(ta)
    # If we cannot resolve the roster for the teaching assignment (common for
    # partially seeded data), avoid hiding published marks entirely. In DBs
    # that lack `teaching_assignment_id`, marks cannot be scoped reliably
    # anyway; prefer returning unscoped rows over an empty result.
    if not student_ids:
        return qs
    return qs.filter(student_id__in=student_ids)


def _safe_marks_map_for_subject(model, *, subject, ta, strict_scope: bool = False) -> dict[str, str | None]:
    """Return {student_id: mark_str_or_none} without selecting missing TA columns.

    On partially migrated DBs, the Django model may include a `teaching_assignment`
    FK while the DB column `teaching_assignment_id` is absent. Fetching model
    instances would then raise ProgrammingError because Django selects all model
    columns by default.
    """
    has_ta_field = any(getattr(f, 'name', '') == 'teaching_assignment' for f in model._meta.get_fields())
    has_ta_db_column = False
    if has_ta_field:
        try:
            has_ta_db_column = _db_table_has_column(model._meta.db_table, 'teaching_assignment_id')
        except Exception:
            has_ta_db_column = False

    qs = model.objects.filter(subject=subject)

    # Normal path: DB supports TA scoping.
    if has_ta_field and has_ta_db_column:
        qs = _filter_marks_queryset_for_teaching_assignment(qs, ta, strict_scope=strict_scope)
        pairs = qs.values_list('student_id', 'mark')
        return {str(sid): (str(mark) if mark is not None else None) for sid, mark in pairs}

    # Missing TA column: never reference teaching_assignment in SQL.
    # Try roster-based scoping; if it yields nothing, fall back to unscoped marks.
    if has_ta_field and not has_ta_db_column and ta is not None:
        student_ids = _get_teaching_assignment_student_ids(ta)
        if student_ids:
            try:
                scoped = qs.filter(student_id__in=student_ids)
                scoped_pairs = list(scoped.values_list('student_id', 'mark'))
                if scoped_pairs:
                    return {str(sid): (str(mark) if mark is not None else None) for sid, mark in scoped_pairs}
            except Exception:
                pass

    pairs = qs.values_list('student_id', 'mark')
    return {str(sid): (str(mark) if mark is not None else None) for sid, mark in pairs}


def _safe_formative_marks_map(model, *, subject, ta, strict_scope: bool = False) -> dict[str, dict[str, str | None]]:
    """Return formative marks without selecting missing TA columns."""
    has_ta_field = any(getattr(f, 'name', '') == 'teaching_assignment' for f in model._meta.get_fields())
    has_ta_db_column = False
    if has_ta_field:
        try:
            has_ta_db_column = _db_table_has_column(model._meta.db_table, 'teaching_assignment_id')
        except Exception:
            has_ta_db_column = False

    qs = model.objects.filter(subject=subject)
    if has_ta_field and has_ta_db_column:
        qs = _filter_marks_queryset_for_teaching_assignment(qs, ta, strict_scope=strict_scope)
    elif has_ta_field and not has_ta_db_column and ta is not None:
        student_ids = _get_teaching_assignment_student_ids(ta)
        if student_ids:
            try:
                scoped = qs.filter(student_id__in=student_ids)
                # If scoping yields rows, use them; otherwise fall back unscoped.
                if scoped.exists():
                    qs = scoped
            except Exception:
                pass

    cols = ['student_id', 'skill1', 'skill2', 'att1', 'att2', 'total']
    out: dict[str, dict[str, str | None]] = {}
    for sid, skill1, skill2, att1, att2, total in qs.values_list(*cols):
        out[str(sid)] = {
            'skill1': str(skill1) if skill1 is not None else None,
            'skill2': str(skill2) if skill2 is not None else None,
            'att1': str(att1) if att1 is not None else None,
            'att2': str(att2) if att2 is not None else None,
            'total': str(total) if total is not None else None,
        }
    return out


def _upsert_scoped_mark(model, *, subject, student, mark_defaults: dict, teaching_assignment=None):
    has_ta_field = any(getattr(f, 'name', '') == 'teaching_assignment' for f in model._meta.get_fields())
    has_ta_db_column = False
    if has_ta_field:
        try:
            has_ta_db_column = _db_table_has_column(model._meta.db_table, 'teaching_assignment_id')
        except Exception:
            has_ta_db_column = False

    # If the Django model has `teaching_assignment` but the DB is missing the column
    # (partially migrated production DB), ORM queries will fail because Django will
    # try to SELECT/INSERT the non-existent column. Fall back to raw SQL that only
    # touches columns that exist.
    if has_ta_field and not has_ta_db_column:
        table = model._meta.db_table
        subj_id = getattr(subject, 'id', subject)
        stud_id = getattr(student, 'id', student)

        # Keep only columns that exist in the DB table.
        to_set: dict[str, object] = {}
        for k, v in (mark_defaults or {}).items():
            try:
                if _db_table_has_column(table, str(k)):
                    to_set[str(k)] = v
            except Exception:
                # Assume column exists; if it doesn't, the SQL will error and be
                # caught by the caller's transaction.
                to_set[str(k)] = v

        # Some mark tables have NOT NULL timestamps; ORM would fill these.
        # Ensure they are set in raw SQL mode.
        now = timezone.now()
        try:
            if _db_table_has_column(table, 'updated_at') and 'updated_at' not in to_set:
                to_set['updated_at'] = now
        except Exception:
            pass

        if not to_set:
            return (None, False)

        with connection.cursor() as cursor:
            cursor.execute(
                f'SELECT id FROM "{table}" WHERE subject_id = %s AND student_id = %s LIMIT 1',
                [subj_id, stud_id],
            )
            row = cursor.fetchone()
            if row:
                row_id = row[0]
                set_cols = list(to_set.keys())
                set_clause = ', '.join([f'"{c}" = %s' for c in set_cols])
                params = [to_set[c] for c in set_cols] + [row_id]
                cursor.execute(
                    f'UPDATE "{table}" SET {set_clause} WHERE id = %s',
                    params,
                )
                return (None, False)

            ins_to_set = dict(to_set)
            try:
                if _db_table_has_column(table, 'created_at') and 'created_at' not in ins_to_set:
                    ins_to_set['created_at'] = now
            except Exception:
                pass
            try:
                if _db_table_has_column(table, 'updated_at') and 'updated_at' not in ins_to_set:
                    ins_to_set['updated_at'] = now
            except Exception:
                pass

            ins_cols = ['subject_id', 'student_id'] + list(ins_to_set.keys())
            placeholders = ', '.join(['%s'] * len(ins_cols))
            ins_clause = ', '.join([f'"{c}"' for c in ins_cols])
            params = [subj_id, stud_id] + [ins_to_set[c] for c in ins_to_set.keys()]
            cursor.execute(
                f'INSERT INTO "{table}" ({ins_clause}) VALUES ({placeholders})',
                params,
            )
        return (None, True)

    lookup = {'subject': subject, 'student': student}
    if has_ta_field and has_ta_db_column:
        lookup['teaching_assignment'] = teaching_assignment
    return model.objects.update_or_create(**lookup, defaults=mark_defaults)


def _delete_scoped_mark(model, *, subject, student, teaching_assignment=None):
    has_ta_field = any(getattr(f, 'name', '') == 'teaching_assignment' for f in model._meta.get_fields())
    has_ta_db_column = False
    if has_ta_field:
        try:
            has_ta_db_column = _db_table_has_column(model._meta.db_table, 'teaching_assignment_id')
        except Exception:
            has_ta_db_column = False

    # Same schema-mismatch fallback as _upsert_scoped_mark.
    if has_ta_field and not has_ta_db_column:
        table = model._meta.db_table
        subj_id = getattr(subject, 'id', subject)
        stud_id = getattr(student, 'id', student)
        with connection.cursor() as cursor:
            cursor.execute(
                f'DELETE FROM "{table}" WHERE subject_id = %s AND student_id = %s',
                [subj_id, stud_id],
            )
            deleted = int(getattr(cursor, 'rowcount', 0) or 0)
        # ORM delete returns (count, details).
        return (deleted, {})

    filters = {'subject': subject, 'student': student}
    if has_ta_field and has_ta_db_column:
        filters['teaching_assignment'] = teaching_assignment
    return model.objects.filter(**filters).delete()


def _get_scoped_obe_json_row(model, *, subject, teaching_assignment=None, strict_scope: bool = False, assessment: str | None = None):
    qs = model.objects.filter(subject=subject)
    if assessment is not None:
        qs = qs.filter(assessment=assessment)

    has_ta_field = any(getattr(f, 'name', '') == 'teaching_assignment' for f in model._meta.get_fields())
    if not has_ta_field:
        return qs.order_by('-updated_at', '-pk').first()

    if teaching_assignment is not None:
        row = qs.filter(teaching_assignment=teaching_assignment).order_by('-updated_at', '-pk').first()
        if row is not None:
            return row
        if strict_scope:
            return None
        row = qs.filter(teaching_assignment__isnull=True).order_by('-updated_at', '-pk').first()
        if row is not None:
            return row

    return qs.filter(teaching_assignment__isnull=True).order_by('-updated_at', '-pk').first()


def _upsert_scoped_obe_json_row(model, *, subject, defaults: dict, teaching_assignment=None, assessment: str | None = None):
    lookup = {'subject': subject}
    if assessment is not None:
        lookup['assessment'] = assessment

    has_ta_field = any(getattr(f, 'name', '') == 'teaching_assignment' for f in model._meta.get_fields())
    if has_ta_field:
        lookup['teaching_assignment'] = teaching_assignment

    return model.objects.update_or_create(**lookup, defaults=defaults)


def _delete_scoped_obe_json_rows(model, *, subject, teaching_assignment=None, strict_scope: bool = False, assessment: str | None = None) -> int:
    qs = model.objects.filter(subject=subject)
    if assessment is not None:
        qs = qs.filter(assessment=assessment)

    has_ta_field = any(getattr(f, 'name', '') == 'teaching_assignment' for f in model._meta.get_fields())
    if not has_ta_field:
        return int(qs.delete()[0] or 0)

    if teaching_assignment is not None:
        deleted = int(qs.filter(teaching_assignment=teaching_assignment).delete()[0] or 0)
        if strict_scope:
            return deleted
        deleted += int(qs.filter(teaching_assignment__isnull=True).delete()[0] or 0)
        return deleted

    return int(qs.filter(teaching_assignment__isnull=True).delete()[0] or 0)


def _delete_marks_rows_for_reset(qs, *, ta=None, strict_scope: bool = False) -> int:
    """Delete mark rows for reset, including legacy fallback rows when non-strict.

    Reset must be deterministic: once reset succeeds, no TA/legacy fallback should
    repopulate marks for the same subject+assessment on refresh.
    """
    model = getattr(qs, 'model', None)
    has_ta_field = bool(model) and any(getattr(f, 'name', '') == 'teaching_assignment' for f in model._meta.get_fields())
    has_ta_db_column = False
    if has_ta_field and model is not None:
        try:
            has_ta_db_column = _db_table_has_column(model._meta.db_table, 'teaching_assignment_id')
        except Exception:
            has_ta_db_column = False

    if has_ta_field and has_ta_db_column:
        if ta is not None:
            deleted = int(qs.filter(teaching_assignment=ta).delete()[0] or 0)
            if strict_scope:
                return deleted
            deleted += int(qs.filter(teaching_assignment__isnull=True).delete()[0] or 0)
            return deleted
        return int(qs.filter(teaching_assignment__isnull=True).delete()[0] or 0)

    if has_ta_field and not has_ta_db_column:
        if ta is None:
            return int(qs.delete()[0] or 0)
        student_ids = _get_teaching_assignment_student_ids(ta)
        if not student_ids:
            return int(qs.delete()[0] or 0)
        return int(qs.filter(student_id__in=student_ids).delete()[0] or 0)

    if ta is None:
        return int(qs.delete()[0] or 0)

    student_ids = _get_teaching_assignment_student_ids(ta)
    if not student_ids:
        return int(qs.delete()[0] or 0)
    return int(qs.filter(student_id__in=student_ids).delete()[0] or 0)


def _get_cia_questions_for_export(*, subject, assessment_key: str, teaching_assignment=None, strict_scope: bool = False, class_type: str = '', question_paper_type: str = '') -> list[dict]:
    """Return question definitions for CIA export.

    Priority (most accurate first):
    1) Saved draft sheet questions (exactly what the CIA entry UI is using)
    2) Published sheet snapshot questions
    3) IQAC QP pattern config (marks [+ optional CO mapping])
    4) OBE master config questions
    5) Hard fallback (common template)

        Output shape:
            [{ key: 'q1', label: 'Q1', max: 2.0 }, ...]
    """
    assessment_key = str(assessment_key or '').strip().lower()
    class_type = _normalize_obe_class_type(class_type)
    question_paper_type = str(question_paper_type or '').strip().upper()

    def _coerce_questions(raw):
        if not isinstance(raw, list):
            return []
        out = []
        for idx, q in enumerate(raw):
            if not isinstance(q, dict):
                continue
            key = str(q.get('key') or f"q{idx + 1}").strip() or f"q{idx + 1}"
            label = str(q.get('label') or q.get('key') or f"Q{idx + 1}")
            try:
                mx = float(q.get('max') or 0)
            except Exception:
                mx = 0.0
            if not label:
                label = f"Q{idx + 1}"
            out.append({'key': key, 'label': label, 'max': mx})
        return out

    # 1) Draft sheet questions (frontend saves full sheet JSON in AssessmentDraft)
    try:
        from .models import AssessmentDraft

        d = _get_scoped_obe_json_row(
            AssessmentDraft,
            subject=subject,
            teaching_assignment=teaching_assignment,
            strict_scope=strict_scope,
            assessment=assessment_key,
        )
        if d and isinstance(getattr(d, 'data', None), dict):
            qs = d.data.get('questions')
            out = _coerce_questions(qs)
            if out:
                return out
    except Exception:
        pass

    # 2) Published sheet questions (snapshot)
    try:
        if assessment_key == 'cia1':
            from .models import Cia1PublishedSheet

            row = _get_scoped_obe_json_row(
                Cia1PublishedSheet,
                subject=subject,
                teaching_assignment=teaching_assignment,
                strict_scope=strict_scope,
            )
        else:
            from .models import Cia2PublishedSheet

            row = _get_scoped_obe_json_row(
                Cia2PublishedSheet,
                subject=subject,
                teaching_assignment=teaching_assignment,
                strict_scope=strict_scope,
            )
        data = row.data if row and isinstance(getattr(row, 'data', None), dict) else None
        if isinstance(data, dict):
            out = _coerce_questions(data.get('questions'))
            if out:
                return out
    except Exception:
        pass

    # 3) Prefer IQAC QP patterns, matching frontend logic.
    try:
        from .models import ObeQpPatternConfig

        exam = 'CIA1' if assessment_key == 'cia1' else 'CIA2'
        qp_for_db = question_paper_type if (class_type == 'THEORY' and question_paper_type in {'QP1', 'QP2'}) else None
        obj = ObeQpPatternConfig.objects.filter(class_type=class_type, question_paper_type=qp_for_db, exam=exam).first()
        if obj is None and exam in {'CIA1', 'CIA2'}:
            obj = ObeQpPatternConfig.objects.filter(class_type=class_type, question_paper_type=qp_for_db, exam='CIA').first()

        raw = getattr(obj, 'pattern', None) if obj is not None else None
        if isinstance(raw, dict):
            raw_marks = raw.get('marks')
        else:
            raw_marks = raw

        if isinstance(raw_marks, list) and raw_marks:
            out = []
            for i, x in enumerate(raw_marks):
                try:
                    mx = float(x)
                except Exception:
                    mx = 0.0
                out.append({'key': f"q{i + 1}", 'label': f"Q{i + 1}", 'max': mx})
            if out:
                return out
    except Exception:
        pass

    # 4) Fallback: master config (if it has CIA question defs)
    try:
        row = ObeAssessmentMasterConfig.objects.filter(id=1).first()
        cfg = row.config if row and isinstance(getattr(row, 'config', None), dict) else {}
        assessments = cfg.get('assessments') if isinstance(cfg, dict) else None
        if isinstance(assessments, dict):
            key = assessment_key if assessment_key in assessments else 'cia1'
            qs = (assessments.get(key) or {}).get('questions')
            out = _coerce_questions(qs)
            if out:
                return out
    except Exception:
        pass

    # 5) Last resort: match the common CIA sheet shown in the user's template.
    # Return 9 questions ending with Q9=16; QP2 Excel split will convert Q9 into Q9+Q10 (8+8) later.
    return [{'key': f"q{i + 1}", 'label': f"Q{i + 1}", 'max': float(mx)} for i, mx in enumerate([2, 2, 2, 2, 2, 2, 16, 16, 16])]


def _get_cia_sheet_rows_for_export(*, subject, assessment_key: str, teaching_assignment=None, strict_scope: bool = False) -> dict:
    """Return rowsByStudentId map for CIA sheet export.

    Priority:
      1) Draft rows (most likely what staff is currently editing)
      2) Published rows
      3) Empty
    """
    assessment_key = str(assessment_key or '').strip().lower()
    # Draft
    try:
        from .models import AssessmentDraft

        d = _get_scoped_obe_json_row(
            AssessmentDraft,
            subject=subject,
            teaching_assignment=teaching_assignment,
            strict_scope=strict_scope,
            assessment=assessment_key,
        )
        if d and isinstance(getattr(d, 'data', None), dict):
            rows_by = d.data.get('rowsByStudentId')
            if isinstance(rows_by, dict):
                return rows_by
    except Exception:
        pass

    # Published
    try:
        if assessment_key == 'cia1':
            from .models import Cia1PublishedSheet

            row = _get_scoped_obe_json_row(
                Cia1PublishedSheet,
                subject=subject,
                teaching_assignment=teaching_assignment,
                strict_scope=strict_scope,
            )
        else:
            from .models import Cia2PublishedSheet

            row = _get_scoped_obe_json_row(
                Cia2PublishedSheet,
                subject=subject,
                teaching_assignment=teaching_assignment,
                strict_scope=strict_scope,
            )
        data = row.data if row and isinstance(getattr(row, 'data', None), dict) else None
        if isinstance(data, dict):
            rows_by = data.get('rowsByStudentId')
            if isinstance(rows_by, dict):
                return rows_by
    except Exception:
        pass

    return {}


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cia_export_template_xlsx(request, assessment: str, subject_id: str):
    """Download a protected CIA Excel template.

    Columns (matches screenshot):
      Register No | Student Name | Qn (max) ... | Status
    Only the Q columns and Status are editable; the rest is locked.

    Query params:
      teaching_assignment_id (optional, but recommended for correct roster)
    """
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    assessment_key = str(assessment or '').strip().lower()
    if assessment_key not in {'cia1', 'cia2'}:
        return Response({'detail': 'assessment must be cia1 or cia2'}, status=status.HTTP_400_BAD_REQUEST)

    subject = _get_subject(subject_id, request)
    ta_id = _get_teaching_assignment_id_from_request(request)

    # Enforce enabled assessment for SPECIAL courses.
    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject.code, assessment=assessment_key, teaching_assignment_id=ta_id)
    if gate is not None:
        return gate

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
    strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
    row = _resolve_curriculum_row_for_subject(request, subject_code=subject.code, teaching_assignment_id=ta_id)

    class_type = str(getattr(row, 'class_type', '') or '').strip().upper() if row else ''
    qp_type = str(getattr(row, 'question_paper_type', '') or '').strip().upper() if row else ''

    # Allow caller to specify QP type explicitly (useful for IQAC flows where curriculum row
    # resolution is not reliable, but the UI knows the question paper type).
    try:
        req_qp = None
        if hasattr(request, 'query_params'):
            req_qp = request.query_params.get('question_paper_type')
        if req_qp is None:
            req_qp = request.GET.get('question_paper_type')
        if req_qp:
            qp_type = str(req_qp).strip().upper()
    except Exception:
        pass

    q_defs = _get_cia_questions_for_export(
        subject=subject,
        assessment_key=assessment_key,
        teaching_assignment=ta,
        strict_scope=strict_scope,
        class_type=class_type,
        question_paper_type=qp_type,
    )
    rows_by_student_id = _get_cia_sheet_rows_for_export(
        subject=subject,
        assessment_key=assessment_key,
        teaching_assignment=ta,
        strict_scope=strict_scope,
    )

    # QP2-specific Excel-only tweak:
    # Split the last question (typically Q9 out of 16) into two columns Q9 and Q10
    # each out of 8. This does NOT change the UI/table or persisted keys; it only
    # affects the downloaded template.
    qp2_excel_split = str(qp_type or '').strip().upper() == 'QP2'

    q_headers = []
    q_keys: list[str] = []

    def _split_qp2_defs(defs: list[dict]) -> list[dict]:
        if not qp2_excel_split:
            return defs
        if not isinstance(defs, list) or not defs:
            return defs

        # Prefer splitting the question with key 'q9', else split the last question.
        split_idx = None
        for i, d in enumerate(defs):
            k = str((d or {}).get('key') or '').strip().lower()
            if k == 'q9':
                split_idx = i
                break
        if split_idx is None:
            split_idx = len(defs) - 1

        base = defs[split_idx] or {}
        try:
            mx = float(base.get('max') or 0)
        except Exception:
            mx = 0.0

        # Requirement: for QP2, split 16 -> 8 and 8.
        half = 8.0 if abs(mx - 16.0) < 1e-6 else (mx / 2.0 if mx else 0.0)

        left = {**base, 'key': 'q9', 'label': 'Q9', 'max': half}
        right = {**base, 'key': 'q10', 'label': 'Q10', 'max': half}

        return list(defs[:split_idx]) + [left, right] + list(defs[split_idx + 1 :])

    q_defs_for_excel = _split_qp2_defs(list(q_defs or []))

    for i, q in enumerate(q_defs_for_excel):
        key = str((q or {}).get('key') or f"q{i + 1}").strip() or f"q{i + 1}"
        label = str((q or {}).get('label') or f"Q{i + 1}").strip() or f"Q{i + 1}"
        try:
            mx = float((q or {}).get('max') or 0)
        except Exception:
            mx = 0.0
        q_keys.append(key)
        q_headers.append(f"{label} ({mx:.2f})")

    # Determine roster.
    students = _get_students_for_teaching_assignment(ta) if ta is not None else []
    if not students:
        students = (
            StudentProfile.objects.select_related('user', 'section')
            .filter(section__semester=subject.semester)
            .order_by('reg_no')
        )
        if not students.exists():
            students = StudentProfile.objects.select_related('user', 'section').all().order_by('reg_no')

    # Sort roster alphabetically by student name (matches UI display order).
    def _sort_name(sp):
        u = getattr(sp, 'user', None)
        if not u:
            return ''
        return ' '.join([
            str(getattr(u, 'first_name', '') or '').strip(),
            str(getattr(u, 'last_name', '') or '').strip(),
        ]).strip().upper() or str(getattr(u, 'username', '') or '').strip().upper()
    students = sorted(students, key=_sort_name)

    # Build workbook
    import openpyxl
    from openpyxl.styles import Font, Protection

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = assessment_key.upper()

    headers = ['Register No', 'Student Name'] + q_headers + ['Status']
    ws.append(headers)

    # Style header
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)

    # Column widths (roughly matching screenshot)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 28
    # Q columns start at C
    for i in range(len(q_headers)):
        col_letter = openpyxl.utils.get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(3 + len(q_headers))].width = 12  # Status

    # Fill rows
    start_row = 2
    def _cell_value(raw):
        if raw is None:
            return ''
        if isinstance(raw, str) and raw.strip() == '':
            return ''
        # Keep ints as ints when possible.
        try:
            if isinstance(raw, bool):
                return '1' if raw else '0'
            if isinstance(raw, (int, float, Decimal)):
                return float(raw)
            s = str(raw).strip()
            if s == '':
                return ''
            # If numeric string, keep numeric.
            try:
                return float(s)
            except Exception:
                return s
        except Exception:
            return ''

    for idx, sp in enumerate(list(students), start=start_row):
        reg_no = str(getattr(sp, 'reg_no', '') or '')
        name = _student_display_name(getattr(sp, 'user', None))

        sid = getattr(sp, 'id', None)
        row_obj = rows_by_student_id.get(str(sid)) if sid is not None else None
        if not isinstance(row_obj, dict):
            row_obj = {}
        absent = bool(row_obj.get('absent'))
        q_map = row_obj.get('q') if isinstance(row_obj.get('q'), dict) else {}

        q_vals = []
        # QP2 Excel split: extract q9 total once before the loop.
        # When exporting existing marks, split q9 into q9 (max 8) and q10 (max 8).
        q9_total = None
        if qp2_excel_split:
            try:
                q9_raw = q_map.get('q9')
                if q9_raw not in (None, ''):
                    q9_total = float(q9_raw)
            except Exception:
                q9_total = None

        q9_half = None
        if qp2_excel_split and q9_total is not None:
            try:
                # Requirement: split equally (e.g., 16->8+8, 12->6+6, 10->5+5).
                q9_half = float(q9_total) / 2.0
                # Cap each column by its max (normally 8).
                q9_half = max(0.0, min(8.0, q9_half))
                # Keep the sheet neat.
                q9_half = round(q9_half, 2)
            except Exception:
                q9_half = None

        for k in q_keys:
            lk = str(k or '').strip().lower()
            # QP2: split q9 value across q9 and q10 columns
            if qp2_excel_split and lk == 'q10':
                if q9_half is None:
                    q_vals.append('')
                else:
                    q_vals.append(_cell_value(q9_half))
                continue
            if qp2_excel_split and lk == 'q9':
                if q9_half is None:
                    q_vals.append('')
                else:
                    q_vals.append(_cell_value(q9_half))
                continue
            # All other questions: use normal value
            q_vals.append(_cell_value(q_map.get(k)))

        status_val = 'absent' if absent else 'present'
        row_values = [reg_no, name] + q_vals + [status_val]
        ws.append(row_values)

    # Freeze header
    ws.freeze_panes = 'C2'

    # Protect sheet: lock everything except Q columns + Status column.
    ws.protection.sheet = True
    try:
        # Use SECRET_KEY-derived password so users can't trivially unprotect.
        ws.protection.set_password(str(getattr(settings, 'SECRET_KEY', 'obe'))[:32])
    except Exception:
        pass

    # Unlock editable cells for all data rows.
    last_row = ws.max_row
    q_start_col = 3
    q_end_col = 2 + len(q_headers)  # inclusive
    status_col = 3 + len(q_headers)

    for r in range(2, last_row + 1):
        for c in range(q_start_col, status_col + 1):
            ws.cell(row=r, column=c).protection = Protection(locked=False)
        # Explicitly lock Reg No + Student Name
        ws.cell(row=r, column=1).protection = Protection(locked=True)
        ws.cell(row=r, column=2).protection = Protection(locked=True)

    # Save to response
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"{subject.code}_{assessment_key.upper()}_template.xlsx"
    from django.http import HttpResponse
    resp = HttpResponse(
        bio.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _parse_int(value):
    try:
        if value is None:
            return None
        s = str(value).strip()
        if not s:
            return None
        return int(s)
    except Exception:
        return None


def _get_query_params(request):
    return getattr(request, 'query_params', {}) if hasattr(request, 'query_params') else request.GET


def _get_teaching_assignment_id_from_request(request, body: dict | None = None) -> int | None:
    if body and isinstance(body, dict) and body.get('teaching_assignment_id') is not None:
        ta_id = _parse_int(body.get('teaching_assignment_id'))
        if ta_id is not None:
            return ta_id
    qp = _get_query_params(request)
    return _parse_int(qp.get('teaching_assignment_id'))


def _normalize_cqi_assessment_type(value) -> str | None:
    s = str(value or '').strip().lower()
    return s or None


def _normalize_cqi_co_numbers(value) -> list[int]:
    nums: list[int] = []

    def _push(raw):
        try:
            n = int(str(raw).strip())
        except Exception:
            return
        if 1 <= n <= 20 and n not in nums:
            nums.append(n)

    if isinstance(value, (list, tuple, set)):
        for item in value:
            if isinstance(item, str):
                parts = re.findall(r'\d+', item)
                if parts:
                    for part in parts:
                        _push(part)
                    continue
            _push(item)
    elif isinstance(value, str):
        for part in re.findall(r'\d+', value):
            _push(part)
    elif value is not None:
        _push(value)

    nums.sort()
    return nums


def _make_cqi_page_key(assessment_type: str | None, co_numbers: list[int], explicit_key=None) -> str | None:
    explicit = str(explicit_key or '').strip()
    if explicit:
        return explicit
    parts = []
    if assessment_type:
        parts.append(str(assessment_type).strip().lower())
    if co_numbers:
        parts.append(','.join(str(n) for n in co_numbers))
    return ':'.join(parts) if parts else None


def _resolve_cqi_page_context(request, body: dict | None = None) -> tuple[str | None, str | None, list[int]]:
    qp = _get_query_params(request)
    raw_page_key = None
    raw_assessment = None
    raw_co_numbers = None

    if isinstance(body, dict):
        raw_page_key = body.get('pageKey', body.get('page_key'))
        raw_assessment = body.get('assessmentType', body.get('assessment_type'))
        raw_co_numbers = body.get('coNumbers', body.get('co_numbers'))

    if raw_page_key in (None, ''):
        raw_page_key = qp.get('page_key') or qp.get('pageKey')
    if raw_assessment in (None, ''):
        raw_assessment = qp.get('assessment_type') or qp.get('assessmentType')
    if raw_co_numbers in (None, '', []):
        raw_co_numbers = qp.get('co_numbers') or qp.get('coNumbers')

    assessment_type = _normalize_cqi_assessment_type(raw_assessment)
    co_numbers = _normalize_cqi_co_numbers(raw_co_numbers)
    page_key = _make_cqi_page_key(assessment_type, co_numbers, raw_page_key)
    return page_key, assessment_type, co_numbers


def _split_cqi_entries_payload(raw_entries) -> tuple[dict, dict]:
    merged_entries: dict = {}
    pages: dict = {}

    if not isinstance(raw_entries, dict):
        return merged_entries, pages

    raw_pages = raw_entries.get('__pages')
    if isinstance(raw_pages, dict):
        pages = {str(k): v for k, v in raw_pages.items() if isinstance(v, dict)}

    for key, value in raw_entries.items():
        if str(key).startswith('__'):
            continue
        if isinstance(value, dict):
            merged_entries[str(key)] = value

    return merged_entries, pages


def _find_cqi_page_snapshot(pages: dict, page_key: str | None, assessment_type: str | None, co_numbers: list[int]):
    if not isinstance(pages, dict) or not pages:
        return None

    if page_key and isinstance(pages.get(page_key), dict):
        return pages.get(page_key)

    for snapshot in pages.values():
        if not isinstance(snapshot, dict):
            continue
        snap_assessment = _normalize_cqi_assessment_type(snapshot.get('assessmentType', snapshot.get('assessment_type')))
        snap_co_numbers = _normalize_cqi_co_numbers(snapshot.get('coNumbers', snapshot.get('co_numbers')))
        if assessment_type and snap_assessment and snap_assessment != assessment_type:
            continue
        if co_numbers and snap_co_numbers and snap_co_numbers == co_numbers:
            return snapshot

    for snapshot in pages.values():
        if not isinstance(snapshot, dict):
            continue
        snap_assessment = _normalize_cqi_assessment_type(snapshot.get('assessmentType', snapshot.get('assessment_type')))
        snap_co_numbers = _normalize_cqi_co_numbers(snapshot.get('coNumbers', snapshot.get('co_numbers')))
        if assessment_type and snap_assessment and snap_assessment != assessment_type:
            continue
        if co_numbers and snap_co_numbers and all(c in snap_co_numbers for c in co_numbers):
            return snapshot

    return None


def _extract_cqi_page_state(raw_entries, page_key: str | None, assessment_type: str | None, co_numbers: list[int], legacy_co_numbers=None):
    merged_entries, pages = _split_cqi_entries_payload(raw_entries)
    if pages:
        snapshot = _find_cqi_page_snapshot(pages, page_key, assessment_type, co_numbers)
        if not isinstance(snapshot, dict):
            return None
        entries = snapshot.get('entries') if isinstance(snapshot.get('entries'), dict) else {}
        return {
            'entries': entries,
            'co_numbers': _normalize_cqi_co_numbers(snapshot.get('coNumbers', snapshot.get('co_numbers'))),
            'assessment_type': _normalize_cqi_assessment_type(snapshot.get('assessmentType', snapshot.get('assessment_type'))),
            'updated_at': snapshot.get('updatedAt'),
            'updated_by': snapshot.get('updatedBy'),
            'published_at': snapshot.get('publishedAt'),
            'published_by': snapshot.get('publishedBy'),
        }

    if not isinstance(raw_entries, dict):
        return None

    legacy_nums = _normalize_cqi_co_numbers(legacy_co_numbers)
    if page_key and co_numbers and legacy_nums:
        if not all(c in legacy_nums for c in co_numbers):
            return None

    return {
        'entries': merged_entries,
        'co_numbers': legacy_nums,
        'assessment_type': assessment_type,
        'updated_at': None,
        'updated_by': None,
        'published_at': None,
        'published_by': None,
    }


def _merge_cqi_page_entries(pages: dict) -> tuple[dict, list[int]]:
    merged: dict = {}
    all_co_numbers: list[int] = []

    for snapshot in pages.values():
        if not isinstance(snapshot, dict):
            continue
        entries = snapshot.get('entries')
        if isinstance(entries, dict):
            for student_id, student_entries in entries.items():
                if not isinstance(student_entries, dict):
                    continue
                student_key = str(student_id)
                bucket = merged.setdefault(student_key, {})
                bucket.update(student_entries)
        for co_num in _normalize_cqi_co_numbers(snapshot.get('coNumbers', snapshot.get('co_numbers'))):
            if co_num not in all_co_numbers:
                all_co_numbers.append(co_num)

    all_co_numbers.sort()
    return merged, all_co_numbers


def _build_cqi_entries_payload(existing_entries, page_key: str | None, assessment_type: str | None, co_numbers: list[int], entries: dict, *, meta_kind: str, user_id=None, legacy_co_numbers=None):
    if not page_key:
        return entries or {}, co_numbers, None

    legacy_entries, pages = _split_cqi_entries_payload(existing_entries)
    if not pages and legacy_entries:
        legacy_page_key = _make_cqi_page_key(None, _normalize_cqi_co_numbers(legacy_co_numbers)) or page_key
        pages[legacy_page_key] = {
            'entries': legacy_entries,
            **({'coNumbers': _normalize_cqi_co_numbers(legacy_co_numbers)} if _normalize_cqi_co_numbers(legacy_co_numbers) else {}),
        }

    snapshot = dict(pages.get(page_key) or {})
    snapshot['entries'] = entries or {}
    if assessment_type:
        snapshot['assessmentType'] = assessment_type
    if co_numbers:
        snapshot['coNumbers'] = co_numbers

    now_iso = timezone.now().isoformat()
    if meta_kind == 'draft':
        snapshot['updatedAt'] = now_iso
        snapshot['updatedBy'] = user_id
    else:
        snapshot['publishedAt'] = now_iso
        snapshot['publishedBy'] = user_id

    pages[page_key] = snapshot
    merged_entries, merged_co_numbers = _merge_cqi_page_entries(pages)

    payload = {
        '__version': 2,
        '__pages': pages,
    }
    payload.update(merged_entries)
    return payload, merged_co_numbers or co_numbers, snapshot


def _resolve_section_name_from_ta(ta) -> str:
    if not ta:
        return ''
    sec = getattr(ta, 'section', None)
    if not sec:
        return ''
    return str(getattr(sec, 'name', None) or str(sec) or '').strip()


def _resolve_semester_from_ta(ta):
    """Resolve Semester FK from a TeachingAssignment.

    Some datasets (often older/migrated) have Section.semester unset even though
    the assignment clearly belongs to a semester via curriculum/subject.
    Due schedules and global publish controls are semester-scoped, so we attempt
    multiple fallbacks.
    """
    if not ta:
        return None

    # Primary: section.semester
    try:
        sem = getattr(getattr(ta, 'section', None), 'semester', None)
        if sem is not None:
            return sem
    except Exception:
        pass

    # Last resort: compute from section.batch.start_year + academic_year parity
    # (mirrors academics.models.Section.save()). This is needed for older/migrated
    # sections where `semester` FK was left null.
    try:
        sec = getattr(ta, 'section', None)
        batch = getattr(sec, 'batch', None) if sec else None
        ay = getattr(ta, 'academic_year', None)
        if batch is not None:
            start_year = getattr(batch, 'start_year', None)
            if start_year is None:
                try:
                    start_year = int(str(getattr(batch, 'name', '')).split('-')[0])
                except Exception:
                    start_year = None

            acad_start = None
            if ay is not None:
                try:
                    acad_start = int(str(getattr(ay, 'name', '')).split('-')[0])
                except Exception:
                    acad_start = None

            if start_year is not None and acad_start is not None:
                delta = int(acad_start) - int(start_year)
                parity = str(getattr(ay, 'parity', '') or '').upper()
                offset = 1 if parity == 'ODD' else 2
                sem_number = int(delta) * 2 + int(offset)
                if sem_number >= 1 and sem_number <= 20:
                    from academics.models import Semester
                    sem_obj, _ = Semester.objects.get_or_create(number=sem_number)
                    return sem_obj
    except Exception:
        pass

    # Fallbacks: curriculum row / elective / subject
    try:
        sem = getattr(getattr(ta, 'curriculum_row', None), 'semester', None)
        if sem is not None:
            return sem
    except Exception:
        pass
    try:
        row = getattr(ta, 'curriculum_row', None)
        sem = getattr(getattr(row, 'master', None), 'semester', None)
        if sem is not None:
            return sem
    except Exception:
        pass
    try:
        sem = getattr(getattr(ta, 'elective_subject', None), 'semester', None)
        if sem is not None:
            return sem
    except Exception:
        pass
    try:
        sem = getattr(getattr(ta, 'subject', None), 'semester', None)
        if sem is not None:
            return sem
    except Exception:
        pass

    return None


def _get_mark_table_lock_if_exists(*, staff_user, subject_code: str, assessment: str, teaching_assignment=None, academic_year=None, section_name: str = ''):
    from .models import ObeMarkTableLock

    if teaching_assignment is not None:
        lock = ObeMarkTableLock.objects.filter(teaching_assignment=teaching_assignment, assessment=str(assessment).lower()).first()
        if lock is not None:
            return lock

        # If the lock was created without a TA (e.g., UI published without passing
        # teaching_assignment_id), fall back to any matching unscoped lock.
        # Do NOT restrict by staff_user here: published state should be visible to
        # HOD/IQAC/master viewers too.
        try:
            fallback = (
                ObeMarkTableLock.objects.filter(
                    teaching_assignment__isnull=True,
                    subject_code=str(subject_code),
                    assessment=str(assessment).lower(),
                )
                .order_by('-updated_at')
                .first()
            )
            if fallback is not None:
                return fallback
        except Exception:
            pass

    qs = ObeMarkTableLock.objects.filter(
        teaching_assignment__isnull=True,
        staff_user=staff_user,
        subject_code=str(subject_code),
        assessment=str(assessment).lower(),
        section_name=str(section_name or ''),
    )
    if academic_year is not None:
        qs = qs.filter(academic_year=academic_year)
    return qs.order_by('-updated_at').first()


def _upsert_mark_table_lock(
    *,
    staff_user,
    subject_code: str,
    subject_name: str,
    assessment: str,
    teaching_assignment=None,
    academic_year=None,
    section_name: str = '',
    updated_by: int | None = None,
):
    from .models import ObeMarkTableLock

    assessment_key = str(assessment).lower()
    section_name = str(section_name or '').strip()
    subject_code = str(subject_code or '').strip()
    subject_name = str(subject_name or '').strip()

    defaults = {
        'staff_user': staff_user,
        'academic_year': academic_year,
        'subject_code': subject_code,
        'subject_name': subject_name,
        'section_name': section_name,
        'updated_by': updated_by,
    }

    if teaching_assignment is not None:
        obj, _created = ObeMarkTableLock.objects.update_or_create(
            teaching_assignment=teaching_assignment,
            assessment=assessment_key,
            defaults=defaults,
        )
        return obj

    obj, _created = ObeMarkTableLock.objects.update_or_create(
        teaching_assignment=None,
        staff_user=staff_user,
        subject_code=subject_code,
        assessment=assessment_key,
        section_name=section_name,
        academic_year=academic_year,
        defaults=defaults,
    )
    return obj


def _touch_lock_after_publish(request, *, subject_code: str, subject_name: str, assessment: str, teaching_assignment_id: int | None = None):
    from .models import ObeMarkTableLock

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject_code, teaching_assignment_id=teaching_assignment_id)
    academic_year = getattr(ta, 'academic_year', None) if ta else None
    section_name = _resolve_section_name_from_ta(ta)

    lock = _upsert_mark_table_lock(
        staff_user=getattr(request, 'user', None),
        subject_code=subject_code,
        subject_name=subject_name,
        assessment=assessment,
        teaching_assignment=ta,
        academic_year=academic_year,
        section_name=section_name,
        updated_by=getattr(getattr(request, 'user', None), 'id', None),
    )

    lock.is_published = True
    lock.mark_manager_locked = True
    lock.mark_entry_unblocked_until = None
    lock.mark_manager_unlocked_until = None
    lock.recompute_blocks()
    lock.save(
        update_fields=[
            'is_published',
            'published_blocked',
            'mark_entry_blocked',
            'mark_manager_locked',
            'mark_entry_unblocked_until',
            'mark_manager_unlocked_until',
            'updated_by',
            'updated_at',
        ]
    )
    return lock


def _enforce_mark_entry_not_blocked(
    request,
    *,
    subject_code: str,
    subject_name: str,
    assessment: str,
    teaching_assignment_id: int | None = None,
):
    """Block edits after publish unless an IQAC approval window is active."""
    # Check master config to see if publish lock is active globally
    master_cfg_qs = ObeAssessmentMasterConfig.objects.filter(id=1).first()
    master_cfg = master_cfg_qs.config if master_cfg_qs and getattr(master_cfg_qs, 'config', None) else {}
    if not master_cfg.get('edit_requests_enabled', True):
        return None

    user = getattr(request, 'user', None)
    if _has_obe_master_permission(user):
        return None

    try:
        ta = _resolve_staff_teaching_assignment(request, subject_code=subject_code, teaching_assignment_id=teaching_assignment_id)
        academic_year = getattr(ta, 'academic_year', None) if ta else None
        section_name = _resolve_section_name_from_ta(ta)
        lock = _get_mark_table_lock_if_exists(
            staff_user=user,
            subject_code=str(subject_code),
            assessment=str(assessment).lower(),
            teaching_assignment=ta,
            academic_year=academic_year,
            section_name=section_name,
        )
    except Exception:
        lock = None

    if lock is None:
        return None

    try:
        lock.recompute_blocks()
    except Exception:
        pass

    if bool(getattr(lock, 'is_published', False)) and bool(getattr(lock, 'mark_entry_blocked', False)):
        return Response(
            {
                'detail': 'Marks entry is locked after publish. Request IQAC approval to re-enter marks.',
                'assessment': str(assessment).lower(),
                'subject_code': str(subject_code),
                'how_to_fix': [
                    'Create an OBE Edit Request with scope=MARK_ENTRY (or MARK_MANAGER) for this assessment.',
                    'Wait for IQAC approval; then re-publish within the approved window.',
                ],
            },
            status=423,
        )

    return None


def _has_obe_master_access(user) -> bool:
    if user is None:
        return False
    if getattr(user, 'is_superuser', False) or getattr(user, 'is_staff', False):
        return True
    try:
        perms = get_user_permissions(user)
    except Exception:
        perms = set()
    return 'obe.master.manage' in {str(p).lower() for p in (perms or set())}


def _require_obe_master(request):
    if not _has_obe_master_access(getattr(request, 'user', None)):
        return Response({'detail': 'OBE Master access only.'}, status=status.HTTP_403_FORBIDDEN)
    return None


def _has_obe_master_permission(user) -> bool:
    if user is None:
        return False
    if getattr(user, 'is_superuser', False):
        return True

    def _normalized_names(items) -> set[str]:
        out: set[str] = set()
        for item in items or []:
            name = str(getattr(item, 'name', item) or '').strip().upper()
            if name:
                out.add(name)
        return out

    # IQAC/HAA users are treated as OBE masters for IQAC-managed configuration endpoints.
    # Note: this codebase uses BOTH Django auth groups (user.groups) AND custom roles (user.roles).
    # Allow either so an "IQAC role" user doesn't get blocked by missing Django group membership.
    if getattr(user, 'is_authenticated', False):
        try:
            if {'IQAC', 'HAA'} & _normalized_names(user.groups.all()):
                return True
        except Exception:
            pass
        try:
            if getattr(user, 'roles', None) is not None:
                if {'IQAC', 'HAA'} & _normalized_names(user.roles.all()):
                    return True
        except Exception:
            try:
                role_names = {str(r.name or '').strip().upper() for r in user.roles.all()}
                if 'IQAC' in role_names or 'HAA' in role_names:
                    return True
            except Exception:
                pass

    try:
        perms = get_user_permissions(user)
    except Exception:
        perms = set()
    return 'obe.master.manage' in {str(p).lower() for p in (perms or set())}


def _require_obe_master_permission(request):
    if not _has_obe_master_permission(getattr(request, 'user', None)):
        return Response({'detail': 'OBE Master permission required.'}, status=status.HTTP_403_FORBIDDEN)
    return None


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def iqac_sync_final_internal_marks(request):
    """Recompute and persist final internal marks (out of 40) for assigned subjects.

    Optional body filters:
      - teaching_assignment_id
      - subject_code
      - semester
    """
    denied = _require_obe_master_permission(request)
    if denied is not None:
        return denied

    data = request.data if isinstance(request.data, dict) else {}
    filters = {}

    if data.get('teaching_assignment_id') not in (None, ''):
        try:
            filters['teaching_assignment_id'] = int(data.get('teaching_assignment_id'))
        except Exception:
            return Response({'detail': 'Invalid teaching_assignment_id'}, status=status.HTTP_400_BAD_REQUEST)

    if data.get('subject_code') not in (None, ''):
        filters['subject_code'] = str(data.get('subject_code')).strip()

    if data.get('semester') not in (None, ''):
        try:
            filters['semester'] = int(data.get('semester'))
        except Exception:
            return Response({'detail': 'Invalid semester'}, status=status.HTTP_400_BAD_REQUEST)

    result = recompute_final_internal_marks(actor_user_id=getattr(request.user, 'id', None), filters=filters)
    return Response({'detail': 'Final internal marks synced successfully.', **result}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def final_internal_marks_by_student(request, student_id: int):
    """Return student-wise stored final internal marks grouped by assigned subject."""
    denied = _require_obe_master_permission(request)
    if denied is not None:
        return denied

    try:
        sid = int(student_id)
    except Exception:
        return Response({'detail': 'Invalid student_id'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import FinalInternalMark

    rows = (
        FinalInternalMark.objects.filter(student_id=sid)
        .select_related('subject', 'teaching_assignment', 'student', 'student__user')
        .order_by('subject__code', 'teaching_assignment_id')
    )

    out = []
    for r in rows:
        ta = getattr(r, 'teaching_assignment', None)
        sec = getattr(getattr(ta, 'section', None), 'name', None) if ta else None
        out.append(
            {
                'course_code': getattr(getattr(r, 'subject', None), 'code', None),
                'course_name': getattr(getattr(r, 'subject', None), 'name', None),
                'teaching_assignment_id': getattr(r, 'teaching_assignment_id', None),
                'section': sec,
                'final_internal_mark': float(r.final_mark) if getattr(r, 'final_mark', None) is not None else None,
                'max_mark': float(r.max_mark) if getattr(r, 'max_mark', None) is not None else None,
                'computed_at': getattr(r, 'computed_at', None),
            }
        )

    student_obj = rows.first().student if rows.exists() else None
    student_name = _student_display_name(getattr(student_obj, 'user', None)) if student_obj else None
    reg_no = getattr(student_obj, 'reg_no', None) if student_obj else None

    return Response(
        {
            'student': {'id': sid, 'reg_no': reg_no, 'name': student_name},
            'courses': out,
        },
        status=status.HTTP_200_OK,
    )


def _require_publish_owner(request):
    """If `settings.OBE_PUBLISH_ALLOWED_USERNAME` is set, only that user (username or email)
    or superusers may perform publish actions. Return a Response on denial or None to allow.
    """
    allowed = getattr(settings, 'OBE_PUBLISH_ALLOWED_USERNAME', None)
    if not allowed:
        return None
    user = getattr(request, 'user', None)
    if not user or not getattr(user, 'is_authenticated', False):
        return Response({'detail': 'Authentication required.'}, status=status.HTTP_403_FORBIDDEN)
    if getattr(user, 'is_superuser', False):
        return None
    try:
        allowed_norm = str(allowed or '').strip().lower()
        uname = str(getattr(user, 'username', '') or '').strip().lower()
        email = str(getattr(user, 'email', '') or '').strip().lower()
        if uname == allowed_norm or email == allowed_norm:
            return None
    except Exception:
        pass
    return Response({'detail': 'Only the designated user may publish marks.'}, status=status.HTTP_403_FORBIDDEN)


# ---------------------------------------------------------------------------
# Shared helper: canonical slot-length enforcement for internal_mark_weights
# ---------------------------------------------------------------------------
_EXPECTED_INTERNAL_WEIGHTS_SLOTS = {'TCPL': 21}  # all other class types: 17
_DEFAULT_INTERNAL_WEIGHTS_SLOTS = 17

_TCPL_DEFAULT_21 = [
    1.0, 3.25, 3.5, 0.0,  # CO1 SSA, CIA, LAB, CIA-Exam
    1.0, 3.25, 3.5, 0.0,  # CO2
    1.0, 3.25, 3.5, 0.0,  # CO3
    1.0, 3.25, 3.5, 0.0,  # CO4
    3.0, 3.0, 3.0, 3.0, 7.0,  # ME CO1-CO5
]
_THEORY_DEFAULT_17 = [1.5, 3.0, 2.5, 1.5, 3.0, 2.5, 1.5, 3.0, 2.5, 1.5, 3.0, 2.5, 2.0, 2.0, 2.0, 2.0, 4.0]


def _normalise_class_type_weights_array(class_type: str, arr):
    """Return a properly-sized list of floats (or a structured dict for LAB/PROJECT).

    * TCPL expects 21 slots – old 17-slot arrays are automatically upgraded by
      inserting a 0.0 CIA-Exam slot after every CO's LAB position.
    * LAB/PRACTICAL/PROJECT may use a structured dict format (type: lab_cycles / project_reviews / project_prbl)
      – these are passed through as-is.
    * All other class types expect 17 slots.
    * If ``arr`` is None / not a list/dict the canonical defaults are returned.
    """
    ct = str(class_type or '').strip().upper()
    expected = _EXPECTED_INTERNAL_WEIGHTS_SLOTS.get(ct, _DEFAULT_INTERNAL_WEIGHTS_SLOTS)
    defaults = _TCPL_DEFAULT_21 if ct == 'TCPL' else _THEORY_DEFAULT_17

    # Structured format for LAB/PRACTICAL/PROJECT/SPECIAL/ENGLISH/FOREIGN_LANG – pass through as-is
    if isinstance(arr, dict) and arr.get('type') in ('lab_cycles', 'project_reviews', 'project_prbl', 'special_exam_weights', 'english_exam_weights', 'foreign_lang_exam_weights'):
        return arr

    if not isinstance(arr, list):
        return list(defaults)

    if ct == 'TCPL':
        if len(arr) == 17:
            # Upgrade: insert CIA-Exam (0.0) slot after each CO's LAB value
            out = []
            for co in range(4):
                base = co * 3
                out.append(float(arr[base]) if base < len(arr) else 0.0)
                out.append(float(arr[base + 1]) if base + 1 < len(arr) else 0.0)
                out.append(float(arr[base + 2]) if base + 2 < len(arr) else 0.0)
                out.append(0.0)  # CIA-Exam – default 0 for upgraded rows
            for i in range(12, 17):
                out.append(float(arr[i]) if i < len(arr) else 0.0)
            return out  # always 21 elements
        # Already 21-slot (or wrong length): pad / truncate
        out = [float(x) for x in arr]
        while len(out) < 21:
            out.append(0.0)
        return out[:21]

    # Non-TCPL: 17 slots
    out = [float(x) for x in arr]
    while len(out) < expected:
        out.append(defaults[len(out)] if len(out) < len(defaults) else 0.0)
    return out[:expected]


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([AllowAny])
def class_type_weights_list(request):
    """Return current class-type weights as a mapping keyed by normalized class type."""
    try:
        from .models import ClassTypeWeights
        try:
            objs = ClassTypeWeights.objects.all().order_by('id')
        except Exception:
            objs = []
    except ImportError:
        # Model missing (e.g., migrations or model removed). Return empty mapping
        objs = []

    out = {}
    for o in objs:
        ct = str(o.class_type or '').strip().upper()
        if not ct:
            continue
        prev = out.get(ct)
        current_updated = getattr(o, 'updated_at', None)
        prev_updated = prev.get('updated_at_obj') if isinstance(prev, dict) else None
        if prev is not None and prev_updated is not None and current_updated is not None and current_updated <= prev_updated:
            continue
        out[ct] = {
            'ssa1': float(o.ssa1) if o.ssa1 is not None else None,
            'cia1': float(o.cia1) if o.cia1 is not None else None,
            'formative1': float(o.formative1) if o.formative1 is not None else None,
            # Return structured dicts as-is; arrays get canonically-sized (upgrades legacy 17-slot TCPL rows).
            'internal_mark_weights': (
                o.internal_mark_weights
                if isinstance(getattr(o, 'internal_mark_weights', None), dict)
                   and o.internal_mark_weights.get('type') in ('lab_cycles', 'project_reviews', 'project_prbl')
                else _normalise_class_type_weights_array(
                    ct, o.internal_mark_weights if isinstance(getattr(o, 'internal_mark_weights', None), list) else None
                )
            ),
            'updated_at': (o.updated_at.isoformat() if getattr(o, 'updated_at', None) else None),
            'updated_by': o.updated_by,
            'updated_at_obj': current_updated,
        }
    for v in out.values():
        if isinstance(v, dict):
            v.pop('updated_at_obj', None)
    resp = Response({'results': out})
    try:
        resp['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp['Pragma'] = 'no-cache'
        resp['Expires'] = '0'
    except Exception:
        pass
    return resp


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def iqac_cqi_get(request):
    """Return global CQI configuration (managed by IQAC)."""
    try:
        from .models import ObeCqiConfig
        cfg = ObeCqiConfig.objects.first()
    except Exception:
        cfg = None

    if not cfg:
        return Response({'options': [], 'divider': 2.0, 'multiplier': 0.15})

    return Response({
        'options': cfg.options or [],
        'divider': float(cfg.divider or 2.0),
        'multiplier': float(cfg.multiplier or 0.15),
        'updated_at': cfg.updated_at.isoformat() if getattr(cfg, 'updated_at', None) else None,
        'updated_by': cfg.updated_by,
    })


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def iqac_cqi_upsert(request):
    """Create or update global CQI configuration. Restricted to OBE master (IQAC/HAA) users."""
    auth = _require_obe_master_permission(request)
    if auth:
        return auth

    body = request.data or {}
    opts = body.get('options', [])
    try:
        d = float(body.get('divider', 2.0))
    except Exception:
        d = 2.0
    try:
        m = float(body.get('multiplier', 0.15))
    except Exception:
        m = 0.15

    try:
        from .models import ObeCqiConfig
        obj, _ = ObeCqiConfig.objects.get_or_create(id=1, defaults={'options': opts, 'divider': d, 'multiplier': m, 'updated_by': getattr(request.user, 'id', None)})
        obj.options = opts or []
        obj.divider = d
        obj.multiplier = m
        obj.updated_by = getattr(request.user, 'id', None)
        obj.save()
    except Exception as e:
        return Response({'detail': 'Failed to save CQI config', 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response({'status': 'ok'})


@api_view(['GET', 'PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cqi_draft(request, subject_id: str):
    """Get/Upsert CQI draft for a subject + teaching assignment."""
    _staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)
    ta_id = _get_teaching_assignment_id_from_request(request, request.data if request.method == 'PUT' else None)
    if ta_id is None:
        return Response({'detail': 'teaching_assignment_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
    if ta is None:
        return Response({'detail': 'Invalid teaching_assignment_id.'}, status=status.HTTP_400_BAD_REQUEST)

    body = request.data if isinstance(request.data, dict) else None
    page_key, assessment_type, requested_co_numbers = _resolve_cqi_page_context(request, body)

    from .models import ObeCqiDraft

    if request.method == 'GET':
        obj = ObeCqiDraft.objects.filter(subject=subject, teaching_assignment=ta).first()
        if obj is None:
            return Response({'draft': None})
        snapshot = _extract_cqi_page_state(obj.entries, page_key, assessment_type, requested_co_numbers)
        if snapshot is None:
            return Response({'draft': None})
        return Response(
            {
                'draft': {'entries': snapshot.get('entries') or {}},
                'updated_at': snapshot.get('updated_at') or (obj.updated_at.isoformat() if getattr(obj, 'updated_at', None) else None),
                'updated_by': snapshot.get('updated_by', getattr(obj, 'updated_by', None)),
            }
        )

    body = body or {}
    entries = body.get('entries', None)
    if entries is None:
        return Response({'detail': 'Missing entries.'}, status=status.HTTP_400_BAD_REQUEST)
    if not isinstance(entries, dict):
        return Response({'detail': 'entries must be an object.'}, status=status.HTTP_400_BAD_REQUEST)

    user_id = getattr(getattr(request, 'user', None), 'id', None)
    existing = ObeCqiDraft.objects.filter(subject=subject, teaching_assignment=ta).first()
    stored_entries, _merged_co_numbers, snapshot = _build_cqi_entries_payload(
        existing.entries if existing is not None else None,
        page_key,
        assessment_type,
        requested_co_numbers,
        entries,
        meta_kind='draft',
        user_id=user_id,
    )
    obj, _created = ObeCqiDraft.objects.update_or_create(
        subject=subject,
        teaching_assignment=ta,
        defaults={
            'entries': stored_entries,
            'updated_by': user_id,
        },
    )

    return Response(
        {
            'status': 'ok',
            'updated_at': (snapshot or {}).get('updatedAt') or (obj.updated_at.isoformat() if getattr(obj, 'updated_at', None) else None),
            'updated_by': (snapshot or {}).get('updatedBy', getattr(obj, 'updated_by', None)),
        }
    )


@api_view(['PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cqi_save(request, subject_id: str):
    """Alias for CQI draft save (kept for frontend compatibility)."""
    return cqi_draft(request, subject_id)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cqi_published(request, subject_id: str):
    """Fetch published CQI snapshot for Internal Marks."""
    _staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)
    ta_id = _get_teaching_assignment_id_from_request(request)
    if ta_id is None:
        return Response({'detail': 'teaching_assignment_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
    if ta is None:
        return Response({'detail': 'Invalid teaching_assignment_id.'}, status=status.HTTP_400_BAD_REQUEST)

    page_key, assessment_type, requested_co_numbers = _resolve_cqi_page_context(request)

    qp = _get_query_params(request)
    want_page_entries = str(qp.get('include_page_entries', '')).strip().lower() in ('1', 'true', 'yes')

    from .models import ObeCqiPublished

    obj = ObeCqiPublished.objects.filter(subject=subject, teaching_assignment=ta).first()
    if obj is None:
        return Response({'published': None})

    snapshot = _extract_cqi_page_state(obj.entries, page_key, assessment_type, requested_co_numbers, legacy_co_numbers=obj.co_numbers)

    def _build_pages_info(raw_entries, *, include_entries=False):
        """Return list of published page summaries from paged entries."""
        _, pgs = _split_cqi_entries_payload(raw_entries or {})
        out = []
        for pk, snap in pgs.items():
            if not isinstance(snap, dict) or not snap.get('publishedAt'):
                continue
            item = {
                'key': str(pk),
                'assessmentType': snap.get('assessmentType'),
                'coNumbers': _normalize_cqi_co_numbers(snap.get('coNumbers', snap.get('co_numbers'))),
                'publishedAt': snap.get('publishedAt'),
            }
            if include_entries:
                item['entries'] = snap.get('entries') if isinstance(snap.get('entries'), dict) else {}
            out.append(item)
        return out

    if snapshot is None:
        # When no specific page params requested and the record uses the new paged format,
        # merge entries across all published pages so the Internal Mark view can consume them.
        if not page_key and not assessment_type and not requested_co_numbers:
            _, pages = _split_cqi_entries_payload(obj.entries or {})
            if pages:
                all_merged, merged_co_nums = _merge_cqi_page_entries(pages)
                pub_dates = [s.get('publishedAt', '') for s in pages.values() if isinstance(s, dict) and s.get('publishedAt')]
                latest_pub = max(pub_dates) if pub_dates else (obj.published_at.isoformat() if getattr(obj, 'published_at', None) else None)
                return Response({'published': {
                    'publishedAt': latest_pub,
                    'coNumbers': merged_co_nums,
                    'entries': all_merged,
                    'pages': _build_pages_info(obj.entries, include_entries=want_page_entries),
                }})
        return Response({'published': None})

    return Response(
        {
            'published': {
                'publishedAt': snapshot.get('published_at') or (obj.published_at.isoformat() if getattr(obj, 'published_at', None) else None),
                'coNumbers': snapshot.get('co_numbers') or obj.co_numbers or [],
                'entries': snapshot.get('entries') or {},
                'pages': _build_pages_info(obj.entries, include_entries=want_page_entries),
            }
        }
    )


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cqi_publish(request, subject_id: str):
    """Publish CQI snapshot to DB for a subject + teaching assignment."""
    _staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)
    ta_id = _get_teaching_assignment_id_from_request(request, request.data if isinstance(request.data, dict) else None)
    if ta_id is None:
        return Response({'detail': 'teaching_assignment_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
    if ta is None:
        return Response({'detail': 'Invalid teaching_assignment_id.'}, status=status.HTTP_400_BAD_REQUEST)

    body = request.data if isinstance(request.data, dict) else {}
    page_key, assessment_type, requested_co_numbers = _resolve_cqi_page_context(request, body)
    entries = body.get('entries')
    co_numbers = body.get('coNumbers', body.get('co_numbers'))
    nums = _normalize_cqi_co_numbers(co_numbers)
    if not nums:
        nums = requested_co_numbers

    # Backward compatibility: if frontend doesn't send entries, publish latest draft.
    if entries is None:
        try:
            from .models import ObeCqiDraft

            d = ObeCqiDraft.objects.filter(subject=subject, teaching_assignment=ta).first()
            if d is not None:
                if page_key:
                    snapshot = _extract_cqi_page_state(d.entries, page_key, assessment_type, nums)
                    entries = snapshot.get('entries') if snapshot else None
                    if not nums and snapshot:
                        nums = snapshot.get('co_numbers') or []
                else:
                    entries = d.entries
            else:
                entries = None
        except Exception:
            entries = None

    if entries is None:
        return Response({'detail': 'Missing entries (send in body or save a draft first).'}, status=status.HTTP_400_BAD_REQUEST)
    if not isinstance(entries, dict):
        return Response({'detail': 'entries must be an object.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import ObeCqiPublished

    user_id = getattr(getattr(request, 'user', None), 'id', None)
    existing = ObeCqiPublished.objects.filter(subject=subject, teaching_assignment=ta).first()
    stored_entries, merged_nums, snapshot = _build_cqi_entries_payload(
        existing.entries if existing is not None else None,
        page_key,
        assessment_type,
        nums,
        entries,
        meta_kind='published',
        user_id=user_id,
        legacy_co_numbers=existing.co_numbers if existing is not None else None,
    )
    obj, _created = ObeCqiPublished.objects.update_or_create(
        subject=subject,
        teaching_assignment=ta,
        defaults={
            'co_numbers': merged_nums,
            'entries': stored_entries,
            'published_by': user_id,
        },
    )

    try:
        _touch_lock_after_publish(
            request,
            subject_code=subject.code,
            subject_name=str(getattr(subject, 'name', '') or subject.code),
            assessment=_build_cqi_assessment_key(page_key=page_key, assessment_type=assessment_type, co_numbers=nums or requested_co_numbers),
            teaching_assignment_id=ta_id,
        )
    except Exception:
        pass

    try:
        recompute_final_internal_marks(
            actor_user_id=getattr(getattr(request, 'user', None), 'id', None),
            filters={
                'subject_code': subject.code,
                'teaching_assignment_id': ta_id,
            },
        )
    except Exception:
        logger.exception('cqi_publish: recompute_final_internal_marks failed for subject=%s ta=%s', subject.code, ta_id)

    return Response(
        {
            'status': 'ok',
            'published_at': (snapshot or {}).get('publishedAt') or (obj.published_at.isoformat() if getattr(obj, 'published_at', None) else None),
            'published_by': (snapshot or {}).get('publishedBy', getattr(obj, 'published_by', None)),
        }
    )


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def class_type_weights_upsert(request):
    """Upsert multiple class-type weights. Requires OBE master permission (IQAC).

    Body: { CLASS_TYPE: { ssa1: number, cia1: number, formative1: number, internal_mark_weights?: number[] }, ... }
    """
    auth = _require_obe_master_permission(request)
    if auth:
        return auth

    data = request.data if isinstance(request.data, dict) else {}
    if not isinstance(data, dict):
        return Response({'detail': 'Invalid payload'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import ClassTypeWeights

    user_id = getattr(getattr(request, 'user', None), 'id', None)
    out = {}
    with transaction.atomic():
        for k, v in data.items():
            try:
                ct = str(k or '').strip().upper()
                ssa = float(v.get('ssa1')) if v and v.get('ssa1') is not None else None
                cia = float(v.get('cia1')) if v and v.get('cia1') is not None else None
                f1 = float(v.get('formative1')) if v and v.get('formative1') is not None else None
            except Exception:
                continue

            im = None
            try:
                im_raw = v.get('internal_mark_weights') if isinstance(v, dict) else None
                # Structured format (dict) for LAB/PRACTICAL/PROJECT/SPECIAL/ENGLISH/FOREIGN_LANG – store as-is
                if isinstance(im_raw, dict) and im_raw.get('type') in ('lab_cycles', 'project_reviews', 'project_prbl', 'special_exam_weights', 'english_exam_weights', 'foreign_lang_exam_weights'):
                    im = im_raw
                elif isinstance(im_raw, list):
                    im = []
                    for x in im_raw:
                        try:
                            im.append(float(x))
                        except Exception:
                            im.append(0)
            except Exception:
                im = None

            if not ct:
                continue

            # Server-side slot normalisation: TCPL must always be 21 slots, others 17.
            if im is not None:
                im = _normalise_class_type_weights_array(ct, im)

            existing = None
            duplicates = []
            try:
                duplicates = list(ClassTypeWeights.objects.filter(class_type__iexact=ct).order_by('-updated_at', '-id'))
                existing = duplicates[0] if duplicates else None
            except Exception:
                existing = None
                duplicates = []
            existing_im = getattr(existing, 'internal_mark_weights', None) if existing is not None else None
            if not isinstance(existing_im, list):
                existing_im = []

            if existing is not None:
                obj = existing
                obj.class_type = ct
                obj.ssa1 = ssa if ssa is not None else 0
                obj.cia1 = cia if cia is not None else 0
                obj.formative1 = f1 if f1 is not None else 0
                obj.internal_mark_weights = im if im is not None else existing_im
                obj.updated_by = user_id
                obj.save()
                extra_ids = [d.id for d in duplicates[1:] if getattr(d, 'id', None) != getattr(obj, 'id', None)]
                if extra_ids:
                    ClassTypeWeights.objects.filter(id__in=extra_ids).delete()
            else:
                obj = ClassTypeWeights.objects.create(
                    class_type=ct,
                    ssa1=ssa if ssa is not None else 0,
                    cia1=cia if cia is not None else 0,
                    formative1=f1 if f1 is not None else 0,
                    internal_mark_weights=im if im is not None else existing_im,
                    updated_by=user_id,
                )
            im_val = getattr(obj, 'internal_mark_weights', None)
            # Return structured dicts as-is; arrays as lists; fallback to empty list
            if isinstance(im_val, dict) and im_val.get('type') in ('lab_cycles', 'project_reviews', 'project_prbl'):
                im_out = im_val
            elif isinstance(im_val, list):
                im_out = im_val
            else:
                im_out = []
            out[ct] = {
                'ssa1': float(obj.ssa1),
                'cia1': float(obj.cia1),
                'formative1': float(obj.formative1),
                'internal_mark_weights': im_out,
            }

    return Response({'results': out})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def qp_pattern_get(request):
    """Fetch the configured QP pattern for a key.

    Query params:
      class_type=THEORY|TCPR|TCPL|LAB|...
      question_paper_type=QP1|QP2 (optional)
    exam=CIA|CIA1|CIA2|MODEL

        Returns:
            { class_type, question_paper_type, exam, pattern: { marks: number[], cos?: (number|string)[] }, updated_at, updated_by }
    """
    from .models import ObeQpPatternConfig

    qp = _get_query_params(request)
    class_type = str(qp.get('class_type') or '').strip().upper()
    question_paper_type = _normalize_qp_type_key(qp.get('question_paper_type'))
    exam = str(qp.get('exam') or '').strip().upper()

    if not class_type:
        return Response({'detail': 'class_type is required'}, status=status.HTTP_400_BAD_REQUEST)
    if exam not in {'CIA', 'CIA1', 'CIA2', 'MODEL', 'SSA1', 'SSA2', 'FORMATIVE1', 'FORMATIVE2'}:
        return Response({'detail': 'exam must be CIA, CIA1, CIA2, MODEL, SSA1, SSA2, FORMATIVE1, or FORMATIVE2'}, status=status.HTTP_400_BAD_REQUEST)

    qp_type_val = question_paper_type if question_paper_type else None

    obj = ObeQpPatternConfig.objects.filter(class_type=class_type, question_paper_type=qp_type_val, exam=exam).first()
    # Backward-compat: CIA1/CIA2 fall back to legacy CIA config if not present.
    if obj is None and exam in {'CIA1', 'CIA2'}:
        obj = ObeQpPatternConfig.objects.filter(class_type=class_type, question_paper_type=qp_type_val, exam='CIA').first()
    pattern = {'marks': []}
    updated_at = None
    updated_by = None
    if obj is not None:
        raw = getattr(obj, 'pattern', None)
        if isinstance(raw, dict):
            # New shape: { marks: [...], cos: [...] }
            pattern = raw
        elif isinstance(raw, list):
            # Legacy shape: [marks...]
            pattern = {'marks': raw}
        else:
            pattern = {'marks': []}
        updated_at = obj.updated_at.isoformat() if getattr(obj, 'updated_at', None) else None
        updated_by = obj.updated_by

    return Response({
        'class_type': class_type,
        'question_paper_type': qp_type_val,
        'exam': exam,
        'pattern': pattern,
        'updated_at': updated_at,
        'updated_by': updated_by,
    })


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def qp_pattern_upsert(request):
    """Upsert QP pattern. IQAC/OBE master only.

        Body:
            { class_type: string, question_paper_type?: 'QP1'|'QP2'|null, exam: 'CIA'|'CIA1'|'CIA2'|'MODEL', pattern: number[] | { marks: number[], cos?: (number|string)[] } }
    """
    auth = _require_obe_master_permission(request)
    if auth:
        return auth

    from .models import ObeQpPatternConfig

    data = request.data if isinstance(request.data, dict) else {}
    class_type = str(data.get('class_type') or '').strip().upper()
    question_paper_type = _normalize_qp_type_key(data.get('question_paper_type'))
    exam = str(data.get('exam') or '').strip().upper()
    pattern_raw = data.get('pattern')

    if not class_type:
        return Response({'detail': 'class_type is required'}, status=status.HTTP_400_BAD_REQUEST)
    if exam not in {'CIA', 'CIA1', 'CIA2', 'MODEL', 'SSA1', 'SSA2', 'FORMATIVE1', 'FORMATIVE2'}:
        return Response({'detail': 'exam must be CIA, CIA1, CIA2, MODEL, SSA1, SSA2, FORMATIVE1, or FORMATIVE2'}, status=status.HTTP_400_BAD_REQUEST)

    qp_type_val = question_paper_type if question_paper_type else None

    # Accept legacy list-only shape OR new object shape with CO mapping.
    marks_raw = None
    cos_raw = None
    if isinstance(pattern_raw, list):
        marks_raw = pattern_raw
    elif isinstance(pattern_raw, dict):
        marks_raw = pattern_raw.get('marks')
        cos_raw = pattern_raw.get('cos')
    else:
        return Response({'detail': 'pattern must be a list or an object with marks'}, status=status.HTTP_400_BAD_REQUEST)

    if not isinstance(marks_raw, list):
        return Response({'detail': 'pattern.marks must be a list'}, status=status.HTTP_400_BAD_REQUEST)

    marks: list[float] = []
    for x in marks_raw:
        try:
            v = float(x)
            if v < 0:
                return Response({'detail': 'pattern marks must be non-negative'}, status=status.HTTP_400_BAD_REQUEST)
            marks.append(v)
        except Exception:
            return Response({'detail': 'pattern marks must be numbers'}, status=status.HTTP_400_BAD_REQUEST)

    cos = None
    if cos_raw is not None:
        if not isinstance(cos_raw, list):
            return Response({'detail': 'pattern.cos must be a list if provided'}, status=status.HTTP_400_BAD_REQUEST)
        if len(cos_raw) != len(marks):
            return Response({'detail': 'pattern.cos length must match pattern.marks length'}, status=status.HTTP_400_BAD_REQUEST)
        cleaned_cos = []
        for x in cos_raw:
            # Allow numbers (1..n), or a small set of split markers.
            if isinstance(x, str):
                s = x.strip()
                if s in {'both', '1&2', '3&4', '1&2&3&4&5'}:
                    cleaned_cos.append(s)
                    continue
                try:
                    x = int(float(s))
                except Exception:
                    return Response({'detail': 'pattern.cos values must be numbers or split strings'}, status=status.HTTP_400_BAD_REQUEST)
            if isinstance(x, (int, float)):
                try:
                    n = int(x)
                except Exception:
                    return Response({'detail': 'pattern.cos values must be numbers or split strings'}, status=status.HTTP_400_BAD_REQUEST)
                if n <= 0:
                    return Response({'detail': 'pattern.cos values must be positive'}, status=status.HTTP_400_BAD_REQUEST)
                cleaned_cos.append(n)
                continue
            return Response({'detail': 'pattern.cos values must be numbers or split strings'}, status=status.HTTP_400_BAD_REQUEST)
        cos = cleaned_cos

    user_id = getattr(getattr(request, 'user', None), 'id', None)
    stored_pattern = {'marks': marks}
    if cos is not None:
        stored_pattern['cos'] = cos

    obj, _created = ObeQpPatternConfig.objects.update_or_create(
        class_type=class_type,
        question_paper_type=qp_type_val,
        exam=exam,
        defaults={'pattern': stored_pattern, 'updated_by': user_id},
    )

    return Response({
        'status': 'ok',
        'class_type': class_type,
        'question_paper_type': qp_type_val,
        'exam': exam,
        'pattern': stored_pattern,
        'updated_at': obj.updated_at.isoformat() if getattr(obj, 'updated_at', None) else None,
        'updated_by': obj.updated_by,
    })


def _normalize_exam_key(s) -> str:
    e = str(s or '').strip().upper()
    return e


def _normalize_qp_type_key(value) -> str:
    qp = str(value or '').strip().upper().replace(' ', '')
    return qp if qp in {'QP1', 'QP2', 'CSD', 'QP1FINAL'} else ''


def _validate_qp_pattern_payload(pattern_raw):
    """Shared validator for QP pattern payloads.

    Accepts:
      - list[marks]
      - { marks: [...], cos?: [...] }
    Returns:
      (pattern_dict, error_response)
    """
    marks_raw = None
    cos_raw = None
    if isinstance(pattern_raw, list):
        marks_raw = pattern_raw
    elif isinstance(pattern_raw, dict):
        marks_raw = pattern_raw.get('marks')
        cos_raw = pattern_raw.get('cos')
    else:
        return None, Response({'detail': 'pattern must be a list or an object with marks'}, status=status.HTTP_400_BAD_REQUEST)

    if not isinstance(marks_raw, list):
        return None, Response({'detail': 'pattern.marks must be a list'}, status=status.HTTP_400_BAD_REQUEST)

    marks: list[float] = []
    for x in marks_raw:
        try:
            v = float(x)
            if v < 0:
                return None, Response({'detail': 'pattern marks must be non-negative'}, status=status.HTTP_400_BAD_REQUEST)
            marks.append(v)
        except Exception:
            return None, Response({'detail': 'pattern marks must be numbers'}, status=status.HTTP_400_BAD_REQUEST)

    cos = None
    if cos_raw is not None:
        if not isinstance(cos_raw, list):
            return None, Response({'detail': 'pattern.cos must be a list if provided'}, status=status.HTTP_400_BAD_REQUEST)
        if len(cos_raw) != len(marks):
            return None, Response({'detail': 'pattern.cos length must match pattern.marks length'}, status=status.HTTP_400_BAD_REQUEST)
        cleaned_cos = []
        for x in cos_raw:
            if isinstance(x, str):
                s = x.strip()
                if s in {'both', '1&2', '3&4', '1&2&3&4&5'}:
                    cleaned_cos.append(s)
                    continue
                try:
                    x = int(float(s))
                except Exception:
                    return None, Response({'detail': 'pattern.cos must contain numbers or split markers'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                n = int(x)
            except Exception:
                return None, Response({'detail': 'pattern.cos must contain numbers or split markers'}, status=status.HTTP_400_BAD_REQUEST)
            if n <= 0:
                return None, Response({'detail': 'pattern.cos values must be positive integers'}, status=status.HTTP_400_BAD_REQUEST)
            cleaned_cos.append(n)
        cos = cleaned_cos

    out = {'marks': marks}
    if cos is not None:
        out['cos'] = cos
    return out, None


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def iqac_custom_exam_batches(request):
    auth = _require_permissions(request, {'obe.view'})
    if auth:
        return auth

    from academics.models import Batch

    # Return cohort batches directly from DB, de-duplicated by their cohort range.
    # Some databases can contain multiple Batch rows that display the same cohort
    # (e.g., repeated 2024-2028). The UI only needs a single entry per cohort.
    qs = (
        Batch.objects.filter(is_active=True)
        .values('id', 'name', 'start_year', 'end_year')
        .order_by('-start_year', '-end_year', 'name', '-id')
    )

    out = []
    seen_labels = set()
    for row in qs[:2000]:
        sy = row.get('start_year')
        ey = row.get('end_year')

        label = (row.get('name') or '').strip()
        if sy and ey:
            label = f"{sy}-{ey}"
        elif sy and not ey:
            try:
                label = f"{int(sy)}-{int(sy) + 4}"
            except Exception:
                pass

        label = label.strip() or str(row.get('id') or '')

        if label in seen_labels:
            continue
        seen_labels.add(label)

        out.append({
            'id': row['id'],
            'label': label,
            'start_year': sy,
            'end_year': ey,
        })

    return Response({'batches': out})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def iqac_batch_qp_pattern_get(request):
    from .models import ObeBatchQpPatternOverride, ObeQpPatternConfig

    qp = _get_query_params(request)
    batch_id = _parse_int(qp.get('batch_id'))
    class_type = str(qp.get('class_type') or '').strip().upper()
    question_paper_type = _normalize_qp_type_key(qp.get('question_paper_type'))
    exam = _normalize_exam_key(qp.get('exam'))

    if not batch_id:
        return Response({'detail': 'batch_id is required'}, status=status.HTTP_400_BAD_REQUEST)
    if not class_type:
        return Response({'detail': 'class_type is required'}, status=status.HTTP_400_BAD_REQUEST)
    if not exam:
        return Response({'detail': 'exam is required'}, status=status.HTTP_400_BAD_REQUEST)

    qp_type_val = question_paper_type if question_paper_type else None

    obj = ObeBatchQpPatternOverride.objects.filter(batch_id=batch_id, class_type=class_type, question_paper_type=qp_type_val, exam=exam).first()
    if obj is not None:
        pattern = getattr(obj, 'pattern', None)
        if not isinstance(pattern, dict):
            pattern = {'marks': []}
        return Response({
            'batch_id': batch_id,
            'class_type': class_type,
            'question_paper_type': qp_type_val,
            'exam': exam,
            'pattern': pattern,
            'is_override': True,
            'updated_at': obj.updated_at.isoformat() if getattr(obj, 'updated_at', None) else None,
            'updated_by': obj.updated_by,
        })

    # Fallback: reuse global IQAC qp-pattern only for supported global exam keys.
    fallback_pattern = {'marks': []}
    fallback_updated_at = None
    fallback_updated_by = None
    if exam in {'CIA', 'CIA1', 'CIA2', 'MODEL'}:
        g = ObeQpPatternConfig.objects.filter(class_type=class_type, question_paper_type=qp_type_val, exam=exam).first()
        if g is None and exam in {'CIA1', 'CIA2'}:
            g = ObeQpPatternConfig.objects.filter(class_type=class_type, question_paper_type=qp_type_val, exam='CIA').first()
        if g is not None:
            raw = getattr(g, 'pattern', None)
            if isinstance(raw, dict):
                fallback_pattern = raw
            elif isinstance(raw, list):
                fallback_pattern = {'marks': raw}
            fallback_updated_at = g.updated_at.isoformat() if getattr(g, 'updated_at', None) else None
            fallback_updated_by = g.updated_by

    return Response({
        'batch_id': batch_id,
        'class_type': class_type,
        'question_paper_type': qp_type_val,
        'exam': exam,
        'pattern': fallback_pattern,
        'is_override': False,
        'updated_at': fallback_updated_at,
        'updated_by': fallback_updated_by,
    })


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def iqac_batch_qp_pattern_upsert(request):
    auth = _require_obe_master_permission(request)
    if auth:
        return auth

    from .models import ObeBatchQpPatternOverride
    from academics.models import Batch

    data = request.data if isinstance(request.data, dict) else {}
    batch_id = _parse_int(data.get('batch_id'))
    class_type = str(data.get('class_type') or '').strip().upper()
    question_paper_type = _normalize_qp_type_key(data.get('question_paper_type'))
    exam = _normalize_exam_key(data.get('exam'))
    pattern_raw = data.get('pattern')

    if not batch_id:
        return Response({'detail': 'batch_id is required'}, status=status.HTTP_400_BAD_REQUEST)
    if not class_type:
        return Response({'detail': 'class_type is required'}, status=status.HTTP_400_BAD_REQUEST)
    if not exam:
        return Response({'detail': 'exam is required'}, status=status.HTTP_400_BAD_REQUEST)

    qp_type_val = question_paper_type if question_paper_type else None

    if not Batch.objects.filter(id=batch_id).exists():
        return Response({'detail': 'batch not found'}, status=status.HTTP_404_NOT_FOUND)

    pattern, err = _validate_qp_pattern_payload(pattern_raw)
    if err is not None:
        return err

    obj, _ = ObeBatchQpPatternOverride.objects.update_or_create(
        batch_id=batch_id,
        class_type=class_type,
        question_paper_type=qp_type_val,
        exam=exam,
        defaults={'pattern': pattern, 'updated_by': getattr(request.user, 'id', None)},
    )

    return Response({
        'status': 'ok',
        'batch_id': batch_id,
        'class_type': class_type,
        'question_paper_type': qp_type_val,
        'exam': exam,
        'pattern': obj.pattern if isinstance(getattr(obj, 'pattern', None), dict) else {'marks': []},
        'is_override': True,
        'updated_at': obj.updated_at.isoformat() if getattr(obj, 'updated_at', None) else None,
        'updated_by': obj.updated_by,
    })


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def internal_mark_mapping_get(request, subject_id: str):
    """Get IQAC-managed internal mark mapping for a subject.

    Returns:
      { subject: { code, name }, mapping: object|null, updated_at, updated_by }
    """
    from .models import InternalMarkMapping

    subject = _get_subject(subject_id, request)
    obj = None
    try:
        obj = InternalMarkMapping.objects.select_related('subject').filter(subject=subject).first()
    except Exception:
        obj = None

    return Response({
        'subject': {'code': getattr(subject, 'code', str(subject_id)), 'name': getattr(subject, 'name', '')},
        'mapping': (obj.mapping if obj else None),
        'updated_at': (obj.updated_at.isoformat() if obj and getattr(obj, 'updated_at', None) else None),
        'updated_by': (obj.updated_by if obj else None),
    })


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def internal_mark_mapping_upsert(request, subject_id: str):
    """Upsert internal mark mapping for a subject. Requires OBE master permission (IQAC).

    Body: { mapping: object }
    """
    auth = _require_obe_master_permission(request)
    if auth:
        return auth

    payload = request.data if isinstance(request.data, dict) else {}
    mapping = payload.get('mapping') if isinstance(payload, dict) else None
    if mapping is None:
        return Response({'detail': 'mapping is required'}, status=status.HTTP_400_BAD_REQUEST)
    if not isinstance(mapping, dict):
        return Response({'detail': 'mapping must be an object'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import InternalMarkMapping

    subject = _get_subject(subject_id, request)
    user_id = getattr(getattr(request, 'user', None), 'id', None)

    with transaction.atomic():
        obj, _created = InternalMarkMapping.objects.update_or_create(
            subject=subject,
            defaults={'mapping': mapping, 'updated_by': user_id},
        )

    return Response({
        'status': 'ok',
        'subject': {'code': getattr(subject, 'code', str(subject_id)), 'name': getattr(subject, 'name', '')},
        'mapping': obj.mapping,
        'updated_at': (obj.updated_at.isoformat() if getattr(obj, 'updated_at', None) else None),
        'updated_by': obj.updated_by,
    })


# ---------------------------------------------------------------------------
# Template Apply — helpers
# ---------------------------------------------------------------------------

_ALLOWED_CO_STRINGS = {'both', '1&2', '3&4', '1&2&3&4&5'}
_MAX_DECIMAL_PLACES = 2


def _decimal_places(v: float) -> int:
    s = f'{v:.10f}'.rstrip('0')
    if '.' in s:
        return len(s.split('.')[1])
    return 0


def _validate_internal_mark_mapping_for_apply(mapping: dict, class_type: str) -> list:
    errors = []
    if not isinstance(mapping, dict):
        return ['internal_mark_mapping must be an object']
    weights = mapping.get('weights')
    if not isinstance(weights, list):
        errors.append('internal_mark_mapping.weights must be an array')
        return errors
    ct = str(class_type or '').strip().upper()
    expected = _EXPECTED_INTERNAL_WEIGHTS_SLOTS.get(ct, _DEFAULT_INTERNAL_WEIGHTS_SLOTS)
    if len(weights) != expected:
        errors.append(
            f'internal_mark_mapping.weights has {len(weights)} slots; '
            f'{ct} requires exactly {expected}'
        )
    for i, w in enumerate(weights):
        try:
            fv = float(w)
        except (TypeError, ValueError):
            errors.append(f'weights[{i}] is not a number')
            continue
        if fv < 0:
            errors.append(f'weights[{i}] is negative ({fv})')
        if _decimal_places(fv) > _MAX_DECIMAL_PLACES:
            errors.append(f'weights[{i}] has more than {_MAX_DECIMAL_PLACES} decimal places')
    if not errors and sum(float(w) for w in weights) <= 0:
        errors.append('internal_mark_mapping.weights must sum to a positive value')
    return errors


def _validate_qp_pattern_for_apply(pattern: dict) -> list:
    errors = []
    if not isinstance(pattern, dict):
        return ['pattern must be an object']
    marks_raw = pattern.get('marks')
    if not isinstance(marks_raw, list) or not marks_raw:
        return ['pattern.marks must be a non-empty array']
    marks = []
    for i, m in enumerate(marks_raw):
        try:
            fv = float(m)
        except (TypeError, ValueError):
            errors.append(f'pattern.marks[{i}] is not a number')
            continue
        if fv < 0:
            errors.append(f'pattern.marks[{i}] is negative ({fv})')
        if _decimal_places(fv) > _MAX_DECIMAL_PLACES:
            errors.append(f'pattern.marks[{i}] has more than {_MAX_DECIMAL_PLACES} decimal places')
        marks.append(fv)
    cos_raw = pattern.get('cos')
    if cos_raw is not None:
        if not isinstance(cos_raw, list):
            errors.append('pattern.cos must be an array when provided')
        elif len(cos_raw) != len(marks_raw):
            errors.append(
                f'pattern.cos length ({len(cos_raw)}) must match '
                f'pattern.marks length ({len(marks_raw)})'
            )
        else:
            for i, c in enumerate(cos_raw):
                if isinstance(c, str) and c in _ALLOWED_CO_STRINGS:
                    continue
                try:
                    iv = int(c)
                    if iv <= 0:
                        raise ValueError
                except (TypeError, ValueError):
                    errors.append(
                        f'pattern.cos[{i}] must be a positive integer or one of '
                        f'{sorted(_ALLOWED_CO_STRINGS)}'
                    )
    return errors


def _build_config_snapshot(subject, class_type: str, relevant_exams: list) -> dict:
    """Return current DB state for snapshot recording (non-locking read)."""
    from .models import InternalMarkMapping, ObeQpPatternConfig
    try:
        imm = InternalMarkMapping.objects.filter(subject=subject).first()
        mapping_val = imm.mapping if imm else None
    except Exception:
        mapping_val = None
    ct = str(class_type or '').strip().upper()
    qp_rows = []
    for exam, qp_type in relevant_exams:
        try:
            row = ObeQpPatternConfig.objects.filter(
                class_type__iexact=ct,
                question_paper_type=qp_type,
                exam__iexact=exam,
            ).first()
            qp_rows.append({
                'exam': exam,
                'question_paper_type': qp_type,
                'pattern': row.pattern if row else None,
            })
        except Exception:
            qp_rows.append({'exam': exam, 'question_paper_type': qp_type, 'pattern': None})
    return {'internal_mark_mapping': mapping_val, 'qp_patterns': qp_rows}


# ---------------------------------------------------------------------------
# Template Apply — endpoint
# ---------------------------------------------------------------------------

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def obe_template_apply(request):
    """Apply an OBE template preset to one or more courses.

    Body: { "template_id": int, "subject_codes": list[str], "dry_run": bool }

    Only targets InternalMarkMapping and ObeQpPatternConfig (config models).
    Never writes to student mark tables or final_internal_marks.
    """
    denied = _require_obe_master_permission(request)
    if denied is not None:
        return denied

    data = request.data if isinstance(request.data, dict) else {}
    template_id = data.get('template_id')
    subject_codes = data.get('subject_codes')
    dry_run = bool(data.get('dry_run', True))

    if not template_id:
        return Response({'detail': 'template_id is required'}, status=status.HTTP_400_BAD_REQUEST)
    if not isinstance(subject_codes, list) or not subject_codes:
        return Response(
            {'detail': 'subject_codes must be a non-empty list'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    from .models import ObeTemplatePreset, ObeAuditChange, InternalMarkMapping, ObeQpPatternConfig
    from academics.models import Subject

    try:
        template = ObeTemplatePreset.objects.get(pk=template_id)
    except ObeTemplatePreset.DoesNotExist:
        return Response(
            {'detail': f'Template {template_id} not found'},
            status=status.HTTP_404_NOT_FOUND,
        )

    payload = template.payload if isinstance(template.payload, dict) else {}
    tpl_mapping = payload.get('internal_mark_mapping')
    tpl_qp_patterns = payload.get('qp_patterns') or []
    tpl_class_types_raw = payload.get('target_class_types')
    tpl_class_types = (
        {str(ct).strip().upper() for ct in tpl_class_types_raw if ct}
        if isinstance(tpl_class_types_raw, list)
        else None
    )

    user = getattr(request, 'user', None)
    user_id = getattr(user, 'id', None)
    results = []

    for raw_code in subject_codes:
        code = str(raw_code or '').strip()
        if not code:
            results.append({'subject_code': raw_code, 'status': 'error', 'errors': ['empty subject code']})
            continue

        try:
            subject = Subject.objects.get(code=code)
        except Subject.DoesNotExist:
            results.append({'subject_code': code, 'status': 'error', 'errors': ['subject not found']})
            continue

        class_type = str(getattr(subject, 'class_type', '') or '').strip().upper() or 'THEORY'

        if tpl_class_types and class_type not in tpl_class_types:
            results.append({
                'subject_code': code,
                'status': 'skipped',
                'reason': f'template targets {sorted(tpl_class_types)}, subject class_type is {class_type}',
            })
            continue

        # Determine applicable QP patterns for this class_type
        applicable_patterns = []
        for pat in tpl_qp_patterns:
            if not isinstance(pat, dict):
                continue
            pat_ct = str(pat.get('class_type') or '').strip().upper()
            if pat_ct and pat_ct != class_type:
                continue
            applicable_patterns.append(pat)

        # Validate
        all_errors = []
        if tpl_mapping is not None:
            all_errors.extend(_validate_internal_mark_mapping_for_apply(tpl_mapping, class_type))
        for pat in applicable_patterns:
            errs = _validate_qp_pattern_for_apply(pat.get('pattern') or {})
            if errs:
                exam_label = pat.get('exam', '?')
                all_errors.extend(f'qp_pattern({exam_label}): {e}' for e in errs)

        if all_errors:
            results.append({'subject_code': code, 'status': 'error', 'errors': all_errors})
            continue

        relevant_exams = [
            (str(p.get('exam', '')).strip().upper(), p.get('question_paper_type'))
            for p in applicable_patterns
            if p.get('exam')
        ]

        before_snapshot = _build_config_snapshot(subject, class_type, relevant_exams)

        # Build the after snapshot from template values (not yet written)
        after_qp = []
        for p in applicable_patterns:
            exam_val = str(p.get('exam', '')).strip().upper()
            if exam_val:
                after_qp.append({
                    'exam': exam_val,
                    'question_paper_type': p.get('question_paper_type'),
                    'pattern': p.get('pattern'),
                })
        after_snapshot = {
            'internal_mark_mapping': tpl_mapping if tpl_mapping is not None else before_snapshot['internal_mark_mapping'],
            'qp_patterns': after_qp if after_qp else before_snapshot['qp_patterns'],
        }

        if dry_run:
            results.append({
                'subject_code': code,
                'status': 'dry_run',
                'class_type': class_type,
                'before': before_snapshot,
                'after': after_snapshot,
            })
            continue

        # Live run — per-subject atomic block so one failure doesn't roll back others
        try:
            with transaction.atomic():
                if tpl_mapping is not None:
                    InternalMarkMapping.objects.select_for_update().filter(subject=subject)
                    InternalMarkMapping.objects.update_or_create(
                        subject=subject,
                        defaults={'mapping': tpl_mapping, 'updated_by': user_id},
                    )

                for pat in applicable_patterns:
                    exam_val = str(pat.get('exam', '')).strip().upper()
                    if not exam_val:
                        continue
                    qp_type = pat.get('question_paper_type')
                    pattern_val = pat.get('pattern')
                    ct_key = class_type
                    ObeQpPatternConfig.objects.select_for_update().filter(
                        class_type__iexact=ct_key,
                        question_paper_type=qp_type,
                        exam__iexact=exam_val,
                    )
                    ObeQpPatternConfig.objects.update_or_create(
                        class_type=ct_key,
                        question_paper_type=qp_type,
                        exam=exam_val,
                        defaults={'pattern': pattern_val, 'updated_by': user_id},
                    )

                ObeAuditChange.objects.create(
                    subject_code=code,
                    template=template,
                    before_snapshot=before_snapshot,
                    after_snapshot=after_snapshot,
                    changed_by=user,
                )

            results.append({'subject_code': code, 'status': 'applied', 'class_type': class_type})
        except Exception as exc:
            results.append({'subject_code': code, 'status': 'error', 'errors': [str(exc)]})

    return Response({
        'status': 'ok',
        'dry_run': dry_run,
        'template_id': template_id,
        'results': results,
    })


def _reset_assessment_rows(*, request, assessment_key: str, subject, ta, create_notification: bool = False) -> dict:
    from .models import AssessmentDraft
    from .models import LabPublishedSheet, Cia1PublishedSheet, Cia2PublishedSheet
    from .models import ObeCqiDraft, ObeCqiPublished
    from .models import Ssa1Mark, Ssa2Mark, Review1Mark, Review2Mark, Formative1Mark, Formative2Mark, Cia1Mark, Cia2Mark, ProjectMark
    from .models import ObeMarkTableLock

    deleted = {
        'draft': 0,
        'published': 0,
        'lock': 0,
    }
    strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
    student_ids = _get_teaching_assignment_student_ids(ta)

    with transaction.atomic():
        try:
            deleted['draft'] = _delete_scoped_obe_json_rows(
                AssessmentDraft,
                subject=subject,
                teaching_assignment=ta,
                strict_scope=strict_scope,
                assessment=assessment_key,
            )
        except Exception:
            deleted['draft'] = 0

        try:
            if assessment_key == 'ssa1':
                deleted['published'] += _delete_marks_rows_for_reset(
                    Ssa1Mark.objects.filter(subject=subject, student_id__in=student_ids),
                    ta=ta,
                    strict_scope=strict_scope,
                )
            elif assessment_key == 'review1':
                review1_qs = Review1Mark.objects.filter(subject=subject)
                if student_ids:
                    review1_qs = review1_qs.filter(student_id__in=student_ids)
                deleted['published'] += _delete_marks_rows_for_reset(
                    review1_qs,
                    ta=ta,
                    strict_scope=strict_scope,
                )
                deleted['published'] += _delete_scoped_obe_json_rows(
                    LabPublishedSheet,
                    subject=subject,
                    teaching_assignment=ta,
                    strict_scope=strict_scope,
                    assessment='review1',
                )
                project_qs = ProjectMark.objects.filter(subject=subject)
                if student_ids:
                    project_qs = project_qs.filter(student_id__in=student_ids)
                deleted['published'] += int(project_qs.delete()[0] or 0)
            elif assessment_key == 'ssa2':
                deleted['published'] += _delete_marks_rows_for_reset(
                    Ssa2Mark.objects.filter(subject=subject, student_id__in=student_ids),
                    ta=ta,
                    strict_scope=strict_scope,
                )
            elif assessment_key == 'review2':
                review2_qs = Review2Mark.objects.filter(subject=subject)
                if student_ids:
                    review2_qs = review2_qs.filter(student_id__in=student_ids)
                deleted['published'] += _delete_marks_rows_for_reset(
                    review2_qs,
                    ta=ta,
                    strict_scope=strict_scope,
                )
                deleted['published'] += _delete_scoped_obe_json_rows(
                    LabPublishedSheet,
                    subject=subject,
                    teaching_assignment=ta,
                    strict_scope=strict_scope,
                    assessment='review2',
                )
                project_qs = ProjectMark.objects.filter(subject=subject)
                if student_ids:
                    project_qs = project_qs.filter(student_id__in=student_ids)
                deleted['published'] += int(project_qs.delete()[0] or 0)
            elif assessment_key == 'formative1':
                deleted['published'] += _delete_marks_rows_for_reset(
                    Formative1Mark.objects.filter(subject=subject, student_id__in=student_ids),
                    ta=ta,
                    strict_scope=strict_scope,
                )
                deleted['published'] += _delete_scoped_obe_json_rows(
                    LabPublishedSheet,
                    subject=subject,
                    teaching_assignment=ta,
                    strict_scope=strict_scope,
                    assessment='formative1',
                )
            elif assessment_key == 'formative2':
                deleted['published'] += _delete_marks_rows_for_reset(
                    Formative2Mark.objects.filter(subject=subject, student_id__in=student_ids),
                    ta=ta,
                    strict_scope=strict_scope,
                )
                deleted['published'] += _delete_scoped_obe_json_rows(
                    LabPublishedSheet,
                    subject=subject,
                    teaching_assignment=ta,
                    strict_scope=strict_scope,
                    assessment='formative2',
                )
            elif assessment_key == 'cia1':
                deleted['published'] += _delete_scoped_obe_json_rows(
                    Cia1PublishedSheet,
                    subject=subject,
                    teaching_assignment=ta,
                    strict_scope=strict_scope,
                )
                deleted['published'] += _delete_marks_rows_for_reset(
                    Cia1Mark.objects.filter(subject=subject, student_id__in=student_ids),
                    ta=ta,
                    strict_scope=strict_scope,
                )
                deleted['published'] += _delete_scoped_obe_json_rows(
                    LabPublishedSheet,
                    subject=subject,
                    teaching_assignment=ta,
                    strict_scope=strict_scope,
                    assessment='cia1',
                )
            elif assessment_key == 'cia2':
                deleted['published'] += _delete_scoped_obe_json_rows(
                    Cia2PublishedSheet,
                    subject=subject,
                    teaching_assignment=ta,
                    strict_scope=strict_scope,
                )
                deleted['published'] += _delete_marks_rows_for_reset(
                    Cia2Mark.objects.filter(subject=subject, student_id__in=student_ids),
                    ta=ta,
                    strict_scope=strict_scope,
                )
                deleted['published'] += _delete_scoped_obe_json_rows(
                    LabPublishedSheet,
                    subject=subject,
                    teaching_assignment=ta,
                    strict_scope=strict_scope,
                    assessment='cia2',
                )
            elif assessment_key == 'model':
                try:
                    from .models import ModelPublishedSheet, ModelExamMark, ModelExamCOMark
                    deleted['published'] += _delete_scoped_obe_json_rows(
                        ModelPublishedSheet,
                        subject=subject,
                        teaching_assignment=ta,
                        strict_scope=strict_scope,
                    )
                    deleted['published'] += _delete_marks_rows_for_reset(
                        ModelExamMark.objects.filter(subject=subject, student_id__in=student_ids),
                        ta=ta,
                        strict_scope=strict_scope,
                    )
                    deleted['published'] += _delete_marks_rows_for_reset(
                        ModelExamCOMark.objects.filter(subject=subject, student_id__in=student_ids),
                        ta=ta,
                        strict_scope=strict_scope,
                    )
                except Exception:
                    pass
                deleted['published'] += _delete_scoped_obe_json_rows(
                    LabPublishedSheet,
                    subject=subject,
                    teaching_assignment=ta,
                    strict_scope=strict_scope,
                    assessment='model',
                )
            elif assessment_key == 'cqi':
                deleted['draft'] += _delete_scoped_obe_json_rows(
                    ObeCqiDraft,
                    subject=subject,
                    teaching_assignment=ta,
                    strict_scope=strict_scope,
                )
                deleted['published'] += _delete_scoped_obe_json_rows(
                    ObeCqiPublished,
                    subject=subject,
                    teaching_assignment=ta,
                    strict_scope=strict_scope,
                )
        except Exception:
            pass

        try:
            if assessment_key == 'cqi':
                deleted['lock'] = int(
                    ObeMarkTableLock.objects.filter(teaching_assignment=ta, assessment__startswith='cqi_').delete()[0] or 0
                )
            else:
                deleted['lock'] = int(ObeMarkTableLock.objects.filter(teaching_assignment=ta, assessment=assessment_key).delete()[0] or 0)
        except Exception:
            deleted['lock'] = 0

        if create_notification:
            try:
                from .models import IqacResetNotification
                IqacResetNotification.objects.create(
                    teaching_assignment=ta,
                    assessment=assessment_key,
                    reset_by=request.user if hasattr(request, 'user') else None,
                )
            except Exception:
                pass

    return deleted


def _normalize_reset_assessment_key(value: str) -> str:
    raw = str(value or '').strip().lower()
    if not raw:
        return ''

    token = re.sub(r'[^a-z0-9]+', '', raw)
    aliases = {
        'ssa1': 'ssa1',
        'asmt1': 'ssa1',
        'assessment1': 'ssa1',
        'review1': 'review1',
        'reviewone': 'review1',
        'ssa2': 'ssa2',
        'asmt2': 'ssa2',
        'assessment2': 'ssa2',
        'review2': 'review2',
        'reviewtwo': 'review2',
        'cia1': 'cia1',
        'cycle1': 'cia1',
        'cia2': 'cia2',
        'cycle2': 'cia2',
        'formative1': 'formative1',
        'fa1': 'formative1',
        'formative2': 'formative2',
        'fa2': 'formative2',
        'model': 'model',
        'modelexam': 'model',
        'cqi': 'cqi',
        'cdap': 'cdap',
        'articulation': 'articulation',
        'lca': 'lca',
    }
    return aliases.get(token, raw)


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def iqac_reset_assessment(request, assessment: str, subject_id: str):
    """IQAC/OBE Master: reset a single assessment for a course.

    Clears:
    - Draft JSON (AssessmentDraft)
    - Published data for that assessment (marks tables / published snapshots)
    - Mark table lock row for the teaching assignment + assessment

    Does NOT affect other assessments.

    Query/body:
    - teaching_assignment_id (required for deterministic lock reset)
    """
    auth = _require_obe_master_permission(request)
    if auth:
        return auth

    assessment_key = _normalize_reset_assessment_key(assessment)
    if assessment_key not in {'ssa1', 'review1', 'ssa2', 'review2', 'cia1', 'cia2', 'formative1', 'formative2', 'model', 'cqi', 'cdap', 'articulation', 'lca'}:
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    subject = _get_subject(subject_id, request)
    ta_id = _get_teaching_assignment_id_from_request(request, request.data if isinstance(request.data, dict) else None)
    if ta_id is None:
        return Response({'detail': 'teaching_assignment_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        ta = TeachingAssignment.objects.select_related('section', 'academic_year').filter(id=int(ta_id), is_active=True).first()
    except Exception:
        ta = None

    if ta is None:
        return Response({'detail': 'Invalid teaching_assignment_id.'}, status=status.HTTP_400_BAD_REQUEST)

    deleted = _reset_assessment_rows(
        request=request,
        assessment_key=assessment_key,
        subject=subject,
        ta=ta,
        create_notification=True,
    )

    return Response({'status': 'reset', 'assessment': assessment_key, 'subject_code': subject.code, 'deleted': deleted})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def faculty_reset_assessment(request, assessment: str, subject_id: str):
    """Faculty reset for CIA marks in own scoped teaching assignment.

    Supports: ssa1, ssa2, cia1, cia2, review1, review2, formative1, formative2, model.
    """
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    assessment_key = _normalize_reset_assessment_key(assessment)
    if assessment_key not in {'ssa1', 'ssa2', 'cia1', 'cia2', 'review1', 'review2', 'formative1', 'formative2', 'model'}:
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    subject = _get_subject(subject_id, request)
    ta_id = _get_teaching_assignment_id_from_request(request, request.data if isinstance(request.data, dict) else None)
    if ta_id is None:
        return Response({'detail': 'teaching_assignment_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    gate = _enforce_assessment_enabled_for_course(
        request,
        subject_code=subject.code,
        assessment=assessment_key,
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate

    gate = _enforce_publish_window(request, subject.code, assessment_key)
    if gate is not None:
        return gate

    gate = _enforce_mark_entry_not_blocked(
        request,
        subject_code=subject.code,
        subject_name=subject.name,
        assessment=assessment_key,
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
    if ta is None:
        return Response({'detail': 'Teaching assignment not found for this course.'}, status=status.HTTP_403_FORBIDDEN)

    deleted = _reset_assessment_rows(
        request=request,
        assessment_key=assessment_key,
        subject=subject,
        ta=ta,
        create_notification=False,
    )

    return Response({'status': 'reset', 'assessment': assessment_key, 'subject_code': subject.code, 'deleted': deleted})


def _parse_due_at(value):
    if value is None:
        return None
    if isinstance(value, str):
        dt = parse_datetime(value)
        if dt is None:
            return None
        if timezone.is_naive(dt):
            try:
                dt = timezone.make_aware(dt, timezone.get_current_timezone())
            except Exception:
                pass
        return dt
    return None


def _parse_open_from(value):
    # Same parsing rules as due_at
    return _parse_due_at(value)


def _normalize_obe_class_type(value) -> str:
    raw = str(value or '').strip().upper()
    compact = re.sub(r'[^A-Z0-9]+', '', raw)

    if not compact:
        return 'THEORY'
    if 'TCPR' in compact:
        return 'TCPR'
    if 'TCPL' in compact:
        return 'TCPL'
    if compact == 'THEORYPMBL' or compact == 'THEORY' or compact.startswith('THEORY'):
        return 'THEORY'
    if compact == 'PRBL' or compact == 'PROJECT' or 'PROJECT' in compact:
        return 'PROJECT'
    if compact == 'LAB' or compact == 'L' or compact.startswith('LAB'):
        return 'LAB'
    if compact == 'PRACTICAL' or compact.startswith('PRACT'):
        return 'PRACTICAL'
    if compact == 'AUDIT':
        return 'AUDIT'
    if compact == 'SPECIAL':
        return 'SPECIAL'
    return raw or 'THEORY'


def _resolve_teaching_assignment_class_type(teaching_assignment) -> str:
    if teaching_assignment is None:
        return 'THEORY'

    try:
        row = getattr(teaching_assignment, 'curriculum_row', None)
        if row is not None:
            class_type = getattr(row, 'class_type', None) or getattr(getattr(row, 'master', None), 'class_type', None)
            if class_type:
                return _normalize_obe_class_type(class_type)
    except Exception:
        pass

    try:
        elective_subject = getattr(teaching_assignment, 'elective_subject', None)
        class_type = getattr(elective_subject, 'class_type', None) if elective_subject is not None else None
        if class_type:
            return _normalize_obe_class_type(class_type)
    except Exception:
        pass

    return 'THEORY'


def _resolve_staff_teaching_assignment(request, subject_code: str, teaching_assignment_id: int | None = None):
    user = getattr(request, 'user', None)
    staff_profile = getattr(user, 'staff_profile', None)

    qs = TeachingAssignment.objects.select_related('academic_year', 'subject', 'curriculum_row', 'section').filter(is_active=True)
    if teaching_assignment_id is not None:
        ta = None
        if staff_profile is not None:
            ta = qs.filter(id=teaching_assignment_id, staff=staff_profile).first()
            # For IQAC/HOD/master roles that also have a staff_profile, allow scoping
            # by explicit teaching_assignment_id even if they are not the assigned staff.
            if ta is None and _has_obe_master_access(user):
                ta = qs.filter(id=teaching_assignment_id).first()
        elif _has_obe_master_access(user):
            ta = qs.filter(id=teaching_assignment_id).first()
        if ta is not None:
            return ta

    if not staff_profile:
        return None

    # fallback: match by subject code
    qs = qs.filter(staff=staff_profile).filter(
        Q(subject__code=subject_code)
        | Q(curriculum_row__course_code=subject_code)
        | Q(curriculum_row__master__course_code=subject_code)
    )
    # prefer active academic year
    if qs.filter(academic_year__is_active=True).exists():
        qs = qs.filter(academic_year__is_active=True)
    return qs.order_by('-id').first()


def _resolve_curriculum_row_for_subject(request, subject_code: str, teaching_assignment_id: int | None = None):
    """Best-effort curriculum row lookup for a subject code.

    Preference:
    1) TeachingAssignment.curriculum_row (most accurate)
    2) CurriculumDepartment match for staff user's department
    3) Any CurriculumDepartment match
    """
    try:
        ta = _resolve_staff_teaching_assignment(request, subject_code=subject_code, teaching_assignment_id=teaching_assignment_id)
        row = getattr(ta, 'curriculum_row', None)
        if row is not None:
            return row
    except Exception:
        row = None

    try:
        from curriculum.models import CurriculumDepartment

        staff_profile = getattr(getattr(request, 'user', None), 'staff_profile', None)
        dept = getattr(staff_profile, 'current_department', None) if staff_profile else None
        qs = CurriculumDepartment.objects.all().select_related('master', 'department')
        if dept is not None:
            qs = qs.filter(department=dept)
        code = str(subject_code or '').strip()
        if code:
            qs = qs.filter(Q(course_code__iexact=code) | Q(master__course_code__iexact=code))
        return qs.order_by('-updated_at').first()
    except Exception:
        return None


def _is_cqi_assessment_key(assessment: str | None) -> bool:
    return str(assessment or '').strip().lower().startswith('cqi_')


def _assessment_enablement_key(assessment: str | None) -> str:
    assessment_key = str(assessment or '').strip().lower()
    if _is_cqi_assessment_key(assessment_key):
        parts = [p for p in assessment_key.split('_') if p]
        if len(parts) >= 2:
            return str(parts[1]).strip().lower()
        return 'model'
    return assessment_key


def _is_valid_mark_assessment_key(assessment: str | None, *, allow_documents: bool = True, allow_cqi: bool = False) -> bool:
    assessment_key = str(assessment or '').strip().lower()
    base = {'ssa1', 'review1', 'ssa2', 'review2', 'cia1', 'cia2', 'formative1', 'formative2', 'model'}
    if allow_documents:
        base.update({'cdap', 'articulation', 'lca'})
    if assessment_key in base:
        return True
    if allow_cqi and _is_cqi_assessment_key(assessment_key):
        return True
    return False


def _normalize_cqi_page_key(page_key=None, assessment_type=None, co_numbers=None) -> str:
    raw = str(page_key or '').strip().lower()
    if raw:
        return raw

    assessment = str(assessment_type or '').strip().lower()
    nums: list[int] = []
    if isinstance(co_numbers, list):
        for item in co_numbers:
            try:
                n = int(item)
            except Exception:
                continue
            if n not in nums:
                nums.append(n)
    if assessment and nums:
        return f"{assessment}:{','.join(str(n) for n in nums)}"
    if assessment:
        return assessment
    return ''


def _build_cqi_assessment_key(page_key=None, assessment_type=None, co_numbers=None) -> str:
    normalized_page_key = _normalize_cqi_page_key(page_key=page_key, assessment_type=assessment_type, co_numbers=co_numbers)
    if not normalized_page_key:
        return 'cqi_model'

    if ':' in normalized_page_key:
        assessment, raw_numbers = normalized_page_key.split(':', 1)
    else:
        assessment, raw_numbers = normalized_page_key, ''

    assessment_slug = re.sub(r'[^a-z0-9]+', '_', str(assessment or '').strip().lower()).strip('_') or 'model'
    nums: list[str] = []
    for piece in str(raw_numbers or '').split(','):
        piece = str(piece or '').strip()
        if not piece:
            continue
        if piece.isdigit() and piece not in nums:
            nums.append(piece)
    suffix = f"_{'_'.join(nums)}" if nums else ''
    return f"cqi_{assessment_slug}{suffix}"


def _is_cqi_pages_container(raw) -> bool:
    return isinstance(raw, dict) and isinstance(raw.get('__pages'), dict)


def _merge_cqi_entries(raw_entries) -> dict:
    if not isinstance(raw_entries, dict):
        return {}
    if not _is_cqi_pages_container(raw_entries):
        return raw_entries

    merged: dict = {}
    for payload in (raw_entries.get('__pages') or {}).values():
        if not isinstance(payload, dict):
            continue
        page_entries = payload.get('entries')
        if not isinstance(page_entries, dict):
            continue
        for student_id, student_entries in page_entries.items():
            if not isinstance(student_entries, dict):
                continue
            bucket = merged.setdefault(str(student_id), {})
            for co_key, value in student_entries.items():
                bucket[str(co_key)] = value
    return merged


def _extract_cqi_page_payload(raw_entries, *, page_key=None, assessment_type=None, co_numbers=None) -> dict | None:
    normalized_page_key = _normalize_cqi_page_key(page_key=page_key, assessment_type=assessment_type, co_numbers=co_numbers)
    if not isinstance(raw_entries, dict):
        return None
    if _is_cqi_pages_container(raw_entries):
        if not normalized_page_key:
            return None
        payload = (raw_entries.get('__pages') or {}).get(normalized_page_key)
        return payload if isinstance(payload, dict) else None
    return {
        'page_key': normalized_page_key,
        'assessment_type': str(assessment_type or '').strip().lower() or None,
        'co_numbers': co_numbers if isinstance(co_numbers, list) else [],
        'entries': raw_entries,
    }


def _upsert_cqi_page_payload(raw_entries, *, page_key=None, assessment_type=None, co_numbers=None, entries=None, published_at=None):
    normalized_page_key = _normalize_cqi_page_key(page_key=page_key, assessment_type=assessment_type, co_numbers=co_numbers)
    if not normalized_page_key:
        return entries if isinstance(entries, dict) else {}

    pages = {}
    if _is_cqi_pages_container(raw_entries):
        pages = dict(raw_entries.get('__pages') or {})

    page_payload = {
        'page_key': normalized_page_key,
        'assessment_type': str(assessment_type or '').strip().lower() or None,
        'co_numbers': [int(x) for x in (co_numbers or []) if str(x).strip().isdigit()],
        'entries': entries if isinstance(entries, dict) else {},
    }
    if published_at is not None:
        page_payload['published_at'] = published_at.isoformat() if hasattr(published_at, 'isoformat') else str(published_at)
    pages[normalized_page_key] = page_payload
    return {'__pages': pages}


def _collect_cqi_co_numbers(raw_entries, fallback=None) -> list[int]:
    nums: list[int] = []
    if _is_cqi_pages_container(raw_entries):
        for payload in (raw_entries.get('__pages') or {}).values():
            if not isinstance(payload, dict):
                continue
            for item in payload.get('co_numbers') or []:
                try:
                    n = int(item)
                except Exception:
                    continue
                if 1 <= n <= 20 and n not in nums:
                    nums.append(n)
    elif isinstance(fallback, list):
        for item in fallback:
            try:
                n = int(item)
            except Exception:
                continue
            if 1 <= n <= 20 and n not in nums:
                nums.append(n)
    nums.sort()
    return nums


def _enforce_assessment_enabled_for_course(request, *, subject_code: str, assessment: str, teaching_assignment_id: int | None = None):
    """Reject requests for disabled assessments on SPECIAL courses."""
    assessment_key = _assessment_enablement_key(assessment)
    if not assessment_key:
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    row = _resolve_curriculum_row_for_subject(request, subject_code=subject_code, teaching_assignment_id=teaching_assignment_id)
    class_type = str(getattr(row, 'class_type', '') or '').strip().upper() if row else ''
    if class_type != 'SPECIAL':
        return None

    enabled = None
    # Prefer the globally locked SPECIAL selection (if present), otherwise fall
    # back to curriculum-row configuration.
    try:
        from academics.models import SpecialCourseAssessmentSelection

        ta = _resolve_staff_teaching_assignment(request, subject_code=subject_code, teaching_assignment_id=teaching_assignment_id)
        academic_year = getattr(ta, 'academic_year', None) if ta else None
        if row is not None and academic_year is not None and getattr(row, 'master_id', None) is not None:
            sel = SpecialCourseAssessmentSelection.objects.filter(curriculum_row__master_id=row.master_id, academic_year=academic_year).order_by('id').first()
            if sel is not None:
                enabled = getattr(sel, 'enabled_assessments', None)
    except Exception:
        enabled = None

    if enabled is None:
        enabled = getattr(row, 'enabled_assessments', None) or []
    enabled_set = {str(x).strip().lower() for x in enabled if str(x).strip()}
    if assessment_key not in enabled_set:
        return Response(
            {
                'detail': 'Assessment not enabled for this Special course.',
                'how_to_fix': [
                    'Edit Curriculum Master → set Class type = Special and enable this assessment.',
                    'Propagate the master to departments if needed.',
                ],
            },
            status=status.HTTP_400_BAD_REQUEST,
        )
    return None


def _get_due_schedule_for_request(request, subject_code: str, assessment: str, teaching_assignment_id: int | None = None):
    from .models import ObeAssessmentControl, ObeDueSchedule, ObePublishRequest, ObeGlobalPublishControl
    now = timezone.now()

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject_code, teaching_assignment_id=teaching_assignment_id)
    academic_year = getattr(ta, 'academic_year', None) if ta else None
    semester = _resolve_semester_from_ta(ta)

    if semester is None:
        try:
            row = _resolve_curriculum_row_for_subject(request, subject_code=subject_code, teaching_assignment_id=teaching_assignment_id)
            semester = getattr(row, 'semester', None) if row is not None else None
        except Exception:
            semester = None

    # Subject code matching:
    # IQAC schedules are configured using curriculum `course_code`, while faculty mark-entry
    # screens sometimes call APIs using Subject.code (or master course_code) depending on
    # how the TeachingAssignment is mapped. Treat these as equivalent identifiers and
    # match due schedules / controls / approvals using any available variant.
    subject_code_variants: list[str] = []

    def _add_code(v):
        s = str(v or '').strip()
        if not s:
            return
        if s not in subject_code_variants:
            subject_code_variants.append(s)
        su = s.upper()
        if su and su not in subject_code_variants:
            subject_code_variants.append(su)

    _add_code(subject_code)
    if ta is not None:
        try:
            _add_code(getattr(getattr(ta, 'subject', None), 'code', None))
        except Exception:
            pass
        try:
            row = getattr(ta, 'curriculum_row', None)
            _add_code(getattr(row, 'course_code', None))
            _add_code(getattr(getattr(row, 'master', None), 'course_code', None))
        except Exception:
            pass
        try:
            _add_code(getattr(getattr(ta, 'elective_subject', None), 'course_code', None))
        except Exception:
            pass

    schedule = None
    # Prefer Semester-based schedules; fall back to AcademicYear for older rows.
    if semester is not None:
        schedule = ObeDueSchedule.objects.filter(
            semester=semester,
            subject_code__in=subject_code_variants or [str(subject_code)],
            assessment=str(assessment).lower(),
            is_active=True,
        ).order_by('-updated_at').first()
    if schedule is None and academic_year is not None:
        schedule = ObeDueSchedule.objects.filter(
            academic_year=academic_year,
            semester__isnull=True,
            subject_code__in=subject_code_variants or [str(subject_code)],
            assessment=str(assessment).lower(),
            is_active=True,
        ).order_by('-updated_at').first()

    due_at = getattr(schedule, 'due_at', None)
    open_from = getattr(schedule, 'open_from', None)

    # Assessment enable + open state (IQAC-controlled)
    ctrl = None
    ctrl_active = False
    ctrl_enabled = None
    ctrl_open = None
    ctrl_query_failed = False
    if semester is not None:
        try:
            ctrl = ObeAssessmentControl.objects.filter(
                semester=semester,
                subject_code__in=subject_code_variants or [str(subject_code)],
                assessment=str(assessment).lower(),
            ).order_by('-updated_at').first()
        except OperationalError:
            ctrl = None
            ctrl_query_failed = True

    # Backward compatibility: older controls could be stored against AcademicYear only.
    if ctrl is None and academic_year is not None:
        try:
            ctrl = ObeAssessmentControl.objects.filter(
                academic_year=academic_year,
                semester__isnull=True,
                subject_code__in=subject_code_variants or [str(subject_code)],
                assessment=str(assessment).lower(),
            ).order_by('-updated_at').first()
        except OperationalError:
            ctrl = None
            ctrl_query_failed = True

    if ctrl is not None:
        ctrl_active = True
        ctrl_enabled = bool(getattr(ctrl, 'is_enabled', True))
        ctrl_open = bool(getattr(ctrl, 'is_open', True))

    # If the controls storage is unavailable (e.g., migrations not applied), fail-open.
    # If no explicit control row exists, default to ENABLED so faculty can always
    # access exams until IQAC explicitly disables them.
    if (not ctrl_active) and ctrl_query_failed:
        assessment_enabled = True
        assessment_open = True
    elif ctrl_active:
        assessment_enabled = bool(ctrl_enabled)
        assessment_open = bool(ctrl_open)
    else:
        # No control row exists yet → default to enabled & open.
        assessment_enabled = True
        assessment_open = True

    allowed_by_due = True
    remaining_seconds = None
    starts_in_seconds = None
    if (not assessment_enabled) or (not assessment_open):
        allowed_by_due = False
        remaining_seconds = None
    else:
        if open_from is not None:
            try:
                starts_in_seconds = int(max(0, (open_from - now).total_seconds()))
            except Exception:
                starts_in_seconds = None

        # Not started yet → closed
        if open_from is not None and now < open_from:
            allowed_by_due = False
            remaining_seconds = None
        elif due_at is not None:
            allowed_by_due = now < due_at
            remaining_seconds = int(max(0, (due_at - now).total_seconds()))

    approval = None
    if academic_year is not None:
        approval = ObePublishRequest.objects.filter(
            staff_user=getattr(request, 'user', None),
            academic_year=academic_year,
            subject_code__in=subject_code_variants or [str(subject_code)],
            assessment=str(assessment).lower(),
            status='APPROVED',
            approved_until__gt=now,
        ).order_by('-updated_at').first()

    allowed_by_approval = approval is not None
    approval_until = getattr(approval, 'approved_until', None)

    # Check for an optional global override (takes precedence)
    global_override = None
    global_override_active = False
    global_is_open = None
    global_updated_at = None
    global_updated_by = None
    allowed_by_global = None
    if semester is not None:
        try:
            global_override = ObeGlobalPublishControl.objects.filter(
                semester=semester,
                assessment=str(assessment).lower(),
            ).order_by('-updated_at').first()
        except OperationalError:
            global_override = None
        if global_override is not None:
            global_override_active = True
            global_is_open = bool(getattr(global_override, 'is_open', True))
            global_updated_at = getattr(global_override, 'updated_at', None)
            global_updated_by = getattr(global_override, 'updated_by', None)
            allowed_by_global = global_is_open

    # Backward compatibility: older global overrides stored against AcademicYear.
    if (not global_override_active) and academic_year is not None:
        try:
            global_override = ObeGlobalPublishControl.objects.filter(
                academic_year=academic_year,
                semester__isnull=True,
                assessment=str(assessment).lower(),
            ).order_by('-updated_at').first()
        except OperationalError:
            # Missing table or DB error — treat as no global override so UI won't 500.
            global_override = None
        if global_override is not None:
            global_override_active = True
            global_is_open = bool(getattr(global_override, 'is_open', True))
            global_updated_at = getattr(global_override, 'updated_at', None)
            global_updated_by = getattr(global_override, 'updated_by', None)
            allowed_by_global = global_is_open

    # Final decision:
    # - If a global override exists and it is OPEN, publishing is allowed regardless of due time.
    # - If a global override exists and it is CLOSED, publishing is allowed only when an explicit
    #   IQAC approval exists (so approving a request actually enables publishing).
    # - If no global override exists, normal due/approval logic applies.
    master_cfg_qs = ObeAssessmentMasterConfig.objects.filter(id=1).first()
    master_cfg = master_cfg_qs.config if master_cfg_qs and getattr(master_cfg_qs, 'config', None) else {}
    unlimited_publish = not master_cfg.get('edit_requests_enabled', True)

    if unlimited_publish:
        publish_allowed = True
    elif (not assessment_enabled) or (not assessment_open):
        publish_allowed = False
    elif global_override_active:
        if bool(global_is_open):
            publish_allowed = True
        else:
            publish_allowed = bool(allowed_by_approval)
    else:
        publish_allowed = bool(allowed_by_due or allowed_by_approval)

    return {
        'academic_year': academic_year,
        'semester': semester,
        'teaching_assignment': ta,
        'schedule': schedule,
        'open_from': open_from,
        'due_at': due_at,
        'now': now,
        'remaining_seconds': remaining_seconds,
        'starts_in_seconds': starts_in_seconds,
        'allowed_by_due': allowed_by_due,
        'allowed_by_approval': allowed_by_approval,
        'approval_until': approval_until,
        'assessment_control_active': ctrl_active,
        'assessment_enabled': assessment_enabled,
        'assessment_open': assessment_open,
        'publish_allowed': publish_allowed,
        'global_override_active': global_override_active,
        'global_is_open': global_is_open,
        'global_updated_at': global_updated_at,
        'global_updated_by': global_updated_by,
        'allowed_by_global': allowed_by_global,
    }


def _auto_publish_from_draft_if_due(request, *, subject, assessment_key: str, teaching_assignment_id: int | None = None) -> bool:
    """Best-effort auto publish when due time has passed.

    This is intentionally idempotent; it will no-op if already published.
    """
    assessment_key = str(assessment_key or '').strip().lower()
    if assessment_key not in {'ssa1', 'review1', 'ssa2', 'review2', 'cia1', 'cia2', 'formative1', 'formative2', 'model'}:
        return False

    info = _get_due_schedule_for_request(request, subject_code=getattr(subject, 'code', ''), assessment=assessment_key, teaching_assignment_id=teaching_assignment_id)
    due_at = info.get('due_at')
    now = info.get('now')
    # No timer, or not yet due.
    if due_at is None or now is None or now < due_at:
        return False
    # Unlimited/global-open means no auto publish.
    if info.get('global_override_active') and bool(info.get('global_is_open')):
        return False

    ta = info.get('teaching_assignment')
    academic_year = info.get('academic_year')
    section_name = _resolve_section_name_from_ta(ta)
    strict_scope = _strict_assignment_scope(subject_code=getattr(subject, 'code', ''), teaching_assignment=ta)
    lock = _get_mark_table_lock_if_exists(
        staff_user=getattr(request, 'user', None),
        subject_code=str(getattr(subject, 'code', '') or ''),
        assessment=assessment_key,
        teaching_assignment=ta,
        academic_year=academic_year,
        section_name=section_name,
    )
    if lock is not None and bool(getattr(lock, 'is_published', False)):
        return False

    from .models import (
        AssessmentDraft,
        Cia1Mark,
        Cia1PublishedSheet,
        Cia2Mark,
        Cia2PublishedSheet,
        Formative1Mark,
        Formative2Mark,
        ModelPublishedSheet,
        Review1Mark,
        Review2Mark,
        Ssa1Mark,
        Ssa2Mark,
    )

    draft = _get_scoped_obe_json_row(
        AssessmentDraft,
        subject=subject,
        teaching_assignment=ta,
        strict_scope=strict_scope,
        assessment=assessment_key,
    )
    data = getattr(draft, 'data', None) if draft else None
    if not isinstance(data, dict):
        return False

    # Frontend drafts sometimes wrap the actual sheet under `sheet`.
    payload = data.get('sheet') if isinstance(data.get('sheet'), dict) else data

    # Publish implementations mirror existing publish endpoints but use stored draft.
    try:
        if assessment_key in {'ssa1', 'ssa2'}:
            rows = payload.get('rows', [])
            if not isinstance(rows, list):
                rows = []
            model = Ssa1Mark if assessment_key == 'ssa1' else Ssa2Mark
            with transaction.atomic():
                for item in rows:
                    try:
                        sid = int(item.get('studentId'))
                    except Exception:
                        continue
                    student = StudentProfile.objects.filter(id=sid).first()
                    if not student:
                        continue
                    raw_total = item.get('total')
                    mark = _coerce_decimal_or_none(raw_total)
                    # Match publish endpoint behavior: blank => delete, invalid => skip.
                    if raw_total not in (None, '',) and mark is None:
                        continue
                    if mark is None:
                        _delete_scoped_mark(model, subject=subject, student=student, teaching_assignment=ta)
                    else:
                        _upsert_scoped_mark(
                            model,
                            subject=subject,
                            student=student,
                            teaching_assignment=ta,
                            mark_defaults={'mark': mark},
                        )

        elif assessment_key in {'review1', 'review2'}:
            model = Review1Mark if assessment_key == 'review1' else Review2Mark
            with transaction.atomic():
                rows = payload.get('rows', [])
                if not isinstance(rows, list):
                    rows = []
                for item in rows:
                    try:
                        sid = int(item.get('studentId'))
                    except Exception:
                        continue
                    student = StudentProfile.objects.filter(id=sid).first()
                    if not student:
                        continue
                    raw_total = item.get('total')
                    mark = _coerce_decimal_or_none(raw_total)
                    if raw_total not in (None, '',) and mark is None:
                        continue
                    if mark is None:
                        _delete_scoped_mark(model, subject=subject, student=student, teaching_assignment=ta)
                    else:
                        _upsert_scoped_mark(
                            model,
                            subject=subject,
                            student=student,
                            teaching_assignment=ta,
                            mark_defaults={'mark': mark},
                        )

        elif assessment_key in {'formative1', 'formative2'}:
            rows_by = payload.get('rowsByStudentId', {})
            if not isinstance(rows_by, dict):
                rows_by = {}
            model = Formative1Mark if assessment_key == 'formative1' else Formative2Mark
            with transaction.atomic():
                for sid_str, item in rows_by.items():
                    try:
                        sid = int(sid_str)
                    except Exception:
                        continue
                    student = StudentProfile.objects.filter(id=sid).first()
                    if not student:
                        continue
                    skill1 = _coerce_decimal_or_none((item or {}).get('skill1'))
                    skill2 = _coerce_decimal_or_none((item or {}).get('skill2'))
                    att1 = _coerce_decimal_or_none((item or {}).get('att1'))
                    att2 = _coerce_decimal_or_none((item or {}).get('att2'))
                    if skill1 is None or skill2 is None or att1 is None or att2 is None:
                        _delete_scoped_mark(model, subject=subject, student=student, teaching_assignment=ta)
                        continue
                    total = skill1 + skill2 + att1 + att2
                    _upsert_scoped_mark(
                        model,
                        subject=subject,
                        student=student,
                        teaching_assignment=ta,
                        mark_defaults={'skill1': skill1, 'skill2': skill2, 'att1': att1, 'att2': att2, 'total': total},
                    )

        elif assessment_key in {'cia1', 'cia2'}:
            if assessment_key == 'cia1':
                pub_model = Cia1PublishedSheet
                mark_model = Cia1Mark
            else:
                pub_model = Cia2PublishedSheet
                mark_model = Cia2Mark

            _upsert_scoped_obe_json_row(
                pub_model,
                subject=subject,
                teaching_assignment=ta,
                defaults={'data': payload, 'updated_by': getattr(getattr(request, 'user', None), 'id', None)},
            )
            questions = payload.get('questions', [])
            if not isinstance(questions, list):
                questions = []
            rows_by = payload.get('rowsByStudentId', {})
            if not isinstance(rows_by, dict):
                rows_by = {}
            qkeys = [str(q.get('key')) for q in questions if isinstance(q, dict) and q.get('key')]
            with transaction.atomic():
                for sid_str, row in rows_by.items():
                    try:
                        sid = int(sid_str)
                    except Exception:
                        continue
                    student = StudentProfile.objects.filter(id=sid).first()
                    if not student:
                        continue
                    absent = bool((row or {}).get('absent'))
                    if absent:
                        total_dec = Decimal('0')
                    else:
                        q = (row or {}).get('q', {})
                        if not isinstance(q, dict):
                            q = {}
                        total = Decimal('0')
                        for k in qkeys:
                            dec = _coerce_decimal_or_none(q.get(k))
                            if dec is None:
                                continue
                            total += dec
                        total_dec = total
                    _upsert_scoped_mark(
                        mark_model,
                        subject=subject,
                        student=student,
                        teaching_assignment=ta,
                        mark_defaults={'mark': total_dec},
                    )

        elif assessment_key == 'model':
            _upsert_scoped_obe_json_row(
                ModelPublishedSheet,
                subject=subject,
                teaching_assignment=ta,
                defaults={'data': payload, 'updated_by': getattr(getattr(request, 'user', None), 'id', None)},
            )
        else:
            return False
    except Exception:
        return False

    try:
        _touch_lock_after_publish(
            request,
            subject_code=str(getattr(subject, 'code', '') or ''),
            subject_name=str(getattr(subject, 'name', '') or str(getattr(subject, 'code', '') or '')),
            assessment=assessment_key,
            teaching_assignment_id=teaching_assignment_id,
        )
    except Exception:
        pass
    return True


def _enforce_mark_entry_window(request, *, subject, assessment_key: str, teaching_assignment_id: int | None = None):
    """Enforce IQAC-configured start/end window for editing (draft saves and marks PUT).

    If the window has ended, attempts best-effort auto-publish.
    """
    assessment_key = str(assessment_key or '').strip().lower()
    info = _get_due_schedule_for_request(request, subject_code=getattr(subject, 'code', ''), assessment=assessment_key, teaching_assignment_id=teaching_assignment_id)

    # Unlimited/global-open: allow edits regardless of due time (but still respect enable/open flags).
    if info.get('global_override_active') and bool(info.get('global_is_open')):
        return None

    if bool(info.get('assessment_enabled')) is False or bool(info.get('assessment_open')) is False:
        return Response({'detail': 'Assessment is disabled by IQAC.'}, status=status.HTTP_403_FORBIDDEN)

    open_from = info.get('open_from')
    due_at = info.get('due_at')
    now = info.get('now')

    # Before start
    if open_from is not None and now is not None and now < open_from:
        return Response(
            {
                'detail': 'Mark entry window has not started yet.',
                'assessment': assessment_key,
                'subject_code': str(getattr(subject, 'code', '') or ''),
                'open_from': open_from.isoformat() if open_from else None,
                'due_at': due_at.isoformat() if due_at else None,
                'now': now.isoformat() if now else None,
            },
            status=status.HTTP_403_FORBIDDEN,
        )

    # After end
    if due_at is not None and now is not None and now >= due_at:
        auto_published = False
        try:
            auto_published = _auto_publish_from_draft_if_due(
                request,
                subject=subject,
                assessment_key=assessment_key,
                teaching_assignment_id=teaching_assignment_id,
            )
        except Exception:
            auto_published = False
        return Response(
            {
                'detail': 'Mark entry window ended. Table is now read-only.',
                'assessment': assessment_key,
                'subject_code': str(getattr(subject, 'code', '') or ''),
                'open_from': open_from.isoformat() if open_from else None,
                'due_at': due_at.isoformat() if due_at else None,
                'now': now.isoformat() if now else None,
                'auto_published': bool(auto_published),
            },
            status=status.HTTP_403_FORBIDDEN,
        )

    return None


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def assessment_controls(request):
    """IQAC: list per-subject assessment controls for selected semesters."""
    auth = _require_obe_master(request)
    if auth:
        return auth

    raw = (
        getattr(request, 'query_params', {}).get('semester_ids')
        if hasattr(request, 'query_params')
        else request.GET.get('semester_ids')
    )
    sem_ids: list[int] = []
    for part in str(raw or '').split(','):
        part = part.strip()
        if not part:
            continue
        try:
            sem_ids.append(int(part))
        except Exception:
            continue

    if not sem_ids:
        return Response({'detail': 'semester_ids is required.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import ObeAssessmentControl

    try:
        qs = ObeAssessmentControl.objects.select_related('semester').filter(semester_id__in=sem_ids).order_by('semester_id', 'subject_code', 'assessment')
    except OperationalError:
        return Response(
            {
                'detail': 'Assessment controls are not available on this server (database table missing).',
                'how_to_fix': [
                    'Run backend migrations (includes OBE migration 0035_obeassessmentcontrol).',
                    'Restart the backend after migrating.',
                ],
            },
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )
    out = []
    for r in qs:
        out.append(
            {
                'id': r.id,
                'semester': {
                    'id': r.semester_id,
                    'number': getattr(getattr(r, 'semester', None), 'number', None),
                }
                if r.semester_id
                else None,
                'subject_code': r.subject_code,
                'subject_name': r.subject_name,
                'assessment': r.assessment,
                'is_enabled': bool(r.is_enabled),
                'is_open': bool(r.is_open),
                'updated_at': r.updated_at.isoformat() if r.updated_at else None,
                'updated_by': r.updated_by,
            }
        )

    return Response({'results': out})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def assessment_controls_bulk_set(request):
    """IQAC: bulk upsert assessment controls for many subjects and/or assessments.

    Body: semester_id, subject_codes (list), assessments (list), is_enabled? (bool), is_open? (bool)
    At least one of is_enabled/is_open must be provided.
    """
    auth = _require_obe_master(request)
    if auth:
        return auth

    body = request.data or {}
    sem_id = body.get('semester_id')
    assessments = body.get('assessments') or []
    subject_codes = body.get('subject_codes') or []

    has_is_enabled = 'is_enabled' in body
    has_is_open = 'is_open' in body
    if not (has_is_enabled or has_is_open):
        return Response({'detail': 'Provide is_enabled and/or is_open.'}, status=status.HTTP_400_BAD_REQUEST)

    if not sem_id:
        return Response({'detail': 'semester_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    if not isinstance(assessments, list) or not assessments:
        return Response({'detail': 'assessments must be a non-empty list.'}, status=status.HTTP_400_BAD_REQUEST)

    if not isinstance(subject_codes, list) or not subject_codes:
        return Response({'detail': 'subject_codes must be a non-empty list.'}, status=status.HTTP_400_BAD_REQUEST)

    norm_assessments = [str(a).strip().lower() for a in assessments]
    allowed = {'ssa1', 'review1', 'ssa2', 'review2', 'cia1', 'cia2', 'formative1', 'formative2', 'model'}
    bad = [a for a in norm_assessments if a not in allowed]
    if bad:
        return Response({'detail': f'Invalid assessments: {bad}'}, status=status.HTTP_400_BAD_REQUEST)

    from academics.models import Semester, Subject
    from .models import ObeAssessmentControl

    sem = Semester.objects.filter(id=int(sem_id)).first()
    if not sem:
        return Response({'detail': 'Semester not found.'}, status=status.HTTP_404_NOT_FOUND)

    updated = 0
    req_user_id = getattr(getattr(request, 'user', None), 'id', None)

    try:
        for code in [str(s).strip() for s in subject_codes]:
            if not code:
                continue
            subj = Subject.objects.filter(code=code).first()
            name = getattr(subj, 'name', '') if subj else ''
            for a in norm_assessments:
                defaults = {
                    'subject_name': name,
                    'updated_by': req_user_id,
                    'created_by': req_user_id,
                }
                if has_is_enabled:
                    defaults['is_enabled'] = bool(body.get('is_enabled'))
                if has_is_open:
                    defaults['is_open'] = bool(body.get('is_open'))

                ObeAssessmentControl.objects.update_or_create(
                    semester=sem,
                    subject_code=code,
                    assessment=a,
                    defaults=defaults,
                )
                updated += 1
    except OperationalError:
        return Response(
            {
                'detail': 'Failed to save assessment controls (database table missing).',
                'how_to_fix': [
                    'Run backend migrations (includes OBE migration 0035_obeassessmentcontrol).',
                    'Restart the backend after migrating.',
                ],
            },
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    return Response({'status': 'ok', 'updated': int(updated)})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def obe_semesters(request):
    """IQAC helper: list all semesters used by sections/curriculum."""
    auth = _require_obe_master(request)
    if auth:
        return auth

    from academics.models import Semester

    out = []
    for s in Semester.objects.all().order_by('number'):
        out.append({'id': s.id, 'number': getattr(s, 'number', None)})
    return Response({'results': out})


def _enforce_publish_window(request, subject_code: str, assessment: str):
    # Accept optional TA id to correctly resolve academic year
    ta_id_raw = getattr(request, 'query_params', {}).get('teaching_assignment_id') if hasattr(request, 'query_params') else request.GET.get('teaching_assignment_id')
    ta_id = None
    try:
        if ta_id_raw:
            ta_id = int(str(ta_id_raw))
    except Exception:
        ta_id = None

    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject_code, assessment=assessment, teaching_assignment_id=ta_id)
    if gate is not None:
        return gate

    info = _get_due_schedule_for_request(request, subject_code=subject_code, assessment=assessment, teaching_assignment_id=ta_id)
    if info.get('publish_allowed'):
        return None

    due_at = info.get('due_at')
    now = info.get('now')
    # If a global override exists and it is closed, return a distinct message
    if info.get('global_override_active') and (info.get('global_is_open') is False):
        return Response(
            {
                'detail': 'Publishing is disabled globally by IQAC for this assessment.',
                'assessment': str(assessment).lower(),
                'subject_code': str(subject_code),
                'due_at': due_at.isoformat() if due_at else None,
                'now': now.isoformat() if now else None,
                'remaining_seconds': info.get('remaining_seconds'),
                'how_to_fix': [
                    'Contact IQAC to enable global publishing for this assessment or reset the global override.',
                ],
            },
            status=status.HTTP_403_FORBIDDEN,
        )
    return Response(
        {
            'detail': 'Publish window closed. Please request approval from IQAC.',
            'assessment': str(assessment).lower(),
            'subject_code': str(subject_code),
            'due_at': due_at.isoformat() if due_at else None,
            'now': now.isoformat() if now else None,
            'remaining_seconds': info.get('remaining_seconds'),
            'how_to_fix': [
                'Use the Request button in the OBE mark entry page',
                'Or ask IQAC to approve your publish request',
            ],
        },
        status=status.HTTP_403_FORBIDDEN,
    )


def _faculty_only(request):
    user = request.user
    staff_profile = getattr(user, 'staff_profile', None)
    if not staff_profile:
        # Allow OBE Master / IQAC users to act as faculty for OBE flows (they may pass
        # a `teaching_assignment_id` to identify the section). This permits IQAC users
        # with the `obe.master.manage` permission (or superusers) to access and edit
        # mark-entry endpoints without requiring a staff_profile.
        if _has_obe_master_permission(user):
            return None, None
        return None, Response({'detail': 'Faculty access only.'}, status=status.HTTP_403_FORBIDDEN)
    return staff_profile, None


def _get_subject(subject_id: str, request=None):
    """Resolve a Subject by code.

    If the Subject row does not exist (common when curriculum is used without
    seeding academics.Subject), create a minimal Subject record so OBE drafts/
    publish flows do not 404.
    """
    code = str(subject_id or '').strip()
    if not code:
        return get_object_or_404(Subject, code=code)

    existing = Subject.objects.filter(code=code).select_related('semester').first()
    if existing:
        return existing

    # Best-effort: derive name + semester from CurriculumDepartment for the staff's department.
    name = code
    semester = None
    try:
        from curriculum.models import CurriculumDepartment

        staff_profile = getattr(getattr(request, 'user', None), 'staff_profile', None)
        dept = getattr(staff_profile, 'current_department', None) if staff_profile else None
        if dept:
            row = (
                CurriculumDepartment.objects.filter(department=dept).filter(Q(course_code__iexact=code) | Q(master__course_code__iexact=code))
                .select_related('semester')
                .order_by('-updated_at')
                .first()
            )
            if row:
                name = row.course_name or name
                semester = row.semester
    except Exception:
        # Fall back to any semester if curriculum lookup fails.
        pass

    if semester is None:
        semester = Semester.objects.order_by('id').first()

    if semester is None:
        # No semester exists in the DB; cannot create a Subject.
        return get_object_or_404(Subject, code=code)

    obj, _created = Subject.objects.get_or_create(
        code=code,
        defaults={
            'name': str(name)[:128],
            'semester': semester,
            'course': None,
        },
    )
    return obj


def _coerce_decimal_or_none(raw):
    if raw is None:
        return None
    if isinstance(raw, str) and raw.strip() == '':
        return None
    try:
        return Decimal(str(raw))
    except (InvalidOperation, ValueError):
        return None


@api_view(['GET', 'PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def assessment_draft(request, assessment: str, subject_id: str):
    """Shared draft endpoint.

    - GET: returns draft JSON (or null)
    - PUT: saves draft JSON

    Assessment: ssa1 | ssa2 | cia1 | cia2 | formative1 | formative2 | model
    """
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    if assessment not in {'ssa1', 'review1', 'ssa2', 'review2', 'cia1', 'cia2', 'formative1', 'formative2', 'model', 'cdap', 'articulation', 'lca'}:
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    subject = _get_subject(subject_id, request)
    ta_id = _get_teaching_assignment_id_from_request(request, request.data if request.method == 'PUT' else None)
    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
    strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)

    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject.code, assessment=assessment, teaching_assignment_id=ta_id)
    if gate is not None:
        return gate

    from .models import AssessmentDraft

    def _looks_like_lab_sheet_payload(payload):
        if not isinstance(payload, dict):
            return False
        sheet = payload.get('sheet') if isinstance(payload.get('sheet'), dict) else payload
        if not isinstance(sheet, dict):
            return False
        rows = sheet.get('rowsByStudentId')
        if not isinstance(rows, dict):
            return False
        if isinstance(sheet.get('coConfigs'), dict):
            return True
        for row in rows.values():
            if not isinstance(row, dict):
                continue
            if isinstance(row.get('marksByCo'), dict):
                return True
            if isinstance(row.get('marksA'), list) or isinstance(row.get('marksB'), list):
                return True
        return False

    if request.method == 'PUT':
        gate = _enforce_mark_entry_window(request, subject=subject, assessment_key=assessment, teaching_assignment_id=ta_id)
        if gate is not None:
            return gate

        body = request.data or {}
        data = body.get('data', None)
        if data is None:
            return Response({'detail': 'Missing draft data.'}, status=status.HTTP_400_BAD_REQUEST)
        if not isinstance(data, (dict, list)):
            return Response({'detail': 'Invalid draft data.'}, status=status.HTTP_400_BAD_REQUEST)

        _upsert_scoped_obe_json_row(
            AssessmentDraft,
            subject=subject,
            teaching_assignment=ta,
            assessment=assessment,
            defaults={'data': data, 'updated_by': getattr(request.user, 'id', None)},
        )

        # Keep dedicated LabExamMark rows in sync for lab-style sheets on each save,
        # so DB/Admin portal reflects student marks immediately (not only after publish).
        if assessment in {'cia1', 'cia2', 'model', 'formative1', 'formative2', 'review1', 'review2'} and _looks_like_lab_sheet_payload(data):
            try:
                from .services.exam_mark_persistence import persist_lab_exam_marks
                persist_lab_exam_marks(
                    subject=subject,
                    teaching_assignment=ta,
                    assessment=assessment,
                    data=data,
                )
            except Exception:
                # Never fail draft save because of mirror-sync issues.
                pass
        return Response({'status': 'draft_saved'})

    draft = _get_scoped_obe_json_row(
        AssessmentDraft,
        subject=subject,
        teaching_assignment=ta,
        strict_scope=strict_scope,
        assessment=assessment,
    )
    draft_data = draft.data if draft else None
    updated_at = draft.updated_at.isoformat() if draft and getattr(draft, 'updated_at', None) else None
    updated_by = None
    if draft and getattr(draft, 'updated_by', None):
        try:
            User = apps.get_model(settings.AUTH_USER_MODEL)
            u = User.objects.filter(id=getattr(draft, 'updated_by')).first()
            if u:
                updated_by = {'id': getattr(u, 'id', None), 'username': getattr(u, 'username', None), 'name': ' '.join(filter(None, [getattr(u, 'first_name', ''), getattr(u, 'last_name', '')])).strip() or getattr(u, 'username', None)}
        except Exception:
            updated_by = {'id': getattr(draft, 'updated_by', None)}

    return Response({'subject': {'code': subject.code, 'name': subject.name}, 'draft': draft_data, 'updated_at': updated_at, 'updated_by': updated_by})


def _extract_review_co_splits(subject, assessment_key, co_keys, ta=None):
    co_splits = {}
    try:
        from .models import AssessmentDraft
        drafts = AssessmentDraft.objects.filter(
            subject=subject,
            assessment=assessment_key,
        ).order_by('-updated_at')
        if ta is not None:
            draft = drafts.filter(teaching_assignment=ta).first() or drafts.first()
        else:
            draft = drafts.first()
        if draft and isinstance(draft.data, dict):
            sheet = draft.data.get('sheet', draft.data)
            rows = sheet.get('rows', []) if isinstance(sheet, dict) else []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                sid = str(row.get('studentId', '')).strip()
                if not sid:
                    continue
                entry = {}
                for ck in co_keys:
                    val = None
                    reviewCoMarks = row.get('reviewCoMarks')
                    if isinstance(reviewCoMarks, dict):
                        raw_arr = reviewCoMarks.get(ck)
                        if isinstance(raw_arr, list):
                            try:
                                val = sum([float(x) for x in raw_arr if x not in ('', None)])
                            except (ValueError, TypeError):
                                pass
                    if val is None:
                        dir_val = row.get(ck)
                        if dir_val not in ('', None):
                            try:
                                val = float(dir_val)
                            except (ValueError, TypeError):
                                pass
                    if val is not None:
                        entry[ck] = val
                if entry:
                    co_splits[sid] = entry
    except Exception:
        pass
    return co_splits

def _extract_ssa_co_splits(subject, assessment_key, co_keys, ta=None):
    """Extract per-CO split marks from SSA draft data.

    Returns dict: { studentId_str: { co_key: value, ... } }
    """
    co_splits = {}
    try:
        from .models import AssessmentDraft
        drafts = AssessmentDraft.objects.filter(
            subject=subject,
            assessment=assessment_key,
        ).order_by('-updated_at')
        if ta is not None:
            draft = drafts.filter(teaching_assignment=ta).first() or drafts.first()
        else:
            draft = drafts.first()
        if draft and isinstance(draft.data, dict):
            sheet = draft.data.get('sheet', draft.data)
            rows = sheet.get('rows', []) if isinstance(sheet, dict) else []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                sid = str(row.get('studentId', '')).strip()
                if not sid:
                    continue
                entry = {}
                all_present = True
                for ck in co_keys:
                    val = row.get(ck)
                    if val == '' or val is None:
                        all_present = False
                        break
                    try:
                        entry[ck] = float(val)
                    except (TypeError, ValueError):
                        all_present = False
                        break
                if all_present and entry:
                    co_splits[sid] = entry
    except Exception:
        pass
    return co_splits


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def ssa1_published(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject.code, assessment='ssa1', teaching_assignment_id=ta_id)
    if gate is not None:
        return gate
    from .models import Ssa1Mark
    try:
        ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
        strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
        marks = _safe_marks_map_for_subject(Ssa1Mark, subject=subject, ta=ta, strict_scope=strict_scope)
    except (OperationalError, ProgrammingError):
        marks = {}
        ta = None
    co_splits = _extract_ssa_co_splits(subject, 'ssa1', ['co1', 'co2'], ta=ta)
    return Response({'subject': {'code': subject.code, 'name': subject.name}, 'marks': marks, 'co_splits': co_splits})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def ssa1_publish(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err
    subject = _get_subject(subject_id, request)
    pub_auth = _require_publish_owner(request)
    if pub_auth:
        return pub_auth

    gate = _enforce_publish_window(request, subject.code, 'ssa1')
    if gate is not None:
        return gate

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_mark_entry_not_blocked(
        request,
        subject_code=subject.code,
        subject_name=subject.name,
        assessment='ssa1',
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate

    body = request.data or {}
    data = body.get('data', None)
    if data is None or not isinstance(data, dict):
        return Response({'detail': 'Invalid payload.'}, status=status.HTTP_400_BAD_REQUEST)

    rows = data.get('rows', [])
    if not isinstance(rows, list):
        return Response({'detail': 'Invalid rows.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import Ssa1Mark

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)

    errors: list[str] = []
    with transaction.atomic():
        for item in rows:
            try:
                sid = int(item.get('studentId'))
            except Exception:
                continue
            student = StudentProfile.objects.filter(id=sid).first()
            if not student:
                continue

            mark = _coerce_decimal_or_none(item.get('total'))
            if item.get('total') not in (None, '',) and mark is None:
                errors.append(f'Invalid mark for student {sid}: {item.get("total")}')
                continue

            if mark is None:
                _delete_scoped_mark(Ssa1Mark, subject=subject, student=student, teaching_assignment=ta)
            else:
                _upsert_scoped_mark(
                    Ssa1Mark,
                    subject=subject,
                    student=student,
                    teaching_assignment=ta,
                    mark_defaults={'mark': mark},
                )

    if errors:
        return Response({'detail': 'Validation error.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    try:
        _touch_lock_after_publish(
            request,
            subject_code=subject.code,
            subject_name=subject.name,
            assessment='ssa1',
            teaching_assignment_id=ta_id,
        )
    except OperationalError:
        pass

    return Response({'status': 'published'})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def review1_published(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject.code, assessment='review1', teaching_assignment_id=ta_id)
    if gate is not None:
        return gate
    from .models import Review1Mark
    try:
        ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
        strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
        marks = _safe_marks_map_for_subject(Review1Mark, subject=subject, ta=ta, strict_scope=strict_scope)
    except (OperationalError, ProgrammingError):
        marks = {}
        ta = None
    co_splits = _extract_review_co_splits(subject, 'review1', ['co1', 'co2'], ta=ta)
    return Response({'subject': {'code': subject.code, 'name': subject.name}, 'marks': marks, 'co_splits': co_splits})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def review1_publish(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err
    subject = _get_subject(subject_id, request)
    pub_auth = _require_publish_owner(request)
    if pub_auth:
        return pub_auth

    gate = _enforce_publish_window(request, subject.code, 'review1')
    if gate is not None:
        return gate

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_mark_entry_not_blocked(
        request,
        subject_code=subject.code,
        subject_name=subject.name,
        assessment='review1',
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate

    body = request.data or {}
    data = body.get('data', None)
    if data is None or not isinstance(data, dict):
        return Response({'detail': 'Invalid payload.'}, status=status.HTTP_400_BAD_REQUEST)

    rows = data.get('rows', [])
    if not isinstance(rows, list):
        return Response({'detail': 'Invalid rows.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import Review1Mark, ProjectMark

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
    is_project_course = _resolve_teaching_assignment_class_type(ta) == 'PROJECT'

    errors: list[str] = []
    with transaction.atomic():
        for item in rows:
            try:
                sid = int(item.get('studentId'))
            except Exception:
                continue
            student = StudentProfile.objects.filter(id=sid).first()
            if not student:
                continue

            mark = _coerce_decimal_or_none(item.get('total'))
            if item.get('total') not in (None, '',) and mark is None:
                errors.append(f'Invalid mark for student {sid}: {item.get("total")}')
                continue

            if mark is None:
                _delete_scoped_mark(Review1Mark, subject=subject, student=student, teaching_assignment=ta)
                if is_project_course:
                    _delete_scoped_mark(ProjectMark, subject=subject, student=student, teaching_assignment=ta)
            else:
                _upsert_scoped_mark(
                    Review1Mark,
                    subject=subject,
                    student=student,
                    teaching_assignment=ta,
                    mark_defaults={'mark': mark},
                )
                if is_project_course:
                    _upsert_scoped_mark(
                        ProjectMark,
                        subject=subject,
                        student=student,
                        teaching_assignment=ta,
                        mark_defaults={'mark': mark},
                    )

    if errors:
        return Response({'detail': 'Validation error.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    try:
        _touch_lock_after_publish(
            request,
            subject_code=subject.code,
            subject_name=subject.name,
            assessment='review1',
            teaching_assignment_id=ta_id,
        )
    except OperationalError:
        pass

    return Response({'status': 'published'})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def ssa2_published(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject.code, assessment='ssa2', teaching_assignment_id=ta_id)
    if gate is not None:
        return gate
    from .models import Ssa2Mark
    try:
        ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
        strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
        marks = _safe_marks_map_for_subject(Ssa2Mark, subject=subject, ta=ta, strict_scope=strict_scope)
    except (OperationalError, ProgrammingError):
        marks = {}
        ta = None
    co_splits = _extract_ssa_co_splits(subject, 'ssa2', ['co3', 'co4'], ta=ta)
    return Response({'subject': {'code': subject.code, 'name': subject.name}, 'marks': marks, 'co_splits': co_splits})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def ssa2_publish(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err
    subject = _get_subject(subject_id, request)
    pub_auth = _require_publish_owner(request)
    if pub_auth:
        return pub_auth

    gate = _enforce_publish_window(request, subject.code, 'ssa2')
    if gate is not None:
        return gate

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_mark_entry_not_blocked(
        request,
        subject_code=subject.code,
        subject_name=subject.name,
        assessment='ssa2',
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate

    body = request.data or {}
    data = body.get('data', None)
    if data is None or not isinstance(data, dict):
        return Response({'detail': 'Invalid payload.'}, status=status.HTTP_400_BAD_REQUEST)

    rows = data.get('rows', [])
    if not isinstance(rows, list):
        return Response({'detail': 'Invalid rows.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import Ssa2Mark

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)

    errors: list[str] = []
    with transaction.atomic():
        for item in rows:
            try:
                sid = int(item.get('studentId'))
            except Exception:
                continue
            student = StudentProfile.objects.filter(id=sid).first()
            if not student:
                continue

            mark = _coerce_decimal_or_none(item.get('total'))
            if item.get('total') not in (None, '',) and mark is None:
                errors.append(f'Invalid mark for student {sid}: {item.get("total")}')
                continue

            if mark is None:
                _delete_scoped_mark(Ssa2Mark, subject=subject, student=student, teaching_assignment=ta)
            else:
                _upsert_scoped_mark(
                    Ssa2Mark,
                    subject=subject,
                    student=student,
                    teaching_assignment=ta,
                    mark_defaults={'mark': mark},
                )

    if errors:
        return Response({'detail': 'Validation error.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    try:
        _touch_lock_after_publish(
            request,
            subject_code=subject.code,
            subject_name=subject.name,
            assessment='ssa2',
            teaching_assignment_id=ta_id,
        )
    except OperationalError:
        pass

    return Response({'status': 'published'})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def review2_published(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject.code, assessment='review2', teaching_assignment_id=ta_id)
    if gate is not None:
        return gate
    from .models import Review2Mark
    try:
        ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
        strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
        marks = _safe_marks_map_for_subject(Review2Mark, subject=subject, ta=ta, strict_scope=strict_scope)
    except (OperationalError, ProgrammingError):
        marks = {}
        ta = None
    co_splits = _extract_review_co_splits(subject, 'review2', ['co3', 'co4'], ta=ta)
    return Response({'subject': {'code': subject.code, 'name': subject.name}, 'marks': marks, 'co_splits': co_splits})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def review2_publish(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err
    subject = _get_subject(subject_id, request)
    pub_auth = _require_publish_owner(request)
    if pub_auth:
        return pub_auth

    gate = _enforce_publish_window(request, subject.code, 'review2')
    if gate is not None:
        return gate

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_mark_entry_not_blocked(
        request,
        subject_code=subject.code,
        subject_name=subject.name,
        assessment='review2',
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate

    body = request.data or {}
    data = body.get('data', None)
    if data is None or not isinstance(data, dict):
        return Response({'detail': 'Invalid payload.'}, status=status.HTTP_400_BAD_REQUEST)

    rows = data.get('rows', [])
    if not isinstance(rows, list):
        return Response({'detail': 'Invalid rows.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import Review2Mark, ProjectMark

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
    is_project_course = _resolve_teaching_assignment_class_type(ta) == 'PROJECT'

    errors: list[str] = []
    with transaction.atomic():
        for item in rows:
            try:
                sid = int(item.get('studentId'))
            except Exception:
                continue
            student = StudentProfile.objects.filter(id=sid).first()
            if not student:
                continue

            mark = _coerce_decimal_or_none(item.get('total'))
            if item.get('total') not in (None, '',) and mark is None:
                errors.append(f'Invalid mark for student {sid}: {item.get("total")}')
                continue

            if mark is None:
                _delete_scoped_mark(Review2Mark, subject=subject, student=student, teaching_assignment=ta)
                if is_project_course:
                    _delete_scoped_mark(ProjectMark, subject=subject, student=student, teaching_assignment=ta)
            else:
                _upsert_scoped_mark(
                    Review2Mark,
                    subject=subject,
                    student=student,
                    teaching_assignment=ta,
                    mark_defaults={'mark': mark},
                )
                if is_project_course:
                    _upsert_scoped_mark(
                        ProjectMark,
                        subject=subject,
                        student=student,
                        teaching_assignment=ta,
                        mark_defaults={'mark': mark},
                    )

    if errors:
        return Response({'detail': 'Validation error.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    try:
        _touch_lock_after_publish(
            request,
            subject_code=subject.code,
            subject_name=subject.name,
            assessment='review2',
            teaching_assignment_id=ta_id,
        )
    except OperationalError:
        pass

    return Response({'status': 'published'})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def formative1_published(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject.code, assessment='formative1', teaching_assignment_id=ta_id)
    if gate is not None:
        return gate
    from .models import Formative1Mark
    try:
        ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
        strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
        marks = _safe_formative_marks_map(Formative1Mark, subject=subject, ta=ta, strict_scope=strict_scope)
    except (OperationalError, ProgrammingError):
        marks = {}
    return Response({'subject': {'code': subject.code, 'name': subject.name}, 'marks': marks})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def formative1_publish(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err
    subject = _get_subject(subject_id, request)
    pub_auth = _require_publish_owner(request)
    if pub_auth:
        return pub_auth

    gate = _enforce_publish_window(request, subject.code, 'formative1')
    if gate is not None:
        return gate

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_mark_entry_not_blocked(
        request,
        subject_code=subject.code,
        subject_name=subject.name,
        assessment='formative1',
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate

    body = request.data or {}
    data = body.get('data', None)
    if data is None or not isinstance(data, dict):
        return Response({'detail': 'Invalid payload.'}, status=status.HTTP_400_BAD_REQUEST)

    rows_by = data.get('rowsByStudentId', {})
    if not isinstance(rows_by, dict):
        return Response({'detail': 'Invalid rowsByStudentId.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import Formative1Mark

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)

    errors: list[str] = []
    with transaction.atomic():
        for sid_str, item in rows_by.items():
            try:
                sid = int(sid_str)
            except Exception:
                continue
            student = StudentProfile.objects.filter(id=sid).first()
            if not student:
                continue

            skill1 = _coerce_decimal_or_none((item or {}).get('skill1'))
            skill2 = _coerce_decimal_or_none((item or {}).get('skill2'))
            att1 = _coerce_decimal_or_none((item or {}).get('att1'))
            att2 = _coerce_decimal_or_none((item or {}).get('att2'))

            # If any is missing, treat as no published mark.
            if skill1 is None or skill2 is None or att1 is None or att2 is None:
                _delete_scoped_mark(Formative1Mark, subject=subject, student=student, teaching_assignment=ta)
                continue

            total = skill1 + skill2 + att1 + att2

            _upsert_scoped_mark(
                Formative1Mark,
                subject=subject,
                student=student,
                teaching_assignment=ta,
                mark_defaults={'skill1': skill1, 'skill2': skill2, 'att1': att1, 'att2': att2, 'total': total},
            )

    if errors:
        return Response({'detail': 'Validation error.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    try:
        _touch_lock_after_publish(
            request,
            subject_code=subject.code,
            subject_name=subject.name,
            assessment='formative1',
            teaching_assignment_id=ta_id,
        )
    except OperationalError:
        pass
    return Response({'status': 'published'})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def formative2_published(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject.code, assessment='formative2', teaching_assignment_id=ta_id)
    if gate is not None:
        return gate
    from .models import Formative2Mark
    try:
        ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
        strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
        marks = _safe_formative_marks_map(Formative2Mark, subject=subject, ta=ta, strict_scope=strict_scope)
    except (OperationalError, ProgrammingError):
        marks = {}
    return Response({'subject': {'code': subject.code, 'name': subject.name}, 'marks': marks})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def formative2_publish(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err
    subject = _get_subject(subject_id, request)
    pub_auth = _require_publish_owner(request)
    if pub_auth:
        return pub_auth

    gate = _enforce_publish_window(request, subject.code, 'formative2')
    if gate is not None:
        return gate

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_mark_entry_not_blocked(
        request,
        subject_code=subject.code,
        subject_name=subject.name,
        assessment='formative2',
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate

    body = request.data or {}
    data = body.get('data', None)
    if data is None or not isinstance(data, dict):
        return Response({'detail': 'Invalid payload.'}, status=status.HTTP_400_BAD_REQUEST)

    rows_by = data.get('rowsByStudentId', {})
    if not isinstance(rows_by, dict):
        return Response({'detail': 'Invalid rowsByStudentId.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import Formative2Mark

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)

    errors: list[str] = []
    with transaction.atomic():
        for sid_str, item in rows_by.items():
            try:
                sid = int(sid_str)
            except Exception:
                continue
            student = StudentProfile.objects.filter(id=sid).first()
            if not student:
                continue

            skill1 = _coerce_decimal_or_none((item or {}).get('skill1'))
            skill2 = _coerce_decimal_or_none((item or {}).get('skill2'))
            att1 = _coerce_decimal_or_none((item or {}).get('att1'))
            att2 = _coerce_decimal_or_none((item or {}).get('att2'))

            # If any is missing, treat as no published mark.
            if skill1 is None or skill2 is None or att1 is None or att2 is None:
                _delete_scoped_mark(Formative2Mark, subject=subject, student=student, teaching_assignment=ta)
                continue

            total = skill1 + skill2 + att1 + att2

            _upsert_scoped_mark(
                Formative2Mark,
                subject=subject,
                student=student,
                teaching_assignment=ta,
                mark_defaults={'skill1': skill1, 'skill2': skill2, 'att1': att1, 'att2': att2, 'total': total},
            )

    if errors:
        return Response({'detail': 'Validation error.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    try:
        _touch_lock_after_publish(
            request,
            subject_code=subject.code,
            subject_name=subject.name,
            assessment='formative2',
            teaching_assignment_id=ta_id,
        )
    except OperationalError:
        pass
    return Response({'status': 'published'})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cia1_published_sheet(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    pub_auth = _require_publish_owner(request)
    if pub_auth:
        return pub_auth

    subject = _get_subject(subject_id, request)


    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject.code, assessment='cia1', teaching_assignment_id=ta_id)
    if gate is not None:
        return gate
    from .models import Cia1PublishedSheet
    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
    strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
    row = _get_scoped_obe_json_row(
        Cia1PublishedSheet,
        subject=subject,
        teaching_assignment=ta,
        strict_scope=strict_scope,
    )
    return Response({'subject': {'code': subject.code, 'name': subject.name}, 'data': row.data if row else None})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cia1_publish_sheet(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)

    gate = _enforce_publish_window(request, subject.code, 'cia1')
    if gate is not None:
        return gate

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_mark_entry_not_blocked(
        request,
        subject_code=subject.code,
        subject_name=subject.name,
        assessment='cia1',
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate
    body = request.data or {}
    data = body.get('data', None)
    if data is None or not isinstance(data, dict):
        return Response({'detail': 'Invalid payload.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import Cia1PublishedSheet, Cia1Mark

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)

    _upsert_scoped_obe_json_row(
        Cia1PublishedSheet,
        subject=subject,
        teaching_assignment=ta,
        defaults={'data': data, 'updated_by': getattr(request.user, 'id', None)},
    )

    # Also upsert totals into the existing CIA1 totals table.
    questions = data.get('questions', [])
    if not isinstance(questions, list):
        questions = []
    rows_by = data.get('rowsByStudentId', {})
    if not isinstance(rows_by, dict):
        rows_by = {}

    qkeys = [str(q.get('key')) for q in questions if isinstance(q, dict) and q.get('key')]

    with transaction.atomic():
        for sid_str, row in rows_by.items():
            try:
                sid = int(sid_str)
            except Exception:
                continue
            student = StudentProfile.objects.filter(id=sid).first()
            if not student:
                continue

            absent = bool((row or {}).get('absent'))
            if absent:
                total_dec = Decimal('0')
            else:
                q = (row or {}).get('q', {})
                if not isinstance(q, dict):
                    q = {}
                total = Decimal('0')
                for k in qkeys:
                    v = q.get(k)
                    dec = _coerce_decimal_or_none(v)
                    if dec is None:
                        continue
                    total += dec
                total_dec = total

            _upsert_scoped_mark(
                Cia1Mark,
                subject=subject,
                student=student,
                teaching_assignment=ta,
                mark_defaults={'mark': total_dec},
            )

    try:
        _touch_lock_after_publish(
            request,
            subject_code=subject.code,
            subject_name=subject.name,
            assessment='cia1',
            teaching_assignment_id=ta_id,
        )
    except OperationalError:
        pass

    return Response({'status': 'published'})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def model_published_sheet(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    pub_auth = _require_publish_owner(request)
    if pub_auth:
        return pub_auth

    subject = _get_subject(subject_id, request)

    ta_id = _get_teaching_assignment_id_from_request(request)
    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
    strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
    # No special gating here; mirror CIA1 behaviour
    from .models import ModelPublishedSheet
    row = _get_scoped_obe_json_row(
        ModelPublishedSheet,
        subject=subject,
        teaching_assignment=ta,
        strict_scope=strict_scope,
    )
    return Response({'subject': {'code': subject.code, 'name': subject.name}, 'data': row.data if row else None})


def _resolve_model_student_id_from_row(*, row_key, row_payload):
    sid_raw = (row_payload or {}).get('studentId') if isinstance(row_payload, dict) else None
    try:
        sid = int(sid_raw)
        if sid > 0:
            return sid
    except Exception:
        pass

    key = str(row_key or '').strip()
    if key.isdigit():
        try:
            sid = int(key)
            return sid if sid > 0 else None
        except Exception:
            return None

    if key.startswith('id:'):
        try:
            sid = int(key.split(':', 1)[1])
            return sid if sid > 0 else None
        except Exception:
            return None

    if key.startswith('reg:'):
        reg = str(key.split(':', 1)[1]).strip()
        if not reg:
            return None
        student = StudentProfile.objects.filter(reg_no=reg).only('id').first()
        return int(student.id) if student else None

    return None


def _extract_model_totals_from_payload(data: dict):
    if not isinstance(data, dict):
        return {}

    class_type = str(data.get('classType') or '').strip().upper()
    if class_type in {'TCPL', 'TCPR'}:
        is_tcpl_like = True
    elif class_type in {'THEORY', 'LAB', 'PRACTICAL'}:
        is_tcpl_like = False
    else:
        raw_tcpl = data.get('tcplLikeKind')
        if isinstance(raw_tcpl, bool):
            is_tcpl_like = raw_tcpl
        else:
            tcpl_key = str(raw_tcpl or '').strip().upper()
            is_tcpl_like = tcpl_key in {'1', 'TRUE', 'YES', 'TCPL', 'TCPR'}

    primary_key = 'tcplSheet' if is_tcpl_like else 'theorySheet'
    fallback_key = 'theorySheet' if is_tcpl_like else 'tcplSheet'

    rows_by = data.get(primary_key, {})
    if not isinstance(rows_by, dict) or not rows_by:
        alt = data.get(fallback_key, {})
        rows_by = alt if isinstance(alt, dict) else {}

    totals_by_sid = {}
    for row_key, row in rows_by.items():
        if not isinstance(row, dict):
            continue

        sid = _resolve_model_student_id_from_row(row_key=row_key, row_payload=row)
        if not sid:
            continue

        absent = bool(row.get('absent'))
        absent_kind = str(row.get('absentKind') or 'AL').strip().upper()
        if absent and absent_kind == 'AL':
            totals_by_sid[sid] = Decimal('0')
            continue

        total = Decimal('0')
        q_obj = row.get('q', {})
        if isinstance(q_obj, dict):
            for val in q_obj.values():
                dec = _coerce_decimal_or_none(val)
                if dec is not None:
                    total += dec

        lab_dec = _coerce_decimal_or_none(row.get('lab'))
        if lab_dec is not None:
            total += lab_dec

        totals_by_sid[sid] = total

    return totals_by_sid


def _extract_project_totals_from_lab_payload(data: dict) -> dict[int, Decimal]:
    if not isinstance(data, dict):
        return {}

    payload = data.get('sheet') if isinstance(data.get('sheet'), dict) else data
    if not isinstance(payload, dict):
        return {}

    rows_by = payload.get('rowsByStudentId', {})
    if not isinstance(rows_by, dict):
        return {}

    totals_by_sid: dict[int, Decimal] = {}
    for sid_key, row in rows_by.items():
        if not isinstance(row, dict):
            continue

        sid = None
        try:
            sid = int(sid_key)
        except Exception:
            sid_raw = row.get('studentId')
            try:
                sid = int(sid_raw)
            except Exception:
                sid = None
        if not sid:
            continue

        if bool(row.get('absent')):
            totals_by_sid[sid] = Decimal('0')
            continue

        mark_total = None

        review_component_marks = row.get('reviewComponentMarks', {})
        if isinstance(review_component_marks, dict) and review_component_marks:
            total = Decimal('0')
            has_value = False
            for value in review_component_marks.values():
                dec = _coerce_decimal_or_none(value)
                if dec is None:
                    continue
                total += dec
                has_value = True
            if has_value:
                mark_total = total

        if mark_total is None:
            caa_by_co = row.get('caaExamByCo', {})
            if isinstance(caa_by_co, dict) and caa_by_co:
                total = Decimal('0')
                has_value = False
                for value in caa_by_co.values():
                    dec = _coerce_decimal_or_none(value)
                    if dec is None:
                        continue
                    total += dec
                    has_value = True
                if has_value:
                    mark_total = total

        if mark_total is None:
            mark_total = _coerce_decimal_or_none(row.get('ciaExam'))

        if mark_total is None:
            mark_total = _coerce_decimal_or_none(row.get('total'))

        if mark_total is None:
            continue

        totals_by_sid[sid] = mark_total

    return totals_by_sid


def _sync_project_marks_from_totals(*, subject, teaching_assignment, totals_by_sid: dict[int, Decimal]) -> int:
    from .models import ProjectMark

    if not isinstance(totals_by_sid, dict) or not totals_by_sid:
        return 0

    updated = 0
    with transaction.atomic():
        for sid, mark in totals_by_sid.items():
            try:
                sid_int = int(sid)
            except Exception:
                continue
            student = StudentProfile.objects.filter(id=sid_int).first()
            if not student:
                continue
            _upsert_scoped_mark(
                ProjectMark,
                subject=subject,
                student=student,
                teaching_assignment=teaching_assignment,
                mark_defaults={'mark': mark},
            )
            updated += 1

    return updated


def _backfill_project_marks_from_lab_published(*, subject, teaching_assignment=None, strict_scope: bool = False) -> int:
    from .models import LabPublishedSheet

    candidates = []
    for assessment in ('review1', 'review2'):
        row = _get_scoped_obe_json_row(
            LabPublishedSheet,
            subject=subject,
            teaching_assignment=teaching_assignment,
            strict_scope=strict_scope,
            assessment=assessment,
        )
        if row is not None:
            candidates.append(row)

    def _sort_key(item):
        updated_at = getattr(item, 'updated_at', None)
        return (updated_at or timezone.make_aware(datetime.min), int(getattr(item, 'pk', 0) or 0))

    candidates.sort(key=_sort_key, reverse=True)

    for row in candidates:
        data = row.data if isinstance(getattr(row, 'data', None), dict) else None
        if not isinstance(data, dict):
            continue
        totals_by_sid = _extract_project_totals_from_lab_payload(data)
        if not totals_by_sid:
            continue
        return _sync_project_marks_from_totals(
            subject=subject,
            teaching_assignment=teaching_assignment,
            totals_by_sid=totals_by_sid,
        )

    return 0


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def model_publish_sheet(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)

    gate = _enforce_publish_window(request, subject.code, 'model')
    if gate is not None:
        return gate

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_mark_entry_not_blocked(
        request,
        subject_code=subject.code,
        subject_name=subject.name,
        assessment='model',
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate
    body = request.data or {}
    data = body.get('data', None)
    if data is None or not isinstance(data, dict):
        return Response({'detail': 'Invalid payload.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import ModelPublishedSheet, ModelExamMark, ModelExamCOMark

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)

    _upsert_scoped_obe_json_row(
        ModelPublishedSheet,
        subject=subject,
        teaching_assignment=ta,
        defaults={'data': data, 'updated_by': getattr(request.user, 'id', None)},
    )
    
    co_marks_array = body.get('coMarks', data.get('coMarks', []))
    totals_by_sid = {}
    co_breakdown_by_sid = {}
    delete_sid = set()

    if isinstance(co_marks_array, list):
        for student_item in co_marks_array:
            if not isinstance(student_item, dict):
                continue
            try:
                sid = int(student_item.get('studentId'))
            except (ValueError, TypeError):
                continue
            if sid <= 0:
                continue

            total_dec = _coerce_decimal_or_none(student_item.get('total'))
            if total_dec is None:
                delete_sid.add(sid)
                continue

            totals_by_sid[sid] = total_dec
            co_payload = student_item.get('coBreakdown')
            if isinstance(co_payload, dict):
                co_breakdown_by_sid[sid] = co_payload

    if not totals_by_sid:
        totals_by_sid = _extract_model_totals_from_payload(data)

    with transaction.atomic():
        for sid in delete_sid:
            student = StudentProfile.objects.filter(id=sid).first()
            if not student:
                continue
            ModelExamMark.objects.filter(subject=subject, student=student, teaching_assignment=ta).delete()

        for sid, total_dec in totals_by_sid.items():
            student = StudentProfile.objects.filter(id=sid).first()
            if not student:
                continue

            mark_parent, _ = ModelExamMark.objects.update_or_create(
                subject=subject,
                student=student,
                teaching_assignment=ta,
                defaults={'total_mark': total_dec},
            )

            co_payload = co_breakdown_by_sid.get(sid)
            ModelExamCOMark.objects.filter(model_exam_mark=mark_parent).delete()
            if not isinstance(co_payload, dict):
                continue

            for c_k, c_data in sorted(co_payload.items(), key=lambda kv: str(kv[0])):
                c_num_str = str(c_k).replace('co', '').strip()
                try:
                    c_num = int(c_num_str)
                except Exception:
                    continue
                if c_num <= 0:
                    continue

                c_obj = c_data if isinstance(c_data, dict) else {}
                c_val = _coerce_decimal_or_none(c_obj.get('mark'))
                c_pct = _coerce_decimal_or_none(c_obj.get('percentage'))

                ModelExamCOMark.objects.create(
                    model_exam_mark=mark_parent,
                    co_num=c_num,
                    mark=c_val,
                    percentage=c_pct,
                )

    try:
        _touch_lock_after_publish(
            request,
            subject_code=subject.code,
            subject_name=subject.name,
            assessment='model',
            teaching_assignment_id=ta_id,
        )
    except OperationalError:
        pass

    try:
        recompute_final_internal_marks(
            actor_user_id=getattr(getattr(request, 'user', None), 'id', None),
            filters={
                'subject_code': subject.code,
                'teaching_assignment_id': ta_id,
            },
        )
    except Exception:
        logger.exception('model_publish_sheet: recompute_final_internal_marks failed for subject=%s ta=%s', subject.code, ta_id)

    return Response({'status': 'published'})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cia2_published_sheet(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    pub_auth = _require_publish_owner(request)
    if pub_auth:
        return pub_auth

    subject = _get_subject(subject_id, request)


    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject.code, assessment='cia2', teaching_assignment_id=ta_id)
    if gate is not None:
        return gate
    from .models import Cia2PublishedSheet
    try:
        ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
        strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
        row = _get_scoped_obe_json_row(
            Cia2PublishedSheet,
            subject=subject,
            teaching_assignment=ta,
            strict_scope=strict_scope,
        )
        data = row.data if row else None
    except OperationalError:
        data = None
    return Response({'subject': {'code': subject.code, 'name': subject.name}, 'data': data})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cia2_publish_sheet(request, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    subject = _get_subject(subject_id, request)

    gate = _enforce_publish_window(request, subject.code, 'cia2')
    if gate is not None:
        return gate

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_mark_entry_not_blocked(
        request,
        subject_code=subject.code,
        subject_name=subject.name,
        assessment='cia2',
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate
    body = request.data or {}
    data = body.get('data', None)
    if data is None or not isinstance(data, dict):
        return Response({'detail': 'Invalid payload.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import Cia2PublishedSheet, Cia2Mark

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)

    _upsert_scoped_obe_json_row(
        Cia2PublishedSheet,
        subject=subject,
        teaching_assignment=ta,
        defaults={'data': data, 'updated_by': getattr(request.user, 'id', None)},
    )

    # Also upsert totals into the existing CIA2 totals table.
    questions = data.get('questions', [])
    if not isinstance(questions, list):
        questions = []
    rows_by = data.get('rowsByStudentId', {})
    if not isinstance(rows_by, dict):
        rows_by = {}

    qkeys = [str(q.get('key')) for q in questions if isinstance(q, dict) and q.get('key')]

    with transaction.atomic():
        for sid_str, row in rows_by.items():
            try:
                sid = int(sid_str)
            except Exception:
                continue
            student = StudentProfile.objects.filter(id=sid).first()
            if not student:
                continue

            absent = bool((row or {}).get('absent'))
            if absent:
                total_dec = Decimal('0')
            else:
                q = (row or {}).get('q', {})
                if not isinstance(q, dict):
                    q = {}
                total = Decimal('0')
                for k in qkeys:
                    v = q.get(k)
                    dec = _coerce_decimal_or_none(v)
                    if dec is None:
                        continue
                    total += dec
                total_dec = total

            _upsert_scoped_mark(
                Cia2Mark,
                subject=subject,
                student=student,
                teaching_assignment=ta,
                mark_defaults={'mark': total_dec},
            )

    try:
        _touch_lock_after_publish(
            request,
            subject_code=subject.code,
            subject_name=subject.name,
            assessment='cia2',
            teaching_assignment_id=ta_id,
        )
    except OperationalError:
        pass

    return Response({'status': 'published'})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def lab_published_sheet(request, assessment: str, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    pub_auth = _require_publish_owner(request)
    if pub_auth:
        return pub_auth

    if assessment not in ('cia1', 'cia2', 'model', 'formative1', 'formative2', 'review1', 'review2'):
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    subject = _get_subject(subject_id, request)
    from .models import LabPublishedSheet
    try:
        ta_id = _get_teaching_assignment_id_from_request(request)
        ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)
        strict_scope = _strict_assignment_scope(subject_code=subject.code, teaching_assignment=ta)
        row = _get_scoped_obe_json_row(
            LabPublishedSheet,
            subject=subject,
            teaching_assignment=ta,
            strict_scope=strict_scope,
            assessment=assessment,
        )
        data = row.data if row else None
    except OperationalError:
        data = None
    return Response({'subject': {'code': subject.code, 'name': subject.name}, 'assessment': assessment, 'data': data})
    
@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def obe_progress_overview(request):
    """Aggregated OBE progress overview for HOD and Advisor roles.

    Response shape (simplified):

    {
      "role": "HOD" | "ADVISOR" | "FACULTY",
      "academic_year": { id, name } | null,
      "department": { id, code, name, short_name } | null,
      "sections": [
        {
          "id": number,
          "name": string,
          "batch": { id, name },
          "course": { id, name, department: { ... } },
          "staff": [
            {
              "id": number,        # StaffProfile id
              "name": string,
              "user_id": number | null,
              "teaching_assignments": [
                {
                  "id": number,
                  "subject_code": string,
                  "subject_name": string,
                  "enabled_assessments": string[],
                  "exam_progress": [
                    {
                      "assessment": string,   # e.g. ssa1, cia1
                      "rows_filled": number,
                      "total_students": number,
                      "percentage": number,
                      "published": boolean,
                    },
                    ...
                  ],
                }
              ],
            }
          ],
        }
      ],
    }

    HODs see all sections in their department for the active AcademicYear;
    Advisors see only their advised sections.

    Access is restricted to HOD/AHOD/Advisor users (superusers allowed).
    """

    from academics.models import AcademicYear, Department, DepartmentRole, SectionAdvisor, TeachingAssignment, Subject
    from .models import (
        ObeMarkTableLock,
        Ssa1Mark,
        Ssa2Mark,
        Review1Mark,
        Review2Mark,
        Formative1Mark,
        Formative2Mark,
        Cia1Mark,
        Cia2Mark,
        ModelPublishedSheet,
    )

    user = getattr(request, 'user', None)
    if not user or not getattr(user, 'is_authenticated', False):
        return Response({'detail': 'Authentication required.'}, status=status.HTTP_401_UNAUTHORIZED)

    # Resolve active academic year (fallback to latest).
    # Use latest active year deterministically; some DBs may have multiple active rows.
    active_ay = AcademicYear.objects.filter(is_active=True).order_by('-id').first()
    ay = active_ay or AcademicYear.objects.order_by('-id').first()

    # Helper: normalize role names for this user
    role_names: set[str] = set()
    try:
        role_names = {str(r.name or '').strip().upper() for r in user.roles.all()}
    except Exception:
        try:
            # Fallback via Role model m2m if available
            from accounts.models import Role
            role_qs = Role.objects.filter(userrole__user=user)
            role_names = {str(r.name or '').strip().upper() for r in role_qs}
        except Exception:
            role_names = set()

    has_hod = 'HOD' in role_names or 'AHOD' in role_names
    is_iqac = 'IQAC' in role_names
    is_iqac_main = False
    try:
        is_iqac_main = bool(is_iqac and str(getattr(user, 'username', '') or '').strip() == '000000')
    except Exception:
        is_iqac_main = False

    # Advisors are primarily represented by academics.SectionAdvisor mappings.
    # Some deployments may not assign an explicit 'ADVISOR' role, so detect via DB too.
    staff_profile = None
    try:
        staff_profile = getattr(user, 'staff_profile', None)
    except Exception:
        staff_profile = None

    advisor_rows = None
    advisor_year = None
    try:
        advisor_rows_base = SectionAdvisor.objects.filter(is_active=True)
        if staff_profile is not None:
            advisor_rows_base = advisor_rows_base.filter(advisor=staff_profile)
        else:
            # Fallback for inconsistent data: try through user relation
            advisor_rows_base = advisor_rows_base.filter(advisor__user=user)

        # Prefer SectionAdvisor mappings for the current active year (most common).
        if active_ay is not None and advisor_rows_base.filter(academic_year=active_ay).exists():
            advisor_year = active_ay
        else:
            # Otherwise, prefer any mapping where the academic year is marked active.
            active_rows = advisor_rows_base.filter(academic_year__is_active=True).select_related('academic_year')
            if active_rows.exists():
                advisor_year = active_rows.order_by('-academic_year_id').first().academic_year
            else:
                # Last resort: use the most recent mapping's academic year.
                last_row = advisor_rows_base.select_related('academic_year').order_by('-academic_year_id').first()
                advisor_year = last_row.academic_year if last_row else None

        advisor_rows = advisor_rows_base.filter(academic_year=advisor_year) if advisor_year is not None else advisor_rows_base.none()
    except Exception:
        advisor_rows = None
        advisor_year = None

    has_advisor = 'ADVISOR' in role_names
    if not has_advisor and advisor_rows is not None:
        try:
            has_advisor = bool(advisor_rows.exists())
        except Exception:
            has_advisor = False

    # Enforce access control: HOD/AHOD/Advisor can view this aggregated progress.
    # Additionally, allow the *main IQAC account* (username 000000) to view all departments/sections.
    if not getattr(user, 'is_superuser', False) and not has_hod and not has_advisor and not is_iqac_main:
        return Response({'detail': 'Progress view is only available for HOD/Advisor.'}, status=status.HTTP_403_FORBIDDEN)

    # Determine primary context: IQAC_MAIN > HOD > ADVISOR > FACULTY
    primary_role = 'FACULTY'
    department = None
    sections_qs = None

    # IQAC main account: can view progress across departments/sections.
    departments_for_iqac = None
    selected_department_id = None
    if is_iqac_main and ay is not None:
        primary_role = 'IQAC'
        try:
            from academics.models import Section

            departments_for_iqac = list(Department.objects.all().order_by('short_name', 'code', 'name', 'id'))

            raw_dept = None
            try:
                raw_dept = request.query_params.get('department_id')
            except Exception:
                raw_dept = request.GET.get('department_id')

            raw_dept = str(raw_dept or '').strip()
            if raw_dept.lower() in ('all', '*', '0', 'none') or raw_dept == '':
                selected_department_id = None
                sections_qs = (
                    Section.objects.select_related('batch', 'batch__course', 'batch__course__department', 'semester')
                    .all()
                )
                department = None
            else:
                try:
                    selected_department_id = int(raw_dept)
                except Exception:
                    selected_department_id = None

                if selected_department_id is not None:
                    department = Department.objects.filter(id=selected_department_id).first()
                else:
                    department = None

                # If an invalid department id is provided, return no sections rather than a potentially huge dataset.
                if department is None:
                    sections_qs = Section.objects.none()
                else:
                    sections_qs = (
                        Section.objects.select_related('batch', 'batch__course', 'batch__course__department')
                        .filter(batch__course__department=department)
                    )
        except Exception:
            sections_qs = None

    if has_hod and ay is not None:
        # Find department from DepartmentRole for this academic year
        dept_role = DepartmentRole.objects.filter(staff__user=user, academic_year=ay, is_active=True).select_related('department').first()
        department = getattr(dept_role, 'department', None)
        if department is not None:
            primary_role = 'HOD'
            # All sections whose batch.course.department matches HOD department
            from academics.models import Section

            sections_qs = (
                Section.objects.select_related('batch', 'batch__course', 'batch__course__department', 'semester')
                .filter(batch__course__department=department)
            )

    if sections_qs is None and has_advisor:
        # Advisor sees only their advised sections for the year
        primary_role = 'ADVISOR'
        # Use the advisor mapping year (can differ from global ay if AcademicYear flags are inconsistent)
        if advisor_year is not None:
            ay = advisor_year
        if advisor_rows is None:
            # Very defensive fallback
            if staff_profile is not None:
                advisor_rows = SectionAdvisor.objects.filter(advisor=staff_profile, is_active=True)
            else:
                advisor_rows = SectionAdvisor.objects.filter(advisor__user=user, is_active=True)
            if ay is not None:
                advisor_rows = advisor_rows.filter(academic_year=ay)
            else:
                advisor_rows = advisor_rows.filter(academic_year__is_active=True)

        advisor_rows = advisor_rows.select_related('section', 'section__batch', 'section__batch__course', 'section__batch__course__department', 'academic_year')
        from academics.models import Section

        section_ids = list(advisor_rows.values_list('section_id', flat=True))
        sections_qs = Section.objects.filter(id__in=section_ids).select_related('batch', 'batch__course', 'batch__course__department', 'semester')
        # Derive department from the first section if not already set
        if department is None:
            first_sec = advisor_rows.first()
            if first_sec and getattr(first_sec.section.batch.course, 'department', None) is not None:
                department = first_sec.section.batch.course.department

    if sections_qs is None:
        # Fallback: treat as regular faculty; show sections from their TeachingAssignments
        from academics.models import Section

        ta_qs = TeachingAssignment.objects.select_related('section', 'section__batch', 'section__batch__course', 'section__batch__course__department').filter(
            staff__user=user,
            is_active=True,
        )
        if ay is not None:
            ta_qs = ta_qs.filter(academic_year=ay)
        section_ids = {ta.section_id for ta in ta_qs if ta.section_id}
        sections_qs = Section.objects.filter(id__in=section_ids).select_related('batch', 'batch__course', 'batch__course__department', 'semester')
        if department is None:
            first_ta = ta_qs.first()
            if first_ta and getattr(first_ta.section.batch.course, 'department', None) is not None:
                department = first_ta.section.batch.course.department

    sections = list(sections_qs or [])

    def _normalize_class_type(val) -> str:
        return _normalize_obe_class_type(val)

    def _resolve_ta_class_type(ta) -> str:
        try:
            row = getattr(ta, 'curriculum_row', None)
            if row is not None:
                ct = getattr(row, 'class_type', None) or getattr(getattr(row, 'master', None), 'class_type', None)
                if ct:
                    return _normalize_class_type(ct)
        except Exception:
            pass
        try:
            es = getattr(ta, 'elective_subject', None)
            ct = getattr(es, 'class_type', None) if es is not None else None
            if ct:
                return _normalize_class_type(ct)
        except Exception:
            pass
        return 'THEORY'

    def _resolve_curriculum_row_for_ta(ta):
        row = getattr(ta, 'curriculum_row', None)
        if row is not None:
            return row
        # Legacy: TA may have only Subject. Try to resolve curriculum row.
        try:
            from curriculum.models import CurriculumDepartment

            dept = None
            try:
                dept = ta.section.batch.course.department
            except Exception:
                dept = None

            subj = getattr(ta, 'subject', None)
            code = getattr(subj, 'code', None)
            name = getattr(subj, 'name', None)

            qs = CurriculumDepartment.objects.all().select_related('master', 'department')
            if dept is not None:
                qs = qs.filter(department=dept)

            if code:
                code_s = str(code).strip()
                qs = qs.filter(Q(course_code__iexact=code_s) | Q(master__course_code__iexact=code_s))
            elif name:
                name_s = str(name).strip()
                qs = qs.filter(Q(course_name__iexact=name_s) | Q(master__course_name__iexact=name_s))
            else:
                return None

            return qs.order_by('-updated_at', '-id').first()
        except Exception:
            return None

    def _resolve_subject_meta_for_ta(ta):
        """Return (subject_code, subject_name, subject_obj_or_none)."""
        subj = getattr(ta, 'subject', None)
        if subj is not None:
            return getattr(subj, 'code', None), getattr(subj, 'name', None), subj

        code = None
        name = None

        try:
            es = getattr(ta, 'elective_subject', None)
            if es is not None:
                code = getattr(es, 'course_code', None)
                name = getattr(es, 'course_name', None)
        except Exception:
            pass

        if not code and not name:
            try:
                row = getattr(ta, 'curriculum_row', None)
                if row is not None:
                    code = getattr(row, 'course_code', None) or getattr(getattr(row, 'master', None), 'course_code', None)
                    name = getattr(row, 'course_name', None) or getattr(getattr(row, 'master', None), 'course_name', None)
            except Exception:
                pass

        if not code and not name:
            try:
                cs = getattr(ta, 'custom_subject', None)
                if cs:
                    code = str(cs).strip()
                    try:
                        name = ta.get_custom_subject_display()
                    except Exception:
                        name = code
            except Exception:
                pass

        code = str(code).strip() if code else None
        name = str(name).strip() if name else None

        subj_obj = None
        if code:
            try:
                subj_obj = Subject.objects.filter(code=code).first()
            except Exception:
                subj_obj = None

        # Best effort name fallback from Subject record
        if not name and subj_obj is not None:
            name = getattr(subj_obj, 'name', None)

        return code, name, subj_obj

    def _clean_special_enabled(vals):
        allowed = ['ssa1', 'ssa2', 'formative1', 'formative2', 'cia1', 'cia2']
        out = []
        for v in (vals or []):
            s = str(v or '').strip().lower()
            if s and s in allowed and s not in out:
                out.append(s)
        # Keep stable UI order
        ordered = [k for k in allowed if k in out]
        # preserve any additional allowed keys (shouldn't happen)
        for k in out:
            if k not in ordered:
                ordered.append(k)
        return ordered

    def _resolve_special_enabled_assessments(ta) -> list[str]:
        """Mirrors academics.views.TeachingAssignmentViewSet.enabled_assessments (GET) for SPECIAL courses."""
        row = _resolve_curriculum_row_for_ta(ta)
        master_id = getattr(row, 'master_id', None) if row is not None else None

        enabled = None
        try:
            from academics.models import SpecialCourseAssessmentSelection

            if master_id is not None and getattr(ta, 'academic_year', None) is not None:
                sel = (
                    SpecialCourseAssessmentSelection.objects.filter(
                        curriculum_row__master_id=master_id,
                        academic_year=ta.academic_year,
                    )
                    .select_related('curriculum_row')
                    .order_by('id')
                    .first()
                )
                if sel is not None:
                    enabled = getattr(sel, 'enabled_assessments', None)
        except Exception:
            enabled = None

        if enabled is None and row is not None:
            enabled = getattr(row, 'enabled_assessments', None)
        if enabled is None:
            enabled = []

        return _clean_special_enabled(enabled)

    def _assessments_for_ta(ta) -> list[str]:
        """Determine which assessment keys should be shown for progress."""
        ct = _resolve_ta_class_type(ta)
        if ct == 'SPECIAL':
            return _resolve_special_enabled_assessments(ta)
        if ct == 'PRACTICAL':
            return ['cia1', 'cia2', 'model']
        if ct == 'PROJECT':
            return ['review1', 'review2', 'model']
        if ct == 'TCPR':
            return ['ssa1', 'review1', 'cia1', 'ssa2', 'review2', 'cia2', 'model']
        if ct == 'LAB':
            return ['cia1', 'cia2', 'model']
        # THEORY / TCPL / unknown: show the standard theory keys
        return ['ssa1', 'formative1', 'cia1', 'ssa2', 'formative2', 'cia2', 'model']

    def _assessment_label_for_progress(*, class_type: str, assessment_key: str) -> str | None:
        """Return a UI label for an assessment key based on class type.

        Keep this narrowly scoped: only cases where the label differs from the raw key.
        """
        ct = _normalize_class_type(class_type)
        key = str(assessment_key or '').strip().lower()
        if not key:
            return None

        # TCPL: formative keys are used but shown as LAB 1 / LAB 2
        if ct == 'TCPL':
            if key == 'formative1':
                return 'LAB 1'
            if key == 'formative2':
                return 'LAB 2'

        # LAB: assessment names should explicitly say LAB
        if ct == 'LAB':
            if key == 'cia1':
                return 'CIA 1 LAB'
            if key == 'cia2':
                return 'CIA 2 LAB'
            if key == 'model':
                return 'MODEL LAB'

        return None

    # Utility: compute exam progress for a teaching assignment and assessment key
    def _exam_progress_for_ta(ta, *, assessment_key: str, subject_obj, students, class_type: str) -> dict:
        total_students = len(students)

        subject = subject_obj

        rows_filled = 0

        if subject is not None and total_students > 0:
            student_ids = [s.id for s in students]

            if assessment_key == 'ssa1':
                rows_filled = _filter_marks_queryset_for_teaching_assignment(Ssa1Mark.objects.filter(subject=subject, student_id__in=student_ids, mark__isnull=False), ta, strict_scope=_strict_assignment_scope(subject_code=getattr(subject, 'code', ''), teaching_assignment=ta)).count()
            elif assessment_key == 'ssa2':
                rows_filled = _filter_marks_queryset_for_teaching_assignment(Ssa2Mark.objects.filter(subject=subject, student_id__in=student_ids, mark__isnull=False), ta, strict_scope=_strict_assignment_scope(subject_code=getattr(subject, 'code', ''), teaching_assignment=ta)).count()
            elif assessment_key == 'review1':
                rows_filled = _filter_marks_queryset_for_teaching_assignment(Review1Mark.objects.filter(subject=subject, student_id__in=student_ids, mark__isnull=False), ta, strict_scope=_strict_assignment_scope(subject_code=getattr(subject, 'code', ''), teaching_assignment=ta)).count()
            elif assessment_key == 'review2':
                rows_filled = _filter_marks_queryset_for_teaching_assignment(Review2Mark.objects.filter(subject=subject, student_id__in=student_ids, mark__isnull=False), ta, strict_scope=_strict_assignment_scope(subject_code=getattr(subject, 'code', ''), teaching_assignment=ta)).count()
            elif assessment_key == 'formative1':
                rows_filled = _filter_marks_queryset_for_teaching_assignment(Formative1Mark.objects.filter(subject=subject, student_id__in=student_ids).exclude(total__isnull=True), ta, strict_scope=_strict_assignment_scope(subject_code=getattr(subject, 'code', ''), teaching_assignment=ta)).count()
            elif assessment_key == 'formative2':
                rows_filled = _filter_marks_queryset_for_teaching_assignment(Formative2Mark.objects.filter(subject=subject, student_id__in=student_ids).exclude(total__isnull=True), ta, strict_scope=_strict_assignment_scope(subject_code=getattr(subject, 'code', ''), teaching_assignment=ta)).count()
            elif assessment_key == 'cia1':
                rows_filled = _filter_marks_queryset_for_teaching_assignment(Cia1Mark.objects.filter(subject=subject, student_id__in=student_ids, mark__isnull=False), ta, strict_scope=_strict_assignment_scope(subject_code=getattr(subject, 'code', ''), teaching_assignment=ta)).count()
            elif assessment_key == 'cia2':
                rows_filled = _filter_marks_queryset_for_teaching_assignment(Cia2Mark.objects.filter(subject=subject, student_id__in=student_ids, mark__isnull=False), ta, strict_scope=_strict_assignment_scope(subject_code=getattr(subject, 'code', ''), teaching_assignment=ta)).count()
            elif assessment_key == 'model':
                # MODEL snapshot: treat presence of a published sheet as fully filled
                try:
                    has_sheet = _get_scoped_obe_json_row(
                        ModelPublishedSheet,
                        subject=subject,
                        teaching_assignment=ta,
                        strict_scope=_strict_assignment_scope(subject_code=getattr(subject, 'code', ''), teaching_assignment=ta),
                    ) is not None
                except Exception:
                    has_sheet = False
                rows_filled = total_students if has_sheet else 0

        percentage = float(rows_filled) / float(total_students) * 100.0 if total_students > 0 and rows_filled > 0 else 0.0

        # Published flag via ObeMarkTableLock (preferred), falling back to False
        published = False
        try:
            lock = ObeMarkTableLock.objects.filter(teaching_assignment=ta, assessment=str(assessment_key).lower()).first()
            published = bool(getattr(lock, 'is_published', False)) if lock is not None else False
        except Exception:
            published = False

        return {
            'assessment': assessment_key,
            'label': _assessment_label_for_progress(class_type=class_type, assessment_key=assessment_key),
            'rows_filled': int(rows_filled),
            'total_students': int(total_students),
            'percentage': round(percentage, 1),
            'published': published,
        }

    # Build response per section
    section_results = []

    for sec in sections:
        # Teaching assignments in this section for the academic year.
        # Note: electives are often stored as sectionless teaching assignments.
        # To avoid missing courses in section-wise progress, we also include
        # elective teaching assignments that apply to students in this section
        # (based on curriculum.ElectiveChoice mappings).
        ta_qs = TeachingAssignment.objects.select_related(
            'staff',
            'staff__user',
            'subject',
            'section',
            'academic_year',
            'curriculum_row',
            'curriculum_row__master',
            'elective_subject',
        ).filter(is_active=True)
        if ay is not None:
            ta_qs = ta_qs.filter(academic_year=ay)

        ta_filter = Q(section=sec)
        try:
            from curriculum.models import ElectiveChoice

            # Elective choices can be stored on StudentProfile.section or via StudentSectionAssignment.
            choice_qs = ElectiveChoice.objects.filter(is_active=True).exclude(elective_subject__isnull=True).filter(
                Q(student__section=sec)
                | Q(student__section_assignments__section=sec, student__section_assignments__end_date__isnull=True)
            )
            if ay is not None:
                # Some data uses ODD/EVEN parity-split AcademicYear rows. Prefer exact match,
                # but fall back to matching by AcademicYear.name so electives don't disappear.
                ay_name = str(getattr(ay, 'name', '') or '').strip()
                name_q = Q()
                if ay_name:
                    name_q = Q(academic_year__name=ay_name)
                # Some older data may have academic_year NULL; include it so we don't hide electives.
                choice_qs = choice_qs.filter(Q(academic_year=ay) | name_q | Q(academic_year__isnull=True))

            elective_subject_ids = list(choice_qs.values_list('elective_subject_id', flat=True).distinct())
            if elective_subject_ids:
                ta_filter |= Q(section__isnull=True, elective_subject_id__in=elective_subject_ids)
        except Exception:
            pass

        ta_qs = ta_qs.filter(ta_filter)

        tas = list(ta_qs)
        if not tas:
            # For Advisor view, still include sections even if no TAs are configured yet.
            if primary_role == 'ADVISOR':
                batch = getattr(sec, 'batch', None)
                course = getattr(batch, 'course', None) if batch is not None else None
                course_dept = getattr(course, 'department', None) if course is not None else None
                sem_obj = getattr(sec, 'semester', None)
                section_results.append(
                    {
                        'id': getattr(sec, 'id', None),
                        'name': getattr(sec, 'name', None),
                        'batch': {
                            'id': getattr(batch, 'id', None) if batch is not None else None,
                            'name': getattr(batch, 'name', None) if batch is not None else None,
                        },
                        'course': {
                            'id': getattr(course, 'id', None) if course is not None else None,
                            'name': getattr(course, 'name', None) if course is not None else None,
                        },
                        'department': {
                            'id': getattr(course_dept, 'id', None) if course_dept is not None else None,
                            'code': getattr(course_dept, 'code', None) if course_dept is not None else None,
                            'name': getattr(course_dept, 'name', None) if course_dept is not None else None,
                            'short_name': getattr(course_dept, 'short_name', None) if course_dept is not None else None,
                        },
                        'semester': getattr(sem_obj, 'number', None) if sem_obj is not None else None,
                        'staff': [],
                    }
                )
            continue

        # Group by staff
        staff_map: dict[int, dict] = {}

        for ta in tas:
            staff = getattr(ta, 'staff', None)
            if not staff:
                continue
            staff_id = getattr(staff, 'id', None)
            if staff_id is None:
                continue

            if staff_id not in staff_map:
                user_obj = getattr(staff, 'user', None)
                full_name = ''
                try:
                    full_name = str(getattr(user_obj, 'get_full_name', lambda: '')() or '').strip() if user_obj is not None else ''
                except Exception:
                    full_name = ''
                if not full_name:
                    try:
                        full_name = ' '.join(filter(None, [getattr(user_obj, 'first_name', ''), getattr(user_obj, 'last_name', '')])).strip() if user_obj is not None else ''
                    except Exception:
                        full_name = ''
                if not full_name:
                    try:
                        full_name = str(getattr(user_obj, 'username', '') or '').strip() if user_obj is not None else ''
                    except Exception:
                        full_name = ''

                staff_map[staff_id] = {
                    'id': staff_id,
                    'name': full_name or str(getattr(staff, 'staff_id', '') or '').strip() or str(getattr(staff, 'id', '') or '').strip(),
                    'user_id': getattr(user_obj, 'id', None),
                    'teaching_assignments': [],
                }

            # Resolve subject code/name (supports curriculum_row / elective_subject)
            subject_code, subject_name, subject_obj = _resolve_subject_meta_for_ta(ta)

            # Determine which exams should be shown for this TA
            class_type = _resolve_ta_class_type(ta)
            assessment_keys = _assessments_for_ta(ta)

            # Get roster once per TA
            try:
                students = _get_students_for_teaching_assignment(ta)
            except Exception:
                students = []

            exam_progress = []
            for assess in assessment_keys:
                if not assess:
                    continue
                exam_progress.append(
                    _exam_progress_for_ta(
                        ta,
                        assessment_key=assess,
                        subject_obj=subject_obj,
                        students=students,
                        class_type=class_type,
                    )
                )

            staff_map[staff_id]['teaching_assignments'].append(
                {
                    'id': getattr(ta, 'id', None),
                    'subject_code': subject_code,
                    'subject_name': subject_name,
                    'class_type': class_type,
                    # Kept for UI/debug: for SPECIAL this is the enabled list; for others it's the computed visible keys
                    'enabled_assessments': assessment_keys,
                    'exam_progress': exam_progress,
                }
            )

        # Skip if no staff entries
        if not staff_map:
            continue

        batch = getattr(sec, 'batch', None)
        course = getattr(batch, 'course', None) if batch is not None else None
        course_dept = getattr(course, 'department', None) if course is not None else None
        sem_obj = getattr(sec, 'semester', None)

        section_results.append(
            {
                'id': getattr(sec, 'id', None),
                'name': getattr(sec, 'name', None),
                'batch': {
                    'id': getattr(batch, 'id', None) if batch is not None else None,
                    'name': getattr(batch, 'name', None) if batch is not None else None,
                },
                'course': {
                    'id': getattr(course, 'id', None) if course is not None else None,
                    'name': getattr(course, 'name', None) if course is not None else None,
                },
                'department': {
                    'id': getattr(course_dept, 'id', None) if course_dept is not None else None,
                    'code': getattr(course_dept, 'code', None) if course_dept is not None else None,
                    'name': getattr(course_dept, 'name', None) if course_dept is not None else None,
                    'short_name': getattr(course_dept, 'short_name', None) if course_dept is not None else None,
                },
                'semester': getattr(sem_obj, 'number', None) if sem_obj is not None else None,
                'staff': list(staff_map.values()),
            }
        )

    resp = {
        'role': primary_role,
        'academic_year': {'id': getattr(ay, 'id', None), 'name': getattr(ay, 'name', None)} if ay is not None else None,
        'department': None,
        'sections': section_results,
    }

    if is_iqac_main:
        resp['selected_department_id'] = selected_department_id
        resp['departments'] = [
            {
                'id': getattr(d, 'id', None),
                'code': getattr(d, 'code', None),
                'name': getattr(d, 'name', None),
                'short_name': getattr(d, 'short_name', None),
            }
            for d in (departments_for_iqac or [])
        ]

    if department is not None:
        resp['department'] = {
            'id': getattr(department, 'id', None),
            'code': getattr(department, 'code', None),
            'name': getattr(department, 'name', None),
            'short_name': getattr(department, 'short_name', None),
        }

    return Response(resp)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def obe_progress_departments(request):
    """
    Lightweight endpoint for IQAC main progress view.
    Returns all departments with section_count and course_count for the current AY.
    """
    user = request.user
    try:
        from academics.models import AcademicYear, Department, RoleAssignment, TeachingAssignment, Section

        # Verify IQAC main
        role_names: set = set()
        try:
            rqs = RoleAssignment.objects.filter(staff__user=user, is_active=True).values_list('role__name', flat=True)
            role_names = {str(n or '').strip().upper() for n in rqs}
        except Exception:
            pass

        is_iqac = 'IQAC' in role_names
        is_iqac_main = is_iqac and str(getattr(user, 'username', '') or '').strip() == '000000'

        if not is_iqac_main and not getattr(user, 'is_superuser', False):
            return Response({'detail': 'IQAC main account required.'}, status=status.HTTP_403_FORBIDDEN)

        ay = AcademicYear.objects.filter(is_active=True).first()

        departments = list(Department.objects.all().order_by('short_name', 'code', 'name', 'id'))

        result = []
        for dept in departments:
            # Count sections in this department
            sec_qs = Section.objects.filter(batch__course__department=dept)
            section_count = sec_qs.count()

            # Count distinct subjects (TAs) for the dept in the current AY
            try:
                ta_qs = TeachingAssignment.objects.filter(section__batch__course__department=dept)
                if ay is not None:
                    ta_qs = ta_qs.filter(academic_year=ay)
                course_count = ta_qs.values('subject_id').distinct().count()
            except Exception:
                course_count = 0

            result.append({
                'id': dept.id,
                'code': getattr(dept, 'code', None),
                'name': dept.name,
                'short_name': getattr(dept, 'short_name', None),
                'section_count': section_count,
                'course_count': course_count,
            })

        return Response({
            'departments': result,
            'academic_year': {
                'id': getattr(ay, 'id', None),
                'name': getattr(ay, 'name', None),
            } if ay else None,
        })
    except Exception as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def lab_publish_sheet(request, assessment: str, subject_id: str):
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    assessment = str(assessment or '').lower().strip()
    if assessment not in ('cia1', 'cia2', 'model', 'formative1', 'formative2', 'review1', 'review2'):
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    subject = _get_subject(subject_id, request)

    gate = _enforce_publish_window(request, subject.code, assessment)
    if gate is not None:
        return gate

    ta_id = _get_teaching_assignment_id_from_request(request)
    gate = _enforce_mark_entry_not_blocked(
        request,
        subject_code=subject.code,
        subject_name=subject.name,
        assessment=str(assessment).lower(),
        teaching_assignment_id=ta_id,
    )
    if gate is not None:
        return gate

    body = request.data or {}
    data = body.get('data', None)
    if data is None or not isinstance(data, dict):
        return Response({'detail': 'Invalid payload.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import AssessmentDraft, LabPublishedSheet
    from .services.exam_mark_persistence import persist_lab_exam_marks

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject.code, teaching_assignment_id=ta_id)

    with transaction.atomic():
        _upsert_scoped_obe_json_row(
            LabPublishedSheet,
            subject=subject,
            teaching_assignment=ta,
            assessment=assessment,
            defaults={'data': data, 'updated_by': getattr(request.user, 'id', None)},
        )

        # Keep draft aligned with the just-published payload so refresh/reload
        # never falls back to stale or empty draft state.
        _upsert_scoped_obe_json_row(
            AssessmentDraft,
            subject=subject,
            teaching_assignment=ta,
            assessment=assessment,
            defaults={'data': data, 'updated_by': getattr(request.user, 'id', None)},
        )

        # Persist dedicated per-student rows for DB/reporting visibility.
        persist_lab_exam_marks(
            subject=subject,
            teaching_assignment=ta,
            assessment=assessment,
            data=data,
        )

    if assessment in ('review1', 'review2'):
        try:
            totals_by_sid = _extract_project_totals_from_lab_payload(data)
            if totals_by_sid:
                _sync_project_marks_from_totals(
                    subject=subject,
                    teaching_assignment=ta,
                    totals_by_sid=totals_by_sid,
                )
        except Exception:
            pass

    try:
        _touch_lock_after_publish(
            request,
            subject_code=subject.code,
            subject_name=subject.name,
            assessment=str(assessment).lower(),
            teaching_assignment_id=ta_id,
        )
    except OperationalError:
        pass

    return Response({'status': 'published'})


@api_view(['GET', 'PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cia1_marks(request, subject_id):
    """CIA1 marks API used by the React Faculty → OBE → Mark Entry → CIA1 screen.

    - GET: returns roster + current marks
    - PUT: upserts/clears marks
    """
    user = request.user
    staff_profile = getattr(user, 'staff_profile', None)
    role_names = {r.name.upper() for r in user.roles.all()}
    try:
        user_perms = {str(p or '').lower() for p in (get_user_permissions(user) or [])}
    except Exception:
        user_perms = set()

    # Academic Controller (IQAC / OBE Master) needs to view/edit any staff's roster.
    is_obe_master = ('obe.master.manage' in user_perms) or ('IQAC' in role_names) or getattr(user, 'is_superuser', False)
    if not staff_profile and not is_obe_master and not getattr(user, 'is_staff', False):
        return Response({'detail': 'Faculty access only.'}, status=status.HTTP_403_FORBIDDEN)

    # Subject may not exist when teaching assignments reference curriculum rows only.
    subject = Subject.objects.filter(code=subject_id).first()
    if subject is None:
        try:
            subject = _get_subject(subject_id, request)
        except Exception:
            subject = None

    try:
        tas = TeachingAssignment.objects.select_related('section', 'academic_year', 'curriculum_row').filter(is_active=True)
        if subject is not None:
            tas = tas.filter(Q(subject=subject) | Q(curriculum_row__course_code=subject.code))
        else:
            tas = tas.filter(Q(subject__code=subject_id) | Q(curriculum_row__course_code=subject_id))
    except (ValueError, FieldError) as e:
        return Response(
            {
                'detail': 'CIA1 teaching assignment query failed.',
                'error': str(e),
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    # Staff: only their teaching assignments; HOD/ADVISOR: within their department.
    # IQAC/OBE Master: do not restrict to staff-owned assignments (Academic Controller viewer).
    try:
        if is_obe_master:
            pass
        elif 'HOD' in role_names or 'ADVISOR' in role_names:
            if staff_profile and getattr(staff_profile, 'department_id', None):
                # Semester is canonical (no course FK). Department lives on Course -> Batch.
                tas = tas.filter(section__batch__course__department=staff_profile.department)
        else:
            if staff_profile:
                tas = tas.filter(staff=staff_profile)
    except (ValueError, FieldError) as e:
        return Response(
            {
                'detail': 'CIA1 teaching assignment filtering failed.',
                'error': str(e),
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    # Prefer active academic year assignments when present
    if tas.filter(academic_year__is_active=True).exists():
        tas = tas.filter(academic_year__is_active=True)

    # If UI provided an explicit teaching assignment id, scope roster to that section.
    ta_id_raw = getattr(request, 'query_params', {}).get('teaching_assignment_id') if hasattr(request, 'query_params') else request.GET.get('teaching_assignment_id')
    ta_id = None
    try:
        if ta_id_raw:
            ta_id = int(str(ta_id_raw))
    except Exception:
        ta_id = None

    # For OBE Master/IQAC flows, require explicit TA id to scope the roster.
    if is_obe_master and ta_id is None:
        return Response({'detail': 'teaching_assignment_id is required for IQAC / OBE Master roster view.'}, status=status.HTTP_400_BAD_REQUEST)

    # Resolve roster sources:
    # - Normal TAs: section roster
    # - Elective TAs (often sectionless): elective-choices roster
    selected_ta = None
    if ta_id is not None:
        # If the UI provided a TA id, prefer scoping to that TA when it belongs
        # to the filtered `tas` queryset (i.e., the current user's / HOD's assignments).
        # If the provided TA id is not in `tas`, do NOT raise immediately; instead
        # fall back to the user's available TAs so the UI can still display a roster.
        selected_ta = tas.filter(id=ta_id).first()
        if selected_ta:
            # Narrow `tas` so elective roster detection matches the selected TA only.
            tas = tas.filter(id=selected_ta.id)
            section_ids = [selected_ta.section_id]
        else:
            # If the selected TA isn't present in the subject-filtered queryset (data mismatch),
            # try resolving the TA by id directly. This prevents returning an empty roster
            # with 200 OK when the UI explicitly selected a section.
            fallback_ta = TeachingAssignment.objects.select_related(
                'section', 'academic_year', 'curriculum_row', 'subject', 'staff',
            ).filter(is_active=True, id=ta_id).first()

            def _ta_matches_subject(ta_obj) -> bool:
                try:
                    want = str(subject_id or '').strip().upper()
                    have = None
                    try:
                        have = getattr(getattr(ta_obj, 'subject', None), 'code', None)
                    except Exception:
                        have = None
                    if not have:
                        try:
                            have = getattr(getattr(ta_obj, 'curriculum_row', None), 'course_code', None)
                        except Exception:
                            have = None
                    have = str(have or '').strip().upper()
                    return bool(want and have and want == have)
                except Exception:
                    return False

            def _can_use_fallback_ta(ta_obj) -> bool:
                if is_obe_master:
                    return True
                # HOD/ADVISOR: allow within department
                if ('HOD' in role_names or 'ADVISOR' in role_names) and staff_profile and getattr(staff_profile, 'department_id', None):
                    try:
                        dept_id = getattr(staff_profile, 'department_id', None)
                        ta_dept_id = getattr(getattr(getattr(getattr(ta_obj, 'section', None), 'batch', None), 'course', None), 'department_id', None)
                        return dept_id is not None and ta_dept_id == dept_id
                    except Exception:
                        return False
                # Staff: only their own TA
                if staff_profile:
                    try:
                        return getattr(ta_obj, 'staff_id', None) == getattr(staff_profile, 'id', None)
                    except Exception:
                        return False
                return False

            if fallback_ta and _ta_matches_subject(fallback_ta) and _can_use_fallback_ta(fallback_ta):
                selected_ta = fallback_ta
                tas = TeachingAssignment.objects.select_related('section', 'academic_year', 'curriculum_row').filter(id=selected_ta.id)
                section_ids = [selected_ta.section_id]
            else:
                # For OBE Master/IQAC, an invalid TA id should be a hard error (since they explicitly picked it).
                if is_obe_master:
                    return Response({'detail': 'Teaching assignment not found for this course.'}, status=status.HTTP_404_NOT_FOUND)
                # Otherwise ignore invalid/unowned TA id and use the user's own assignments instead.
                section_ids = list(tas.values_list('section_id', flat=True).distinct())
    else:
        section_ids = list(tas.values_list('section_id', flat=True).distinct())

    # Clean up null section ids (elective TAs may not have a section)
    section_ids = [sid for sid in section_ids if sid]

    elective_student_ids: list[int] = []
    try:
        # When TA is elective (has elective_subject_id and no section), roster should come
        # from curriculum.ElectiveChoice rather than section.
        from curriculum.models import ElectiveChoice

        elective_subject_ids: list[int] = []
        elective_ay_id = None
        if selected_ta and getattr(selected_ta, 'elective_subject_id', None) and not getattr(selected_ta, 'section_id', None):
            elective_subject_ids = [int(selected_ta.elective_subject_id)]
            elective_ay_id = getattr(selected_ta, 'academic_year_id', None)
        else:
            elective_subject_ids = list(
                tas.filter(section__isnull=True)
                .exclude(elective_subject__isnull=True)
                .values_list('elective_subject_id', flat=True)
                .distinct()
            )

        if elective_subject_ids:
            eqs = ElectiveChoice.objects.filter(is_active=True, elective_subject_id__in=elective_subject_ids).exclude(student__isnull=True)
            if elective_ay_id:
                eqs = eqs.filter(academic_year_id=elective_ay_id)
            elective_student_ids = list(eqs.values_list('student_id', flat=True).distinct())
    except Exception:
        elective_student_ids = []

    # Roster: use the same TA roster logic as SSA/Model (section assignments + legacy section, and elective choices).
    # This avoids over-including students via broad StudentProfile joins.
    ta = selected_ta or _resolve_staff_teaching_assignment(
        request,
        subject_code=(subject.code if subject is not None else str(subject_id)),
        teaching_assignment_id=ta_id,
    )
    students = _get_studentprofile_queryset_for_teaching_assignment(ta)
    strict_scope = _strict_assignment_scope(subject_code=(subject.code if subject is not None else str(subject_id)), teaching_assignment=ta)

    from .models import Cia1Mark

    # If the DB hasn't been migrated yet, avoid returning Django's HTML 500 page.
    try:
        # Quick touch to ensure table exists.
        Cia1Mark.objects.none().count()
    except OperationalError:
        return Response(
            {
                'detail': 'CIA1 marks table is missing. Run database migrations first.',
                'how_to_fix': [
                    'cd backend',
                    'python manage.py migrate',
                ],
            },
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    if request.method == 'PUT':
        if subject is not None:
            gate = _enforce_mark_entry_window(request, subject=subject, assessment_key='cia1', teaching_assignment_id=ta_id)
            if gate is not None:
                return gate

        payload = request.data or {}
        incoming = payload.get('marks', {})
        if incoming is None:
            incoming = {}

        marks_map: dict[int, object] = {}
        if isinstance(incoming, list):
            for item in incoming:
                try:
                    sid = int(item.get('student_id'))
                except Exception:
                    continue
                marks_map[sid] = item.get('mark')
        elif isinstance(incoming, dict):
            for k, v in incoming.items():
                try:
                    marks_map[int(k)] = v
                except Exception:
                    continue
        else:
            return Response({'detail': 'Invalid marks payload.'}, status=status.HTTP_400_BAD_REQUEST)

        errors: list[str] = []
        try:
            with transaction.atomic():
                for s in students:
                    if s.id not in marks_map:
                        continue
                    raw = marks_map.get(s.id)

                    if raw is None or (isinstance(raw, str) and raw.strip() == ''):
                        _delete_scoped_mark(Cia1Mark, subject=subject, student=s, teaching_assignment=ta)
                        continue

                    try:
                        mark = Decimal(str(raw))
                    except (InvalidOperation, ValueError):
                        errors.append(f'Invalid mark for {s.reg_no}: {raw}')
                        continue

                    _upsert_scoped_mark(
                        Cia1Mark,
                        subject=subject,
                        student=s,
                        teaching_assignment=ta,
                        mark_defaults={'mark': mark},
                    )
        except OperationalError:
            return Response(
                {
                    'detail': 'CIA1 marks table is missing. Run database migrations first.',
                    'how_to_fix': [
                        'cd backend',
                        'python manage.py migrate',
                    ],
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        if errors:
            return Response({'detail': 'Validation error.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = _filter_marks_queryset_for_teaching_assignment(
            Cia1Mark.objects.filter(subject=subject, student__in=students),
            ta,
            strict_scope=strict_scope,
        )
        existing = {
            m.student_id: (str(m.mark) if m.mark is not None else None)
            for m in rows
        }
    except OperationalError:
        existing = {}

    out_students = []
    for s in students:
        full_name = ''
        try:
            full_name = s.user.get_full_name()
        except Exception:
            full_name = ''

        out_students.append({
            'id': s.id,
            'reg_no': s.reg_no,
            'name': full_name or s.user.username,
            'section': str(s.section) if s.section else None,
        })

    subject_payload = None
    if subject is not None:
        subject_payload = {'code': subject.code, 'name': subject.name}
    else:
        # Fall back to curriculum_row info if available.
        any_ta = tas.first()
        cr = getattr(any_ta, 'curriculum_row', None) if any_ta else None
        subject_payload = {
            'code': str(subject_id),
            'name': (getattr(cr, 'course_name', None) or str(subject_id)),
        }

    return Response({
        'subject': subject_payload,
        'students': out_students,
        'marks': existing,
    })


@api_view(['GET', 'PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cia2_marks(request, subject_id):
    """CIA2 marks API used by the React Faculty → OBE → Mark Entry → CIA2 screen.

    - GET: returns roster + current marks
    - PUT: upserts/clears marks
    """
    user = request.user
    staff_profile = getattr(user, 'staff_profile', None)
    role_names = {r.name.upper() for r in user.roles.all()}
    try:
        user_perms = {str(p or '').lower() for p in (get_user_permissions(user) or [])}
    except Exception:
        user_perms = set()

    is_obe_master = ('obe.master.manage' in user_perms) or ('IQAC' in role_names) or getattr(user, 'is_superuser', False)
    if not staff_profile and not is_obe_master and not getattr(user, 'is_staff', False):
        return Response({'detail': 'Faculty access only.'}, status=status.HTTP_403_FORBIDDEN)

    # Subject may not exist when teaching assignments reference curriculum rows only.
    subject = Subject.objects.filter(code=subject_id).first()
    if subject is None:
        try:
            subject = _get_subject(subject_id, request)
        except Exception:
            subject = None

    try:
        tas = TeachingAssignment.objects.select_related('section', 'academic_year', 'curriculum_row').filter(is_active=True)
        if subject is not None:
            tas = tas.filter(Q(subject=subject) | Q(curriculum_row__course_code=subject.code))
        else:
            tas = tas.filter(Q(subject__code=subject_id) | Q(curriculum_row__course_code=subject_id))
    except (ValueError, FieldError) as e:
        return Response(
            {
                'detail': 'CIA2 teaching assignment query failed.',
                'error': str(e),
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    # Staff: only their teaching assignments; HOD/ADVISOR: within their department.
    # IQAC/OBE Master: do not restrict to staff-owned assignments.
    try:
        if is_obe_master:
            pass
        elif 'HOD' in role_names or 'ADVISOR' in role_names:
            if staff_profile and getattr(staff_profile, 'department_id', None):
                tas = tas.filter(section__batch__course__department=staff_profile.department)
        else:
            if staff_profile:
                tas = tas.filter(staff=staff_profile)
    except (ValueError, FieldError) as e:
        return Response(
            {
                'detail': 'CIA2 teaching assignment filtering failed.',
                'error': str(e),
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    # Prefer active academic year assignments when present
    if tas.filter(academic_year__is_active=True).exists():
        tas = tas.filter(academic_year__is_active=True)

    # If UI provided an explicit teaching assignment id, scope roster to that section.
    ta_id_raw = getattr(request, 'query_params', {}).get('teaching_assignment_id') if hasattr(request, 'query_params') else request.GET.get('teaching_assignment_id')
    ta_id = None
    try:
        if ta_id_raw:
            ta_id = int(str(ta_id_raw))
    except Exception:
        ta_id = None

    if is_obe_master and ta_id is None:
        return Response({'detail': 'teaching_assignment_id is required for IQAC / OBE Master roster view.'}, status=status.HTTP_400_BAD_REQUEST)

    selected_ta = None
    if ta_id is not None:
        # If the UI provided a TA id, prefer scoping to that TA when it belongs
        # to the filtered `tas` queryset. Fall back to user's TAs if not found.
        selected_ta = tas.filter(id=ta_id).first()
        if selected_ta:
            tas = tas.filter(id=selected_ta.id)
            section_ids = [selected_ta.section_id]
        else:
            fallback_ta = TeachingAssignment.objects.select_related(
                'section', 'academic_year', 'curriculum_row', 'subject', 'staff',
            ).filter(is_active=True, id=ta_id).first()

            def _ta_matches_subject(ta_obj) -> bool:
                try:
                    want = str(subject_id or '').strip().upper()
                    have = None
                    try:
                        have = getattr(getattr(ta_obj, 'subject', None), 'code', None)
                    except Exception:
                        have = None
                    if not have:
                        try:
                            have = getattr(getattr(ta_obj, 'curriculum_row', None), 'course_code', None)
                        except Exception:
                            have = None
                    have = str(have or '').strip().upper()
                    return bool(want and have and want == have)
                except Exception:
                    return False

            def _can_use_fallback_ta(ta_obj) -> bool:
                if is_obe_master:
                    return True
                if ('HOD' in role_names or 'ADVISOR' in role_names) and staff_profile and getattr(staff_profile, 'department_id', None):
                    try:
                        dept_id = getattr(staff_profile, 'department_id', None)
                        ta_dept_id = getattr(getattr(getattr(getattr(ta_obj, 'section', None), 'batch', None), 'course', None), 'department_id', None)
                        return dept_id is not None and ta_dept_id == dept_id
                    except Exception:
                        return False
                if staff_profile:
                    try:
                        return getattr(ta_obj, 'staff_id', None) == getattr(staff_profile, 'id', None)
                    except Exception:
                        return False
                return False

            if fallback_ta and _ta_matches_subject(fallback_ta) and _can_use_fallback_ta(fallback_ta):
                selected_ta = fallback_ta
                tas = TeachingAssignment.objects.select_related('section', 'academic_year', 'curriculum_row').filter(id=selected_ta.id)
                section_ids = [selected_ta.section_id]
            else:
                if is_obe_master:
                    return Response({'detail': 'Teaching assignment not found for this course.'}, status=status.HTTP_404_NOT_FOUND)
                section_ids = list(tas.values_list('section_id', flat=True).distinct())
    else:
        section_ids = list(tas.values_list('section_id', flat=True).distinct())

    section_ids = [sid for sid in section_ids if sid]

    elective_student_ids: list[int] = []
    try:
        from curriculum.models import ElectiveChoice

        elective_subject_ids: list[int] = []
        elective_ay_id = None
        if selected_ta and getattr(selected_ta, 'elective_subject_id', None) and not getattr(selected_ta, 'section_id', None):
            elective_subject_ids = [int(selected_ta.elective_subject_id)]
            elective_ay_id = getattr(selected_ta, 'academic_year_id', None)
        else:
            elective_subject_ids = list(
                tas.filter(section__isnull=True)
                .exclude(elective_subject__isnull=True)
                .values_list('elective_subject_id', flat=True)
                .distinct()
            )

        if elective_subject_ids:
            eqs = ElectiveChoice.objects.filter(is_active=True, elective_subject_id__in=elective_subject_ids).exclude(student__isnull=True)
            if elective_ay_id:
                eqs = eqs.filter(academic_year_id=elective_ay_id)
            elective_student_ids = list(eqs.values_list('student_id', flat=True).distinct())
    except Exception:
        elective_student_ids = []

    # Roster: use the same TA roster logic as SSA/Model (section assignments + legacy section, and elective choices).
    ta = selected_ta or _resolve_staff_teaching_assignment(
        request,
        subject_code=(subject.code if subject is not None else str(subject_id)),
        teaching_assignment_id=ta_id,
    )
    students = _get_studentprofile_queryset_for_teaching_assignment(ta)
    strict_scope = _strict_assignment_scope(subject_code=(subject.code if subject is not None else str(subject_id)), teaching_assignment=ta)

    from .models import Cia2Mark

    # If the DB hasn't been migrated yet, avoid returning Django's HTML 500 page.
    try:
        Cia2Mark.objects.none().count()
    except OperationalError:
        return Response(
            {
                'detail': 'CIA2 marks table is missing. Run database migrations first.',
                'how_to_fix': [
                    'cd backend',
                    'python manage.py migrate',
                ],
            },
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    if request.method == 'PUT':
        if subject is not None:
            gate = _enforce_mark_entry_window(request, subject=subject, assessment_key='cia2', teaching_assignment_id=ta_id)
            if gate is not None:
                return gate

        payload = request.data or {}
        incoming = payload.get('marks', {})
        if incoming is None:
            incoming = {}

        marks_map: dict[int, object] = {}
        if isinstance(incoming, list):
            for item in incoming:
                try:
                    sid = int(item.get('student_id'))
                except Exception:
                    continue
                marks_map[sid] = item.get('mark')
        elif isinstance(incoming, dict):
            for k, v in incoming.items():
                try:
                    marks_map[int(k)] = v
                except Exception:
                    continue
        else:
            return Response({'detail': 'Invalid marks payload.'}, status=status.HTTP_400_BAD_REQUEST)

        errors: list[str] = []
        try:
            with transaction.atomic():
                for s in students:
                    if s.id not in marks_map:
                        continue
                    raw = marks_map.get(s.id)

                    if raw is None or (isinstance(raw, str) and raw.strip() == ''):
                        _delete_scoped_mark(Cia2Mark, subject=subject, student=s, teaching_assignment=ta)
                        continue

                    try:
                        mark = Decimal(str(raw))
                    except (InvalidOperation, ValueError):
                        errors.append(f'Invalid mark for {s.reg_no}: {raw}')
                        continue

                    _upsert_scoped_mark(
                        Cia2Mark,
                        subject=subject,
                        student=s,
                        teaching_assignment=ta,
                        mark_defaults={'mark': mark},
                    )
        except OperationalError:
            return Response(
                {
                    'detail': 'CIA2 marks table is missing. Run database migrations first.',
                    'how_to_fix': [
                        'cd backend',
                        'python manage.py migrate',
                    ],
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        if errors:
            return Response({'detail': 'Validation error.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = _filter_marks_queryset_for_teaching_assignment(
            Cia2Mark.objects.filter(subject=subject, student__in=students),
            ta,
            strict_scope=strict_scope,
        )
        existing = {
            m.student_id: (str(m.mark) if m.mark is not None else None)
            for m in rows
        }
    except OperationalError:
        existing = {}

    out_students = []
    for s in students:
        full_name = ''
        try:
            full_name = s.user.get_full_name()
        except Exception:
            full_name = ''

        out_students.append({
            'id': s.id,
            'reg_no': s.reg_no,
            'name': full_name or s.user.username,
            'section': str(s.section) if s.section else None,
        })

    subject_payload = None
    if subject is not None:
        subject_payload = {'code': subject.code, 'name': subject.name}
    else:
        any_ta = tas.first()
        cr = getattr(any_ta, 'curriculum_row', None) if any_ta else None
        subject_payload = {
            'code': str(subject_id),
            'name': (getattr(cr, 'course_name', None) or str(subject_id)),
        }

    return Response({
        'subject': subject_payload,
        'students': out_students,
        'marks': existing,
    })


def _require_permissions(request, required_codes: set[str]):
    user = getattr(request, 'user', None)
    if not user or not user.is_authenticated:
        return Response({'detail': 'Authentication required.'}, status=status.HTTP_401_UNAUTHORIZED)

    if getattr(user, 'is_superuser', False):
        return None

    user_perms = set(get_user_permissions(user))
    if user_perms.intersection(required_codes):
        return None

    needed = ', '.join(sorted(required_codes))
    return Response({'detail': f'Permission required: {needed}.'}, status=status.HTTP_403_FORBIDDEN)


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def upload_cdap(request):
    auth = _require_permissions(request, {'obe.cdap.upload'})
    if auth:
        return auth

    # Optional: enforce per-course lock if subject_id is provided (frontend sends it).
    try:
        subject_code = str((request.data or {}).get('subject_id') or '').strip()
    except Exception:
        subject_code = ''
    if subject_code:
        try:
            subject = _get_subject(subject_code, request)
            gate = _enforce_mark_entry_not_blocked(
                request,
                subject_code=subject_code,
                subject_name=getattr(subject, 'name', '') or subject_code,
                assessment='cdap',
            )
            if gate is not None:
                return gate
        except Exception:
            # If subject can't be resolved, proceed with parse only.
            pass

    if 'file' not in request.FILES:
        return Response({'detail': 'Missing file'}, status=status.HTTP_400_BAD_REQUEST)
    parsed = parse_cdap_excel(request.FILES['file'])
    return Response(parsed)


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def upload_articulation_matrix(request):
    auth = _require_permissions(request, {'obe.cdap.upload'})
    if auth:
        return auth
    if 'file' not in request.FILES:
        return Response({'detail': 'Missing file'}, status=status.HTTP_400_BAD_REQUEST)
    parsed = parse_articulation_matrix_excel(request.FILES['file'])
    return Response(parsed)


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def upload_docx(request):
    """Accept a .docx file upload for Exam Management. Saves file and returns stored path."""
    auth = _require_permissions(request, {'obe.cdap.upload'})
    if auth:
        return auth

    if 'file' not in request.FILES:
        return Response({'detail': 'Missing file'}, status=status.HTTP_400_BAD_REQUEST)

    f = request.FILES['file']
    # Ensure upload directory exists under MEDIA_ROOT if configured
    upload_dir = os.path.join('obe', 'uploads')
    try:
        saved_path = default_storage.save(os.path.join(upload_dir, f.name), f)
    except Exception as e:
        return Response({'detail': f'Failed to save file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    file_url = None
    try:
        file_url = default_storage.url(saved_path)
    except Exception:
        file_url = saved_path

    return Response({'saved_path': saved_path, 'file_url': file_url})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def list_uploads(request):
    """List files under the OBE uploads folder. Returns name, path and url for each file."""
    auth = _require_permissions(request, {'obe.cdap.upload'})
    if auth:
        return auth

    upload_dir = os.path.join('obe', 'uploads')
    files_list = []
    try:
        # default_storage.listdir returns (dirs, files)
        dirs, files = default_storage.listdir(upload_dir)
        for fname in files:
            rel_path = os.path.join(upload_dir, fname)
            try:
                url = default_storage.url(rel_path)
            except Exception:
                url = rel_path
            files_list.append({'name': fname, 'path': rel_path, 'url': url})
    except Exception:
        # If directory doesn't exist or other error, return empty list
        files_list = []

    return Response({'files': files_list})


@api_view(['GET', 'PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cdap_revision(request, subject_id):
    if request.method == 'PUT':
        required = {'obe.cdap.upload'}
        auth = _require_permissions(request, required)
        if auth:
            return auth

    if request.method == 'GET':
        rev = CdapRevision.objects.filter(subject_id=subject_id).first()
        if not rev:
            return Response({
                'subject_id': str(subject_id),
                'status': 'draft',
                'rows': [],
                'books': {'textbook': '', 'reference': ''},
                'active_learning': {'grid': [], 'dropdowns': []},
            })
        return Response({
            'subject_id': str(rev.subject_id),
            'status': rev.status,
            'rows': rev.rows,
            'books': rev.books,
            'active_learning': rev.active_learning,
        })

    body = request.data or {}
    if body is None:
        return Response({'detail': 'Invalid JSON'}, status=status.HTTP_400_BAD_REQUEST)

    incoming_status = str(body.get('status', 'draft') or 'draft').strip().lower()

    # Fetch existing revision early so we can correct legacy lock rows.
    obj = CdapRevision.objects.filter(subject_id=subject_id).first()

    # Legacy safety: older logic incorrectly marked CDAP as "published" (lock row) even for draft saves.
    # If the stored revision isn't actually published, clear any stale CDAP lock BEFORE enforcing gates.
    try:
        current_status = str(getattr(obj, 'status', '') or '').strip().lower() if obj else ''
        if current_status != 'published':
            ta = _resolve_staff_teaching_assignment(
                request,
                subject_code=str(subject_id),
                teaching_assignment_id=_get_teaching_assignment_id_from_request(request, body),
            )
            academic_year = getattr(ta, 'academic_year', None) if ta else None
            section_name = _resolve_section_name_from_ta(ta)
            lock = _get_mark_table_lock_if_exists(
                staff_user=getattr(request, 'user', None),
                subject_code=str(subject_id),
                assessment='cdap',
                teaching_assignment=ta,
                academic_year=academic_year,
                section_name=section_name,
            )
            if lock is not None and bool(getattr(lock, 'is_published', False)):
                lock.is_published = False
                lock.mark_manager_locked = False
                lock.mark_entry_unblocked_until = None
                lock.mark_manager_unlocked_until = None
                lock.recompute_blocks()
                lock.save(
                    update_fields=[
                        'is_published',
                        'published_blocked',
                        'mark_entry_blocked',
                        'mark_manager_locked',
                        'mark_entry_unblocked_until',
                        'mark_manager_unlocked_until',
                        'updated_at',
                    ]
                )
    except Exception:
        pass

    # Enforce lock after publish unless an IQAC approval window is active.
    try:
        subject = _get_subject(subject_id, request)
        gate = _enforce_mark_entry_not_blocked(
            request,
            subject_code=str(subject_id),
            subject_name=getattr(subject, 'name', '') or str(subject_id),
            assessment='cdap',
            teaching_assignment_id=_get_teaching_assignment_id_from_request(request, body),
        )
        if gate is not None:
            return gate
    except Exception:
        # best-effort; do not block on subject resolution failures
        pass

    defaults = {
        'rows': body.get('rows', []),
        'books': body.get('books', {}),
        'active_learning': body.get('active_learning', {}),
        'status': incoming_status or 'draft',
        'updated_by': getattr(request.user, 'id', None),
    }

    if obj:
        for k, v in defaults.items():
            setattr(obj, k, v)
        obj.save(update_fields=list(defaults.keys()) + ['updated_at'])
    else:
        obj = CdapRevision(subject_id=subject_id, created_by=getattr(request.user, 'id', None), **defaults)
        obj.save()

    # Lock only on publish (draft saves should remain editable).
    if incoming_status == 'published':
        try:
            _touch_lock_after_publish(
                request,
                subject_code=str(subject_id),
                subject_name=str(subject_id),
                assessment='cdap',
                teaching_assignment_id=_get_teaching_assignment_id_from_request(request, body),
            )
        except Exception:
            pass

    return Response({
        'subject_id': str(obj.subject_id),
        'status': obj.status,
        'rows': obj.rows,
        'books': obj.books,
        'active_learning': obj.active_learning,
    })


@api_view(['GET', 'PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def lca_revision(request, subject_id):
    from .models import LcaRevision

    required = {'obe.view'} if request.method == 'GET' else {'obe.cdap.upload'}
    auth = _require_permissions(request, required)
    if auth:
        return auth

    if request.method == 'GET':
        rev = LcaRevision.objects.filter(subject_id=subject_id).first()
        if not rev:
            return Response({'subject_id': str(subject_id), 'status': 'draft', 'data': {}})
        return Response({'subject_id': str(rev.subject_id), 'status': rev.status, 'data': rev.data})

    body = request.data or {}
    if body is None:
        return Response({'detail': 'Invalid JSON'}, status=status.HTTP_400_BAD_REQUEST)

    # Enforce lock unless an edit approval window exists.
    try:
        subject = _get_subject(subject_id, request)
        gate = _enforce_mark_entry_not_blocked(
            request,
            subject_code=str(subject_id),
            subject_name=getattr(subject, 'name', '') or str(subject_id),
            assessment='lca',
            teaching_assignment_id=_get_teaching_assignment_id_from_request(request, body),
        )
        if gate is not None:
            return gate
    except Exception:
        pass

    defaults = {
        'data': body.get('data', {}),
        'status': body.get('status', 'draft'),
        'updated_by': getattr(request.user, 'id', None),
    }

    obj = LcaRevision.objects.filter(subject_id=subject_id).first()
    if obj:
        for k, v in defaults.items():
            setattr(obj, k, v)
        obj.save(update_fields=list(defaults.keys()) + ['updated_at'])
    else:
        obj = LcaRevision(subject_id=subject_id, created_by=getattr(request.user, 'id', None), **defaults)
        obj.save()

    # Lock only on publish (draft saves must remain editable).
    try:
        incoming_status = str(defaults.get('status') or '').strip().lower()
    except Exception:
        incoming_status = 'draft'
    if incoming_status == 'published':
        try:
            _touch_lock_after_publish(
                request,
                subject_code=str(subject_id),
                subject_name=str(subject_id),
                assessment='lca',
                teaching_assignment_id=_get_teaching_assignment_id_from_request(request, body),
            )
        except Exception:
            pass

    return Response({'subject_id': str(obj.subject_id), 'status': obj.status, 'data': obj.data})


@api_view(['GET', 'PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def co_target_revision(request, subject_id):
    """Save/load CO Target page entries.

    Stored as a JSON blob keyed by subject_id.
    """
    from .models import CoTargetRevision

    required = {'obe.view'} if request.method == 'GET' else {'obe.cdap.upload'}
    auth = _require_permissions(request, required)
    if auth:
        return auth

    if request.method == 'GET':
        rev = CoTargetRevision.objects.filter(subject_id=subject_id).first()
        if not rev:
            return Response({'subject_id': str(subject_id), 'status': 'draft', 'data': {}})
        return Response({'subject_id': str(rev.subject_id), 'status': rev.status, 'data': rev.data})

    body = request.data or {}
    if body is None:
        return Response({'detail': 'Invalid JSON'}, status=status.HTTP_400_BAD_REQUEST)

    # Enforce lock unless an edit approval window exists.
    # Reuse the existing 'lca' assessment key to avoid expanding the lock system.
    try:
        subject = _get_subject(subject_id, request)
        gate = _enforce_mark_entry_not_blocked(
            request,
            subject_code=str(subject_id),
            subject_name=getattr(subject, 'name', '') or str(subject_id),
            assessment='lca',
            teaching_assignment_id=_get_teaching_assignment_id_from_request(request, body),
        )
        if gate is not None:
            return gate
    except Exception:
        pass

    defaults = {
        'data': body.get('data', {}),
        'status': body.get('status', 'draft'),
        'updated_by': getattr(request.user, 'id', None),
    }

    obj = CoTargetRevision.objects.filter(subject_id=subject_id).first()
    if obj:
        for k, v in defaults.items():
            setattr(obj, k, v)
        obj.save(update_fields=list(defaults.keys()) + ['updated_at'])
    else:
        obj = CoTargetRevision(subject_id=subject_id, created_by=getattr(request.user, 'id', None), **defaults)
        obj.save()

    # Lock only on publish (draft saves must remain editable).
    try:
        incoming_status = str(defaults.get('status') or '').strip().lower()
    except Exception:
        incoming_status = 'draft'
    if incoming_status == 'published':
        try:
            _touch_lock_after_publish(
                request,
                subject_code=str(subject_id),
                subject_name=str(subject_id),
                assessment='lca',
                teaching_assignment_id=_get_teaching_assignment_id_from_request(request, body),
            )
        except Exception:
            pass

    return Response({'subject_id': str(obj.subject_id), 'status': obj.status, 'data': obj.data})


@api_view(['GET', 'PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def active_learning_mapping(request):
    required = {'obe.view'} if request.method == 'GET' else {'obe.master.manage'}
    auth = _require_permissions(request, required)
    if auth:
        return auth

    row = CdapActiveLearningAnalysisMapping.objects.filter(id=1).first()

    if request.method == 'GET':
        return Response({
            'mapping': row.mapping if row else {},
            'updated_at': row.updated_at.isoformat() if row and row.updated_at else None,
        })

    body = request.data or {}
    if body is None:
        return Response({'detail': 'Invalid JSON'}, status=status.HTTP_400_BAD_REQUEST)

    mapping = body.get('mapping', {})
    if row:
        row.mapping = mapping
        row.updated_by = getattr(request.user, 'id', None)
        row.save(update_fields=['mapping', 'updated_by', 'updated_at'])
    else:
        row = CdapActiveLearningAnalysisMapping(id=1, mapping=mapping, updated_by=getattr(request.user, 'id', None))
        row.save()

    return Response({'mapping': row.mapping, 'updated_at': row.updated_at.isoformat()})


@api_view(['GET', 'PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def assessment_master_config(request):
    required = {'obe.view'} if request.method == 'GET' else {'obe.master.manage'}
    auth = _require_permissions(request, required)
    if auth:
        return auth

    row = ObeAssessmentMasterConfig.objects.filter(id=1).first()

    if request.method == 'GET':
        return Response({
            'config': row.config if row else {},
            'updated_at': row.updated_at.isoformat() if row and row.updated_at else None,
        })

    body = request.data or {}
    if body is None:
        return Response({'detail': 'Invalid JSON'}, status=status.HTTP_400_BAD_REQUEST)

    config = body.get('config', {})
    if row:
        row.config = config
        row.updated_by = getattr(request.user, 'id', None)
        row.save(update_fields=['config', 'updated_by', 'updated_at'])
    else:
        row = ObeAssessmentMasterConfig(id=1, config=config, updated_by=getattr(request.user, 'id', None))
        row.save()

    return Response({'config': row.config, 'updated_at': row.updated_at.isoformat()})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def articulation_matrix(request, subject_id: str):
    auth = _require_permissions(request, {'obe.view'})
    if auth:
        return auth

    rev = CdapRevision.objects.filter(subject_id=subject_id).first()
    rows = []
    extras = {}
    if rev and isinstance(rev.rows, list):
        rows = rev.rows

    if rev and isinstance(getattr(rev, 'active_learning', None), dict):
        maybe = rev.active_learning.get('articulation_extras')
        if isinstance(maybe, dict):
            extras = maybe

    matrix = build_articulation_matrix_from_revision_rows(rows)

    # Get global mapping from OBE Master for the last 3 rows
    global_mapping_row = CdapActiveLearningAnalysisMapping.objects.filter(id=1).first()
    global_mapping = global_mapping_row.mapping if global_mapping_row and isinstance(global_mapping_row.mapping, dict) else {}

    # Apply to Units 1-4: replace the last 3 rows with OBE Master mapping
    if isinstance(matrix.get('units'), list):
        for u in matrix['units']:
            unit_idx = u.get('unit_index', 0)
            
            # For Units 1-4, use OBE Master mapping; for Unit 5+, use articulation extras
            if unit_idx in [1, 2, 3, 4] and global_mapping:
                # Get the extras for this unit to determine activity names and hours
                unit_label = str(u.get('unit') or '')
                picked = extras.get(unit_label, [])
                
                base_rows = u.get('rows') or []
                next_serial = 0
                try:
                    next_serial = max(int(r.get('s_no') or 0) for r in base_rows) if base_rows else 0
                except Exception:
                    next_serial = len(base_rows)
                
                # Add the last 3 rows based on saved extras but with OBE Master PO mapping
                for rr in picked:
                    next_serial += 1
                    
                    # Get the activity name from topic_name
                    activity_name = str(rr.get('topic_name') or rr.get('topic') or '').strip()
                    co_mapped = rr.get('co_mapped') or ''
                    hours_value = rr.get('hours') or rr.get('class_session_hours') or 2
                    
                    # Try to convert hours to number
                    try:
                        hours_value = int(hours_value) if hours_value != '-' else 2
                    except:
                        hours_value = 2
                    
                    # Get PO mapping from global mapping using activity name as key
                    po_mapping = global_mapping.get(activity_name, [])
                    if not isinstance(po_mapping, list):
                        po_mapping = []
                    
                    # Build PO values: if checked, use hours; else '-'
                    po_vals = []
                    for i in range(11):
                        is_checked = po_mapping[i] if i < len(po_mapping) else False
                        po_vals.append(hours_value if is_checked else '-')
                    
                    # PSO values remain as '-' for now
                    pso_vals = ['-', '-', '-']
                    
                    u.setdefault('rows', []).append({
                        'excel_row': rr.get('excel_row'),
                        's_no': next_serial,
                        'co_mapped': co_mapped,
                        'topic_no': rr.get('topic_no') or '-',
                        'topic_name': activity_name,
                        'po': po_vals,
                        'pso': pso_vals,
                        'hours': hours_value,
                    })
                    
            else:
                # For other units, use articulation extras as-is
                unit_label = str(u.get('unit') or '')
                picked = extras.get(unit_label)
                if not isinstance(picked, list) or not picked:
                    continue
                base_rows = u.get('rows') or []
                next_serial = 0
                try:
                    next_serial = max(int(r.get('s_no') or 0) for r in base_rows) if base_rows else 0
                except Exception:
                    next_serial = len(base_rows)
                for rr in picked:
                    next_serial += 1
                    u.setdefault('rows', []).append({
                        'excel_row': rr.get('excel_row'),
                        's_no': next_serial,
                        'co_mapped': rr.get('co_mapped') or rr.get('co_mapped'.upper()) or rr.get('co') or rr.get('label') or '',
                        'topic_no': rr.get('topic_no') or '',
                        'topic_name': rr.get('topic_name') or rr.get('topic') or '',
                        'po': rr.get('po') or [],
                        'pso': rr.get('pso') or [],
                        'hours': rr.get('hours') or rr.get('class_session_hours') or '',
                    })

    matrix['meta'] = {**(matrix.get('meta') or {}), 'subject_id': str(subject_id)}
    return Response(matrix)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def publish_window(request, assessment: str, subject_id: str):
    """Return publish window status for a staff user (used by UI to disable publish + show countdown)."""
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    assessment_key = str(assessment or '').strip().lower()
    if not _is_valid_mark_assessment_key(assessment_key, allow_documents=True, allow_cqi=True):
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    subject_code = str(subject_id or '').strip()
    ta_id_raw = getattr(request, 'query_params', {}).get('teaching_assignment_id') if hasattr(request, 'query_params') else request.GET.get('teaching_assignment_id')
    ta_id = None
    try:
        if ta_id_raw:
            ta_id = int(str(ta_id_raw))
    except Exception:
        ta_id = None

    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject_code, assessment=assessment_key, teaching_assignment_id=ta_id)
    if gate is not None:
        return gate

    info = _get_due_schedule_for_request(request, subject_code=subject_code, assessment=assessment_key, teaching_assignment_id=ta_id)

    due_at = info.get('due_at')
    open_from = info.get('open_from')
    now = info.get('now')
    academic_year = info.get('academic_year')
    ta = info.get('teaching_assignment')

    # Determine edit/window status for mark entry tables.
    unlimited = bool(info.get('global_override_active') and info.get('global_is_open'))
    assessment_enabled = bool(info.get('assessment_enabled'))
    assessment_open = bool(info.get('assessment_open'))
    within_window = bool(info.get('allowed_by_due'))
    edit_allowed = bool(assessment_enabled and assessment_open and (unlimited or within_window))

    if not assessment_enabled or not assessment_open:
        window_state = 'DISABLED'
    elif unlimited:
        window_state = 'UNLIMITED'
    elif open_from is not None and now is not None and now < open_from:
        window_state = 'NOT_STARTED'
    elif due_at is not None and now is not None and now >= due_at:
        window_state = 'ENDED'
    elif due_at is not None:
        window_state = 'OPEN'
    else:
        window_state = 'OPEN'

    auto_published = False
    if window_state == 'ENDED':
        try:
            subj = _get_subject(subject_code, request)
            auto_published = _auto_publish_from_draft_if_due(
                request,
                subject=subj,
                assessment_key=assessment_key,
                teaching_assignment_id=ta_id,
            )
        except Exception:
            auto_published = False

    # If publish is restricted to a single user, and the caller is not that user,
    # present the same response shape but with publishing disabled so the UI
    # will hide/disable the Publish button.
    pub_auth = _require_publish_owner(request)
    publish_allowed = bool(info.get('publish_allowed'))
    if pub_auth:
        publish_allowed = False

    return Response(
        {
            'assessment': assessment_key,
            'subject_code': subject_code,
            'assessment_control_active': bool(info.get('assessment_control_active')),
            'assessment_enabled': assessment_enabled,
            'assessment_open': assessment_open,
            'edit_allowed': edit_allowed,
            'window_state': window_state,
            'publish_allowed': publish_allowed,
            'allowed_by_due': bool(info.get('allowed_by_due')),
            'allowed_by_approval': bool(info.get('allowed_by_approval')),
            'global_override_active': bool(info.get('global_override_active')),
            'global_is_open': bool(info.get('global_is_open')) if info.get('global_override_active') else None,
            'allowed_by_global': bool(info.get('allowed_by_global')) if info.get('global_override_active') else None,
            'auto_published': bool(auto_published),
            'open_from': open_from.isoformat() if open_from else None,
            'due_at': due_at.isoformat() if due_at else None,
            'now': now.isoformat() if now else None,
            'remaining_seconds': info.get('remaining_seconds'),
            'starts_in_seconds': info.get('starts_in_seconds'),
            'approval_until': (info.get('approval_until').isoformat() if info.get('approval_until') else None),
            'academic_year': {
                'id': getattr(academic_year, 'id', None),
                'name': getattr(academic_year, 'name', None),
            }
            if academic_year
            else None,
            'teaching_assignment_id': getattr(ta, 'id', None) if ta else None,
        }
    )


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def auto_publish_due(request, assessment: str, subject_id: str):
    """Auto publish (best-effort) from stored draft when the due time has passed.

    This is intended to be called by the frontend exactly when the countdown reaches 0.
    """
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    assessment_key = str(assessment or '').strip().lower()
    if assessment_key not in {'ssa1', 'review1', 'ssa2', 'review2', 'cia1', 'cia2', 'formative1', 'formative2', 'model'}:
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    subject = _get_subject(subject_id, request)
    ta_id = _get_teaching_assignment_id_from_request(request, request.data if isinstance(request.data, dict) else None)

    info = _get_due_schedule_for_request(request, subject_code=subject.code, assessment=assessment_key, teaching_assignment_id=ta_id)
    due_at = info.get('due_at')
    now = info.get('now')
    unlimited = bool(info.get('global_override_active') and info.get('global_is_open'))
    if unlimited:
        return Response({'status': 'skipped', 'reason': 'unlimited', 'auto_published': False})
    if due_at is None or now is None or now < due_at:
        return Response({'status': 'skipped', 'reason': 'not_due', 'auto_published': False, 'due_at': due_at.isoformat() if due_at else None, 'now': now.isoformat() if now else None})

    ok = _auto_publish_from_draft_if_due(request, subject=subject, assessment_key=assessment_key, teaching_assignment_id=ta_id)
    return Response({'status': 'ok', 'auto_published': bool(ok)})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_window(request, assessment: str, subject_id: str):
    """Return edit window status for a staff user after publishing.

    This is separate from `publish_window` and is scoped:
    - MARK_ENTRY: unblock the marks entry table after publish
    - MARK_MANAGER: allow changing the mark manager configuration after confirmation
    """
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    assessment_key = str(assessment or '').strip().lower()
    if not _is_valid_mark_assessment_key(assessment_key, allow_documents=True, allow_cqi=True):
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    qp = getattr(request, 'query_params', {}) if hasattr(request, 'query_params') else request.GET
    scope_raw = str(qp.get('scope') or '').strip().upper()
    if scope_raw in {'MARKS', 'MARKS_ENTRY', 'MARK_ENTRY', 'TABLE'}:
        scope = 'MARK_ENTRY'
    elif scope_raw in {'MARK_MANAGER', 'MANAGER'}:
        scope = 'MARK_MANAGER'
    else:
        return Response({'detail': 'scope is required (MARK_ENTRY or MARK_MANAGER).'}, status=status.HTTP_400_BAD_REQUEST)

    subject_code = str(subject_id or '').strip()
    ta_id_raw = qp.get('teaching_assignment_id')
    ta_id = None
    try:
        if ta_id_raw:
            ta_id = int(str(ta_id_raw))
    except Exception:
        ta_id = None

    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject_code, assessment=assessment_key, teaching_assignment_id=ta_id)
    if gate is not None:
        return gate

    # Reuse schedule resolver to determine academic year for this staff+subject.
    info = _get_due_schedule_for_request(request, subject_code=subject_code, assessment=assessment_key, teaching_assignment_id=ta_id)
    academic_year = info.get('academic_year')
    ta = info.get('teaching_assignment')
    now = timezone.now()
    user = getattr(request, 'user', None)

    master_cfg_qs = ObeAssessmentMasterConfig.objects.filter(id=1).first()
    master_cfg = master_cfg_qs.config if master_cfg_qs and getattr(master_cfg_qs, 'config', None) else {}
    unlimited_publish = not master_cfg.get('edit_requests_enabled', True)

    allowed_by_approval = False
    approval_until = None
    if academic_year is not None:
        from .models import ObeEditRequest

        qs = ObeEditRequest.objects.filter(
            staff_user=getattr(request, 'user', None),
            academic_year=academic_year,
            subject_code=str(subject_code),
            assessment=str(assessment_key).lower(),
            scope=str(scope),
            status='APPROVED',
            approved_until__gt=now,
        )
        if ta is not None:
            qs = qs.filter(teaching_assignment=ta)
        approval = qs.order_by('-updated_at').first()
        if approval is not None:
            allowed_by_approval = True
            approval_until = getattr(approval, 'approved_until', None)

    allowed = bool(allowed_by_approval or unlimited_publish or _has_obe_master_permission(user))

    return Response(
        {
            'assessment': assessment_key,
            'subject_code': subject_code,
            'scope': scope,
            'allowed_by_approval': bool(allowed_by_approval),
            'allowed': bool(allowed),
            'allowed_by_unlimited': bool(unlimited_publish),
            'approval_until': approval_until.isoformat() if approval_until else None,
            'now': now.isoformat() if now else None,
            'academic_year': {
                'id': getattr(academic_year, 'id', None),
                'name': getattr(academic_year, 'name', None),
            }
            if academic_year
            else None,
            'teaching_assignment_id': getattr(ta, 'id', None) if ta else None,
        }
    )


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def mark_table_lock_status(request, assessment: str, subject_id: str):
    """Faculty: return authoritative lock state for a mark-entry table.

    Query params:
    - teaching_assignment_id (optional but recommended)

    Derived semantics:
    - entry_open is True only when mark_entry is unblocked AND mark manager is locked/confirmed.
    """
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    assessment_key = str(assessment or '').strip().lower()
    if not _is_valid_mark_assessment_key(assessment_key, allow_documents=True, allow_cqi=True):
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    subject_code = str(subject_id or '').strip()
    qp = _get_query_params(request)
    ta_id = _parse_int(qp.get('teaching_assignment_id'))

    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject_code, assessment=assessment_key, teaching_assignment_id=ta_id)
    if gate is not None:
        return gate

    master_cfg_qs = ObeAssessmentMasterConfig.objects.filter(id=1).first()
    master_cfg = master_cfg_qs.config if master_cfg_qs and getattr(master_cfg_qs, 'config', None) else {}
    unlimited_publish = not master_cfg.get('edit_requests_enabled', True)

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject_code, teaching_assignment_id=ta_id)
    academic_year = getattr(ta, 'academic_year', None) if ta else None
    section_name = _resolve_section_name_from_ta(ta)

    try:
        lock = _get_mark_table_lock_if_exists(
            staff_user=getattr(request, 'user', None),
            subject_code=subject_code,
            assessment=assessment_key,
            teaching_assignment=ta,
            academic_year=academic_year,
            section_name=section_name,
        )
    except OperationalError:
        lock = None

    # Default for pre-publish flows: Mark Manager is editable, but the marks table is NOT open
    # until Mark Manager is confirmed/locked.
    if lock is None:
        # For OBE Master / IQAC users we want to avoid showing the "Table Locked" overlay;
        # treat the table as open and mark-manager editable so IQAC can view/edit without
        # restrictions. Non-master users keep the conservative default of entry_closed.
        if _has_obe_master_permission(getattr(request, 'user', None)):
            return Response(
                {
                    'assessment': assessment_key,
                    'subject_code': subject_code,
                    'teaching_assignment_id': getattr(ta, 'id', None) if ta else None,
                    'academic_year': {
                        'id': getattr(academic_year, 'id', None),
                        'name': getattr(academic_year, 'name', None),
                    }
                    if academic_year
                    else None,
                    'section_name': section_name or None,
                    'exists': False,
                    'is_published': False,
                    'published_blocked': False,
                    'mark_entry_blocked': False,
                    'mark_manager_locked': False,
                    'mark_entry_unblocked_until': None,
                    'mark_manager_unlocked_until': None,
                    'entry_open': True,
                    'mark_manager_editable': True,
                    'unlimited_publish': bool(unlimited_publish),
                }
            )

        return Response(
            {
                'assessment': assessment_key,
                'subject_code': subject_code,
                'teaching_assignment_id': getattr(ta, 'id', None) if ta else None,
                'academic_year': {
                    'id': getattr(academic_year, 'id', None),
                    'name': getattr(academic_year, 'name', None),
                }
                if academic_year
                else None,
                'section_name': section_name or None,
                'exists': False,
                'is_published': False,
                'published_blocked': False,
                'mark_entry_blocked': False,
                'mark_manager_locked': False,
                'mark_entry_unblocked_until': None,
                'mark_manager_unlocked_until': None,
                'entry_open': False,
                'mark_manager_editable': True,
                'unlimited_publish': bool(unlimited_publish),
            }
        )

    # Effective values (apply time windows without needing cron)
    try:
        lock.recompute_blocks()
    except Exception:
        pass

    is_published = bool(getattr(lock, 'is_published', False))
    # During an approved MARK_MANAGER window, Mark Manager is unlocked and editable.
    # In that window, we also allow the marks table to open if mark-entry is unblocked
    # (so both can be edited without the published overlay blocking the page).
    now = timezone.now()
    manager_unlocked_active = bool(getattr(lock, 'mark_manager_unlocked_until', None) and getattr(lock, 'mark_manager_unlocked_until', None) > now)
    entry_open = (not bool(getattr(lock, 'mark_entry_blocked', False))) and (bool(getattr(lock, 'mark_manager_locked', False)) or manager_unlocked_active)
    mark_manager_editable = not bool(getattr(lock, 'mark_manager_locked', False))

    # IQAC / OBE master users should not be blocked by lock rows; present the table as open.
    if _has_obe_master_permission(getattr(request, 'user', None)) or unlimited_publish:
        entry_open = True
        mark_manager_editable = True

    return Response(
        {
            'assessment': assessment_key,
            'subject_code': subject_code,
            'teaching_assignment_id': getattr(getattr(lock, 'teaching_assignment', None), 'id', None),
            'academic_year': {
                'id': getattr(getattr(lock, 'academic_year', None), 'id', None),
                'name': getattr(getattr(lock, 'academic_year', None), 'name', None),
            }
            if getattr(lock, 'academic_year', None)
            else None,
            'section_name': getattr(lock, 'section_name', None) or None,
            'exists': True,
            'is_published': is_published,
            'published_blocked': bool(getattr(lock, 'published_blocked', False)),
            'mark_entry_blocked': bool(getattr(lock, 'mark_entry_blocked', False)),
            'mark_manager_locked': bool(getattr(lock, 'mark_manager_locked', False)),
            'mark_entry_unblocked_until': getattr(lock, 'mark_entry_unblocked_until', None).isoformat() if getattr(lock, 'mark_entry_unblocked_until', None) else None,
            'mark_manager_unlocked_until': getattr(lock, 'mark_manager_unlocked_until', None).isoformat() if getattr(lock, 'mark_manager_unlocked_until', None) else None,
            'entry_open': bool(entry_open),
            'mark_manager_editable': bool(mark_manager_editable),
            'unlimited_publish': bool(unlimited_publish),
            'updated_at': getattr(lock, 'updated_at', None).isoformat() if getattr(lock, 'updated_at', None) else None,
        }
    )


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def mark_table_lock_confirm_mark_manager(request, assessment: str, subject_id: str):
    """Faculty: confirm/re-lock Mark Manager in DB.

    This is used after an IQAC MARK_MANAGER approval window: when the staff
    presses Confirm, we re-lock Mark Manager immediately so the marks table can
    open (if MARK_ENTRY is unblocked).
    """
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    assessment_key = str(assessment or '').strip().lower()
    if assessment_key not in {'ssa1', 'review1', 'ssa2', 'review2', 'cia1', 'cia2', 'formative1', 'formative2', 'model'}:
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    subject_code = str(subject_id or '').strip()
    ta_id = _get_teaching_assignment_id_from_request(request, request.data if isinstance(request.data, dict) else None)

    gate = _enforce_assessment_enabled_for_course(request, subject_code=subject_code, assessment=assessment_key, teaching_assignment_id=ta_id)
    if gate is not None:
        return gate

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject_code, teaching_assignment_id=ta_id)
    academic_year = getattr(ta, 'academic_year', None) if ta else None
    section_name = _resolve_section_name_from_ta(ta)

    subject = _get_subject(subject_code, request)

    try:
        lock = _upsert_mark_table_lock(
            staff_user=getattr(request, 'user', None),
            subject_code=subject_code,
            subject_name=getattr(subject, 'name', '') or subject_code,
            assessment=assessment_key,
            teaching_assignment=ta,
            academic_year=academic_year,
            section_name=section_name,
            updated_by=getattr(getattr(request, 'user', None), 'id', None),
        )

        lock.mark_manager_locked = True
        lock.mark_manager_unlocked_until = None
        lock.recompute_blocks()
        lock.save(
            update_fields=[
                'mark_manager_locked',
                'mark_manager_unlocked_until',
                'published_blocked',
                'mark_entry_blocked',
                'updated_by',
                'updated_at',
            ]
        )
    except OperationalError:
        return Response({'detail': 'Database unavailable.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    entry_open = (not bool(getattr(lock, 'mark_entry_blocked', False))) and bool(getattr(lock, 'mark_manager_locked', False))
    return Response(
        {
            'status': 'ok',
            'assessment': assessment_key,
            'subject_code': subject_code,
            'exists': True,
            'is_published': bool(getattr(lock, 'is_published', False)),
            'mark_entry_blocked': bool(getattr(lock, 'mark_entry_blocked', False)),
            'mark_manager_locked': bool(getattr(lock, 'mark_manager_locked', False)),
            'mark_entry_unblocked_until': getattr(lock, 'mark_entry_unblocked_until', None).isoformat() if getattr(lock, 'mark_entry_unblocked_until', None) else None,
            'mark_manager_unlocked_until': getattr(lock, 'mark_manager_unlocked_until', None).isoformat() if getattr(lock, 'mark_manager_unlocked_until', None) else None,
            'entry_open': bool(entry_open),
        }
    )


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def due_schedule_subjects(request):
    """IQAC helper: list subject codes for the selected semesters."""
    auth = _require_obe_master(request)
    if auth:
        return auth

    raw = (
        getattr(request, 'query_params', {}).get('semester_ids')
        if hasattr(request, 'query_params')
        else request.GET.get('semester_ids')
    ) or (
        getattr(request, 'query_params', {}).get('semester_id')
        if hasattr(request, 'query_params')
        else request.GET.get('semester_id')
    )

    sem_ids: list[int] = []
    for part in str(raw or '').split(','):
        part = part.strip()
        if not part:
            continue
        try:
            sem_ids.append(int(part))
        except Exception:
            continue

    if not sem_ids:
        return Response({'detail': 'semester_ids is required.'}, status=status.HTTP_400_BAD_REQUEST)

    # Use curriculum mapping by Semester so IQAC can set schedules without relying on TA presence.
    from curriculum.models import CurriculumDepartment, CurriculumMaster, ElectiveSubject

    out: dict[str, dict[str, dict]] = {}
    dept_qs = CurriculumDepartment.objects.select_related('semester').filter(semester_id__in=sem_ids)
    for r in dept_qs:
        sem_id = str(getattr(r, 'semester_id', '') or '')
        code = str(getattr(r, 'course_code', '') or '').strip()
        name = str(getattr(r, 'course_name', '') or '').strip()
        class_type = str(getattr(r, 'class_type', '') or '').strip() or 'THEORY'
        enabled_assessments = getattr(r, 'enabled_assessments', None)
        if not isinstance(enabled_assessments, list):
            enabled_assessments = []
        enabled_assessments = [str(x or '').strip().lower() for x in enabled_assessments if str(x or '').strip()]
        if not sem_id or not code:
            continue
        out.setdefault(sem_id, {})
        if code not in out[sem_id]:
            out[sem_id][code] = {
                'subject_code': code,
                'subject_name': name,
                'class_type': class_type,
                'enabled_assessments': enabled_assessments,
            }

    # Fallback to master curriculum if department rows are missing.
    master_qs = CurriculumMaster.objects.select_related('semester').filter(semester_id__in=sem_ids)
    for r in master_qs:
        sem_id = str(getattr(r, 'semester_id', '') or '')
        code = str(getattr(r, 'course_code', '') or '').strip()
        name = str(getattr(r, 'course_name', '') or '').strip()
        class_type = str(getattr(r, 'class_type', '') or '').strip() or 'THEORY'
        enabled_assessments = getattr(r, 'enabled_assessments', None)
        if not isinstance(enabled_assessments, list):
            enabled_assessments = []
        enabled_assessments = [str(x or '').strip().lower() for x in enabled_assessments if str(x or '').strip()]
        if not sem_id or not code:
            continue
        out.setdefault(sem_id, {})
        if code not in out[sem_id]:
            out[sem_id][code] = {
                'subject_code': code,
                'subject_name': name,
                'class_type': class_type,
                'enabled_assessments': enabled_assessments,
            }

    # Include elective subjects that are NOT already in curriculum department/master rows.
    elective_qs = ElectiveSubject.objects.select_related('semester').filter(semester_id__in=sem_ids)
    for r in elective_qs:
        sem_id = str(getattr(r, 'semester_id', '') or '')
        code = str(getattr(r, 'course_code', '') or '').strip()
        name = str(getattr(r, 'course_name', '') or '').strip()
        class_type = str(getattr(r, 'class_type', '') or '').strip() or 'THEORY'
        if not sem_id or not code:
            continue
        out.setdefault(sem_id, {})
        if code not in out[sem_id]:
            out[sem_id][code] = {
                'subject_code': code,
                'subject_name': name,
                'class_type': class_type,
                'enabled_assessments': [],
            }

    return Response(
        {
            'semester_ids': sem_ids,
            'subjects_by_semester': {
                sem: sorted(list(items.values()), key=lambda x: (x.get('subject_code') or ''))
                for sem, items in out.items()
            },
        }
    )


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def due_schedules(request):
    """IQAC: list due schedules for semesters."""
    auth = _require_obe_master(request)
    if auth:
        return auth

    raw = (
        getattr(request, 'query_params', {}).get('semester_ids')
        if hasattr(request, 'query_params')
        else request.GET.get('semester_ids')
    )
    sem_ids: list[int] = []
    for part in str(raw or '').split(','):
        part = part.strip()
        if not part:
            continue
        try:
            sem_ids.append(int(part))
        except Exception:
            continue

    from .models import ObeDueSchedule
    qs = ObeDueSchedule.objects.select_related('semester').filter(is_active=True).order_by('semester_id', 'subject_code', 'assessment')
    if sem_ids:
        qs = qs.filter(semester_id__in=sem_ids)

    items = []
    try:
        for r in qs:
            items.append(
                {
                    'id': r.id,
                    'semester': {
                        'id': r.semester_id,
                        'number': getattr(getattr(r, 'semester', None), 'number', None),
                    }
                    if r.semester_id
                    else None,
                    'subject_code': r.subject_code,
                    'subject_name': r.subject_name,
                    'assessment': r.assessment,
                    'open_from': r.open_from.isoformat() if getattr(r, 'open_from', None) else None,
                    'due_at': r.due_at.isoformat() if r.due_at else None,
                    'is_active': bool(r.is_active),
                    'updated_at': r.updated_at.isoformat() if r.updated_at else None,
                }
            )
    except Exception as e:
        # Common cause: DB schema not migrated yet (missing `open_from` column).
        try:
            from django.db.utils import OperationalError, ProgrammingError

            if isinstance(e, (ProgrammingError, OperationalError)) and 'open_from' in str(e).lower():
                return Response(
                    {
                        'detail': 'Database schema out of date for OBE due schedules. Apply migrations (OBE 0046_obedueschedule_open_from).',
                        'hint': 'Run: python manage.py migrate',
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
        except Exception:
            pass
        raise

    return Response({'results': items})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def due_schedule_delete(request):
    """IQAC: soft-delete a due schedule row (sets is_active=False).

    Body: semester_id, subject_code, assessment
    """
    auth = _require_obe_master(request)
    if auth:
        return auth

    body = request.data or {}
    sem_id = body.get('semester_id')
    subject_code = str(body.get('subject_code') or '').strip()
    assessment = str(body.get('assessment') or '').strip().lower()

    if not sem_id or not subject_code or not assessment:
        return Response({'detail': 'semester_id, subject_code, assessment are required.'}, status=status.HTTP_400_BAD_REQUEST)

    if assessment not in {'ssa1', 'review1', 'ssa2', 'review2', 'cia1', 'cia2', 'formative1', 'formative2', 'model', 'cdap', 'articulation', 'lca'}:
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    from academics.models import Semester
    from .models import ObeDueSchedule

    sem = Semester.objects.filter(id=int(sem_id)).first()
    if not sem:
        return Response({'detail': 'Semester not found.'}, status=status.HTTP_404_NOT_FOUND)

    obj = ObeDueSchedule.objects.filter(
        semester=sem,
        subject_code=subject_code,
        assessment=assessment,
        is_active=True,
    ).first()

    if obj is None:
        return Response({'status': 'ok', 'deleted': 0})

    obj.is_active = False
    obj.updated_by = getattr(request.user, 'id', None)
    obj.save(update_fields=['is_active', 'updated_by', 'updated_at'])

    return Response({'status': 'ok', 'deleted': 1})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def due_schedule_upsert(request):
    """IQAC: create/update a due schedule row."""
    auth = _require_obe_master(request)
    if auth:
        return auth

    body = request.data or {}
    sem_id = body.get('semester_id')
    subject_code = str(body.get('subject_code') or '').strip()
    subject_name = str(body.get('subject_name') or '').strip()
    assessment = str(body.get('assessment') or '').strip().lower()
    open_from = _parse_open_from(body.get('open_from'))
    due_at = _parse_due_at(body.get('due_at'))

    if not sem_id or not subject_code or not assessment or due_at is None:
        return Response({'detail': 'semester_id, subject_code, assessment, due_at are required.'}, status=status.HTTP_400_BAD_REQUEST)

    if open_from is not None and due_at is not None and open_from >= due_at:
        return Response({'detail': 'open_from must be before due_at.'}, status=status.HTTP_400_BAD_REQUEST)

    if assessment not in {
        'ssa1',
        'review1',
        'ssa2',
        'review2',
        'cia1',
        'cia2',
        'formative1',
        'formative2',
        'model',
        # Document-level assessments that still use the same publish/lock gate
        'cdap',
        'lca',
        'articulation',
    }:
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)

    from academics.models import Semester, Subject
    from .models import ObeDueSchedule

    sem = Semester.objects.filter(id=int(sem_id)).first()
    if not sem:
        return Response({'detail': 'Semester not found.'}, status=status.HTTP_404_NOT_FOUND)

    subj = Subject.objects.filter(code=subject_code).first()
    if not subject_name and subj is not None:
        subject_name = subj.name

    try:
        obj, _created = ObeDueSchedule.objects.update_or_create(
            semester=sem,
            subject_code=subject_code,
            assessment=assessment,
            defaults={
                'subject': subj,
                'subject_name': subject_name,
                'open_from': open_from,
                'due_at': due_at,
                'is_active': True,
                'updated_by': getattr(request.user, 'id', None),
                'created_by': getattr(request.user, 'id', None),
            },
        )
    except Exception as e:
        try:
            from django.db.utils import OperationalError, ProgrammingError

            if isinstance(e, (ProgrammingError, OperationalError)) and 'open_from' in str(e).lower():
                return Response(
                    {
                        'detail': 'Database schema out of date for OBE due schedules. Apply migrations (OBE 0046_obedueschedule_open_from).',
                        'hint': 'Run: python manage.py migrate',
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
        except Exception:
            pass
        raise

    return Response(
        {
            'id': obj.id,
            'semester_id': obj.semester_id,
            'subject_code': obj.subject_code,
            'subject_name': obj.subject_name,
            'assessment': obj.assessment,
            'open_from': obj.open_from.isoformat() if getattr(obj, 'open_from', None) else None,
            'due_at': obj.due_at.isoformat() if obj.due_at else None,
            'updated_at': obj.updated_at.isoformat() if obj.updated_at else None,
        }
    )


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def due_schedule_bulk_upsert(request):
    """IQAC: bulk upsert due schedules for many subjects and/or assessments."""
    auth = _require_obe_master(request)
    if auth:
        return auth

    body = request.data or {}
    sem_id = body.get('semester_id')
    assessments = body.get('assessments') or []
    subject_codes = body.get('subject_codes') or []
    open_from = _parse_open_from(body.get('open_from'))
    due_at = _parse_due_at(body.get('due_at'))

    if not sem_id or due_at is None:
        return Response({'detail': 'semester_id and due_at are required.'}, status=status.HTTP_400_BAD_REQUEST)

    if open_from is not None and due_at is not None and open_from >= due_at:
        return Response({'detail': 'open_from must be before due_at.'}, status=status.HTTP_400_BAD_REQUEST)

    if not isinstance(assessments, list) or not assessments:
        return Response({'detail': 'assessments must be a non-empty list.'}, status=status.HTTP_400_BAD_REQUEST)

    if not isinstance(subject_codes, list) or not subject_codes:
        return Response({'detail': 'subject_codes must be a non-empty list.'}, status=status.HTTP_400_BAD_REQUEST)

    norm_assessments = [str(a).strip().lower() for a in assessments]
    bad = [a for a in norm_assessments if a not in {'ssa1', 'review1', 'ssa2', 'review2', 'cia1', 'cia2', 'formative1', 'formative2', 'model'}]
    if bad:
        return Response({'detail': f'Invalid assessments: {bad}'}, status=status.HTTP_400_BAD_REQUEST)

    from academics.models import Semester, Subject
    from .models import ObeDueSchedule

    sem = Semester.objects.filter(id=int(sem_id)).first()
    if not sem:
        return Response({'detail': 'Semester not found.'}, status=status.HTTP_404_NOT_FOUND)

    updated = 0
    try:
        for code in [str(s).strip() for s in subject_codes]:
            if not code:
                continue
            subj = Subject.objects.filter(code=code).first()
            name = getattr(subj, 'name', '') if subj else ''
            for a in norm_assessments:
                ObeDueSchedule.objects.update_or_create(
                    semester=sem,
                    subject_code=code,
                    assessment=a,
                    defaults={
                        'subject': subj,
                        'subject_name': name,
                        'open_from': open_from,
                        'due_at': due_at,
                        'is_active': True,
                        'updated_by': getattr(request.user, 'id', None),
                        'created_by': getattr(request.user, 'id', None),
                    },
                )
                updated += 1
    except Exception as e:
        try:
            from django.db.utils import OperationalError, ProgrammingError

            if isinstance(e, (ProgrammingError, OperationalError)) and 'open_from' in str(e).lower():
                return Response(
                    {
                        'detail': 'Database schema out of date for OBE due schedules. Apply migrations (OBE 0046_obedueschedule_open_from).',
                        'hint': 'Run: python manage.py migrate',
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
        except Exception:
            pass
        raise

    return Response({'status': 'ok', 'updated': updated})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def due_schedule_bulk_delete(request):
    """IQAC: bulk soft-delete due schedules for many subjects and/or assessments.

    Body: semester_id, assessments (list), subject_codes (list)
    """
    auth = _require_obe_master(request)
    if auth:
        return auth

    body = request.data or {}
    sem_id = body.get('semester_id')
    assessments = body.get('assessments') or []
    subject_codes = body.get('subject_codes') or []

    if not sem_id:
        return Response({'detail': 'semester_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    if not isinstance(assessments, list) or not assessments:
        return Response({'detail': 'assessments must be a non-empty list.'}, status=status.HTTP_400_BAD_REQUEST)

    if not isinstance(subject_codes, list) or not subject_codes:
        return Response({'detail': 'subject_codes must be a non-empty list.'}, status=status.HTTP_400_BAD_REQUEST)

    norm_assessments = [str(a).strip().lower() for a in assessments]
    allowed = {'ssa1', 'review1', 'ssa2', 'review2', 'cia1', 'cia2', 'formative1', 'formative2', 'model'}
    bad = [a for a in norm_assessments if a not in allowed]
    if bad:
        return Response({'detail': f'Invalid assessments: {bad}'}, status=status.HTTP_400_BAD_REQUEST)

    from academics.models import Semester
    from .models import ObeDueSchedule

    sem = Semester.objects.filter(id=int(sem_id)).first()
    if not sem:
        return Response({'detail': 'Semester not found.'}, status=status.HTTP_404_NOT_FOUND)

    codes = [str(s).strip() for s in subject_codes if str(s).strip()]
    if not codes:
        return Response({'status': 'ok', 'deleted': 0})

    now = timezone.now()
    deleted = (
        ObeDueSchedule.objects.filter(
            semester=sem,
            subject_code__in=codes,
            assessment__in=norm_assessments,
            is_active=True,
        ).update(
            is_active=False,
            updated_by=getattr(request.user, 'id', None),
            updated_at=now,
        )
        or 0
    )

    return Response({'status': 'ok', 'deleted': int(deleted)})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def global_publish_controls(request):
    """IQAC: list global publish overrides for selected semesters and assessments.

    Query params: semester_ids (comma separated), assessments (comma separated)
    """
    auth = _require_obe_master(request)
    if auth:
        return auth

    raw_ay = (getattr(request, 'query_params', {}).get('semester_ids') if hasattr(request, 'query_params') else request.GET.get('semester_ids')) or ''
    raw_assess = (getattr(request, 'query_params', {}).get('assessments') if hasattr(request, 'query_params') else request.GET.get('assessments')) or ''

    ay_ids: list[int] = []
    for part in str(raw_ay or '').split(','):
        part = part.strip()
        if not part:
            continue
        try:
            ay_ids.append(int(part))
        except Exception:
            continue

    assessments = [a.strip().lower() for a in str(raw_assess or '').split(',') if a and a.strip()]

    from .models import ObeGlobalPublishControl

    qs = ObeGlobalPublishControl.objects.select_related('semester')
    if ay_ids:
        qs = qs.filter(semester_id__in=ay_ids)
    if assessments:
        qs = qs.filter(assessment__in=assessments)

    out = []
    for r in qs.order_by('semester_id', 'assessment'):
        out.append({
            'id': r.id,
            'semester': {'id': getattr(r, 'semester_id', None), 'number': getattr(getattr(r, 'semester', None), 'number', None)} if r.semester_id else None,
            'assessment': r.assessment,
            'is_open': bool(r.is_open),
            'updated_at': r.updated_at.isoformat() if r.updated_at else None,
            'updated_by': r.updated_by,
        })

    return Response({'results': out})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def global_publish_controls_bulk_set(request):
    """IQAC: set (create/update) global publish override for combinations of semester_ids and assessments.

    Body: { semester_ids: [1,2], assessments: ['ssa1'], is_open: true }
    """
    auth = _require_obe_master(request)
    if auth:
        return auth

    body = request.data or {}
    ay_ids = body.get('semester_ids') or []
    assessments = body.get('assessments') or []
    is_open = bool(body.get('is_open') if body.get('is_open') is not None else True)

    try:
        ay_ids = [int(a) for a in ay_ids]
    except Exception:
        ay_ids = []
    assessments = [str(a).strip().lower() for a in assessments if str(a).strip()]

    if not ay_ids or not assessments:
        return Response({'detail': 'semester_ids and assessments are required.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import ObeGlobalPublishControl
    updated = 0
    for ay in ay_ids:
        for a in assessments:
            obj, created = ObeGlobalPublishControl.objects.update_or_create(
                semester_id=ay,
                assessment=a,
                defaults={
                    'is_open': is_open,
                    'updated_by': getattr(request.user, 'id', None),
                },
            )
            updated += 1

    return Response({'status': 'ok', 'updated': updated})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def global_publish_controls_bulk_reset(request):
    """IQAC: remove global publish override rows to fall back to due dates.

    Body: { semester_ids: [1,2], assessments: ['ssa1'] }
    """
    auth = _require_obe_master(request)
    if auth:
        return auth

    body = request.data or {}
    ay_ids = body.get('semester_ids') or []
    assessments = body.get('assessments') or []

    try:
        ay_ids = [int(a) for a in ay_ids]
    except Exception:
        ay_ids = []
    assessments = [str(a).strip().lower() for a in assessments if str(a).strip()]

    if not ay_ids or not assessments:
        return Response({'detail': 'semester_ids and assessments are required.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import ObeGlobalPublishControl
    qs = ObeGlobalPublishControl.objects.filter(semester_id__in=ay_ids, assessment__in=assessments)
    deleted, _ = qs.delete()
    return Response({'status': 'ok', 'deleted': int(deleted)})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def publish_request_create(request):
    """Faculty: create a publish request after the due time is closed."""
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    body = request.data or {}
    assessment = str(body.get('assessment') or '').strip().lower()
    subject_code = str(body.get('subject_code') or body.get('subject_id') or '').strip()
    reason = str(body.get('reason') or '').strip()
    force = bool(body.get('force'))
    ta_id = None
    try:
        if body.get('teaching_assignment_id') is not None:
            ta_id = int(str(body.get('teaching_assignment_id')))
    except Exception:
        ta_id = None

    if not _is_valid_mark_assessment_key(assessment, allow_documents=True, allow_cqi=True):
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)
    if not subject_code:
        return Response({'detail': 'subject_code is required.'}, status=status.HTTP_400_BAD_REQUEST)
    if not reason:
        return Response({'detail': 'reason is required.'}, status=status.HTTP_400_BAD_REQUEST)

    info = _get_due_schedule_for_request(request, subject_code=subject_code, assessment=assessment, teaching_assignment_id=ta_id)
    if (not force) and info.get('publish_allowed') and info.get('allowed_by_due'):
        return Response({'detail': 'Publish is still open; request is not needed.'}, status=status.HTTP_400_BAD_REQUEST)

    academic_year = info.get('academic_year')
    if academic_year is None:
        return Response({'detail': 'Unable to resolve academic year for this subject.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import ObePublishRequest

    # Resolve department HOD for this staff member + academic year (same logic as edit requests).
    hod_user = None
    dept = getattr(staff_profile, 'current_department', None) or getattr(staff_profile, 'department', None)
    if dept is None:
        return Response(
            {
                'detail': 'Unable to resolve your department for HOD routing.',
                'how_to_fix': [
                    'Ensure this staff has an active department assignment (or department set) in Academics.',
                    'Then re-try submitting the publish request.',
                ],
            },
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        from academics.models import DepartmentRole

        if dept is not None and academic_year is not None:
            hod_role = (
                DepartmentRole.objects.filter(
                    department=dept,
                    academic_year=academic_year,
                    role='HOD',
                    is_active=True,
                )
                .select_related('staff__user')
                .first()
            )
            if hod_role is not None:
                hod_user = getattr(getattr(hod_role, 'staff', None), 'user', None)

        # Backward-compat: if a HOD is configured but not mapped for this academic year, pick latest active.
        if hod_user is None and dept is not None:
            hod_role_any = (
                DepartmentRole.objects.filter(
                    department=dept,
                    role='HOD',
                    is_active=True,
                )
                .select_related('staff__user', 'academic_year')
                .order_by('-academic_year_id', '-id')
                .first()
            )
            if hod_role_any is not None:
                hod_user = getattr(getattr(hod_role_any, 'staff', None), 'user', None)
    except Exception:
        hod_user = None

    if hod_user is None and dept is not None:
        try:
            from academics.models import StaffProfile, RoleAssignment
            from django.db.models import Q

            dept_staff_qs = StaffProfile.objects.filter(
                Q(department=dept)
                | Q(department_assignments__department=dept, department_assignments__end_date__isnull=True)
            ).select_related('user').distinct()

            ra = (
                RoleAssignment.objects.filter(
                    staff__in=dept_staff_qs,
                    role_name__iexact='HOD',
                    end_date__isnull=True,
                )
                .select_related('staff__user')
                .order_by('-start_date')
                .first()
            )
            if ra is not None:
                hod_user = getattr(getattr(ra, 'staff', None), 'user', None)
        except Exception:
            hod_user = None

    routing_warning = None
    if hod_user is None:
        dept_name = (
            str(getattr(dept, 'short_name', '') or '').strip()
            or str(getattr(dept, 'name', '') or '').strip()
            or str(getattr(dept, 'code', '') or '').strip()
            or 'your department'
        )
        routing_warning = f'No active HOD configured for {dept_name}. Request will be sent directly to IQAC.'

    if hod_user is None and dept is not None:
        try:
            from academics.models import StaffProfile
            from django.db.models import Q

            dept_staff_qs = StaffProfile.objects.filter(
                Q(department=dept)
                | Q(department_assignments__department=dept, department_assignments__end_date__isnull=True)
            ).select_related('user').distinct()

            sp_hod = (
                dept_staff_qs.filter(user__roles__name__iexact='HOD')
                .exclude(user=request.user)
                .order_by('id')
                .first()
            )
            if sp_hod is not None:
                hod_user = getattr(sp_hod, 'user', None)
        except Exception:
            hod_user = None

    # If a pending request already exists, update reason and keep status pending.
    existing = ObePublishRequest.objects.filter(
        staff_user=request.user,
        academic_year=academic_year,
        subject_code=subject_code,
        assessment=assessment,
        status='PENDING',
    ).order_by('-created_at').first()

    if existing:
        existing.reason = reason
        existing.subject_name = existing.subject_name or (getattr(info.get('schedule'), 'subject_name', '') or '')

        # If this is an older request with no HOD routing info yet, assign it now.
        try:
            if getattr(existing, 'hod_user_id', None) is None and hod_user is not None:
                existing.hod_user = hod_user
                existing.hod_approved = False
        except Exception:
            pass

        existing.save(update_fields=['reason', 'subject_name', 'hod_user', 'hod_approved', 'updated_at'])
        req = existing
    else:
        req = ObePublishRequest.objects.create(
            staff_user=request.user,
            academic_year=academic_year,
            subject_code=subject_code,
            subject_name=(getattr(info.get('schedule'), 'subject_name', '') or ''),
            assessment=assessment,
            reason=reason,
            hod_user=hod_user,
            hod_approved=(False if hod_user is not None else True),
        )

    routed_to = 'HOD' if getattr(req, 'hod_user_id', None) is not None and not bool(getattr(req, 'hod_approved', True)) else 'IQAC'

    hod_u = getattr(req, 'hod_user', None)
    hod_name = (
        ' '.join([
            str(getattr(hod_u, 'first_name', '') or '').strip(),
            str(getattr(hod_u, 'last_name', '') or '').strip(),
        ]).strip()
        if hod_u
        else ''
    )
    if not hod_name:
        hod_name = str(getattr(hod_u, 'username', '') or '').strip() if hod_u else ''

    return Response(
        {
            'id': req.id,
            'status': req.status,
            'created_at': req.created_at.isoformat() if req.created_at else None,
            'routed_to': routed_to,
            'routing_warning': routing_warning,
            'department': (
                {
                    'id': getattr(dept, 'id', None),
                    'code': getattr(dept, 'code', None),
                    'name': getattr(dept, 'name', None),
                    'short_name': getattr(dept, 'short_name', None),
                }
                if dept is not None
                else None
            ),
            'hod': (
                {
                    'id': getattr(hod_u, 'id', None),
                    'username': getattr(hod_u, 'username', None),
                    'name': hod_name or None,
                }
                if getattr(req, 'hod_user_id', None) is not None
                else None
            ),
        }
    )


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def publish_requests_pending(request):
    """IQAC: list pending publish requests."""
    auth = _require_obe_master(request)
    if auth:
        return auth

    from .models import ObePublishRequest

    qs = (
        ObePublishRequest.objects.select_related('staff_user', 'academic_year')
        .filter(status='PENDING', hod_approved=True)
        .order_by('-created_at')
    )

    def staff_name(u):
        if not u:
            return None
        try:
            full = ' '.join([str(getattr(u, 'first_name', '') or '').strip(), str(getattr(u, 'last_name', '') or '').strip()]).strip()
            return full or getattr(u, 'username', None)
        except Exception:
            return getattr(u, 'username', None)

    out = []
    for r in qs:
        u = getattr(r, 'staff_user', None)
        out.append(
            {
                'id': r.id,
                'status': r.status,
                'assessment': r.assessment,
                'subject_code': r.subject_code,
                'subject_name': r.subject_name,
                'reason': r.reason,
                'requested_at': r.created_at.isoformat() if r.created_at else None,
                'academic_year': {
                    'id': r.academic_year_id,
                    'name': getattr(getattr(r, 'academic_year', None), 'name', None),
                }
                if r.academic_year_id
                else None,
                'staff': {
                    'id': getattr(u, 'id', None),
                    'username': getattr(u, 'username', None),
                    'name': staff_name(u),
                },
            }
        )

    return Response({'results': out})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def publish_requests_history(request):
    """IQAC: list reviewed publish requests (approved/rejected)."""
    auth = _require_obe_master(request)
    if auth:
        return auth

    from .models import ObePublishRequest

    qp = getattr(request, 'query_params', {}) if hasattr(request, 'query_params') else request.GET
    statuses_raw = str(qp.get('statuses') or '').strip()
    if statuses_raw:
        statuses = [s.strip().upper() for s in statuses_raw.split(',') if s.strip()]
    else:
        statuses = ['APPROVED', 'REJECTED']
    statuses = [s for s in statuses if s in {'APPROVED', 'REJECTED'}]
    if not statuses:
        statuses = ['APPROVED', 'REJECTED']

    try:
        limit = int(qp.get('limit') or 200)
    except Exception:
        limit = 200
    limit = max(1, min(500, limit))

    qs = (
        ObePublishRequest.objects.select_related('staff_user', 'academic_year', 'reviewed_by')
        .filter(status__in=statuses, hod_approved=True)
        .order_by('-updated_at')
    )[:limit]

    def user_name(u):
        if not u:
            return None
        try:
            full = ' '.join([
                str(getattr(u, 'first_name', '') or '').strip(),
                str(getattr(u, 'last_name', '') or '').strip(),
            ]).strip()
            return full or getattr(u, 'username', None)
        except Exception:
            return getattr(u, 'username', None)

    def staff_department(u):
        try:
            sp = getattr(u, 'staff_profile', None)
            if not sp:
                return None
            dept = None
            try:
                dept = sp.get_current_department()
            except Exception:
                dept = getattr(sp, 'department', None)
            return getattr(dept, 'name', None) if dept else None
        except Exception:
            return None

    out = []
    for r in qs:
        u = getattr(r, 'staff_user', None)
        reviewer = getattr(r, 'reviewed_by', None)
        out.append(
            {
                'id': r.id,
                'status': r.status,
                'assessment': r.assessment,
                'subject_code': r.subject_code,
                'subject_name': r.subject_name,
                'reason': r.reason,
                'requested_at': r.created_at.isoformat() if r.created_at else None,
                'academic_year': {
                    'id': r.academic_year_id,
                    'name': getattr(getattr(r, 'academic_year', None), 'name', None),
                }
                if r.academic_year_id
                else None,
                'staff': {
                    'id': getattr(u, 'id', None),
                    'username': getattr(u, 'username', None),
                    'name': user_name(u),
                    'department': staff_department(u),
                },
                'reviewed_by': {
                    'id': getattr(reviewer, 'id', None),
                    'username': getattr(reviewer, 'username', None),
                    'name': user_name(reviewer),
                }
                if reviewer
                else None,
                'reviewed_at': r.reviewed_at.isoformat() if getattr(r, 'reviewed_at', None) else None,
                'approved_until': r.approved_until.isoformat() if getattr(r, 'approved_until', None) else None,
            }
        )

    return Response({'results': out})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def publish_requests_pending_count(request):
    try:
        auth = _require_obe_master(request)
        if auth:
            return auth
        from .models import ObePublishRequest
        return Response({'pending': int(ObePublishRequest.objects.filter(status='PENDING', hod_approved=True).count())})
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception('Error in publish_requests_pending_count: %s', e)
        return Response({'detail': 'Internal error', 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def publish_requests_hod_pending(request):
    """HOD: list pending publish requests assigned to this HOD (pre-approval stage)."""
    user = getattr(request, 'user', None)
    staff_profile = getattr(user, 'staff_profile', None)
    if staff_profile is None:
        return Response({'detail': 'Not a staff member.'}, status=status.HTTP_403_FORBIDDEN)

    from .models import ObePublishRequest

    qs = ObePublishRequest.objects.select_related('staff_user', 'academic_year').filter(
        status='PENDING',
        hod_approved=False,
        hod_user=user,
    ).order_by('-created_at')

    def staff_name(u):
        if not u:
            return None
        try:
            full = ' '.join([str(getattr(u, 'first_name', '') or '').strip(), str(getattr(u, 'last_name', '') or '').strip()]).strip()
            return full or getattr(u, 'username', None)
        except Exception:
            return getattr(u, 'username', None)

    out = []
    for r in qs:
        u = getattr(r, 'staff_user', None)
        out.append(
            {
                'id': r.id,
                'status': r.status,
                'assessment': r.assessment,
                'subject_code': r.subject_code,
                'subject_name': r.subject_name,
                'reason': r.reason,
                'requested_at': r.created_at.isoformat() if r.created_at else None,
                'academic_year': {
                    'id': r.academic_year_id,
                    'name': getattr(getattr(r, 'academic_year', None), 'name', None),
                }
                if r.academic_year_id
                else None,
                'staff': {
                    'id': getattr(u, 'id', None),
                    'username': getattr(u, 'username', None),
                    'name': staff_name(u),
                },
            }
        )
    return Response({'results': out})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def publish_requests_hod_pending_count(request):
    """HOD: count pending publish requests assigned to this HOD."""
    user = getattr(request, 'user', None)
    staff_profile = getattr(user, 'staff_profile', None)
    if staff_profile is None:
        return Response({'pending': 0})

    from .models import ObePublishRequest
    return Response({'pending': int(ObePublishRequest.objects.filter(status='PENDING', hod_approved=False, hod_user=user).count())})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def publish_request_hod_approve(request, req_id: int):
    """HOD: approve/forward a publish request to IQAC."""
    user = getattr(request, 'user', None)
    staff_profile = getattr(user, 'staff_profile', None)
    if staff_profile is None:
        return Response({'detail': 'Not a staff member.'}, status=status.HTTP_403_FORBIDDEN)

    from .models import ObePublishRequest
    row = ObePublishRequest.objects.filter(id=req_id).first()
    if not row:
        return Response({'detail': 'Request not found.'}, status=status.HTTP_404_NOT_FOUND)

    if getattr(row, 'hod_user_id', None) != getattr(user, 'id', None):
        return Response({'detail': 'Not assigned to you.'}, status=status.HTTP_403_FORBIDDEN)

    if str(getattr(row, 'status', '') or '').upper() != 'PENDING':
        return Response({'detail': 'Request is not pending.'}, status=status.HTTP_400_BAD_REQUEST)

    if bool(getattr(row, 'hod_approved', False)):
        return Response({'status': 'forwarded'})

    row.hod_approved = True
    row.hod_reviewed_by = user
    row.hod_reviewed_at = timezone.now()
    row.save(update_fields=['hod_approved', 'hod_reviewed_by', 'hod_reviewed_at', 'updated_at'])
    return Response({'status': 'forwarded'})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def publish_request_hod_reject(request, req_id: int):
    """HOD: reject a publish request (does not reach IQAC)."""
    user = getattr(request, 'user', None)
    staff_profile = getattr(user, 'staff_profile', None)
    if staff_profile is None:
        return Response({'detail': 'Not a staff member.'}, status=status.HTTP_403_FORBIDDEN)

    from .models import ObePublishRequest
    row = ObePublishRequest.objects.filter(id=req_id).first()
    if not row:
        return Response({'detail': 'Request not found.'}, status=status.HTTP_404_NOT_FOUND)

    if getattr(row, 'hod_user_id', None) != getattr(user, 'id', None):
        return Response({'detail': 'Not assigned to you.'}, status=status.HTTP_403_FORBIDDEN)

    if str(getattr(row, 'status', '') or '').upper() != 'PENDING':
        return Response({'detail': 'Request is not pending.'}, status=status.HTTP_400_BAD_REQUEST)

    # Mark as rejected at HOD stage.
    row.status = 'REJECTED'
    row.hod_approved = False
    row.hod_reviewed_by = user
    row.hod_reviewed_at = timezone.now()
    row.save(update_fields=['status', 'hod_approved', 'hod_reviewed_by', 'hod_reviewed_at', 'updated_at'])
    return Response({'status': 'rejected'})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def publish_request_approve(request, req_id: int):
    auth = _require_obe_master(request)
    if auth:
        return auth

    from .models import ObePublishRequest
    row = ObePublishRequest.objects.filter(id=req_id).first()
    if not row:
        return Response({'detail': 'Request not found.'}, status=status.HTTP_404_NOT_FOUND)

    minutes = (request.data or {}).get('window_minutes')
    try:
        minutes_int = int(minutes) if minutes is not None else 120
    except Exception:
        minutes_int = 120

    row.mark_approved(request.user, window_minutes=minutes_int)
    row.save(update_fields=['status', 'approved_until', 'reviewed_by', 'reviewed_at', 'updated_at'])
    return Response({'status': 'approved', 'approved_until': row.approved_until.isoformat() if row.approved_until else None})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def publish_request_reject(request, req_id: int):
    auth = _require_obe_master(request)
    if auth:
        return auth

    from .models import ObePublishRequest
    row = ObePublishRequest.objects.filter(id=req_id).first()
    if not row:
        return Response({'detail': 'Request not found.'}, status=status.HTTP_404_NOT_FOUND)

    row.mark_rejected(request.user)
    row.save(update_fields=['status', 'approved_until', 'reviewed_by', 'reviewed_at', 'updated_at'])
    return Response({'status': 'rejected'})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_request_create(request):
    """Faculty: create a scoped edit request after publishing."""
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    body = request.data if isinstance(request.data, dict) else {}
    assessment = str(body.get('assessment') or '').strip().lower()
    subject_code = str(body.get('subject_code') or '').strip()
    reason = str(body.get('reason') or '').strip()
    scope_raw = str(body.get('scope') or '').strip().upper()
    ta_id = None
    try:
        if body.get('teaching_assignment_id') is not None:
            ta_id = int(str(body.get('teaching_assignment_id')))
    except Exception:
        ta_id = None

    if not _is_valid_mark_assessment_key(assessment, allow_documents=True, allow_cqi=True):
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)
    if not subject_code:
        return Response({'detail': 'subject_code is required.'}, status=status.HTTP_400_BAD_REQUEST)

    if scope_raw in {'MARKS', 'MARKS_ENTRY', 'MARK_ENTRY', 'TABLE'}:
        scope = 'MARK_ENTRY'
    elif scope_raw in {'MARK_MANAGER', 'MANAGER'}:
        scope = 'MARK_MANAGER'
    else:
        return Response({'detail': 'scope is required (MARK_ENTRY or MARK_MANAGER).'}, status=status.HTTP_400_BAD_REQUEST)

    info = _get_due_schedule_for_request(request, subject_code=subject_code, assessment=assessment, teaching_assignment_id=ta_id)
    academic_year = info.get('academic_year')
    if academic_year is None:
        return Response({'detail': 'Unable to resolve academic year for this subject.'}, status=status.HTTP_400_BAD_REQUEST)

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject_code, teaching_assignment_id=ta_id)
    section_name = _resolve_section_name_from_ta(ta)

    from .models import ObeEditRequest

    # Resolve department HOD for this staff member + academic year.
    # Primary: academics.DepartmentRole (department + academic_year).
    # Fallbacks: RoleAssignment('HOD') and accounts.Role('HOD') for staff in same department.
    hod_user = None
    dept = getattr(staff_profile, 'current_department', None) or getattr(staff_profile, 'department', None)
    if dept is None:
        return Response(
            {
                'detail': 'Unable to resolve your department for HOD routing.',
                'how_to_fix': [
                    'Ensure this staff has an active department assignment (or department set) in Academics.',
                    'Then re-try submitting the edit request.',
                ],
            },
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        from academics.models import DepartmentRole

        if dept is not None and academic_year is not None:
            hod_role = (
                DepartmentRole.objects.filter(
                    department=dept,
                    academic_year=academic_year,
                    role='HOD',
                    is_active=True,
                )
                .select_related('staff__user')
                .first()
            )
            if hod_role is not None:
                hod_user = getattr(getattr(hod_role, 'staff', None), 'user', None)

        # Backward-compat / convenience: if a HOD is configured for the department
        # but not explicitly mapped for this academic year, pick the latest active HOD.
        if hod_user is None and dept is not None:
            hod_role_any = (
                DepartmentRole.objects.filter(
                    department=dept,
                    role='HOD',
                    is_active=True,
                )
                .select_related('staff__user', 'academic_year')
                .order_by('-academic_year_id', '-id')
                .first()
            )
            if hod_role_any is not None:
                hod_user = getattr(getattr(hod_role_any, 'staff', None), 'user', None)
    except Exception:
        hod_user = None

    if hod_user is None and dept is not None:
        try:
            from academics.models import StaffProfile, RoleAssignment
            from django.db.models import Q

            dept_staff_qs = StaffProfile.objects.filter(
                Q(department=dept)
                | Q(department_assignments__department=dept, department_assignments__end_date__isnull=True)
            ).select_related('user').distinct()

            # Prefer explicit time-bound RoleAssignment first.
            ra = (
                RoleAssignment.objects.filter(
                    staff__in=dept_staff_qs,
                    role_name__iexact='HOD',
                    end_date__isnull=True,
                )
                .select_related('staff__user')
                .order_by('-start_date')
                .first()
            )
            if ra is not None:
                hod_user = getattr(getattr(ra, 'staff', None), 'user', None)
        except Exception:
            hod_user = None

    routing_warning = None
    if hod_user is None:
        dept_name = (
            str(getattr(dept, 'short_name', '') or '').strip()
            or str(getattr(dept, 'name', '') or '').strip()
            or str(getattr(dept, 'code', '') or '').strip()
            or 'your department'
        )

        # LCA-related requests must go to HOD first.
        if assessment == 'lca':
            return Response(
                {
                    'detail': f'No active HOD configured for {dept_name}. This request must be routed to HOD first.',
                    'how_to_fix': [
                        'Configure an active HOD for this department in Academics (DepartmentRole / RoleAssignment).',
                        'Then re-submit the request.',
                    ],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        routing_warning = f'No active HOD configured for {dept_name}. Request will be sent directly to IQAC.'

    if hod_user is None and dept is not None:
        try:
            from academics.models import StaffProfile
            from django.db.models import Q

            dept_staff_qs = StaffProfile.objects.filter(
                Q(department=dept)
                | Q(department_assignments__department=dept, department_assignments__end_date__isnull=True)
            ).select_related('user').distinct()

            sp_hod = (
                dept_staff_qs.filter(user__roles__name__iexact='HOD')
                .exclude(user=request.user)
                .order_by('id')
                .first()
            )
            if sp_hod is not None:
                hod_user = getattr(sp_hod, 'user', None)
        except Exception:
            hod_user = None

    existing = ObeEditRequest.objects.filter(
        staff_user=request.user,
        academic_year=academic_year,
        subject_code=subject_code,
        assessment=assessment,
        scope=scope,
        status='PENDING',
    ).order_by('-created_at').first()

    if existing:
        existing.reason = reason
        existing.subject_name = existing.subject_name or (getattr(info.get('schedule'), 'subject_name', '') or '')
        if getattr(existing, 'teaching_assignment_id', None) is None and ta is not None:
            existing.teaching_assignment = ta
        if not getattr(existing, 'section_name', None) and section_name:
            existing.section_name = section_name
        # If this is an older request with no HOD routing info yet, assign it now.
        try:
            if getattr(existing, 'hod_user_id', None) is None and hod_user is not None:
                existing.hod_user = hod_user
                existing.hod_approved = False
        except Exception:
            pass
        existing.save(update_fields=['reason', 'subject_name', 'teaching_assignment', 'section_name', 'hod_user', 'hod_approved', 'updated_at'])
        req = existing
    else:
        req = ObeEditRequest.objects.create(
            staff_user=request.user,
            academic_year=academic_year,
            subject_code=subject_code,
            subject_name=(getattr(info.get('schedule'), 'subject_name', '') or ''),
            assessment=assessment,
            scope=scope,
            reason=reason,
            teaching_assignment=ta,
            section_name=section_name,
            hod_user=hod_user,
            hod_approved=(False if hod_user is not None else True),
        )

    routed_to = 'HOD' if getattr(req, 'hod_user_id', None) is not None and not bool(getattr(req, 'hod_approved', True)) else 'IQAC'

    hod_u = getattr(req, 'hod_user', None)
    hod_name = (
        ' '.join([
            str(getattr(hod_u, 'first_name', '') or '').strip(),
            str(getattr(hod_u, 'last_name', '') or '').strip(),
        ]).strip()
        if hod_u
        else ''
    )
    if not hod_name:
        hod_name = str(getattr(hod_u, 'username', '') or '').strip() if hod_u else ''

    # Resolve the submitting staff's display name for the approver message
    _staff_display = (
        ' '.join([
            str(getattr(request.user, 'first_name', '') or '').strip(),
            str(getattr(request.user, 'last_name', '') or '').strip(),
        ]).strip()
        or str(getattr(request.user, 'username', '') or '').strip()
    )
    sp = getattr(staff_profile, 'name', None) or getattr(staff_profile, 'full_name', None) if staff_profile else None
    if sp:
        _staff_display = str(sp).strip() or _staff_display

    # Notify the HOD (or IQAC) via WhatsApp using their profile mobile number
    try:
        from OBE.services.edit_request_notifications import notify_approver_of_new_request
        notify_approver_of_new_request(
            req,
            hod_user=hod_u,
            routed_to=routed_to,
            department=dept,
            staff_name=_staff_display,
        )
    except Exception:
        logger.exception('Failed to send approver WhatsApp notification for edit_request=%s', req.id)

    return Response(
        {
            'id': req.id,
            'status': req.status,
            'scope': req.scope,
            'created_at': req.created_at.isoformat() if req.created_at else None,
            'routed_to': routed_to,
            'routing_warning': routing_warning,
            'department': (
                {
                    'id': getattr(dept, 'id', None),
                    'code': getattr(dept, 'code', None),
                    'name': getattr(dept, 'name', None),
                    'short_name': getattr(dept, 'short_name', None),
                }
                if dept is not None
                else None
            ),
            'hod': (
                {
                    'id': getattr(hod_u, 'id', None),
                    'username': getattr(hod_u, 'username', None),
                    'name': hod_name or None,
                }
                if getattr(req, 'hod_user_id', None) is not None
                else None
            ),
        }
    )


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_requests_my_latest(request):
    """Faculty: fetch the latest edit request for this staff user (for status polling).

    Query params:
    - assessment (required)
    - subject_code (required)
    - scope (required: MARK_ENTRY or MARK_MANAGER)
    - teaching_assignment_id (optional)
    """
    staff_profile, err = _faculty_only(request)
    if err:
        return err

    qp = getattr(request, 'query_params', {}) if hasattr(request, 'query_params') else request.GET
    assessment = str(qp.get('assessment') or '').strip().lower()
    subject_code = str(qp.get('subject_code') or '').strip()
    scope_raw = str(qp.get('scope') or '').strip().upper()
    ta_id = None
    try:
        if qp.get('teaching_assignment_id') is not None and str(qp.get('teaching_assignment_id')).strip() != '':
            ta_id = int(str(qp.get('teaching_assignment_id')))
    except Exception:
        ta_id = None

    if not _is_valid_mark_assessment_key(assessment, allow_documents=True, allow_cqi=True):
        return Response({'detail': 'Invalid assessment.'}, status=status.HTTP_400_BAD_REQUEST)
    if not subject_code:
        return Response({'detail': 'subject_code is required.'}, status=status.HTTP_400_BAD_REQUEST)

    if scope_raw in {'MARKS', 'MARKS_ENTRY', 'MARK_ENTRY', 'TABLE'}:
        scope = 'MARK_ENTRY'
    elif scope_raw in {'MARK_MANAGER', 'MANAGER'}:
        scope = 'MARK_MANAGER'
    else:
        return Response({'detail': 'scope is required (MARK_ENTRY or MARK_MANAGER).'}, status=status.HTTP_400_BAD_REQUEST)

    info = _get_due_schedule_for_request(request, subject_code=subject_code, assessment=assessment, teaching_assignment_id=ta_id)
    academic_year = info.get('academic_year')
    if academic_year is None:
        return Response({'detail': 'Unable to resolve academic year for this subject.'}, status=status.HTTP_400_BAD_REQUEST)

    ta = _resolve_staff_teaching_assignment(request, subject_code=subject_code, teaching_assignment_id=ta_id)

    from .models import ObeEditRequest

    qs = ObeEditRequest.objects.filter(
        staff_user=request.user,
        academic_year=academic_year,
        subject_code=subject_code,
        assessment=assessment,
        scope=scope,
    )
    if ta is not None:
        qs = qs.filter(teaching_assignment=ta)
    req = qs.order_by('-updated_at').first()
    if not req:
        return Response({'result': None})

    reviewer = getattr(req, 'reviewed_by', None)
    now = timezone.now()
    approved_until = getattr(req, 'approved_until', None)

    return Response(
        {
            'result': {
                'id': req.id,
                'status': req.status,
                'assessment': req.assessment,
                'scope': req.scope,
                'subject_code': req.subject_code,
                'reason': req.reason,
                'requested_at': req.created_at.isoformat() if req.created_at else None,
                'updated_at': req.updated_at.isoformat() if req.updated_at else None,
                'reviewed_at': req.reviewed_at.isoformat() if req.reviewed_at else None,
                'approved_until': approved_until.isoformat() if approved_until else None,
                'is_active': bool(approved_until and approved_until > now),
                'reviewed_by': {
                    'id': getattr(reviewer, 'id', None),
                    'username': getattr(reviewer, 'username', None),
                    'name': (
                        ' '.join([
                            str(getattr(reviewer, 'first_name', '') or '').strip(),
                            str(getattr(reviewer, 'last_name', '') or '').strip(),
                        ]).strip()
                        if reviewer
                        else None
                    )
                    or (getattr(reviewer, 'username', None) if reviewer else None),
                }
                if reviewer
                else None,
            }
        }
    )


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_requests_pending(request):
    """IQAC: list pending edit requests."""
    auth = _require_obe_master(request)
    if auth:
        return auth

    from .models import ObeEditRequest

    qp = getattr(request, 'query_params', {}) if hasattr(request, 'query_params') else request.GET
    scope_raw = str(qp.get('scope') or '').strip().upper()
    scope_filter = None
    if scope_raw in {'MARK_ENTRY', 'MARK_MANAGER'}:
        scope_filter = scope_raw

    qs = ObeEditRequest.objects.select_related('staff_user', 'academic_year').filter(status='PENDING', hod_approved=True)
    if scope_filter:
        qs = qs.filter(scope=scope_filter)
    qs = qs.order_by('-created_at')

    def staff_name(u):
        if not u:
            return None
        try:
            full = ' '.join([str(getattr(u, 'first_name', '') or '').strip(), str(getattr(u, 'last_name', '') or '').strip()]).strip()
            return full or getattr(u, 'username', None)
        except Exception:
            return getattr(u, 'username', None)

    out = []
    for r in qs:
        u = getattr(r, 'staff_user', None)
        out.append(
            {
                'id': r.id,
                'status': r.status,
                'assessment': r.assessment,
                'scope': r.scope,
                'subject_code': r.subject_code,
                'subject_name': r.subject_name,
                'reason': r.reason,
                'requested_at': r.created_at.isoformat() if r.created_at else None,
                'academic_year': {
                    'id': r.academic_year_id,
                    'name': getattr(getattr(r, 'academic_year', None), 'name', None),
                }
                if r.academic_year_id
                else None,
                'staff': {
                    'id': getattr(u, 'id', None),
                    'username': getattr(u, 'username', None),
                    'name': staff_name(u),
                },
            }
        )

    return Response({'results': out})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_requests_history(request):
    """IQAC: list reviewed edit requests (approved/rejected)."""
    auth = _require_obe_master(request)
    if auth:
        return auth

    from .models import ObeEditRequest

    qp = getattr(request, 'query_params', {}) if hasattr(request, 'query_params') else request.GET
    statuses_raw = str(qp.get('statuses') or '').strip()
    if statuses_raw:
        statuses = [s.strip().upper() for s in statuses_raw.split(',') if s.strip()]
    else:
        statuses = ['APPROVED', 'REJECTED']
    statuses = [s for s in statuses if s in {'APPROVED', 'REJECTED'}]
    if not statuses:
        statuses = ['APPROVED', 'REJECTED']

    scope_raw = str(qp.get('scope') or '').strip().upper()
    scope_filter = None
    if scope_raw in {'MARK_ENTRY', 'MARK_MANAGER'}:
        scope_filter = scope_raw

    try:
        limit = int(qp.get('limit') or 200)
    except Exception:
        limit = 200
    limit = max(1, min(500, limit))

    qs = ObeEditRequest.objects.select_related('staff_user', 'academic_year', 'reviewed_by').filter(status__in=statuses, hod_approved=True)
    if scope_filter:
        qs = qs.filter(scope=scope_filter)
    qs = qs.order_by('-updated_at')[:limit]

    def user_name(u):
        if not u:
            return None
        try:
            full = ' '.join([
                str(getattr(u, 'first_name', '') or '').strip(),
                str(getattr(u, 'last_name', '') or '').strip(),
            ]).strip()
            return full or getattr(u, 'username', None)
        except Exception:
            return getattr(u, 'username', None)

    out = []
    for r in qs:
        u = getattr(r, 'staff_user', None)
        reviewer = getattr(r, 'reviewed_by', None)
        out.append(
            {
                'id': r.id,
                'status': r.status,
                'assessment': r.assessment,
                'scope': r.scope,
                'subject_code': r.subject_code,
                'subject_name': r.subject_name,
                'reason': r.reason,
                'requested_at': r.created_at.isoformat() if r.created_at else None,
                'academic_year': {
                    'id': r.academic_year_id,
                    'name': getattr(getattr(r, 'academic_year', None), 'name', None),
                }
                if r.academic_year_id
                else None,
                'staff': {
                    'id': getattr(u, 'id', None),
                    'username': getattr(u, 'username', None),
                    'name': user_name(u),
                },
                'reviewed_by': {
                    'id': getattr(reviewer, 'id', None),
                    'username': getattr(reviewer, 'username', None),
                    'name': user_name(reviewer),
                }
                if reviewer
                else None,
                'reviewed_at': r.reviewed_at.isoformat() if getattr(r, 'reviewed_at', None) else None,
                'approved_until': r.approved_until.isoformat() if getattr(r, 'approved_until', None) else None,
            }
        )

    return Response({'results': out})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_requests_hod_pending(request):
    """HOD: list pending edit requests assigned to this HOD (pre-approval stage)."""
    user = getattr(request, 'user', None)
    staff_profile = getattr(user, 'staff_profile', None)
    if staff_profile is None:
        return Response({'detail': 'Not a staff member.'}, status=status.HTTP_403_FORBIDDEN)

    # Authorization is assignment-based: a user can only see rows where hod_user=request.user.

    from .models import ObeEditRequest

    qp = getattr(request, 'query_params', {}) if hasattr(request, 'query_params') else request.GET
    scope_raw = str(qp.get('scope') or '').strip().upper()
    scope_filter = None
    if scope_raw in {'MARK_ENTRY', 'MARK_MANAGER'}:
        scope_filter = scope_raw

    qs = ObeEditRequest.objects.select_related('staff_user', 'academic_year').filter(
        status='PENDING',
        hod_approved=False,
        hod_user=user,
    )
    if scope_filter:
        qs = qs.filter(scope=scope_filter)
    qs = qs.order_by('-created_at')

    def staff_name(u):
        if not u:
            return None
        try:
            full = ' '.join([str(getattr(u, 'first_name', '') or '').strip(), str(getattr(u, 'last_name', '') or '').strip()]).strip()
            return full or getattr(u, 'username', None)
        except Exception:
            return getattr(u, 'username', None)

    out = []
    for r in qs:
        u = getattr(r, 'staff_user', None)
        out.append(
            {
                'id': r.id,
                'status': r.status,
                'assessment': r.assessment,
                'scope': r.scope,
                'subject_code': r.subject_code,
                'subject_name': r.subject_name,
                'reason': r.reason,
                'requested_at': r.created_at.isoformat() if r.created_at else None,
                'academic_year': {
                    'id': r.academic_year_id,
                    'name': getattr(getattr(r, 'academic_year', None), 'name', None),
                }
                if r.academic_year_id
                else None,
                'staff': {
                    'id': getattr(u, 'id', None),
                    'username': getattr(u, 'username', None),
                    'name': staff_name(u),
                },
            }
        )

    return Response({'results': out})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_requests_pending_count(request):
    auth = _require_obe_master(request)
    if auth:
        return auth
    from .models import ObeEditRequest
    qp = getattr(request, 'query_params', {}) if hasattr(request, 'query_params') else request.GET
    scope_raw = str(qp.get('scope') or '').strip().upper()
    qs = ObeEditRequest.objects.filter(status='PENDING', hod_approved=True)
    if scope_raw in {'MARK_ENTRY', 'MARK_MANAGER'}:
        qs = qs.filter(scope=scope_raw)
    return Response({'pending': int(qs.count())})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_requests_hod_pending_count(request):
    """HOD: count pending edit requests assigned to this HOD."""
    user = getattr(request, 'user', None)
    staff_profile = getattr(user, 'staff_profile', None)
    if staff_profile is None:
        return Response({'pending': 0})

    # Authorization is assignment-based: count is scoped to hod_user=request.user.

    from .models import ObeEditRequest
    qp = getattr(request, 'query_params', {}) if hasattr(request, 'query_params') else request.GET
    scope_raw = str(qp.get('scope') or '').strip().upper()

    qs = ObeEditRequest.objects.filter(status='PENDING', hod_approved=False, hod_user=user)
    if scope_raw in {'MARK_ENTRY', 'MARK_MANAGER'}:
        qs = qs.filter(scope=scope_raw)
    return Response({'pending': int(qs.count())})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_request_hod_approve(request, req_id: int):
    """HOD: approve/forward an edit request to IQAC."""
    user = getattr(request, 'user', None)
    staff_profile = getattr(user, 'staff_profile', None)
    if staff_profile is None:
        return Response({'detail': 'Not a staff member.'}, status=status.HTTP_403_FORBIDDEN)

    # Authorization is enforced by assignment check below (row.hod_user == request.user).

    from .models import ObeEditRequest
    row = ObeEditRequest.objects.filter(id=req_id).first()
    if not row:
        return Response({'detail': 'Request not found.'}, status=status.HTTP_404_NOT_FOUND)

    if getattr(row, 'hod_user_id', None) != getattr(user, 'id', None):
        return Response({'detail': 'Not assigned to you.'}, status=status.HTTP_403_FORBIDDEN)

    if str(getattr(row, 'status', '') or '').upper() != 'PENDING':
        return Response({'detail': 'Request is not pending.'}, status=status.HTTP_400_BAD_REQUEST)

    if bool(getattr(row, 'hod_approved', False)):
        return Response({'status': 'forwarded'})

    row.hod_approved = True
    row.hod_reviewed_by = user
    row.hod_reviewed_at = timezone.now()
    row.save(update_fields=['hod_approved', 'hod_reviewed_by', 'hod_reviewed_at', 'updated_at'])
    return Response({'status': 'forwarded'})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_request_hod_reject(request, req_id: int):
    """HOD: reject an edit request (does not reach IQAC)."""
    user = getattr(request, 'user', None)
    staff_profile = getattr(user, 'staff_profile', None)
    if staff_profile is None:
        return Response({'detail': 'Not a staff member.'}, status=status.HTTP_403_FORBIDDEN)

    # Authorization is enforced by assignment check below (row.hod_user == request.user).

    from .models import ObeEditRequest
    row = ObeEditRequest.objects.filter(id=req_id).first()
    if not row:
        return Response({'detail': 'Request not found.'}, status=status.HTTP_404_NOT_FOUND)

    if getattr(row, 'hod_user_id', None) != getattr(user, 'id', None):
        return Response({'detail': 'Not assigned to you.'}, status=status.HTTP_403_FORBIDDEN)

    if str(getattr(row, 'status', '') or '').upper() != 'PENDING':
        return Response({'detail': 'Request is not pending.'}, status=status.HTTP_400_BAD_REQUEST)

    if bool(getattr(row, 'hod_approved', False)):
        return Response({'detail': 'Already forwarded to IQAC.'}, status=status.HTTP_400_BAD_REQUEST)

    row.mark_rejected(user)
    row.hod_reviewed_by = user
    row.hod_reviewed_at = timezone.now()
    row.save(update_fields=['status', 'approved_until', 'reviewed_by', 'reviewed_at', 'hod_reviewed_by', 'hod_reviewed_at', 'updated_at'])
    return Response({'status': 'rejected'})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_request_approve(request, req_id: int):
    auth = _require_obe_master(request)
    if auth:
        return auth

    from .models import ObeEditRequest
    row = ObeEditRequest.objects.filter(id=req_id).first()
    if not row:
        return Response({'detail': 'Request not found.'}, status=status.HTTP_404_NOT_FOUND)

    # Enforce HOD pre-approval when the request is routed to an HOD.
    if getattr(row, 'hod_user_id', None) is not None and not bool(getattr(row, 'hod_approved', True)):
        return Response({'detail': 'Awaiting HOD approval.'}, status=status.HTTP_400_BAD_REQUEST)

    minutes = (request.data or {}).get('window_minutes')
    try:
        minutes_int = int(minutes) if minutes is not None else 120
    except Exception:
        minutes_int = 120

    row.mark_approved(request.user, window_minutes=minutes_int)
    row.save(update_fields=['status', 'approved_until', 'reviewed_by', 'reviewed_at', 'updated_at'])

    try:
        from .services.edit_request_notifications import notify_edit_request_approved

        notify_edit_request_approved(row)
    except Exception:
        logger.exception('Failed to send approval notifications for OBE edit request id=%s', getattr(row, 'id', None))

    # SPECIAL course enabled-assessment selection edit requests are surfaced in the central OBE queue
    # as assessment='model' + scope='MARK_MANAGER' with a distinct reason. When IQAC approves here,
    # also approve the underlying SPECIAL selection edit request so faculty can edit the "Select Exams" panel.
    try:
        assessment_key = str(getattr(row, 'assessment', '') or '').strip().lower()
        scope_key = str(getattr(row, 'scope', '') or '').strip().upper()
        reason_txt = str(getattr(row, 'reason', '') or '')
        is_special_selection_req = (
            assessment_key == 'model'
            and scope_key == 'MARK_MANAGER'
            and 'enabled assessments (special course global selection)' in reason_txt.lower()
        )
    except Exception:
        is_special_selection_req = False

    if is_special_selection_req:
        try:
            from academics.models import SpecialCourseAssessmentEditRequest
            from django.db.models import Q

            staff_profile = getattr(getattr(row, 'staff_user', None), 'staff_profile', None)
            academic_year = getattr(row, 'academic_year', None)
            subject_code = str(getattr(row, 'subject_code', '') or '').strip()

            if staff_profile is not None and academic_year is not None and subject_code:
                latest = (
                    SpecialCourseAssessmentEditRequest.objects.filter(
                        requested_by=staff_profile,
                        selection__academic_year=academic_year,
                    )
                    .filter(
                        Q(selection__curriculum_row__course_code__iexact=subject_code)
                        | Q(selection__curriculum_row__master__course_code__iexact=subject_code)
                    )
                    .order_by('-requested_at')
                    .first()
                )
                if latest is not None:
                    latest.status = SpecialCourseAssessmentEditRequest.STATUS_APPROVED
                    latest.reviewed_by = request.user
                    latest.reviewed_at = timezone.now()
                    latest.used_at = None
                    latest.can_edit_until = getattr(row, 'approved_until', None)
                    latest.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'used_at', 'can_edit_until'])
        except Exception:
            # best-effort only
            pass

        # Do NOT update OBE mark-table locks for special-selection edit requests.
        return Response({'status': 'approved', 'scope': row.scope, 'approved_until': row.approved_until.isoformat() if row.approved_until else None})

    # Sync to authoritative lock row so the UI can use one source of truth.
    try:
        ta = getattr(row, 'teaching_assignment', None)
        academic_year = getattr(row, 'academic_year', None)
        section_name = str(getattr(row, 'section_name', '') or '').strip() or _resolve_section_name_from_ta(ta)

        lock = _upsert_mark_table_lock(
            staff_user=getattr(row, 'staff_user', None),
            subject_code=str(getattr(row, 'subject_code', '') or ''),
            subject_name=str(getattr(row, 'subject_name', '') or ''),
            assessment=str(getattr(row, 'assessment', '') or ''),
            teaching_assignment=ta,
            academic_year=academic_year,
            section_name=section_name,
            updated_by=getattr(getattr(request, 'user', None), 'id', None),
        )

        lock.is_published = True

        if str(getattr(row, 'scope', '')).upper() == 'MARK_ENTRY':
            lock.mark_entry_unblocked_until = getattr(row, 'approved_until', None)
        elif str(getattr(row, 'scope', '')).upper() == 'MARK_MANAGER':
            approved_until = getattr(row, 'approved_until', None)
            lock.mark_manager_unlocked_until = approved_until
            # New logic: when MARK_MANAGER is approved for edit, also unlock the marks table
            # from the published lock for the same window.
            lock.mark_entry_unblocked_until = approved_until
            lock.mark_manager_locked = False

        lock.recompute_blocks()
        lock.save(
            update_fields=[
                'is_published',
                'published_blocked',
                'mark_entry_blocked',
                'mark_manager_locked',
                'mark_entry_unblocked_until',
                'mark_manager_unlocked_until',
                'updated_by',
                'updated_at',
            ]
        )
    except OperationalError:
        pass

    return Response({'status': 'approved', 'scope': row.scope, 'approved_until': row.approved_until.isoformat() if row.approved_until else None})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_request_reject(request, req_id: int):
    auth = _require_obe_master(request)
    if auth:
        return auth

    from .models import ObeEditRequest
    row = ObeEditRequest.objects.filter(id=req_id).first()
    if not row:
        return Response({'detail': 'Request not found.'}, status=status.HTTP_404_NOT_FOUND)

    # Enforce HOD pre-approval when the request is routed to an HOD.
    if getattr(row, 'hod_user_id', None) is not None and not bool(getattr(row, 'hod_approved', True)):
        return Response({'detail': 'Awaiting HOD approval.'}, status=status.HTTP_400_BAD_REQUEST)

    row.mark_rejected(request.user)
    row.save(update_fields=['status', 'approved_until', 'reviewed_by', 'reviewed_at', 'updated_at'])

    # Mirror rejection into SPECIAL-course enabled-assessment selection edit requests when they were
    # routed through the OBE edit-request queue.
    try:
        assessment_key = str(getattr(row, 'assessment', '') or '').strip().lower()
        scope_key = str(getattr(row, 'scope', '') or '').strip().upper()
        reason_txt = str(getattr(row, 'reason', '') or '')
        is_special_selection_req = (
            assessment_key == 'model'
            and scope_key == 'MARK_MANAGER'
            and 'enabled assessments (special course global selection)' in reason_txt.lower()
        )
    except Exception:
        is_special_selection_req = False

    if is_special_selection_req:
        try:
            from academics.models import SpecialCourseAssessmentEditRequest
            from django.db.models import Q

            staff_profile = getattr(getattr(row, 'staff_user', None), 'staff_profile', None)
            academic_year = getattr(row, 'academic_year', None)
            subject_code = str(getattr(row, 'subject_code', '') or '').strip()

            if staff_profile is not None and academic_year is not None and subject_code:
                latest = (
                    SpecialCourseAssessmentEditRequest.objects.filter(
                        requested_by=staff_profile,
                        selection__academic_year=academic_year,
                    )
                    .filter(
                        Q(selection__curriculum_row__course_code__iexact=subject_code)
                        | Q(selection__curriculum_row__master__course_code__iexact=subject_code)
                    )
                    .order_by('-requested_at')
                    .first()
                )
                if latest is not None:
                    latest.status = SpecialCourseAssessmentEditRequest.STATUS_REJECTED
                    latest.reviewed_by = request.user
                    latest.reviewed_at = timezone.now()
                    latest.used_at = None
                    latest.can_edit_until = None
                    latest.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'used_at', 'can_edit_until'])
        except Exception:
            pass

    return Response({'status': 'rejected', 'scope': row.scope})


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_reset_notifications(request):
    """Get unread reset notifications for the authenticated staff member.
    
    Query params:
    - teaching_assignment_id (optional): filter for specific teaching assignment
    """
    user = getattr(request, 'user', None)
    staff_profile = getattr(user, 'staff_profile', None)
    if staff_profile is None:
        return Response({'notifications': []})

    from .models import IqacResetNotification
    from academics.models import TeachingAssignment

    ta_id = request.query_params.get('teaching_assignment_id')
    
    qs = IqacResetNotification.objects.filter(
        teaching_assignment__staff=staff_profile,
        teaching_assignment__is_active=True,
        is_read=False
    ).select_related('teaching_assignment__subject', 'teaching_assignment__section', 'reset_by')

    if ta_id:
        try:
            qs = qs.filter(teaching_assignment_id=int(ta_id))
        except (ValueError, TypeError):
            pass

    notifications = []
    for n in qs:
        ta = n.teaching_assignment
        notifications.append({
            'id': n.id,
            'teaching_assignment_id': ta.id,
            'subject_code': ta.subject.code if ta.subject else '',
            'subject_name': ta.subject.name if ta.subject else '',
            'section_name': getattr(ta.section, 'name', '') if ta.section else '',
            'assessment': n.assessment,
            'reset_at': n.reset_at.isoformat() if n.reset_at else None,
            'reset_by': getattr(n.reset_by, 'username', None) if n.reset_by else None,
        })

    return Response({'notifications': notifications})


@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def dismiss_reset_notifications(request):
    """Mark reset notifications as read.
    
    Body:
    - notification_ids: list of notification IDs to mark as read
    """
    user = getattr(request, 'user', None)
    staff_profile = getattr(user, 'staff_profile', None)
    if staff_profile is None:
        return Response({'detail': 'Not a staff member.'}, status=status.HTTP_403_FORBIDDEN)

    notification_ids = request.data.get('notification_ids', [])
    if not isinstance(notification_ids, list):
        return Response({'detail': 'notification_ids must be a list.'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import IqacResetNotification

    updated = IqacResetNotification.objects.filter(
        id__in=notification_ids,
        teaching_assignment__staff=staff_profile,
        is_read=False
    ).update(is_read=True, read_at=timezone.now())

    return Response({'status': 'ok', 'marked_read': updated})


# ---------------------------------------------------------------------------
# IQAC – Special Courses list (for CO Weights panel)
# ---------------------------------------------------------------------------
@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([AllowAny])
def iqac_special_courses_list(request):
    """Return all active teaching-assignments whose curriculum row has class_type = SPECIAL.

    The frontend SpecialCoWeightsPanel expects a list of objects with:
      id, subject_code, subject_name, section_name, academic_year, department,
      staff_name, co_weights
    """
    from academics.models import TeachingAssignment, AcademicYear
    from curriculum.models import CurriculumDepartment
    from .models import SpecialCourseCoWeights

    # Fetch active teaching assignments linked to SPECIAL curriculum rows
    qs = (
        TeachingAssignment.objects
        .filter(is_active=True, curriculum_row__isnull=False, curriculum_row__class_type='SPECIAL')
        .select_related(
            'curriculum_row', 'curriculum_row__department',
            'section', 'section__batch',
            'academic_year', 'staff', 'staff__user',
        )
        .order_by('-academic_year__name', 'curriculum_row__course_code')
    )

    # Pre-fetch CO weights
    ta_ids = [ta.id for ta in qs]
    co_weights_map = {}
    if ta_ids:
        for scw in SpecialCourseCoWeights.objects.filter(teaching_assignment_id__in=ta_ids):
            co_weights_map[scw.teaching_assignment_id] = scw.weights or {}

    results = []
    for ta in qs:
        cr = ta.curriculum_row
        sec = ta.section
        staff = ta.staff
        ay = ta.academic_year

        results.append({
            'id': ta.id,
            'subject_code': cr.course_code or '',
            'subject_name': cr.course_name or '',
            'section_name': f"{sec.batch.name if sec and sec.batch else ''} {getattr(sec, 'name', '')}".strip() if sec else '',
            'academic_year': str(ay) if ay else '',
            'department': cr.department.code if cr.department else '',
            'staff_name': staff.user.get_full_name() if staff and staff.user else (str(staff) if staff else ''),
            'co_weights': co_weights_map.get(ta.id, {}),
        })

    return Response({'results': results})


# ═══════════════════════════════════════════════════════════════════════
# IQAC Special Exam Config — manage which exams are enabled for SPECIAL
# ═══════════════════════════════════════════════════════════════════════

# Cycle definitions for each assessment group
_SPECIAL_EXAM_CYCLES = {
    'SSA': ['SSA1', 'SSA2'],
    'FA': ['FORMATIVE1', 'FORMATIVE2'],
    'CIA': ['CIA1', 'CIA2'],
    'MODEL': ['MODEL'],
}

# All valid SPECIAL exam keys
_ALL_SPECIAL_EXAM_KEYS = {'SSA1', 'SSA2', 'FORMATIVE1', 'FORMATIVE2', 'CIA1', 'CIA2', 'MODEL'}

# Canonical display order for exam keys
_SPECIAL_EXAM_ORDER = ['SSA1', 'SSA2', 'FORMATIVE1', 'FORMATIVE2', 'CIA1', 'CIA2', 'MODEL']
_SPECIAL_EXAM_ORDER_MAP = {k: i for i, k in enumerate(_SPECIAL_EXAM_ORDER)}

def _sort_special_exams(exams):
    """Sort exam keys in canonical cycle order (SSA1, SSA2, FA1, FA2, CIA1, CIA2, MODEL)."""
    return sorted(exams, key=lambda e: _SPECIAL_EXAM_ORDER_MAP.get(e, 99))

# Map exam keys → enabled_assessments keys
_EXAM_TO_EA_KEY = {
    'SSA1': 'ssa1',
    'SSA2': 'ssa2',
    'FORMATIVE1': 'formative1',
    'FORMATIVE2': 'formative2',
    'CIA1': 'cia1',
    'CIA2': 'cia2',
    'MODEL': 'model',
}


def _propagate_special_enabled_assessments(ea_keys: list):
    """Update enabled_assessments on all SPECIAL CurriculumMaster and CurriculumDepartment rows."""
    from curriculum.models import CurriculumMaster, CurriculumDepartment
    try:
        CurriculumMaster.objects.filter(
            class_type__iexact='SPECIAL',
        ).update(enabled_assessments=ea_keys)
        CurriculumDepartment.objects.filter(
            class_type__iexact='SPECIAL',
        ).update(enabled_assessments=ea_keys)
    except Exception:
        pass


@api_view(['GET', 'POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def iqac_special_exam_config(request):
    """GET/POST: manage which exams are enabled for SPECIAL class type.

    GET ?question_paper_type=CSD
        → { exams: ['SSA1', 'CIA1', ...] }

    POST { question_paper_type?: 'CSD', action: 'add'|'remove', exam_group: 'SSA'|'CIA'|'FA'|'MODEL' }
        → auto-resolves cycle (SSA→SSA1 or SSA2) and updates config.
        → also propagates to curriculum enabled_assessments.

    POST { question_paper_type?: 'CSD', action: 'set', exams: ['SSA1', 'CIA1', ...] }
        → directly set the full list.
    """
    auth = _require_obe_master_permission(request)
    if auth:
        return auth

    from .models import ObeQpPatternConfig

    qp_raw = str(request.GET.get('question_paper_type', '') or request.data.get('question_paper_type', '') or '').strip().upper()
    # For SPECIAL, patterns are stored with question_paper_type=NULL (CSD normalizes to empty)
    qp_type_val = None

    def _get_current_exams():
        """Return sorted list of exam keys that have ObeQpPatternConfig rows for SPECIAL."""
        rows = ObeQpPatternConfig.objects.filter(
            class_type='SPECIAL',
            question_paper_type=qp_type_val,
        ).values_list('exam', flat=True)
        return _sort_special_exams(set(e.upper() for e in rows if str(e or '').upper() in _ALL_SPECIAL_EXAM_KEYS))

    if request.method == 'GET':
        return Response({'exams': _get_current_exams()})

    # POST
    data = request.data if isinstance(request.data, dict) else {}
    action = str(data.get('action', '')).strip().lower()
    user_id = getattr(getattr(request, 'user', None), 'id', None)

    if action == 'set':
        # Directly set full exam list
        raw_exams = data.get('exams', [])
        if not isinstance(raw_exams, list):
            return Response({'detail': 'exams must be a list'}, status=400)
        new_exams = _sort_special_exams(set(
            e.upper() for e in [str(x or '').strip() for x in raw_exams]
            if e.upper() in _ALL_SPECIAL_EXAM_KEYS
        ))

        current = set(_get_current_exams())
        to_add = set(new_exams) - current
        to_remove = current - set(new_exams)

        # Remove deleted exams
        if to_remove:
            ObeQpPatternConfig.objects.filter(
                class_type='SPECIAL',
                question_paper_type=qp_type_val,
                exam__in=to_remove,
            ).delete()

        # Add new exams with empty patterns
        for ex in to_add:
            ObeQpPatternConfig.objects.update_or_create(
                class_type='SPECIAL',
                question_paper_type=qp_type_val,
                exam=ex,
                defaults={'pattern': {'marks': [], 'cos': []}, 'updated_by': user_id},
            )

        # Propagate to curriculum rows
        ea_keys = [_EXAM_TO_EA_KEY[e] for e in new_exams if e in _EXAM_TO_EA_KEY]
        _propagate_special_enabled_assessments(ea_keys)

        return Response({'exams': new_exams, 'enabled_assessments': ea_keys})

    elif action in ('add', 'remove'):
        exam_group = str(data.get('exam_group', '')).strip().upper()
        if exam_group not in _SPECIAL_EXAM_CYCLES:
            return Response({'detail': f"exam_group must be one of: {', '.join(_SPECIAL_EXAM_CYCLES.keys())}"}, status=400)

        current = _get_current_exams()
        current_set = set(current)
        cycle = _SPECIAL_EXAM_CYCLES[exam_group]

        if action == 'add':
            # Find the next exam in the cycle that isn't already present
            target = None
            for ex in cycle:
                if ex not in current_set:
                    target = ex
                    break
            if target is None:
                return Response({'detail': f"All {exam_group} exams are already configured."}, status=400)

            ObeQpPatternConfig.objects.update_or_create(
                class_type='SPECIAL',
                question_paper_type=qp_type_val,
                exam=target,
                defaults={'pattern': {'marks': [], 'cos': []}, 'updated_by': user_id},
            )
            new_exams = _sort_special_exams(current_set | {target})

        else:  # remove
            # Find the highest cycle exam in the group that exists
            target = None
            for ex in reversed(cycle):
                if ex in current_set:
                    target = ex
                    break
            if target is None:
                return Response({'detail': f"No {exam_group} exams to remove."}, status=400)

            ObeQpPatternConfig.objects.filter(
                class_type='SPECIAL',
                question_paper_type=qp_type_val,
                exam=target,
            ).delete()
            new_exams = _sort_special_exams(current_set - {target})

        # Propagate to curriculum rows
        ea_keys = [_EXAM_TO_EA_KEY[e] for e in new_exams if e in _EXAM_TO_EA_KEY]
        _propagate_special_enabled_assessments(ea_keys)

        return Response({
            'exams': new_exams,
            'added': target if action == 'add' else None,
            'removed': target if action == 'remove' else None,
            'enabled_assessments': ea_keys,
        })

    else:
        return Response({'detail': "action must be 'add', 'remove', or 'set'"}, status=400)
