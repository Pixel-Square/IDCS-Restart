from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, timedelta, date
import csv
import io

from .models import (
    AttendanceRecord,
    UploadLog,
    HalfDayRequest,
    Holiday,
    AttendanceSettings,
    DepartmentAttendanceSettings,
    SpecialDepartmentDateAttendanceLimit,
    StaffAttendanceTimeLimitOverride,
)
from .serializers import (
    AttendanceRecordSerializer,
    UploadLogSerializer,
    CSVUploadSerializer,
    HalfDayRequestSerializer,
    HalfDayRequestCreateSerializer,
    HalfDayRequestReviewSerializer,
    HolidaySerializer,
    HolidayCreateSerializer,
    AttendanceSettingsSerializer,
    DepartmentAttendanceSettingsSerializer,
    SpecialDepartmentDateAttendanceLimitSerializer,
    StaffAttendanceTimeLimitOverrideSerializer,
)
from .permissions import (
    StaffAttendanceUploadPermission,
    StaffAttendanceViewPermission,
    StaffAttendanceConfigPermission,
)


class AttendanceRecordViewSet(viewsets.ModelViewSet):
    """ViewSet for managing staff attendance records"""
    queryset = AttendanceRecord.objects.all()
    serializer_class = AttendanceRecordSerializer
    permission_classes = [IsAuthenticated, StaffAttendanceViewPermission]
    filterset_fields = ['user', 'date', 'status']
    search_fields = ['user__username', 'user__first_name', 'user__last_name']

    @action(detail=False, methods=['get'])
    def today_status(self, request):
        """Get current user's attendance status for today"""
        today = timezone.now().date()
        record = AttendanceRecord.objects.filter(
            user=request.user,
            date=today
        ).first()
        
        if record:
            return Response({
                'status': record.status,
                'morning_in': record.morning_in,
                'evening_out': record.evening_out,
                'fn_status': record.fn_status,
                'an_status': record.an_status,
            })
        return Response({'status': 'no_record', 'morning_in': None, 'evening_out': None})

    @action(detail=False, methods=['get'])
    def available_departments(self, request):
        """Get departments that have staff with attendance records"""
        from academics.models import Department
        depts = Department.objects.filter(
            staff__isnull=False
        ).distinct().values('id', 'name', 'code', 'short_name').order_by('code')
        return Response({'departments': list(depts)})

    @action(detail=False, methods=['get'])
    def monthly_records(self, request):
        """Get monthly attendance records with summary"""
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        department_id = request.query_params.get('department_id')

        # Build query
        if from_date and to_date:
            records = AttendanceRecord.objects.filter(
                date__gte=from_date,
                date__lte=to_date
            )
        else:
            records = AttendanceRecord.objects.filter(
                date__year=year,
                date__month=month
            )

        if department_id:
            records = records.filter(user__staff_profile__department_id=department_id)

        # Filter to current user or allowed view
        if not request.user.is_superuser and not request.user.has_perm('staff_attendance.view_attendance_records'):
            records = records.filter(user=request.user)

        records = records.order_by('-date', 'user')
        serializer = AttendanceRecordSerializer(records, many=True)

        # Calculate summary
        summary = {
            'year': year,
            'month': month,
            'total_records': records.count(),
            'present_count': records.filter(status='present').count(),
            'absent_count': records.filter(status='absent').count(),
            'partial_count': records.filter(Q(status='partial') | Q(status='half_day')).count(),
        }

        return Response({
            'records': serializer.data,
            'summary': summary
        })

    @action(detail=False, methods=['get'])
    def organization_analytics(self, request):
        """Get organization-wide attendance analytics"""
        report_type = request.query_params.get('report_type', '1')
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        month = request.query_params.get('month')
        department_id = request.query_params.get('department_id')
        export_format = request.query_params.get('export', 'json')

        # Build date range
        if report_type == '1':
            if from_date:
                start_date = datetime.fromisoformat(from_date).date()
                end_date = datetime.fromisoformat(to_date).date() if to_date else start_date
            else:
                return Response({'error': 'from_date required for report_type 1'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            if month:
                year, month_num = month.split('-')
                start_date = datetime(int(year), int(month_num), 1).date()
                if int(month_num) == 12:
                    end_date = datetime(int(year) + 1, 1, 1).date() - timedelta(days=1)
                else:
                    end_date = datetime(int(year), int(month_num) + 1, 1).date() - timedelta(days=1)
            else:
                return Response({'error': 'month required for report_type > 1'}, status=status.HTTP_400_BAD_REQUEST)

        if report_type != '1':
            from academics.models import StaffProfile
            
            # 1. Fetch holidays mapping
            holidays_qs = Holiday.objects.filter(date__gte=start_date, date__lte=end_date)
            global_holidays = set(h.date for h in holidays_qs if not h.departments.exists())
            dept_holidays = {}
            for h in holidays_qs:
                for d in h.departments.all():
                    dept_holidays.setdefault(d.id, set()).add(h.date)

            def _is_hol(d_date, d_id):
                return d_date in global_holidays or (d_id and d_id in dept_holidays and d_date in dept_holidays[d_id])

            # 2. Fetch staff profiles
            staff_qs = StaffProfile.objects.filter(status='ACTIVE', user__isnull=False).select_related('user', 'department')
            if department_id:
                staff_qs = staff_qs.filter(department_id=department_id)
            staff_qs = staff_qs.order_by('department__name', 'user__first_name', 'user__username')

            # 3. Group attendance records by user_id
            records = AttendanceRecord.objects.filter(date__gte=start_date, date__lte=end_date)
            if department_id:
                records = records.filter(user__staff_profile__department_id=department_id)
                
            rec_map = {}
            for r in records:
                rec_map.setdefault(r.user_id, {})[r.date] = r

            def _short_s(s_str):
                if not s_str:
                    return '-'
                sl = s_str.lower()
                if sl == 'present': return 'P'
                if sl == 'absent': return 'A'
                return s_str.upper()

            def _cell(r_obj, hol_flag, r_type):
                if not r_obj:
                    return {'value': '-' if hol_flag else 'A', 'is_holiday': hol_flag}
                
                st_val = r_obj.status or 'absent'
                leave_set = {'cl', 'od', 'ml', 'col', 'leave'}
                is_lv = st_val.lower() in leave_set or st_val.lower() not in {'present', 'absent', 'partial', 'half_day', ''}

                if r_type == '5':
                    if is_lv or st_val.lower() == 'present':
                        v = '0'
                    elif st_val.lower() in ('half_day', 'partial'):
                        v = '0.5'
                    else:
                        v = '1'
                    return {'value': '-' if hol_flag and v == '1' else v, 'is_holiday': hol_flag}

                if is_lv:
                    return {'value': st_val.upper(), 'is_holiday': hol_flag}

                if st_val.lower() == 'absent':
                    return {'value': '-' if hol_flag else 'A', 'is_holiday': hol_flag}

                fs = _short_s(r_obj.fn_status)
                ans = _short_s(r_obj.an_status)
                hdr = f"FN: {fs} | AN: {ans}"

                min_str = r_obj.morning_in.strftime('%H:%M') if r_obj.morning_in else '-'
                eout_str = r_obj.evening_out.strftime('%H:%M') if r_obj.evening_out else '-'

                eff = ""
                if r_obj.morning_in and r_obj.evening_out:
                    dt1 = datetime.combine(start_date, r_obj.morning_in)
                    dt2 = datetime.combine(start_date, r_obj.evening_out)
                    if dt2 >= dt1:
                        tmins = int((dt2 - dt1).total_seconds() / 60)
                        eff = f"{tmins // 60}h {tmins % 60}m"
                    else:
                        eff = "-"
                elif r_obj.morning_in:
                    eff = f"In: {min_str}"
                elif r_obj.evening_out:
                    eff = f"Out: {eout_str}"

                if r_type == '2':
                    v = f"{hdr}\n{eff}" if eff else hdr
                elif r_type == '3':
                    v = f"{hdr}\nIn: {min_str}\nOut: {eout_str}"
                else: # Type 4
                    if eff and r_obj.morning_in and r_obj.evening_out:
                        v = f"{hdr}\nIn: {min_str} | Out: {eout_str}\n{eff}"
                    else:
                        v = f"{hdr}\nIn: {min_str} | Out: {eout_str}"

                return {'value': v.strip(), 'is_holiday': hol_flag}

            working_days_count = (end_date - start_date).days + 1
            day_columns = [str(d) for d in range(1, working_days_count + 1)]
            columns = ['days']

            staff_rows = []
            for sp in staff_qs:
                u_id = sp.user_id
                d_id = sp.department_id if sp.department else None
                user_recs = rec_map.get(u_id, {})
                
                d_count = 0.0
                for r_obj in user_recs.values():
                    sv = (r_obj.status or '').lower()
                    if sv in ('present', 'cl', 'od', 'ml', 'col', 'leave') or sv not in ('absent', 'half_day', 'partial', ''):
                        d_count += 1.0
                    elif sv in ('half_day', 'partial'):
                        d_count += 0.5

                val_dict = {}
                for d_num in range(1, working_days_count + 1):
                    cur_d = date(start_date.year, start_date.month, d_num)
                    hol = _is_hol(cur_d, d_id)
                    r_obj = user_recs.get(cur_d)
                    val_dict[str(d_num)] = _cell(r_obj, hol, report_type)

                staff_rows.append({
                    'staff_user_id': u_id,
                    'staff_id': sp.staff_id,
                    'staff_name': sp.user.get_full_name() or sp.user.username,
                    'department': sp.department.name if sp.department else '',
                    'days': d_count,
                    'values': val_dict,
                })

            if export_format == 'excel':
                from django.http import HttpResponse
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = f"Matrix Type {report_type}"

                headers = ["Staff ID", "Staff Name", "Days"] + day_columns
                ws.append(headers)

                header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
                header_font = Font(name="Calibri", size=11, bold=True)
                center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
                hol_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

                for col_idx in range(1, len(headers) + 1):
                    c = ws.cell(row=1, column=col_idx)
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = center_align

                for r_idx, s_row in enumerate(staff_rows, start=2):
                    row_vals = [s_row['staff_id'], s_row['staff_name'], round(s_row['days'], 1)]
                    for d_str in day_columns:
                        row_vals.append(s_row['values'][d_str]['value'])
                    ws.append(row_vals)

                    for c_idx in range(1, len(row_vals) + 1):
                        c = ws.cell(row=r_idx, column=c_idx)
                        if c_idx in (1, 2):
                            c.alignment = left_align
                        else:
                            c.alignment = center_align
                            
                        if c_idx > 3:
                            d_str = day_columns[c_idx - 4]
                            if s_row['values'][d_str]['is_holiday']:
                                c.fill = hol_fill

                ws.row_dimensions[1].height = 25
                for r_idx in range(2, len(staff_rows) + 2):
                    ws.row_dimensions[r_idx].height = 45 if report_type in ('2', '3', '4') else 22

                for col in ws.columns:
                    col_letter = get_column_letter(col[0].column)
                    if col[0].column in (1, 2):
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                    elif col[0].column == 3:
                        ws.column_dimensions[col_letter].width = 8
                    else:
                        ws.column_dimensions[col_letter].width = 16 if report_type in ('2', '3', '4') else 6

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                resp = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                filename = f"organization_matrix_type_{report_type}_{month}.xlsx"
                resp['Content-Disposition'] = f'attachment; filename="{filename}"'
                return resp

            matrix_data = {
                'report_type': report_type,
                'month': month,
                'date_range': {
                    'from_date': str(start_date),
                    'to_date': str(end_date),
                    'working_days': working_days_count,
                },
                'columns': columns,
                'day_columns': day_columns,
                'total_staff': len(staff_rows),
                'staff_rows': staff_rows,
            }
            return Response(matrix_data)

        records = AttendanceRecord.objects.filter(date__gte=start_date, date__lte=end_date)
        if department_id:
            records = records.filter(user__staff_profile__department_id=department_id)

        # Calculate working days
        working_days = (end_date - start_date).days + 1

        # Build analytics
        staff_data = {}
        for record in records:
            user_id = record.user_id
            if user_id not in staff_data:
                staff = record.user
                profile = getattr(staff, 'staff_profile', None)
                dept = getattr(profile, 'department', None) if profile else None
                staff_data[user_id] = {
                    'staff_id': getattr(profile, 'staff_id', '') if profile else '',
                    'name': staff.get_full_name() or staff.username,
                    'email': staff.email,
                    'department': getattr(dept, 'name', '') if dept else '',
                    'present': 0,
                    'absent': 0,
                    'no_record': 0,
                    'cl_count': 0,
                    'od_count': 0,
                    'late_entry_count': 0,
                    'col_count': 0,
                    'others_count': 0,
                }
            
            status_val = record.status or 'absent'
            if status_val == 'present':
                staff_data[user_id]['present'] += 1
            elif status_val == 'absent':
                staff_data[user_id]['absent'] += 1
            elif status_val in ('CL', 'cl'):
                staff_data[user_id]['cl_count'] += 1
            elif status_val in ('OD', 'od'):
                staff_data[user_id]['od_count'] += 1
            elif status_val in ('COL', 'col'):
                staff_data[user_id]['col_count'] += 1
            else:
                staff_data[user_id]['others_count'] += 1

        if export_format == 'excel':
            from django.http import HttpResponse
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Summary"

            headers = [
                "Staff ID", "Staff Name", "Department", "Present", "Absent",
                "CL", "OD", "Late Entry", "COL", "Others", "Attendance %"
            ]
            ws.append(headers)

            header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            for idx, staff in enumerate(staff_data.values(), start=2):
                wd = working_days
                pct = round((staff['present'] / wd) * 100, 2) if wd > 0 else 0.0
                row_vals = [
                    staff['staff_id'],
                    staff['name'],
                    staff['department'],
                    staff['present'],
                    staff['absent'],
                    staff['cl_count'],
                    staff['od_count'],
                    staff['late_entry_count'],
                    staff['col_count'],
                    staff['others_count'],
                    f"{pct}%"
                ]
                ws.append(row_vals)
                
                for col_idx in range(1, len(row_vals) + 1):
                    c = ws.cell(row=idx, column=col_idx)
                    if col_idx in (1, 2, 3):
                        c.alignment = left_align
                    else:
                        c.alignment = center_align

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            resp = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            filename = f"organization_attendance_{start_date}.xlsx"
            resp['Content-Disposition'] = f'attachment; filename="{filename}"'
            return resp

        analytics = {
            'date_range': {
                'from_date': str(start_date),
                'to_date': str(end_date),
                'working_days': working_days,
            },
            'summary': {
                'total_staff': len(staff_data),
                'total_records': records.count(),
                'total_present': sum(d['present'] for d in staff_data.values()),
                'total_absent': sum(d['absent'] for d in staff_data.values()),
                'staff_present_count': sum(1 for d in staff_data.values() if d['present'] > 0),
                'staff_absent_count': sum(1 for d in staff_data.values() if d['absent'] > 0),
                'staff_cl_count': sum(d['cl_count'] for d in staff_data.values()),
                'staff_od_count': sum(d['od_count'] for d in staff_data.values()),
                'staff_col_count': sum(d['col_count'] for d in staff_data.values()),
                'staff_late_entry_count': 0,
                'staff_others_count': sum(d['others_count'] for d in staff_data.values()),
            },
            'staff_analytics': list(staff_data.values()),
        }

        return Response(analytics)

    # Alias for compatibility
    @action(detail=False, methods=['get'], url_path='organization_analytics')
    def organization_analytics_legacy(self, request):
        """Alias for organization-analytics endpoint"""
        return self.organization_analytics(request)


class UploadLogViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing staff attendance upload logs"""
    queryset = UploadLog.objects.all()
    serializer_class = UploadLogSerializer
    permission_classes = [IsAuthenticated, StaffAttendanceConfigPermission]
    ordering_fields = ['uploaded_at']
    ordering = ['-uploaded_at']


class CSVUploadViewSet(viewsets.ViewSet):
    """ViewSet for CSV file uploads and management"""
    permission_classes = [IsAuthenticated, StaffAttendanceUploadPermission]

    @action(detail=False, methods=['post'])
    def upload(self, request):
        """Upload and process attendance CSV file"""
        serializer = CSVUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        file_obj = serializer.validated_data['file']
        dry_run = serializer.validated_data.get('dry_run', False)
        overwrite = serializer.validated_data.get('overwrite_existing', False)

        try:
            # Parse CSV
            content = file_obj.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(content))
            
            success_count = 0
            error_count = 0
            errors = []
            
            for row_num, row in enumerate(csv_reader, 1):
                try:
                    # Process each row
                    staff_id = row.get('staff_id') or row.get('Staff ID')
                    date_str = row.get('date') or row.get('Date')
                    morning_in = row.get('morning_in') or row.get('Morning In')
                    evening_out = row.get('evening_out') or row.get('Evening Out')
                    
                    if not staff_id or not date_str:
                        error_count += 1
                        errors.append({
                            'row': row_num,
                            'user_id': staff_id,
                            'error': 'Missing staff_id or date'
                        })
                        continue

                    # Find user by staff_id
                    from academics.models import StaffProfile
                    profile = StaffProfile.objects.filter(staff_id=staff_id).first()
                    if not profile or not profile.user:
                        error_count += 1
                        errors.append({
                            'row': row_num,
                            'user_id': staff_id,
                            'error': 'Staff not found'
                        })
                        continue

                    if not dry_run:
                        record, created = AttendanceRecord.objects.get_or_create(
                            user=profile.user,
                            date=datetime.fromisoformat(date_str).date(),
                            defaults={
                                'morning_in': morning_in or None,
                                'evening_out': evening_out or None,
                                'uploaded_by': request.user,
                            }
                        )
                        if overwrite or created:
                            record.morning_in = morning_in or None
                            record.evening_out = evening_out or None
                            record.uploaded_by = request.user
                            record.save()
                        success_count += 1
                    else:
                        success_count += 1

                except Exception as e:
                    error_count += 1
                    errors.append({
                        'row': row_num,
                        'error': str(e)
                    })

            if not dry_run:
                upload_log = UploadLog.objects.create(
                    uploader=request.user,
                    filename=file_obj.name,
                    target_date=timezone.now().date(),
                    processed_rows=success_count + error_count,
                    success_count=success_count,
                    error_count=error_count,
                    errors=errors,
                    file=file_obj
                )

            return Response({
                'success': True,
                'processed_rows': success_count + error_count,
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:100],
                'upload_log_id': upload_log.id if not dry_run else None,
            })

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def bulk_delete_month(self, request):
        """Delete all attendance records for a month"""
        month = request.data.get('month')
        year = request.data.get('year')
        confirm = request.data.get('confirm', False)

        if not month or not year:
            return Response({'error': 'month and year required'}, status=status.HTTP_400_BAD_REQUEST)

        records = AttendanceRecord.objects.filter(date__year=year, date__month=month)
        count = records.count()

        if not confirm:
            return Response({
                'message': f'Ready to delete {count} records for {year}-{month:02d}',
                'count': count,
                'confirm': False
            })

        records.delete()
        return Response({
            'message': f'Deleted {count} records for {year}-{month:02d}',
            'count': count,
        })

    @action(detail=False, methods=['get'])
    def essl_settings(self, request):
        """Get eSSL device settings"""
        return Response({'devices': []})

    @action(detail=False, methods=['post'])
    def retrieve_essl_data(self, request):
        """Retrieve and process eSSL data"""
        return Response({'success': True, 'message': 'eSSL data retrieval not yet implemented'})


class HalfDayRequestViewSet(viewsets.ModelViewSet):
    """ViewSet for period attendance access requests"""
    queryset = HalfDayRequest.objects.all()
    serializer_class = HalfDayRequestSerializer
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return HalfDayRequestCreateSerializer
        if self.action in ['review_request']:
            return HalfDayRequestReviewSerializer
        return HalfDayRequestSerializer

    @action(detail=False, methods=['get'])
    def check_period_attendance_access(self, request):
        """Check if user can mark period attendance for a date"""
        date_str = request.query_params.get('date')
        if not date_str:
            return Response({'error': 'date parameter required'}, status=status.HTTP_400_BAD_REQUEST)

        request_obj = HalfDayRequest.objects.filter(
            staff_user=request.user,
            attendance_date=date_str
        ).first()

        if request_obj:
            return Response({
                'can_mark_attendance': request_obj.status == 'approved',
                'reason': f'Request status: {request_obj.status}',
                'attendance_record': None,
            })

        return Response({
            'can_mark_attendance': False,
            'reason': 'No approved period attendance request for this date',
            'attendance_record': None,
        })

    @action(detail=False, methods=['get'])
    def pending_for_review(self, request):
        """Get pending period attendance requests for HOD/AHOD"""
        # Check if user is HOD or AHOD
        if not (request.user.is_superuser or request.user.has_perm('staff_attendance.manage_attendance')):
            return Response([], status=status.HTTP_403_FORBIDDEN)

        requests = HalfDayRequest.objects.filter(status='pending').order_by('-requested_at')
        serializer = HalfDayRequestSerializer(requests, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def review_request(self, request, pk=None):
        """Review a period attendance access request"""
        half_day_req = self.get_object()
        
        # Check if user is HOD or AHOD
        if not (request.user.is_superuser or request.user.has_perm('staff_attendance.manage_attendance')):
            return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        action_val = request.data.get('action')
        review_notes = request.data.get('review_notes', '')

        if action_val not in ['approve', 'reject']:
            return Response({'error': 'action must be approve or reject'}, status=status.HTTP_400_BAD_REQUEST)

        half_day_req.status = 'approved' if action_val == 'approve' else 'rejected'
        half_day_req.reviewed_by = request.user
        half_day_req.reviewed_at = timezone.now()
        half_day_req.review_notes = review_notes
        half_day_req.save()

        serializer = HalfDayRequestSerializer(half_day_req)
        return Response(serializer.data)


class HolidayViewSet(viewsets.ModelViewSet):
    """ViewSet for managing holidays"""
    queryset = Holiday.objects.all()
    serializer_class = HolidaySerializer
    permission_classes = [IsAuthenticated, StaffAttendanceConfigPermission]
    ordering = ['-date']

    def get_permissions(self):
        """
        Override permissions per action:
        - Safe read-only actions (list, retrieve, my_holidays, departments):
          any authenticated user can access so staff calendars show holidays.
        - All write/admin actions: require StaffAttendanceConfigPermission.
        """
        read_only_actions = {'list', 'retrieve', 'my_holidays', 'departments'}
        if self.action in read_only_actions:
            return [IsAuthenticated()]
        return [IsAuthenticated(), StaffAttendanceConfigPermission()]

    def get_serializer_class(self):
        if self.action == 'create':
            return HolidayCreateSerializer
        return HolidaySerializer


    @action(detail=False, methods=['get'])
    def departments(self, request):
        """Get list of all departments (teaching and non-teaching) for holiday assignment"""
        from academics.models import Department
        depts = Department.objects.all().values('id', 'name', 'code', 'short_name', 'is_teaching').order_by('is_teaching', 'code')
        return Response(list(depts))

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def my_holidays(self, request):
        """Get holidays that apply to the requesting user's department.

        Rules:
          - A holiday with no departments assigned applies to ALL staff.
          - A holiday with specific departments assigned only applies to staff
            whose department is in that list.
        """
        # Determine the requesting user's department
        user_department = None
        try:
            profile = getattr(request.user, 'staff_profile', None)
            if profile:
                if hasattr(profile, 'get_current_department'):
                    user_department = profile.get_current_department()
                if not user_department:
                    user_department = getattr(profile, 'department', None)
        except Exception:
            user_department = None

        if user_department:
            # Return holidays that are either global (no dept assigned, i.e. empty M2M)
            # or include this specific department.
            # For M2M, departments__isnull=True matches rows with no related dept records.
            qs = Holiday.objects.filter(
                Q(departments__isnull=True) | Q(departments=user_department)
            ).distinct().order_by('-date')
        else:
            # No department info — only return global (all-department) holidays
            qs = Holiday.objects.filter(departments__isnull=True).distinct().order_by('-date')

        serializer = HolidaySerializer(qs, many=True)
        return Response(serializer.data)


    @action(detail=False, methods=['post'])
    def generate_sundays(self, request):
        """Generate Sunday holidays for a month"""
        year = request.data.get('year')
        month = request.data.get('month')

        if not year or not month:
            return Response({'error': 'year and month required'}, status=status.HTTP_400_BAD_REQUEST)

        created = 0
        already_exists = 0

        from calendar import monthrange
        days_in_month = monthrange(year, month)[1]
        
        for day in range(1, days_in_month + 1):
            date_obj = datetime(year, month, day).date()
            if date_obj.weekday() == 6:  # Sunday
                holiday, created_flag = Holiday.objects.get_or_create(
                    date=date_obj,
                    defaults={
                        'name': f'Sunday {date_obj.strftime("%B %d")}',
                        'is_sunday': True,
                        'is_removable': True,
                        'created_by': request.user,
                    }
                )
                if created_flag:
                    created += 1
                else:
                    already_exists += 1

        return Response({
            'created': created,
            'already_exists': already_exists,
            'message': f'Generated {created} Sunday holidays for {year}-{month:02d}'
        })

    @action(detail=False, methods=['post'])
    def remove_sundays(self, request):
        """Remove Sunday holidays for a month"""
        year = request.data.get('year')
        month = request.data.get('month')

        if not year or not month:
            return Response({'error': 'year and month required'}, status=status.HTTP_400_BAD_REQUEST)

        deleted, _ = Holiday.objects.filter(
            date__year=year,
            date__month=month,
            is_sunday=True,
            is_removable=True
        ).delete()

        return Response({
            'deleted_count': deleted,
            'message': f'Removed {deleted} Sunday holidays for {year}-{month:02d}'
        })


class AttendanceSettingsViewSet(viewsets.ModelViewSet):
    """ViewSet for attendance settings"""
    queryset = AttendanceSettings.objects.all()
    serializer_class = AttendanceSettingsSerializer
    permission_classes = [IsAuthenticated, StaffAttendanceConfigPermission]

    @action(detail=False, methods=['get'])
    def current(self, request):
        """Get current global attendance settings"""
        settings = AttendanceSettings.objects.first()
        if not settings:
            settings = AttendanceSettings.objects.create(
                updated_by=request.user
            )
        serializer = AttendanceSettingsSerializer(settings)
        return Response(serializer.data)


class DepartmentAttendanceSettingsViewSet(viewsets.ModelViewSet):
    """ViewSet for department-specific attendance settings"""
    queryset = DepartmentAttendanceSettings.objects.all()
    serializer_class = DepartmentAttendanceSettingsSerializer
    permission_classes = [IsAuthenticated, StaffAttendanceConfigPermission]
    ordering = ['name']


class SpecialDepartmentDateAttendanceLimitViewSet(viewsets.ModelViewSet):
    """ViewSet for special department date-range attendance limits"""
    queryset = SpecialDepartmentDateAttendanceLimit.objects.all()
    serializer_class = SpecialDepartmentDateAttendanceLimitSerializer
    permission_classes = [IsAuthenticated, StaffAttendanceConfigPermission]
    ordering = ['-from_date', '-id']

    @action(detail=True, methods=['post'])
    def reapply(self, request, pk=None):
        """Reapply special attendance limits to affected records"""
        special_limit = self.get_object()
        # Mark records in date range to be recalculated
        records = AttendanceRecord.objects.filter(
            user__staff_profile__department__in=special_limit.departments.all(),
            date__gte=special_limit.from_date,
            date__lte=special_limit.to_date or special_limit.from_date
        )
        updated = 0
        for record in records:
            record.update_status()
            record.save()
            updated += 1

        return Response({
            'message': f'Reapplied special limits to {updated} records',
            'updated': updated
        })


class StaffAttendanceTimeLimitOverrideViewSet(viewsets.ModelViewSet):
    """ViewSet for staff-specific attendance time limit overrides"""
    queryset = StaffAttendanceTimeLimitOverride.objects.all()
    serializer_class = StaffAttendanceTimeLimitOverrideSerializer
    permission_classes = [IsAuthenticated, StaffAttendanceConfigPermission]
    ordering = ['-updated_at', '-id']

    @action(detail=False, methods=['get'])
    def staff_options(self, request):
        """Get list of staff for override assignment"""
        from academics.models import StaffProfile
        
        department_id = request.query_params.get('department_id')
        search_q = request.query_params.get('q', '').strip()

        staff_qs = StaffProfile.objects.select_related('user', 'department')
        
        if department_id:
            staff_qs = staff_qs.filter(department_id=department_id)
        
        if search_q:
            staff_qs = staff_qs.filter(
                Q(user__username__icontains=search_q) |
                Q(user__first_name__icontains=search_q) |
                Q(user__last_name__icontains=search_q) |
                Q(staff_id__icontains=search_q)
            )

        staff_list = []
        for profile in staff_qs[:100]:
            full_name = profile.user.get_full_name() or profile.user.username
            dept = profile.department
            staff_list.append({
                'user_id': profile.user_id,
                'username': profile.user.username,
                'full_name': full_name,
                'staff_id': profile.staff_id,
                'department': {
                    'id': dept.id,
                    'code': dept.code,
                    'short_name': dept.short_name,
                    'name': dept.name,
                } if dept else None,
            })

        return Response(staff_list)

    @action(detail=False, methods=['post'])
    def upsert(self, request):
        """Create or update a staff override"""
        user_id = request.data.get('user')
        if not user_id:
            return Response({'error': 'user required'}, status=status.HTTP_400_BAD_REQUEST)

        override, created = StaffAttendanceTimeLimitOverride.objects.get_or_create(
            user_id=user_id,
            defaults={'created_by': request.user}
        )

        # Update fields
        override.attendance_in_time_limit = request.data.get('attendance_in_time_limit', override.attendance_in_time_limit)
        override.attendance_out_time_limit = request.data.get('attendance_out_time_limit', override.attendance_out_time_limit)
        override.mid_time_split = request.data.get('mid_time_split', override.mid_time_split)
        override.lunch_from = request.data.get('lunch_from')
        override.lunch_to = request.data.get('lunch_to')
        override.apply_time_based_absence = request.data.get('apply_time_based_absence', override.apply_time_based_absence)
        override.enabled = request.data.get('enabled', override.enabled)
        override.updated_by = request.user
        override.save()

        serializer = StaffAttendanceTimeLimitOverrideSerializer(override)
        return Response(serializer.data)
