import logging
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated

logger = logging.getLogger(__name__)
from django.db import transaction
from django.db.models import Q, Case, When, Value, IntegerField, Count
from django.shortcuts import get_object_or_404
from django.utils import timezone

from .models import FeedbackForm, FeedbackQuestion, FeedbackQuestionOption, FeedbackResponse, FeedbackFormSubmission
from .serializers import (
    FeedbackFormCreateSerializer,
    FeedbackFormSerializer,
    FeedbackSubmissionSerializer,
    get_subject_feedback_completion,
)
from accounts.utils import get_user_permissions
from academics.models import StaffProfile


def get_normalized_permissions(user):
    if not user or not getattr(user, 'is_authenticated', False):
        return set()
    try:
        return {str(p).lower() for p in get_user_permissions(user)}
    except Exception:
        return set()


def user_has_permission(user, permission_code: str) -> bool:
    return permission_code.lower() in get_normalized_permissions(user)


def user_has_any_permission(user, permission_codes) -> bool:
    user_perms = get_normalized_permissions(user)
    return any(str(code).lower() in user_perms for code in permission_codes)


def get_feedback_department_scope(user):
    """Resolve department scope for feedback create users.

    Returns dict with keys:
    - allowed: bool
    - all_departments: bool
    - department_ids: list[int]
    """
    permissions = get_normalized_permissions(user)
    can_create = 'feedback.create' in permissions
    can_all = 'feedback.all_departments_access' in permissions
    can_own = 'feedback.own_department_access' in permissions

    if not can_create or not (can_all or can_own):
        return {
            'allowed': False,
            'all_departments': False,
            'department_ids': [],
        }

    if can_all:
        return {
            'allowed': True,
            'all_departments': True,
            'department_ids': [],
        }

    department_ids = []
    try:
        from academics.models import AcademicYear, DepartmentRole

        staff_profile = StaffProfile.objects.get(user=user)
        active_ay = AcademicYear.objects.filter(is_active=True).first()

        if active_ay:
            department_ids = list(
                DepartmentRole.objects.filter(
                    staff=staff_profile,
                    role='HOD',
                    is_active=True,
                    academic_year=active_ay,
                ).values_list('department_id', flat=True)
            )

        if not department_ids and getattr(staff_profile, 'department_id', None):
            department_ids = [staff_profile.department_id]
    except Exception:
        department_ids = []

    return {
        'allowed': True,
        'all_departments': False,
        'department_ids': [int(d) for d in department_ids if d],
    }


def apply_department_scope_filter(queryset, scope, field_name='department_id'):
    if scope.get('all_departments'):
        return queryset

    department_ids = scope.get('department_ids') or []
    if not department_ids:
        return queryset.none()

    return queryset.filter(**{f'{field_name}__in': department_ids})


def calculate_feedback_response_metrics(feedback_form):
    """Return response_count, expected_count, and percentage for a feedback form."""
    from academics.models import StudentProfile, Section, AcademicYear, DepartmentRole

    response_count = FeedbackResponse.objects.filter(
        feedback_form=feedback_form
    ).values('user_id').distinct().count()

    expected_count = 0

    if feedback_form.target_type in {'STAFF', 'HOD'}:
        if feedback_form.target_type == 'HOD':
            active_ay = AcademicYear.objects.filter(is_active=True).first()
            if active_ay:
                expected_count = DepartmentRole.objects.filter(
                    role='HOD',
                    is_active=True,
                    academic_year=active_ay,
                    department_id=feedback_form.department_id,
                ).values('staff_id').distinct().count()
        else:
            expected_count = StaffProfile.objects.filter(
                department_id=feedback_form.department_id
            ).count()
    elif feedback_form.target_type == 'STUDENT':
        if feedback_form.all_classes:
            expected_count = StudentProfile.objects.filter(
                section__batch__course__department_id=feedback_form.department_id
            ).count()
        else:
            sections_to_query = []

            current_ay = AcademicYear.objects.filter(is_active=True).first()
            current_acad_year = None
            if current_ay:
                try:
                    current_acad_year = int(str(current_ay.name).split('-')[0])
                except Exception:
                    current_acad_year = None

            if feedback_form.sections:
                sections_to_query = list(feedback_form.sections)
            elif feedback_form.section_id:
                sections_to_query = [feedback_form.section_id]
            else:
                sections_filter = Q(batch__course__department_id=feedback_form.department_id)

                if feedback_form.years:
                    year_filters = Q()
                    for year in feedback_form.years:
                        if current_acad_year:
                            batch_start_year = current_acad_year - year + 1
                            year_filters |= Q(batch__start_year=str(batch_start_year))
                    sections_filter &= year_filters
                elif feedback_form.year:
                    if current_acad_year:
                        batch_start_year = current_acad_year - feedback_form.year + 1
                        sections_filter &= Q(batch__start_year=str(batch_start_year))

                if feedback_form.semesters:
                    sections_filter &= Q(semester_id__in=feedback_form.semesters)
                elif feedback_form.semester_id:
                    sections_filter &= Q(semester_id=feedback_form.semester_id)

                sections_to_query = list(
                    Section.objects.filter(sections_filter).values_list('id', flat=True)
                )

            if sections_to_query:
                expected_count = StudentProfile.objects.filter(
                    section_id__in=sections_to_query
                ).count()

    percentage = round((response_count / expected_count * 100) if expected_count > 0 else 0, 1)
    return response_count, expected_count, percentage


def has_principal_scope_permissions(user):
    perms = get_normalized_permissions(user)
    required = {
        'feedback.principal_feedback_page',
        'feedback.principal_all_departments_access',
    }
    return required.issubset(perms)


class CreateFeedbackFormView(APIView):
    """
    API 1: Create Feedback Form (HOD)
    POST /api/feedback/create
    
    HODs can create feedback forms with questions for their department.
    Uses active department from session for HODs with multiple departments.
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            # Check if user has permission to create feedback forms
            user_permissions = get_normalized_permissions(request.user)
            if 'feedback.create' not in user_permissions:
                return Response({
                    'detail': 'You do not have permission to create feedback forms.'
                }, status=status.HTTP_403_FORBIDDEN)

            can_all_departments_access = 'feedback.all_departments_access' in user_permissions
            can_own_department_access = 'feedback.own_department_access' in user_permissions

            if not (can_all_departments_access or can_own_department_access):
                return Response({
                    'detail': 'You do not have department access permission.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            from academics.models import AcademicYear, Semester, DepartmentRole, Department
            
            active_ay = AcademicYear.objects.filter(is_active=True).first()
            if not active_ay:
                return Response({
                    'detail': 'No active academic year found.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            selected_department_ids = []
            departments_payload = request.data.get('departments', []) or request.data.get('department_ids', [])
            all_departments_flag = request.data.get('all_departments', False)
            
            if can_all_departments_access:
                # Users with all department permission can choose any department.
                if all_departments_flag:
                    selected_department_ids = list(
                        Department.objects.values_list('id', flat=True)
                    )
                elif departments_payload and len(departments_payload) > 0:
                    # Validate all departments exist
                    for dept_id in departments_payload:
                        dept = Department.objects.filter(id=int(dept_id)).first()
                        if not dept:
                            return Response({
                                'detail': f'Department ID {dept_id} not found.'
                            }, status=status.HTTP_400_BAD_REQUEST)
                    selected_department_ids = [int(d) for d in departments_payload]
                else:
                    return Response({
                        'detail': 'Please select at least one department or select all departments.'
                    }, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Users with own department permission are restricted to assigned departments.
                try:
                    staff_profile = StaffProfile.objects.get(user=request.user)
                except StaffProfile.DoesNotExist:
                    return Response({
                        'detail': 'Staff profile not found.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Get HOD department roles
                department_roles = DepartmentRole.objects.select_related('department').filter(
                    staff=staff_profile,
                    role='HOD',
                    is_active=True,
                    academic_year=active_ay
                )
                
                departments_count = department_roles.count()
                
                if departments_count > 1:
                    # Multiple departments available to HOD - verify departments payload if provided
                    if departments_payload and len(departments_payload) > 0:
                        # HOD selected specific departments
                        dept_ids = [dr.department.id for dr in department_roles]
                        
                        # Verify all selected departments belong to this HOD
                        for dept_id in departments_payload:
                            if int(dept_id) not in dept_ids:
                                return Response({
                                    'detail': f'You do not have HOD access to department ID {dept_id}.'
                                }, status=status.HTTP_403_FORBIDDEN)
                        
                        selected_department_ids = [int(d) for d in departments_payload]
                    else:
                        # No departments selected, require at least one
                        return Response({
                            'detail': 'Please select at least one department.'
                        }, status=status.HTTP_400_BAD_REQUEST)
                elif departments_count == 1:
                    # Single department - use it automatically
                    selected_department_ids = [department_roles.first().department.id]
                else:
                    # No department roles - fall back to staff profile department
                    if staff_profile.department:
                        selected_department_ids = [staff_profile.department.id]
                    else:
                        return Response({
                            'detail': 'No department assigned to your profile.'
                        }, status=status.HTTP_400_BAD_REQUEST)
            
            # Automatically determine current semester(s) based on academic year parity
            years = request.data.get('years', [])
            semesters_to_use = []
            
            if active_ay and active_ay.parity and years:
                # Parity: ODD = semesters 1,3,5,7 | EVEN = semesters 2,4,6,8
                is_odd_semester = active_ay.parity.upper() == 'ODD'
                
                for year in years:
                    # Year 1 → Sem 1 or 2, Year 2 → Sem 3 or 4, Year 3 → Sem 5 or 6, Year 4 → Sem 7 or 8
                    base_semester = (year - 1) * 2 + 1  # Odd semester for this year
                    current_semester_num = base_semester if is_odd_semester else base_semester + 1
                    
                    # Get the Semester model instance with this number
                    try:
                        semester = Semester.objects.get(number=current_semester_num)
                        semesters_to_use.append(semester.id)
                    except Semester.DoesNotExist:
                        pass
            
            # Create feedback forms for each selected department
            created_forms = []
            errors = []
            
            for dept_id in selected_department_ids:
                # Update request data with automatically determined semesters and current department
                mutable_data = request.data.copy()
                mutable_data['semesters'] = semesters_to_use
                mutable_data['department'] = dept_id
                
                serializer = FeedbackFormCreateSerializer(data=mutable_data)
                if serializer.is_valid():
                    # Set the created_by field to current user
                    feedback_form = serializer.save(created_by=request.user)
                    created_forms.append(feedback_form)
                else:
                    errors.append({
                        'department_id': dept_id,
                        'errors': serializer.errors
                    })
            
            # Check results
            if len(created_forms) == 0:
                # All failed
                return Response({
                    'detail': 'Failed to create feedback forms.',
                    'errors': errors
                }, status=status.HTTP_400_BAD_REQUEST)
            elif len(errors) > 0:
                # Partial success
                return Response({
                    'detail': f'Created {len(created_forms)} feedback form(s), but {len(errors)} failed.',
                    'created_count': len(created_forms),
                    'created_forms': [FeedbackFormSerializer(form).data for form in created_forms],
                    'errors': errors
                }, status=status.HTTP_207_MULTI_STATUS)
            else:
                # Full success
                return Response({
                    'detail': f'Successfully created {len(created_forms)} feedback form(s).',
                    'created_count': len(created_forms),
                    'created_forms': [FeedbackFormSerializer(form).data for form in created_forms]
                }, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            # Catch any unexpected errors and return JSON
            return Response({
                'detail': f'An error occurred while creating the feedback form: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetFeedbackFormsView(APIView):
    """
    API 2: Get Forms
    GET /api/feedback/forms
    
    Returns feedback forms based on user's role:
    - HOD: Forms they created in their department
    - STAFF/STUDENT: Active forms targeted to them in their department
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Check if user has permission to view feedback page
        user_permissions = get_user_permissions(request.user)
        if 'feedback.feedback_page' not in user_permissions:
            return Response({
                'detail': 'You do not have permission to view feedback forms.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        user = request.user
        
        # Check if user is HOD (has CREATE permission)
        is_hod = 'feedback.create' in user_permissions
        
        if is_hod:
            # HOD: Return forms they created, sorted by status priority
            # Priority: Active & active=True (0) -> Draft (1) -> Deactivated (2) -> Closed (3)
            forms = FeedbackForm.objects.filter(
                created_by=user
            ).annotate(
                status_priority=Case(
                    # Active forms that are actually active (not deactivated)
                    When(status='ACTIVE', active=True, then=Value(0)),
                    # Draft forms
                    When(status='DRAFT', then=Value(1)),
                    # Deactivated forms (active=False)
                    When(status='ACTIVE', active=False, then=Value(2)),
                    # Closed forms
                    When(status='CLOSED', then=Value(3)),
                    default=Value(4),
                    output_field=IntegerField()
                )
            ).order_by('status_priority', '-created_at')
        else:
            # STAFF/STUDENT: Return active forms for their role and department
            try:
                # Try to get staff profile first
                staff_profile = StaffProfile.objects.get(user=user)
                department_id = staff_profile.department_id
                target_type = 'STAFF'
                
                # For staff: filter by target_type and department (only active forms)
                forms = FeedbackForm.objects.filter(
                    department_id=department_id,
                    target_type=target_type,
                    status='ACTIVE',
                    active=True  # Only show active forms to staff
                ).order_by('-created_at')
                
            except StaffProfile.DoesNotExist:
                # Try student profile
                try:
                    from academics.models import StudentProfile
                    student_profile = StudentProfile.objects.get(user=user)
                    section = student_profile.section
                    
                    # Get student's year, semester, section
                    batch = section.batch
                    student_year = None
                    student_semester = section.semester
                    student_section = section
                    
                    # Calculate year from batch
                    if batch.start_year:
                        from academics.models import AcademicYear
                        current_ay = AcademicYear.objects.filter(is_active=True).first()
                        if current_ay:
                            try:
                                acad_start = int(str(current_ay.name).split('-')[0])
                                delta = acad_start - int(batch.start_year)
                                student_year = delta + 1
                            except:
                                pass
                    
                    # Get department from section -> batch -> course -> department
                    department_id = batch.course.department_id if batch.course else batch.department_id
                    
                    # For students: filter by target_type, department, and class info
                    # Include forms where:
                    # - all_classes=True, OR
                    # - year matches (either in years list or legacy year field) AND
                    #   (semester is null/empty OR semester matches) AND
                    #   (section is null/empty OR section matches)
                    from django.db.models import Q
                    
                    # Build query filters
                    base_filter = Q(
                        department_id=department_id,
                        target_type='STUDENT',
                        status='ACTIVE',
                        active=True  # Only show active forms to students
                    )
                    
                    # All classes filter
                    all_classes_filter = Q(all_classes=True)
                    
                    # Specific class filter - check both multi-class and legacy fields
                    class_filter = Q()
                    
                    # Year matching: check if student's year is in the years list OR matches legacy year field
                    if student_year:
                        year_filter = (
                            Q(years__contains=[student_year]) |  # Multi-class: year in list
                            Q(year=student_year) |  # Legacy: single year
                            Q(years=[]) & Q(year__isnull=True)  # Empty/null = all years
                        )
                        class_filter &= year_filter
                    
                    # Semester matching: check if student's semester is in semesters list OR matches legacy semester
                    if student_semester:
                        semester_filter = (
                            Q(semesters__contains=[student_semester.id]) |  # Multi-class
                            Q(semester=student_semester) |  # Legacy
                            Q(semesters=[]) & Q(semester__isnull=True)  # Empty/null = all semesters
                        )
                        class_filter &= semester_filter
                    
                    # Section matching: check if student's section is in sections list OR matches legacy section
                    if student_section:
                        section_filter = (
                            Q(sections__contains=[student_section.id]) |  # Multi-class
                            Q(section=student_section) |  # Legacy
                            Q(sections=[]) & Q(section__isnull=True)  # Empty/null = all sections
                        )
                        class_filter &= section_filter
                    
                    # Combine filters
                    filters = base_filter & (all_classes_filter | class_filter)
                    
                    forms = FeedbackForm.objects.filter(filters).order_by('-created_at')
                    
                except Exception as e:
                    return Response({
                        'detail': 'Profile not found or invalid.'
                    }, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = FeedbackFormSerializer(forms, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)


class SubmitFeedbackView(APIView):
    """
    API 3: Submit Feedback
    POST /api/feedback/submit
    
    Staff and students submit answers to feedback forms.
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        # Check if user has permission to reply to feedback
        user_permissions = get_user_permissions(request.user)
        if 'feedback.reply' not in user_permissions:
            return Response({
                'detail': 'You do not have permission to submit feedback.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = FeedbackSubmissionSerializer(data=request.data)
        if serializer.is_valid():
            feedback_form_id = serializer.validated_data['feedback_form_id']
            teaching_assignment_id = serializer.validated_data.get('teaching_assignment_id')
            
            # Get the feedback form to check type
            feedback_form = get_object_or_404(FeedbackForm, id=feedback_form_id)
            
            # Check for duplicate submission
            if feedback_form.type == 'SUBJECT_FEEDBACK':
                # For subject feedback: check if user has already submitted for this subject
                if not teaching_assignment_id:
                    return Response({
                        'detail': 'Teaching assignment ID is required for subject feedback.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Handle pseudo-assignments (negative IDs for electives without teaching assignments)
                if teaching_assignment_id < 0:
                    return Response({
                        'detail': 'Subject mapping not found. Please contact your HOD to map this subject before submitting feedback.'
                    }, status=status.HTTP_400_BAD_REQUEST)

                existing_response = FeedbackResponse.objects.filter(
                    feedback_form_id=feedback_form_id,
                    user=request.user,
                    teaching_assignment_id=teaching_assignment_id
                ).exists()

                if existing_response:
                    return Response({
                        'detail': 'Feedback already submitted'
                    }, status=status.HTTP_400_BAD_REQUEST)
            else:
                # For open feedback: check if user has submitted for this form
                existing_response = FeedbackResponse.objects.filter(
                    feedback_form_id=feedback_form_id,
                    user=request.user,
                    teaching_assignment__isnull=True
                ).exists()
                
                if existing_response:
                    return Response({
                        'detail': 'You have already submitted feedback for this form.'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Save responses
            try:
                feedback_form = serializer.save(user=request.user)

                if feedback_form.type == 'SUBJECT_FEEDBACK':
                    completion = get_subject_feedback_completion(feedback_form, request.user)
                    total_subjects = completion['total_subjects']
                    responded_subjects = completion['responded_subjects']
                    all_completed = completion['all_completed']

                    status_value = 'SUBMITTED' if all_completed else 'PENDING'
                    submitted_at = timezone.now() if all_completed else None

                    FeedbackFormSubmission.objects.update_or_create(
                        feedback_form=feedback_form,
                        user=request.user,
                        defaults={
                            'submission_status': status_value,
                            'total_subjects': total_subjects,
                            'responded_subjects': responded_subjects,
                            'submitted_at': submitted_at,
                        }
                    )

                    if all_completed:
                        return Response({
                            'message': 'Feedback Submitted Successfully',
                            'submission_status': 'SUBMITTED',
                            'total_subjects': total_subjects,
                            'responded_subjects': responded_subjects,
                        }, status=status.HTTP_200_OK)

                    return Response({
                        'message': 'Pending – Complete all subjects',
                        'submission_status': 'PENDING',
                        'total_subjects': total_subjects,
                        'responded_subjects': responded_subjects,
                    }, status=status.HTTP_200_OK)
                
                return Response({
                    'message': 'Feedback submitted successfully'
                }, status=status.HTTP_200_OK)
            except Exception as e:
                logger.exception('[FEEDBACK SUBMIT] Error saving responses')
                return Response({
                    'detail': 'Error saving feedback. Please try again.'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        
        # Format validation errors for better frontend display
        error_messages = []
        for field, errors in serializer.errors.items():
            if isinstance(errors, list):
                for error in errors:
                    if isinstance(error, str):
                        error_messages.append(error)
                    elif isinstance(error, dict):
                        error_messages.extend(error.values())
            else:
                error_messages.append(str(errors))
        
        return Response({
            'detail': error_messages[0] if error_messages else 'Invalid feedback data',
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


class GetUserDepartmentView(APIView):
    """
    API 4: Get User's Department(s) for HOD
    GET /api/feedback/department/
    
    Returns the department(s) for HOD users based on DepartmentRole.
    - If HOD has one department: returns single department (no switch needed)
    - If HOD has multiple departments: returns all departments (enable switch)
    - Supports active_department_id parameter to set active department in session
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user

        user_permissions = get_normalized_permissions(user)
        can_create_feedback = 'feedback.create' in user_permissions
        can_all_departments_access = 'feedback.all_departments_access' in user_permissions
        can_own_department_access = 'feedback.own_department_access' in user_permissions

        if not can_create_feedback:
            return Response({
                'detail': 'You do not have permission to create feedback forms.'
            }, status=status.HTTP_403_FORBIDDEN)

        if not (can_all_departments_access or can_own_department_access):
            return Response({
                'success': False,
                'department': None,
                'detail': 'Department access permission is required.'
            }, status=status.HTTP_200_OK)

        # Users with all department access can view/switch all departments.
        if can_all_departments_access:
            try:
                from academics.models import Department
                
                # Get all departments (no is_active filter - Department model doesn't have that field)
                departments = Department.objects.all().order_by('name')
                departments_list = [{
                    'id': dept.id,
                    'name': dept.name,
                    'code': dept.code
                } for dept in departments]
                
                if not departments_list:
                    return Response({
                        'success': False,
                        'detail': 'No departments found.'
                    }, status=status.HTTP_200_OK)
                
                # Handle active department selection
                session_key = 'active_feedback_department_id'
                active_department_id = request.GET.get('active_department_id')
                if active_department_id:
                    try:
                        active_department_id = int(active_department_id)
                        active_dept = next((d for d in departments_list if d['id'] == active_department_id), None)
                        if active_dept:
                            request.session[session_key] = active_department_id
                            active_department = active_dept
                        else:
                            active_department = departments_list[0]
                            request.session[session_key] = departments_list[0]['id']
                    except ValueError:
                        active_department = departments_list[0]
                        request.session[session_key] = departments_list[0]['id']
                else:
                    # Check session for previously selected department
                    session_dept_id = request.session.get(session_key)
                    if session_dept_id:
                        active_dept = next((d for d in departments_list if d['id'] == session_dept_id), None)
                        if active_dept:
                            active_department = active_dept
                        else:
                            active_department = departments_list[0]
                            request.session[session_key] = departments_list[0]['id']
                    else:
                        active_department = departments_list[0]
                        request.session[session_key] = departments_list[0]['id']
                
                response_data = {
                    'success': True,
                    'has_multiple_departments': len(departments_list) > 1,
                    'departments': departments_list,
                    'active_department': active_department,
                }
                response_data['is_all_departments_access'] = True
                return Response(response_data, status=status.HTTP_200_OK)
            except Exception as e:
                import traceback
                traceback.print_exc()
                return Response({
                    'detail': f'Error retrieving departments: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        if not can_own_department_access:
            # For non-HOD users, return single department from staff profile
            try:
                staff_profile = StaffProfile.objects.get(user=user)
                if staff_profile.department:
                    return Response({
                        'success': True,
                        'has_multiple_departments': False,
                        'departments': [{
                            'id': staff_profile.department.id,
                            'name': staff_profile.department.name,
                            'code': staff_profile.department.code
                        }],
                        'active_department': {
                            'id': staff_profile.department.id,
                            'name': staff_profile.department.name,
                            'code': staff_profile.department.code
                        }
                    }, status=status.HTTP_200_OK)
            except StaffProfile.DoesNotExist:
                pass
            
            return Response({
                'success': False,
                'department': None
            }, status=status.HTTP_200_OK)
        
        # For HOD: Get departments from DepartmentRole
        try:
            from academics.models import DepartmentRole, AcademicYear, StaffProfile
            
            # Get staff profile
            staff_profile = StaffProfile.objects.get(user=user)
            
            # Get active academic year
            active_ay = AcademicYear.objects.filter(is_active=True).first()
            if not active_ay:
                return Response({
                    'detail': 'No active academic year found.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get HOD department roles for current academic year
            department_roles = DepartmentRole.objects.select_related('department').filter(
                staff=staff_profile,
                role='HOD',
                is_active=True,
                academic_year=active_ay
            ).order_by('department__name')
            
            departments_count = department_roles.count()
            
            if departments_count == 0:
                # Fall back to staff profile department
                if staff_profile.department:
                    return Response({
                        'success': True,
                        'has_multiple_departments': False,
                        'departments': [{
                            'id': staff_profile.department.id,
                            'name': staff_profile.department.name,
                            'code': staff_profile.department.code
                        }],
                        'active_department': {
                            'id': staff_profile.department.id,
                            'name': staff_profile.department.name,
                            'code': staff_profile.department.code
                        }
                    }, status=status.HTTP_200_OK)
                else:
                    return Response({
                        'success': False,
                        'department': None
                    }, status=status.HTTP_200_OK)
            
            # Build departments list
            departments = [{
                'id': dr.department.id,
                'name': dr.department.name,
                'code': dr.department.code
            } for dr in department_roles]
            
            # Handle active department selection
            active_department_id = request.GET.get('active_department_id')
            if active_department_id:
                try:
                    active_department_id = int(active_department_id)
                    # Verify this department belongs to the HOD
                    active_dept = next((d for d in departments if d['id'] == active_department_id), None)
                    if active_dept:
                        # Store in session
                        request.session['active_hod_department_id'] = active_department_id
                        active_department = active_dept
                    else:
                        # Invalid department ID, use first one
                        active_department = departments[0]
                        request.session['active_hod_department_id'] = departments[0]['id']
                except ValueError:
                    active_department = departments[0]
                    request.session['active_hod_department_id'] = departments[0]['id']
            else:
                # Check session for previously selected department
                session_dept_id = request.session.get('active_hod_department_id')
                if session_dept_id:
                    active_dept = next((d for d in departments if d['id'] == session_dept_id), None)
                    if active_dept:
                        active_department = active_dept
                    else:
                        # Session dept not found, use first
                        active_department = departments[0]
                        request.session['active_hod_department_id'] = departments[0]['id']
                else:
                    # No session, use first department
                    active_department = departments[0]
                    request.session['active_hod_department_id'] = departments[0]['id']
            
            return Response({
                'success': True,
                'has_multiple_departments': departments_count > 1,
                'departments': departments,
                'active_department': active_department
            }, status=status.HTTP_200_OK)
            
        except StaffProfile.DoesNotExist:
            return Response({
                'success': False,
                'department': None
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'detail': f'Error retrieving departments: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetClassOptionsView(APIView):
    """
    API 5: Get Class Options
    GET /api/feedback/class-options/
    
    Returns available years, semesters, and sections with year-section mappings.
    Used to populate dropdowns when creating feedback forms.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            from academics.models import Semester, Section, AcademicYear

            user_permissions = get_normalized_permissions(request.user)
            can_create_feedback = 'feedback.create' in user_permissions
            can_all_departments_access = 'feedback.all_departments_access' in user_permissions
            can_own_department_access = 'feedback.own_department_access' in user_permissions

            if not can_create_feedback:
                return Response({
                    'error': 'You do not have permission to create feedback forms.',
                    'success': False
                }, status=status.HTTP_403_FORBIDDEN)

            if not (can_all_departments_access or can_own_department_access):
                return Response({
                    'error': 'You do not have department access permission.',
                    'success': False
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get all academic years (1-4)
            years = [
                {"value": 1, "label": "1st Year"},
                {"value": 2, "label": "2nd Year"},
                {"value": 3, "label": "3rd Year"},
                {"value": 4, "label": "4th Year"}
            ]
            
            # Get all semesters from database
            semesters_qs = Semester.objects.all().order_by('number')
            semesters = [
                {"value": sem.id, "label": f"Semester {sem.number}", "number": sem.number}
                for sem in semesters_qs
            ]
            
            # Get sections that belong to HOD's active department(s)
            sections_filter = {}
            user_departments = []
            
            # Check if departments parameter is provided (for multi-select case)
            departments_param = request.GET.getlist('departments[]')
            if not departments_param:
                departments_param = request.GET.get('departments', '').split(',') if request.GET.get('departments') else []
            
            if can_all_departments_access:
                if departments_param:
                    sections_filter['batch__course__department_id__in'] = [int(d) for d in departments_param if str(d).strip()]
            else:
                try:
                    staff_profile = StaffProfile.objects.get(user=request.user)
                    from academics.models import DepartmentRole

                    active_ay = AcademicYear.objects.filter(is_active=True).first()
                    if active_ay:
                        department_roles = DepartmentRole.objects.select_related('department').filter(
                            staff=staff_profile,
                            role='HOD',
                            is_active=True,
                            academic_year=active_ay
                        )

                        departments_count = department_roles.count()

                        if departments_count > 1 and departments_param:
                            available_dept_ids = [dr.department.id for dr in department_roles]
                            for dept_id_str in departments_param:
                                if dept_id_str:
                                    dept_id = int(dept_id_str)
                                    if dept_id in available_dept_ids:
                                        from academics.models import Department
                                        try:
                                            dept = Department.objects.get(id=dept_id)
                                            user_departments.append(dept)
                                        except Department.DoesNotExist:
                                            pass
                        elif departments_count > 1:
                            user_departments = [dr.department for dr in department_roles]
                        elif departments_count == 1:
                            user_departments = [department_roles.first().department]
                        else:
                            if staff_profile.department:
                                user_departments = [staff_profile.department]
                    else:
                        if staff_profile.department:
                            user_departments = [staff_profile.department]

                    if user_departments:
                        sections_filter['batch__course__department__in'] = user_departments
                except StaffProfile.DoesNotExist:
                    pass
            
            # Get current academic year to calculate student years
            current_ay = AcademicYear.objects.filter(is_active=True).first()
            current_acad_year = None
            if current_ay:
                try:
                    current_acad_year = int(str(current_ay.name).split('-')[0])
                except:
                    pass
            
            # Build year-section mappings (1-4)
            year_sections = {1: [], 2: [], 3: [], 4: []}
            sections_all = []
            seen_section_ids = set()
            
            sections_qs = Section.objects.filter(**sections_filter).select_related('batch').order_by('name')
            
            for sec in sections_qs:
                if sec.id in seen_section_ids:
                    continue
                    
                seen_section_ids.add(sec.id)
                
                # Calculate year from batch
                student_year = None
                batch = sec.batch
                if batch and batch.start_year and current_acad_year:
                    delta = current_acad_year - int(batch.start_year)
                    student_year = delta + 1
                    # Include years 1-4
                    if student_year < 1 or student_year > 4:
                        continue  # Skip invalid years only
                
                section_data = {
                    "value": sec.id,
                    "label": f"Section {sec.name}",
                    "name": sec.name,
                    "year": student_year
                }
                
                sections_all.append(section_data)
                
                # Add to year mapping
                if student_year and student_year in year_sections:
                    year_sections[student_year].append(section_data)
            
            return Response({
                'years': years,
                'semesters': semesters,
                'sections': sections_all,
                'year_sections': year_sections,  # NEW: Year to sections mapping
                'success': True
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'Failed to fetch class options: {str(e)}',
                'success': False
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DeactivateFeedbackFormView(APIView):
    """
    API 6: Deactivate/Activate Feedback Form
    POST /api/feedback/<id>/toggle-active
    
    HOD can toggle the active status of a feedback form.
    When deactivated, form is hidden from students/staff but data is preserved.
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request, form_id):
        # Check if user has permission to create feedback (HOD only)
        user_permissions = get_user_permissions(request.user)
        if 'feedback.create' not in user_permissions:
            return Response({
                'detail': 'You do not have permission to deactivate feedback forms.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get the feedback form
        feedback_form = get_object_or_404(FeedbackForm, id=form_id)
        
        # Ensure the form was created by the current user
        if feedback_form.created_by != request.user:
            return Response({
                'detail': 'You can only deactivate forms you created.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Toggle active status
        feedback_form.active = not feedback_form.active
        feedback_form.save()
        
        return Response({
            'message': f'Feedback form {"activated" if feedback_form.active else "deactivated"} successfully',
            'active': feedback_form.active
        }, status=status.HTTP_200_OK)


class DeactivateAllFeedbackFormsView(APIView):
    """
    API: Deactivate All Active Feedback Forms
    POST /api/feedback/deactivate-all/

    IQAC/Admin can deactivate all currently active feedback forms in one action.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        scope = get_feedback_department_scope(request.user)
        if not scope.get('allowed'):
            return Response({
                'detail': 'You do not have permission to deactivate all feedback forms.'
            }, status=status.HTTP_403_FORBIDDEN)

        with transaction.atomic():
            qs = FeedbackForm.objects.filter(status='ACTIVE', active=True)
            qs = apply_department_scope_filter(qs, scope, field_name='department_id')
            updated = qs.update(active=False)

        return Response({
            'message': 'All active feedback forms deactivated',
            'count': updated,
        }, status=status.HTTP_200_OK)


class DeactivateFilteredFeedbackFormsView(APIView):
    """
    API: Deactivate filtered active feedback forms
    POST /api/feedback/deactivate-filtered/

    IQAC/Admin can deactivate active feedback forms filtered by
    selected departments and years.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        scope = get_feedback_department_scope(request.user)
        if not scope.get('allowed'):
            return Response({
                'detail': 'You do not have permission to deactivate feedback forms.'
            }, status=status.HTTP_403_FORBIDDEN)

        all_departments = bool(request.data.get('all_departments', False))
        department_ids = request.data.get('department_ids', []) or []
        all_years = bool(request.data.get('all_years', False))
        years = request.data.get('years', []) or []

        qs = FeedbackForm.objects.filter(status='ACTIVE', active=True)
        qs = apply_department_scope_filter(qs, scope, field_name='department_id')

        if not scope.get('all_departments'):
            # Own-department users are always scoped by server-side department filter.
            all_departments = True

        if not all_departments and department_ids:
            try:
                normalized_department_ids = [int(d) for d in department_ids]
                qs = qs.filter(department_id__in=normalized_department_ids)
            except Exception:
                return Response({
                    'detail': 'Invalid department_ids payload.'
                }, status=status.HTTP_400_BAD_REQUEST)

        if not all_years and years:
            year_filter = Q()
            for year in years:
                try:
                    y = int(year)
                except Exception:
                    continue
                year_filter |= Q(year=y)
                year_filter |= Q(years__contains=[y])

            if year_filter:
                qs = qs.filter(year_filter)

        with transaction.atomic():
            updated = qs.update(active=False)

        return Response({
            'message': 'Filtered active feedback forms deactivated',
            'count': updated,
        }, status=status.HTTP_200_OK)


class ActivateAllFeedbackFormsView(APIView):
    """
    API: Activate all deactivated feedback forms
    POST /api/feedback/activate-all/

    IQAC/Admin can activate all currently deactivated feedback forms.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        scope = get_feedback_department_scope(request.user)
        if not scope.get('allowed'):
            return Response({
                'detail': 'You do not have permission to activate all feedback forms.'
            }, status=status.HTTP_403_FORBIDDEN)

        with transaction.atomic():
            qs = FeedbackForm.objects.filter(status='ACTIVE', active=False)
            qs = apply_department_scope_filter(qs, scope, field_name='department_id')
            updated = qs.update(active=True)

        return Response({
            'message': 'All forms activated',
            'count': updated,
        }, status=status.HTTP_200_OK)


class ActivateFilteredFeedbackFormsView(APIView):
    """
    API: Activate filtered deactivated feedback forms
    POST /api/feedback/activate-filtered/

    IQAC/Admin can activate deactivated feedback forms filtered by
    selected departments and years.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        scope = get_feedback_department_scope(request.user)
        if not scope.get('allowed'):
            return Response({
                'detail': 'You do not have permission to activate feedback forms.'
            }, status=status.HTTP_403_FORBIDDEN)

        all_departments = bool(request.data.get('all_departments', False))
        department_ids = request.data.get('department_ids', []) or []
        all_years = bool(request.data.get('all_years', False))
        years = request.data.get('years', []) or []

        qs = FeedbackForm.objects.filter(status='ACTIVE', active=False)
        qs = apply_department_scope_filter(qs, scope, field_name='department_id')

        if not scope.get('all_departments'):
            # Own-department users are always scoped by server-side department filter.
            all_departments = True

        if not all_departments and department_ids:
            try:
                normalized_department_ids = [int(d) for d in department_ids]
                qs = qs.filter(department_id__in=normalized_department_ids)
            except Exception:
                return Response({
                    'detail': 'Invalid department_ids payload.'
                }, status=status.HTTP_400_BAD_REQUEST)

        if not all_years and years:
            year_filter = Q()
            for year in years:
                try:
                    y = int(year)
                except Exception:
                    continue
                year_filter |= Q(year=y)
                year_filter |= Q(years__contains=[y])

            if year_filter:
                qs = qs.filter(year_filter)

        with transaction.atomic():
            updated = qs.update(active=True)

        return Response({
            'message': 'Forms activated successfully',
            'count': updated,
        }, status=status.HTTP_200_OK)


class PublishFeedbackFormView(APIView):
    """
    API: Publish Feedback Form
    POST /api/feedback/<id>/publish
    
    HOD can publish a draft feedback form, changing status from DRAFT to ACTIVE.
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request, form_id):
        # Check if user has permission to create feedback (HOD only)
        user_permissions = get_user_permissions(request.user)
        if 'feedback.create' not in user_permissions:
            return Response({
                'detail': 'You do not have permission to publish feedback forms.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get the feedback form
        feedback_form = get_object_or_404(FeedbackForm, id=form_id)
        
        # Ensure the form was created by the current user
        if feedback_form.created_by != request.user:
            return Response({
                'detail': 'You can only publish forms you created.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check if already published
        if feedback_form.status == 'ACTIVE':
            return Response({
                'detail': 'This form is already published.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Publish the form (change status to ACTIVE)
        feedback_form.status = 'ACTIVE'
        feedback_form.save()
        
        return Response({
            'message': 'Feedback form published successfully',
            'status': feedback_form.status
        }, status=status.HTTP_200_OK)


class GetResponseStatisticsView(APIView):
    """
    API 7: Get Response Statistics
    GET /api/feedback/<id>/statistics
    
    HOD can view response count for a feedback form.
    Returns: total responses / expected responses (based on target audience)
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, form_id):
        # Check if user has permission to create feedback (HOD only)
        user_permissions = get_user_permissions(request.user)
        if 'feedback.create' not in user_permissions:
            return Response({
                'detail': 'You do not have permission to view response statistics.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get the feedback form
        feedback_form = get_object_or_404(FeedbackForm, id=form_id)
        
        # Ensure the form was created by the current user
        if feedback_form.created_by != request.user:
            return Response({
                'detail': 'You can only view statistics for forms you created.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get unique users who responded
        responded_users = FeedbackResponse.objects.filter(
            feedback_form=feedback_form
        ).values_list('user', flat=True).distinct()
        
        response_count = len(set(responded_users))
        
        # Calculate expected responses based on target type and class info
        expected_count = 0
        
        if feedback_form.target_type == 'STAFF':
            # Count staff in department
            expected_count = StaffProfile.objects.filter(
                department=feedback_form.department
            ).count()
        elif feedback_form.target_type == 'STUDENT':
            from academics.models import StudentProfile, Section, AcademicYear
            
            if feedback_form.all_classes:
                # All students in department
                expected_count = StudentProfile.objects.filter(
                    section__batch__course__department=feedback_form.department
                ).count()
            else:
                # Students matching year/semester/section criteria
                sections_to_query = []
                
                # Get current academic year for year calculation
                current_ay = AcademicYear.objects.filter(is_active=True).first()
                current_acad_year = None
                if current_ay:
                    try:
                        current_acad_year = int(str(current_ay.name).split('-')[0])
                    except:
                        pass
                
                # Build list of matching sections
                if feedback_form.sections:  # Multi-class: use sections list
                    sections_to_query = list(feedback_form.sections)
                elif feedback_form.section_id:  # Legacy: single section
                    sections_to_query = [feedback_form.section_id]
                else:
                    # No specific sections: query by year and semester
                    sections_filter = Q(batch__course__department=feedback_form.department)
                    
                    # Filter by years
                    if feedback_form.years:
                        year_filters = Q()
                        for year in feedback_form.years:
                            if current_acad_year:
                                batch_start_year = current_acad_year - year + 1
                                year_filters |= Q(batch__start_year=str(batch_start_year))
                        sections_filter &= year_filters
                    elif feedback_form.year:
                        if current_acad_year:
                            batch_start_year = current_acad_year - feedback_form.year + 1
                            sections_filter &= Q(batch__start_year=str(batch_start_year))
                    
                    # Filter by semesters
                    if feedback_form.semesters:
                        sections_filter &= Q(semester_id__in=feedback_form.semesters)
                    elif feedback_form.semester_id:
                        sections_filter &= Q(semester=feedback_form.semester)
                    
                    matching_sections = Section.objects.filter(sections_filter)
                    sections_to_query = [s.id for s in matching_sections]
                
                # Count students in matching sections
                if sections_to_query:
                    expected_count = StudentProfile.objects.filter(
                        section_id__in=sections_to_query
                    ).count()
        
        return Response({
            'feedback_form_id': form_id,
            'response_count': response_count,
            'expected_count': expected_count,
            'percentage': round((response_count / expected_count * 100) if expected_count > 0 else 0, 1)
        }, status=status.HTTP_200_OK)


class PrincipalCreateFeedbackView(APIView):
    """Create principal feedback forms from unified page payload."""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        user_permissions = get_normalized_permissions(request.user)
        required = {
            'feedback.principal_feedback_page',
            'feedback.principal_create',
            'feedback.principal_all_departments_access',
        }
        if not required.issubset(user_permissions):
            return Response({
                'detail': 'You do not have permission to create principal feedback.'
            }, status=status.HTTP_403_FORBIDDEN)

        target_audience = request.data.get('target_audience', [])
        if isinstance(target_audience, str):
            target_audience = [target_audience]
        if not isinstance(target_audience, list) or not target_audience:
            return Response({
                'detail': 'Please select at least one target audience.'
            }, status=status.HTTP_400_BAD_REQUEST)

        supported_targets = {'STUDENT', 'STAFF'}
        invalid_targets = [str(t) for t in target_audience if str(t).upper() not in supported_targets]
        if invalid_targets:
            return Response({
                'detail': 'Unsupported target audience selected.',
                'invalid_targets': invalid_targets,
            }, status=status.HTTP_400_BAD_REQUEST)

        questions_payload = request.data.get('questions', [])
        if not isinstance(questions_payload, list) or not questions_payload:
            return Response({
                'detail': 'At least one question is required.'
            }, status=status.HTTP_400_BAD_REQUEST)

        from academics.models import Department

        department_ids = list(Department.objects.values_list('id', flat=True))
        if not department_ids:
            return Response({
                'detail': 'No departments found to create principal feedback.'
            }, status=status.HTTP_400_BAD_REQUEST)

        normalized_questions = []
        for idx, question in enumerate(questions_payload):
            question_text = str(question.get('question_text', '')).strip()
            if not question_text:
                return Response({
                    'detail': f'Question {idx + 1}: question text is required.'
                }, status=status.HTTP_400_BAD_REQUEST)

            allow_rating = bool(question.get('allow_rating', True))
            allow_comment = bool(question.get('allow_comment', True))
            question_type = str(question.get('question_type', '')).strip().lower()
            if not question_type:
                if bool(question.get('allow_own_type', False)):
                    question_type = 'radio'
                elif allow_comment and not allow_rating:
                    question_type = 'text'
                else:
                    question_type = 'rating'

            options_payload = question.get('options', []) or []
            options = []
            if question_type in {'radio', 'rating_radio_comment'}:
                for opt in options_payload:
                    if isinstance(opt, dict):
                        text = str(opt.get('option_text', '')).strip()
                    else:
                        text = str(opt).strip()
                    if text:
                        options.append({'option_text': text})

            normalized_questions.append({
                'question': question_text,
                'question_type': question_type,
                'allow_rating': allow_rating,
                'allow_comment': allow_comment,
                'is_mandatory': bool(question.get('is_mandatory', True)),
                'order': idx + 1,
                'options': options,
            })

        created_forms = []
        errors = []
        is_anonymous = bool(request.data.get('is_anonymous', False))

        with transaction.atomic():
            for target in [str(t).upper() for t in target_audience]:
                for department_id in department_ids:
                    payload = {
                        'target_type': target,
                        'type': 'OPEN_FEEDBACK',
                        'is_subject_based': False,
                        'department': department_id,
                        'status': 'DRAFT',
                        # Preserve legacy behavior: anonymous flag maps to common comment capture setting.
                        'common_comment_enabled': bool(is_anonymous and target == 'STUDENT'),
                        'questions': normalized_questions,
                    }

                    serializer = FeedbackFormCreateSerializer(data=payload, context={'request': request})
                    if serializer.is_valid():
                        created_forms.append(serializer.save(created_by=request.user))
                    else:
                        errors.append({
                            'target_type': target,
                            'department_id': department_id,
                            'errors': serializer.errors,
                        })

        if not created_forms:
            return Response({
                'detail': 'Failed to create principal feedback forms.',
                'errors': errors,
            }, status=status.HTTP_400_BAD_REQUEST)

        response_status = status.HTTP_201_CREATED if not errors else status.HTTP_207_MULTI_STATUS
        return Response({
            'detail': (
                f'Created {len(created_forms)} principal feedback form(s).'
                if not errors else
                f'Created {len(created_forms)} form(s), {len(errors)} failed.'
            ),
            'created_count': len(created_forms),
            'feedback_form_id': created_forms[0].id,
            'created_forms': [
                {
                    'id': form.id,
                    'target_type': form.target_type,
                    'department_id': form.department_id,
                    'status': form.status,
                }
                for form in created_forms
            ],
            'errors': errors,
        }, status=response_status)


class PrincipalAnalyticsDashboardView(APIView):
    """List principal-created feedback forms with summary analytics."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        user_permissions = get_normalized_permissions(request.user)
        required = {
            'feedback.principal_feedback_page',
            'feedback.principal_analytics',
            'feedback.principal_all_departments_access',
        }
        if not required.issubset(user_permissions):
            return Response({
                'detail': 'You do not have permission to view principal analytics dashboard.'
            }, status=status.HTTP_403_FORBIDDEN)

        forms = FeedbackForm.objects.filter(
            created_by=request.user,
            is_subject_based=False,
        ).prefetch_related('questions')

        items = []
        for form in forms.order_by('-created_at'):
            response_count, expected_count, percentage = calculate_feedback_response_metrics(form)
            items.append({
                'id': form.id,
                'feedback_type': 'PRINCIPAL',
                'target_audience': form.target_type,
                'is_anonymous': False,
                'status': form.status,
                'created_at': form.created_at,
                'response_count': response_count,
                'expected_count': expected_count,
                'percentage': percentage,
                'questions_count': form.questions.count(),
            })

        return Response({'items': items}, status=status.HTTP_200_OK)


class PrincipalFormAnalyticsView(APIView):
    """Get principal analytics details for one feedback form."""

    permission_classes = [IsAuthenticated]

    def get(self, request, form_id):
        user_permissions = get_normalized_permissions(request.user)
        required = {
            'feedback.principal_feedback_page',
            'feedback.principal_analytics',
            'feedback.principal_all_departments_access',
        }
        if not required.issubset(user_permissions):
            return Response({
                'detail': 'You do not have permission to view principal analytics.'
            }, status=status.HTTP_403_FORBIDDEN)

        form = get_object_or_404(
            FeedbackForm,
            id=form_id,
            created_by=request.user,
            is_subject_based=False,
        )

        response_count, expected_count, percentage = calculate_feedback_response_metrics(form)

        questions = []
        question_qs = FeedbackQuestion.objects.filter(feedback_form=form).order_by('order', 'id')
        option_map = {}
        question_ids = list(question_qs.values_list('id', flat=True))
        if question_ids:
            for opt in FeedbackQuestionOption.objects.filter(question_id__in=question_ids).values('id', 'question_id', 'option_text'):
                option_map.setdefault(opt['question_id'], []).append({
                    'id': opt['id'],
                    'option_text': opt['option_text'],
                })

        response_counts_by_question = {
            row['question_id']: row['count']
            for row in FeedbackResponse.objects.filter(
                feedback_form=form,
                question_id__in=question_ids,
            ).values('question_id').annotate(count=Count('id'))
        }

        for q in question_qs:
            questions.append({
                'id': q.id,
                'question_text': q.question,
                'question_type': q.question_type,
                'is_mandatory': bool(getattr(q, 'is_mandatory', False)),
                'responses_count': int(response_counts_by_question.get(q.id, 0)),
                'options': option_map.get(q.id, []),
            })

        return Response({
            'feedback_form_id': form.id,
            'feedback_type': 'PRINCIPAL',
            'target_audience': form.target_type,
            'is_anonymous': False,
            'status': form.status,
            'created_at': form.created_at,
            'response_count': response_count,
            'expected_count': expected_count,
            'percentage': percentage,
            'questions': questions,
        }, status=status.HTTP_200_OK)


class GetResponseListView(APIView):
    """
    API 8: Get Response List
    GET /api/feedback/<id>/responses
    
    HOD can view detailed response list for a feedback form.
    Shows: responded users with answers, and non-responded users.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, form_id):
        import logging
        logger = logging.getLogger(__name__)
        
        logger.info(f"[GetResponseListView] User {request.user.id} ({request.user.username}) requesting responses for form {form_id}")
        
        # Check if user has permission to create feedback (HOD only)
        user_permissions = get_user_permissions(request.user)
        logger.info(f"[GetResponseListView] User permissions: {user_permissions}")
        
        if 'feedback.create' not in user_permissions:
            logger.warning(f"[GetResponseListView] User {request.user.username} lacks feedback.create permission")
            return Response({
                'detail': 'You do not have permission to view responses. Only HODs and advisors can view feedback responses.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get the feedback form
        try:
            feedback_form = get_object_or_404(FeedbackForm, id=form_id)
            logger.info(f"[GetResponseListView] Found feedback form {form_id}, created by user {feedback_form.created_by.id}")
        except Exception as e:
            logger.error(f"[GetResponseListView] Error finding form: {e}")
            return Response({
                'detail': f'Feedback form not found.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Ensure the form was created by the current user
        if feedback_form.created_by != request.user:
            logger.warning(f"[GetResponseListView] User {request.user.id} tried to access form created by user {feedback_form.created_by.id}")
            return Response({
                'detail': f'You can only view responses for forms you created. This form was created by {feedback_form.created_by.get_full_name() or feedback_form.created_by.username}.'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get all responses for this form, grouped by user
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            logger.info(f"[GetResponseListView] Processing responses for form {form_id}")
            
            responses_by_user = {}
            all_responses = FeedbackResponse.objects.filter(
                feedback_form=feedback_form
            ).select_related(
                'user', 
                'question',
                'teaching_assignment',
                'teaching_assignment__staff',
                'teaching_assignment__staff__user',
                'teaching_assignment__curriculum_row',
                'teaching_assignment__subject',
                'teaching_assignment__elective_subject'
            )
            
            logger.info(f"[GetResponseListView] Found {all_responses.count()} total responses")
            
            for response in all_responses:
                user_id = response.user.id
                if user_id not in responses_by_user:
                    # Get user details
                    user = response.user
                    user_name = user.get_full_name() or user.username
                    register_number = None
                    
                    # Get register number from profile
                    if feedback_form.target_type == 'STUDENT':
                        try:
                            from academics.models import StudentProfile
                            student_profile = StudentProfile.objects.get(user=user)
                            register_number = getattr(student_profile, 'reg_no', None) or user.username
                        except StudentProfile.DoesNotExist:
                            register_number = user.username
                    elif feedback_form.target_type == 'STAFF':
                        try:
                            staff_profile = StaffProfile.objects.get(user=user)
                            register_number = getattr(staff_profile, 'staff_id', None) or user.username
                        except StaffProfile.DoesNotExist:
                            register_number = user.username
                    
                    responses_by_user[user_id] = {
                        'user_id': user_id,
                        'user_name': user_name,
                        'register_number': register_number,
                        'submitted_at': response.created_at.isoformat(),
                        'answers': []
                    }
                
                # Get teaching assignment details if present (for subject feedback)
                teaching_assignment_data = None
                if response.teaching_assignment:
                    ta = response.teaching_assignment
                    
                    # Get subject name
                    subject_name = None
                    subject_code = None
                    if ta.curriculum_row:
                        subject_name = ta.curriculum_row.course_name
                        subject_code = ta.curriculum_row.course_code
                    elif ta.subject:
                        subject_name = ta.subject.name
                        subject_code = ta.subject.code
                    elif ta.elective_subject:
                        subject_name = ta.elective_subject.course_name
                        subject_code = ta.elective_subject.course_code
                    elif ta.custom_subject:
                        subject_name = ta.get_custom_subject_display()
                        subject_code = ta.custom_subject
                    
                    # Get staff name
                    staff_name = ta.staff.user.get_full_name() or ta.staff.user.username if ta.staff else None
                    
                    teaching_assignment_data = {
                        'teaching_assignment_id': ta.id,
                        'subject_name': subject_name,
                        'subject_code': subject_code,
                        'staff_name': staff_name
                    }
                
                # Add answer (runs for every response)
                responses_by_user[user_id]['answers'].append({
                    'question_id': response.question.id,
                    'question_text': response.question.question,
                    'answer_type': response.question.answer_type,
                    'answer_star': response.answer_star,
                    # Keep both key styles for frontend compatibility.
                    'question_comment': response.answer_text,
                    'answer_text': response.answer_text,
                    'common_comment': response.common_comment,
                    'selected_option': response.selected_option_text,
                    'selected_option_text': response.selected_option_text,
                    'teaching_assignment': teaching_assignment_data
                })
            
            # Get list of users who should have responded but didn't
            non_responders = []
            expected_users = []
            
            if feedback_form.target_type == 'STAFF':
                expected_users = User.objects.filter(
                    staff_profile__department=feedback_form.department
                ).exclude(id=feedback_form.created_by.id).values_list('id', flat=True)
            elif feedback_form.target_type == 'STUDENT':
                from academics.models import StudentProfile, Section, AcademicYear
                
                if feedback_form.all_classes:
                    expected_users = User.objects.filter(
                        student_profile__section__batch__course__department=feedback_form.department
                    ).exclude(id=feedback_form.created_by.id).values_list('id', flat=True)
                else:
                    # Build query for matching students
                    sections_to_query = []
                    
                    current_ay = AcademicYear.objects.filter(is_active=True).first()
                    current_acad_year = None
                    if current_ay:
                        try:
                            current_acad_year = int(str(current_ay.name).split('-')[0])
                        except:
                            pass
                    
                    if feedback_form.sections:
                        sections_to_query = list(feedback_form.sections)
                    elif feedback_form.section_id:
                        sections_to_query = [feedback_form.section_id]
                    else:
                        sections_filter = Q(batch__course__department=feedback_form.department)
                        
                        if feedback_form.years:
                            year_filters = Q()
                            for year in feedback_form.years:
                                if current_acad_year:
                                    batch_start_year = current_acad_year - year + 1
                                    year_filters |= Q(batch__start_year=str(batch_start_year))
                            sections_filter &= year_filters
                        elif feedback_form.year:
                            if current_acad_year:
                                batch_start_year = current_acad_year - feedback_form.year + 1
                                sections_filter &= Q(batch__start_year=str(batch_start_year))
                        
                        if feedback_form.semesters:
                            sections_filter &= Q(semester_id__in=feedback_form.semesters)
                        elif feedback_form.semester_id:
                            sections_filter &= Q(semester=feedback_form.semester)
                        
                        matching_sections = Section.objects.filter(sections_filter)
                        sections_to_query = [s.id for s in matching_sections]
                    
                    if sections_to_query:
                        expected_users = User.objects.filter(
                            student_profile__section_id__in=sections_to_query
                        ).exclude(id=feedback_form.created_by.id).values_list('id', flat=True)
            
            # Find non-responders
            responded_user_ids = set(responses_by_user.keys())
            for user_id in expected_users:
                if user_id not in responded_user_ids:
                    user = User.objects.get(id=user_id)
                    user_name = user.get_full_name() or user.username
                    register_number = None
                    
                    if feedback_form.target_type == 'STUDENT':
                        try:
                            from academics.models import StudentProfile
                            student_profile = StudentProfile.objects.get(user=user)
                            register_number = getattr(student_profile, 'reg_no', None) or user.username
                        except StudentProfile.DoesNotExist:
                            register_number = user.username
                    elif feedback_form.target_type == 'STAFF':
                        try:
                            staff_profile = StaffProfile.objects.get(user=user)
                            register_number = getattr(staff_profile, 'staff_id', None) or user.username
                        except StaffProfile.DoesNotExist:
                            register_number = user.username
                    
                    non_responders.append({
                        'user_id': user_id,
                        'user_name': user_name,
                        'register_number': register_number
                    })
            
            logger.info(f"[GetResponseListView] Successfully processed: {len(responses_by_user)} responded, {len(non_responders)} non-responders")
            
            return Response({
                'feedback_form_id': form_id,
                'target_type': feedback_form.target_type,
                'responded': list(responses_by_user.values()),
                'non_responders': non_responders,
                'total_responded': len(responses_by_user),
                'total_non_responded': len(non_responders)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[GetResponseListView] Unexpected error: {str(e)}", exc_info=True)
            return Response({
                'detail': f'Error processing responses: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetStudentSubjectsView(APIView):
    """
    API 9: Get Student's Subjects for Subject Feedback
    GET /api/feedback/<form_id>/subjects/
    
    Returns list of subjects (teaching assignments) for the student to rate.
    Excludes 1st year students from subject feedback.
    
    Student View: Shows ONLY core subjects + electives the student has chosen via ElectiveChoice.
    This ensures students don't see electives they haven't selected.
    Staff names are taken from TeachingAssignment where available.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, form_id):
        print(f"[GetStudentSubjectsView] User: {request.user.username}, Form ID: {form_id}")
        try:
            # Get the feedback form
            feedback_form = get_object_or_404(FeedbackForm, id=form_id)
            
            print(f"[GetStudentSubjectsView] Feedback form found: {feedback_form.type}, Status: {feedback_form.status}, Active: {feedback_form.active}")
            
            # Check if this is subject feedback
            if feedback_form.type != 'SUBJECT_FEEDBACK':
                return Response({
                    'detail': 'This endpoint is only for subject feedback forms.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if form is active and published
            if feedback_form.status != 'ACTIVE':
                return Response({
                    'detail': f'This feedback form has not been published yet. Current status: {feedback_form.get_status_display()}',
                    'subjects': [],
                    'total_subjects': 0,
                    'completed_subjects': 0,
                    'all_completed': False,
                    'form_status': feedback_form.status
                }, status=status.HTTP_200_OK)
            
            if not feedback_form.active:
                return Response({
                    'detail': 'This feedback form has been deactivated.',
                    'subjects': [],
                    'total_subjects': 0,
                    'completed_subjects': 0,
                    'all_completed': False
                }, status=status.HTTP_200_OK)
            
            # Get student profile
            try:
                from academics.models import StudentProfile, TeachingAssignment, AcademicYear
                student_profile = StudentProfile.objects.get(user=request.user)
                print(f"[GetStudentSubjectsView] ===== STUDENT INFO =====")
                print(f"[GetStudentSubjectsView] Student: {request.user.username} (ID: {request.user.id})")
                print(f"[GetStudentSubjectsView] StudentProfile ID: {student_profile.id}")
                print(f"[GetStudentSubjectsView] Section: {student_profile.section}")
                
                # Quick check for ElectiveChoice records
                from curriculum.models import ElectiveChoice
                total_choices = ElectiveChoice.objects.filter(student=student_profile).count()
                active_choices = ElectiveChoice.objects.filter(student=student_profile, is_active=True).count()
                print(f"[GetStudentSubjectsView] ElectiveChoice records: {total_choices} total, {active_choices} active")
                
                if total_choices == 0:
                    print(f"[GetStudentSubjectsView] ⚠⚠⚠ NO ElectiveChoice records found for this student!")
                    print(f"[GetStudentSubjectsView] ⚠⚠⚠ Student has not chosen any electives yet!")
                
            except StudentProfile.DoesNotExist:
                return Response({
                    'detail': 'Student profile not found.'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Calculate student's current year
            section = student_profile.section
            if not section:
                return Response({
                    'detail': 'Section information not found.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            batch = section.batch
            student_year = None
            
            # Calculate year from batch
            if batch and batch.start_year:
                current_ay = AcademicYear.objects.filter(is_active=True).first()
                if current_ay:
                    try:
                        acad_start = int(str(current_ay.name).split('-')[0])
                        delta = acad_start - int(batch.start_year)
                        student_year = delta + 1
                        print(f"[GetStudentSubjectsView] Student year calculated: {student_year} (batch start: {batch.start_year}, current AY: {current_ay.name})")
                    except Exception as calc_err:
                        print(f"[GetStudentSubjectsView] Error calculating year: {calc_err}")
                        pass
            
            if not student_year:
                print(f"[GetStudentSubjectsView] WARNING: Could not determine student year")
            
            # Exclude 1st year students
            if student_year == 1:
                return Response({
                    'detail': 'Subject feedback is not applicable for 1st year students.',
                    'is_first_year': True,
                    'subjects': [],
                    'total_subjects': 0,
                    'completed_subjects': 0,
                    'all_completed': False
                }, status=status.HTTP_200_OK)
            
            # Get student's department
            student_department = None
            if section.batch and section.batch.course:
                student_department = section.batch.course.department
            
            print(f"[GetStudentSubjectsView] Student info - Year: {student_year}, Section: {section.name} (ID: {section.id}), Department: {student_department.code if student_department else 'None'}")
            print(f"[GetStudentSubjectsView] Feedback targets - Years: {feedback_form.years}, Sections: {feedback_form.sections}, Department: {feedback_form.department.code}")
            
            # Validate if student's year and section match feedback form's targets
            # Check if feedback targets specific years
            if feedback_form.years and len(feedback_form.years) > 0:
                if student_year not in feedback_form.years:
                    return Response({
                        'detail': f'This feedback is not for your year (Year {student_year}).',
                        'subjects': [],
                        'total_subjects': 0,
                        'completed_subjects': 0,
                        'all_completed': False
                    }, status=status.HTTP_200_OK)
            
            # Check if feedback targets specific sections
            if feedback_form.sections and len(feedback_form.sections) > 0:
                if section.id not in feedback_form.sections:
                    return Response({
                        'detail': f'This feedback is not for your section ({section.name}).',
                        'subjects': [],
                        'total_subjects': 0,
                        'completed_subjects': 0,
                        'all_completed': False
                    }, status=status.HTTP_200_OK)
            
            # Check if feedback targets specific department
            if student_department and feedback_form.department.id != student_department.id:
                return Response({
                    'detail': 'This feedback is not for your department.',
                    'subjects': [],
                    'total_subjects': 0,
                    'completed_subjects': 0,
                    'all_completed': False
                }, status=status.HTTP_200_OK)
            
            # ========== STEP 1: Load CORE subjects from curriculum ==========
            # Fetch core subjects directly from CurriculumDepartment to avoid missing subjects
            # due to strict section filtering or missing TeachingAssignment records
            print(f"[GetStudentSubjectsView] STEP 1: Loading CORE subjects from curriculum")
            
            current_ay = AcademicYear.objects.filter(is_active=True).first()
            
            from curriculum.models import CurriculumDepartment
            
            # Get semesters from feedback form (if specified)
            target_semesters = []
            if hasattr(feedback_form, 'semesters') and feedback_form.semesters:
                target_semesters = feedback_form.semesters
                print(f"[GetStudentSubjectsView] Feedback form targets specific semesters: {target_semesters}")
            
            # If no semesters specified, use student's batch to determine
            if not target_semesters and batch and batch.start_year:
                # Determine semester based on year
                try:
                    acad_start = int(str(current_ay.name).split('-')[0])
                    delta = acad_start - int(batch.start_year)
                    calculated_year = delta + 1
                    
                    # Map year to semesters (e.g., Year 3 = Sem 5 or 6)
                    if calculated_year == 2:
                        target_semesters = [3, 4]  # 2nd year = Sem 3 or 4
                    elif calculated_year == 3:
                        target_semesters = [5, 6]  # 3rd year = Sem 5 or 6
                    elif calculated_year == 4:
                        target_semesters = [7, 8]  # 4th year = Sem 7 or 8
                    
                    print(f"[GetStudentSubjectsView] Calculated year {calculated_year}, using semesters: {target_semesters}")
                except Exception as e:
                    print(f"[GetStudentSubjectsView] Error calculating semesters: {e}")
            
            # Fallback: if still no semesters, fetch all for department
            if not target_semesters:
                print(f"[GetStudentSubjectsView] No semesters determined, will fetch all core subjects for department")
            
            # Fetch core curriculum subjects
            curriculum_filter = {
                'department': student_department,
                'is_elective': False,  # Core subjects only
            }
            
            if target_semesters:
                curriculum_filter['semester__number__in'] = target_semesters
            
            curriculum_subjects = CurriculumDepartment.objects.filter(
                **curriculum_filter
            ).select_related(
                'semester',
                'department'
            ).order_by('course_code')
            
            print(f"[GetStudentSubjectsView] Found {curriculum_subjects.count()} core curriculum subjects")
            
            # Now try to match each curriculum subject with a TeachingAssignment
            core_teaching_assignments = []
            subjects_without_ta = []
            
            for curr_subj in curriculum_subjects:
                # Try to find TeachingAssignment for this subject
                # First try: Match by section (preferred)
                ta_query = TeachingAssignment.objects.filter(
                    curriculum_row=curr_subj,
                    section=section,
                    academic_year=current_ay,
                    is_active=True
                ).select_related('staff', 'staff__user', 'curriculum_row')
                
                ta = ta_query.first()
                
                # Second try: Match by department (if section-specific TA not found)
                if not ta:
                    ta_query = TeachingAssignment.objects.filter(
                        curriculum_row=curr_subj,
                        academic_year=current_ay,
                        is_active=True
                    ).select_related('staff', 'staff__user', 'curriculum_row')
                    ta = ta_query.first()
                
                if ta:
                    core_teaching_assignments.append(ta)
                    staff_name = ta.staff.user.get_full_name() if ta.staff and ta.staff.user else "No staff"
                    print(f"[GetStudentSubjectsView]   ✓ {curr_subj.course_code}: {curr_subj.course_name} (Staff: {staff_name})")
                else:
                    # No TA found - will create pseudo-assignment
                    subjects_without_ta.append(curr_subj)
                    print(f"[GetStudentSubjectsView]   ⚠ {curr_subj.course_code}: {curr_subj.course_name} (No TA - will create pseudo)")
            
            print(f"[GetStudentSubjectsView] Matched {len(core_teaching_assignments)} subjects with TAs, {len(subjects_without_ta)} without TAs")
            
            # ========== STEP 2: Load Student's Elective Choices ==========
            print(f"[GetStudentSubjectsView] STEP 2: Loading student's elective choices")
            
            from curriculum.models import ElectiveChoice, ElectiveSubject
            
            # Get student's chosen electives
            # CRITICAL: Get ALL active choices without academic year restriction
            # Elective subjects may span multiple semesters within the same academic year
            # Students should only have active ElectiveChoice records for their current selections
            elective_choices = ElectiveChoice.objects.filter(
                student=student_profile,
                is_active=True
            ).select_related(
                'elective_subject', 
                'elective_subject__parent', 
                'elective_subject__department'
            )
            
            print(f"[GetStudentSubjectsView] Total elective choices (all AYs, is_active=True): {elective_choices.count()}")
            
            if elective_choices.count() == 0:
                print(f"[GetStudentSubjectsView] ⚠⚠⚠ NO active ElectiveChoice records found for this student!")
                print(f"[GetStudentSubjectsView] ⚠⚠⚠ Student will see NO electives in feedback form!")
            
            for i, choice in enumerate(elective_choices, 1):
                subj_code = choice.elective_subject.course_code if choice.elective_subject else 'N/A'
                subj_name = choice.elective_subject.course_name if choice.elective_subject else 'N/A'
                ay_name = choice.academic_year.name if choice.academic_year else 'No AY'
                sem_num = choice.elective_subject.semester.number if (choice.elective_subject and choice.elective_subject.semester) else 'N/A'
                print(f"[GetStudentSubjectsView]   Choice {i}: {subj_code} - {subj_name} (AY: {ay_name}, Sem: {sem_num}, Active: {choice.is_active})")
            
            # Check if filtering by current AY would reduce the count
            ay_filtered_choices = elective_choices.filter(academic_year=current_ay)
            ay_filtered_count = ay_filtered_choices.count()
            total_count = elective_choices.count()
            
            print(f"[GetStudentSubjectsView] AY filtering analysis: Total={total_count}, Current AY={ay_filtered_count}")
            
            # DECISION: Use ALL active choices without AY filtering
            # This ensures electives from different semesters within the year are all included
            # Students should only have is_active=True for their current selections
            if ay_filtered_count > 0 and ay_filtered_count < total_count:
                print(f"[GetStudentSubjectsView] ⚠️ WARNING: AY filter would exclude {total_count - ay_filtered_count} choices")
                print(f"[GetStudentSubjectsView] DECISION: Using ALL {total_count} active choices (no AY filter)")
            elif ay_filtered_count == total_count:
                print(f"[GetStudentSubjectsView] INFO: All choices belong to current AY")
            else:
                print(f"[GetStudentSubjectsView] INFO: Using all {total_count} active choices")
            
            # Do NOT filter by academic_year - use all active choices
            # This ensures 2nd year electives from both Sem 3 and Sem 4 are included
            
            # Build list of chosen course codes AND department mapping
            # CRITICAL: Map course_code -> department_id to filter staff by chosen department
            chosen_elective_codes = []
            elective_dept_map = {}  # Map course_code -> department_id from student's choice
            seen_codes_for_building = set()  # Track to avoid duplicates in chosen_elective_codes
            
            for choice in elective_choices:
                if choice.elective_subject and choice.elective_subject.course_code:
                    code = choice.elective_subject.course_code
                    dept = choice.elective_subject.department
                    
                    # Only add if not already added (avoid duplicate codes)
                    if code not in seen_codes_for_building:
                        chosen_elective_codes.append(code)
                        elective_dept_map[code] = dept.id if dept else None
                        seen_codes_for_building.add(code)
                        
                        dept_name = dept.code if dept else "No Dept"
                        print(f"[GetStudentSubjectsView]   - Student chose: {code} - {choice.elective_subject.course_name} (from {dept_name} dept)")
                    else:
                        print(f"[GetStudentSubjectsView]   - Skipping duplicate code: {code}")
            
            print(f"[GetStudentSubjectsView] Student's chosen elective codes ({len(chosen_elective_codes)} total): {chosen_elective_codes}")
            print(f"[GetStudentSubjectsView] Elective department mapping: {elective_dept_map}")
            
            if not chosen_elective_codes:
                print(f"[GetStudentSubjectsView] ⚠️ WARNING: No elective codes extracted from {elective_choices.count()} choices!")
            
            # ========== STEP 3: Load Elective Teaching Assignments WITH department preference ==========
            # CRITICAL: Show ALL student-chosen electives, prefer staff from chosen department
            # Student should see staff from the department they chose the elective from when available
            # But if not available, still show the elective with any available staff or "Staff Not Assigned"
            print(f"[GetStudentSubjectsView] STEP 3: Loading elective teaching assignments (prefer chosen department)")
            
            elective_teaching_assignments = []
            staff_map = {}  # Map course_code -> list of staff names
            
            if chosen_elective_codes:
                # Query each elective individually to apply correct department preference
                
                for code in chosen_elective_codes:
                    dept_id = elective_dept_map.get(code)
                    print(f"[GetStudentSubjectsView]   Processing {code} (preferred department: {dept_id})")
                    
                    found_staff = False
                    
                    # ATTEMPT 1: Try with department filter (preferred)
                    if dept_id:
                        ta_query = TeachingAssignment.objects.filter(
                            elective_subject__course_code=code,
                            elective_subject__department_id=dept_id,
                            academic_year=current_ay,
                            is_active=True
                        ).select_related('staff', 'staff__user', 'elective_subject', 'elective_subject__department')
                        
                        for ta in ta_query:
                            if ta.staff and ta.staff.user:
                                staff_name = ta.staff.user.get_full_name() or ta.staff.user.username or "Unknown"
                            else:
                                staff_name = "Unknown"
                            
                            if code not in staff_map:
                                staff_map[code] = []
                            if staff_name and staff_name.strip() and staff_name not in staff_map[code]:
                                staff_map[code].append(staff_name)
                                dept_code = ta.elective_subject.department.code if ta.elective_subject.department else "N/A"
                                print(f"[GetStudentSubjectsView]     ✓ Found TA (preferred dept): {code} - {staff_name} (dept: {dept_code})")
                                found_staff = True
                    
                    # ATTEMPT 2: Try WITHOUT department filter (fallback)
                    if not found_staff:
                        ta_query = TeachingAssignment.objects.filter(
                            elective_subject__course_code=code,
                            academic_year=current_ay,
                            is_active=True
                        ).select_related('staff', 'staff__user', 'elective_subject', 'elective_subject__department')
                        
                        for ta in ta_query:
                            if ta.staff and ta.staff.user:
                                staff_name = ta.staff.user.get_full_name() or ta.staff.user.username or "Unknown"
                            else:
                                staff_name = "Unknown"
                            
                            if code not in staff_map:
                                staff_map[code] = []
                            if staff_name and staff_name.strip() and staff_name not in staff_map[code]:
                                staff_map[code].append(staff_name)
                                dept_code = ta.elective_subject.department.code if ta.elective_subject.department else "N/A"
                                print(f"[GetStudentSubjectsView]     ✓ Found TA (any dept): {code} - {staff_name} (dept: {dept_code})")
                                found_staff = True
                    
                    # ATTEMPT 3: Try curriculum_row method (for elective placeholders)
                    if not found_staff:
                        ta_query_base = TeachingAssignment.objects.filter(
                            curriculum_row__course_code=code,
                            curriculum_row__is_elective=True,
                            academic_year=current_ay,
                            is_active=True
                        )
                        
                        # First try with dept filter if available
                        if dept_id:
                            ta_query = ta_query_base.filter(curriculum_row__department_id=dept_id)
                            tas = ta_query.select_related('staff', 'staff__user', 'curriculum_row', 'curriculum_row__department')
                        else:
                            tas = ta_query_base.select_related('staff', 'staff__user', 'curriculum_row', 'curriculum_row__department')
                        
                        # If nothing found with dept filter, try without
                        if not tas.exists() and dept_id:
                            tas = ta_query_base.select_related('staff', 'staff__user', 'curriculum_row', 'curriculum_row__department')
                        
                        for ta in tas:
                            if ta.staff and ta.staff.user:
                                staff_name = ta.staff.user.get_full_name() or ta.staff.user.username or "Unknown"
                            else:
                                staff_name = "Unknown"
                            if code not in staff_map:
                                staff_map[code] = []
                            if staff_name and staff_name.strip() and staff_name not in staff_map[code]:
                                staff_map[code].append(staff_name)
                                dept_code = ta.curriculum_row.department.code if ta.curriculum_row.department else "N/A"
                                print(f"[GetStudentSubjectsView]     ✓ Found TA (curriculum_row): {code} - {staff_name} (dept: {dept_code})")
                                found_staff = True
                    
                    # ATTEMPT 4: Try legacy subject table (last resort)
                    if not found_staff:
                        from academics.models import Subject
                        tas_by_subject = TeachingAssignment.objects.filter(
                            subject__code=code,
                            academic_year=current_ay,
                            is_active=True
                        ).select_related('staff', 'staff__user', 'subject')
                        
                        for ta in tas_by_subject:
                            if ta.staff and ta.staff.user:
                                staff_name = ta.staff.user.get_full_name() or ta.staff.user.username or "Unknown"
                            else:
                                staff_name = "Unknown"
                            if code not in staff_map:
                                staff_map[code] = []
                            if staff_name and staff_name.strip() and staff_name not in staff_map[code]:
                                staff_map[code].append(staff_name)
                                print(f"[GetStudentSubjectsView]     ✓ Found TA (subject): {code} - {staff_name}")
                                found_staff = True
                    
                    # If still no staff found, mark for later pseudo-assignment creation
                    if not found_staff:
                        print(f"[GetStudentSubjectsView]     ⚠ No TA found for {code}, will create pseudo-assignment")
                
                # Collect unique TAs for electives (store first TA for each course_code for feedback submission)
                # Query again to get actual TA objects for each code
                seen_codes = set()
                for code in chosen_elective_codes:
                    if code in seen_codes:
                        continue
                    
                    dept_id = elective_dept_map.get(code)
                    ta = None
                    
                    # Try to find a TA for this code (prefer department filter, fallback to any)
                    if dept_id:
                        ta = TeachingAssignment.objects.filter(
                            elective_subject__course_code=code,
                            elective_subject__department_id=dept_id,
                            academic_year=current_ay,
                            is_active=True
                        ).select_related('staff', 'staff__user', 'elective_subject', 'elective_subject__department').first()
                    
                    # Fallback: try without department filter
                    if not ta:
                        ta = TeachingAssignment.objects.filter(
                            elective_subject__course_code=code,
                            academic_year=current_ay,
                            is_active=True
                        ).select_related('staff', 'staff__user', 'elective_subject', 'elective_subject__department').first()
                    
                    if ta:
                        elective_teaching_assignments.append(ta)
                        seen_codes.add(code)
                    else:
                        # Try curriculum_row method (prefer dept, fallback to any)
                        ta = None
                        if dept_id:
                            ta = TeachingAssignment.objects.filter(
                                curriculum_row__course_code=code,
                                curriculum_row__is_elective=True,
                                curriculum_row__department_id=dept_id,
                                academic_year=current_ay,
                                is_active=True
                            ).select_related('staff', 'staff__user', 'curriculum_row').first()
                        
                        # Fallback: try without department filter
                        if not ta:
                            ta = TeachingAssignment.objects.filter(
                                curriculum_row__course_code=code,
                                curriculum_row__is_elective=True,
                                academic_year=current_ay,
                                is_active=True
                            ).select_related('staff', 'staff__user', 'curriculum_row').first()
                        
                        if ta:
                            elective_teaching_assignments.append(ta)
                            seen_codes.add(code)
                        else:
                            # Try subject method as last resort
                            from academics.models import Subject
                            ta = TeachingAssignment.objects.filter(
                                subject__code=code,
                                academic_year=current_ay,
                                is_active=True
                            ).select_related('staff', 'staff__user', 'subject').first()
                            
                            if ta:
                                elective_teaching_assignments.append(ta)
                                seen_codes.add(code)
            
            print(f"[GetStudentSubjectsView] Found {len(elective_teaching_assignments)} elective teaching assignments")
            print(f"[GetStudentSubjectsView] Expected {len(chosen_elective_codes)} electives, got {len(elective_teaching_assignments)} TAs")
            print(f"[GetStudentSubjectsView] Staff mapping for {len(staff_map)} elective codes:")
            for code, staff_list in staff_map.items():
                print(f"[GetStudentSubjectsView]   {code}: {staff_list}")
            
            # Check for missing electives
            ta_codes = set()
            for ta in elective_teaching_assignments:
                if ta.elective_subject:
                    ta_codes.add(ta.elective_subject.course_code)
                elif ta.curriculum_row:
                    ta_codes.add(ta.curriculum_row.course_code)
                elif ta.subject:
                    ta_codes.add(ta.subject.code)
                elif hasattr(ta, 'id') and isinstance(ta.id, str) and ta.id.startswith('pseudo_'):
                    if ta.elective_subject:
                        ta_codes.add(ta.elective_subject.course_code)
            
            missing_codes = set(chosen_elective_codes) - ta_codes
            if missing_codes:
                print(f"[GetStudentSubjectsView] ⚠️ WARNING: Missing TAs for codes: {missing_codes}")
            
            # ========== STEP 4: Load Elective Subject Details ==========
            print(f"[GetStudentSubjectsView] STEP 4: Loading elective subject details")
            
            elective_subjects_map = {}  # Map course_code -> ElectiveSubject
            
            if chosen_elective_codes:
                # Load each elective subject (prefer department filter, fallback to any)  
                for code in chosen_elective_codes:
                    dept_id = elective_dept_map.get(code)
                    elec_subj = None
                    
                    # Try with department filter first (preferred)
                    if dept_id:
                        elec_subj = ElectiveSubject.objects.filter(
                            course_code=code,
                            department_id=dept_id
                        ).select_related('parent', 'department').first()
                    
                    # Fallback: try without department filter to ensure ALL chosen electives appear
                    if not elec_subj:
                        elec_subj = ElectiveSubject.objects.filter(
                            course_code=code
                        ).select_related('parent', 'department').first()
                    
                    if elec_subj:
                        elective_subjects_map[code] = elec_subj
                        dept_name = elec_subj.department.code if elec_subj.department else "No Dept"
                        print(f"[GetStudentSubjectsView]   - Elective subject: {elec_subj.course_code} - {elec_subj.course_name} (dept: {dept_name})")
                    else:
                        print(f"[GetStudentSubjectsView]   ⚠ WARNING: Elective subject not found in database for code: {code}")
                        print(f"[GetStudentSubjectsView]   ⚠ Student chose this elective but ElectiveSubject record doesn't exist!")
            
            print(f"[GetStudentSubjectsView] Loaded {len(elective_subjects_map)} elective subjects from {len(chosen_elective_codes)} chosen codes")
            if len(elective_subjects_map) < len(chosen_elective_codes):
                missing_codes = set(chosen_elective_codes) - set(elective_subjects_map.keys())
                print(f"[GetStudentSubjectsView] ⚠⚠⚠ MISSING {len(missing_codes)} elective records: {missing_codes}")
                print(f"[GetStudentSubjectsView] ⚠⚠⚠ These electives will NOT appear unless corrected!")
            
            # ========== STEP 5: Build pseudo-assignments for subjects without TAs ==========
            # CRITICAL: Show ALL subjects (core + elective) even if no TeachingAssignment exists
            print(f"[GetStudentSubjectsView] STEP 5: Creating pseudo-assignments for subjects without TAs")
            
            class PseudoAssignment:
                def __init__(self, subject_obj, subject_type='elective'):
                    """
                    Create a pseudo teaching assignment for subjects without real TAs.
                    
                    Args:
                        subject_obj: Either ElectiveSubject or CurriculumDepartment instance
                        subject_type: 'elective' or 'core'
                    """
                    # Use negative ID to avoid collision with real TeachingAssignment IDs
                    # This ensures teaching_assignment_id is an integer for frontend compatibility
                    self.id = -subject_obj.id
                    
                    if subject_type == 'elective':
                        self.elective_subject = subject_obj
                        self.curriculum_row = None
                    else:  # core
                        self.elective_subject = None
                        self.curriculum_row = subject_obj
                    
                    self.subject = None
                    self.custom_subject = None
                    self.staff = None
                    self.is_pseudo = True  # Flag to identify pseudo-assignments
            
            # Create pseudo-assignments for CORE subjects without TAs
            for curr_subj in subjects_without_ta:
                pseudo = PseudoAssignment(curr_subj, subject_type='core')
                core_teaching_assignments.append(pseudo)
                print(f"[GetStudentSubjectsView]   ⚠ Created pseudo for core subject: {curr_subj.course_code} - {curr_subj.course_name}")
            
            print(f"[GetStudentSubjectsView] Total core assignments (including pseudo): {len(core_teaching_assignments)}")
            
            # Iterate over ALL chosen elective codes (not just those in elective_subjects_map)
            for code in chosen_elective_codes:
                # Get the ElectiveSubject if it exists
                elec_subj = elective_subjects_map.get(code)
                
                if not elec_subj:
                    # ElectiveSubject record missing - can't create pseudo-assignment
                    print(f"[GetStudentSubjectsView]   ⚠⚠⚠ CRITICAL: Cannot create assignment for {code} - ElectiveSubject record missing!")
                    print(f"[GetStudentSubjectsView]   ⚠⚠⚠ This elective will NOT appear in feedback form!")
                    print(f"[GetStudentSubjectsView]   ⚠⚠⚠ Action needed: Verify ElectiveSubject exists for course_code='{code}'")
                    continue
                
                # Check if we have a teaching assignment for this code
                has_ta = any(
                    ta for ta in elective_teaching_assignments 
                    if (ta.elective_subject and ta.elective_subject.course_code == code) or
                       (ta.curriculum_row and ta.curriculum_row.course_code == code) or
                       (ta.subject and ta.subject.code == code)
                )
                
                if not has_ta:
                    # No teaching assignment found - create pseudo for elective
                    pseudo = PseudoAssignment(elec_subj, subject_type='elective')
                    elective_teaching_assignments.append(pseudo)
                    staff_map[code] = ["Staff Not Assigned"]
                    print(f"[GetStudentSubjectsView]   ⚠ No TA for elective {code}, created pseudo-assignment")
                elif code in staff_map and not staff_map[code]:
                    # TA exists but no valid staff names found - set fallback
                    staff_map[code] = ["Staff Not Assigned"]
                    print(f"[GetStudentSubjectsView]   ⚠ TA found for {code} but no valid staff names, using fallback")
                elif code not in staff_map:
                    # TA exists but not in staff_map (shouldn't happen, but handle it)
                    staff_map[code] = ["Staff Not Assigned"]
                    print(f"[GetStudentSubjectsView]   ⚠ TA found for {code} but not in staff_map, using fallback")
            
            # ========== STEP 6: Combine Core + Elective Lists ==========
            print(f"[GetStudentSubjectsView] STEP 6: Combining core and elective subjects")
            
            all_teaching_assignments = list(core_teaching_assignments) + elective_teaching_assignments
            
            print(f"[GetStudentSubjectsView] Total: {len(all_teaching_assignments)} subjects ({len(core_teaching_assignments)} core + {len(elective_teaching_assignments)} electives)")
            print(f"[GetStudentSubjectsView] Elective breakdown:")
            for i, ta in enumerate(elective_teaching_assignments, 1):
                if hasattr(ta, 'is_pseudo') and ta.is_pseudo:
                    code = ta.elective_subject.course_code if ta.elective_subject else 'N/A'
                    name = ta.elective_subject.course_name if ta.elective_subject else 'N/A'
                    print(f"[GetStudentSubjectsView]   {i}. {code} - {name} [PSEUDO]")
                elif ta.elective_subject:
                    print(f"[GetStudentSubjectsView]   {i}. {ta.elective_subject.course_code} - {ta.elective_subject.course_name} [TA]")
                elif ta.curriculum_row:
                    print(f"[GetStudentSubjectsView]   {i}. {ta.curriculum_row.course_code} - {ta.curriculum_row.course_name} [CURRICULUM]")
                elif ta.subject:
                    print(f"[GetStudentSubjectsView]   {i}. {ta.subject.code} - {ta.subject.name} [SUBJECT]")
            
            # ========== STEP 7: Build subject response list ==========
            print(f"[GetStudentSubjectsView] STEP 7: Building subject response list")
            
            subjects = []
            for assignment in all_teaching_assignments:
                # Get subject name and code from different possible sources
                subject_name = None
                subject_code = None
                is_elective = False
                
                # Priority 1: curriculum_row (most common for regular subjects)
                if assignment.curriculum_row:
                    subject_name = assignment.curriculum_row.course_name
                    subject_code = assignment.curriculum_row.course_code
                    is_elective = assignment.curriculum_row.is_elective
                # Priority 2: elective_subject (elective courses)
                elif assignment.elective_subject:
                    subject_name = assignment.elective_subject.course_name
                    subject_code = assignment.elective_subject.course_code
                    is_elective = True
                # Priority 3: subject (legacy subjects)
                elif assignment.subject:
                    subject_name = assignment.subject.name
                    subject_code = assignment.subject.code
                # Priority 4: custom_subject (special subjects like Sports, Yoga)
                elif assignment.custom_subject:
                    subject_name = assignment.get_custom_subject_display()
                    subject_code = assignment.custom_subject
                
                if not subject_name:
                    print(f"[GetStudentSubjectsView] Skipping assignment {assignment.id} - no subject name found")
                    continue
                
                # Get staff name
                staff_id = None
                
                if is_elective and subject_code and subject_code in staff_map and staff_map[subject_code]:
                    # For ELECTIVES: Use staff_map which may have multiple staff members
                    staff_list = [s for s in staff_map[subject_code] if s and s.strip()]  # Filter empty names
                    if staff_list:
                        staff_name = ", ".join(staff_list)
                        print(f"[GetStudentSubjectsView] Elective {subject_code}: {len(staff_list)} staff - {staff_name}")
                    else:
                        staff_name = "Staff Not Assigned"
                        print(f"[GetStudentSubjectsView] Elective {subject_code}: Empty staff list, showing 'Staff Not Assigned'")
                elif assignment.staff:
                    # For CORE subjects: Use the staff from teaching assignment
                    staff_name = assignment.staff.user.get_full_name() or assignment.staff.user.username
                    staff_id = assignment.staff.id
                    print(f"[GetStudentSubjectsView] Core subject {subject_code}: {staff_name}")
                else:
                    # No staff assigned
                    staff_name = "Staff Not Assigned"
                    print(f"[GetStudentSubjectsView] No staff for {subject_code}")
                
                # Check if student has already submitted feedback for this subject
                # For pseudo-assignments (electives without TAs), check using negative ID
                if hasattr(assignment, 'is_pseudo') and assignment.is_pseudo:
                    # Pseudo-assignments use negative IDs - feedback submissions for these
                    # will have NULL teaching_assignment, so we can't check completion reliably.
                    # For now, always show as not completed for pseudo-assignments
                    existing_responses = False
                else:
                    existing_responses = FeedbackResponse.objects.filter(
                        feedback_form=feedback_form,
                        user=request.user,
                        teaching_assignment=assignment
                    ).exists()
                
                subjects.append({
                    'teaching_assignment_id': assignment.id,
                    'subject_name': subject_name,
                    'subject_code': subject_code,
                    'staff_name': staff_name,
                    'staff_id': staff_id,
                    'is_completed': existing_responses,
                    'type': 'ELECTIVE' if is_elective else 'CORE'
                })
            
            # Calculate completion status
            total_subjects = len(subjects)
            completed_subjects = sum(1 for s in subjects if s['is_completed'])
            
            # Count by type for debugging
            core_count = sum(1 for s in subjects if s['type'] == 'CORE')
            elective_count = sum(1 for s in subjects if s['type'] == 'ELECTIVE')
            
            print(f"[GetStudentSubjectsView] ========== FINAL SUMMARY ==========")
            print(f"[GetStudentSubjectsView] Returning {total_subjects} total subjects:")
            print(f"[GetStudentSubjectsView]   - CORE: {core_count}")
            print(f"[GetStudentSubjectsView]   - ELECTIVE: {elective_count}")
            print(f"[GetStudentSubjectsView]   - Completed: {completed_subjects}")
            print(f"[GetStudentSubjectsView] Elective subjects returned:")
            for s in subjects:
                if s['type'] == 'ELECTIVE':
                    print(f"[GetStudentSubjectsView]   → {s['subject_code']}: {s['subject_name']} - {s['staff_name']}")
            
            if elective_count == 0:
                print(f"[GetStudentSubjectsView] ⚠⚠⚠ NO ELECTIVES IN RESPONSE!")
                print(f"[GetStudentSubjectsView] ⚠⚠⚠ Possible reasons:")
                print(f"[GetStudentSubjectsView] ⚠⚠⚠ 1. Student has no ElectiveChoice records (check earlier log)")
                print(f"[GetStudentSubjectsView] ⚠⚠⚠ 2. ElectiveChoice records exist but elective_subject is NULL")
                print(f"[GetStudentSubjectsView] ⚠⚠⚠ 3. ElectiveSubject records missing for chosen course codes")
                print(f"[GetStudentSubjectsView] ⚠⚠⚠ 4. All chosen electives were skipped in final loop")
                print(f"[GetStudentSubjectsView] ⚠⚠⚠ Run: python manage.py shell < check_elective_data.py")
            
            print(f"[GetStudentSubjectsView] ===================================")
            
            if total_subjects == 0:
                print(f"[GetStudentSubjectsView] WARNING: No subjects found!")
                print(f"  - Student section: {section.name} (ID: {section.id})")
                print(f"  - Student year: {student_year}")
                print(f"  - Student department: {student_department.code if student_department else 'None'}")
                print(f"  - Feedback form years: {feedback_form.years}")
                print(f"  - Feedback form sections: {feedback_form.sections}")
                print(f"  - Feedback form department: {feedback_form.department.code}")
            
            return Response({
                'feedback_form_id': form_id,
                'subjects': subjects,
                'total_subjects': total_subjects,
                'completed_subjects': completed_subjects,
                'all_completed': completed_subjects == total_subjects and total_subjects > 0
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.exception('[GetStudentSubjectsView] ERROR')
            return Response({
                'detail': 'Error fetching subjects. Please try again.',
                'subjects': [],
                'total_subjects': 0,
                'completed_subjects': 0,
                'all_completed': False
            }, status=status.HTTP_200_OK)


class DiagnosticElectiveChoicesView(APIView):
    """
    Diagnostic API: Check ElectiveChoice records for current user
    GET /api/feedback/diagnostic/elective-choices/
    
    Returns all ElectiveChoice records for the logged-in student to debug why electives aren't showing
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            from curriculum.models import ElectiveChoice
            from academics.models import StudentProfile, AcademicYear
            
            # Get student profile
            try:
                student_profile = StudentProfile.objects.select_related(
                    'current_section', 
                    'current_section__department'
                ).get(user=request.user)
            except StudentProfile.DoesNotExist:
                return Response({
                    'detail': 'Student profile not found',
                    'is_student': False
                }, status=status.HTTP_200_OK)
            
            # Get current academic year
            current_ay = AcademicYear.objects.filter(is_active=True).first()
            
            # Fetch all elective choices (with and without AY filter)
            all_choices = ElectiveChoice.objects.filter(
                student=student_profile
            ).select_related(
                'elective_subject', 
                'elective_subject__parent',
                'academic_year'
            ).order_by('-created_at')
            
            # Separate active and inactive
            active_choices = all_choices.filter(is_active=True)
            
            # Categorize by parent category
            categorized = {}
            for choice in active_choices:
                category = choice.elective_subject.parent.category if choice.elective_subject.parent else 'Unknown'
                if category not in categorized:
                    categorized[category] = []
                
                categorized[category].append({
                    'id': choice.id,
                    'course_code': choice.elective_subject.course_code,
                    'course_name': choice.elective_subject.course_name,
                    'semester': choice.elective_subject.semester,
                    'academic_year': choice.academic_year.name if choice.academic_year else None,
                    'is_active': choice.is_active,
                    'created_at': choice.created_at.isoformat(),
                    'matches_current_ay': choice.academic_year == current_ay if choice.academic_year else False
                })
            
            return Response({
                'student': {
                    'name': student_profile.user.get_full_name(),
                    'roll_number': student_profile.roll_number,
                    'section': student_profile.current_section.name if student_profile.current_section else None,
                    'year': student_profile.current_year
                },
                'current_academic_year': current_ay.name if current_ay else None,
                'total_choices': all_choices.count(),
                'active_choices': active_choices.count(),
                'categorized_choices': categorized,
                'debug_info': {
                    'note': 'If OE subjects not showing, check if ElectiveChoice records exist with is_active=True',
                    'expected_category': 'Open Elective (should contain "Open Elective" in parent.category)',
                }
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.exception('[DiagnosticElectiveChoicesView] ERROR')
            return Response({
                'detail': 'An error occurred. Please try again.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DeleteFeedbackFormView(APIView):
    """
    API 11: Delete Feedback Form (HOD)
    DELETE /api/feedback/<form_id>/delete/
    
    Allows HOD to permanently delete a deactivated feedback form.
    Only deactivated forms (active=False) can be deleted.
    """
    permission_classes = [IsAuthenticated]
    
    def delete(self, request, form_id):
        try:
            # Check if user has permission to delete feedback forms
            user_permissions = get_user_permissions(request.user)
            if 'feedback.create' not in user_permissions:
                return Response({
                    'detail': 'You do not have permission to delete feedback forms.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get the feedback form
            try:
                feedback_form = FeedbackForm.objects.get(id=form_id)
            except FeedbackForm.DoesNotExist:
                return Response({
                    'detail': 'Feedback form not found.'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Verify the user is the creator or has HOD access to the department
            from academics.models import StaffProfile, DepartmentRole, AcademicYear
            try:
                staff_profile = StaffProfile.objects.get(user=request.user)
            except StaffProfile.DoesNotExist:
                return Response({
                    'detail': 'Staff profile not found.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if user created this form OR is HOD of the department
            is_creator = feedback_form.created_by == request.user
            is_hod = False
            
            active_ay = AcademicYear.objects.filter(is_active=True).first()
            if active_ay:
                is_hod = DepartmentRole.objects.filter(
                    staff=staff_profile,
                    department=feedback_form.department,
                    role='HOD',
                    is_active=True,
                    academic_year=active_ay
                ).exists()
            
            if not (is_creator or is_hod):
                return Response({
                    'detail': 'You do not have permission to delete this feedback form.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Only allow deletion of deactivated forms
            if feedback_form.active:
                return Response({
                    'detail': 'Cannot delete active feedback forms. Please deactivate the form first.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Store form info for response
            form_type = feedback_form.get_type_display()
            
            # Delete the form (cascade will delete questions and responses)
            feedback_form.delete()
            
            return Response({
                'detail': f'{form_type} feedback form deleted successfully.',
                'deleted_form_id': form_id
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'detail': f'Error deleting feedback form: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetSubjectsByYearView(APIView):
    """
    API 10: Get Subjects by Year (Supports Multiple Years and Section Filtering)
    GET /api/feedback/subjects-by-year/?years=2,3&sections=1,2&department_id=1
    
    Returns list of subjects (teaching assignments) for given years, sections, and department.
    Used by HOD when creating subject feedback to preview what subjects will be included.
    Supports multiple years via comma-separated values or multiple year parameters.
    Optionally filters by specific section IDs.
    
    HOD View: Shows ALL subjects (core + ALL electives) with staff names from TeachingAssignment.
    This is different from student view which only shows electives the student has chosen.
    
    Parameters:
    - years: Comma-separated year numbers (e.g., "2,3,4")
    - sections: Comma-separated section IDs (optional, e.g., "1,2,3")
    - department_id: Department ID (optional)
    - semester: Semester ID (optional)
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            # Support both 'years' (comma-separated) and multiple 'year' parameters
            years_param = request.GET.get('years')
            semester_id = request.GET.get('semester')
            department_id = request.GET.get('department_id')
            sections_param = request.GET.get('sections')  # NEW: section IDs filter
            
            # Parse years
            years = []
            if years_param:
                # Comma-separated years: "2,3,4"
                try:
                    years = [int(y.strip()) for y in years_param.split(',') if y.strip()]
                except ValueError:
                    return Response({
                        'detail': 'Years must be comma-separated integers.'
                    }, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Fall back to single 'year' parameter for backward compatibility
                year_param = request.GET.get('year')
                if year_param:
                    try:
                        years = [int(year_param)]
                    except ValueError:
                        return Response({
                            'detail': 'Year must be a valid integer.'
                        }, status=status.HTTP_400_BAD_REQUEST)
            
            if not years:
                return Response({
                    'detail': 'Years parameter is required.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Parse section IDs if provided
            section_ids_filter = None
            if sections_param:
                try:
                    section_ids_filter = [int(s.strip()) for s in sections_param.split(',') if s.strip()]
                except ValueError:
                    return Response({
                        'detail': 'Sections must be comma-separated integers (section IDs).'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Import required models
            from academics.models import (
                TeachingAssignment, 
                AcademicYear, 
                Section,
                Semester,
                Department,
                DepartmentRole
            )
            from curriculum.models import ElectiveSubject, CurriculumDepartment
            
            # Determine active department for HOD
            if not department_id:
                try:
                    staff_profile = StaffProfile.objects.get(user=request.user)
                    active_ay = AcademicYear.objects.filter(is_active=True).first()
                    
                    if active_ay:
                        department_roles = DepartmentRole.objects.select_related('department').filter(
                            staff=staff_profile,
                            role='HOD',
                            is_active=True,
                            academic_year=active_ay
                        )
                        
                        departments_count = department_roles.count()
                        
                        if departments_count > 1:
                            # Multiple departments - use session
                            active_dept_id = request.session.get('active_hod_department_id')
                            if active_dept_id:
                                department_id = active_dept_id
                            else:
                                # Use first department
                                department_id = department_roles.first().department.id
                        elif departments_count == 1:
                            # Single department
                            department_id = department_roles.first().department.id
                        else:
                            # No department roles, fall back to staff profile
                            if staff_profile.department:
                                department_id = staff_profile.department.id
                    else:
                        # No active AY, fall back to staff profile
                        if staff_profile.department:
                            department_id = staff_profile.department.id
                except StaffProfile.DoesNotExist:
                    pass
            
            # Get active academic year
            current_ay = AcademicYear.objects.filter(is_active=True).first()
            if not current_ay:
                return Response({
                    'detail': 'No active academic year found.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Calculate batch start years for all given years
            try:
                acad_start = int(str(current_ay.name).split('-')[0])
                batch_start_years = [acad_start - year + 1 for year in years]
                print(f"[GetSubjectsByYearView] Academic year: {current_ay.name}, acad_start: {acad_start}")
                print(f"[GetSubjectsByYearView] Requested years: {years}, calculated batch_start_years: {batch_start_years}")
            except Exception as e:
                print(f"[GetSubjectsByYearView] Error calculating batch years: {str(e)}")
                return Response({
                    'detail': 'Error calculating batch years.'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # Build filters for sections (multiple batch start years)
            section_filters = Q(
                batch__start_year__in=batch_start_years
            )
            
            # Filter by department if provided
            if department_id:
                try:
                    department_id = int(department_id)
                    section_filters &= Q(batch__course__department_id=department_id)
                    print(f"[GetSubjectsByYearView] Filtering by department_id: {department_id}")
                except ValueError:
                    pass
            
            # Filter by semester if provided
            if semester_id:
                try:
                    semester_id = int(semester_id)
                    section_filters &= Q(semester_id=semester_id)
                    print(f"[GetSubjectsByYearView] Filtering by semester_id: {semester_id}")
                except ValueError:
                    pass
            
            # Filter by specific section IDs if provided
            if section_ids_filter:
                section_filters &= Q(id__in=section_ids_filter)
                print(f"[GetSubjectsByYearView] Filtering by section IDs: {section_ids_filter}")
            
            # Get matching sections
            sections = Section.objects.filter(section_filters).distinct()
            print(f"[GetSubjectsByYearView] Found {sections.count()} sections")
            
            # Create a mapping of section_id to year for display
            section_to_year = {}
            section_ids = []
            
            if sections.exists():
                for section in sections:
                    if section.batch and section.batch.start_year:
                        # Calculate which year this section belongs to
                        for year in years:
                            expected_start = acad_start - year + 1
                            if section.batch.start_year == expected_start:
                                section_to_year[section.id] = year
                                break
                
                section_ids = [s.id for s in sections]
            else:
                # Debug: Check what sections exist for these batch years without department filter
                all_sections = Section.objects.filter(batch__start_year__in=batch_start_years).distinct()
                print(f"[GetSubjectsByYearView] DEBUG: Total sections with these batch years (no dept filter): {all_sections.count()}")
                if all_sections.exists():
                    for sec in all_sections[:5]:
                        print(f"  - Section: {sec.name}, Batch: {sec.batch.name if sec.batch else 'None'}, Dept: {sec.batch.course.department.code if sec.batch and sec.batch.course else 'None'}")
                
                print(f"[GetSubjectsByYearView] No sections found, but will still fetch elective subjects from curriculum")
            
            # Calculate semesters for the selected years based on AY parity
            semesters_for_years = []
            if current_ay and current_ay.parity:
                is_odd_semester = current_ay.parity.upper() == 'ODD'
                print(f"[GetSubjectsByYearView] Academic year parity: {current_ay.parity}")
                
                for year in years:
                    # Year 1 → Sem 1 or 2, Year 2 → Sem 3 or 4, Year 3 → Sem 5 or 6, Year 4 → Sem 7 or 8
                    base_semester = (year - 1) * 2 + 1  # Odd semester for this year
                    current_semester_num = base_semester if is_odd_semester else base_semester + 1
                    
                    # Get the Semester model instance with this number
                    try:
                        semester_obj = Semester.objects.get(number=current_semester_num)
                        semesters_for_years.append(semester_obj.id)
                        print(f"[GetSubjectsByYearView] Year {year} → Semester {current_semester_num} (ID: {semester_obj.id})")
                    except Semester.DoesNotExist:
                        print(f"[GetSubjectsByYearView] WARNING: Semester {current_semester_num} not found for year {year}")
            
            print(f"[GetSubjectsByYearView] Semesters for years {years}: {semesters_for_years}")
            
            # Debug: Show which semester numbers we're searching for
            if semesters_for_years:
                semester_numbers = []
                for sem_id in semesters_for_years:
                    try:
                        sem_obj = Semester.objects.get(id=sem_id)
                        semester_numbers.append(sem_obj.number)
                    except:
                        pass
                print(f"[GetSubjectsByYearView] Semester numbers: {semester_numbers}")
            
            print(f"[GetSubjectsByYearView] Fetching teaching assignments for {len(section_ids)} sections")
            
            # Build subject list with unique subjects
            subjects_dict = {}
            elective_groups = {}  # Group electives by parent category
            
            # Fetch teaching assignments for these sections (if sections exist)
            if section_ids:
                teaching_assignments = TeachingAssignment.objects.filter(
                    section_id__in=section_ids,
                    academic_year=current_ay,
                    is_active=True
                ).select_related(
                    'staff', 
                    'staff__user', 
                    'subject', 
                    'curriculum_row',
                    'elective_subject',
                    'elective_subject__parent',  # Load parent for elective grouping
                    'elective_subject__department',  # Load department for OE subjects
                    'section'
                ).distinct()
                
                print(f"[GetSubjectsByYearView] Found {teaching_assignments.count()} teaching assignments")
            else:
                teaching_assignments = []
                print(f"[GetSubjectsByYearView] No sections found, skipping teaching assignments")
            
            for assignment in teaching_assignments:
                # Get subject name and code
                subject_name = None
                subject_code = None
                subject_key = None
                is_elective = False
                elective_category = None
                department_name = None
                department_code = None
                
                if assignment.elective_subject:
                    # This is an elective subject
                    is_elective = True
                    subject_name = assignment.elective_subject.course_name
                    subject_code = assignment.elective_subject.course_code
                    subject_key = f"elec_{assignment.elective_subject.id}"
                    
                    # Get department info for elective subjects (important for OE)
                    if assignment.elective_subject.department:
                        department_name = assignment.elective_subject.department.name
                        department_code = assignment.elective_subject.department.code
                    
                    # Get parent category for grouping
                    if assignment.elective_subject.parent:
                        elective_category = assignment.elective_subject.parent.category or 'Other Electives'
                    else:
                        elective_category = 'Other Electives'
                        
                elif assignment.curriculum_row:
                    subject_name = assignment.curriculum_row.course_name
                    subject_code = assignment.curriculum_row.course_code
                    subject_key = f"curr_{assignment.curriculum_row.id}"
                    is_elective = assignment.curriculum_row.is_elective
                    if is_elective:
                        elective_category = assignment.curriculum_row.category or 'Other Electives'
                        
                elif assignment.subject:
                    subject_name = assignment.subject.name
                    subject_code = assignment.subject.code
                    subject_key = f"subj_{assignment.subject.id}"
                    
                elif assignment.custom_subject:
                    subject_name = assignment.get_custom_subject_display()
                    subject_code = assignment.custom_subject
                    subject_key = f"cust_{assignment.custom_subject}"
                
                if not subject_name or not subject_key:
                    continue
                
                # Get staff name
                staff_name = assignment.staff.user.get_full_name() or assignment.staff.user.username
                section_name = assignment.section.name if assignment.section else 'N/A'
                
                # Get year for this assignment
                year_for_display = section_to_year.get(assignment.section.id) if assignment.section else None
                
                # Build subject data
                subject_data = {
                    'subject_name': subject_name,
                    'subject_code': subject_code,
                    'staff': set(),
                    'sections': set(),
                    'years': set(),
                    'teaching_assignment_ids': [],
                    'is_elective': is_elective,
                    'elective_category': elective_category,
                    'department_name': department_name,
                    'department_code': department_code
                }
                
                # Track unique subjects and their staff
                if subject_key not in subjects_dict:
                    subjects_dict[subject_key] = subject_data
                
                subjects_dict[subject_key]['staff'].add(staff_name)
                subjects_dict[subject_key]['sections'].add(section_name)
                if year_for_display:
                    subjects_dict[subject_key]['years'].add(year_for_display)
                subjects_dict[subject_key]['teaching_assignment_ids'].append(assignment.id)
                
                # Group electives by category
                if is_elective and elective_category:
                    if elective_category not in elective_groups:
                        elective_groups[elective_category] = []
                    if subject_key not in [s['key'] for s in elective_groups[elective_category]]:
                        elective_groups[elective_category].append({
                            'key': subject_key,
                            'subject_name': subject_name,
                            'subject_code': subject_code
                        })
            
            # FETCH ELECTIVE SUBJECTS WITH STAFF NAMES FROM TEACHING ASSIGNMENTS
            # HOD view should show actual staff names for all electives (not "Multiple Staff")
            # 
            # IMPORTANT: Open Electives (OE) are cross-department subjects stored in ElectiveSubject table
            # They may have placeholder codes (like XXC13XX) in DepartmentCurricula but actual data is here
            # Staff mapping must check THREE methods since TeachingAssignments may reference them via:
            #   1. elective_subject (ForeignKey to ElectiveSubject)
            #   2. curriculum_row.course_code (ForeignKey to CurriculumDepartment)
            #   3. subject.code (ForeignKey to legacy Subject table)
            # 
            # HOD must see ALL OE subjects for current semester across ALL departments (no dept filtering)
            # Students see only their chosen OE via ElectiveChoice (handled separately)
            print(f"[GetSubjectsByYearView] Fetching elective subjects from curriculum and matching with teaching assignments...")
            
            if department_id and semesters_for_years:
                # Fetch PE/EE elective subjects for the department only
                dept_electives = ElectiveSubject.objects.filter(
                    department_id=department_id,
                    semester_id__in=semesters_for_years,
                    approval_status='APPROVED'
                ).select_related('parent', 'semester', 'department').distinct()
                
                print(f"[GetSubjectsByYearView] Found {dept_electives.count()} department elective subjects")
                
                # Fetch ALL Open Elective subjects across departments for the semesters
                # OE subjects are cross-department - HOD should see ALL options regardless of department
                # DO NOT filter by department_id for Open Electives
                # Match various category patterns: "Open Elective", "OE", "Open Elective I", etc.
                # 
                # Use TWO approaches to maximize OE subject discovery:
                # 1. Filter by calculated semester IDs (primary method)
                # 2. Also include subjects where semester number matches year range (fallback)
                oe_query = Q(parent__category__icontains='Open Elective') | Q(parent__category__istartswith='OE')
                oe_query = oe_query & Q(approval_status='APPROVED')
                
                # Primary: Use calculated semester IDs
                if semesters_for_years:
                    oe_electives = ElectiveSubject.objects.filter(
                        oe_query,
                        semester_id__in=semesters_for_years
                    ).select_related('parent', 'semester', 'department').distinct()
                else:
                    # Fallback: If no semester IDs, try to match by semester number range
                    # Year 2 = Sem 3-4, Year 3 = Sem 5-6, Year 4 = Sem 7-8
                    semester_numbers = []
                    for year in years:
                        semester_numbers.extend([(year - 1) * 2 + 1, (year - 1) * 2 + 2])
                    
                    oe_electives = ElectiveSubject.objects.filter(
                        oe_query,
                        semester__number__in=semester_numbers
                    ).select_related('parent', 'semester', 'department').distinct()
                
                print(f"[GetSubjectsByYearView] Found {oe_electives.count()} Open Elective subjects across all departments")
                if oe_electives.exists():
                    for oe in oe_electives[:5]:  # Show first 5 for debugging
                        print(f"  - OE: {oe.course_code} - {oe.course_name} ({oe.department.code if oe.department else 'No Dept'}) - Category: {oe.parent.category if oe.parent else 'No Parent'}")
                
                # Combine both querysets
                curriculum_electives = list(dept_electives) + [oe for oe in oe_electives if oe not in dept_electives]
                
                print(f"[GetSubjectsByYearView] Total elective subjects to process: {len(curriculum_electives)}")
                print(f"[GetSubjectsByYearView] Breakdown: {dept_electives.count()} dept electives + {len([oe for oe in oe_electives if oe not in dept_electives])} unique OE subjects")
                
                # Build a mapping of elective_subject_id to teaching assignments with staff
                # This ensures ALL electives (OE, PE, EE) show correct staff names
                # 
                # THREE MAPPING METHODS (try all to maximize staff resolution):
                # 1. By elective_subject_id: Direct link in TeachingAssignment.elective_subject
                # 2. By course_code via curriculum_row: TeachingAssignment.curriculum_row.course_code
                #    (needed for OE subjects with placeholder slots like XXC13XX)
                # 3. By subject code: TeachingAssignment.subject.code
                #    (legacy subjects table)
                all_elective_ids = [e.id for e in curriculum_electives]
                all_elective_codes = [e.course_code for e in curriculum_electives if e.course_code]
                
                elective_teaching_assignments = {}
                elective_teaching_by_code = {}  # Fallback mapping by course code
                
                if all_elective_ids:
                    # Fetch by elective_subject_id
                    elective_tas = TeachingAssignment.objects.filter(
                        elective_subject_id__in=all_elective_ids,
                        academic_year=current_ay,
                        is_active=True
                    ).select_related('staff', 'staff__user', 'elective_subject')
                    
                    for ta in elective_tas:
                        if ta.elective_subject_id not in elective_teaching_assignments:
                            elective_teaching_assignments[ta.elective_subject_id] = []
                        elective_teaching_assignments[ta.elective_subject_id].append(ta)
                    
                    print(f"[GetSubjectsByYearView] Found teaching assignments for {len(elective_teaching_assignments)} elective subjects (by ID)")
                
                # Also fetch by course_code (for OE subjects that might be mapped via curriculum_row)
                if all_elective_codes:
                    code_tas = TeachingAssignment.objects.filter(
                        curriculum_row__course_code__in=all_elective_codes,
                        curriculum_row__is_elective=True,
                        academic_year=current_ay,
                        is_active=True
                    ).select_related('staff', 'staff__user', 'curriculum_row')
                    
                    for ta in code_tas:
                        if ta.curriculum_row and ta.curriculum_row.course_code:
                            code = ta.curriculum_row.course_code
                            if code not in elective_teaching_by_code:
                                elective_teaching_by_code[code] = []
                            elective_teaching_by_code[code].append(ta)
                    
                    print(f"[GetSubjectsByYearView] Found teaching assignments for {len(elective_teaching_by_code)} elective subjects (by course code via curriculum_row)")
                    
                    # Also try mapping via legacy Subject table (subject.code)
                    subject_tas = TeachingAssignment.objects.filter(
                        subject__code__in=all_elective_codes,
                        academic_year=current_ay,
                        is_active=True
                    ).select_related('staff', 'staff__user', 'subject')
                    
                    for ta in subject_tas:
                        if ta.subject and ta.subject.code:
                            code = ta.subject.code
                            if code not in elective_teaching_by_code:
                                elective_teaching_by_code[code] = []
                            elective_teaching_by_code[code].append(ta)
                    
                    print(f"[GetSubjectsByYearView] Total teaching assignments by code (after adding subject.code): {len(elective_teaching_by_code)}")
                
                # Process each curriculum elective
                processed_count = 0
                merged_count = 0
                for elective in curriculum_electives:
                    subject_key = f"elec_{elective.id}"
                    
                    # Check if already added from teaching assignment section
                    already_exists = subject_key in subjects_dict
                    if already_exists:
                        merged_count += 1
                        print(f"[GetSubjectsByYearView] Merging data for existing elective: {elective.course_name}")
                    else:
                        processed_count += 1
                    
                    # Get parent category
                    elective_category = 'Other Electives'
                    if elective.parent:
                        elective_category = elective.parent.category or 'Other Electives'
                    
                    # Calculate which year this elective belongs to
                    elective_year = None
                    if elective.semester:
                        elective_year = ((elective.semester.number + 1) // 2)
                        if elective_year in years:
                            print(f"[GetSubjectsByYearView] Processing elective: {elective.course_name} (Category: {elective_category}, Year {elective_year})")
                    
                    # Check if this is an Open Elective
                    # Match various OE category patterns: "Open Elective", "OE", "OE I", "OE II", etc.
                    is_open_elective = (
                        'Open Elective' in elective_category or 
                        'OE' in elective_category or 
                        elective_category.startswith('OE')
                    )
                    
                    # Get actual staff names from teaching assignments (for ALL electives)
                    # Try three methods: by ID, by course code (curriculum_row), by subject code
                    staff_names = set()
                    
                    # Method 1: Fetch by elective_subject_id
                    if elective.id in elective_teaching_assignments:
                        for ta in elective_teaching_assignments[elective.id]:
                            if ta.staff:
                                staff_name = ta.staff.user.get_full_name() or ta.staff.user.username
                                staff_names.add(staff_name)
                        print(f"[GetSubjectsByYearView] Elective {elective.course_name} has staff (by ID): {staff_names}")
                    
                    # Methods 2 & 3: Fetch by course_code (curriculum_row or subject)
                    if not staff_names and elective.course_code and elective.course_code in elective_teaching_by_code:
                        for ta in elective_teaching_by_code[elective.course_code]:
                            if ta.staff:
                                staff_name = ta.staff.user.get_full_name() or ta.staff.user.username
                                staff_names.add(staff_name)
                        print(f"[GetSubjectsByYearView] Elective {elective.course_name} has staff (by code): {staff_names}")
                    
                    # If no staff found, use placeholder
                    if not staff_names:
                        staff_names = set(['To be assigned'])
                        print(f"[GetSubjectsByYearView] No staff assigned for elective: {elective.course_name} - using placeholder")
                    
                    # Get department information (important for OE subjects offered by different departments)
                    department_name = None
                    department_code = None
                    if elective.department:
                        department_name = elective.department.name
                        department_code = elective.department.code
                    
                    # Collect all teaching assignment IDs (from both mappings)
                    ta_ids = []
                    ta_ids.extend([ta.id for ta in elective_teaching_assignments.get(elective.id, [])])
                    if elective.course_code and elective.course_code in elective_teaching_by_code:
                        ta_ids.extend([ta.id for ta in elective_teaching_by_code[elective.course_code]])
                    # Remove duplicates
                    ta_ids = list(set(ta_ids))
                    
                    # Add or merge data in subjects_dict
                    if already_exists:
                        # MERGE: Update existing entry with additional data
                        print(f"[GetSubjectsByYearView] Merging additional data for: {elective.course_name}")
                        subjects_dict[subject_key]['staff'].update(staff_names)
                        subjects_dict[subject_key]['sections'].update(set(['All Sections']))
                        if elective_year:
                            subjects_dict[subject_key]['years'].add(elective_year)
                        # Merge teaching assignment IDs (avoid duplicates)
                        existing_ids = set(subjects_dict[subject_key]['teaching_assignment_ids'])
                        new_ids = set(ta_ids)
                        subjects_dict[subject_key]['teaching_assignment_ids'] = list(existing_ids | new_ids)
                        print(f"[GetSubjectsByYearView]   Merged staff: {subjects_dict[subject_key]['staff']}")
                        print(f"[GetSubjectsByYearView]   Total TAs: {len(subjects_dict[subject_key]['teaching_assignment_ids'])}")
                    else:
                        # ADD: Create new entry
                        subject_data = {
                            'subject_name': elective.course_name or 'Unknown Elective',
                            'subject_code': elective.course_code or '',
                            'staff': staff_names,
                            'sections': set(['All Sections']),
                            'years': set([elective_year]) if elective_year else set(),
                            'teaching_assignment_ids': ta_ids,
                            'is_elective': True,
                            'elective_category': elective_category,
                            'department_name': department_name,
                            'department_code': department_code
                        }
                        
                        subjects_dict[subject_key] = subject_data
                    
                    # Group electives by category
                    if elective_category not in elective_groups:
                        elective_groups[elective_category] = []
                    if subject_key not in [s['key'] for s in elective_groups[elective_category]]:
                        elective_groups[elective_category].append({
                            'key': subject_key,
                            'subject_name': elective.course_name or 'Unknown Elective',
                            'subject_code': elective.course_code or ''
                        })
                
                print(f"[GetSubjectsByYearView] Elective processing summary: {processed_count} new added, {merged_count} merged with existing data")
                print(f"[GetSubjectsByYearView] Elective groups created: {list(elective_groups.keys())}")
                for cat, items in elective_groups.items():
                    print(f"  - {cat}: {len(items)} subjects")
                
                # Also fetch core subjects from curriculum if no teaching assignments found
                if not section_ids or len(subjects_dict) == 0:
                    print(f"[GetSubjectsByYearView] Fetching core subjects from curriculum department rows...")
                    curriculum_rows = CurriculumDepartment.objects.filter(
                        department_id=department_id,
                        semester_id__in=semesters_for_years,
                        is_elective=False,
                        approval_status='APPROVED'
                    ).select_related('semester').distinct()
                    
                    print(f"[GetSubjectsByYearView] Found {curriculum_rows.count()} core subjects in curriculum")
                    
                    for row in curriculum_rows:
                        subject_key = f"curr_{row.id}"
                        
                        # Skip if already added from teaching assignment
                        if subject_key in subjects_dict:
                            continue
                        
                        # Calculate which year this subject belongs to
                        subject_year = None
                        if row.semester:
                            subject_year = ((row.semester.number + 1) // 2)
                            if subject_year in years:
                                print(f"[GetSubjectsByYearView] Adding core subject: {row.course_name} (Sem {row.semester.number}, Year {subject_year})")
                        
                        # Add to subjects_dict
                        subject_data = {
                            'subject_name': row.course_name or 'Unknown Subject',
                            'subject_code': row.course_code or '',
                            'staff': set(['To be assigned']),
                            'sections': set(['All Sections']),
                            'years': set([subject_year]) if subject_year else set(),
                            'teaching_assignment_ids': [],
                            'is_elective': False,
                            'elective_category': None,
                            'department_name': None,
                            'department_code': None
                        }
                        
                        subjects_dict[subject_key] = subject_data
            
            print(f"[GetSubjectsByYearView] Total subjects after adding curriculum subjects: {len(subjects_dict)}")
            
            # Format response - separate regular subjects and electives
            regular_subjects = []
            elective_subjects = []
            
            # For HOD creation view: Show only elective category headings (not individual subjects)
            # Individual subjects will be shown to students and in response view
            elective_categories_only = []
            
            for subject_key, data in subjects_dict.items():
                subject_info = {
                    'subject_name': data['subject_name'],
                    'subject_code': data['subject_code'],
                    'staff_names': ', '.join(sorted(data['staff'])),
                    'sections': ', '.join(sorted(data['sections'])),
                    'years': sorted(list(data['years'])),  # List of years this subject appears in
                    'teaching_assignment_ids': data['teaching_assignment_ids'],
                    'assignment_count': len(data['teaching_assignment_ids']),
                    'is_elective': data.get('is_elective', False),
                    'elective_category': data.get('elective_category'),
                    'department_name': data.get('department_name'),
                    'department_code': data.get('department_code')
                }
                
                if data.get('is_elective'):
                    elective_subjects.append(subject_info)
                else:
                    regular_subjects.append(subject_info)
            
            # Sort by subject name
            regular_subjects.sort(key=lambda x: x['subject_name'])
            elective_subjects.sort(key=lambda x: (x.get('elective_category', ''), x['subject_name']))
            
            # Format elective CATEGORIES for HOD view (not individual subjects)
            # HOD sees: "Professional Elective IV", "Emerging Elective I", etc.
            formatted_elective_categories = []
            for category in sorted(elective_groups.keys()):
                # Calculate total subjects in this category
                category_count = len(elective_groups[category])
                
                # Get all years for this category
                category_years = set()
                for subj_key_info in elective_groups[category]:
                    subject_key = subj_key_info['key']
                    if subject_key in subjects_dict:
                        category_years.update(subjects_dict[subject_key]['years'])
                
                formatted_elective_categories.append({
                    'category': category,
                    'count': category_count,
                    'years': sorted(list(category_years)),
                    'display_name': category  # e.g., "Professional Elective IV"
                })
            
            # Format elective groups WITH individual subjects (for student/response views)
            formatted_elective_groups = []
            for category, subjects in sorted(elective_groups.items()):
                category_subjects = []
                for subj_key_info in subjects:
                    subject_key = subj_key_info['key']
                    if subject_key in subjects_dict:
                        data = subjects_dict[subject_key]
                        category_subjects.append({
                            'subject_name': data['subject_name'],
                            'subject_code': data['subject_code'],
                            'staff_names': ', '.join(sorted(data['staff'])),
                            'sections': ', '.join(sorted(data['sections'])),
                            'years': sorted(list(data['years'])),
                            'teaching_assignment_ids': data['teaching_assignment_ids'],
                            'assignment_count': len(data['teaching_assignment_ids'])
                        })
                
                # Sort subjects within category
                category_subjects.sort(key=lambda x: x['subject_name'])
                
                formatted_elective_groups.append({
                    'category': category,
                    'subjects': category_subjects,
                    'count': len(category_subjects)
                })
            
            return Response({
                'regular_subjects': regular_subjects,
                'elective_subjects': elective_subjects,
                'elective_categories': formatted_elective_categories,
                'elective_groups': formatted_elective_groups,
                'total_subjects': len(subjects_dict),
                'has_electives': len(formatted_elective_categories) > 0,
                'success': True
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            print(f"[GetSubjectsByYearView] ERROR: {str(e)}")
            print(traceback.format_exc())
            return Response({
                'detail': f'Error fetching subjects: {str(e)}',
                'regular_subjects': [],
                'elective_subjects': [],
                'elective_categories': [],
                'elective_groups': [],
                'total_subjects': 0,
                'success': False
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================================
# IQAC Export Filter APIs
# ============================================================================

class IQACExportOptionsView(APIView):
    """
    API: Get IQAC Common Export Filter Options
    GET /api/feedback/common-export/options/
    
    Returns ALL departments and data for IQAC export filter dropdown.
    Used by IQAC users to filter feedback export by department.
    
    IQAC has no department restrictions - returns all active departments.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            scope = get_feedback_department_scope(request.user)
            if not scope.get('allowed'):
                return Response({
                    'detail': 'You do not have permission to view all departments.'
                }, status=status.HTTP_403_FORBIDDEN)

            from academics.models import Department

            departments_qs = Department.objects.all().order_by('name')
            if not scope.get('all_departments'):
                department_ids = scope.get('department_ids') or []
                departments_qs = departments_qs.filter(id__in=department_ids)

            departments = departments_qs.values('id', 'name', 'code', 'short_name')
            
            return Response({
                'departments': list(departments),
                'all_departments_access': bool(scope.get('all_departments')),
                'success': True
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'detail': f'Error fetching export options: {str(e)}',
                'departments': [],
                'success': False
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class IQACExportYearsView(APIView):
    """
    API: Get IQAC Export Years List
    GET /api/feedback/export-years/
    
    Returns all available academic years (1-4) for IQAC export filter.
    Used by IQAC users to filter feedback export by year.
    
    Returns: [1, 2, 3, 4] or from database if available.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            scope = get_feedback_department_scope(request.user)
            if not scope.get('allowed'):
                return Response({
                    'detail': 'You do not have permission to view export years.'
                }, status=status.HTTP_403_FORBIDDEN)

            from academics.models import Section
            
            # Try to get distinct years from active sections
            section_qs = Section.objects.filter(batch__start_year__isnull=False)
            if not scope.get('all_departments'):
                department_ids = scope.get('department_ids') or []
                section_qs = section_qs.filter(batch__course__department_id__in=department_ids)

            years = section_qs.values_list('batch__start_year', flat=True).distinct()
            
            # Calculate academic years from batch start years
            from academics.models import AcademicYear
            current_ay = AcademicYear.objects.filter(is_active=True).first()
            
            years_list = []
            if current_ay:
                try:
                    acad_start = int(str(current_ay.name).split('-')[0])
                    for start_year in years:
                        if start_year:
                            year = acad_start - int(start_year) + 1
                            if 1 <= year <= 4:
                                years_list.append(year)
                except:
                    pass
            
            # Remove duplicates and sort
            years_list = sorted(list(set(years_list)))
            
            # Fallback to default years if none found
            if not years_list:
                years_list = [1, 2, 3, 4]
            
            return Response({
                'years': years_list,
                'success': True
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            # Fallback on error
            return Response({
                'years': [1, 2, 3, 4],
                'success': True
            }, status=status.HTTP_200_OK)


class IQACCommonExportView(APIView):
    """
    API: IQAC Common Export (Download Feedback Responses)
    POST /api/feedback/common-export/
    
    Allows IQAC users to export feedback responses with filters.
    Returns Excel file with feedback data filtered by departments and years.
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            scope = get_feedback_department_scope(request.user)
            if not scope.get('allowed'):
                return Response({
                    'detail': 'You do not have permission to export feedback.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get filter parameters
            all_departments = request.data.get('all_departments', False)
            department_ids = request.data.get('department_ids', [])
            years = request.data.get('years', [])
            
            # Build query
            from academics.models import AcademicYear, StudentProfile
            
            # Start with all responses with proper select_related for all relations
            qs = FeedbackResponse.objects.select_related(
                'feedback_form',
                'feedback_form__department',
                'question',
                'user',
                'teaching_assignment',
                'teaching_assignment__subject',
                'teaching_assignment__curriculum_row',
                'teaching_assignment__elective_subject',
                'teaching_assignment__staff',
                'teaching_assignment__staff__user',
                'teaching_assignment__section',
                'teaching_assignment__section__batch'
            ).prefetch_related(
                'user__student_profile'
            ).all()

            qs = apply_department_scope_filter(qs, scope, field_name='feedback_form__department_id')

            if not scope.get('all_departments'):
                # Own-department users are already scoped by server-side filter.
                all_departments = True
            
            # Filter by department if specified
            if not all_departments and department_ids:
                qs = qs.filter(feedback_form__department_id__in=department_ids)
            
            # Filter by year if specified
            if years:
                # Year filter: get students in those years
                # For forms targeting specific years
                year_filter = Q()
                for year in years:
                    year_filter |= Q(feedback_form__years__contains=[year])
                    year_filter |= Q(feedback_form__year=year)
                qs = qs.filter(year_filter)
            
            # Get data for Excel with strict column structure
            responses_data = []
            current_ay = AcademicYear.objects.filter(is_active=True).first()
            current_acad_year = None
            if current_ay:
                try:
                    current_acad_year = int(str(current_ay.name).split('-')[0])
                except Exception:
                    pass
            
            for response in qs:
                student_name = ""
                register_number = ""
                department_name = ""
                year_section = ""
                subject_code = ""
                subject_name = ""
                staff_name = ""
                comment_value = ""
                overall_comment_value = ""
                
                # Get student info
                if response.user:
                    student_name = response.user.get_full_name() or response.user.username
                    try:
                        student_profile = response.user.student_profile
                        register_number = student_profile.reg_no or ""
                        
                        # Get department with fallback chain
                        if student_profile.home_department:
                            department_name = student_profile.home_department.name or ""
                        elif student_profile.section and student_profile.section.batch and student_profile.section.batch.course:
                            department_name = student_profile.section.batch.course.department.name or ""
                        
                        # Calculate year and get section
                        if student_profile.section and current_acad_year:
                            section_name = student_profile.section.name or ""
                            batch = student_profile.section.batch
                            if batch and batch.start_year:
                                try:
                                    calculated_year = current_acad_year - int(batch.start_year) + 1
                                    year_section = f"{calculated_year} / {section_name}"
                                except:
                                    year_section = f"/ {section_name}"
                    except (AttributeError, StudentProfile.DoesNotExist):
                        pass
                
                # Fallback department from teaching assignment section for cases
                # where student profile department is unavailable.
                if not department_name and response.teaching_assignment and response.teaching_assignment.section:
                    ta_section = response.teaching_assignment.section
                    ta_batch = getattr(ta_section, 'batch', None)
                    ta_course = getattr(ta_batch, 'course', None) if ta_batch else None
                    ta_department = getattr(ta_course, 'department', None) if ta_course else None
                    if ta_department:
                        department_name = ta_department.name or ""

                # Fallback department from the form this response belongs to
                if not department_name and response.feedback_form and response.feedback_form.department:
                    department_name = response.feedback_form.department.name or ""
                
                # Get subject and staff from teaching assignment with multi-source fallback
                if response.teaching_assignment:
                    ta = response.teaching_assignment
                    
                    # Extract subject code and name from multiple sources
                    if ta.curriculum_row:
                        subject_code = ta.curriculum_row.course_code or ""
                        subject_name = ta.curriculum_row.course_name or ""
                    elif ta.elective_subject:
                        subject_code = ta.elective_subject.course_code or ""
                        subject_name = ta.elective_subject.course_name or ""
                    elif ta.subject:
                        subject_code = ta.subject.code or ""
                        subject_name = ta.subject.name or ""
                    elif ta.custom_subject:
                        subject_code = ta.custom_subject
                        subject_name = dict(ta._meta.get_field('custom_subject').choices).get(ta.custom_subject, ta.custom_subject)
                    
                    # Extract staff name
                    if ta.staff and ta.staff.user:
                        staff_name = ta.staff.user.get_full_name() or ta.staff.user.username or ""

                    # Fallback year/section from teaching assignment section.
                    if not year_section and ta.section:
                        ta_section_name = ta.section.name or ""
                        ta_batch = getattr(ta.section, 'batch', None)
                        ta_year_text = ""
                        if ta_batch and getattr(ta_batch, 'start_year', None) and current_acad_year:
                            try:
                                ta_year_text = str(current_acad_year - int(ta_batch.start_year) + 1)
                            except Exception:
                                ta_year_text = ""
                        year_section = f"{ta_year_text} / {ta_section_name}" if (ta_year_text or ta_section_name) else ""
                
                # Get question text
                question_text = response.question.question if response.question else ""
                
                # Apply conditional display logic: show either question-wise comment
                # or overall comment, never both.
                question_comment = (response.answer_text or "").strip()
                common_comment = (response.common_comment or "").strip()
                if question_comment:
                    comment_value = question_comment
                    overall_comment_value = ""
                elif common_comment:
                    comment_value = ""
                    overall_comment_value = common_comment
                else:
                    comment_value = ""
                    overall_comment_value = ""

                # Show selected option only when real value is available.
                selected_option_value = (response.selected_option_text or "").strip()
                
                # Collect data row
                responses_data.append({
                    'student_name': student_name,
                    'register_number': register_number,
                    'department': department_name,
                    'year_section': year_section,
                    'subject_code': subject_code,
                    'subject_name': subject_name,
                    'staff_name': staff_name,
                    'question_text': question_text,
                    'rating_value': response.answer_star or "",
                    'comment': comment_value,
                    'overall_comment': overall_comment_value,
                    'selected_option': selected_option_value,
                })
            
            # Generate Excel file
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Feedback Responses'
            
            # Build optional columns dynamically based on actual feedback content.
            has_question_comment = any((row.get('comment') or '').strip() for row in responses_data)
            has_overall_comment = any((row.get('overall_comment') or '').strip() for row in responses_data)
            has_selected_option = any((row.get('selected_option') or '').strip() for row in responses_data)

            headers = [
                "Student Name",
                "Register Number",
                "Department",
                "Year / Section",
                "Subject Code",
                "Subject Name",
                "Staff Name",
                "Question Text",
                "Rating Value",
            ]
            if has_question_comment:
                headers.append("Comment")
            if has_overall_comment:
                headers.append("Overall Comment")
            if has_selected_option:
                headers.append("Selected Option")
            ws.append(headers)
            
            # Data rows - align with dynamic headers.
            for row_data in responses_data:
                row = [
                    row_data['student_name'],
                    row_data['register_number'],
                    row_data['department'],
                    row_data['year_section'],
                    row_data['subject_code'],
                    row_data['subject_name'],
                    row_data['staff_name'],
                    row_data['question_text'],
                    row_data['rating_value'],
                ]
                if has_question_comment:
                    row.append(row_data['comment'])
                if has_overall_comment:
                    row.append(row_data['overall_comment'])
                if has_selected_option:
                    row.append(row_data['selected_option'])
                ws.append(row)
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Return as file download
            from django.http import FileResponse
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="Feedback_Export.xlsx"'
            return response
            
        except Exception as e:
            return Response({
                'detail': f'Error exporting feedback: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class FormExportExcelView(APIView):
    """
    API: Export Form Responses to Excel
    GET /api/feedback/<form_id>/export-excel/
    
    HOD can export all responses for a specific feedback form.
    Returns Excel file with clean IQAC report format.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, form_id):
        try:
            # Check if user has permission to create feedback (HOD only)
            user_permissions = get_user_permissions(request.user)
            if 'feedback.create' not in user_permissions:
                return Response({
                    'detail': 'You do not have permission to export feedback.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get the feedback form
            try:
                feedback_form = get_object_or_404(FeedbackForm, id=form_id)
            except:
                return Response({
                    'detail': 'Feedback form not found.'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Ensure the form was created by the current user
            if feedback_form.created_by != request.user:
                return Response({
                    'detail': 'You can only export responses for forms you created.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Build query with proper joins
            from academics.models import AcademicYear, StudentProfile
            
            qs = FeedbackResponse.objects.filter(
                feedback_form=feedback_form
            ).select_related(
                'feedback_form',
                'feedback_form__department',
                'question',
                'user',
                'teaching_assignment',
                'teaching_assignment__subject',
                'teaching_assignment__curriculum_row',
                'teaching_assignment__elective_subject',
                'teaching_assignment__staff',
                'teaching_assignment__staff__user',
                'teaching_assignment__section',
                'teaching_assignment__section__batch'
            ).prefetch_related(
                'user__student_profile'
            )
            
            # Get data for Excel with strict column structure
            responses_data = []
            current_ay = AcademicYear.objects.filter(is_active=True).first()
            current_acad_year = None
            if current_ay:
                try:
                    current_acad_year = int(str(current_ay.name).split('-')[0])
                except Exception:
                    pass
            
            for response in qs:
                student_name = ""
                register_number = ""
                department_name = ""
                year_section = ""
                subject_code = ""
                subject_name = ""
                staff_name = ""
                comment_value = ""
                overall_comment_value = ""
                
                # Get student info
                if response.user:
                    student_name = response.user.get_full_name() or response.user.username
                    try:
                        student_profile = response.user.student_profile
                        register_number = student_profile.reg_no or ""
                        
                        # Get department with fallback chain
                        if student_profile.home_department:
                            department_name = student_profile.home_department.name or ""
                        elif student_profile.section and student_profile.section.batch and student_profile.section.batch.course:
                            department_name = student_profile.section.batch.course.department.name or ""
                        
                        # Calculate year and get section
                        if student_profile.section and current_acad_year:
                            section_name = student_profile.section.name or ""
                            batch = student_profile.section.batch
                            if batch and batch.start_year:
                                try:
                                    calculated_year = current_acad_year - int(batch.start_year) + 1
                                    year_section = f"{calculated_year} / {section_name}"
                                except:
                                    year_section = f"/ {section_name}"
                    except (AttributeError, StudentProfile.DoesNotExist):
                        pass
                
                # Fallback department from teaching assignment section for cases
                # where student profile department is unavailable.
                if not department_name and response.teaching_assignment and response.teaching_assignment.section:
                    ta_section = response.teaching_assignment.section
                    ta_batch = getattr(ta_section, 'batch', None)
                    ta_course = getattr(ta_batch, 'course', None) if ta_batch else None
                    ta_department = getattr(ta_course, 'department', None) if ta_course else None
                    if ta_department:
                        department_name = ta_department.name or ""

                # Fallback department from form
                if not department_name and feedback_form.department:
                    department_name = feedback_form.department.name or ""
                
                # Get subject and staff from teaching assignment with multi-source fallback
                if response.teaching_assignment:
                    ta = response.teaching_assignment
                    
                    # Extract subject code and name from multiple sources
                    if ta.curriculum_row:
                        subject_code = ta.curriculum_row.course_code or ""
                        subject_name = ta.curriculum_row.course_name or ""
                    elif ta.elective_subject:
                        subject_code = ta.elective_subject.course_code or ""
                        subject_name = ta.elective_subject.course_name or ""
                    elif ta.subject:
                        subject_code = ta.subject.code or ""
                        subject_name = ta.subject.name or ""
                    elif ta.custom_subject:
                        subject_code = ta.custom_subject
                        subject_name = dict(ta._meta.get_field('custom_subject').choices).get(ta.custom_subject, ta.custom_subject)
                    
                    # Extract staff name
                    if ta.staff and ta.staff.user:
                        staff_name = ta.staff.user.get_full_name() or ta.staff.user.username or ""

                    # Fallback year/section from teaching assignment section.
                    if not year_section and ta.section:
                        ta_section_name = ta.section.name or ""
                        ta_batch = getattr(ta.section, 'batch', None)
                        ta_year_text = ""
                        if ta_batch and getattr(ta_batch, 'start_year', None) and current_acad_year:
                            try:
                                ta_year_text = str(current_acad_year - int(ta_batch.start_year) + 1)
                            except Exception:
                                ta_year_text = ""
                        year_section = f"{ta_year_text} / {ta_section_name}" if (ta_year_text or ta_section_name) else ""
                
                # Get question text
                question_text = response.question.question if response.question else ""
                
                # Apply conditional display logic: show either question-wise comment
                # or overall comment, never both.
                question_comment = (response.answer_text or "").strip()
                common_comment = (response.common_comment or "").strip()
                if question_comment:
                    comment_value = question_comment
                    overall_comment_value = ""
                elif common_comment:
                    comment_value = ""
                    overall_comment_value = common_comment
                else:
                    comment_value = ""
                    overall_comment_value = ""

                # Show selected option only when real value is available.
                selected_option_value = (response.selected_option_text or "").strip()
                
                # Collect data row
                responses_data.append({
                    'student_name': student_name,
                    'register_number': register_number,
                    'department': department_name,
                    'year_section': year_section,
                    'subject_code': subject_code,
                    'subject_name': subject_name,
                    'staff_name': staff_name,
                    'question_text': question_text,
                    'rating_value': response.answer_star or "",
                    'comment': comment_value,
                    'overall_comment': overall_comment_value,
                    'selected_option': selected_option_value,
                })
            
            # Generate Excel file
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Feedback Responses'
            
            # Build optional columns dynamically based on actual feedback content.
            has_question_comment = any((row.get('comment') or '').strip() for row in responses_data)
            has_overall_comment = any((row.get('overall_comment') or '').strip() for row in responses_data)
            has_selected_option = any((row.get('selected_option') or '').strip() for row in responses_data)

            headers = [
                "Student Name",
                "Register Number",
                "Department",
                "Year / Section",
                "Subject Code",
                "Subject Name",
                "Staff Name",
                "Question Text",
                "Rating Value",
            ]
            if has_question_comment:
                headers.append("Comment")
            if has_overall_comment:
                headers.append("Overall Comment")
            if has_selected_option:
                headers.append("Selected Option")
            ws.append(headers)
            
            # Data rows - align with dynamic headers.
            for row_data in responses_data:
                row = [
                    row_data['student_name'],
                    row_data['register_number'],
                    row_data['department'],
                    row_data['year_section'],
                    row_data['subject_code'],
                    row_data['subject_name'],
                    row_data['staff_name'],
                    row_data['question_text'],
                    row_data['rating_value'],
                ]
                if has_question_comment:
                    row.append(row_data['comment'])
                if has_overall_comment:
                    row.append(row_data['overall_comment'])
                if has_selected_option:
                    row.append(row_data['selected_option'])
                ws.append(row)
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Return as file download
            from django.http import FileResponse
            filename = f"Feedback_{feedback_form.id}_{feedback_form.get_type_display()}.xlsx"
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            return Response({
                'detail': f'Error exporting feedback: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class SubjectWiseReportView(APIView):
    """
    API: Subject Wise Report
    GET /api/feedback/subject-wise-report/
    
    Returns aggregated feedback report grouped by subject + staff.
    Shows average rating and total rating count for each subject-staff combination.
    Supports filtering by department_ids, years, and optionally form_id.
    
    Example response:
    [
        {
            'subject_code': 'ADB1322',
            'subject_name': 'Big Data Analytics',
            'staff_name': 'Reetha Jeyarani',
            'average_rating': 4.25,
            'total_ratings': 63
        }
    ]
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            scope = get_feedback_department_scope(request.user)
            if not scope.get('allowed'):
                return Response({
                    'detail': 'You do not have permission to view feedback reports.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get filter parameters
            all_departments = request.GET.get('all_departments', 'false').lower() == 'true'
            department_ids = request.GET.getlist('department_ids[]')
            if not department_ids:
                department_ids_str = request.GET.get('department_ids', '')
                if department_ids_str:
                    department_ids = [int(d) for d in department_ids_str.split(',') if d.strip()]
            else:
                department_ids = [int(d) for d in department_ids if d.strip()]
            
            years_str = request.GET.get('years', '')
            years = [int(y) for y in years_str.split(',') if y.strip()] if years_str else []
            
            form_id = request.GET.get('form_id')
            
            # Start with responses that have ratings
            qs = FeedbackResponse.objects.filter(
                answer_star__isnull=False
            ).select_related(
                'feedback_form',
                'feedback_form__department',
                'user',
                'teaching_assignment',
                'teaching_assignment__subject',
                'teaching_assignment__curriculum_row',
                'teaching_assignment__elective_subject',
                'teaching_assignment__staff',
                'teaching_assignment__staff__user',
                'teaching_assignment__section',
                'teaching_assignment__section__batch'
            ).prefetch_related(
                'user__student_profile'
            )
            
            # Apply department scope filter
            qs = apply_department_scope_filter(qs, scope, field_name='feedback_form__department_id')
            
            # Filter by specific departments if requested
            if not scope.get('all_departments'):
                # Own-department users are already scoped
                all_departments = True
            
            if not all_departments and department_ids:
                qs = qs.filter(feedback_form__department_id__in=department_ids)
            
            # Filter by years if specified
            if years:
                year_filter = Q()
                for year in years:
                    year_filter |= Q(feedback_form__years__contains=[year])
                    year_filter |= Q(feedback_form__year=year)
                qs = qs.filter(year_filter)
            
            # Filter by form_id if specified
            if form_id:
                try:
                    form_id = int(form_id)
                    qs = qs.filter(feedback_form_id=form_id)
                except (ValueError, TypeError):
                    pass
            
            # Group by subject and staff, calculate average and count
            from django.db.models import Avg, Count
            
            report_data = []
            
            # Create a dictionary to aggregate data (subject_code + staff_name -> data)
            aggregated = {}
            
            for response in qs:
                # Skip if no teaching assignment
                if not response.teaching_assignment:
                    continue
                
                ta = response.teaching_assignment
                
                # Get subject code and name
                subject_code = None
                subject_name = None
                
                if ta.curriculum_row:
                    subject_code = ta.curriculum_row.course_code
                    subject_name = ta.curriculum_row.course_name
                elif ta.elective_subject:
                    subject_code = ta.elective_subject.course_code
                    subject_name = ta.elective_subject.course_name
                elif ta.subject:
                    subject_code = ta.subject.code
                    subject_name = ta.subject.name
                elif ta.custom_subject:
                    subject_code = ta.custom_subject
                    subject_name = ta.get_custom_subject_display()
                
                if not subject_code or not subject_name:
                    continue
                
                # Get staff name
                staff_name = None
                if ta.staff and ta.staff.user:
                    staff_name = ta.staff.user.get_full_name() or ta.staff.user.username
                
                if not staff_name:
                    staff_name = "Unknown Staff"
                
                # Create a unique key
                key = f"{subject_code}|{subject_name}|{staff_name}"
                
                # Initialize if not exists
                if key not in aggregated:
                    aggregated[key] = {
                        'subject_code': subject_code,
                        'subject_name': subject_name,
                        'staff_name': staff_name,
                        'ratings': []
                    }
                
                # Add rating
                aggregated[key]['ratings'].append(response.answer_star)
            
            # Calculate averages and counts
            for key, data in aggregated.items():
                if data['ratings']:
                    total_ratings = len(data['ratings'])
                    average_rating = sum(data['ratings']) / total_ratings
                    # Round to 2 decimals
                    average_rating = round(average_rating, 2)
                    
                    report_data.append({
                        'subject_code': data['subject_code'],
                        'subject_name': data['subject_name'],
                        'staff_name': data['staff_name'],
                        'average_rating': average_rating,
                        'total_ratings': total_ratings
                    })
            
            # Sort by subject name, then by staff name
            report_data.sort(key=lambda x: (x['subject_name'], x['staff_name']))
            
            # Generate Excel file
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Subject Wise Report'
            
            # Add headers
            headers = [
                'Subject Code',
                'Subject Name',
                'Staff Name',
                'Average Rating',
                'Total Ratings'
            ]
            ws.append(headers)
            
            # Add data rows
            for item in report_data:
                ws.append([
                    item['subject_code'],
                    item['subject_name'],
                    item['staff_name'],
                    item['average_rating'],
                    item['total_ratings']
                ])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 15
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Return as file download
            from django.http import FileResponse
            from datetime import datetime
            filename = f"Subject_Wise_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            logger.exception('[SubjectWiseReportView] ERROR')
            return Response({
                'detail': f'Error generating subject wise report: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class SubjectsFilterView(APIView):
    """
    API: Get Subjects for filtering
    GET /api/feedback/subjects-filter/?dept_ids=1,2&years=2,3
    
    Returns distinct subjects from TeachingAssignment based on selected departments and years.
    Independent of feedback submission status - returns all assigned subjects.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            scope = get_feedback_department_scope(request.user)
            if not scope.get('allowed'):
                return Response({
                    'detail': 'You do not have permission to view feedback data.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get filter parameters
            dept_ids_str = request.GET.get('dept_ids', '')
            dept_ids = [int(d) for d in dept_ids_str.split(',') if d.strip()] if dept_ids_str else []
            
            years_str = request.GET.get('years', '')
            years = [int(y) for y in years_str.split(',') if y.strip()] if years_str else []
            
            # Query TeachingAssignment directly, not FeedbackResponse
            # This returns all assigned subjects regardless of feedback submission status
            from academics.models import TeachingAssignment
            
            qs = TeachingAssignment.objects.select_related(
                'subject',
                'curriculum_row',
                'elective_subject',
                'department'
            )
            
            # Filter by specific departments if provided
            if dept_ids:
                qs = qs.filter(department_id__in=dept_ids)
            
            # Filter by years if provided
            if years:
                qs = qs.filter(year__in=years)
            
            # Get distinct subjects
            subjects_data = []
            seen_codes = set()
            
            # Use values_list to get distinct subjects efficiently
            subjects_qs = qs.values_list(
                'subject__code',
                'subject__name',
                'curriculum_row__course_code',
                'curriculum_row__course_name',
                'elective_subject__course_code',
                'elective_subject__course_name'
            ).distinct()
            
            for row in subjects_qs:
                subject_code = row[0] or row[2] or row[4]  # subject code, curriculum code, or elective code
                subject_name = row[1] or row[3] or row[5]   # subject name, curriculum name, or elective name
                
                if subject_code and subject_code not in seen_codes:
                    seen_codes.add(subject_code)
                    subjects_data.append({
                        'code': subject_code,
                        'name': subject_name or 'Unknown'
                    })
            
            # Sort by name
            subjects_data.sort(key=lambda x: x['name'])
            
            return Response({
                'success': True,
                'subjects': subjects_data
            })
            
        except Exception as e:
            logger.exception('[SubjectsFilterView] ERROR: %s', str(e))
            return Response({
                'detail': f'Error fetching subjects: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class BulkSubjectWiseReportView(APIView):
    """
    API: Bulk Subject Wise Report (with modal filters)
    GET /api/feedback/bulk-subject-wise-report/
    
    Returns aggregated feedback report grouped by department, year, staff, and subject.
    Supports filtering by department_ids, years, and subject_codes.
    
    Query Parameters:
    - all_departments=true/false
    - department_ids[]=1,2,3
    - years[]=1,2
    - subject_codes[]=ADB1322,CS101 (optional)
    
    Output columns:
    - Department
    - Year
    - Staff Name
    - Subject Code
    - Subject Name
    - Total Students (Count of distinct students)
    - Total Stars Given (Sum of all ratings)
    - Average Rating (Avg of ratings, rounded to 2 decimals)
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            from django.db.models import Sum, Count, Avg
            
            scope = get_feedback_department_scope(request.user)
            if not scope.get('allowed'):
                return Response({
                    'detail': 'You do not have permission to view feedback reports.'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get filter parameters
            all_departments = request.GET.get('all_departments', 'false').lower() == 'true'
            
            department_ids = request.GET.getlist('department_ids[]')
            if not department_ids:
                department_ids_str = request.GET.get('department_ids', '')
                if department_ids_str:
                    department_ids = [int(d) for d in department_ids_str.split(',') if d.strip()]
            else:
                department_ids = [int(d) for d in department_ids if d.strip()]
            
            years_str = request.GET.get('years', '')
            years = [int(y) for y in years_str.split(',') if y.strip()] if years_str else []
            
            subject_codes = request.GET.getlist('subject_codes[]')
            if not subject_codes:
                subject_codes_str = request.GET.get('subject_codes', '')
                if subject_codes_str:
                    subject_codes = [s.strip() for s in subject_codes_str.split(',') if s.strip()]
            
            # Start with responses that have ratings
            qs = FeedbackResponse.objects.filter(
                answer_star__isnull=False
            ).select_related(
                'feedback_form',
                'feedback_form__department',
                'user',
                'teaching_assignment',
                'teaching_assignment__subject',
                'teaching_assignment__curriculum_row',
                'teaching_assignment__elective_subject',
                'teaching_assignment__staff',
                'teaching_assignment__staff__user'
            ).prefetch_related(
                'user__student_profile',
                'feedback_form__years'
            )
            
            # Apply department scope filter
            qs = apply_department_scope_filter(qs, scope, field_name='feedback_form__department_id')
            
            # Filter by specific departments if requested
            if not all_departments and department_ids:
                qs = qs.filter(feedback_form__department_id__in=department_ids)
            
            # Filter by years if specified
            if years:
                year_filter = Q()
                for year in years:
                    year_filter |= Q(feedback_form__years__contains=[year])
                qs = qs.filter(year_filter)
            
            # Build subject filter based on curriculum_row and elective_subject codes
            if subject_codes:
                subject_filter = Q()
                subject_filter |= Q(teaching_assignment__curriculum_row__course_code__in=subject_codes)
                subject_filter |= Q(teaching_assignment__elective_subject__course_code__in=subject_codes)
                qs = qs.filter(subject_filter)
            
            # Process responses manually to handle department and year properly
            report_data = []
            processed_keys = set()
            
            for response in qs:
                if not response.teaching_assignment:
                    continue
                
                ta = response.teaching_assignment
                
                # Get department name
                dept_name = response.feedback_form.department.name if response.feedback_form.department else 'Unknown'
                
                # Get year from form
                form_years = response.feedback_form.years or [1]
                years_to_process = years if years else form_years
                
                # Get staff name
                if ta.staff and ta.staff.user:
                    staff_name = ta.staff.user.get_full_name() or ta.staff.user.username
                else:
                    staff_name = 'Unknown Staff'
                
                # Get subject info
                subject_code = None
                subject_name = None
                
                if ta.curriculum_row:
                    subject_code = ta.curriculum_row.course_code
                    subject_name = ta.curriculum_row.course_name
                elif ta.elective_subject:
                    subject_code = ta.elective_subject.course_code
                    subject_name = ta.elective_subject.course_name
                elif ta.subject:
                    subject_code = ta.subject.code
                    subject_name = ta.subject.name
                
                if not subject_code:
                    continue
                
                # Process each year
                for year in years_to_process:
                    if isinstance(year, (list, tuple)):
                        year = year[0] if year else 1
                    
                    key = f"{dept_name}|{year}|{staff_name}|{subject_code}|{subject_name}"
                    
                    if key not in processed_keys:
                        processed_keys.add(key)
                        
                        # Aggregate data for this combination
                        subset = qs.filter(
                            feedback_form__department_id=response.feedback_form.department_id,
                            teaching_assignment__staff_id=ta.staff_id,
                            teaching_assignment__curriculum_row_id=ta.curriculum_row_id if ta.curriculum_row else None,
                            teaching_assignment__elective_subject_id=ta.elective_subject_id if ta.elective_subject else None
                        )
                        
                        # Apply year filter to subset
                        year_filter = Q(feedback_form__years__contains=[year])
                        subset = subset.filter(year_filter)
                        
                        agg = subset.aggregate(
                            total_students=Count('user_id', distinct=True),
                            total_stars=Sum('answer_star'),
                            avg_rating=Avg('answer_star')
                        )
                        
                        total_students = agg.get('total_students') or 0
                        total_stars = agg.get('total_stars') or 0
                        avg_rating = agg.get('avg_rating') or 0
                        
                        if avg_rating:
                            avg_rating = round(float(avg_rating), 2)
                        
                        report_data.append({
                            'department': dept_name,
                            'year': year,
                            'staff_name': staff_name,
                            'subject_code': subject_code,
                            'subject_name': subject_name,
                            'total_students': total_students,
                            'total_stars': total_stars,
                            'average_rating': avg_rating
                        })
            
            # Remove duplicates and sort
            final_data = []
            seen = set()
            for item in report_data:
                key = (item['department'], item['year'], item['staff_name'], item['subject_code'])
                if key not in seen:
                    seen.add(key)
                    final_data.append(item)
            
            final_data.sort(key=lambda x: (x['department'], x['year'], x['staff_name'], x['subject_name']))
            
            # Generate Excel file
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Subject Wise Report'
            
            # Add headers
            headers = [
                'Department',
                'Year',
                'Staff Name',
                'Subject Code',
                'Subject Name',
                'Total Students',
                'Total Stars',
                'Average Rating'
            ]
            ws.append(headers)
            
            # Add data rows
            for item in final_data:
                ws.append([
                    item['department'],
                    f"Year {item['year']}" if isinstance(item['year'], int) else item['year'],
                    item['staff_name'],
                    item['subject_code'],
                    item['subject_name'],
                    item['total_students'],
                    item['total_stars'],
                    item['average_rating']
                ])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 14
            ws.column_dimensions['H'].width = 16
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Return as file download
            from django.http import FileResponse
            from datetime import datetime
            filename = f"Subject_Wise_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            logger.exception('[BulkSubjectWiseReportView] ERROR')
            return Response({
                'detail': f'Error generating subject wise report: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

